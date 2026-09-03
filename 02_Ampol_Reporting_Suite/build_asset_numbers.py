#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=====================================================================
COATES | AMPOL - ASSET NUMBERS (button 17)
=====================================================================
Author: Andrew Fisher | POWERED BY SITEIQ

WHY (03 Sep 2026, Andrew): when new gear is added to SiteIQ the store
needs to know which asset number to give it - the main number (AMP200)
and the next /001 under it - and what that number is already used for.
This workbook answers that from the register itself, fresh every
morning, so a number is never guessed and never reused:

    Asset numbers   one row per main number (the part before the /):
                    the description as SiteIQ writes it, how many items
                    carry the number, the highest /NNN used, the next
                    /NNN to use, the next ten, and any unused numbers
                    below the highest
    Next by family  for each family of numbers (AMP, WG, CTX, SD ...):
                    the highest main number used and the next new one
    All barcodes    every barcode as written in SiteIQ with its main
                    number, /NNN, description and status
    Other formats   barcodes that do not follow MAIN/NNN (the Coates
                    fleet numbers) - listed so nothing is missed
    New to address  every item on the register that Ampol_Master.xlsx
                    does not know yet - no corrected description for
                    its barcode, or no price for its description - by
                    the same rules the reports use. Add the row to the
                    master and the item drops off this tab.
    Read me         how to use it

Sources: RENTAL_STOCK (the register, every status) plus every barcode
the TRANSACTIONS log has seen this year that is no longer on the
register - a number that has left the register is still taken. Nothing
is typed in, nothing is estimated. The file is GENERATED - edit
Ampol_Master.xlsx, never this one.
"""
import os
import re
import sys
from collections import Counter, OrderedDict, defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import ampol_master
import ampol_names
import ampol_paths
import pull_diff
import txn_insights

SLASH_RE = re.compile(r"^(.+?)/(\d+)$")
FAMILY_RE = re.compile(r"^([A-Za-z&]+?)(\d+)$")


def load():
    reg_path = ampol_paths.find_data("RENTAL_STOCK*.xlsx")
    if not reg_path:
        sys.exit("ERROR: no RENTAL_STOCK*.xlsx in Data\\SiteIQ - run 12_PULL_SITEIQ_EXPORTS first.")
    reg = pull_diff.load_register(reg_path)
    asat = pull_diff.reference_pull_time(reg_path)
    rows = []
    for r in reg.values():
        rows.append({"barcode": r["barcode"], "desc": r["desc_raw"], "status": r["status"],
                     "company": r["company"], "source": "register"})
    on_reg = {r["barcode"].upper() for r in rows}
    tx_path = ampol_paths.find_data("TRANSACTIONS*.xlsx")
    log_only = 0
    if tx_path:
        tx, window = txn_insights.load_transactions(tx_path)
        seen = {}
        for t in tx:
            bc = (t["bc"] or "").strip()
            if bc and bc.upper() not in on_reg:
                seen.setdefault(bc.upper(), (bc, t["desc_raw"]))
        for bc, desc in seen.values():
            rows.append({"barcode": bc, "desc": desc, "status": "not on the register now",
                         "company": "", "source": "log only (this year)"})
        log_only = len(seen)
    return rows, asat, reg_path, tx_path, log_only


def new_to_address(rows, reg_path):
    """Items the master does not know yet, judged the way the stocktake
    engine judges them (build_stocktake_compliance_tool.load_corrections /
    load_pricing / price_for), so this tab and the reports never
    disagree. 'New since last pull' marks a barcode absent from the
    previous register pull in Data\\SiteIQ\\previous."""
    import build_stocktake_compliance_tool as eng
    fx_path, fx_sheet = ampol_master.locate("descriptions", "New_Descriptions*.xlsx", "*Descriptions*.xlsx")
    pr_path, pr_sheet = ampol_master.locate("pricing", "*Pricing*.xlsx")
    bc_map, desc_map = eng.load_corrections(fx_path, fx_sheet) if fx_path else ({}, {})
    exact, stripped, _ = eng.load_pricing(pr_path, pr_sheet) if pr_path else ({}, {}, [])
    prev_bcs = set()
    try:
        cur_time = pull_diff.reference_pull_time(reg_path)
        prev_path, _t = pull_diff.find_previous_pull(reg_path, cur_time)
        if prev_path:
            prev_bcs = {b.upper() for b in pull_diff.load_register(prev_path)}
    except Exception:
        prev_bcs = set()
    out = []
    for r in rows:
        if r["source"] != "register":
            continue
        raw = r["desc"]
        fixed = bc_map.get(eng.norm(r["barcode"])) or desc_map.get(eng.norm(raw))
        has_desc = bool(fixed) or ampol_names.product_name(raw) is not None
        priced = eng.price_for(raw, fixed or raw, exact, stripped) is not None
        if has_desc and priced:
            continue
        needs = "Description + Price" if not has_desc and not priced else ("Description" if not has_desc else "Price")
        out.append({"barcode": r["barcode"], "prefix": r.get("prefix") or "", "desc": raw, "status": r["status"],
                    "who": r["company"] if r["status"].lower() == "on hire" else "",
                    "new": "Yes" if (prev_bcs and r["barcode"].upper() not in prev_bcs) else "",
                    "needs": needs,
                    "where": ("Descriptions tab (barcode + corrected description)" if needs == "Description" else
                              "Pricing tab (description exactly as SiteIQ writes it + price)" if needs == "Price" else
                              "Descriptions tab and Pricing tab")})
    # new gear first, then what touches the value figures (a price), then the tidy-ups
    order = {"Description + Price": 0, "Price": 1, "Description": 2}
    out.sort(key=lambda r: (r["new"] != "Yes", order[r["needs"]], ampol_names.sort_key(r["prefix"] or r["barcode"]),
                            ampol_names.sort_key(r["barcode"])))
    return out, bool(prev_bcs)


def split(barcode):
    """('AMP200', 4, '004') for AMP200/004; (None, None, None) otherwise."""
    m = SLASH_RE.match(barcode.strip())
    if not m:
        return None, None, None
    return m.group(1).upper(), int(m.group(2)), m.group(2)


def fmt(prefix, n, width):
    return f"{prefix}/{n:0{width}d}"


def build(rows):
    by_prefix = defaultdict(list)
    other = []
    for r in rows:
        p, n, raw = split(r["barcode"])
        r["prefix"], r["n"], r["suffix"] = p, n, raw
        if p is None:
            other.append(r)
        else:
            by_prefix[p].append(r)
    prefix_rows = []
    for p in sorted(by_prefix, key=ampol_names.sort_key):
        items = by_prefix[p]
        reg_items = [r for r in items if r["source"] == "register"]
        descs = Counter(r["desc"] for r in reg_items if r["desc"]) or Counter(r["desc"] for r in items if r["desc"])
        top, top_n = (descs.most_common(1)[0] if descs else ("", 0))
        used = sorted({r["n"] for r in items})
        width = Counter(len(r["suffix"]) for r in items).most_common(1)[0][0]
        hi = used[-1]
        nxt = hi + 1
        gaps = [n for n in range(1, hi) if n not in set(used)]
        prefix_rows.append({
            "prefix": p, "desc": top, "other_descs": len(descs) - 1 if descs else 0,
            "items": len(reg_items), "log_only": len(items) - len(reg_items),
            "highest": fmt(p, hi, width), "next": fmt(p, nxt, width),
            "next_ten": f"{fmt(p, nxt, width)} to {fmt(p, nxt + 9, width)}",
            "gaps_n": len(gaps),
            "gaps": ", ".join(f"/{g:0{width}d}" for g in gaps[:10]) + (f" and {len(gaps) - 10} more" if len(gaps) > 10 else ""),
            "statuses": ", ".join(f"{s} {n}" for s, n in Counter(r["status"] for r in reg_items).most_common()),
        })
    # families: the letters of the main number and its digits
    fam = defaultdict(list)
    for p in by_prefix:
        m = FAMILY_RE.match(p)
        if m:
            fam[m.group(1).upper()].append((int(m.group(2)), len(m.group(2)), p))
    fam_rows = []
    for f in sorted(fam, key=ampol_names.sort_key):
        nums = fam[f]
        width = Counter(w for _, w, _ in nums).most_common(1)[0][0]
        run = sorted(n for n, w, _ in nums if w == width)
        hi = run[-1] if run else max(n for n, _, _ in nums)
        used = set(run)
        free = [n for n in range(1, hi) if n not in used]
        fam_rows.append({
            "family": f, "prefixes": len(nums), "items": sum(len(by_prefix[p]) for _, _, p in nums),
            "highest": f"{f}{hi:0{width}d}", "next": f"{f}{hi + 1:0{width}d}",
            "free_n": len(free),
            "free": ", ".join(f"{f}{n:0{width}d}" for n in free[:10]) + (f" and {len(free) - 10} more" if len(free) > 10 else ""),
            "odd": ", ".join(sorted(p for n, w, p in nums if w != width))[:120],
        })
    return prefix_rows, fam_rows, other


def write(path, prefix_rows, fam_rows, rows, other, asat, reg_path, tx_path, log_only, todo, have_prev):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hf, hfill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F2A37")

    def sheet(name, headers, data, widths, wrap=()):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for c in ws[1]:
            c.font, c.fill = hf, hfill
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for row in data:
            ws.append(row)
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w
        for i in wrap:
            for cell in ws.iter_rows(min_row=2, min_col=i + 1, max_col=i + 1):
                cell[0].alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        if data:
            ws.auto_filter.ref = ws.dimensions
        return ws

    rm = wb.create_sheet("Read me")
    for line in [
        "ASSET NUMBERS - which number to give new gear, and what every number is already used for",
        "Author: Andrew Fisher | POWERED BY SITEIQ",
        "",
        f"Generated {datetime.now():%d %b %Y %H:%M} from the SiteIQ register pulled {asat:%d %b %Y %H:%M}"
        + (" and the TRANSACTIONS log for this year." if tx_path else "."),
        "This file is GENERATED every run - do not edit it. Descriptions and prices live in Ampol_Master.xlsx.",
        "",
        f"New to address - {len(todo):,} items on the register that Ampol_Master.xlsx does not know yet: no corrected",
        "                 description for the barcode, or no price for the description (judged exactly as the",
        "                 reports judge it). Add the row to the master and the item drops off this tab next run."
        + ("" if have_prev else " No earlier pull was found, so 'New since last pull' is blank this run."),
        "Asset numbers  - one row per main number (the part before the /). 'Next number' is the highest /NNN",
        "                 used plus one, in the width that number uses (/001, /002 ...). 'Unused below highest'",
        "                 lists numbers that were skipped - use them only if you are sure they were never issued.",
        "Next by family - the letters of the main number (AMP, WG, CTX, SD ...): the highest main number in use",
        "                 and the next new one. Main numbers with an odd digit width are listed, not counted.",
        "All barcodes   - every barcode exactly as SiteIQ writes it, with its description and status.",
        "Other formats  - barcodes without a / (the Coates fleet numbers) - listed so nothing is missed.",
        "",
        "A number seen in the transaction log this year but no longer on the register is still counted as",
        "used - it may come back, and SiteIQ remembers it. Nothing here is typed in or estimated.",
        f"Register rows: {sum(1 for r in rows if r['source'] == 'register'):,}. Log-only barcodes: {log_only:,}. "
        f"Main numbers: {len(prefix_rows):,}. Other formats: {len(other):,}.",
    ]:
        rm.append([line])
    rm.column_dimensions["A"].width = 118
    rm["A1"].font = Font(bold=True, size=13)

    sheet("New to address",
          ["Barcode (as written)", "Asset number", "Description (as in RENTAL_STOCK)", "Status", "Company",
           "New since last pull", "Needs", "Where to add it in Ampol_Master.xlsx"],
          [[r["barcode"], r["prefix"], r["desc"], r["status"], r["who"], r["new"], r["needs"], r["where"]] for r in todo],
          [20, 16, 52, 22, 22, 12, 20, 52], wrap=(2, 7))
    sheet("Asset numbers",
          ["Asset number", "Description (as in RENTAL_STOCK)", "Other descriptions under it", "Items on register",
           "In log only", "Highest used", "Next number", "Next ten", "Unused below highest", "Which ones", "Status of items"],
          [[r["prefix"], r["desc"], r["other_descs"] or "", r["items"], r["log_only"] or "", r["highest"], r["next"],
            r["next_ten"], r["gaps_n"] or "", r["gaps"], r["statuses"]] for r in prefix_rows],
          [16, 46, 12, 10, 10, 16, 16, 30, 12, 40, 34], wrap=(1, 9, 10))
    sheet("Next by family",
          ["Family", "Main numbers in use", "Items", "Highest main number", "Next new main number",
           "Unused main numbers below it", "Which ones", "Odd widths (listed, not counted)"],
          [[r["family"], r["prefixes"], r["items"], r["highest"], r["next"], r["free_n"] or "", r["free"], r["odd"]]
           for r in fam_rows],
          [14, 12, 10, 18, 20, 14, 50, 40], wrap=(6, 7))
    allrows = sorted((r for r in rows if r["prefix"]), key=lambda r: (ampol_names.sort_key(r["prefix"]), r["n"]))
    sheet("All barcodes",
          ["Barcode (as written)", "Asset number", "/NNN", "Description (as in RENTAL_STOCK)", "Status", "Company", "Source"],
          [[r["barcode"], r["prefix"], r["suffix"], r["desc"], r["status"], r["company"], r["source"]] for r in allrows],
          [20, 16, 8, 52, 24, 24, 20])
    sheet("Other formats",
          ["Barcode (as written)", "Description (as in RENTAL_STOCK)", "Status", "Company", "Source"],
          [[r["barcode"], r["desc"], r["status"], r["company"], r["source"]]
           for r in sorted(other, key=lambda r: ampol_names.sort_key(r["barcode"]))],
          [22, 52, 24, 24, 20])
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


def main():
    print("=" * 66)
    print(" COATES | AMPOL - ASSET NUMBERS")
    print("=" * 66)
    rows, asat, reg_path, tx_path, log_only = load()
    prefix_rows, fam_rows, other = build(rows)
    todo, have_prev = new_to_address(rows, reg_path)
    out_dir = ampol_paths.day_folder("Asset_Numbers")
    stem = ampol_names.report_stem("asset_numbers")
    path = os.path.join(out_dir, f"{stem}.xlsx")
    write(path, prefix_rows, fam_rows, rows, other, asat, reg_path, tx_path, log_only, todo, have_prev)
    print(f"Register           : {os.path.relpath(reg_path, ampol_paths.suite_dir())}  (pulled {asat:%d %b %Y %H:%M})")
    print(f"Transactions       : {os.path.relpath(tx_path, ampol_paths.suite_dir()) if tx_path else 'not found - log-only barcodes not counted'}")
    print(f"Barcodes           : {sum(1 for r in rows if r['source'] == 'register'):,} on the register, {log_only:,} in the log only")
    print(f"Main numbers       : {len(prefix_rows):,} (MAIN/NNN); other formats {len(other):,}")
    print(f"New to address     : {len(todo):,} items the master does not know yet "
          f"({sum(1 for t in todo if t['needs'] == 'Description'):,} need a description, "
          f"{sum(1 for t in todo if t['needs'] == 'Price'):,} a price, "
          f"{sum(1 for t in todo if t['needs'] == 'Description + Price'):,} both; "
          f"{sum(1 for t in todo if t['new'] == 'Yes'):,} new since the last pull)")
    for f in fam_rows[:6]:
        print(f"  {f['family']:10} highest {f['highest']:12} next {f['next']}")
    print(f"Workbook           : {path}")
    print("Done. The Coates Way - consistent execution, every day.")


if __name__ == "__main__":
    main()
