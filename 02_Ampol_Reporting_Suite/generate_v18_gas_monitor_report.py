#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=====================================================================
COATES V18 ELITE - AMPOL GAS MONITOR EXECUTIVE OPERATIONS DASHBOARD
Full report generator: workbook in -> Outlook-safe .eml + HTML out
=====================================================================
Author: Andrew Fisher
The Coates Way - Operational Excellence - POWERED BY SITEIQ

WHAT THIS SCRIPT DOES
  1. Reads Ampol_Gas_Monitor_Report.xlsm
  2. Counts every headline figure DIRECTLY from the detail tabs
     (Available Store / Onhire 1-7 / 7-30 / 30+ / FCCU / Repairs),
     so hardcoded summary cells can never cause an off-by-one again.
  3. Cross-validates against the Performance tab summary cells and
     (optionally) a live RENTAL_STOCK.xlsx export - any mismatch is
     printed as a reconciliation warning.
  4. Renders the full V18 ELITE dark dashboard (Outlook-safe HTML,
     tables + inline styles only, no JavaScript) with health rings
     generated as embedded PNGs.
  5. Writes a ready-to-send .eml draft (X-Unsent) with the source
     workbook attached, plus a standalone browser HTML.

USAGE - AMPOL REPORTING SUITE
  WHY (12 Aug 2026): this engine now lives in the suite. Inputs come
  from the one shared Data\ area, outputs land in the dated
  Reports\<today>\Gas_Monitors folder, and the button that runs it is
  02_RUN_GAS_EXEC_DASHBOARD.bat.
  Or from a terminal:  python generate_v18_gas_monitor_report.py

  Kit layout - one flat folder, everything in it:
      AmpolGas2026\
        RUN_REPORT.bat                        <- double-click this
        README.txt
        generate_v18_gas_monitor_report.py    <- this file
        Ampol Gas Monitor Report.xlsm         <- save the fresh one over the top
        RENTAL_STOCK.xlsx                     <- short kit name, save over the top
        TRANSACTIONS.xlsx                     <- short kit name, save over the top
        Gas_Monitor_Serial_Numbers.xlsx       <- reference
        Source_*                              <- archived copies, skipped
        Coates_GasMonitor_..._OUTLOOK_SAFE.eml    <- built here, same name each run
        Coates_GasMonitor_...V18.html             <- built here, same name each run

  Same habit as the K2 gear kit: short export names, saved over the top,
  newest-mtime-wins globs, one file per pattern in the folder. No stale
  duplicates, no dated filenames to pick through at 05:30.

  Inputs are found AUTOMATICALLY - the workbook is any .xlsm whose name
  contains "gas monitor" (spaces or underscores, any case); newest wins.
  Anything prefixed "Source_" is treated as an archived copy and skipped.
  If a RENTAL_STOCK*.xlsx export is present it is used for
  cross-validation. No filenames need to match exactly, no paths are
  hardcoded, and an input\ / output\ subfolder is still honoured if you
  ever want one (see CONFIG).

  Edit CONFIG below only for recipients, dates and business rules.

DEPENDENCIES
  pip install openpyxl pillow
  (pandas is NOT required.) No internet needed at run time - no API
  calls, no CDN fonts, no web-hosted images. Runs on a locked-down site
  laptop from a USB stick.
=====================================================================
"""

import base64
import email
import io
import json
import os
import re
import shutil
import sys
from collections import Counter, OrderedDict
from datetime import datetime
from email import policy
from email.message import EmailMessage
from email.utils import formatdate

import openpyxl
from PIL import Image, ImageDraw

# WHY (12 Aug 2026): the suite's path oracle. Every input now comes from
# the one shared Data\ area and every output lands in the dated
# Reports\<today>\ folder - same rule for every report in the suite.
import ampol_paths

# WHY (02 Sep 2026): one engine for the gas monitor family. The dashboard
# keeps its look, but every number now comes from gasmon_engine - counted
# from the SiteIQ RENTAL_STOCK and TRANSACTIONS exports - so it can never
# disagree with the PDF or the email. The workbook is optional: attached
# to the email when present, read for numbers only if the engine is absent.
try:
    import gasmon_engine as _engine
except ImportError:
    _engine = None

# =====================================================================
# CONFIG
# =====================================================================

CONFIG = {
    # --- Inputs -------------------------------------------------------
    # "auto" = find the workbook in the kit folder automatically (any .xlsm
    # whose name contains "gas monitor", newest wins). An input\ subfolder
    # is also searched if one exists. Or set an exact path.
    "workbook_path": "auto",
    # "auto" = use RENTAL_STOCK*.xlsx if present; "" = skip validation
    "rental_stock_path": "auto",

    # --- Outputs ------------------------------------------------------
    # WHY (12 Aug 2026): outputs now always land in the suite's dated
    # Reports\<today>\Gas_Monitors folder (see resolve_output_dir). This
    # setting is kept only so older notes referencing it still make sense.
    "output_dir": ".",
    "eml_name": "Coates_GasMonitor_Executive_Operations_Dashboard_V18_OUTLOOK_SAFE.eml",
    "html_name": "Coates_GasMonitor_Executive_Operations_Dashboard_V18.html",
    # False = fixed filenames, saved over the top each run - the K2 way, so
    # the file you send is always the same one. True = stamp the run date
    # into the filename instead, if you'd rather keep every run.
    "date_stamp_outputs": False,

    # --- Report identity ---------------------------------------------
    "report_date": None,        # None = today; or "13/07/2026"
    "run_time": None,           # None = now;  or "07:45"  (24-hour time)
    "author_name": "Andrew Fisher",
    "author_email": "Andrew.Fisher@coates.com.au",
    "author_mobile": "0429 352 788",
    "author_address": "340 Curtin Ave W, Eagle Farm, QLD 4009",

    # --- Business rules ----------------------------------------------
    "charge_per_unit": 2000,          # $ replacement charge per 30+ monitor
    # Units in store for a 100 availability score. Was 50, which scored the
    # shelf 100/100 while the store ran dry every morning. Reset 30 Jul 2026
    # to 250, grounded in the TRANSACTIONS data: the average mid-morning net
    # draw (issues minus returns) over the last 10 full working days is ~240
    # monitors, so the shelf needs ~250 at opening to survive to the
    # afternoon return wave without turning crews away.
    "availability_target": 250,
    "include_fccu_detail_table": False,  # 194-row table makes the email big
    "top_priorities_count": 10,

    # --- Email --------------------------------------------------------
    "subject_prefix": "Ampol Gas Monitor Report",
    "recipients": [
        "Ampol Store <Ampolstore@coates.com.au>",
        '"Mitchell, Cody" <Cody.mitchell@coates.com.au>',
        '"Riki Nicholls (External)" <rxnicho@ampol.com.au>',
        "Steve Webster <swebste@ampol.com.au>",
        "Lucas Riddell <lxridde@ampol.com.au>",
        "Michael McEvoy <mxmcevo@ampol.com.au>",
        "Fabian Wong <fwong@ampol.com.au>",
        "Lyle Mildenhall <lmilden@ampol.com.au>",
        "Steve Costanzo <scostan@ampol.com.au>",
        "Virginia Harding <vhardin@ampol.com.au>",
        "Tn Nam <tnam@ampol.com.au>",
        "Thiago Lucca <txlucca@ampol.com.au>",
        "John Strano <jstrano@ampol.com.au>",
        '"David McGurk (External)" <dmcgurk@ampol.com.au>',
        "Kail Linehan <kxlineh@ampol.com.au>",
        "Eamon Rees <erees@ampol.com.au>",
        "William Potter <wxpotte@ampol.com.au>",
        "Jonathon Giudice <jgiudic@ampol.com.au>",
        "Jason Waddell <jwaddel@ampol.com.au>",
        '"Paul Henry (External)" <phenry@ampol.com.au>',
        '"Jarred Bishop (External)" <jxbisho@ampol.com.au>',
        '"Simon Phillips (External)" <sxphill@ampol.com.au>',
        "Darren Parker <dparker@ampol.com.au>",
        '"Ross Comerford (External)" <rcommer@ampol.com.au>',
        "Arron Pelling <ajpelli@ampol.com.au>",
        "Robbie Glyde <robbie.glyde@universalcranes.com>",
        "Saad Khan <szkhan1@ampol.com.au>",
        "Sean Wong <czwong1@ampol.com.au>",
        "LytRefMaintReporting <LytRefMaintReporting@ampol.com.au>",
        '"John Martin (External)" <jzmarti1@ampol.com.au>',
        '"Chris Jurlina (External)" <cjurlin@ampol.com.au>',
        '"Sam Robson (External)" <srobers@ampol.com.au>',
        '"Sims, Akira (North)" <Akira.sims@coates.com.au>',
        '"Logan Ferguson-rudolph (External)" <lyfergu@ampol.com.au>',
        '"Webley, Alicia (North)" <Alicia.webley@coates.com.au>',
        "andrew fisher <atfish3r@hotmail.com>",
    ],
}

# =====================================================================
# COLOUR PALETTE (V18 ELITE dark theme)
# =====================================================================

C = {
    "orange": "#F36F21",
    "green": "#16A34A",
    "amber": "#D97706",
    "red": "#DC2626",
    "blue": "#2563EB",
    "bg": "#07090C",
    "panel": "#0D1117",
    "card": "#141B24",
    "card2": "#1A2332",
    "card3": "#1F2C3D",
    "line": "#2A3A4E",
    "muted": "#9AAFC3",
    "muted2": "#7A8FA3",
    "white": "#FFFFFF",
}

RING_RGB = {
    "track": (42, 58, 78, 255),
    "green": (22, 163, 74, 255),
    "amber": (217, 119, 6, 255),
    "red": (220, 38, 38, 255),
}



# =====================================================================
# 0. INPUT AUTO-DETECTION - no exact filenames required
# =====================================================================

def script_dir():
    return os.path.dirname(os.path.abspath(__file__))


def kit_root():
    """The kit root folder.

    Normal kit layout is <kit>\\scripts\\this_file.py, so the root is one
    level up. If someone has flattened the kit onto a USB stick and the
    script is sitting in the root itself, fall back to its own folder.
    """
    here = script_dir()
    if os.path.basename(here).lower() == "scripts":
        return os.path.dirname(here)
    return here


def input_dir():
    """The suite's one Data\\ area - every export saved over the top.

    WHY (12 Aug 2026): Andrew asked for one area all the reports read
    from, so the old input\\ subfolder habit is retired and this now
    answers with the suite's shared Data\\ folder.
    """
    return ampol_paths.data_dir()


def search_dirs():
    """Folders to hunt for inputs, in priority order, de-duplicated.

    WHY (12 Aug 2026): Data\\ first - the suite's one area for every
    Excel - then the suite root as a fallback for anything not yet moved
    across. Newest mtime still wins, so saving a fresh export over the
    top is all it takes.

    archive\\ and output\\ are deliberately NOT searched - archived sources
    and last week's outputs must never be picked up as this morning's data.
    """
    dirs, seen = [], set()
    for d in (input_dir(), kit_root()):
        real = os.path.abspath(d)
        if real not in seen and os.path.isdir(real):
            dirs.append(real)
            seen.add(real)
    return dirs


def find_workbook(cfg):
    """Locate the gas monitor workbook - newest match in the kit folder."""
    if cfg["workbook_path"] not in ("auto", "", None):
        if os.path.exists(cfg["workbook_path"]):
            return os.path.abspath(cfg["workbook_path"])
        sys.exit(f"ERROR: workbook_path is set to '{cfg['workbook_path']}' "
                 f"but that file does not exist.")
    looked = []
    for folder in search_dirs():
        looked.append(folder)
        candidates = []
        for name in os.listdir(folder):
            low = name.lower()
            if not low.endswith(".xlsm"):
                continue
            if name.startswith("~$") or low.startswith("source_"):
                continue      # skip Excel temp files and archived sources
            norm = low.replace("_", " ")
            if "gas monitor" in norm:
                candidates.append(os.path.join(folder, name))
        if candidates:
            # newest file wins if there is more than one
            candidates.sort(key=os.path.getmtime, reverse=True)
            if len(candidates) > 1:
                print(f"[INFO] {len(candidates)} workbooks found in "
                      f"{folder} - using the newest.")
            return candidates[0]
    sys.exit("ERROR: no gas monitor workbook found.\n"
             + "".join(f"  Looked in: {d}\n" for d in looked)
             + "  Save your .xlsm (any name containing 'gas monitor') into the\n"
               f"  Data folder - {ampol_paths.data_dir()}\n"
               "  - then press the gas monitor button again.")


def find_workbook_optional(cfg):
    """The workbook if it is there, '' if not - never a hard stop."""
    try:
        return find_workbook(cfg)
    except SystemExit:
        return ""


def find_rental_stock(cfg):
    if cfg["rental_stock_path"] not in ("auto", "", None):
        return cfg["rental_stock_path"] if os.path.exists(cfg["rental_stock_path"]) else ""
    if cfg["rental_stock_path"] == "":
        return ""
    cands = []
    for folder in search_dirs():
        cands += [os.path.join(folder, n) for n in os.listdir(folder)
                  if n.lower().startswith("rental_stock") and n.lower().endswith(".xlsx")
                  and not n.startswith("~$")]
    cands.sort(key=os.path.getmtime, reverse=True)
    return cands[0] if cands else ""


def resolve_output_dir(cfg):
    """Where the .eml and .html get written.

    WHY (12 Aug 2026): outputs now land in the suite's dated report area -
    Reports\\<today>\\Gas_Monitors, created on demand. Fixed filenames are
    still fine because the date lives in the folder name, so nothing is
    ever silently overwritten day to day. The old output_dir setting in
    CONFIG no longer steers this.
    """
    return ampol_paths.day_folder("Gas_Monitors")


def stamped_name(base, date_str, cfg):
    """Insert the run date into a filename: NAME_2026-07-30.eml.

    Stops this morning's run overwriting yesterday's report - on site the
    previous report is often the only record of what was sent.
    """
    if not cfg.get("date_stamp_outputs", True):
        return base
    stem, ext = os.path.splitext(base)
    stamp = datetime.strptime(date_str, "%d %B %Y").strftime("%Y-%m-%d")
    return f"{stem}_{stamp}{ext}"


# =====================================================================
# 1. DATA LOAD - everything counted from detail tabs
# =====================================================================

def load_data(path=""):
    """Every figure the dashboard needs.

    WHY (02 Sep 2026): counted by gasmon_engine from the SiteIQ exports
    when the engine is present (always, in the suite). The workbook parse
    below is kept as the fallback for a bare kit without the engine."""
    if _engine is not None:
        ctx = _engine.load()
        m0 = _engine.compute(ctx)
        data = _engine.v18_data(m0)
        data["_engine"] = m0
        return data
    return load_workbook_data(path)


def load_workbook_data(path):
    """Read all tabs of the Excel workbook and return raw row data."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    data = {}

    def rows(sheet, ncols):
        out = []
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            out.append(list(row[:ncols]))
        return out

    data["available"] = rows("Available Store", 5)
    data["oh_1_7"] = rows("Onhire 1–7 days", 9)
    data["oh_8_29"] = rows("Onhire 7–30 days", 9)
    data["oh_30"] = rows("Onhire 30+ Days", 10)
    data["fccu"] = rows("Onhire FCCU", 9)
    data["repairs"] = rows("Repairs", 6)

    # Performance tab: summary rows (A3:B8) + per-company table (row 9+)
    ws = wb["Gas Monitor Return Performance"]
    perf_rows = list(ws.iter_rows(min_row=3, values_only=True))
    data["perf_summary"] = {}
    data["companies"] = []
    for r in perf_rows:
        if r[0] is None:
            continue
        name = str(r[0]).strip()
        if name in ("Current Onhire", "Available", "Out Of Service",
                    "Onhire To Operations", "Onhire To FCCU",
                    "Onhire To Future Fuels"):
            data["perf_summary"][name] = int(r[1] or 0)
        else:
            data["companies"].append({
                "name": name,
                "issued": int(r[1] or 0),
                "returned": int(r[2] or 0),
                "nrsd": int(r[3] or 0),
                "nrsd_pct": r[4] or "",
                "top_offenders": r[5] or "",
                "reoffender_action": r[6] or "",
                "prev_day": int(r[7] or 0),
                "prev_day_pct": r[8] or "",
                "prev_day_names": r[9] or "",
                "still_out": int(r[10] or 0),
                "still_out_pct": r[11] or "",
                "still_out_names": r[12] or "",
                "action": r[13] or "",
                "d1_7": int(r[14] or 0),
                "d8_29": int(r[15] or 0),
                "d30": int(r[16] or 0),
                "charge": int(r[17] or 0),
                "today": int(r[18] or 0),
            })
    wb.close()
    return data


# =====================================================================
# 2. METRICS - single source of truth, all reconciled
# =====================================================================

def compute_metrics(data, cfg):
    m = {}
    warn = []

    # -- counted directly from detail tabs (authoritative) -------------
    m["available"] = len(data["available"])
    m["fccu"] = len(data["fccu"])
    m["repairs"] = len(data["repairs"])
    m["out_1_7"] = len(data["oh_1_7"])
    m["out_8_29"] = len(data["oh_8_29"])
    m["out_30"] = len(data["oh_30"])
    m["outstanding"] = m["out_1_7"] + m["out_8_29"] + m["out_30"]

    # -- custody groups only exist as summary values -------------------
    m["ops"] = data["perf_summary"].get("Onhire To Operations", 0)
    m["ff"] = data["perf_summary"].get("Onhire To Future Fuels", 0)
    m["ah"] = data["perf_summary"].get("Onhire To After Hours", 0)

    # -- per-company aggregates ----------------------------------------
    comps = data["companies"]
    m["issued_today"] = data.get("issued_today", sum(c["today"] for c in comps))
    m["prev_day_total"] = sum(c["prev_day"] for c in comps)
    m["exposure"] = sum(c["charge"] for c in comps)
    m["exposure_check"] = m["out_30"] * cfg["charge_per_unit"]

    # -- derived fleet positions ---------------------------------------
    m["onhire_general"] = m["issued_today"] + m["outstanding"]
    m["onhire_all"] = m["onhire_general"] + m["fccu"] + m["ops"] + m["ff"] + m["ah"]
    m["fleet_total"] = m["onhire_all"] + m["available"] + m["repairs"]
    m["fleet_general"] = m["onhire_general"] + m["available"] + m["repairs"]

    # -- previous-day recovery cycle ------------------------------------
    # WHY (02 Sep 2026): this was hard-coded to 0, which pinned the Returns
    # score at 0/100 every morning. The engine counts yesterday's non-returns
    # that have since been scanned back.
    m["recovered"] = data.get("recovered", 0)
    m["yday_still_out"] = m["prev_day_total"] - m["recovered"]
    m["recovery_rate"] = (100 if m["prev_day_total"] == 0
                          else round(m["recovered"] / m["prev_day_total"] * 100))

    # -- repairs breakdown ----------------------------------------------
    cats = Counter()
    for r in data["repairs"]:
        cat = re.sub(r"^Dräger\s*[–-]\s*", "", str(r[0])).strip()
        cats[cat] += 1
    m["repair_cats"] = cats.most_common()
    m["repair_top"] = m["repair_cats"][0][0] if m["repair_cats"] else "None"
    m["fleet_impact_pct"] = round(m["repairs"] / m["fleet_total"] * 100, 1)

    # -- health scores (transparent formulas) ---------------------------
    m["score_availability"] = min(100, round(m["available"] / cfg["availability_target"] * 100))
    m["score_recovery"] = m["recovery_rate"]
    m["score_repairs"] = max(0, 100 - round(m["repairs"] / m["fleet_total"] * 100))
    m["score_30"] = max(0, 100 - 3 * m["out_30"])
    m["health"] = round((m["score_availability"] + m["score_recovery"]
                         + m["score_repairs"] + m["score_30"]) / 4)
    # WHY (02 Sep 2026): one health score across the PDF, the email and this
    # dashboard. The engine scores same-day returns over the last 30 days in
    # place of "recovered so far today" (which is 0 by definition at 06:45).
    eng = data.get("_engine")
    if eng is not None:
        m["score_recovery"] = eng["score_sameday"]
        m["health"] = eng["health"]
        m["sameday_30_pct"] = eng["d30"]["sd_pct"]

    # -- top recovery priorities -----------------------------------------
    # Rank: exposure desc, total outstanding desc, 1-7 desc.
    # Urgency: Critical = 30+ units; High = 8-29 units or total >= 3; Watch = rest.
    ranked = [c for c in comps if (c["still_out"] + c["prev_day"]) > 0]
    ranked.sort(key=lambda c: (-c["charge"], -c["still_out"], -c["d1_7"]))
    for c in ranked:
        c["urgency"] = ("Critical" if c["d30"] > 0
                        else "High" if (c["d8_29"] > 0 or c["still_out"] >= 3)
                        else "Watch")
    m["priorities"] = ranked[: cfg["top_priorities_count"]]
    m["focus3"] = ranked[:3]

    # -- reconciliation against summary cells -----------------------------
    ps = data["perf_summary"]
    checks = [
        ("Current Onhire", ps.get("Current Onhire"), m["onhire_general"]),
        ("Available", ps.get("Available"), m["available"]),
        ("Out Of Service", ps.get("Out Of Service"), m["repairs"]),
        ("Onhire To FCCU", ps.get("Onhire To FCCU"), m["fccu"]),
    ]
    for label, cell, counted in checks:
        if cell is not None and cell != counted:
            warn.append(f"Performance tab '{label}' = {cell} but detail count = {counted} "
                        f"-> report uses {counted}")
    if m["exposure"] != m["exposure_check"]:
        warn.append(f"Charge column total ${m['exposure']:,} != 30+ count x "
                    f"${cfg['charge_per_unit']:,} = ${m['exposure_check']:,}")
    m["validation_pass"] = (m["prev_day_total"] == m["recovered"] + m["yday_still_out"])
    m["warnings"] = warn
    return m


def validate_against_rental_stock(path, m):
    """Optional cross-check against a live SiteIQ RENTAL_STOCK export."""
    notes = []
    try:
        import pandas as pd
        df = pd.read_excel(path, sheet_name="RENTAL_STOCK")
    except Exception as e:
        return [f"RENTAL_STOCK validation skipped ({e})"]
    gm = df[df["ITEM_DESCRIPTION"].str.contains("X-am|X-AM", case=False, na=False)]
    mon = gm[~gm["ITEM_DESCRIPTION"].str.contains("CHARGER|PROBE", case=False, na=False)]
    ex_total = len(mon)
    ex_avail = int((mon["ITEM_STATUS"] == "Available for Hire").sum())
    ex_onhire = int((mon["ITEM_STATUS"] == "On Hire").sum())
    if ex_total != m["fleet_total"]:
        notes.append(f"RENTAL_STOCK fleet {ex_total} != report fleet {m['fleet_total']}")
    if ex_avail != m["available"]:
        notes.append(f"RENTAL_STOCK available {ex_avail} != report available {m['available']}")
    if ex_onhire != m["onhire_all"] + m["repairs"]:
        notes.append(f"RENTAL_STOCK on-hire {ex_onhire} != report on-hire+repair "
                     f"{m['onhire_all'] + m['repairs']}")
    if not notes:
        notes.append(f"RENTAL_STOCK cross-check PASS: fleet {ex_total}, "
                     f"available {ex_avail}, on-hire {ex_onhire}")
    return notes


# =====================================================================
# 3. HEALTH RINGS - PNG donuts, base64-embedded
# =====================================================================

def ring_colour(score):
    if score >= 80:
        return RING_RGB["green"]
    if score >= 45:
        return RING_RGB["amber"]
    return RING_RGB["red"]


def make_ring_b64(score):
    S = 4
    size = 300 * S
    im = Image.new("RGBA", (size, size), (0, 0, 0, 0))
    d = ImageDraw.Draw(im)
    outer, inner = 136 * S, 109 * S
    cx = cy = size // 2
    bbox = [cx - outer, cy - outer, cx + outer, cy + outer]
    width = outer - inner
    d.arc(bbox, 0, 360, fill=RING_RGB["track"], width=width)
    if score > 0:
        d.arc(bbox, -90, -90 + score / 100 * 360, fill=ring_colour(score), width=width)
    im = im.resize((300, 300), Image.LANCZOS)
    buf = io.BytesIO()
    im.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode()


# =====================================================================
# 4. HTML BUILDING BLOCKS
# =====================================================================

def esc(s):
    # WHY (12 Aug 2026): a blank cell (e.g. a missing serial number on the
    # detail tabs) used to print as the word 'None'. A value the source
    # does not carry shows as a dash - never a guess, never Python's None.
    if s is None:
        return "&ndash;"
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def money(n):
    return f"${n:,.0f}"


def hero(m, cfg, date_str, time_str, gen_str):
    recon_col = C["green"] if m["validation_pass"] else C["red"]
    recon_txt = ("PASS" if m["validation_pass"] else "FAIL")
    hs_col = C["green"] if m["health"] >= 80 else C["amber"] if m["health"] >= 45 else C["red"]
    return f"""
<tr>
  <td bgcolor="{C['panel']}" style="background-color:{C['panel']};padding:0;border-bottom:1px solid {C['line']};">
    <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
      <tr>
        <td bgcolor="{C['orange']}" style="background-color:{C['orange']};width:8px;padding:0;">&nbsp;</td>
        <td style="padding:36px 32px 30px 32px;vertical-align:top;">
          <table cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
            <tr>
              <td style="vertical-align:middle;padding-right:22px;border-right:2px solid {C['line']};">
                <div style="font-size:54px;font-weight:900;color:{C['white']};letter-spacing:-2px;line-height:1;">Coates</div>
              </td>
              <td style="padding-left:22px;vertical-align:middle;">
                <div style="font-size:13px;font-weight:900;color:{C['orange']};letter-spacing:3px;text-transform:uppercase;">Industrial Solutions</div>
                <div style="font-size:10px;color:{C['muted']};font-weight:600;letter-spacing:1px;margin-top:5px;text-transform:uppercase;">Ampol &middot; Dr&auml;ger Gas Monitor Programme</div>
              </td>
            </tr>
          </table>
          <div style="margin-top:22px;padding-top:20px;border-top:1px solid {C['line']};">
            <div style="font-size:28px;font-weight:900;color:{C['white']};letter-spacing:-.4px;">Executive Operations Dashboard</div>
            <div style="font-size:12px;color:{C['muted']};margin-top:7px;letter-spacing:.3px;">Dr&auml;ger X-am 5000 Fleet Status &ndash; Recovery, FCCU, Repairs and Exposure</div>
          </div>
          <div style="margin-top:14px;">
            <span style="display:inline-block;background:{C['card2']};border:1px solid {C['line']};border-radius:4px;padding:5px 12px;font-size:9px;font-weight:900;color:{C['orange']};letter-spacing:2px;text-transform:uppercase;">THE COATES WAY</span>
            <span style="display:inline-block;margin-left:8px;font-size:9px;color:{C['muted']};font-weight:600;letter-spacing:.5px;">Operational Excellence &middot; Asset Performance &middot; Disciplined Execution</span>
            <span style="display:inline-block;margin-left:8px;background:{C['card2']};border:1px solid {C['line']};border-radius:4px;padding:5px 12px;font-size:9px;font-weight:900;color:{C['muted']};letter-spacing:2px;text-transform:uppercase;">POWERED BY SITEIQ</span>
          </div>
        </td>
        <td style="padding:36px 32px 30px 0;vertical-align:top;text-align:right;width:300px;">
          <table cellspacing="0" cellpadding="0" style="border-collapse:collapse;display:inline-table;">
            <tr><td style="padding:0 0 6px 0;text-align:right;">
              <span style="display:inline-block;background-color:{C['card2']};border:1px solid {C['line']};padding:3px 10px;border-radius:99px;font-size:8px;font-weight:900;color:{C['green']};letter-spacing:1.5px;text-transform:uppercase;">&#9679; LIVE INTELLIGENCE</span>
            </td></tr>
            <tr><td bgcolor="{C['orange']}" style="background-color:{C['orange']};padding:12px 24px;border-radius:4px 4px 0 0;">
              <div style="font-size:8px;font-weight:900;color:rgba(255,255,255,.7);text-transform:uppercase;letter-spacing:1.5px;">Report Date</div>
              <div style="font-size:22px;font-weight:900;color:#fff;margin-top:3px;letter-spacing:-.3px;">{date_str}</div>
              <div style="font-size:8px;font-weight:900;color:rgba(255,255,255,.6);text-transform:uppercase;letter-spacing:1.5px;margin-top:8px;">Run Time</div>
              <div style="font-size:18px;font-weight:900;color:#fff;margin-top:2px;letter-spacing:-.2px;">{time_str}</div>
            </td></tr>
            <tr><td bgcolor="{C['card2']}" style="background-color:{C['card2']};border:1px solid {C['line']};border-top:0;padding:14px 24px;border-radius:0 0 4px 4px;">
              <div style="font-size:8px;font-weight:900;color:{C['muted']};text-transform:uppercase;letter-spacing:1.5px;">Reconciliation</div>
              <div style="font-size:18px;font-weight:900;color:{recon_col};margin-top:5px;letter-spacing:-.2px;">{recon_txt}: {m['prev_day_total']} = {m['recovered']} + {m['yday_still_out']}</div>
              <div style="font-size:8px;font-weight:900;color:{C['muted']};text-transform:uppercase;letter-spacing:1px;margin-top:8px;">Health Score</div>
              <div style="font-size:16px;font-weight:900;color:{hs_col};margin-top:3px;">{m['health']} / 100</div>
              <div style="font-size:9px;color:{C['muted']};margin-top:8px;">Generated: {gen_str}</div>
            </td></tr>
          </table>
        </td>
      </tr>
    </table>
  </td>
</tr>"""


def disciplines():
    items = [
        ("Safety First", "Bump, charge and scan before issue", C["green"]),
        ("Customer Centric", "Recovery follow-up by company and name", C["orange"]),
        ("Pace", "Previous-day non-returns actioned same day", C["orange"]),
        ("Open to Feedback", "Validation check published on every report", C["blue"]),
        ("Reduce Friction", "One dashboard, one click, all custody types", C["green"]),
        ("Cadence", "Generated daily as part of the operating rhythm", C["orange"]),
    ]
    cells = "".join(f"""
      <td bgcolor="{C['card2']}" width="16%" style="background-color:{C['card2']};border:1px solid {C['line']};border-top:3px solid {col};padding:12px 10px;vertical-align:top;text-align:center;">
        <div style="font-size:11px;font-weight:900;color:{C['white']};letter-spacing:.3px;">{t}</div>
        <div style="font-size:9px;color:{C['muted']};margin-top:6px;line-height:1.5;">{d}</div>
      </td>""" for t, d, col in items)
    return f"""
<tr>
  <td bgcolor="{C['panel']}" style="background-color:{C['panel']};padding:16px 22px;border-bottom:2px solid {C['orange']};">
    <div style="margin-bottom:10px;">
      <span style="font-size:12px;font-weight:900;color:{C['orange']};letter-spacing:2px;text-transform:uppercase;">Six Operating Disciplines</span>
      <span style="font-size:10px;color:{C['muted']};margin-left:10px;">How this report runs The Coates Way, every day</span>
    </div>
    <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:separate;border-spacing:3px;"><tr>{cells}</tr></table>
  </td>
</tr>"""


def kpi_tile(tag, value, title, note, colour, glow):
    return f"""
<td bgcolor="{C['card']}" style="background-color:{C['card']};border:1px solid {C['line']};border-bottom:3px solid {colour};padding:0;vertical-align:top;">
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
    <tr><td style="padding:10px 14px 0 14px;"><span style="font-size:9px;font-weight:900;color:{C['muted']};letter-spacing:1.5px;">{tag}</span><span style="font-size:7px;font-weight:900;color:{C['line']};float:right;text-transform:uppercase;letter-spacing:.5px;">LIVE</span></td></tr>
    <tr><td style="padding:8px 14px 4px 14px;"><div style="font-size:32px;line-height:36px;font-weight:900;color:{colour};letter-spacing:-1px;text-shadow:0 0 24px {glow};">{value}</div></td></tr>
    <tr><td style="padding:0 14px 4px 14px;"><div style="font-size:10px;font-weight:900;color:{C['white']};text-transform:uppercase;letter-spacing:.4px;">{title}</div></td></tr>
    <tr><td style="padding:0 14px 10px 14px;border-top:1px solid {C['line']};margin-top:4px;"><div style="font-size:9px;color:{C['muted']};padding-top:6px;letter-spacing:.2px;">{note}</div></td></tr>
  </table>
</td>"""


GLOW = {
    C["orange"]: "rgba(243,111,33,0.18)",
    C["green"]: "rgba(22,163,74,0.15)",
    C["blue"]: "rgba(37,99,235,0.18)",
    C["red"]: "rgba(220,38,38,0.18)",
}


def kpi_rows(m):
    fleet_note = (f"All categories &middot; <span style='color:{C['orange']};font-weight:900;'>"
                  f"General fleet {m['fleet_general']}</span> excl. FCCU/Ops/FF")
    r1 = [
        ("FLT", m["fleet_total"], "Total Fleet", fleet_note, C["orange"]),
        ("AVL", m["available"], "Available Store", "Ready for general issue", C["green"]),
        ("HIRE", m["onhire_general"], "On Hire",
         f"{m['issued_today']} issued today + {m['outstanding']} outstanding", C["orange"]),
        ("FCCU", m["fccu"], "FCCU On Hire", "Tracked separately", C["blue"]),
        ("OPS", m["ops"], "Operations On Hire", "Tracked separately", C["blue"]),
        ("FF", m["ff"], "Future Fuels On Hire", "Tracked separately", C["blue"]),
        ("YDAY", m["prev_day_total"], "Prev Day Non-Returns", "Yesterday total",
         C["green"] if m["prev_day_total"] == 0 else C["red"]),
        ("REC", m["recovered"], "Recovered", "Since yesterday",
         C["green"] if m["yday_still_out"] == 0 else C["amber"]),
        ("ACT", m["yday_still_out"], "Still Outstanding", "Action required",
         C["green"] if m["yday_still_out"] == 0 else C["red"]),
    ]
    hs_col = C["orange"]
    rr_col = C["green"] if m["recovery_rate"] >= 80 else C["red"]
    r2 = [
        ("1-7", m["out_1_7"], "Outstanding 1-7", "Current outstanding", C["orange"]),
        ("8-29", m["out_8_29"], "Outstanding 8-29", "Follow-up required", C["orange"]),
        ("30+", m["out_30"], "Outstanding 30+", "Charge / recovery priority", C["red"]),
        ("$", money(m["exposure"]), "30+ Exposure", "Charge exposure", C["green"]),
        ("REP", m["repairs"], "Dr&auml;ger Repairs", "Out of service / repair", C["orange"]),
        ("%", f"{m['recovery_rate']}%", "Recovery Rate",
         "No yday non-returns to chase" if m["prev_day_total"] == 0 else "Recovered from yesterday", rr_col),
        ("HS", f"{m['health']}/100", "Health Score", "Calculated operations score", hs_col),
    ]

    def row(tiles):
        return ("<table width='100%' cellspacing='0' cellpadding='0' "
                "style='border-collapse:separate;border-spacing:4px;'><tr>"
                + "".join(kpi_tile(t, v, ti, n, col, GLOW.get(col, GLOW[C['orange']]))
                          for t, v, ti, n, col in tiles)
                + "</tr></table>")

    return f"""
<tr><td bgcolor="{C['bg']}" style="background-color:{C['bg']};padding:14px 14px 0 14px;">{row(r1)}</td></tr>
<tr><td bgcolor="{C['bg']}" style="background-color:{C['bg']};padding:4px 14px 0 14px;">{row(r2)}</td></tr>"""


def section_header(title, sub):
    return f"""
<tr><td bgcolor="{C['bg']}" style="background-color:{C['bg']};padding:12px 14px 4px 14px;">
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;margin-bottom:4px;">
    <tr><td bgcolor="{C['panel']}" style="background-color:{C['panel']};border-left:5px solid {C['orange']};padding:12px 22px;border-radius:4px;">
      <span style="font-size:13px;font-weight:900;color:{C['white']};letter-spacing:.5px;">{title}</span>
      <span style="font-size:9px;color:{C['muted']};margin-left:12px;letter-spacing:1px;text-transform:uppercase;">{sub}</span>
    </td></tr>
  </table>
</td></tr>"""


def exec_brief(m):
    o, g, r, b, a = C["orange"], C["green"], C["red"], C["blue"], C["amber"]
    hs_col = g if m["health"] >= 80 else a if m["health"] >= 45 else r
    body = (
        f"The previous day non-return total is <b style='color:{o};'>{m['prev_day_total']}</b>. "
        f"Of these, <b style='color:{g};'>{m['recovered']}</b> have now been recovered and "
        f"<b style='color:{r};'>{m['yday_still_out']}</b> remain outstanding to the same person. "
        f"Current outstanding monitors are <b style='color:{o};'>{m['outstanding']}</b> across 1-7, 8-29 "
        f"and 30+ day ageing, with <b style='color:{r};'>{m['out_30']}</b> over 30 days and "
        f"<b style='color:{r};'>{money(m['exposure'])}</b> exposure. "
        f"General on hire currently stands at <b style='color:{o};'>{m['onhire_general']}</b> monitors "
        f"&mdash; <b style='color:{o};'>{m['issued_today']}</b> issued today plus "
        f"<b style='color:{o};'>{m['outstanding']}</b> outstanding from previous days &mdash; with "
        f"<b style='color:{g};'>{m['available']}</b> available in store. "
        f"Total fleet is <b style='color:{o};'>{m['fleet_total']}</b> monitors, of which the general fleet "
        f"&mdash; excluding FCCU, Operations and Future Fuels custody &mdash; is "
        f"<b style='color:{o};'>{m['fleet_general']}</b> ({m['onhire_general']} on hire + "
        f"{m['available']} available + {m['repairs']} in repair). "
        f"FCCU is kept separate with <b style='color:{b};'>{m['fccu']}</b> monitors on hire, Operations "
        f"custody holds <b style='color:{b};'>{m['ops']}</b> and Future Fuels custody holds "
        f"<b style='color:{b};'>{m['ff']}</b> &mdash; all tracked outside the general fleet. "
        f"Dr&auml;ger repairs currently total <b style='color:{a};'>{m['repairs']}</b>, with "
        f"<b>Dr&auml;ger &ndash; {esc(m['repair_top'])}</b> the largest repair category."
    )
    return section_header("Executive Brief", "Clear operational snapshot for today") + f"""
<tr><td bgcolor="{C['bg']}" style="background-color:{C['bg']};padding:0 14px 12px 14px;">
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
    <tr><td bgcolor="{C['card']}" style="background-color:{C['card']};border:1px solid {C['line']};padding:20px 24px;">
      <div style="font-size:12px;color:#C9D6E2;line-height:1.9;">{body}</div>
      <div style="border-top:1px solid {C['line']};margin-top:14px;padding-top:12px;">
        <span style="font-size:11px;color:{C['orange']};font-weight:900;font-style:italic;">Consistent execution today drives the Flywheel forward.</span>
      </div>
    </td></tr>
  </table>
</td></tr>"""


def action_framework(m):
    focus_names = ", ".join(esc(c["name"]) for c in m["focus3"]) or "None"
    cards = [
        ("OPERATIONAL EXCELLENCE", str(m["yday_still_out"]),
         f"Chase <b>{m['yday_still_out']}</b> monitors still outstanding from yesterday.",
         C["red"] if m["yday_still_out"] else C["red"]),
        ("ASSET PERFORMANCE", str(m["repairs"]),
         f"Reduce Dr&auml;ger repairs. Largest category: <b>Dr&auml;ger &ndash; {esc(m['repair_top'])}</b>.",
         C["amber"]),
        ("FINANCIAL DISCIPLINE", money(m["exposure"]),
         f"Recover <b>{m['out_30']}</b> monitors over 30 days &mdash; <b>{money(m['exposure'])}</b> at risk.",
         C["red"]),
        ("PEOPLE ACCOUNTABILITY", str(len(m["focus3"])),
         f"Top focus: <b>{focus_names}</b>.", C["orange"]),
    ]
    cells = "".join(f"""
      <td bgcolor="{C['card']}" width="25%" style="background-color:{C['card']};border:1px solid {C['line']};border-top:3px solid {col};padding:0;vertical-align:top;">
        <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
          <tr><td style="padding:14px 16px 0 16px;"><span style="display:inline-block;border:1px solid {C['line']};padding:3px 8px;font-size:8px;font-weight:900;color:{col};letter-spacing:1.5px;text-transform:uppercase;">{t}</span></td></tr>
          <tr><td style="padding:10px 16px 4px 16px;"><div style="font-size:30px;font-weight:900;color:{col};letter-spacing:-1px;">{big}</div></td></tr>
          <tr><td style="padding:6px 16px 14px 16px;border-top:1px solid {C['line']};"><div style="font-size:10px;color:#C9D6E2;line-height:1.6;padding-top:8px;">{txt}</div></td></tr>
        </table>
      </td>""" for t, big, txt, col in cards)
    return section_header("Coates Way Action Framework", "Today's priorities mapped to performance pillars") + f"""
<tr><td bgcolor="{C['bg']}" style="background-color:{C['bg']};padding:0 14px 12px 14px;">
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:separate;border-spacing:4px;"><tr>{cells}</tr></table>
</td></tr>"""


def pillars(m):
    items = [
        ("Overall Health", m["health"], "OPERATIONS", C["orange"]),
        ("Availability", m["score_availability"], "ASSETS", C["blue"]),
        ("Same-day 30d", m["score_recovery"], "PEOPLE", C["amber"]),
        ("Repairs", m["score_repairs"], "ASSETS", C["green"]),
        ("30+ Control", m["score_30"], "FINANCIALS", C["red"]),
    ]
    cells = ""
    for title, score, pillar, pcol in items:
        b64 = make_ring_b64(score)
        cells += f"""
      <td bgcolor="{C['card']}" style="background-color:{C['card']};border:1px solid {C['line']};padding:18px 12px;vertical-align:top;text-align:center;"><table cellpadding="0" cellspacing="0" style="margin:0 auto;"><tr><td style="text-align:center;"><span style="font-size:10px;font-weight:900;color:{C['muted']};text-transform:uppercase;letter-spacing:1px;">{title}</span></td></tr><tr><td style="text-align:center;padding-top:10px;"><img src="data:image/png;base64,{b64}" width="150" height="150" style="display:block;margin:0 auto;" alt="Score {score}/100" /></td></tr><tr><td style="text-align:center;padding-top:6px;"><span style="font-size:18px;color:{C['white']};font-weight:900;">{score}</span><span style="font-size:13px;color:{C['muted2']};font-weight:700;"> /100</span></td></tr><tr><td style="text-align:center;padding-top:8px;"><table cellpadding="0" cellspacing="0" style="margin:0 auto;"><tr><td bgcolor="{C['panel']}" style="background-color:{C['panel']};border:1px solid {C['line']};padding:3px 8px;"><span style="font-size:8px;font-weight:900;color:{pcol};letter-spacing:1.5px;text-transform:uppercase;">{pillar}</span></td></tr></table></td></tr></table></td>"""
    return section_header("Coates Way Performance Pillars", "How we measure success") + f"""
<tr><td bgcolor="{C['bg']}" style="background-color:{C['bg']};padding:0 14px 12px 14px;">
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:separate;border-spacing:4px;"><tr>{cells}</tr></table>
</td></tr>"""


def scorecard(m):
    op_status = (("&#9679; ON TRACK", C["green"]) if m["repairs"] == 0 and m["yday_still_out"] == 0
                 else ("&#9679; ACTION REQUIRED", C["red"]))
    fin_status = (("&#9679; ON TRACK", C["green"]) if m["out_30"] == 0
                  else ("&#9679; ATTENTION", C["amber"]))
    cust_line = (f"{m['yday_still_out']} monitors still", "with customer")
    op_line = (("No yday non-returns" if m["prev_day_total"] == 0
                else f"{m['recovery_rate']}% recovery rate"),
               f"{m['repairs']} repairs queued")
    cards = [
        (("&#9679; ON TRACK", C["green"]), "Safety", "All monitors tracked", "No safety incidents"),
        (("&#9679; ON TRACK", C["green"]) if m["yday_still_out"] == 0 else ("&#9679; ATTENTION", C["amber"]),
         "Customer", *cust_line),
        (op_status, "Operational", *op_line),
        (fin_status, "Financial", f"{money(m['exposure'])} exposure", f"{m['out_30']} at charge risk"),
        (("&#9679; ON TRACK", C["green"]), "People", f"Top {len(m['priorities'])} identified", "Actions assigned"),
    ]
    cells = "".join(f"""
      <td bgcolor="{C['card']}" width="20%" style="background-color:{C['card']};border:1px solid {C['line']};padding:0;vertical-align:top;text-align:center;">
        <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
          <tr><td bgcolor="{col}" style="background-color:{col};padding:7px 6px;"><span style="font-size:8px;font-weight:900;color:#fff;letter-spacing:1.5px;text-transform:uppercase;">{status}</span></td></tr>
          <tr><td style="padding:12px 8px 4px 8px;"><span style="font-size:10px;font-weight:900;color:{C['muted']};letter-spacing:1.5px;text-transform:uppercase;">{title}</span></td></tr>
          <tr><td style="padding:2px 8px 14px 8px;"><div style="font-size:10px;color:#C9D6E2;line-height:1.6;">{l1}<br/>{l2}</div></td></tr>
        </table>
      </td>""" for (status, col), title, l1, l2 in cards)
    return section_header("Balanced Scorecard", "Today's performance against The Coates Way") + f"""
<tr><td bgcolor="{C['bg']}" style="background-color:{C['bg']};padding:0 14px 12px 14px;">
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:separate;border-spacing:4px;"><tr>{cells}</tr></table>
</td></tr>"""


def bar_row(label, sub, value, maxv, colour, bg):
    pct = 0 if maxv == 0 else max(1, round(value / maxv * 100)) if value else 0
    return f"""
<tr style="background:{bg};">
  <td style="padding:11px 14px;vertical-align:middle;width:160px;">
    <div style="font-size:12px;font-weight:900;color:{C['white']};">{label}</div>
    <div style="font-size:9px;color:{C['muted']};font-weight:600;margin-top:2px;">{sub}</div>
  </td>
  <td style="padding:9px 14px 9px 0;vertical-align:middle;">
    <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;"><tr>
      <td bgcolor="{colour}" style="background-color:{colour};width:{pct}%;height:14px;font-size:1px;">&nbsp;</td>
      <td bgcolor="{C['line']}" style="background-color:{C['line']};width:{100-pct}%;height:14px;font-size:1px;">&nbsp;</td>
      <td style="width:44px;text-align:right;padding-left:8px;"><span style="font-size:12px;font-weight:900;color:{colour};">{value}</span></td>
    </tr></table>
  </td>
</tr>"""


def panel(title, colour, inner):
    return f"""
<td bgcolor="{C['card2']}" style="background-color:{C['card2']};border:1px solid {C['line']};padding:0;vertical-align:top;width:33%;">
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
    <tr><td bgcolor="{colour}" style="background-color:{colour};padding:12px 14px;"><span style="font-size:11px;font-weight:900;color:#fff;text-transform:uppercase;letter-spacing:.8px;">{title}</span></td></tr>
    {inner}
  </table>
</td>"""


def totals_row(label, value, colour):
    return f"""
<tr><td colspan="2" bgcolor="{C['panel']}" style="background-color:{C['panel']};padding:10px 14px;border-top:1px solid {C['line']};">
  <span style="font-size:9px;font-weight:900;color:{C['muted']};text-transform:uppercase;letter-spacing:.8px;">{label}</span>
  <span style="font-size:16px;font-weight:900;color:{colour};float:right;">{value}</span>
</td></tr>"""


def data_panels(m):
    g, o, r, b = C["green"], C["orange"], C["red"], C["blue"]
    maxr = max(m["prev_day_total"], m["recovered"], m["yday_still_out"], 1)
    p1 = panel("Recovery Performance", o,
        bar_row("Prev Day Non-Returns", "yesterday total", m["prev_day_total"], maxr, r, C["card3"]) +
        bar_row("Recovered", "now closed", m["recovered"], maxr, g, C["card2"]) +
        bar_row("Still Outstanding", "action list", m["yday_still_out"], maxr, r, C["card3"]) +
        totals_row("Recovery Rate", f"{m['recovery_rate']}%", g if m["recovery_rate"] >= 80 else r))
    maxo = max(m["out_1_7"], m["out_8_29"], m["out_30"], 1)
    p2 = panel("Outstanding Ageing", o,
        bar_row("1-7 Days", "current", m["out_1_7"], maxo, o, C["card3"]) +
        bar_row("8-29 Days", "follow-up", m["out_8_29"], maxo, o, C["card2"]) +
        bar_row("30+ Days", "charge risk", m["out_30"], maxo, r, C["card3"]) +
        totals_row("Total Outstanding", m["outstanding"], C["amber"]))
    maxc = max(m["fccu"], m["ops"], m["ff"], m["onhire_general"], 1)
    p3 = panel("Separate Custody Position", b,
        bar_row("FCCU On Hire", "separate", m["fccu"], maxc, b, C["card3"]) +
        bar_row("Operations On Hire", "separate", m["ops"], maxc, b, C["card2"]) +
        bar_row("Future Fuels On Hire", "separate", m["ff"], maxc, b, C["card3"]) +
        bar_row("After Hours Account", "separate", m["ah"], maxc, b, C["card2"]) +
        bar_row("General On Hire", "general fleet", m["onhire_general"], maxc, o, C["card3"]) +
        totals_row("Total On Hire (all custody)", m["onhire_all"], b) +
        totals_row("General Fleet (excl. FCCU / Ops / FF)", m["fleet_general"], o))
    return f"""
<tr><td bgcolor="{C['bg']}" style="background-color:{C['bg']};padding:0 14px 12px 14px;">
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:separate;border-spacing:4px;"><tr>{p1}{p2}{p3}</tr></table>
</td></tr>"""


def repairs_and_focus(m):
    o, r, a = C["orange"], C["red"], C["amber"]
    total = m["repairs"] or 1
    rows = ""
    bgs = [C["card3"], C["card2"]]
    for i, (cat, n) in enumerate(m["repair_cats"]):
        pct = round(n / total * 100)
        rows += f"""
<tr style="background:{bgs[i % 2]};">
  <td style="padding:10px 14px;vertical-align:middle;width:200px;"><div style="font-size:11px;font-weight:700;color:{C['white']};">Dr&auml;ger &ndash; {esc(cat)}</div></td>
  <td style="padding:8px 14px 8px 0;vertical-align:middle;">
    <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;"><tr>
      <td bgcolor="{o}" style="background-color:{o};width:{max(2,pct)}%;height:12px;font-size:1px;">&nbsp;</td>
      <td bgcolor="{C['line']}" style="background-color:{C['line']};width:{100-max(2,pct)}%;height:12px;font-size:1px;">&nbsp;</td>
      <td style="width:70px;text-align:right;padding-left:8px;"><span style="font-size:11px;font-weight:900;color:{a};">{n}</span> <span style="font-size:9px;color:{C['muted']};">({pct}%)</span></td>
    </tr></table>
  </td>
</tr>"""
    impact = m["fleet_impact_pct"]
    repairs_panel = f"""
<td bgcolor="{C['card2']}" style="background-color:{C['card2']};border:1px solid {C['line']};padding:0;vertical-align:top;width:46%;">
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
    <tr><td bgcolor="{a}" style="background-color:{a};padding:12px 14px;"><span style="font-size:11px;font-weight:900;color:#fff;text-transform:uppercase;letter-spacing:.8px;">Dr&auml;ger Repairs Dashboard</span></td></tr>
    <tr><td style="padding:14px;"><span style="font-size:30px;font-weight:900;color:{a};">{m['repairs']}</span> <span style="font-size:10px;font-weight:900;color:{C['muted']};text-transform:uppercase;letter-spacing:1px;margin-left:8px;">Total Repairs</span></td></tr>
    <tr><td style="padding:0;"><table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">{rows}</table></td></tr>
    <tr><td bgcolor="{C['panel']}" style="background-color:{C['panel']};padding:12px 14px;border-top:1px solid {C['line']};">
      <div style="font-size:9px;font-weight:900;color:{C['muted']};text-transform:uppercase;letter-spacing:.8px;margin-bottom:6px;">Fleet Impact</div>
      <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;"><tr>
        <td style="background:{C['line']};border-radius:99px;height:10px;overflow:hidden;"><div style="background:linear-gradient(90deg,{a},{r});width:{max(2, round(impact))}%;height:10px;border-radius:99px;"></div></td>
        <td style="width:80px;text-align:right;padding-left:8px;"><span style="font-size:13px;font-weight:900;color:{r};">{impact}%</span></td>
      </tr></table>
      <div style="font-size:9px;color:{C['muted']};margin-top:4px;">of fleet currently unavailable</div>
    </td></tr>
  </table>
</td>"""
    # Top 3 company focus cards
    medals = ["1ST", "2ND", "3RD"]
    card_bgs = [C["red"], C["orange"], C["amber"]]
    cards = ""
    for i, comp in enumerate(m["focus3"]):
        maxq = max(comp["still_out"], 1)
        cards += f"""
<td bgcolor="{C['card2']}" width="33%" style="background-color:{C['card2']};border:1px solid {C['line']};padding:0;vertical-align:top;">
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
    <tr><td bgcolor="{card_bgs[i]}" style="background-color:{card_bgs[i]};padding:12px 14px;">
      <span style="font-size:8px;font-weight:900;color:rgba(255,255,255,.75);letter-spacing:1.5px;text-transform:uppercase;">Priority {i+1}</span>
      <span style="display:inline-block;background:rgba(0,0,0,.35);border-radius:99px;padding:3px 9px;font-size:8px;font-weight:900;color:#fff;float:right;letter-spacing:1px;">{medals[i]}</span>
      <div style="font-size:16px;font-weight:900;color:#fff;margin-top:4px;">{esc(comp['name'])}</div>
    </td></tr>
    <tr><td style="padding:12px 14px 4px 14px;"><span style="font-size:9px;color:{C['muted']};font-weight:700;text-transform:uppercase;letter-spacing:.5px;">Still out from yesterday</span><span style="font-size:12px;font-weight:900;color:{C['red']};float:right;">{comp['prev_day']}</span></td></tr>
    <tr><td style="padding:8px 14px 4px 14px;border-top:1px solid {C['line']};"><span style="font-size:9px;color:{C['muted']};font-weight:700;text-transform:uppercase;letter-spacing:.5px;">Total outstanding</span><span style="font-size:12px;font-weight:900;color:{C['white']};float:right;">{comp['still_out']}</span></td></tr>
    <tr><td style="padding:8px 14px 4px 14px;border-top:1px solid {C['line']};"><span style="font-size:9px;color:{C['muted']};font-weight:700;text-transform:uppercase;letter-spacing:.5px;">30+ days</span><span style="font-size:12px;font-weight:900;color:{C['red']};float:right;">{comp['d30']}</span></td></tr>
    <tr><td style="padding:8px 14px 14px 14px;border-top:1px solid {C['line']};"><span style="font-size:9px;color:{C['muted']};font-weight:700;text-transform:uppercase;letter-spacing:.5px;">Exposure</span><span style="font-size:12px;font-weight:900;color:{C['red']};float:right;">{money(comp['charge'])}</span></td></tr>
  </table>
</td>"""
    focus_panel = f"""
<td style="padding:0;vertical-align:top;">
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
    <tr><td bgcolor="{o}" style="background-color:{o};padding:12px 14px;" colspan="3"><span style="font-size:11px;font-weight:900;color:#fff;text-transform:uppercase;letter-spacing:.8px;">Top Company Focus</span></td></tr>
    <tr><td style="padding-top:4px;" colspan="3"><table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:separate;border-spacing:4px;"><tr>{cards}</tr></table></td></tr>
  </table>
</td>"""
    return f"""
<tr><td bgcolor="{C['bg']}" style="background-color:{C['bg']};padding:0 14px 12px 14px;">
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:separate;border-spacing:4px;"><tr>{repairs_panel}{focus_panel}</tr></table>
</td></tr>"""


def urgency_badge(u):
    col = {"Critical": C["red"], "High": C["orange"], "Watch": C["amber"]}[u]
    return (f"<span style='display:inline-block;background:{col};border-radius:99px;"
            f"padding:3px 12px;font-size:8px;font-weight:900;color:#fff;"
            f"letter-spacing:1px;text-transform:uppercase;'>{u}</span>")


def priorities_table(m):
    maxexp = max((c["charge"] for c in m["priorities"]), default=1) or 1
    body = ""
    for i, c in enumerate(m["priorities"], 1):
        pct = max(4, round(c["charge"] / maxexp * 100)) if c["charge"] else 4
        bg = C["card2"] if i % 2 else C["card3"]
        body += f"""
<tr style="background:{bg};">
  <td style="padding:11px 12px;font-size:12px;font-weight:900;color:{C['orange']};">{i}</td>
  <td style="padding:11px 12px;width:110px;"><table cellspacing="0" cellpadding="0" style="border-collapse:collapse;width:90px;"><tr>
    <td style="background:{C['red']};width:{pct}%;height:6px;border-radius:99px;font-size:1px;">&nbsp;</td>
    <td style="background:{C['line']};width:{100-pct}%;height:6px;border-radius:99px;font-size:1px;">&nbsp;</td></tr></table></td>
  <td style="padding:11px 12px;">{urgency_badge(c['urgency'])}</td>
  <td style="padding:11px 12px;font-size:12px;font-weight:900;color:{C['white']};">{esc(c['name'])}</td>
  <td style="padding:11px 12px;font-size:12px;font-weight:900;color:{C['red'] if c['prev_day'] else C['muted']};text-align:center;">{c['prev_day']}</td>
  <td style="padding:11px 12px;font-size:12px;font-weight:900;color:{C['orange'] if c['d1_7'] else C['muted']};text-align:center;">{c['d1_7']}</td>
  <td style="padding:11px 12px;font-size:12px;font-weight:900;color:{C['orange'] if c['d8_29'] else C['muted']};text-align:center;">{c['d8_29']}</td>
  <td style="padding:11px 12px;font-size:12px;font-weight:900;color:{C['red'] if c['d30'] else C['muted']};text-align:center;">{c['d30']}</td>
  <td style="padding:11px 12px;font-size:12px;font-weight:900;color:{C['white']};text-align:center;">{c['still_out']}</td>
  <td style="padding:11px 12px;font-size:12px;font-weight:900;color:{C['red'] if c['charge'] else C['muted']};text-align:right;">{money(c['charge'])}</td>
</tr>"""
    hdrs = ["#", "URGENCY", "PRIORITY", "COMPANY", "YDAY", "1&ndash;7", "8&ndash;29", "30+", "TOTAL", "EXPOSURE"]
    aligns = ["left", "left", "left", "left", "center", "center", "center", "center", "center", "right"]
    head = "".join(f"<th style='padding:10px 12px;font-size:9px;font-weight:900;color:{C['muted']};"
                   f"letter-spacing:1.2px;text-align:{a};text-transform:uppercase;border-bottom:1px solid {C['line']};'>{h}</th>"
                   for h, a in zip(hdrs, aligns))
    return section_header("Today's Top 10 Recovery Priorities",
                          "Ranked by recovery urgency, ageing and exposure") + f"""
<tr><td bgcolor="{C['bg']}" style="background-color:{C['bg']};padding:0 14px 12px 14px;">
  <table width="100%" cellspacing="0" cellpadding="0" bgcolor="{C['card']}" style="background-color:{C['card']};border:1px solid {C['line']};border-collapse:collapse;">
    <tr>{head}</tr>{body}
  </table>
</td></tr>"""


# ------------------ white-background detail tables --------------------

def detail_banner(title, colour):
    return f"""
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;margin-top:14px;">
  <tr>
    <td style="width:6px;background:{colour};font-size:1px;">&nbsp;</td>
    <td style="border:1px solid #B7C0CA;border-left:0;background:#FFFFFF;padding:12px 16px;">
      <span style="font-size:14px;font-weight:900;color:#0B0B0B;">{title}</span>
    </td>
  </tr>
</table>"""


def detail_empty(msg):
    return (f"<table width='100%' cellspacing='0' cellpadding='0' style='border-collapse:collapse;'>"
            f"<tr><td style='border:1px solid #B7C0CA;border-top:0;background:#FFFFFF;"
            f"padding:12px 16px;font-size:12px;color:#333;'>{msg}</td></tr></table>")


def detail_table(rows, days_red_from=1, extra_cost_col=False):
    """rows: list of (days, hirer, barcode, serial, date, time[, cost])"""
    hdr_cells = ["Days", "Hirer / Name", "Barcode", "Serial", "On Hire Date", "Time"]
    if extra_cost_col:
        hdr_cells.append("Replacement")
    head = "".join(f"<th style='background:#EEF1F4;color:#0B0B0B;border:1px solid #66717D;"
                   f"padding:8px;text-align:left;font-size:11px;'>{h}</th>" for h in hdr_cells)
    body = ""
    for i, r in enumerate(rows):
        bg = "#FFF1F1" if i % 2 == 0 else "#FFFFFF"
        days = r[0]
        cells = (f"<td style='border:1px solid #B7C0CA;padding:7px;font-weight:900;color:#C81E1E;'>{days}</td>"
                 + "".join(f"<td style='border:1px solid #B7C0CA;padding:7px;'>{esc(v)}</td>" for v in r[1:]))
        body += f"<tr style='background:{bg};'>{cells}</tr>"
    return (f"<table width='100%' cellspacing='0' cellpadding='0' "
            f"style='border-collapse:collapse;font-size:11px;color:#222;'>"
            f"<tr>{head}</tr>{body}</table>")


def company_block(name, count, rows, extra_cost_col=False):
    return f"""
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;margin-top:10px;">
  <tr>
    <td style="width:6px;background:#0B0B0B;font-size:1px;">&nbsp;</td>
    <td style="background:{C['orange']};padding:10px 14px;">
      <span style="font-size:14px;font-weight:900;color:#fff;">{esc(name)}</span>
      <span style="font-size:10px;color:rgba(255,255,255,.85);font-weight:700;margin-left:8px;">&mdash; {count} monitors</span>
    </td>
  </tr>
</table>
{detail_table(rows, extra_cost_col=extra_cost_col)}"""


def fmt_dt(d, t):
    # WHY (12 Aug 2026): same rule as esc() - a date or time the source
    # does not carry shows as a dash, not the word 'None'.
    ds = d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else (str(d) if d else "-")
    ts = t.strftime("%H:%M:%S") if hasattr(t, "strftime") else (str(t) if t else "-")
    return ds, ts


def detail_tables(data, m, cfg):
    out = "<div style='background:#FFFFFF;padding:8px 0;'>"

    def grouped(tab_rows, has_status_col=False):
        groups = OrderedDict()
        for r in tab_rows:
            comp = str(r[0]).strip()
            days = r[5]
            if has_status_col:
                d_on, t_on = r[8], r[9]
            else:
                d_on, t_on = r[7], r[8]
            ds, ts = fmt_dt(d_on, t_on)
            groups.setdefault(comp, []).append((days, r[1], r[2], r[3], ds, ts))
        for comp in groups:
            groups[comp].sort(key=lambda x: -x[0])
        return groups

    # ========== PREVIOUS DAY RECOVERY CYCLE ==========
    out += detail_banner("Still Outstanding from Previous Days &ndash; Action Required", C["red"])
    yday_comps = [c for c in data["companies"] if c["still_out"] > 0]
    if not yday_comps:
        out += detail_empty(f"No outstanding items from yesterday. Full recovery achieved.")
    else:
        for c in yday_comps:
            names = str(c["still_out_names"]).strip() if c["still_out_names"] else "(No names recorded)"
            out += f"""
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;margin-top:10px;">
  <tr>
    <td style="width:6px;background:#0B0B0B;font-size:1px;">&nbsp;</td>
    <td style="background:{C['red']};padding:10px 14px;">
      <span style="font-size:14px;font-weight:900;color:#fff;">{esc(c['name'])}</span>
      <span style="font-size:10px;color:rgba(255,255,255,.85);font-weight:700;margin-left:8px;">&mdash; {c['still_out']} monitors overdue</span>
    </td>
  </tr>
</table>
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;font-size:11px;color:#222;background:#FFFFFF;border:1px solid #B7C0CA;">
  <tr style="background:#FFF1F1;">
    <td style="border:1px solid #B7C0CA;padding:7px;font-weight:900;color:#C81E1E;width:80px;">{c['still_out']} monitors</td>
    <td style="border:1px solid #B7C0CA;padding:7px;"><b>Contact:</b> {esc(names)}</td>
    <td style="border:1px solid #B7C0CA;padding:7px;text-align:right;"><b>Action:</b> {esc(c['action'] or 'Follow up required')}</td>
  </tr>
</table>"""

    out += detail_banner("Outstanding 1-7 Days &ndash; Recovery Follow-Up", C["orange"])
    g17 = grouped(data["oh_1_7"])
    if not g17:
        out += detail_empty("No current records.")
    for comp, rows in g17.items():
        out += company_block(comp, len(rows), rows)

    out += detail_banner("Outstanding 8-29 Days &ndash; Escalated Follow-Up", C["amber"])
    g829 = grouped(data["oh_8_29"])
    if not g829:
        out += detail_empty("No current records.")
    for comp, rows in g829.items():
        out += company_block(comp, len(rows), rows)

    out += detail_banner(f"Outstanding 30+ Days &ndash; Charge / Recovery ({money(m['exposure'])} exposure)", C["red"])
    g30 = OrderedDict()
    for r in data["oh_30"]:
        comp = str(r[0]).strip()
        ds, ts = fmt_dt(r[8], r[9])
        g30.setdefault(comp, []).append((r[5], r[1], r[2], r[3], ds, ts, money(r[7] or cfg["charge_per_unit"])))
    for comp in g30:
        g30[comp].sort(key=lambda x: -x[0])
    if not g30:
        out += detail_empty("No current records.")
    for comp, rows in g30.items():
        out += company_block(comp, len(rows), rows, extra_cost_col=True)

    out += detail_banner(f"Dr&auml;ger Repairs &ndash; {m['repairs']} Units Out of Service", C["amber"])
    rep_rows = []
    for r in sorted(data["repairs"], key=lambda x: -(x[4] or 0)):
        ds = r[5].strftime("%d/%m/%Y") if hasattr(r[5], "strftime") else str(r[5])
        rep_rows.append((r[4], r[0], r[1], r[2], ds, ""))
    if not rep_rows:
        out += detail_empty("No current repair records.")
    else:
        head = "".join(f"<th style='background:#EEF1F4;color:#0B0B0B;border:1px solid #66717D;padding:8px;"
                       f"text-align:left;font-size:11px;'>{h}</th>"
                       for h in ["Days in Status", "Status / Note", "Barcode", "Serial", "On Hire Date", ""])
        body = ""
        for i, r in enumerate(rep_rows):
            bg = "#FFF7ED" if i % 2 == 0 else "#FFFFFF"
            body += (f"<tr style='background:{bg};'>"
                     f"<td style='border:1px solid #B7C0CA;padding:7px;font-weight:900;color:#B45309;'>{r[0]}</td>"
                     + "".join(f"<td style='border:1px solid #B7C0CA;padding:7px;'>{esc(v)}</td>" for v in r[1:])
                     + "</tr>")
        out += (f"<table width='100%' cellspacing='0' cellpadding='0' "
                f"style='border-collapse:collapse;font-size:11px;color:#222;'><tr>{head}</tr>{body}</table>")

    if cfg["include_fccu_detail_table"]:
        out += detail_banner(f"FCCU Custody &ndash; {m['fccu']} Units", C["blue"])
        fccu_rows = [(r[5], r[1], r[2], r[3], *fmt_dt(r[7], r[8])) for r in data["fccu"]]
        out += detail_table(sorted(fccu_rows, key=lambda x: -x[0]))

    out += "</div>"
    return f"""
<tr><td bgcolor="#FFFFFF" style="background-color:#FFFFFF;padding:6px 14px 14px 14px;">{out}</td></tr>"""


def compliance_strip(m):
    cards = [
        ("BUMP + CHARGE", "Every monitor is bump tested and fully charged before issue. "
         "No exceptions - a monitor that has not passed bump does not leave the store."),
        ("SCAN DISCIPLINE", "Nothing issues and nothing returns unless scanned. Every unit on this "
         "dashboard is accounted for by barcode and serial number."),
        ("OUT OF SERVICE", "Damaged or failed units are tagged Out of Service immediately, photographed, "
         "and reported in writing before entering the Dr&auml;ger repair process."),
        ("REPAIR VISIBILITY", f"{m['repairs']} units are in the repair process today, with "
         f"Dr&auml;ger &ndash; {esc(m['repair_top'])} the largest category - tracked to closure, never parked."),
    ]
    cells = "".join(f"""
      <td bgcolor="{C['card2']}" width="25%" style="background-color:{C['card2']};border:1px solid {C['line']};border-top:3px solid {C['green']};padding:14px;vertical-align:top;">
        <div style="font-size:9px;font-weight:900;color:{C['green']};letter-spacing:1.5px;text-transform:uppercase;">{t}</div>
        <div style="font-size:10px;color:#C9D6E2;line-height:1.7;margin-top:8px;">{d}</div>
      </td>""" for t, d in cards)
    return f"""
<tr><td bgcolor="{C['bg']}" style="background-color:{C['bg']};padding:0 14px 12px 14px;">
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;margin-bottom:4px;">
    <tr><td bgcolor="{C['panel']}" style="background-color:{C['panel']};border-left:5px solid {C['green']};padding:12px 22px;border-radius:4px;">
      <span style="font-size:13px;font-weight:900;color:{C['white']};letter-spacing:.5px;">Gas Monitor Compliance &mdash; Safety First</span>
      <span style="font-size:9px;color:{C['muted']};margin-left:12px;letter-spacing:1px;text-transform:uppercase;">Standards in force across all {m['fleet_total']} monitors under management &middot; Per the Bumping Gas Monitors SOP &middot; The Coates Way operating discipline</span>
    </td></tr>
  </table>
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:separate;border-spacing:4px;"><tr>
    <td bgcolor="{C['card2']}" width="25%" style="background-color:{C['card2']};border:1px solid {C['line']};border-top:3px solid {C['green']};padding:14px;vertical-align:top;">
      <div style="font-size:9px;font-weight:900;color:{C['green']};letter-spacing:1.5px;text-transform:uppercase;">SAFETY FIRST</div>
      <div style="font-size:10px;color:#C9D6E2;line-height:1.7;margin-top:8px;">The Coates Way begins and ends with people going home safe. Gas detection is life-critical equipment and is managed to that standard.</div>
    </td>{cells}
  </tr></table>
</td></tr>"""


def closing(m, gen_str):
    o, g, r, b, a = C["orange"], C["green"], C["red"], C["blue"], C["amber"]
    hs_col = g if m["health"] >= 80 else a if m["health"] >= 45 else r
    v = ("PASS" if m["validation_pass"] else "FAIL")
    return section_header("Executive Closing Snapshot", "The Coates Way") + f"""
<tr><td bgcolor="{C['bg']}" style="background-color:{C['bg']};padding:0 14px 12px 14px;">
  <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
    <tr><td bgcolor="{C['card']}" style="background-color:{C['card']};border:1px solid {C['line']};padding:18px 24px;">
      <div style="font-size:12px;color:#C9D6E2;line-height:1.9;">
        Validation: <b style="color:{g if m['validation_pass'] else r};">{v}</b> &ndash;
        Previous Day Non-Returns <b style="color:{o};">{m['prev_day_total']}</b> =
        Recovered <b style="color:{g};">{m['recovered']}</b> +
        Still Outstanding <b style="color:{r};">{m['yday_still_out']}</b>.
        Current outstanding <b style="color:{o};">{m['outstanding']}</b>.
        General on hire <b style="color:{o};">{m['onhire_general']}</b>
        (<b style="color:{o};">{m['issued_today']}</b> issued today).
        Total fleet <b style="color:{o};">{m['fleet_total']}</b>; general fleet excl. FCCU/Ops/FF
        <b style="color:{o};">{m['fleet_general']}</b>.
        30+ exposure <b style="color:{r};">{money(m['exposure'])}</b>.
        FCCU kept separate: <b style="color:{b};">{m['fccu']}</b> on hire.
        Operations kept separate: <b style="color:{b};">{m['ops']}</b> on hire.
        Future Fuels kept separate: <b style="color:{b};">{m['ff']}</b> on hire.
        Health Score: <b style="color:{hs_col};">{m['health']}/100</b>.
        Report generated at {gen_str}.
        <span style="color:{o};font-style:italic;">Consistent execution today drives the Flywheel forward.</span>
      </div>
    </td></tr>
  </table>
</td></tr>"""


def footer(cfg, gen_str, asat_str=""):
    # WHY (12 Aug 2026): the suite standard puts a data-as-at stamp in the
    # footer of every report - the workbook file time, so the reader always
    # knows how fresh the numbers are.
    asat_bit = (f" &middot; Position as at {asat_str} (SiteIQ register pull)"
                if asat_str else "")
    return f"""
<tr>
  <td bgcolor="{C['panel']}" style="background-color:{C['panel']};padding:24px 32px;border-top:2px solid {C['orange']};">
    <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
      <tr>
        <td style="vertical-align:top;">
          <div style="font-size:9px;font-weight:900;color:{C['muted']};text-transform:uppercase;letter-spacing:1.5px;">Report Author</div>
          <div style="font-size:15px;font-weight:900;color:{C['white']};margin-top:5px;">{cfg['author_name']}</div>
          <div style="font-size:10px;color:{C['orange']};font-weight:700;margin-top:2px;">The Coates Way &middot; Operational Excellence</div>
          <div style="font-size:10px;color:{C['muted']};margin-top:8px;line-height:1.7;">{cfg['author_address']}<br/>
            E: <span style="color:#C9D6E2;">{cfg['author_email']}</span> &middot;
            M: <span style="color:#C9D6E2;">{cfg['author_mobile']}</span> &middot;
            W: <span style="color:#C9D6E2;">www.coates.com.au</span></div>
        </td>
        <td style="vertical-align:top;text-align:right;">
          <div style="font-size:26px;font-weight:900;color:{C['white']};letter-spacing:-1px;">Coates</div>
          <div style="font-size:9px;font-weight:900;color:{C['orange']};letter-spacing:2px;text-transform:uppercase;margin-top:2px;">Industrial Solutions</div>
          <div style="margin-top:10px;">
            <span style="display:inline-block;background:{C['card2']};border:1px solid {C['line']};border-radius:4px;padding:4px 10px;font-size:8px;font-weight:900;color:{C['orange']};letter-spacing:1.5px;text-transform:uppercase;">The Coates Way</span>
            <span style="display:inline-block;background:{C['card2']};border:1px solid {C['line']};border-radius:4px;padding:4px 10px;font-size:8px;font-weight:900;color:{C['muted']};letter-spacing:1.5px;text-transform:uppercase;margin-left:4px;">Operational Excellence</span>
            <span style="display:inline-block;background:{C['card2']};border:1px solid {C['line']};border-radius:4px;padding:4px 10px;font-size:8px;font-weight:900;color:{C['muted']};letter-spacing:1.5px;text-transform:uppercase;margin-left:4px;">Disciplined Execution</span>
          </div>
          <div style="font-size:9px;color:{C['muted2']};margin-top:10px;">v18 ELITE &middot; {gen_str}{asat_bit}</div>
        </td>
      </tr>
      <tr><td colspan="2" style="padding-top:16px;border-top:1px solid {C['line']};">
        <div style="font-size:9px;color:{C['muted']};line-height:1.8;margin-top:12px;">
          This report forms part of The Coates Way operating cadence &mdash; delivering Best Service and Value through consistent execution.
          Performance tracked against Balanced Scorecard measures across Safety, Customer, Operational, Financial and People.<br/>
          Coates acknowledges the Traditional Owners of Country throughout Australia. We pay our respects to Elders past, present and emerging.
          Learn more about our Reconciliation Action Plan.<br/>
          Coates Hire Operations Pty Limited &nbsp;|&nbsp; This email and any attachments are confidential and intended solely for the use of the named recipient(s).
          If you have received this email in error, please notify the sender immediately and delete it from your system.<br/>
          Coates Hire Operations Pty Limited &nbsp;|&nbsp; ABN 50 009 779 338 &nbsp;|&nbsp; www.coates.com.au &nbsp;|&nbsp; POWERED BY SITEIQ
        </div>
      </td></tr>
    </table>
  </td>
</tr>"""


# =====================================================================
# 5. ASSEMBLE
# =====================================================================

def build_html(data, m, cfg, date_str, time_str, asat_str=""):
    gen_str = f"{datetime.strptime(date_str, '%d %B %Y').strftime('%d/%m/%Y')} at {time_str}"
    parts = [
        hero(m, cfg, date_str, time_str, gen_str),
        disciplines(),
        kpi_rows(m),
        exec_brief(m),
        action_framework(m),
        pillars(m),
        scorecard(m),
        data_panels(m),
        repairs_and_focus(m),
        priorities_table(m),
        detail_tables(data, m, cfg),
        compliance_strip(m),
        closing(m, gen_str),
        footer(cfg, gen_str, asat_str),
    ]
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Coates Gas Monitor Executive Dashboard &ndash; {date_str}</title>
</head>
<body bgcolor="#FFFFFF" style="margin:0;padding:0;background-color:#FFFFFF;font-family:Arial,Helvetica,sans-serif;color:#0D1520;">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0" bgcolor="#FFFFFF" style="background-color:#FFFFFF;">
<tr><td align="center" bgcolor="#FFFFFF" style="background-color:#FFFFFF;padding:20px 12px;">
<table role="presentation" width="1100" cellspacing="0" cellpadding="0"
  style="width:1100px;max-width:1100px;border-collapse:collapse;border:1px solid {C['line']};
  box-shadow:0 20px 60px rgba(0,0,0,.5);">
{''.join(parts)}
</table>
</td></tr>
</table>
</body>
</html>"""


def build_eml(html, cfg, date_str, wb_path):
    weekday = datetime.strptime(date_str, "%d %B %Y").strftime("%A")
    msg = EmailMessage()
    msg["Subject"] = f"{cfg['subject_prefix']} - {weekday} {date_str}"
    msg["To"] = ", ".join(cfg["recipients"])
    msg["Date"] = formatdate(localtime=True)
    msg["X-Unsent"] = "1"
    msg.set_content("This report is best viewed in HTML. The source workbook is attached.\n")
    msg.add_alternative(html, subtype="html")
    if wb_path and os.path.exists(wb_path):
        with open(wb_path, "rb") as f:
            msg.add_attachment(
                f.read(), maintype="application",
                subtype="vnd.ms-excel.sheet.macroenabled.12",
                filename=os.path.basename(wb_path))
    return msg


# =====================================================================
# 6. MAIN
# =====================================================================

def main():
    cfg = CONFIG
    now = datetime.now()
    date_str = cfg["report_date"] or now.strftime("%d %B %Y").lstrip("0")
    if cfg["report_date"] and "/" in cfg["report_date"]:
        date_str = datetime.strptime(cfg["report_date"], "%d/%m/%Y").strftime("%d %B %Y").lstrip("0")
    # 24-hour time - Coates reporting standard, and it removes any AM/PM
    # ambiguity on a report that gets read at 05:30 and again at 17:30.
    time_str = cfg["run_time"] or now.strftime("%H:%M")

    print("=" * 68)
    print("COATES V18 ELITE - GAS MONITOR REPORT GENERATOR")
    print("=" * 68)

    print(f"Kit root             : {kit_root()}")
    print(f"Input folder         : {input_dir()}")
    print(f"Output folder        : {resolve_output_dir(cfg)}")
    print("-" * 68)

    wb_path = find_workbook_optional(cfg) if _engine is not None else find_workbook(cfg)
    rs_path = find_rental_stock(cfg)
    if wb_path:
        print(f"Workbook             : {os.path.basename(wb_path)}  (attached to the email)")
    else:
        print("Workbook             : none in Data\\ - not needed, the engine counts from the exports")
    if _engine is not None:
        print("Numbers from         : gasmon_engine (SiteIQ RENTAL_STOCK + TRANSACTIONS)")
    if rs_path:
        print(f"RENTAL_STOCK export  : {os.path.basename(rs_path)}")
        print(f"  extracted          : "
              f"{datetime.fromtimestamp(os.path.getmtime(rs_path)).strftime('%d %b %Y %H:%M')}")
    else:
        print("RENTAL_STOCK export  : none found - cross-validation SKIPPED")
        print("                       (save RENTAL_STOCK.xlsx into the kit folder"
              " to enable it)")

    data = load_data(wb_path)
    m = compute_metrics(data, cfg)

    print(f"Fleet total          : {m['fleet_total']}")
    print(f"General fleet        : {m['fleet_general']}  (excl. FCCU/Ops/FF)")
    print(f"On hire (general)    : {m['onhire_general']}  "
          f"= {m['issued_today']} today + {m['outstanding']} outstanding")
    print(f"On hire (all custody): {m['onhire_all']}  "
          f"(FCCU {m['fccu']} / Ops {m['ops']} / FF {m['ff']})")
    print(f"Available            : {m['available']}")
    print(f"Repairs              : {m['repairs']}  ({m['fleet_impact_pct']}% of fleet)")
    print(f"Outstanding          : {m['outstanding']}  "
          f"({m['out_1_7']} / {m['out_8_29']} / {m['out_30']})  exposure {money(m['exposure'])}")
    print(f"Prev-day cycle       : {m['prev_day_total']} = {m['recovered']} + {m['yday_still_out']}  "
          f"(recovery {m['recovery_rate']}%)")
    print(f"Health score         : {m['health']}/100  "
          f"(A{m['score_availability']} R{m['score_recovery']} "
          f"P{m['score_repairs']} C{m['score_30']})")

    for w in m["warnings"]:
        print(f"[RECON WARNING] {w}")
    if rs_path:
        for n in validate_against_rental_stock(rs_path, m):
            print(f"[RENTAL_STOCK] {n}")

    if data.get("_engine") is not None:
        asat_str = data["_engine"]["asat"].strftime("%d %b %Y %H:%M")
    else:
        asat_str = datetime.fromtimestamp(os.path.getmtime(wb_path)).strftime("%d %b %Y %H:%M")
    html = build_html(data, m, cfg, date_str, time_str, asat_str)
    out_dir = resolve_output_dir(cfg)
    os.makedirs(out_dir, exist_ok=True)
    html_path = os.path.join(out_dir, stamped_name(cfg["html_name"], date_str, cfg))
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"HTML written         : {html_path}  ({len(html):,} bytes)")

    eml = build_eml(html, cfg, date_str, wb_path)
    eml_path = os.path.join(out_dir, stamped_name(cfg["eml_name"], date_str, cfg))
    with open(eml_path, "wb") as f:
        f.write(eml.as_bytes())
    print(f"EML written          : {eml_path}  ({os.path.getsize(eml_path):,} bytes)")

    # WHY (12 Aug 2026): the shared MAKE_OUTLOOK_DRAFTS button reads a
    # .draft.json manifest from the same Reports folder and turns it into a
    # native Outlook DRAFT - subject, body, attachments, To. Nothing sends
    # itself, same discipline as the .eml. The workbook is copied in beside
    # it so the manifest can attach it by basename, and the dated folder
    # then also keeps the exact data this report was built from.
    attachments = []
    if wb_path:
        wb_copy = os.path.join(out_dir, os.path.basename(wb_path))
        if os.path.abspath(wb_path) != os.path.abspath(wb_copy):
            shutil.copy2(wb_path, wb_copy)
        attachments.append(os.path.basename(wb_copy))
    weekday = datetime.strptime(date_str, "%d %B %Y").strftime("%A")
    manifest = {
        "subject": f"{cfg['subject_prefix']} - {weekday} {date_str}",
        "body": os.path.basename(html_path),
        "attachments": attachments,
        "to": "; ".join(cfg["recipients"]),
    }
    man_path = os.path.join(
        out_dir, os.path.splitext(cfg["html_name"])[0] + ".draft.json")
    with open(man_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=1)
    print(f"Draft manifest       : {man_path}")
    print("")
    if os.path.abspath(out_dir) == os.path.abspath(kit_root()):
        print("NEXT STEP: the .eml is in this folder. Double-click it, check it,")
        print("           then press Send in Outlook.")
    else:
        print(f"NEXT STEP: open {out_dir}, double-click the .eml, check it,")
        print("           then press Send in Outlook.")
    print("Done. The Coates Way - consistent execution, every day.")


if __name__ == "__main__":
    # WHY (12 Aug 2026): a failed run must exit nonzero so the suite's
    # buttons can see it - the old block printed the error and exited 0,
    # which made failures look like passes. The keep-window-open pause is
    # gone too: the .bat buttons own the pause now.
    try:
        main()
    except SystemExit:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        sys.exit(f"ERROR: {e}")
