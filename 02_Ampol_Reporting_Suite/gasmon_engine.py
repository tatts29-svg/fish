#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=====================================================================
COATES | AMPOL GAS MONITOR ENGINE - every number from the SiteIQ exports
=====================================================================
Author: Andrew Fisher | POWERED BY SITEIQ

WHY (02 Sep 2026)
  The gas monitor report was being questioned on accuracy. The cause:
  the headline figures were read off the Excel workbook's summary tab,
  whose "Issued / Returned / Not returned same day" columns run over a
  window that is neither year-to-date nor the last 30 days, while the
  report labelled them as both. "Recovered today" was hard-coded to
  zero, which pinned the Returns score at 0/100 every morning. Custody
  accounts (After Hours, Dräger) were named as people.

  This module replaces all of that with one engine that counts every
  figure from the two SiteIQ exports Andrew already pulls each morning:

    RENTAL_STOCK.xlsx   - WHERE every monitor is right now (the position)
    TRANSACTIONS.xlsx   - every issue and return with a timestamp (the flow)

  The Excel workbook is no longer a source of numbers. It stays in
  Data\\ for the email attachment and as a cross-check that is printed
  on the report's data page, never hidden.

THE RULES (printed on the report's method page, kept in one place here)
  Gas monitor     description contains "X-am" or "gas monitor"; chargers
                  and probes are not monitors.
  Crew draw       a transaction to a named person. Custody and workflow
                  accounts are NOT crew: Dräger service statuses (Failed
                  Bump Test, Out of Calibration, Out of Service ...),
                  Dräger FCCU / FCCU T&I, Ampol Operations, Future Fuels,
                  and the After Hours account. They are reported on their
                  own lines, never as a person.
  Same day        returned on the calendar day it went out. A draw made
                  at or after NIGHT_SHIFT_FROM (15:00) counts as same day
                  if it is back by NIGHT_SHIFT_BACK_BY (08:00) next morning.
  Not same day    everything else, including draws still open.
  Outstanding     on hire to a person with an on-hire date before the
                  report date (1+ days). Gear issued today is "out today",
                  not outstanding.
  Windows         YTD = the export window (01 Jan to the pull);
                  30 days = the 30 days up to the pull; 7 days likewise;
                  yesterday = the last complete day before the pull.
  Australian dates, parsed explicitly. Never the US default.

  Nothing is estimated. If a value is not in the source it is left out
  and said so.
=====================================================================
"""

import os
import re
from collections import Counter, defaultdict
from datetime import datetime, timedelta, date, time as dtime
from statistics import median

import openpyxl

import ampol_paths
import ampol_names

# ---------------------------------------------------------------------
# Business rules - the only knobs. Everything else is arithmetic.
# ---------------------------------------------------------------------

RULES = {
    "charge_per_unit": 2000,       # $ replacement charge for a 30+ day monitor
                                   # (the workbook's REPLACEMENT COST column)
    "availability_target": 250,    # units on the shelf for a 100 availability score
    "night_shift_from": dtime(15, 0),   # a draw at/after this is a night-shift draw
    "night_shift_back_by": dtime(8, 0),  # ...and counts as same day if back by this
    # WHY (03 Sep 2026, Andrew): monitors dropped in the return box after
    # the counter closes are scanned by the first shift between 04:00 and
    # 05:30 next morning, before the store opens at 07:00. A return scanned
    # in that window was back on the day it went out - the scan is the
    # store's, not the crew's - so it counts as same day. Disclosed on the
    # data page and counted separately ("via the return box").
    "return_box_scan_until": dtime(5, 30),
    "repeat_weeks": 3,             # weeks with a non-return in the last 30 days = repeat
    "stale_repair_days": 180,      # a repair older than this is dead fleet
}

# Words in a HIRER name that mark a custody / workflow account, not a person.
_ACCOUNT = re.compile(
    r"dr[aä]ger|fccu|t&i$|^t&i|operations?$|future\s*-?\s*fuels|after\s*hours",
    re.I)

# Company acronyms that stay upper-case when everything else is title-cased.
_ACRONYMS = {"HIS", "ARL", "IPS", "CSA", "BMD", "UGL", "NDE", "CXC", "AGM",
             "WSP", "BLJ", "IPCQ", "FSACE", "UGL"}

GAS_RE = re.compile(r"x-am|gas monitor", re.I)
NOT_GAS_RE = re.compile(r"charger|probe|pump|calibration gas|dock|cradle", re.I)


# =====================================================================
# helpers
# =====================================================================

def is_gas_monitor(desc):
    d = str(desc or "")
    return bool(GAS_RE.search(d)) and not NOT_GAS_RE.search(d)


def parse_dt(dv, tv):
    """SiteIQ dates: DD/MM/YYYY plus HH:MM:SS or hh:mm AM/PM. Explicit
    Australian parse - a US default would produce plausible, wrong reports.
    Excel datetimes come back as datetime objects and are passed through."""
    if dv in (None, ""):
        return None
    if isinstance(dv, datetime):
        base = dv
        if isinstance(tv, dtime):
            return datetime.combine(base.date(), tv)
        if isinstance(tv, datetime):
            return datetime.combine(base.date(), tv.time())
        if tv in (None, ""):
            return base
        tv = str(tv).strip()
        for f in ("%H:%M:%S", "%I:%M %p", "%H:%M", "%I:%M:%S %p"):
            try:
                return datetime.combine(base.date(), datetime.strptime(tv, f).time())
            except ValueError:
                continue
        return base
    ds = str(dv).strip()
    ts = "" if tv in (None, "") else (tv.strftime("%H:%M:%S") if isinstance(tv, (dtime, datetime)) else str(tv).strip())
    for f in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %I:%M %p", "%d/%m/%Y %H:%M",
              "%d/%m/%Y %I:%M:%S %p", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
              "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime((ds + " " + ts).strip(), f)
        except ValueError:
            continue
    return None


def parse_stamp(s):
    """REFERENCE_INFO stamps: '02/09/2026 06:30 PM' or '02/09/2026 05:59:59'."""
    s = str(s or "").strip()
    for f in ("%d/%m/%Y %I:%M %p", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
              "%d/%m/%Y"):
        try:
            return datetime.strptime(s, f)
        except ValueError:
            continue
    return None


def account_kind(who, co):
    """Custody / workflow account kind, or 'crew' for a person.
    Only the HIRER name decides - a person working under the employer
    'Contract Resources FCCU' is still a person."""
    w = str(who or "").strip()
    wl = w.lower()
    cl = str(co or "").strip().lower()
    if "after hours" in wl or "after-hours" in wl:
        return "afterhours"
    if "future" in wl and "fuel" in wl:
        return "ff"
    if cl == "future fuels" and not re.search(r"[a-z]+\s+-?\s*[a-z]+", wl.replace("future", "").replace("fuels", "").strip()):
        return "ff"
    if "fccu" in wl:
        return "fccu"
    if re.search(r"dr[aä]ger", wl):
        return "repair"
    if re.search(r"\boperations?\b", wl) and re.search(r"ampol|site|refinery|^operations?$", wl):
        return "ops"
    if re.fullmatch(r"\s*(fccu\s*)?t&i\s*", wl):
        return "fccu"
    return "crew"


def repair_status(who):
    """'Dräger - Dräger – Out of Calibration' -> 'Out of Calibration'."""
    s = str(who or "")
    s = re.sub(r"dr[aä]ger\s*[–\-]\s*", "", s, flags=re.I)
    s = re.sub(r"dr[aä]ger\s*[–\-]\s*", "", s, flags=re.I)
    s = re.sub(r"^\s*dr[aä]ger\s*", "", s, flags=re.I)
    return s.strip(" –-") or "Dräger"


def norm_person(name):
    """'David - McGurk', 'David McGurk', 'Rangi Tamihana-Shutdown',
    'JAY PURCELL T&I' -> ('david mcgurk', 'David McGurk')."""
    s = str(name or "").strip()
    s = re.sub(r"\s*-\s*shutdown\s*$", "", s, flags=re.I)
    s = re.sub(r"\s+t&i\s*$", "", s, flags=re.I)
    s = re.sub(r"\s+-\s+", " ", s)          # 'First - Last' -> 'First Last'
    s = re.sub(r"\s+", " ", s).strip()
    key = s.lower()
    # display: the suite-wide rule (ampol_names.display_person) on the
    # merged 'First Last' form - inner capitals like McGurk / O'Flynn kept
    return key, ampol_names.display_person(s)


def norm_company(co):
    """One customer, one name - the shared rule in ampol_names (the former
    site-name account and the refinery legal name both read Ampol; project
    accounts roll up to their parent; acronym companies stay upper-case)."""
    return ampol_names.display_company(co)


def same_day(st, en):
    """The return rule. None (still open) is never same day."""
    if en is None:
        return False
    if en.date() == st.date():
        return True
    if via_return_box(st, en):
        return True
    if st.time() >= RULES["night_shift_from"]:
        back_by = datetime.combine(st.date() + timedelta(days=1),
                                   RULES["night_shift_back_by"])
        return en <= back_by
    return False


def via_return_box(st, en):
    """A return scanned the next morning inside the return-box window
    (04:00 to return_box_scan_until) for a draw made the day before: the
    crew put it in the box on the day; the store scanned it at shift start."""
    if en is None or en.date() != st.date() + timedelta(days=1):
        return False
    return dtime(4, 0) <= en.time() <= RULES["return_box_scan_until"]


def hm(minutes):
    return f"{int(minutes) // 60:02d}:{int(minutes) % 60:02d}"


# =====================================================================
# LOAD - the two exports (+ the serial list for display only)
# =====================================================================

def find_export(prefix):
    return ampol_paths.find_data(f"{prefix}*.xlsx", f"{prefix.lower()}*.xlsx")


def load_serials():
    """barcode -> serial, from Gas_Monitor_Serial_Numbers.xlsx. Display only."""
    path = ampol_paths.find_data("Gas_Monitor_Serial*.xlsx", "*serial*.xlsx")
    out = {}
    if not path:
        return out, ""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for s in wb.sheetnames:
        for r in wb[s].iter_rows(min_row=2, values_only=True):
            if r and r[0] and len(r) > 1 and r[1]:
                out.setdefault(str(r[0]).strip().upper(), str(r[1]).strip())
    wb.close()
    return out, path


def load_rental_stock(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    requested = None
    if "REFERENCE_INFO" in wb.sheetnames:
        rows = list(wb["REFERENCE_INFO"].iter_rows(values_only=True))
        if len(rows) > 1:
            hdr = [str(h or "").upper() for h in rows[0]]
            for i, h in enumerate(hdr):
                if "REQUESTED_DATE" in h:
                    requested = parse_stamp(rows[1][i])
    ws = wb["RENTAL_STOCK"]
    hdr = [str(h or "").strip().upper() for h in
           next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {h: i for i, h in enumerate(hdr)}

    def g(r, name, default=""):
        i = col.get(name)
        return r[i] if i is not None and i < len(r) else default

    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        # the site is Ampol: shown under the current name (ampol_names);
        # the monitor test below is on words the rename never touches
        raw_desc = str(g(r, "ITEM_DESCRIPTION") or "")
        desc = ampol_names.display_desc(raw_desc, barcode=g(r, "ITEM_BARCODE"))
        if not is_gas_monitor(raw_desc):
            continue
        co, who = g(r, "COMPANY_NAME"), g(r, "HIRER_NAME")
        status = str(g(r, "ITEM_STATUS") or "").strip()
        on_dt = parse_dt(g(r, "ON_HIRE_DATE"), g(r, "ON_HIRE_TIME"))
        out.append({
            "co_raw": str(co or "").strip(), "who_raw": str(who or "").strip(),
            "co": norm_company(co) if co else "",
            "who": norm_person(who)[1] if who else "",
            "who_key": norm_person(who)[0] if who else "",
            "bc": str(g(r, "ITEM_BARCODE") or "").strip(),
            "desc": str(desc).strip(),
            "status": status,
            "on_hire": status.lower() == "on hire",
            "on_dt": on_dt,
            "unit": str(g(r, "STORAGE_UNIT") or "").strip(),
            "kind": account_kind(who, co) if status.lower() == "on hire" else "",
        })
    wb.close()
    return out, requested


def load_transactions(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    window = (None, None)
    requested = None
    if "REFERENCE_INFO" in wb.sheetnames:
        rows = list(wb["REFERENCE_INFO"].iter_rows(values_only=True))
        if len(rows) > 1:
            hdr = [str(h or "").upper() for h in rows[0]]
            for i, h in enumerate(hdr):
                v = rows[1][i] if i < len(rows[1]) else None
                if "REPORT_PERIOD" in h and v:
                    bits = [b.strip() for b in str(v).split(" - ")]
                    if len(bits) == 2:
                        window = (parse_stamp(bits[0]), parse_stamp(bits[1]))
                if "REQUESTED_DATE" in h:
                    requested = parse_stamp(v)
    ws = wb["CUSTOMER_CONTRACTOR_EQUIP"]
    hdr = [str(h or "").strip().upper() for h in
           next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {h: i for i, h in enumerate(hdr)}

    def g(r, name):
        i = col.get(name)
        return r[i] if i is not None and i < len(r) else None

    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        desc = ampol_names.display_desc(str(g(r, "PRODUCT_VARIANT") or "") + " " + str(g(r, "SKU/ITEM DESCRIPTION") or ""))
        if not is_gas_monitor(desc):
            continue
        st = parse_dt(g(r, "TRAN_START_DATE"), g(r, "TRAN_START_TIME"))
        if st is None:
            continue
        en = parse_dt(g(r, "TRAN_END_DATE"), g(r, "TRAN_END_TIME"))
        co, who = g(r, "EMPLOYER_NAME"), g(r, "HIRER_NAME")
        key, disp = norm_person(who)
        out.append({
            "co_raw": str(co or "").strip(), "who_raw": str(who or "").strip(),
            "co": norm_company(co), "who": disp, "who_key": key,
            "bc": str(g(r, "LATEST_BARCODE") or "").strip(),
            "st": st, "en": en,
            "kind": account_kind(who, co),
            "sd": same_day(st, en),
            "box": via_return_box(st, en),
        })
    wb.close()
    return out, window, requested


def load(data_dir=None):
    """Everything the report needs, from Data\\. Fails loudly and plainly."""
    rs_path = find_export("RENTAL_STOCK")
    tx_path = find_export("TRANSACTIONS")
    missing = [n for n, p in (("RENTAL_STOCK.xlsx", rs_path),
                              ("TRANSACTIONS.xlsx", tx_path)) if not p]
    if missing:
        raise SystemExit(
            "ERROR: the gas monitor report needs these SiteIQ exports in Data\\ "
            "and could not find them:\n" +
            "".join(f"  - {m}\n" for m in missing) +
            "  Download them from SiteIQ, run 12_PULL_SITEIQ_EXPORTS, and press "
            "the button again.")
    rs, rs_req = load_rental_stock(rs_path)
    tx, tx_win, tx_req = load_transactions(tx_path)
    serials, ser_path = load_serials()
    for r in rs:
        r["serial"] = serials.get(r["bc"].upper(), "")
    asat = rs_req or datetime.fromtimestamp(os.path.getmtime(rs_path))
    return {
        "rs": rs, "tx": tx, "serials": serials,
        "rs_path": rs_path, "tx_path": tx_path, "ser_path": ser_path,
        "asat": asat,                      # position as at (RENTAL_STOCK pull)
        "tx_window": tx_win,               # (start, end) issues covered
        "tx_requested": tx_req or datetime.fromtimestamp(os.path.getmtime(tx_path)),
    }


# =====================================================================
# COMPUTE
# =====================================================================

def _flow(crew, start, end, now):
    """Flow aggregates for crew draws issued in [start, end]."""
    sel = [t for t in crew if start <= t["st"] <= end]
    closed = [t for t in sel if t["en"]]
    sd = sum(1 for t in sel if t["sd"])
    box = sum(1 for t in sel if t.get("box"))
    nsd = len(sel) - sd
    open_now = sum(1 for t in sel if not t["en"])
    imins = sorted(t["st"].hour * 60 + t["st"].minute for t in sel)
    rmins = sorted(t["en"].hour * 60 + t["en"].minute for t in closed)
    days = sorted({t["st"].date() for t in sel})
    wdays = [d for d in days if d.weekday() < 5]
    hh = Counter((t["st"].date(), t["st"].hour) for t in sel)
    if hh:
        (rd, rh), rn = hh.most_common(1)[0]
        record = {"date": rd, "hour": rh, "n": rn, "every_s": 3600 // rn if rn else 0}
    else:
        record = {"date": None, "hour": 0, "n": 0, "every_s": 0}
    return {
        "start": start, "end": end,
        "draws": len(sel), "same_day": sd, "not_same_day": nsd,
        "sd_pct": round(sd / len(sel) * 100, 1) if sel else 0.0,
        "box": box,
        "nsd_pct": round(nsd / len(sel) * 100, 1) if sel else 0.0,
        "open_now": open_now,
        "people": len({t["who_key"] for t in sel}),
        "companies": len({t["co"] for t in sel}),
        "days": len(days), "working_days": len(wdays),
        "per_working_day": round(len(sel) / len(wdays)) if wdays else 0,
        "median_issue": hm(imins[len(imins) // 2]) if imins else "-",
        "median_return": hm(rmins[len(rmins) // 2]) if rmins else "-",
        "pct_before_6": round(sum(1 for t in sel if t["st"].hour < 6) / len(sel) * 100) if sel else 0,
        "hour_issues": Counter(t["st"].hour for t in sel),
        "hour_returns": Counter(t["en"].hour for t in closed),
        # the shift rhythm: weekday (Mon=0) x hour grids for draws and returns
        "heat_draws": [[sum(1 for t in sel if t["st"].weekday() == wd and t["st"].hour == h)
                        for h in range(24)] for wd in range(7)],
        "heat_returns": [[sum(1 for t in closed if t["en"].weekday() == wd and t["en"].hour == h)
                          for h in range(24)] for wd in range(7)],
        "record_hour": record,
        "sel": sel,
    }


def _league(sel, now):
    """Per-person aggregates over a selection of crew draws."""
    ppl = defaultdict(lambda: {"draws": 0, "sd": 0, "nsd": 0, "open": 0,
                               "maxd": 0, "co": Counter(), "disp": Counter(),
                               "weeks": set(), "last_nsd": None})
    for t in sel:
        p = ppl[t["who_key"]]
        p["draws"] += 1
        p["co"][t["co"]] += 1
        p["disp"][t["who"]] += 1
        if t["sd"]:
            p["sd"] += 1
        else:
            p["nsd"] += 1
            iso = t["st"].isocalendar()
            p["weeks"].add((iso[0], iso[1]))
            if p["last_nsd"] is None or t["st"] > p["last_nsd"]:
                p["last_nsd"] = t["st"]
        if t["en"]:
            p["maxd"] = max(p["maxd"], (t["en"] - t["st"]).days)
        else:
            p["open"] += 1
            p["maxd"] = max(p["maxd"], (now - t["st"]).days)
    out = {}
    for k, p in ppl.items():
        out[k] = {
            "key": k, "name": p["disp"].most_common(1)[0][0],
            "co": p["co"].most_common(1)[0][0],
            "draws": p["draws"], "sd": p["sd"], "nsd": p["nsd"],
            "open": p["open"], "maxd": p["maxd"],
            "sd_pct": round(p["sd"] / p["draws"] * 100) if p["draws"] else 0,
            "weeks": len(p["weeks"]), "last_nsd": p["last_nsd"],
        }
    return out


def _company_league(sel, now):
    cos = defaultdict(lambda: {"draws": 0, "sd": 0, "nsd": 0, "open": 0,
                               "people": set(), "ppl": Counter()})
    for t in sel:
        c = cos[t["co"]]
        c["draws"] += 1
        c["people"].add(t["who_key"])
        if t["sd"]:
            c["sd"] += 1
        else:
            c["nsd"] += 1
            c["ppl"][t["who"]] += 1
        if not t["en"]:
            c["open"] += 1
    out = {}
    for k, c in cos.items():
        out[k] = {
            "name": k, "draws": c["draws"], "sd": c["sd"], "nsd": c["nsd"],
            "open": c["open"], "people": len(c["people"]),
            "sd_pct": round(c["sd"] / c["draws"] * 100, 1) if c["draws"] else 0.0,
            "nsd_pct": round(c["nsd"] / c["draws"] * 100, 1) if c["draws"] else 0.0,
            "top": c["ppl"].most_common(),
        }
    return out


def compute(ctx, rules=None):
    R = dict(RULES)
    if rules:
        R.update(rules)
    rs, tx = ctx["rs"], ctx["tx"]
    asat = ctx["asat"]
    today = asat.date()
    yday = today - timedelta(days=1)
    m = {"asat": asat, "today": today, "yesterday": yday, "rules": R,
         "tx_window": ctx["tx_window"], "tx_requested": ctx["tx_requested"],
         "sources": {"rental_stock": os.path.basename(ctx["rs_path"]),
                     "transactions": os.path.basename(ctx["tx_path"]),
                     "serials": os.path.basename(ctx["ser_path"]) if ctx["ser_path"] else ""}}

    # ------------------------------------------------------------------
    # 1. THE POSITION - from RENTAL_STOCK
    # ------------------------------------------------------------------
    on = [r for r in rs if r["on_hire"]]
    avail = [r for r in rs if not r["on_hire"]]
    m["fleet_total"] = len(rs)
    m["available"] = len(avail)
    m["on_hire_total"] = len(on)
    by_kind = Counter(r["kind"] for r in on)
    m["kind_counts"] = dict(by_kind)
    m["fccu"] = by_kind.get("fccu", 0)
    m["ops"] = by_kind.get("ops", 0)
    m["ff"] = by_kind.get("ff", 0)
    m["afterhours"] = by_kind.get("afterhours", 0)
    m["repairs"] = by_kind.get("repair", 0)
    m["crew_out"] = by_kind.get("crew", 0)
    m["custody_total"] = m["fccu"] + m["ops"] + m["ff"] + m["afterhours"]

    def age_days(r):
        return (today - r["on_dt"].date()).days if r["on_dt"] else 0

    crew_items = []
    for r in on:
        if r["kind"] != "crew":
            continue
        d = age_days(r)
        crew_items.append({
            "who": r["who"], "who_key": r["who_key"], "co": r["co"],
            "bc": r["bc"], "serial": r.get("serial", ""), "desc": r["desc"],
            "on_dt": r["on_dt"], "days": d,
            "bucket": ("today" if d < 1 else "1" if d < 2 else "2-7" if d < 8
                       else "8-29" if d < 30 else "30+"),
            "cost": R["charge_per_unit"],
        })
    crew_items.sort(key=lambda x: (-x["days"], x["co"], x["who"]))
    m["crew_items"] = crew_items
    m["out_today"] = sum(1 for x in crew_items if x["bucket"] == "today")
    m["out_1"] = sum(1 for x in crew_items if x["bucket"] == "1")
    m["out_2_7"] = sum(1 for x in crew_items if x["bucket"] == "2-7")
    m["out_1_7"] = m["out_1"] + m["out_2_7"]
    m["out_8_29"] = sum(1 for x in crew_items if x["bucket"] == "8-29")
    m["out_30"] = sum(1 for x in crew_items if x["bucket"] == "30+")
    m["outstanding"] = m["out_1_7"] + m["out_8_29"] + m["out_30"]
    m["outstanding_items"] = [x for x in crew_items if x["days"] >= 1]
    m["exposure"] = m["out_30"] * R["charge_per_unit"]
    m["exposure_all_outstanding"] = m["outstanding"] * R["charge_per_unit"]

    # custody lines (not people) with their ages
    cust = {}
    for kind, label in (("fccu", "FCCU turnaround custody"),
                        ("ops", "Ampol Operations"),
                        ("ff", "Future Fuels"),
                        ("afterhours", "After Hours Hire account")):
        items = [r for r in on if r["kind"] == kind]
        ages = [age_days(r) for r in items]
        cust[kind] = {"label": label, "n": len(items),
                      "min_days": min(ages) if ages else 0,
                      "max_days": max(ages) if ages else 0,
                      "hirers": Counter(r["who_raw"] for r in items).most_common(3)}
    m["custody"] = cust

    # repairs by status, with ages
    rep = []
    for r in on:
        if r["kind"] != "repair":
            continue
        rep.append({"status": repair_status(r["who_raw"]), "bc": r["bc"],
                    "serial": r.get("serial", ""), "desc": r["desc"],
                    "days": age_days(r), "on_dt": r["on_dt"]})
    rep.sort(key=lambda x: -x["days"])
    m["repair_items"] = rep
    m["repair_cats"] = Counter(x["status"] for x in rep).most_common()
    m["repair_top"] = m["repair_cats"][0][0] if m["repair_cats"] else "None"
    m["repair_stale"] = [x for x in rep if x["days"] >= R["stale_repair_days"]]
    m["repair_age_buckets"] = Counter(
        ("under 30 days" if x["days"] < 30 else "30-89 days" if x["days"] < 90
         else "90-179 days" if x["days"] < 180 else "180+ days") for x in rep)
    m["fleet_impact_pct"] = round(m["repairs"] / m["fleet_total"] * 100, 1) if m["fleet_total"] else 0.0
    m["usable_fleet"] = m["fleet_total"] - m["repairs"]

    # where the monitors are, by company (crew) - the "help Ampol" table
    byco = defaultdict(lambda: {"today": 0, "1": 0, "2-7": 0, "8-29": 0, "30+": 0,
                                "total": 0, "people": set(), "names": Counter()})
    for x in crew_items:
        c = byco[x["co"]]
        c[x["bucket"]] += 1
        c["total"] += 1
        c["people"].add(x["who_key"])
        if x["days"] >= 1:
            c["names"][x["who"]] += 1
    where = []
    for k, c in byco.items():
        outst = c["1"] + c["2-7"] + c["8-29"] + c["30+"]
        where.append({"name": k, "today": c["today"], "d1": c["1"], "d2_7": c["2-7"],
                      "d8_29": c["8-29"], "d30": c["30+"], "total": c["total"],
                      "outstanding": outst, "people": len(c["people"]),
                      "exposure": c["30+"] * R["charge_per_unit"],
                      "names": c["names"].most_common(),
                      "urgency": ("Critical" if c["30+"] else "High"
                                  if (c["8-29"] or outst >= 3) else "Watch"
                                  if outst else "Clear")})
    where.sort(key=lambda w: (-w["d30"], -w["outstanding"], -w["total"], w["name"]))
    m["where"] = where
    m["priorities"] = [w for w in where if w["outstanding"]]
    m["focus3"] = m["priorities"][:3]

    # ------------------------------------------------------------------
    # 2. THE FLOW - from TRANSACTIONS
    # ------------------------------------------------------------------
    win_start, win_end = ctx["tx_window"]
    now = ctx["tx_requested"]
    if win_start is None:
        win_start = min(t["st"] for t in tx) if tx else asat
    if win_end is None:
        win_end = max(t["st"] for t in tx) if tx else asat
    crew = [t for t in tx if t["kind"] == "crew"]
    m["tx_all"] = len(tx)
    m["tx_crew"] = len(crew)
    m["tx_accounts"] = len(tx) - len(crew)
    m["tx_account_kinds"] = dict(Counter(t["kind"] for t in tx if t["kind"] != "crew"))

    # Windows are whole calendar days: the last 30 days run from midnight
    # 30 days before the report day up to where the export closes.
    # "Last 3 months" = the 13 full weeks before this one, plus this week so
    # far - Monday-aligned so the week-by-week chart sums exactly to the
    # window total (a reader adding up the bars must land on the headline).
    this_monday = today - timedelta(days=today.weekday())
    d90_start = datetime.combine(this_monday - timedelta(weeks=13), dtime(0, 0))
    d30_start = datetime.combine(today - timedelta(days=30), dtime(0, 0))
    d7_start = datetime.combine(today - timedelta(days=7), dtime(0, 0))
    ytd = _flow(crew, win_start, win_end, now)
    d90 = _flow(crew, d90_start, win_end, now)
    d30 = _flow(crew, d30_start, win_end, now)
    d7 = _flow(crew, d7_start, win_end, now)
    ytd_label = f'{win_start.strftime("%d %b")} - {win_end.strftime("%d %b %Y")}'
    d90_label = f'{d90_start.strftime("%d %b")} - {win_end.strftime("%d %b %Y")}'
    d30_label = f'{d30_start.strftime("%d %b")} - {win_end.strftime("%d %b %Y")}'
    m["ytd"] = ytd
    m["d90"] = d90
    m["d30"] = d30
    m["d7"] = d7
    m["ytd_label"] = ytd_label
    m["d90_label"] = d90_label
    m["d30_label"] = d30_label
    m["tx_window_complete_to"] = win_end

    # yesterday - the last complete day
    yd_sel = [t for t in crew if t["st"].date() == yday]
    yd_nsd = [t for t in yd_sel if not t["sd"]]
    yd_rec = [t for t in yd_nsd if t["en"]]
    yd_open = [t for t in yd_nsd if not t["en"]]
    m["yday_draws"] = len(yd_sel)
    m["yday_people"] = len({t["who_key"] for t in yd_sel})
    m["yday_companies"] = len({t["co"] for t in yd_sel})
    m["yday_nsd"] = len(yd_nsd)
    m["yday_recovered"] = len(yd_rec)
    m["yday_still_out"] = len(yd_open)
    m["yday_recovery_pct"] = (round(len(yd_rec) / len(yd_nsd) * 100)
                              if yd_nsd else 100)
    m["yday_sd_pct"] = round((len(yd_sel) - len(yd_nsd)) / len(yd_sel) * 100, 1) if yd_sel else 0.0
    yco = defaultdict(lambda: {"nsd": 0, "rec": 0, "open": 0, "ppl": Counter(),
                               "ppl_open": Counter()})
    for t in yd_nsd:
        c = yco[t["co"]]
        c["nsd"] += 1
        c["ppl"][t["who"]] += 1
        if t["en"]:
            c["rec"] += 1
        else:
            c["open"] += 1
            c["ppl_open"][t["who"]] += 1
    m["yday_by_company"] = sorted(
        [{"name": k, "nsd": c["nsd"], "recovered": c["rec"], "open": c["open"],
          "people": c["ppl"].most_common(), "people_open": c["ppl_open"].most_common()}
         for k, c in yco.items()],
        key=lambda c: (-c["open"], -c["nsd"], c["name"]))
    # cross-check: the register should show the same names still out since yesterday
    m["yday_register_still_out"] = m["out_1"]

    # today so far (the export closes at the window end - say so)
    td_sel = [t for t in crew if t["st"].date() == today]
    m["today_draws_in_export"] = len(td_sel)
    m["today_export_closes"] = win_end

    # leagues - one entry per person who drew this year, with the last
    # 30 days, the last 3 months and the year side by side
    lg_ytd = _league(ytd["sel"], now)
    lg_90 = _league(d90["sel"], now)
    lg_30 = _league(d30["sel"], now)
    lg_7 = _league(d7["sel"], now)
    open_by_person = Counter(x["who_key"] for x in crew_items if x["days"] >= 1)
    yday_by_person = Counter(t["who_key"] for t in yd_open)
    Z = {"draws": 0, "nsd": 0, "sd_pct": 0, "maxd": 0, "weeks": 0, "last_nsd": None}
    everyone = []
    for k, py in lg_ytd.items():
        p90 = lg_90.get(k, Z)
        p30 = lg_30.get(k, Z)
        p7 = lg_7.get(k, Z)
        co = (p30 if p30 is not Z else p90 if p90 is not Z else py)["co"]
        everyone.append({
            "key": k, "name": py["name"], "co": co,
            "d30_draws": p30["draws"], "d30_nsd": p30["nsd"],
            "d30_sd_pct": p30["sd_pct"], "d30_weeks": p30["weeks"],
            "d90_draws": p90["draws"], "d90_nsd": p90["nsd"], "d90_sd_pct": p90["sd_pct"],
            "d7_nsd": p7["nsd"], "yday_open": yday_by_person.get(k, 0),
            "open_now": open_by_person.get(k, 0),
            "ytd_draws": py["draws"], "ytd_nsd": py["nsd"], "ytd_sd_pct": py["sd_pct"],
            "maxd": max(py["maxd"], p30["maxd"]),
            "repeat": p30["weeks"] >= R["repeat_weeks"],
            "last_nsd": py["last_nsd"],
        })
    m["league"] = sorted([x for x in everyone if x["d30_draws"]],
                         key=lambda x: (-x["d30_nsd"], -x["open_now"], -x["ytd_nsd"]))
    m["league_90"] = sorted([x for x in everyone if x["d90_draws"]],
                            key=lambda x: (-x["d90_nsd"], -x["d30_nsd"], -x["ytd_nsd"]))
    m["league_ytd"] = sorted(everyone, key=lambda x: (-x["ytd_nsd"], -x["d90_nsd"], -x["d30_nsd"]))
    m["repeat_offenders"] = [x for x in m["league"] if x["repeat"]]
    m["people_with_nsd_30"] = sum(1 for x in m["league"] if x["d30_nsd"])
    m["people_active_30"] = len(m["league"])
    m["people_with_nsd_90"] = sum(1 for x in everyone if x["d90_nsd"])
    m["people_active_90"] = sum(1 for x in everyone if x["d90_draws"])
    m["people_with_nsd_ytd"] = sum(1 for x in everyone if x["ytd_nsd"])
    m["people_active_ytd"] = len(everyone)

    # company leagues - the same three windows side by side
    co30 = _company_league(d30["sel"], now)
    co90 = _company_league(d90["sel"], now)
    coytd = _company_league(ytd["sel"], now)
    open_by_co = Counter(x["co"] for x in crew_items if x["days"] >= 1)
    ZC = {"draws": 0, "sd": 0, "nsd": 0, "open": 0, "people": 0,
          "sd_pct": 0.0, "nsd_pct": 0.0, "top": []}
    comp = []
    for k, cy in coytd.items():
        c = co30.get(k, ZC)
        c9 = co90.get(k, ZC)
        comp.append({
            "name": k,
            "d30_draws": c["draws"], "d30_nsd": c["nsd"], "d30_nsd_pct": c["nsd_pct"],
            "d30_sd_pct": c["sd_pct"], "d30_people": c["people"], "d30_top": c["top"],
            "d90_draws": c9["draws"], "d90_nsd": c9["nsd"], "d90_nsd_pct": c9["nsd_pct"],
            "d90_sd_pct": c9["sd_pct"], "d90_people": c9["people"], "d90_top": c9["top"],
            "ytd_draws": cy["draws"], "ytd_nsd": cy["nsd"], "ytd_nsd_pct": cy["nsd_pct"],
            "ytd_sd_pct": cy["sd_pct"], "ytd_people": cy["people"], "ytd_top": cy["top"],
            "open_now": open_by_co.get(k, 0),
        })
    m["companies"] = sorted(comp, key=lambda x: (-x["d30_nsd"], -x["d30_draws"]))
    m["companies_90"] = sorted(comp, key=lambda x: (-x["d90_nsd"], -x["d90_draws"]))
    m["companies_ytd"] = sorted(comp, key=lambda x: (-x["ytd_nsd"], -x["ytd_draws"]))

    # ------------------------------------------------------------------
    # 3. TRENDS
    # ------------------------------------------------------------------
    monthly = defaultdict(lambda: {"draws": 0, "sd": 0, "people": set()})
    for t in ytd["sel"]:
        k = t["st"].strftime("%Y-%m")
        monthly[k]["draws"] += 1
        monthly[k]["people"].add(t["who_key"])
        if t["sd"]:
            monthly[k]["sd"] += 1
    m["monthly"] = []
    for k in sorted(monthly):
        v = monthly[k]
        m["monthly"].append({
            "key": k, "label": datetime.strptime(k, "%Y-%m").strftime("%b"),
            "draws": v["draws"], "sd": v["sd"], "nsd": v["draws"] - v["sd"],
            "sd_pct": round(v["sd"] / v["draws"] * 100, 1) if v["draws"] else 0,
            "people": len(v["people"]),
            "partial": k == win_end.strftime("%Y-%m") and win_end.day < 28,
        })

    weekly = defaultdict(lambda: [0, 0])
    for t in ytd["sel"]:
        iso = t["st"].isocalendar()
        weekly[(iso[0], iso[1])][0] += 1
        if t["sd"]:
            weekly[(iso[0], iso[1])][1] += 1
    m["weekly"] = []
    for k in sorted(weekly):
        n, s = weekly[k]
        mon = datetime.fromisocalendar(k[0], k[1], 1)
        m["weekly"].append({"label": mon.strftime("%d %b"), "n": n, "sd": s, "nsd": n - s,
                            "draws": n,
                            "pct": round(s / n * 100, 1) if n else 0,
                            "in_30": mon.date() >= d30_start.date(),
                            "in_90": mon.date() >= d90_start.date(),
                            "partial": k == (win_end.isocalendar()[0], win_end.isocalendar()[1]),
                            "weekend": False})
    cur_iso = win_end.isocalendar()
    m["current_week_partial"] = (cur_iso[0], cur_iso[1]) in weekly
    m["weekly90"] = [w for w in m["weekly"] if w["in_90"]]

    daily = defaultdict(lambda: [0, 0])
    for t in d30["sel"]:
        daily[t["st"].date()][0] += 1
        if not t["sd"]:
            daily[t["st"].date()][1] += 1
    # every calendar day in the window, zeros included - a quiet weekend is
    # part of the picture, not a gap in it
    m["daily30"] = []
    d = d30_start.date()
    while d <= win_end.date():
        v = daily.get(d, [0, 0])
        m["daily30"].append({"date": d, "label": d.strftime("%d %b"), "draws": v[0],
                             "nsd": v[1], "weekend": d.weekday() >= 5,
                             "partial": d == win_end.date()})
        d += timedelta(days=1)
    if m["daily30"]:
        m["daily30_busiest"] = max(m["daily30"], key=lambda x: x["draws"])
        m["daily30_worst_nsd"] = max(m["daily30"], key=lambda x: x["nsd"])
    else:
        m["daily30_busiest"] = m["daily30_worst_nsd"] = None

    # concurrency - every open transaction incl. custody (a unit in FCCU custody
    # is still a unit off the shelf)
    events = []
    for t in tx:
        events.append((t["st"], 1))
        events.append((t["en"] if t["en"] else now, -1))
    events.sort(key=lambda e: (e[0], -e[1]))
    cur, day_peak = 0, {}
    for ts, d in events:
        cur += d
        dt_ = ts.date()
        if dt_ not in day_peak or cur > day_peak[dt_][0]:
            day_peak[dt_] = (cur, ts)
    m["day_peaks"] = [{"date": d, "peak": day_peak[d][0], "at": day_peak[d][1]}
                      for d in sorted(day_peak)]
    m["record_peak"] = max(m["day_peaks"], key=lambda p: p["peak"]) if m["day_peaks"] else None
    p30 = [p for p in m["day_peaks"] if p["date"] >= d30_start.date()]
    m["record_peak_30"] = max(p30, key=lambda p: p["peak"]) if p30 else None

    # intraday net-draw curve, last 10 complete working days
    wdays = [d for d in sorted({t["st"].date() for t in tx})
             if d.weekday() < 5 and d < win_end.date()][-10:]
    half_hours = [x / 2 for x in range(6, 41)]     # 03:00 .. 20:00
    curves = defaultdict(list)
    for d in wdays:
        ev = []
        for t in tx:
            if t["st"].date() == d:
                ev.append((t["st"].hour + t["st"].minute / 60, 1))
            if t["en"] and t["en"].date() == d:
                ev.append((t["en"].hour + t["en"].minute / 60, -1))
        ev.sort()
        i, cur = 0, 0
        for hh2 in half_hours:
            while i < len(ev) and ev[i][0] <= hh2:
                cur += ev[i][1]
                i += 1
            curves[hh2].append(cur)
    m["curve_days"] = wdays
    m["net_curve"] = [(h, round(sum(v) / len(v), 1)) for h, v in sorted(curves.items())] if wdays else []
    m["net_curve_worst"] = [(h, max(v)) for h, v in sorted(curves.items())] if wdays else []
    m["net_plateau"] = max((v for _, v in m["net_curve"]), default=0)
    m["net_plateau_worst"] = max((v for _, v in m["net_curve_worst"]), default=0)

    # duration buckets, 30 days and YTD (closed crew draws)
    def buckets(sel):
        b = Counter()
        for t in sel:
            if not t["en"]:
                continue
            h = (t["en"] - t["st"]).total_seconds() / 3600
            b["Same shift (under 12h)" if h < 12 else "Overnight (12-24h)" if h < 24
              else "1-3 days" if h < 72 else "3-7 days" if h < 168 else "Over a week"] += 1
        order = ["Same shift (under 12h)", "Overnight (12-24h)", "1-3 days",
                 "3-7 days", "Over a week"]
        return [(k, b.get(k, 0)) for k in order]
    m["dur_buckets_30"] = buckets(d30["sel"])
    m["dur_buckets_ytd"] = buckets(ytd["sel"])
    closed_ytd = [t for t in ytd["sel"] if t["en"]]
    m["longest_kept"] = max(closed_ytd, key=lambda t: t["en"] - t["st"]) if closed_ytd else None

    # ------------------------------------------------------------------
    # 4. HEALTH - four plain scores, formulas printed on the page
    # ------------------------------------------------------------------
    m["score_availability"] = min(100, round(m["available"] / R["availability_target"] * 100)) if R["availability_target"] else 100
    m["score_sameday"] = round(d30["sd_pct"])
    m["score_repairs"] = max(0, 100 - round(m["repairs"] / m["fleet_total"] * 100)) if m["fleet_total"] else 100
    m["score_30"] = max(0, 100 - 3 * m["out_30"])
    m["health"] = round((m["score_availability"] + m["score_sameday"]
                         + m["score_repairs"] + m["score_30"]) / 4)

    # ------------------------------------------------------------------
    # 5. RECONCILIATION - printed on the data page, never hidden
    # ------------------------------------------------------------------
    notes = []
    open_tx_bc = {t["bc"] for t in tx if not t["en"]}
    rs_on_bc = {r["bc"] for r in on}
    both = len(open_tx_bc & rs_on_bc)
    only_rs = rs_on_bc - open_tx_bc
    only_tx = open_tx_bc - rs_on_bc
    notes.append(f"RENTAL_STOCK shows {len(on)} monitors on hire; TRANSACTIONS holds "
                 f"{len(open_tx_bc)} open transactions; {both} match by barcode.")
    if only_rs:
        kinds = Counter(r["kind"] for r in on if r["bc"] in only_rs)
        pre = sum(1 for r in on if r["bc"] in only_rs and r["on_dt"] and r["on_dt"] < win_start)
        late = sum(1 for r in on if r["bc"] in only_rs and r["on_dt"] and r["on_dt"] > win_end)
        notes.append(f"{len(only_rs)} on-hire rows have no open transaction: {pre} were issued "
                     f"before the transactions window opened ({win_start.strftime('%d %b %Y')}) and "
                     f"{late} after it closed ({win_end.strftime('%d %b %Y %H:%M')}) - "
                     + ", ".join(f"{cust.get(k, {}).get('label', k) if k != 'crew' else 'crew'} {v}"
                                 for k, v in kinds.most_common()) + ".")
    if only_tx:
        notes.append(f"{len(only_tx)} open transactions have no on-hire row in RENTAL_STOCK "
                     f"- returned between the two pulls, or a barcode change.")
    if m["yday_still_out"] != m["out_1"]:
        notes.append(f"Yesterday's unreturned draws still open in TRANSACTIONS: "
                     f"{m['yday_still_out']}; RENTAL_STOCK rows on hire since yesterday: "
                     f"{m['out_1']}. The difference is gear issued yesterday after the "
                     f"transactions window closed, or returned between the two pulls.")
    else:
        notes.append(f"Yesterday's still-out count agrees between the two exports: "
                     f"{m['yday_still_out']}.")
    ser = ctx["serials"]
    fleet_bcs = [r["bc"].upper() for r in rs]
    missing = sorted(b for b in fleet_bcs if not ser.get(b))
    by_serial = defaultdict(list)
    for b, s_ in ser.items():
        if s_:
            by_serial[s_].append(b)
    dups = {s_: bcs for s_, bcs in by_serial.items()
            if len(bcs) > 1 and any(b in set(fleet_bcs) for b in bcs)}
    m["serial_stats"] = {"fleet": len(fleet_bcs), "with": len(fleet_bcs) - len(missing),
                         "missing": missing, "dups": dups, "list_size": len(ser)}
    if ser:
        pre = Counter(b.split("/")[0] for b in missing).most_common(1)
        notes.append(f"Serial list: {len(fleet_bcs) - len(missing)} of the {len(fleet_bcs)} monitors on the "
                     f"register carry a serial on it; {len(missing)} do not"
                     + (f" ({pre[0][1]} of them in the {pre[0][0]} range)" if pre else "")
                     + ". Barcode is the identity SiteIQ scans; the serial is display only."
                     + (f" {len(dups)} serial(s) appear against two barcodes ("
                        + "; ".join(f"{s_} = {', '.join(bcs)}" for s_, bcs in list(dups.items())[:3])
                        + ") - usually a re-labelled unit whose old barcode was never retired from the list."
                        if dups else ""))
    tx_bcs = {t["bc"].upper() for t in tx}
    departed = tx_bcs - set(fleet_bcs)
    dep_rows = sum(1 for t in tx if t["bc"].upper() in departed)
    m["departed"] = {"n": len(departed), "rows": dep_rows}
    if departed:
        notes.append(f"{len(departed)} barcodes moved this year but are no longer on the register "
                     f"(returned to Dräger, re-labelled or written off): their {dep_rows:,} draws are counted "
                     f"in the flow figures, and they are not in the position.")
    m["recon_notes"] = notes
    return m


# =====================================================================
# V18 SHIM - the executive dashboard keeps its shape, the numbers move
# onto this engine so the PDF, the email and the dashboard agree.
# =====================================================================

def v18_data(m):
    """The `data` dict generate_v18_gas_monitor_report.load_data() used to
    read off the workbook, rebuilt from the engine's figures."""
    R = m["rules"]

    def row(x, with_status=False):
        d = x["on_dt"]
        base = [x["co"], x["who"], x["bc"], x.get("serial", ""), x["desc"], x["days"]]
        if with_status:
            base.append("On Hire")
        base += [R["charge_per_unit"], d.date() if d else None, d.time() if d else None]
        return base

    items = m["crew_items"]
    data = {
        "available": [],   # V18 only counts these; the list itself is not rendered
        "oh_1_7": [row(x) for x in items if 1 <= x["days"] <= 7],
        "oh_8_29": [row(x) for x in items if 8 <= x["days"] <= 29],
        "oh_30": [row(x, True) for x in items if x["days"] >= 30],
        "fccu": [],
        "repairs": [[f"Dräger – {x['status']}", x["bc"], x["serial"], x["desc"],
                     x["days"], x["on_dt"]] for x in m["repair_items"]],
        "perf_summary": {
            "Current Onhire": m["crew_out"],
            "Available": m["available"],
            "Out Of Service": m["repairs"],
            "Onhire To Operations": m["ops"],
            "Onhire To FCCU": m["fccu"],
            "Onhire To Future Fuels": m["ff"],
            "Onhire To After Hours": m["afterhours"],
        },
        "companies": [],
        "recovered": m["yday_recovered"],
        "issued_today": m["out_today"],
        "available_count": m["available"],
        "fccu_count": m["fccu"],
    }
    data["available"] = [["", "", "", "Available for Hire", None]] * m["available"]
    data["fccu"] = [["Dräger", "Dräger - FCCU", "", "", "", 0, R["charge_per_unit"], None, None]] * m["fccu"]

    yday = {c["name"]: c for c in m["yday_by_company"]}
    where = {w["name"]: w for w in m["where"]}
    names = {c["name"] for c in m["companies"]} | set(yday) | set(where)
    for n in sorted(names):
        c = next((x for x in m["companies"] if x["name"] == n), None)
        y = yday.get(n)
        w = where.get(n)
        top = ", ".join(f"{p} ({k})" for p, k in (c["d30_top"][:10] if c else []))
        flag = ("High Reoffender - Supervisor Follow Up" if c and c["d30_nsd"] >= 40
                else "Repeat Same Day Offender - Review Behaviour" if c and c["d30_nsd"] >= 10
                else "Watchlist - Repeat Offender" if c and c["d30_nsd"] >= 3
                else "No Reoffender Action")
        data["companies"].append({
            "name": n,
            "issued": c["d30_draws"] if c else 0,
            "returned": (c["d30_draws"] - c["open_now"]) if c else 0,
            "nrsd": c["d30_nsd"] if c else 0,
            "nrsd_pct": f'{c["d30_nsd_pct"]}%' if c else "0.0%",
            "top_offenders": top,
            "reoffender_action": flag,
            # V18 reads prev_day as yesterday's non-return TOTAL and subtracts
            # what was recovered to get "still out" - so hand it the total.
            "prev_day": y["nsd"] if y else 0,
            "prev_day_pct": "",
            "prev_day_names": ", ".join(f"{p} ({k})" for p, k in y["people"]) if y else "",
            "still_out": w["outstanding"] if w else 0,
            "still_out_pct": "",
            "still_out_names": ", ".join(f"{p} ({k})" for p, k in w["names"]) if w else "",
            "action": ("Charge / Recovery Action" if w and w["d30"] else
                       "Follow Up" if w and w["outstanding"] else "No Action"),
            "d1_7": (w["d1"] + w["d2_7"]) if w else 0,
            "d8_29": w["d8_29"] if w else 0,
            "d30": w["d30"] if w else 0,
            "charge": w["exposure"] if w else 0,
            "today": w["today"] if w else 0,
        })
    return data


# =====================================================================
# console summary - the same lines every button prints
# =====================================================================

def print_summary(m):
    print(f"Position as at       : {m['asat'].strftime('%d %b %Y %H:%M')}  (RENTAL_STOCK pull)")
    ws, we = m["tx_window"]
    if ws and we:
        print(f"Transactions window  : {ws.strftime('%d %b %Y %H:%M')} -> {we.strftime('%d %b %Y %H:%M')}")
    print(f"Fleet                : {m['fleet_total']}  = {m['available']} available + "
          f"{m['crew_out']} out to crew + {m['custody_total']} custody + {m['repairs']} repair")
    print(f"Out to crew          : {m['out_today']} today, {m['out_1']} since yesterday, "
          f"{m['out_2_7']} 2-7d, {m['out_8_29']} 8-29d, {m['out_30']} 30+d  "
          f"-> {m['outstanding']} outstanding, exposure ${m['exposure']:,}")
    print(f"Yesterday ({m['yesterday'].strftime('%d %b')})  : {m['yday_draws']} crew draws, "
          f"{m['yday_nsd']} not back same day, {m['yday_recovered']} recovered, "
          f"{m['yday_still_out']} still out")
    print(f"Last 30 days         : {m['d30']['draws']:,} draws, {m['d30']['sd_pct']}% same day, "
          f"{m['d30']['not_same_day']:,} not, {m['people_with_nsd_30']} people with a non-return")
    print(f"Last 3 months        : {m['d90']['draws']:,} draws, {m['d90']['sd_pct']}% same day, "
          f"{m['d90']['not_same_day']:,} not, {m['people_with_nsd_90']} people with a non-return")
    print(f"Year to date         : {m['ytd']['draws']:,} draws, {m['ytd']['sd_pct']}% same day, "
          f"{m['ytd']['not_same_day']:,} not, {m['people_with_nsd_ytd']} people with a non-return")
    print(f"Health score         : {m['health']}/100  (A{m['score_availability']} "
          f"S{m['score_sameday']} R{m['score_repairs']} C{m['score_30']})")
    for n in m["recon_notes"]:
        print(f"[RECON] {n}")


if __name__ == "__main__":
    ctx = load()
    mm = compute(ctx)
    print_summary(mm)
    print("")
    print("Top 10, last 30 days - not returned same day:")
    for x in mm["league"][:10]:
        print(f"  {x['name']:26s} {x['co']:22s} 30d {x['d30_nsd']:3d}/{x['d30_draws']:<4d} "
              f"open {x['open_now']:2d}  YTD {x['ytd_nsd']:4d}/{x['ytd_draws']:<5d} "
              f"{'REPEAT' if x['repeat'] else ''}")
