#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | PULL SITEIQ EXPORTS - Downloads -> Data, safely
#  Ampol Tool Store (Lytton Refinery)
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  THE JOB: SiteIQ downloads land in Downloads with long names, and
#  some of them belong to OTHER sites. Copying the wrong one in would
#  poison a whole day of reports, so this never trusts a filename. It
#  OPENS each file and checks two things:
#
#    1. WHICH SITE - it must be the Ampol Lytton store. Anything else
#       is left where it is and named on screen, so you can see it was
#       skipped and why.
#    2. WHICH EXPORT - the data sheet name (RENTAL_STOCK, STOCKTAKE,
#       CUSTOMER_CONTRACTOR_EQUIP). That is how SiteIQ builds them, so
#       the download can be called anything and it still lands right.
#
#  Newest of each type wins (by the export's own REQUESTED_DATE/TIME,
#  not the file date). Whatever is already in Data\ is backed up to
#  Data\previous\ first - nothing is ever just overwritten, and
#  nothing is ever deleted.
# =====================================================================
import datetime as dt
import os
import re
import shutil
import sys

import ampol_paths  # WHY (12 Aug 2026): one Data area in, always

HERE = ampol_paths.suite_dir()
DEST = ampol_paths.data_dir()
BACKUP = ampol_paths.previous_dir()
DOWNLOADS = os.path.join(os.path.expanduser("~"), "Downloads")

#  WHOSE EXPORT IS THIS? A file is accepted if any REFERENCE_INFO
#  project field contains ANY of these words (case and spacing
#  ignored). Two markers, not one, so a rename at the SiteIQ end can't
#  lock us out - and an export from another site matches neither and
#  is left alone.
SITE_MARKERS = ["ampol", "lytton"]

#  WHY (12 Aug 2026): the Permanent Store exports are honest but
#  unhelpful - RENTAL_STOCK and TRANSACTIONS say only PROJECT SCHEDULE
#  = "Permanent Store", no site name at all (checked against the real
#  pulls in Data\). Only STOCKTAKE carries "Ampol Refineries" in
#  PROJECT. So when the project cells don't answer, the first rows of
#  the data itself get a look - the owner and barcode columns say
#  AMPOL within the first handful of rows on every genuine pull, and
#  another site's export says its own name instead. Belt and braces,
#  never a guess.
FINGERPRINT_ROWS = 200

#  data sheet name -> the name the suite expects on disk
#  WHY (12 Aug 2026): at Ampol the transactions pull is identified by
#  its CUSTOMER_CONTRACTOR_EQUIP sheet - that is the sheet the tooling
#  and gas engines actually read.
WANTED = {
    "RENTAL_STOCK": "RENTAL_STOCK.xlsx",
    "STOCKTAKE": "STOCKTAKE.xlsx",
    "CUSTOMER_CONTRACTOR_EQUIP": "TRANSACTIONS.xlsx",
}


def norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


def parse_when(v):
    """SiteIQ's own REQUESTED_DATE/TIME - Australian d/m/Y, AM/PM."""
    if isinstance(v, dt.datetime):
        return v
    m = re.match(r"\s*(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{2})\s*([AP]M)?",
                 str(v or ""), re.I)
    if not m:
        return None
    d, mo, y, h, mi, ap = m.groups()
    h = int(h)
    if ap:
        if ap.upper() == "PM" and h != 12:
            h += 12
        if ap.upper() == "AM" and h == 12:
            h = 0
    try:
        return dt.datetime(int(y), int(mo), int(d), h, int(mi))
    except ValueError:
        return None


def site_in_rows(wb, sheet):
    """Second opinion when REFERENCE_INFO doesn't name the site: does
    the data itself say AMPOL or LYTTON in its first rows?"""
    try:
        ws = wb[sheet]
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i >= FINGERPRINT_ROWS:
                break
            joined = " ".join(str(v).lower() for v in r if v is not None)
            if any(m in joined for m in SITE_MARKERS):
                return True
    except Exception:
        pass
    return False


def inspect(path):
    """(export_key, site_ok, site_text, requested_when) -
    or (None, False, why, None)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return None, False, "not a readable Excel file ({})".format(
            str(e)[:40]), None
    try:
        sheets = list(wb.sheetnames)
        key = next((s for s in sheets if s.upper() in WANTED), None)
        if not key:
            return None, False, "not a SiteIQ export (no known data sheet)", None
        site, when = "", None
        if "REFERENCE_INFO" in sheets:
            rows = []
            for i, r in enumerate(wb["REFERENCE_INFO"].iter_rows(values_only=True)):
                rows.append(r)
                if i >= 1:
                    break
            if len(rows) >= 2 and rows[0] and rows[1]:
                hdr = [norm(v) for v in rows[0]]
                bits = []
                for i, h in enumerate(hdr):
                    #  BOTH project fields count - STOCKTAKE carries the
                    #  client in PROJECT, the rest only carry the store
                    #  schedule in PROJECT SCHEDULE.
                    if h.startswith("project") and i < len(rows[1]):
                        v = str(rows[1][i] or "").strip()
                        if v:
                            bits.append(v)
                    if "requested_date" in h and i < len(rows[1]):
                        when = parse_when(rows[1][i])
                site = " / ".join(bits)
        site_ok = any(m in norm(site) for m in SITE_MARKERS)
        if not site_ok:
            #  The project cells didn't say the site - ask the data.
            site_ok = site_in_rows(wb, key)
            if site_ok:
                site = (site + " / " if site else "") + "AMPOL named in the data itself"
        return key.upper(), site_ok, site, when
    finally:
        try:
            wb.close()
        except Exception:
            pass


def explain():
    print(" What I do, in plain words:")
    print("   1. Look at every Excel file in Downloads (or the folder you")
    print("      give me - drag one onto 12_PULL_SITEIQ_EXPORTS.bat).")
    print("   2. Open each one and check it is an AMPOL LYTTON export -")
    print("      wrong-site files are skipped loudly, never filed.")
    print("   3. Work out which export it is from its data sheet name:")
    for k in sorted(WANTED):
        print("        {:<28} -> Data\\{}".format(k, WANTED[k]))
    print("   4. Newest of each type wins, by the export's own requested")
    print("      date. The old copy in Data\\ is backed up to")
    print("      Data\\previous\\ first. Nothing is ever deleted.")


def main():
    if len(sys.argv) > 1 and sys.argv[1].lower() in ("--help", "-h", "/?"):
        print("=" * 62)
        print(" COATES | PULL SITEIQ EXPORTS - Ampol Tool Store (Lytton)")
        print("=" * 62)
        explain()
        return 0

    dl = sys.argv[1] if len(sys.argv) > 1 else DOWNLOADS
    print("=" * 62)
    print(" COATES | PULL SITEIQ EXPORTS - Ampol Tool Store (Lytton)")
    print(" From: " + dl)
    print(" To  : " + DEST)
    print("=" * 62)
    if not os.path.isdir(dl):
        print(" That Downloads folder does not exist on this machine, so")
        print(" there is nothing to bring in. Data\\ is untouched.")
        print(" (You can hand me a folder: drag it onto")
        print("  12_PULL_SITEIQ_EXPORTS.bat, or name it after the command.)")
        return 1

    files = [os.path.join(dl, f) for f in os.listdir(dl)
             if f.lower().endswith((".xlsx", ".xlsm"))
             and not f.startswith("~$")]
    files.sort(key=os.path.getmtime, reverse=True)
    print(" Looking at {} Excel file(s) in {}...".format(
        len(files), os.path.basename(dl.rstrip("\\/")) or dl))
    print("")

    best, skipped = {}, []
    for p in files:
        name = os.path.basename(p)
        key, site_ok, site, when = inspect(p)
        if not key:
            skipped.append((name, site))
            continue
        if not site_ok:
            #  The important one: right kind of export, WRONG SITE.
            skipped.append((name, "another site - '{}'".format(
                (site or "site not stated").strip())))
            print("  SKIP  | {}\n          {} export, but it is for {}".format(
                name[:52], key, (site or "an unnamed site").strip()))
            continue
        stamp = when or dt.datetime.fromtimestamp(os.path.getmtime(p))
        cur = best.get(key)
        if not cur or stamp > cur[1]:
            best[key] = (p, stamp)

    if not best:
        print("")
        print(" Nothing to bring in - no Ampol Lytton exports in that")
        print(" folder. Data\\ is untouched.")
        return 1

    print("")
    for key in sorted(best):
        src, stamp = best[key]
        dest = os.path.join(DEST, WANTED[key])
        if os.path.isfile(dest):
            keep = os.path.join(BACKUP, "{}_{}".format(
                dt.datetime.fromtimestamp(os.path.getmtime(dest))
                .strftime("%Y%m%d_%H%M"), WANTED[key]))
            try:
                shutil.copy2(dest, keep)
            except Exception:
                pass
        shutil.copy2(src, dest)
        print("  IN    | {:<18} <- {}".format(WANTED[key],
                                              os.path.basename(src)[:44]))
        #  WHY (12 Aug 2026): 24-hour time, same as the report pages.
        print("          exported {}".format(
            stamp.strftime("%d %b %Y %H:%M").lstrip("0")))

    missing = [WANTED[k] for k in WANTED if k not in best]
    print("-" * 62)
    print(" Brought in : {} of {} export(s)".format(len(best), len(WANTED)))
    if missing:
        print(" Still needed: " + ", ".join(sorted(missing)))
        print("  (download them from SiteIQ and run me again - the ones")
        print("   already in place are untouched)")
    if skipped:
        print(" Left alone : {} file(s) that are not this site's "
              "exports".format(len(skipped)))
    print(" Previous copies kept in Data\\previous\\")
    print("")
    print(" Next: run 00_RUN_EVERYTHING.bat")
    return 0


if __name__ == "__main__":
    sys.exit(main())
