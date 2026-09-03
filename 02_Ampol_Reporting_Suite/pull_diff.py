"""COATES | AMPOL TOOL STORE - what changed since the last pull.

Author: Andrew Fisher | POWERED BY SITEIQ

WHY (03 Sep 2026): a snapshot says where everything is; a client also
wants to know what MOVED. Two honest sources, both already in Data\\:

  1. Pull against pull - Data\\previous\\ keeps every earlier RENTAL_STOCK
     export (button 28 parks the old one as YYYYMMDD_HHMM_RENTAL_STOCK.xlsx
     before saving the new one). Compare the newest earlier pull with the
     current one, item by item: came back, went out, changed hands, and
     crossed 30 / 60 / 90 days while still out. Companies cleared, new.
  2. The 24 hours before the pull - TRANSACTIONS (CUSTOMER_CONTRACTOR_EQUIP)
     holds every issue and return with a time, so the last day's traffic
     is always countable, even on the first report.

No earlier pull = no pull-against-pull rows and a plain note saying so.
Nothing is estimated; every row here is an item with a barcode.

A family passes its own scope (a predicate on the register row) so the
tooling report sees tooling, the gas report sees monitors, and so on.
"""

import glob
import os
import re
from datetime import datetime, timedelta

import openpyxl

import ampol_names
import ampol_paths

ON_HIRE = "on hire"


def parse_dt(d, t=None):
    """dd/mm/yyyy [+ hh:mm(:ss) [AM/PM]] -> datetime, or None."""
    if d is None or str(d).strip() == "":
        return None
    if isinstance(d, datetime):
        base = d
    else:
        s = str(d).strip()
        base = None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %I:%M %p"):
            try:
                base = datetime.strptime(s, fmt)
                break
            except ValueError:
                pass
        if base is None:
            return None
    if t is not None and str(t).strip():
        ts = str(t).strip()
        for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"):
            try:
                tt = datetime.strptime(ts, fmt)
                return base.replace(hour=tt.hour, minute=tt.minute, second=tt.second)
            except ValueError:
                pass
    return base


def reference_pull_time(path):
    """The REQUESTED_DATE/TIME on the export's REFERENCE_INFO tab."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "REFERENCE_INFO" not in wb.sheetnames:
            return None
        ws = wb["REFERENCE_INFO"]
        hdr = None
        for row in ws.iter_rows(values_only=True):
            vals = [str(c).strip() if c is not None else "" for c in row]
            if hdr is None:
                if any("REQUESTED_DATE" in v.upper() for v in vals):
                    hdr = vals
                continue
            for h, v in zip(hdr, vals):
                if "REQUESTED_DATE" in h.upper() and v:
                    for fmt in ("%d/%m/%Y %I:%M %p", "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S"):
                        try:
                            return datetime.strptime(v, fmt)
                        except ValueError:
                            pass
                    return None
        return None
    finally:
        wb.close()


def load_register(path):
    """{barcode: row} for every RENTAL_STOCK row with a barcode."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["RENTAL_STOCK"] if "RENTAL_STOCK" in wb.sheetnames else wb[wb.sheetnames[-1]]
        it = ws.iter_rows(values_only=True)
        hdr = [str(c).strip() if c is not None else "" for c in next(it)]
        ix = {h.upper(): i for i, h in enumerate(hdr)}

        def g(r, k):
            i = ix.get(k)
            return r[i] if i is not None and i < len(r) else None
        out = {}
        for r in it:
            bc = str(g(r, "ITEM_BARCODE") or "").strip()
            if not bc:
                continue
            out[bc] = {
                "barcode": bc,
                "item": str(g(r, "ITEM_NUMBER") or "").strip(),
                "desc": ampol_names.display_desc(str(g(r, "ITEM_DESCRIPTION") or "").strip()),
                "status": str(g(r, "ITEM_STATUS") or "").strip(),
                "company_raw": str(g(r, "COMPANY_NAME") or "").strip(),
                "company": ampol_names.display_company(str(g(r, "COMPANY_NAME") or "").strip()),
                "hirer": str(g(r, "HIRER_NAME") or "").strip(),
                "on_dt": parse_dt(g(r, "ON_HIRE_DATE"), g(r, "ON_HIRE_TIME")),
                "family": str(g(r, "PRODUCT_FAMILY") or "").strip(),
                "product": str(g(r, "PRODUCT") or "").strip(),
                "unit": str(g(r, "STORAGE_UNIT") or "").strip(),
            }
        return out
    finally:
        wb.close()


def find_previous_pull(current_path, current_time):
    """The newest earlier RENTAL_STOCK export in Data\\previous with a
    pull time before the current one. (path, pull_time) or (None, None)."""
    prev_dir = ampol_paths.previous_dir()
    cands = []
    for f in glob.glob(os.path.join(prev_dir, "*RENTAL_STOCK*.xlsx")):
        if os.path.basename(f).startswith("~$"):
            continue
        try:
            t = reference_pull_time(f)
        except Exception:
            t = None
        if t is None:
            m = re.match(r"(\d{8})_(\d{4})", os.path.basename(f))
            if m:
                t = datetime.strptime(m.group(1) + m.group(2), "%Y%m%d%H%M")
        if t is None or (current_time and t >= current_time):
            continue
        cands.append((t, f))
    if not cands:
        return None, None
    cands.sort()
    t, f = cands[-1]
    return f, t


def _is_out(row):
    return row["status"].lower() == ON_HIRE


def _days(row, at):
    if row["on_dt"] is None or at is None:
        return None
    return max(0, (at.date() - row["on_dt"].date()).days)


def diff(cur, prev, cur_time, prev_time, scope=None, thresholds=(30, 60, 90)):
    """Pull against pull. cur/prev: {barcode: row}. scope(row) keeps a
    row in the family (checked on whichever pull has the row)."""
    def keep(bc):
        # WHY (03 Sep 2026): the family rule is checked on BOTH pulls' rows -
        # an item that came back into a holding account still counts as
        # the family's if it was the family's when it went out
        rows = [r for r in (cur.get(bc), prev.get(bc)) if r is not None]
        return bool(rows) and (scope is None or any(scope(r) for r in rows))
    codes = sorted(set(cur) | set(prev), key=lambda b: b.upper())
    returned, issued, moved = [], [], []
    crossed = {t: [] for t in thresholds}
    co_prev, co_cur = set(), set()
    n_out_prev = n_out_cur = 0
    for bc in codes:
        if not keep(bc):
            continue
        c, p = cur.get(bc), prev.get(bc)
        c_out = c is not None and _is_out(c)
        p_out = p is not None and _is_out(p)
        if p_out:
            n_out_prev += 1
            co_prev.add(p["company"])
        if c_out:
            n_out_cur += 1
            co_cur.add(c["company"])
        if p_out and not c_out:
            r = dict(p)
            r["days_out"] = _days(p, cur_time)
            r["now"] = c["status"] if c else "not on the register"
            returned.append(r)
        elif c_out and not p_out:
            r = dict(c)
            r["days_out"] = _days(c, cur_time)
            issued.append(r)
        elif c_out and p_out:
            if (c["hirer"], c["company"]) != (p["hirer"], p["company"]):
                r = dict(c)
                r["from_hirer"], r["from_company"] = p["hirer"], p["company"]
                r["days_out"] = _days(c, cur_time)
                moved.append(r)
            dp, dc = _days(p, prev_time), _days(c, cur_time)
            if dp is not None and dc is not None:
                for t in thresholds:
                    if dp < t <= dc:
                        r = dict(c)
                        r["days_out"] = dc
                        crossed[t].append(r)
    for lst in (returned, issued, moved):
        lst.sort(key=lambda r: (ampol_names.sort_key(r["company"]), ampol_names.sort_key(r["hirer"]),
                                -(r["days_out"] or 0), r["desc"].upper()))
    for t in thresholds:
        crossed[t].sort(key=lambda r: (-(r["days_out"] or 0), ampol_names.sort_key(r["company"])))
    return {
        "have_previous": True, "prev_time": prev_time, "cur_time": cur_time,
        "returned": returned, "issued": issued, "moved": moved, "crossed": crossed,
        "companies_new": sorted(co_cur - co_prev, key=ampol_names.sort_key),
        "companies_cleared": sorted(co_prev - co_cur, key=ampol_names.sort_key),
        "out_prev": n_out_prev, "out_cur": n_out_cur,
    }


def no_previous(cur_time):
    return {"have_previous": False, "prev_time": None, "cur_time": cur_time,
            "returned": [], "issued": [], "moved": [], "crossed": {30: [], 60: [], 90: []},
            "companies_new": [], "companies_cleared": [], "out_prev": None, "out_cur": None}


def last_24h(tx_path, cur_time, scope_barcodes=None, hours=24):
    """Issues and returns in the `hours` before the pull, from the
    TRANSACTIONS export. scope_barcodes: set of barcodes in the family
    (None = every row). Returns {"issued": [...], "returned": [...],
    "window": (start, end)} - each row: barcode, desc, company, hirer, at."""
    start = cur_time - timedelta(hours=hours)
    wb = openpyxl.load_workbook(tx_path, read_only=True, data_only=True)
    try:
        if "CUSTOMER_CONTRACTOR_EQUIP" not in wb.sheetnames:
            return {"issued": [], "returned": [], "window": (start, cur_time), "available": False}
        ws = wb["CUSTOMER_CONTRACTOR_EQUIP"]
        it = ws.iter_rows(values_only=True)
        hdr = [str(c).strip().upper() if c is not None else "" for c in next(it)]
        ix = {h: i for i, h in enumerate(hdr)}

        def g(r, k):
            i = ix.get(k)
            return r[i] if i is not None and i < len(r) else None
        issued, returned = [], []
        for r in it:
            bc = str(g(r, "LATEST_BARCODE") or "").strip()
            if not bc or (scope_barcodes is not None and bc not in scope_barcodes):
                continue
            row = {"barcode": bc,
                   "desc": ampol_names.display_desc(str(g(r, "SKU/ITEM DESCRIPTION") or "").strip()),
                   "company": ampol_names.display_company(str(g(r, "EMPLOYER_NAME") or "").strip()),
                   "hirer": str(g(r, "HIRER_NAME") or "").strip()}
            s = parse_dt(g(r, "TRAN_START_DATE"), g(r, "TRAN_START_TIME"))
            e = parse_dt(g(r, "TRAN_END_DATE"), g(r, "TRAN_END_TIME"))
            if s and start < s <= cur_time:
                issued.append(dict(row, at=s))
            if e and start < e <= cur_time:
                returned.append(dict(row, at=e))
        for lst in (issued, returned):
            lst.sort(key=lambda x: x["at"])
        return {"issued": issued, "returned": returned, "window": (start, cur_time), "available": True}
    finally:
        wb.close()


def changes(scope=None, scope_barcodes=None, data_dir=None):
    """Everything a 'since the last report' page needs, in one call:
    the pull-against-pull diff (or the honest no-previous shape) and the
    last 24 hours of traffic from TRANSACTIONS."""
    data_dir = data_dir or ampol_paths.data_dir()
    cur_path = ampol_paths.find_data("RENTAL_STOCK*.xlsx")
    if not cur_path:
        raise FileNotFoundError("RENTAL_STOCK.xlsx is not in Data\\")
    cur_time = reference_pull_time(cur_path) or datetime.fromtimestamp(os.path.getmtime(cur_path))
    cur = load_register(cur_path)
    # WHY (03 Sep 2026): a family that hands in its barcode set gets the
    # diff on THAT set - before this, only the 24-hour block was scoped
    # and the pull-against-pull tables showed the whole store's movement
    # on every family's page (Andrew caught a gas monitor on the radio
    # page). The predicate is built from the set and applied to the diff.
    if scope is None and scope_barcodes is not None:
        _bcs = {str(b).strip().upper() for b in scope_barcodes}
        scope = lambda r: str(r.get("barcode", "")).strip().upper() in _bcs   # noqa: E731
    prev_path, prev_time = find_previous_pull(cur_path, cur_time)
    if prev_path:
        d = diff(cur, load_register(prev_path), cur_time, prev_time, scope)
        d["prev_path"] = prev_path
    else:
        d = no_previous(cur_time)
        d["prev_path"] = None
    tx_path = ampol_paths.find_data("TRANSACTIONS*.xlsx")
    if scope_barcodes is None and scope is not None:
        scope_barcodes = {bc for bc, r in cur.items() if scope(r)}
    d["last24"] = last_24h(tx_path, cur_time, scope_barcodes) if tx_path else \
        {"issued": [], "returned": [], "window": (cur_time - timedelta(hours=24), cur_time), "available": False}
    return d


if __name__ == "__main__":
    d = changes()
    print("current pull :", d["cur_time"])
    print("previous pull:", d["prev_time"], d.get("prev_path"))
    if d["have_previous"]:
        print(f"returned {len(d['returned'])}  issued {len(d['issued'])}  moved {len(d['moved'])}  "
              f"crossed 30/60/90 {len(d['crossed'][30])}/{len(d['crossed'][60])}/{len(d['crossed'][90])}")
        print("companies new:", d["companies_new"], " cleared:", d["companies_cleared"])
    else:
        print("no earlier pull in Data\\previous - pull-against-pull rows start with the next pull")
    l = d["last24"]
    print(f"last 24 h before the pull: issued {len(l['issued'])}, returned {len(l['returned'])}")
