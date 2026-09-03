#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=====================================================================
COATES CALIBRATION REGISTER - HOUSE-STYLE REPORT (the K2 look)
One run -> client-ready PDF + Outlook email: the human-maintained
calibration list joined to the live SiteIQ RENTAL_STOCK pull.
=====================================================================
Author: Andrew Fisher
The Coates Way - Operational Excellence - POWERED BY SITEIQ

WHAT THIS IS
  The calibration register's report of its own - the K2 house style,
  same shell as the stocktake family. One run produces:

  (file names come from ampol_names.report_stem - one rule for the
  whole suite, the day the button was pressed on the end, e.g. 03Sep2026)

   1. Coates_Ampol_Calibration_Register_<day>.pdf   (THE report - the
      position at the SiteIQ pull time, what fell due since the list was
      last maintained, the chase list, the honest No Date story, and the
      "on hire, not in the register" set)
      + Coates_Ampol_Calibration_Register_<day>.html (the same pages -
        kept beside the PDF so the report still exists when no PDF
        engine is on the machine; VERIFY_NUMBERS reads it)
   2. Coates_Ampol_Calibration_Register_<day>_OUTLOOK.eml (DRAFT -
      never sends; the position card shows inline under the header and
      rides as a file beside the PDF)
      + _OUTLOOK.body.html and _OUTLOOK.draft.json so
        MAKE_OUTLOOK_DRAFTS keeps working (the card is an attachment
        there).
   3. Coates_Ampol_Calibration_Register_<day>_PositionCard.png - the
      phone card: the position-page tiles, the RAG line and four scores.
  The PDF is finished with its document properties (Author: Andrew
  Fisher) and a bookmark per section (pdf_finish). Once seven days are
  on the History scoreboard it gains a trend page before the close.
  The PDF opens on a dark cover (the one number of the day), the position
  page carries the movement since the previous recorded pull and a RAG
  band with an owner and a dated next action, and the closing page shows
  the Coates Way. Each run writes its figures to History\report_history
  .json keyed on the pull day; the next run reads them back for the
  movement notes. No earlier day on record means no arrow - never a guess.

WHERE THE NUMBERS COME FROM
  Two workbooks in the suite's Data area:
   - Ampol_Calibration_Register.xlsx - read as the HUMAN-MAINTAINED
     LIST only: the 'Register Entry' sheet (asset, description, serial,
     certificate, calibration due). Nothing on these pages needs the
     workbook refreshed. Its refreshed sheets ('Live Register', 'On
     Hire Audit') are read only to print the register's own view at
     its Last Refresh as a labelled comparison.
   - RENTAL_STOCK.xlsx - the live SiteIQ pull (request time on its own
     REFERENCE_INFO sheet). Everything else is computed from it: where
     each asset is, who has it, the hire start date, the bay, the
     status buckets at the pull time, the chase list, and the "on
     hire, not in the register" set.

DATA RULES
  Data as at = the SiteIQ pull's request time. Status is computed from
  each asset's Calibration Due at that time with the register's own
  thresholds (overdue / due 0-30 / 31-60 / 61-90 / current / no date);
  the same rule is cross-checked against the register's own status at
  its Last Refresh and the result is printed.
  "Due dates last maintained" = when the register workbook was last
  saved (its own file properties, UTC turned into site time). When
  that is more than STALE_DAYS before the pull, page 1 and the console
  say so - the point being that a certificate issued since then is not
  on the register yet, not that anything needs refreshing.
  WHY (02 Sep 2026): Andrew's rule - the suite must never require a
  workbook refresh. The old report printed the register's own
  refreshed columns as today's position; they were 19 days old.
  Nothing is invented: a blank cell renders as a dash with a plain-
  words note; an asset SiteIQ cannot find is "not in SiteIQ", never
  "in store".
=====================================================================
"""

import html as _html
import json
import os
import re
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone
from email.message import EmailMessage
from email.utils import formatdate
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import ampol_paths
import build_stocktake_compliance_tool as eng
import ampol_names   # write_pdf_robust + find_workbook + parse_dt
import k2shell as sh
import pdf_finish
import report_history as rh   # the movement scoreboard - recorded days only
from k2shell import esc, money, num, K

import openpyxl

BASE = Path(__file__).resolve().parent
# Outputs land in the suite's dated Reports area - Reports\<today>\
# Calibrations - one folder per day, nothing ever silently overwritten.
OUT = Path(ampol_paths.day_folder("Calibrations"))

# Lytton is Brisbane time - AEST, no daylight saving. Used to turn the
# workbook's own UTC "modified" stamp into the time Andrew saw on screen.
SITE_TZ = timezone(timedelta(hours=10))
# Due dates last maintained more than this many days before the SiteIQ
# pull get the page-1 banner and the console warning. Three days covers
# a weekend.
STALE_DAYS = 3

CONFIG = {
    "client": "Ampol",
    "title": "Calibration Register",
    "kicker": "COATES · TOOL STORE - CALIBRATION REGISTER REPORT",
    "project": "Ampol Lytton Refinery · Permanent Tool Store",
    # WHY (03 Sep 2026): the one output name comes from ampol_names.report_stem
    # - the client reads the attachment name before a single figure, so the
    # whole suite shares one shape with the day on the end. Everything else
    # (.pdf, .html, _OUTLOOK.eml, _PositionCard.png, the manifest) hangs off it.
    "stem": ampol_names.report_stem("calibration"),
    # filled in by main() once both source stamps are known - the shared
    # page shell prints it after the as-at time on page 1
    "asat_note": "",
    # ---- the RAG line on the position page (default line - set here) ----
    # Red when this many overdue assets, or more, are out on hire (a lapsed
    # certificate in a customer's hands); Amber when the overdue gear is
    # off hire only (shelf, repairs, not in SiteIQ); Green when nothing is
    # overdue. The band prints the rule beside the result.
    "rag_owner": "Andrew Fisher, Shutdown Manager",
    "rag_red_onhire": 1,
    # the next action falls due this many days after the SiteIQ pull date -
    # anchored on the pull, never on the day the report happens to be run
    "action_days": 7,
    "team": [
        {"name": "Andrew Fisher", "role": "Shutdown Manager",
         "shift": "", "email": "andrew.fisher@coates.com.au",
         "blurb": "Owns the register and the chase list - anything at all, start here",
         "lead": True},
    ],
    "key_items": [
        ("orange", "IN DATE", "certificate current at the SiteIQ pull time"),
        ("amber", "DUE 30 DAYS", "book in before it lapses"),
        ("blue", "NO DATE", "calibration status unknown until the certificate is entered"),
    ],
}

STEM = CONFIG["stem"]
CONFIG.update({
    "pdf_name": f"{STEM}.pdf", "page_html": f"{STEM}.html",
    "eml_name": f"{STEM}_OUTLOOK.eml", "email_html": f"{STEM}_OUTLOOK.body.html",
    "draft_json": f"{STEM}_OUTLOOK.draft.json", "card_name": f"{STEM}_PositionCard.png",
})
# the scoreboard's path as the pages print it (a backslash cannot sit inside
# an f-string expression on the store laptops' Python)
HIST_NAME = "History\\report_history.json"

# The "looks like calibration-type gear" keyword rule for the on-hire,
# not-in-register set. A keyword match, not a judgement - the page says
# so. Gas monitors have their own programme and report; slings and chain
# blocks are lifting gear on the Rigging & Lifting Register's beat. Both
# are counted and disclosed, not silently dropped.
CAL_KEYWORDS = re.compile(
    r"FLUKE|CALIBRAT|TORQUE|GAUGE|MULTIMETER|CLAMP METER|PROCESS METER|"
    r"VIBRATION METER|WEATHER METER|TESTER|FLOW ?METER|CALIPER|MICROMETER|"
    r"VERNIER|THERMOMETER|MANOMETER|TACHOMETER|MEGGER", re.I)
CAL_EXCLUDE = re.compile(r"DR[ÄA]GER|X-AM|GAS MONITOR|GAS DETECT|\bSLING|CHAIN BLOCK", re.I)
CAL_RULE_WORDS = ("Fluke, calibrat-, torque, gauge, multimeter, clamp / process / "
                  "vibration / weather meter, tester, flowmeter, caliper, "
                  "micrometer, vernier, thermometer, manometer, tachometer, megger")
GAS_WORDS = re.compile(r"DR[ÄA]GER|X-AM|GAS MONITOR|GAS DETECT", re.I)
LIFT_WORDS = re.compile(r"\bSLING|CHAIN BLOCK", re.I)


def item_family(desc):
    """Item family for the on-hire, not-in-register set - a plain grouping
    by description keywords so the chart shows WHAT is on hire
    unregistered, not who."""
    u = str(desc or "").upper()
    if "RADIO" in u or "MOTOROLA" in u:
        return "Radios"
    if GAS_WORDS.search(u):
        return "Gas monitors"
    if re.search(r"SAMSUNG|GALAXY|PHONE|IPAD|TABLET", u):
        return "Phones and tablets"
    if "MILWAUKEE" in u or re.search(r"\bM1[28]\b", u):
        return "Milwaukee power tools"
    if re.search(r"\bSLING|SHACKLE|CHAIN BLOCK|LEVER HOIST|LEVER BLOCK|BEAM CLAMP|LIFTING", u):
        return "Rigging and lifting"
    if re.search(r"USB|CHARGER", u):
        return "Chargers and USB"
    if "RUBBISH CHUTE" in u:
        return "Rubbish chutes"
    if re.search(r"CRITIC?TAL RISK", u):
        return "Critical risk signage"
    return "Other general tooling"


def cal_kind(desc):
    """Sub-grouping inside the keyword slice, for the summary table."""
    u = str(desc or "").upper()
    if "FLUKE" in u or "CALIBRAT" in u:
        return "Fluke test and calibration gear"
    if "TORQUE" in u:
        return "Torque wrenches and torque tools"
    if "FLOW" in u:
        return "Argon flowmeters and regulators"
    if re.search(r"CALIPER|MICROMETER|VERNIER|FEELER", u):
        return "Calipers and feeler gauges"
    if "GAUGE" in u:
        return "Pressure and other gauges"
    if "METER" in u or "TESTER" in u:
        return "Meters and testers"
    return "Other keyword matches"


def is_repairs(hirer, company=""):
    """The repairs queue hires gear like a person does - SiteIQ company
    'Repairs', hirer names ending 'Repairs'. Those are a repair-queue
    conversation, not a chase."""
    return ("repair" in str(hirer or "").lower()
            or str(company or "").strip().lower() == "repairs")


def status_at(due, ref):
    """The register's own status thresholds, applied at any date."""
    if due is None:
        return "No Date"
    n = (due.date() - ref).days
    if n < 0:
        return "Overdue"
    if n <= 30:
        return "Due 0-30 Days"
    if n <= 60:
        return "Due 31-60 Days"
    if n <= 90:
        return "Due 61-90 Days"
    return "Current"


def first_words(desc, n=2):
    return " ".join(str(desc or "").split()[:n])


def as_dt(v):
    """A cell that should be a date: Excel date, or a typed dd/mm/yyyy."""
    if isinstance(v, datetime):
        return v
    if v in (None, ""):
        return None
    return eng.parse_dt(v)


# =====================================================================
# load - the human-maintained list, the live pull, the register's own
# view (comparison only)
# =====================================================================

def load_entry(path):
    """The human-maintained list: 'Register Entry' - asset, description,
    serial, certificate, calibration due. Falls back to the same
    columns on 'Live Register' if the entry sheet is missing."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    name = next((n for n in ("Register Entry", "Live Register") if n in wb.sheetnames), None)
    if not name:
        sys.exit("ERROR: the calibration workbook has no 'Register Entry' (or "
                 "'Live Register') sheet - is this the right file in Data\\?")
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h or "").strip() for h in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    need = ["New Asset No", "Description", "Serial No", "Certificate No.",
            "Calibration Date", "Calibration Due"]
    missing = [c for c in need if c not in ix]
    if missing:
        sys.exit(f"ERROR: '{name}' is missing columns {missing} - "
                 "the register layout has changed; report needs updating.")

    def cell(r, col):
        v = r[ix[col]] if col in ix else None
        return "" if v is None else v

    rows = []
    for r in it:
        if not any(c not in (None, "") for c in r):
            continue
        asset = str(cell(r, "New Asset No")).strip()
        desc = ampol_names.display_desc(str(cell(r, "Description")).strip())
        if not asset and not desc:
            continue      # template padding - a formula, no asset
        rows.append({
            "asset": asset,
            "desc": desc,
            "serial": str(cell(r, "Serial No")).strip(),
            "cert": str(cell(r, "Certificate No.")).strip(),
            "cal_date": as_dt(cell(r, "Calibration Date")),
            "due": as_dt(cell(r, "Calibration Due")),
            "notes": str(cell(r, "Notes")).strip(),
        })
    wb.close()
    return rows, name


def load_register_view(path):
    """The register's own refreshed view - 'Live Register' - read ONLY
    for the labelled comparison. Returns (view by asset, blank template
    rows, Last Refresh) - or ({}, 0, None) when the sheet is absent."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Live Register" not in wb.sheetnames:
        wb.close()
        return {}, 0, None
    ws = wb["Live Register"]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h or "").strip() for h in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}

    def cell(r, col):
        v = r[ix[col]] if col in ix else None
        return "" if v is None else v

    view, blanks, refresh = {}, 0, None
    for r in it:
        if not any(c not in (None, "") for c in r):
            continue
        lr = cell(r, "Last Refresh")
        if isinstance(lr, datetime) and (refresh is None or lr > refresh):
            refresh = lr
        asset = str(cell(r, "New Asset No")).strip()
        desc = ampol_names.display_desc(str(cell(r, "Description")).strip())
        if not asset and not desc:
            blanks += 1   # register template padding - a status formula, no asset
            continue
        days = cell(r, "Days Remaining")
        view[asset.upper()] = {
            "status": str(cell(r, "Calibration Status")).strip(),
            "days": int(days) if isinstance(days, (int, float)) else None,
            "onhire": str(cell(r, "On Hire")).strip(),   # Yes / No / blank
            "hirer": str(cell(r, "Hirer Name")).strip(),
            "unit": str(cell(r, "Storage Unit")).strip(),
            "due": as_dt(cell(r, "Calibration Due")),
        }
    wb.close()
    return view, blanks, refresh


def load_audit_count(path):
    """How many rows the register's own 'On Hire Audit' sweep holds -
    printed as a comparison to the set computed here, nothing more."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "On Hire Audit" not in wb.sheetnames:
        wb.close()
        return 0
    it = wb["On Hire Audit"].iter_rows(values_only=True)
    next(it, None)
    n = sum(1 for r in it if any(c not in (None, "") for c in r))
    wb.close()
    return n


def load_rental_stock(path):
    """The live SiteIQ pull: barcode -> where it is right now and who has
    it, plus the pull's own request time from REFERENCE_INFO."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    pulled = None
    if "REFERENCE_INFO" in wb.sheetnames:
        it = wb["REFERENCE_INFO"].iter_rows(values_only=True)
        hdr = [str(h or "").strip().upper() for h in next(it, [])]
        col = next((i for i, h in enumerate(hdr) if "REQUESTED_DATE" in h), None)
        if col is not None:
            for r in it:
                if r and r[col] not in (None, ""):
                    pulled = eng.parse_dt(r[col])
                    break
    name = "RENTAL_STOCK" if "RENTAL_STOCK" in wb.sheetnames else wb.sheetnames[-1]
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h or "").strip() for h in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    need = ["ITEM_BARCODE", "ITEM_DESCRIPTION", "ITEM_STATUS", "HIRER_NAME",
            "COMPANY_NAME", "ON_HIRE_DATE", "ON_HIRE_TIME", "STORAGE_UNIT"]
    missing = [c for c in need if c not in ix]
    if missing:
        sys.exit(f"ERROR: RENTAL_STOCK is missing columns {missing} - "
                 "the SiteIQ export layout has changed; report needs updating.")

    def col(r, name):
        return str(r[ix[name]] or "").strip()

    live, n_rows = {}, 0
    for r in it:
        bc = col(r, "ITEM_BARCODE").upper()
        if not bc:
            continue
        n_rows += 1
        out_d, out_t = col(r, "ON_HIRE_DATE"), col(r, "ON_HIRE_TIME")
        out = eng.parse_dt(f"{out_d} {out_t}".strip()) or eng.parse_dt(out_d)
        live[bc] = {
            "barcode": col(r, "ITEM_BARCODE"),
            "status": col(r, "ITEM_STATUS"),
            "hirer": col(r, "HIRER_NAME"),
            "company": col(r, "COMPANY_NAME"),
            "out": out,
            "unit": col(r, "STORAGE_UNIT"),
            "desc": ampol_names.display_desc(col(r, "ITEM_DESCRIPTION")),
            "former_name": ampol_names.carries_former_name(col(r, "ITEM_DESCRIPTION")),
        }
    wb.close()
    return live, pulled, n_rows


def workbook_saved(path):
    """When the workbook was last saved, from its own docProps/core.xml
    (dcterms:modified, UTC) turned into site time. The file's time on
    disk is only when it was copied here, so it is the fallback, and
    the label says which one was used."""
    try:
        with zipfile.ZipFile(path) as z:
            xml = z.read("docProps/core.xml").decode("utf-8", "replace")
        m = re.search(r"<dcterms:modified[^>]*>([^<]+)</dcterms:modified>", xml)
        if m:
            stamp = datetime.strptime(m.group(1).strip()[:19], "%Y-%m-%dT%H:%M:%S")
            local = stamp.replace(tzinfo=timezone.utc).astimezone(SITE_TZ)
            return local.replace(tzinfo=None), "from the file's own properties"
    except Exception:
        pass
    return datetime.fromtimestamp(os.path.getmtime(path)), "file time on disk"


# =====================================================================
# derive - the position at the SiteIQ pull time, computed
# =====================================================================

def derive(rows, view, live, asat_d, maint_d, refresh, audit_n):
    d = {}
    d["total"] = len(rows)
    d["former_live"] = sum(1 for v in live.values() if v.get("former_name"))
    regset = {r["asset"].upper() for r in rows if r["asset"]}
    d["dup_assets"] = len([r for r in rows if r["asset"]]) - len(regset)

    # ---- every asset: status at the pull time + where SiteIQ has it -----
    for r in rows:
        r["now_status"] = status_at(r["due"], asat_d)
        r["now_days"] = (r["due"].date() - asat_d).days if r["due"] else None
        l = live.get(r["asset"].upper())
        r["live"] = l
        r["where"] = l["status"] if l else "not in SiteIQ"
        r["live_hirer"] = l["hirer"] if l else ""
        r["company"] = l["company"] if l else ""
        r["repairs_now"] = is_repairs(l["hirer"], l["company"]) if l else False
        r["out"] = l["out"] if l else None
        r["issued_after_due"] = bool(l and l["status"] == "On Hire" and l["out"]
                                     and r["due"] and l["out"].date() > r["due"].date())
        r["reg"] = view.get(r["asset"].upper())
    d["now"] = Counter(r["now_status"] for r in rows)
    d["now_overdue"] = sorted([r for r in rows if r["now_status"] == "Overdue"],
                              key=lambda r: r["now_days"])
    d["now_due30"] = sorted([r for r in rows if r["now_status"] == "Due 0-30 Days"],
                            key=lambda r: r["due"])
    d["now_incal"] = [r for r in rows if r["now_status"] in
                      ("Current", "Due 31-60 Days", "Due 61-90 Days")]
    d["nodate"] = [r for r in rows if r["due"] is None]
    d["dated"] = [r for r in rows if r["due"] is not None]
    d["now_dated_ok"] = [r for r in d["dated"] if r["now_status"] != "Overdue"]
    d["lapsed"] = sorted([r for r in rows if r["due"] and maint_d <= r["due"].date() < asat_d],
                         key=lambda r: r["due"])
    d["now_od_onhire"] = [r for r in d["now_overdue"] if r["where"] == "On Hire"]
    d["now_od_avail"] = [r for r in d["now_overdue"] if r["where"] == "Available for Hire"]
    d["now_od_missing"] = [r for r in d["now_overdue"] if r["live"] is None]
    d["now_od_other"] = [r for r in d["now_overdue"]
                         if r["live"] and r["where"] not in ("On Hire", "Available for Hire")]
    d["issued_after"] = [r for r in d["now_od_onhire"] if r["issued_after_due"]]
    d["not_in_siteiq"] = [r for r in rows if r["live"] is None]
    d["onhire_live"] = [r for r in rows if r["where"] == "On Hire"]

    # the chase list - person first, worst first; the repairs queue split
    # out because that is a different conversation
    def grouped(items, key, days_key):
        # WHY (02 Sep 2026): hirers A-Z - a name is found by eye - and the
        # worst item first inside a name (the row sort in chase_rows).
        g = defaultdict(list)
        for r in items:
            g[key(r) or "(no hirer recorded)"].append(r)
        return sorted(g.items(), key=lambda kv: ampol_names.sort_key(kv[0]))
    d["now_chase_hirer"] = grouped([r for r in d["now_od_onhire"] if not r["repairs_now"]],
                                   lambda r: r["live_hirer"], "now_days")
    d["now_chase_repairs"] = grouped([r for r in d["now_od_onhire"] if r["repairs_now"]],
                                     lambda r: r["live_hirer"], "now_days")

    # overdue on the shelf, grouped by description and SiteIQ bay for the
    # calibration run
    grp = defaultdict(list)
    for r in d["now_od_avail"]:
        grp[(r["desc"] or "(no description)", r["live"]["unit"] or "(no bay recorded)")].append(r["asset"])
    d["now_od_avail_groups"] = sorted(grp.items(), key=lambda kv: (-len(kv[1]), kv[0]))
    d["now_od_avail_units"] = Counter(r["live"]["unit"] or "(no bay recorded)" for r in d["now_od_avail"])
    d["now_od_family"] = Counter(first_words(r["desc"]) for r in d["now_overdue"] if r["desc"])

    # the No Date position - where SiteIQ has them, what they are, on hire
    nd = d["nodate"]
    d["nd_units"] = Counter(("(not in SiteIQ)" if r["live"] is None
                             else (r["live"]["unit"] or "(no bay recorded)")) for r in nd)
    d["nd_family"] = Counter(first_words(r["desc"]) for r in nd if r["desc"])
    d["nd_serial"] = sum(1 for r in nd if r["serial"] and r["serial"] != "-")
    d["nd_onhire_live"] = [r for r in nd if r["where"] == "On Hire"]
    d["nd_onhire_family"] = Counter(first_words(r["desc"]) for r in d["nd_onhire_live"] if r["desc"])
    d["nd_missing"] = [r for r in nd if r["live"] is None]

    # on hire, not in the register - computed from the pull and the list
    d["live_n"] = len(live)
    d["live_onhire"] = sum(1 for l in live.values() if l["status"] == "On Hire")
    nir = [l for bc, l in live.items() if l["status"] == "On Hire" and bc not in regset]
    d["nir"] = nir
    d["nir_hirers"] = len({l["hirer"] for l in nir})
    d["nir_family"] = Counter(item_family(l["desc"]) for l in nir)
    d["kw_gas"] = sum(1 for l in nir if GAS_WORDS.search(l["desc"]))
    d["kw_lift"] = sum(1 for l in nir if LIFT_WORDS.search(l["desc"]))
    d["kw_chain"] = sum(1 for l in nir if re.search(r"CHAIN BLOCK", l["desc"], re.I))
    kw = [l for l in nir if CAL_KEYWORDS.search(l["desc"]) and not CAL_EXCLUDE.search(l["desc"])]
    d["kw"] = kw
    d["kw_hirers"] = len({l["hirer"] for l in kw})
    kinds = defaultdict(list)
    for l in kw:
        kinds[cal_kind(l["desc"])].append(l)
    d["kw_kinds"] = sorted(kinds.items(), key=lambda kv: (-len(kv[1]), kv[0]))
    groups = defaultdict(list)
    for l in kw:
        groups[(cal_kind(l["desc"]), " ".join(l["desc"].split()),
                l["hirer"] or "(no hirer recorded)", l["company"])].append(l["barcode"])
    d["kw_groups"] = sorted(groups.items(), key=lambda kv: (kv[0][0], kv[0][1].upper(), kv[0][2]))
    d["audit_n"] = audit_n

    # ---- the register's own view at its Last Refresh - comparison only --
    d["have_view"] = bool(view) and refresh is not None
    withv = [r for r in rows if r["reg"]]
    d["view_n"] = len(withv)
    d["reg_status"] = Counter(r["reg"]["status"] or "(no status)" for r in withv)
    d["reg_incal"] = [r for r in withv if r["reg"]["status"] in ("Current", "Due 31-60 Days", "Due 61-90 Days")]
    d["reg_due30"] = [r for r in withv if r["reg"]["status"] == "Due 0-30 Days"]
    d["reg_overdue"] = [r for r in withv if r["reg"]["status"] == "Overdue"]
    d["reg_nodate"] = [r for r in withv if r["reg"]["status"] == "No Date"]
    d["reg_dated_ok"] = [r for r in withv if r["due"] and r["reg"]["status"] != "Overdue"]
    d["reg_od_onhire"] = sorted([r for r in d["reg_overdue"] if r["reg"]["onhire"] == "Yes"],
                                key=lambda r: r["reg"]["days"] or 0)
    d["reg_zero_day"] = [r for r in withv if r["reg"]["days"] == 0]
    d["reg_due30_lapsed"] = [r for r in d["reg_due30"] if r["due"] and r["due"].date() < asat_d]
    d["reg_due30_still"] = [r for r in d["reg_due30"] if r["now_status"] == "Due 0-30 Days"]
    d["reg_due30_new"] = [r for r in d["now_due30"] if not (r["reg"] and r["reg"]["status"] == "Due 0-30 Days")]
    refresh_d = refresh.date() if refresh else asat_d
    d["rule_mismatch"] = sum(1 for r in withv
                             if status_at(r["due"], refresh_d) != (r["reg"]["status"] or "No Date"))
    d["view_due_mismatch"] = sum(1 for r in withv if r["reg"]["due"] != r["due"])

    def grouped_reg(items):
        g = defaultdict(list)
        for r in items:
            g[r["reg"]["hirer"] or "(no hirer recorded)"].append(r)
        return sorted(g.items(), key=lambda kv: min(x["reg"]["days"] or 0 for x in kv[1]))
    d["reg_chase_hirer"] = grouped_reg([r for r in d["reg_od_onhire"] if not is_repairs(r["reg"]["hirer"])])
    d["reg_chase_repairs"] = grouped_reg([r for r in d["reg_od_onhire"] if is_repairs(r["reg"]["hirer"])])
    d["reg_chase_back"] = [r for r in d["reg_od_onhire"] if r["where"] == "Available for Hire"]
    d["reg_chase_still"] = [r for r in d["reg_od_onhire"] if r["where"] == "On Hire"]
    return d


# =====================================================================
# PDF rendering via the engine's robust writer
# =====================================================================

_MEASURE_JS = r"""<script>
document.fonts.ready.then(function(){
  var out = [];
  var pages = document.querySelectorAll('.page');
  for (var i = 0; i < pages.length; i++) {
    var pg = pages[i];
    var body = pg.querySelector('.body');
    var foot = pg.querySelector('.foot');
    if (!body || !foot) { out.push({page: i + 1, over: null, wide: 0}); continue; }   // the cover: no footer by design
    var ft = foot.getBoundingClientRect();
    var bottom = body.getBoundingClientRect().top;
    var els = body.querySelectorAll('*');
    for (var j = 0; j < els.length; j++) {
      var r = els[j].getBoundingClientRect();
      if (r.bottom > bottom) bottom = r.bottom;
    }
    out.push({page: i + 1, over: Math.round(bottom - ft.top),
              wide: Math.round(body.scrollWidth - body.clientWidth)});
  }
  var d = document.createElement('div');
  d.id = 'layout-report';
  d.setAttribute('data-lato', document.fonts.check('12px Lato') ? '1' : '0');
  d.textContent = JSON.stringify(out);
  document.body.appendChild(d);
});
</script>"""


def _browser():
    try:
        from generate_k2style_gas_monitor_report import find_browser
        return find_browser()
    except Exception:
        import shutil
        for name in ("msedge", "chrome", "chromium", "chromium-browser", "google-chrome"):
            if shutil.which(name):
                return shutil.which(name)
        return None


def fit_check(doc, css, label):
    """Measure every page in the browser with the house face loaded.
    WHY (03 Sep 2026): k2style.css now embeds Lato, and the shared fit
    check measured before the font files had loaded - it saw the fallback
    face, up to 9 px off on a full page, on pages that had 0 px to spare.
    This one waits on document.fonts.ready under a virtual-time budget,
    with the page written beside the PDF so the relative font URLs
    resolve, so the pixels it reports are the pixels the PDF prints.
    Returns (ok, worst_spare_px, rows); ok is None when nothing could
    measure - never a reason not to build."""
    import subprocess
    import tempfile
    browser = _browser()
    if not browser:
        print(f"Fit check            : skipped - no browser to measure with ({label})")
        return None, None, []
    doc2 = (doc.replace("</head>", f"<style>{css}</style></head>", 1)
               .replace("</body>", _MEASURE_JS + "</body>", 1))
    tmp = OUT / f"__measure_{os.getpid()}__.html"
    profile = os.path.join(tempfile.gettempdir(), f"coates_fit_{os.getpid()}")
    try:
        tmp.write_text(doc2, encoding="utf-8")
        res = subprocess.run([browser, "--headless", "--disable-gpu", "--no-sandbox",
                              "--no-first-run", "--user-data-dir=" + profile,
                              "--virtual-time-budget=10000", "--dump-dom", tmp.as_uri()],
                             capture_output=True, timeout=240)
        dom = (res.stdout or b"").decode("utf-8", "ignore")
    except Exception as e:
        print(f"Fit check            : skipped ({type(e).__name__}) ({label})")
        return None, None, []
    finally:
        try:
            tmp.unlink()
        except OSError:
            pass
    m = re.search(r'id="layout-report" data-lato="(\d)">(\[.*?\])</div>', dom, re.S)
    if not m:
        print(f"Fit check            : skipped - the page could not be measured ({label})")
        return None, None, []
    rows = [(r["page"], r["over"], r["wide"]) for r in json.loads(m.group(2))]
    face = "Lato" if m.group(1) == "1" else "FALLBACK FACE - Lato did not load"
    rows = [r for r in rows if r[1] is not None]      # the cover carries no footer
    bad = [r for r in rows if r[1] > 0 or r[2] > 0]
    worst = max((r[1] for r in rows), default=-9999)
    if bad:
        print("*" * 68)
        print(f"WARNING: CONTENT DOES NOT FIT in {label} - do not send as is ({face}).")
        for pg, over, wide in bad:
            print(f"  page {pg:2d}: {over:+d}px past the footer" + (f", {wide}px too wide" if wide > 0 else ""))
        print("*" * 68)
        return False, -worst, rows
    print(f"Fit check            : PASS - tightest page has {-worst}px to spare, measured in {face} ({label})")
    return True, -worst, rows


def render_k2_pdf(doc, pdf_path, authored, css):
    """Measures every page with the house face loaded (fit_check), writes
    the page HTML beside the PDF (kept - it IS the report when no PDF
    engine is on the machine, and VERIFY_NUMBERS reads it), renders the
    PDF, and returns (pdf_ok, layout_ok) - layout_ok covers the fit
    check, the page count and the footer clearance. Never exits: the
    email step still runs."""
    fit_ok, _, _ = fit_check(doc, css, Path(pdf_path).name)
    doc = doc.replace("</head>", f"<style>{css}</style></head>", 1)
    page_html = OUT / CONFIG["page_html"]
    page_html.write_text(doc, encoding="utf-8")
    # pre-delete: write_pdf_robust treats an EXISTING file as success, so a
    # stale copy from a previous run must never be able to masquerade
    try:
        Path(pdf_path).unlink()
    except OSError:
        pass
    ok = eng.write_pdf_robust(str(page_html), str(pdf_path))
    if not ok or not Path(pdf_path).exists():
        print("*" * 68)
        print(f"WARNING: could not render {Path(pdf_path).name} - no PDF engine "
              "on this machine (Edge is standard on Coates laptops).")
        print(f"         The report pages are in {page_html.name} beside it; "
              "the email is built without the PDF attached.")
        print("*" * 68)
        return False, False
    return True, layout_check(pdf_path, authored) and fit_ok is not False


def layout_check(pdf_path, authored):
    """PASS only when the page count matches AND nothing on any page has
    run down into the footer. The K2 page is a fixed-height box with
    overflow hidden, so a page count alone cannot see a band or a table
    that has grown past the footer - PyMuPDF (where installed) reads
    every page and looks for body text or drawings overlapping it. The
    cover is page 1 and carries no footer by design, so it is skipped."""
    problems = []
    raw = open(pdf_path, "rb").read()
    counts = re.findall(rb"/Count\s+(\d+)", raw)
    actual = max(int(c) for c in counts) if counts else -1
    if actual == -1:
        problems.append(f"page count unreadable (authored {authored})")
    elif actual != authored:
        problems.append(f"authored {authored} pages, PDF has {actual}")
    try:
        import pymupdf
    except ImportError:
        pymupdf = None
    if pymupdf is not None:
        doc = pymupdf.open(str(pdf_path))
        for pno, page in enumerate(doc, start=1):
            if pno <= COVER_PAGES:
                continue           # the dark cover has no footer by design
            h = page.rect.height
            blocks = page.get_text("blocks")
            # letter-spaced footer text extracts as "Y O U R  C O A T E S ..." -
            # compare with the whitespace stripped
            foot = [b for b in blocks
                    if "COATESTOOLSTORETEAM" in re.sub(r"\s+", "", str(b[4])).upper()]
            if not foot:
                problems.append(f"page {pno}: footer not found")
                continue
            foot_top = min(b[1] for b in foot) - 7   # the rule above the footer text
            for b in blocks:
                if b[1] >= foot_top - 1:
                    continue           # the footer's own lines
                if b[3] > foot_top:
                    txt = re.sub(r"\s+", " ", str(b[4]))[:40]
                    problems.append(f"page {pno}: text runs into the footer ({txt!r})")
                    break
            for dr in page.get_drawings():
                r = dr.get("rect")
                if r is None or r.height > 0.8 * h or r.height < 3:
                    continue           # the orange frame / hairlines
                if r.y0 < foot_top - 2 and r.y1 > foot_top + 2:
                    problems.append(f"page {pno}: a chart or panel runs into the footer")
                    break
        doc.close()
        checked = f"{actual} pages, footer clearance checked on every page after the cover"
    else:
        checked = f"{actual} pages (page count only - PyMuPDF not installed)"
    if problems:
        print("*" * 68)
        print(f"Layout check         : FAIL - {Path(pdf_path).name}")
        for pr in problems:
            print(f"   - {pr}")
        print("   Do not send as is.")
        print("*" * 68)
        return False
    print(f"Layout check         : PASS - {checked} ({Path(pdf_path).name})")
    return True


# =====================================================================
# shared page fragments (PDF)
# =====================================================================

def psect(title):
    return f'<div class="sect"><h3>{title}</h3></div>'


def pcallout(inner, tight=True):
    cls = "callout tight" if tight else "callout"
    return f'<div class="{cls}">{inner}</div>'


def pnote(inner):
    return f'<div class="note">{inner}</div>'


def psubh(title, thin=""):
    t = f' <span class="thin">{thin}</span>' if thin else ""
    return f'<div class="sub-h">{title}{t}</div>'


def prowlab(text):
    return f'<div class="rowlab">{text}</div>'


def chartpanel(inner):
    return f'<div class="chartpanel">{inner}</div>'


def dash():
    return '<span class="tbc">&ndash;</span>'


def tag(text, cls):
    return f'<span class="tag {cls}">{esc(text)}</span>'


def fmt_d(dt):
    return dt.strftime("%d %b %Y") if dt else dash()


def fmt_due(r):
    return fmt_d(r["due"])


def fmt_days(days, overdue=False):
    if days is None:
        return dash()
    if overdue:
        return f'<span class="rd">{num(abs(days))}d over</span>'
    cls = "rd" if days < 0 else "a" if days <= 30 else "g"
    return f'<span class="{cls}">{num(days)}d</span>'


def where_cell(r):
    """Where SiteIQ says the asset is at the pull time."""
    if r["live"] is None:
        return tag("not in SiteIQ", "red")
    if r["where"] == "On Hire":
        return '<span class="a">On hire</span>'
    if r["where"] == "Available for Hire":
        return "In store"
    return esc(r["where"])


def who_cell(r):
    if r["live"] is None or r["where"] != "On Hire":
        return dash()
    return esc(r["live_hirer"]) if r["live_hirer"] else dash()


def out_cell(r):
    """SiteIQ hire start, flagged when it is later than the due date."""
    if r["live"] is None or r["where"] != "On Hire" or not r["out"]:
        return dash()
    s = esc(r["out"].strftime("%d %b %Y"))
    if r["issued_after_due"]:
        s += tag("after due", "red")
    return s


def plural(n, one, many=None):
    return one if n == 1 else (many or one + "s")


def isare(n):
    return "is" if n == 1 else "are"


# the dark cover is page 1 of the pack, so the position page is page 2 and
# every cross-reference counts from there
COVER_PAGES = 1
# tile note colours as the phone card draws them (the page uses CSS classes)
NOTE_HEX = {"green": "#1FA75A", "amber": "#F5A623", "red": "#EF4444", "grey": "#8A9AAC"}


def plain(fragment):
    """The words of a page fragment with the markup taken off - for the
    phone card, which draws text, not HTML."""
    return _html.unescape(re.sub(r"<[^>]+>", "", fragment)).replace("\xa0", " ")


def moved(key, asat_dt, value, good, note, ncls):
    """(note, class) for a tile: the movement since the previous recorded
    pull when there is one (report_history), else the tile's own note."""
    mv, mcls = rh.movement("calibration", key, asat_dt, value, good=good)
    return (mv, mcls) if mv else (note, ncls)


def spark_of(key, asat_dt, value):
    """30-day sparkline values - every earlier recorded day, then today.
    None until an earlier day exists: no history means no line."""
    past = [v for dd, v in rh.series("calibration", key, asat_dt, 30) if dd < asat_dt.date()]
    return past + [value] if past else None


def history_days(family, asat):
    """Days the scoreboard holds for a family, today's counted in - the
    entry is written after the pages, keyed on the pull day."""
    return len(set(rh.load().get(family, {})) | {asat.date().isoformat()})


def trend_rows(family, asat, today):
    """(labels, {key: [values]}, days) over the 30-day window ending at the
    pull day: every recorded earlier day plus today's own figures. A day
    with no run for a key leaves None - the chart draws a gap."""
    window = {}
    for key in today:
        for dd, v in rh.series(family, key, asat, days=30):
            if dd < asat.date():
                window.setdefault(dd, {})[key] = v
    window.setdefault(asat.date(), {}).update(today)
    days = sorted(window)
    labels = [dd.strftime("%d %b") for dd in days]
    return labels, {k: [window[dd].get(k) for dd in days] for k in today}, days


def trend_page(asat_dt, d):
    """The fixed trend page: overdue, due inside 30 days and in calibration
    over the last 30 days, from the History scoreboard. Empty until seven
    days are on record - the closing page says so - and every point is a
    figure a report printed on that day."""
    if history_days("calibration", asat_dt) < 7:
        return ""
    today = {"overdue": len(d["now_overdue"]), "due30": len(d["now_due30"]),
             "incal": len(d["now_incal"])}
    labels, ser, days = trend_rows("calibration", asat_dt, today)
    if len(days) < 2:
        return ""
    first, last = days[0].strftime("%d %b"), days[-1].strftime("%d %b %Y")

    def cell(v):
        return num(v) if v is not None else '<span class="tbc">no run</span>'
    trows = [[dd.strftime("%d %b %Y"), cell(ser["overdue"][i]), cell(ser["due30"][i]), cell(ser["incal"][i])]
             for i, dd in enumerate(days)][-10:]
    return f"""{psect("The trend - last 30 days")}
{pcallout(f'<b>{num(len(days))} days on record</b> between {esc(first)} and {esc(last)}, read back from {HIST_NAME} - every point is a figure a report printed on that day, nothing interpolated. A day with no run leaves a gap. The lines answer the one question a single pull cannot: is the overdue tail shrinking, and is the due-30 list being booked in before it lapses?', False)}
{psubh("Overdue, due inside 30 days and in calibration", "- assets at each pull, computed from Calibration Due")}
{chartpanel(sh.line_chart(labels, [("Overdue", ser["overdue"]), ("Due inside 30 days", ser["due30"]), ("In calibration", ser["incal"])], y_label="assets", h=220))}
{sh.dtable(["Pull day (last " + str(len(trows)) + " on record)", "Overdue", "Due inside 30 days", "In calibration"], trows, ["", "r", "r", "r"], "cp")}
{pnote('The scoreboard holds exactly what each day&rsquo;s report printed - the same figures as the position page of that day&rsquo;s PDF. Numbers, not lines, are the record.')}"""


def cal_rag(n_od, n_od_oh):
    """The position-page RAG - the rule in CONFIG applied to the counts at
    the pull. Returns (status, headline, card_headline): the headline is
    HTML for the band; the card headline says the same in fewer words,
    because the phone card draws a fixed box (two headline lines)."""
    if n_od_oh >= CONFIG["rag_red_onhire"]:
        return "red", (f'<b class="o">{num(n_od_oh)}</b> of the <b>{num(n_od)}</b> overdue assets '
                       f'{isare(n_od_oh)} out on hire at the pull - a lapsed certificate in a '
                       f'customer&rsquo;s hands until each one is swapped or recalled.'), (
                       f'{num(n_od_oh)} of the {num(n_od)} overdue assets {isare(n_od_oh)} out on '
                       f'hire - lapsed certificates in customers\' hands.')
    if n_od:
        return "amber", (f'<b>{num(n_od)}</b> overdue {plural(n_od, "asset")}, none on hire - '
                         f'the shelf and the not-in-SiteIQ list: a calibration run, not a chase.'), (
                         f'{num(n_od)} overdue {plural(n_od, "asset")}, none on hire - a calibration '
                         f'run, not a chase.')
    return "green", ('Nothing on the register is overdue at the pull - every dated asset is '
                     'inside its certificate.'), 'Nothing on the register is overdue at the pull.'


# =====================================================================
# the PDF pages
# =====================================================================

STATUS_ORDER = [("Current", K["green"]), ("Due 61-90 Days", "#4CC38A"),
                ("Due 31-60 Days", K["blue"]), ("Due 0-30 Days", K["amber"]),
                ("Overdue", K["red"]), ("No Date", "#5F7183")]

ROWS_STD = 16   # what a page holds cleanly under the K2 shell
ROWS_CP = 26    # the compact table variant
ROWS_KW = 20    # the keyword list - long descriptions wrap, so fewer per page


def table_pages(title, lead, headers, aligns, rows, per_page, note, cls="cp"):
    """One table across as many pages as it needs. The lead callout goes
    on the first page, the note on every page, the title gets
    '(continued)' - an overflowing page never goes to a client."""
    chunks = [rows[i:i + per_page] for i in range(0, len(rows), per_page)] or [[]]
    pages = []
    for k, chunk in enumerate(chunks):
        t = title if k == 0 else f"{title} (continued)"
        run = (pnote(f'Rows {k * per_page + 1}&ndash;{k * per_page + len(chunk)} of {len(rows)}.')
               if len(chunks) > 1 else "")
        body = (sh.dtable(headers, chunk, aligns, cls) if chunk
                else pnote("Nothing to list - the source holds no rows for this table."))
        pages.append(f"{psect(t)}{lead if k == 0 else ''}{body}{run}{note}")
    return pages


def build_pages(rows, d, S):
    """S = the stamps dict from main(). Returns (pages, marks): marks maps
    a section key to the page it starts on, so every cross-reference in
    the text is filled in after pagination - never hard-coded."""
    P, marks = [], {}

    def mark(key):
        # +1 for the page about to be appended, +COVER_PAGES for the cover
        marks[key] = len(P) + 1 + COVER_PAGES
    total = d["total"]
    dated = len(d["dated"])
    asat_s, asat_day, maint_short = S["asat_s"], S["asat_day"], S["maint_short"]
    refresh_short = S["refresh_short"]
    now_pct = len(d["now_dated_ok"]) / dated * 100 if dated else 0
    n_od, n_od_oh = len(d["now_overdue"]), len(d["now_od_onhire"])
    n_av, n_mi, n_ot = len(d["now_od_avail"]), len(d["now_od_missing"]), len(d["now_od_other"])
    n_l = len(d["lapsed"])

    # ---- P1 the position at the pull time -------------------------------
    if S["stale_days"] > STALE_DAYS:
        banner = (f'<div class="stale"><span class="h">Due dates are {num(S["stale_days"])} days old - check the paperwork before sending</span>'
                  f'Due dates last maintained <b>{esc(maint_short)}</b> (the register workbook&rsquo;s last save), '
                  f'<b>{num(S["stale_days"])} days</b> before the SiteIQ pull this report is built on. A certificate issued '
                  f'since then is not on the register yet - {num(n_l)} {plural(n_l, "asset")} fell due in that window '
                  f'(page @@P:lapsed@@). New certificates go in the <b>Register Entry</b> tab; save and run again.</div>')
    else:
        banner = ""
    fam = d["now_od_family"].most_common(2)
    if fam:
        fam_n = sum(n for _, n in fam)
        cause = (f'The overdue tail is <b>{esc(" and ".join(f0.lower() for f0, _ in fam))}</b> '
                 f'({num(fam_n)} of the {num(n_od)}); <b class="rd">{num(n_od_oh)}</b> {isare(n_od_oh)} out on hire '
                 f'right now (the chase list, page @@P:chase@@) and {num(n_l)} fell due after the due dates were '
                 f'last maintained on {esc(maint_short)} (page @@P:lapsed@@)')
    else:
        cause = 'Nothing is overdue - the chase list on page @@P:chase@@ is empty'
    # where the overdue and No Date gear is - printed on page 2 beside the
    # status mix, where it has the room (page 1 clipped its own note when
    # this row sat there too)
    asat_dt = S["asat_dt"]
    n_nd = len(d["nodate"])
    # every tile: the count, then the movement since the previous recorded
    # pull where one exists (report_history), else the tile's own note
    where_spec = [
        ("people", "chase", n_od_oh, "Overdue, on hire", "the chase list - page @@P:chase@@",
         "red" if n_od_oh else "green", "down"),
        ("wrench", "overdue_shelf", n_av, "Overdue, on the shelf", "the calibration run - page @@P:shelf@@",
         "amber" if n_av else "green", "down"),
        ("diamond", "not_in_siteiq", n_mi, "Overdue, not in SiteIQ", "whereabouts unknown - page @@P:missing@@",
         "red" if n_mi else "green", "down"),
        ("layers", "nodate", n_nd, "No date",
         f"{num(len(d['nd_onhire_live']))} of them out on hire - page @@P:nodate@@",
         "amber" if d["nd_onhire_live"] else "grey", "down"),
    ]
    where_items = []
    for ico, key, val, lab, note, ncls, good in where_spec:
        note, ncls = moved(key, asat_dt, val, good, note, ncls)
        where_items.append((ico, num(val), lab, note, ncls))
    where_tiles = (prowlab("Where the overdue and No Date gear is - SiteIQ RENTAL_STOCK")
                   + sh.tiles_plus(where_items))
    # the position-page tiles carry a 30-day sparkline once there is an
    # earlier day on record
    tiles1 = []
    for ico, key, val, lab, note, ncls, good in [
            ("box", "assets", total, "Assets on the register", "the human-maintained list", "grey", "up"),
            ("check", "incal", len(d["now_incal"]), "In calibration", "current, next due 31+ days", "green", "up"),
            ("clock", "due30", len(d["now_due30"]), "Due inside 30 days", "book them in now - page @@P:due30@@",
             "amber" if d["now_due30"] else "green", "down"),
            ("warn", "overdue", n_od, "Overdue", f"{num(n_l)} fell due since {maint_short}",
             "red" if n_od else "green", "down")]:
        note, ncls = moved(key, asat_dt, val, good, note, ncls)
        tiles1.append((ico, num(val), lab, note, ncls, spark_of(key, asat_dt, val)))
    # the RAG line: the rule lives in CONFIG and is printed beside the result;
    # the next action is dated from the pull, not from today's clock
    status, headline, card_headline = cal_rag(n_od, n_od_oh)
    rule = (f'Red when {num(CONFIG["rag_red_onhire"])} or more overdue assets are out on hire; '
            f'Amber when overdue assets are off hire only (shelf, repairs or not in SiteIQ); '
            f'Green when nothing is overdue. <b>Default line - set in CONFIG.</b>')
    action = (f'Chase list to each hirer on next touch; calibration run for the {num(n_av)} shelf '
              f'{plural(n_av, "item")} booked - by <b>{esc(S["action_due_s"])}</b>. Certificates '
              f'issued since {esc(maint_short)} entered in Register Entry.')
    band = sh.rag_band(status, headline, rule, esc(CONFIG["rag_owner"]), action)
    # the same facts in fewer words for the phone card, which draws a fixed
    # box: two headline lines, two action lines
    card_action = (f'Chase list to each hirer on next touch; the {num(n_av)} shelf '
                   f'{plural(n_av, "item")} to the calibrator - by {S["action_due_s"]}.')
    S["band"] = (status, card_headline, CONFIG["rag_owner"], card_action)
    # the same facts in words for the History scoreboard (extra) - one
    # status, never a second opinion
    S["rag_words"] = {"status": status, "headline": plain(headline), "rule": plain(rule),
                      "action": plain(action)}
    S["card_tiles"] = [(v, lab, re.sub(r"\s*-\s*page @@P:\w+@@", "", note), NOTE_HEX.get(ncls, "#8A9AAC"))
                       for _, v, lab, note, ncls, _ in tiles1] + \
                      [(v, lab, re.sub(r"\s*-\s*page @@P:\w+@@", "", note), NOTE_HEX.get(ncls, "#8A9AAC"))
                       for _, v, lab, note, ncls in (where_items[0], where_items[3])]
    mark("position")
    P.append(f"""<div class="pos">{banner}{pcallout(
        f'<span class="lead">The position.</span> As at <b>{esc(asat_s)}</b> - the SiteIQ pull - the register holds '
        f'<b>{num(total)} assets on the calibration register</b>: '
        f'<b class="g">{num(len(d["now_incal"]))} in calibration</b>, '
        f'<b class="a">{num(len(d["now_due30"]))} due inside 30 days</b>, '
        f'<b class="rd">{num(n_od)} overdue</b> and '
        f'<b>{num(len(d["nodate"]))} with no certificate date</b>. {cause}. '
        f'No Date means status unknown until the certificate is entered - page @@P:nodate@@.', False)}
<table class="two" style="margin-top:8px"><tr>
  <td style="width:31%"><div class="donut-wrap">
    {sh.donut(round(now_pct), sh.health_hex(round(now_pct)), f"{now_pct:.0f}%", "IN DATE")}
    <div class="donut-cap">Dated assets inside their due date at {esc(asat_day)}</div></div></td>
  <td style="padding-left:10px">{sh.score_rows([
        ("Dated fleet in date", round(now_pct),
         f"{num(len(d['now_dated_ok']))} of {num(dated)} assets holding a due date are inside it at "
         f"{asat_s} - computed from Calibration Due"),
        ("Register certified", round(dated / total * 100 if total else 0),
         f"{num(dated)} of {num(total)} assets carry a calibration due date - "
         f"the other {num(len(d['nodate']))} have no certificate entered yet"),
    ])}</td>
</tr></table>
{prowlab(f"At {esc(asat_s)} - computed from Calibration Due and the SiteIQ pull")}
{sh.tiles_plus(tiles1)}
{band}
{pnote(S["source_note"])}</div>""")

    # ---- P2 status mix, and the register's own view for comparison ------
    segs_now = [(lab, d["now"].get(lab, 0), col)
                for lab, col in STATUS_ORDER if d["now"].get(lab, 0)]
    if d["have_view"]:
        cmp_rows = []
        for lab, _ in STATUS_ORDER:
            a, b = d["now"].get(lab, 0), d["reg_status"].get(lab, 0)
            delta = a - b
            dcell = (f'<span class="rd">+{num(delta)}</span>' if delta > 0 and lab == "Overdue"
                     else f'<span class="g">+{num(delta)}</span>' if delta > 0
                     else f'{num(delta)}' if delta < 0 else dash())
            cmp_rows.append([esc(lab), num(a), num(b), dcell])
        cmp_rows.append(["<b>All assets on the register</b>", f"<b>{num(total)}</b>",
                         f"<b>{num(d['view_n'])}</b>", dash()])
        zero = len(d["reg_zero_day"])
        check = (f'Cross-check: applying the same thresholds at {esc(refresh_short)} reproduces the register&rsquo;s own status on all {num(d["view_n"])} of its lines (0 differences) - the rule used here is the register&rsquo;s rule.'
                 if d["rule_mismatch"] == 0 else
                 f'Cross-check: applying the same thresholds at {esc(refresh_short)} differs from the register&rsquo;s own status on {num(d["rule_mismatch"])} of {num(d["view_n"])} lines - read the comparison with that in mind.')
        view_words = (f'The register also carries its own status columns, last refreshed <b>{esc(S["refresh_s"])}</b>; they are printed '
                      f'here, labelled, so the two can be read side by side. Between the two dates nothing in the register changed; '
                      f'the calendar did: {num(len(d["reg_due30_lapsed"]))} of the register&rsquo;s '
                      f'{num(len(d["reg_due30"]))} &ldquo;due 0-30 days&rdquo; have since fallen due'
                      + (f', {num(zero)} of them printed as 0 days remaining on {esc(refresh_short)}' if zero else '')
                      + '.')
        body2 = f"""{psubh(f"Computed at {asat_day} against the register&rsquo;s own view at {refresh_short}", "- by status")}
{sh.dtable(["Status", f"At {asat_day} (computed)", f"Register's view, {refresh_short}", "Difference"], cmp_rows, ["", "r", "r", "r"], "cp")}
{pnote(f'Thresholds as the register applies them: overdue = past the due date; due 0-30 / 31-60 / 61-90 days; current = more than 90 days out; no date = no Calibration Due entered. In calibration = current + due 31-60 + due 61-90. {check} The register&rsquo;s Live Register sheet also pads {num(S["blanks"])} blank template rows (a status formula, no asset) - excluded.')}"""
    else:
        view_words = ("The register&rsquo;s own status columns (its Live Register sheet) were not found, so no "
                      "comparison to its own view is printed - nothing on this report needs them.")
        body2 = pnote("Thresholds as the register applies them: overdue = past the due date; due 0-30 / 31-60 / 61-90 days; current = more than 90 days out; no date = no Calibration Due entered.")
    mark("mix")
    P.append(f"""{psect(f"Status mix at {asat_day} - and the register&rsquo;s own view, for comparison")}
{pcallout(f'Every figure on this report is computed from the register&rsquo;s due dates and the SiteIQ pull of {esc(asat_s)} - nothing needs the workbook refreshed. {view_words}', False)}
{where_tiles}
{psubh(f"Status mix at {asat_day}", "- computed from Calibration Due, every asset on the register")}
{chartpanel(sh.stackband(segs_now))}
{body2}""")

    # ---- P3 fell due since the due dates were last maintained -----------
    lrows = []
    for r in d["lapsed"]:
        lrows.append([esc(r["asset"]) or dash(), esc(r["desc"]) or dash(), fmt_due(r),
                      fmt_days(r["now_days"], overdue=True), where_cell(r), who_cell(r)])
    l_where = Counter(r["where"] for r in d["lapsed"])
    l_words = ", ".join(
        f'{num(n)} {"on hire" if w == "On Hire" else "in store" if w == "Available for Hire" else "not in SiteIQ" if w == "not in SiteIQ" else w.lower()}'
        for w, n in l_where.most_common())
    win_end = (S["asat_d"] - timedelta(days=1)).strftime("%d %b %Y")
    mark("lapsed")
    P.extend(table_pages(
        f"Fell due since the due dates were last maintained - {maint_short} to {win_end}",
        pcallout(f'<b class="rd">{num(n_l)} {plural(n_l, "asset")}</b> passed the calibration due date between {esc(maint_short)}, when the register workbook was last saved, and the SiteIQ pull. {"It counts" if n_l == 1 else "They count"} in the {num(n_od)} overdue. A certificate issued in that window would not be on the register yet, so check the paperwork first: recalibrated means enter it in the Register Entry tab and this list shrinks on the next run; not recalibrated means a calibration run (in store) or a swap on next touch (on hire). Where SiteIQ has {"it" if n_l == 1 else "them"} at {esc(asat_s)}: {esc(l_words) if l_words else "nothing to place"}.', False),
        ["Asset", "Description", "Was due", "Over by", f"Where at {asat_day}", "Who has it"],
        ["", "", "r", "r", "c", ""], lrows, ROWS_CP,
        pnote(f'Over by = days past the due date at {esc(asat_s)}, computed. Where = SiteIQ RENTAL_STOCK status: &ldquo;In store&rdquo; is SiteIQ&rsquo;s &ldquo;Available for Hire&rdquo;; &ldquo;not in SiteIQ&rdquo; means the asset number is not in the pull at all - whereabouts unknown until it is found and scanned. Nothing here is estimated.')))

    # ---- P4 due inside 30 days ------------------------------------------
    d30 = d["now_due30"]
    drows = [[esc(r["asset"]) or dash(), esc(r["desc"]) or dash(), fmt_due(r),
              fmt_days(r["now_days"]), where_cell(r), who_cell(r)] for r in d30]
    recon = ""
    if d["have_view"]:
        recon = (f' For comparison, the register&rsquo;s own view at {esc(refresh_short)} had {num(len(d["reg_due30"]))} inside 30 days: '
                 f'{num(len(d["reg_due30_lapsed"]))} have since fallen due (page @@P:lapsed@@), {num(len(d["reg_due30_still"]))} '
                 f'{"remains" if len(d["reg_due30_still"]) == 1 else "remain"} inside 30 days, and {num(len(d["reg_due30_new"]))} '
                 f'{"has" if len(d["reg_due30_new"]) == 1 else "have"} moved in from 31-60 days - {num(len(d30))} today.')
    mark("due30")
    P.extend(table_pages(
        f"Due inside 30 days at {asat_day} - book these in now",
        pcallout(f'<b class="a">{num(len(d30))} {plural(len(d30), "asset")}</b> fall due inside 30 days of {esc(asat_day)}, soonest first, computed from Calibration Due. Booked in before the date lapses, these never touch the overdue list - that is the whole game. An asset out on hire gets its swap organised at the counter on next touch; the who-has-it column is SiteIQ&rsquo;s answer at {esc(asat_s)}.{recon}', False),
        ["Asset", "Description", "Calibration due", "Days left", f"Where at {asat_day}", "Who has it"],
        ["", "", "r", "r", "c", ""], drows, ROWS_STD,
        pnote('Days left are computed from Calibration Due at the SiteIQ pull time. A dash means the source holds no value for that cell - nothing here is estimated.'),
        cls=""))

    # ---- P5 the chase list ----------------------------------------------
    def chase_rows(groups):
        out = []
        for hirer, items in groups:
            for j, r in enumerate(sorted(items, key=lambda x: x["now_days"] or 0)):
                who = (f'<b>{esc(hirer)}</b><span class="s2">{esc(r["company"]) + " &middot; " if r["company"] else ""}'
                       f'{len(items)} overdue {plural(len(items), "item")}</span>' if j == 0 else "")
                out.append([who, esc(r["asset"]) or dash(), esc(r["desc"]) or dash(),
                            fmt_due(r), fmt_days(r["now_days"], overdue=True), out_cell(r)])
        return out
    n_h = sum(len(v) for _, v in d["now_chase_hirer"])
    n_r = sum(len(v) for _, v in d["now_chase_repairs"])
    n_after = len(d["issued_after"])
    after_words, after_detail = "", ""
    if n_after:
        bits = []
        for r in sorted(d["issued_after"], key=lambda x: x["out"]):
            gap = (r["out"].date() - r["due"].date()).days
            bits.append(f'<b>{esc(r["asset"])}</b> (due {esc(fmt_d(r["due"]))}) went out to '
                        f'{esc(r["live_hirer"])} on {esc(r["out"].strftime("%d %b %Y %H:%M"))}, '
                        f'{num(gap)} {plural(gap, "day")} after its due date')
        after_words = ("SiteIQ&rsquo;s hire start date is later than the calibration due date on "
                       f"<b>{num(n_after)} of the {num(n_od_oh)}</b> (red tag): either the item went out the counter "
                       "after its date lapsed, or the register&rsquo;s due date is wrong - check both; the note below names them.")
        after_detail = " Hire started after the due date: " + "; ".join(bits) + "."
    hdr5 = ["Who has it (SiteIQ)", "Asset", "Description", "Was due", "Overdue by", "Out since"]
    al5 = ["", "", "", "r", "r", ""]
    body5 = ""
    if d["now_chase_hirer"]:
        body5 += psubh("With a hirer", f"- {num(n_h)} {plural(n_h, 'item')} across {num(len(d['now_chase_hirer']))} {plural(len(d['now_chase_hirer']), 'name or account', 'names and accounts')}")
        body5 += sh.dtable(hdr5, chase_rows(d["now_chase_hirer"]), al5, "cp")
    if d["now_chase_repairs"]:
        body5 += psubh("On the repairs account", f"- {num(n_r)} {plural(n_r, 'item')} - a repair-queue conversation, not a chase")
        body5 += sh.dtable(hdr5, chase_rows(d["now_chase_repairs"]), al5, "cp")
    if not body5:
        body5 = pnote("Nothing overdue is on hire in SiteIQ at the pull time.")
    mark("chase")
    P.append(f"""{psect(f"Overdue and on hire at {asat_day} - the chase list, by name")}
{pcallout(f'<b class="rd">{num(n_od_oh)} overdue {plural(n_od_oh, "asset")} {isare(n_od_oh)} out on hire at {esc(asat_s)}</b> - register due dates joined to the SiteIQ RENTAL_STOCK pull (the register&rsquo;s own on-hire flag is page @@P:regchase@@). {num(n_h)} {isare(n_h)} with a hirer - a swap or recall on next touch - and {num(n_r)} {"sits" if n_r == 1 else "sit"} on the repairs account. Names A to Z, worst item first inside a name. {after_words}', False)}
{body5}
{pnote(f'Hirers A to Z, worst item first inside a name; overdue by = days past the due date at {esc(asat_s)}, computed. Out since = SiteIQ ON_HIRE_DATE for the current hire; the red tag marks a hire that started after the calibration due date. Every one of these is a counter conversation on next touch - swap organised, certificate sorted, no drama.' + after_detail)}""")

    # ---- P6 the register's own chase list - comparison ------------------
    if d["have_view"]:
        hdr6 = ["Who had it (register)", "Asset", "Description", "Was due", f"Over at {refresh_short}", f"Where at {asat_day}"]
        al6 = ["", "", "", "r", "r", "c"]

        def reg_rows(groups):
            out = []
            for hirer, items in groups:
                for j, r in enumerate(sorted(items, key=lambda x: x["reg"]["days"] or 0)):
                    who = (f'<b>{esc(hirer)}</b><span class="s2">{len(items)} '
                           f'overdue {plural(len(items), "item")}</span>' if j == 0 else "")
                    now = where_cell(r)
                    if r["where"] == "On Hire" and r["live_hirer"] and r["live_hirer"] != r["reg"]["hirer"]:
                        now += f'<span class="s2">now {esc(r["live_hirer"])}</span>'
                    out.append([who, esc(r["asset"]) or dash(), esc(r["desc"]) or dash(),
                                fmt_due(r), fmt_days(r["reg"]["days"], overdue=True), now])
            return out
        rh = sum(len(v) for _, v in d["reg_chase_hirer"])
        rr = sum(len(v) for _, v in d["reg_chase_repairs"])
        body6 = ""
        if d["reg_chase_hirer"]:
            body6 += psubh("With a hirer", f"- {num(rh)} {plural(rh, 'item')}")
            body6 += sh.dtable(hdr6, reg_rows(d["reg_chase_hirer"]), al6, "cp")
        if d["reg_chase_repairs"]:
            body6 += psubh("On the repairs account", f"- {num(rr)} {plural(rr, 'item')}")
            body6 += sh.dtable(hdr6, reg_rows(d["reg_chase_repairs"]), al6, "cp")
        if not body6:
            body6 = pnote("The register listed nothing overdue and on hire at its Last Refresh.")
        n_ro = len(d["reg_od_onhire"])
        mark("regchase")
        P.append(f"""{psect(f"The register&rsquo;s own chase list at {refresh_short} - for comparison")}
{pcallout(f'At its Last Refresh the register&rsquo;s own columns flagged <b>{num(n_ro)} overdue {plural(n_ro, "asset")} as on hire</b> - {num(rh)} with a hirer, {num(rr)} on the repairs account. Checked against SiteIQ at {esc(asat_s)}: <b>{num(len(d["reg_chase_back"]))}</b> {isare(len(d["reg_chase_back"]))} back in store and <b>{num(len(d["reg_chase_still"]))}</b> still out. Printed so the register&rsquo;s figure and the computed one can be read side by side; the list to work is page @@P:chase@@.', False)}
{body6}
{pnote(f'&ldquo;Over at {esc(refresh_short)}&rdquo; is the register&rsquo;s own Days Remaining at its Last Refresh, verbatim. &ldquo;Where at {esc(asat_day)}&rdquo; is SiteIQ RENTAL_STOCK at {esc(asat_s)}. An item back in store is still overdue - it moves to the calibration run on page @@P:shelf@@, not off the list.')}""")

    # ---- P7 overdue, not on hire ----------------------------------------
    unit_words = ", ".join(f'{esc(u)} {num(n)}' for u, n in d["now_od_avail_units"].most_common(4))
    grows = [[esc(desc), esc(unit), num(len(assets)),
              esc(", ".join(sorted(assets)))] for (desc, unit), assets in d["now_od_avail_groups"]]
    mrows = [[esc(r["asset"]) or dash(), esc(r["desc"]) or dash(), fmt_due(r),
              fmt_days(r["now_days"], overdue=True),
              ((esc(r["reg"]["unit"]) if r["reg"]["unit"] else dash())
               + f'<span class="s2">register on hire: {esc(r["reg"]["onhire"] or "blank")}</span>') if r["reg"] else dash()]
             for r in d["now_od_missing"]]
    orows = [[esc(r["asset"]) or dash(), esc(r["desc"]) or dash(), fmt_due(r),
              fmt_days(r["now_days"], overdue=True), esc(r["where"])] for r in d["now_od_other"]]
    lead7 = pcallout(
        f'<b>{num(n_av + n_mi + n_ot)} of the {num(n_od)} overdue assets are not on hire</b> at {esc(asat_s)}. '
        f'<b>{num(n_av)}</b> {isare(n_av)} on the shelf in SiteIQ (&ldquo;Available for Hire&rdquo;'
        + (f' - {unit_words}' if unit_words else "") + ') - a calibration run, not a chase: batch them to the calibrator and each comes back onto the dated fleet. '
        f'<b>{num(n_mi)}</b> {isare(n_mi)} <b>not in the SiteIQ RENTAL_STOCK at all - whereabouts unknown</b> until '
        f'{"it is" if n_mi == 1 else "they are"} found and scanned.'
        + (f' {num(n_ot)} {"carries" if n_ot == 1 else "carry"} another SiteIQ status - listed as written.' if n_ot else ""), False)
    note7 = pnote(f'Shelf list grouped by description and SiteIQ bay so a calibration run can be batched; every asset number is printed. Overdue by = days past the due date at {esc(asat_s)}, computed. Not-in-SiteIQ rows show the register&rsquo;s own bay and on-hire flag at {esc(refresh_short)} where it has one - nothing newer exists for them.')
    tail7 = ""
    if mrows:
        tail7 += psubh("Not in SiteIQ - whereabouts unknown", f"- {num(n_mi)} {plural(n_mi, 'asset')}")
        tail7 += sh.dtable(["Asset", "Description", "Was due", "Overdue by", f"Register's own bay, {refresh_short}"], mrows, ["", "", "r", "r", ""], "cp")
    if orows:
        tail7 += psubh("Another SiteIQ status", f"- {num(n_ot)} {plural(n_ot, 'asset')}")
        tail7 += sh.dtable(["Asset", "Description", "Was due", "Overdue by", "SiteIQ status"], orows, ["", "", "r", "r", ""], "cp")
    shelf_h = ["Description", "SiteIQ bay", "Qty", "Assets"]
    shelf_a = ["", "", "r", ""]
    # the shelf table and the not-in-SiteIQ table share a page when they fit;
    # otherwise the shelf list paginates and the rest gets a page of its own
    shelf_sub = psubh("On the shelf in SiteIQ", f"- {num(n_av)} {plural(n_av, 'asset')} in {num(len(grows))} {plural(len(grows), 'line')}")
    mark("shelf")
    marks["missing"] = marks["shelf"]
    if len(grows) + len(mrows) + len(orows) <= 20:
        P.append(f"""{psect(f"Overdue, not on hire at {asat_day} - the calibration run")}
{lead7}
{shelf_sub}
{sh.dtable(shelf_h, grows, shelf_a, "cp") if grows else pnote("Nothing overdue is on the shelf in SiteIQ at the pull time.")}
{tail7}
{note7}""")
    else:
        P.extend(table_pages(f"Overdue, not on hire at {asat_day} - the calibration run", lead7 + shelf_sub,
                             shelf_h, shelf_a, grows, ROWS_CP, note7))
        if tail7:
            mark("missing")
            P.append(f"""{psect(f"Overdue, not on hire at {asat_day} - not in SiteIQ")}
{tail7}
{note7}""")

    # ---- P8 the No Date position ----------------------------------------
    nd = d["nodate"]
    nd_named = [r for r in nd if r["desc"]]
    units_top = d["nd_units"].most_common(10)
    fam_top = d["nd_family"].most_common(7)
    frows = [[esc(f), num(n), num(d["nd_onhire_family"].get(f, 0)) if d["nd_onhire_family"].get(f, 0) else dash()]
             for f, n in fam_top]
    nd_unnamed = len(nd) - len(nd_named)
    unnamed_note = (f'{num(nd_unnamed)} No Date lines carry an asset number '
                    f'but no description yet - counted above, family unknown '
                    f'until the register line is completed. '
                    if nd_unnamed else '')
    n_ndl = len(d["nd_onhire_live"])
    nd_fam_words = ", ".join(f'{esc(f.lower())} {num(n)}' for f, n in d["nd_onhire_family"].most_common(4))
    mark("nodate")
    P.append(f"""{psect("No Date - calibration status unknown until the certificate is entered")}
{pcallout(f'<b>{num(len(nd))} assets</b> sit on the register with no calibration due date entered. Say it straight: their <b>calibration status is unknown</b> - the register cannot say whether they are in date, and neither can this report, until the certificate details are entered. They are not counted as failed and not counted as in date. <b class="a">{num(n_ndl)} of them {isare(n_ndl)} out on hire at {esc(asat_s)}</b>' + (f' ({nd_fam_words})' if nd_fam_words else "") + f'. {num(len(d["nd_missing"]))} {isare(len(d["nd_missing"]))} not in SiteIQ at all. {num(d["nd_serial"])} {"carries" if d["nd_serial"] == 1 else "carry"} a serial number. Each certificate entered moves an asset from this page onto the dated fleet on page @@P:position@@ - that is the fill-in work under way, and the on-hire ones come first.', False)}
{psubh("Where they sit", f"- No Date assets by SiteIQ bay at {asat_day}, top {len(units_top)} of {len(d['nd_units'])}")}
{chartpanel(sh.hbars([(u, n) for u, n in units_top], colour=K["blue"]))}
{psubh("What they are", f"- by item family, top {len(fam_top)} of {len(d['nd_family'])}")}
{sh.dtable(["Item family (first words of description)", "Assets", f"On hire at {asat_day}"], frows, ["", "r", "r"], "cp")}
{pnote(unnamed_note + 'Family = the first words of the register description - a plain grouping, not a new classification. Bay = SiteIQ STORAGE_UNIT at the pull; an asset SiteIQ does not hold is counted under &ldquo;not in SiteIQ&rdquo;. Nothing is guessed.')}""")

    # ---- P9 on hire, not in the register - computed from SiteIQ ---------
    n_nir, n_kw = len(d["nir"]), len(d["kw"])
    famrows = d["nir_family"].most_common()
    krows = []
    for kind, items in d["kw_kinds"]:
        hirers = Counter(l["hirer"] or "(no hirer recorded)" for l in items)
        ex = Counter(" ".join(l["desc"].split()) for l in items).most_common(1)
        krows.append([esc(kind), num(len(items)), num(len(hirers)),
                      esc("; ".join(f'{e} ({n})' for e, n in ex))])
    audit_words = (f' (The register&rsquo;s own sweep - its On Hire Audit sheet - listed {num(d["audit_n"])} at its Last Refresh, {esc(refresh_short)}: a different day and a different on-hire fleet, printed for comparison only.)'
                   if d["audit_n"] else "")
    audit_words = audit_words.replace("a different day and a different on-hire fleet, printed for comparison only", "a different day, printed for comparison only")
    mark("nir")
    P.append(f"""{psect("On hire, not in the register - computed from the SiteIQ pull")}
{pcallout(f'<b>{num(n_nir)} SiteIQ lines</b> are on hire at {esc(asat_s)} with no line on the calibration register - {num(d["live_onhire"])} on hire in the pull, less the {num(len(d["onhire_live"]))} register assets among them - across {num(d["nir_hirers"])} hirers.{audit_words} By construction that is the whole on-hire fleet less what the register holds; radios, gas monitors, phones and general tooling make up most of it and need no calibration certificate. The slice worth a look is the <b>{num(n_kw)} rows</b> whose description reads like test or measuring gear - a keyword match, listed from page @@P:kw@@.', False)}
{psubh("What is on hire unregistered", f"- {num(n_nir)} rows by item family")}
{chartpanel(sh.hbars([(f, n, num(n)) for f, n in famrows], colour=K["orange"]))}
{psubh("Looks like calibration-type gear", f"- keyword match, {num(n_kw)} rows across {num(d['kw_hirers'])} {plural(d['kw_hirers'], 'hirer')} - verify before acting")}
{sh.dtable(["Kind", "Rows", "Hirers", "Most common description"], krows, ["", "r", "r", ""], "cp") if krows else pnote("No row on hire matches the keyword rule.")}
{pnote(f'Keyword rule: the SiteIQ description contains any of <b>{esc(CAL_RULE_WORDS)}</b>, and none of Dr&auml;ger / X-am / gas monitor, sling, chain block. Excluded and counted separately: {num(d["kw_gas"])} gas-monitor rows (their own programme and report) and {num(d["kw_lift"])} sling and chain-block rows ({num(d["kw_chain"])} chain blocks - lifting gear, the Rigging &amp; Lifting Register report). A match is a prompt to check, not a finding.')}""")

    # ---- P10+ the keyword slice, every row ------------------------------
    kwrows = [[esc(desc), esc(hirer) + (f" ({esc(co)})" if co else ""),
               num(len(bcs)), esc(", ".join(sorted(bcs)))]
              for (kind, desc, hirer, co), bcs in d["kw_groups"]]
    mark("kw")
    P.extend(table_pages(
        "Looks like calibration-type gear, on hire and not in the register",
        pcallout(f'<b>{num(n_kw)} rows</b> on hire at {esc(asat_s)} match the keyword rule on page @@P:nir@@, grouped by description and hirer - {num(len(kwrows))} lines, every barcode printed, in the order of the kinds table on that page. Work it like the register&rsquo;s dashboard says: add it to the register or confirm it needs no calibration certificate. <b>Keyword match, verify before acting</b> - a Fluke connector or a hand pump matches the word, not the need.', False),
        ["SiteIQ description", "Who has it (company)", "Qty", "Barcodes"],
        ["", "", "r", ""], kwrows, ROWS_KW,
        pnote(f'Source: SiteIQ RENTAL_STOCK at {esc(asat_s)} - description, hirer and company as SiteIQ has them. Nothing here is a compliance finding until the item is checked.')))

    # ---- the trend - only once seven days are on record ------------------
    tp = trend_page(asat_dt, d)
    if tp:
        mark("trend")
        P.append(tp)

    # ---- close --------------------------------------------------------
    cards = sh.info_cards([
        ("In date or it stays in",
         "The standard: nothing goes out the counter with a lapsed "
         "certificate - Life Saving Rule 5, Tools and Equipment (SEQ-GL-009). "
         "This report checks the standard against the live SiteIQ pull and "
         "says so plainly when a hire started after the due date."),
        ("Computed, never refreshed",
         "The register is the human-maintained list - asset, description, "
         "certificate, due date. Everything else on these pages is computed "
         "from the SiteIQ pull at the time printed on every page. No "
         "workbook refresh, ever."),
        ("Names and accounts, not categories",
         "An overdue asset on hire is chased with whoever SiteIQ says holds "
         "it - a person, a project account or the repairs queue - and each "
         "kind gets its own conversation at the counter."),
        ("Honest about the gaps",
         "No Date means the calibration status is unknown until the "
         "certificate is entered - said plainly, counted separately, never "
         "blended into a compliance score. Blanks print as dashes, "
         "<b>never guesses</b>."),
    ])
    former_live = d["former_live"]
    n_days = history_days("calibration", asat_dt)
    trend_line = (f'Trend page: appears once seven days are on record ({num(n_days)} today).'
                  if n_days < 7 else
                  f'Trend page: {num(n_days)} days on record - the 30-day lines are on page @@P:trend@@.')
    P.append(f"""{psect("How the calibrated fleet is run")}
{cards}
{pnote(f'Names as shown: the site is Ampol. SiteIQ still carries the site&rsquo;s former name on {num(former_live)} live-register lines and the register on some descriptions; every one is shown here under the current name. Asset numbers and barcodes are identifiers and never change. Each run writes its figures to {HIST_NAME} keyed on the pull day; the next run reads them back for the movement notes. {trend_line}')}
{sh.coates_way_panel()}
{psect("Meet the tool store team")}
{pnote(f'The crew running your {esc(CONFIG["client"])} store - keeping the register true, the certificates current and the gear ready. Something not right? Tell us and we&rsquo;ll sort it.')}
{sh.team_cards(CONFIG["team"])}""")
    # fill in every cross-reference from the pagination that actually happened
    P = [re.sub(r"@@P:(\w+)@@", lambda m: str(marks[m.group(1)]), p) for p in P]
    return P, marks


# extra styling this report needs on top of the shared sheet - the
# staleness banner, the small row labels above the two tile rows, and the
# tighter tiles on the position page (class "pos" - it sits behind the
# cover, so it is no longer the shell's page1)
EXTRA_CSS = """
.stale { background:#FDE8E8; border-left:4px solid #DC2626; border-radius:0 10px 10px 0;
         padding:9px 18px; margin-top:12px; font-size:10.4px; line-height:1.65; color:#7F1D1D; }
.stale b { color:#B91C1C; font-weight:700; }
.stale .h { display:block; color:#B91C1C; font-weight:700; text-transform:uppercase;
            letter-spacing:1.4px; font-size:8.6px; margin-bottom:3px; }
.rowlab { margin:9px 0 -5px 0; font-size:8.4px; font-weight:700; letter-spacing:1.5px;
          text-transform:uppercase; color:#5A6875; }
.pos .tiles td { padding:10px 9px 9px 9px; }
.pos .tiles { margin-top:7px; }
.pos .note { margin-top:7px; }
"""


def render_doc(pages, cover, gen_s, asat_s):
    """The cover is page 1 of the pack; the position page is page 2 and
    every page number and cross-reference counts from there."""
    tot = len(pages) + COVER_PAGES
    body = cover + "".join(sh.render_page(CONFIG, p, i + 1 + COVER_PAGES, tot, gen_s, asat_s)
                           for i, p in enumerate(pages))
    doc = (f'<!doctype html><html lang="en"><head><meta charset="utf-8">'
           f'<title>Coates {esc(CONFIG["client"])} {esc(CONFIG["title"])} - '
           f'{esc(asat_s)}</title><style>{EXTRA_CSS}</style></head><body>{body}</body></html>')
    return doc, tot


# =====================================================================
# the Outlook email (draft - never sends)
# =====================================================================

def card_block(cid):
    """The position card inline, under the header: a cid image the .eml
    carries in its own related part. Outlook honours the width attribute;
    the style caps it for everything else."""
    FONT = sh.FONT
    return (f'<table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr>'
            f'<td align="center" style="padding:16px 0 2px 0;">'
            f'<img src="cid:{cid}" width="420" alt="The position on one card" '
            f'style="display:block;width:420px;max-width:420px;height:auto;border:0;">'
            f'<div style="{FONT}font-size:10px;color:#7A8A9A;padding-top:6px;">The position on one card - '
            f'the same figures as the position page of the PDF; the PNG is attached for your phone.</div>'
            f'</td></tr></table>')


def build_email_html(d, gen_s, S, marks, card_cid=""):
    W = 1000
    FONT = sh.FONT
    total = d["total"]
    dated = len(d["dated"])
    asat_s, asat_day, maint_short = S["asat_s"], S["asat_day"], S["maint_short"]
    refresh_short = S["refresh_short"]
    now_pct = len(d["now_dated_ok"]) / dated * 100 if dated else 0
    n_od, n_od_oh = len(d["now_overdue"]), len(d["now_od_onhire"])
    n_av, n_mi = len(d["now_od_avail"]), len(d["now_od_missing"])
    n_l = len(d["lapsed"])
    parts = []

    parts.append(f"""<table role="presentation" width="100%" cellspacing="0" cellpadding="0">
<tr><td bgcolor="#1A2430" style="padding:22px 24px 19px 24px;">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr>
<td>
<div style="{FONT}font-size:11px;font-weight:bold;letter-spacing:2.5px;color:#F36F21;text-transform:uppercase;">{esc(CONFIG['kicker'])}</div>
<div style="{FONT}font-size:26px;font-weight:bold;color:#FFFFFF;padding-top:8px;">Ampol Calibration Register - The Position</div>
<div style="{FONT}font-size:13px;color:#A7B6C4;padding-top:7px;">{esc(CONFIG['project'])}</div>
</td>
<td width="185" align="right" style="vertical-align:top;">
<div style="{FONT}font-size:11px;font-weight:bold;letter-spacing:1.5px;color:#FFFFFF;">POWERED BY <span style="color:#F36F21;">SITEIQ</span></div>
<div style="{FONT}font-size:11.5px;color:#8395A6;padding-top:5px;">Equipped for anything</div>
</td></tr></table>
<div style="{FONT}font-size:11px;color:#8395A6;padding-top:10px;line-height:1.6;">Generated: <b style="color:#FFFFFF;">{esc(gen_s)}</b> &nbsp;|&nbsp; Data as at: <b style="color:#FFFFFF;">{esc(asat_s)}</b> (SiteIQ RENTAL_STOCK request time) &nbsp;|&nbsp; Due dates last maintained: <b style="color:#FFFFFF;">{esc(S['maint_s'])}</b> &nbsp;|&nbsp; Author: <b style="color:#FFFFFF;">Andrew Fisher</b></div>
</td></tr></table>""")
    if card_cid:
        parts.append(card_block(card_cid))

    if S["stale_days"] > STALE_DAYS:
        parts.append(f"""<table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr>
<td style="padding:16px 0 0 0;"><table role="presentation" width="100%" cellspacing="0" cellpadding="0">
<tr><td width="5" bgcolor="#DC2626" style="font-size:0;">&nbsp;</td>
<td bgcolor="#FDE8E8" style="{FONT}padding:13px 20px;font-size:12.5px;line-height:1.7;color:#7F1D1D;">
<div style="{FONT}font-size:10px;font-weight:bold;letter-spacing:1.5px;color:#B91C1C;text-transform:uppercase;padding-bottom:3px;">Due dates are {num(S['stale_days'])} days old - check the paperwork before sending</div>
Due dates last maintained <b style="color:#B91C1C;">{esc(maint_short)}</b>, <b style="color:#B91C1C;">{num(S['stale_days'])} days</b> before the SiteIQ pull this report is built on. Any certificate issued since then is not on the register yet - {num(n_l)} {"asset" if n_l == 1 else "assets"} fell due in that window. New certificates go in the Register Entry tab; save and run again. Nothing needs the workbook refreshed - every status and location here is computed from SiteIQ.</td>
</tr></table></td></tr></table>""")

    reg_words = ""
    if d["have_view"]:
        reg_words = (f' For comparison, the register&rsquo;s own status columns at their Last Refresh ({esc(refresh_short)}) '
                     f'said {num(len(d["reg_incal"]))} in calibration, {num(len(d["reg_due30"]))} due inside 30 days, '
                     f'{num(len(d["reg_overdue"]))} overdue ({num(len(d["reg_od_onhire"]))} on hire) - page {marks["mix"]} of the PDF.')
    parts.append(sh.ecallout(
        f'<span style="color:#D95F14;font-weight:bold;text-transform:uppercase;">'
        f'The position.</span> As at <b>{esc(asat_s)}</b> (the SiteIQ pull): '
        f'{sh.eo(num(total) + " assets on the calibration register")} - '
        f'<b>{num(len(d["now_incal"]))}</b> in calibration, '
        f'{sh.eo(num(len(d["now_due30"])) + " due inside 30 days")}, '
        f'{sh.eo(num(n_od) + " overdue")} '
        f'({num(n_od_oh)} on hire - the chase list below; {num(n_av)} on the shelf; {num(n_mi)} not in SiteIQ; '
        f'{num(n_l)} fell due since the due dates were last maintained on {esc(maint_short)}) and <b>{num(len(d["nodate"]))}</b> '
        f'with no certificate date - calibration status unknown until it is entered. Status is computed from each '
        f'asset&rsquo;s Calibration Due; where it is and who has it comes from SiteIQ.{reg_words} Full detail attached.'))

    parts.append(sh.esubh(f"At {esc(asat_s)}", "- computed from Calibration Due and the SiteIQ pull"))
    parts.append(sh.etiles([
        (num(total), "ASSETS ON THE REGISTER", "the human-maintained list", "#8A9AAC"),
        (num(len(d["now_incal"])), "IN CALIBRATION", "current, due 31+ days", "#22C55E"),
        (num(len(d["now_due30"])), "DUE INSIDE 30 DAYS",
         "book them in now", "#EFA82B" if d["now_due30"] else "#22C55E"),
        (num(n_od), "OVERDUE",
         f'{num(n_l)} fell due since {esc(maint_short)}',
         "#F0603E" if d["now_overdue"] else "#22C55E"),
    ]))
    parts.append(sh.esubh("Where the overdue and No Date gear is", "- SiteIQ RENTAL_STOCK"))
    parts.append(sh.etiles([
        (num(n_od_oh), "OVERDUE, ON HIRE", "the chase list", "#F0603E" if n_od_oh else "#22C55E"),
        (num(n_av), "OVERDUE, ON THE SHELF", "the calibration run", "#EFA82B" if n_av else "#22C55E"),
        (num(n_mi), "OVERDUE, NOT IN SITEIQ", "whereabouts unknown", "#F0603E" if n_mi else "#22C55E"),
        (num(len(d["nodate"])), "NO DATE",
         f'{num(len(d["nd_onhire_live"]))} of them out on hire', "#EFA82B"),
    ]))

    parts.append(f"""<table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="margin-top:16px;"><tr>
<td width="200" align="center" style="vertical-align:top;padding-top:6px;">
{sh.donut_png(round(now_pct), sh.health_hex(round(now_pct)), f"{now_pct:.0f}%", "IN DATE")}
<div style="{FONT}font-size:10.5px;color:#98A6B4;padding-top:8px;">Dated assets inside their due date at {esc(asat_day)}</div></td>
<td style="vertical-align:top;padding-left:16px;">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0">
{sh.score_bar_row("Dated fleet in date", round(now_pct), f"{num(len(d['now_dated_ok']))} of {num(dated)} assets holding a due date are inside it at {esc(asat_s)} - computed")}
{sh.score_bar_row("Register certified", round(dated / total * 100 if total else 0), f"{num(dated)} of {num(total)} assets carry a due date - the rest have no certificate entered yet, status unknown")}
</table></td></tr></table>""")

    cap = 8
    chase_rows = []
    for hirer, items in d["now_chase_hirer"] + d["now_chase_repairs"]:
        for r in sorted(items, key=lambda x: x["now_days"] or 0):
            out = r["out"].strftime("%d %b %Y") if r["out"] else "&ndash;"
            if r["issued_after_due"]:
                out += ' <span style="color:#C81E1E;font-weight:bold;font-size:10px;">AFTER DUE</span>'
            chase_rows.append([esc(hirer) + (' <span style="color:#8A9AAC;font-size:10px;">repairs</span>' if r["repairs_now"] else ""),
                               esc(r["asset"]) or "&ndash;",
                               esc(r["desc"]) or "&ndash;",
                               r["due"].strftime("%d %b %Y") if r["due"] else "&ndash;",
                               f'{num(abs(r["now_days"]))}d' if r["now_days"] is not None else "&ndash;",
                               out])
    shown = chase_rows[:cap]
    parts.append(sh.esect(f"Overdue and on hire at {esc(asat_day)} - the chase list, by name"))
    parts.append(sh.enote(f'Register due dates joined to SiteIQ RENTAL_STOCK at {esc(asat_s)}. '
                          f'{num(sum(len(v) for _, v in d["now_chase_hirer"]))} with a hirer, '
                          f'{num(sum(len(v) for _, v in d["now_chase_repairs"]))} on the repairs account. '
                          f'AFTER DUE = SiteIQ hire start later than the calibration due date - check the item and the register date.'))
    parts.append(sh.edtable(["Who has it", "Asset", "Description", "Was due", "Over by", "Out since"],
                            shown, ["", "", "", "r", "r", ""]))
    if len(chase_rows) > cap:
        parts.append(sh.enote(f'Showing {cap} of {len(chase_rows)} - the PDF '
                              f'attached carries the full list.'))
    if not chase_rows:
        parts.append(sh.enote("Nothing overdue is on hire in SiteIQ at the pull time."))

    cap2 = 8
    d30 = d["now_due30"]
    drows = [[esc(r["asset"]) or "&ndash;", esc(r["desc"]) or "&ndash;",
              r["due"].strftime("%d %b %Y") if r["due"] else "&ndash;",
              f'{num(r["now_days"])}d' if r["now_days"] is not None else "&ndash;",
              ("On hire" if r["where"] == "On Hire" else "In store" if r["where"] == "Available for Hire" else esc(r["where"])),
              esc(r["live_hirer"]) if r["where"] == "On Hire" and r["live_hirer"] else "&ndash;"]
             for r in d30[:cap2]]
    parts.append(sh.esect(f"Due inside 30 days at {esc(asat_day)} - soonest first"))
    parts.append(sh.edtable(["Asset", "Description", "Due", "Days left", "Where", "Who has it"],
                            drows, ["", "", "r", "r", "c", ""]))
    if len(d30) > cap2:
        parts.append(sh.enote(f'Showing {cap2} of {len(d30)} - full list in the PDF.'))

    parts.append(sh.enote(
        f'{num(len(d["nodate"]))} assets carry no certificate date - calibration status unknown until it is '
        f'entered; {num(len(d["nd_onhire_live"]))} of them are out on hire at {esc(asat_day)}. '
        f'The PDF gives that position a page of its own. {num(len(d["nir"]))} SiteIQ lines are on hire and not on '
        f'the calibration register at {esc(asat_s)} - mostly radios, gas monitors, phones and general tooling; '
        f'{num(len(d["kw"]))} rows read like test or measuring gear (keyword match, verify before acting) - '
        f'pages {marks["nir"]} and {marks["kw"]} of the PDF.'))

    team_line = " &middot; ".join(
        f'<b style="color:#16202C;">{esc(p["name"])}</b> '
        f'<span style="color:#8A9AAC;">{esc(p["role"])}</span>'
        for p in CONFIG["team"])
    parts.append(f"""<table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="margin-top:26px;">
<tr><td style="border-top:1px solid #E4E8EC;padding-top:10px;">
<div style="{FONT}font-size:10px;font-weight:bold;letter-spacing:2px;color:#F36F21;text-transform:uppercase;">Your Coates Tool Store Team</div>
<div style="{FONT}font-size:11px;color:#8A9AAC;padding-top:5px;line-height:1.7;">{team_line}</div>
<div style="{FONT}font-size:10px;color:#98A6B4;padding-top:9px;line-height:1.7;">
Coates Hire &middot; {S['source_note']} {S['view_words']} The Coates Way - consistent execution, every day. <b style="color:#16202C;">POWERED BY SITEIQ</b></div>
</td></tr></table>""")

    body = "".join(
        f'<tr><td style="padding:0 24px;">{p}</td></tr>'
        if "bgcolor=\"#1A2430\"" not in p[:120] else f'<tr><td>{p}</td></tr>'
        for p in parts)
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<title>Ampol Calibration Register - {esc(asat_day)}</title></head>
<body bgcolor="#EEF1F4" style="margin:0;padding:0;background-color:#EEF1F4;">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0" bgcolor="#EEF1F4">
<tr><td align="center" style="padding:18px 10px;">
<table role="presentation" width="{W}" cellspacing="0" cellpadding="0" bgcolor="#FFFFFF" style="width:{W}px;max-width:{W}px;">
<tr><td height="6" bgcolor="#F36F21" style="font-size:0;">&nbsp;</td></tr>
{body}
<tr><td height="24" style="font-size:0;">&nbsp;</td></tr>
</table></td></tr></table></body></html>"""


# =====================================================================
# MAIN
# =====================================================================

def main():
    print("=" * 68)
    print("COATES CALIBRATION REGISTER - HOUSE-STYLE REPORT (the K2 look)")
    print("=" * 68)
    src = eng.find_workbook(["Ampol_Calibration_Register*.xlsx"],
                            sys.argv[1] if len(sys.argv) > 1 else None)
    if not src:
        sys.exit("ERROR: no Ampol_Calibration_Register*.xlsx in the suite's "
                 "Data folder - save the register there and run again.")
    rs_path = eng.find_workbook(["RENTAL_STOCK*.xlsx"],
                                sys.argv[2] if len(sys.argv) > 2 else None)
    if not rs_path:
        sys.exit("ERROR: no RENTAL_STOCK*.xlsx in the suite's Data folder - "
                 "this report is computed from the live SiteIQ pull. "
                 "Run 12_PULL_SITEIQ_EXPORTS (or save the export there) and run again.")
    print(f"Calibration register : {src}")
    print(f"SiteIQ RENTAL_STOCK  : {rs_path}")

    rows, entry_sheet = load_entry(src)
    view, blanks, refresh = load_register_view(src)
    audit_n = load_audit_count(src)
    live, pulled, rs_rows = load_rental_stock(rs_path)

    # ---- the stamps -------------------------------------------------------
    # data as at = the SiteIQ pull's own request time
    if pulled is None:
        pulled = datetime.fromtimestamp(os.path.getmtime(rs_path))
        pulled_how = "file time on disk - no REFERENCE_INFO request time in the export"
        print(f"NOTE: RENTAL_STOCK has no REFERENCE_INFO request time - using the "
              f"file's time on disk ({pulled:%d %b %Y %H:%M}) as the data-as-at stamp.")
    else:
        pulled_how = "SiteIQ request time"
    asat_d = pulled.date()
    asat_s = pulled.strftime("%d %b %Y %H:%M")
    asat_day = pulled.strftime("%d %b %Y")
    # due dates last maintained = when the register workbook was last saved
    maint, maint_how = workbook_saved(src)
    if maint_how == "file time on disk" and refresh is not None:
        maint, maint_how = refresh, "register Last Refresh - no file properties"
    maint_s = maint.strftime("%d %b %Y %H:%M")
    maint_short = maint.strftime("%d %b %Y")
    refresh_s = refresh.strftime("%d %b %Y %H:%M") if refresh else "(none)"
    refresh_short = refresh.strftime("%d %b %Y") if refresh else "(none)"
    gen_dt = datetime.now()
    gen_s = gen_dt.strftime("%d %b %Y %H:%M")
    stale_days = (asat_d - maint.date()).days
    d = derive(rows, view, live, asat_d, maint.date(), refresh, audit_n)

    view_words = ""
    if refresh is not None:
        gap = maint - refresh
        view_words = (f"The register&rsquo;s own status columns were last refreshed {esc(refresh_s)}"
                      + (f" - {abs(gap).days} days {'before' if gap > timedelta(0) else 'after'} the last save, so the two are not the same moment"
                         if abs(gap) > timedelta(hours=1) else "")
                      + "; they appear only as a labelled comparison.")
    source_note = (f"Data as at <b>{esc(asat_s)}</b> - RENTAL_STOCK.xlsx, {esc(pulled_how)} "
                   f"({num(rs_rows)} lines, {num(d['live_onhire'])} on hire). Register: "
                   f"Ampol_Calibration_Register.xlsx, &lsquo;{esc(entry_sheet)}&rsquo; sheet - the human-maintained "
                   f"list; due dates last maintained <b>{esc(maint_s)}</b> ({esc(maint_how)}). Status at the pull "
                   f"time is computed from Calibration Due; location, hirer and hire start are SiteIQ&rsquo;s. "
                   f"Nothing here needs the workbook refreshed.")
    # the next action is dated from the pull day - the wall clock never sets it
    action_due_s = (asat_d + timedelta(days=CONFIG["action_days"])).strftime("%d %b %Y")
    S = {"asat_s": asat_s, "asat_day": asat_day, "asat_d": asat_d, "asat_dt": pulled,
         "maint_s": maint_s, "maint_short": maint_short, "refresh_s": refresh_s,
         "refresh_short": refresh_short, "stale_days": stale_days, "blanks": blanks,
         "source_note": source_note, "view_words": view_words,
         "action_due_s": action_due_s}
    # the shared page shell prints this after the as-at time on page 1
    CONFIG["asat_note"] = (f"(SiteIQ RENTAL_STOCK request time · due dates last maintained "
                           f"{maint_short}, {stale_days} days earlier)")

    print(f"Data as at           : {asat_s}  ({pulled_how}; {rs_rows:,} lines, "
          f"{d['live_onhire']:,} on hire)")
    print(f"Due dates maintained : {maint_s}  ({maint_how}; {stale_days} days before the pull)")
    if refresh is not None:
        print(f"Register's own view  : Last Refresh {refresh_s}  (comparison only)")
    if stale_days > STALE_DAYS:
        print("*" * 68)
        print(f"WARNING: due dates last maintained {maint_short}, {stale_days} days before "
              f"the SiteIQ pull.")
        print(f"         {len(d['lapsed']):,} assets fell due in that window - a certificate issued")
        print("         since then is not on the register yet. Check the paperwork; new")
        print("         certificates go in the Register Entry tab. No refresh needed.")
        print("*" * 68)
    print(f"Assets on register   : {d['total']:,}  ('{entry_sheet}' sheet"
          + (f"; {d['dup_assets']:,} duplicate asset numbers" if d["dup_assets"] else "")
          + (f"; Live Register pads {blanks:,} blank template rows, excluded" if blanks else "") + ")")
    print(f"At {asat_d:%d %b} (computed): in-cal {len(d['now_incal']):,} | "
          f"due-30 {len(d['now_due30']):,} | overdue {len(d['now_overdue']):,} | "
          f"no-date {len(d['nodate']):,}  ({len(d['lapsed']):,} fell due since {maint_short})")
    if d["have_view"]:
        print(f"Register's view {refresh.strftime('%d %b')}: in-cal {len(d['reg_incal']):,} | "
              f"due-30 {len(d['reg_due30']):,} | overdue {len(d['reg_overdue']):,} | "
              f"no-date {len(d['reg_nodate']):,}  (its own status; comparison only)")
        print(f"Status rule check    : {d['rule_mismatch']:,} of {d['view_n']:,} lines differ from the "
              f"register's own status at its refresh; {d['view_due_mismatch']:,} due dates differ "
              f"between the two sheets")
    print(f"Chase list           : {len(d['now_od_onhire']):,} overdue on hire in SiteIQ across "
          f"{len(d['now_chase_hirer']) + len(d['now_chase_repairs']):,} hirers "
          f"({sum(len(v) for _, v in d['now_chase_repairs']):,} on the repairs account; "
          f"{len(d['issued_after']):,} hired after the due date)")
    if d["have_view"]:
        print(f"Register's chase list: {len(d['reg_od_onhire']):,} overdue on hire at "
              f"{refresh.strftime('%d %b')} across "
              f"{len(d['reg_chase_hirer']) + len(d['reg_chase_repairs']):,} hirers "
              f"({len(d['reg_chase_back']):,} since back in store)")
    print(f"Overdue not on hire  : {len(d['now_od_avail']):,} on the shelf in SiteIQ | "
          f"{len(d['now_od_missing']):,} not in SiteIQ"
          + (f" | {len(d['now_od_other']):,} other SiteIQ status" if d['now_od_other'] else ""))
    print(f"Register vs SiteIQ   : {len(d['not_in_siteiq']):,} register assets not in RENTAL_STOCK at all; "
          f"{len(d['onhire_live']):,} on hire")
    print(f"No Date              : {len(d['nodate']):,} | {len(d['nd_onhire_live']):,} on hire in SiteIQ | "
          f"{len(d['nd_missing']):,} not in SiteIQ")
    print(f"On hire, not in reg. : {len(d['nir']):,} SiteIQ lines across {d['nir_hirers']:,} hirers "
          f"(computed); {len(d['kw']):,} match the calibration keyword rule"
          + (f"; register's own sweep listed {audit_n:,} at its refresh" if audit_n else ""))

    # ---- 1. the PDF -----------------------------------------------------
    css_path = BASE / "k2style.css"
    if not css_path.exists():
        sys.exit("ERROR: k2style.css is missing from the suite folder - the "
                 "house-style PDF cannot render without it.")
    css = css_path.read_text(encoding="utf-8")
    print("-" * 68)
    print("[1/2] Calibration register PDF (house style)...")
    pages, marks = build_pages(rows, d, S)
    total, dated, n_nd = d["total"], len(d["dated"]), len(d["nodate"])
    n_od_oh = len(d["now_od_onhire"])
    # the cover: the one number of the day and three true lines under it -
    # wearing the SAME status as the band on the position page, and saying
    # how old the data was when the pack was built
    key_value, key_label = num(len(d["now_overdue"])), "assets overdue at the pull"
    cover = sh.cover_page(CONFIG, key_value, key_label, [
        f'<b>{num(len(d["now_incal"]))}</b> in calibration - current, next due 31+ days',
        f'<b>{num(len(d["now_due30"]))}</b> due inside 30 days - book them in now',
        f'<b>{num(n_nd)}</b> with no certificate date - status unknown until it is entered',
    ], gen_s, asat_s, rag=S["band"][0], fresh=sh.freshness_line(pulled, gen_dt))
    doc, n_pages = render_doc(pages, cover, gen_s, asat_s)
    pdf_path = OUT / CONFIG["pdf_name"]
    pdf_ok, layout_ok = render_k2_pdf(doc, pdf_path, n_pages, css)
    print(f"Page HTML kept       : {OUT / CONFIG['page_html']}")
    if pdf_ok:
        print("PDF finish           : " + pdf_finish.finish(
            pdf_path, f"{CONFIG['client']} {CONFIG['title']} - as at {asat_s}",
            "Calibration status of the Ampol tool store register at the SiteIQ pull - overdue, "
            "due inside 30 days, the chase list by name and the No Date position.",
            doc, keywords="calibration, register", has_cover=True, family="Calibration"))

    # ---- the phone card: the position-page values, the band, four scores --
    # "chase list cleared" reads the previous recorded pull's chase count
    # against today's; with no earlier day on record it is 0, said so.
    card_path = OUT / CONFIG["card_name"]
    now_pct = len(d["now_dated_ok"]) / dated * 100 if dated else 0
    prev_chase = rh.previous("calibration", "chase", pulled)
    if prev_chase and prev_chase[1]:
        cleared = round(max(0, prev_chase[1] - n_od_oh) / prev_chase[1] * 100)
        cleared_lab = f"Chase list cleared since {prev_chase[0]:%d %b}"
    else:
        cleared, cleared_lab = 0, "Chase list cleared (no earlier day yet)"
    scores = [("Dated fleet in date", round(now_pct)),
              ("Register certified", round(dated / total * 100) if total else 0),
              (cleared_lab, cleared),
              ("No Date entered", round((total - n_nd) / total * 100) if total else 0)]
    sh.position_card_png(CONFIG, asat_s, S["card_tiles"], S["band"], scores, str(card_path),
                         foot=f"Counted from the SiteIQ pull of {asat_s} and the register's "
                              f"due dates - nothing estimated.")
    print(f"Position card        : {card_path}")

    # ---- 2. the email (draft - never sends) -----------------------------
    print("[2/2] Outlook email (house style, draft only)...")
    # WHY (03 Sep 2026): the .eml shows the position card inline under the
    # header (a cid part) and still carries it as a file; the native-draft
    # manifest lists it as an attachment only, so its body is written
    # without the inline image.
    body_html = build_email_html(d, gen_s, S, marks)
    (OUT / CONFIG["email_html"]).write_text(body_html, encoding="utf-8")
    html = build_email_html(d, gen_s, S, marks,
                            card_cid="positioncard" if card_path.exists() else "")
    msg = EmailMessage()
    subject = (f"Ampol Tool Store - Calibration Register Report - "
               f"as at {pulled.strftime('%d/%m/%Y %H:%M')}")
    msg["Subject"] = subject
    msg["To"] = eng.STAFF_EMAIL_TO
    msg["Date"] = formatdate(localtime=True)
    msg["X-Unsent"] = "1"
    msg.set_content("This report is best viewed in HTML. The calibration "
                    "register PDF is attached.\n" if pdf_ok else
                    "This report is best viewed in HTML. The calibration register "
                    "PDF could not be rendered on this machine - the page HTML is "
                    "in the report folder.\n")
    msg.add_alternative(html, subtype="html")
    if card_path.exists():
        with open(card_path, "rb") as f:
            msg.get_payload()[1].add_related(f.read(), maintype="image", subtype="png",
                                             cid="<positioncard>", filename=card_path.name,
                                             disposition="inline")
    # the PDF is attached only when it exists - a missing engine must not
    # kill the email, and the manifest must never promise a file that is
    # not there
    attach = [str(pdf_path)] if pdf_ok and os.path.exists(pdf_path) else []
    if card_path.exists():
        attach.append(str(card_path))      # the phone card rides beside the PDF
    for p in attach:
        with open(p, "rb") as f:
            if p.lower().endswith(".png"):
                msg.add_attachment(f.read(), maintype="image", subtype="png",
                                   filename=os.path.basename(p))
            else:
                msg.add_attachment(f.read(), maintype="application",
                                   subtype="pdf", filename=os.path.basename(p))
    eml_path = OUT / CONFIG["eml_name"]
    with open(eml_path, "wb") as f:
        f.write(msg.as_bytes())
    print(f"EML written          : {eml_path}  "
          f"({os.path.getsize(eml_path):,} bytes; attached: "
          + (", ".join(os.path.basename(p) for p in attach) if attach else "NO PDF") + ")")
    # manifest so MAKE_OUTLOOK_DRAFTS keeps working - recipients derive
    # from the engine's STAFF_EMAIL_TO, one source of truth.
    to_line = "; ".join(re.findall(r"<([^>]+)>", eng.STAFF_EMAIL_TO))
    (OUT / CONFIG["draft_json"]).write_text(json.dumps({
        "subject": subject,
        "to": to_line,
        "body": CONFIG["email_html"],
        "attachments": [os.path.basename(p) for p in attach],
    }, indent=1), encoding="utf-8")

    # ---- the scoreboard: today's figures, keyed on the pull day ----------
    # A re-run on the same pull replaces the day's entry; the next pull
    # reads it back and prints the movement. Real recorded days only.
    # WHY (03 Sep 2026): the entry carries the position in words as well as
    # figures (extra) - the status, the headline, the rule, the owner, the
    # dated action and the cover number - so the daily position page can
    # quote the report without opening it.
    rw = S["rag_words"]
    hist = rh.record("calibration", pulled, {
        "assets": total, "incal": len(d["now_incal"]), "due30": len(d["now_due30"]),
        "overdue": len(d["now_overdue"]), "nodate": n_nd, "chase": n_od_oh,
        "chase_hirers": len(d["now_chase_hirer"]) + len(d["now_chase_repairs"]),
        "overdue_shelf": len(d["now_od_avail"]), "not_in_siteiq": len(d["not_in_siteiq"]),
        "onhire_not_in_register": len(d["nir"])},
        extra={"rag": rw["status"], "headline": rw["headline"], "rule": rw["rule"],
               "owner": CONFIG["rag_owner"], "action": rw["action"], "due": S["action_due_s"],
               "key_value": key_value, "key_label": key_label,
               "second_value": num(n_od_oh), "second_label": "overdue and out on hire - the chase list",
               "title": f"{CONFIG['client']} {CONFIG['title']}", "folder": "Calibrations",
               "pdf": CONFIG["pdf_name"], "card": CONFIG["card_name"]})
    prev = rh.previous("calibration", "overdue", pulled)
    print(f"History              : {asat_d:%d %b %Y} figures written to {hist.parent.name}/{hist.name}"
          + (f" - movement shown against {prev[0]:%d %b %Y}" if prev
             else " - first day on record; movement notes start with the next pull"))
    print("")
    if stale_days > STALE_DAYS:
        print(f"BEFORE SENDING: due dates are {stale_days} days old - check for certificates "
              "issued since, enter them in Register Entry, save, run again.")
    print(f"NEXT STEP: double-click the .eml in {OUT}, check it, press Send.")
    print("Done. The Coates Way - consistent execution, every day.")
    if not pdf_ok:
        sys.exit("\nERROR: PDF not rendered - the email and page HTML were written, "
                 "but there is no PDF to send. Edge is standard on Coates laptops.")
    if not layout_ok:
        sys.exit("\nWARNING: the PDF failed its layout check - see above. Do "
                 "not send it as is.")


if __name__ == "__main__":
    # Failures leave a nonzero exit so the bat button tells the truth;
    # the bat owns the end-of-run pause - no input() here.
    try:
        main()
    except Exception as e:
        import traceback
        traceback.print_exc()
        sys.exit(f"\nERROR: {e}")
