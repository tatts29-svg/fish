# =============================================================================
#  COATES | AMPOL TOOL STORE - ON-HIRE, UTILISATION & COMPLIANCE REPORT KIT
#  Author: Andrew Fisher  |  The Coates Way  |  POWERED BY SITEIQ
#
#  Reads every input from the suite's one Data area (ampol_paths; never
#  modifies the workbook). Reports: Executive Summary, Tooling On-Hire
#  Report (every company A to Z in one document), Quarterly on-hire charge
#  reports, Utilisation & What-to-Buy, Compliance & Trends. A one-off
#  Company on-hire report is still available (--company NAME) but is no
#  longer part of --everything. Every report ships as PDF + HTML + X-Unsent
#  .eml + Outlook-safe email draft (08_MAKE_OUTLOOK_DRAFTS.bat -> native
#  drafts with full To-field search).
#
#  WHY (02 Sep 2026): every printed figure is now computed here, from the
#  raw SiteIQ exports (RENTAL_STOCK, TRANSACTIONS, STOCKTAKE), the pricing
#  file and the corrected-descriptions mapping. The Excel workbook's Power
#  Query tabs are no longer the source of any number - an audit found them
#  quoting a 06 Aug refresh on a 02 Sep report, and four different on-hire
#  totals on one page because each tab filtered differently. The workbook
#  rules (its M code) are ported below so the numbers still mean the same
#  thing; the workbook itself is only read, when present, for a console
#  cross-check that tells Andrew whether it is stale.
#
#  WHY (02 Sep 2026, later): Andrew asked for ONE clean tooling on-hire
#  report instead of a report per company ("there are hundreds of
#  companies"), the quarterly charge reports kept, the site's former name
#  gone from every page, and everything in alphabetical order. Names now
#  come from the shared ampol_names module (one customer, one name; project
#  accounts roll up to their parent and are shown as a sub-label), every
#  table is A to Z unless its heading says "ranked by ...", and the
#  Tooling On-Hire Report carries the full register: company A to Z,
#  hirer A to Z, items longest-held first.
#
#  WHY (02 Sep 2026, presentation pass): the eight reports now wear the
#  Coates house frame every other family in the suite wears (k2flow +
#  k2shell: orange frame, dark hero, KEY strip, running header, pinned
#  footer band, page N of M, dark KPI tiles, house tables and charts).
#  Every report is authored ONCE as a list of content blocks; the frame
#  dresses them for the PDF / HTML and the proven Outlook-safe markup
#  dresses the same blocks for the email body - same words, same figures,
#  same order in both. No count, rule or ordering changed.
#
#  WHY (03 Sep 2026, day anchoring): every day-based figure - days out,
#  days elapsed, ageing bands, quarter guards, partial-month words - was
#  counted to the machine's clock (TODAY). Built on the pull day it was
#  right by luck; built the next morning on the same exports it moved every
#  one of those figures by a day (the utilisation page would have said 246
#  days elapsed against a 02 Sep pull). They now count to ASAT_DAY, the
#  SiteIQ pull date read from RENTAL_STOCK's REFERENCE_INFO; TODAY names
#  only the output folder. Same exports in, same figures out, any day.
#
#  WHY (03 Sep 2026, the wow pass): the family now carries the movement
#  scoreboard (report_history: a page-1 tile says how it moved since the
#  last recorded day, and says nothing when there is no earlier day), a
#  RAG band with an owner and a dated next action on the Executive Summary,
#  the Tooling On-Hire Report and the quarterlies (the rule and its lines
#  printed on the band, set in CONFIG), a cover page on the two client-
#  facing reports, company mini-scorecards in the register (share of fleet
#  and an ageing strip with the counts printed beside it), the Coates Way
#  panel on every closing page, and a phone-sized position card PNG beside
#  the PDF, attached to the draft. Same numbers, same rules, same order.
#
#  WHY (03 Sep 2026, the 10/10 pass): one file-name rule for every output
#  (ampol_names.report_stem - Coates_Ampol_<Report>_03Sep2026 and its .pdf,
#  .html, .body.html, _OUTLOOK.eml, _PositionCard.png and draft manifest);
#  PDF properties and bookmarks stamped after every print (pdf_finish:
#  Author, Subject, the navigation pane); the cover carries the page-1
#  status stripe and the freshness line; three things to do today sit
#  under the Executive Summary's band; a "Since the last pull" section
#  reads the register pull against the previous one and the 24 hours of
#  traffic before it (pull_diff - an honest note until a previous pull is
#  parked, never an invented row); an ageing-by-company panel under the
#  company bars; an APPENDIX divider before every complete register; a
#  trend page once seven days are on the scoreboard; the family's position
#  written to the scoreboard for the daily position page; and the position
#  card inlined in the two client emails. Same numbers, same rules, same
#  order - the new pages are added, nothing already printed moved.
#
#  WHY (03 Sep 2026, the layout pass): the position page now reads the same
#  way on the Executive Summary, the Tooling On-Hire Report and every
#  quarterly - hero and key strip, the band, the tiles, three things to do
#  today, then a three-line story - and everything else starts on the page
#  after. The quarterlies get a cover; every cover lists what is inside with
#  real page numbers read off the printed PDF (printed twice, same page
#  count asserted). The register is one group table per company, so a
#  company that runs over a page opens the next page with its name again.
#  Ageing panels chart the companies with ten or more items and list the
#  rest in one line. Headings are sentence case and the long dash is gone
#  from every string this file writes. Same numbers, same rules, same order.
#
#  WHY (03 Sep 2026, the insights pass): the TRANSACTIONS export holds every
#  issue and return since 01 Jan - 90,000-odd real movements - and the family
#  read a slice of it. The shared txn_insights engine now reads it once per
#  build and every report asks it about its own population: the Executive
#  Summary gains four store-wide pages (the year in movements, who holds
#  what, the counter's rhythm, what the log and the register disagree on),
#  the Tooling On-Hire Report gains the quarter-close look forward (arithmetic
#  on the on-hire date, never a forecast), the return windows by product and
#  the year in movements, every quarterly gains its own quarter-close look
#  forward, and the Utilisation report gains dead stock, headroom and the fast
#  movers. Every new section carries a one-line So what; every new page names
#  its source on the data-and-method page. Page 1 keeps its grammar - the one
#  addition is a sentence under the band's rule on the Tooling On-Hire Report
#  stating the data behind the 90-day line. Same numbers on every page that
#  already existed.
# =============================================================================
import datetime as dt
import html as _html
import json
import os
import re
import subprocess
import sys
import time
from collections import Counter, defaultdict

import openpyxl

import ampol_names as N  # WHY (02 Sep 2026): one place for how names are shown
import ampol_paths  # WHY (12 Aug 2026): one Data area in, dated Reports folder out
import k2flow       # WHY (02 Sep 2026): the Coates house frame for flowing reports
import k2shell      # WHY (12 Aug 2026): the shared K2 chart kit - self-contained SVG
import k2shell as sh
import pdf_finish   # WHY (03 Sep 2026): PDF properties and bookmarks after every print
import pull_diff    # WHY (03 Sep 2026): what moved since the last register pull
import report_history as rh  # WHY (03 Sep 2026): the movement scoreboard
import txn_insights as ti    # WHY (03 Sep 2026): what the transaction log already knows

HERE = os.path.dirname(os.path.abspath(__file__))
# WHY (12 Aug 2026): outputs now land in the suite's dated Reports area -
# Reports\<today>\Tooling - created on demand, one folder per day.
OUT_DIR = ampol_paths.day_folder("Tooling")
WORKBOOK = "Ampol_Onhire_Tooling_Report.xlsm"

TODAY = dt.date.today()
# WHY (12 Aug 2026): Australian date style (02 Sep 2026) and 24-hour time,
# same as the rest of the suite. WHY (02 Sep 2026): the leading zero stays
# - the house style is "02 Sep 2026", and every other date on the page
# (fmt_date, stamp) already prints it that way.
GENERATED_DT = dt.datetime.now()
GENERATED = GENERATED_DT.strftime("%d %b %Y %H:%M")
# WHY (03 Sep 2026): every output file is named by ampol_names.report_stem
# (Coates_Ampol_<Report>_03Sep2026); the day tag is the day the button was
# pressed, which is also the Reports\ folder the file sits in. The console
# names that folder relative to the suite and nothing else carries a date
# of its own.
OUT_REL = os.path.relpath(OUT_DIR, HERE)
# WHY (03 Sep 2026): TODAY names the output folder and nothing else. Every
# day-based figure (days out, days elapsed, ageing bands, quarter guards,
# partial-month words, the sort tie-break) counts to ASAT_DAY - the date of
# the SiteIQ RENTAL_STOCK pull, set by load_all from the export's own
# REFERENCE_INFO. Until the export is read it falls back to the clock.
ASAT_DAY = TODAY
ASAT_DT = None       # the full pull timestamp (datetime) once read
# WHY (02 Sep 2026): the data-as-at stamp is the SiteIQ pull time written
# inside the RENTAL_STOCK export (REFERENCE_INFO), never a file's mtime.
DATA_ASAT = "TBC"
# the same stamp without the 'SiteIQ pull' words - the house hero prints
# "Data as at: <stamp> (SiteIQ register pull)" and the running header "AS AT <stamp>"
ASAT_SHORT = "TBC"

ORANGE = "#F26222"
DARK = "#1D1D1B"
GREY = "#555555"
LIGHT = "#F5F1EC"
RED = "#B3261E"
GREEN = "#1E7B34"
AMBER = "#9A6A00"

HIGH_VALUE = 1000.0          # replacement-cost threshold for the high-value chase
NOT_SIGHTED_DAYS = 90        # "not seen in 3 months = potentially out of test date"
OUT_OF_TAG_HIRER = "RIGGING & 240V - OUT OF TAG DATE"

# WHY (03 Sep 2026): the page-1 RAG band and its owner. The rating is the
# share of this year's tooling on hire that has been out more than 90 days:
# Green under the amber line, Amber from it, Red from the red line. The
# band prints these lines beside the rating, so a reader can see the rule
# that produced it; change them here, never on the page.
CONFIG = {
    "over90_amber_pct": 10.0,
    "over90_red_pct": 25.0,
    "rag_owner": "Andrew Fisher, Shutdown Manager",
    "rag_action": ("Over-90-day list to each company's supervisor at the quarter-close "
                   "run; returns or rescans reported on the next report"),
    "rag_due_days": 7,          # next-action due date = pull date + this many days
}

QUARTERS = {
    "Q1": ("Q1 Jan-Mar Onhire", "Q1 (Jan-Mar)", (1, 2, 3)),
    "Q2": ("Q2 Apr-Jun Onhire", "Q2 (Apr-Jun)", (4, 5, 6)),
    "Q3": ("Q3 Jul-Sep Onhire", "Q3 (Jul-Sep)", (7, 8, 9)),
    "Q4": ("Q4 Oct-Dec Onhire", "Q4 (Oct-Dec)", (10, 11, 12)),
}
QTR_WORD = {"Q1": "1st Qtr", "Q2": "2nd Qtr", "Q3": "3rd Qtr", "Q4": "4th Qtr"}
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

ELECTRICAL_KW = ["240V", "110V", " LEAD", "EXTENSION LEAD", "RCD", "POWER BOARD",
                 "POWERBOARD", "TRANSFORMER", "DISTRIBUTION BOARD"]
RIGGING_KW = ["SLING", "SHACKLE", "CHAIN BLOCK", "LEVER HOIST", "HOIST", "BEAM CLAMP",
              "BEAM TROLLEY", "TIRFOR", "TURFER", "SPREADER", "LIFTING", "RIGGING",
              "PLATE CLAMP", "DRUM LIFTER"]
HIGH_TORQUE_KW = ["TORQUE", "TENSIONER", "HYTORC", "ENERPAC", "IMPACT WRENCH",
                  "RATTLE GUN", "NUT RUNNER"]
# WHY (02 Sep 2026): BLJ / NDE / UGL / FCCU / SATGAS / MOL were printing as
# "Blj", "Nde Solutions", "Ugl Fccu", "Satgas/mol" - they are acronyms.
ACRONYMS = {"CR", "HIS", "IPS", "CSA", "BMD", "ARL", "JLG", "PPE", "MMG",
            "BLJ", "NDE", "UGL", "FCCU", "SATGAS", "MOL", "SFI", "T&I", "AGM",
            "IPCQ", "WSP", "CXC"}

# ----------------------------------------------------------------------------
# ONE exclusion list for both engines (on-hire and utilisation).
# WHY (02 Sep 2026): the workbook excluded LANYARD from on-hire but the
# utilisation tab still raised "Retractable Lanyard" as a buy signal. These
# families are reported by the Radio / Gas Monitor / Rigging kits instead.
FAMILY_EXCLUSIONS = ("RADIO", "GAS MONITOR", "GAS DETECTOR", "MULTI GAS", "MULTIGAS",
                     "DRAGER", "LANYARD", "STEEL COIL CLAMP")
# Extra group-NAME hygiene the utilisation engine applies to corrected
# descriptions (an uncorrected "Ampol ..." or a brand name is not a group).
UTIL_NAME_BANS = ("AMPOL", "COATES", "MOTOROLA", "DRAEGER", "X-AM", "XAM")
UTIL_EXCLUSIONS = FAMILY_EXCLUSIONS + UTIL_NAME_BANS
# The workbook's own utilisation list, kept only so the port can be verified
# against the workbook tab like-for-like (see verify note in compute_utilisation).
UTIL_M_BANNED = ("AMPOL", "COATES", "RADIO", "MOTOROLA", "DRAGER", "DRAEGER",
                 "X-AM", "XAM", "GAS MONITOR")

# Hirer accounts the workbook's Master Onhire query drops (exact, after Proper case)
MASTER_HIRER_EXCLUSIONS = {"BULK - YARD", "LOADING BAY - OUT OF SERVICE",
                           "OUT OF - CALIBRATION", "RIGGING & 240V - OUT OF TAG DATE"}
# ... and the two it drops by substring
MASTER_HIRER_SUBSTRING_EXCLUSIONS = ("T&I - TOOL STORE", "ALL-AROUND - REPAIRS")

# Custody accounts that are NOT customers: never a company chip, never a
# recovery row, never a company report. Reported on their own labelled line.
INTERNAL_CUSTODY = {"REPAIRS", "DRÄGER", "DRAGER"}

# WHY (02 Sep 2026): shutdown / workflow hirer accounts ("Fccu - 2026 (Sfi)",
# "Atlas Chains - Offsite Repairs", "Loading Bay - Out Of Service") were being
# listed and counted as people. This pattern spots them in both the register
# ("First - Last" style) and the transactions export ("First Last" style).
# A bare "T&I" after a person's name (e.g. "Jay Purcell T&I") stays a person.
CUSTODY_HIRER_RE = re.compile(
    r"repairs|off\s?site|out of service|out of tag|calibration|bulk\s*-?\s*yard"
    r"|^t&i\b|t&i\s*-?\s*(shutdown|tool\s?store)|fccu|\(sfi\)|operations|after hours"
    r"|future\s*-?\s*fuels|dr[äa]e?ger|coates\s*-?\s*out", re.I)

LSR_LINE = ("Life Saving Rule 5 - Tools and Equipment (SEQ-GL-009): nothing damaged, "
            "defective or out of test date is ever reissued.")
CANON = ["Two scans. Two looks. One standard.",
         "Nothing goes on the shelf unless it is ready for hire.",
         "Right first time. Every item. Every time."]


# ---------------------------------------------------------------- helpers ---
def esc(v):
    return _html.escape(str(v)) if v is not None else ""


def money(v):
    try:
        return "${:,.2f}".format(float(v))
    except (TypeError, ValueError):
        return "-"


def pct(v):
    try:
        return "{:.0f}%".format(float(v) * 100)
    except (TypeError, ValueError):
        return "-"


def n_fmt(v):
    """Thousands-separated integer for tiles and prose (1,500 not 1500)."""
    try:
        return f"{int(round(float(v))):,}"
    except (TypeError, ValueError):
        return "-"


def plural(n, word="item"):
    return f"{n_fmt(n)} {word}{'' if n == 1 else 's'}"


def clean(v):
    return str(v).strip() if v is not None else ""


def fmt_date(d):
    return d.strftime("%d %b %Y") if d else "-"


def m_proper(text):
    """Power Query Text.Proper: every run of letters gets a capital first
    letter and lower-case rest, and ANY non-letter starts a new run
    ("240v" -> "240V", "Plumber's" -> "Plumber'S", "DRÄGER" -> "Dräger").
    Ported so descriptions and hirers read exactly as the workbook tabs did."""
    return re.sub(r"[^\W\d_]+",
                  lambda m: m.group(0)[0].upper() + m.group(0)[1:].lower(),
                  clean(text))


def clean_text(v):
    """The workbook's CleanText: Proper(upper(trim)) with the site's former
    name read as the current one (ampol_names.display_desc)."""
    if v is None:
        return ""
    return m_proper(N.display_desc(clean(v).upper()))


def proper_clean(v):
    """The workbook utilisation query's ProperCleanText (Text.Clean strips
    control characters; small words lowered; 300Mm -> 300mm)."""
    t = "".join(ch for ch in clean(v) if ch >= " ")
    t = m_proper(t.strip())
    for a, b in ((" And ", " and "), (" Of ", " of "), (" To ", " to "),
                 (" For ", " for "), (" With ", " with "), ("Mm", "mm")):
        t = t.replace(a, b)
    return t


def desc_key(v):
    """Pricing match key (workbook CleanDescription): former site name read
    as the current one, collapse spaces, upper-case."""
    s = N.display_desc(clean(v))
    return " ".join(t for t in s.split(" ") if t).upper()


def desc_key_noampol(v):
    return " ".join(t for t in desc_key(v).split(" ") if t != "AMPOL").strip()


def harmonise_barcode(v):
    """Workbook HarmonizeBarcode: upper, drop spaces/dashes/slashes, strip
    leading zeros - so 'AMP028/002' and 'AMP028-002' meet as one item."""
    s = clean(v).upper().replace(" ", "").replace("-", "").replace("/", "")
    return s.lstrip("0")


def acronym_case(text, capitalise=True):
    out = []
    for w in re.split(r"(\s+|/|\(|\))", clean(text)):
        if not w or not w.strip() or w in "/()":
            out.append(w)
        elif w.upper() in ACRONYMS:
            out.append(w.upper())
        else:
            out.append(w.capitalize() if capitalise else w)
    return "".join(out)


def display_company(name):
    """The customer's one name, from the shared ampol_names rule.
    WHY (02 Sep 2026): the kit used to keep its own casing rules and its
    own idea of which accounts merge; the suite now has one rule for every
    report, so a company can never print under two names."""
    return N.display_company(name) if clean(name) else ""


def canonical_company(name):
    """ONE company rule for the register AND the transactions export -
    delegated to ampol_names.display_company so the two can never disagree.
    WHY (02 Sep 2026): the former site-name account and the refinery legal
    name both read Ampol; CR reads Contract Resources; FCCU and SATGAS/MOL
    project accounts roll up to their parent company (the SiteIQ account is
    kept beside every row as account_label, so nothing is lost)."""
    return display_company(name)


def account_of(name):
    """The SiteIQ account a row is booked to, shown under its company, e.g.
    'Wood (FCCU project account)' or 'Ampol (refinery account)'."""
    return N.account_label(name) if clean(name) else ""


def display_hirer(name):
    """Hirer name as SiteIQ records it. The only touch is whitespace, plus
    title case for an entry typed wholly in capitals or wholly in lower
    case - the register mixes 'David - McGurk' with 'ROBERT - MCGREGOR'
    and 'leonard - atterwell', and a mixed-case entry is kept exactly
    (proper-casing it would turn McGurk into Mcgurk)."""
    s = re.sub(r"\s+", " ", clean(name))
    letters = "".join(ch for ch in s if ch.isalpha())
    if letters and (letters.isupper() or letters.islower()):
        return acronym_case(m_proper(s), capitalise=False)
    return s


def is_custody_hirer(name):
    return bool(CUSTODY_HIRER_RE.search(clean(name)))


def safe_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip().split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def safe_datetime(v):
    """SiteIQ REFERENCE_INFO stamps look like '02/09/2026 06:30 PM'."""
    if isinstance(v, dt.datetime):
        return v
    s = clean(v)
    for fmt in ("%d/%m/%Y %I:%M %p", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def stamp(d):
    return d.strftime("%d %b %Y %H:%M") if d else "TBC"


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def matches_any(text, kws):
    t = " " + clean(text).upper() + " "
    return any(k in t for k in kws)


def family_hit(desc, words=FAMILY_EXCLUSIONS):
    """First excluded family word found in a description, or None."""
    u = clean(desc).upper()
    for w in words:
        if w in u:
            return w
    return None


def category_of(desc):
    if matches_any(desc, HIGH_TORQUE_KW):
        return "High Torque"
    if matches_any(desc, RIGGING_KW):
        return "Rigging"
    if matches_any(desc, ELECTRICAL_KW):
        return "Electrical"
    return None


def quarter_of(d):
    return "Q" + str((d.month - 1) // 3 + 1) if d else None


def quarter_started(qk):
    return ASAT_DAY >= dt.date(ASAT_DAY.year, QUARTERS[qk][2][0], 1)


# ---------------------------------------------------------------- loading ---
def find_file(*names):
    # WHY (12 Aug 2026): inputs now come from the suite's one Data area via
    # ampol_paths - newest match wins, Excel ~$ lock files and archived
    # Source_ pulls are never candidates. Same name, same call sites.
    return ampol_paths.find_data(*names) or None


def sheet_rows(ws, probe):
    """Read a sheet as dicts, locating the header row by a probe column name."""
    rows = list(ws.iter_rows(values_only=True))
    hdr_i = None
    for i, r in enumerate(rows[:6]):
        if r and any(clean(c) == probe for c in r if c is not None):
            hdr_i = i
            break
    if hdr_i is None:
        return []
    headers = [clean(c) for c in rows[hdr_i]]
    out = []
    for r in rows[hdr_i + 1:]:
        if r is None or all(c is None or clean(c) == "" for c in r):
            continue
        out.append({h: v for h, v in zip(headers, r) if h})
    return out


def reference_info(wb):
    """The SiteIQ export's own REQUESTED_DATE/TIME and report period.
    WHY (02 Sep 2026): this is the true data-as-at - the moment SiteIQ
    answered - not when a file was saved or a workbook refreshed."""
    info = {"pulled": None, "period": ""}
    if "REFERENCE_INFO" not in wb.sheetnames:
        return info
    rows = [r for r in wb["REFERENCE_INFO"].iter_rows(values_only=True)
            if r and any(c is not None for c in r)]
    if len(rows) < 2:
        return info
    hdr = [clean(c) for c in rows[0]]
    vals = rows[1]
    for h, v in zip(hdr, vals):
        if "REQUESTED_DATE" in h.upper():
            info["pulled"] = safe_datetime(v)
        elif h.upper() == "REPORT_PERIOD":
            info["period"] = clean(v)
    return info


def load_all():
    global DATA_ASAT, ASAT_SHORT, ASAT_DAY, ASAT_DT
    d = {"asat": {}}

    # ---- RENTAL_STOCK: the live register, the source of every on-hire figure
    rs_path = find_file("RENTAL_STOCK.xlsx")
    if not rs_path:
        sys.exit("RENTAL_STOCK.xlsx not found in the Data folder - run "
                 "12_PULL_SITEIQ_EXPORTS.bat (or save the SiteIQ export there) "
                 "and run again.")
    rwb = openpyxl.load_workbook(rs_path, read_only=True, data_only=True)
    d["asat"]["stock"] = reference_info(rwb)
    ws = rwb["RENTAL_STOCK"]
    rows = ws.iter_rows(values_only=True)
    headers = [clean(c) for c in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}

    def col(r, name, default=None):
        i = idx.get(name)
        return r[i] if i is not None and i < len(r) else default

    d["stock"] = []
    for r in rows:
        if r is None or all(c is None for c in r):
            continue
        d["stock"].append({
            "barcode": clean(col(r, "ITEM_BARCODE")),
            "desc_raw": clean(col(r, "ITEM_DESCRIPTION")),
            "status": clean(col(r, "ITEM_STATUS")),
            "company_raw": clean(col(r, "COMPANY_NAME")),
            "hirer_raw": clean(col(r, "HIRER_NAME")),
            "date": safe_date(col(r, "ON_HIRE_DATE")),
            "time": clean(col(r, "ON_HIRE_TIME")),
            "unit": clean(col(r, "STORAGE_UNIT")),
        })
    rwb.close()
    pulled = d["asat"]["stock"]["pulled"]
    DATA_ASAT = ("SiteIQ pull " + stamp(pulled)) if pulled else "TBC (no REFERENCE_INFO)"
    ASAT_SHORT = stamp(pulled) if pulled else "TBC"
    # WHY (03 Sep 2026): from here on every day count is to the pull date
    ASAT_DT = pulled
    ASAT_DAY = pulled.date() if pulled else TODAY

    # ---- corrected descriptions (Tooling_Description_Mapping.xlsx, sheet 'Use this')
    # WHY (02 Sep 2026): the workbook's Master query looked for a column named
    # "CorrectedDescriptionsTable.Corrected Description" that the file does not
    # have ("Corrected Description"), so no correction ever reached the tabs
    # and "Critictal Risk Signage" printed as-is. Read here by the real header.
    d["corr"] = {}          # barcode (upper) -> corrected description, first wins
    d["corr_rows"] = []     # raw rows for the utilisation engine
    # WHY (02 Sep 2026): the mapping is keyed by barcode, so "Critictal Risk
    # Signage" was corrected for AMP368/052 only while its two sister signs on
    # hire kept the typo. When every mapping row for a register description
    # agrees on one corrected name, that name is used for any barcode carrying
    # the same description - display only; utilisation stays barcode-keyed.
    d["corr_by_desc"] = {}  # register description key -> corrected description
    by_desc = defaultdict(set)
    mp_path = find_file("Tooling_Description_Mapping.xlsx")
    if mp_path:
        mwb = openpyxl.load_workbook(mp_path, read_only=True, data_only=True)
        sheet = next((s for s in ("Use this", "In use", "CorrectedDescriptionsTable")
                      if s in mwb.sheetnames), None)
        if sheet:
            mrows = list(mwb[sheet].iter_rows(values_only=True))
            mh = [clean(c) for c in mrows[0]] if mrows else []
            bi = next((i for i, h in enumerate(mh) if h == "ITEM_BARCODE"), None)
            ci = next((i for i, h in enumerate(mh)
                       if h in ("Corrected Description",
                                "CorrectedDescriptionsTable.Corrected Description")), None)
            di = next((i for i, h in enumerate(mh) if h == "ITEM_DESCRIPTION"), None)
            if bi is not None and ci is not None:
                for r in mrows[1:]:
                    if not r:
                        continue
                    bc = clean(r[bi] if bi < len(r) else None).upper()
                    corr = clean(r[ci] if ci < len(r) else None)
                    d["corr_rows"].append((bc, corr))
                    if bc and corr:
                        d["corr"].setdefault(bc, corr)
                    if di is not None and corr:
                        raw = clean(r[di] if di < len(r) else None)
                        if raw:
                            by_desc[desc_key(clean_text(raw))].add(corr)
        mwb.close()
    d["corr_by_desc"] = {k: next(iter(v)) for k, v in by_desc.items() if len(v) == 1}
    d["mapping_n"] = len(d["corr_rows"]) if mp_path else None

    # ---- pricing (Avg Buy Price (New) / where to buy)
    # WHY (12 Aug 2026): the underscore name is how the file actually lands in
    # Data, so it is tried first; the spaced name stays as the fallback.
    # WHY (02 Sep 2026): the file carries 831 duplicate descriptions (16 with
    # conflicting prices). The workbook keeps the FIRST row per description
    # (Table.Distinct) - so does every dictionary below (setdefault), so the
    # utilisation buy prices and the replacement costs agree with each other.
    d["pricing"] = {}        # raw description (upper) -> {buy, source}
    d["price_normal"] = {}   # workbook match key -> Avg Buy Price (New)
    d["price_noampol"] = {}  # workbook match key without the AMPOL token
    d["pricing_noampol"] = {}  # group-name lookup without the AMPOL token (priced rows)
    d["pricing_rows"] = 0
    pr_path = find_file("Ampol_ToolStore_Pricing.xlsx", "Ampol ToolStore Pricing.xlsx")
    if pr_path:
        pwb = openpyxl.load_workbook(pr_path, read_only=True, data_only=True)
        ws = pwb["RENTAL_STOCK"]
        rows = ws.iter_rows(values_only=True)
        ph = [clean(c) for c in next(rows)]
        pi = {h: i for i, h in enumerate(ph)}
        di, bi_, si = pi.get("ITEM_DESCRIPTION", 0), pi.get("Avg Buy Price (New)", 1), \
            pi.get("Source to Buy", 5)
        for r in rows:
            if r is None or not clean(r[di]):
                continue
            d["pricing_rows"] += 1
            buy = num(r[bi_])
            src = clean(r[si]) if si < len(r) else ""
            d["pricing"].setdefault(clean(r[di]).upper(), {"buy": buy, "source": src})
            if buy is not None:
                d["pricing_noampol"].setdefault(desc_key_noampol(r[di]),
                                                {"buy": buy, "source": src})
            ct = clean_text(r[di])
            d["price_normal"].setdefault(desc_key(ct), buy)
            d["price_noampol"].setdefault(desc_key_noampol(ct), buy)
        pwb.close()

    # ---- the on-hire master list, straight from the register
    build_master(d)

    # ---- transactions (year to date)
    # WHY (12 Aug 2026): the SiteIQ pull can land as TRANSACTIONS_Full.xlsx or
    # plain TRANSACTIONS.xlsx - try both, newest wins, same sheet either way.
    d["tx"] = []
    d["asat"]["tx"] = {"pulled": None, "period": ""}
    tx_path = find_file("TRANSACTIONS_Full.xlsx", "TRANSACTIONS.xlsx")
    if tx_path:
        twb = openpyxl.load_workbook(tx_path, read_only=True, data_only=True)
        d["asat"]["tx"] = reference_info(twb)
        ws = twb["CUSTOMER_CONTRACTOR_EQUIP"]
        rows = ws.iter_rows(values_only=True)
        headers = [clean(c) for c in next(rows)]
        idx = {h: i for i, h in enumerate(headers)}
        bc_col = next((c for c in ("LATEST_BARCODE", "BARCODE", "ITEM_BARCODE")
                       if c in idx), "LATEST_BARCODE")
        for r in rows:
            if r is None or r[0] is None:
                continue
            company_raw = clean(r[idx.get("EMPLOYER_NAME", 0)])
            hirer = clean(r[idx.get("HIRER_NAME", 1)])
            desc = clean(r[idx.get("SKU/ITEM DESCRIPTION", 4)])
            company = canonical_company(company_raw)
            d["tx"].append({
                "company": company,
                "company_raw": company_raw,
                "hirer": hirer,
                "desc": desc,
                "barcode": clean(r[idx.get(bc_col, 5)]),
                "cat": clean(r[idx.get("PRODUCT_CATEGORY", 6)]),
                "start": safe_date(r[idx.get("TRAN_START_DATE", 9)]),
                "end": safe_date(r[idx.get("TRAN_END_DATE", 11)]),
                "custody": (company.upper() in INTERNAL_CUSTODY) or is_custody_hirer(hirer),
                "family": family_hit(desc),
            })
        twb.close()

    # ---- stocktake (last sighted)
    d["stocktake"] = {}
    d["asat"]["stocktake"] = {"pulled": None, "period": ""}
    st_path = find_file("STOCKTAKE.xlsx", "STOCKTAKE (1).xlsx")
    if st_path:
        swb = openpyxl.load_workbook(st_path, read_only=True, data_only=True)
        d["asat"]["stocktake"] = reference_info(swb)
        ws = swb["STOCKTAKE"]
        rows = ws.iter_rows(values_only=True)
        headers = [clean(c) for c in next(rows)]
        idx = {h: i for i, h in enumerate(headers)}
        for r in rows:
            if r is None:
                continue
            bc = clean(r[idx.get("LATEST_BARCODE", 4)])
            if not bc:
                continue
            seen = r[idx.get("LAST_SIGHTED_DATE_TIME", 9)]
            seen_d = safe_date(seen)
            prev = d["stocktake"].get(bc)
            if prev is None or (seen_d and (prev["seen"] is None or seen_d > prev["seen"])):
                d["stocktake"][bc] = {"seen": seen_d,
                                      "by": clean(r[idx.get("LAST_SIGHTED_BY", 10)]),
                                      "unit": clean(r[idx.get("STORAGE_UNIT", 1)])}
        swb.close()

    # ---- out-of-tag parked gear straight from RENTAL_STOCK
    d["out_of_tag"] = [{"barcode": s["barcode"], "desc": s["desc_raw"], "status": s["status"]}
                       for s in d["stock"] if s["hirer_raw"].upper() == OUT_OF_TAG_HIRER]

    # ---- utilisation, computed here (the workbook tab is not used)
    d["util"] = compute_utilisation(d, ASAT_DAY, UTIL_EXCLUSIONS)

    # ---- workbook: optional console cross-check only
    d["wb"] = workbook_crosscheck(d)
    return d


def build_master(d):
    """The on-hire master list - the workbook's 'Master Onhire' rules, applied
    to the live register:
      * register row with a company (= On Hire), not COMPANY REPAIRS
      * hirer not 'T&I - Tool store' / 'All-Around - Repairs'
      * on-hire date in the report year (older rows are 'legacy', counted
        separately and stated on the page)
      * not a radio / gas monitor / Drager / lanyard / steel coil clamp
      * not Bulk - Yard, Loading Bay - Out Of Service, Out Of - Calibration,
        Rigging & 240V - Out Of Tag Date
    Every quarter count, company count and recovery row is derived from
    THIS list by on-hire month, so they all tie to one total."""
    master, legacy, holding = [], [], []
    custody_family = defaultdict(int)
    year = ASAT_DAY.year

    def row_for(s, desc_clean, fam):
        """One register row in the shape every page uses. WHY (02 Sep 2026):
        the same shape serves the on-hire master, the tool store's own
        holding accounts and the legacy list, so one renderer lists them all
        the same way. 'account' keeps the SiteIQ account beside the rolled-up
        company name; 'days' is calendar days from the on-hire date to today."""
        company = canonical_company(s["company_raw"])
        corrected = (d["corr"].get(s["barcode"].upper())
                     or d["corr_by_desc"].get(desc_key(desc_clean)))
        hirer = display_hirer(s["hirer_raw"])
        return {
            "company": company,
            "company_raw": s["company_raw"],
            "account": account_of(s["company_raw"]),
            "hirer": hirer,
            "barcode": s["barcode"],
            "desc": corrected or desc_clean,
            "desc_raw": s["desc_raw"],
            "corrected": ("barcode" if d["corr"].get(s["barcode"].upper())
                          else ("description" if corrected else "")),
            "date": s["date"],
            "time": s["time"],
            "days": (ASAT_DAY - s["date"]).days if s["date"] else None,
            "month": s["date"].strftime("%B") if s["date"] else "",
            "q": quarter_of(s["date"]),
            "cost": price_for_description(d, desc_clean),
            "family": fam,
            "custody": company.upper() in INTERNAL_CUSTODY,
            "custody_hirer": is_custody_hirer(hirer),
            # category keys off the register description (as audited);
            # see the note in compliance limits
            "cat": category_of(desc_clean),
        }

    for s in d["stock"]:
        if not s["company_raw"] or not s["barcode"]:
            continue
        desc_clean = clean_text(s["desc_raw"])
        fam = family_hit(desc_clean)
        this_year = bool(s["date"] and s["date"].year == year)
        hu = s["hirer_raw"].upper()
        # WHY (02 Sep 2026): the tool store's own holding accounts (T&I -
        # Tool store, All-Around - Repairs, Bulk - Yard, Loading Bay - Out
        # Of Service, Out Of - Calibration, Rigging & 240V - Out Of Tag Date)
        # stay OUT of the on-hire count exactly as before, but this year's
        # rows in them are kept so the Tooling On-Hire Report can list
        # them under "not customer hire" instead of leaving them unseen.
        if (s["company_raw"].upper() == "COMPANY REPAIRS"
                or any(x in hu for x in MASTER_HIRER_SUBSTRING_EXCLUSIONS)):
            if this_year:
                holding.append(row_for(s, desc_clean, fam))
            continue
        if s["date"] and s["date"].year < year:
            legacy.append({**s, "family": fam, "company": canonical_company(s["company_raw"]),
                           "account": account_of(s["company_raw"])})
            continue
        if not this_year:
            continue
        if fam:
            co = canonical_company(s["company_raw"])
            if co.upper() in INTERNAL_CUSTODY:
                custody_family[(co, fam)] += 1
            continue
        if m_proper(s["hirer_raw"]).upper() in MASTER_HIRER_EXCLUSIONS:
            holding.append(row_for(s, desc_clean, fam))
            continue
        master.append(row_for(s, desc_clean, fam))
    master.sort(key=item_order)
    d["master"] = master
    d["legacy"] = legacy
    # this year's tooling rows in the store's own holding accounts (outside
    # the on-hire count) and the radio / gas family rows sitting in the
    # custody companies (Repairs, Dräger) - both stated on the on-hire report
    d["holding"] = sorted([r for r in holding if not r["family"]], key=item_order)
    d["holding_family"] = defaultdict(int)
    for r in holding:
        if r["family"]:
            d["holding_family"][(r["hirer"], r["family"])] += 1
    d["custody_family"] = dict(custody_family)
    # register-level counts for the executive tiles
    st = defaultdict(int)
    for s in d["stock"]:
        st[s["status"]] += 1
    d["status_counts"] = dict(st)
    avail = [s for s in d["stock"] if "AVAILABLE" in s["status"].upper()]
    d["available_all_n"] = len(avail)
    d["available_n"] = sum(1 for s in avail if not family_hit(clean_text(s["desc_raw"])))
    d["available_family"] = defaultdict(int)
    for s in avail:
        f = family_hit(clean_text(s["desc_raw"]))
        if f:
            d["available_family"][f] += 1
    d["repairs_rows"] = [s for s in d["stock"] if s["company_raw"].upper() == "REPAIRS"]
    d["repairs_n"] = len(d["repairs_rows"])
    return master


def price_for_description(d, desc_clean):
    """Replacement cost the workbook way: exact match key first, then the
    key with the AMPOL token dropped; None (a dash) when neither hits."""
    v = d["price_normal"].get(desc_key(desc_clean))
    if v is None:
        v = d["price_noampol"].get(desc_key_noampol(desc_clean))
    return v


def price_for_group(d, group):
    """Catalogue buy price for a utilisation group name: exact description
    first, then the description with the AMPOL token dropped (first priced
    row wins either way, matching the replacement-cost rule)."""
    p = d["pricing"].get(clean(group).upper())
    if p and p.get("buy") is not None:
        return p
    p = d["pricing_noampol"].get(desc_key_noampol(group))
    return p if p else {"buy": None, "source": ""}


def compute_utilisation(d, today, banned):
    """Port of the workbook's 'Tooling Utilisation' query.

    Inputs: the corrected-descriptions mapping (barcode -> group), the live
    register (qty / on hire / available per group) and every transaction
    since 1 Jan. Hire days = distinct calendar days each barcode was out
    between 1 Jan and `today` (transactions plus current open hires),
    summed per group. YTD % = hire days / (qty x days elapsed this year).

    VERIFIED (02 Sep 2026): run with today = 06 Aug 2026 and the workbook's
    own banned list, every group's hire days, qty and YTD % that the 06 Aug
    tab shows are reproduced except where the inputs themselves moved after
    06 Aug (items added / relabelled / returned and re-issued) - see the
    verification note in the console output. The report runs it at the
    real date with the shared FAMILY_EXCLUSIONS, so radios / lanyards never
    appear as buy signals.
    """
    y0 = dt.date(today.year, 1, 1)
    y1 = min(max(today, y0), dt.date(today.year, 12, 31))
    days_elapsed = (y1 - y0).days + 1

    def is_banned(t):
        u = clean(t).upper()
        return any(b in u for b in banned)

    # corrected mapping: harmonised barcode -> group (first wins)
    mapping = {}
    for bc, corr in d["corr_rows"]:
        bcu = clean(bc).upper()
        g = proper_clean(corr)
        if not bcu or not g or is_banned(g):
            continue
        mapping.setdefault(harmonise_barcode(bcu), g)

    # stock, one row per barcode (highest priority, then latest on-hire date)
    best = {}
    for s in d["stock"]:
        bc = s["barcode"].upper()
        if not bc:
            continue
        pri = 2 if (s["company_raw"] or s["hirer_raw"]) else (1 if s["status"] else 0)
        key = (pri, s["date"] or dt.date(1, 1, 1))
        if bc not in best or key > best[bc][0]:
            best[bc] = (key, s)
    groups = defaultdict(lambda: {"qty": set(), "oh": 0, "av": 0})
    open_hires = []
    for bc, (_k, s) in best.items():
        g = mapping.get(harmonise_barcode(bc))
        if not g:
            continue
        G = groups[g]
        G["qty"].add(bc)
        on_hire = bool(s["company_raw"] or s["hirer_raw"])
        if on_hire:
            G["oh"] += 1
            open_hires.append((harmonise_barcode(bc), g, proper_clean(s["hirer_raw"]), s["date"]))
        else:
            su = proper_clean(s["status"]).upper()
            if not any(x in su for x in ("OUT", "REPAIR", "DAMAGED", "FAILED",
                                          "CALIBRATION", "MISSING")):
                G["av"] += 1

    # transactions + open hires -> per barcode: count, distinct days, hirers
    per = defaultdict(lambda: {"n": 0, "days": set(), "hirers": set()})
    n_tx_used = 0
    for t in d["tx"]:
        if not t["barcode"] or not t["start"]:
            continue
        s, e = t["start"], t["end"]
        if s > y1 or (e is not None and e < y0):
            continue
        g = mapping.get(harmonise_barcode(t["barcode"]))
        if not g:
            continue
        n_tx_used += 1
        P = per[(harmonise_barcode(t["barcode"]), g)]
        P["n"] += 1
        h = proper_clean(t["hirer"])
        if h:
            P["hirers"].add(h)
        a = max(s, y0)
        z = y1 if (e is None or e > y1) else e
        if a <= z:
            P["days"].update(range(a.toordinal(), z.toordinal() + 1))
    for hb, g, h, dte in open_hires:
        P = per[(hb, g)]
        if h:
            P["hirers"].add(h)
        a = y0 if (dte is None or dte < y0) else dte
        if a <= y1:
            P["days"].update(range(a.toordinal(), y1.toordinal() + 1))
    tg = defaultdict(lambda: {"tx": 0, "hirers": set(), "days": 0})
    for (_hb, g), P in per.items():
        T = tg[g]
        T["tx"] += P["n"]
        T["hirers"] |= P["hirers"]
        T["days"] += len(P["days"])

    rows = []
    for g in sorted(set(groups) | set(tg)):
        q = len(groups[g]["qty"]) if g in groups else 0
        oh = groups[g]["oh"] if g in groups else 0
        av = groups[g]["av"] if g in groups else 0
        tx = tg[g]["tx"] if g in tg else 0
        hr = len(tg[g]["hirers"]) if g in tg else 0
        dys = tg[g]["days"] if g in tg else 0
        live = 0.0 if q == 0 else oh / q
        ytd = 0.0 if q == 0 else dys / (q * days_elapsed)
        if q == 0 and tx > 0:
            rec = "Review - Used in Transactions but Not Found in Current Stock"
        elif q == 0:
            rec = "No Current Stock"
        elif av == 0 and oh > 0:
            rec = "High Priority - No Available Stock"
        elif live >= 0.85:
            rec = "Consider Additional Stock"
        elif ytd >= 0.70:
            rec = "Strong Usage - Monitor Closely"
        elif tx == 0 and q > 0:
            rec = "Low Usage - Review Need"
        elif av > oh and ytd < 0.20:
            rec = "Potential Overstock"
        else:
            rec = "Stock Level Looks OK"
        price = price_for_group(d, g)
        rows.append({"group": g, "qty": q, "on_hire": oh, "avail": av, "live": live,
                     "tx": tx, "hirers": hr, "days": dys, "ytd": ytd, "rec": rec,
                     "buy": price.get("buy"), "source": price.get("source", "")})
    buy = sorted([r for r in rows if r["rec"] in
                  ("High Priority - No Available Stock", "Consider Additional Stock")],
                 key=lambda r: (-(r["live"] or 0), -r["days"]))
    overstock = sorted([r for r in rows if r["rec"] == "Potential Overstock"],
                       key=lambda r: (r["ytd"] or 0))
    top_used = sorted([r for r in rows if r["days"] > 0], key=lambda r: -r["days"])[:15]
    return {"rows": rows, "buy": buy, "overstock": overstock, "top_used": top_used,
            "today": today, "days_elapsed": days_elapsed,
            "mapping_usable": len(mapping), "mapping_rows": len(d["corr_rows"]),
            "tx_used": n_tx_used, "tx_total": len(d["tx"]),
            "groups_with_days": sum(1 for r in rows if r["days"] > 0)}


def workbook_crosscheck(d):
    """If the Excel workbook is beside the exports, say on the console whether
    its tabs still agree with the live register - nothing on any page reads
    from it any more."""
    wb_path = find_file(WORKBOOK)
    if not wb_path:
        return None
    info = {"mtime": dt.datetime.fromtimestamp(os.path.getmtime(wb_path)),
            "master_rows": None, "master_ties": None, "util_refresh": None}
    try:
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
        if "Master Onhire" in wb.sheetnames:
            tab = sheet_rows(wb["Master Onhire"], "ITEM_BARCODE")
            tab_bc = {clean(r.get("ITEM_BARCODE")).upper() for r in tab if clean(r.get("ITEM_BARCODE"))}
            mine = {r["barcode"].upper() for r in d["master"]}
            info["master_rows"] = len(tab_bc)
            info["master_ties"] = (tab_bc == mine)
        if "Tooling Utilisation" in wb.sheetnames:
            tab = sheet_rows(wb["Tooling Utilisation"], "Equipment Group")
            ests = []
            for r in tab:
                q, dys, y = num(r.get("Total Qty")), num(r.get("Total Hire Days")), \
                    num(r.get("YTD Utilisation %"))
                if q and dys and y:
                    ests.append(dys / (q * y))
            if ests:
                ests.sort()
                implied = int(round(ests[len(ests) // 2]))
                info["util_refresh"] = dt.date(ASAT_DAY.year, 1, 1) + dt.timedelta(days=implied - 1)
        wb.close()
    except Exception as e:  # the cross-check must never stop a build
        info["error"] = str(e)
    return info


# ----------------------------------------------------------------- models ---
def item_order(r):
    """The one order every register on every page uses: company A to Z,
    people before the company's project / workflow accounts, hirer A to Z,
    then within a hirer the longest-held item first and A to Z by
    description (barcode as the final tie-break so the order is stable)."""
    return (N.sort_key(r["company"]), 1 if r["custody_hirer"] else 0,
            N.sort_key(r["hirer"]), N.sort_key(r["account"]),
            r["date"] or ASAT_DAY, N.sort_key(r["desc"]), r["barcode"])


def hirer_groups(rows):
    """Rows of one company -> hirer groups A to Z (people first, then the
    company's project / workflow accounts), items longest-held first then
    A to Z by description. The SiteIQ account travels with each group."""
    by = defaultdict(list)
    for r in rows:
        by[(bool(r["custody_hirer"]), r["hirer"], r["account"])].append(r)
    groups = []
    for (is_acc, hirer, account), items in sorted(
            by.items(), key=lambda kv: (kv[0][0], N.sort_key(kv[0][1]),
                                        N.sort_key(kv[0][2]))):
        items.sort(key=item_order)
        groups.append({"hirer": hirer, "account": account, "is_account": is_acc,
                       "items": items, "vb": value_bits(items)})
    return groups


def company_groups(rows):
    """Customer rows -> companies A to Z, each with its hirer groups, its
    SiteIQ accounts (with counts) and its value bits. Custody rows (the
    Repairs account) are not companies and are left out - callers list
    them separately under their own heading."""
    by_co = defaultdict(list)
    for r in rows:
        if not r["custody"]:
            by_co[r["company"]].append(r)
    out = []
    for co in sorted(by_co, key=N.sort_key):
        items = sorted(by_co[co], key=item_order)
        groups = hirer_groups(items)
        accounts = defaultdict(int)
        for r in items:
            accounts[r["account"]] += 1
        out.append({"name": co, "items": items, "groups": groups,
                    "vb": value_bits(items),
                    "people": [g for g in groups if not g["is_account"]],
                    "accounts_g": [g for g in groups if g["is_account"]],
                    "n_people": sum(1 for g in groups if not g["is_account"]),
                    "n_account_items": sum(len(g["items"]) for g in groups
                                           if g["is_account"]),
                    "oldest": min((r["date"] for r in items if r["date"]), default=None),
                    "siteiq_accounts": sorted(accounts.items(),
                                              key=lambda kv: N.sort_key(kv[0]))})
    return out


def custody_blocks(rows, company_note=True):
    """Custody / holding rows -> the same structure company_groups returns,
    one block per account, A to Z, so the register renderer lists them the
    same way as customers (clearly marked as not customer hire)."""
    by = defaultdict(list)
    for r in rows:
        by[r["hirer"]].append(r)
    out = []
    for acc in sorted(by, key=N.sort_key):
        items = sorted(by[acc], key=item_order)
        cos = defaultdict(int)
        for r in items:
            cos[r["company"]] += 1
        g = {"hirer": acc, "account": acc, "is_account": True, "items": items,
             "vb": value_bits(items)}
        out.append({"name": acc, "items": items, "groups": [g], "vb": value_bits(items),
                    "people": [], "accounts_g": [g], "n_people": 0,
                    "n_account_items": len(items), "custody": True,
                    "oldest": min((r["date"] for r in items if r["date"]), default=None),
                    "siteiq_accounts": sorted(cos.items(), key=lambda kv: N.sort_key(kv[0]))})
    return out


def partial_note(d, months):
    """'September is partial ...' when the window includes the current
    month and the month has not finished - said once, in plain words."""
    if ASAT_DAY.month in months and ASAT_DAY.day < 28:
        nxt = (dt.date(ASAT_DAY.year, ASAT_DAY.month, 28) + dt.timedelta(days=4)).replace(day=1)
        if ASAT_DAY < nxt:
            return (f"{ASAT_DAY.strftime('%B')} is partial - the register was pulled "
                    f"{stamp(d['asat']['stock']['pulled'])}, so it holds only the "
                    f"first {ASAT_DAY.day} days of the month.")
    return ""


def rows_for(d, qk):
    if qk == "YEAR":
        return list(d["master"])
    return [r for r in d["master"] if r["q"] == qk]


def value_bits(rows):
    priced = [r for r in rows if r["cost"] is not None]
    return {"n": len(rows), "value": sum(r["cost"] for r in priced),
            "priced": len(priced), "unpriced": len(rows) - len(priced)}


def quarter_model(d, qk):
    label = QUARTERS[qk][1] if qk != "YEAR" else "Year " + str(ASAT_DAY.year)
    rows = rows_for(d, qk)
    custody = sorted([r for r in rows if r["custody"]], key=item_order)
    client = [r for r in rows if not r["custody"]]
    # WHY (02 Sep 2026): companies A to Z, hirers A to Z inside a company,
    # items longest-held first - the same order as the Tooling On-Hire
    # Report, built by the same grouping code.
    companies = company_groups(client)
    vb = value_bits(rows)
    months = QUARTERS[qk][2] if qk != "YEAR" else tuple(range(1, 13))
    return {"key": qk, "label": label, "rows": rows, "companies": companies,
            "custody": custody, "total_val": vb["value"], "priced": vb["priced"],
            "unpriced": vb["unpriced"], "n_companies": len(companies),
            "n_accounts": sum(1 for r in client if r["custody_hirer"]),
            "partial": partial_note(d, months)}


def company_model(d, name):
    disp = display_company(name)
    mine = sorted([r for r in d["master"] if r["company"] == disp], key=item_order)
    per_q = {qk: [r for r in mine if r["q"] == qk] for qk in QUARTERS}
    tx = [t for t in d["tx"] if t["company"] == disp]
    tx_ytd = len(tx)
    tx_custody = sum(1 for t in tx if t["custody"])
    same_day = sum(1 for t in tx if t["end"] and t["start"] and t["end"] == t["start"])
    people = sorted({t["hirer"] for t in tx if t["hirer"] and not t["custody"]},
                    key=N.sort_key)
    accounts = sorted({t["hirer"] for t in tx if t["hirer"] and t["custody"]},
                      key=N.sort_key)
    high_val = sorted([r for r in mine if (r["cost"] or 0) >= HIGH_VALUE],
                      key=lambda r: -(r["cost"] or 0))
    compliance = sorted([r for r in mine if r["cat"]], key=item_order)
    vb = value_bits(mine)
    # WHY (02 Sep 2026): project accounts (FCCU, SATGAS/MOL) now roll into
    # their company, so a company report carries every SiteIQ account of
    # that customer and says which account each hirer is booked to.
    blocks = company_groups(mine)
    return {"name": disp, "items": mine,
            "people_items": [r for r in mine if not r["custody_hirer"]],
            "account_items": [r for r in mine if r["custody_hirer"]],
            "blocks": blocks,
            "siteiq_accounts": blocks[0]["siteiq_accounts"] if blocks else [],
            "per_q": per_q, "tx_ytd": tx_ytd, "tx_custody": tx_custody,
            "same_day": same_day,
            "same_day_pct": (same_day / tx_ytd) if tx_ytd else None,
            "people": people, "accounts": accounts, "high_val": high_val,
            "compliance": compliance, "total_val": vb["value"], "priced": vb["priced"],
            "unpriced": vb["unpriced"]}


def util_model(d):
    return d["util"]


def compliance_model(d):
    cutoff = ASAT_DAY - dt.timedelta(days=NOT_SIGHTED_DAYS)
    chase = []
    for r in d["master"]:
        if not r["cat"]:
            continue
        st = d["stocktake"].get(r["barcode"])
        seen = st["seen"] if st else None
        if seen is None or seen < cutoff:
            chase.append({**r, "seen": seen})
    chase.sort(key=lambda r: (r["cat"], r["seen"] or dt.date(2000, 1, 1)))
    by_cat = defaultdict(list)
    for c in chase:
        by_cat[c["cat"]].append(c)
    high_val = sorted([r for r in d["master"] if (r["cost"] or 0) >= HIGH_VALUE],
                      key=lambda r: -(r["cost"] or 0))
    # monthly movement trend from transactions (every account, every family -
    # the page says so; the current month is partial and is labelled)
    trend = defaultdict(int)
    for t in d["tx"]:
        if t["start"] and t["start"].year == ASAT_DAY.year:
            trend[t["start"].month] += 1
    trend_rows = []
    for m in range(1, 13):
        if trend.get(m, 0) or m < ASAT_DAY.month:
            lab = MONTHS[m - 1] + (" (partial)" if m == ASAT_DAY.month else "")
            trend_rows.append((lab, trend.get(m, 0)))
    out_of_tag = sorted(d["out_of_tag"], key=lambda r: (N.sort_key(r["desc"]), r["barcode"]))
    return {"chase": chase, "by_cat": dict(by_cat), "out_of_tag": out_of_tag,
            "high_val": high_val, "trend": trend_rows}


def recovery_rows(d):
    """Per-company recovery table from the master list by on-hire quarter.
    Custody accounts (Repairs) come back on their own line, never as a
    company. Everything ties to len(master)."""
    by_co = defaultdict(list)
    custody = []
    for r in d["master"]:
        (custody if r["custody"] else by_co[r["company"]]).append(r)

    def line(name, items):
        row = {"company": name, "q": {}, "items": items}
        for qk in QUARTERS:
            row["q"][qk] = value_bits([r for r in items if r["q"] == qk])
        row["total"] = value_bits(items)
        return row
    companies = [line(co, by_co[co]) for co in sorted(by_co, key=N.sort_key)]
    custody_line = line("Repairs custody account (internal)", custody) if custody else None
    return companies, custody_line


def legacy_model(d):
    leg = d["legacy"]
    fam = defaultdict(int)
    for r in leg:
        fam[r["family"] or "Tooling"] += 1
    by_co = defaultdict(lambda: defaultdict(int))
    oldest_co = {}
    for r in leg:
        by_co[r["company"]][r["family"] or "Tooling"] += 1
        by_co[r["company"]]["_n"] += 1
        if r["date"] and (r["company"] not in oldest_co or r["date"] < oldest_co[r["company"]]):
            oldest_co[r["company"]] = r["date"]
    top = sorted(by_co.items(), key=lambda kv: (-kv[1]["_n"], N.sort_key(kv[0])))[:8]
    az = [(co, by_co[co], oldest_co.get(co)) for co in sorted(by_co, key=N.sort_key)]
    return {"n": len(leg), "oldest": min((r["date"] for r in leg if r["date"]), default=None),
            "families": dict(fam), "tooling_n": fam.get("Tooling", 0), "top": top,
            "az": az, "n_companies": len(by_co)}


def exec_model(d):
    master = d["master"]
    client = [r for r in master if not r["custody"]]
    custody = [r for r in master if r["custody"]]
    vb = value_bits(master)
    companies = sorted({r["company"] for r in client if r["company"]}, key=N.sort_key)
    um = util_model(d)
    cm = compliance_model(d)
    qcounts = {qk: sum(1 for r in master if r["q"] == qk) for qk in QUARTERS}
    tx = d["tx"]
    tx_ytd = len(tx)
    same_day = sum(1 for t in tx if t["end"] and t["start"] and t["end"] == t["start"])
    tx_custody = sum(1 for t in tx if t["custody"])
    tx_family = sum(1 for t in tx if t["family"] and not t["custody"])
    tx_by_custody_co = defaultdict(int)
    for t in tx:
        if t["company"].upper() in INTERNAL_CUSTODY:
            tx_by_custody_co[t["company"]] += 1
    comp_rows, custody_line = recovery_rows(d)
    buy_priced = [r for r in um["buy"] if r["buy"] is not None]
    return {"on_hire": len(master), "total_val": vb["value"], "priced": vb["priced"],
            "unpriced": vb["unpriced"], "companies": companies,
            "custody_n": len(custody), "n_accounts": sum(1 for r in client if r["custody_hirer"]),
            "available": d["available_n"], "available_all": d["available_all_n"],
            "available_family": dict(d["available_family"]),
            "repairs": d["repairs_n"], "status_counts": d["status_counts"],
            "qcounts": qcounts, "tx_ytd": tx_ytd, "same_day": same_day,
            "same_day_pct": (same_day / tx_ytd) if tx_ytd else None,
            "tx_custody": tx_custody, "tx_family": tx_family,
            "tx_by_custody_co": dict(tx_by_custody_co),
            "tx_client_tooling": tx_ytd - tx_custody - tx_family,
            "buy_n": len(um["buy"]), "buy_priced_n": len(buy_priced),
            "overstock_n": len(um["overstock"]),
            "chase_n": len(cm["chase"]), "out_of_tag_n": len(cm["out_of_tag"]),
            "high_val_n": len(cm["high_val"]),
            "high_val_sum": sum(r["cost"] or 0 for r in cm["high_val"]),
            "recovery": comp_rows, "custody_line": custody_line,
            "legacy": legacy_model(d)}


AGE_BANDS = (("0-30 days", 0, 30), ("31-60 days", 31, 60), ("61-90 days", 61, 90),
             ("91-180 days", 91, 180), ("Over 180 days", 181, None))
CAT_ORDER = ("General", "Rigging", "Electrical", "High Torque")


def onhire_model(d):
    """Everything the Tooling On-Hire Report prints, counted once from the
    master list so every page ties to the same total."""
    master = d["master"]
    client = [r for r in master if not r["custody"]]
    custody = sorted([r for r in master if r["custody"]], key=item_order)
    companies = company_groups(client)
    vb = value_bits(master)
    people_rows = [r for r in client if not r["custody_hirer"]]
    people_pairs = {(r["company"], r["hirer"]) for r in people_rows}
    people_names = {h for _, h in people_pairs}
    project_accounts = defaultdict(int)
    for r in client:
        if r["custody_hirer"]:
            project_accounts[(r["hirer"], r["company"])] += 1
    oldest = min(master, key=lambda r: (r["date"] or ASAT_DAY, r["barcode"])) if master else None
    over90 = [r for r in master if r["days"] is not None and r["days"] > 90]
    ageing = []
    for lab, lo, hi in AGE_BANDS:
        band = [r for r in master if r["days"] is not None and lo <= r["days"]
                and (hi is None or r["days"] <= hi)]
        ageing.append((lab, value_bits(band)))
    cats = []
    for c in CAT_ORDER:
        rows = [r for r in master if (r["cat"] or "General") == c]
        cats.append((c, value_bits(rows)))
    months = []
    for m in range(1, ASAT_DAY.month + 1):
        rows = [r for r in master if r["date"] and r["date"].month == m]
        months.append((MONTHS[m - 1] + (" (partial)" if m == ASAT_DAY.month else ""),
                       value_bits(rows)))
    by_pair = defaultdict(list)
    for r in people_rows:
        by_pair[(r["hirer"], r["company"])].append(r)
    top = sorted(by_pair.items(), key=lambda kv: (-len(kv[1]), N.sort_key(kv[0][0]),
                                                  N.sort_key(kv[0][1])))[:15]
    top_hirers = [(h, co, value_bits(rows)) for (h, co), rows in top]
    # the site's former name, as SiteIQ still carries it (counted, never printed)
    former_desc_all = sum(1 for s in d["stock"] if N.carries_former_name(s["desc_raw"]))
    former_desc_here = sum(1 for r in master if N.carries_former_name(r["desc_raw"]))
    former_accounts = len({s["company_raw"] for s in d["stock"]
                           if N.carries_former_name(s["company_raw"])})
    return {"master": master, "client": client, "custody": custody,
            "companies": companies, "n_companies": len(companies),
            "total_val": vb["value"], "priced": vb["priced"], "unpriced": vb["unpriced"],
            "people_items": len(people_rows), "people_pairs": len(people_pairs),
            "people_names": len(people_names),
            "project_accounts": sorted(project_accounts.items(),
                                       key=lambda kv: (N.sort_key(kv[0][0]),
                                                       N.sort_key(kv[0][1]))),
            "account_items": sum(project_accounts.values()),
            "oldest": oldest, "over90": over90, "over90_vb": value_bits(over90),
            "ageing": ageing, "cats": cats, "months": months, "top_hirers": top_hirers,
            "holding": d["holding"], "holding_blocks": custody_blocks(d["holding"]),
            "custody_blocks": custody_blocks(custody),
            "holding_family": dict(d["holding_family"]),
            "custody_family": d["custody_family"],
            "former_desc_all": former_desc_all, "former_desc_here": former_desc_here,
            "former_accounts": former_accounts,
            "corrected_bc": sum(1 for r in master if r["corrected"] == "barcode"),
            "corrected_desc": sum(1 for r in master if r["corrected"] == "description"),
            "legacy": legacy_model(d), "partial": partial_note(d, (ASAT_DAY.month,))}


# ------------------------------------------------- 10/10 pass helpers ---
# WHY (03 Sep 2026): the pieces the 10/10 pass added, each counted from the
# same master list (or the same SiteIQ exports) as page 1, so nothing here
# can disagree with a figure already printed.
NO_PREVIOUS_NOTE = ("No earlier register pull is saved in Data\\previous yet - movement pull "
                    "against pull starts with the next pull (button 28 parks the old "
                    "export automatically).")
CHANGE_CAP = 50      # rows printed per pull-against-pull table (the rest is counted)
LAST24_CAP = 25      # rows printed for the 24 hours before the pull
TREND_MIN_DAYS = 7   # the trend page needs this many recorded days
SCOPE_WORDS = ("tooling items on the SiteIQ register - radios, gas monitors, Dräger "
               "equipment, lanyards and steel coil clamps are excluded, and so are the "
               "tool store's holding accounts and the Company Repairs account")


def tooling_scope(row):
    """pull_diff scope: a register row is in this family when its description
    is tooling - no radio / gas / Dräger / lanyard / steel coil clamp family
    word - the same words build_master applies."""
    return not family_hit(clean_text(row["desc"]))


def _acct_norm(text):
    """'Loading Bay - Out Of Service' and 'Loading Bay Out Of Service' are
    the same account: the register writes the dash, the transactions
    export does not. Upper case, dashes to spaces, one space between words."""
    return re.sub(r"\s+", " ", re.sub(r"\s*-\s*", " ", clean(text).upper())).strip()


_HOLDING_EXACT = {_acct_norm(x) for x in MASTER_HIRER_EXCLUSIONS}
_HOLDING_SUB = tuple(_acct_norm(x) for x in MASTER_HIRER_SUBSTRING_EXCLUSIONS)


def is_holding_account(hirer):
    hu = _acct_norm(hirer)
    return hu in _HOLDING_EXACT or any(x in hu for x in _HOLDING_SUB)


def in_population(row, dated=True):
    """The rest of build_master's rule, applied to a pull_diff row from
    whichever pull it came from: not a holding account, not the COMPANY
    REPAIRS account and (when dated) an on-hire date in the report year."""
    if is_holding_account(row.get("hirer", "")):
        return False
    if clean(row.get("company_raw", row.get("company", ""))).upper() == "COMPANY REPAIRS":
        return False
    if dated:
        od = row.get("on_dt")
        if od is None or od.year != ASAT_DAY.year:
            return False
    return True


def get_changes(d):
    """pull_diff.changes for the tooling population, read once per run and
    kept on d. Every list is then trimmed to build_master's rule (holding
    accounts, COMPANY REPAIRS and pre-year issues out) so the movement ties
    to the population on page 1. None when the exports could not be read -
    the section then says so instead of printing nothing."""
    if "changes" in d:
        return d["changes"]
    try:
        c = pull_diff.changes(scope=tooling_scope)
    except Exception as e:  # a missing export must not stop the build
        print(f"  NOTE: since-the-last-pull could not be read ({e}) - the section says so")
        c = None
    d["changes"] = trim_changes(c) if c is not None else None
    return d["changes"]


def trim_changes(c):
    """Trim every pull_diff list to build_master's rule (see in_population)."""
    for k in ("returned", "issued", "moved"):
        c[k] = [r for r in c[k] if in_population(r)]
    for t in list(c["crossed"]):
        c["crossed"][t] = [r for r in c["crossed"][t] if in_population(r)]
    for k in ("issued", "returned"):
        c["last24"][k] = [r for r in c["last24"][k] if in_population(r, dated=False)]
    return c


def diff_desc(d, r):
    """A movement row's description the way the register prints it: the
    corrected name where the mapping has one, else the cleaned register
    description; the site's former name never prints."""
    corr = (d["corr"].get(r["barcode"].upper())
            or d["corr_by_desc"].get(desc_key(clean_text(r["desc"]))))
    return N.display_desc(corr or clean_text(r["desc"]))


def since_last_pull(bl, d, full=True):
    """The 'Since the last pull' section. full=True is the Tooling On-Hire
    Report's page (the two pull times, four counts, the tables, the 24
    hours before the pull); full=False is the Executive Summary's one
    paragraph (counts and the last-24-hours line)."""
    c = get_changes(d)
    if c is None:
        bl.note("<b>Since the last pull:</b> the register and transactions exports could "
                "not be compared on this run - see the console.")
        return
    l = c["last24"]
    ws, we = l["window"]
    n_i, n_r = len(l["issued"]), len(l["returned"])
    if l["available"]:
        last24_line = (f"In the 24 hours before the pull ({stamp(ws)} to {stamp(we)}): "
                       f"<b>{n_fmt(n_i)}</b> tooling items issued, <b>{n_fmt(n_r)}</b> returned.")
    else:
        last24_line = ("The TRANSACTIONS export was not in Data, so the 24 hours before the "
                       "pull could not be counted.")
    n_back, n_out = len(c["returned"]), len(c["issued"])
    n_moved, n_90 = len(c["moved"]), len(c["crossed"][90])
    if not full:
        if c["have_previous"]:
            bl.note(f"<b>Since the last pull</b> ({stamp(c['prev_time'])} to "
                    f"{stamp(c['cur_time'])}): <b>{n_fmt(n_back)}</b> came back, "
                    f"<b>{n_fmt(n_out)}</b> went out, <b>{n_fmt(n_moved)}</b> changed hands, "
                    f"<b>{n_fmt(n_90)}</b> crossed 90 days while out. {last24_line}")
        else:
            bl.note(f"<b>Since the last pull:</b> {esc(NO_PREVIOUS_NOTE)} {last24_line}")
        return

    year = ASAT_DAY.year
    bl.h2("Since the last pull", pb=True)
    if c["have_previous"]:
        bl.story(f"Pull against pull: the SiteIQ register pulled <b>{stamp(c['prev_time'])}</b> "
                 f"against the register pulled <b>{stamp(c['cur_time'])}</b>, item by item, "
                 f"for the population of this report ({SCOPE_WORDS}; on-hire dates in "
                 f"{year}). Every row below is an item with a barcode - nothing is estimated.")
        bl.tiles([tile(n_fmt(n_back), "Came back", "check", "", "green"),
                  tile(n_fmt(n_out), "Went out", "swap"),
                  tile(n_fmt(n_moved), "Changed hands", "people"),
                  tile(n_fmt(n_90), "Crossed 90 days", "warn", "", "amber")])
        bits = []
        if c["companies_new"]:
            bits.append("new on the register: " + ", ".join(esc(x) for x in c["companies_new"]))
        if c["companies_cleared"]:
            bits.append("cleared (nothing left on hire): "
                        + ", ".join(esc(x) for x in c["companies_cleared"]))
        bl.note("Companies " + ("; ".join(bits) if bits else "- none new, none cleared")
                + f". Crossed 30 / 60 / 90 days while still out: "
                f"{n_fmt(len(c['crossed'][30]))} / {n_fmt(len(c['crossed'][60]))} / "
                f"{n_fmt(n_90)}.")

        def capped(rows, what):
            if len(rows) > CHANGE_CAP:
                bl.note(f"Showing {CHANGE_CAP} of {n_fmt(len(rows))} {what}; the count above "
                        "is the full figure.")
            return rows[:CHANGE_CAP]
        if c["returned"]:
            bl.h2("Came back, company A to Z")
            bl.table(["Company", "Hirer", "Description", "Barcode", "Days it was out"],
                     [[r["company"], display_hirer(r["hirer"]), diff_desc(d, r), r["barcode"],
                       r["days_out"] if r["days_out"] is not None else "-"]
                      for r in capped(c["returned"], "items that came back")],
                     cls="tight", aligns=["", "", "", "", "r"])
        if c["issued"]:
            bl.h2("Went out, company A to Z")
            bl.table(["Company", "Hirer", "Description", "Barcode", "On hire since"],
                     [[r["company"], display_hirer(r["hirer"]), diff_desc(d, r), r["barcode"],
                       fmt_date(r["on_dt"].date() if r["on_dt"] else None)]
                      for r in capped(c["issued"], "items that went out")], cls="tight")
        if c["moved"]:
            bl.h2("Changed hands, company A to Z")
            bl.table(["Company", "Hirer now", "Was with", "Description", "Barcode", "Days out"],
                     [[r["company"], display_hirer(r["hirer"]),
                       f"{display_hirer(r['from_hirer'])}, {r['from_company']}",
                       diff_desc(d, r), r["barcode"],
                       r["days_out"] if r["days_out"] is not None else "-"]
                      for r in capped(c["moved"], "items that changed hands")],
                     cls="tight", aligns=["", "", "", "", "", "r"])
        if c["crossed"][90]:
            bl.h2("Crossed 90 days since the last pull (ranked by days out, oldest first)")
            bl.table(["Company", "Hirer", "Description", "Barcode", "On hire since", "Days out"],
                     [[r["company"], display_hirer(r["hirer"]), diff_desc(d, r), r["barcode"],
                       fmt_date(r["on_dt"].date() if r["on_dt"] else None),
                       r["days_out"] if r["days_out"] is not None else "-"]
                      for r in capped(c["crossed"][90], "items that crossed 90 days")],
                     cls="tight", aligns=["", "", "", "", "", "r"])
    else:
        bl.note(esc(NO_PREVIOUS_NOTE))
    bl.h2("The 24 hours before the pull")
    if l["available"]:
        bl.story(f"Between <b>{stamp(ws)}</b> and <b>{stamp(we)}</b> the SiteIQ TRANSACTIONS "
                 f"export records <b>{n_fmt(n_i)}</b> tooling items issued and "
                 f"<b>{n_fmt(n_r)}</b> returned ({SCOPE_WORDS}). This part needs no "
                 "earlier pull - it is counted from the day's transactions.")
        moves = ([dict(r, kind="Issued") for r in l["issued"]]
                 + [dict(r, kind="Returned") for r in l["returned"]])
        moves.sort(key=lambda r: (N.sort_key(r["company"]), r["at"], r["barcode"]))
        rows = [[r["company"] or "-", display_hirer(r["hirer"]) or "-", diff_desc(d, r),
                 r["barcode"], r["kind"], r["at"].strftime("%d %b %H:%M")]
                for r in moves[:LAST24_CAP]]
        if rows:
            bl.table(["Company", "Hirer", "Description", "Barcode", "Movement", "Time"], rows,
                     cls="tight")
        if len(moves) > LAST24_CAP:
            bl.note(f"Showing {LAST24_CAP} of {n_fmt(len(moves))} movements - company A to Z, "
                    "then time; the counts above are the full figures.")
    else:
        bl.note(last24_line)


def three_things_for(d, rows=None, window=""):
    """The three things to do today, each drawn from the list handed in -
    the master list, or one quarterly window's rows: the largest over-90
    holder, the oldest item out, and the unpriced items (or, when
    everything is priced, the second-largest over-90 holder). Fewer than
    three true items prints fewer - never a made-up one.
    WHY (03 Sep 2026): the Tooling On-Hire Report prints the same three as
    the Executive Summary; each quarterly prints its own from its own
    window (rows, with window naming it, e.g. 'Q3 (Jul-Sep)')."""
    rows = d["master"] if rows is None else rows
    client = [r for r in rows if not r["custody"]]
    due = fmt_date(ASAT_DAY + dt.timedelta(days=CONFIG["rag_due_days"]))
    who = f"Andrew Fisher \u00b7 by {due}"
    win = f" in {window}" if window else ""
    items = []
    by_co = defaultdict(list)
    for r in client:
        if r["days"] is not None and r["days"] > 90:
            by_co[r["company"]].append(r)
    ranked = sorted(by_co.items(), key=lambda kv: (-len(kv[1]), N.sort_key(kv[0])))

    def chase(rank, word):
        co, held = ranked[rank]
        vb = value_bits(held)
        val = (f"{money(vb['value'])} at replacement" if vb["priced"] else "nothing priced")
        if vb["unpriced"]:
            val += f" ({vb['unpriced']} unpriced)"
        return (f"Chase {co} for {plural(len(held))} over 90 days",
                f"the {word} over-90 holder{win}; {val}", who)
    if ranked:
        items.append(chase(0, "largest"))
    oldest = min(client, key=lambda r: (r["date"] or ASAT_DAY, r["barcode"])) if client else None
    if oldest and oldest["date"]:
        items.append((f"Recover {oldest['barcode']} ({N.display_desc(oldest['desc'])}) from "
                      f"{oldest['hirer']}, {oldest['company']}",
                      f"the oldest item out{win}: {oldest['days']} days, since "
                      f"{fmt_date(oldest['date'])}", who))
    unpriced = [r for r in rows if r["cost"] is None]
    if unpriced:
        by_desc = Counter(N.display_desc(r["desc"]) for r in unpriced)
        desc, m = sorted(by_desc.items(), key=lambda kv: (-kv[1], N.sort_key(kv[0])))[0]
        items.append((f"Price the {n_fmt(len(unpriced))} unpriced items{win}",
                      f"{desc} carries {m} of them; the value on page 1 is understated "
                      "until then", who))
    elif len(ranked) > 1:
        items.append(chase(1, "second-largest"))
    return items[:3]


def ageing_panel(bl, companies, heading):
    """'On-hire ageing by company': one stacked bar per company (A to Z,
    custody and holding accounts excluded - the same population as the
    company bars) in the four shared ageing bands, with the counts in a
    table beneath so every segment reconciles to the company's total.
    WHY (03 Sep 2026): the chart carries the companies with ten or more
    items (sh.split_long_tail); the rest are one A to Z line under it with
    their counts, so a forty-row chart never shrinks to the unreadable.
    The table beneath still carries every company and the full total."""
    rows = []
    for co in companies:
        segs = [0, 0, 0, 0]
        for r in co["items"]:
            segs[sh.age_band_index(r["days"])] += 1
        rows.append((co["name"], segs))
    if not rows:
        return
    big, tail = sh.split_long_tail(rows, 10)
    bl.h2(heading)
    if tail:
        scope = (f"{len(big)} of the {len(rows)} companies, A to Z, not ranked - those with ten "
                 f"or more items; the {len(tail)} with fewer are listed under the chart. The "
                 "legend totals are for the charted companies; the table carries every company "
                 "and the full total.")
    else:
        scope = f"All {len(rows)} companies A to Z, not ranked."
    bl.chart(sh.stacked_hbars(big),
             f"Each company's items on hire by days out at the pull ({fmt_date(ASAT_DAY)}): "
             f"0-30, 31-60, 61-90 and over 90 days. {scope} A row's total is the company's "
             "items in the chart above. The Repairs custody account is not a company and is "
             "not drawn.")
    if tail:
        bl.note("Companies with fewer than ten items: "
                + ", ".join(f"{esc(co)} ({sum(segs)})" for co, segs in tail) + ".")
    tot = [sum(s[i] for _c, s in rows) for i in range(4)]
    bl.table(["Company", "0-30 days", "31-60 days", "61-90 days", "Over 90 days", "Items"],
             [[co] + segs + [sum(segs)] for co, segs in rows]
             + [["Total"] + tot + [sum(tot)]],
             cls="tight", aligns=["", "r", "r", "r", "r", "r"])


def trend_days_on_record():
    return len(rh.load().get("tooling", {}))


def trend_page(bl, d, live):
    """'The trend - last 30 days': only when the family has TREND_MIN_DAYS
    recorded days on the scoreboard. live holds today's figures (on_hire,
    over90, value) so the last point is today even on the first build of
    the day, before the build records it. Returns True when drawn."""
    if trend_days_on_record() < TREND_MIN_DAYS:
        return False
    keys = ("on_hire", "over90", "value")
    series = {k: dict(rh.series("tooling", k, ASAT_DT or ASAT_DAY, days=30)) for k in keys}
    for k in keys:
        if live.get(k) is not None:
            series[k][ASAT_DAY] = live[k]
    days = sorted(set().union(*[set(s) for s in series.values()]))
    if len(days) < 2:
        return False
    labels = [dd.strftime("%d %b") for dd in days]
    vals = {k: [series[k].get(dd) for dd in days] for k in keys}
    thousands = [round(v / 1000.0, 1) if v is not None else None for v in vals["value"]]
    bl.h2("The trend - last 30 days", pb=True)
    bl.story(f"Every point is the figure this report printed on that day, read back from "
             f"the movement scoreboard (History\\report_history.json) - {len(days)} days on "
             f"record in the 30 days to {fmt_date(ASAT_DAY)}. A day with no build shows a "
             "gap; nothing is interpolated.")
    bl.chart(sh.line_chart(labels, [("Items on hire", vals["on_hire"]),
                                    ("Out more than 90 days", vals["over90"])],
                           y_label="items"),
             f"Tooling items on hire (issued this year) and the part of them out more than "
             "90 days, by pull day.")
    bl.chart(sh.line_chart(labels, [("Replacement value, priced items", thousands)],
                           y_label="$'000"),
             "Replacement value of the priced items on hire, in $'000, by pull day - "
             "unpriced items are never in it.")
    bl.table(["Day", "Items on hire", "Out more than 90 days", "Replacement value"],
             [[dd.strftime("%d %b %Y"),
               n_fmt(vals["on_hire"][i]) if vals["on_hire"][i] is not None else "-",
               n_fmt(vals["over90"][i]) if vals["over90"][i] is not None else "-",
               money(vals["value"][i]) if vals["value"][i] is not None else "-"]
              for i, dd in enumerate(days)], cls="tight", aligns=["", "r", "r", "r"])
    return True


def trend_line(shown):
    n = trend_days_on_record()
    if shown:
        return (f"Trend page: {n} days on record in History\\report_history.json; "
                "the last 30 are drawn.")
    return f"Trend page: appears once seven days are on record ({n} today)."


# -------------------------------------------------------------- rendering ---
def page(title, subtitle, body, limits, standard_break=False):
    """WHY (02 Sep 2026): standard_break puts Our Standard / Honest Limits on
    a page of their own - the Tooling On-Hire Report's data-and-method page
    is full, and without the break the footer landed alone on a last page."""
    lim = "".join("<li>" + esc(x) + "</li>" for x in limits)
    canon = "  |  ".join(CANON)
    return f"""<!doctype html><html><head><meta charset="utf-8"><title>{esc(title)}</title>
<style>
 @page{{size:A4;margin:10mm 9mm 12mm 9mm;}}
 body{{font-family:Segoe UI,Arial,sans-serif;color:{DARK};margin:0;background:#fff;}}
 .band{{background:{DARK};color:#fff;padding:26px 34px;}}
 .band h1{{margin:0;font-size:24px;letter-spacing:.5px;}}
 .band .sub{{color:{ORANGE};font-weight:700;margin-top:6px;font-size:13px;}}
 .band .meta{{color:#bbb;font-size:11px;margin-top:8px;}}
 .wrap{{padding:22px 34px;}}
 h2{{font-size:15px;border-left:5px solid {ORANGE};padding-left:10px;margin:26px 0 10px 0;text-transform:uppercase;letter-spacing:1px;}}
 p.story{{font-size:12.5px;line-height:1.75;margin:8px 0;}}
 p.note{{font-size:11px;line-height:1.6;margin:6px 0;color:{GREY};}}
 table{{border-collapse:collapse;width:100%;font-size:11px;margin:8px 0;}}
 th{{background:{DARK};color:#fff;text-align:left;padding:6px 8px;font-size:10px;text-transform:uppercase;letter-spacing:.5px;}}
 td{{padding:5px 8px;border-bottom:1px solid #e5e0da;}}
 tr:nth-child(even) td{{background:{LIGHT};}}
 table.tight{{font-size:10px;}} table.tight th{{padding:5px 5px;font-size:9px;}} table.tight td{{padding:4px 5px;}}
 .tiles{{display:flex;gap:10px;flex-wrap:wrap;margin:14px 0;}}
 .tiles.three{{display:grid;grid-template-columns:repeat(3,1fr);}}
 .tiles.three .tile{{min-width:0;}}
 .tile{{border:1px solid #e0dad2;border-bottom:3px solid {ORANGE};padding:12px 16px;min-width:130px;}}
 .tile .v{{font-size:22px;font-weight:800;color:{ORANGE};}}
 .tile .l{{font-size:10px;color:{GREY};text-transform:uppercase;letter-spacing:1px;margin-top:3px;}}
 .good{{color:{GREEN};font-weight:700;}} .warn{{color:{AMBER};font-weight:700;}} .bad{{color:{RED};font-weight:700;}}
 .footer{{margin-top:30px;border-top:3px solid {ORANGE};padding:14px 34px;font-size:10px;color:{GREY};line-height:1.7;}}
 .cochip{{background:{DARK};color:#fff;padding:4px 10px;font-size:11px;font-weight:700;display:inline-block;margin-top:14px;}}
 .subhead{{font-size:11px;font-weight:700;color:{AMBER};margin:10px 0 2px 0;}}
 .chartpanel{{background:#171F2B;border-radius:8px;padding:12px 12px 8px 12px;margin:10px 0 4px 0;max-width:660px;}}
 .chartcap{{font-size:10px;color:{GREY};margin:2px 0 10px 2px;}}
 h2{{break-after:avoid;page-break-after:avoid;}}
 .chartpanel,.tile,.tiles,.cochip,.subhead{{break-inside:avoid;page-break-inside:avoid;}}
 .chartpanel{{break-before:avoid;page-break-before:avoid;}}
 tr{{break-inside:avoid;page-break-inside:avoid;}}
 .pb{{break-before:page;page-break-before:always;}}
 .defbox{{border:1px solid #e0dad2;border-left:5px solid {ORANGE};background:{LIGHT};padding:10px 14px;font-size:11px;line-height:1.65;margin:12px 0;break-inside:avoid;}}
 .defbox b{{display:block;font-size:11.5px;margin-bottom:4px;text-transform:uppercase;letter-spacing:.5px;}}
 .defbox ul{{margin:4px 0 0 18px;padding:0;}}
 .kicker{{font-size:10px;color:{GREY};text-transform:uppercase;letter-spacing:1px;margin:4px 0 6px 0;}}
 table.reg thead{{display:table-header-group;}}
 tr.cohead td{{background:{DARK} !important;color:#fff;padding:7px 10px;border-top:14px solid #fff;}}
 tr.cohead .nm{{font-size:13px;font-weight:800;}}
 tr.cohead .st{{font-size:10.5px;color:#dddddd;margin-left:10px;}}
 tr.cohead .ac{{font-size:10px;color:#f5c9b3;margin-top:3px;}}
 tr.hirer td{{background:#EADFD4 !important;font-weight:700;color:{DARK};font-size:10.5px;padding:5px 10px;}}
 tr.hirer .acc{{font-weight:400;color:{AMBER};}}
 tr.cohead,tr.hirer{{break-after:avoid;page-break-after:avoid;}}
 tbody.keep{{break-inside:avoid;page-break-inside:avoid;}}
 td.num{{text-align:right;white-space:nowrap;}}
 th.num{{text-align:right;}}
</style></head><body>
<div class="band"><h1>Coates &nbsp;|&nbsp; Ampol Tool Store</h1>
<div class="sub">{esc(subtitle)}</div>
<div class="meta">Generated {esc(GENERATED)} &nbsp;|&nbsp; Data as at {esc(DATA_ASAT)}</div>
<div class="meta">The Coates Way &nbsp;|&nbsp; POWERED BY SITEIQ &nbsp;|&nbsp; Author: Andrew Fisher</div></div>
<div class="wrap"><!--BODY-START-->{body}
<h2{' class="pb"' if standard_break else ''}>Our standard</h2>
<p class="story">{esc(LSR_LINE)} Every issue and every return runs through the double scan
- {esc(canon)} Daily stocktakes keep eyes on the fleet: nothing in the store goes
over 30 days without being scanned, and anything damaged is tagged Out of Service on the
spot. We are here to help - if gear is finished with, hand it back to the tool store
and it comes off your list the same day.</p>
<h2>Honest limits</h2>
<ul style="font-size:11px;color:{GREY};line-height:1.7;">{lim}</ul>
<!--BODY-END--></div>
<div class="footer">Data as at {esc(DATA_ASAT)} |
Coates Hire Operations Pty Limited | ABN 50 009 779 338 | www.coates.com.au |
POWERED BY SITEIQ | Care Deeply &middot; Customer Focused &middot; Be Our Best &middot; One Team &middot; Competitive Spirit</div>
</body></html>"""


def tiles(pairs, cls=""):
    """WHY (02 Sep 2026): cls='three' lays the tiles in an even three-column
    grid (labels wrap) - six tiles then fill two rows instead of 2 + 3 + 1."""
    return (f'<div class="tiles{" " + cls if cls else ""}">' +
            "".join(f'<div class="tile"><div class="v">{esc(v)}</div>'
                    f'<div class="l">{esc(l)}</div></div>' for v, l in pairs) +
            "</div>")


def table(headers, rows, cls=""):
    h = "".join("<th>" + esc(x) + "</th>" for x in headers)
    b = "".join("<tr>" + "".join("<td>" + (c if str(c).startswith("<span") else esc(c)) +
                                 "</td>" for c in r) + "</tr>" for r in rows)
    tag = f'<table class="{cls}">' if cls else "<table>"
    return f"{tag}<tr>{h}</tr>{b}</table>"


def chart_block(svg, caption=""):
    """A k2shell SVG chart on its dark panel, sized to sit inside the page
    shell (charts are drawn 636px wide; the wrap leaves ~726px, so they fit
    with room to spare).

    WHY (12 Aug 2026): charts are new to this kit - they ride between CHART
    markers so email_html() can strip them, because Outlook's Word engine
    cannot draw SVG. PDF and HTML get the chart; the email keeps its proven
    inline-table style. The tables the charts sit beside are all still there.
    """
    cap = f'<div class="chartcap">{esc(caption)}</div>' if caption else ""
    return f'<!--CHART--><div class="chartpanel">{svg}</div>{cap}<!--/CHART-->'


def item_rows(items, with_company=False, limit=None):
    out = []
    for r in items[:limit] if limit else items:
        row = []
        if with_company:
            row.append(r["company"] or "-")
        row += [r["hirer"] or "-", r["barcode"], N.display_desc(r["desc"]),
                fmt_date(r["date"]), money(r["cost"]) if r["cost"] is not None else "-"]
        out.append(row)
    return out


ITEM_HDR = ["Hirer", "Barcode", "Description", "On Hire Since", "Replacement"]


def source_limits(d):
    """The as-at lines every page carries - each export's own pull stamp."""
    a = d["asat"]
    tx_bits = "SiteIQ TRANSACTIONS pull " + stamp(a["tx"]["pulled"])
    if a["tx"]["period"]:
        tx_bits += " (report period " + a["tx"]["period"] + ")"
    return [("Data as at: SiteIQ RENTAL_STOCK pull " + stamp(a["stock"]["pulled"])
             + "; " + tx_bits + "; SiteIQ STOCKTAKE pull " + stamp(a["stocktake"]["pulled"])
             + ". Every figure is computed from these exports by this kit - the Excel "
             "workbook is not read for any printed number."),
            "Replacement value is the catalogue new-buy average (Avg Buy Price (New) in "
            "Ampol_ToolStore_Pricing.xlsx); where a description is listed more than once "
            "the first price is used. Unpriced items show a dash and are excluded from "
            "value totals - never estimated."]


FAM_WORDS = {"RADIO": "radios", "DRAGER": "Dräger gas monitors", "GAS MONITOR": "gas monitors",
             "GAS DETECTOR": "gas detectors", "MULTI GAS": "multi-gas monitors",
             "MULTIGAS": "multi-gas monitors", "LANYARD": "lanyards",
             "STEEL COIL CLAMP": "steel coil clamps"}
REG_HDR = ["Description", "Barcode", "On Hire Since", "Days", "Replacement"]


def value_words(vb):
    """'$1,234.00 priced (12 items) | 3 unpriced' - the same words on every
    company header so priced and unpriced are never confused."""
    s = (f'{money(vb["value"])} priced ({plural(vb["priced"])})' if vb["priced"]
         else "nothing priced")
    return s + f' | {vb["unpriced"]} unpriced'


def co_stats(co):
    """The company row's figures: items, hirers, priced value, unpriced count.
    WHY (02 Sep 2026): one place for these words - the PDF register and the
    email register both print exactly this."""
    vb = co["vb"]
    if co.get("custody"):
        return f'{plural(vb["n"])} | {value_words(vb)}'
    stats = f'{plural(vb["n"])} | {plural(co["n_people"], "hirer")}'
    if co["accounts_g"]:
        stats += f' + {plural(len(co["accounts_g"]), "project / workflow account")}'
    return stats + " | " + value_words(vb)


def co_accounts(co):
    """'SiteIQ accounts: a (n) · b (n)' under the company name, or '' when
    the company is booked to one account under its own name."""
    accts = co["siteiq_accounts"]
    if co.get("custody") or len(accts) > 1 or (accts and accts[0][0] != co["name"]):
        return (("SiteIQ company" if co.get("custody") else "SiteIQ accounts")
                + ": " + " &middot; ".join(f"{esc(a)} ({n})" for a, n in accts))
    return ""


def hirer_bits(g, multi_account, co_name=""):
    """The hirer row's figures and its account tag (plain text, no markup)."""
    vb = g["vb"]
    bits = plural(vb["n"])
    if vb["priced"]:
        bits += f' | {money(vb["value"])} priced ({vb["priced"]})'
    if vb["unpriced"]:
        bits += f' | {vb["unpriced"]} unpriced'
    tag = ""
    if g["is_account"]:
        tag = "project / workflow account, not a person"
    elif multi_account and g["account"] != co_name:
        tag = f'&middot; SiteIQ account: {esc(g["account"])}'
    return bits, tag


def co_header_tr(co):
    """The company row of the register (email skin): name, items, hirers,
    priced value, unpriced count, and the SiteIQ accounts rolled into it."""
    acc = co_accounts(co)
    acc = f'<div class="ac">{acc}</div>' if acc else ""
    return (f'<tr class="cohead"><td colspan="{len(REG_HDR)}"><span class="nm">{esc(co["name"])}'
            f'</span><span class="st">{co_stats(co)}</span>{acc}</td></tr>')


def hirer_tr(g, multi_account, co_name=""):
    bits, tag = hirer_bits(g, multi_account, co_name)
    tag = f' <span class="acc">{tag}</span>' if tag else ""
    return (f'<tr class="hirer"><td colspan="{len(REG_HDR)}">{esc(g["hirer"])}{tag} '
            f'- {bits}</td></tr>')


def reg_item_tr(r):
    return ("<tr><td>" + esc(N.display_desc(r["desc"])) + "</td><td>" + esc(r["barcode"])
            + "</td><td>" + fmt_date(r["date"]) + '</td><td class="num">'
            + (str(r["days"]) if r["days"] is not None else "-") + '</td><td class="num">'
            + (money(r["cost"]) if r["cost"] is not None else "-") + "</td></tr>")


def register_table(blocks, keep_rows=3):
    """The register: one table, column headings repeated on every printed
    page, a dark company row, a shaded hirer row, then the items.

    WHY (02 Sep 2026): each hirer's header row and its first rows sit in a
    tbody that will not split, and a company row shares that tbody with the
    first hirer under it - so a company or hirer heading can never be left
    alone at the foot of a page with its items on the next."""
    hdr = "".join(f'<th{" class=num" if h in ("Days", "Replacement") else ""}>{esc(h)}</th>'
                  for h in REG_HDR)
    out = [f'<table class="reg"><thead><tr>{hdr}</tr></thead>']
    for co in blocks:
        multi = len(co["siteiq_accounts"]) > 1
        first = True
        for g in co["groups"]:
            head = co_header_tr(co) if first else ""
            if not (co.get("custody") and len(co["groups"]) == 1):
                head += hirer_tr(g, multi, co["name"])
            first = False
            rows = [reg_item_tr(r) for r in g["items"]]
            out.append('<tbody class="keep">' + head + "".join(rows[:keep_rows]) + "</tbody>")
            if len(rows) > keep_rows:
                out.append("<tbody>" + "".join(rows[keep_rows:]) + "</tbody>")
    out.append("</table>")
    return "".join(out)


def hbar_chart(rows, lab_w=210, col_w=(70, 100), w=636, rowh=22, colour=ORANGE):
    """Horizontal bars on the dark panel with an unclipped label and one or
    two right-hand figures (a count and a value). rows: (label, value, texts).
    WHY (02 Sep 2026): the shared kit's hbars clips labels at 34 characters
    and prints one figure; the register charts need the full company or
    hirer name and a value column beside the count."""
    if not rows:
        return '<div class="note">Nothing recorded in the source.</div>'
    h = len(rows) * rowh + 10
    mx = max(r[1] for r in rows) or 1
    n_cols = max(len(r[2]) for r in rows)
    right_w = sum(col_w[:n_cols]) + 10
    bar_w = w - lab_w - right_w
    font = "Lato, Calibri, sans-serif"
    out = [f'<svg width="{w}" height="{h}" viewBox="0 0 {w} {h}">']
    for i, (lab, v, texts) in enumerate(rows):
        y = 5 + i * rowh
        out.append(f'<text x="0" y="{y + 12}" fill="#C9D6E2" font-family="{font}" '
                   f'font-size="8.8">{esc(lab)}</text>')
        out.append(f'<rect x="{lab_w}" y="{y + 3.5}" width="{bar_w}" height="10" rx="5" '
                   f'fill="#26313D"/>')
        if v > 0:
            out.append(f'<rect x="{lab_w}" y="{y + 3.5}" width="{max(bar_w * v / mx, 6):.1f}" '
                       f'height="10" rx="5" fill="{colour}"/>')
        x = w
        for j, t in enumerate(reversed(list(texts))):
            col = "#FFFFFF" if j == len(texts) - 1 else "#C9D6E2"
            out.append(f'<text x="{x}" y="{y + 12}" text-anchor="end" fill="{col}" '
                       f'font-family="{font}" font-size="9.2" font-weight="700">{esc(t)}</text>')
            x -= col_w[len(texts) - 1 - j] + 10
    out.append("</svg>")
    return "".join(out)


# ------------------------------------------------ one content, two skins ---
# WHY (02 Sep 2026): every report is authored once as a list of content
# blocks. The Coates house frame (k2flow + k2shell) dresses them for the
# PDF and the HTML; the kit's proven Outlook-safe markup dresses the same
# blocks for the email body. Same words, same figures, same order in
# both - only the clothes differ.
TEAM = [{"name": "Andrew Fisher", "role": "Shutdown Manager", "shift": "",
         "email": "andrew.fisher@coates.com.au",
         "blurb": "Runs the Ampol tool store the Coates Way - every issue and every "
                  "return double-scanned, every figure counted from SiteIQ.",
         "lead": True}]
PROJECT = "Ampol Lytton Refinery · Permanent Tool Store"
ASAT_NOTE = "(SiteIQ register pull)"
K_ORANGE = k2shell.K["orange"]

# KEY strip terms per report (colour, TERM, tail) - plain words, no figures,
# so the strip never repeats a number the body already carries.
KEY_EXEC = [("orange", "ON HIRE", "tooling issued this year and still out at the pull"),
            ("blue", "REPLACEMENT", "catalogue new-buy average; a dash is unpriced, never estimated"),
            ("amber", "SIGNALS", "buy, right-size and test-date sightings, each with its evidence")]
KEY_REGISTER = [("orange", "ON HIRE", "issued this year and still out at the pull"),
                ("blue", "A TO Z", "companies, then hirers, items longest-held first"),
                ("amber", "DAYS", "calendar days out since the on-hire date")]
KEY_QUARTER = [("orange", "ON HIRE", "issued in this window and still out"),
               ("blue", "BILLABLE", "not home by quarter close is charged at replacement"),
               ("green", "EASY FIX", "hand it in and it is off the list the same day")]
KEY_UTIL = [("orange", "LIVE", "share of a group on hire at the pull"),
            ("blue", "YTD", "hire days against days available this year"),
            ("amber", "BUY SIGNAL", "demand-backed, with the evidence beside it")]
KEY_COMPLIANCE = [("amber", "SIGHT", "not through our hands in three months"),
                  ("green", "CAUGHT", "parked out of tag by our checks"),
                  ("orange", "HIGH VALUE", "the costliest gear in the field")]

# closing-page cards - plain words, no figures (the figures live in the body)
CARD_HONEST = ("Every figure on these pages is counted by this kit from the SiteIQ exports "
               "named under Data and method - never from a workbook tab or a typed-in "
               "summary. Unpriced items show a dash and never enter a total. Where a rule "
               "or a limit matters, it is written out there in plain words.")
CARD_NAMES = ("The site is Ampol on every page: where SiteIQ still carries the former name "
              "it is shown under the current one, and the Tooling On-Hire Report's data "
              "page says how many lines do. One customer, one name: project accounts roll "
              "up to their company and the SiteIQ account is shown against the hirer. "
              "Hirer names are as SiteIQ records them.")
HOW_EXEC = ("The tiles are the position; the tables beneath carry the same figures by "
            "quarter and by company, A to Z. A ranked table says so in its heading. The "
            "detail sits in the companion reports named on the last page.")
HOW_REGISTER = ("Every table is A to Z unless its heading says ranked by. Days are calendar "
                "days from the on-hire date to the pull. A dash in a value column means no "
                "catalogue price - never an estimate. Anything finished with: hand it to "
                "the counter and it is off the list the same day.")
HOW_UTIL = ("Live is the share of a group on hire at the pull; YTD is hire days against "
            "days available this year. Buy signals are ranked by live utilisation then "
            "hire days; right-size candidates by YTD utilisation, lowest first. A dash "
            "means no catalogue price.")
HOW_COMPLIANCE = ("Each family lists the gear we have not had hands on for three months, "
                  "oldest sighting first. Caught and parked means our own checks found it "
                  "before it could go out. High-value items are ranked by replacement cost.")

# extra CSS on top of the house frame: the tooling family's own pieces
# (a body paragraph, the rules list, the dark definition panel, a compact
# register table with dark company rows and shaded hirer rows)
EXTRA_CSS = """
.k2body p.para { font-size: 10.6px; line-height: 1.7; color: #35404E; margin: 10px 0 0 0; }
.k2body p.para b { color: #16202C; font-weight: 700; }
.k2body ul.rul { margin: 8px 0 0 16px; padding: 0; font-size: 9.5px; color: #47566A;
                 line-height: 1.7; }
.k2body ul.rul li { margin: 0 0 4px 0; }
.k2body ul.rul { break-inside: avoid; page-break-inside: avoid; }
.k2body .alerts .al-b { color: #C9D6E2; font-size: 9.6px; line-height: 1.55; }
.k2body .alerts .al td { padding-bottom: 8px; }
.k2body .chart-cap { break-before: avoid; page-break-before: avoid; }
.k2body .sect + .note, .k2body .sect + p.para { break-after: avoid; page-break-after: avoid; }
table.dt.xs th { padding: 6px 5px; font-size: 7.4px; letter-spacing: 0.6px; }
table.dt.xs td { padding: 4px 5px; font-size: 8.3px; line-height: 1.25; }
table.dt.xs th:first-child, table.dt.xs td:first-child { padding-left: 9px; }
table.dt.xs th:last-child, table.dt.xs td:last-child { padding-right: 9px; }
table.dt.reg th { padding: 7px 8px; font-size: 7.8px; letter-spacing: 1px; }
table.dt.reg td { padding: 3.5px 8px; font-size: 8.9px; line-height: 1.25; }
table.dt.reg td:first-child { padding-left: 11px; }
table.dt.reg td:last-child { padding-right: 11px; }
/* the register as group tables (03 Sep 2026): the company title row sits in
   the thead with the column row, so it repeats on every page the company
   runs over; the hirer rows are shaded rows in the body */
.k2body table.dt.grp.reg { margin-top: 12px; }
.k2body table.dt.grp.reg thead { break-after: avoid; page-break-after: avoid; }
.k2body table.dt.grp.reg thead tr.gt th { background: #1B2532; padding: 7px 11px 6px 11px;
                                          line-height: 1.35; }
.k2body table.dt.grp.reg thead tr.gt th .gn { font-size: 11.4px; font-weight: 700; }
.k2body table.dt.grp.reg thead tr.gt th .gm { margin-left: 0; font-size: 8.4px; color: #C9D6E2; }
.k2body table.dt.grp.reg thead tr.gt th .st { margin-left: 10px; }
.k2body table.dt.grp.reg thead tr.gt th .ac { display: block; font-size: 8.1px; color: #F5B58A;
                                              margin-top: 2px; }
table.dt.reg tr.hr td { background: #E6EAEF; color: #16202C; font-weight: 700;
                        font-size: 8.9px; padding: 5px 11px; }
table.dt.reg tr.hr .acc { font-weight: 400; color: #B45309; }
table.dt.reg tr.hr { break-after: avoid; page-break-after: avoid; }
table.dt.reg tbody.keep { break-inside: avoid; page-break-inside: avoid; }
table.dt.reg tr.z td { background: #F7F8FA; }
/* company mini-scorecard in the title row (03 Sep 2026): share-of-fleet bar
   and a four-band ageing strip, counts printed beside */
.k2body table.dt.grp.reg thead tr.gt th .sc { display: block; margin-top: 5px; font-size: 7.4px;
                                              color: #C9D6E2; line-height: 1; white-space: nowrap; }
.k2body table.dt.grp.reg thead tr.gt th .sc .sck { color: #8A9AAC; font-weight: 700; letter-spacing: 1px;
                                                   text-transform: uppercase; font-size: 6.4px;
                                                   margin-right: 5px; }
.k2body table.dt.grp.reg thead tr.gt th .sc .scv { margin: 0 14px 0 5px; }
.k2body table.dt.grp.reg thead tr.gt th .sc .scv b { color: #FFFFFF; }
.k2body table.dt.grp.reg thead tr.gt th .sc svg { vertical-align: middle; }
.k2body .ragband { margin-top: 10px; }
/* WHY (03 Sep 2026): the flowing cover's "What's inside" block sits 40 mm
   from the foot of the cover; with ten rows it reached the freshness line
   under the as-at stamp. The cover text starts 8 mm higher here so the two
   never meet - a shared-frame change is asked for in the layout notes. */
.k2body .fcover .cover-in { padding-top: 22mm; }
/* the Executive Summary position page (03 Sep 2026): compact tiles so the
   band and the three things share the page with them */
.k2body .cpt .tiles { border-spacing: 7px 5px; }
.k2body .cpt .tiles td { padding: 9px 8px 8px 8px; }
.k2body .cpt .t-ico { margin-bottom: 5px; }
.k2body .cpt .t-num { font-size: 23px; }
.k2body .cpt .t-num.sm { font-size: 18px; }
.k2body .cpt .t-lab { margin-top: 5px; }
.k2body .cpt .t-note { margin-top: 3px; }
.k2body .cpt .t-spark { margin-top: 2px; }
.k2body .cpt + .ragband { margin-top: 6px; }
.k2body .callout.tight { line-height: 1.75; }
.k2body .three { margin-top: 6px; padding: 5px 12px 3px 12px; }
.k2body .three .t3 { padding: 3px 0; }
/* the insights pass (03 Sep 2026): a sub-heading's lead-in note stays with
   the table under it, so a data-quality block never leaves its heading and
   its one-line explanation alone at the foot of a page */
.k2body .sub-h + .note { break-after: avoid; page-break-after: avoid; }
"""


def over90_words(over_rows):
    """The cover's over-90 line (03 Sep 2026): the count, then the priced
    value in brackets - no brackets at all when nothing is over 90 days,
    and 'nothing priced' when none of the items carries a price."""
    vb = value_bits(over_rows)
    if not over_rows:
        return f"<b>{n_fmt(0)}</b> items out more than 90 days"
    tail = f"{money(vb['value'])} priced" if vb["priced"] else "nothing priced"
    return f"<b>{n_fmt(len(over_rows))}</b> items out more than 90 days ({tail})"


def k2cfg(title, kicker, key_items):
    return {"client": "Ampol", "title": title, "kicker": kicker, "project": PROJECT,
            "asat_note": ASAT_NOTE, "key_items": key_items, "team": TEAM}


NOTE_HEX = {"grey": "#7A8A9A", "green": "#22C55E", "red": "#F0603E", "amber": "#EFA82B"}


def tile(value, label, icon="box", note="", ncls="grey", email_label=None,
         key=None, raw=None, good="down", spark=None):
    """One KPI tile. The page shows label + a small note under the figure;
    the email keeps its one-line label (email_label, or label when not given).
    WHY (03 Sep 2026): key + raw name the figure in the movement scoreboard
    (report_history). When an earlier day holds the same key the note
    becomes the movement ("▲ 16 since 02 Sep"); when there is none the
    note given here stays - a report never invents an arrow. good says
    which way is the good way ("down", "up", or None for a figure with no
    good direction - it moves in grey). spark is an optional series drawn
    under the note (only where a report has a genuine series to show)."""
    if key is not None and raw is not None:
        mv, mcls = rh.movement("tooling", key, ASAT_DAY, raw, good,
                               money=key.startswith("value") or key.endswith("exposure"))
        if mv:
            note, ncls = mv, mcls
    return (value, label, icon, note, ncls,
            label if email_label is None else email_label, spark)


class Blocks:
    """The content of one report, in order, skin-free."""

    def __init__(self):
        self.items = []

    def story(self, html):
        self.items.append(("story", html))

    def note(self, html):
        self.items.append(("note", html))

    def tiles(self, items, cls=""):
        self.items.append(("tiles", items, cls))

    def h2(self, text, pb=False):
        self.items.append(("h2", text, pb))

    def subh(self, text):
        # WHY (03 Sep 2026, insights pass): a plain sub-heading inside a
        # section (the four data-quality blocks) - not a section, so it
        # never lands on the cover or in the bookmarks
        self.items.append(("subh", text))

    def table(self, headers, rows, cls="", aligns=None):
        self.items.append(("table", headers, rows, cls, aligns))

    def chart(self, svg, caption=""):
        self.items.append(("chart", svg, caption))

    def register(self, blocks, fleet=None):
        # fleet: the on-hire total the company mini-scorecards are a share of
        self.items.append(("register", blocks, fleet))

    def rag(self, band, tight=False):
        # the page-1 RAG band (a rag_over90 dict) - page skin only;
        # tight is the shorter form (03 Sep 2026: the Executive Summary,
        # so the three things fit under it on the position page)
        self.items.append(("rag", band, tight))

    def defbox(self, title, items):
        self.items.append(("defbox", title, items))

    def ul(self, items):
        self.items.append(("ul", items))

    def three(self, items):
        # WHY (03 Sep 2026): the three-things block, under the band
        self.items.append(("three", items))

    def divider(self, title, sub, note=""):
        # WHY (03 Sep 2026): the full-page APPENDIX divider before a register
        self.items.append(("divider", title, sub, note))

    def page_break(self):
        # WHY (03 Sep 2026, layout pass): closes the position page - the
        # next block starts on the page after (page skin only; the email
        # has no pages)
        self.items.append(("page_break",))

    def email_end(self):
        self.items.append(("email_end",))


def email_three(items):
    """The three things in the email's plain table markup."""
    rows = "".join(f"<tr><td>{i}</td><td><b>{esc(w)}</b><br>{esc(y)}</td><td>{esc(o)}</td></tr>"
                   for i, (w, y, o) in enumerate(items, 1))
    return ('<div class="subhead">Three things to do today</div>'
            f'<table><tr><th>#</th><th>What and why</th><th>Owner / by</th></tr>{rows}</table>')


def email_body(bl):
    """The blocks in the kit's proven Outlook-safe markup - what the email
    body carried before the house frame arrived, byte for byte."""
    out = []
    for b in bl.items:
        k = b[0]
        if k == "story":
            out.append(f"<p class='story'>{b[1]}</p>")
        elif k == "note":
            out.append(f"<p class='note'>{b[1]}</p>")
        elif k == "tiles":
            out.append(tiles([(t[0], t[5]) for t in b[1]], b[2] if b[2] == "three" else ""))
        elif k == "h2":
            out.append(f'<h2 class="pb">{b[1]}</h2>' if b[2] else f"<h2>{b[1]}</h2>")
        elif k == "subh":
            out.append(f'<div class="subhead">{esc(b[1])}</div>')
        elif k == "table":
            out.append(table(b[1], b[2], b[3]))
        elif k == "chart":
            out.append(chart_block(b[1], b[2]))
        elif k == "register":
            out.append(register_table(b[1]))
        elif k == "rag":
            pass  # the band is page-1 chrome; the attached position card carries it
        elif k == "defbox":
            out.append(f'<div class="defbox"><b>{b[1]}</b><ul>'
                       + "".join(f"<li>{x}</li>" for x in b[2]) + "</ul></div>")
        elif k == "ul":
            out.append("<ul style='font-size:11px;line-height:1.7;'>"
                       + "".join("<li>" + esc(r) + "</li>" for r in b[1]) + "</ul>")
        elif k == "three":
            out.append(email_three(b[1]))
        elif k == "divider":
            out.append(f"<h2>Appendix - {esc(b[1])}</h2><p class='note'>{b[2]}"
                       + (f" {esc(b[3])}." if b[3] else "") + "</p>")
        elif k == "email_end":
            out.append("<!--EMAIL-END-->")
    return "".join(out)


# ---- the house skin (k2flow frame + k2shell primitives)
RIGHT_COLS = {"items", "value", "days", "hire days", "qty", "on hire", "avail", "live %",
              "ytd %", "hirers", "buy price", "replacement", "transactions",
              "priced / unpriced", "priced value", "legacy items", "items still on hire",
              "total items", "total value"}


def k2_aligns(headers):
    """Figures sit on the right, words on the left - by column heading."""
    out = []
    for h in headers:
        u = str(h).strip().lower()
        out.append("r" if (u in RIGHT_COLS or u.endswith(" items") or u.endswith(" value"))
                   else "")
    return out


def k2_cell(c):
    s = str(c)
    return s if s.startswith("<span") else esc(c)


def k2_table(headers, rows, cls="", aligns=None):
    aligns = aligns or k2_aligns(headers)
    if len(headers) >= 10:
        k = "xs"
    elif cls == "tight" or len(headers) >= 7 or len(rows) > 14:
        k = "cp"
    else:
        k = ""
    body = []
    for r in rows:
        cells = [k2_cell(c) for c in r]
        if cells and cells[0].strip().upper() == "TOTAL":
            cells = [f"<b>{c}</b>" for c in cells]
        body.append(cells)
    if len(body) > 14:
        return k2flow.dtable_flow(headers, body, aligns, k)
    # a short table moves to the next page whole rather than splitting
    # away from its header row
    return f'<div class="keep">{sh.dtable(headers, body, aligns, k)}</div>'


def k2_tiles(items):
    n = len(items)
    per = 4 if n <= 4 else 3
    # WHY (03 Sep 2026): tiles_plus draws the optional sparkline (7th slot)
    return sh.tiles_plus([(t[2], t[0], t[1], t[3], t[4]) + ((t[6],) if len(t) > 6 and t[6] else ())
                          for t in items], per_row=per)


def card_tiles(items):
    """The page-1 tiles as the position card wants them: value, label,
    note, note colour - the same values the page prints."""
    return [(t[0], t[1], t[3], NOTE_HEX.get(t[4], NOTE_HEX["grey"])) for t in items]


def rag_over90(n_over, n_all, what):
    """The page-1 RAG band on the share of items out more than 90 days.
    Returns the pieces both skins use: status, the headline in words (HTML
    and plain), the rule with its lines, the owner, the next action with
    its due date (pull date + CONFIG days). Lines come from CONFIG."""
    share = (n_over / n_all * 100.0) if n_all else 0.0
    amber, red = CONFIG["over90_amber_pct"], CONFIG["over90_red_pct"]
    status = sh.rag_of(share, amber, red)
    line = {"green": f"under the {amber:g}% amber line",
            "amber": f"above the {amber:g}% amber line, below the {red:g}% red line",
            "red": f"above the {red:g}% red line"}[status]
    head_txt = (f"{n_fmt(n_over)} of the {n_fmt(n_all)} {what} have been out more than "
                f"90 days ({share:.1f}%) - {line}.")
    head_html = (f'<span class="o">{n_fmt(n_over)}</span> of the {n_fmt(n_all)} {esc(what)} '
                 f'have been out more than 90 days (<span class="o">{share:.1f}%</span>) - '
                 f'{line}.')
    rule = (f"Share of {what} out more than 90 days: Green under {amber:g}%, Amber from "
            f"{amber:g}%, Red from {red:g}% (default line - set in CONFIG).")
    due = fmt_date(ASAT_DAY + dt.timedelta(days=CONFIG["rag_due_days"]))
    # the phone card's band is a fixed height and wraps its headline to two
    # lines and its action to two - so the card gets the short form of the
    # same sentence, the due date always on it
    card_head = (f"{n_fmt(n_over)} of {n_fmt(n_all)} on hire out more than 90 days "
                 f"({share:.1f}%) - {line}.")
    card_action = (f"Over-90-day list to each company's supervisor at the quarter-close "
                   f"run - by {due}.")
    return {"status": status, "share": share, "head_html": head_html, "head_txt": head_txt,
            "rule": rule, "owner": CONFIG["rag_owner"],
            "action_html": f"{esc(CONFIG['rag_action'])} - by <b>{due}</b>.",
            "action_txt": f"{CONFIG['rag_action']} - by {due}.",
            "card_head": card_head, "card_action": card_action}


def card_scores(on_hire, priced, same_day_pct, over90_n, buy_n, buy_priced_n):
    """Up to four 0-100 scores for the position card, each a share the
    page already prints: priced on-hire lines, same-day returns, items
    inside 90 days, buy signals priced."""
    def share(a, b):
        return int(round(a / b * 100.0)) if b else 0
    out = [("Priced on-hire lines", share(priced, on_hire))]
    if same_day_pct is not None:
        out.append(("Same-day returns, year to date", int(round(same_day_pct * 100.0))))
    out.append(("Items inside 90 days", share(on_hire - over90_n, on_hire)))
    out.append(("Buy signals priced", share(buy_priced_n, buy_n)))
    return out


def k2_chart(svg, caption=""):
    cap = f'<div class="chart-cap">{esc(caption)}</div>' if caption else ""
    return f'<div class="keep"><div class="chartpanel">{svg}</div>{cap}</div>'


def k2_defbox(title, items):
    rows = "".join(f'<tr><td class="al-dot d-amber">●</td><td><div class="al-b">{x}</div></td></tr>'
                   for x in items)
    return f'<div class="alerts"><div class="ah">{title}</div><table class="al">{rows}</table></div>'


def k2_ul(items):
    return '<ul class="rul">' + "".join("<li>" + esc(r) + "</li>" for r in items) + "</ul>"


REG_ALIGN = ["", "", "", "r", "r"]
# the register scorecard's ageing strip: label, low, high, colour
AGE_STRIP = (("0-30", 0, 30, "#1FA75A"), ("31-90", 31, 90, "#F5A623"),
             ("91-180", 91, 180, "#F36F21"), ("180+", 181, None, "#EF4444"))
_SC_ID = [0]


def co_scorecard(co, fleet, w=90, h=7):
    """WHY (03 Sep 2026): the company row's mini-scorecard - a share-of-fleet
    bar (the company's items as a share of the on-hire total) and a four-band
    ageing strip (0-30 / 31-90 / 91-180 / over 180 days). The counts and the
    share are printed beside the bars, so nothing depends on colour alone."""
    n = co["vb"]["n"]
    share = (n / fleet) if fleet else 0.0
    _SC_ID[0] += 1
    cid = f"sc{_SC_ID[0]}"
    clip = (f'<defs><clipPath id="{cid}"><rect x="0" y="0" width="{w}" height="{h}" '
            f'rx="3.5"/></clipPath></defs>')
    track = f'<rect x="0" y="0" width="{w}" height="{h}" rx="3.5" fill="#33414F"/>'
    fill = (f'<rect x="0" y="0" width="{max(w * share, 1.5):.1f}" height="{h}" '
            f'fill="#F36F21" clip-path="url(#{cid})"/>' if n else "")
    share_svg = (f'<svg width="{w}" height="{h}" viewBox="0 0 {w} {h}">{clip}{track}{fill}</svg>')
    counts = []
    for lab, lo, hi, col in AGE_STRIP:
        c = sum(1 for r in co["items"] if r["days"] is not None and lo <= r["days"]
                and (hi is None or r["days"] <= hi))
        counts.append((lab, c, col))
    tot = sum(c for _l, c, _c in counts)
    x, segs = 0.0, []
    for lab, c, col in counts:
        if c:
            sw = w * c / tot
            segs.append(f'<rect x="{x:.1f}" y="0" width="{sw:.1f}" height="{h}" fill="{col}" '
                        f'clip-path="url(#{cid}a)"/>')
            x += sw
    aclip = (f'<defs><clipPath id="{cid}a"><rect x="0" y="0" width="{w}" height="{h}" '
             f'rx="3.5"/></clipPath></defs>')
    age_svg = (f'<svg width="{w}" height="{h}" viewBox="0 0 {w} {h}">{aclip}{track}'
               f'{"".join(segs)}</svg>')
    txt = " &middot; ".join(f"{lab}: <b>{c}</b>" for lab, c, _c in counts)
    # WHY (03 Sep 2026): a span, not a div - it now sits inside the group
    # table's title cell (an inline span); the CSS shows it as a block
    return (f'<span class="sc"><span class="sck">Share of fleet</span>{share_svg}'
            f'<span class="scv"><b>{share * 100:.1f}%</b> of {n_fmt(fleet)} on hire</span>'
            f'<span class="sck">Ageing, days</span>{age_svg}<span class="scv">{txt}</span></span>')


def k2_reg_item_tr(r, i):
    z = ' class="z"' if i % 2 else ""
    return (f"<tr{z}><td>{esc(N.display_desc(r['desc']))}</td>"
            f'<td class="nw">{esc(r["barcode"])}</td>'
            f'<td class="nw">{fmt_date(r["date"])}</td>'
            f'<td class="r">{r["days"] if r["days"] is not None else "-"}</td>'
            f'<td class="r nw">{money(r["cost"]) if r["cost"] is not None else "-"}</td></tr>')


def co_meta(co, fleet):
    """The company title row's meta (03 Sep 2026): the stats, the SiteIQ
    accounts and the mini-scorecard, as inline pieces the group table's
    title cell carries beside the name."""
    acc = co_accounts(co)
    acc = f'<span class="ac">{acc}</span>' if acc else ""
    sc = co_scorecard(co, fleet) if (fleet and not co.get("custody")) else ""
    return f'<span class="st">{co_stats(co)}</span>{acc}{sc}'


def k2_register_table(blocks, keep_rows=3, fleet=None):
    """The register in house clothes: one k2flow.group_table per company,
    so the company's name sits in the thead beside the column row and is
    printed again at the top of every page the company runs over - a
    forty-page register never needs a flip back to find whose items these
    are. Inside it the hirer rows are shaded rows and the items sit
    longest-held first.
    WHY (03 Sep 2026): the keep rule is unchanged - a hirer heading and its
    first rows share a tbody that will not split, so a heading is never
    left alone at the foot of a page. group_table builds the thead (and an
    empty tbody); that tbody is swapped for these keep tbodies, which is
    the only way to have both the repeating title and the keep rule.
    fleet: when given, every company row carries the mini-scorecard (share
    of that total, the ageing strip); custody and holding blocks are not
    companies and never get one."""
    out = []
    empty = "<tbody></tbody></table>"
    for co in blocks:
        multi = len(co["siteiq_accounts"]) > 1
        shell = k2flow.group_table(co["name"], co_meta(co, fleet), REG_HDR, [], REG_ALIGN, "reg")
        if not shell.endswith(empty):
            raise RuntimeError("k2flow.group_table changed shape - the register cannot be built")
        bodies = []
        for g in co["groups"]:
            head = ""
            if not (co.get("custody") and len(co["groups"]) == 1):
                bits, tag = hirer_bits(g, multi, co["name"])
                tag = f' <span class="acc">{tag}</span>' if tag else ""
                head = (f'<tr class="hr"><td colspan="{len(REG_HDR)}">{esc(g["hirer"])}{tag} '
                        f'- {bits}</td></tr>')
            rows = [k2_reg_item_tr(r, i) for i, r in enumerate(g["items"])]
            bodies.append('<tbody class="keep">' + head + "".join(rows[:keep_rows]) + "</tbody>")
            if len(rows) > keep_rows:
                bodies.append("<tbody>" + "".join(rows[keep_rows:]) + "</tbody>")
        out.append(shell[:-len(empty)] + "".join(bodies) + "</table>")
    return "".join(out)


def k2_tail(limits, how, data_heading=True):
    """Every report ends the same way: Data and method (the honest limits,
    with the pull stamps and rules), then a closing page of cards - Our
    standard / Honest limits / Names as shown / How to read this - and the
    tool store team."""
    out = []
    if data_heading:
        out.append('<div class="sect"><h3>Data and method</h3></div>')
    out.append('<div class="sub-h">Honest limits</div>')
    out.append(k2_ul(limits))
    canon = "  |  ".join(CANON)
    standard = (esc(LSR_LINE) + " Every issue and every return runs through the double scan "
                "- " + esc(canon) + " Daily stocktakes keep eyes on the fleet: nothing in "
                "the store goes over 30 days without being scanned, and anything damaged is "
                "tagged Out of Service on the spot. We are here to help - if gear is "
                "finished with, hand it back to the tool store and it comes off your list the "
                "same day.")
    cards = sh.info_cards([("Our standard", standard), ("Honest limits", CARD_HONEST),
                           ("Names as shown", CARD_NAMES), ("How to read this", how)])
    # WHY (03 Sep 2026): the Coates Way panel sits above the team on every
    # closing page - the cog, the objective, the values, the counting promise
    out.append('<div class="pb"><div class="sect"><h3>Our standard, honest limits and how to '
               'read this report</h3></div>' + cards + sh.coates_way_panel(traits=(3, 4), disciplines=(2, 6), line="every over-90-day item has a named holder; the quarterly recovery cycle is the cadence that keeps the fleet honest")
               + '<div class="sub-h">Your Coates tool store team</div>' + sh.team_cards(TEAM)
               + '<div class="note" style="margin-top:16px;text-align:center">Coates Hire '
               'Operations Pty Limited &middot; ABN 50 009 779 338 &middot; www.coates.com.au '
               '&middot; Care Deeply &middot; Customer Focused &middot; Be Our Best &middot; '
               'One Team &middot; Competitive Spirit</div></div>')
    return "".join(out)


def k2_body(bl, limits, how, data_heading=True):
    """The blocks in the house skin. The first story is 'The position' -
    the answer in one breath - in the peach callout; later stories are body
    paragraphs; headings are section panels; tables, tiles and charts are
    the shared k2shell pieces.
    WHY (03 Sep 2026, layout pass): a page_break block (the Executive
    Summary, the Tooling On-Hire Report and the quarterlies emit one after
    their story) makes whatever block follows it start on the page after -
    a heading becomes a page-break section, anything else is wrapped in
    one - so the position page holds band, tiles, three things and the
    story, nothing more."""
    out, first, pb_next = [], True, False
    for b in bl.items:
        k = b[0]
        piece = ""
        if k == "story":
            if first:
                cls = "callout tight" if len(b[1]) > 700 else "callout"
                out.append(f'<div class="{cls}"><span class="lead">The position.</span> '
                           f'{b[1]}</div>')
                first = False
                continue
            piece = f'<p class="para">{b[1]}</p>'
        elif k == "note":
            piece = f'<div class="note">{b[1]}</div>'
        elif k == "tiles":
            t = k2_tiles(b[1])
            piece = f'<div class="cpt">{t}</div>' if b[2] == "compact" else t
        elif k == "h2":
            piece = f'<div class="sect{" pb" if (b[2] or pb_next) else ""}"><h3>{b[1]}</h3></div>'
            pb_next = False
        elif k == "subh":
            piece = f'<div class="sub-h">{esc(b[1])}</div>'
        elif k == "table":
            piece = k2_table(b[1], b[2], b[3], b[4])
        elif k == "chart":
            piece = k2_chart(b[1], b[2])
        elif k == "register":
            piece = k2_register_table(b[1], fleet=b[2])
        elif k == "rag":
            r = b[1]
            piece = sh.rag_band(r["status"], r["head_html"], esc(r["rule"]),
                                esc(r["owner"]), r["action_html"],
                                tight=bool(b[2]) if len(b) > 2 else False)
        elif k == "defbox":
            piece = k2_defbox(b[1], b[2])
        elif k == "ul":
            piece = k2_ul(b[1])
        elif k == "three":
            piece = sh.three_things(b[1])
        elif k == "divider":
            piece = k2flow.divider_block(b[1], b[2], b[3])
        elif k == "page_break":
            pb_next = True
            continue
        else:
            continue          # email_end marks where the email stops - nothing on the page
        if pb_next and piece:
            piece = f'<div class="pb">{piece}</div>'
            pb_next = False
        out.append(piece)
    out.append(k2_tail(limits, how, data_heading))
    return "".join(out)


def report_outputs(title, subtitle, subject, bl, limits, page_title, page_sub, cfg, how,
                   standard_break=False, data_heading=True, cover=None, pdf_subject="",
                   card=None):
    """One report, ready to write: the page document wears the house frame;
    the email document is the kit's legacy page, kept only so email_html()
    can lift the Outlook-safe body from it. cover (03 Sep 2026): a function
    of the cover's contents list (the (title, page) rows) returning a
    k2flow.cover_block - the page is built once without it, printed, and
    built again with the page numbers read off that print (build). pdf_subject:
    the one plain sentence stamped into the PDF's properties; card: the phone
    position card's values."""
    email_doc = page(page_title, page_sub, email_body(bl), limits, standard_break)
    body_html = k2_body(bl, limits, how, data_heading)

    def build(contents=None):
        return k2flow.flow_doc(cfg, GENERATED, ASAT_SHORT, body_html, extra_css=EXTRA_CSS,
                               cover=cover(contents) if cover else None)
    return {"title": title, "subtitle": subtitle, "doc": build(), "build": build,
            "mail": email_doc, "subject": subject, "pdf_subject": pdf_subject or subject,
            "has_cover": cover is not None, "card": card,
            # a report with a page_break block keeps its position page to
            # band, tiles, three things and the story - that page must hold
            # room for tomorrow's movement notes; a report without one
            # simply flows full
            "position_page": any(b[0] == "page_break" for b in bl.items)}


# ------------------------------------------------ the insights pass ---
# WHY (03 Sep 2026, the insights pass): the TRANSACTIONS export holds every
# issue and return since 01 Jan and the family was reading a slice of it.
# txn_insights reads it once per build and answers, for this report's own
# population (the master rows' barcodes, a quarterly window's barcodes, or
# the whole store), the questions a client asks next: what crosses 90 days
# by the quarter close unless it comes back (arithmetic on the on-hire
# date, never a forecast), how long a product is normally out, what never
# moved, where the fleet has headroom, who holds what across the store, the
# counter's rhythm by hour, and what the log and the register do not agree
# on. Every figure is a count or a sum over rows in the exports; every new
# section carries a one-line "So what" under its chart or table; every new
# page's source is named on the data-and-method page. Page 1 keeps its
# grammar - the new material sits on pages after the position.
QC_CAP_ONHIRE = 60      # quarter-close rows printed on the Tooling On-Hire Report
QC_CAP_QUARTER = 40     # ... and on a quarterly
QUALITY_CAP = 12        # sample rows per data-quality block
WEEKDAYS = ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun")


def insights(d):
    """The txn_insights context - register, log and pricing - loaded once
    per build (about 12 s) and kept on d for every report that follows."""
    if "ctx" not in d:
        t0 = time.time()
        ctx = ti.load_all()
        d["ctx"] = ctx
        w0, w1 = ctx["tx_window"]
        print(f"Transaction log     : {len(ctx['tx']):,} movements ({stamp(w0)} to {stamp(w1)}) "
              f"against {len(ctx['reg']):,} register barcodes - read once, {time.time() - t0:.0f} s")
    return d["ctx"]


def log_period(ctx):
    w0, w1 = ctx["tx_window"]
    return f"{stamp(w0)} to {stamp(w1)}" if (w0 and w1) else "year to date"


def master_barcodes(d):
    """This family's on-hire population - the master rows' barcodes."""
    return {r["barcode"] for r in d["master"] if r["barcode"]}


def tooling_barcodes(d):
    """Every register barcode whose description is tooling - the same
    family words build_master applies (radios, gas monitors, Dräger
    equipment, lanyards and steel coil clamps out) - every status, every
    account. The utilisation pages' population."""
    return {s["barcode"] for s in d["stock"]
            if s["barcode"] and not family_hit(clean_text(s["desc_raw"]))}


def next_close(day):
    """The last day of the quarter that holds `day`."""
    q = (day.month - 1) // 3
    return dt.date(day.year, q * 3 + 3, [31, 30, 30, 31][q])


def window_close(qk):
    """A quarterly window's own quarter end; the Year's is the pull
    quarter's end (the next close)."""
    if qk == "YEAR":
        return next_close(ASAT_DAY)
    m = QUARTERS[qk][2][-1]
    return dt.date(ASAT_DAY.year, m, [31, 30, 30, 31][(m - 1) // 3])


def product_labels(d):
    """product key -> the corrected name's product key, where every
    corrected register row under that key agrees on one (display only -
    the grouping stays the engine's). WHY (03 Sep 2026): the engine keys
    products off the SiteIQ description, so a typo the mapping corrects
    ('Critictal Risk Signage') would otherwise print as SiteIQ has it."""
    if "product_labels" in d:
        return d["product_labels"]
    seen = defaultdict(set)
    for s in d["stock"]:
        if not s["barcode"]:
            continue
        corr = (d["corr"].get(s["barcode"].upper())
                or d["corr_by_desc"].get(desc_key(clean_text(s["desc_raw"]))))
        if corr:
            seen[ti.product_key(N.display_desc(s["desc_raw"]))].add(ti.product_key(corr))
    d["product_labels"] = {k: next(iter(v)) for k, v in seen.items() if len(v) == 1}
    return d["product_labels"]


def product_show(d, key):
    """A product key the way the register pages print descriptions:
    the corrected name where the mapping has one for every item under
    it, proper case, the former site name as Ampol."""
    return m_proper(N.display_desc(product_labels(d).get(key, key)))


def hirer_show(name):
    """A person as SiteIQ records them; a custody, holding or project
    account named as an account, never as a person."""
    h = display_hirer(name) or "-"
    if is_holding_account(name) or is_custody_hirer(name):
        return f"{h} (account)"
    return h


def desc_show(d, barcode, desc):
    """A description the way the register prints it: the corrected name
    where the mapping has one for the barcode or for every row of that
    description, else the cleaned SiteIQ text."""
    corr = (d["corr"].get(clean(barcode).upper())
            or d["corr_by_desc"].get(desc_key(clean_text(desc))))
    return N.display_desc(corr or clean_text(desc))


def dfmt(x):
    """Days: one decimal under ten, whole days from ten."""
    if x is None:
        return "-"
    return f"{x:.1f}" if x < 10 else f"{x:,.0f}"


def so_what(bl, text):
    """The one-line caption every new section carries under its chart or
    table: what it shows and what we do about it."""
    bl.note(f"<b>So what:</b> {text}")


def line_chart_signed(labels, series, y_label="", w=636, h=210, colours=None,
                      partial=()):
    """Lines over time on the dark panel with a y axis that runs from the
    lowest value (or zero) to the highest, and a zero line - so a week
    where returns beat issues (net below zero) is drawn where it belongs.
    Same look as sh.line_chart, which starts its axis at zero; this file
    must not change the shared chart, so the signed one lives here.
    partial: indexes drawn as hollow points (a week that is not complete)."""
    n = len(labels)
    if n < 2 or not series:
        return '<div class="note">Not enough weeks on record yet for a trend line.</div>'
    colours = colours or [sh.K["orange"], "#22C55E", "#5DADE2", "#EFA82B"]
    top, base, pad_r = 26, h - 26, 44
    allv = [v for _, vs in series for v in vs if v is not None]
    if not allv:
        return '<div class="note">Not enough weeks on record yet for a trend line.</div>'
    hi, lo = max(max(allv), 0), min(min(allv), 0)
    if hi == lo:
        hi = lo + 1
    ticks = [hi - (hi - lo) * k / 4 for k in range(5)]
    widest = max(len(f"{round(t):,}") for t in ticks)
    pad_l = max(34, int(widest * 4.6) + 10)
    step = max(1, (n - 1) // 8 or 1)

    def x_of(i):
        return pad_l + (w - pad_l - pad_r) * i / (n - 1)

    def y_of(v):
        return top + (base - top) * (hi - v) / (hi - lo)
    out = [f'<svg width="{w}" height="{h}" viewBox="0 0 {w} {h}">']
    for t in ticks:
        y = y_of(t)
        out.append(f'<line x1="{pad_l}" y1="{y:.1f}" x2="{w - pad_r}" y2="{y:.1f}" '
                   f'stroke="#2A3644" stroke-width="1"/>')
        out.append(f'<text x="{pad_l - 5}" y="{y + 3:.1f}" text-anchor="end" fill="#8A9AAC" '
                   f'font-family="Lato, Calibri, sans-serif" font-size="7.6">{round(t):,}</text>')
    if lo < 0:
        y0 = y_of(0)
        out.append(f'<line x1="{pad_l}" y1="{y0:.1f}" x2="{w - pad_r}" y2="{y0:.1f}" '
                   f'stroke="#8A9AAC" stroke-width="1.2"/>')
    for i, lab in enumerate(labels):
        if i % step == 0 or i == n - 1:
            out.append(f'<text x="{x_of(i):.1f}" y="{base + 14}" text-anchor="middle" fill="#8A9AAC" '
                       f'font-family="Lato, Calibri, sans-serif" font-size="7.6">{esc(str(lab))}</text>')
    lx = pad_l
    ends = []
    for (name, vals), c in zip(series, colours):
        pts = [(i, x_of(i), y_of(v)) for i, v in enumerate(vals) if v is not None]
        if len(pts) > 1:
            out.append('<polyline points="' + " ".join(f"{x:.1f},{y:.1f}" for _i, x, y in pts)
                       + f'" fill="none" stroke="{c}" stroke-width="2.2" stroke-linejoin="round" '
                       'stroke-linecap="round"/>')
        for i, x, y in pts:
            if i in partial:
                out.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" fill="#171F2B" stroke="{c}" '
                           'stroke-width="1.6"/>')
            else:
                out.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="2.4" fill="{c}"/>')
        if pts:
            last = [v for v in vals if v is not None][-1]
            ends.append([pts[-1][2] + 3.5, pts[-1][1] + 6, c, f"{last:,}"])
        out.append(f'<rect x="{lx}" y="4" width="9" height="9" rx="2" fill="{c}"/>')
        out.append(f'<text x="{lx + 13}" y="12" fill="#C9D6E2" font-family="Lato, Calibri, sans-serif" '
                   f'font-size="8.2">{esc(name)}</text>')
        lx += 13 + 5.6 * len(name) + 16
    # the last values, nudged apart where two lines end at the same height
    prev = None
    for e in sorted(ends, key=lambda e: e[0]):
        y = e[0] if prev is None or e[0] - prev >= 9 else prev + 9
        prev = y
        out.append(f'<text x="{e[1]:.1f}" y="{y:.1f}" fill="{e[2]}" '
                   f'font-family="Lato, Calibri, sans-serif" font-size="8.6" font-weight="700">{e[3]}</text>')
    if y_label:
        out.append(f'<text x="{w - pad_r}" y="12" text-anchor="end" fill="#8A9AAC" '
                   f'font-family="Lato, Calibri, sans-serif" font-size="7.6">{esc(y_label)}</text>')
    out.append("</svg>")
    return "".join(out)


def weekly_block(bl, ctx, ws, what):
    """The two weekly charts and their table: issues, returns and net out
    by ISO week, then the same-day share. The partial weeks - the first
    when the log begins inside it, and the current one - are starred on
    the axis, drawn hollow, and named in the caption."""
    w0, w1 = ctx["tx_window"]
    partial = {i for i, w in enumerate(ws) if w["partial"]}
    if w0 and ws and ws[0]["start"] < w0.date():
        partial.add(0)
    labels = [w["week"] + ("*" if i in partial else "") for i, w in enumerate(ws)]
    bits = []
    if 0 in partial:
        bits.append(f"{ws[0]['week']} (the log begins {fmt_date(w0.date())})")
    cur = [w for i, w in enumerate(ws) if i in partial and i != 0]
    if cur:
        bits.append(f"{cur[-1]['week']} (runs to the pull, {stamp(w1)})")
    star = (" * partial weeks: " + "; ".join(bits) + ".") if bits else ""
    peak = max(ws, key=lambda w: (w["issues"], w["start"]))
    bl.chart(line_chart_signed(labels, [("Issues", [w["issues"] for w in ws]),
                                        ("Returns", [w["returns"] for w in ws]),
                                        ("Net out", [w["net"] for w in ws])],
                               y_label="movements per week", partial=partial),
             f"{what} by ISO week (week starting): issues by the start time, returns by the end "
             f"time, net out = issues less returns in the week (below the zero line when returns "
             f"beat issues). Busiest week {peak['week']}: {n_fmt(peak['issues'])} issues.{star}")
    sd_ok = [w for w in ws if w["sd_pct"] is not None]
    hi = max(sd_ok, key=lambda w: (w["sd_pct"], w["start"])) if sd_ok else None
    lo = min(sd_ok, key=lambda w: (w["sd_pct"], w["start"])) if sd_ok else None
    bl.chart(sh.line_chart(labels, [("Same-day %", [w["sd_pct"] for w in ws])],
                           y_label="% of the week's returned issues", pct=True),
             "Same-day share by week: of the week's issues that have come back, the share "
             "that came back the day they went out."
             + (f" Highest {hi['week']} ({hi['sd_pct']:g}%), lowest {lo['week']} "
                f"({lo['sd_pct']:g}%)." if hi and lo else "") + star)
    bl.table(["Week starting", "Issues", "Returns", "Net out", "Same-day %"],
             [[w["week"] + (" (partial)" if i in partial else ""), n_fmt(w["issues"]),
               n_fmt(w["returns"]), f"{w['net']:+,}" if w["net"] else "0",
               f"{w['sd_pct']:g}%" if w["sd_pct"] is not None else "-"]
              for i, w in enumerate(ws)]
             + [["Total", n_fmt(sum(w["issues"] for w in ws)), n_fmt(sum(w["returns"] for w in ws)),
                 f"{sum(w['net'] for w in ws):+,}", "-"]],
             cls="tight", aligns=["", "r", "r", "r", "r"])


def crossing_tables(bl, d, qc, cap):
    """The quarter-close company table (A to Z) and the item rows (company
    A to Z, then the day they cross), capped at `cap` with the count said."""
    by_bc = {r["barcode"].upper(): r for r in d["master"]}

    def co_name(r):
        m = by_bc.get(r["barcode"].upper())
        co = (m["company"] if m else r["company"]) or "-"
        return f"{co} (custody account, internal)" if co.upper() in INTERNAL_CUSTODY else co
    per = defaultdict(lambda: {"n": 0, "value": 0.0, "unpriced": 0, "first": None})
    for r in qc["rows"]:
        p = per[co_name(r)]
        p["n"] += 1
        if r["price"]:
            p["value"] += r["price"]
        else:
            p["unpriced"] += 1
        if p["first"] is None or r["crosses"] < p["first"]:
            p["first"] = r["crosses"]
    co_rows = [[co, n_fmt(p["n"]), fmt_date(p["first"]),
                money(p["value"]) if p["value"] else "-", n_fmt(p["unpriced"])]
               for co, p in sorted(per.items(), key=lambda kv: N.sort_key(kv[0]))]
    co_rows.append(["Total", n_fmt(qc["n"]), fmt_date(qc["rows"][0]["crosses"]) if qc["rows"] else "-",
                    money(qc["value"]) if qc["value"] else "-", n_fmt(qc["unpriced"])])
    bl.table(["Company", "Items crossing", "Earliest crossing", "Priced value", "Unpriced"],
             co_rows, aligns=["", "r", "", "r", "r"])
    rows = sorted(qc["rows"], key=lambda r: (N.sort_key(co_name(r)), r["crosses"], r["barcode"]))
    shown = rows[:cap]
    if len(rows) > cap:
        bl.note(f"Showing {cap} of {n_fmt(len(rows))} items - company A to Z, then the day they "
                "cross; the counts above are the full figures.")
    if shown:
        bl.table(["Company", "Hirer", "Barcode", "Description", "Out since", "Days now", "Crosses on"],
                 [[co_name(r),
                   hirer_show(by_bc[r["barcode"].upper()]["hirer"] if r["barcode"].upper() in by_bc
                              else r["hirer"]),
                   r["barcode"], desc_show(d, r["barcode"], r["desc"]),
                   fmt_date(r["on_dt"].date()), r["days"], fmt_date(r["crosses"])]
                  for r in shown], cls="tight", aligns=["", "", "", "", "", "r", ""])
    return per


def exec_insights(bl, d):
    """The Executive Summary's four store-wide pages (after the position,
    before the data page): the store's year in movements, who holds what,
    the counter's rhythm, and what the log and the register disagree on.
    Every figure here is the whole store - every account, every family -
    and says so. Returns what the data page quotes."""
    ctx = insights(d)
    w0, w1 = ctx["tx_window"]
    period = log_period(ctx)

    # ---- 1a. the store's year in movements
    ws = ti.weekly_series(ctx, None)
    tot_i, tot_r = sum(w["issues"] for w in ws), sum(w["returns"] for w in ws)
    peak = max(ws, key=lambda w: (w["issues"], w["start"]))
    pos_weeks = sum(1 for w in ws if w["net"] > 0)
    bl.h2("The store's year in movements", pb=True)
    bl.story(f"Every issue and every return in the SiteIQ TRANSACTIONS export (CUSTOMER_CONTRACTOR_"
             f"EQUIP, {esc(period)}) - every account and every family, not only the tooling on "
             f"page 1 - by the ISO week it happened: <b>{n_fmt(tot_i)}</b> issues and "
             f"<b>{n_fmt(tot_r)}</b> returns over {len(ws)} weeks. A return is counted in the week "
             "it came back, so a week's net out is what went out and stayed out that week.")
    weekly_block(bl, ctx, ws, "The whole store's movements")
    so_what(bl, f"{peak['week']} was the busiest week ({n_fmt(peak['issues'])} issues) and net out "
                f"ran positive in {pos_weeks} of {len(ws)} weeks - the year's issues are "
                f"{n_fmt(tot_i - tot_r)} ahead of its returns, which is the gear the quarter-close "
                "list brings home. Roster the counter to the peak weeks, not the average week.")

    # ---- 1b. who holds what
    ho = ti.holders(ctx, None)
    bl.h2("Who holds what - ranked by items held, top 20 of the store's holders", pb=True)
    bl.story(f"Every item On Hire on the SiteIQ RENTAL_STOCK register at the pull, every family, "
             f"grouped by hirer and company: <b>{n_fmt(ho['holders'])}</b> holders carry "
             f"<b>{n_fmt(ho['items'])}</b> items ({money(ho['value'])} priced). This table is "
             "ranked by items held; custody, holding and project accounts are named as accounts - "
             "they are workflows, not people. Oldest is the days out of the holder's longest-held "
             f"item at {fmt_date(ASAT_DAY)}; families are gas, radio and tooling.")
    bl.table(["#", "Hirer", "Company", "Items", "Priced value", "Unpriced", "Oldest (days)", "Families held"],
             [[i, hirer_show(h["hirer"]), h["company"], n_fmt(h["items"]),
               money(h["value"]) if h["value"] else "-", n_fmt(h["unpriced"]), n_fmt(h["oldest"]),
               ", ".join(sorted(h["families"]))]
              for i, h in enumerate(ho["top"], 1)],
             cls="tight", aligns=["r", "", "", "r", "r", "r", "r", ""])
    bl.note(f"<b>{n_fmt(ho['n80_items'])}</b> of the {n_fmt(ho['holders'])} holders carry 80% of the "
            f"items on hire; <b>{n_fmt(ho['n80_value'])}</b> carry 80% of the priced value.")
    cross = ho["cross_family"]
    three = sum(1 for h in cross if len(h["families"]) >= 3)
    bl.subh(f"{n_fmt(len(cross))} holders hold two or three families at once "
            f"({n_fmt(three)} hold all three) - the top five by items")
    bl.table(["Hirer", "Company", "Items", "Families held"],
             [[hirer_show(h["hirer"]), h["company"], n_fmt(h["items"]), ", ".join(sorted(h["families"]))]
              for h in cross[:5]], aligns=["", "", "r", ""])
    so_what(bl, f"one conversation per holder recovers more than one per item: the "
                f"{n_fmt(ho['n80_items'])} holders carrying 80% of the items are the list to walk "
                f"first, and the {n_fmt(len(cross))} cross-family holders get one combined list, "
                "not a radio list, a gas list and a tooling list.")

    # ---- 1c. the counter's rhythm
    cr = ti.counter_rhythm(ctx, None)
    ret_total = sum(v for row in cr["returns"] for v in row)
    bl.h2("The counter's rhythm", pb=True)
    bl.story(f"Every movement in the TRANSACTIONS export by the weekday and hour it was scanned: "
             f"<b>{n_fmt(cr['total'])}</b> draws by their start time and <b>{n_fmt(ret_total)}</b> "
             "returns by their end time. The count is printed in every cell; the shade follows "
             "the count against the grid's busiest cell, and an empty cell is a zero.")
    hours = [f"{h:02d}" for h in range(24)]
    bl.chart(sh.heatgrid(cr["draws"], list(WEEKDAYS), hours),
             "Draws by weekday and hour of the start scan, Monday to Sunday, 00 to 23 hours "
             "(24-hour clock, the hour the movement started in).")
    bl.chart(sh.heatgrid(cr["returns"], list(WEEKDAYS), hours, colour=(34, 197, 94)),
             "Returns by weekday and hour of the end scan - the same grid, the hour the movement "
             "closed in.")
    busy = ", ".join(f"<b>{h:02d}:00</b> ({n_fmt(c)} draws)" for h, c in cr["busiest"])
    dawn = sum(cr["draws"][dd][hh] for dd in range(7) for hh in (4, 5))
    tot = cr["total"] or 1
    bl.note(f"The three busiest hours for draws: {busy}. Day (06:00 to 17:59): "
            f"<b>{n_fmt(cr['day'])}</b> draws ({cr['day'] / tot * 100:.0f}%); night (18:00 to 05:59): "
            f"<b>{n_fmt(cr['night'])}</b> ({cr['night'] / tot * 100:.0f}%). The two hours 04:00 to "
            f"05:59 alone carry {n_fmt(dawn)} draws ({dawn / tot * 100:.0f}%).")
    so_what(bl, f"the counter's peak is the pre-dawn shift start - {dawn / tot * 100:.0f}% of all "
                "draws land between 04:00 and 05:59 - so that is the window to have the double scan "
                "fully staffed; the quiet afternoon hours are where a bay-by-bay stocktake sweep fits.")

    # ---- 1d. what the log and the register disagree on
    dq = ti.data_quality(ctx, None)
    bl.h2("What the log and the register disagree on", pb=True)
    bl.story("Four things the SiteIQ TRANSACTIONS log and the RENTAL_STOCK register do not agree "
             "on, or that look like a scan habit rather than a hire - the whole store, every account "
             "and family. Nothing here is corrected; it is shown so the reader knows what the "
             f"figures on every other page are built on. The log begins {stamp(w0)} and ends "
             f"{stamp(w1)}; the register was pulled {stamp(ASAT_DT)}.")
    # short hires
    bl.subh(f"{n_fmt(dq['short_n'])} movements closed inside {dq['short_minutes']} minutes")
    bl.note(f"Out and back inside {dq['short_minutes']} minutes - a scan to test a tag, a wrong item "
            "handed straight back, or a barcode scanned twice at the counter. They sit inside the "
            "transactions count and the same-day figure on page 1. The 12 most recent of the "
            f"{n_fmt(dq['short_n'])}:")
    short = sorted(dq["short"], key=lambda t: t["st"], reverse=True)[:QUALITY_CAP]
    bl.table(["Hirer", "Company", "Description", "Barcode", "Out", "Back", "Minutes"],
             [[hirer_show(t["who"]), t["co"] or "-", desc_show(d, t["bc"], t["desc"]), t["bc"],
               t["st"].strftime("%d %b %H:%M"), t["en"].strftime("%d %b %H:%M"),
               f"{t['hours'] * 60:.1f}"] for t in short],
             cls="tight", aligns=["", "", "", "", "", "", "r"])
    # mass draws
    bl.subh(f"{n_fmt(dq['mass_n'])} mass draws - one person, {dq['mass_threshold']} or more items in one hour")
    bl.note("A mass draw is a kit or a crew lead drawing for a whole crew in one visit - every "
            "item scanned to one name in one hour. It is a scan habit to know about, not an error: "
            "the gear is on hire to the person who drew it, not to the people carrying it. The 12 "
            f"largest of the {n_fmt(dq['mass_n'])}, ranked by items:")
    bl.table(["Hirer", "Company", "Day", "Hour", "Items"],
             [[hirer_show(k[0]), k[1] or "-", fmt_date(k[2]), f"{k[3]:02d}:00 to {k[3]:02d}:59", n_fmt(v)]
              for k, v in dq["mass"][:QUALITY_CAP]],
             aligns=["", "", "", "", "r"])
    # on hire, no movement in the log
    nolog = sorted(dq["onhire_no_log"], key=lambda r: (r["on_dt"] or ASAT_DT, r["barcode"]))
    after = sum(1 for r in nolog if r["on_dt"] and w1 and r["on_dt"] > w1)
    bl.subh(f"{n_fmt(len(nolog))} items on hire with no movement in the log since it began")
    if nolog:
        tail = (f" All {n_fmt(after)} were issued on {fmt_date(w1.date())} after {w1:%H:%M}, when the "
                "TRANSACTIONS export's period ends - they are newer than the log, not missing from it."
                if after == len(nolog) else
                (f" {n_fmt(after)} of them were issued after the export's period ends ({stamp(w1)}) "
                 "and are newer than the log, not missing from it." if after else ""))
        bl.note(f"The register shows these On Hire, and the log holds no movement for their barcode "
                f"at all.{tail}")
        bl.table(["Company", "Hirer", "Barcode", "Description", "On hire since"],
                 [[r["company"], hirer_show(r["hirer"]), r["barcode"], desc_show(d, r["barcode"], r["desc"]),
                   stamp(r["on_dt"])] for r in nolog[:QUALITY_CAP]], cls="tight")
    else:
        bl.note("None - every item the register shows On Hire has a movement in the log.")
    before = sorted(dq["onhire_before_log"], key=lambda r: (r["on_dt"] or ASAT_DT, r["barcode"]))
    fam = Counter(ti.report_family(r["desc"]) for r in before)
    fam_bits = ", ".join(f"{n_fmt(n)} {f}" for f, n in sorted(fam.items(), key=lambda kv: (-kv[1], kv[0])))
    bl.subh(f"{n_fmt(len(before))} items on hire issued before the log begins - history, not a gap")
    bl.note(f"Separately, {n_fmt(len(before))} items on hire were issued before {fmt_date(w0.date())}, "
            f"where the TRANSACTIONS export starts ({fam_bits}) - the register carries their on-hire "
            "date, the log cannot. They are the legacy rows the Radio and Gas Monitor reports chase. "
            f"The 12 oldest, ranked by on-hire date:")
    bl.table(["Company", "Hirer", "Barcode", "Description", "On hire since"],
             [[r["company"], hirer_show(r["hirer"]), r["barcode"], desc_show(d, r["barcode"], r["desc"]),
               stamp(r["on_dt"])] for r in before[:QUALITY_CAP]], cls="tight")
    # available, but the latest movement is open
    oa = dq["open_but_available"]
    bl.subh(f"{n_fmt(len(oa))} register-available items whose latest movement is still open")
    if oa:
        bl.table(["Barcode", "Description", "Bay", "Hirer on the open movement", "Company", "Out since"],
                 [[r["barcode"], desc_show(d, r["barcode"], r["desc"]), r["unit"] or "-",
                   hirer_show(t["who"]), t["co"] or "-", t["st"].strftime("%d %b %Y %H:%M")]
                  for r, t in oa[:QUALITY_CAP]], cls="tight")
    else:
        bl.note("None - every item the register shows Available for Hire has its newest movement "
                "closed in the log. The register and the log agree on every available item.")
    so_what(bl, "the six-minute closes and the mass draws are habits, not errors - a word at the "
                "counter about scanning a kit as a kit keeps the same-day figure honest; the register "
                "and the log agree on every available item, so the on-hire pages can be trusted to "
                "the item.")
    return {"period": period, "tx_n": len(ctx["tx"]), "reg_n": len(ctx["reg"]),
            "qc_all": ti.quarter_close(ctx, None),
            "qc_tool": ti.quarter_close(ctx, master_barcodes(d))}


def onhire_insights(bl, d, n):
    """The Tooling On-Hire Report's three log-based sections, after 'Since
    the last pull': the quarter close look forward, the return windows by
    product, and the year in movements - each for this report's population
    (the master rows' barcodes), so every figure ties to page 1. Returns
    the one sentence page 1's band adds under its rule (or '')."""
    ctx = insights(d)
    pop = master_barcodes(d)
    w0, w1 = ctx["tx_window"]

    # ---- 2a. quarter close - the look forward
    qc = ti.quarter_close(ctx, pop)
    qa = ti.quarter_close(ctx, None)
    qend = qc["qend"]
    bl.h2("Quarter close - the look forward", pb=True)
    bl.story(f"Of the <b>{n_fmt(n)}</b> tooling items on hire, <b>{n_fmt(qc['n'])}</b> will have been "
             f"out 90 days or more by <b>{fmt_date(qend)}</b> unless they come back - "
             f"{money(qc['value'])} at replacement ({n_fmt(qc['n'] - qc['unpriced'])} priced / "
             f"{n_fmt(qc['unpriced'])} unpriced). This is arithmetic on each item's on-hire date and "
             f"nothing more: an item under 90 days out at the pull ({fmt_date(ASAT_DAY)}) whose "
             f"on-hire date is on or before {fmt_date(qend - dt.timedelta(days=90))} crosses the "
             f"line before the close. A further <b>{n_fmt(qc['already_over'])}</b> are already over "
             "90 days today - the band on page 1 - and are not counted again here. Across the whole "
             f"register, every family and every account, {n_fmt(qa['n'])} items will have been out "
             f"90 days or more by {fmt_date(qa['qend'])} unless they come back.")
    bl.tiles([tile(n_fmt(qc["n"]), f"Cross 90 days by {qend:%d %b}", "warn", "", "amber"),
              tile(money(qc["value"]), "Replacement value", "shield",
                   f"{n_fmt(qc['n'] - qc['unpriced'])} priced / {n_fmt(qc['unpriced'])} unpriced"),
              tile(n_fmt(qc["already_over"]), "Already over 90 days", "clock", "", "amber"),
              tile(n_fmt(len(qc["by_company"])), "Companies", "layers")])
    per = crossing_tables(bl, d, qc, QC_CAP_ONHIRE)
    if per:
        top_co, top_p = max(per.items(), key=lambda kv: (kv[1]["n"], N.sort_key(kv[0])))
        so_what(bl, f"these {n_fmt(qc['n'])} items are the quarter-close charge in waiting - "
                    f"{top_co} carries the most ({n_fmt(top_p['n'])}, from {fmt_date(top_p['first'])}); "
                    f"a supervisor's walk of this list before {fmt_date(qend)} brings them home "
                    "before they become a charge at replacement cost.")
    else:
        so_what(bl, f"nothing on hire crosses 90 days by {fmt_date(qend)} - the over-90 list on page 1 "
                    "is the whole recovery task for this close.")

    # ---- 2b. return windows by product
    rw = ti.return_windows(ctx, pop)
    a = rw["all"]
    rule_extra = ""
    bl.h2(f"Return windows by product - ranked by completed hires, top 25 of {n_fmt(len(rw['rows']))} products",
          pb=True)
    if a:
        bl.story(f"How long the items on hire today are normally out: the TRANSACTIONS log holds "
                 f"<b>{n_fmt(a['n'])}</b> completed hires this year (out and back, any hirer) for the "
                 f"{n_fmt(n)} barcodes on page 1. Median hold <b>{dfmt(a['median'])} days</b>; nine "
                 f"in ten were back inside <b>{dfmt(a['p90'])} days</b>; {a['sd_pct']:g}% came back the "
                 "same day. By product - the SiteIQ description with its size and serial tail removed - "
                 f"for the products with ten or more completed hires; {n_fmt(rw['pooled_n'])} hires "
                 f"sit in products with fewer than {rw['min_n']} each and are pooled into the total, "
                 "not listed.")
        bl.table(["Product", "Hires", "Median days", "90th-percentile days", "Same-day %"],
                 [[product_show(d, r["product"]), n_fmt(r["n"]), dfmt(r["median"]), dfmt(r["p90"]),
                   f"{r['sd_pct']:g}%"] for r in rw["rows"][:25]]
                 + [["Total", n_fmt(a["n"]), dfmt(a["median"]), dfmt(a["p90"]), f"{a['sd_pct']:g}%"]],
                 cls="tight", aligns=["", "r", "r", "r", "r"])
        bl.note(f"Total = every completed hire of this population this year ({n_fmt(a['n'])}), the "
                f"pooled {n_fmt(rw['pooled_n'])} included.")
        if a["p90"] > 0 and 90 / a["p90"] >= 1:
            times = round(90 / a["p90"], 1)
            rule_extra = (f"Nine in ten completed tooling hires this year were back inside "
                          f"{dfmt(a['p90'])} days; the 90-day line is {times:.1f} times that.")
            so_what(bl, f"nine in ten of these items come back inside {dfmt(a['p90'])} days, so the "
                        f"90-day line is {times:.1f} times the normal window - an item past it is not "
                        "late by a little, it is parked. Chase the over-90 list as parked gear, "
                        "not overdue gear.")
        else:
            so_what(bl, "the normal window is longer than the 90-day line for this population - "
                        "read the over-90 list against the product's own window above.")
    else:
        bl.note("The log holds no completed hire this year for the items on hire today.")

    # ---- 2c. the year in movements
    ws = ti.weekly_series(ctx, pop)
    bl.h2("The year in movements - the items on hire today", pb=True)
    if ws:
        tot_i, tot_r = sum(w["issues"] for w in ws), sum(w["returns"] for w in ws)
        peak = max(ws, key=lambda w: (w["issues"], w["start"]))
        bl.story(f"The history of the {n_fmt(n)} tooling items on hire today, from the TRANSACTIONS "
                 f"export ({esc(log_period(ctx))}): every time one of them went out this year and "
                 f"every time one came back, by ISO week - <b>{n_fmt(tot_i)}</b> issues and "
                 f"<b>{n_fmt(tot_r)}</b> returns. It is these items' year, not the whole store's "
                 "traffic (that picture is on the Executive Summary); the movement scoreboard's "
                 "trend page, when it appears, reads this report's own printed figures instead.")
        weekly_block(bl, ctx, ws, "The items on hire today")
        so_what(bl, f"the week of {peak['week']} put the most of today's on-hire gear into the field "
                    f"({n_fmt(peak['issues'])} issues) - the weeks with the highest net out are where "
                    "the over-90 tail was born, and the same-day line is the habit to hold: the higher "
                    "it sits, the shorter the tail.")
    else:
        bl.note("The log holds no movement this year for the items on hire today.")
    return rule_extra


def quarter_close_section(bl, d, m):
    """A quarterly's 'Quarter close - the look forward': the window's
    barcodes against the window's own close (a close that has already
    passed rolls forward to the next one and the page says so). When
    every item in the window is already past 90 days it says exactly that
    and prints nothing else."""
    ctx = insights(d)
    rows = m["rows"]
    pop = {r["barcode"] for r in rows if r["barcode"]}
    own = window_close(m["key"])
    rolled = own < ASAT_DAY
    qend = next_close(ASAT_DAY) if rolled else own
    qc = ti.quarter_close(ctx, pop, qend=qend)
    n_all = len(rows)
    bl.h2("Quarter close - the look forward")
    if not n_all:
        bl.story("Nothing is on hire from this window, so nothing can cross 90 days by the close.")
        return
    if qc["already_over"] >= n_all:
        bl.story(f"Every one of the <b>{n_fmt(n_all)}</b> items on hire from {esc(m['label'])} is "
                 f"already past 90 days at the pull ({fmt_date(ASAT_DAY)}). There is nothing left to "
                 "cross a quarter close; the over-90 position is this window's register itself.")
        return
    bl.story(f"<b>{n_fmt(qc['n'])}</b> of the {n_fmt(n_all)} items on hire from {esc(m['label'])} "
             f"will have been out 90 days or more by <b>{fmt_date(qend)}</b> unless they come back - "
             f"{money(qc['value'])} at replacement ({n_fmt(qc['n'] - qc['unpriced'])} priced / "
             f"{n_fmt(qc['unpriced'])} unpriced); <b>{n_fmt(qc['already_over'])}</b> are already over "
             "90 days today. "
             + (f"This window's own close ({fmt_date(own)}) has passed, so the look forward is to the "
                f"next close, {fmt_date(qend)}. " if rolled else "")
             + f"Arithmetic on each item's on-hire date and nothing more: an item under 90 days out "
             f"at the pull whose on-hire date is on or before {fmt_date(qend - dt.timedelta(days=90))} "
             "crosses the line before the close.")
    per = crossing_tables(bl, d, qc, QC_CAP_QUARTER)
    if per:
        top_co, top_p = max(per.items(), key=lambda kv: (kv[1]["n"], N.sort_key(kv[0])))
        so_what(bl, f"{n_fmt(qc['n'])} items from this window become a charge at replacement cost on "
                    f"{fmt_date(qend)} unless they are back - {top_co} carries the most "
                    f"({n_fmt(top_p['n'])}); this is the list to walk with each supervisor before "
                    "the quarter ticks over.")
    else:
        so_what(bl, f"nothing from this window crosses 90 days by {fmt_date(qend)}; the over-90 items "
                    "already counted are the whole recovery task for this close.")


def util_insights(bl, d):
    """The Utilisation report's three log-based pages: dead stock (never
    moved this year - store-wide and tooling), headroom (fleet against the
    peak out at once, with the buy and cut rule printed), and the fast
    movers. Population: every register barcode with a tooling description.
    Returns what the data page quotes."""
    ctx = insights(d)
    pop = tooling_barcodes(d)
    w0, w1 = ctx["tx_window"]

    # ---- 4a. dead stock
    ds_all = ti.dead_stock(ctx, None)
    ds = ti.dead_stock(ctx, pop)
    bl.h2("Dead stock - never moved this year", pb=True)
    bl.story(f"An item is dead stock here when the SiteIQ RENTAL_STOCK register shows it Available "
             f"for Hire and the TRANSACTIONS log holds no movement for its barcode at all since the "
             f"log began ({stamp(w0)}). Across the whole register, every family, "
             f"<b>{n_fmt(ds_all['n'])}</b> of the {n_fmt(ds_all['available'])} available items with a "
             f"barcode never moved - {money(ds_all['value'])} at replacement "
             f"({n_fmt(ds_all['unpriced'])} unpriced). For tooling - this report's scope, radios, gas "
             f"monitors, Dräger equipment, lanyards and steel coil clamps out - <b>{n_fmt(ds['n'])}</b> "
             f"of the {n_fmt(ds['available'])} available barcoded items never moved.")
    bl.tiles([tile(n_fmt(ds["n"]), "Never moved (tooling)", "box", "", "amber"),
              tile(n_fmt(ds["available"]), "Available base (tooling, barcoded)", "check"),
              tile(money(ds["value"]), "Replacement value, never moved", "shield",
                   f"{n_fmt(ds['n'] - ds['unpriced'])} priced"),
              tile(n_fmt(ds["unpriced"]), "Unpriced, never moved", "warn", "", "amber")])
    top = ds["by_product"][:20]
    top_n = sum(p["n"] for _k, p in top)
    bl.h2("Never moved by product - ranked by items, top 20")
    bl.table(["Product", "Items never moved", "Priced value", "Unpriced"],
             [[product_show(d, k), n_fmt(p["n"]), money(p["value"]) if p["value"] else "-",
               n_fmt(p["unpriced"])] for k, p in top]
             + [["Total", n_fmt(ds["n"]), money(ds["value"]) if ds["value"] else "-", n_fmt(ds["unpriced"])]],
             cls="tight", aligns=["", "r", "r", "r"])
    bl.note(f"The top 20 products hold {n_fmt(top_n)} of the {n_fmt(ds['n'])}; the Total row is every "
            f"product ({n_fmt(len(ds['by_product']))} in all).")
    bays = Counter((x["unit"] or "(no bay)") for x in ds["rows"]).most_common()
    bl.h2("Never moved by bay - ranked by items, top 12")
    bl.chart(sh.hbars([(b, c) for b, c in bays[:12]], colour=K_ORANGE),
             f"Tooling items never moved this year by the storage unit the register shows them in - "
             f"top {min(12, len(bays))} of {len(bays)} bays, ranked by items.")
    if len(bays) > 12:
        bl.note("The other bays: " + ", ".join(f"{esc(b)} ({c})" for b, c in bays[12:]) + ".")
    b1 = bays[0] if bays else ("-", 0)
    so_what(bl, f"{n_fmt(ds['n'])} tooling items have not left the shelf all year - "
                f"{money(ds['value'])} of replacement value sitting still, {n_fmt(top_n)} of them in "
                f"the top 20 products. These are candidates to cut, not a decision: walk {b1[0]} "
                f"({n_fmt(b1[1])} items) with this list first, and cut nothing a shutdown needs.")

    # ---- 4b. headroom
    hr = ti.headroom(ctx, pop)

    def mark(r):
        if r["fleet"] and r["peak"] >= 0.9 * r["fleet"]:
            return "BUY candidate"
        if r["fleet"] and r["headroom"] >= 0.5 * r["fleet"] and r["never_moved"] > 0:
            return "CUT candidate"
        return ""
    buy_n = sum(1 for r in hr if mark(r) == "BUY candidate")
    cut_n = sum(1 for r in hr if mark(r) == "CUT candidate")
    bl.h2("Headroom - fleet against the peak out at once", pb=True)
    bl.story(f"Per product: the fleet on the register today, the most out at once this year (counted "
             "from the log's start and end times, every account), out now, never moved, and the "
             "headroom - fleet less peak, the part of the fleet the store has never needed at one "
             "time. <b>The rule, printed so the marks can be checked:</b> BUY candidate when the peak "
             "out at once is 90% or more of the fleet; CUT candidate when the headroom is 50% or more "
             "of the fleet and at least one item never moved. The marks are the rule applied, not a "
             f"decision. Of the {n_fmt(len(hr))} tooling products, {n_fmt(buy_n)} meet the buy rule and "
             f"{n_fmt(cut_n)} the cut rule; the top 30 by fleet are listed, ranked by fleet.")
    bl.table(["Product", "Fleet", "Peak out at once", "Peak on", "Out now", "Never moved", "Headroom", "Mark"],
             [[product_show(d, r["product"]), n_fmt(r["fleet"]), n_fmt(r["peak"]),
               r["peak_at"].strftime("%d %b %H:%M") if r["peak_at"] else "-", n_fmt(r["out_now"]),
               n_fmt(r["never_moved"]), n_fmt(r["headroom"]), mark(r)] for r in hr[:30]],
             cls="tight", aligns=["", "r", "r", "", "r", "r", "r", ""])
    buys = [r for r in hr[:30] if mark(r) == "BUY candidate"]
    cuts = [r for r in hr[:30] if mark(r) == "CUT candidate"]
    so_what(bl, f"in the top 30 by fleet, {n_fmt(len(buys))} products ran at or near their whole fleet at "
                f"once ({', '.join(product_show(d, r['product']) for r in buys[:3]) or 'none'}) - the "
                f"demand-backed buy list above is where they belong - and {n_fmt(len(cuts))} carry "
                "headroom the store never used; a right-size call on those starts with the never-moved "
                "count beside them, not with the fleet number.")

    # ---- 4c. fast movers
    fm = ti.fast_movers(ctx, pop, 20)
    bl.h2("Fast movers - ranked by movements, top 20", pb=True)
    bl.story(f"The tooling items with the most movements in the log this year - every issue counts "
             f"one - with the bay the register shows them in now, their status at the pull and their "
             f"last movement. <b>{n_fmt(fm['items_moved'])}</b> tooling barcodes moved at all this year "
             f"against the {n_fmt(ds['n'])} available ones that never did.")
    bl.table(["#", "Barcode", "Description", "Bay", "Status", "Moves", "Last movement"],
             [[i, r["barcode"], desc_show(d, r["barcode"], r["desc"]), r["unit"] or "-", r["status"] or "-",
               n_fmt(r["moves"]), r["last"].strftime("%d %b %Y %H:%M")]
              for i, r in enumerate(fm["rows"], 1)],
             cls="tight", aligns=["r", "", "", "", "", "r", ""])
    bl.chart(sh.hbars([(u, c) for u, c in fm["by_unit"]], colour=K_ORANGE),
             f"Movements this year by the bay the item sits in now - top {len(fm['by_unit'])} bays, "
             "ranked by movements, every tooling barcode that moved.")
    bays2 = [u for u, _c in fm["by_unit"][:2]]
    so_what(bl, f"the busiest tooling lives in {' and '.join(bays2) if bays2 else 'a few bays'} - keep "
                "those bays nearest the counter and check their tags first; the fast movers are the "
                "items whose test dates come round quickest.")
    return {"period": log_period(ctx), "tx_n": len(ctx["tx"]), "pop_n": len(pop), "store_dead": ds_all["n"]}


def render_quarter(d, qk):
    m = quarter_model(d, qk)
    n = len(m["rows"])
    # WHY (03 Sep 2026, layout pass): the position page reads band, tiles,
    # three things, then a three-line story - what the report is, the one
    # number, where the detail is. The rest of the old opening is the first
    # paragraph of the page after, figure-free.
    story = (f"This is the {esc(m['label'])} on-hire position for the Ampol tool store: "
             f"<b>{n_fmt(n)}</b> items issued in this window and still out. Gear not home by "
             "quarter close is billable at replacement cost; the register, company A to Z, "
             "sits behind the appendix divider.")
    bl = Blocks()
    band = None
    if m["rows"]:
        # WHY (03 Sep 2026): the band on this window's own over-90 share
        q_over90 = sum(1 for r in m["rows"] if r["days"] is not None and r["days"] > 90)
        band = rag_over90(q_over90, n, f"items on hire from {m['label']}")
        bl.rag(band)
    bl.tiles([tile(n_fmt(n), "Items on hire", "box",
                   key="on_hire_" + qk, raw=n, good=None),
              tile(m["n_companies"], "Companies", "layers",
                   key="companies_" + qk, raw=m["n_companies"], good=None),
              tile(money(m["total_val"]), "Replacement value", "shield",
                   key="value_" + qk, raw=round(m["total_val"], 2), good=None),
              tile(f"{m['priced']} / {m['unpriced']}", "Items priced / unpriced", "bars")])
    # WHY (03 Sep 2026): this window's own three things, from its own rows
    bl.three(three_things_for(d, m["rows"], m["label"]))
    bl.story(story)
    bl.page_break()
    bl.story("Gear on hire in this window that has not come home is billable at replacement "
             "cost at quarter close, and gear no longer needed is removed from hire at the "
             "same point - so this list is the one to walk before the quarter ticks over. "
             "Everything here is easy to fix: bring it back to the counter, it is double-scanned "
             "off your name in seconds, and it disappears from this report the same day. We are "
             "not chasing blame - we are helping everyone finish the quarter clean.")
    extras = []
    if m["n_accounts"]:
        extras.append(f"{plural(m['n_accounts'])} sit under shutdown / custody accounts "
                      "(listed under their company, not as people)")
    if m["custody"]:
        extras.append(f"{plural(len(m['custody']))} sit in the Repairs custody account "
                      "(internal, listed last, not a customer)")
    if m["partial"]:
        extras.append(m["partial"].rstrip("."))
    if extras:
        bl.note("Inside the count: " + "; ".join(extras) + ".")
    if not m["rows"]:
        bl.story("Nothing is on hire from this window"
                 + (" - the quarter has not started." if not quarter_started(qk) else "."))
    # WHY (03 Sep 2026, insights pass): the window's own quarter-close look
    # forward - arithmetic on the on-hire date, before the ageing panel
    quarter_close_section(bl, d, m)
    if m["companies"]:
        # WHY (03 Sep 2026): the ageing-by-company panel, then the APPENDIX
        # divider - everything after it is the complete register
        ageing_panel(bl, m["companies"], "On-hire ageing by company, A to Z")
        n_client = sum(co["vb"]["n"] for co in m["companies"])
        bl.divider("The complete register, A to Z",
                   f"Everything from here is the full register for {esc(m['label'])}: every "
                   "company A to Z, every hirer under it, every item with its start date and "
                   "days out. The story - the position, the band and the ageing - is on the "
                   "pages before this one.",
                   f"{plural(n_client)} across {m['n_companies']} "
                   f"{'company' if m['n_companies'] == 1 else 'companies'}"
                   + (f", then the Repairs custody account ({plural(len(m['custody']))})"
                      if m["custody"] else ""))
        bl.h2("On hire by company, A to Z")
        bl.note("Companies A to Z; under each company the hirers A to Z "
                "(people first, then the company's project / workflow accounts); under "
                "each hirer the longest-held item first. Hirer names are as SiteIQ "
                "records them. Where a customer has more than one SiteIQ account, the "
                "account is shown against the hirer. Days = calendar days from the "
                f"on-hire date to {fmt_date(ASAT_DAY)}.")
        bl.register(m["companies"], fleet=len(d["master"]))
    if m["custody"]:
        bl.h2("Internal custody - Repairs account (not a customer)")
        bl.note("Items booked to the Repairs custody account are inside "
                "the on-hire count above but are Coates' own workflow - tagged, tracked "
                "and never reissued until right. They are not chased with any company.")
        bl.register(custody_blocks(m["custody"]))
    limits = ["Counts come straight from the SiteIQ RENTAL_STOCK register: items On Hire "
              f"whose on-hire date falls in {m['label']}. Radios, gas monitors, Dräger "
              "equipment, lanyards and steel coil clamps are excluded - they are reported "
              "separately. Items on hire since before 01 Jan " + str(ASAT_DAY.year) +
              " are not in this list (see the Executive Summary and the Tooling On-Hire "
              "Report).",
              "Company names follow the suite's one naming rule: the site's former-name "
              "account and the refinery legal name both read Ampol; CR reads Contract "
              "Resources; FCCU and SATGAS/MOL project accounts roll up to their company "
              "and the account is shown against the hirer.",
              "Descriptions use the corrected name from Tooling_Description_Mapping.xlsx "
              "where one exists for the barcode, or where every mapping row for that "
              "register description agrees on one corrected name; otherwise the register "
              "description. The site's former name is shown as Ampol wherever SiteIQ "
              "still carries it.",
              "Three things to do today are drawn from this window's own rows: the company "
              "holding the most of its items over 90 days, the oldest item out, and the "
              "unpriced items (or the second-largest over-90 holder when everything is "
              "priced). Fewer than three true items prints fewer.",
              "Quarter close - the look forward is arithmetic on the SiteIQ RENTAL_STOCK "
              "on-hire dates of this window's own items: an item under 90 days out at the "
              "pull whose on-hire date is 90 days or more before the close crosses the line "
              "unless it comes back - a count, never a forecast. The window's own close is "
              "used; a close that has already passed rolls forward to the next one and the "
              "section says so. Where every item in the window is already past 90 days the "
              "section says exactly that. Replacement value as everywhere on this page; the "
              "TRANSACTIONS export is read for no figure in this report.",
              ] + source_limits(d)
    cfg = k2cfg("Quarterly On-Hire Report - " + m["label"],
                "COATES · TOOL STORE · QUARTERLY ON-HIRE REPORT", KEY_QUARTER)
    # WHY (03 Sep 2026): the cover - the window's one number, three true
    # lines, the stripe from this window's own band, the freshness line and
    # (second pass) what is inside with real page numbers
    q_over = [r for r in m["rows"] if r["days"] is not None and r["days"] > 90]
    big_label = (f"items on hire from {m['label']}" if qk != "YEAR"
                 else f"items on hire, issued in {ASAT_DAY.year}")

    def cover(contents):
        return k2flow.cover_block(cfg, n_fmt(n), big_label, [
            f"<b>{money(m['total_val'])}</b> of replacement value in the field "
            f"({n_fmt(m['priced'])} priced / {n_fmt(m['unpriced'])} unpriced)",
            f"<b>{n_fmt(m['n_companies'])}</b> companies, A to Z",
            over90_words(q_over)], GENERATED, ASAT_SHORT,
            rag=band["status"] if band else None,
            fresh=sh.freshness_line(ASAT_DT, GENERATED_DT), contents=contents)
    return report_outputs(
        "Quarterly On-Hire Report - " + m["label"],
        f"Quarterly On-Hire &amp; Recovery | {esc(m['label'])} | "
        f"{plural(n)} | {money(m['total_val'])}",
        f"Ampol Tool Store - Quarterly On-Hire Report - {m['label']} - "
        f"{plural(n)}",
        bl, limits, "Ampol Tooling - " + m["label"],
        "Quarterly On-Hire Report - " + m["label"], cfg, HOW_REGISTER, cover=cover,
        pdf_subject=(f"Tooling issued from the Ampol tool store in {m['label']} and still on "
                     f"hire at the SiteIQ pull of {ASAT_SHORT}, by company A to Z, with the "
                     "charge position at quarter close."))


def render_company(d, name):
    m = company_model(d, name)
    sd = pct(m["same_day_pct"]) if m["same_day_pct"] is not None else "-"
    story = (f"Thanks for working with the Coates tool store - here is exactly "
             f"where things stand for <b>{esc(m['name'])}</b>. Your crews have run "
             f"{n_fmt(m['tx_ytd'])} tool store transactions this year, and "
             f"{n_fmt(m['same_day'])} of those came back the same day ({sd}) - that "
             f"is the habit that keeps gear available for everyone. Right now "
             f"{plural(len(m['items']))} {'are' if len(m['items']) != 1 else 'is'} on hire "
             f"to your team. Anything not needed: hand it in, and it is off your list the "
             f"same day. If it IS needed, perfect - this list is simply your record "
             f"of where it all is.")
    bl = Blocks()
    bl.story(story)
    bl.tiles([tile(n_fmt(len(m["items"])), "Items on hire", "box"),
              tile(money(m["total_val"]), "Replacement value", "shield"),
              tile(n_fmt(m["tx_ytd"]), "Transactions YTD", "swap"),
              tile(sd, "Same-day returns", "clock"),
              tile(n_fmt(len(m["people"])), "Hirer names using store", "people")])
    qbits = " &nbsp;|&nbsp; ".join(f"{QUARTERS[qk][1]}: <b>{len(m['per_q'][qk])}</b>"
                                   for qk in QUARTERS)
    bl.story(f"On hire by quarter issued: {qbits}")
    notes = []
    if m["unpriced"]:
        notes.append(f"{m['unpriced']} of the {len(m['items'])} items have no catalogue "
                     "price and are not in the replacement value")
    if m["accounts"]:
        notes.append(f"{n_fmt(m['tx_custody'])} of the transactions and "
                     f"{plural(len(m['account_items']))} on hire are booked to shutdown / "
                     f"custody accounts ({', '.join(m['accounts'][:4])}"
                     f"{', ...' if len(m['accounts']) > 4 else ''}) rather than to a person; "
                     "the hirer-names tile leaves those out")
    if len(m["siteiq_accounts"]) > 1:
        notes.append("SiteIQ accounts rolled into this company: "
                     + ", ".join(f"{a} ({n})" for a, n in m["siteiq_accounts"])
                     + " - the account is shown against each hirer below")
    if notes:
        bl.note(". ".join(notes) + ".")
    if m["blocks"]:
        bl.h2("On hire now (hirers A to Z, longest-held first)")
        bl.register(m["blocks"])
    if m["high_val"]:
        bl.h2("High-value items in your care (ranked by replacement value)")
        bl.story("These carry the highest replacement cost - worth a "
                 "quick check they are secure and still needed.")
        bl.table(ITEM_HDR, item_rows(m["high_val"]))
    if m["compliance"]:
        bl.h2("Electrical, rigging and high-torque gear in your care (hirers A to Z)")
        bl.story("Test and tag gear: if it has been out a while, swing "
                 "it past the counter for a quick check - we will make sure the tags are "
                 "current and hand it straight back if you still need it.")
        bl.table(ITEM_HDR, item_rows(m["compliance"]))
    limits = ["On-hire lines are the SiteIQ RENTAL_STOCK register (items On Hire with a "
              f"{ASAT_DAY.year} on-hire date; radios, gas monitors and Dräger equipment "
              "reported separately). Transactions are the SiteIQ CUSTOMER_CONTRACTOR_EQUIP "
              "export for the year to date, matched to this company after standardising "
              "the employer name.",
              "'Hirer names using store' counts distinct hirer names on this year's "
              "transactions, leaving out shutdown / custody / workflow accounts. Names are "
              "as recorded in SiteIQ and are not verified as individuals.",
              ] + source_limits(d)
    cfg = k2cfg("Company On-Hire Report - " + m["name"],
                "COATES · TOOL STORE · COMPANY ON-HIRE REPORT", KEY_REGISTER)
    return report_outputs(
        "Company On-Hire Report - " + m["name"],
        f"Company On-Hire Report | {esc(m['name'])} | {plural(len(m['items']))}",
        f"Ampol Tool Store - {m['name']} - On-Hire Report - "
        f"{plural(len(m['items']))}",
        bl, limits, "Ampol Tooling - " + m["name"],
        "Company On-Hire Report - " + m["name"], cfg, HOW_REGISTER,
        pdf_subject=(f"Tooling on hire to {m['name']} from the Ampol tool store at the SiteIQ "
                     f"pull of {ASAT_SHORT}, by hirer A to Z."))


def render_onhire(d):
    """The Tooling On-Hire Report - one document for every company.
    WHY (02 Sep 2026): Andrew asked for one clean on-hire report instead of
    a report per company. Page 1 is the position; then the definition and
    where the count sits; then the pictures; then the register (company A
    to Z, hirer A to Z, longest held first); then custody / holding
    accounts, the legacy rows, and a data-and-method page. Every number is
    counted from the master list."""
    x = onhire_model(d)
    n = len(x["master"])
    year = ASAT_DAY.year
    pulled = stamp(d["asat"]["stock"]["pulled"])
    oldest = x["oldest"]
    acct_names = ", ".join(f"{h} under {co}" for (h, co), _n in x["project_accounts"])
    # WHY (03 Sep 2026, layout pass): the position page reads band, tiles,
    # three things, then a three-line story - what the report is, the one
    # number, where the detail is. Every other figure the old opening
    # carried lives in the tiles or in "Where the count sits" on the page
    # after; the rest of its words open that page, figure-free.
    story = (f"This is the register of every tooling item on hire from the Ampol tool store "
             f"at the SiteIQ pull of {esc(pulled)}: <b>{n_fmt(n)}</b> items issued since "
             f"01 Jan {year}. The pictures follow; the complete register, company A to Z, "
             "sits behind the appendix divider.")
    bl = Blocks()
    # WHY (03 Sep 2026): the on-hire tile carries the by-month-started series
    # as its sparkline (the same counts the month table below prints)
    p1 = [tile(n_fmt(n), f"Items on hire (tooling, {year})", "box",
               key="on_hire", raw=n, good=None, spark=[vb["n"] for _lab, vb in x["months"]]),
          tile(money(x["total_val"]), "Replacement value", "shield",
               f"{n_fmt(x['priced'])} priced / {n_fmt(x['unpriced'])} unpriced", "grey",
               f"Replacement value - {n_fmt(x['priced'])} priced / "
               f"{n_fmt(x['unpriced'])} unpriced",
               key="value", raw=round(x["total_val"], 2), good=None),
          tile(n_fmt(x["n_companies"]), "Companies", "layers",
               key="companies", raw=x["n_companies"], good=None),
          tile(n_fmt(x["people_names"]), "Hirers (people)", "people",
               key="hirers", raw=x["people_names"], good=None),
          tile(fmt_date(oldest["date"]) if oldest else "-", "Oldest hire", "clock",
               f"{oldest['days']} days" if oldest else "", "amber",
               f"Oldest hire ({oldest['days']} days)" if oldest else "Oldest hire",
               key="oldest_days", raw=oldest["days"] if oldest else None, good="down"),
          tile(n_fmt(len(x["over90"])), "Items over 90 days", "warn", "", "amber",
               key="over90", raw=len(x["over90"]), good="down")]
    band = rag_over90(len(x["over90"]), n, f"tooling items on hire ({year})")
    bl.rag(band)
    bl.tiles(p1, cls="three")
    # WHY (03 Sep 2026): the same three things the Executive Summary prints
    bl.three(three_things_for(d))
    bl.story(story)
    bl.page_break()
    bl.story("Every company A to Z, every hirer under it A to Z, every item with its start "
             "date and days out - so anyone can find their gear and hand back what is "
             "finished with. The quarterly charge reports carry the billing position at "
             "quarter close; this is the field position today.")
    fams = ", ".join(f"{n_fmt(v)} {FAM_WORDS.get(k, k.lower())}"
                     for k, v in sorted(x["legacy"]["families"].items(),
                                        key=lambda kv: N.sort_key(kv[0])) if k != "Tooling")
    bl.defbox('What "on hire" means in this report', [
        f"An item is on hire when the SiteIQ RENTAL_STOCK register showed it "
        f"On Hire to a company at the pull ({esc(pulled)}) with an on-hire date in "
        f"{year}. Days = calendar days from that date to {fmt_date(ASAT_DAY)}.",
        "Tooling only: radios, gas monitors, Dräger equipment, lanyards and "
        "steel coil clamps are excluded because they have their own reports "
        "(Radio, Gas Monitor, Rigging).",
        f"Custody accounts are not customers. The Repairs custody account "
        f"({plural(len(x['custody']))}) is inside the count and the tool store's own "
        f"holding accounts ({plural(len(x['holding']))}) are outside it; both are "
        "listed in their own section at the back, marked not customer hire.",
        f"{n_fmt(x['legacy']['n'])} register rows on hire since before 01 Jan "
        f"{year} are excluded from the count and listed at the back by company - "
        f"they are {fams}, carried by their own reports.",
        "Every table is A to Z (companies, hirers, accounts) with items "
        "longest-held first. The one ranked table - the top 15 hirers - says so "
        "in its heading."])
    bl.h2("Where the count sits")
    pairs_note = ""
    if x["people_pairs"] != x["people_names"]:
        extra = x["people_pairs"] - x["people_names"]
        pairs_note = (f" ({n_fmt(x['people_pairs'])} company-hirer pairs: {n_fmt(extra)} "
                      f"name{'s' if extra != 1 else ''} recorded under two companies)")
    sit = [["Customer hire to named people", n_fmt(x["people_items"]),
            f"{x['n_companies']} companies, {n_fmt(x['people_names'])} hirer names{pairs_note}"]]
    for (h, co), cnt in x["project_accounts"]:
        sit.append([f"Customer project / workflow account: {h} (under {co})", n_fmt(cnt),
                    "booked to an account, not a person"])
    sit.append(["Repairs custody account (internal)", n_fmt(len(x["custody"])),
                "Coates workflow, not customer hire"])
    sit.append(["Total on hire", n_fmt(n), money(x["total_val"]) + " priced"])
    bl.table(["Where", "Items", "Note"], sit)
    # the figures the old opening carried that no tile holds: the oldest
    # hire, the over-90 value and the project / workflow account items
    o_vb = x["over90_vb"]
    note = ((f"The oldest hire started {fmt_date(oldest['date'])} ({oldest['days']} days ago); "
             if oldest else "")
            + f"{n_fmt(len(x['over90']))} items have been out more than 90 days "
            f"({money(o_vb['value']) if o_vb['priced'] else '-'} priced).")
    if x["account_items"]:
        note += (f" {n_fmt(x['account_items'])} items are booked to customer project / workflow "
                 f"accounts ({esc(acct_names)}).")
    bl.note(note)

    # ---- since the last pull (03 Sep 2026): pull against pull, then the
    # 24 hours of traffic before the pull
    since_last_pull(bl, d, full=True)

    # ---- the insights pass (03 Sep 2026): the quarter-close look forward,
    # the return windows by product and the year in movements, all for this
    # report's own population; the one sentence page 1 gains sits under the
    # band's rule and states the data behind the 90-day line
    rule_extra = onhire_insights(bl, d, n)
    if rule_extra:
        band["rule"] += " " + rule_extra

    # ---- pictures: companies A to Z
    bl.h2("Items on hire by company, A to Z", pb=True)
    bl.chart(
        hbar_chart([(co["name"], co["vb"]["n"],
                     [plural(co["vb"]["n"]),
                      money(co["vb"]["value"]) if co["vb"]["priced"] else "-"])
                    for co in x["companies"]], lab_w=200),
        f"Items on hire by company with the priced replacement value beside each bar - "
        f"all {x['n_companies']} companies, A to Z, not ranked. The Repairs custody "
        f"account ({plural(len(x['custody']))}) is not a company and is not drawn.")
    co_rows = []
    for co in x["companies"]:
        accts = "; ".join(f"{a} ({cnt})" for a, cnt in co["siteiq_accounts"])
        co_rows.append([co["name"], accts, co["vb"]["n"], co["n_people"]
                        + (0 if not co["accounts_g"] else 0),
                        money(co["vb"]["value"]) if co["vb"]["priced"] else "-",
                        f'{co["vb"]["priced"]} / {co["vb"]["unpriced"]}',
                        fmt_date(co["oldest"])])
    bl.table(["Company", "SiteIQ accounts (items)", "Items", "Hirers", "Priced value",
              "Priced / Unpriced", "Oldest hire"], co_rows, cls="tight")
    bl.note("Hirers counts people only; a company's project / workflow "
            "account is listed under the company in the register, not counted as a "
            "hirer.")
    # WHY (03 Sep 2026): the same companies, split into the four ageing bands
    ageing_panel(bl, x["companies"], "On-hire ageing by company, A to Z")
    bl.email_end()

    # ---- pictures: ageing and category
    bl.h2("On-hire ageing profile", pb=True)
    bl.chart(
        hbar_chart([(lab, vb["n"], [plural(vb["n"]), money(vb["value"]) if vb["priced"] else "-"])
                    for lab, vb in x["ageing"]], lab_w=150),
        f"Days since the on-hire date, to {fmt_date(ASAT_DAY)}, with the priced replacement "
        "value in each band. The long tail is what the quarterly recovery cycle brings home.")
    bl.table(["Band", "Items", "Priced value", "Priced / Unpriced"],
             [[lab, vb["n"], money(vb["value"]) if vb["priced"] else "-",
               f'{vb["priced"]} / {vb["unpriced"]}'] for lab, vb in x["ageing"]]
             + [["Total", n, money(x["total_val"]), f"{x['priced']} / {x['unpriced']}"]])
    bl.h2("Category split")
    bl.chart(
        hbar_chart([(c, vb["n"], [plural(vb["n"]), money(vb["value"]) if vb["priced"] else "-"])
                    for c, vb in x["cats"]], lab_w=150),
        "Items on hire by description family - a keyword match on the register "
        "description for Rigging, Electrical and High Torque; General is everything else. "
        "Fixed order, not ranked.")
    bl.table(["Category", "Items", "Priced value", "Priced / Unpriced"],
             [[c, vb["n"], money(vb["value"]) if vb["priced"] else "-",
               f'{vb["priced"]} / {vb["unpriced"]}'] for c, vb in x["cats"]])

    # ---- pictures: top 15 hirers (the only ranked table) and by month
    bl.h2("Top 15 hirers by items (ranked by items - the only ranked "
          "table in this report)", pb=True)
    bl.chart(
        hbar_chart([(f"{h} ({co})", vb["n"],
                     [plural(vb["n"]), money(vb["value"]) if vb["priced"] else "-"])
                    for h, co, vb in x["top_hirers"]], lab_w=250),
        f"The 15 people holding the most items, of {n_fmt(x['people_pairs'])} company-hirer "
        "pairs on the register; project / workflow accounts and the Repairs custody "
        "account are left out. Ranked by items, then A to Z on a tie.")
    bl.table(["Hirer", "Company", "Items", "Priced value", "Priced / Unpriced"],
             [[h, co, vb["n"], money(vb["value"]) if vb["priced"] else "-",
               f'{vb["priced"]} / {vb["unpriced"]}'] for h, co, vb in x["top_hirers"]])
    bl.h2("Items on hire by month started")
    bl.chart(
        k2shell.grouped_bars([{"label": lab, "n": vb["n"]} for lab, vb in x["months"]],
                             series=(("n", K_ORANGE, "Items still on hire"),)),
        f"Items still on hire by the month their hire started, Jan to {ASAT_DAY.strftime('%b')} "
        f"{year}. " + (x["partial"] or ""))
    bl.table(["Month started", "Items still on hire", "Priced value", "Priced / Unpriced"],
             [[lab, vb["n"], money(vb["value"]) if vb["priced"] else "-",
               f'{vb["priced"]} / {vb["unpriced"]}'] for lab, vb in x["months"]]
             + [["Total", n, money(x["total_val"]), f"{x['priced']} / {x['unpriced']}"]])

    # ---- the trend page (03 Sep 2026): only once seven days are on record
    trend_shown = trend_page(bl, d, {"on_hire": n, "over90": len(x["over90"]),
                                     "value": round(x["total_val"], 2)})

    # ---- the register, behind its APPENDIX divider (03 Sep 2026)
    n_client = len(x["client"])
    bl.divider("The complete register, A to Z",
               "Everything from here is the full register: every company A to Z, every hirer "
               "under it, every item with its start date, days out and replacement value, then "
               "the custody and holding accounts and the legacy rows. The story - the position, "
               "the movement, the pictures - is on the pages before this one.",
               f"{plural(n_client)} across {x['n_companies']} companies"
               + (f", then the Repairs custody account ({plural(len(x['custody']))})"
                  if x["custody"] else ""))
    bl.h2("The register - every company A to Z")
    bl.note("Companies A to Z (the SiteIQ account is shown under the "
            "company name where a customer has more than one); under each company the "
            "hirers A to Z, people first and then the company's project / workflow "
            "accounts; under each hirer the longest-held item first, then A to Z by "
            "description. Hirer names are as SiteIQ records them. Descriptions are the "
            "corrected names where the mapping has one. Replacement is the catalogue "
            "new-buy average; a dash means no catalogue price, never $0. Days = calendar "
            f"days from the on-hire date to {fmt_date(ASAT_DAY)}.")
    bl.register(x["companies"], fleet=n)

    # ---- custody / holding accounts
    bl.h2("Custody and holding accounts - not customer hire", pb=True)
    bl.story(f"Two kinds of internal account hold {year} tooling. The "
             f"<b>Repairs custody account</b> ({plural(len(x['custody']))}) is inside the "
             f"{n_fmt(n)} above: gear sent off site for repair, tagged and tracked, never "
             f"reissued until right. The tool store's own <b>holding accounts</b> "
             f"({plural(len(x['holding']))}) are outside the count exactly as the "
             "workbook always treated them: the counter, the yard, the loading bay, "
             "all-around repairs and the out-of-tag park. Neither is chased with any "
             "company; both are listed here so nothing on the register is unseen.")
    if x["custody_blocks"]:
        bl.h2(f"Repairs custody account - inside the {n_fmt(n)}, A to Z")
        bl.register(x["custody_blocks"])
    if x["holding_blocks"]:
        bl.h2(f"Tool store holding accounts - outside the {n_fmt(n)}, A to Z")
        bl.register(x["holding_blocks"])
    fam_rows = []
    for (co, fam), cnt in x["custody_family"].items():
        fam_rows.append([co, FAM_WORDS.get(fam, fam.lower()).capitalize(), cnt,
                         "Gas Monitor report" if "GAS" in fam or fam == "DRAGER"
                         else ("Radio report" if fam == "RADIO" else "Rigging report")])
    for (acc, fam), cnt in x["holding_family"].items():
        fam_rows.append([acc, FAM_WORDS.get(fam, fam.lower()).capitalize(), cnt,
                         "Gas Monitor report" if "GAS" in fam or fam == "DRAGER"
                         else ("Radio report" if fam == "RADIO" else "Rigging report")])
    if fam_rows:
        fam_rows.sort(key=lambda r: (N.sort_key(r[0]), N.sort_key(r[1])))
        bl.h2(f"Radio and gas families in custody accounts ({year}, not tooling)")
        bl.note("Counted here so the custody picture is complete; these "
                "items are not tooling and are carried by their own reports.")
        bl.table(["Account", "Family", "Items", "Reported by"], fam_rows)
    rep_rows = d["repairs_n"]
    bl.note(f"The Repairs account holds {n_fmt(rep_rows)} register rows "
            f"in all: {n_fmt(len(x['custody']))} inside the count (off-site repair custody) "
            f"and {n_fmt(sum(1 for r in x['holding'] if r['company'] == 'Repairs'))} in the "
            "holding accounts above"
            + (f"; {n_fmt(rep_rows - len(x['custody']) - sum(1 for r in x['holding'] if r['company'] == 'Repairs'))} "
               f"are radio / gas family rows or pre-{year} rows"
               if rep_rows - len(x['custody']) - sum(1 for r in x['holding'] if r['company'] == 'Repairs') else "")
            + ".")

    # ---- legacy
    leg = x["legacy"]
    bl.h2(f"Legacy on hire - before 01 Jan {year}, A to Z", pb=True)
    bl.story(f"{n_fmt(leg['n'])} register rows are still on hire from "
             f"issues before 01 Jan {year} (oldest {fmt_date(leg['oldest'])}). They are "
             f"not in the {n_fmt(n)} because they are {fams}"
             + (f" - and {n_fmt(leg['tooling_n'])} tooling" if leg["tooling_n"] else
                "; no tooling item on hire pre-dates this year")
             + ". The Radio and Gas Monitor reports chase them; they are listed here by "
             f"company so the recovery story is complete - {leg['n_companies']} companies, "
             "A to Z.")
    fams_present = [f for f in ("RADIO", "DRAGER", "GAS MONITOR", "GAS DETECTOR", "MULTI GAS",
                                "MULTIGAS", "LANYARD", "STEEL COIL CLAMP", "Tooling")
                    if f in leg["families"]]
    hdr = ["Company", "Legacy items"] + [FAM_WORDS.get(f, f).capitalize() for f in fams_present] \
        + ["Oldest on-hire date"]
    leg_rows = [[co, n_fmt(c["_n"])] + [n_fmt(c.get(f, 0)) for f in fams_present]
                + [fmt_date(old)] for co, c, old in leg["az"]]
    leg_rows.append(["TOTAL", n_fmt(leg["n"])]
                    + [n_fmt(leg["families"].get(f, 0)) for f in fams_present]
                    + [fmt_date(leg["oldest"])])
    bl.table(hdr, leg_rows, aligns=["", "r"] + ["r"] * len(fams_present) + [""])

    # ---- data and method
    a = d["asat"]
    bl.h2("Data and method", pb=True)
    bl.table(["Source", "As at / size", "Used for"], [
        ["SiteIQ RENTAL_STOCK.xlsx", "SiteIQ pull " + stamp(a["stock"]["pulled"])
         + f"; {n_fmt(len(d['stock']))} register rows",
         "every on-hire line: company, hirer, barcode, description, on-hire date, status"],
        ["Tooling_Description_Mapping.xlsx",
         f"{n_fmt(d['util']['mapping_rows'])} rows; {n_fmt(len(d['corr']))} barcodes, "
         f"{n_fmt(len(d['corr_by_desc']))} unambiguous descriptions",
         f"corrected descriptions: {n_fmt(x['corrected_bc'])} lines by barcode and "
         f"{n_fmt(x['corrected_desc'])} by matching description of the {n_fmt(n)}"],
        ["Ampol_ToolStore_Pricing.xlsx",
         f"{n_fmt(d['pricing_rows'])} rows; {n_fmt(len(d['pricing']))} distinct descriptions",
         "replacement value = Avg Buy Price (New); first price wins where a description "
         "is listed more than once"],
        ["SiteIQ TRANSACTIONS.xlsx (CUSTOMER_CONTRACTOR_EQUIP)",
         "SiteIQ pull " + stamp(a["tx"]["pulled"]) + f"; report period {log_period(insights(d))}; "
         f"{n_fmt(len(insights(d)['tx']))} movements",
         "the 24 hours before the pull, the return windows by product and the year in "
         "movements - each for this report's population (the on-hire barcodes above)"],
        ["SiteIQ STOCKTAKE.xlsx", "SiteIQ pull " + stamp(a["stocktake"]["pulled"]),
         "not used for any figure on this report (it feeds the Compliance report)"]])
    bl.h2("Rules applied")
    rules = [
        f"On hire = a RENTAL_STOCK row with a company, a {year} on-hire date and a "
        "tooling description; radios, gas monitors, Dräger equipment, lanyards and steel "
        "coil clamps are excluded (their own reports); the tool store's holding accounts "
        "(T&I - Tool store, All-Around - Repairs, Bulk - Yard, Loading Bay - Out Of "
        "Service, Out Of - Calibration, Rigging & 240V - Out Of Tag Date) are excluded "
        "from the count and listed in the custody section.",
        "One customer, one name (shared suite rule): the site's former-name account and "
        "the refinery legal name both read Ampol; CR reads Contract Resources; FCCU and "
        "SATGAS/MOL project accounts roll up to their company and the SiteIQ account is "
        "shown under the company and against the hirer.",
        f"SiteIQ still carries the site's former name on {n_fmt(x['former_desc_all'])} "
        f"item descriptions ({n_fmt(x['former_desc_here'])} of the {n_fmt(n)} lines in this "
        f"report) and {n_fmt(x['former_accounts'])} customer account; shown here under the "
        "current name. Barcodes are identifiers and are never changed.",
        "Hirer names are as SiteIQ records them (an entry typed wholly in capitals or "
        "wholly in lower case is shown in title case). A hirer is a person; project / "
        "workflow accounts and custody accounts are counted separately and never as people.",
        f"Replacement value = catalogue Avg Buy Price (New) from Ampol_ToolStore_Pricing.xlsx, "
        f"first price wins on duplicates; {n_fmt(x['priced'])} of the {n_fmt(n)} lines are "
        f"priced ({money(x['total_val'])}) and {n_fmt(x['unpriced'])} are unpriced - a dash, "
        "never $0, never estimated, never in a total.",
        f"Days = calendar days from the on-hire date to {fmt_date(ASAT_DAY)}; the ageing bands "
        "are 0-30, 31-60, 61-90, 91-180 and over 180 days. Category is a keyword match on "
        "the register description (Rigging / Electrical / High Torque; General otherwise).",
        "Order: every table is A to Z (companies, hirers, accounts, legacy companies) with "
        "items longest-held first and then A to Z by description; the ranked tables - the "
        "top-15 hirers and the items that crossed 90 days - say so in their headings. Every "
        "figure is counted by this kit from the exports named above - the Excel workbook is "
        "not read for any printed number.",
        "Since the last pull: the SiteIQ register is compared item by item with the newest "
        "earlier RENTAL_STOCK export parked in Data\\previous (came back, went out, changed "
        "hands, crossed 30 / 60 / 90 days while out); the 24 hours before the pull are "
        "counted from the TRANSACTIONS export's start and end times. Both use this report's "
        f"population ({SCOPE_WORDS}). No earlier pull means no pull-against-pull rows and a "
        "plain note saying so.",
        trend_line(trend_shown),
        "Quarter close - the look forward: arithmetic on the RENTAL_STOCK on-hire dates of "
        "the items on hire above - an item under 90 days out at the pull whose on-hire date "
        "is 90 days or more before the quarter close crosses the line unless it comes back; "
        "a count, never a forecast. Items already over 90 days are stated beside it, not "
        "counted again. The store-wide figure in that section counts every family and "
        "account on the register the same way.",
        "Return windows and the year in movements: the TRANSACTIONS export's rows for this "
        "report's barcodes - a completed hire is a row with a start and an end time; the "
        "median and the 90th percentile are read off the sorted hold times in days; same-day "
        "means the end date equals the start date. A product is the SiteIQ description with "
        "its size and serial tail removed (the corrected name is shown where the mapping has "
        "one for every item under it); products with fewer than ten completed hires are pooled. "
        "Weeks are ISO weeks by start time (issues) and end time (returns); the partial weeks "
        "are marked. Every figure is a count or a sum over rows in the export.",
    ]
    bl.ul(rules)
    limits = source_limits(d)
    cfg = k2cfg("Tooling On-Hire Report", "COATES · TOOL STORE · TOOLING ON-HIRE REPORT",
                KEY_REGISTER)
    # WHY (03 Sep 2026): the cover - the one number, three true lines, the
    # status stripe from the band on page 1, the freshness line and (second
    # pass) what is inside with real page numbers
    def cover(contents):
        return k2flow.cover_block(cfg, n_fmt(n), "tooling items on hire", [
            f"<b>{money(x['total_val'])}</b> of replacement value in the field "
            f"({n_fmt(x['priced'])} priced / {n_fmt(x['unpriced'])} unpriced)",
            f"<b>{n_fmt(x['n_companies'])}</b> companies, A to Z",
            over90_words(x["over90"])], GENERATED, ASAT_SHORT,
            rag=band["status"], fresh=sh.freshness_line(ASAT_DT, GENERATED_DT),
            contents=contents)
    um = util_model(d)
    tx_all = d["tx"]
    same_day = sum(1 for t in tx_all if t["end"] and t["start"] and t["end"] == t["start"])
    card = {"cfg": cfg, "tiles": card_tiles(p1),
            "band": (band["status"], band["card_head"], band["owner"], band["card_action"]),
            "scores": card_scores(n, x["priced"], (same_day / len(tx_all)) if tx_all else None,
                                  len(x["over90"]), len(um["buy"]),
                                  sum(1 for r in um["buy"] if r["buy"] is not None)),
            "foot": f"Counted from the SiteIQ exports of {ASAT_SHORT} - nothing estimated."}
    return report_outputs(
        "Tooling On-Hire Report",
        f"Tooling On-Hire Report | {plural(n)} on hire | {x['n_companies']} companies | "
        f"{money(x['total_val'])}",
        f"Ampol Tool Store - Tooling On-Hire Report - {plural(n)} on hire across "
        f"{x['n_companies']} companies",
        bl, limits, "Ampol Tooling - On-Hire Report", "Tooling On-Hire Report", cfg,
        HOW_REGISTER, standard_break=True, data_heading=False, cover=cover, card=card,
        pdf_subject=(f"Every tooling item on hire from the Ampol tool store at the SiteIQ "
                     f"pull of {ASAT_SHORT}: the position, what moved since the last pull, "
                     "the ageing by company and the complete register, company A to Z."))


def render_util(d):
    um = util_model(d)
    buy_priced = [r for r in um["buy"] if r["buy"] is not None]
    buy_spend = sum(r["buy"] for r in buy_priced)
    unpriced_n = len(um["buy"]) - len(buy_priced)
    story = ("Utilisation is measured two ways: <b>live</b> (what share of a group is on "
             "hire right now) and <b>year-to-date</b> (actual hire days against days "
             "available). Groups running hot are where crews wait for gear; groups "
             "running cold are capital sitting still. The buy list below is evidence-"
             "based - every line shows the demand that justifies it. Figures are computed "
             f"from the {stamp(d['asat']['stock']['pulled'])} SiteIQ exports: "
             f"{um['days_elapsed']} days elapsed this year to {fmt_date(um['today'])}.")
    bl = Blocks()
    bl.story(story)
    bl.tiles([tile(n_fmt(len(um["rows"])), "Equipment groups", "bars",
                   key="groups", raw=len(um["rows"]), good=None),
              tile(n_fmt(len(um["buy"])), "Buy signals", "zap",
                   key="buy_signals", raw=len(um["buy"]), good="down"),
              tile(money(buy_spend) if buy_priced else "-", "Est. buy spend, 1 each", "shield",
                   f"{len(buy_priced)} of {len(um['buy'])} priced", "grey",
                   f"Est. buy spend, 1 each ({len(buy_priced)} of {len(um['buy'])} priced)",
                   key="buy_spend", raw=round(buy_spend, 2) if buy_priced else None, good="down"),
              tile(n_fmt(unpriced_n), "Buy signals unpriced", "warn", "", "amber",
                   key="buy_signals_unpriced", raw=unpriced_n, good="down"),
              tile(n_fmt(len(um["overstock"])), "Right-size candidates", "box",
                   key="right_size", raw=len(um["overstock"]), good="down")])
    bl.h2("What to buy - demand-backed (ranked by live utilisation, then hire days)")
    bl.note(f"The spend estimate covers only the {len(buy_priced)} "
            f"priced groups; {unpriced_n} buy signals have no catalogue price yet and "
            "show a dash - they are not $0.")
    rows = []
    for r in um["buy"]:
        rows.append([r["group"], int(r["qty"]), int(r["on_hire"]), int(r["avail"]),
                     pct(r["live"]), pct(r["ytd"]), int(r["hirers"]), int(r["days"]),
                     money(r["buy"]) if r["buy"] is not None else "-",
                     r["source"] or "-", r["rec"]])
    bl.table(["Equipment Group", "Qty", "On Hire", "Avail", "Live %", "YTD %",
              "Hirers", "Hire Days", "Buy Price", "Source", "Recommendation"], rows,
             cls="tight")
    bl.h2("Working hardest (ranked by hire days)")
    # WHY (12 Aug 2026): the top-used list now gets a bar chart beside it -
    # same numbers the table carries (total hire days YTD), nothing new invented.
    n_with_days = um["groups_with_days"]
    if um["top_used"]:
        bl.chart(
            k2shell.hbars([(r["group"], int(r["days"])) for r in um["top_used"]],
                          colour=K_ORANGE),
            f"Total hire days year to date - showing {len(um['top_used'])} of "
            f"{n_with_days} equipment groups with hire days recorded.")
    bl.table(["Equipment Group", "Qty", "Live %", "YTD %", "Hire Days", "Hirers"],
             [[r["group"], int(r["qty"]), pct(r["live"]), pct(r["ytd"]),
               int(r["days"]), int(r["hirers"])] for r in um["top_used"]])
    if len(um["top_used"]) < n_with_days:
        bl.story(f"Showing {len(um['top_used'])} of {n_with_days} "
                 "groups with hire days recorded - the busiest first.")
    bl.h2("Potential overstock - right-size candidates (ranked by YTD utilisation, lowest first)")
    bl.table(["Equipment Group", "Qty", "On Hire", "Avail", "YTD %"],
             [[r["group"], int(r["qty"]), int(r["on_hire"]), int(r["avail"]),
               pct(r["ytd"])] for r in um["overstock"][:40]])
    if len(um["overstock"]) > 40:
        bl.story(f"Showing 40 of {len(um['overstock'])} right-size "
                 "candidates - lowest YTD utilisation first.")
    # ---- the insights pass (03 Sep 2026): dead stock, headroom, fast movers
    ui = util_insights(bl, d)
    # WHY (02 Sep 2026): coverage is stated both ways - how much of the
    # mapping is usable, and how much of the year's movement it explains.
    if d.get("mapping_n") is not None:
        share = (um["tx_used"] / um["tx_total"] * 100) if um["tx_total"] else 0
        map_bit = (f"Tooling_Description_Mapping.xlsx has {n_fmt(um['mapping_rows'])} rows; "
                   f"{n_fmt(um['mapping_usable'])} distinct barcodes are usable after the "
                   f"excluded-family filter, and {n_fmt(um['tx_used'])} of "
                   f"{n_fmt(um['tx_total'])} transactions ({share:.1f}%) feed utilisation. "
                   "Coverage grows as descriptions are corrected")
    else:
        map_bit = ("mapping coverage TBC - Tooling_Description_Mapping.xlsx was "
                   "not in the Data folder for this run")
    limits = ["Utilisation is computed by this kit from the SiteIQ TRANSACTIONS and "
              "RENTAL_STOCK exports, grouped by corrected description, using the same "
              "rules as the workbook's Tooling Utilisation query (hire days = distinct "
              "calendar days a barcode was out since 1 Jan; YTD % = hire days / (qty x "
              "days elapsed)). Items without a corrected description are not yet "
              "included - " + map_bit + ".",
              "Radios, gas monitors, Dräger equipment, lanyards and steel coil clamps are "
              "excluded here exactly as they are from the on-hire reports - one list for "
              "both.",
              "Buy prices are catalogue averages from Ampol_ToolStore_Pricing.xlsx (first "
              "price where a description is listed more than once); groups without a "
              "price show a dash.",
              "The dead stock, headroom and fast-mover pages are counted from the SiteIQ "
              "TRANSACTIONS export (CUSTOMER_CONTRACTOR_EQUIP, report period "
              f"{ui['period']}; {n_fmt(ui['tx_n'])} movements) against the RENTAL_STOCK "
              f"register. Their population is every register barcode with a tooling description "
              f"({n_fmt(ui['pop_n'])} barcodes, every status and account) - radios, gas monitors, "
              "Dräger equipment, lanyards and steel coil clamps out, as everywhere in this family; "
              "the store-wide never-moved figure counts every family. Rules: dead stock is an "
              "Available for Hire barcode with no movement in the log at all; a product is the "
              "SiteIQ description with its size and serial tail removed (the corrected name is "
              "shown where the mapping has one for every item under it); the peak out at once is "
              "counted from the log's start and end times; headroom is fleet less peak; BUY and "
              "CUT are the printed rule applied, not a decision; movements count every issue.",
              ] + source_limits(d)[:1]
    cfg = k2cfg("Utilisation & What-To-Buy Report",
                "COATES · TOOL STORE · UTILISATION & WHAT TO BUY", KEY_UTIL)
    return report_outputs(
        "Utilisation & What To Buy",
        f"Utilisation &amp; What-To-Buy | {len(um['buy'])} buy signals",
        f"Ampol Tool Store - Utilisation & What-To-Buy - "
        f"{len(um['buy'])} buy signals",
        bl, limits, "Ampol Tooling - Utilisation", "Utilisation & What-To-Buy Report",
        cfg, HOW_UTIL,
        pdf_subject=(f"Utilisation of the Ampol tool store's equipment groups at the SiteIQ "
                     f"pull of {ASAT_SHORT}, with the demand-backed buy list and the "
                     "right-size candidates."))


def render_compliance(d):
    cm = compliance_model(d)
    story = ("Electrical, rigging and high-torque gear carries test and inspection "
             "dates. Anything in those families that has not been through our hands "
             f"in {NOT_SIGHTED_DAYS} days is potentially out of test date - not a "
             "problem, just a reason for a five-minute visit to the counter. We check "
             "the tags, and if it is current it goes straight back with you. That is "
             "us doing the compliance work so your crews do not have to think about it.")
    bl = Blocks()
    bl.story(story)
    bl.tiles([tile(n_fmt(len(cm["chase"])), "Not seen 3+ months", "warn", "", "amber",
                   key="sightings_due", raw=len(cm["chase"]), good="down"),
              tile(n_fmt(len(cm["out_of_tag"])), "Caught & parked (out of tag)", "check",
                   "", "green", key="out_of_tag", raw=len(cm["out_of_tag"]), good=None),
              tile(n_fmt(len(cm["high_val"])), "High-value items out", "box",
                   key="high_value_items", raw=len(cm["high_val"]), good="down"),
              tile(money(sum(r['cost'] or 0 for r in cm['high_val'])),
                   "High-value exposure", "shield",
                   key="high_value_exposure",
                   raw=round(sum(r['cost'] or 0 for r in cm['high_val']), 2), good="down")])
    for cat in ("Electrical", "Rigging", "High Torque"):
        items = cm["by_cat"].get(cat, [])
        if not items:
            continue
        bl.h2(f"{esc(cat.capitalize())} - not sighted in {NOT_SIGHTED_DAYS}+ days "
              "(ranked by last sighted, oldest first)")
        rows = []
        for r in items:
            rows.append([r["company"] or "-", r["hirer"] or "-", r["barcode"],
                         N.display_desc(r["desc"]),
                         fmt_date(r["seen"]) if r["seen"] else "Never sighted",
                         money(r["cost"]) if r["cost"] is not None else "-"])
        bl.table(["Company", "Hirer", "Barcode", "Description", "Last Sighted",
                  "Replacement"], rows)
    if cm["out_of_tag"]:
        bl.h2("Already caught - parked out of tag (our checks working; A to Z)")
        bl.story("These items were caught by our checks and parked "
                 "under 'Rigging &amp; 240V - Out Of Tag Date' so they cannot be "
                 "issued. Nothing goes on the shelf unless it is ready for hire.")
        bl.table(["Barcode", "Description", "Status"],
                 [[r["barcode"], N.display_desc(r["desc"]), r["status"] or "-"]
                  for r in cm["out_of_tag"]])
    bl.h2("High-value items on hire (ranked by replacement value)")
    bl.table(["Company", "Hirer", "Barcode", "Description", "On Hire Since",
              "Replacement"], item_rows(cm["high_val"], with_company=True, limit=40))
    if len(cm["high_val"]) > 40:
        bl.story(f"Showing 40 of {len(cm['high_val'])} high-value "
                 "items - highest replacement cost first.")
    if cm["trend"]:
        bl.h2("Tool store activity by month (transactions)")
        # WHY (12 Aug 2026): 'Trends' finally gets a trend line - the same
        # monthly transaction counts the table below carries, drawn with the
        # shared K2 chart kit. The table stays; the chart is added beside it.
        period = d["asat"]["tx"]["period"] or "year to date"
        if len(cm["trend"]) >= 2:
            # WHY (03 Sep 2026): the shared line chart now takes (name, values)
            # series - same monthly counts as the table below
            bl.chart(
                sh.line_chart([mth for mth, _ in cm["trend"]],
                              [("Transactions", [n for _, n in cm["trend"]])],
                              y_label="transactions per month"),
                f"Tool store transactions per month by SiteIQ start date, export period "
                f"{period}. Every account and family is counted (custody and radio / gas "
                f"movements included); the current month is partial.")
        bl.table(["Month", "Transactions"],
                 [[m, f"{n:,}"] for m, n in cm["trend"]])
        bl.note("The current month covers only the days inside the "
                "export period and is labelled partial - it is not a full month's "
                "activity.")
    limits = [f"'Not sighted' uses the SiteIQ STOCKTAKE export (last sighted date per "
              f"barcode). An item on hire cannot be sighted in store - which is exactly "
              f"the point: {NOT_SIGHTED_DAYS}+ days out means we have not had hands on "
              f"it to verify tags.",
              "Category matching is by keywords in the register description (electrical "
              "/ rigging / high torque) - anything misfiled, tell us and we correct the "
              "mapping. The corrected name is printed where one exists.",
              "High-value = replacement cost " + money(HIGH_VALUE) + " or more.",
              "The monthly activity counts every transaction in the export - customer "
              "issues and returns plus custody movements (Repairs, Dräger) and the radio "
              "/ gas families the on-hire pages leave out.",
              ] + source_limits(d)
    cfg = k2cfg("Compliance & Trends Report", "COATES · TOOL STORE · COMPLIANCE & TRENDS",
                KEY_COMPLIANCE)
    return report_outputs(
        "Compliance & Trends",
        f"Compliance &amp; Trends | {len(cm['chase'])} to sight | "
        f"{len(cm['out_of_tag'])} caught",
        f"Ampol Tool Store - Compliance & Trends - {len(cm['chase'])} items to "
        f"sight",
        bl, limits, "Ampol Tooling - Compliance", "Compliance & Trends Report", cfg,
        HOW_COMPLIANCE,
        pdf_subject=(f"Electrical, rigging and high-torque tooling on hire from the Ampol tool "
                     f"store not sighted in {NOT_SIGHTED_DAYS} days at the SiteIQ pull of "
                     f"{ASAT_SHORT}, the gear caught out of tag, the high-value items and "
                     "the monthly activity."))


def recovery_table(x):
    """Recovery-by-company table from the master list; ties to the on-hire
    tile. Returns (headers, rows) for whichever skin draws it."""
    hdrs = ["Company"]
    for qk in QUARTERS:
        hdrs += [QTR_WORD[qk] + " Items", QTR_WORD[qk] + " Value"]
    hdrs += ["Total Items", "Total Value", "Priced / Unpriced"]

    def cells(row):
        out = [row["company"]]
        for qk in QUARTERS:
            vb = row["q"][qk]
            out += [vb["n"], money(vb["value"]) if vb["priced"] else "-"]
        t = row["total"]
        out += [t["n"], money(t["value"]) if t["priced"] else "-",
                f"{t['priced']} / {t['unpriced']}"]
        return out
    rows = [cells(r) for r in x["recovery"]]
    if x["custody_line"]:
        rows.append(cells(x["custody_line"]))
    # totals row, summed from the lines above so it cannot disagree with them
    lines = x["recovery"] + ([x["custody_line"]] if x["custody_line"] else [])
    tot = ["TOTAL"]
    for qk in QUARTERS:
        n = sum(r["q"][qk]["n"] for r in lines)
        v = sum(r["q"][qk]["value"] for r in lines)
        p = sum(r["q"][qk]["priced"] for r in lines)
        tot += [n, money(v) if p else "-"]
    n = sum(r["total"]["n"] for r in lines)
    v = sum(r["total"]["value"] for r in lines)
    p = sum(r["total"]["priced"] for r in lines)
    u = sum(r["total"]["unpriced"] for r in lines)
    tot += [n, money(v) if p else "-", f"{p} / {u}"]
    rows.append(tot)
    return hdrs, rows


def render_exec(d):
    x = exec_model(d)
    sd = pct(x["same_day_pct"]) if x["same_day_pct"] is not None else "-"
    leg = x["legacy"]
    fam_words = FAM_WORDS
    leg_fams = ", ".join(f"{n_fmt(n)} {fam_words.get(f, f.lower())}"
                         for f, n in sorted(leg["families"].items(), key=lambda kv: -kv[1])
                         if f != "Tooling")
    avail_fams = ", ".join(f"{n_fmt(n)} {fam_words.get(f, f.lower())}"
                           for f, n in sorted(x["available_family"].items(),
                                              key=lambda kv: -kv[1]))
    pulled = stamp(d["asat"]["stock"]["pulled"])
    # WHY (03 Sep 2026, layout pass): the position page reads band, tiles,
    # three things, then a three-line story - what the report is, the one
    # number, where the detail is. Every other figure the old opening
    # carried lives in the tiles or in the note on the page after; the rest
    # of its words open that page, figure-free.
    story = (f"The Ampol tool store position at the SiteIQ pull of {esc(pulled)}: "
             f"<b>{n_fmt(x['on_hire'])}</b> tooling items issued since 01 Jan {ASAT_DAY.year} "
             "are on hire. The tiles carry the figures; the quarterly recovery by company, "
             "the signals and the data page follow on the pages after.")
    bl = Blocks()
    # WHY (03 Sep 2026): the on-hire tile's sparkline is items still on hire
    # by the month their hire started (the Tooling On-Hire Report's month
    # table, counted the same way); the band is the over-90 share.
    master = d["master"]
    spark = [sum(1 for r in master if r["date"] and r["date"].month == m)
             for m in range(1, ASAT_DAY.month + 1)]
    over90_n = sum(1 for r in master if r["days"] is not None and r["days"] > 90)
    over90_vb = value_bits([r for r in master if r["days"] is not None and r["days"] > 90])
    p1 = [tile(n_fmt(x["on_hire"]), f"On hire (tooling, {ASAT_DAY.year})", "box",
               key="on_hire", raw=x["on_hire"], good=None, spark=spark),
          tile(money(x["total_val"]), "Value out (replacement)", "shield",
               key="value", raw=round(x["total_val"], 2), good=None),
          tile(n_fmt(x["available"]), "Available for hire (tooling)", "check",
               key="available", raw=x["available"], good="up"),
          tile(n_fmt(x["repairs"]), "In Repairs custody (register)", "wrench",
               key="repairs", raw=x["repairs"], good="down"),
          tile(n_fmt(x["tx_ytd"]), "Transactions YTD (all accounts)", "swap",
               key="transactions_ytd", raw=x["tx_ytd"], good=None),
          tile(sd, "Same-day returns", "clock",
               key="same_day_pct", good="up",
               raw=round(x["same_day_pct"] * 100.0, 1) if x["same_day_pct"] is not None else None)]
    band = rag_over90(over90_n, x["on_hire"], f"tooling items on hire ({ASAT_DAY.year})")
    bl.rag(band, tight=True)
    bl.tiles(p1, cls="compact")
    # WHY (03 Sep 2026): three things to do today, each drawn from the master
    # list; then the story, and on the page after what moved since the last pull
    bl.three(three_things_for(d))
    bl.story(story)
    bl.page_break()
    bl.story("The Ampol tool store is running the Coates Way: every issue and every return "
             "is double-scanned, and the quarterly recovery cycle keeps the fleet honest - gear "
             "not returned by quarter close is billable at replacement cost, and this report "
             "gives every company the list to beat before that happens.")
    since_last_pull(bl, d, full=False)
    bl.note(f"Available for hire counts tooling only: {n_fmt(x['available_all'])} register rows "
            f"are Available for Hire in all; the other {n_fmt(x['available_all'] - x['available'])} "
            f"are {avail_fams}, reported separately. "
            f"On hire excludes {n_fmt(leg['n'])} register rows on hire since before "
            f"01 Jan {ASAT_DAY.year} (oldest {fmt_date(leg['oldest'])}) - "
            + (f"all of them {leg_fams}, chased by the Radio and Gas Monitor reports; "
               "no tooling item on hire pre-dates this year"
               if leg["tooling_n"] == 0 else
               f"{leg_fams}; {n_fmt(leg['tooling_n'])} tooling")
            + f". Inside the {n_fmt(x['on_hire'])}: {n_fmt(x['custody_n'])} in the Repairs "
            f"custody account (internal) and {n_fmt(x['n_accounts'])} booked to shutdown / "
            f"custody accounts of a company rather than to a person. "
            f"Transactions YTD counts every movement in the export: "
            f"{n_fmt(x['tx_custody'])} are custody movements ("
            + ", ".join(f"{co} {n_fmt(n)}" for co, n in sorted(x["tx_by_custody_co"].items()))
            + f" and company custody accounts) and {n_fmt(x['tx_family'])} are radio / gas "
            f"/ lanyard family movements the on-hire pages leave out; "
            f"{n_fmt(x['tx_client_tooling'])} are customer tooling movements. "
            f"Same-day returns are measured on all {n_fmt(x['tx_ytd'])}.")
    bl.h2("On hire by quarter issued")
    qrows = []
    for qk in QUARTERS:
        n = x["qcounts"][qk]
        part = partial_note(d, QUARTERS[qk][2]) if quarter_started(qk) else ""
        qrows.append([QUARTERS[qk][1],
                      n_fmt(n) + ("" if quarter_started(qk) else " (quarter not started)")
                      + (f" ({ASAT_DAY.strftime('%B')} partial)" if part else "")])
    qrows.append(["Total", n_fmt(x["on_hire"])])
    bl.table(["Quarter", "Items Still On Hire"], qrows)
    if x["recovery"]:
        bl.h2("Quarterly recovery summary, by company A to Z")
        bl.note(f"{len(x['recovery'])} companies A to Z"
                + (" plus the Repairs custody account on its own line" if x["custody_line"] else "")
                + f"; the TOTAL ties to the {n_fmt(x['on_hire'])} on hire and "
                f"{money(x['total_val'])} above. Values are replacement (catalogue new-buy "
                "average); a dash means nothing priced in that cell. Project accounts "
                "(FCCU, SATGAS/MOL) roll into their company; the Tooling On-Hire Report "
                "shows the account against each hirer.")
        # WHY (12 Aug 2026): the recovery table leads with a bar chart -
        # total replacement value by company (top 12, and it says so).
        vals = sorted(((r["company"], r["total"]["value"]) for r in x["recovery"]),
                      key=lambda t: -t[1])
        top12 = [(co, round(v / 1000.0, 1)) for co, v in vals[:12]]
        bl.chart(
            k2shell.hbars(top12, colour=K_ORANGE),
            f"Total replacement value by company, in $'000 - ranked by replacement "
            f"value, top {len(top12)} of {len(vals)} companies; the table below is A to Z.")
        hdrs, rows = recovery_table(x)
        bl.table(hdrs, rows, cls="tight")
    if leg["n"]:
        bl.h2(f"Legacy on hire - before 01 Jan {ASAT_DAY.year} (ranked by legacy "
              f"items - top {len(leg['top'])} of {leg['n_companies']} companies)")
        bl.note(f"{n_fmt(leg['n'])} register rows are still on hire from "
                f"{ASAT_DAY.year - 3}-{ASAT_DAY.year - 1} issues (oldest {fmt_date(leg['oldest'])}). "
                "They are not in the tooling recovery cycle above because they are "
                + (leg_fams or "other families") + " - the Radio and Gas Monitor reports "
                "carry them. Shown here so the recovery story is complete; the full "
                "A-to-Z list by company is in the Tooling On-Hire Report.")
        fams = [f for f in ("RADIO", "DRAGER", "GAS MONITOR") if f in leg["families"]]
        bl.table(["Company", "Legacy Items"] + [fam_words[f].capitalize() for f in fams],
                 [[co, n_fmt(c["_n"])] + [n_fmt(c.get(f, 0)) for f in fams]
                  for co, c in leg["top"]],
                 aligns=["", "r"] + ["r"] * len(fams))
    # WHY (12 Aug 2026): two pictures drawn only from fields the kit already
    # loads - what families the on-hire gear falls into (the same keyword
    # classifier the compliance report uses) and how long it has been out.
    master = d["master"]
    cats = {"High Torque": 0, "Rigging": 0, "Electrical": 0, "General": 0}
    for r in master:
        cats[r["cat"] or "General"] += 1
    bl.h2("What is out - category split")
    bl.chart(
        k2shell.hbars(list(cats.items()), colour=K_ORANGE),
        "Items on hire by description family (High Torque / Rigging / Electrical "
        "keyword match on the register description; General is everything else).")
    age_rows = []
    for lab, lo, hi in (("0-30 days", 0, 30), ("31-60 days", 31, 60),
                        ("61-90 days", 61, 90), ("91-180 days", 91, 180),
                        ("Over 180 days", 181, None)):
        age_rows.append((lab, sum(
            1 for r in master if r["date"]
            and lo <= (ASAT_DAY - r["date"]).days
            and (hi is None or (ASAT_DAY - r["date"]).days <= hi))))
    undated = sum(1 for r in master if not r["date"])
    if undated:
        age_rows.append(("No on-hire date recorded", undated))
    bl.h2("On-hire ageing profile")
    bl.chart(
        k2shell.hbars(age_rows, colour=K_ORANGE),
        "How long the current on-hire items have been out - days since their "
        "on-hire date. The quarterly recovery cycle is what brings the long "
        "tail home.")
    bl.h2("Signals - buy, right-size and compliance")
    bl.tiles([tile(n_fmt(x["buy_n"]), "Buy signals (demand-backed)", "zap",
                   key="buy_signals", raw=x["buy_n"], good="down"),
              tile(n_fmt(x["overstock_n"]), "Right-size candidates", "bars",
                   key="right_size", raw=x["overstock_n"], good="down"),
              tile(n_fmt(x["chase_n"]), "Test-date sightings due", "warn", "", "amber",
                   key="sightings_due", raw=x["chase_n"], good="down"),
              tile(n_fmt(x["out_of_tag_n"]), "Caught & parked out-of-tag", "check", "", "green",
                   key="out_of_tag", raw=x["out_of_tag_n"], good=None),
              tile(money(x["high_val_sum"]), "High-value exposure", "shield",
                   key="high_value_exposure", raw=round(x["high_val_sum"], 2), good="down")])
    bl.story("Detail sits in the companion reports: the Tooling "
             "On-Hire Report (every company, hirer and item, A to Z), the Quarterly "
             "On-Hire charge reports, Utilisation &amp; What-To-Buy, and Compliance "
             "&amp; Trends - each one client-ready, each one emailable from this "
             "kit.")
    # ---- the trend page (03 Sep 2026): only once seven days are on record
    trend_shown = trend_page(bl, d, {"on_hire": x["on_hire"], "over90": over90_n,
                                     "value": round(x["total_val"], 2)})
    # ---- the insights pass (03 Sep 2026): four store-wide pages from the
    # transaction log and the register, after the position, before the data page
    ins = exec_insights(bl, d)
    qa, qt = ins["qc_all"], ins["qc_tool"]
    limits = ["Every figure is counted from the SiteIQ exports beside this report - never "
              "from a workbook tab or a hardcoded summary cell. On hire = register rows "
              f"On Hire with a {ASAT_DAY.year} on-hire date, less radios, gas monitors, "
              "Dräger equipment, lanyards and steel coil clamps (reported separately) and "
              "less the tool store's own holding accounts (T&I - Tool store, All-Around - "
              "Repairs, Bulk - Yard, Loading Bay - Out Of Service, Out Of - Calibration, "
              "Rigging & 240V - Out Of Tag Date).",
              "Company names follow the suite's one naming rule: the site's former-name "
              "account and the refinery legal name both read Ampol; CR reads Contract "
              "Resources; FCCU and SATGAS/MOL project accounts roll up to their company "
              "(the Tooling On-Hire Report shows the account against each hirer). Repairs "
              "is an internal custody account, never a company.",
              "Three things to do today are drawn from the same on-hire list: the company "
              "holding the most items over 90 days, the oldest item out, and the unpriced "
              "items (or the second-largest over-90 holder when everything is priced). Since "
              "the last pull compares the register with the newest earlier export parked in "
              "Data\\previous, and the 24 hours before the pull are counted from the "
              "TRANSACTIONS export - both for this report's population "
              f"({SCOPE_WORDS}).",
              trend_line(trend_shown),
              f"{n_fmt(qa['n'])} items on hire across the whole register will have been out 90 "
              f"days or more by {fmt_date(qa['qend'])} unless they come back - arithmetic on each "
              "item's on-hire date (an item under 90 days out at the pull whose on-hire date is "
              f"90 days or more before the close), never a forecast; {n_fmt(qa['already_over'])} "
              f"are already over 90 days. The tooling share - {n_fmt(qt['n'])} of the "
              f"{n_fmt(x['on_hire'])} on hire on page 1 - is on the Tooling On-Hire Report's "
              "quarter-close page.",
              "The store-wide pages (the store's year in movements, who holds what, the "
              "counter's rhythm, what the log and the register disagree on) are counted from "
              "the SiteIQ TRANSACTIONS export (CUSTOMER_CONTRACTOR_EQUIP, report period "
              f"{ins['period']}; {n_fmt(ins['tx_n'])} movements) and the RENTAL_STOCK register "
              f"({n_fmt(ins['reg_n'])} barcodes), every account and every family - the whole "
              "store, not the tooling population on page 1. Rules: a short hire is a movement "
              "closed inside 6 minutes; a mass draw is one person with 15 or more items in one "
              "hour; a 90-day crossing is as above; a product is the SiteIQ description with its "
              "size and serial tail removed; weeks are ISO weeks by start time (issues) and end "
              "time (returns). Nothing is corrected, weighted or forecast.",
              ] + source_limits(d)
    cfg = k2cfg("Executive Summary", "COATES · TOOL STORE · EXECUTIVE SUMMARY", KEY_EXEC)
    # WHY (03 Sep 2026): the cover and the phone position card - the same
    # page-1 values, nothing counted twice; the cover's stripe is the band's
    # own status and the freshness line is the honest age of the data
    def cover(contents):
        return k2flow.cover_block(cfg, n_fmt(x["on_hire"]), "tooling items on hire", [
            f"<b>{money(x['total_val'])}</b> of replacement value in the field "
            f"({n_fmt(x['priced'])} priced / {n_fmt(x['unpriced'])} unpriced)",
            f"<b>{n_fmt(len(x['companies']))}</b> companies, A to Z",
            over90_words([r for r in master if r["days"] is not None and r["days"] > 90])],
            GENERATED, ASAT_SHORT, rag=band["status"],
            fresh=sh.freshness_line(ASAT_DT, GENERATED_DT), contents=contents)
    card = {"cfg": cfg, "tiles": card_tiles(p1),
            "band": (band["status"], band["card_head"], band["owner"], band["card_action"]),
            "scores": card_scores(x["on_hire"], x["priced"], x["same_day_pct"], over90_n,
                                  x["buy_n"], x["buy_priced_n"]),
            "foot": f"Counted from the SiteIQ exports of {ASAT_SHORT} - nothing estimated."}
    return report_outputs(
        "Executive Summary",
        f"Executive Summary | {n_fmt(x['on_hire'])} on hire | {money(x['total_val'])}",
        f"Ampol Tool Store - Executive Summary - {plural(x['on_hire'])} on hire",
        bl, limits, "Ampol Tooling - Executive Summary", "Executive Summary", cfg,
        HOW_EXEC, cover=cover, card=card,
        pdf_subject=(f"The Ampol tool store position at the SiteIQ pull of {ASAT_SHORT}: "
                     "tooling on hire, its replacement value, the three things to do today, "
                     "what moved since the last pull, the quarterly recovery by company and "
                     "the buy, right-size and compliance signals."))


# ------------------------------------------------------------------ email ---
def email_html(subtitle, inner_note, html_doc, card_html=""):
    """Outlook-safe like-for-like body: reuse the report body inside a bordered
    680px card. We inline the full report HTML converted to nested-table-safe
    markup by keeping our simple structure (tables + divs render acceptably in
    Outlook's Word engine because all styling is inline-safe).
    card_html (03 Sep 2026): the inline position card (a cid image) printed
    under the opening line - the .eml only; the draft manifest's body stays
    without it because an Outlook draft made by PowerShell cannot resolve a
    cid part."""
    # extract body content between the marker comments page() now writes
    # WHY (12 Aug 2026): the old wrap/footer regex silently made the WHOLE
    # document the email body the moment page() formatting shifted. The
    # markers survive any styling change; the old regex stays as a fallback
    # so a marker-less document still extracts the same way it always did.
    m = re.search(r"<!--BODY-START-->(.*)<!--BODY-END-->", html_doc, re.S)
    if not m:
        m = re.search(r"<div class=\"wrap\">(.*)</div>\n<div class=\"footer\">",
                      html_doc, re.S)
    inner = m.group(1) if m else html_doc
    # WHY (12 Aug 2026): the new SVG charts do not survive Outlook's Word
    # engine, so they are stripped here - the email keeps the tables that
    # carry the same numbers, exactly as it always has.
    inner = re.sub(r"<!--CHART-->.*?<!--/CHART-->", "", inner, flags=re.S)
    # WHY (02 Sep 2026): a report can mark where its email should stop
    # (the Tooling On-Hire Report's register runs to many pages - the email
    # carries the position and the A-to-Z company table, the PDF the rest).
    inner = inner.split("<!--EMAIL-END-->")[0]
    inner = inner.replace('<h2 class="pb">', '<h2>')
    # the register's own elements, inlined for Outlook's Word engine
    inner = inner.replace('<tr class="cohead"><td colspan="5">',
                          '<tr><td colspan="5" style="background:' + DARK + ';color:#fff;'
                          'padding:7px 10px;font-size:11px;">')
    inner = inner.replace('<tr class="hirer"><td colspan="5">',
                          '<tr><td colspan="5" style="background:#EADFD4;padding:5px 8px;'
                          'font-size:10px;font-weight:700;color:' + DARK + ';">')
    inner = inner.replace('<span class="nm">', '<span style="font-size:12px;font-weight:800;">')
    inner = inner.replace('<span class="st">',
                          '<span style="font-size:10px;color:#dddddd;margin-left:8px;">')
    inner = inner.replace('<div class="ac">',
                          '<div style="font-size:9.5px;color:#f5c9b3;margin-top:3px;">')
    inner = inner.replace('<span class="acc">',
                          '<span style="font-weight:400;color:' + AMBER + ';">')
    inner = inner.replace('<td class="num">',
                          '<td style="padding:4px 7px;border-bottom:1px solid #e5e0da;'
                          'font-size:10px;color:' + DARK + ';text-align:right;'
                          'white-space:nowrap;">')
    inner = inner.replace('<th class=num>',
                          '<th style="background:' + DARK + ';color:#fff;text-align:right;'
                          'padding:5px 7px;font-size:9px;text-transform:uppercase;">')
    inner = inner.replace('<table class="reg">', '<table>')
    inner = inner.replace('<div class="defbox">',
                          '<div style="border:1px solid #e0dad2;border-left:5px solid '
                          + ORANGE + ';background:' + LIGHT + ';padding:10px 14px;'
                          'font-size:11px;line-height:1.65;margin:12px 0;color:' + DARK + ';">')
    inner = re.sub(r"<tbody( class=\"keep\")?>", "<tbody>", inner)
    # convert class-styled elements to inline styles for Outlook
    inner = inner.replace('<div class="tiles three">', '<div class="tiles">')
    inner = inner.replace('<div class="tiles">',
                          '<div style="margin:12px 0;">')
    inner = inner.replace('<div class="tile">',
                          '<div style="display:inline-block;border:1px solid #e0dad2;'
                          'border-bottom:3px solid ' + ORANGE + ';padding:10px 14px;'
                          'margin:0 6px 6px 0;">')
    inner = inner.replace('<div class="v">',
                          '<div style="font-size:20px;font-weight:800;color:'
                          + ORANGE + ';">')
    inner = inner.replace('<div class="l">',
                          '<div style="font-size:9px;color:' + GREY +
                          ';text-transform:uppercase;letter-spacing:1px;">')
    inner = inner.replace('<p class=\'story\'>',
                          '<p style="font-size:12px;line-height:1.7;color:' + DARK +
                          ';">').replace('<p class="story">',
                          '<p style="font-size:12px;line-height:1.7;color:' + DARK +
                          ';">')
    inner = inner.replace('<p class=\'note\'>',
                          '<p style="font-size:10px;line-height:1.6;color:' + GREY +
                          ';">')
    inner = inner.replace('<div class="subhead">',
                          '<div style="font-size:10px;font-weight:700;color:' + AMBER +
                          ';margin:8px 0 2px 0;">')
    inner = inner.replace('<h2>', '<h2 style="font-size:13px;border-left:5px solid '
                          + ORANGE + ';padding-left:9px;text-transform:uppercase;'
                          'letter-spacing:1px;color:' + DARK + ';margin:22px 0 8px 0;">')
    inner = inner.replace('<table class="tight">', '<table>')
    inner = inner.replace('<table>', '<table cellspacing="0" cellpadding="0" '
                          'width="100%" style="border-collapse:collapse;font-size:10px;'
                          'margin:8px 0;">')
    inner = inner.replace('<th>', '<th style="background:' + DARK + ';color:#fff;'
                          'text-align:left;padding:5px 7px;font-size:9px;'
                          'text-transform:uppercase;">')
    inner = inner.replace('<td>', '<td style="padding:4px 7px;border-bottom:1px solid '
                          '#e5e0da;font-size:10px;color:' + DARK + ';">')
    inner = inner.replace('<div class="cochip">',
                          '<div style="background:' + DARK + ';color:#fff;padding:4px '
                          '10px;font-size:11px;font-weight:700;display:inline-block;'
                          'margin-top:12px;">')
    # WHY (12 Aug 2026): a literal no-op replace('<ul style=', '<ul style=')
    # used to sit here - dead code, removed.
    return f"""<table role="presentation" width="100%" cellspacing="0" cellpadding="0"
 style="background:#EFECE7;"><tr><td align="center" style="padding:18px 8px;">
<table role="presentation" width="680" cellspacing="0" cellpadding="0"
 style="width:680px;max-width:680px;background:#ffffff;border:2px solid {DARK};
 border-collapse:collapse;">
<tr><td style="background:{DARK};padding:20px 26px;">
  <div style="font-family:Arial,sans-serif;font-size:22px;font-weight:900;color:#fff;">Coates &nbsp;|&nbsp; Ampol Tool Store</div>
  <div style="font-family:Arial,sans-serif;font-size:12px;font-weight:700;color:{ORANGE};margin-top:5px;">{subtitle}</div>
  <div style="font-family:Arial,sans-serif;font-size:10px;color:#bbbbbb;margin-top:6px;">Generated {esc(GENERATED)} &nbsp;|&nbsp; Data as at {esc(DATA_ASAT)}</div>
  <div style="font-family:Arial,sans-serif;font-size:10px;color:#bbbbbb;margin-top:3px;">The Coates Way &nbsp;|&nbsp; POWERED BY SITEIQ &nbsp;|&nbsp; Author: Andrew Fisher</div>
</td></tr>
<tr><td style="padding:18px 26px;font-family:Arial,sans-serif;">
  <p style="font-size:11px;color:{GREY};margin:0 0 10px 0;">{inner_note}</p>
  {card_html}
  {inner}
</td></tr>
<tr><td style="background:{LIGHT};border-top:3px solid {ORANGE};padding:12px 26px;
 font-family:Arial,sans-serif;font-size:9px;color:{GREY};line-height:1.7;">
 Data as at {esc(DATA_ASAT)} |
 Coates Hire Operations Pty Limited | ABN 50 009 779 338 | www.coates.com.au |
 POWERED BY SITEIQ<br/>Care Deeply &middot; Customer Focused &middot; Be Our Best &middot;
 One Team &middot; Competitive Spirit</td></tr>
</table></td></tr></table>"""


def edge_pdf(html_path, pdf_path):
    import shutil
    import tempfile
    import time
    # WHY (12 Aug 2026): a stale PDF already sitting on the target name used
    # to make a failed print look like a success - clear it first, so the only
    # PDF that can pass the exists-check is the one this run actually wrote.
    if os.path.exists(pdf_path):
        try:
            os.remove(pdf_path)
        except OSError as e:
            sys.exit(f"Cannot replace {os.path.basename(pdf_path)} - close it "
                     f"if it is open, then run again. ({e})")
    # WHY (12 Aug 2026): the fixed Edge paths only exist on Windows, so they
    # stay behind an os.name guard; any Edge/Chrome/Chromium found on PATH is
    # now a candidate too, so the kit prints PDFs on whatever machine runs it.
    candidates = []
    if os.name == "nt":
        candidates += [r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
                       r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"]
    for name in ("msedge", "chrome", "chromium", "chromium-browser",
                 "google-chrome"):
        p = shutil.which(name)
        if p:
            candidates.append(p)
    # WHY (12 Aug 2026): throwaway profiles used to pile up beside the script
    # when TEMP was unset - they now go to the system temp folder everywhere.
    base_profile = os.path.join(tempfile.gettempdir(), "coates_edge_pdf")
    url = "file:///" + os.path.abspath(html_path).replace("\\", "/").lstrip("/")
    edge_pdf._n = getattr(edge_pdf, "_n", 0) + 1
    for exe in candidates:
        if not os.path.exists(exe):
            continue
        for attempt in range(3):
            profile = f"{base_profile}_{os.getpid()}_{edge_pdf._n}_{attempt}"
            try:
                subprocess.run([exe, "--headless", "--disable-gpu",
                                "--user-data-dir=" + profile,
                                "--no-first-run", "--no-pdf-header-footer",
                                "--print-to-pdf=" + pdf_path, url],
                               timeout=120, capture_output=True)
            except FileNotFoundError:
                break        # not runnable here - on to the next candidate
            except (subprocess.SubprocessError, OSError):
                pass         # retry with a fresh profile
            if os.path.exists(pdf_path):
                return True
            time.sleep(1.5)
        # WHY (12 Aug 2026): the old unconditional return here meant only the
        # first browser found ever got a go - every candidate gets one now.
    return False


CARD_CID = "positioncard"
CARD_IMG = (f'<img src="cid:{CARD_CID}" alt="The position card" width="420" '
            'style="display:block;max-width:420px;width:100%;height:auto;'
            'margin:4px 0 14px 0;border-radius:8px;">')


def write_eml(eml_path, subject, body_html, attach_paths, inline_png=None):
    """An X-Unsent .eml beside the manifest - double-click it and Outlook
    opens an editable DRAFT. No To line: Andrew addresses it himself.

    WHY (12 Aug 2026): the kit now ships the .eml as well as the manifest,
    so the drafts flow works even on a machine without the PowerShell step -
    and nothing can ever send itself either way.
    WHY (03 Sep 2026): takes a list - the PDF and, for the client-facing
    reports, the position card PNG beside it. inline_png embeds that card
    in the body as a cid part (the body carries <img src="cid:...">), and
    the same PNG still rides as a normal attachment so it can be saved.
    """
    from email.mime.application import MIMEApplication
    from email.mime.image import MIMEImage
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    if isinstance(attach_paths, str):
        attach_paths = [attach_paths]
    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["X-Unsent"] = "1"
    if inline_png and os.path.exists(inline_png):
        related = MIMEMultipart("related")
        related.attach(MIMEText(body_html, "html", "utf-8"))
        with open(inline_png, "rb") as f:
            img = MIMEImage(f.read(), _subtype="png")
        img.add_header("Content-ID", f"<{CARD_CID}>")
        img.add_header("Content-Disposition", "inline",
                       filename=os.path.basename(inline_png))
        related.attach(img)
        msg.attach(related)
    else:
        msg.attach(MIMEText(body_html, "html", "utf-8"))
    for attach_path in attach_paths or []:
        if not (attach_path and os.path.exists(attach_path)):
            continue
        with open(attach_path, "rb") as f:
            data = f.read()
        if attach_path.lower().endswith(".png"):
            part = MIMEImage(data, _subtype="png")
        else:
            sub = "pdf" if attach_path.lower().endswith(".pdf") else "octet-stream"
            part = MIMEApplication(data, _subtype=sub)
        part.add_header("Content-Disposition", "attachment",
                        filename=os.path.basename(attach_path))
        msg.attach(part)
    with open(eml_path, "wb") as f:
        f.write(msg.as_bytes())


CLOSING_HEADING = "Our standard, honest limits and how to read this report"
# WHY (03 Sep 2026): the headings the cover's "What's inside" leaves out -
# the closing page, the team, and the sub-sections that sit under a listed
# section - so the ten rows the cover holds are the sections a reader
# turns to (matched on the start of the heading, so a count in it does
# not matter)
CONTENTS_SKIP_PREFIXES = (CLOSING_HEADING, "Meet the team", "Your Coates tool store team",
                          "Where the count sits", "Came back", "Went out", "Changed hands",
                          "Crossed 90 days", "The 24 hours before the pull", "Category split",
                          "Items on hire by month started", "Rules applied",
                          "Repairs custody account", "Tool store holding accounts",
                          "Radio and gas families")
PAGE1_SPARE_MIN = 15      # px the position page must have left (movement notes land under the tiles)


def contents_skip(doc):
    return tuple(t for _lv, t in pdf_finish.headings_from_html(doc)
                 if any(t.startswith(p) for p in CONTENTS_SKIP_PREFIXES))


def pdf_pages(pdf_path):
    """Page count of a printed PDF (PyMuPDF, then pypdf); None with neither."""
    try:
        import pymupdf
        with pymupdf.open(str(pdf_path)) as doc:
            return len(doc)
    except ImportError:
        try:
            from pypdf import PdfReader
            return len(PdfReader(str(pdf_path)).pages)
        except ImportError:
            return None
    except Exception:
        return None


def page1_spare(pdf_path, has_cover):
    """The room left on the position page, measured on the PRINTED PDF
    (03 Sep 2026): the lowest text, drawing or image on page 1 (page 2
    behind a cover) against the line the print engine breaks the page at
    - the k2flow page area less the body's cloned 6 mm bottom padding
    (checked against a full register page, whose content ends exactly
    there). In CSS pixels; None when nothing could be measured."""
    try:
        import pymupdf
    except ImportError:
        return None
    try:
        mm = 72 / 25.4
        with pymupdf.open(str(pdf_path)) as doc:
            page = doc[1 if (has_cover and len(doc) > 1) else 0]
            H, W = page.rect.height, page.rect.width
            top, bottom = 21 * mm, H - 17.5 * mm          # the @page margins in k2flow
            limit = bottom - 6 * mm                        # the body's bottom padding
            bots = [b[3] for b in page.get_text("blocks") if b[1] >= top - 1 and b[3] <= bottom + 1]
            for dr in page.get_drawings():
                r = dr["rect"]
                if r.height > 0.8 * H or (r.y0 < top + 2 and r.y1 > bottom - 2):
                    continue                               # the frame itself
                if r.y0 >= top - 1 and r.y1 <= bottom + 1:
                    bots.append(r.y1)
            for im in page.get_image_info():
                _x0, y0, _x1, y1 = im["bbox"]
                if y0 >= top - 1 and y1 <= bottom + 1:
                    bots.append(y1)
        if not bots:
            return None
        return int(round((limit - max(bots)) * 96 / 72))
    except Exception:
        return None


def _write_text(path, text):
    with open(path, "w", encoding="utf-8") as f:
        f.write(text)


def write_outputs(key, out, note=None):
    """Write one report under the suite's one file-name rule. key is an
    ampol_names.REPORT_STEMS key (exec, onhire, q1..q4, year, util,
    compliance) - or, for the one-off company report, a ready-made stem in
    the same shape. out is what report_outputs returned.

    Every file derives from the stem: <stem>.html (the page in the house
    frame, printed to <stem>.pdf), <stem>.body.html (the Outlook-safe email
    body the draft manifest <stem>.draft.json points at), <stem>_OUTLOOK.eml
    and, for the client-facing reports, <stem>_PositionCard.png.
    WHY (02 Sep 2026): the email builder strips a page by its BODY markers,
    which the house frame does not carry - so it is handed the legacy page
    instead of the printed one.
    WHY (03 Sep 2026): after every print pdf_finish stamps the PDF's
    properties (Author: Andrew Fisher, Subject) and builds its bookmarks,
    BEFORE the .eml is written, so the attached PDF carries them. The
    position card is drawn from the page-1 values, saved beside the PDF,
    inlined in the .eml body (cid) and attached as well; the manifest body
    stays without the cid part."""
    os.makedirs(OUT_DIR, exist_ok=True)
    stem = N.report_stem(key) if key.lower() in N.REPORT_STEMS else key
    base = os.path.join(OUT_DIR, stem)
    title, subtitle, subject = out["title"], out["subtitle"], out["subject"]
    html_path = base + ".html"
    pdf_path = base + ".pdf"
    doc = out["doc"]
    _write_text(html_path, doc)
    pdf_ok = edge_pdf(html_path, pdf_path)
    if pdf_ok and out["has_cover"]:
        # WHY (03 Sep 2026): the cover's "What's inside" block carries REAL
        # page numbers - read off the printed PDF, never guessed. The page is
        # printed once, the level-1 headings are located on its pages, the
        # page is rebuilt with that list and printed again. The cover is a
        # fixed-height block, so the pagination cannot move: the page count
        # is asserted equal and the console says so.
        n1 = pdf_pages(pdf_path)
        contents = pdf_finish.contents_from_pdf(pdf_path, doc, has_cover=True,
                                                skip=contents_skip(doc))
        if contents:
            doc = out["build"](contents)
            _write_text(html_path, doc)
            pdf_ok = edge_pdf(html_path, pdf_path)
            n2 = pdf_pages(pdf_path) if pdf_ok else None
            if pdf_ok and n1 != n2:
                sys.exit(f"ERROR: {stem}.pdf went from {n1} to {n2} pages when the cover "
                         "contents were added - the contents would be wrong. Not written.")
            print(f"  Cover contents: {len(contents)} rows read off the printed pages "
                  f"({n1} pages before, {n2} after - the same)")
        else:
            print("  Cover contents: nothing could be read off the printed pages - the cover "
                  "prints no What's inside block")
    if pdf_ok:
        n_pages = pdf_pages(pdf_path)
        spare = page1_spare(pdf_path, out["has_cover"])
        line = f"  Pages: {n_pages}"
        if spare is not None and out.get("position_page"):
            line += f"; position page has {spare} px to spare"
            if spare < PAGE1_SPARE_MIN:
                line += f" - WARNING: under the {PAGE1_SPARE_MIN} px the movement notes need"
        print(line)
        print("  " + pdf_finish.finish(pdf_path, f"Ampol {title} - as at {ASAT_SHORT}",
                                       out["pdf_subject"], doc, has_cover=out["has_cover"],
                                       family="Tooling"))
    note = note or ("The full report PDF is attached. This email is the same report, "
                    "formatted for reading in place.")
    body = email_html(subtitle, note, out["mail"])
    with open(base + ".body.html", "w", encoding="utf-8") as f:
        f.write(body)
    attach_paths = [pdf_path if pdf_ok else html_path]
    card_ok, png_path = False, None
    card = out.get("card")
    if card:
        png_path = base + "_PositionCard.png"
        sh.position_card_png(card["cfg"], ASAT_SHORT, card["tiles"], card["band"],
                             card["scores"], png_path, card.get("foot", ""))
        card_ok = os.path.exists(png_path)
        if card_ok:
            attach_paths.append(png_path)
    attach = [os.path.basename(p) for p in attach_paths]
    with open(base + ".draft.json", "w", encoding="utf-8") as f:
        json.dump({"subject": subject, "body": os.path.basename(base + ".body.html"),
                   "attachments": attach}, f, indent=1)
    eml_body = email_html(subtitle, note, out["mail"], card_html=CARD_IMG) if card_ok else body
    write_eml(base + "_OUTLOOK.eml", subject, eml_body, attach_paths,
              inline_png=png_path if card_ok else None)
    print(f"  {title}: {stem} - HTML" + (" + PDF" if pdf_ok else " (PDF skipped)")
          + (" + position card PNG (inlined in the .eml)" if card_ok else "")
          + " + .eml + email draft manifest")
    return base


# ------------------------------------------------------------------- main ---
def company_list(d):
    names = sorted({r["company"] for r in d["master"] if r["company"]})
    return [n for n in names if n.upper() not in INTERNAL_CUSTODY]


def run_quarter(d, qk):
    # WHY (02 Sep 2026): a quarter that has not started used to ship a report
    # and an email saying "0 items" - now it is skipped and the console says so.
    if qk != "YEAR" and not quarter_started(qk):
        start = dt.date(ASAT_DAY.year, QUARTERS[qk][2][0], 1)
        print(f"  Quarterly On-Hire Report - {QUARTERS[qk][1]}: quarter not started "
              f"(begins {fmt_date(start)}) - skipped")
        return
    write_outputs("year" if qk == "YEAR" else qk.lower(), render_quarter(d, qk))


def run_company(d, name):
    # WHY (03 Sep 2026): the one-off company report has no key in the shared
    # REPORT_STEMS table, so its stem is built here in the same shape
    safe = re.sub(r"[^\w]+", "_", name).strip("_")[:40]
    write_outputs(f"Coates_Ampol_Company_On_Hire_{safe}_{N.day_tag()}", render_company(d, name))


def run_onhire(d):
    # WHY (02 Sep 2026): the one tooling on-hire report for every company.
    # Its email carries page 1 and the A-to-Z company table; the full
    # register is the attached PDF. Same recipients as the Executive
    # Summary: the draft is addressed in Outlook, exactly as that one is.
    write_outputs("onhire", render_onhire(d),
                  note="The full Tooling On-Hire Report PDF is attached - every company "
                       "A to Z, every hirer, every item. This email carries the position, "
                       "what moved since the last pull, the quarter-close look forward, the "
                       "return windows and the company tables; the register and the charts "
                       "are in the PDF.")


def run_util(d):
    write_outputs("util", render_util(d))


def run_compliance(d):
    write_outputs("compliance", render_compliance(d))


def run_exec(d):
    write_outputs("exec", render_exec(d))


def history_figures(d):
    """The day's key figures for the movement scoreboard - the same values
    the pages print, keyed by the names the tiles use - and (03 Sep 2026)
    the family's position for the daily position page: the Tooling On-Hire
    Report's band, its cover figures and the files it wrote. Returns
    (figures, extra)."""
    x = exec_model(d)
    oh = onhire_model(d)
    um = d["util"]
    f = {"on_hire": x["on_hire"], "value": round(x["total_val"], 2),
         "priced": x["priced"], "unpriced": x["unpriced"],
         "available": x["available"], "repairs": x["repairs"],
         "companies": len(x["companies"]), "hirers": oh["people_names"],
         "over90": len(oh["over90"]), "over90_value": round(oh["over90_vb"]["value"], 2),
         "oldest_days": oh["oldest"]["days"] if oh["oldest"] else None,
         "transactions_ytd": x["tx_ytd"],
         "same_day_pct": (round(x["same_day_pct"] * 100.0, 1)
                          if x["same_day_pct"] is not None else None),
         "buy_signals": x["buy_n"], "buy_signals_priced": x["buy_priced_n"],
         "buy_signals_unpriced": x["buy_n"] - x["buy_priced_n"],
         "buy_spend": round(sum(r["buy"] for r in um["buy"] if r["buy"] is not None), 2),
         "right_size": x["overstock_n"], "groups": len(um["rows"]),
         "sightings_due": x["chase_n"], "out_of_tag": x["out_of_tag_n"],
         "high_value_items": x["high_val_n"], "high_value_exposure": round(x["high_val_sum"], 2)}
    for qk in list(QUARTERS) + ["YEAR"]:
        m = quarter_model(d, qk)
        f["on_hire_" + qk] = len(m["rows"])
        f["companies_" + qk] = m["n_companies"]
        f["value_" + qk] = round(m["total_val"], 2)
    band = rag_over90(len(oh["over90"]), x["on_hire"], f"tooling items on hire ({ASAT_DAY.year})")
    stem = N.report_stem("onhire")
    extra = {"rag": band["status"], "headline": band["head_txt"], "rule": band["rule"],
             "owner": band["owner"], "action": band["action_txt"],
             "due": fmt_date(ASAT_DAY + dt.timedelta(days=CONFIG["rag_due_days"])),
             "key_value": n_fmt(x["on_hire"]), "key_label": "tooling items on hire",
             "second_value": n_fmt(len(oh["over90"])),
             "second_label": "items out more than 90 days",
             "title": "Ampol Tooling On-Hire Report", "folder": "Tooling",
             "pdf": stem + ".pdf", "card": stem + "_PositionCard.png"}
    # WHY (03 Sep 2026, insights pass): the third figure for the daily position
    # - the tooling items that cross 90 days by the quarter close unless returned
    qt = ti.quarter_close(insights(d), master_barcodes(d))
    extra["third_value"] = n_fmt(qt["n"])
    extra["third_label"] = f"cross 90 days by {qt['qend']:%d %b} unless returned"
    return f, extra


def record_history(d):
    """WHY (03 Sep 2026): written once at the end of a build, keyed by the
    pull day - tomorrow's tiles read it back as the movement, and the daily
    position page reads the extra (band, cover figures, files). A re-run on
    the same pull day replaces the entry; it never doubles up."""
    figures, extra = history_figures(d)
    path = rh.record("tooling", ASAT_DT or ASAT_DAY, figures, extra=extra)
    print(f"  Movement scoreboard: {os.path.relpath(path, HERE)} holds tooling "
          f"{ASAT_DAY.isoformat()} (band {extra['rag']}, {extra['key_value']} "
          f"{extra['key_label']})")


def run_everything(d):
    # WHY (02 Sep 2026): the per-company reports are no longer part of the
    # everything run - the Tooling On-Hire Report covers every company in
    # one document. A one-off company report is still --company NAME.
    run_exec(d)
    for q in list(QUARTERS) + ["YEAR"]:
        run_quarter(d, q)
    run_onhire(d)
    run_util(d)
    run_compliance(d)


def console_summary(d):
    a = d["asat"]
    print(f"SiteIQ pull (as at) : RENTAL_STOCK {stamp(a['stock']['pulled'])} | "
          f"TRANSACTIONS {stamp(a['tx']['pulled'])}"
          + (f" (period {a['tx']['period']})" if a['tx']['period'] else "")
          + f" | STOCKTAKE {stamp(a['stocktake']['pulled'])}")
    sc = d["status_counts"]
    print(f"Register rows       : {len(d['stock']):,} ("
          + ", ".join(f"{v:,} {k}" for k, v in sorted(sc.items(), key=lambda kv: -kv[1]))
          + ")")
    qc = {qk: sum(1 for r in d["master"] if r["q"] == qk) for qk in QUARTERS}
    cust = sum(1 for r in d["master"] if r["custody"])
    leg = legacy_model(d)
    print(f"On hire (tooling {ASAT_DAY.year}) : {len(d['master']):,}  "
          f"(Q1 {qc['Q1']:,} / Q2 {qc['Q2']:,} / Q3 {qc['Q3']:,} / Q4 {qc['Q4']:,}; "
          f"{len(company_list(d))} companies; Repairs custody {cust}; "
          f"legacy pre-{ASAT_DAY.year} {leg['n']:,} of which tooling {leg['tooling_n']})")
    vb = value_bits(d["master"])
    print(f"Replacement value   : {money(vb['value'])} over {vb['priced']:,} priced "
          f"({vb['unpriced']} unpriced)")
    x = onhire_model(d)
    print(f"On-hire report      : {x['n_companies']} companies A to Z; "
          f"{x['people_names']:,} hirers (people); {x['account_items']} items on project "
          f"accounts; Repairs custody {len(x['custody'])}; holding accounts outside the "
          f"count {len(x['holding'])}; oldest hire {fmt_date(x['oldest']['date']) if x['oldest'] else '-'}; "
          f"over 90 days {len(x['over90'])}")
    print(f"Former site name    : still on {x['former_desc_all']:,} register descriptions "
          f"({x['former_desc_here']} on-hire lines) and {x['former_accounts']} account - "
          "shown as Ampol on every page")
    cb = sum(1 for r in d["master"] if r["corrected"] == "barcode")
    cd = sum(1 for r in d["master"] if r["corrected"] == "description")
    print(f"Corrected names     : {cb:,} by barcode + {cd:,} by matching description "
          f"of {len(d['master']):,} on-hire lines (mapping: {len(d['corr']):,} barcodes, "
          f"{len(d['corr_by_desc']):,} unambiguous descriptions)")
    print(f"Available for hire  : {d['available_n']:,} tooling of {d['available_all_n']:,} "
          f"register rows | Repairs custody rows {d['repairs_n']}")
    print(f"Transactions YTD    : {len(d['tx']):,}")
    print(f"Stocktake barcodes  : {len(d['stocktake']):,}")
    print(f"Priced descriptions : {len(d['pricing']):,} distinct of {d['pricing_rows']:,} "
          "rows (first price wins)")
    um = d["util"]
    print(f"Utilisation         : {len(um['rows']):,} groups, {len(um['buy'])} buy signals, "
          f"{len(um['overstock'])} right-size, {um['groups_with_days']} with hire days, "
          f"{um['tx_used']:,} of {um['tx_total']:,} transactions mapped, "
          f"{um['days_elapsed']} days elapsed")
    c = get_changes(d)
    if c is None:
        print("Since the last pull : could not be read - see the note above")
    else:
        l = c["last24"]
        prev = (f"previous pull {stamp(c['prev_time'])}: {len(c['returned'])} came back, "
                f"{len(c['issued'])} went out, {len(c['moved'])} changed hands, "
                f"{len(c['crossed'][90])} crossed 90 days" if c["have_previous"]
                else "no earlier pull in Data\\previous (pull against pull starts next pull)")
        print(f"Since the last pull : {prev}; 24 h before the pull: "
              f"{len(l['issued'])} issued, {len(l['returned'])} returned (tooling)"
              if l["available"] else
              f"Since the last pull : {prev}; TRANSACTIONS not in Data - 24 h not counted")
    wb = d.get("wb")
    if wb:
        tie = ("ties to the live register" if wb["master_ties"]
               else "DOES NOT tie to the live register - stale, not used")
        util_bit = (f"; Utilisation tab implied refresh {fmt_date(wb['util_refresh'])}"
                    if wb.get("util_refresh") else "")
        print(f"Workbook (info only): file saved {stamp(wb['mtime'])}; Master Onhire tab "
              f"{wb['master_rows'] if wb['master_rows'] is not None else '-'} rows - {tie}"
              f"{util_bit}. No page reads the workbook.")
    else:
        print("Workbook (info only): not in Data - not needed; nothing reads it.")


def main():
    print("=" * 68)
    print("COATES | AMPOL TOOL STORE REPORT KIT | The Coates Way")
    print("=" * 68)
    d = load_all()
    console_summary(d)

    args = [a for a in sys.argv[1:]]
    # WHY (02 Sep 2026): run from a scheduler or a pipe (no keyboard) with no
    # arguments, the menu used to die on EOF and exit red - it now builds
    # everything, which is what every unattended run wants.
    if not args and not sys.stdin.isatty():
        args = ["--everything"]
    if args:
        if "--exec" in args:
            run_exec(d)
        if "--quarter" in args:
            # WHY (12 Aug 2026): --quarter as the last argument used to die
            # with a raw IndexError - now it says what it needs, and exits
            # nonzero so the .bat button shows red, not green.
            i = args.index("--quarter")
            if i + 1 >= len(args):
                sys.exit("--quarter needs a value: Q1, Q2, Q3, Q4, YEAR or ALL.")
            qv = args[i + 1].upper()
            for q in ([k for k in QUARTERS] + ["YEAR"] if qv == "ALL" else [qv]):
                if q not in QUARTERS and q != "YEAR":
                    sys.exit(f"Unknown quarter '{q}' - use Q1, Q2, Q3, Q4, "
                             "YEAR or ALL.")
                run_quarter(d, q)
        if "--onhire" in args:
            run_onhire(d)
        if "--utilisation" in args or "--util" in args:
            run_util(d)
        if "--compliance" in args:
            run_compliance(d)
        if "--company" in args:
            i = args.index("--company")
            if i + 1 >= len(args):
                sys.exit("--company needs a company name after it.")
            run_company(d, args[i + 1])
        if "--all" in args:
            # every company report, only when asked for by name - not part
            # of --everything since 02 Sep 2026
            for n in company_list(d):
                run_company(d, n)
        if "--everything" in args:
            run_everything(d)
        record_history(d)
        print(f"\nDone. Output: {OUT_REL}. Run 08_MAKE_OUTLOOK_DRAFTS.bat to load the "
              "emails into Outlook Drafts.")
        return

    companies = company_list(d)
    while True:
        print("\n--- REPORT MENU -------------------------------------------")
        print("  E = Executive Summary          U = Utilisation & What-To-Buy")
        print("  T = Compliance & Trends        Y = Year on-hire report")
        print("  1 = Q1 (Jan-Mar)   2 = Q2 (Apr-Jun)   3 = Q3 (Jul-Sep)   4 = Q4 (Oct-Dec)")
        print("  O = Tooling On-Hire Report     X = everything (no per-company reports)")
        print("  Or pick ONE company for a one-off company report (off by default -")
        print("  the Tooling On-Hire Report covers every company):")
        for i, n in enumerate(companies, 1):
            print(f"   C{i:<3} {n}")
        print("  Q = quit")
        choice = input("> ").strip().upper()
        if choice == "Q":
            break
        elif choice == "E":
            run_exec(d)
        elif choice == "U":
            run_util(d)
        elif choice == "T":
            run_compliance(d)
        elif choice == "Y":
            run_quarter(d, "YEAR")
        elif choice in ("1", "2", "3", "4"):
            run_quarter(d, "Q" + choice)
        elif choice == "O":
            run_onhire(d)
        elif choice == "X":
            run_everything(d)
        elif choice.startswith("C") and choice[1:].isdigit() \
                and 1 <= int(choice[1:]) <= len(companies):
            run_company(d, companies[int(choice[1:]) - 1])
        else:
            print("  (not recognised)")
            choice = ""
        if choice:
            record_history(d)
        print(f"\nOutputs in {OUT_REL}. 08_MAKE_OUTLOOK_DRAFTS.bat loads the")
        print("emails into Outlook Drafts - full To search, attachments included.")


if __name__ == "__main__":
    # WHY (12 Aug 2026): the old handler swallowed every error (including
    # SystemExit) and finished with exit code 0, so a failed run looked green
    # to the .bat buttons. Failures now exit nonzero with a plain-English
    # message; the button owns the pause.
    try:
        main()
    except Exception as e:
        print("\nERROR - the tooling report run failed: " + str(e))
        sys.exit(1)
