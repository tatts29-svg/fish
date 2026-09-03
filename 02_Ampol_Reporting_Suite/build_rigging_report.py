#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=====================================================================
COATES RIGGING & LIFTING REGISTER - HOUSE-STYLE REPORT (the K2 look)
One run -> client-ready PDF + Outlook email, off the rigging register
workbook joined barcode-for-barcode to the live SiteIQ export.
=====================================================================
Author: Andrew Fisher
The Coates Way - Operational Excellence - POWERED BY SITEIQ

WHAT THIS IS
  The rigging and height-safety register gets a report of its own -
  the K2 house style, same shell as the stocktake family. One run
  produces:

  (file names come from ampol_names.report_stem - one rule for the
  whole suite, the day the button was pressed on the end, e.g. 03Sep2026)

   1. Coates_Ampol_Rigging_Register_<day>.pdf   (THE report - the live
      position, where the gear is by company and hirer, what is not
      found in SiteIQ, the test-status truth, and the register's own
      identity gaps)
      + Coates_Ampol_Rigging_Register_<day>.html (the same pages, kept
      beside the PDF so a machine with no PDF engine still has the
      report; VERIFY_NUMBERS reads it)
   2. Coates_Ampol_Rigging_Register_<day>_OUTLOOK.eml (DRAFT - never
      sends; the position card shows inline under the header and rides
      as a file beside the PDF)
      + _OUTLOOK.body.html and _OUTLOOK.draft.json so
      MAKE_OUTLOOK_DRAFTS keeps working (the card is an attachment there).
   3. Coates_Ampol_Rigging_Register_<day>_PositionCard.png - the phone
      card: the position-page tiles, the RAG line and four scores.
  The PDF is finished with its document properties (Author: Andrew
  Fisher) and a bookmark per section (pdf_finish). Once seven days are
  on the History scoreboard it gains a trend page before the close.
  The PDF opens on a dark cover (the one number of the day), the position
  page carries the movement since the previous recorded pull and a RAG
  band with an owner and a dated next action, and the closing page shows
  the Coates Way. Each run writes its figures to History\report_history
  .json keyed on the pull day; the next run reads them back for the
  movement notes. No earlier day on record means no arrow - never a guess.

WHERE THE NUMBERS COME FROM (changed 02 Sep 2026)
  Two files in the suite's Data area:
   - Rigging Register.xlsx : WHICH barcodes are on the register.
     'Rigging register Master' gives category, description, serial and
     the eight test columns; 'Extracted Register' is the certificate
     extract disclosed on the test page. The 'To Help Locate' sheet is
     a static join to a June SiteIQ pull and is read for one thing
     only - the last-known status of barcodes SiteIQ no longer returns.
   - RENTAL_STOCK.xlsx     : WHERE every barcode is right now - status,
     company, hirer, on-hire date and storage unit - as at the export's
     own request time (REFERENCE_INFO). Every "on hire", every "in the
     store" and every "held N days" on these pages comes from here.
  An audit on 02 Sep 2026 found the old build stamping the June
  snapshot as today's position: 210 of the barcodes had changed status
  since, and every "held" count was measured to the wrong date. The
  join to the live export fixes that; the audit's other findings
  (barcode duplicates, the certificate extract, repairs printed as a
  customer, rigging gear not on the register) each get their own
  honest line or page below.

DATA RULES
  Real data only. Nothing is invented - a blank test cell prints as a
  dash with the fill-in noted. A register barcode the live export does
  not return is printed as "not found in SiteIQ - whereabouts unknown"
  and is never counted as accounted for. Repairs, off-site, out-of-tag
  and out-of-service custody lines are kept apart from customer hire.
  Barcodes are upper-cased and de-duplicated for the join, and the
  page says so.
=====================================================================
"""

import html as _html
import json
import math
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from email.message import EmailMessage
from email.utils import formatdate
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import ampol_paths
import build_stocktake_compliance_tool as eng   # write_pdf_robust, find_workbook, parse_dt
import gasmon_engine as ge                       # norm_company, parse_stamp
import ampol_names                               # how names are SHOWN


def hl(name):
    """A hirer as printed: a shared booking account under the suite-wide
    label (ampol_names.hirer_label), a person as SiteIQ spells them."""
    return esc(ampol_names.hirer_label(name))
import k2shell as sh
import pdf_finish
import report_history as rh                      # the movement scoreboard - recorded days only
import txn_insights as ti                        # the transaction log, read once
from k2shell import esc, num, K

import openpyxl

BASE = Path(__file__).resolve().parent
# Outputs land in the suite's dated Reports area - Reports\<today>\
# Rigging - one folder per day, nothing ever silently overwritten.
OUT = Path(ampol_paths.day_folder("Rigging"))

CONFIG = {
    "client": "Ampol",
    # WHY (02 Sep 2026): data-as-at is the SiteIQ export's own request
    # time (REFERENCE_INFO), not the register file's saved time - the
    # register tells us WHAT is on it, SiteIQ tells us WHERE it is now.
    "asat_note": "(SiteIQ register pull)",
    "title": "Rigging & Lifting Register",
    "kicker": "COATES · TOOL STORE - RIGGING & LIFTING REGISTER REPORT",
    "project": "Ampol Lytton Refinery · Permanent Tool Store",
    # WHY (03 Sep 2026): the one output name comes from ampol_names.report_stem
    # - the client reads the attachment name before a single figure, so the
    # whole suite shares one shape with the day on the end. Everything else
    # (.pdf, .html, _OUTLOOK.eml, _PositionCard.png, the manifest) hangs off it.
    "stem": ampol_names.report_stem("rigging"),
    # ---- the RAG line on the position page (default line - set here) ----
    # Share of the distinct register barcodes the SiteIQ pull does not
    # return: Green only at 0%, Amber below rag_red_pct, Red from it. The
    # band prints the rule beside the result.
    "rag_owner": "Andrew Fisher, Shutdown Manager",
    "rag_red_pct": 10,
    # the next action falls due this many days after the SiteIQ pull date -
    # anchored on the pull, never on the day the report happens to be run
    "action_days": 7,
    "team": [
        {"name": "Andrew Fisher", "role": "Shutdown Manager",
         "shift": "", "email": "andrew.fisher@coates.com.au",
         "blurb": "Owns the register and the fill-in push - anything at all, start here",
         "lead": True},
    ],
    "key_items": [
        ("orange", "LIVE POSITION", "status, holder and days held from the SiteIQ pull, not the workbook snapshot"),
        ("blue", "NOT FOUND", "register barcodes SiteIQ does not return - whereabouts unknown, never assumed"),
        ("amber", "LONGEST HELD FIRST", "customer hire chased oldest first; repairs and quarantine kept separate"),
    ],
}

STEM = CONFIG["stem"]
CONFIG.update({
    "pdf_name": f"{STEM}.pdf", "pdf_html": f"{STEM}.html",
    "eml_name": f"{STEM}_OUTLOOK.eml", "email_html": f"{STEM}_OUTLOOK.body.html",
    "draft_json": f"{STEM}_OUTLOOK.draft.json", "card_name": f"{STEM}_PositionCard.png",
})
# the scoreboard's path as the pages print it (a backslash cannot sit inside
# an f-string expression on the store laptops' Python)
HIST_NAME = "History\\report_history.json"

# Master test columns - a row with ANY of these filled counts as having
# test details started. Used for the honest fill-in progress numbers.
TEST_COLS = ["Last Test Date", "Next Test Due", "Test Status",
             "Test Tag Colour", "Tested By", "Tester Licence No",
             "Certificate No", "Inspection Comments"]

# Custody lines that are NOT a customer holding the gear: SiteIQ's
# "Repairs" company and its parking hirers (offsite repairs, off site,
# out of tag date, out of service). They print on their own line.
REPAIR_RE = re.compile(r"repair|off\s*-?\s*site|out\s*of\s*tag|out\s*of\s*service", re.I)
REPAIR_RULE = ("SiteIQ COMPANY_NAME 'Repairs', or a hirer name containing "
               "Repairs, Off Site, Out of Tag or Out of Service")

# "Possibly rigging gear not on the register" - the keyword rule, printed
# on the page exactly as applied. A keyword match is a lead to verify,
# never a finding.
KW_INCLUDE = re.compile(
    r"SLING|SHACKLE|CHAIN\s*BLOCK|LEVER\s*(?:BLOCK|HOIST)|CUM-?A-?LONG|COME-?A-?LONG|"
    r"TIRFOR|BEAM\s*CLAMP|GIRDER\s*CLAMP|PLATE\s*CLAMP|EYE\s*BOLT|SPREADER\s*BAR|"
    r"SWIVEL|LIFTING\s*BAG|BUCKET\s*LIFTER|WINCH|HARNESS|LANYARD|INERTIA\s*REEL|"
    r"ROPE\s*GRAB|ANCHOR|GOTCHA|RESCUE|DAVIT|TRIPOD", re.I)
KW_EXCLUDE = re.compile(r"FLANGE\s*SPREADER|HOOK\s*WRENCH|DUMPY|CUTTING\s*TROLLEY", re.I)
KW_RULE = ("ITEM_DESCRIPTION contains sling, shackle, chain block, lever block or "
           "hoist, cumalong, tirfor, beam / girder / plate clamp, eyebolt, spreader "
           "bar, swivel, lifting bag, bucket lifter, winch, harness, lanyard, "
           "inertia reel, rope grab, anchor, Gotcha, rescue, davit or tripod - "
           "excluding flange spreaders, hook wrenches, dumpy-level tripods and "
           "cutting trolleys")


# =====================================================================
# load - the register workbook (membership, test columns, extract,
# last-known snapshot) and the live SiteIQ export (where it is now)
# =====================================================================

def _sheet_rows(wb, name, need):
    if name not in wb.sheetnames:
        sys.exit(f"ERROR: the rigging workbook has no '{name}' sheet - "
                 "is this the right file in Data\\?")
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h or "").strip() for h in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    missing = [c for c in need if c not in ix]
    if missing:
        sys.exit(f"ERROR: '{name}' is missing columns {missing} - the "
                 "register layout has changed; report needs updating.")
    rows = [r for r in it if any(c not in (None, "") for c in r)]
    return rows, ix


def _s(v):
    return "" if v is None else str(v).strip()


def load_register(path):
    """Rigging Register.xlsx - three things, each labelled for what it is.

    rows    : every Master row verbatim (category, serial, barcode as typed
              and upper-cased, description, the eight test columns)
    snap    : barcode -> what the 'To Help Locate' sheet said (a static
              June SiteIQ join) - used ONLY for barcodes the live export
              no longer returns, and labelled as the snapshot on the page
    extract : the 'Extracted Register' certificate rows, or None
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    raw, ix = _sheet_rows(wb, "Rigging register Master",
                          ["REGISTER_CATEGORY", "ITEM_BARCODE",
                           "REGISTER_DESCRIPTION"] + TEST_COLS)
    rows = []
    for r in raw:
        bc_raw = _s(r[ix["ITEM_BARCODE"]])
        tests = {c: _s(r[ix[c]]) for c in TEST_COLS}
        rows.append({
            "cat": _s(r[ix["REGISTER_CATEGORY"]]),
            "serial": _s(r[ix["Serial Number"]]) if "Serial Number" in ix else "",
            "bc_raw": bc_raw,
            "barcode": bc_raw.upper(),
            "desc": ampol_names.display_desc(_s(r[ix["REGISTER_DESCRIPTION"]])),
            "former_name": ampol_names.carries_former_name(_s(r[ix["REGISTER_DESCRIPTION"]])),
            "tests": tests,
            "has_test": any(tests.values()),
        })

    snap, snap_max = {}, None
    if "To Help Locate" in wb.sheetnames:
        raw2, ix2 = _sheet_rows(wb, "To Help Locate",
                                ["ITEM_BARCODE", "ITEM_STATUS", "COMPANY_NAME",
                                 "HIRER_NAME", "ON_HIRE_DATE"])
        for r in raw2:
            bc = _s(r[ix2["ITEM_BARCODE"]]).upper()
            ohd = r[ix2["ON_HIRE_DATE"]]
            ohd = ohd if isinstance(ohd, datetime) else None
            if ohd and (snap_max is None or ohd > snap_max):
                snap_max = ohd
            if bc and bc not in snap:
                snap[bc] = {"status": _s(r[ix2["ITEM_STATUS"]]),
                            "company": _s(r[ix2["COMPANY_NAME"]]),
                            "hirer": _s(r[ix2["HIRER_NAME"]]),
                            "since": ohd}

    extract = None
    if "Extracted Register" in wb.sheetnames:
        ws = wb["Extracted Register"]
        hdr_ix, extract = None, []
        for r in ws.iter_rows(values_only=True):
            cells = [_s(c) for c in r]
            if hdr_ix is None:
                # the sheet carries a throwaway 'Column1..' line above the
                # real header - find the header by its own column names
                if any("next insp" in c.lower() for c in cells):
                    hdr_ix = {c.lower(): i for i, c in enumerate(cells)}
                continue
            if not any(cells):
                continue

            def col(name):
                for k, i in hdr_ix.items():
                    if name in k:
                        return cells[i] if i < len(cells) else ""
                return ""
            extract.append({"serial": col("serial"), "barcode": col("barcode").upper(),
                            "desc": col("description"), "wll": col("wll"),
                            "next_due": col("next insp"), "parse": col("parse")})
        if hdr_ix is None:
            extract = None
    wb.close()
    return rows, snap, snap_max, extract


def load_live(path):
    """RENTAL_STOCK.xlsx - every line of the live SiteIQ export plus the
    export's own request time from REFERENCE_INFO."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    asat, asat_note = None, "(SiteIQ register pull)"
    if "REFERENCE_INFO" in wb.sheetnames:
        ref = list(wb["REFERENCE_INFO"].iter_rows(values_only=True))
        if len(ref) > 1:
            for i, h in enumerate(ref[0]):
                if h and "REQUESTED_DATE" in str(h).upper():
                    asat = ge.parse_stamp(ref[1][i])
    if asat is None:
        asat = datetime.fromtimestamp(os.path.getmtime(path))
        asat_note = "(RENTAL_STOCK file saved - no request stamp found)"
    if "RENTAL_STOCK" not in wb.sheetnames:
        sys.exit("ERROR: RENTAL_STOCK.xlsx has no RENTAL_STOCK sheet - is "
                 "this the SiteIQ rental stock export?")
    ws = wb["RENTAL_STOCK"]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h or "").strip().upper() for h in next(it)]
    col = {h: i for i, h in enumerate(hdr)}
    for need in ("ITEM_BARCODE", "ITEM_DESCRIPTION", "ITEM_STATUS",
                 "COMPANY_NAME", "HIRER_NAME", "ON_HIRE_DATE"):
        if need not in col:
            sys.exit(f"ERROR: RENTAL_STOCK is missing column {need} - the "
                     "export layout has changed; report needs updating.")

    def g(r, name):
        i = col.get(name)
        return r[i] if i is not None and i < len(r) else None

    rows = []
    for r in it:
        if not r or all(c in (None, "") for c in r):
            continue
        bc = _s(g(r, "ITEM_BARCODE"))
        since = eng.parse_dt(g(r, "ON_HIRE_DATE"))
        rows.append({
            "bc_raw": bc, "barcode": bc.upper(),
            "desc": ampol_names.display_desc(_s(g(r, "ITEM_DESCRIPTION"))),
            "former_name": ampol_names.carries_former_name(_s(g(r, "ITEM_DESCRIPTION"))),
            "status": _s(g(r, "ITEM_STATUS")),
            "company_raw": _s(g(r, "COMPANY_NAME")),
            "hirer": _s(g(r, "HIRER_NAME")),
            "since": since,
            "unit": _s(g(r, "STORAGE_UNIT")),
        })
    wb.close()
    return rows, asat, asat_note


# =====================================================================
# rules
# =====================================================================

def company_name(raw):
    """One customer, one name: Ampol / Ampol Refineries (Qld) / Caltex are
    the client; project accounts (FCCU, SATGAS/MOL) fold into the
    company. The raw SiteIQ account is kept alongside for the page."""
    u = str(raw or "").strip().upper()
    if u.startswith("CALTEX"):
        return "Ampol"
    return ge.norm_company(raw)


def is_repair(lr):
    return (lr["company_raw"].strip().lower() == "repairs"
            or bool(REPAIR_RE.search(lr["hirer"])))


def norm_desc(s):
    return re.sub(r"\s+", " ", str(s or "").replace("\xa0", " ")).strip()


# =====================================================================
# derive - the positions the pages print
# =====================================================================

def derive(reg_rows, snap, snap_max, extract, live_rows, asat_dt):
    d = {"asat": asat_dt, "snap_max": snap_max}

    # ---- register identity (row basis) ----------------------------------
    d["rows"] = len(reg_rows)
    d["former_reg"] = sum(1 for r in reg_rows if r.get("former_name"))
    d["former_live"] = sum(1 for r in live_rows if r.get("former_name"))
    d["blank"] = [r for r in reg_rows if not r["barcode"]]
    d["lower"] = sum(1 for r in reg_rows if r["bc_raw"] and r["bc_raw"] != r["barcode"])
    cnt = Counter(r["barcode"] for r in reg_rows if r["barcode"])
    d["dups"] = sorted([(b, n) for b, n in cnt.items() if n > 1],
                       key=lambda x: (-x[1], x[0]))
    d["dup_rows"] = sum(n for _, n in d["dups"])
    d["cats_rows"] = Counter(r["cat"] or "(no category)" for r in reg_rows)
    serials = [r["serial"] for r in reg_rows if r["serial"]]
    d["serial_rows"] = len(serials)
    d["serial_distinct"] = len(set(serials))
    d["with_test"] = [r for r in reg_rows if r["has_test"]]
    d["col_fill"] = [(c, sum(1 for r in reg_rows if r["tests"][c])) for c in TEST_COLS]
    d["test_status"] = Counter(r["tests"]["Test Status"] for r in reg_rows
                               if r["tests"]["Test Status"])

    # ---- one item per distinct barcode (first Master row wins) ----------
    items = {}
    for r in reg_rows:
        if r["barcode"] and r["barcode"] not in items:
            it = dict(r)
            it["rows"] = cnt[r["barcode"]]
            items[r["barcode"]] = it
    d["distinct"] = len(items)

    # ---- join to the live export on ITEM_BARCODE ------------------------
    # exact first, then upper-cased/trimmed (the register carries 22
    # lower-case barcodes; SiteIQ's are upper-case)
    live_exact = {}
    live_upper = {}
    for lr in live_rows:
        if lr["bc_raw"] and lr["bc_raw"] not in live_exact:
            live_exact[lr["bc_raw"]] = lr
        if lr["barcode"] and lr["barcode"] not in live_upper:
            live_upper[lr["barcode"]] = lr
    d["join_exact"] = 0
    for bc, it in items.items():
        lr = live_exact.get(it["bc_raw"])
        if lr is not None:
            d["join_exact"] += 1
        else:
            lr = live_upper.get(bc)
        it["live"] = lr
        it["snap"] = snap.get(bc)
        it["held"] = None
        it["company"] = ""
        it["account"] = ""
        it["hirer"] = ""
        it["since"] = None
        it["unit"] = ""
        if lr is None:
            it["bucket"] = "missing"
            continue
        it["account"] = lr["company_raw"]
        it["hirer"] = lr["hirer"]
        it["since"] = lr["since"]
        it["unit"] = lr["unit"]
        st = lr["status"].lower()
        if st == "on hire":
            it["held"] = ((asat_dt.date() - lr["since"].date()).days
                          if lr["since"] else None)
            if is_repair(lr):
                it["bucket"] = "repair"
            else:
                it["bucket"] = "onhire"
                it["company"] = company_name(lr["company_raw"])
        elif st == "available for hire":
            it["bucket"] = "store"
        else:
            it["bucket"] = "other"
    d["items"] = items
    d["found"] = [it for it in items.values() if it["bucket"] != "missing"]
    d["onhire"] = [it for it in items.values() if it["bucket"] == "onhire"]
    d["repair"] = [it for it in items.values() if it["bucket"] == "repair"]
    d["store"] = [it for it in items.values() if it["bucket"] == "store"]
    d["other"] = [it for it in items.values() if it["bucket"] == "other"]
    d["missing"] = [it for it in items.values() if it["bucket"] == "missing"]
    d["other_status"] = Counter(it["live"]["status"] for it in d["other"])
    d["found_pct"] = round(len(d["found"]) / d["distinct"] * 100) if d["distinct"] else 0

    # ---- where the customer hire is: company (merged) -> hirers ---------
    co = defaultdict(list)
    for it in d["onhire"]:
        co[it["company"] or "(no company recorded)"].append(it)
    comp = []
    for c, lst in co.items():
        lst.sort(key=lambda it: (it["since"] is None, it["since"] or datetime.max))
        comp.append({"co": c, "n": len(lst),
                     "accounts": sorted({it["account"] for it in lst}),
                     "hirers": len({it["hirer"] for it in lst}),
                     "oldest": lst[0], "items": lst})
    # WHY (02 Sep 2026): companies A-Z - the counter finds a name by eye; the
    # ranked view of the same data is the longest-held page.
    comp.sort(key=lambda x: ampol_names.sort_key(x["co"]))
    d["companies"] = comp
    d["oh_longest"] = sorted([it for it in d["onhire"] if it["since"]],
                             key=lambda it: it["since"])
    d["oh_nodate"] = [it for it in d["onhire"] if not it["since"]]
    d["oh_hirers"] = len({it["hirer"] for it in d["onhire"]})
    d["oh_after_snap"] = (sum(1 for it in d["onhire"] if it["since"] and snap_max
                              and it["since"] > snap_max) if snap_max else 0)

    # ---- repairs / quarantine custody lines -----------------------------
    rp = defaultdict(list)
    for it in d["repair"]:
        rp[it["hirer"] or "(no hirer recorded)"].append(it)
    rep = []
    for h, lst in rp.items():
        lst.sort(key=lambda it: (it["since"] is None, it["since"] or datetime.max))
        rep.append({"hirer": h, "n": len(lst),
                    "accounts": sorted({it["account"] for it in lst}),
                    "oldest": lst[0], "items": lst})
    rep.sort(key=lambda x: ampol_names.sort_key(x["hirer"]))
    d["repair_lines"] = rep
    d["out_of_tag"] = [it for it in d["repair"]
                       if re.search(r"out\s*of\s*tag", it["hirer"], re.I)]

    # ---- not found in SiteIQ - last-known from the workbook snapshot ----
    def snap_key(it):
        s = it["snap"] or {}
        oh = 0 if s.get("status", "").lower() == "on hire" else 1
        return (oh, s.get("hirer", ""), it["desc"], it["barcode"])
    d["missing"].sort(key=snap_key)
    d["missing_snap"] = Counter(
        (it["snap"] or {}).get("status", "") or "(no snapshot row)" for it in d["missing"])
    d["missing_cats"] = Counter(it["cat"] or "(no category)" for it in d["missing"])

    # ---- the certificate extract ----------------------------------------
    ex = {"present": extract is not None}
    if extract is not None:
        ex["rows"] = len(extract)
        exb = {}
        for r in extract:
            if r["barcode"] and r["barcode"] not in exb:
                exb[r["barcode"]] = r
        ex["distinct"] = len(exb)
        onreg = [b for b in exb if b in items]
        ex["on_register"] = len(onreg)
        ex["reg_buckets"] = Counter(items[b]["bucket"] for b in onreg)
        ex["next_due"] = Counter(r["next_due"] or "(blank)" for r in extract)
        ex["parse"] = Counter(r["parse"] or "(blank)" for r in extract)
        ex["with_due"] = sum(1 for r in extract if r["next_due"])
    d["extract"] = ex

    # ---- possibly rigging gear not on the register (keyword rule) ------
    kw = [lr for lr in live_rows
          if lr["barcode"] and lr["barcode"] not in items
          and KW_INCLUDE.search(lr["desc"]) and not KW_EXCLUDE.search(lr["desc"])]
    grp = defaultdict(list)
    for lr in kw:
        grp[norm_desc(lr["desc"])].append(lr)
    kwg = []
    for desc, lst in grp.items():
        st = Counter(lr["status"].lower() for lr in lst)
        kwg.append({"desc": desc, "n": len(lst),
                    "onhire": st.get("on hire", 0),
                    "store": st.get("available for hire", 0),
                    "other": len(lst) - st.get("on hire", 0) - st.get("available for hire", 0),
                    "eg": sorted(lr["barcode"] for lr in lst)[0]})
    kwg.sort(key=lambda x: (-x["n"], x["desc"].upper()))
    d["kw"] = kw
    d["kw_groups"] = kwg
    d["kw_onhire"] = sum(g["onhire"] for g in kwg)

    # ---- category detail ------------------------------------------------
    cats = []
    for c, n_rows in d["cats_rows"].most_common():
        grp_items = [it for it in items.values() if (it["cat"] or "(no category)") == c]
        b = Counter(it["bucket"] for it in grp_items)
        oh_dated = sorted([it for it in grp_items if it["bucket"] == "onhire" and it["since"]],
                          key=lambda it: it["since"])
        cats.append({"cat": c, "rows": n_rows, "distinct": len(grp_items),
                     "onhire": b.get("onhire", 0), "repair": b.get("repair", 0),
                     "store": b.get("store", 0), "other": b.get("other", 0),
                     "missing": b.get("missing", 0),
                     "oldest": oh_dated[0] if oh_dated else None})
    d["cat_detail"] = cats
    return d


# =====================================================================
# PDF rendering via the engine's robust writer + a real layout check
# =====================================================================

def layout_check(pdf_path, authored):
    """PASS only when the page count matches AND nothing on any page has
    run down into the footer. The K2 page is a fixed-height box with
    overflow hidden, so a page count alone cannot see a table that has
    grown past the footer - PyMuPDF (where installed) reads every page
    and looks for body text or drawings overlapping the footer. The cover
    is page 1 and carries no footer by design, so it is skipped."""
    problems = []
    raw = open(pdf_path, "rb").read()
    counts = re.findall(rb"/Count\s+(\d+)", raw)
    actual = max(int(c) for c in counts) if counts else -1
    if actual == -1:
        problems.append(f"page count unreadable (authored {authored})")
    elif actual != authored:
        problems.append(f"authored {authored} pages, PDF has {actual}")
    try:
        import pymupdf
    except ImportError:
        pymupdf = None
    if pymupdf is not None:
        doc = pymupdf.open(str(pdf_path))
        for pno, page in enumerate(doc, start=1):
            if pno <= COVER_PAGES:
                continue           # the dark cover has no footer by design
            h = page.rect.height
            blocks = page.get_text("blocks")
            # letter-spaced footer text extracts as "Y O U R  C O A T E S ..." -
            # compare with the whitespace stripped
            foot = [b for b in blocks
                    if "COATESTOOLSTORETEAM" in re.sub(r"\s+", "", str(b[4])).upper()]
            if not foot:
                problems.append(f"page {pno}: footer not found")
                continue
            foot_top = min(b[1] for b in foot) - 7   # the rule above the footer text
            for b in blocks:
                if b[1] >= foot_top - 1:
                    continue           # the footer's own lines
                if b[3] > foot_top:
                    txt = re.sub(r"\s+", " ", str(b[4]))[:40]
                    problems.append(f"page {pno}: text runs into the footer ({txt!r})")
                    break
            for dr in page.get_drawings():
                r = dr.get("rect")
                if r is None or r.height > 0.8 * h or r.height < 3:
                    continue           # the orange frame / hairlines
                if r.y0 < foot_top - 2 and r.y1 > foot_top + 2:
                    problems.append(f"page {pno}: a chart or panel runs into the footer")
                    break
        doc.close()
        checked = f"{actual} pages, footer clearance checked on every page after the cover"
    else:
        checked = f"{actual} pages (page count only - PyMuPDF not installed)"
    if problems:
        print("*" * 68)
        print(f"Layout check         : FAIL - {Path(pdf_path).name}")
        for p in problems:
            print(f"   - {p}")
        print("   Do not send as is.")
        print("*" * 68)
        return False
    print(f"Layout check         : PASS - {checked} ({Path(pdf_path).name})")
    return True


_MEASURE_JS = r"""<script>
document.fonts.ready.then(function(){
  var out = [];
  var pages = document.querySelectorAll('.page');
  for (var i = 0; i < pages.length; i++) {
    var pg = pages[i];
    var body = pg.querySelector('.body');
    var foot = pg.querySelector('.foot');
    if (!body || !foot) { out.push({page: i + 1, over: null, wide: 0}); continue; }   // the cover: no footer by design
    var ft = foot.getBoundingClientRect();
    var bottom = body.getBoundingClientRect().top;
    var els = body.querySelectorAll('*');
    for (var j = 0; j < els.length; j++) {
      var r = els[j].getBoundingClientRect();
      if (r.bottom > bottom) bottom = r.bottom;
    }
    out.push({page: i + 1, over: Math.round(bottom - ft.top),
              wide: Math.round(body.scrollWidth - body.clientWidth)});
  }
  var d = document.createElement('div');
  d.id = 'layout-report';
  d.setAttribute('data-lato', document.fonts.check('12px Lato') ? '1' : '0');
  d.textContent = JSON.stringify(out);
  document.body.appendChild(d);
});
</script>"""


def _browser():
    try:
        from generate_k2style_gas_monitor_report import find_browser
        return find_browser()
    except Exception:
        import shutil
        for name in ("msedge", "chrome", "chromium", "chromium-browser", "google-chrome"):
            if shutil.which(name):
                return shutil.which(name)
        return None


def fit_check(doc, css, label):
    """Measure every page in the browser with the house face loaded.
    WHY (03 Sep 2026): k2style.css now embeds Lato, and the shared fit
    check measured before the font files had loaded - it saw the fallback
    face, up to 9 px off on a full page, on pages that had 0 px to spare.
    This one waits on document.fonts.ready under a virtual-time budget,
    with the page written beside the PDF so the relative font URLs
    resolve, so the pixels it reports are the pixels the PDF prints.
    Returns (ok, worst_spare_px, rows); ok is None when nothing could
    measure - never a reason not to build."""
    import subprocess
    import tempfile
    browser = _browser()
    if not browser:
        print(f"Fit check            : skipped - no browser to measure with ({label})")
        return None, None, []
    doc2 = (doc.replace("</head>", f"<style>{css}</style></head>", 1)
               .replace("</body>", _MEASURE_JS + "</body>", 1))
    tmp = OUT / f"__measure_{os.getpid()}__.html"
    profile = os.path.join(tempfile.gettempdir(), f"coates_fit_{os.getpid()}")
    try:
        tmp.write_text(doc2, encoding="utf-8")
        res = subprocess.run([browser, "--headless", "--disable-gpu", "--no-sandbox",
                              "--no-first-run", "--user-data-dir=" + profile,
                              "--virtual-time-budget=10000", "--dump-dom", tmp.as_uri()],
                             capture_output=True, timeout=240)
        dom = (res.stdout or b"").decode("utf-8", "ignore")
    except Exception as e:
        print(f"Fit check            : skipped ({type(e).__name__}) ({label})")
        return None, None, []
    finally:
        try:
            tmp.unlink()
        except OSError:
            pass
    m = re.search(r'id="layout-report" data-lato="(\d)">(\[.*?\])</div>', dom, re.S)
    if not m:
        print(f"Fit check            : skipped - the page could not be measured ({label})")
        return None, None, []
    rows = [(r["page"], r["over"], r["wide"]) for r in json.loads(m.group(2))]
    face = "Lato" if m.group(1) == "1" else "FALLBACK FACE - Lato did not load"
    rows = [r for r in rows if r[1] is not None]      # the cover carries no footer
    bad = [r for r in rows if r[1] > 0 or r[2] > 0]
    worst = max((r[1] for r in rows), default=-9999)
    if bad:
        print("*" * 68)
        print(f"WARNING: CONTENT DOES NOT FIT in {label} - do not send as is ({face}).")
        for pg, over, wide in bad:
            print(f"  page {pg:2d}: {over:+d}px past the footer" + (f", {wide}px too wide" if wide > 0 else ""))
        print("*" * 68)
        return False, -worst, rows
    print(f"Fit check            : PASS - tightest page has {-worst}px to spare, measured in {face} ({label})")
    return True, -worst, rows


def render_k2_pdf(doc, pdf_path, authored, css):
    """Measures every page with the house face loaded (fit_check), writes
    the HTML beside the PDF (kept, not a temp file - VERIFY_NUMBERS reads
    it) and renders the PDF through the engine's writer. Returns
    (pdf_ok, layout_ok) - layout_ok covers the fit check, the page count
    and the footer clearance. No PDF engine is reported, not fatal - the
    email step still runs."""
    fit_ok, _, _ = fit_check(doc, css, Path(pdf_path).name)
    doc = doc.replace("</head>", f"<style>{css}</style></head>", 1)
    html_path = OUT / CONFIG["pdf_html"]
    html_path.write_text(doc, encoding="utf-8")
    print(f"HTML written         : {html_path}")
    # pre-delete: write_pdf_robust treats an EXISTING file as success, so a
    # stale copy from a previous run must never be able to masquerade
    try:
        Path(pdf_path).unlink()
    except OSError:
        pass
    ok = eng.write_pdf_robust(str(html_path), str(pdf_path))
    if not ok or not Path(pdf_path).exists():
        print("*" * 68)
        print(f"WARNING: could not render {Path(pdf_path).name} - no PDF engine "
              "available on this machine (Edge is standard on Coates laptops).")
        print(f"         The report pages are in {html_path.name}; the email "
              "draft goes out without the PDF attached.")
        print("*" * 68)
        return False, False
    return True, layout_check(pdf_path, authored) and fit_ok is not False


# =====================================================================
# shared page fragments (PDF)
# =====================================================================

def psect(title):
    return f'<div class="sect"><h3>{title}</h3></div>'


def pcallout(inner, tight=True):
    cls = "callout tight" if tight else "callout"
    return f'<div class="{cls}">{inner}</div>'


def pnote(inner):
    return f'<div class="note">{inner}</div>'


def psubh(title, thin=""):
    t = f' <span class="thin">{thin}</span>' if thin else ""
    return f'<div class="sub-h">{title}{t}</div>'


def chartpanel(inner):
    return f'<div class="chartpanel">{inner}</div>'


def dash():
    return '<span class="tbc">&ndash;</span>'


def held_cell(it):
    if it.get("held") is None:
        return dash()
    cls = "rd" if it["held"] > 90 else "a" if it["held"] > 30 else "g"
    return f'<span class="{cls}">{num(it["held"])}d</span>'


def since_cell(it):
    return it["since"].strftime("%d %b %Y") if it.get("since") else dash()


def fmt_dt(dt):
    return dt.strftime("%d %b %Y") if dt else "&ndash;"


# the dark cover is page 1 of the pack, so the position page is page 2 and
# every cross-reference counts from there
COVER_PAGES = 1
# tile note colours as the phone card draws them (the page uses CSS classes)
NOTE_HEX = {"green": "#1FA75A", "amber": "#F5A623", "red": "#EF4444", "grey": "#8A9AAC"}


def plain(fragment):
    """The words of a page fragment with the markup taken off - for the
    phone card, which draws text, not HTML."""
    return _html.unescape(re.sub(r"<[^>]+>", "", fragment)).replace("\xa0", " ")


def moved(key, asat_dt, value, good, note, ncls):
    """(note, class) for a tile: the movement since the previous recorded
    pull when there is one (report_history), else the tile's own note."""
    mv, mcls = rh.movement("rigging", key, asat_dt, value, good=good)
    return (mv, mcls) if mv else (note, ncls)


def spark_of(key, asat_dt, value):
    """30-day sparkline values - every earlier recorded day, then today.
    None until an earlier day exists: no history means no line."""
    past = [v for dd, v in rh.series("rigging", key, asat_dt, 30) if dd < asat_dt.date()]
    return past + [value] if past else None


def history_days(family, asat):
    """Days the scoreboard holds for a family, today's counted in - the
    entry is written after the pages, keyed on the pull day."""
    return len(set(rh.load().get(family, {})) | {asat.date().isoformat()})


def trend_rows(family, asat, today):
    """(labels, {key: [values]}, days) over the 30-day window ending at the
    pull day: every recorded earlier day plus today's own figures. A day
    with no run for a key leaves None - the chart draws a gap."""
    window = {}
    for key in today:
        for dd, v in rh.series(family, key, asat, days=30):
            if dd < asat.date():
                window.setdefault(dd, {})[key] = v
    window.setdefault(asat.date(), {}).update(today)
    days = sorted(window)
    labels = [dd.strftime("%d %b") for dd in days]
    return labels, {k: [window[dd].get(k) for dd in days] for k in today}, days


def trend_page(d):
    """The fixed trend page: not found in SiteIQ and on hire to customers
    over the last 30 days, from the History scoreboard. Empty until seven
    days are on record - the closing page says so - and every point is a
    figure a report printed on that day."""
    asat_dt = d["asat"]
    if history_days("rigging", asat_dt) < 7:
        return ""
    today = {"not_found": len(d["missing"]), "on_hire": len(d["onhire"])}
    labels, ser, days = trend_rows("rigging", asat_dt, today)
    if len(days) < 2:
        return ""
    first, last = days[0].strftime("%d %b"), days[-1].strftime("%d %b %Y")

    def cell(v):
        return num(v) if v is not None else '<span class="tbc">no run</span>'
    trows = [[dd.strftime("%d %b %Y"), cell(ser["not_found"][i]), cell(ser["on_hire"][i])]
             for i, dd in enumerate(days)][-10:]
    return f"""{psect("The trend - last 30 days")}
{pcallout(f'<b>{num(len(days))} days on record</b> between {esc(first)} and {esc(last)}, read back from {HIST_NAME} - every point is a figure a report printed on that day, nothing interpolated. A day with no run leaves a gap. The lines answer the one question a single pull cannot: is the not-found list being walked down, and is the customer hire coming home?', False)}
{psubh("Not found in SiteIQ, and on hire to customers", "- distinct register barcodes at each pull")}
{chartpanel(sh.line_chart(labels, [("Not found in SiteIQ", ser["not_found"]), ("On hire to customers", ser["on_hire"])], y_label="barcodes", h=220))}
{sh.dtable(["Pull day (last " + str(len(trows)) + " on record)", "Not found in SiteIQ", "On hire to customers"], trows, ["", "r", "r"], cls="cp")}
{pnote('The scoreboard holds exactly what each day&rsquo;s report printed - the same figures as the position page of that day&rsquo;s PDF. Numbers, not lines, are the record.')}"""


# =====================================================================
# THE YEAR IN MOVEMENTS - the transaction log for the register gear
# =====================================================================
# WHY (03 Sep 2026): txn_insights reads the whole TRANSACTIONS export once
# and, for the register barcodes SiteIQ returns, gives the year's issues
# and returns week by week, who holds the gear, and the log's own quality
# counts. Every figure is a count over rows in the export - nothing is
# modelled or forecast.

def weekly_chart(labels, series, w=636, h=200, partial_last=False):
    """Issues, returns and net out by week on the dark panel - the shell's
    line chart with a signed axis. WHY (03 Sep 2026): k2shell.line_chart
    draws from zero up, and net out (issues less returns) goes below zero
    in a week more comes back than goes out; the shell has no signed axis
    yet, so this local drawing keeps the shell's look and adds the zero
    line. series: [(name, colour, values)]. With partial_last the final
    point is the current, partial week - drawn hollow and starred."""
    n = len(labels)
    if n < 2 or not series:
        return '<div class="note">Not enough weeks in the log for a line.</div>'
    top, base, pad_r = 26, h - 26, 44
    allv = [v for _, _, vs in series for v in vs if v is not None]
    hi = max(allv + [1])
    lo = min(allv + [0])
    # a round grid: the step is 1, 2, 2.5 or 5 times a power of ten, the
    # floor and ceiling sit on it, so the axis reads -40, 0, 40, 80, 120
    raw = (hi - lo) / 4 or 1
    mag = 10 ** math.floor(math.log10(raw))
    step = min(x * mag for x in (1, 2, 2.5, 5, 10) if x * mag >= raw)
    lo = math.floor(lo / step) * step
    hi = math.ceil(hi / step) * step
    ticks = []
    t = lo
    while t <= hi + step / 2:
        ticks.append(t)
        t += step
    span = (hi - lo) or 1
    widest = max(len(num(round(t))) for t in ticks)
    pad_l = max(36, int(widest * 4.6) + 12)
    F = 'font-family="Lato, Calibri, sans-serif"'

    def x_of(i):
        return pad_l + (w - pad_l - pad_r) * i / (n - 1)

    def y_of(v):
        return base - (base - top) * (v - lo) / span
    out = [f'<svg width="{w}" height="{h}" viewBox="0 0 {w} {h}">']
    for gv in ticks:
        y = y_of(gv)
        zero = lo < 0 and abs(gv) < step / 2
        out.append(f'<line x1="{pad_l}" y1="{y:.1f}" x2="{w - pad_r}" y2="{y:.1f}" '
                   f'stroke="{"#5A6875" if zero else "#2A3644"}" stroke-width="{1.4 if zero else 1}"/>')
        out.append(f'<text x="{pad_l - 5}" y="{y + 3:.1f}" text-anchor="end" fill="#8A9AAC" {F} '
                   f'font-size="7.6">{num(round(gv))}</text>')
    step = max(1, (n - 1) // 8 or 1)
    for i, lab in enumerate(labels):
        if i % step == 0 or i == n - 1:
            star = "*" if partial_last and i == n - 1 else ""
            out.append(f'<text x="{x_of(i):.1f}" y="{base + 14}" text-anchor="middle" fill="#8A9AAC" {F} '
                       f'font-size="7.6">{esc(str(lab))}{star}</text>')
    lx = pad_l
    for name, c, vals in series:
        pts = [(x_of(i), y_of(v)) for i, v in enumerate(vals) if v is not None]
        if len(pts) > 1:
            out.append('<polyline points="' + " ".join(f"{x:.1f},{y:.1f}" for x, y in pts)
                       + f'" fill="none" stroke="{c}" stroke-width="2.2" stroke-linejoin="round" stroke-linecap="round"/>')
        for j, (x, y) in enumerate(pts):
            hollow = partial_last and j == len(pts) - 1
            out.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="2.4" fill="{"#171F2B" if hollow else c}" '
                       f'stroke="{c}" stroke-width="1.4"/>')
        if pts:
            last = [v for v in vals if v is not None][-1]
            out.append(f'<text x="{pts[-1][0] + 6:.1f}" y="{pts[-1][1] + 3.5:.1f}" fill="{c}" {F} '
                       f'font-size="8.6" font-weight="700">{num(last)}</text>')
        out.append(f'<rect x="{lx}" y="4" width="9" height="9" rx="2" fill="{c}"/>')
        out.append(f'<text x="{lx + 13}" y="12" fill="#C9D6E2" {F} font-size="8.2">{esc(name)}</text>')
        lx += 13 + 5.6 * len(name) + 16
    out.append(f'<text x="{w - pad_r}" y="12" text-anchor="end" fill="#8A9AAC" {F} font-size="7.6">movements</text>')
    out.append("</svg>")
    return "".join(out)


def movements_page(d, asat_s):
    """The year in movements (03 Sep 2026): every issue and return of a
    register barcode in the log, week by week, with net out; and who
    holds the register gear at the pull, ranked, with the 80/20 line.
    One fixed A4 page; the holders list is the top ten and says so."""
    L = d["log"]
    wk, ho = L["weekly"], L["holders"]
    w0, w1 = L["window"] if L["window"][0] else (None, None)
    labels = [w["week"] for w in wk]
    partial = bool(wk) and wk[-1]["partial"]
    issues = sum(w["issues"] for w in wk)
    returns = sum(w["returns"] for w in wk)
    busiest = max(wk, key=lambda w: w["issues"]) if wk else None
    top = ho["top"][:10]
    hrows = []
    for i, r in enumerate(top, 1):
        custody = bool(REPAIR_RE.search(r["hirer"] or "")) or (r["company"] or "").strip().lower() == "repairs"
        co = esc(r["company"]) or dash()
        if custody:
            co += '<span class="s2">custody line - not customer hire</span>'
        hrows.append([num(i), hl(r["hirer"]) or dash(), co, num(r["items"]), num(r["oldest"])])
    if not hrows:
        hrows = [['<span class="tbc">no register item is on hire</span>', dash(), dash(), "0", dash()]]
    top_items = sum(r["items"] for r in top)
    share = round(top_items / ho["items"] * 100) if ho["items"] and top else 0
    chart = weekly_chart(labels, [("Issues", K["orange"], [w["issues"] for w in wk]),
                                  ("Returns", K["green"], [w["returns"] for w in wk]),
                                  ("Net out", K["amber"], [w["net"] for w in wk])], h=190, partial_last=partial)
    first = wk[0]["start"].strftime("%d %b %Y") if wk else "-"
    busy_s = (f'; the busiest week for issues began <b>{esc(busiest["week"])}</b> with {num(busiest["issues"])}'
              if busiest and busiest["issues"] else "")
    part_s = (f' The current week ({esc(wk[-1]["week"])}) is partial - to the pull.' if partial else "")
    period = f"report period {w0:%d %b %Y %H:%M} to {w1:%d %b %Y %H:%M}" if w0 else "the whole export"
    return f"""{psect("The year in movements - register gear through the counter, week by week")}
{pcallout(f'<span class="lead">The log, week by week.</span> The <b>{num(L["scope_n"])} register barcodes SiteIQ returns</b> moved <b class="o">{num(issues)} times out</b> and <b>{num(returns)} back</b> across {num(len(wk))} weeks, from the week beginning {esc(first)} to the pull at {esc(asat_s)}{busy_s}. Net out is the week&rsquo;s issues less its returns - above zero, gear building up on site; below, gear coming home.{part_s}')}
{psubh("Issues, returns and net out by week", "- every movement of a register barcode; * marks the partial week")}
{chartpanel(chart)}
{psubh("Who holds the register gear", "- top 10, ranked by items on hire at the pull")}
{sh.dtable(["Rank", "Hirer", "Company", "Items", "Oldest (days)"], hrows, ["r", "", "", "r", "r"], cls="cp")}
{pnote(f'<b>So what:</b> <b>{num(ho["n80_items"])} of the {num(ho["holders"])} holders</b> carry 80% of the {num(ho["items"])} register items out, and the ten names above hold {num(share)}% - the net-out line says which weeks to walk that list.')}
{pnote(f'Counted from TRANSACTIONS (sheet CUSTOMER_CONTRACTOR_EQUIP, {period}) for the register barcodes SiteIQ returns; holders from RENTAL_STOCK at the pull, oldest = days from the on-hire date. Rules: short hire = closed inside 6 minutes; mass draw = one person drawing 15 or more items inside one hour; product key = the description with its size and serial tail removed.')}"""

def rig_tiles(d):
    """The position-page tiles: the count, the movement since the previous
    recorded pull where one exists (else the tile's own note), and a
    sparkline once there is an earlier day. Same tuples feed the card."""
    asat_dt = d["asat"]
    oh_n, rp_n, st_n, ms_n = (len(d["onhire"]), len(d["repair"]),
                              len(d["store"]), len(d["missing"]))
    fill_n, rows_n, n_co = len(d["with_test"]), d["rows"], len(d["companies"])
    spec = [
        ("box", "distinct", d["distinct"], num(d["distinct"]), "Distinct barcodes",
         f"{num(rows_n)} register rows", "grey", "up"),
        ("swap", "on_hire", oh_n, num(oh_n), "On hire to customers",
         f"across {num(n_co)} companies", "amber", "up"),
        ("check", "in_store", st_n, num(st_n), "In the store", "available for hire", "green", "up"),
        ("wrench", "repairs", rp_n, num(rp_n), "At repairs / quarantined", "not customer hire",
         "amber", "down"),
        ("warn", "not_found", ms_n, num(ms_n), "Not found in SiteIQ", "whereabouts unknown",
         "red", "down"),
        ("shield", "tests_filled", fill_n, f"{num(fill_n)}/{num(rows_n)}", "Test records entered",
         "fill-in not started" if not fill_n else "fill-in under way",
         "red" if not fill_n else "amber", "up"),
    ]
    out = []
    for ico, key, val, shown, lab, note, ncls, good in spec:
        note, ncls = moved(key, asat_dt, val, good, note, ncls)
        out.append((ico, shown, lab, note, ncls, spark_of(key, asat_dt, val)))
    return out


def rig_rag(d):
    """The position-page RAG - the rule in CONFIG applied to the not-found
    share at the pull. Returns a dict: the rag_band arguments (status,
    headline, rule, owner, action) as HTML, plus card_headline and
    card_action - the same facts in fewer words, because the phone card
    draws a fixed box (two headline lines, two action lines)."""
    ms_n, distinct = len(d["missing"]), d["distinct"]
    share = ms_n / distinct * 100 if distinct else 0
    red_at = CONFIG["rag_red_pct"]
    if ms_n == 0:
        status = "green"
        headline = ('Every distinct register barcode is returned by the SiteIQ pull - the '
                    'fleet is accounted for.')
        card_headline = 'Every register barcode is returned by the SiteIQ pull - accounted for.'
    elif share >= red_at:
        status = "red"
        headline = (f'<b class="o">{num(ms_n)}</b> of the <b>{num(distinct)}</b> distinct register '
                    f'barcodes - <b class="o">{share:.0f}%</b> - are not found in SiteIQ: '
                    f'whereabouts unknown, not accounted for.')
        card_headline = (f'{num(ms_n)} of the {num(distinct)} register barcodes ({share:.0f}%) are '
                         f'not found in SiteIQ - whereabouts unknown.')
    else:
        status = "amber"
        headline = (f'<b>{num(ms_n)}</b> of the <b>{num(distinct)}</b> distinct register barcodes - '
                    f'{share:.0f}% - are not found in SiteIQ: under the {num(red_at)}% red line, '
                    f'still whereabouts unknown.')
        card_headline = (f'{num(ms_n)} of the {num(distinct)} register barcodes ({share:.0f}%) are '
                         f'not found in SiteIQ - under the {num(red_at)}% red line.')
    rule = (f'Share of the distinct register barcodes the SiteIQ pull does not return: Green at 0%, '
            f'Amber under {num(red_at)}%, Red from {num(red_at)}%. <b>Default line - set in CONFIG.</b>')
    ms_oh = d["missing_snap"].get("On Hire", 0)
    # dated from the pull day - the wall clock never sets it
    due_s = (d["asat"].date() + timedelta(days=CONFIG["action_days"])).strftime("%d %b %Y")
    action = (f'Hunt list of the {num(ms_n)} walked; the {num(ms_oh)} last seen on hire raised '
              f'with their hirers - by <b>{esc(due_s)}</b>.')
    return {"status": status, "headline": headline, "rule": rule,
            "owner": esc(CONFIG["rag_owner"]), "action": action,
            "card_headline": card_headline, "card_action": plain(action), "due": due_s}


def paginate(rows, first_budget, next_budget):
    """rows: (cells, est_height_px). Fills pages by height so a long
    table never runs into the footer - the K2 page clips silently."""
    pages, cur, used, budget = [], [], 0, first_budget
    for cells, h in rows:
        if cur and used + h > budget:
            pages.append(cur)
            cur, used, budget = [], 0, next_budget
        cur.append(cells)
        used += h
    if cur or not pages:
        pages.append(cur)
    return pages


# =====================================================================
# the position page (03 Sep 2026): three things to do today, the story
# =====================================================================

def story_words(fragment):
    """Word count of a callout with the markup off. The position page's
    story is held to three lines - about 45 words - and the console
    prints the count so a longer one is caught before it is sent."""
    return len(plain(fragment).split())


def three_things_rigging(d, rag):
    """The three actions on the position page, each read from the data and
    dated by the band's own rule (action_days after the pull). Candidates
    in priority order; the block prints the first three that exist and
    never invents one."""
    who = f"Andrew Fisher · by {rag['due']}"
    cands = []
    # 1. the hunt list - who was last holding the most of it
    ms = d["missing"]
    if ms:
        holders = Counter((it["snap"] or {}).get("hirer", "") for it in ms
                          if (it["snap"] or {}).get("status", "").lower() == "on hire")
        holders.pop("", None)
        snap = f" of {d['snap_max']:%d %b %Y}" if d["snap_max"] else ""
        if holders:
            h, k = holders.most_common(1)[0]
            why = (f"{num(k)} last seen on hire with {h}, per the workbook snapshot{snap} - "
                   f"the first person to ask")
        else:
            c, k = d["missing_cats"].most_common(1)[0]
            why = f"{num(k)} of them {c.lower()} - whereabouts unknown until SiteIQ returns them"
        cands.append((f"Hunt the {num(len(ms))} register items SiteIQ cannot see", why, who))
    # 2. the oldest customer hire
    if d["oh_longest"]:
        o = d["oh_longest"][0]
        cands.append((f"Chase {o['company'] or o['hirer']} for {o['barcode']}, out {num(o['held'])} days",
                      f"{o['desc']} - {o['hirer']}, on hire since {o['since']:%d %b %Y}", who))
    # 3. the custody lines; the test-record gap when there are none
    if d["repair"]:
        lines = ", ".join(f"{r['hirer']} {num(r['n'])}" for r in d["repair_lines"][:2])
        cands.append((f"Clear {num(len(d['repair']))} items at repairs or quarantine",
                      f"custody lines, not customer hire - {lines}", who))
    fill_n, rows_n = len(d["with_test"]), d["rows"]
    if fill_n < rows_n:
        cands.append((f"Start the test records - {num(fill_n)} of {num(rows_n)} rows entered",
                      "the eight test columns on the Master are blank; a dash is not a certificate", who))
    return cands[:3]


# =====================================================================
# the PDF pages
# =====================================================================

def build_pages(d, asat_s):
    """Returns the list of page bodies. Page-number references inside
    the text are tokens, resolved once the page order is final."""
    P = []
    distinct, rows_n = d["distinct"], d["rows"]
    oh_n, rp_n, st_n, ms_n = (len(d["onhire"]), len(d["repair"]),
                              len(d["store"]), len(d["missing"]))
    ot_n = len(d["other"])
    found_n = len(d["found"])
    n_co = len(d["companies"])
    fill_n = len(d["with_test"])
    ex = d["extract"]
    snap_s = fmt_dt(d["snap_max"])
    cat_bits = " and ".join(f'<b>{num(n)} {esc(c.lower())}</b>'
                            for c, n in d["cats_rows"].most_common())
    other_bit = (f', {num(ot_n)} in another SiteIQ status '
                 f'({esc(", ".join(f"{k} {v}" for k, v in d["other_status"].most_common()))})'
                 if ot_n else "")

    # ---- P1 the position - the band, the tiles, three things, the story --
    # WHY (03 Sep 2026): one grammar for every position page in the suite -
    # the RAG band first, the tiles with their movement, the three things
    # to do today, then the story in three lines. The arithmetic behind it
    # (the donut, the ladders, the long paragraph, the where-the-barcodes-
    # are band) sits on the scorecard page after this one.
    rag = rig_rag(d)
    story = (f'<span class="lead">The story.</span> <b>{num(rows_n)}</b> register rows, <b>{num(distinct)}</b> '
             f'distinct barcodes, joined to SiteIQ at the pull: <b>{num(found_n)} found ({d["found_pct"]}%)</b> - '
             f'{num(oh_n)} on hire to customers across {num(n_co)} companies, {num(st_n)} in the store, '
             f'{num(rp_n)} at repairs - and <b class="o">{num(ms_n)} not found, whereabouts unknown</b>. '
             f'Test records: {num(fill_n)} of {num(rows_n)} rows.')
    print(f"Story callout        : {story_words(story)} words (position page, three lines at most)")
    P.append(f"""<div class="pos">{sh.rag_band(rag["status"], rag["headline"], rag["rule"], rag["owner"], rag["action"])}
{sh.tiles_plus(rig_tiles(d), per_row=6)}
{sh.three_things(three_things_rigging(d, rag))}
{pcallout(story)}</div>""")

    # ---- P1b the scorecard - the arithmetic behind the position ---------
    ladders = sh.score_rows([
        ("Found in SiteIQ", d["found_pct"],
         f"{num(found_n)} of {num(distinct)} distinct register barcodes are returned "
         f"by the SiteIQ pull as at {asat_s}; {num(ms_n)} are not - listed from "
         f"page __PG_MISSING__"),
        ("Test records filled in", round(fill_n / rows_n * 100) if rows_n else 0,
         f"{num(fill_n)} of {num(rows_n)} register rows have any of the eight test "
         f"columns entered - see page __PG_TEST__"),
    ])
    where_band = sh.stackband([
        ("On hire to customers", oh_n, K["orange"]),
        ("At repairs / quarantined", rp_n, K["amber"]),
        ("In the store", st_n, K["green"]),
    ] + ([("Other SiteIQ status", ot_n, K["blue"])] if ot_n else []) + [
        ("Not found in SiteIQ", ms_n, K["red"]),
    ])
    P.append(f"""{psect("The scorecard - the arithmetic behind the position")}
{pcallout(
        f'<span class="lead">The position, in full.</span> One honest answer: '
        f'is the {esc(CONFIG["client"])} rigging and height-safety fleet on the '
        f'register and accounted for? The register carries <b>{num(rows_n)} rows</b> '
        f'- <b>{num(distinct)} distinct barcodes</b> ({num(len(d["blank"]))} blank, '
        f'{num(len(d["dups"]))} barcodes repeated on {num(d["dup_rows"])} rows) - split '
        f'{cat_bits}. Joined barcode-for-barcode to the SiteIQ register as at '
        f'<b>{esc(asat_s)}</b>: <b>{num(found_n)} found ({d["found_pct"]}%)</b> - '
        f'<b class="o">{num(oh_n)} on hire to customers</b> across {num(n_co)} '
        f'companies, {num(rp_n)} at repairs or quarantined, {num(st_n)} in the '
        f'store{other_bit} - and <b class="o">{num(ms_n)} not found in SiteIQ - '
        f'whereabouts unknown</b>. The test columns are still empty: '
        f'<b>{num(fill_n)} of {num(rows_n)}</b> rows carry any test detail (page __PG_TEST__).', False)}
<table class="two" style="margin-top:12px"><tr>
  <td style="width:31%"><div class="donut-wrap">
    {sh.donut(d["found_pct"], K["orange"], f'{d["found_pct"]}%', "FOUND IN SITEIQ", size=118)}
    <div class="donut-cap">Share of the {num(distinct)} distinct register barcodes the SiteIQ pull returns</div></div></td>
  <td style="padding-left:10px">{ladders}</td>
</tr></table>
{psubh("Where the barcodes are", f"- every one of the {num(distinct)}, per the SiteIQ pull")}
{chartpanel(where_band)}""")

    # ---- P2 where the gear is - by company ------------------------------
    pg_company = len(P) + 1 + COVER_PAGES
    comp = d["companies"]
    crows = []
    for c in comp:
        o = c["oldest"]
        same = (len(c["accounts"]) == 1
                and re.sub(r"[^a-z0-9]", "", c["accounts"][0].lower())
                == re.sub(r"[^a-z0-9]", "", c["co"].lower()))
        acc = ("" if same else
               f'<span class="s2">{esc(" · ".join(c["accounts"]))}</span>')
        crows.append([
            f'{esc(c["co"])}{acc}',
            num(c["n"]),
            num(c["hirers"]),
            (f'{esc(o["desc"])} &middot; {esc(o["barcode"])} &middot; {hl(o["hirer"])}'
             if o else dash()),
            since_cell(o) if o else dash(),
            held_cell(o) if o else dash()])
    rrows = []
    for r in d["repair_lines"]:
        o = r["oldest"]
        rrows.append([
            hl(r["hirer"]),
            esc(" · ".join(r["accounts"])),
            num(r["n"]),
            f'{esc(o["desc"])} &middot; {esc(o["barcode"])}' if o else dash(),
            since_cell(o) if o else dash(),
            held_cell(o) if o else dash()])
    if not rrows:
        rrows = [['<span class="tbc">no register item is at repairs or quarantined</span>',
                  dash(), "0", dash(), dash(), dash()]]
    P.append(f"""{psect("Where the rigging gear is - by company, A to Z")}
{pcallout(f'<b class="o">{num(oh_n)} items</b> are on hire to <b>{num(n_co)} customer companies</b> ({num(d["oh_hirers"])} hirer names) per the SiteIQ pull as at {esc(asat_s)}. Companies A to Z, each with its <b>longest-held item</b> named with barcode and hirer - because the oldest hire is always the first conversation. Project accounts are merged into their company (the SiteIQ account names sit under the company). A further <b>{num(rp_n)}</b> register items sit on repairs or quarantine custody lines - shown separately on page __PG_LONGEST__, never counted as customer hire.', False)}
{sh.dtable(["Company (SiteIQ accounts)", "Items", "Hirers", "Longest-held item · barcode · hirer", "On hire since", "Held"],
           crows, ["", "r", "r", "", "r", "r"], cls="cp")}
{pnote(f'Held = days from the SiteIQ on-hire date to the pull time, {esc(asat_s)}. Hirer names are as SiteIQ records them, shared site accounts included. {num(d["oh_after_snap"])} of the {num(oh_n)} customer-hire items were issued after {snap_s}, the newest date in the workbook&rsquo;s own locate sheet - which is why that sheet is no longer used for status. Repairs and quarantine custody lines are on page __PG_LONGEST__.')}""")

    # ---- P3 the longest-held, item by item -------------------------------
    pg_longest = len(P) + 1 + COVER_PAGES
    cap_l = 16
    lrows = []
    for it in d["oh_longest"][:cap_l]:
        lrows.append([
            esc(it["company"]) or dash(),
            hl(it["hirer"]) or dash(),
            esc(it["desc"]) or dash(),
            esc(it["barcode"]) or dash(),
            since_cell(it),
            held_cell(it)])
    P.append(f"""{psect("The longest-held customer hire, item by item - oldest first")}
{pcallout(f'The front of the queue: the {min(cap_l, len(d["oh_longest"]))} oldest of the {num(len(d["oh_longest"]))} dated customer-hire items, each with its barcode, holder and SiteIQ on-hire date as at {esc(asat_s)}. Work them top down - every return retires the oldest risk first.', False)}
{sh.dtable(["Company", "Hirer", "Item", "Barcode", "On hire since", "Held"],
           lrows, ["", "", "", "", "r", "r"], cls="cp")}
{pnote((f'Showing {cap_l} of {num(len(d["oh_longest"]))} dated customer-hire items - the SiteIQ export carries the lot, same order. ' if len(d["oh_longest"]) > cap_l else '') + ('' if not d["oh_nodate"] else f'{num(len(d["oh_nodate"]))} on-hire items carry no on-hire date in SiteIQ - shown as dashes, not guessed. ') + 'Held = days from the on-hire date to the pull time.')}
{psubh("At repairs / quarantined", f"- {num(rp_n)} items on custody lines, not customer hire")}
{sh.dtable(["Custody line (hirer)", "SiteIQ account", "Items", "Longest-held item", "Since", "Held"],
           rrows, ["", "", "r", "", "r", "r"], cls="cp")}
{pnote(f'Rule: {esc(REPAIR_RULE)}. These lines are excluded from the company count on page __PG_COMPANY__ and from the longest-held list above.')}""")

    # ---- P4 test status - the truth -------------------------------------
    pg_test = len(P) + 1 + COVER_PAGES
    ts = d["test_status"].most_common()
    if ts:
        srows = [[esc(s), num(n)] for s, n in ts]
        srows.append(['<span class="tbc">(blank - not yet entered)</span>',
                      num(rows_n - sum(n for _, n in ts))])
    else:
        srows = [['<span class="tbc">(blank - not yet entered)</span>', num(rows_n)]]
    fill_rows = [(lab, n, rows_n, "f-orange" if n else "f-amber",
                  f"{num(n)} of {num(rows_n)}") for lab, n in d["col_fill"]]
    if ex["present"]:
        rb = ex["reg_buckets"]
        due_txt = " · ".join(f"{esc(v)} on {num(n)} rows" for v, n in ex["next_due"].most_common(3))
        parse_txt = " · ".join(f"{esc(k)} {num(n)}" for k, n in ex["parse"].most_common())
        exrows = [
            ["Rows in the extract / distinct barcodes",
             f'{num(ex["rows"])} / {num(ex["distinct"])}'],
            ["Of which on the register (Master barcodes)",
             f'<b>{num(ex["on_register"])}</b> of {num(distinct)}'],
            ["&nbsp;&nbsp;&rarr; on hire per SiteIQ (customers + custody lines)",
             num(rb.get("onhire", 0) + rb.get("repair", 0))],
            ["&nbsp;&nbsp;&rarr; in the store per SiteIQ", num(rb.get("store", 0) + rb.get("other", 0))],
            ["&nbsp;&nbsp;&rarr; not found in SiteIQ", num(rb.get("missing", 0))],
            ["Next Insp Due values carried", due_txt],
            ["Parse Status (the extract&rsquo;s own column)", parse_txt],
        ]
        ex_block = (psubh("The certificate extract", "- &lsquo;Extracted Register&rsquo; sheet, same workbook, not yet on the Master")
                    + sh.dtable(["What the extract holds", "Count"], exrows, ["", "r"], cls="cp"))
        top_due = ex["next_due"].most_common(1)[0][0] if ex["next_due"] else ""
        ex_sentence = (
            f' A certificate extract in the same workbook (&lsquo;Extracted Register&rsquo;) carries '
            f'<b>Next Insp Due {esc(top_due)} for {num(ex["on_register"])} of the {num(distinct)} register '
            f'barcodes</b> ({num(rb.get("onhire", 0) + rb.get("repair", 0))} on hire per SiteIQ, '
            f'{num(rb.get("store", 0) + rb.get("other", 0))} in the store, {num(rb.get("missing", 0))} not found '
            f'in SiteIQ), not yet transferred to the Master. If {esc(top_due)} means April 2026, those items '
            f'were due before this report - <b>status unconfirmed</b>.')
    else:
        ex_block = pnote("No &lsquo;Extracted Register&rsquo; sheet was found in the workbook - nothing to disclose from a certificate extract.")
        ex_sentence = ""
    oot = d["out_of_tag"]
    oot_bit = (f' SiteIQ already carries a parking hirer, &lsquo;{hl(oot[0]["hirer"])}&rsquo;, holding '
               f'{num(len(oot))} register item{"s" if len(oot) != 1 else ""} (page __PG_COMPANY__).' if oot else "")
    P.append(f"""{psect("Test status - what the register holds today, no varnish")}
{pcallout(f'Straight up: <b>not one of the eight test columns on the Master has been filled in</b> - {num(fill_n)} of {num(rows_n)} rows carry a test date, status, tag colour, tester, licence, certificate or comment.{ex_sentence}{oot_bit}', False)}
{sh.dtable(["Test Status on the Master (as recorded)", "Rows"], srows, ["", "r"], cls="cp")}
{ex_block}
{psubh("Fill-in progress by column", "- how far each test column has got")}
{sh.prog_rows(fill_rows)}
{pnote('A dash anywhere in this report means the register cell is blank - the test details for that item are still to be entered. Nothing is assumed, estimated or copied in from the extract; the extract is disclosed here, not treated as the record.')}""")

    # ---- P5 register identity + category detail -------------------------
    duprows = []
    for b, n in d["dups"]:
        it = d["items"][b]
        lv = it["live"]
        where = ({"onhire": f'On hire · {esc(it["company"])} · {hl(it["hirer"])}',
                  "repair": f'Custody · {hl(it["hirer"])}',
                  "store": "In the store", "other": esc(lv["status"]) if lv else "",
                  "missing": '<span class="or">Not found in SiteIQ</span>'}[it["bucket"]])
        duprows.append([esc(b), esc(it["desc"]) or dash(), num(n), where])
    if not duprows:
        duprows = [['<span class="tbc">no duplicated barcodes</span>', dash(), "0", dash()]]
    blank_bits = "; ".join(f'{esc(r["desc"])} ({esc(r["cat"])})' for r in d["blank"]) or "none"
    catrows2 = []
    for c in d["cat_detail"]:
        o = c["oldest"]
        catrows2.append([
            esc(c["cat"]), num(c["rows"]), num(c["distinct"]), num(c["onhire"]),
            num(c["repair"]), num(c["store"]), num(c["missing"]),
            (f'{esc(o["desc"])}<span class="s2">{esc(o["barcode"])} &middot; '
             f'{esc(o["company"])} &middot; {hl(o["hirer"])}</span>' if o else dash()),
            held_cell(o) if o else dash()])
    P.append(f"""{psect("Register identity - what the barcodes themselves say")}
{pcallout(f'The register is <b>{num(rows_n)} rows</b> but <b>{num(distinct)} distinct barcodes</b>: <b>{num(len(d["blank"]))}</b> row{"s" if len(d["blank"]) != 1 else ""} with no barcode ({blank_bits}), <b>{num(len(d["dups"]))}</b> barcodes repeated across <b>{num(d["dup_rows"])}</b> rows, and <b>{num(d["lower"])}</b> typed in lower case. For the SiteIQ join and on every page here, barcodes are upper-cased and de-duplicated (first Master row wins) - the fixes belong in the Master. Serial numbers: <b>{num(d["serial_rows"])}</b> rows carry one, <b>{num(d["serial_distinct"])}</b> distinct.', False)}
{psubh("Duplicated barcodes", f"- {num(len(d['dups']))} barcodes on {num(d['dup_rows'])} rows")}
{sh.dtable(["Barcode", "Item", "Rows", "Where SiteIQ has it"], duprows,
           ["", "", "r", ""], cls="cp")}
{psubh("Category detail", "- rows, distinct barcodes and the SiteIQ split")}
{sh.dtable(["Category", "Rows", "Barcodes", "On hire", "Repairs", "In store", "Not found", "Longest-held customer item", "Held"],
           catrows2, ["", "r", "r", "r", "r", "r", "r", "", "r"], cls="cp")}
{pnote('Counts are per distinct barcode after upper-casing; a duplicated barcode is one item here whatever the row count says. On hire = customer hire only.')}""")

    # ---- P6.. not found in SiteIQ ---------------------------------------
    def snap_cell(it):
        s = it["snap"]
        if not s:
            return '<span class="tbc">no snapshot row</span>'
        if s["status"].lower() == "on hire":
            return f'<span class="or">On hire</span><span class="s2">{hl(s["hirer"]) or "&ndash;"}</span>'
        if s["status"].lower() == "available for hire":
            return "Available"
        return esc(s["status"] or "&ndash;")

    def snap_h(it):
        # measured on the rendered PDF: 24px a single-line row, 36px two lines
        s = it["snap"]
        h = 36 if len(it["desc"]) > 24 else 24
        if s and s["status"].lower() == "on hire":
            h = max(h, 48 if len(s["hirer"]) > 21 else 36)
        return h

    ms = d["missing"]
    pair_rows = []
    for i in range(0, len(ms), 2):
        a = ms[i]
        b = ms[i + 1] if i + 1 < len(ms) else None
        cells = [esc(a["barcode"]), esc(a["desc"]) or dash(), snap_cell(a)]
        cells += [esc(b["barcode"]), esc(b["desc"]) or dash(), snap_cell(b)] if b else ["", "", ""]
        pair_rows.append((cells, max(snap_h(a), snap_h(b) if b else 0)))
    ms_pages = paginate(pair_rows, 520, 640) if pair_rows else [[]]
    ms_oh = d["missing_snap"].get("On Hire", 0)
    ms_split = " · ".join(f"{esc(k)} {num(v)}" for k, v in d["missing_cats"].most_common())
    snap_split = " · ".join(f"{esc(k)} {num(v)}" for k, v in d["missing_snap"].most_common())
    first_ms = len(P) + 1 + COVER_PAGES
    for i, rows in enumerate(ms_pages):
        n_pages = len(ms_pages)
        head = psect("Not found in SiteIQ - whereabouts unknown"
                     + (f" ({i + 1} of {n_pages})" if n_pages > 1 else ""))
        if i == 0:
            head += pcallout(
                f'<b class="o">{num(ms_n)} register barcodes</b> are not returned by the SiteIQ pull as at '
                f'{esc(asat_s)} - on the register, but nowhere in the live export under that barcode '
                f'(exact match tried first, then upper-cased and trimmed). Split {ms_split}. They are '
                f'<b>not accounted for</b> and are not counted as in the store. The last thing the '
                f'workbook&rsquo;s own locate sheet (a SiteIQ snapshot dated to {snap_s}) said about them: '
                f'{snap_split}. The {num(ms_oh)} last seen on hire carry that hirer&rsquo;s name below - '
                f'the first person to ask.', False)
        if not rows:
            body = pnote("Every register barcode was returned by the SiteIQ pull - nothing to list.")
        else:
            body = sh.dtable(["Barcode", "Item", "Jun snapshot",
                              "Barcode", "Item", "Jun snapshot"], rows,
                             ["", "", "", "", "", ""], cls="cp")
        P.append(head + body + pnote(
            f'Sorted last-seen-on-hire first (by hirer), then by item. &lsquo;Jun snapshot&rsquo; is the '
            f'workbook&rsquo;s To Help Locate sheet, dated to {snap_s} - not the live position. A barcode '
            f'here may have been re-tagged, sold, scrapped or mis-typed on the register; until SiteIQ '
            f'returns it, whereabouts unknown.'))

    # ---- P?.. possibly rigging gear not on the register ------------------
    kwrows = [([esc(g["desc"]), num(g["n"]), num(g["onhire"]), num(g["store"]),
                num(g["other"]), esc(g["eg"])], 24) for g in d["kw_groups"]]
    kw_pages = paginate(kwrows, 520, 640) if kwrows else [[]]
    first_kw = len(P) + 1 + COVER_PAGES
    for i, rows in enumerate(kw_pages):
        n_pages = len(kw_pages)
        head = psect("Possibly rigging gear not on the register - keyword match, verify - ranked by count"
                     + (f" ({i + 1} of {n_pages})" if n_pages > 1 else ""))
        if i == 0:
            head += pcallout(
                f'<b class="o">{num(len(d["kw"]))} live SiteIQ lines</b> ({num(len(d["kw_groups"]))} '
                f'descriptions, {num(d["kw_onhire"])} on hire) carry a rigging or height-safety word in '
                f'their description but their barcode is not on the register. <b>This is a keyword match, '
                f'not a finding</b> - each line needs a look before it is added. Rule applied: '
                f'{esc(KW_RULE)}.', False)
        body = (sh.dtable(["Description (as SiteIQ records it)", "Items", "On hire", "In store",
                           "Other", "Example barcode"], rows, ["", "r", "r", "r", "r", ""], cls="cp")
                if rows else pnote("No live SiteIQ line matched the keyword rule outside the register."))
        P.append(head + body + pnote(
            'Grouped by description exactly as SiteIQ spells it. Counts and statuses are from the same '
            'pull as every other page. Items that are genuinely rigging gear belong on the Master with a '
            'barcode; items that are not can be dismissed - either way, verify before acting.'))

    # ---- the trend - only once seven days are on record ------------------
    tp = trend_page(d)
    if tp:
        P.append(tp)

    # ---- the year in movements - the transaction log (03 Sep 2026) --------
    pg_moves = len(P) + 1 + COVER_PAGES
    if d.get("log"):
        P.append(movements_page(d, asat_s))

    # ---- close --------------------------------------------------------------
    cards = sh.info_cards([
        ("Live position, not a snapshot",
         f"Status, holder, on-hire date and days held come from the SiteIQ "
         f"RENTAL_STOCK pull requested {esc(asat_s)}, joined barcode-for-barcode "
         f"to the register. The workbook decides what is on the register; "
         f"SiteIQ decides where it is."),
        ("Unknown stays unknown",
         f"{num(ms_n)} register barcodes SiteIQ does not return are printed as "
         f"not found - whereabouts unknown - never folded into &lsquo;in the "
         f"store&rsquo;. Blank test cells print as dashes; nothing is assumed, "
         f"estimated or copied in."),
        ("Longest held, first chased",
         "Customer hire is worked oldest first by company and hirer. Repairs, "
         "off-site, out-of-tag and out-of-service custody lines are kept "
         "separate, so a quarantined item never reads as a customer&rsquo;s."),
        ("Identity has gaps too",
         f"{num(rows_n)} rows, {num(distinct)} distinct barcodes: "
         f"{num(len(d['blank']))} blank, {num(len(d['dups']))} barcodes repeated "
         f"on {num(d['dup_rows'])} rows, {num(d['lower'])} lower-case. The join "
         f"upper-cases and de-duplicates; the fixes belong in the Master. Rigging "
         f"gear is Life Saving Rule 5 territory (SEQ-GL-009) - the record has to be real."),
    ])
    former_n, former_live = d["former_reg"], d["former_live"]
    n_days = history_days("rigging", d["asat"])
    trend_line = (f'Trend page: appears once seven days are on record ({num(n_days)} today).'
                  if n_days < 7 else
                  f'Trend page: {num(n_days)} days on record - the 30-day lines are on the page before this one.')
    # WHY (03 Sep 2026): the transaction log is a source now - named here
    # with its report period, and its own check for the register gear in
    # one line; the rules it applies are on the movements page.
    log_note = ""
    L = d.get("log")
    if L:
        lw0, lw1 = L["window"] if L["window"][0] else (None, None)
        period = f", report period {lw0:%d %b %Y %H:%M} to {lw1:%d %b %Y %H:%M}" if lw0 else ""
        dq = L["dq"]
        log_note = (f' The transaction log (TRANSACTIONS, sheet CUSTOMER_CONTRACTOR_EQUIP{period}) feeds page __PG_MOVES__, '
                    f'the year in movements; its own check for the register gear: {num(dq["short_n"])} hires closed inside '
                    f'6 minutes, {num(len(dq["onhire_no_log"]))} on hire with no movement since the log began, '
                    f'{num(len(dq["onhire_before_log"]))} issued before it opened.')
    P.append(f"""{psect("How the rigging fleet is run")}
{cards}
{pnote(f'Names as shown: the site is Ampol. SiteIQ and the register still carry the site&rsquo;s former name on {num(former_n)} register descriptions and {num(former_live)} live-register lines; every one is shown here under the current name. Barcodes are identifiers and never change. Companies are one customer one name (the refinery legal name and the former site account both read Ampol; project accounts roll up to their parent). Each run writes its figures to {HIST_NAME} keyed on the pull day; the next run reads them back for the movement notes. {trend_line}{log_note}')}
{sh.coates_way_panel(traits=(4, 5), disciplines=(1, 5), line="every lifting item is accounted for or hunted; a barcode SiteIQ cannot see is a friction point flagged, not accepted")}
{psect("Meet the tool store team")}
{pnote(f'The crew running your {esc(CONFIG["client"])} store - keeping the register true and the gear ready. Something not right? Tell us and we&rsquo;ll sort it.')}
{sh.team_cards(CONFIG["team"])}""")

    # resolve the page references now the order is final - every number
    # counts the cover as page 1
    refs = {"__PG_MISSING__": first_ms, "__PG_KW__": first_kw, "__PG_COMPANY__": pg_company,
            "__PG_LONGEST__": pg_longest, "__PG_TEST__": pg_test, "__PG_MOVES__": pg_moves}
    for tok, pg in refs.items():
        P = [p.replace(tok, str(pg)) for p in P]
    return P


# extra styling on top of the shared sheet. Six tiles across, so the
# sparkline scales down to its cell instead of the shell's four-up width;
# and the position page (class "pos") runs tighter than the shell's
# defaults - it carries the RAG band under the tiles and must clear the
# footer with the sparkline row still to come.
EXTRA_CSS = """
.tiles .spark { width: 84px; }
.pos .callout { padding: 13px 20px; font-size: 12px; line-height: 1.85; }
.pos .two { margin-top: 6px !important; }
.pos .sub-h { margin: 12px 0 8px 0; }
.pos .tiles { margin-top: 7px; }
.pos .tiles td { padding: 10px 9px 9px 9px; }
.pos .ragband { margin-top: 10px; }
"""


# the pack's closing sections - never listed on the cover
CLOSING_HEADINGS = ("How the rigging fleet is run", "Meet the tool store team")
# a continuation page of a split table - the section is listed once
_CONT_RE = re.compile(r"\((?:continued|(?:[2-9]|[1-9]\d) of \d+)\)$")


def pdf_pages(pdf_path):
    """Page count of a printed PDF, read from its own page tree."""
    raw = open(pdf_path, "rb").read()
    counts = re.findall(rb"/Count\s+(\d+)", raw)
    return max(int(c) for c in counts) if counts else -1


def cover_contents_from_print(doc, pdf_path, html_path, css, closing):
    """First print of the pack, so the cover's "What's inside" block can
    carry page numbers read off the printed pages. WHY (03 Sep 2026): a
    typed-in contents list is a promise pagination can break; these come
    from the PDF itself (pdf_finish.contents_from_pdf). The cover is one
    fixed page, so the second print - the one with the block - paginates
    identically, and main() proves it by comparing the two page counts.
    Continuation pages of a split table and the closing sections stay off
    the block; a long title keeps its first clause so the block holds its
    rows in one column. Returns (rows, page count) - ([], None) when no
    PDF engine printed."""
    heads = [t for lv, t in pdf_finish.headings_from_html(doc) if lv == 1]
    skip = tuple(closing) + tuple(t for t in heads if _CONT_RE.search(t))
    full = doc.replace("</head>", f"<style>{css}</style></head>", 1)
    Path(html_path).write_text(full, encoding="utf-8")
    try:
        Path(pdf_path).unlink()
    except OSError:
        pass
    if not eng.write_pdf_robust(str(html_path), str(pdf_path)) or not Path(pdf_path).exists():
        return [], None
    rows = []
    for t, p in pdf_finish.contents_from_pdf(pdf_path, full, has_cover=True, skip=skip):
        t = re.sub(r"\s*\(1 of \d+\)$", "", t)
        if len(t) > 64 and " - " in t:
            t = t.split(" - ")[0]
        rows.append((t, p))
    return rows, pdf_pages(pdf_path)


def hero_page(cfg, inner, pno, ptot, gen_s, asat_s):
    """The position page behind the cover: page 2 of the pack, but the
    first page of the report proper, so it wears the hero head and the key
    strip. WHY (03 Sep 2026): k2shell.render_page gives the hero to page 1
    only; the footer carries the number now, so this composes the shell's
    own parts (page1_head, key_strip, footer) with the real page number
    and patches nothing. A render_page(..., hero=True) switch in the shell
    would retire this."""
    head = sh.page1_head(cfg, gen_s, asat_s) + sh.key_strip(cfg) + '<div class="grule"></div>'
    return (f'<div class="page page1"><div class="frame">{head}'
            f'<div class="body">{inner}</div>{sh.footer(cfg, pno, ptot)}</div></div>')


def render_doc(pages, cover, gen_s, asat_s):
    """The cover is page 1 of the pack; the position page is page 2, wears
    the hero head, and every page number and cross-reference counts from
    there."""
    tot = len(pages) + COVER_PAGES
    body = cover + "".join(
        (hero_page if i == 0 else sh.render_page)(CONFIG, p, i + 1 + COVER_PAGES, tot, gen_s, asat_s)
        for i, p in enumerate(pages))
    return (f'<!doctype html><html lang="en"><head><meta charset="utf-8">'
            f'<title>Coates {esc(CONFIG["client"])} {esc(CONFIG["title"])} - '
            f'{esc(asat_s)}</title><style>{EXTRA_CSS}</style></head><body>{body}</body></html>'), tot


# =====================================================================
# the Outlook email (draft - never sends)
# =====================================================================

def card_block(cid):
    """The position card inline, under the header: a cid image the .eml
    carries in its own related part. Outlook honours the width attribute;
    the style caps it for everything else."""
    FONT = sh.FONT
    return (f'<table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr>'
            f'<td align="center" style="padding:16px 0 2px 0;">'
            f'<img src="cid:{cid}" width="420" alt="The position on one card" '
            f'style="display:block;width:420px;max-width:420px;height:auto;border:0;">'
            f'<div style="{FONT}font-size:10px;color:#7A8A9A;padding-top:6px;">The position on one card - '
            f'the same figures as the position page of the PDF; the PNG is attached for your phone.</div>'
            f'</td></tr></table>')


def build_email_html(d, gen_s, asat_s, pdf_ok, src_name, live_name, card_cid=""):
    W = 1000
    FONT = sh.FONT
    distinct, rows_n = d["distinct"], d["rows"]
    oh_n, rp_n, st_n, ms_n = (len(d["onhire"]), len(d["repair"]),
                              len(d["store"]), len(d["missing"]))
    n_co = len(d["companies"])
    fill_n = len(d["with_test"])
    ex = d["extract"]
    parts = []

    parts.append(f"""<table role="presentation" width="100%" cellspacing="0" cellpadding="0">
<tr><td bgcolor="#1A2430" style="padding:22px 24px 19px 24px;">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr>
<td>
<div style="{FONT}font-size:11px;font-weight:bold;letter-spacing:2.5px;color:#F36F21;text-transform:uppercase;">{esc(CONFIG['kicker'])}</div>
<div style="{FONT}font-size:26px;font-weight:bold;color:#FFFFFF;padding-top:8px;">Ampol Rigging &amp; Lifting Register - The Position</div>
<div style="{FONT}font-size:13px;color:#A7B6C4;padding-top:7px;">{esc(CONFIG['project'])}</div>
</td>
<td width="185" align="right" style="vertical-align:top;">
<div style="{FONT}font-size:11px;font-weight:bold;letter-spacing:1.5px;color:#FFFFFF;">POWERED BY <span style="color:#F36F21;">SITEIQ</span></div>
<div style="{FONT}font-size:11.5px;color:#8395A6;padding-top:5px;">Equipped for anything</div>
</td></tr></table>
<div style="{FONT}font-size:11px;color:#8395A6;padding-top:10px;line-height:1.6;">Generated: <b style="color:#FFFFFF;">{esc(gen_s)}</b> &nbsp;|&nbsp; Data as at: <b style="color:#FFFFFF;">{esc(asat_s)}</b> {esc(CONFIG["asat_note"])} &nbsp;|&nbsp; Author: <b style="color:#FFFFFF;">Andrew Fisher</b></div>
</td></tr></table>""")
    if card_cid:
        parts.append(card_block(card_cid))

    # nested f-strings reusing the outer quote break on older Pythons on
    # the store laptops - build the fragments first, drop them in after
    fill_s = num(fill_n) + " of " + num(rows_n)
    found_s = num(len(d["found"])) + " found (" + str(d["found_pct"]) + "%)"
    detail_s = "Full detail in the PDF attached." if pdf_ok else \
        "Full detail in the report HTML in today&rsquo;s Rigging folder (no PDF engine on this machine)."
    parts.append(sh.ecallout(
        f'<span style="color:#D95F14;font-weight:bold;text-transform:uppercase;">'
        f'The position.</span> {sh.eo(num(rows_n) + " register rows")}, '
        f'{num(distinct)} distinct barcodes, joined to the SiteIQ register as at '
        f'{esc(asat_s)}: {sh.eo(found_s)}. {sh.eo(num(oh_n) + " on hire to customers")} '
        f'across {num(n_co)} companies, <b>{num(st_n)}</b> in the store, '
        f'<b>{num(rp_n)}</b> at repairs or quarantined, and '
        f'{sh.eo(num(ms_n) + " not found in SiteIQ - whereabouts unknown")}. '
        f'Test records: {sh.eo(fill_s)} rows have any test detail entered. '
        f'{detail_s}'))

    parts.append(sh.etiles([
        (num(distinct), "DISTINCT BARCODES", f"{num(rows_n)} register rows", "#8A9AAC"),
        (num(oh_n), "ON HIRE TO CUSTOMERS", f'across {num(n_co)} companies', "#EFA82B"),
        (num(st_n), "IN THE STORE", "available for hire", "#22C55E"),
        (num(ms_n), "NOT FOUND IN SITEIQ", "whereabouts unknown", "#F0603E"),
    ]))
    parts.append(sh.etiles([
        (num(rp_n), "AT REPAIRS / QUARANTINED", "custody lines, not customer hire", "#EFA82B"),
        (f"{num(fill_n)}/{num(rows_n)}", "TEST RECORDS ENTERED",
         "fill-in not started" if not fill_n else "fill-in under way",
         "#F0603E" if not fill_n else "#EFA82B"),
        (num(len(d["dups"])), "DUPLICATED BARCODES",
         f'on {num(d["dup_rows"])} rows · {num(len(d["blank"]))} blank', "#8A9AAC"),
        (num(len(d["kw"])), "POSSIBLE RIGGING GEAR OFF-REGISTER",
         "keyword match - verify", "#8A9AAC"),
    ]))

    cap = 8
    comp = d["companies"]
    crows = []
    for c in comp[:cap]:
        o = c["oldest"]
        crows.append([esc(c["co"]), num(c["n"]), num(c["hirers"]),
                      (f'{esc(o["desc"])} ({esc(o["barcode"])}) &middot; {hl(o["hirer"])}' if o else "&ndash;"),
                      (f'{num(o["held"])}d' if o and o.get("held") is not None else "&ndash;")])
    parts.append(sh.esect("Where the gear is - customer companies by holding"))
    parts.append(sh.edtable(["Company", "Items", "Hirers", "Longest-held item", "Held"],
                            crows, ["", "r", "r", "", "r"]))
    tail = (f'Showing {cap} of {len(comp)} companies. ' if len(comp) > cap else '')
    parts.append(sh.enote(
        tail + f'Project accounts (FCCU, SATGAS/MOL, Ampol Refineries) are merged into their '
        f'company. {num(rp_n)} items on repairs / quarantine custody lines are excluded from '
        f'this table.'))

    cap2 = 8
    lrows = [[esc(it["company"]) or "&ndash;", hl(it["hirer"]) or "&ndash;",
              esc(it["desc"]) or "&ndash;", esc(it["barcode"]) or "&ndash;",
              it["since"].strftime("%d %b %Y"),
              f'{num(it["held"])}d' if it.get("held") is not None else "&ndash;"]
             for it in d["oh_longest"][:cap2]]
    parts.append(sh.esect("Longest held - oldest first"))
    parts.append(sh.edtable(["Company", "Hirer", "Item", "Barcode", "Since", "Held"],
                            lrows, ["", "", "", "", "r", "r"]))
    if len(d["oh_longest"]) > cap2:
        parts.append(sh.enote(f'Showing {cap2} of {len(d["oh_longest"])} dated '
                              f'customer-hire items - full list in the report.'))

    ms_oh = d["missing_snap"].get("On Hire", 0)
    parts.append(sh.esect("What needs a decision"))
    ex_line = ""
    if ex["present"] and ex["next_due"]:
        top_due = ex["next_due"].most_common(1)[0][0]
        ex_line = (f' A certificate extract in the workbook carries Next Insp Due '
                   f'{esc(top_due)} for {num(ex["on_register"])} register barcodes, not yet on the '
                   f'Master - if that means April 2026 they were due before this report; '
                   f'status unconfirmed.')
    parts.append(sh.ecallout(
        f'<b>{num(ms_n)} register barcodes are not returned by SiteIQ</b> - {num(ms_oh)} of them '
        f'were last seen on hire in the workbook&rsquo;s June snapshot, the rest in the store. '
        f'They are listed barcode by barcode in the report and are not counted as accounted '
        f'for. <b>Test records: {num(fill_n)} of {num(rows_n)}</b> rows carry any test '
        f'detail.{ex_line}', tight=True))

    parts.append(sh.enote(
        'Real data only - a blank register cell renders as a dash with the fill-in noted, '
        'and a barcode SiteIQ does not return is printed as unknown, never assumed. Rigging '
        'gear is Life Saving Rule 5 territory; the record has to be real.'))

    team_line = " &middot; ".join(
        f'<b style="color:#16202C;">{esc(p["name"])}</b> '
        f'<span style="color:#8A9AAC;">{esc(p["role"])}</span>'
        for p in CONFIG["team"])
    parts.append(f"""<table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="margin-top:26px;">
<tr><td style="border-top:1px solid #E4E8EC;padding-top:10px;">
<div style="{FONT}font-size:10px;font-weight:bold;letter-spacing:2px;color:#F36F21;text-transform:uppercase;">Your Coates Tool Store Team</div>
<div style="{FONT}font-size:11px;color:#8A9AAC;padding-top:5px;line-height:1.7;">{team_line}</div>
<div style="{FONT}font-size:10px;color:#98A6B4;padding-top:9px;line-height:1.7;">
Coates Hire &middot; Source: {esc(live_name)} (SiteIQ pull requested {esc(asat_s)}) for status, holder, on-hire date and storage unit, joined on ITEM_BARCODE to {esc(src_name)} (register membership, test columns, certificate extract). Blanks shown as dashes, never guessed. The Coates Way - consistent execution, every day. <b style="color:#16202C;">POWERED BY SITEIQ</b></div>
</td></tr></table>""")

    body = "".join(
        f'<tr><td style="padding:0 24px;">{p}</td></tr>'
        if "bgcolor=\"#1A2430\"" not in p[:120] else f'<tr><td>{p}</td></tr>'
        for p in parts)
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<title>Ampol Rigging &amp; Lifting Register - {esc(asat_s)}</title></head>
<body bgcolor="#EEF1F4" style="margin:0;padding:0;background-color:#EEF1F4;">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0" bgcolor="#EEF1F4">
<tr><td align="center" style="padding:18px 10px;">
<table role="presentation" width="{W}" cellspacing="0" cellpadding="0" bgcolor="#FFFFFF" style="width:{W}px;max-width:{W}px;">
<tr><td height="6" bgcolor="#F36F21" style="font-size:0;">&nbsp;</td></tr>
{body}
<tr><td height="24" style="font-size:0;">&nbsp;</td></tr>
</table></td></tr></table></body></html>"""


# =====================================================================
# MAIN
# =====================================================================

def main():
    print("=" * 68)
    print("COATES RIGGING & LIFTING REGISTER - HOUSE-STYLE REPORT (K2 look)")
    print("=" * 68)
    src = eng.find_workbook(["Rigging Register*.xlsx", "Rigging*Register*.xlsx"],
                            sys.argv[1] if len(sys.argv) > 1 else None)
    if not src:
        sys.exit("ERROR: no Rigging Register*.xlsx in the suite's Data "
                 "folder - save the register there and run again.")
    live_path = eng.find_workbook(["RENTAL_STOCK*.xlsx"],
                                  sys.argv[2] if len(sys.argv) > 2 else None)
    if not live_path:
        sys.exit("ERROR: no RENTAL_STOCK.xlsx in the suite's Data folder - "
                 "download the SiteIQ rental stock export (12_PULL_SITEIQ_EXPORTS) "
                 "and run again. Without it the report cannot say where the gear is.")
    print(f"Rigging register     : {src}")
    print(f"SiteIQ export        : {live_path}")

    reg_rows, snap, snap_max, extract = load_register(src)
    live_rows, asat_dt, asat_note = load_live(live_path)
    CONFIG["asat_note"] = asat_note
    asat_s = asat_dt.strftime("%d %b %Y %H:%M")
    gen_dt = datetime.now()
    gen_s = gen_dt.strftime("%d %b %Y %H:%M")
    d = derive(reg_rows, snap, snap_max, extract, live_rows, asat_dt)
    # WHY (03 Sep 2026): the transaction log (txn_insights) for the register
    # barcodes SiteIQ returns - issues and returns week by week, who holds
    # the gear, and the log's own quality counts. Read once per build.
    log = ti.load_all()
    scope = {it["barcode"] for it in d["found"]}
    d["log"] = {"weekly": ti.weekly_series(log, scope), "holders": ti.holders(log, scope, top=10),
                "dq": ti.data_quality(log, scope), "window": log["tx_window"], "scope_n": len(scope)}

    print(f"Data as at           : {asat_s}  (RENTAL_STOCK request time)")
    print(f"Register rows        : {d['rows']:,}  "
          f"({' | '.join(f'{c} {n:,}' for c, n in d['cats_rows'].most_common())})")
    print(f"Distinct barcodes    : {d['distinct']:,}  (blank {len(d['blank'])}, "
          f"{len(d['dups'])} barcodes repeated on {d['dup_rows']} rows, "
          f"{d['lower']} lower-case)")
    print(f"Found in SiteIQ      : {len(d['found']):,} of {d['distinct']:,} "
          f"({d['found_pct']}%) - exact match {d['join_exact']:,}, "
          f"upper-cased {len(d['found']) - d['join_exact']:,}")
    print(f"On hire / store      : {len(d['onhire']):,} to customers across "
          f"{len(d['companies']):,} companies | {len(d['store']):,} in the store | "
          f"{len(d['repair']):,} at repairs/quarantined"
          + (f" | {len(d['other']):,} other status" if d["other"] else ""))
    print(f"Not found in SiteIQ  : {len(d['missing']):,}  "
          f"(June snapshot said: "
          f"{' | '.join(f'{k} {v}' for k, v in d['missing_snap'].most_common())})")
    print(f"Test records         : {len(d['with_test']):,} of {d['rows']:,} rows "
          f"have any test detail entered")
    if d["extract"]["present"]:
        ex = d["extract"]
        print(f"Certificate extract  : {ex['rows']:,} rows, {ex['distinct']:,} barcodes, "
              f"{ex['on_register']:,} on the register - Next Insp Due "
              f"{' | '.join(f'{v} x{n}' for v, n in ex['next_due'].most_common(3))}")
    print(f"Keyword leads        : {len(d['kw']):,} live lines "
          f"({len(d['kw_groups'])} descriptions) look like rigging gear off the register")
    if d["oh_longest"]:
        o = d["oh_longest"][0]
        print(f"Longest held         : {o['desc']} ({o['barcode']}) - "
              f"{o['company']} / {o['hirer']} since "
              f"{o['since'].strftime('%d %b %Y')} ({o['held']} days)")
    L = d["log"]
    wk, ho, dq = L["weekly"], L["holders"], L["dq"]
    print(f"The year in movements: {len(wk)} weeks, {sum(w['issues'] for w in wk):,} issues / "
          f"{sum(w['returns'] for w in wk):,} returns of {L['scope_n']} register barcodes; "
          f"{ho['n80_items']} of {ho['holders']} holders carry 80% of {ho['items']} on hire; "
          f"log check: {dq['short_n']} closed inside 6 min, {len(dq['onhire_no_log'])} on hire with no log, "
          f"{len(dq['onhire_before_log'])} issued before the log")

    # ---- 1. the PDF -----------------------------------------------------
    css_path = BASE / "k2style.css"
    if not css_path.exists():
        sys.exit("ERROR: k2style.css is missing from the suite folder - the "
                 "house-style PDF cannot render without it.")
    css = css_path.read_text(encoding="utf-8")
    print("-" * 68)
    print("[1/2] Rigging register PDF (house style)...")
    # the cover: the one number of the day and three true lines under it -
    # wearing the SAME status as the band on the position page, and saying
    # how old the data was when the pack was built
    rag = rig_rag(d)
    key_value, key_label = num(len(d["missing"])), "register items SiteIQ cannot see"
    cover_lines = [
        f'<b>{num(len(d["onhire"]))}</b> on hire to customers across <b>{num(len(d["companies"]))}</b> companies',
        f'<b>{num(len(d["store"]))}</b> in the store, available for hire',
        f'<b>{num(len(d["repair"]))}</b> at repairs or quarantined',
    ]

    def cover_for(contents):
        return sh.cover_page(CONFIG, key_value, key_label, cover_lines, gen_s, asat_s,
                             rag=rag["status"], fresh=sh.freshness_line(asat_dt, gen_dt), contents=contents)
    pages = build_pages(d, asat_s)
    pdf_path = OUT / CONFIG["pdf_name"]
    # WHY (03 Sep 2026): two prints - the first to read the section page
    # numbers off the printed pack, the second with them on the cover
    doc_first, _ = render_doc(pages, cover_for(None), gen_s, asat_s)
    contents, n_first = cover_contents_from_print(doc_first, pdf_path, OUT / CONFIG["pdf_html"], css,
                                                  CLOSING_HEADINGS)
    doc, n_pages = render_doc(pages, cover_for(contents), gen_s, asat_s)
    pdf_ok, layout_ok = render_k2_pdf(doc, pdf_path, n_pages, css)
    if pdf_ok and n_first is not None:
        n_second = pdf_pages(pdf_path)
        print(f"Cover contents       : {len(contents)} rows - page numbers read off the first print "
              f"({n_first} pages); the second print has {n_second} pages - "
              f"{'the same pagination' if n_first == n_second else 'NOT THE SAME'}")
        if n_first != n_second:
            print("*" * 68)
            print("WARNING: the pack paginated differently once the cover carried its contents -")
            print("         the page numbers on the cover cannot be trusted. Do not send as is.")
            print("*" * 68)
            layout_ok = False
    if pdf_ok:
        print("PDF finish           : " + pdf_finish.finish(
            pdf_path, f"{CONFIG['client']} {CONFIG['title']} - as at {asat_s}",
            "Where the rigging and lifting register gear is at the SiteIQ pull - on hire by "
            "company, in the store, at repairs, not found in SiteIQ, and the test-record truth.",
            doc, keywords="rigging, lifting, register", has_cover=True, family="Rigging"))

    # ---- the phone card: the position-page tiles, the band, four scores --
    card_path = OUT / CONFIG["card_name"]
    found_n, rows_n = len(d["found"]), d["rows"]
    scores = [("Found in SiteIQ", d["found_pct"]),
              ("Test records filled in", round(len(d["with_test"]) / rows_n * 100) if rows_n else 0),
              ("Customer hire share of found", round(len(d["onhire"]) / found_n * 100) if found_n else 0),
              ("Repairs share of found", round(len(d["repair"]) / found_n * 100) if found_n else 0)]
    card_tiles = [(shown, lab, plain(note), NOTE_HEX.get(ncls, "#8A9AAC"))
                  for _, shown, lab, note, ncls, _ in rig_tiles(d)]
    sh.position_card_png(CONFIG, asat_s, card_tiles,
                         (rag["status"], rag["card_headline"], CONFIG["rag_owner"], rag["card_action"]),
                         scores, str(card_path),
                         foot=f"Counted from the SiteIQ pull of {asat_s} joined to the register - "
                              f"nothing estimated.")
    print(f"Position card        : {card_path}")

    # ---- 2. the email (draft - never sends) -----------------------------
    print("[2/2] Outlook email (house style, draft only)...")
    # WHY (03 Sep 2026): the .eml shows the position card inline under the
    # header (a cid part) and still carries it as a file; the native-draft
    # manifest lists it as an attachment only, so its body is written
    # without the inline image.
    body_html = build_email_html(d, gen_s, asat_s, pdf_ok,
                                 Path(src).name, Path(live_path).name)
    (OUT / CONFIG["email_html"]).write_text(body_html, encoding="utf-8")
    html = build_email_html(d, gen_s, asat_s, pdf_ok, Path(src).name, Path(live_path).name,
                            card_cid="positioncard" if card_path.exists() else "")
    msg = EmailMessage()
    subject = (f"Ampol Tool Store - Rigging & Lifting Register Report - "
               f"as at {asat_dt.strftime('%d/%m/%Y %H:%M')}")
    msg["Subject"] = subject
    msg["To"] = eng.STAFF_EMAIL_TO
    msg["Date"] = formatdate(localtime=True)
    msg["X-Unsent"] = "1"
    msg.set_content("This report is best viewed in HTML. The rigging and "
                    "lifting register PDF is attached.\n" if pdf_ok else
                    "This report is best viewed in HTML. No PDF could be "
                    "rendered on this machine - the report pages are in the "
                    "day's Rigging folder as HTML.\n")
    msg.add_alternative(html, subtype="html")
    if card_path.exists():
        with open(card_path, "rb") as f:
            msg.get_payload()[1].add_related(f.read(), maintype="image", subtype="png",
                                             cid="<positioncard>", filename=card_path.name,
                                             disposition="inline")
    attach = [str(pdf_path)] if pdf_ok and pdf_path.exists() else []
    if card_path.exists():
        attach.append(str(card_path))      # the phone card rides beside the PDF
    for p in attach:
        with open(p, "rb") as f:
            if p.lower().endswith(".png"):
                msg.add_attachment(f.read(), maintype="image", subtype="png",
                                   filename=os.path.basename(p))
            else:
                msg.add_attachment(f.read(), maintype="application",
                                   subtype="pdf", filename=os.path.basename(p))
    eml_path = OUT / CONFIG["eml_name"]
    with open(eml_path, "wb") as f:
        f.write(msg.as_bytes())
    print(f"EML written          : {eml_path}  "
          f"({os.path.getsize(eml_path):,} bytes; attached: "
          + (", ".join(os.path.basename(p) for p in attach) if attach else "NO PDF") + ")")
    # manifest so MAKE_OUTLOOK_DRAFTS keeps working - recipients derive
    # from the engine's STAFF_EMAIL_TO, one source of truth.
    to_line = "; ".join(re.findall(r"<([^>]+)>", eng.STAFF_EMAIL_TO))
    (OUT / CONFIG["draft_json"]).write_text(json.dumps({
        "subject": subject,
        "to": to_line,
        "body": CONFIG["email_html"],
        "attachments": [os.path.basename(p) for p in attach],
    }, indent=1), encoding="utf-8")

    # ---- the scoreboard: today's figures, keyed on the pull day ----------
    # A re-run on the same pull replaces the day's entry; the next pull
    # reads it back and prints the movement. Real recorded days only.
    oldest = d["oh_longest"][0]["held"] if d["oh_longest"] else None
    # WHY (03 Sep 2026): the entry carries the position in words as well as
    # figures (extra) - the status, the headline, the rule, the owner, the
    # dated action and the cover number - so the daily position page can
    # quote the report without opening it.
    hist = rh.record("rigging", asat_dt, {
        "rows": d["rows"], "distinct": d["distinct"], "found": len(d["found"]),
        "not_found": len(d["missing"]), "on_hire": len(d["onhire"]),
        "in_store": len(d["store"]), "repairs": len(d["repair"]),
        "companies": len(d["companies"]), "tests_filled": len(d["with_test"]),
        "oldest_days": oldest},
        extra={"rag": rag["status"], "headline": plain(rag["headline"]), "rule": plain(rag["rule"]),
               "owner": CONFIG["rag_owner"], "action": rag["card_action"], "due": rag["due"],
               "key_value": key_value, "key_label": key_label,
               "second_value": num(len(d["onhire"])), "second_label": "on hire to customers",
               "title": f"{CONFIG['client']} {CONFIG['title']}", "folder": "Rigging",
               "pdf": CONFIG["pdf_name"], "card": CONFIG["card_name"]})
    prev = rh.previous("rigging", "not_found", asat_dt)
    print(f"History              : {asat_dt:%d %b %Y} figures written to {hist.parent.name}/{hist.name}"
          + (f" - movement shown against {prev[0]:%d %b %Y}" if prev
             else " - first day on record; movement notes start with the next pull"))
    print("")
    print(f"NEXT STEP: double-click the .eml in {OUT}, check it, press Send.")
    print("Done. The Coates Way - consistent execution, every day.")
    if not pdf_ok:
        sys.exit("\nWARNING: finished without a PDF - see above. The HTML and "
                 "the email draft are written; the bat reports this run as "
                 "incomplete on purpose.")
    if not layout_ok:
        sys.exit("\nWARNING: the PDF failed its layout check - see above. Do "
                 "not send it as is.")


if __name__ == "__main__":
    # Failures leave a nonzero exit so the bat button tells the truth;
    # the bat owns the end-of-run pause - no input() here.
    try:
        main()
    except Exception as e:
        import traceback
        traceback.print_exc()
        sys.exit(f"\nERROR: {e}")
