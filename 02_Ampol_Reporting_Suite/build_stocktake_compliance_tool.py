#!/usr/bin/env python3
"""
Coates Stocktake Compliance Tool V3 - The Coates Way
Author: Andrew Fisher / POWERED BY SITEIQ

From the SiteIQ STOCKTAKE export (plus RENTAL_STOCK, the pricing workbook
and the corrected-descriptions workbook), one run produces:

 1. STAFF COUNT WORKLIST (PDF+HTML) - sectioned the way the store works:
      0. Possible missed returns (resolve first)
      1. Gas monitors            (Priority 1 - 7-day target cycle)
      2. Radios & radio batteries(Priority 1 - 7-day target cycle)
      3. Milwaukee power tools   (Priority 2 - 14-day target cycle)
      4. Store walk - all other tooling (Priority 3 - 30-day SOP cycle),
         grouped by storage unit so the team clears a bay at a time
      5. On-hire verification by company (verify on return, don't hunt)
 2. STAFF EXCEL WORKLIST (.xlsx) - same sections as sortable/filterable
    tabs for tablet use, plus a Pricing Gaps tab
 3. CLIENT COMPLIANCE REPORT (PDF+HTML) - proof of compliance with the
    priority determination stated, dollarised coverage, and what has
    been completed (activity in the last 7/30 days)
 4. STAFF EMAIL (.eml) - opens in Outlook, PDFs + Excel attached

DATA JOINS
 - Avg Buy Price (New): pricing workbook matched on ITEM_DESCRIPTION
   (case-insensitive, whitespace-collapsed; fallback via the corrected
   description and with CALTEX/AMPOL site prefixes stripped).
   Duplicate pricing lines resolved by the most common value; conflicts
   are listed on the Pricing Gaps tab. No price is ever invented.
 - Corrected Description: corrections workbook matched on ITEM_BARCODE.
 - System status / hirer / home storage unit: RENTAL_STOCK matched on
   ITEM_BARCODE.

PRIORITY DETERMINATION (stated on every output)
 - P1 GAS MONITORS + RADIOS/BATTERIES: safety-critical, high-value,
   high-churn. Target: sighted every 7 days (bump/charge cycle covers
   monitors; idle units surface here fast).
 - P2 MILWAUKEE: high-value, high-shrinkage-risk power tools.
   Target: sighted every 14 days.
 - P3 ALL OTHER TOOLING: 30-day full-coverage SOP cycle.
 - Movement is the mechanism: issue/return scans reset the clock, so
   the worklist naturally surfaces IDLE stock - exactly the gear that
   goes missing quietly.
 - Severity vs SOP is unchanged: OK <=30 days, DUE 31-60, CRITICAL over 60
   or never sighted. Client compliance is measured on the 30-day SOP;
   the tighter P1/P2 cycles are our internal standard above the SOP.
 - Departed stock (Pending Branch Receipt, or a Departure scan with the
   item no longer on the live register) is excluded.
 - On-hire items verified via return rescans and shutdown checks.
"""
import sys, re
from pathlib import Path
from datetime import datetime, date
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import ampol_paths
import ampol_names

BASE = Path(__file__).resolve().parent
# WHY (12 Aug 2026): outputs land in the suite's dated Reports area now -
# Reports\<today>\Stocktake - one folder per day, nothing silently
# overwritten. Fixed filenames are fine; the date lives in the folder.
OUT = Path(ampol_paths.day_folder("Stocktake"))
NOW = datetime.now()

COATES_PURPOSE = "Supporting Australia's growth with leading equipment solutions"
COATES_OBJECTIVE = "Australia's most trusted equipment partner \u2014 delivering Best Service & Value"
COATES_VALUES = "Care Deeply &nbsp;\u2022&nbsp; Customer Focused &nbsp;\u2022&nbsp; Be Our Best &nbsp;\u2022&nbsp; One Team &nbsp;\u2022&nbsp; Competitive Spirit"

PRIORITY_PREFIXES = ["AMPMR/", "AMPMRB", "AMP037", "SATGASRAD", "SATGASBAT",
                     "AMPMAINGAS001", "AMPT&I001", "CTX011"]
ONHIRE_UNITS = {"ON HIRE", "SHUTDOWN", "TURNAROUND"}
STAFF_EMAIL_TO = '"Ampol Store" <Ampolstore@coates.com.au>, "Mitchell, Cody" <Cody.mitchell@coates.com.au>'

# Priority tiers: key -> (label, target cycle days, rationale)
TIERS = {
    "gas":       ("Gas Monitors",             7,  "Safety-critical; bumped & charged before issue"),
    "radio":     ("Radios & Radio Batteries", 7,  "Safety-critical comms; high churn"),
    "milwaukee": ("Milwaukee Power Tools",    14, "High value, high shrinkage risk"),
    "general":   ("All Other Tooling",        30, "30-day full-coverage SOP cycle"),
}

# ---------------------------------------------------------------- utilities
def parse_dt(x):
    if isinstance(x, datetime): return x
    if isinstance(x, date): return datetime.combine(x, datetime.min.time())
    s = str(x or "").strip()
    for fmt in ["%d/%m/%Y %I:%M %p", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]:
        try: return datetime.strptime(s, fmt)
        except ValueError: pass
    return None

def norm(s):
    s = re.sub(r"\s+", " ", str(s or "").strip().upper())
    return s.translate(str.maketrans("\u00c4\u00d6\u00dc", "AOU"))

def strip_site_prefix(s):
    return re.sub(r"^(?:(?:CALTEX|AMPOL)\s+)+", "", s)

# WHY (02 Sep 2026): the gas tier used to sweep in 177 single chargers, 4
# telescopic probes and the charger joiners because their descriptions carry
# "DRAGER" / "X-AM". The client then saw 1,059 "gas monitors" here against 878
# on the gas monitor report, and the daily-three tile scored the chargers as
# monitors missed. Same exclusion rule as gasmon_engine.NOT_GAS_RE: an
# accessory is general tooling on the 30-day SOP cycle.
GAS_ACCESSORY_RE = re.compile(r"CHARGER|PROBE|PUMP|CALIBRATION GAS|DOCK|CRADLE|JOINER")

def categorise(barcode, desc):
    b = str(barcode or "").upper()
    d = norm(desc)
    if b.startswith(("AMPMRB", "AMP037", "SATGASBAT", "AMPMR/", "SATGASRAD")) or "RADIO" in d:
        return "radio"
    if (b.startswith(("AMPMAINGAS001", "AMPT&I001", "CTX011"))
            or "DRAGER" in d or "DR\u00c4GER" in d or "DRAEGER" in d or "X-AM" in d or "GAS MONITOR" in d):
        if GAS_ACCESSORY_RE.search(d):
            return "general"
        return "gas"
    if "MILWAUKEE" in d or re.search(r"\bM1[28]\b", d):
        return "milwaukee"
    return "general"

def bucket(days):
    """Severity against the 30-day SOP (unchanged from V2)."""
    if days is None: return "never"
    if days <= 30: return "ok"
    if days <= 60: return "due"
    return "critical"

def severity_label(days):
    b = bucket(days)
    return {"ok": "TARGET DUE", "due": "SOP DUE", "critical": "CRITICAL", "never": "CRITICAL"}[b]

# WHY (12 Aug 2026): the em dash lives in a constant so f-string expressions
# can use it - a \u escape inside an f-string expression needs Python 3.12,
# and this way the suite also runs on a laptop still carrying 3.11.
DASH = "\u2014"

def money(v):
    return f"${v:,.0f}" if v is not None else DASH

# ---------------------------------------------------------------- loaders
def load_pricing(path):
    """norm(desc) -> Avg Buy Price (New). Duplicates -> most common value.
    Returns (exact_map, stripped_map, conflicts)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c else "" for c in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    seen = defaultdict(Counter)
    for r in it:
        d = norm(r[ix["ITEM_DESCRIPTION"]])
        p = r[ix["Avg Buy Price (New)"]]
        if d and isinstance(p, (int, float)) and p > 0:
            seen[d][round(float(p), 2)] += 1
    exact, conflicts = {}, []
    for d, prices in seen.items():
        if len(prices) > 1:
            conflicts.append((d, sorted(prices.keys())))
        # WHY (02 Sep 2026): most_common() breaks a 1-vs-1 tie by file order,
        # which is luck, not a rule. Ties now take the LOWER value - the
        # conservative figure - and the Pricing Gaps tab says so.
        top = max(prices.values())
        exact[d] = min(v for v, n in prices.items() if n == top)
    stripped = {}
    for d, p in exact.items():
        sd = strip_site_prefix(d)
        if sd and sd not in exact and sd not in stripped:
            stripped[sd] = p
    return exact, stripped, conflicts

# WHY (02 Sep 2026): 493 gas monitors are registered with the serial in the
# description ("Drager X-am 5000 - T&I -ARSN-0123", "... - Maintenance-ARUB-0045")
# so they never matched the pricing master and about $986,000 of monitor fleet
# fell out of every value figure. When a description ends in a serial token,
# the family ("DRAGER X-AM 5000") is priced from the master's own line for
# that family (Drager X-am 5000 Gas Monitor). Nothing is estimated: the price
# is the master's, applied to the same product. Every output discloses it.
SERIAL_SUFFIX_RE = re.compile(r"\s*-?\s*[A-Z]{4}-\d{3,5}\s*$")
FAMILY_QUALIFIER_RE = re.compile(r"\s*-\s*(T&I|MAINTENANCE|MAINT)\s*$")

def family_keys(desc):
    """Candidate pricing keys for a serial-suffixed description; [] otherwise."""
    s = norm(desc)
    if not s or not SERIAL_SUFFIX_RE.search(s):
        return []
    base = FAMILY_QUALIFIER_RE.sub("", SERIAL_SUFFIX_RE.sub("", s)).strip(" -")
    if not base:
        return []
    return [base + " GAS MONITOR", base]

def price_for(raw_desc, fixed_desc, exact, stripped):
    for key in (norm(raw_desc), norm(fixed_desc),
                strip_site_prefix(norm(raw_desc)), strip_site_prefix(norm(fixed_desc))):
        if key and key in exact: return exact[key]
        if key and key in stripped: return stripped[key]
    for key in family_keys(raw_desc) + family_keys(fixed_desc):
        if key in exact: return exact[key]
        if key in stripped: return stripped[key]
    return None

def priced_by_family(raw_desc, fixed_desc, exact, stripped):
    """True when the price came from the family rule, not an exact match."""
    for key in (norm(raw_desc), norm(fixed_desc),
                strip_site_prefix(norm(raw_desc)), strip_site_prefix(norm(fixed_desc))):
        if key and (key in exact or key in stripped):
            return False
    return any(k in exact or k in stripped for k in family_keys(raw_desc) + family_keys(fixed_desc))

def load_corrections(path):
    """Returns (bc_map, desc_map):
    - bc_map: ITEM_BARCODE -> Corrected Description (exact, wins)
    - desc_map: norm(old ITEM_DESCRIPTION) -> Corrected Description, only
      where unambiguous (a single corrected value). Ambiguous old
      descriptions (e.g. serial/size-specific corrections) stay
      barcode-only so the wrong correction is never applied.
    Rows flagged Changed? = No are skipped; blank Changed? is accepted."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c else "" for c in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    bc_map, desc_votes = {}, defaultdict(set)
    for r in it:
        old = str(r[ix["ITEM_DESCRIPTION"]] or "").strip()
        bc = str(r[ix["ITEM_BARCODE"]] or "").strip()
        new = str(r[ix["Corrected Description"]] or "").strip()
        changed = str(r[ix["Changed?"]] or "").strip().upper() if "Changed?" in ix else ""
        if not new or changed == "NO":
            continue
        if bc:
            bc_map[norm(bc)] = new
        if old:
            desc_votes[norm(old)].add(new)
    desc_map = {d: next(iter(v)) for d, v in desc_votes.items() if len(v) == 1}
    return bc_map, desc_map

def load_master(path):
    """ITEM_BARCODE -> (status, company, hirer, home unit, on-hire date)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["RENTAL_STOCK"]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c else "" for c in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    m = {}
    for r in it:
        if r[ix["ITEM_BARCODE"]]:
            m[str(r[ix["ITEM_BARCODE"]]).strip().upper()] = (
                str(r[ix["ITEM_STATUS"]] or "").strip(),
                str(r[ix["COMPANY_NAME"]] or "").strip(),
                str(r[ix["HIRER_NAME"]] or "").strip(),
                str(r[ix["STORAGE_UNIT"]] or "").strip(),
                parse_dt(r[ix["ON_HIRE_DATE"]]))
    return m

def load(path, master, fixes, exact, stripped):
    bc_map, desc_map = fixes
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    export_dt = None
    if "REFERENCE_INFO" in wb.sheetnames:
        for r in wb["REFERENCE_INFO"].iter_rows(min_row=2, max_row=2, values_only=True):
            export_dt = parse_dt(r[5]) or parse_dt(r[3])
    export_dt = export_dt or NOW
    ws = wb["STOCKTAKE"]
    it = ws.iter_rows(values_only=True)
    hdr = [c for c in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    rows, transit = [], 0
    for r in it:
        if not r[ix["TOOL_STORE"]]: continue
        status = str(r[ix["SIGHTED_STATUS"]] or "").upper()
        action = str(r[ix["LAST_SIGHTED_ACTION"]] or "").upper()
        bc = str(r[ix["LATEST_BARCODE"]] or "").strip()
        # WHY (02 Sep 2026): "last action = Departure" alone dropped 17 items
        # the live register still shows on the shelf (RIGGING BAY, departed
        # 10 Mar). Departed stock is only out of scope when the register has
        # let it go too; with no register on hand the old rule still applies.
        on_register = bool(master) and bc.upper() in master
        if status == "PENDING BRANCH RECEIPT" or (action == "DEPARTURE" and not on_register):
            transit += 1; continue
        d = parse_dt(r[ix["LAST_SIGHTED_DATE_TIME"]])
        days = (export_dt - d).days if d else None
        su = str(r[ix["STORAGE_UNIT"]] or "(no storage unit)").strip()
        raw_desc = str(r[ix["DESCRIPTION"]] or "").strip()
        fixed = bc_map.get(norm(bc)) or desc_map.get(norm(raw_desc))
        desc = fixed or raw_desc
        qty = r[ix["SIGHTED_QUANTITY"]] or 1
        price = price_for(raw_desc, fixed, exact, stripped)
        by_family = price is not None and priced_by_family(raw_desc, fixed, exact, stripped)
        mstatus, mcomp, mhirer, mhome, m_ohd = master.get(bc.upper(), ("", "", "", "", None))
        onhire_sys = mstatus == "On Hire"
        home = mhome if mhome and mhome.upper() not in ONHIRE_UNITS else (su if su.upper() not in ONHIRE_UNITS else mhome or su)
        cat = categorise(bc, desc if fixed else raw_desc)
        # WHY (02 Sep 2026): the site is Ampol. SiteIQ still carries the
        # former name on thousands of descriptions; every page shows the
        # current name and the data note says how many lines still carry
        # the old one. Matching and pricing above used the raw text.
        desc = ampol_names.display_desc(desc)
        mcomp_show = ampol_names.display_company(mcomp) if mcomp else ""
        rows.append({"unit": su, "barcode": bc, "desc": desc, "raw_desc": raw_desc,
                     "former_name": ampol_names.carries_former_name(raw_desc),
                     "company": mcomp_show, "account": ampol_names.account_label(mcomp) if mcomp else "",
                     "hirer": mhirer,
                     "fixed": bool(fixed), "price_family": by_family,
                     "qty": qty, "price": price,
                     "value": (price * qty) if price is not None else None,
                     "last": d, "by": str(r[ix["LAST_SIGHTED_BY"]] or "").strip(),
                     "days": days, "cat": cat, "target": TIERS[cat][1],
                     "onhire": onhire_sys,
                     "onhire_to": (f"{mcomp_show} \u2013 {mhirer}" if mcomp or mhirer else "").strip(" \u2013"),
                     "home": home or "(no storage unit)",
                     "status": mstatus or "Not in master", "ohd": m_ohd})
    return rows, transit, export_dt

# ---------------------------------------------------------------- shared derivations
def derive(rows, export_dt):
    d = {}
    d["countable"] = len(rows)
    d["ok30"] = sum(1 for r in rows if bucket(r["days"]) == "ok")
    d["comp30"] = d["ok30"] / d["countable"] * 100 if d["countable"] else 0
    d["instore"] = [r for r in rows if not r["onhire"]]
    d["onhire"] = [r for r in rows if r["onhire"]]
    # target-cycle due: in-store and outside the item's tier target cycle
    d["due"] = {k: [r for r in d["instore"] if r["cat"] == k and (r["days"] is None or r["days"] > r["target"])]
                for k in TIERS}
    d["onhire_due30"] = [r for r in d["onhire"] if bucket(r["days"]) != "ok"]
    d["missed_returns"] = [r for r in rows if r["onhire"] and r["unit"].upper() not in ONHIRE_UNITS
                           and r["last"] and r["ohd"] and r["last"].date() > r["ohd"].date()]
    # values (Avg Buy Price New x qty; unpriced excluded and disclosed)
    d["val_total"] = sum(r["value"] for r in rows if r["value"])
    d["val_ok30"] = sum(r["value"] for r in rows if r["value"] and bucket(r["days"]) == "ok")
    d["val_onhire"] = sum(r["value"] for r in d["onhire"] if r["value"])
    d["priced_lines"] = sum(1 for r in rows if r["price"] is not None)
    d["priced_family"] = sum(1 for r in rows if r.get("price_family"))
    d["unpriced_lines"] = d["countable"] - d["priced_lines"]
    # WHY (02 Sep 2026): the SOP headline blends two very different things -
    # shelf items the team can count and on-hire items whose only "sighting"
    # is the hire-out scan. Both halves are kept so every page can print
    # the split instead of one blended number.
    d["ok30_instore"] = sum(1 for r in d["instore"] if bucket(r["days"]) == "ok")
    d["ok30_onhire"] = sum(1 for r in d["onhire"] if bucket(r["days"]) == "ok")
    d["late_onhire"] = sum(1 for r in d["onhire"] if bucket(r["days"]) != "ok")
    d["late_instore"] = sum(1 for r in d["instore"] if bucket(r["days"]) != "ok")
    d["crit_instore"] = sum(1 for r in d["instore"] if bucket(r["days"]) in ("critical", "never"))
    d["crit_onhire"] = sum(1 for r in d["onhire"] if bucket(r["days"]) in ("critical", "never"))
    d["not_on_register"] = [r for r in rows if r["status"] == "Not in master"]
    d["former_name_lines"] = sum(1 for r in rows if r.get("former_name"))
    d["awaiting_arrival"] = [r for r in rows if r["status"].upper() == "AWAITING ARRIVAL"]
    # activity - what's been done
    d["done7"] = [r for r in rows if r["days"] is not None and r["days"] <= 7]
    d["done30"] = [r for r in rows if r["days"] is not None and r["days"] <= 30]
    d["counters7"] = Counter(r["by"] for r in d["done7"] if r["by"]).most_common(6)
    # per-tier stats: coverage rated on IN-STORE items (on-hire items are
    # verified via return rescans/shutdown checks, not shelf counting)
    d["tier_stats"] = []
    for k, (label, target, why) in TIERS.items():
        tr = [r for r in rows if r["cat"] == k]
        ins = [r for r in tr if not r["onhire"]]
        wt = sum(1 for r in ins if r["days"] is not None and r["days"] <= target)
        w30 = sum(1 for r in ins if bucket(r["days"]) == "ok")
        d["tier_stats"].append((k, label, target, why, len(tr), len(tr) - len(ins),
                                len(ins), wt, w30, (w30 / len(ins) * 100) if ins else 100.0))
    return d

# ---------------------------------------------------------------- CSS / branding
CSS = """
@page { size: A4; margin: 12mm 11mm 16mm 11mm;
        @bottom-right { content: "Page " counter(page) " of " counter(pages); font-size:8pt; color:#666; }
        @bottom-left  { content: "__FOOT__"; font-size:8pt; color:#666; } }
body { font-family:Calibri, Arial, sans-serif; color:#1a1a1a; font-size:9.5pt; margin:0; }
h1 { font-size:17pt; margin:0 0 2pt; }
.purpose { color:#e07000; font-size:9pt; font-weight:bold; margin:0 0 2pt; }
.sub { color:#555; font-size:9pt; margin-bottom:8pt; }
.banner { border-left:5pt solid #e07000; padding-left:8pt; margin-bottom:8pt; }
.totals { width:100%; border-collapse:collapse; margin-bottom:8pt; }
.totals td { border:1pt solid #ccc; padding:6pt 8pt; }
.totals .lbl { font-size:7pt; text-transform:uppercase; letter-spacing:0.5pt; color:#666; }
.totals .val { font-size:12.5pt; font-weight:bold; color:#e07000; }
.totals .val.g { color:#1e7d32; } .totals .val.r { color:#c00000; } .totals .val.a { color:#b07700; }
.charge { border:1pt solid #f0d5b8; background:#fdf4ea; padding:7pt 10pt; margin-bottom:8pt; font-size:9pt; }
.charge .ntitle { font-weight:bold; color:#b35a00; font-size:9.5pt; margin-bottom:3pt; letter-spacing:0.5pt; }
.charge ul { margin:0; padding-left:14pt; } .charge li { margin:2pt 0; }
.section { page-break-before: always; }
h2.sec { color:#e07000; font-size:13pt; text-transform:uppercase; letter-spacing:0.5pt;
         border-bottom:2pt solid #e07000; padding-bottom:3pt; margin:0 0 4pt; }
.secsub { color:#555; font-size:9pt; margin-bottom:8pt; }
.cblock { page-break-inside: avoid; margin-bottom:9pt; }
.cblock.big { page-break-inside: auto; }
.cbar { background:#e07000; border-left:8pt solid #1a1a1a; padding:5pt 12pt; }
.cbar.red { background:#c00000; } .cbar.blue { background:#1f5c99; } .cbar.dark { background:#333; }
.cname { color:#fff; font-weight:bold; font-size:10.5pt; }
.cmeta { color:#fff; font-size:8pt; font-weight:bold; }
table.items, table.matrix { width:100%; border-collapse:collapse; font-size:8.5pt; border:1pt solid #ccc; }
th { text-align:left; font-size:7.5pt; text-transform:uppercase; color:#555;
     padding:3pt 8pt; border-bottom:1pt solid #ccc; background:#fafafa; }
td { padding:3pt 8pt; border-bottom:0.5pt solid #e3e3e3; vertical-align:top; }
td.num, th.num { text-align:right; }
.overdue { color:#c00000; font-weight:bold; }
.duec { color:#b07700; font-weight:bold; }
.okc { color:#1e7d32; font-weight:bold; }
.chip { display:inline-block; padding:1pt 7pt; border-radius:8pt; font-weight:bold; font-size:7.5pt; }
.hi { background:#fdf0f0; }
.notice { border:1pt solid #b8d4f0; background:#eef5fc; padding:7pt 10pt; margin-bottom:8pt; font-size:9pt; }
.notice .ntitle { font-weight:bold; color:#1f5c99; font-size:9.5pt; margin-bottom:3pt; letter-spacing:0.5pt; }
.foot { margin-top:8pt; color:#555; font-size:8pt; border-top:1pt solid #ccc; padding-top:5pt; }
.cway { margin-top:5pt; color:#e07000; font-size:8pt; font-weight:bold; text-align:center; }
.cway .obj { color:#555; font-weight:normal; display:block; margin-top:2pt; }
"""

LSR_LINE = ("Life Saving Rule 5 \u2013 Tools and Equipment (SEQ-GL-009): nothing damaged, "
            "defective or out of test date is ever reissued. Two scans. Two looks. One "
            "standard. Stock takes are done daily. No exceptions.")


def sig_footer():
    return f"""<div class="cway"><span style="font-size:8pt;color:#555">{LSR_LINE}</span><br>{COATES_VALUES}<span class="obj">{COATES_OBJECTIVE}</span>
<span class="obj"><b style="color:#1a1a1a">Author: Andrew Fisher</b> &nbsp;\u2022&nbsp; <b style="color:#e07000">POWERED BY SITEIQ</b></span></div>"""

def sev_chip(days):
    lbl = severity_label(days)
    style = {"TARGET DUE": ("color:#1f5c99;background:#eef5fc"),
             "SOP DUE":    ("color:#b07700;background:#fdf3dd"),
             "CRITICAL":   ("color:#c00000;background:#fbe9e9")}[lbl]
    return f'<span class="chip" style="{style}">{lbl}</span>'

def determination_panel():
    rows = "".join(
        f"""<tr><td><b>{label}</b></td><td>Every {t} days</td><td>{why}</td></tr>"""
        for k, (label, t, why) in TIERS.items())
    return f"""<div class="notice"><div class="ntitle">HOW PRIORITY IS DETERMINED (AND WHAT COUNTS AS DONE)</div>
<table class="matrix" style="margin-bottom:5pt"><tr><th>Priority Tier</th><th>Target Cycle</th><th>Why</th></tr>{rows}</table>
An item is <b>DONE</b> when it is physically scanned \u2014 the scan resets its clock. Issue and return scans count, so this
worklist naturally surfaces <b>idle stock</b> (gear that hasn't moved), which is exactly what goes missing quietly.
Severity against the 30-day SOP: <span class="chip" style="color:#1f5c99;background:#eef5fc">TARGET DUE</span> = outside its
tier target but still inside the SOP window &nbsp; <span class="chip" style="color:#b07700;background:#fdf3dd">SOP DUE</span>
= 31\u201360 days &nbsp; <span class="chip" style="color:#c00000;background:#fbe9e9">CRITICAL</span> = over 60 days or never sighted.
Client compliance is measured on the 30-day SOP; the 7/14-day cycles are the Coates internal standard, set above the SOP.</div>"""

# ---------------------------------------------------------------- staff worklist
def item_table(items, show_value=True):
    parts = [f"""<table class="items"><tr><th>Barcode</th><th>Description</th><th class="num">Qty</th>
<th class="num">Unit Price</th><th>Last Sighted</th><th class="num">Days</th><th>Severity</th><th>Counted \u2611</th></tr>"""]
    for r in items:
        last = r["last"].strftime("%d/%m/%Y") if r["last"] else "Never"
        dcls = "overdue" if (r["days"] is None or r["days"] > 60) else ("duec" if r["days"] > 30 else "")
        parts.append(f"""<tr><td>{r['barcode']}</td><td>{r['desc'][:58]}</td><td class="num">{r['qty']}</td>
<td class="num">{money(r['price'])}</td><td>{last}</td><td class="num {dcls}">{r['days'] if r['days'] is not None else DASH}</td>
<td>{sev_chip(r['days'])}</td><td style="width:40pt">&nbsp;</td></tr>""")
    parts.append("</table>")
    return "".join(parts)

def grouped_by_unit(items, bar_class=""):
    parts = []
    by_unit = defaultdict(list)
    for r in items: by_unit[r["home"]].append(r)
    # WHY (02 Sep 2026): bays A-Z - the order you walk the store and find a
    # bay by eye - items inside a bay oldest first, A-Z breaking ties.
    for unit in sorted(by_unit, key=ampol_names.sort_key):
        its = sorted(by_unit[unit], key=lambda r: (-(r["days"] if r["days"] is not None else 99999), ampol_names.sort_key(r["desc"])))
        val = sum(r["value"] for r in its if r["value"])
        big = ' big' if len(its) > 30 else ''
        parts.append(f"""<div class="cblock{big}"><div class="cbar {bar_class}"><span class="cname">{unit}</span>
<span class="cmeta">&mdash; {len(its)} item{'s' if len(its) != 1 else ''} due &nbsp;\u2022&nbsp; {(money(val) + ' at risk') if val else 'unpriced'}</span></div>""")
        parts.append(item_table(its))
        parts.append("</div>")
    return "".join(parts)

WORKLIST_CSS = """
/* floor worklist item tables, toned to the house palette */
.wl table.items, .wl table.matrix { width:100%; border-collapse:collapse; font-size:8.6px; border:1px solid #E4E8EC; }
.wl table.items th, .wl table.matrix th { text-align:left; font-size:7.6px; text-transform:uppercase; letter-spacing:.6px; color:#FFFFFF; background:#1A2430; padding:5px 8px; }
.wl table.items td, .wl table.matrix td { padding:4px 8px; border-bottom:1px solid #EEF1F4; vertical-align:top; color:#35404E; }
.wl table.items tr:nth-child(even) td, .wl table.matrix tr:nth-child(even) td { background:#F7F8FA; }
.wl td.num, .wl th.num { text-align:right; }
.wl .overdue { color:#DC2626; font-weight:700; }
.wl .duec { color:#B45309; font-weight:700; }
.wl .okc { color:#16A34A; font-weight:700; }
.wl .chip { display:inline-block; padding:1px 7px; border-radius:8px; font-weight:700; font-size:7.4px; }
.wl .hi td { background:#FDF0E7; }
.wl .cblock { margin:8px 0 12px; break-inside:avoid; page-break-inside:avoid; }
.wl .cblock.big { break-inside:auto; page-break-inside:auto; }
.wl .cbar { background:#F36F21; border-left:8px solid #1A2430; padding:5px 12px; border-radius:0 8px 8px 0; break-after:avoid; }
.wl .cbar.red { background:#DC2626; } .wl .cbar.blue { background:#2F7FD0; } .wl .cbar.dark { background:#1A2430; }
.wl .cname { color:#FFFFFF; font-weight:700; font-size:10.5px; }
.wl .cmeta { color:#FFFFFF; font-size:8.4px; font-weight:700; }
.wl table.matrix { margin-bottom:6px; }
.wl table.matrix th { background:#1A2430; }
.wl thead { display:table-header-group; }
.wl .notice { background:#F6F7F9; border-left:4px solid #F36F21; border-radius:0 10px 10px 0; padding:10px 14px 8px; font-size:9.6px; line-height:1.6; color:#35404E; margin:10px 0 12px; break-inside:avoid; }
.wl .notice .ntitle { color:#16202C; font-weight:700; font-size:10.5px; letter-spacing:.4px; margin-bottom:6px; }
.wl .notice table.matrix { margin:4px 0 8px; }
.wl .notice b { color:#16202C; }
.wl .sect + .note { break-after:avoid; page-break-after:avoid; }
"""


def _house_worklist(cfg_title, kicker, key_items, body, export_dt, refresh_s):
    import k2flow as kf
    cfg = {"client": "Ampol", "title": cfg_title, "kicker": kicker,
           "project": "Ampol Lytton Refinery \u00b7 Permanent Tool Store",
           "asat_note": "(SiteIQ stocktake export request time)",
           "key_items": key_items,
           "team": [{"name": "Andrew Fisher", "role": "Shutdown Manager", "shift": "",
                     "email": "andrew.fisher@coates.com.au", "blurb": "", "lead": True}]}
    return kf.flow_doc(cfg, refresh_s, export_dt.strftime("%d %b %Y %H:%M"),
                       f'<div class="wl">{body}</div>', extra_css=WORKLIST_CSS)


def build_staff_worklist(rows, transit, export_dt, d):
    # WHY (02 Sep 2026): the floor worklist now wears the same house frame
    # as every other PDF in the suite (k2flow): frame, hero, key strip,
    # running header and footer, page numbers. The list itself - bays A-Z,
    # oldest first, tick boxes - is unchanged.
    import k2shell as sh
    from k2shell import esc, num
    refresh = NOW.strftime("%d %b %Y %H:%M")
    due_gas, due_radio = d["due"]["gas"], d["due"]["radio"]
    due_milw, due_gen = d["due"]["milwaukee"], d["due"]["general"]
    all_due = due_gas + due_radio + due_milw + due_gen
    crit = d["crit_instore"]
    val_due = sum(r["value"] for r in all_due if r["value"])
    counters = " \u00b7 ".join(f"{n} \u00d7{c:,}" for n, c in d["counters7"]) or "-"
    parts = [f"""<div class="callout"><span class="lead">Today\u2019s count.</span> <b>{num(len(all_due))} items</b> are due on their tier targets across the store, worth <b class="o">{money(val_due) if val_due else 'unpriced'}</b>; <b>{num(len(d['done7']))}</b> items were scanned in the last 7 days and SOP compliance sits at <b class="o">{d['comp30']:.1f}%</b>. Work each section oldest first - every scan resets that item\u2019s clock. Rows marked <b class="o">ON HIRE</b> are legitimately out: do not hunt shelves for them.</div>
{sh.tiles([
    ("box", num(d['countable']), "Countable assets", f"{num(len(d['instore']))} in store, {num(len(d['onhire']))} on hire", "grey"),
    ("check", f"{d['comp30']:.1f}%", "SOP compliance (30d)", "", "green" if d['comp30'] >= 90 else "amber"),
    ("clock", num(len(d['done7'])), "Scanned last 7 days", "", "green"),
    ("warn", num(len(all_due)), "In-store due (all tiers)", "the list below", "amber" if all_due else "green"),
])}
{sh.tiles([
    ("shield", money(val_due) if val_due else "unpriced", "Value awaiting count", "", "amber"),
    ("warn", num(crit), "In-store over 60 days", f"+{num(d['crit_onhire'])} on hire", "red" if crit else "green"),
    ("swap", num(len(d['missed_returns'])), "Possible missed returns", "resolve first", "red" if d['missed_returns'] else "green"),
    ("bars", num(transit), "Departed lines excluded", "no longer on the register", "grey"),
])}
<div class="alerts"><div class="ah">Team - how to work this list (per the daily stocktake SOP)</div><table class="al">
<tr><td class="al-dot d-amber">&#9679;</td><td><div class="al-t">Stock takes are completed daily, without exception</div><div class="al-s">Both shifts as one team. Work each section oldest first; every scan resets that item\u2019s clock.</div></td></tr>
<tr><td class="al-dot d-amber">&#9679;</td><td><div class="al-t">Sections 1-2 first (gas monitors, radios and batteries), every day</div><div class="al-s">Priority 1, 7-day cycle. Monitors are bumped and charged before issue. Then section 3 (Milwaukee, 14-day), then the store walk bay by bay (30-day SOP).</div></td></tr>
<tr><td class="al-dot d-amber">&#9679;</td><td><div class="al-t">Discrepancies are actioned on the spot</div><div class="al-s">Missing - escalate same day. Misplaced - relocate and correct in the system. Count mismatch - recount and report, never adjust away. Damaged - Out of Service tag, photo, report.</div></td></tr>
<tr><td class="al-dot d-amber">&#9679;</td><td><div class="al-t">ON HIRE rows are legitimately out - do not hunt shelves for them</div><div class="al-s">If you physically sight an on-hire item in the store, flag it same day and process it through the double-check return process.</div></td></tr>
</table></div>
<div class="note">Scanned last 7 days: <b>{num(len(d['done7']))} items</b> ({counters}). Good work - keep the cadence.</div>
{determination_panel()}"""]
    # Section 0
    if d["missed_returns"]:
        parts.append(f"""<div class="sect"><h3>0. Possible missed returns - resolve first ({len(d['missed_returns'])})</h3></div>
<div class="note">Physically sighted in a store bay on a later day than their recorded on-hire date, yet still showing ON HIRE -
likely missed return scans. Locate, confirm, and process through the double-check return process today so hirers are not held for gear that is home.</div>
<table class="items"><tr><th>Sighted In</th><th>Barcode</th><th>Description</th><th>On Hire To</th><th>On-Hire Date</th><th>Sighted</th><th class="num">Unit Price</th></tr>""")
        for r in sorted(d["missed_returns"], key=lambda r: r["last"], reverse=True):
            parts.append(f"""<tr class="hi"><td><b>{r['unit']}</b></td><td>{r['barcode']}</td><td>{r['desc'][:40]}</td>
<td>{r['onhire_to'] or DASH}</td><td>{r['ohd'].strftime('%d/%m/%Y') if r['ohd'] else DASH}</td>
<td>{r['last'].strftime('%d/%m/%Y') if r['last'] else DASH}</td><td class="num">{money(r['price'])}</td></tr>""")
        parts.append("</table>")
    # Sections 1-3
    for num, key, blurb in [
        ("1", "gas", "Priority 1 \u2014 7-day target. Safety-critical: every monitor bumped and charged before issue. Grouped by home bay."),
        ("2", "radio", "Priority 1 \u2014 7-day target. Radios and radio batteries; charging station cycle should keep these current \u2014 anything here is idle."),
        ("3", "milwaukee", "Priority 2 \u2014 14-day target. High-value cordless tools; count includes batteries and chargers.")]:
        items = d["due"][key]
        label = TIERS[key][0]
        val = sum(r["value"] for r in items if r["value"])
        parts.append(f"""<div class="sect"><h3>{num}. {label} due - {len(items)} items \u00b7 {money(val) if val else 'unpriced'}</h3></div>
<div class="note">{blurb}</div>""")
        if items:
            parts.append(grouped_by_unit(items, bar_class="red" if key != "milwaukee" else "dark"))
        else:
            parts.append('<div class="note"><b class="g">All clear</b> - every item in this tier is inside its target cycle. Nice work.</div>')
    # Section 4
    items = d["due"]["general"]
    val = sum(r["value"] for r in items if r["value"])
    parts.append(f"""<div class="sect"><h3>4. Store walk - all other tooling due - {len(items)} items \u00b7 {money(val) if val else 'unpriced'}</h3></div>
<div class="note">Priority 3 - 30-day SOP cycle. Grouped by storage unit, bays A-Z, so one person clears a bay at a time. Tick as counted; scanning resets the clock.</div>""")
    parts.append(grouped_by_unit(items) if items else
                 '<div class="note"><b class="g">All clear</b> - full 30-day coverage.</div>')
    # Section 5: on-hire verification
    onhire_due = d["onhire_due30"]
    parts.append(f"""<div class="pb"></div><div class="sect"><h3>5. On-hire verification - {len(onhire_due)} items outside 30 days (do not shelf count)</h3></div>
<div class="note">ON HIRE in the system - with contractors, not on shelves. Verified through the double-check return process
and shutdown checks. Grouped by company (A-Z, one customer one name; hirers A-Z inside a company) so we know who holds what and the value of it; each line shows the home bay it returns to.</div>""")
    by_co = defaultdict(list)
    for r in onhire_due:
        comp = (r["onhire_to"].split(" \u2013 ")[0] if r["onhire_to"] else "(unknown company)")
        by_co[comp].append(r)
    for comp in sorted(by_co, key=ampol_names.sort_key):
        its = by_co[comp]
        cval = sum(r["value"] for r in its if r["value"])
        hirers = defaultdict(lambda: defaultdict(int)); oldest = defaultdict(int)
        homes = defaultdict(lambda: defaultdict(int)); hval = defaultdict(float)
        for r in its:
            h = r["onhire_to"].split(" \u2013 ", 1)[1] if " \u2013 " in r["onhire_to"] else "(unknown)"
            hirers[h][TIERS[r["cat"]][0] if r["cat"] != "general" else "Tooling & Equipment"] += 1
            oldest[h] = max(oldest[h], r["days"] or 0)
            homes[h][r["home"]] += 1
            if r["value"]: hval[h] += r["value"]
        parts.append(f"""<div class="cblock{' big' if len(hirers) > 25 else ''}"><div class="cbar blue"><span class="cname">{comp}</span>
<span class="cmeta">&mdash; {len(its)} item{'s' if len(its) != 1 else ''} on hire 30+ days since last sighting &nbsp;\u2022&nbsp; {money(cval) if cval else 'unpriced'}</span></div>
<table class="items"><tr><th>Hirer</th><th>Items (by category)</th><th>Returns To (Home Bay)</th><th class="num">Value</th><th class="num">Oldest (Days)</th></tr>""")
        for h in sorted(hirers, key=ampol_names.sort_key):
            cats = ", ".join(f"{c} \u00d7{n}" for c, n in sorted(hirers[h].items(), key=lambda kv: -kv[1]))
            hm = ", ".join(f"{u} \u00d7{n}" for u, n in sorted(homes[h].items(), key=lambda kv: -kv[1])[:3])
            parts.append(f"""<tr><td><b>{h}</b></td><td>{cats}</td><td style="color:#1f5c99">{hm}</td>
<td class="num">{money(hval[h]) if hval[h] else DASH}</td><td class="num overdue">{oldest[h]}</td></tr>""")
        parts.append("</table></div>")
    parts.append(f"""<div class="pb"></div><div class="sect"><h3>Data and method</h3></div>
<div class="note">Countable assets exclude {transit:,} lines that have departed the store (Pending Branch Receipt, or a Departure scan with the item
no longer on the live register). Prices are Avg Buy Price (New) matched on item description (corrected descriptions applied by barcode;
{d['priced_family']:,} serial-numbered gas monitors priced by their family line); unpriced items show a dash and are listed on the Pricing Gaps tab of the
Excel worklist - no price is estimated. SiteIQ still carries the site\u2019s former name on {d['former_name_lines']:,} of these descriptions; they are shown under the current name (Ampol). Status, hirer and home bay joined from RENTAL_STOCK by barcode. Sighting an ON HIRE item in the store
is a missed return - flag and process same day.</div>
<div class="note">{LSR_LINE}</div>""")
    body = "".join(parts)
    return _house_worklist("Stocktake Count Worklist",
                           "COATES \u00b7 STORES STOCKTAKE \u00b7 FLOOR WORKLIST",
                           [("orange", "OLDEST FIRST", "every scan resets the clock"),
                            ("blue", "P1 DAILY", "gas monitors, radios and batteries first"),
                            ("amber", "ON HIRE", "legitimately out - do not hunt shelves")],
                           body, export_dt, refresh)

# ---------------------------------------------------------------- client report
def build_client_report(rows, transit, export_dt, d):
    refresh = NOW.strftime("%d/%m/%Y %I:%M %p")
    countable = d["countable"]
    b = Counter(bucket(r["days"]) for r in rows)
    comp = d["comp30"]
    instore = d["instore"]
    instore_ok = sum(1 for r in instore if bucket(r["days"]) == "ok")
    instore_comp = instore_ok / len(instore) * 100 if instore else 0
    val_cov = d["val_ok30"] / d["val_total"] * 100 if d["val_total"] else 0
    priced_pct = d["priced_lines"] / countable * 100 if countable else 0
    su = defaultdict(lambda: [0, 0])
    for r in instore:
        su[r["unit"]][0] += 1
        if bucket(r["days"]) == "ok": su[r["unit"]][1] += 1
    su_rows = sorted(((u, t, k, k / t * 100) for u, (t, k) in su.items() if t >= 10), key=lambda x: x[3])
    css = CSS.replace("__FOOT__", f"Coates \u2014 Stocktake Compliance Report \u2013 {NOW.strftime('%d/%m/%Y')} \u2014 POWERED BY SITEIQ")
    rag = lambda p: ('#1e7d32', '#e7f3e9', 'GREEN') if p >= 90 else (('#b07700', '#fdf3dd', 'AMBER') if p >= 75 else ('#c00000', '#fbe9e9', 'RED'))
    c, bg, lbl = rag(comp)
    tier_rows = ""
    for k, label, target, why, n, oh, ins, wt, w30, cov in d["tier_stats"]:
        c2, bg2, lbl2 = rag(cov)
        tier_rows += f"""<tr><td><b>{label}</b></td><td>Every {target} days</td><td class="num">{n:,}</td>
<td class="num">{oh:,}</td><td class="num">{ins:,}</td><td class="num">{wt:,}</td><td class="num">{w30:,}</td>
<td class="num" style="color:{c2};font-weight:bold">{cov:.0f}%</td>
<td><span class="chip" style="color:{c2};background:{bg2}">{lbl2}</span></td></tr>"""
    parts = [f"""<!DOCTYPE html><html><head><meta charset="utf-8"><style>{css}</style></head><body>
<div class="banner"><div class="purpose">{COATES_PURPOSE}</div>
<h1>STOCKTAKE COMPLIANCE REPORT \u2013 AMPOL TOOL STORE</h1>
<div class="sub">Coates managed &nbsp;|&nbsp; Compliance against the Daily Stocktake SOP (30-day full-coverage cycle, priority tiers above SOP)
&nbsp;|&nbsp; Data as at: {export_dt.strftime('%d/%m/%Y %I:%M %p')} &nbsp;|&nbsp; Generated: {refresh}</div></div>
<table class="totals"><tr>
<td><div class="lbl">Countable Assets (Live Register)</div><div class="val">{countable:,}</div></td>
<td><div class="lbl">Fleet Value (Priced, New)</div><div class="val">{money(d['val_total'])}</div></td>
<td><div class="lbl">SOP Compliance</div><div class="val" style="color:{c}">{comp:.1f}%</div></td>
<td><div class="lbl">Value Verified \u226430 Days</div><div class="val g">{money(d['val_ok30'])} \u00b7 {val_cov:.0f}%</div></td>
<td><div class="lbl">In-Store Compliance</div><div class="val g">{instore_comp:.1f}%</div></td>
<td><div class="lbl">Scanned Last 7 Days</div><div class="val g">{len(d['done7']):,}</div></td>
</tr></table>
<div class="notice"><div class="ntitle">THE STOCKTAKE STANDARD \u2013 HOW THE STORE IS CONTROLLED</div>
Stock takes at the Ampol Tool Store (Coates managed) are completed <b>daily, without exception</b>, by both shifts working as a team,
under a documented Standard Operating Procedure. The store operates a <b>30-day full-coverage cycle</b> \u2014 no asset goes more than 30 days
without being physically scanned and verified \u2014 and Coates runs <b>tighter internal cycles above the SOP</b> for priority equipment
(see determination below). Daily stock takes include <b>on-hire verification</b>, and all returns pass the <b>double-check return process</b>:
inspected and scanned on receipt, then stock-taken to their storage unit before returning to the shelf. Gas monitors are bumped and charged
before issue. Discrepancies are actioned on the spot \u2014 missing items escalated same day, misplaced items relocated and corrected,
count mismatches recounted and reported rather than adjusted away, damaged items tagged Out of Service with photo and report.</div>
<h2 class="sec">Priority Determination & Coverage by Tier</h2>
<div class="secsub">How count priority is set, and coverage achieved in each tier. Coverage and RAG are rated on <b>in-store</b> assets
against the 30-day SOP; "Within Target" shows performance against the tighter Coates internal cycle. On-hire assets are with contractors
and are verified through the double-check return process and shutdown checks, not shelf counting.</div>
<table class="matrix"><tr><th>Priority Tier</th><th>Target Cycle</th><th class="num">Total Assets</th>
<th class="num">On Hire</th><th class="num">In Store</th><th class="num">Within Target</th><th class="num">Within 30d</th>
<th class="num">In-Store Coverage</th><th>Status</th></tr>{tier_rows}</table>
<div class="charge" style="margin-top:8pt"><div class="ntitle">CURRENT POSITION \u2013 WHAT HAS BEEN DONE</div><ul>
<li>Overall SOP compliance is <b style="color:{c}">{comp:.1f}%</b> ({d['ok30']:,} of {countable:,} countable assets sighted within 30 days = {d['ok30_instore']:,} in store + {d['ok30_onhire']:,} on hire); in-store assets sit at <b>{instore_comp:.1f}%</b>. Of the {d['late_instore'] + d['late_onhire']:,} outside 30 days, {d['late_onhire']:,} are on hire.</li>
<li><b>{len(d['done7']):,} assets were physically scanned in the last 7 days</b> and {len(d['done30']):,} in the last 30 \u2014 the daily cadence is being worked, not just scheduled.</li>
<li>In dollar terms, <b>{money(d['val_ok30'])} of {money(d['val_total'])} in fleet value ({val_cov:.0f}%) has been physically verified within 30 days</b> (Avg Buy Price New; {priced_pct:.0f}% of asset lines priced \u2014 valuation coverage improving as the pricing master is completed).</li>
<li>Assets outside the window are loaded on the team's daily count worklist: {b['due']:,} at 31\u201360 days and {b['critical'] + b['never']:,} at 60+ days, worked oldest-first with priority tiers cleared first.</li>
<li>{money(d['val_onhire'])} of fleet value is on hire with contractors \u2014 tracked by company and hirer, verified through return rescans and shutdown checks.</li>
<li>{transit:,} lines that have departed the store (Pending Branch Receipt, or Departure and no longer on the live register) are excluded from counting KPIs.</li>
</ul></div>
<h2 class="sec" style="margin-top:6pt">Coverage by Storage Unit (In-Store)</h2>
<div class="secsub">Every in-store storage unit with 10+ assets, RAG-rated against the 30-day standard, lowest coverage first \u2014 these are the areas the team is walking now.</div>
<table class="matrix"><tr><th>Storage Unit</th><th class="num">Assets</th><th class="num">Sighted \u226430d</th><th class="num">Coverage</th><th>Status</th></tr>"""]
    for u, t, k, p in su_rows:
        c2, bg2, lbl2 = rag(p)
        hi = ' class="hi"' if p < 75 else ''
        parts.append(f"""<tr{hi}><td><b>{u}</b></td><td class="num">{t:,}</td><td class="num">{k:,}</td>
<td class="num" style="color:{c2};font-weight:bold">{p:.0f}%</td>
<td><span class="chip" style="color:{c2};background:{bg2}">{lbl2}</span></td></tr>""")
    parts.append(f"""</table>
<div class="secsub" style="margin-top:5pt">RAG standard: Green 90%+ coverage, Amber 75\u201389%, Red below 75% (shaded \u2014 prioritised on the daily
worklist with owner and same-day escalation of any missing item).</div>
<div class="foot">Compliance measured as assets physically sighted (scanned) within the 30-day SOP window, from the SiteIQ stocktake export of
{export_dt.strftime('%d/%m/%Y')}. Values are Avg Buy Price (New) \u00d7 sighted quantity, matched on item description; unpriced assets are excluded
from value figures, never estimated. Radios, radio batteries, Dr\u00e4ger gas monitors and Milwaukee tools run on tighter internal cycles above the SOP.
Item-level records, scan history and the daily count worklists are auditable and available on request.{sig_footer()}</div></body></html>""")
    return "".join(parts)

# ---------------------------------------------------------------- Excel worklist
ORANGE = "E07000"
def build_excel_worklist(rows, d, conflicts, xlsx_path):
    wb = openpyxl.Workbook()
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor=ORANGE)
    base_font = Font(name="Calibri", size=10)
    thin = Border(bottom=Side(style="thin", color="DDDDDD"))

    def style_sheet(ws, widths, money_cols=(), last_col_letter=None):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for c in ws[1]:
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{last_col_letter or get_column_letter(ws.max_column)}{ws.max_row}"
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font = base_font; c.border = thin
        for col in money_cols:
            for c in ws[col][1:]:
                c.number_format = "$#,##0"

    # README
    ws = wb.active; ws.title = "READ ME"
    ws["A1"] = "COATES STOCKTAKE COUNT WORKLIST \u2013 AMPOL TOOL STORE"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=ORANGE)
    ws["A2"] = f"Author: Andrew Fisher  \u2022  POWERED BY SITEIQ  \u2022  Generated {NOW.strftime('%d/%m/%Y %I:%M %p')}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    lines = [
        "",
        "HOW PRIORITY IS DETERMINED",
        "  P1  Gas Monitors \u2013 7-day target. Safety-critical; bumped & charged before issue.",
        "  P1  Radios & Radio Batteries \u2013 7-day target. Safety-critical comms; high churn.",
        "  P2  Milwaukee Power Tools \u2013 14-day target. High value, high shrinkage risk.",
        "  P3  All Other Tooling \u2013 30-day full-coverage SOP cycle.",
        "",
        "WHAT COUNTS AS DONE",
        "  An item is done when it is physically scanned \u2013 issue and return scans included.",
        "  The scan resets its clock, so these tabs naturally surface idle stock.",
        "",
        "SEVERITY (vs the 30-day SOP)",
        "  TARGET DUE = outside its tier target but inside the SOP window",
        "  SOP DUE = 31\u201360 days   \u2022   CRITICAL = over 60 days or never sighted",
        "",
        "HOW TO USE",
        "  Work oldest first. Sort/filter any column. Mark Y in Counted when scanned.",
        "  Tabs 1\u20133 first (priority tiers), then tab 4 bay by bay.",
        "  Sight an ON HIRE item in the store? Flag same day \u2013 double-check return process.",
        "  Discrepancies actioned on the spot: missing \u2013 escalate today; misplaced \u2013 relocate",
        "  and correct; mismatch \u2013 recount and report; damaged \u2013 OOS tag, photo, report.",
        "",
        "PRICING",
        "  Unit Price = Avg Buy Price (New), matched on item description.",
        "  Blank price = not yet in the pricing master \u2013 see the Pricing Gaps tab.",
        "  Gas monitors registered with a serial in the description are priced by",
        "  their family line (Drager X-am 5000 Gas Monitor) \u2013 never estimated.",
    ]
    for i, t in enumerate(lines, 3):
        ws.cell(row=i, column=1, value=t).font = Font(
            name="Calibri", size=10,
            bold=t and not t.startswith("  "))
    ws.column_dimensions["A"].width = 95

    def due_sheet(title, items):
        ws = wb.create_sheet(title)
        ws.append(["Home Bay", "Barcode", "Description", "Qty", "Unit Price", "Line Value",
                   "Last Sighted", "Days", "Severity", "Counted (Y)", "Notes"])
        r = 2
        for it in sorted(items, key=lambda x: (x["home"], -(x["days"] if x["days"] is not None else 99999))):
            ws.append([it["home"], it["barcode"], it["desc"], it["qty"], it["price"],
                       f'=IF(E{r}="","",D{r}*E{r})',
                       it["last"].strftime("%d/%m/%Y") if it["last"] else "Never",
                       it["days"], severity_label(it["days"]), "", ""])
            r += 1
        style_sheet(ws, [26, 18, 42, 6, 11, 11, 12, 7, 12, 11, 24], money_cols=("E", "F"))
        return ws

    if d["missed_returns"]:
        ws = wb.create_sheet("0 Missed Returns")
        ws.append(["Sighted In", "Barcode", "Description", "On Hire To", "On-Hire Date",
                   "Sighted", "Unit Price", "Resolved (Y)", "Notes"])
        for it in sorted(d["missed_returns"], key=lambda x: x["last"], reverse=True):
            ws.append([it["unit"], it["barcode"], it["desc"], it["onhire_to"],
                       it["ohd"].strftime("%d/%m/%Y") if it["ohd"] else "",
                       it["last"].strftime("%d/%m/%Y") if it["last"] else "",
                       it["price"], "", ""])
        style_sheet(ws, [24, 18, 40, 34, 12, 12, 11, 11, 24], money_cols=("G",))

    due_sheet("1 Gas Monitors", d["due"]["gas"])
    due_sheet("2 Radios & Batteries", d["due"]["radio"])
    due_sheet("3 Milwaukee", d["due"]["milwaukee"])
    due_sheet("4 Store Walk", d["due"]["general"])

    ws = wb.create_sheet("5 On-Hire Verify")
    ws.append(["Company", "Hirer", "Barcode", "Description", "Category", "Home Bay",
               "On-Hire Date", "Days Since Sighted", "Unit Price", "Verified (Y)"])
    for it in sorted(d["onhire_due30"], key=lambda x: (ampol_names.sort_key(x["onhire_to"]), -(x["days"] or 0))):
        comp = it["onhire_to"].split(" \u2013 ")[0] if it["onhire_to"] else "(unknown)"
        hirer = it["onhire_to"].split(" \u2013 ", 1)[1] if " \u2013 " in it["onhire_to"] else ""
        ws.append([comp, hirer, it["barcode"], it["desc"],
                   TIERS[it["cat"]][0] if it["cat"] != "general" else "Tooling & Equipment",
                   it["home"], it["ohd"].strftime("%d/%m/%Y") if it["ohd"] else "",
                   it["days"], it["price"], ""])
    style_sheet(ws, [26, 26, 18, 40, 22, 24, 12, 9, 11, 11], money_cols=("I",))

    # Pricing gaps
    ws = wb.create_sheet("Pricing Gaps")
    ws.append(["Description (as in SiteIQ)", "Asset Lines", "Total Qty", "Add Avg Buy Price (New)"])
    gaps = defaultdict(lambda: [0, 0])
    for r_ in rows:
        if r_["price"] is None:
            g = gaps[r_["desc"]]; g[0] += 1; g[1] += r_["qty"]
    for desc, (n, q) in sorted(gaps.items(), key=lambda kv: -kv[1][0]):
        ws.append([desc, n, q, ""])
    if conflicts:
        ws.append([]); ws.append(["CONFLICTING PRICES IN PRICING MASTER (most common value used; a tie takes the lower value)"])
        ws.cell(row=ws.max_row, column=1).font = Font(name="Calibri", bold=True, color="C00000")
        for dsc, prices in sorted(conflicts):
            ws.append([dsc, "", "", " / ".join(f"${p:,.2f}" for p in prices)])
    style_sheet(ws, [52, 11, 10, 26], last_col_letter="D")

    wb.properties.creator = "Fisher, Andrew (North)"
    wb.properties.lastModifiedBy = "Fisher, Andrew (North)"
    wb.save(xlsx_path)
    print(f"Wrote {xlsx_path}")

# ---------------------------------------------------------------- PDF + email
def write_pdf_robust(html_path, pdf_path):
    try:
        from weasyprint import HTML
        HTML(filename=str(html_path)).write_pdf(str(pdf_path)); return True
    except Exception as e:
        print(f"WeasyPrint unavailable ({type(e).__name__}) - trying Edge/Chrome headless...")
    import subprocess, os, time, tempfile
    write_pdf_robust._n = getattr(write_pdf_robust, "_n", 0) + 1
    for exe in [r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
                r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                "msedge", "chrome", "chromium", "chromium-browser", "google-chrome"]:
        if exe.endswith(".exe") and not os.path.exists(exe): continue
        ran = False
        for attempt in range(3):
            # unique profile per attempt: a running browser with the same
            # profile makes headless print exit 0 without writing the PDF
            # WHY (12 Aug 2026): profile junk goes to the system temp dir,
            # never the suite folder, when TEMP isn't set.
            profile = os.path.join(os.environ.get("TEMP") or tempfile.gettempdir(),
                                   f"coates_edge_pdf_{os.getpid()}_{write_pdf_robust._n}_{attempt}")
            try:
                subprocess.run([exe, "--headless", "--disable-gpu",
                                "--user-data-dir=" + profile, "--no-first-run",
                                f"--print-to-pdf={Path(pdf_path).resolve()}",
                                "--no-pdf-header-footer", Path(html_path).resolve().as_uri()],
                               capture_output=True, timeout=240)
                ran = True
            except FileNotFoundError:
                break   # not on this machine - move on to the next candidate
            except Exception:
                ran = True   # launched but fell over (e.g. timeout) - retry
            if Path(pdf_path).exists():
                print(f"PDF generated via {Path(exe).name}"); return True
            time.sleep(1.5)
        # WHY (12 Aug 2026): only give up after a browser that actually
        # launched - the old unconditional break meant nothing after the
        # first bare name on PATH was ever tried, so the fallback list
        # wasn't really a fallback list.
        if ran:
            break
    print("No PDF engine found - HTML written; PDF skipped."); return False

def write_staff_eml(rows, transit, export_dt, d, attachments, eml_path):
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication
    due_p1 = len(d["due"]["gas"]) + len(d["due"]["radio"])
    due_all = due_p1 + len(d["due"]["milwaukee"]) + len(d["due"]["general"])
    val_due = sum(r["value"] for k in ("gas", "radio", "milwaukee", "general")
                  for r in d["due"][k] if r["value"])
    su = Counter(r["home"] for k in ("gas", "radio", "milwaukee", "general") for r in d["due"][k])
    top = "".join(f'<tr><td style="padding:4px 10px;border-bottom:1px solid #eee"><b>{u}</b></td>'
                  f'<td align="right" style="padding:4px 10px;border-bottom:1px solid #eee;color:#b07700;font-weight:bold">{n}</td></tr>'
                  for u, n in su.most_common(10))
    O = "#e07000"
    body = f"""<!doctype html><html><body style="margin:0;font-family:Calibri,Arial,sans-serif;color:#1a1a1a">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr><td align="center" style="padding:16px 10px">
<table role="presentation" width="700" cellspacing="0" cellpadding="0" style="max-width:700px;border-collapse:collapse">
<tr><td style="border-left:6px solid {O};padding:6px 12px">
<div style="font-size:12px;font-weight:bold;color:{O}">{COATES_PURPOSE}</div>
<div style="font-size:20px;font-weight:bold">STOCKTAKE COUNT WORKLIST \u2013 {NOW.strftime('%d/%m/%Y')}</div>
<div style="font-size:11px;color:#555">Ampol Tool Store \u2013 Coates managed &nbsp;|&nbsp; Data as at {export_dt.strftime('%d/%m/%Y')}</div></td></tr>
<tr><td style="padding:10px 0;font-size:12px;line-height:1.6">Morning team,<br><br>
Attached is today's count worklist \u2014 PDF for the floor, Excel if you're on a tablet (sort, filter, mark counted).<br><br>
<b>Priority order, every day: Sections 1\u20132 first (gas monitors, radios & batteries \u2014 7-day cycle), then Section 3
(Milwaukee \u2014 14-day cycle), then the store walk bay by bay (30-day SOP).</b> Every scan resets that item's clock \u2014
issues and returns count, so what's left on this list is the idle gear. Discrepancies on the spot: missing \u2014 escalate today;
misplaced \u2014 relocate and fix in the system; mismatch \u2014 recount and report; damaged \u2014 OOS tag, photo, report.</td></tr>
<tr><td><table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse">
<tr><td style="border:1px solid #ccc;padding:7px 10px"><div style="font-size:8px;color:#666">SOP COMPLIANCE</div><div style="font-size:15px;font-weight:bold;color:#1e7d32">{d['comp30']:.1f}%</div></td>
<td style="border:1px solid #ccc;padding:7px 10px"><div style="font-size:8px;color:#666">SCANNED LAST 7 DAYS</div><div style="font-size:15px;font-weight:bold;color:#1e7d32">{len(d['done7']):,}</div></td>
<td style="border:1px solid #ccc;padding:7px 10px"><div style="font-size:8px;color:#666">DUE TODAY (ALL TIERS)</div><div style="font-size:15px;font-weight:bold;color:#b07700">{due_all:,}</div></td>
<td style="border:1px solid #ccc;padding:7px 10px"><div style="font-size:8px;color:#666">PRIORITY 1 DUE</div><div style="font-size:15px;font-weight:bold;color:#c00000">{due_p1:,}</div></td>
<td style="border:1px solid #ccc;padding:7px 10px"><div style="font-size:8px;color:#666">VALUE AWAITING COUNT</div><div style="font-size:15px;font-weight:bold;color:#b07700">{money(val_due)}</div></td>
</tr></table></td></tr>
<tr><td style="padding:10px 0 4px;font-size:13px;font-weight:bold;color:{O}">TOP AREAS TO CLEAR TODAY</td></tr>
<tr><td><table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;border:1px solid #ccc;font-size:11px">
<tr bgcolor="#fafafa"><td style="padding:4px 10px;font-size:9px;color:#555;border-bottom:1px solid #ccc">STORAGE UNIT / BAY</td>
<td align="right" style="padding:4px 10px;font-size:9px;color:#555;border-bottom:1px solid #ccc">ITEMS DUE</td></tr>{top}</table></td></tr>
<tr><td style="padding:12px 0;font-size:12px;line-height:1.6">{len(d['done7']):,} items scanned in the last 7 days \u2014 good work, keep the cadence.
The compliance report (client-shareable) is attached too. Thanks team \u2014 daily, without exception. It's what makes this store trusted.<br><br>Andrew</td></tr>
<tr><td style="border-top:1px solid #ccc;padding:8px 0;font-size:10px;color:#555;text-align:center">
<b>Author: Andrew Fisher</b> &nbsp;\u2022&nbsp; <b style="color:{O}">POWERED BY SITEIQ</b><br>
<span style="color:{O};font-weight:bold">{COATES_VALUES}</span></td></tr>
</table></td></tr></table></body></html>"""
    # standard Coates email frame: 2px dark border, dark hero, LSR footer
    body = f"""<table role="presentation" width="100%" cellspacing="0" cellpadding="0"
 style="background:#EFECE7"><tr><td align="center" style="padding:14px 8px">
<table role="presentation" width="740" cellspacing="0" cellpadding="0"
 style="width:740px;max-width:740px;background:#ffffff;border:2px solid #1D1D1B;border-collapse:collapse">
<tr><td style="background:#1D1D1B;padding:16px 22px;font-family:Arial,sans-serif">
 <div style="font-size:20px;font-weight:900;color:#ffffff">Coates &nbsp;|&nbsp; Ampol Tool Store</div>
 <div style="font-size:12px;font-weight:700;color:#F26222;margin-top:4px">Daily Stocktake \u2013 the cadence that keeps this store trusted</div>
 <div style="font-size:10px;color:#bbbbbb;margin-top:5px">The Coates Way &nbsp;|&nbsp; POWERED BY SITEIQ &nbsp;|&nbsp; Author: Andrew Fisher</div>
</td></tr>
<tr><td style="padding:6px 14px">{body}</td></tr>
<tr><td style="background:#F5F1EC;border-top:3px solid #F26222;padding:9px 18px;
 font-family:Arial,sans-serif;font-size:9px;color:#555555;line-height:1.7">{LSR_LINE}<br/>
 Coates Hire Operations Pty Limited | ABN 50 009 779 338 | www.coates.com.au | POWERED BY SITEIQ</td></tr>
</table></td></tr></table>"""
    subject = f"Ampol Tool Store \u2013 Stocktake Count Worklist \u2013 {NOW.strftime('%d/%m/%Y %I:%M %p')}"
    msg = MIMEMultipart("mixed")
    msg["To"] = STAFF_EMAIL_TO
    msg["Subject"] = subject
    msg["X-Unsent"] = "1"
    msg.attach(MIMEText(body, "html", "utf-8"))
    subtypes = {".pdf": "pdf", ".xlsx": "vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    for p in attachments:
        p = Path(p)
        if p.exists():
            part = MIMEApplication(p.read_bytes(), _subtype=subtypes.get(p.suffix, "octet-stream"))
            part.add_header("Content-Disposition", "attachment", filename=p.name)
            msg.attach(part)
    with open(eml_path, "wb") as f:
        f.write(msg.as_bytes())
    # sidecar for MAKE_OUTLOOK_DRAFTS: NATIVE Outlook draft, To pre-filled
    import json
    addrs = "; ".join(re.findall(r"<([^>]+)>", STAFF_EMAIL_TO))
    stem = str(eml_path)[:-4] if str(eml_path).lower().endswith(".eml") else str(eml_path)
    with open(stem + ".body.html", "w", encoding="utf-8") as f:
        f.write(body)
    with open(stem + ".draft.json", "w", encoding="utf-8") as f:
        json.dump({"subject": subject, "to": addrs,
                   "body": Path(stem + ".body.html").name,
                   "attachments": [Path(p).name for p in attachments if Path(p).exists()]},
                  f, indent=1)
    print(f"Wrote {eml_path} + native-draft manifest")

# ---------------------------------------------------------------- main
def find_workbook(patterns, arg=None):
    # WHY (12 Aug 2026): inputs come from the suite's one Data area now,
    # via ampol_paths - newest file wins, Excel lock files (~$...) and
    # archived Source_ pulls are never candidates. A path given on the
    # command line still wins outright, same as always.
    if arg: return arg
    return ampol_paths.find_data(*patterns) or None

def main():
    src = find_workbook(["STOCKTAKE*.xlsx"], sys.argv[1] if len(sys.argv) > 1 else None)
    master_path = find_workbook(["RENTAL_STOCK*.xlsx"], sys.argv[2] if len(sys.argv) > 2 else None)
    if not src: raise FileNotFoundError("Place the STOCKTAKE export (STOCKTAKE*.xlsx) in the suite's Data folder.")
    fixes_path = find_workbook(["New_Descriptions*.xlsx", "*Descriptions*.xlsx"],
                               sys.argv[3] if len(sys.argv) > 3 else None)
    pricing_path = find_workbook(["*Pricing*.xlsx"], sys.argv[4] if len(sys.argv) > 4 else None)
    print(f"Stocktake export:       {src}")
    print(f"Master stock file:      {master_path or 'NOT FOUND - status/hirer/home unit will be blank'}")
    print(f"Corrected descriptions: {fixes_path or 'NOT FOUND - original descriptions used'}")
    print(f"Pricing workbook:       {pricing_path or 'NOT FOUND - values omitted'}")
    OUT.mkdir(exist_ok=True)
    master = load_master(master_path) if master_path else {}
    fixes = load_corrections(fixes_path) if fixes_path else ({}, {})
    exact, stripped, conflicts = load_pricing(pricing_path) if pricing_path else ({}, {}, [])
    rows, transit, export_dt = load(src, master, fixes, exact, stripped)
    d = derive(rows, export_dt)
    applied = sum(1 for r in rows if r["fixed"])
    print(f"Corrected descriptions applied to {applied:,} of {len(rows):,} countable items "
          f"(barcode match first, then unambiguous old-description match)")
    print(f"Priced {d['priced_lines']:,} of {len(rows):,} lines "
          f"({d['priced_lines']/len(rows)*100:.1f}%; {d['priced_family']:,} by family line) "
          f"| fleet value {money(d['val_total'])} "
          f"| pricing conflicts flagged: {len(conflicts)}")
    # WHY (03 Sep 2026): the worklist's names come from ampol_names.report_stem
    # - one file-name rule for the whole suite, the day on the end. The
    # house-style builder (button 05) writes the same worklist under the
    # same stem; this standalone run is the V3 engine on its own.
    w = OUT / ampol_names.report_stem("worklist")
    c = OUT / "Coates_Stocktake_Compliance_Report"
    x = OUT / f"{ampol_names.report_stem('worklist')}.xlsx"
    open(f"{w}.html", "w", encoding="utf-8").write(build_staff_worklist(rows, transit, export_dt, d))
    open(f"{c}.html", "w", encoding="utf-8").write(build_client_report(rows, transit, export_dt, d))
    build_excel_worklist(rows, d, conflicts, x)
    write_pdf_robust(f"{w}.html", f"{w}.pdf")
    write_pdf_robust(f"{c}.html", f"{c}.pdf")
    write_staff_eml(rows, transit, export_dt, d,
                    [f"{w}.pdf", f"{c}.pdf", x], OUT / f"{ampol_names.report_stem('worklist')}_OUTLOOK.eml")
    due_all = sum(len(d["due"][k]) for k in TIERS)
    print(f"{len(rows):,} countable | SOP compliance {d['comp30']:.1f}% | "
          f"due (tiered) {due_all:,} [gas {len(d['due']['gas'])}, radio {len(d['due']['radio'])}, "
          f"milwaukee {len(d['due']['milwaukee'])}, general {len(d['due']['general'])}] | "
          f"missed returns {len(d['missed_returns'])} | onhire verify {len(d['onhire_due30']):,} | transit excl {transit:,}")

if __name__ == "__main__":
    main()
