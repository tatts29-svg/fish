"""COATES | AMPOL TOOL STORE - what the transaction log already knows.

Author: Andrew Fisher | POWERED BY SITEIQ

WHY (03 Sep 2026): the TRANSACTIONS export holds every issue and return
since 1 Jan - 90,000-odd real movements - and the reports were using a
slice of it. This module reads it once and answers, for any family's
barcode set, the questions a client asks next:

  weekly_series    issues, returns, net out and same-day rate by ISO week
  monthly_league   the same by company and month, with movement
  quarter_close    items on hire that cross N days before a date unless
                   they come back - arithmetic on the register, no forecast
  dead_stock       available items with no movement this year, with value
  headroom         fleet against the peak number out at once, by product
  return_windows   median and 90th-percentile hold by product
  holders          who holds what across the store, the 80/20 of exposure
  counter_rhythm   draws and returns by weekday and hour
  data_quality     scan errors and register-versus-log gaps
  fast_movers      the most-drawn items and where they sit

Every figure is a count or a sum over rows that exist in the exports.
Nothing is modelled, weighted or forecast. Where a price is unknown the
value is left out and the unpriced count is returned beside it.
"""

import bisect
import collections
import datetime as dt
import os
import re
import statistics

import openpyxl

import ampol_names
import ampol_paths
import ampol_master
import pull_diff
from gasmon_engine import is_gas_monitor, parse_dt, parse_stamp

# ---------------------------------------------------------------------------
# loading
# ---------------------------------------------------------------------------

_SERIAL = re.compile(r"\s*-?\s*[A-Z]{4}-\d{3,5}\s*$")
# a trailing serial token: letters and digits mixed, six or more characters
# (Motorola radios carry 122TYX0381-style serials in the description)
_SERIAL2 = re.compile(r"[\s-]+(?=[A-Z0-9]{6,}$)(?=[A-Z0-9]*\d)(?=[A-Z0-9]*[A-Z])[A-Z0-9]{6,}$")
_QUALIFIER = re.compile(r"\s*-+\s*(T&I|MAINTENANCE|MAINT|TURNAROUND|SHUTDOWN)\s*-*\s*$")
_NUM = r"(?:\d+(?:[.,]\d+)?(?:\s*\d+/\d+)?|\d+/\d+)"
_SIZE = re.compile(r"(\s*[-x×/]?\s*\(?" + _NUM + r"\s*(MM|CM|M|IN|INCH|\"|''|FT|KG|G|AH|V|W|A|L|PCS?|PIECE|T|TONNE|PSI|BAR|DEG|°)?\)?)+\s*$", re.I)
_SITE = re.compile(r"^(?:(?:CALTEX|AMPOL)\s+)+")
GAS_KEY = "DRAGER X-AM 5000 GAS MONITOR"
_TAIL = re.compile(r"\s*[-,(]\s*$")


def product_key(desc):
    """The product a description belongs to, with its size and serial tail
    removed: '1in Drive Impact Socket 27mm' -> '1IN DRIVE IMPACT SOCKET'.
    Sizes are stripped from the end only; nothing inside a name is touched."""
    if is_gas_monitor(str(desc or "")):
        return GAS_KEY          # one product, three description conventions
    s = re.sub(r"\s+", " ", str(desc or "").strip().upper())
    s = _SITE.sub("", s)
    s = _SERIAL.sub("", s)
    s = _SERIAL2.sub("", s)
    for _ in range(2):
        s = _QUALIFIER.sub("", s)
    s = _SERIAL2.sub("", s)
    for _ in range(3):
        s2 = _SIZE.sub("", s)
        s2 = _TAIL.sub("", s2)
        if s2 == s:
            break
        s = s2
    return s.strip() or "(blank)"


def report_family(desc):
    """gas / radio / tooling - the report family a description belongs to.
    gas uses the gas engine's own test; radio is the radio report's rule
    (Motorola radios and radio batteries); everything else is tooling."""
    d = str(desc or "")
    if is_gas_monitor(d):
        return "gas"
    if re.search(r"\b(MOTOROLA|RADIO)\b", d, re.I) and not re.search(r"CHARG", d, re.I):
        return "radio"
    return "tooling"


def load_transactions(path):
    """Every CUSTOMER_CONTRACTOR_EQUIP row with a parseable start time."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        window = (None, None)
        if "REFERENCE_INFO" in wb.sheetnames:
            rows = list(wb["REFERENCE_INFO"].iter_rows(values_only=True))
            if len(rows) > 1:
                hdr = [str(h or "").upper() for h in rows[0]]
                for i, h in enumerate(hdr):
                    v = rows[1][i] if i < len(rows[1]) else None
                    if "REPORT_PERIOD" in h and v:
                        bits = [b.strip() for b in str(v).split(" - ")]
                        if len(bits) == 2:
                            window = (parse_stamp(bits[0]), parse_stamp(bits[1]))
        ws = wb["CUSTOMER_CONTRACTOR_EQUIP"]
        it = ws.iter_rows(values_only=True)
        hdr = [str(c).strip().upper() if c is not None else "" for c in next(it)]
        ix = {h: i for i, h in enumerate(hdr)}

        def g(r, k):
            i = ix.get(k)
            return r[i] if i is not None and i < len(r) else None
        out = []
        for r in it:
            if not r or r[0] is None:
                continue
            st = parse_dt(g(r, "TRAN_START_DATE"), g(r, "TRAN_START_TIME"))
            if st is None:
                continue
            en = parse_dt(g(r, "TRAN_END_DATE"), g(r, "TRAN_END_TIME"))
            raw_desc = str(g(r, "SKU/ITEM DESCRIPTION") or g(r, "PRODUCT_VARIANT") or "").strip()
            bc = str(g(r, "LATEST_BARCODE") or "").strip()
            desc = ampol_names.display_desc(raw_desc, barcode=bc)
            co_raw = str(g(r, "EMPLOYER_NAME") or "").strip()
            out.append({
                "bc": bc,
                "item": str(g(r, "SKU/ITEM_NUMBER") or "").strip(),
                "desc": desc, "desc_raw": ampol_names.former_to_current(raw_desc), "co_raw": co_raw,
                "co": ampol_names.display_company(co_raw) if co_raw else "",
                "who": str(g(r, "HIRER_NAME") or "").strip(),
                "st": st, "en": en,
                "hours": ((en - st).total_seconds() / 3600) if en else None,
                "sd": bool(en and en.date() == st.date()),
                "qty": g(r, "QUANTITY") or 1,
                "tid": g(r, "TRANSACTION_ID"),
            })
        return out, window
    finally:
        wb.close()


def load_all(data_dir=None):
    """Register, transactions and pricing in one context."""
    data_dir = data_dir or ampol_paths.data_dir()
    reg_path = ampol_paths.find_data("RENTAL_STOCK*.xlsx") or os.path.join(data_dir, "RENTAL_STOCK.xlsx")
    tx_path = ampol_paths.find_data("TRANSACTIONS*.xlsx") or os.path.join(data_dir, "TRANSACTIONS.xlsx")
    reg = pull_diff.load_register(reg_path)
    reg_time = pull_diff.reference_pull_time(reg_path)
    tx, window = load_transactions(tx_path)
    exact, stripped = {}, {}
    try:
        import build_stocktake_compliance_tool as eng
        p, p_sheet = ampol_master.locate("pricing", "Ampol_ToolStore_Pricing*.xlsx", "*Pricing*.xlsx")
        if p:
            exact, stripped, _ = eng.load_pricing(p, p_sheet)
        # WHY (03 Sep 2026): a serial-suffixed description (radios carry
        # 122TYX0381-style serials) never matches the master line by name.
        # The master's own price for the same PRODUCT (description with the
        # size and serial tail removed) is applied - the family rule the
        # stocktake engine already uses, widened to every product. Ties
        # take the lower price; nothing is estimated.
        by_key = {}
        for d0, p0 in exact.items():
            k0 = product_key(d0)
            by_key[k0] = min(by_key.get(k0, p0), p0)

        def price(d):
            p1 = eng.price_for(d, d, exact, stripped)
            return p1 if p1 else by_key.get(product_key(d))
    except Exception:
        price = lambda d: None   # noqa: E731
    return {"reg": reg, "reg_time": reg_time, "tx": tx, "tx_window": window,
            "price": price, "reg_path": reg_path, "tx_path": tx_path}


def _bcs(ctx, scope_barcodes):
    return {b.upper() for b in scope_barcodes} if scope_barcodes is not None else None


def _tx(ctx, scope):
    if scope is None:
        return ctx["tx"]
    return [t for t in ctx["tx"] if t["bc"].upper() in scope]


def _reg(ctx, scope):
    if scope is None:
        return list(ctx["reg"].values())
    return [r for r in ctx["reg"].values() if r["barcode"].upper() in scope]


# ---------------------------------------------------------------------------
# 1. trends from the log
# ---------------------------------------------------------------------------

def weekly_series(ctx, scope_barcodes=None, upto=None):
    """One row per ISO week from the first movement to the pull:
    issues, returns, net (issues - returns), same-day % of the issues that
    have returned. The current, partial week is flagged."""
    scope = _bcs(ctx, scope_barcodes)
    upto = upto or ctx["reg_time"] or dt.datetime.now()
    weeks = collections.OrderedDict()
    for t in _tx(ctx, scope):
        if t["st"] > upto:
            continue
        y, w, _ = t["st"].isocalendar()
        k = (y, w)
        d = weeks.setdefault(k, {"issues": 0, "returns": 0, "sd": 0, "closed": 0})
        d["issues"] += 1
        if t["en"] and t["en"] <= upto:
            d["closed"] += 1
            if t["sd"]:
                d["sd"] += 1
        if t["en"] and t["en"] <= upto:
            ry, rw, _ = t["en"].isocalendar()
            weeks.setdefault((ry, rw), {"issues": 0, "returns": 0, "sd": 0, "closed": 0})["returns"] += 1
    out = []
    cy, cw, _ = upto.isocalendar()
    for (y, w) in sorted(weeks):
        d = weeks[(y, w)]
        start = dt.date.fromisocalendar(y, w, 1)
        out.append({"week": f"{start:%d %b}", "start": start, "issues": d["issues"], "returns": d["returns"],
                    "net": d["issues"] - d["returns"],
                    "sd_pct": round(100 * d["sd"] / d["closed"], 1) if d["closed"] else None,
                    "partial": (y, w) == (cy, cw)})
    return out


def monthly_league(ctx, scope_barcodes=None, upto=None):
    """{company: [{month, issues, returns, sd_pct, late_pct}]} plus, per
    company, the movement of the same-day rate against the month before.
    late = returned after the day it went out (or not yet returned)."""
    scope = _bcs(ctx, scope_barcodes)
    upto = upto or ctx["reg_time"] or dt.datetime.now()
    tab = collections.defaultdict(lambda: collections.defaultdict(lambda: {"issues": 0, "returns": 0, "sd": 0, "closed": 0}))
    for t in _tx(ctx, scope):
        if t["st"] > upto or not t["co"]:
            continue
        m = t["st"].strftime("%Y-%m")
        d = tab[t["co"]][m]
        d["issues"] += 1
        if t["en"] and t["en"] <= upto:
            d["closed"] += 1
            d["returns"] += 1
            if t["sd"]:
                d["sd"] += 1
    out = {}
    for co, months in tab.items():
        rows = []
        for m in sorted(months):
            d = months[m]
            rows.append({"month": dt.datetime.strptime(m, "%Y-%m").strftime("%b"), "key": m, "issues": d["issues"],
                         "returns": d["returns"],
                         "sd_pct": round(100 * d["sd"] / d["closed"], 1) if d["closed"] else None})
        out[co] = rows
    return out


# ---------------------------------------------------------------------------
# 2. quarter close - the look forward that is only arithmetic
# ---------------------------------------------------------------------------

def quarter_close(ctx, scope_barcodes=None, qend=None, threshold=90):
    """Items on hire now that will have been out `threshold` days or more
    by `qend` unless returned, and are not there yet today. Rows carry the
    day they cross. qend defaults to the last day of the pull's quarter."""
    scope = _bcs(ctx, scope_barcodes)
    now = (ctx["reg_time"] or dt.datetime.now()).date()
    if qend is None:
        q = (now.month - 1) // 3
        m = q * 3 + 3
        qend = dt.date(now.year, m, [31, 30, 30, 31][q])
    price = ctx["price"]
    rows = []
    for r in _reg(ctx, scope):
        if r["status"].lower() != "on hire" or not r["on_dt"]:
            continue
        days = (now - r["on_dt"].date()).days
        at_end = (qend - r["on_dt"].date()).days
        if 0 <= days < threshold <= at_end:
            rows.append({"barcode": r["barcode"], "desc": r["desc"], "company": r["company"], "hirer": r["hirer"],
                         "on_dt": r["on_dt"], "days": days,
                         "crosses": r["on_dt"].date() + dt.timedelta(days=threshold),
                         "price": price(r.get("desc_raw", r["desc"]))})
    rows.sort(key=lambda x: (x["crosses"], ampol_names.sort_key(x["company"]), x["barcode"]))
    by_co = collections.Counter(x["company"] for x in rows)
    val = sum(x["price"] for x in rows if x["price"])
    return {"rows": rows, "n": len(rows), "value": val, "unpriced": sum(1 for x in rows if not x["price"]),
            "by_company": sorted(by_co.items(), key=lambda kv: ampol_names.sort_key(kv[0])),
            "qend": qend, "threshold": threshold, "asof": now,
            "already_over": sum(1 for r in _reg(ctx, scope) if r["status"].lower() == "on hire" and r["on_dt"]
                                and (now - r["on_dt"].date()).days >= threshold)}


# ---------------------------------------------------------------------------
# 3. dead stock and headroom
# ---------------------------------------------------------------------------

def dead_stock(ctx, scope_barcodes=None):
    """Available items with no transaction in the log at all."""
    scope = _bcs(ctx, scope_barcodes)
    moved = {t["bc"].upper() for t in ctx["tx"] if t["bc"]}
    price = ctx["price"]
    rows = []
    for r in _reg(ctx, scope):
        if r["status"].lower() != "available for hire":
            continue
        if r["barcode"].upper() in moved:
            continue
        rows.append({"barcode": r["barcode"], "desc": r["desc"], "unit": r["unit"], "product": product_key(r.get("desc_raw", r["desc"])),
                     "price": price(r.get("desc_raw", r["desc"]))})
    rows.sort(key=lambda x: (ampol_names.sort_key(x["product"]), x["barcode"]))
    by_product = collections.defaultdict(lambda: {"n": 0, "value": 0.0, "unpriced": 0})
    for x in rows:
        d = by_product[x["product"]]
        d["n"] += 1
        if x["price"]:
            d["value"] += x["price"]
        else:
            d["unpriced"] += 1
    avail = sum(1 for r in _reg(ctx, scope) if r["status"].lower() == "available for hire")
    return {"rows": rows, "n": len(rows), "available": avail,
            "value": sum(x["price"] for x in rows if x["price"]), "unpriced": sum(1 for x in rows if not x["price"]),
            "by_product": sorted(by_product.items(), key=lambda kv: (-kv[1]["n"], ampol_names.sort_key(kv[0]))),
            "window": ctx["tx_window"]}


def headroom(ctx, scope_barcodes=None, upto=None):
    """Per product: fleet on the register, the most out at once this year
    (from the log's start and end times), out now, and the never-moved
    count. Fleet less peak is the headroom the store has never needed."""
    scope = _bcs(ctx, scope_barcodes)
    upto = upto or ctx["reg_time"] or dt.datetime.now()
    fleet = collections.Counter()
    out_now = collections.Counter()
    for r in _reg(ctx, scope):
        p = product_key(r.get("desc_raw", r["desc"]))
        fleet[p] += 1
        if r["status"].lower() == "on hire":
            out_now[p] += 1
    events = collections.defaultdict(list)
    for t in _tx(ctx, scope):
        p = product_key(t.get("desc_raw", t["desc"]))
        events[p].append((t["st"], 1))
        events[p].append((t["en"] if t["en"] and t["en"] <= upto else upto, -1))
    dead = collections.Counter(x["product"] for x in dead_stock(ctx, scope_barcodes)["rows"])
    rows = []
    for p in set(fleet) | set(events):
        cur = peak = 0
        peak_at = None
        for when, d in sorted(events.get(p, []), key=lambda e: (e[0], e[1])):
            cur += d
            if cur > peak:
                peak, peak_at = cur, when
        rows.append({"product": p, "fleet": fleet.get(p, 0), "peak": peak, "peak_at": peak_at,
                     "out_now": out_now.get(p, 0), "never_moved": dead.get(p, 0),
                     # a peak above today's fleet means barcodes that have since
                     # left the register moved this year - headroom is then 0
                     "headroom": max(0, fleet.get(p, 0) - peak)})
    rows.sort(key=lambda x: (-x["fleet"], ampol_names.sort_key(x["product"])))
    return rows


# ---------------------------------------------------------------------------
# 4. return windows by product - the rule behind the line
# ---------------------------------------------------------------------------

def return_windows(ctx, scope_barcodes=None, min_n=10):
    """Per product (completed hires only): n, median days, 90th-percentile
    days, same-day %. Products with fewer than min_n completed hires are
    pooled into '(fewer than N hires each)'."""
    scope = _bcs(ctx, scope_barcodes)
    byp = collections.defaultdict(list)
    for t in _tx(ctx, scope):
        if t["hours"] is None or t["hours"] < 0:
            continue
        byp[product_key(t.get("desc_raw", t["desc"]))].append((t["hours"] / 24.0, t["sd"]))
    rows, pooled = [], []
    for p, xs in byp.items():
        if len(xs) < min_n:
            pooled.extend(xs)
            continue
        days = sorted(x[0] for x in xs)
        rows.append({"product": p, "n": len(xs), "median": statistics.median(days),
                     "p90": days[min(len(days) - 1, int(len(days) * 0.9))],
                     "sd_pct": round(100 * sum(1 for _, s in xs if s) / len(xs), 1)})
    rows.sort(key=lambda x: (-x["n"], ampol_names.sort_key(x["product"])))
    allx = [x for xs in byp.values() for x in xs]
    total = None
    if allx:
        days = sorted(x[0] for x in allx)
        total = {"n": len(allx), "median": statistics.median(days), "p90": days[min(len(days) - 1, int(len(days) * 0.9))],
                 "sd_pct": round(100 * sum(1 for _, s in allx if s) / len(allx), 1)}
    return {"rows": rows, "pooled_n": len(pooled), "all": total, "min_n": min_n}


# ---------------------------------------------------------------------------
# 5. holders - who has what, and the 80/20
# ---------------------------------------------------------------------------

def holders(ctx, scope_barcodes=None, top=20):
    """Per (hirer, company): items on hire, priced value, unpriced count,
    oldest days, families held (gas / radio / tooling). Sorted by items
    (a RANKED table - say so on the page). Also the 80/20: how many
    holders carry 80% of the items and 80% of the priced value."""
    scope = _bcs(ctx, scope_barcodes)
    now = (ctx["reg_time"] or dt.datetime.now()).date()
    price = ctx["price"]
    h = {}
    for r in _reg(ctx, scope):
        if r["status"].lower() != "on hire":
            continue
        k = (r["hirer"], r["company"])
        d = h.setdefault(k, {"hirer": r["hirer"], "company": r["company"], "items": 0, "value": 0.0, "unpriced": 0,
                             "oldest": 0, "families": set(), "custody": bool(ampol_names.account_label(r["hirer"]) != r["hirer"]) if hasattr(ampol_names, "account_label") else False})
        d["items"] += 1
        p = price(r.get("desc_raw", r["desc"]))
        if p:
            d["value"] += p
        else:
            d["unpriced"] += 1
        if r["on_dt"]:
            d["oldest"] = max(d["oldest"], (now - r["on_dt"].date()).days)
        d["families"].add(report_family(r.get("desc_raw", r["desc"])))
    rows = sorted(h.values(), key=lambda d: (-d["items"], ampol_names.sort_key(d["company"]), ampol_names.sort_key(d["hirer"])))
    tot_i = sum(d["items"] for d in rows) or 1
    tot_v = sum(d["value"] for d in rows) or 1
    def n80(key, tot):
        cum = 0
        for i, d in enumerate(sorted(rows, key=lambda d: -d[key]), 1):
            cum += d[key]
            if cum >= 0.8 * tot:
                return i
        return len(rows)
    cross = [d for d in rows if len(d["families"]) >= 2]
    return {"rows": rows, "top": rows[:top], "holders": len(rows), "items": tot_i, "value": tot_v,
            "n80_items": n80("items", tot_i), "n80_value": n80("value", tot_v),
            "cross_family": sorted(cross, key=lambda d: (-d["items"], ampol_names.sort_key(d["company"])))}


# ---------------------------------------------------------------------------
# 6. counter rhythm and data quality
# ---------------------------------------------------------------------------

# WHY (03 Sep 2026, Andrew): the store runs two shifts, 04:00 to 12:30 and
# 09:00 to 17:30, and opens for business at 07:00. Before opening the first
# shift bumps the gas monitors, scans the return box (04:00 to 05:30) and
# makes up the pre-made packs. The rhythm pages read the hours in those
# windows, not a generic day/night split.
SHIFTS = {
    "preopen": ((4, 0), (6, 59), "before opening (04:00 to 06:59) - bump, return box, pre-made packs"),
    "trading": ((7, 0), (17, 29), "trading hours (07:00 to 17:29)"),
    "after": ((17, 30), (3, 59), "after hours (17:30 to 03:59)"),
    "shift1": ((4, 0), (12, 29), "shift 1 (04:00 to 12:30)"),
    "shift2": ((9, 0), (17, 29), "shift 2 (09:00 to 17:30)"),
    "opens": "07:00",
}


def in_window(t, key):
    (h0, m0), (h1, m1), _ = SHIFTS[key]
    x = t.hour * 60 + t.minute
    a, b = h0 * 60 + m0, h1 * 60 + m1
    return (a <= x <= b) if a <= b else (x >= a or x <= b)


def counter_rhythm(ctx, scope_barcodes=None):
    """7 x 24 matrices (Mon..Sun x hour) of draws and returns, the busiest
    hours, the split of draws by the store's own windows (SHIFTS: preopen,
    trading, after, shift1, shift2) and the older day/night split."""
    scope = _bcs(ctx, scope_barcodes)
    draws = [[0] * 24 for _ in range(7)]
    rets = [[0] * 24 for _ in range(7)]
    hours = collections.Counter()
    day = night = 0
    for t in _tx(ctx, scope):
        draws[t["st"].weekday()][t["st"].hour] += 1
        hours[t["st"].hour] += 1
        if 6 <= t["st"].hour < 18:
            day += 1
        else:
            night += 1
        if t["en"]:
            rets[t["en"].weekday()][t["en"].hour] += 1
    busiest = sorted(hours.items(), key=lambda kv: (-kv[1], kv[0]))[:3]
    win = {k: 0 for k in ("preopen", "trading", "after", "shift1", "shift2")}
    rwin = {k: 0 for k in win}
    for t in _tx(ctx, scope):
        for k in win:
            if in_window(t["st"], k):
                win[k] += 1
            if t["en"] and in_window(t["en"], k):
                rwin[k] += 1
    return {"draws": draws, "returns": rets, "busiest": busiest, "day": day, "night": night,
            "total": day + night, "windows": win, "return_windows": rwin, "shifts": SHIFTS}


def data_quality(ctx, scope_barcodes=None, short_minutes=6, mass=15):
    """What the log and the register do not agree on, and what looks like
    a scan habit rather than a hire:
      short:   hires closed inside `short_minutes` (out and back at once)
      mass:    one person drawing `mass` or more items inside one hour
      open_but_available: the log's newest movement for a barcode is still
               open but the register shows the item available
      onhire_no_log: on hire in the register with no movement in the log
    Counts and capped sample rows; nothing is corrected, only shown."""
    scope = _bcs(ctx, scope_barcodes)
    tx = _tx(ctx, scope)
    short = [t for t in tx if t["hours"] is not None and 0 <= t["hours"] * 60 < short_minutes]
    per = collections.Counter((t["who"], t["co"], t["st"].date(), t["st"].hour) for t in tx)
    mass_rows = sorted(((k, v) for k, v in per.items() if v >= mass), key=lambda kv: (-kv[1], kv[0][2]))
    latest = {}
    for t in tx:
        if not t["bc"]:
            continue
        k = t["bc"].upper()
        if k not in latest or t["st"] > latest[k]["st"]:
            latest[k] = t
    open_avail, onhire_nolog, pre_window = [], [], []
    w0 = ctx["tx_window"][0]
    for r in _reg(ctx, scope):
        k = r["barcode"].upper()
        st = r["status"].lower()
        if st == "available for hire" and k in latest and latest[k]["en"] is None:
            open_avail.append((r, latest[k]))
        if st == "on hire" and k not in latest:
            # issued before the log begins is not a gap - it is history
            if w0 and r["on_dt"] and r["on_dt"] < w0:
                pre_window.append(r)
            else:
                onhire_nolog.append(r)
    return {"short": short, "short_n": len(short), "short_minutes": short_minutes,
            "mass": mass_rows, "mass_n": len(mass_rows), "mass_threshold": mass,
            "open_but_available": open_avail, "onhire_no_log": onhire_nolog,
            "onhire_before_log": pre_window, "log_start": w0, "tx_n": len(tx)}


# ---------------------------------------------------------------------------
# 7. fast movers
# ---------------------------------------------------------------------------

def fast_movers(ctx, scope_barcodes=None, n=20):
    """Items by number of movements in the log, with the bay they sit in
    now (from the register) and their last movement. RANKED by movements."""
    scope = _bcs(ctx, scope_barcodes)
    cnt = collections.Counter()
    last = {}
    for t in _tx(ctx, scope):
        if not t["bc"]:
            continue
        k = t["bc"].upper()
        cnt[k] += 1
        if k not in last or t["st"] > last[k]:
            last[k] = t["st"]
    reg = {b.upper(): r for b, r in ctx["reg"].items()}
    rows = []
    for k, c in cnt.most_common(n):
        r = reg.get(k)
        rows.append({"barcode": r["barcode"] if r else k, "desc": r["desc"] if r else "(not on the register now)",
                     "unit": r["unit"] if r else "", "status": r["status"] if r else "", "moves": c, "last": last[k]})
    by_unit = collections.Counter()
    for k, c in cnt.items():
        r = reg.get(k)
        if r:
            by_unit[r["unit"] or "(no bay)"] += c
    return {"rows": rows, "by_unit": by_unit.most_common(12), "items_moved": len(cnt)}


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import time
    t0 = time.time()
    ctx = load_all()
    print(f"loaded {len(ctx['reg']):,} register rows, {len(ctx['tx']):,} movements in {time.time() - t0:.0f}s; pull {ctx['reg_time']}")
    qc = quarter_close(ctx)
    print(f"quarter close {qc['qend']:%d %b %Y}: {qc['n']} on hire cross {qc['threshold']} days unless returned (already over: {qc['already_over']}); value {qc['value']:,.0f} ({qc['unpriced']} unpriced)")
    ds = dead_stock(ctx)
    print(f"dead stock: {ds['n']:,} of {ds['available']:,} available items never moved; value {ds['value']:,.0f} ({ds['unpriced']} unpriced); top: {ds['by_product'][:3]}")
    rw = return_windows(ctx)
    print(f"return windows: all n={rw['all']['n']:,} median {rw['all']['median']:.2f} d p90 {rw['all']['p90']:.1f} d sd {rw['all']['sd_pct']}%; products with 10+ hires: {len(rw['rows'])}")
    for r in rw["rows"][:5]:
        print(f"   {r['product'][:40]:40s} n={r['n']:6,d} median {r['median']:5.2f} p90 {r['p90']:6.1f} sd {r['sd_pct']}%")
    ho = holders(ctx)
    print(f"holders: {ho['holders']} hold {ho['items']:,} items; 80% of items with {ho['n80_items']} holders, 80% of value with {ho['n80_value']}; cross-family holders {len(ho['cross_family'])}")
    cr = counter_rhythm(ctx)
    print(f"rhythm: busiest {[(f'{h:02d}:00', c) for h, c in cr['busiest']]}; day {cr['day']:,} night {cr['night']:,}")
    dq = data_quality(ctx)
    print(f"quality: short {dq['short_n']}, mass draws {dq['mass_n']}, open-but-available {len(dq['open_but_available'])}, on-hire-no-log {len(dq['onhire_no_log'])} (issued before the log: {len(dq['onhire_before_log'])})")
    fm = fast_movers(ctx)
    print(f"fast movers: {fm['rows'][0]['desc'][:40]} {fm['rows'][0]['moves']} moves; bays {fm['by_unit'][:3]}")
    ws = weekly_series(ctx)
    print(f"weekly series: {len(ws)} weeks, first {ws[0]['week']} issues {ws[0]['issues']}, last {ws[-1]['week']} issues {ws[-1]['issues']} partial={ws[-1]['partial']}")
    hr = headroom(ctx)[:3]
    print("headroom:", [(r['product'][:24], r['fleet'], r['peak'], r['headroom']) for r in hr])
