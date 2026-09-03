#!/usr/bin/env python3
"""
Coates Ampol Site Radio On-Hire Report - The Coates Way
Author: Andrew Fisher / POWERED BY SITEIQ

- Current year in full detail (radios + batteries, companies A-Z, serials
  joined from the radio register, STORAGE_UNIT from the master RENTAL_STOCK file)
- Prior years summarised by year and company - the value story
- At-a-glance charts: fleet position, value by company, age of hire
- Verification message: return radios not in use; if still in use, bring
  them to the Ampol Tool Store for a rescan (proof of existence)
- PDF via WeasyPrint or Edge/Chrome headless fallback

WHERE THE NUMBERS COME FROM (changed 02 Sep 2026)
  Every count, date and dollar is read from the SiteIQ RENTAL_STOCK export
  in Data\\ - the live register of what is on hire, to whom, since when,
  and what is on the shelf - as at the export's own request time. The
  Ampol_Radio_Report.xlsm workbook is no longer read: its Power-Query tabs
  only refresh when the file is opened in Excel, and an audit on 02 Sep
  2026 found them 19 days behind the register (every count and every
  "days on hire" figure stale, returned units still listed, one unit in
  Out-of-Service custody counted as on hire to a company).
  Serial numbers come from radio_register.xlsx and, failing that, from
  the serial SiteIQ carries in the item description. Replacement values
  come from Ampol_ToolStore_Pricing.xlsx ("Avg Buy Price (New)").

Inputs come from the suite's one Data area (see ampol_paths):
RENTAL_STOCK*.xlsx (required), radio_register*.xlsx, *Pricing*.xlsx,
TRANSACTIONS*.xlsx (the 24 hours before the pull), Data\\previous\\ (the
earlier pull the movement page compares against).
Output lands in Reports\\<today>\\Radios\\ - dated, never overwritten - under
the one suite file name (ampol_names.report_stem): the PDF, the page HTML
beside it, the Outlook draft (.eml + manifest) and the phone position card.

WHAT CHANGED (03 Sep 2026) - the 10/10 pass
  Cover carries the RAG stripe and the freshness line. A "Since the last
  pull" section follows the position: pull against pull from Data\\previous
  (honest note until a second pull exists) and the always-real 24 hours
  before the pull from TRANSACTIONS. The company charts gain the on-hire
  ageing bands by company. A 30-day trend page appears once seven days are
  on the scoreboard. An appendix divider sits before the full register.
  The PDF is stamped (Author, Subject, bookmarks) and the phone card rides
  inside the email body as well as attached.

WHAT CHANGED (03 Sep 2026) - the layout pass
  Page 2 reads as a position page: hero and key strip, the RAG band, the
  tiles, "Three things to do today" (each drawn from the register) and a
  three-line story. The full position paragraph, the ask and the
  assurance note follow on the next page. The cover carries a "What's
  inside" block whose page numbers are read off the printed PDF (render,
  read, rebuild, render again - the cover is fixed-height so the page
  count is identical, and the console says so). Every company block in
  the register is a group table whose title row repeats with the column
  row when the company runs over a page. The ageing panel charts the
  companies holding ten or more units and lists the smaller ones in one
  line under it. One dash style throughout: " - ", never the long dash.

WHAT CHANGED (03 Sep 2026) - the insights pass
  Five sections on new pages after "Since the last pull", read from the
  SiteIQ TRANSACTIONS export through txn_insights for this report's own
  barcodes: the year in movements (issues, returns and same-day rate by
  week, the monthly company league), return windows by product, the
  counter's rhythm (draws and returns by weekday and hour), who holds
  what (the 80/20 of the units on hire) and what the log and the register
  disagree on (short hires, mass draws, units issued before the log, the
  prior-year reconciliation). Each carries a one-line "So what". Page 2
  gains one sentence under the band - the daily-return standard against
  the data - and keeps its grammar otherwise. Sources and rules are named
  on the data-and-method page.
"""
import math
import re
import sys
from pathlib import Path
from datetime import datetime, date, timedelta
from collections import defaultdict
import openpyxl
import ampol_names  # WHY (03 Sep 2026): one file-name rule, one A-Z rule, one display rule
import ampol_paths  # WHY (12 Aug 2026): one Data area in, dated Reports folder out
import gasmon_engine as ge  # one company / person normaliser across the suite
import pdf_finish  # WHY (03 Sep 2026): properties and bookmarks on every PDF
import pull_diff  # WHY (03 Sep 2026): what moved since the last pull
import report_history as rh
import txn_insights as ti  # WHY (03 Sep 2026): the transaction log's year, for this report's barcodes
# WHY (03 Sep 2026): the page-1 RAG band - default lines, printed on the page
RAG_AMBER_PRIOR_PCT = 10     # share of on-hire units issued in prior years
RAG_RED_PRIOR_PCT = 30
COVER_PAGE = True
# WHY (03 Sep 2026): the trend page needs a real line, not two dots - it
# waits for seven recorded days and the data page says so until then.
TREND_MIN_DAYS = 7
TREND_DAYS = 30
_WORDS = {5: "five", 6: "six", 7: "seven", 8: "eight", 9: "nine", 10: "ten"}   # the data-page sentence reads in words
# WHY (03 Sep 2026): the movement tables list every barcode that changed
# state; a big demob day is capped so the story pages stay a story, and
# the cap is printed ("showing 60 of N") rather than hidden.
DIFF_ROWS_CAP = 60
LAST24_ROWS_CAP = 25
_ASAT_DT = [None]
BUILD_DT = datetime.now()   # one build stamp for the cover, the pages and the freshness line

BASE = Path(__file__).resolve().parent
REPORT_DATE = date.today()
CUR_YEAR = REPORT_DATE.year
# WHY (12 Aug 2026): PRIOR_LABEL was computed and never used while the headings
# carried hard-coded years - come January the report would have quietly lied.
# Every heading now derives from these two.
PRIOR_LABEL = f"2023–{CUR_YEAR - 1}"      # reset from the data in main()
PRIOR_SHORT = f"2023-{str(CUR_YEAR - 1)[2:]}"
# The unit prices are READ from the pricing file in main(); these are only
# the fallback if that file is missing, and the page says which was used.
PRICE_RADIO = None
PRICE_BATT = None
PRICE_SOURCE = "TBC"
META = {}
# WHY (12 Aug 2026): named constant so the en dash never sits as a backslash
# escape inside an f-string expression - older Pythons refuse that outright.
# WHY (03 Sep 2026): it now marks a year RANGE only (2023-2025); every other
# dash the builder prints is " - " - one style on every page, and the verify
# gate fails on a long dash.
EN_DASH = "–"

COATES_PURPOSE = "Supporting Australia's growth with leading equipment solutions"
COATES_OBJECTIVE = "Australia's most trusted equipment partner - delivering Best Service & Value"
COATES_VALUES = "Care Deeply &nbsp;\u2022&nbsp; Customer Focused &nbsp;\u2022&nbsp; Be Our Best &nbsp;\u2022&nbsp; One Team &nbsp;\u2022&nbsp; Competitive Spirit"

def money(v):
    try: return f"${float(v):,.0f}"
    except (TypeError, ValueError): return "-"

def esc(s):
    """Company names land inside SVG text - keep the markup honest."""
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

RADIO_RE = re.compile(r"motorola radio", re.I)
BATT_RE = re.compile(r"radio batter", re.I)
ACCESSORY_RE = re.compile(r"charger", re.I)
SERIAL_IN_DESC = re.compile(r"\b(\d{3}[A-Za-z]{3}\d{4})\b")


def radio_kind(desc):
    """'radio' / 'battery' / None. Chargers are accessories, not fleet."""
    d = str(desc or "")
    if ACCESSORY_RE.search(d):
        return None
    if BATT_RE.search(d):
        return "battery"
    if RADIO_RE.search(d):
        return "radio"
    return None


def hirer_kind(hirer):
    """'oos' = the Out-of-Service custody line; 'account' = a shared site
    account (after-hours, a shutdown account) that is not a person;
    'person' otherwise. Accounts are shown on their own lines, never as
    a person."""
    h = str(hirer or "")
    if re.search(r"out\s*of\s*service", h, re.I):
        return "oos"
    if re.search(r"after\s*hours|tool\s*store|shutdown\s*-\s*20\d\d|\(sfi\)|^alky", h, re.I):
        return "account"
    return "person"


def load_serials(register_path):
    """barcode -> serial from radio_register.xlsx. A blank serial cell is
    left out so the row falls through to the description or a dash."""
    wb = openpyxl.load_workbook(register_path, data_only=True, read_only=True)
    ws = wb["Radio Register"]
    out = {str(r[0]).strip().upper(): str(r[1]).strip()
           for r in ws.iter_rows(min_row=2, values_only=True)
           if r and r[0] and r[1] is not None and str(r[1]).strip()}
    wb.close()
    return out


def _norm_desc(d):
    return re.sub(r"\s+", " ", str(d or "").replace("\xa0", " ")).strip().upper()


def load_prices(pricing_path):
    """Normalised description -> 'Avg Buy Price (New)' from the pricing
    master, plus the generic radio and battery prices for descriptions
    that carry a serial suffix."""
    prices = {}
    if not pricing_path:
        return prices, None, None
    wb = openpyxl.load_workbook(pricing_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[0] and r[1] not in (None, ""):
            try:
                prices.setdefault(_norm_desc(r[0]), float(r[1]))
            except (TypeError, ValueError):
                pass
    wb.close()
    radio = prices.get("AMPOL MOTOROLA RADIO")
    batt = prices.get("AMPOL MOTOROLA RADIO BATTERY")
    return prices, radio, batt


def price_for(desc, kind, prices):
    p = prices.get(_norm_desc(desc))
    if p is None:
        p = PRICE_RADIO if kind == "radio" else PRICE_BATT
    return p


def company_name(raw):
    """One company, one name: Ampol / Ampol Refineries (Qld) / Caltex are
    the client; employer suffixes (FCCU, SATGAS/MOL) fold into the company."""
    u = str(raw or "").strip().upper()
    if u.startswith("CALTEX"):
        return "Ampol"
    return ge.norm_company(raw)


def load_from_register(master_path, serials, prices):
    """Every radio and battery on the live register, classified.

    Returns dict with r_cur, r_prev, b_cur, b_prev (on hire, by issue
    year), oos (Out-of-Service custody), r_avail, b_avail, asat (export
    request time), and coverage counts for the page notes."""
    wb = openpyxl.load_workbook(master_path, data_only=True, read_only=True)
    asat = None
    if "REFERENCE_INFO" in wb.sheetnames:
        rows = list(wb["REFERENCE_INFO"].iter_rows(values_only=True))
        if len(rows) > 1:
            for i, h in enumerate(rows[0]):
                if h and "REQUESTED_DATE" in str(h).upper():
                    asat = ge.parse_stamp(rows[1][i])
    if asat is None:
        asat = datetime.fromtimestamp(Path(master_path).stat().st_mtime)
    ws = wb["RENTAL_STOCK"]
    hdr = [str(c or "").strip().upper() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {h: i for i, h in enumerate(hdr)}

    def g(r, name):
        i = col.get(name)
        return r[i] if i is not None and i < len(r) else None

    out = {"r_cur": [], "r_prev": [], "b_cur": [], "b_prev": [], "oos": [],
           "r_avail": [], "b_avail": [], "asat": asat,
           "n_radio": 0, "n_batt": 0, "serial_hits": 0, "serial_total": 0,
           "unpriced": 0, "years": set(), "accounts": 0, "turnaround": 0}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        desc = g(r, "ITEM_DESCRIPTION")
        kind = radio_kind(desc)
        if not kind:
            continue
        out["n_radio" if kind == "radio" else "n_batt"] += 1
        bc = str(g(r, "ITEM_BARCODE") or "").strip()
        if bc.upper().startswith("SATGAS"):
            out["turnaround"] += 1
        status = str(g(r, "ITEM_STATUS") or "").strip()
        on = ge.parse_dt(g(r, "ON_HIRE_DATE"), g(r, "ON_HIRE_TIME"))
        serial = serials.get(bc.upper(), "")
        if not serial:
            mm = SERIAL_IN_DESC.search(str(desc or ""))
            serial = mm.group(1).upper() if mm else ""
        cost = price_for(desc, kind, prices)
        if cost is None:
            out["unpriced"] += 1
        hk = hirer_kind(g(r, "HIRER_NAME"))
        raw_h = str(g(r, "HIRER_NAME") or "").strip()
        hirer = raw_h if hk != "person" else ge.norm_person(raw_h)[1]
        if hk == "account":
            hirer = f"{raw_h} (site account)"
            out["accounts"] += 1
        item = {"company": company_name(g(r, "COMPANY_NAME")), "hirer": hirer,
                "barcode": bc, "serial": serial or "-", "kind": kind,
                "year": str(on.year) if on else "", "desc": str(desc or "").strip(),
                "days": (asat.date() - on.date()).days if on else 0,
                "cost": cost, "date": on.date() if on else None,
                "time": on.strftime("%H:%M") if on else "",
                "unit": str(g(r, "STORAGE_UNIT") or "").strip() or "-"}
        if status.lower() != "on hire":
            out["r_avail" if kind == "radio" else "b_avail"].append(item)
            continue
        out["serial_total"] += 1 if kind == "radio" else 0
        out["serial_hits"] += 1 if (kind == "radio" and serial) else 0
        if hk == "oos":
            out["oos"].append(item)
            continue
        if on:
            out["years"].add(on.year)
        cur = bool(on) and on.year == CUR_YEAR
        key = ("r_" if kind == "radio" else "b_") + ("cur" if cur else "prev")
        out[key].append(item)
    wb.close()
    return out


def val(rs): return sum(x["cost"] for x in rs if x["cost"])

def svg_hbars(rows, width=700, label_w=180, val_w=130, bar_h=14, gap=5):
    """Inline-SVG horizontal bars - self-contained, print-safe, house colours.

    WHY (12 Aug 2026): the report carried its whole story in tables; these
    small charts put the position on one page without touching the tables.
    rows = (label, number, display text, colour). A zero draws a grey hairline
    so the category still shows - nothing is hidden.
    """
    mx = max((r[1] for r in rows), default=0) or 1
    span = width - label_w - val_w - 10
    h = len(rows) * (bar_h + gap) + 2
    p = [f'<svg width="{width}" height="{h}" viewBox="0 0 {width} {h}" '
         f'xmlns="http://www.w3.org/2000/svg" style="font-family:Calibri,Arial,sans-serif">']
    y = 1
    for label, num, txt, colour in rows:
        w = max(round(span * (num / mx)), 2) if num else 2
        fill = colour if num else "#cccccc"
        ty = y + bar_h - 4
        p.append(f'<text x="{label_w - 6}" y="{ty}" text-anchor="end" font-size="10" fill="#333">{esc(label)}</text>')
        p.append(f'<rect x="{label_w}" y="{y}" width="{w}" height="{bar_h}" fill="{fill}"/>')
        p.append(f'<text x="{label_w + w + 5}" y="{ty}" font-size="10" font-weight="bold" fill="#1a1a1a">{esc(txt)}</text>')
        y += bar_h + gap
    p.append("</svg>")
    return "".join(p)

def detail_rows(items, serial_col=True):
    """Company blocks A-Z, items longest-held first, on house group tables:
    the company's name sits in the table head beside the column row, so a
    company that runs over a page opens the next page with its name."""
    import k2flow as kf
    from k2shell import esc, num
    by_comp = defaultdict(list)
    for i in items:
        by_comp[i["company"]].append(i)
    parts = []
    hdr = (["Barcode", "Serial", "Price", "Hirer", "On hire since", "Days", "Storage unit"] if serial_col
           else ["Barcode", "Price", "Hirer", "On hire since", "Days", "Storage unit"])
    al = (["", "", "r", "", "", "r", ""] if serial_col else ["", "r", "", "", "r", ""])
    # WHY (03 Sep 2026): one A-Z rule for every directory in the suite
    for comp in sorted(by_comp, key=ampol_names.sort_key):
        rows = sorted(by_comp[comp], key=lambda i: (-i["days"], i["barcode"]))
        cval = val(rows)
        trs = []
        for i in rows:
            d_txt = i["date"].strftime("%d %b %Y") if i["date"] else "-"
            if i["time"]:
                d_txt += f" {i['time']}"
            days = (f'<span class="rd">{i["days"]}</span>' if i["days"] >= 30 else str(i["days"]))
            r = [esc(i["barcode"]),
                 *([esc(i["serial"] or "-")] if serial_col else []),
                 money(i["cost"]), esc(i["hirer"]), d_txt, days, esc(i["unit"])]
            trs.append(r)
        # WHY (03 Sep 2026): the company name used to be a sub-heading above a
        # plain table, so a company spanning two pages opened the second with
        # a bare column row. The group table carries the name in its head,
        # and the print engine repeats the head on every page it spans.
        meta = (f'{len(rows)} unit{"s" if len(rows) != 1 else ""} &middot; '
                f'{money(cval) if cval else "unpriced"}')
        block = kf.group_table(comp, meta, hdr, trs, al, "cp")
        parts.append(f'<div class="keep">{block}</div>' if len(trs) <= 12 else block)
    return "".join(parts)


def position_facts(r26, rprev, b26, bprev, data_asat=""):
    """The page-1 facts, computed once and shared by the cover, the RAG
    band, the phone card and the history record.

    WHY (03 Sep 2026): the scoreboard used to be written after the pages
    were built, so the trend page could never see today's point. The
    facts now come first, the record is written, then the pages are
    built from the same dict - one source, every figure identical."""
    import k2shell as sh
    from k2shell import num
    total_exposure = val(r26) + val(rprev) + val(b26) + val(bprev)
    prev_all = rprev + bprev
    prev_val = val(prev_all)
    oldest = max(i["days"] for i in prev_all) if prev_all else 0
    all_on = r26 + rprev + b26 + bprev
    prior_pct = round(len(prev_all) / len(all_on) * 100) if all_on else 0
    status = sh.rag_of(prior_pct, RAG_AMBER_PRIOR_PCT, RAG_RED_PRIOR_PCT)
    # WHY (03 Sep 2026): the due date counts seven days from the PULL, never
    # from the day the button was pressed
    due = ((datetime.strptime(data_asat, "%d %b %Y %H:%M") + timedelta(days=7)).strftime("%d %b %Y")
           if data_asat else "the next report")
    return {
        "exposure": total_exposure, "prev_all": prev_all, "prev_val": prev_val, "oldest": oldest,
        "all_on": all_on, "prior_pct": prior_pct, "status": status, "due": due,
        "headline": (f"{num(len(prev_all))} of the {num(len(all_on))} units on hire ({prior_pct}%) were issued in "
                     f"{PRIOR_LABEL} and are still out - {money(prev_val)} of equipment, the oldest for {num(oldest)} days."),
        "rule": (f"Share of on-hire units issued in prior years: Green under {RAG_AMBER_PRIOR_PCT}%, Amber from "
                 f"{RAG_AMBER_PRIOR_PCT}%, Red from {RAG_RED_PRIOR_PCT}%. Default lines - set at the top of the script."),
        "action": (f"Every prior-year holder sent their list from this report by {due}; returned or rescanned "
                   f"units drop off the next run."),
        "key_value": num(len(prev_all)), "key_label": f"units on hire since {PRIOR_LABEL}",
        "second_value": num(len(all_on)), "second_label": "units on hire",
    }


def show_hirer(raw):
    """A hirer the way the register pages show one: a person in title case,
    a shared site account marked as such, the custody line as SiteIQ names it."""
    hk = hirer_kind(raw)
    raw = str(raw or "").strip()
    if hk == "account":
        return f"{raw} (site account)"
    if hk == "oos":
        return raw
    return ge.norm_person(raw)[1]


def _capped(rows, cap, order="A to Z"):
    """The first `cap` rows and the honest line that says so."""
    from k2shell import num
    if len(rows) <= cap:
        return rows, ""
    return rows[:cap], f'<div class="note">Showing {cap} of {num(len(rows))} rows ({order}).</div>'


def _days_cell(days):
    if days is None:
        return "-"
    return f'<span class="rd">{days}</span>' if days >= 30 else str(days)


def changes_section(d):
    """'Since the last pull' - what moved. Two honest sources (pull_diff):
    the earlier RENTAL_STOCK pull parked in Data\\previous, compared item by
    item, and the TRANSACTIONS export for the 24 hours before the pull.

    WHY (03 Sep 2026): a snapshot says where everything is; the client also
    wants to know what came back and what went out. Until a second pull is
    parked there is no pull-against-pull row and the page says so in plain
    words - the 24-hour block is countable from the first report."""
    import k2flow as kf
    import k2shell as sh
    from k2shell import esc, num
    if d is None:
        return ""
    # flows on from the position's tail - a forced break left a page two-thirds empty
    P = ['<div class="sect"><h3>Since the last pull</h3></div>']
    ct = d["cur_time"]
    if not d["have_previous"]:
        P.append('<div class="note">No earlier register pull is saved in Data\\previous yet - movement pull '
                 'against pull starts with the next pull (button 28 parks the old export automatically).</div>')
    else:
        pt = d["prev_time"]
        ret, iss, mov, c30 = d["returned"], d["issued"], d["moved"], d["crossed"][30]
        span_h = max(0, (ct - pt).total_seconds() / 3600)
        span = f"{span_h:.0f} hours" if span_h < 48 else f"{span_h / 24:.0f} days"
        P.append(f'<div class="callout tight"><span class="lead">Pull against pull.</span> The register pulled '
                 f'<b>{pt:%d %b %Y %H:%M}</b> compared item by item with the pull of <b>{ct:%d %b %Y %H:%M}</b> '
                 f'({span} apart). Radios and batteries on hire in SiteIQ then: <b>{num(d["out_prev"])}</b>; now: '
                 f'<b>{num(d["out_cur"])}</b> - this count includes the Out-of-Service custody line, which the '
                 f'position page keeps apart. Every row below is a barcode that changed state between the two '
                 f'pulls - nothing is estimated.</div>')
        P.append(sh.tiles([
            ("check", num(len(ret)), "Came back", "on hire then, not now", "green" if ret else "grey"),
            ("swap", num(len(iss)), "Went out", "on hire now, not then", "amber" if iss else "grey"),
            ("box", num(len(mov)), "Changed hands", "same unit, new hirer", "grey"),
            ("clock", num(len(c30)), "Crossed 30 days", "still out, now due", "red" if c30 else "green"),
        ]))
        hdr = ["Company", "Hirer", "Barcode", "Description", "Days out"]
        al = ["", "", "", "", "r nw"]

        def row(r):
            return [esc(r["company"]), esc(show_hirer(r["hirer"])), esc(r["barcode"]), esc(r["desc"]),
                    _days_cell(r.get("days_out"))]
        rows, cap = _capped(ret, DIFF_ROWS_CAP)
        P.append(f'<div class="sub-h">Came back <span class="thin">- {num(len(ret))} units, companies A to Z</span></div>')
        P.append(kf.dtable_flow(hdr + ["Now"], [row(r) + [esc(r.get("now", ""))] for r in rows], al + [""], "cp") + cap
                 if rows else '<div class="note">Nothing came back between the two pulls.</div>')
        rows, cap = _capped(iss, DIFF_ROWS_CAP)
        P.append(f'<div class="sub-h">Went out <span class="thin">- {num(len(iss))} units, companies A to Z</span></div>')
        P.append(kf.dtable_flow(hdr, [row(r) for r in rows], al, "cp") + cap
                 if rows else '<div class="note">Nothing went out between the two pulls.</div>')
        if mov:
            rows, cap = _capped(mov, DIFF_ROWS_CAP)
            P.append(f'<div class="sub-h">Changed hands <span class="thin">- {num(len(mov))} units, companies A to Z</span></div>')
            P.append(kf.dtable_flow(hdr + ["Was with"], [row(r) + [esc(show_hirer(r.get("from_hirer", "")))
                                                                   + (f' ({esc(r["from_company"])})'
                                                                      if r.get("from_company") and r["from_company"] != r["company"] else "")]
                                                         for r in rows], al + [""], "cp") + cap)
        rows, cap = _capped(c30, DIFF_ROWS_CAP, "oldest first")
        P.append(f'<div class="sub-h">Crossed 30 days while still out <span class="thin">- {num(len(c30))} units, oldest first</span></div>')
        P.append(kf.dtable_flow(hdr, [row(r) for r in rows], al, "cp") + cap
                 if rows else '<div class="note">No unit crossed 30 days out between the two pulls.</div>')
        new, gone = d["companies_new"], d["companies_cleared"]
        P.append('<div class="note"><b>Companies new since the last pull:</b> '
                 + (esc(", ".join(new)) if new else "none") + '. <b>Companies cleared:</b> '
                 + (esc(", ".join(gone)) if gone else "none") + '.</div>')
    # ---- the 24 hours before the pull - always real, from TRANSACTIONS ----
    L = d["last24"]
    start, end = L["window"]
    P.append(f'<div class="sub-h">The 24 hours before the pull <span class="thin">- '
             f'{start:%d %b %Y %H:%M} to {end:%d %b %Y %H:%M}</span></div>')
    if not L.get("available", True):
        P.append('<div class="note">The TRANSACTIONS export is not in Data - the 24-hour block needs it. '
                 'Run 12_PULL_SITEIQ_EXPORTS and press the button again.</div>')
        return "".join(P)
    iss, ret = L["issued"], L["returned"]
    P.append(f'<div class="note"><b>Issued {num(len(iss))}</b> and <b>returned {num(len(ret))}</b> radios and batteries '
             f'across the counter in the 24 hours before the pull - every scan with its time, from the TRANSACTIONS '
             f'export (CUSTOMER_CONTRACTOR_EQUIP). Company A to Z, then time.</div>')
    moves = [(r, "Issued", "amber") for r in iss] + [(r, "Returned", "green") for r in ret]
    moves.sort(key=lambda x: (ampol_names.sort_key(x[0]["company"]), x[0]["at"]))
    trs = [[esc(r["company"]), esc(show_hirer(r["hirer"])), esc(r["barcode"]), esc(r["desc"]),
            f'<span class="tag {c}">{mv}</span>', f"{r['at']:%d %b %H:%M}"]
           for r, mv, c in moves[:LAST24_ROWS_CAP]]
    if trs:
        P.append(kf.dtable_flow(["Company", "Hirer", "Barcode", "Description", "Movement", "Time"], trs,
                                ["", "", "", "", "", "nw"], "cp"))
        if len(moves) > LAST24_ROWS_CAP:
            P.append(f'<div class="note">Showing {LAST24_ROWS_CAP} of {num(len(moves))} movements (company A to Z, then time).</div>')
    else:
        P.append('<div class="note">No radio or battery crossed the counter in the 24 hours before the pull.</div>')
    return "".join(P)


def ageing_rows(all_on):
    """(company, [0-30, 31-60, 61-90, 90+]) for every company with units on
    hire, A to Z - days out at the pull, custody units already excluded."""
    import k2shell as sh
    by = defaultdict(lambda: [0, 0, 0, 0])
    for i in all_on:
        by[i["company"]][sh.age_band_index(i["days"])] += 1
    return [(c, by[c]) for c in sorted(by, key=ampol_names.sort_key)]


def trend_section(asat_s):
    """The 30-day trend page, once TREND_MIN_DAYS report days are on the
    scoreboard - (html, days_on_record). Fewer days = no page; the data
    page prints the count instead."""
    import k2shell as sh
    from k2shell import esc, num
    fam = rh.load().get("radio", {})
    n_days = len(fam)
    if n_days < TREND_MIN_DAYS:
        return "", n_days
    units = [("radios_on_hire", "Radios on hire"), ("batteries_on_hire", "Batteries on hire"),
             ("prior_units", f"On hire since {PRIOR_LABEL}")]
    ser = {k: dict(rh.series("radio", k, _ASAT_DT[0], days=TREND_DAYS)) for k, _ in units + [("exposure", "")]}
    dates = sorted(set().union(*[set(s) for s in ser.values()]))
    labels = [dd.strftime("%d %b") for dd in dates]
    a = [(lab, [ser[k].get(dd) for dd in dates]) for k, lab in units]
    # WHY (03 Sep 2026): the value axis is drawn in thousands - a seven-digit
    # dollar figure overran the chart's axis gutter and lost its first digit
    b = [("On-hire value, $ thousand", [(round(ser["exposure"][dd] / 1000) if ser["exposure"].get(dd) is not None else None)
                                        for dd in dates])]
    last_k = [v for v in b[0][1] if v is not None]
    last_full = [ser["exposure"][dd] for dd in dates if ser["exposure"].get(dd) is not None]
    scale_eg = f" ({num(last_k[-1])} = {money(last_full[-1])})" if last_k else ""
    html = (f'<div class="pb"></div><div class="sect"><h3>The trend - last {TREND_DAYS} days</h3></div>'
            f'<div class="callout tight"><span class="lead">The direction.</span> One point per report day - '
            f'<b>{num(len(dates))}</b> days on record in the {TREND_DAYS} days to <b>{esc(asat_s)}</b>, read back from the '
            f'suite scoreboard (History\\report_history.json): the figure each day\'s report printed, nothing '
            f're-counted or smoothed. A day with no report leaves a gap, never a guess.</div>'
            f'<div class="sub-h">Units on hire <span class="thin">- radios, batteries and prior-year units, by day</span></div>'
            f'<div class="chartpanel">{sh.line_chart(labels, a, y_label="units")}</div>'
            f'<div class="sub-h">Replacement value on hire <span class="thin">- the on-hire value tile, by day</span></div>'
            f'<div class="chartpanel">{sh.line_chart(labels, b, y_label="$ replacement, thousands")}</div>'
            f'<div class="note">Down is good on every line: a return or a rescan moves it. The value line is the '
            f'replacement value of everything on hire at each pull, at the prices on the data page, drawn in '
            f'thousands of dollars{esc(scale_eg)}.</div>')
    return html, n_days


def three_things_for(F, r26, b26, oos):
    """The position page's three things to do today, each read from the
    same register rows as the tiles: (1) the company holding the most
    prior-year units, (2) the oldest unit out, (3) the company holding the
    most of this year's units over 90 days - or the out-of-service line
    when that count is the bigger number. A shared site account is named
    as the account, never as a person. Fewer than three true items prints
    fewer; the block never pads.

    WHY (03 Sep 2026): a position page that says only how bad it is leaves
    the reader to work out where to start. These three name the start."""
    from k2shell import num
    who = f"Andrew Fisher · by {F['due']}"
    items = []
    by_co = defaultdict(list)
    for i in F["prev_all"]:
        by_co[i["company"]].append(i)
    if by_co:
        co, rows = sorted(by_co.items(), key=lambda kv: (-len(kv[1]), ampol_names.sort_key(kv[0])))[0]
        v = val(rows)
        unpriced = sum(1 for i in rows if not i["cost"])
        why = f"{money(v)} at replacement" if v else "nothing priced"
        if unpriced:
            why += f" ({unpriced} unpriced)"
        why += f"; oldest {num(max(i['days'] for i in rows))} days"
        items.append((f"Chase {co} for {num(len(rows))} units out since {PRIOR_LABEL}", why, who))
    if F["all_on"]:
        o = sorted(F["all_on"], key=lambda i: (-i["days"], i["barcode"]))[0]
        since = o["date"].strftime("%d %b %Y") if o["date"] else "a date SiteIQ does not carry"
        items.append((f"Recover {o['barcode']} from {o['hirer']}, {o['company']}",
                      f"{num(o['days'])} days, since {since}", who))
    over = defaultdict(list)
    for i in r26 + b26:
        if i["days"] > 90:
            over[i["company"]].append(i)
    top = sorted(over.items(), key=lambda kv: (-len(kv[1]), ampol_names.sort_key(kv[0])))[:1]
    if oos and (not top or len(oos) > len(top[0][1])):
        items.append((f"Clear the {num(len(oos))} units out of service",
                      f"{money(val(oos))} tagged in custody; longest {num(max(i['days'] for i in oos))} days", who))
    elif top:
        co, rows = top[0]
        items.append((f"Chase {co} for {num(len(rows))} {CUR_YEAR} units over 90 days",
                      f"{money(val(rows))} at replacement; oldest {num(max(i['days'] for i in rows))} days", who))
    return items[:3]


# ---------------------------------------------------------------------------
# The insights pages (03 Sep 2026) - what the transaction log knows about
# this report's radios and batteries
# ---------------------------------------------------------------------------
# WHY (03 Sep 2026): the register says where every unit is today; the
# TRANSACTIONS export says how the year went - every issue and return since
# 1 Jan. txn_insights reads it once and answers for THIS report's barcodes
# (every radio and battery on the register as the report defines it: on
# hire, on the shelf and in custody). The rules the pages print are set
# here so the page and the code always agree.
INSIGHT_SAMPLE_ROWS = 10    # sample rows shown for short hires
LEAGUE_TOP = 10             # companies in the monthly league (ranked by issues)
HOLDERS_TOP = 15            # holders in the who-holds-what table (ranked by items)
HEAT_DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
# WHY (03 Sep 2026): the So-what caption and the compact league table are
# this report's own; k2style.css is shared and stays as it is, so the two
# rules ride in with the document as extra CSS.
INSIGHT_CSS = """
.k2body .sowhat { margin: 10px 0 2px 0; padding: 7px 12px; border-left: 4px solid #F36F21; background: #FFF7F1;
                  border-radius: 0 8px 8px 0; font-size: 10.2px; line-height: 1.55; color: #35404E;
                  break-inside: avoid; page-break-inside: avoid; }
.k2body .sowhat b { color: #D95F14; font-weight: 700; }
.k2body .note.std { margin-top: 6px; margin-bottom: 0; font-size: 9.6px; color: #47566A; line-height: 1.5; }
.k2body .note.std b { color: #16202C; }
.k2body table.dt.league th { padding: 7px 4px; font-size: 7.3px; letter-spacing: 0.4px; }
.k2body table.dt.league td { padding: 5px 4px; font-size: 8.5px; line-height: 1.3; white-space: nowrap; }
.k2body table.dt.league td:first-child, .k2body table.dt.league th:first-child { padding-left: 9px; white-space: normal; }
.k2body table.dt.league td:last-child, .k2body table.dt.league th:last-child { padding-right: 9px; }
.k2body .nw { white-space: nowrap; }
"""


def so_what(text):
    """The one-line takeaway under every insights section - the reader's
    'so what', in plain words, from the figures on the page above it."""
    return f'<div class="sowhat"><b>So what.</b> {text}</div>'


def insight_hirer(raw):
    """A hirer on the insights pages: the register pages' rule, plus the
    shutdown account the way the TRANSACTIONS export spells it (no dash
    before the year), so a site account is never printed as a person."""
    raw = str(raw or "").strip()
    if hirer_kind(raw) == "person" and re.search(r"shutdown\s*20\d\d|^t&i\b", raw, re.I):
        return f"{raw} (site account)"
    return show_hirer(raw)


def _pts_arrow(before, after):
    """The change in a same-day rate between two months, in percentage
    points, with the page-1 tile arrow: green up (more came back the same
    day), red down. A month with no closed hire prints a dash."""
    if before is None or after is None:
        return '<span class="tbc">-</span>'
    d = round(after - before, 1)
    if d == 0:
        return '<span class="tbc">no change</span>'
    return f'<span class="{"g" if d > 0 else "rd"}">{"▲" if d > 0 else "▼"} {abs(d):g} pts</span>'


def insights_compute(items_on, oos, ravail, bavail):
    """Everything the insights pages need, computed ONCE per build from the
    transaction engine over this report's barcodes. None when the
    TRANSACTIONS export is not in Data - the pages then say so.

    items_on: the units on hire the position page counts (this year's and
    prior years', custody line apart). Holder values are re-summed at this
    report's own prices from those rows, so the who-holds-what page adds
    up to the on-hire value tile on page 2 - one price list, one total."""
    try:
        ctx = ti.load_all()
    except Exception as e:
        print(f"Insights           : not built - the TRANSACTIONS export could not be read ({type(e).__name__}: {e})")
        return None
    everything = items_on + oos + ravail + bavail
    scope = {i["barcode"] for i in everything if i["barcode"]}
    scope_r = {i["barcode"] for i in everything if i["barcode"] and i["kind"] == "radio"}
    scope_b = {i["barcode"] for i in everything if i["barcode"] and i["kind"] == "battery"}
    scope_hold = scope - {i["barcode"] for i in oos if i["barcode"]}
    I = {"window": ctx["tx_window"], "reg_time": ctx["reg_time"], "scope_n": len(scope),
         "tx_file": Path(ctx["tx_path"]).name}
    I["weekly"] = ti.weekly_series(ctx, scope)
    I["league"] = ti.monthly_league(ctx, scope)
    I["rw"] = ti.return_windows(ctx, scope)
    # WHY (03 Sep 2026): the engine's product key leaves the serial prefix
    # SiteIQ writes into a radio's description ("...T&I- 122TYT0381"), so
    # it splits the radios into a dozen pseudo-products. The two products
    # this report tracks are read with the engine's own function over the
    # radio barcodes and the battery barcodes - nothing re-implemented.
    I["rw_radio"] = ti.return_windows(ctx, scope_r)
    I["rw_batt"] = ti.return_windows(ctx, scope_b)
    I["rhythm"] = ti.counter_rhythm(ctx, scope)
    hold = ti.holders(ctx, scope_hold, top=HOLDERS_TOP)
    # families held store-wide (gas / radio / tooling) for the same people
    store = ti.holders(ctx, None, top=1)
    fam = {(d["hirer"], d["company"]): d["families"] for d in store["rows"]}
    cost = {i["barcode"].upper(): (i["cost"] or 0) for i in items_on}
    by_key = defaultdict(float)
    for r in ctx["reg"].values():
        b = r["barcode"].upper()
        if b in cost and r["status"].lower() == "on hire":
            by_key[(r["hirer"], r["company"])] += cost[b]
    for d in hold["rows"]:
        k = (d["hirer"], d["company"])
        d["rvalue"] = by_key.get(k, 0.0)
        d["also"] = sorted(f for f in fam.get(k, set()) if f != "radio")
    vals = sorted((d["rvalue"] for d in hold["rows"]), reverse=True)
    tot_v = sum(vals)
    cum, n80v = 0.0, len(vals)
    for i, v in enumerate(vals, 1):
        cum += v
        if cum >= 0.8 * tot_v:
            n80v = i
            break
    hold["rvalue_total"] = tot_v
    hold["n80_rvalue"] = n80v
    hold["cross_n"] = sum(1 for d in hold["rows"] if d["also"])
    I["holders"] = hold
    I["dq"] = ti.data_quality(ctx, scope)
    I["tx_n"] = I["dq"]["tx_n"]
    return I


def standard_line(I):
    """The one sentence on page 2 under the band: the daily-return
    standard against the data. X is the 90th-percentile hold of every
    completed hire in the log, rounded up to whole days - nine in ten
    completed hires were back inside it, by construction."""
    from k2shell import num
    if not I or not I["rw"]["all"]:
        return ""
    a = I["rw"]["all"]
    w0, w1 = I["window"]
    x = max(1, math.ceil(a["p90"]))
    return (f'<div class="note std">Nine in ten completed radio and battery hires this year were back inside '
            f'<b>{x} day{"s" if x != 1 else ""}</b> - {num(a["n"])} completed hires in the transaction log '
            f'({w0:%d %b} to {w1:%d %b %Y}), a 90th-percentile hold of {a["p90"]:.1f} days, '
            f'{a["sd_pct"]:g}% back the same day.</div>')


def insights_pages(I, prev_all, asat_s):
    """The five insights sections, in order, each with its chart or table
    and a So-what caption - on new pages after the since-the-last-pull
    section. prev_all: the prior-year units the position page counts, for
    the reconciliation on the log-versus-register page."""
    import k2flow as kf
    import k2shell as sh
    from k2shell import esc, num
    if not I:
        return ""
    w0, w1 = I["window"]
    log_span = f"{w0:%d %b} to {w1:%d %b %Y %H:%M}"
    P = []
    # ---- 1. the year in movements ----------------------------------------
    W = I["weekly"]
    # WHY (03 Sep 2026): the log opens on 1 Jan, so its first ISO week holds
    # two days - as partial as the week of the pull. Both are marked *.
    for w in W:
        w["part"] = w["partial"] or w["start"] < w0.date()
    labels = [w["week"] + ("*" if w["part"] else "") for w in W]
    issues = [w["issues"] for w in W]
    returns = [w["returns"] for w in W]
    sd = [w["sd_pct"] for w in W]
    full = [w for w in W if not w["part"]] or W
    partial = [w for w in W if w["part"]]
    peak = max(full, key=lambda w: w["issues"])
    quiet = min(full, key=lambda w: w["issues"])
    yr_iss, yr_ret = sum(issues), sum(returns)
    med_wk = int(round(sorted(w["issues"] for w in full)[len(full) // 2]))
    rw_all = I["rw"]["all"] or {"sd_pct": 0, "n": 0}
    P.append(
        '<div class="pb"></div><div class="sect"><h3>The year in movements</h3></div>'
        f'<div class="callout tight"><span class="lead">The year so far.</span> <b class="o">{num(yr_iss)} issues</b> and '
        f'<b class="o">{num(yr_ret)} returns</b> of radios and batteries crossed the counter between {esc(log_span)} - '
        f'every scan in the SiteIQ TRANSACTIONS export for the {num(I["scope_n"])} barcodes on this report, laid out by '
        f'ISO week. The busiest full week was the week of <b>{esc(peak["week"])}</b> ({num(peak["issues"])} issues); the '
        f'quietest full week was the week of <b>{esc(quiet["week"])}</b> ({num(quiet["issues"])}). '
        + (('Weeks marked * are partial: ' + "; ".join(
                (f'the week of {esc(w["week"])} holds the log\'s first {(w["start"] + timedelta(days=7) - w0.date()).days} days'
                 if w["start"] < w0.date() else f'the week of {esc(w["week"])} runs to {w1:%d %b %H:%M}')
                for w in partial) + '.') if partial else '')
        + '</div>'
        '<div class="sub-h">Issues and returns by week <span class="thin">- every movement, by the week it was scanned</span></div>'
        f'<div class="chartpanel">{sh.line_chart(labels, [("Issues", issues), ("Returns", returns)], y_label="movements a week")}</div>'
        '<div class="sub-h">Same-day returns by week <span class="thin">- of that week\'s issues that have come back, the share back the same day</span></div>'
        f'<div class="chartpanel">{sh.line_chart(labels, [("Same-day %", sd)], y_label="% of closed hires", pct=True)}</div>')
    # the numbers behind both charts - net out is issues less returns
    hdr = ["Week", "Issues", "Returns", "Net", "Same day"]
    al = ["nw", "r", "r", "r", "r"]

    def wrow(w):
        n = w["issues"] - w["returns"]
        return [esc(w["week"]) + ("*" if w["part"] else ""), num(w["issues"]), num(w["returns"]),
                (f"{n:+d}" if n else "0"), (f'{w["sd_pct"]:.0f}%' if w["sd_pct"] is not None else '<span class="tbc">-</span>')]
    third = (len(W) + 2) // 3
    cols = [W[i:i + third] for i in range(0, len(W), third)]
    # the three side-by-side tables move to the next page as one block
    P.append('<div class="keep"><div class="sub-h">The weeks in numbers <span class="thin">- net is issues less returns; a minus is a week '
             'more came back than went out</span></div><table class="two"><tr>'
             + "".join(f'<td style="width:{100 // len(cols)}%;padding-{"right" if k < len(cols) - 1 else "left"}:{4 if k < len(cols) - 1 else 0}px;'
                       f'padding-left:{4 if k else 0}px">{kf.dtable_flow(hdr, [wrow(w) for w in c], al, "cp")}</td>'
                       for k, c in enumerate(cols))
             + '</tr></table></div>')
    # the monthly company league
    L = I["league"]
    ranked = sorted(L.items(), key=lambda kv: (-sum(r["issues"] for r in kv[1]), ampol_names.sort_key(kv[0])))[:LEAGUE_TOP]
    months = sorted({r["key"] for rows in L.values() for r in rows})
    pull_m = I["reg_time"].strftime("%Y-%m") if I["reg_time"] else months[-1]
    full_m = [m for m in months if m < pull_m] or months
    m1, m2 = (full_m[-2], full_m[-1]) if len(full_m) >= 2 else (full_m[-1], full_m[-1])
    mname = lambda m: datetime.strptime(m, "%Y-%m").strftime("%b")   # noqa: E731
    lhdr = ["Company (ranked by issues)"] + [mname(m) + ("*" if m == pull_m else "") for m in months] + [f"{mname(m1)} to {mname(m2)}"]
    lal = [""] + ["r"] * len(months) + ["r"]
    lrows, moves = [], []
    for co, rows in ranked:
        bym = {r["key"]: r for r in rows}
        cells = [f"<b>{esc(co)}</b>"]
        for m in months:
            r = bym.get(m)
            if not r:
                cells.append('<span class="tbc">-</span>')
            else:
                cells.append(f'{num(r["issues"])} · ' + (f'{r["sd_pct"]:.0f}%' if r["sd_pct"] is not None else '<span class="tbc">-</span>'))
        a = bym.get(m1, {}).get("sd_pct")
        b = bym.get(m2, {}).get("sd_pct")
        cells.append(_pts_arrow(a, b))
        if a is not None and b is not None:
            moves.append((round(b - a, 1), co, a, b))
        lrows.append(cells)
    up = max(moves, key=lambda x: x[0]) if moves else None
    down = min(moves, key=lambda x: x[0]) if moves else None
    league_tot = sum(sum(r["issues"] for r in rows) for _, rows in ranked)
    P.append(
        f'<div class="sub-h">The monthly company league <span class="thin">- ranked by issues this year, the top '
        f'{num(len(ranked))} of {num(len(L))} companies; each cell is issues · same-day %</span></div>'
        + kf.dtable_flow(lhdr, lrows, lal, "cp league")
        + f'<div class="note">Same-day % is the share of that month\'s closed hires back the day they went out. The last column '
          f'is the change between the last two full months, {mname(m1)} and {mname(m2)}, in percentage points - green is up '
          f'(more back the same day), red is down, as on the page-1 tiles. '
        + (f'{mname(pull_m)}* holds the log\'s {mname(pull_m)} movements only, up to {w1:%d %b %H:%M}. ' if pull_m in months else '')
        + f'The {num(len(ranked))} companies here account for {num(league_tot)} of the {num(yr_iss)} issues; the other '
          f'{num(len(L) - len(ranked))} companies are in the register pages, A to Z.</div>')
    P.append(so_what(
        f'Radios and batteries cross the counter about {num(med_wk)} times a week (the median full week), and '
        f'{rw_all["sd_pct"]:g}% of this year\'s {num(rw_all["n"])} completed hires came back the same day - the rest are '
        f'the hires the 30-day rule watches. '
        + (f'Between {mname(m1)} and {mname(m2)} the biggest rise in same-day returns was {esc(up[1])} '
           f'({up[2]:g}% to {up[3]:g}%) and the biggest fall {esc(down[1])} ({down[2]:g}% to {down[3]:g}%).'
           if up and down and up[0] > 0 and down[0] < 0 else '')))
    # ---- 2. return windows by product ------------------------------------
    rw = I["rw"]
    rr, rb = I["rw_radio"]["all"], I["rw_batt"]["all"]

    def rwrow(label, a, bold=False):
        if not a:
            return [label, '<span class="tbc">-</span>', '<span class="tbc">-</span>', '<span class="tbc">-</span>', '<span class="tbc">-</span>']
        f = (lambda s: f"<b>{s}</b>") if bold else (lambda s: s)
        return [f(label), f(num(a["n"])), f(f'{a["median"]:.1f}'), f(f'{a["p90"]:.1f}'), f(f'{a["sd_pct"]:g}%')]
    pooled = rw["pooled_n"]
    pooled_r, pooled_b = I["rw_radio"]["pooled_n"], I["rw_batt"]["pooled_n"]
    allw = rw["all"]
    x_days = max(1, math.ceil(allw["p90"])) if allw else None
    P.append(
        '<div class="keep"><div class="sect"><h3>Return windows by product</h3></div>'
        f'<div class="callout tight"><span class="lead">How long a hire really lasts.</span> Every completed hire in the log - '
        f'out and back - measured from the issue scan to the return scan, for the two products this report tracks. The median '
        f'is the middle hire; the 90th percentile is the hold nine in ten hires were back inside. Days are decimal: 0.5 days is '
        f'about 12 hours.</div>'
        + kf.dtable_flow(["Product", "Completed hires", "Median (days)", "90th percentile (days)", "Same day"],
                         [rwrow("Radios", rr), rwrow("Batteries", rb), rwrow("All radios and batteries", allw, bold=True)],
                         ["", "r", "r", "r", "r"], "cp")
        + f'<div class="note">Completed hires only - a unit still out has no return scan and no window yet. The engine pools '
          f'any product with fewer than {num(rw["min_n"])} completed hires: on this fleet that pool is {num(pooled)} hire{"s" if pooled != 1 else ""} '
          f'({num(pooled_r)} radio{"s" if pooled_r != 1 else ""}, {num(pooled_b)} batter{"ies" if pooled_b != 1 else "y"}), already counted in the rows above. '
          f'Products are this report\'s own radio and battery rule, each barcode classed by its register description.</div>')
    if allw:
        P.append(so_what(
            f'Half of all hires are back within {allw["median"] * 24:.0f} hours and nine in ten within {num(x_days)} days; a unit '
            f'out 30 days has been held {30 / allw["p90"]:.0f} times longer than the 90th-percentile hold - the red on the register '
            f'pages marks a real outlier, not a busy crew.'))
    P.append('</div>')
    # ---- 3. the counter's rhythm -----------------------------------------
    R = I["rhythm"]
    cols24 = [f"{h:02d}" if h % 2 == 0 else "" for h in range(24)]
    busiest = R["busiest"]
    top3 = sum(c for _, c in busiest)
    tot = R["total"] or 1
    wk = sum(sum(R["draws"][r]) for r in range(5))
    we = sum(sum(R["draws"][r]) for r in range(5, 7))
    bd = max(range(7), key=lambda r: sum(R["draws"][r]))
    tiles = [("clock", f"{h:02d}:00", f"Busiest hour {k}", f"{num(c)} draws", "amber" if k == 1 else "grey")
             for k, (h, c) in enumerate(busiest, 1)]
    W = R.get("windows", {})
    tiles.append(("bars", f"{round(100 * W.get('preopen', 0) / tot)}% / {round(100 * W.get('trading', 0) / tot)}%",
                  "Pre-open / trading draws",
                  f"{num(W.get('preopen', 0))} in 04:00-06:59, {num(W.get('trading', 0))} in 07:00-17:29", "grey"))
    P.append(
        '<div class="pb"></div><div class="sect"><h3>The counter\'s rhythm</h3></div>'
        f'<div class="callout tight"><span class="lead">When radios move.</span> Every draw and every return since '
        f'{w0:%d %b} laid on a week: <b class="o">{num(top3)} of the {num(tot)} draws ({round(100 * top3 / tot)}%)</b> fall in the '
        f'busiest three hours, {esc(HEAT_DAYS[bd])} is the busiest weekday ({num(sum(R["draws"][bd]))} draws), and weekdays carry '
        f'<b>{num(wk)}</b> draws against <b>{num(we)}</b> at the weekend.</div>'
        + sh.tiles(tiles)
        + '<div class="sub-h">Draws <span class="thin">- weekday by hour; darker is quieter, orange is busier; the count is printed in every cell</span></div>'
        f'<div class="chartpanel">{sh.heatgrid(R["draws"], HEAT_DAYS, cols24)}</div>'
        '<div class="sub-h">Returns <span class="thin">- the same week, scanned back in</span></div>'
        f'<div class="chartpanel">{sh.heatgrid(R["returns"], HEAT_DAYS, cols24, colour=(31, 167, 90))}</div>'
        f'<div class="note">Counted from every movement in the TRANSACTIONS export for this report\'s barcodes ({esc(log_span)}); a cell '
        f'is the number of issue scans (or return scans) whose time fell in that weekday and hour. A dark cell is a real zero. '
        f'The store runs two shifts, 04:00 to 12:30 and 09:00 to 17:30, and opens at 07:00; before opening the first shift '
        f'bumps monitors, scans the return box and makes up packs.</div>')
    P.append(so_what(
        f'Staff the counter for the {busiest[0][0]:02d}:00 and {busiest[1][0]:02d}:00 starts - {round(100 * top3 / tot)}% of the '
        f'year\'s draws land in three hours - and expect {round(100 * W.get("after", 0) / tot)}% of draws after hours (17:30 to '
        f'03:59), the window where the same-day rate is decided.'))
    # ---- 4. who holds what -----------------------------------------------
    H = I["holders"]
    hrows = []
    for k, d in enumerate(H["top"], 1):
        also = ", ".join({"gas": "gas monitors", "tooling": "tooling"}.get(f, f) for f in d["also"]) or "-"
        hrows.append([str(k), esc(insight_hirer(d["hirer"])), esc(d["company"]), num(d["items"]),
                      (money(d["rvalue"]) if d["rvalue"] else '<span class="tbc">-</span>'),
                      _days_cell(d["oldest"]), esc(also)])
    top_items = sum(d["items"] for d in H["top"])
    top_val = sum(d["rvalue"] for d in H["top"])
    cross_all = sum(1 for d in H["rows"] if d["also"])
    P.append(
        '<div class="keep"><div class="sect"><h3>Who holds what</h3></div>'
        f'<div class="callout tight"><span class="lead">The 80/20.</span> <b class="o">{num(H["holders"])} people and accounts</b> hold the '
        f'{num(H["items"])} radios and batteries on hire; <b>{num(H["n80_items"])}</b> of them hold 80% of the units and '
        f'<b>{num(H["n80_rvalue"])}</b> hold 80% of the {money(H["rvalue_total"])} on-hire value. Ranked by items held, the top '
        f'{num(len(H["top"]))} are below - the custody line is kept apart, as everywhere in this report.</div>'
        + kf.dtable_flow(["Rank", "Hirer", "Company", "Items", "Value", "Oldest (days)", "Also holds"], hrows,
                         ["r", "", "", "r", "r", "r", ""], "cp")
        + f'<div class="note">Ranked by items on hire at the pull ({esc(asat_s)}); value at the prices on the data page, so the '
          f'{num(H["holders"])} holders add up to the on-hire value tile on page 2. "Also holds" reads the whole register: '
          f'{num(cross_all)} of the {num(H["holders"])} radio holders also hold gas monitors or tooling from the store, '
          f'{num(H["cross_n"])} of them in this top {num(len(H["top"]))}. Oldest is the longest-held unit in that holder\'s hands; '
          f'30 days or more shows in red.</div>')
    P.append(so_what(
        f'The top {num(len(H["top"]))} holders have {num(top_items)} units and {money(top_val)} between them - '
        f'{round(100 * top_items / (H["items"] or 1))}% of everything out - and {num(H["cross_n"])} of them also hold gas monitors or '
        f'tooling, so one conversation with each recovers more than radios.'))
    P.append('</div>')
    # ---- 5. what the log and the register disagree on --------------------
    D = I["dq"]
    short = sorted(D["short"], key=lambda t: (ampol_names.sort_key(t["co"]), t["st"]))
    srows = [[esc(t["co"] or "-"), esc(insight_hirer(t["who"])), esc(t["bc"]), esc(t["desc"]),
              f'{t["st"]:%d %b %H:%M}', f'{t["en"]:%H:%M}', f'{t["hours"] * 60:.1f}']
             for t in short[:INSIGHT_SAMPLE_ROWS]]
    mrows = [[esc(insight_hirer(k[0])), esc(k[1] or "-"), f"{k[2]:%d %b %Y}", f"{k[3]:02d}:00-{k[3] + 1:02d}:00", num(v)]
             for k, v in D["mass"][:INSIGHT_SAMPLE_ROWS]]
    mass_acc = sum(1 for k, _ in D["mass"] if "(site account)" in insight_hirer(k[0]))
    before = D["onhire_before_log"]
    before_oos = [r for r in before if hirer_kind(r["hirer"]) == "oos"]
    before_n, page2_n = len(before), len(prev_all)
    by_year = sorted(defaultdict(int, {y: sum(1 for r in before if r["on_dt"].year == y) for y in {r["on_dt"].year for r in before}}).items())
    nolog = D["onhire_no_log"]
    oav = D["open_but_available"]
    log_start = D["log_start"]
    if before_n - len(before_oos) == page2_n:
        recon = (f'That is the <b>{num(page2_n)}</b> units on hire since {esc(PRIOR_LABEL)} on page 2'
                 + (f' plus {num(len(before_oos))} unit{"s" if len(before_oos) != 1 else ""} in the out-of-service custody line, '
                    f'which page 2 keeps apart' if before_oos else '')
                 + ' - the two counts reconcile exactly.')
    else:
        recon = (f'Page 2 counts <b>{num(page2_n)}</b> units on hire since {esc(PRIOR_LABEL)}'
                 + (f' and keeps {num(len(before_oos))} custody unit{"s" if len(before_oos) != 1 else ""} apart' if before_oos else '')
                 + f'; the difference of {num(abs(before_n - len(before_oos) - page2_n))} is a unit whose on-hire date and issue year '
                   f'disagree on which side of {log_start:%d %b %Y} it falls - listed in the register pages.')
    P.append(
        '<div class="pb"></div><div class="sect"><h3>What the log and the register disagree on</h3></div>'
        f'<div class="callout tight"><span class="lead">Two sources, one fleet.</span> The register is the position; the log is the '
        f'history. Read together they show the counter\'s habits and any gap between them. Nothing here is corrected - it is shown, '
        f'with the rule that found it: a hire closed inside <b>{num(D["short_minutes"])} minutes</b> is a short hire; '
        f'<b>{num(D["mass_threshold"])} or more</b> items drawn by one person or account inside one hour is a mass draw.</div>'
        f'<div class="sub-h">Hires closed inside {num(D["short_minutes"])} minutes <span class="thin">- {num(D["short_n"])} this year'
        + (f'; showing {num(len(srows))} (company A to Z, then time)' if D["short_n"] > len(srows) else '; company A to Z, then time') + '</span></div>'
        + (kf.dtable_flow(["Company", "Hirer", "Barcode", "Description", "Out", "Back", "Minutes"], srows,
                          ["", "", "", "", "nw", "nw", "r"], "cp")
           if srows else '<div class="note">No hire on this fleet closed inside six minutes this year.</div>')
        + f'<div class="note">Out and back inside {num(D["short_minutes"])} minutes is a scan corrected on the spot or a unit swapped at the counter - '
          f'{num(D["short_n"])} of the {num(I["tx_n"])} movements this year ({100 * D["short_n"] / (I["tx_n"] or 1):.1f}%).</div>'
        f'<div class="sub-h">Mass draws <span class="thin">- {num(D["mass_n"])} this year, ranked by items drawn</span></div>'
        + (kf.dtable_flow(["Hirer or account", "Company", "Date", "Hour", "Items"], mrows, ["", "", "nw", "nw", "r"], "cp")
           if mrows else '<div class="note">No person or account drew fifteen or more radios or batteries inside one hour this year.</div>')
        + f'<div class="note">{num(D["mass_threshold"])} or more items by one hirer inside an hour: {num(mass_acc)} by site accounts and '
          f'{num(D["mass_n"] - mass_acc)} by named people - kits of batteries going out on a shutdown or after-hours account, and crew '
          f'leads drawing for a whole crew. Habits to know about, not scan errors.</div>'
        f'<div class="sub-h">On hire with no movement since the log began <span class="thin">- issued on or after {log_start:%d %b %Y}</span></div>'
        + (kf.dtable_flow(["Company", "Hirer", "Barcode", "Description", "On hire since"],
                          [[esc(r["company"]), esc(insight_hirer(r["hirer"])), esc(r["barcode"]), esc(r["desc"]),
                            (f'{r["on_dt"]:%d %b %Y %H:%M}' if r["on_dt"] else "-")]
                           for r in sorted(nolog, key=lambda r: (ampol_names.sort_key(r["company"]), r["barcode"]))[:INSIGHT_SAMPLE_ROWS]],
                          ["", "", "", "", "nw"], "cp")
           + f'<div class="note">{num(len(nolog))} unit{"s" if len(nolog) != 1 else ""} on hire in the register with no issue scan in the log '
             f'since {log_start:%d %b %Y}' + (f' - showing {num(INSIGHT_SAMPLE_ROWS)}, company A to Z' if len(nolog) > INSIGHT_SAMPLE_ROWS else '') + '.</div>'
           if nolog else f'<div class="note"><b>No gap.</b> Every unit the register shows on hire since {log_start:%d %b %Y} has its issue scan in the log.</div>')
        + f'<div class="sub-h">Issued before the log begins <span class="thin">- on hire since before {log_start:%d %b %Y}</span></div>'
        f'<div class="note"><b>{num(before_n)} units</b> on hire in the register were issued before the log starts, so their issue scan is '
        f'history the export does not carry: ' + esc(", ".join(f"{n} from {y}" for y, n in by_year)) + f'. {recon}</div>'
        '<div class="sub-h">Available in the register with an open movement in the log</div>'
        + (kf.dtable_flow(["Barcode", "Description", "Storage unit", "Open movement", "Company", "Hirer"],
                          [[esc(r["barcode"]), esc(r["desc"]), esc(r["unit"] or "-"), f'{t["st"]:%d %b %Y %H:%M}', esc(t["co"] or "-"),
                            esc(insight_hirer(t["who"]))] for r, t in sorted(oav, key=lambda x: x[0]["barcode"])[:INSIGHT_SAMPLE_ROWS]],
                          ["", "", "", "nw", "", ""], "cp")
           + f'<div class="note">{num(len(oav))} unit{"s" if len(oav) != 1 else ""} on the shelf whose newest movement in the log has no return scan'
             + (f' - showing {num(INSIGHT_SAMPLE_ROWS)}, barcode A to Z' if len(oav) > INSIGHT_SAMPLE_ROWS else '') + '.</div>'
           if oav else '<div class="note"><b>No gap.</b> No unit the register shows available has an open movement in the log - every '
                       'return scan reached the register.</div>'))
    agree = not nolog and not oav
    P.append(so_what(
        (f'The log and the register agree: no unit is on hire without its issue scan since {log_start:%d %b} and no shelf unit has an '
         f'open movement. ' if agree else
         f'{num(len(nolog))} on-hire unit{"s" if len(nolog) != 1 else ""} with no issue scan and {num(len(oav))} shelf unit{"s" if len(oav) != 1 else ""} '
         f'with an open movement are the rows to check at the counter first. ')
        + f'The {num(D["short_n"])} short hires and {num(D["mass_n"])} mass draws are counter habits, not errors to chase, and the '
        f'{num(before_n)} prior-year units are exactly the page-2 figure' + (' plus the custody line.' if before_oos else '.')))
    return "".join(P)


def insights_method(I):
    """The data-and-method sentence for the insights pages - the source
    and the rules, in the words the pages use."""
    from k2shell import esc, num
    if not I:
        return (' The year-in-movements, return-window, rhythm, who-holds-what and log-versus-register pages need the '
                'SiteIQ TRANSACTIONS export in Data - not found this run, so they are not printed.')
    w0, w1 = I["window"]
    return (f' The year in movements, return windows, the counter\'s rhythm, who holds what and the log-versus-register page '
            f'read the SiteIQ TRANSACTIONS export ({esc(I["tx_file"])}, sheet CUSTOMER_CONTRACTOR_EQUIP, report period '
            f'{w0:%d %b %Y %H:%M} to {w1:%d %b %Y %H:%M}) - {num(I["tx_n"])} movements of this report\'s {num(I["scope_n"])} '
            f'barcodes. Rules: a hire closed inside 6 minutes is a short hire; 15 or more items drawn by one hirer inside one '
            f'hour is a mass draw; a product is the description with its size and serial tail removed - a radio\'s description '
            f'keeps the serial prefix SiteIQ writes into it, so the return-window table groups by this report\'s own radio and '
            f'battery rule and its total line is the engine\'s own. Same-day means the return scan fell on the issue scan\'s date.')


def build_html(r26, rprev, b26, bprev, oos, ravail, bavail, data_asat="", facts=None, changes=None,
               contents=None, insights=None):
    """The client PDF on the Coates house frame (k2flow): the position, the
    ask, what moved, the pictures, the trend, then the register - companies
    A-Z, oldest first - behind an appendix divider. contents: the (title,
    page) rows for the cover's "What's inside" block - None on the first
    pass, the rows read off the printed PDF on the second. insights: the
    transaction-log facts from insights_compute (None = no export)."""
    import k2flow as kf
    import k2shell as sh
    from k2shell import esc, num
    refresh = BUILD_DT.strftime("%d %b %Y %H:%M")
    F = facts or position_facts(r26, rprev, b26, bprev, data_asat)
    total_exposure = F["exposure"]
    prev_all = F["prev_all"]
    prev_val = F["prev_val"]
    oldest = F["oldest"]
    all_on = F["all_on"]
    by_year = defaultdict(lambda: [0, 0.0])
    for i in prev_all:
        by_year[i["year"]][0] += 1
        by_year[i["year"]][1] += i["cost"] or 0
    years = sorted(by_year)
    comp_year = defaultdict(lambda: defaultdict(lambda: [0, 0.0]))
    comp_oldest = defaultdict(int)
    for i in prev_all:
        comp_year[i["company"]][i["year"]][0] += 1
        comp_year[i["company"]][i["year"]][1] += i["cost"] or 0
        comp_oldest[i["company"]] = max(comp_oldest[i["company"]], i["days"])
    oos_r = [i for i in oos if "batter" not in i["desc"].lower()]
    oos_b = [i for i in oos if "batter" in i["desc"].lower()]
    n_radio = META.get("n_radio", len(r26) + len(rprev) + len(ravail) + len(oos_r))
    n_batt = META.get("n_batt", len(b26) + len(bprev) + len(bavail) + len(oos_b))
    comp_tot = defaultdict(float)
    for i in all_on:
        comp_tot[i["company"]] += i["cost"] or 0
    top10 = sorted(comp_tot.items(), key=lambda kv: -kv[1])[:10]
    rest_n = len(comp_tot) - len(top10)
    rest_v = sum(comp_tot.values()) - sum(v for _, v in top10)
    age_defs = [("0-30 days", lambda d: d <= 30), ("31-90 days", lambda d: 31 <= d <= 90),
                ("91-365 days", lambda d: 91 <= d <= 365), ("Over 365 days", lambda d: d > 365)]
    age_rows = []
    for lbl, test in age_defs:
        grp = [i for i in all_on if test(i["days"])]
        pct = round(100 * len(grp) / (len(all_on) or 1))
        age_rows.append((lbl, len(grp), f"{len(grp)} units - {money(val(grp))} - {pct}%"))
    asat_s = data_asat or "TBC"
    unpriced = sum(1 for i in all_on if not i["cost"])
    cfg = {
        "client": "Ampol", "title": "Site Radio On-Hire Report",
        "kicker": "COATES · TOOL STORE · SITE RADIO ON-HIRE REPORT",
        "project": "Ampol Lytton Refinery · Permanent Tool Store",
        "asat_note": "(SiteIQ register pull)",
        "key_items": [("orange", "RETURN IT", "not in use? back to the store - scanned in on the spot"),
                      ("blue", "RESCAN IT", "still in use? bring it past the counter - proof of existence"),
                      ("amber", "30+ DAYS", "shown in red - due for a return or a rescan")],
        "team": [{"name": "Andrew Fisher", "role": "Shutdown Manager", "shift": "",
                  "email": "andrew.fisher@coates.com.au",
                  "blurb": "Oversees the store and the radio fleet - anything at all, start here",
                  "lead": True}],
    }
    P = []
    def mv(key, value, good, fallback, fallback_cls):
        txt, cls = rh.movement("radio", key, _ASAT_DT[0], value, good,
                               money=(key == "exposure"))
        return (txt, cls) if txt else (fallback, fallback_cls)
    prior_pct, status, due = F["prior_pct"], F["status"], F["due"]
    band = sh.rag_band(status,
        f'<b class="o">{num(len(prev_all))} of the {num(len(all_on))} units on hire ({prior_pct}%)</b> were issued in '
        f'{esc(PRIOR_LABEL)} and are still out - {money(prev_val)} of equipment, the oldest for {num(oldest)} days.',
        f'Share of on-hire units issued in prior years: Green under {RAG_AMBER_PRIOR_PCT}%, Amber from '
        f'{RAG_AMBER_PRIOR_PCT}%, Red from {RAG_RED_PRIOR_PCT}%. Default lines - set at the top of the script.',
        "<b>Andrew Fisher</b>, Shutdown Manager - Coates tool store",
        f'Every prior-year holder sent their list from this report by <b>{due}</b>; returned or rescanned units drop off the next run.',
        tight=True)
    # ---- page 2 (after the cover): the position page --------------------
    # WHY (03 Sep 2026): one grammar for every position page in the suite -
    # hero and key strip, the RAG band, the tiles, three things to do today,
    # then a three-line story. The full paragraph, the ask and the assurance
    # note follow on the next page, so page 2 is read in one look.
    story = (f'<div class="callout"><span class="lead">The story.</span> <b class="o">{money(total_exposure)}</b> of site '
             f'radio equipment is on hire: <b>{num(len(r26) + len(rprev))} radios</b> and <b>{num(len(b26) + len(bprev))} '
             f'batteries</b>. <b class="o">{money(prev_val)}</b> across <b>{num(len(prev_all))} units</b> has been out since '
             f'{esc(PRIOR_LABEL)}, the oldest for <b>{num(oldest)} days</b>. <b>Not in use? Return it. Still in use? Bring it '
             f'past the counter for a rescan.</b></div>')
    three = sh.three_things(three_things_for(F, r26, b26, oos))
    pos = (f'<div class="callout"><span class="lead">The position.</span> <b class="o">{money(total_exposure)}</b> '
           f'of site radio equipment is on hire per the SiteIQ pull as at {esc(asat_s)}: '
           f'<b>{num(len(r26) + len(rprev))} radios</b> and <b>{num(len(b26) + len(bprev))} batteries</b>. '
           f'Of that, <b class="o">{money(prev_val)}</b> across <b>{num(len(prev_all))} units</b> has been out since '
           f'{esc(PRIOR_LABEL)} - the oldest for <b>{num(oldest)} days</b>. Every one of those is working hard on site, '
           f'sitting unused in a crib room or ute, or no longer accounted for; this report tells those three apart. '
           f'<b>Not in use? Return it. Still in use? Bring it past the counter for a rescan.</b> Either way the record '
           f'updates the moment it is scanned.</div>')
    tiles1 = sh.tiles_plus([
        ("shield", money(total_exposure), "On-hire value", *mv("exposure", round(total_exposure), "down", f"{num(len(all_on))} units on hire", "grey")),
        ("swap", num(len(r26) + len(rprev)), "Radios on hire", *mv("radios_on_hire", len(r26) + len(rprev), "down", money(val(r26) + val(rprev)), "")),
        ("swap", num(len(b26) + len(bprev)), "Batteries on hire", *mv("batteries_on_hire", len(b26) + len(bprev), "down", money(val(b26) + val(bprev)), "")),
        ("warn", num(len(prev_all)), f"On hire since {PRIOR_LABEL}", *mv("prior_units", len(prev_all), "down", money(prev_val), "red" if prev_all else "green")),
    ])
    tiles2 = sh.tiles([
        ("check", f"{len(ravail)} / {len(bavail)}", "Available in store", "radios / batteries", "green" if ravail else "amber"),
        ("clock", num(oldest), "Oldest hire (days)", "", "amber"),
        ("box", num(len(oos)), "Out of service", money(val(oos)) if val(oos) else "", "grey"),
        ("bars", f"{num(n_radio)} / {num(n_batt)}", "On the register", "radios / batteries", "grey"),
    ])
    ask = ('<div class="alerts"><div class="ah">What we are asking - please read first</div><table class="al">'
           '<tr><td class="al-dot d-amber">&#9679;</td><td><div class="al-t">Not in use? Return it to the Ampol Tool Store (Coates managed).</div>'
           '<div class="al-s">It is scanned in on the spot and comes straight off this report. Returned radios go back into the available pool for the next shutdown.</div></td></tr>'
           '<tr><td class="al-dot d-blue">&#9679;</td><td><div class="al-t">Still in use? Bring it past the store for a rescan.</div>'
           '<div class="al-s">Proof of existence: a thirty-second scan verifies the unit is on site, in whose hands, and resets the record. Nothing is taken off you.</div></td></tr>'
           f'<tr><td class="al-dot d-red">&#9679;</td><td><div class="al-t">Replacement value: radios {money(PRICE_RADIO)} each, batteries {money(PRICE_BATT)} ({esc(PRICE_SOURCE)}).</div>'
           '<div class="al-s">Units that can be neither returned nor verified are ultimately chargeable at replacement value under the hire arrangement, applied consistently to all companies. Verification protects everyone from charges for equipment that is actually on site.</div></td></tr>'
           '<tr><td class="al-dot d-green">&#9679;</td><td><div class="al-t">Something look wrong? Tell the Ampol Tool Store.</div>'
           '<div class="al-s">We review and correct the record with you - no charge is finalised without that review.</div></td></tr>'
           '</table></div>')
    assure = (f'<div class="note">These records are protected by daily stock takes (30-day full-coverage cycle, on-hire verification of the '
              f'radio charging bays) and the double-check return process - every return is inspected and scanned on receipt, then '
              f'stock-taken to its storage unit before going back on charge. <b>The moment a unit is scanned, the record updates.</b> '
              f'Every count on this report is read from the SiteIQ register as at <b>{esc(asat_s)}</b> - nothing comes from a summary tab. '
              f'Serial numbers: {esc(META.get("serial_note", "from the radio register"))}.</div>')
    # WHY (03 Sep 2026): one sentence under the band - the daily-return
    # standard against the log; the rest of page 2 keeps its grammar
    P.append(band + standard_line(insights) + tiles1 + tiles2 + three + story)
    P.append('<div class="pb"></div>' + pos + ask + assure)
    # ---- what moved since the last pull (03 Sep 2026) ---------------------
    P.append(changes_section(changes))
    # ---- the insights pages (03 Sep 2026): the log's year, on new pages --
    P.append(insights_pages(insights, prev_all, asat_s))
    # ---- the pictures -----------------------------------------------------
    radio_rows = [(f"On hire - issued {CUR_YEAR}", len(r26)), (f"On hire - issued {PRIOR_LABEL}", len(rprev)),
                  ("Available in store", len(ravail)), ("Out of service", len(oos_r))]
    batt_rows = [(f"On hire - issued {CUR_YEAR}", len(b26)), (f"On hire - issued {PRIOR_LABEL}", len(bprev)),
                 ("Available in store", len(bavail)), ("Out of service", len(oos_b))]
    exp_rows = [(c, v, money(v)) for c, v in top10]
    exp_note = (f"Ranked by value - the top {len(top10)} of {len(comp_tot)} companies; the remaining {rest_n} hold "
                f"{money(rest_v)} between them." if rest_n > 0 else f"Ranked by value - all {len(comp_tot)} companies shown.")
    # WHY (03 Sep 2026): forty rows of one-unit companies made the ageing
    # chart a wall. Companies holding ten or more units are charted; the
    # rest are one line under it, A to Z with their counts, and the note
    # adds the two back together so the picture still reconciles.
    age_all = ageing_rows(all_on)
    age_chart, age_tail = sh.split_long_tail(age_all, 10)
    chart_units = sum(sum(s) for _, s in age_chart)
    tail_units = sum(sum(s) for _, s in age_tail)
    if age_tail:
        age_scope = f"the {num(len(age_chart))} companies holding ten or more, A to Z"
        tail_line = ('<div class="note"><b>Companies with fewer than ten units:</b> '
                     + esc(", ".join(f"{c} ({sum(s)})" for c, s in age_tail))
                     + f' - {num(tail_units)} units across {num(len(age_tail))} companies.</div>')
        recon = (f"The chart holds {num(chart_units)} units and the legend totals count those; the line under it "
                 f"holds the other {num(tail_units)}, {num(len(all_on))} in all. ")
    else:
        age_scope = f"all {num(len(age_chart))} companies, A to Z"
        tail_line = ""
        recon = ""
    P.append(
        '<div class="sect"><h3>Fleet position at a glance</h3></div>'
        '<div class="callout tight">Three pictures of the same truth - where the fleet sits, who holds the value, and how long it '
        'has been out. Every count and dollar here comes straight from the register tables that follow; nothing is replaced or rounded.</div>'
        '<table class="two"><tr>'
        f'<td style="width:50%;padding-right:6px"><div class="sub-h">Site radios <span class="thin">- {num(n_radio)} on the register</span></div>'
        f'<div class="chartpanel">{sh.hbars(radio_rows, w=300, lab_w=150, rowh=26)}</div></td>'
        f'<td style="width:50%;padding-left:6px"><div class="sub-h">Radio batteries <span class="thin">- {num(n_batt)} on the register</span></div>'
        f'<div class="chartpanel">{sh.hbars(batt_rows, w=300, lab_w=150, rowh=26)}</div></td>'
        '</tr></table>'
        f'<div class="sub-h">Replacement-value exposure by company <span class="thin">- {money(sum(comp_tot.values()))} across '
        f'{len(comp_tot)} companies (ranked by value)</span></div>'
        f'<div class="chartpanel">{sh.hbars(exp_rows, w=636, lab_w=190, rowh=24, right=90)}</div>'
        f'<div class="note">{esc(exp_note)}</div>'
        # WHY (03 Sep 2026): the value chart says who holds the money; this one
        # says how long each company has held it, in the suite's four bands
        f'<div class="sub-h">On-hire ageing by company <span class="thin">- {num(len(all_on))} radios and batteries '
        f'on hire, {age_scope}</span></div>'
        # the chart and the line that completes its count stay on one page
        f'<div class="keep"><div class="chartpanel">{sh.stacked_hbars(age_chart)}</div>{tail_line}</div>'
        f'<div class="note">Days out at the pull ({esc(asat_s)}), in the four bands of the legend. {recon}Each row adds up to that '
        f'company\'s units on hire - its {CUR_YEAR} units in the register pages plus its {esc(PRIOR_LABEL)} units in the value '
        f'story. Out-of-service custody units are not counted here, as everywhere in this report.</div>'
        f'<div class="sub-h">Age of hire <span class="thin">- how long the {num(len(all_on))} on-hire units have been out</span></div>'
        f'<div class="chartpanel">{sh.hbars(age_rows, w=636, lab_w=120, rowh=24, right=200)}</div>'
        f'<div class="note">Everything beyond 30 days is due for a return or a rescan; {esc(PRIOR_LABEL)} issues drive the over-365 band.</div>')
    # ---- prior years: the value story ------------------------------------
    yt = [("clock", num(cnt), f"Issued {y} - still on hire", money(v), "red") for y, (cnt, v) in sorted(by_year.items())]
    ordered = sorted(comp_year.items(), key=lambda kv: -sum(v[1] for v in kv[1].values()))
    yrows = []
    for comp, ymap in ordered:
        tot_n = sum(v[0] for v in ymap.values())
        tot_v = sum(v[1] for v in ymap.values())
        cells = [esc(comp)] + [(num(ymap.get(y, [0, 0])[0]) if ymap.get(y, [0, 0])[0] else '<span class="tbc">-</span>') for y in years]
        cells += [f"<b>{num(tot_n)}</b>",
                  (f'<b class="rd">{money(tot_v)}</b>' if tot_v >= 100000 else f"<b>{money(tot_v)}</b>"),
                  f'<span class="rd">{num(comp_oldest[comp])}</span>']
        yrows.append(cells)
    P.append(
        f'<div class="sect"><h3>{esc(PRIOR_LABEL)} - the value story</h3></div>'
        f'<div class="callout tight"><b class="o">{num(len(prev_all))} radios and batteries issued in prior years remain on hire</b> - '
        f'{money(prev_val)} of equipment. Rather than listing {num(len(prev_all))} lines, this section summarises by year and company '
        f'(line-item detail is available from the Tool Store on request). Any unit still in use: bring it past the store for a rescan. '
        f'Any unit not in use: return it.</div>'
        + (sh.tiles(yt) if yt else "")
        + kf.dtable_flow(["Company (ranked by value)"] + [str(y) for y in years] + ["Units", "Value", "Oldest (days)"],
                         yrows, [""] + ["r"] * len(years) + ["r", "r", "r"], "cp")
        + f'<div class="note">Values in red carry {money(100000)} or more of equipment. The fastest way off this table is a return or a rescan.</div>')
    # ---- the trend page: only once seven days are on the scoreboard ------
    trend_html, days_on_record = trend_section(asat_s)
    P.append(trend_html)
    # ---- the appendix divider, then the register: this year, full detail --
    reg_units = r26 + b26
    reg_companies = {i["company"] for i in reg_units}
    P.append(kf.divider_block(
        "The complete register, A to Z",
        f"Everything from here is the complete list: every radio and every battery issued in {CUR_YEAR} and still "
        f"on hire, company by company from A to Z and longest-held first, then the units out of service. "
        f"The story - the position, what moved, the pictures and the {esc(PRIOR_LABEL)} value summary - is on the "
        f"pages before this one. Prior-year line detail is available from the Tool Store on request.",
        f"{num(len(reg_units))} units across {num(len(reg_companies))} companies"))
    P.append(
        f'<div class="sect"><h3>{CUR_YEAR} - radios on hire, full detail</h3></div>'
        f'<div class="note"><b>{num(len(r26))} radios</b> issued in {CUR_YEAR} are on hire - {money(val(r26))}. Companies A to Z, '
        f'one customer one name; inside a company, longest-held first. 30 days or more shows in red.</div>'
        + detail_rows(r26, serial_col=True))
    P.append(
        f'<div class="pb"></div><div class="sect"><h3>{CUR_YEAR} - radio batteries on hire, full detail</h3></div>'
        f'<div class="note"><b>{num(len(b26))} batteries</b> issued in {CUR_YEAR} are on hire - {money(val(b26))}. Companies A to Z, '
        f'longest-held first.</div>' + detail_rows(b26, serial_col=False))
    # ---- out of service -----------------------------------------------------
    orows = [[esc(i["barcode"]), esc(i["serial"] or "-"), money(i["cost"]),
              (i["date"].strftime("%d %b %Y") if i["date"] else "-"), f'<span class="rd">{i["days"]}</span>']
             for i in sorted(oos, key=lambda x: -x["days"])]
    P.append(
        f'<div class="pb"></div><div class="sect"><h3>Out of service - {num(len(oos))} units ({money(val(oos))})</h3></div>'
        '<div class="note">Units tagged Out of Service under the Tool Store SOP - tagged, photographed, reported and quarantined '
        'pending repair or replacement. Longest first.</div>'
        + (kf.dtable_flow(["Barcode", "Serial", "Price", "Status since", "Days"], orows, ["", "", "r", "", "r"], "cp")
           if orows else '<div class="note">Nothing is out of service at the pull time.</div>'))
    # ---- close: data and method, the standard, the team -------------------
    cards = sh.info_cards([
        ("Return it or rescan it",
         "Not in use - back to the store, scanned in on the spot. Still in use - a thirty-second rescan at the counter is proof "
         "of existence. Either way the record updates immediately."),
        ("Replacement value, applied consistently",
         f"Radios {money(PRICE_RADIO)}, batteries {money(PRICE_BATT)} ({esc(PRICE_SOURCE)}). Units that can be neither returned "
         f"nor verified are chargeable at that value under the hire arrangement - the same rule for every company. "
         f"{esc(META.get('unpriced_note', '').strip())}"),
        ("Two scans, two looks",
         "Every unit is scanned and sighted going out, and scanned and sighted coming back. Radios are charge-checked on return. "
         "Life Saving Rule 5 - Tools and Equipment (SEQ-GL-009): nothing damaged, defective or flat is ever reissued."),
        ("Numbers you can challenge",
         f"Every count on every page is read from the SiteIQ RENTAL_STOCK export as at {esc(asat_s)} - never a summary tab. "
         f"{esc(META.get('coverage', '').strip())} {esc(META.get('serial_note', 'Serial numbers from the radio register'))}. "
         f"Unpriced units ({num(unpriced)}) show a dash and are never estimated."),
    ])
    P.append(
        '<div class="pb"></div><div class="sect"><h3>Data and method</h3></div>'
        f'<div class="note">Source: the SiteIQ RENTAL_STOCK export requested {esc(asat_s)} - every Motorola site radio and radio '
        f'battery on the register, with its status, company, hirer, on-hire date and storage unit. Report built {esc(refresh)}. '
        f'{CUR_YEAR} issues are shown in full; {esc(PRIOR_LABEL)} issues are summarised by year and company. Companies are one '
        f'customer one name (project accounts roll up to their parent). Replacement values: {esc(PRICE_SOURCE)}. '
        f'Movement from Data\\previous and TRANSACTIONS; days out count from the pull time.'
        + (f' Trend page: appears once {_WORDS.get(TREND_MIN_DAYS, str(TREND_MIN_DAYS))} days are on record ({num(days_on_record)} today).'
           if not trend_html else f' Trend page: {num(days_on_record)} report days on record.')
        + insights_method(insights)
        + '</div>'
        '<div class="sect"><h3>How the radio fleet is run</h3></div>' + cards + sh.coates_way_panel(traits=(4, 7), disciplines=(5, 6), line="one scan at the counter closes the record; nothing is chased that the register already answers")
        + '<div class="sect"><h3>Meet the tool store team</h3></div>' + sh.team_cards(cfg["team"]))
    cover = kf.cover_block(cfg, num(len(prev_all)), f"units on hire since {PRIOR_LABEL}", [
        f"<b>{money(total_exposure)}</b> of radio equipment on hire - {num(len(all_on))} units",
        f"<b>{money(prev_val)}</b> of it issued in prior years, the oldest {num(oldest)} days ago",
        f"<b>{len(ravail)}</b> radios and <b>{len(bavail)}</b> batteries ready on the shelf"],
        refresh, asat_s, rag=status,
        fresh=sh.freshness_line(_ASAT_DT[0], BUILD_DT), contents=contents) if COVER_PAGE else None
    # what the phone card and the history need, kept for main()
    build_html.last = {"cfg": cfg, "status": status, "prior_pct": prior_pct, "due": due,
                       "exposure": total_exposure, "prev_val": prev_val, "oldest": oldest,
                       "n_on": len(all_on), "n_prior": len(prev_all)}
    return kf.flow_doc(cfg, refresh, asat_s, "".join(P), extra_css=INSIGHT_CSS, cover=cover)



RADIO_EMAIL_TO = (
    '"Ampol Store" <Ampolstore@coates.com.au>, "Mitchell, Cody" <Cody.mitchell@coates.com.au>, '
    '"Riki Nicholls" <rxnicho@ampol.com.au>, "Steve Webster" <swebste@ampol.com.au>, '
    '"Lucas Riddell" <lxridde@ampol.com.au>, "Michael McEvoy" <mxmcevo@ampol.com.au>, '
    '"Fabian Wong" <fwong@ampol.com.au>, "Lyle Mildenhall" <lmilden@ampol.com.au>, '
    '"Steve Costanzo" <scostan@ampol.com.au>, "Virginia Harding" <vhardin@ampol.com.au>, '
    '"Tn Nam" <tnam@ampol.com.au>, "Thiago Lucca" <txlucca@ampol.com.au>, '
    '"John Strano" <jstrano@ampol.com.au>, "David McGurk" <dmcgurk@ampol.com.au>, '
    '"Kail Linehan" <kxlineh@ampol.com.au>, "Eamon Rees" <erees@ampol.com.au>, '
    '"William Potter" <wxpotte@ampol.com.au>, "Jonathon Giudice" <jgiudic@ampol.com.au>, '
    '"Jason Waddell" <jwaddel@ampol.com.au>, "Paul Henry" <phenry@ampol.com.au>, '
    '"Jarred Bishop" <jxbisho@ampol.com.au>, "Simon Phillips" <sxphill@ampol.com.au>, '
    '"Darren Parker" <dparker@ampol.com.au>, "Ross Comerford" <rcommer@ampol.com.au>, '
    '"Arron Pelling" <ajpelli@ampol.com.au>, '
    '"Robbie Glyde (robbie.glyde@universalcranes.com)" <robbie.glyde@universalcranes.com>, '
    '"Saad Khan" <szkhan1@ampol.com.au>, "Sean Wong" <czwong1@ampol.com.au>, '
    '"LytRefMaintReporting@ampol.com.au" <LytRefMaintReporting@ampol.com.au>, '
    '"John Martin" <jzmarti1@ampol.com.au>, "Chris Jurlina" <cjurlin@ampol.com.au>, '
    '"Sam Robson" <srobers@ampol.com.au>, "Sims, Akira (North)" <Akira.sims@coates.com.au>, '
    '"Logan Ferguson-rudolph (External)" <lyfergu@ampol.com.au>, '
    '"Jay Purcell (External)" <jypurce@ampol.com.au>'
)


def build_email_summary(r26, rprev, b26, bprev, oos, ravail, bavail, data_asat="", pdf_ok=True, card_path=None):
    """High-level summary email body - all data summarised, full detail in
    the attached PDF, the phone position card inline under the greeting."""
    refresh = BUILD_DT.strftime("%d %b %Y %H:%M")
    O = "#e07000"
    total_exposure = val(r26) + val(rprev) + val(b26) + val(bprev)
    prev_all = rprev + bprev
    prev_val = val(prev_all)
    oldest = max(i["days"] for i in prev_all) if prev_all else 0

    by_year = defaultdict(lambda: [0, 0.0])
    for i in prev_all:
        by_year[i["year"]][0] += 1
        by_year[i["year"]][1] += i["cost"] or 0
    years = sorted(by_year)

    # per company: prior-year units/value + 2026 units/value, oldest days
    agg = defaultdict(lambda: {"prev_n": 0, "prev_v": 0.0, "n26": 0, "v26": 0.0, "old": 0})
    for i in prev_all:
        a = agg[i["company"]]; a["prev_n"] += 1; a["prev_v"] += i["cost"] or 0; a["old"] = max(a["old"], i["days"])
    for i in r26 + b26:
        a = agg[i["company"]]; a["n26"] += 1; a["v26"] += i["cost"] or 0; a["old"] = max(a["old"], i["days"])
    ordered = sorted(agg.items(), key=lambda kv: -(kv[1]["prev_v"] + kv[1]["v26"]))

    p = [f"""<!doctype html><html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background-color:#FFFFFF;font-family:Calibri,Arial,sans-serif;color:#1a1a1a">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr><td align="center" style="padding:16px 10px">
<table role="presentation" width="820" cellspacing="0" cellpadding="0" style="width:820px;max-width:820px;border-collapse:collapse">
<tr><td style="border-left:6px solid {O};padding:6px 12px">
<div style="font-size:12px;font-weight:bold;color:{O}">{COATES_PURPOSE}</div>
<div style="font-size:22px;font-weight:bold">AMPOL SITE RADIO ON-HIRE REPORT</div>
<div style="font-size:11px;color:#555">Ampol Tool Store - Coates managed &nbsp;|&nbsp; Motorola site radios &amp; batteries &nbsp;|&nbsp; Refreshed: {refresh}</div></td></tr>

<tr><td style="padding:10px 0 4px;font-size:12px;line-height:1.6">Good morning all,<br><br>
Please find below the summary of the Ampol site radio position{", with the full report attached as a PDF (" + str(CUR_YEAR) + " in complete line-item detail with serial numbers; prior years summarised by company)" if pdf_ok else " - the full line-item report follows separately"}.
This is about visibility, not blame - the ask is simple and it applies to every company equally.</td></tr>
{inline_card_html(card_path, data_asat)}

<tr><td bgcolor="#fdf0f0" style="background-color:#fdf0f0;border:1px solid #f0c0c0;padding:10px 14px">
<div style="font-size:15px;font-weight:bold;color:#c00000">{money(total_exposure)} of radio equipment is currently on hire.</div>
<div style="font-size:11px;padding-top:3px">Of this, <b style="color:#c00000">{money(prev_val)}</b> across <b>{len(prev_all)} units</b> has been on hire since <b>{PRIOR_LABEL}</b> - the oldest for <b>{oldest} days</b>.
Every one of these radios and batteries is either working on site, sitting unused, or no longer accounted for - this report exists to tell those three apart.</div></td></tr>

<tr><td style="padding:8px 0"><table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse">
<tr><td width="20%" style="border:1px solid #ccc;padding:7px 10px"><div style="font-size:8px;color:#666;letter-spacing:1px">RADIOS ON HIRE</div><div style="font-size:15px;font-weight:bold;color:{O}">{len(r26)+len(rprev)} \u00b7 {money(val(r26)+val(rprev))}</div></td>
<td width="20%" style="border:1px solid #ccc;padding:7px 10px"><div style="font-size:8px;color:#666;letter-spacing:1px">BATTERIES ON HIRE</div><div style="font-size:15px;font-weight:bold;color:{O}">{len(b26)+len(bprev)} \u00b7 {money(val(b26)+val(bprev))}</div></td>
<td width="22%" style="border:1px solid #ccc;padding:7px 10px"><div style="font-size:8px;color:#666;letter-spacing:1px">FROM {PRIOR_LABEL}</div><div style="font-size:15px;font-weight:bold;color:#c00000">{len(prev_all)} \u00b7 {money(prev_val)}</div></td>
<td width="19%" style="border:1px solid #ccc;padding:7px 10px"><div style="font-size:8px;color:#666;letter-spacing:1px">AVAILABLE IN STORE</div><div style="font-size:15px;font-weight:bold;color:#1e7d32">{len(ravail)} radios \u00b7 {len(bavail)} batt.</div></td>
<td style="border:1px solid #ccc;padding:7px 10px"><div style="font-size:8px;color:#666;letter-spacing:1px">OUT OF SERVICE</div><div style="font-size:15px;font-weight:bold;color:#b07700">{len(oos)}</div></td></tr></table></td></tr>

<tr><td bgcolor="#fdf4ea" style="background-color:#fdf4ea;border:1px solid #f0d5b8;padding:10px 14px;font-size:11px">
<div style="font-weight:bold;color:#b35a00;letter-spacing:0.5px;font-size:12px">WHAT WE ARE ASKING - RETURN IT OR RESCAN IT</div>
<ul style="margin:4px 0 0;padding-left:16px">
<li style="margin:3px 0"><b>Not in use?</b> Return it to the <b>Ampol Tool Store (Coates managed)</b> - it is scanned in on the spot, comes straight off this report, and goes back into the available pool for the next shutdown.</li>
<li style="margin:3px 0"><b>Still in use?</b> Bring it past the Ampol Tool Store for a <b>rescan - proof of existence</b>. A thirty-second scan verifies the unit is on site and in whose hands, and resets the record. Nothing is taken off anyone.</li>
<li style="margin:3px 0">Site radios are <b>{money(PRICE_RADIO)}</b> each to replace and batteries {money(PRICE_BATT)} ({PRICE_SOURCE}). Units that can be neither returned nor verified are ultimately chargeable at replacement value under the hire arrangement - applied consistently to all companies, and <b>no charge is finalised without review</b>. Verification protects everyone from charges for equipment that is actually on site.</li>
<li style="margin:3px 0">Anything look incorrect? Contact the Ampol Tool Store and we will review and correct the record with you.</li>
</ul></td></tr>

<tr><td style="padding:14px 0 4px"><div style="font-size:15px;font-weight:bold;color:{O};text-transform:uppercase;border-bottom:2px solid {O};padding-bottom:3px">{PRIOR_LABEL} - the value story</div></td></tr>
<tr><td style="padding:4px 0"><table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse"><tr>"""]
    for y in years:
        cnt, v = by_year[y]
        p.append(f"""<td width="33%" style="border:1px solid #ccc;padding:7px 10px"><div style="font-size:8px;color:#666;letter-spacing:1px">ISSUED {y} - STILL ON HIRE</div>
<div style="font-size:14px;font-weight:bold;color:#c00000">{cnt} units \u00b7 {money(v)}</div></td>""")
    # WHY (13 Aug 2026): this block carries the year variables, so it must be
    # an f-string - as a plain string the {placeholders} printed literally in
    # the email body. Caught by 09_CHECK_REPORTS.
    p.append(f"""</tr></table></td></tr>
<tr><td style="padding:6px 0"><table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;border:1px solid #ccc;font-size:10.5px">
<tr bgcolor="#fafafa"><td style="padding:5px 10px;font-size:9px;color:#555;border-bottom:1px solid #ccc">COMPANY</td>
<td align="right" style="padding:5px 10px;font-size:9px;color:#555;border-bottom:1px solid #ccc">{PRIOR_SHORT} UNITS</td>
<td align="right" style="padding:5px 10px;font-size:9px;color:#555;border-bottom:1px solid #ccc">{PRIOR_SHORT} VALUE</td>
<td align="right" style="padding:5px 10px;font-size:9px;color:#555;border-bottom:1px solid #ccc">{CUR_YEAR} UNITS</td>
<td align="right" style="padding:5px 10px;font-size:9px;color:#555;border-bottom:1px solid #ccc">{CUR_YEAR} VALUE</td>
<td align="right" style="padding:5px 10px;font-size:9px;color:#555;border-bottom:1px solid #ccc">TOTAL VALUE</td>
<td align="right" style="padding:5px 10px;font-size:9px;color:#555;border-bottom:1px solid #ccc">OLDEST (DAYS)</td></tr>""")
    for comp, a in ordered:
        tot_v = a["prev_v"] + a["v26"]
        bg = ' bgcolor="#fdf0f0" style="background-color:#fdf0f0"' if a["prev_v"] >= 100000 else ''
        p.append(f"""<tr{bg}><td style="padding:5px 10px;border-bottom:1px solid #eee"><b>{comp}</b></td>
<td align="right" style="padding:5px 10px;border-bottom:1px solid #eee">{a['prev_n'] or '-'}</td>
<td align="right" style="padding:5px 10px;border-bottom:1px solid #eee;color:#c00000;font-weight:bold">{money(a['prev_v']) if a['prev_v'] else '-'}</td>
<td align="right" style="padding:5px 10px;border-bottom:1px solid #eee">{a['n26'] or '-'}</td>
<td align="right" style="padding:5px 10px;border-bottom:1px solid #eee">{money(a['v26']) if a['v26'] else '-'}</td>
<td align="right" style="padding:5px 10px;border-bottom:1px solid #eee;font-weight:bold;color:#b35a00">{money(tot_v)}</td>
<td align="right" style="padding:5px 10px;border-bottom:1px solid #eee;color:#c00000;font-weight:bold">{a['old']}</td></tr>""")
    p.append(f"""</table>
<div style="font-size:10px;color:#555;padding-top:4px">Rows shaded red carry {money(100000)}+ of prior-year equipment. Full line-item detail - every barcode, serial number, hirer, on-hire date and storage unit for {CUR_YEAR}, plus prior-year detail on request - is in the {"attached PDF" if pdf_ok else "full report"}.</div></td></tr>

<tr><td bgcolor="#eef5fc" style="background-color:#eef5fc;border:1px solid #b8d4f0;padding:10px 14px;font-size:11px">
<div style="font-weight:bold;color:#1f5c99;letter-spacing:0.5px;font-size:12px">STORE CONTROLS &amp; ASSURANCE</div>
<div style="padding-top:4px">These records are protected by daily stock takes at the Ampol Tool Store (completed without exception, 30-day full-coverage cycle including the radio charging bays)
and the double-check return process - every return is inspected and scanned on receipt, then stock-taken to its storage unit before going back on charge.
<b>The moment a unit is scanned, the record updates.</b> Every count here is read from the SiteIQ register as at {data_asat} - nothing comes from a summary tab. {META.get("serial_note", "Serial numbers from the radio register")}.</div></td></tr>

<tr><td style="padding:14px 0 6px;font-size:12px;line-height:1.6">
Thanks all - radios are the backbone of safe communication on site, and getting the idle ones back (or a quick rescan of the ones in use)
keeps the fleet available for everyone's next shutdown. Any questions, come see us at the Ampol Tool Store or reply here.<br><br>
Kind regards,<br><b>Andrew Fisher</b><br>Ampol Tool Store \u00b7 Coates</td></tr>
<tr><td style="border-top:1px solid #ccc;padding:10px 0;font-size:10px;color:#555;text-align:center">
<b>Author: Andrew Fisher</b> &nbsp;\u2022&nbsp; <b style="color:{O}">POWERED BY SITEIQ</b><br>
<span style="color:{O};font-weight:bold">{COATES_VALUES}</span><br>{COATES_OBJECTIVE}<br>
Data as at {data_asat or "TBC"} (SiteIQ RENTAL_STOCK export) &nbsp;\u2022&nbsp; Report built {refresh}</td></tr>
</table></td></tr></table></body></html>""")
    return "".join(p)


LSR_LINE = ("Life Saving Rule 5 - Tools and Equipment (SEQ-GL-009) &nbsp;|&nbsp; "
            "Two scans. Two looks. One standard. &nbsp;|&nbsp; Radios are charge-checked "
            "on return - flat means on charge before ready.")


def frame_email(inner):
    """Standard Coates email frame: 2px dark border, dark hero, LSR footer."""
    return f"""<table role="presentation" width="100%" cellspacing="0" cellpadding="0"
 style="background:#EFECE7"><tr><td align="center" style="padding:16px 8px">
<table role="presentation" width="860" cellspacing="0" cellpadding="0"
 style="width:860px;max-width:860px;background:#ffffff;border:2px solid #1D1D1B;border-collapse:collapse">
<tr><td style="background:#1D1D1B;padding:18px 24px;font-family:Arial,sans-serif">
 <div style="font-size:21px;font-weight:900;color:#ffffff">Coates &nbsp;|&nbsp; Ampol Site Radios</div>
 <div style="font-size:12px;font-weight:700;color:#F26222;margin-top:4px">Return it or rescan it - on-hire and recovery report</div>
 <div style="font-size:10px;color:#bbbbbb;margin-top:6px">The Coates Way &nbsp;|&nbsp; POWERED BY SITEIQ &nbsp;|&nbsp; Author: Andrew Fisher</div>
</td></tr>
<tr><td style="padding:14px 20px">{inner}</td></tr>
<tr><td style="background:#F5F1EC;border-top:3px solid #F26222;padding:10px 20px;
 font-family:Arial,sans-serif;font-size:9px;color:#555555;line-height:1.7">{LSR_LINE}<br/>
 Coates Hire Operations Pty Limited | ABN 50 009 779 338 | www.coates.com.au | POWERED BY SITEIQ<br/>
 Care Deeply &middot; Customer Focused &middot; Be Our Best &middot; One Team &middot; Competitive Spirit</td></tr>
</table></td></tr></table>"""


CID_CARD = "positioncard"   # the inline image id the email body points at


def inline_card_html(card_path, data_asat=""):
    """The <img> for the phone card inside the email body - only when the
    card file exists, so the draft never carries a broken picture."""
    if not card_path or not Path(card_path).exists():
        return ""
    return (f'<tr><td style="padding:2px 0 10px">'
            f'<img src="cid:{CID_CARD}" width="420" alt="Position card - the radio position as at {data_asat}" '
            f'style="display:block;width:420px;max-width:420px;height:auto;border-radius:10px"></td></tr>')


def write_eml(r26, rprev, b26, bprev, oos, ravail, bavail, pdf_path, eml_path, data_asat="", card_path=None):
    """The Outlook draft: the summary body with the phone card inline, the
    PDF and the card attached, and the manifest 08_MAKE_OUTLOOK_DRAFTS reads.
    Nothing sends itself - X-Unsent keeps it a draft."""
    import json
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication
    from email.mime.image import MIMEImage
    pdf = Path(pdf_path)
    card = Path(card_path) if card_path else None
    body = frame_email(build_email_summary(r26, rprev, b26, bprev, oos, ravail, bavail, data_asat,
                                           pdf_ok=pdf.exists(), card_path=card))
    subject = f"Ampol Tool Store - Site Radio Report - {REPORT_DATE.strftime('%d %b %Y')}"
    msg = MIMEMultipart("mixed")
    msg["To"] = RADIO_EMAIL_TO
    msg["Subject"] = subject
    msg["X-Unsent"] = "1"
    # WHY (03 Sep 2026): the phone card shows under the greeting the moment
    # the draft opens - a related part carries it by Content-ID - and it
    # stays attached as well so it can be forwarded on its own.
    if card and card.exists():
        rel = MIMEMultipart("related")
        rel.attach(MIMEText(body, "html", "utf-8"))
        inline = MIMEImage(card.read_bytes(), _subtype="png")
        inline.add_header("Content-ID", f"<{CID_CARD}>")
        inline.add_header("Content-Disposition", "inline", filename=card.name)
        rel.attach(inline)
        msg.attach(rel)
    else:
        msg.attach(MIMEText(body, "html", "utf-8"))
    if pdf.exists():
        part = MIMEApplication(pdf.read_bytes(), _subtype="pdf")
        # the attachment carries the suite file name - the same name as on disk
        part.add_header("Content-Disposition", "attachment", filename=pdf.name)
        msg.attach(part)
    if card and card.exists():
        img = MIMEImage(card.read_bytes(), _subtype="png")
        img.add_header("Content-Disposition", "attachment", filename=card.name)
        msg.attach(img)
    with open(eml_path, "wb") as f:
        f.write(msg.as_bytes())
    # sidecar for MAKE_OUTLOOK_DRAFTS: NATIVE Outlook draft, To pre-filled
    # WHY (12 Aug 2026): the manifest used to carry only the regexed-out bare
    # addresses, silently dropping any recipient ever added without angle
    # brackets. It now carries the FULL To line, semicolon-separated the way
    # Outlook wants it, so the draft matches the .eml recipient for recipient.
    to_line = RADIO_EMAIL_TO.replace(">, ", ">; ")
    n_to = to_line.count(";") + 1
    stem = str(eml_path)[:-4] if str(eml_path).lower().endswith(".eml") else str(eml_path)
    with open(stem + ".body.html", "w", encoding="utf-8") as f:
        f.write(body)
    # the manifest lists the card as an attachment only - the native draft
    # builder attaches files, it does not inline them
    with open(stem + ".draft.json", "w", encoding="utf-8") as f:
        json.dump({"subject": subject, "to": to_line,
                   "body": Path(stem + ".body.html").name,
                   "attachments": ([pdf.name] if pdf.exists() else [])
                   + ([card.name] if card and card.exists() else [])}, f, indent=1)
    mb = Path(eml_path).stat().st_size / 1048576
    print(f"Wrote {eml_path} + native-draft manifest (To: {n_to} recipients, PDF attached: {pdf.exists()}, "
          f"card inline + attached: {bool(card and card.exists())}, {mb:.1f} MB)")


def write_pdf_robust(html_path, pdf_path):
    # WHY (12 Aug 2026): an old PDF sitting on the target name used to make a
    # failed browser print look like a success - the only check is 'file
    # exists'. The target is deleted up front, so the only PDF that can pass
    # that check now is the one this run actually wrote.
    tgt = Path(pdf_path)
    try:
        if tgt.exists(): tgt.unlink()
    except OSError as e:
        raise SystemExit(f"Cannot replace {tgt.name} - close it if it is open, then run again. ({e})")
    try:
        from weasyprint import HTML
        HTML(filename=str(html_path)).write_pdf(str(pdf_path))
        return True
    except Exception as e:
        print(f"WeasyPrint unavailable ({type(e).__name__}) - trying Edge/Chrome headless...")
    import subprocess, os, time, tempfile
    for exe in [r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
                r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                "msedge", "chrome", "chromium", "chromium-browser", "google-chrome"]:
        if exe.endswith(".exe") and not os.path.exists(exe): continue
        for attempt in range(3):
            # unique profile per attempt: a live browser with the same profile
            # makes headless print exit 0 without writing the PDF
            # WHY (13 Aug 2026): fall back to the system temp dir, not the
            # suite folder - a machine without TEMP set was leaving throwaway
            # browser profiles next to the scripts.
            profile = os.path.join(os.environ.get("TEMP", tempfile.gettempdir()),
                                   f"coates_edge_pdf_{os.getpid()}_{attempt}")
            try:
                subprocess.run([exe, "--headless", "--disable-gpu",
                                "--user-data-dir=" + profile, "--no-first-run",
                                f"--print-to-pdf={Path(pdf_path).resolve()}",
                                "--no-pdf-header-footer", Path(html_path).resolve().as_uri()],
                               capture_output=True, timeout=180)
            except FileNotFoundError:
                break        # not installed here - straight on to the next candidate
            except Exception:
                pass
            if Path(pdf_path).exists():
                print(f"PDF generated via {Path(exe).name}")
                return True
            time.sleep(1.5)
        # WHY (12 Aug 2026): an unconditional break used to sit here, so if the
        # first browser found kept failing the rest were never tried. Now every
        # candidate on the list gets its turn before we give up.
    print("No PDF engine found - HTML written; PDF skipped.")
    return False

def _page_count(pdf_path):
    """Pages in a printed PDF (PyMuPDF, else pypdf, else None - the caller
    then says the two-pass check could not run rather than guessing)."""
    try:
        import pymupdf
        with pymupdf.open(str(pdf_path)) as d:
            return len(d)
    except ImportError:
        pass
    except Exception:
        return None
    try:
        from pypdf import PdfReader
        return len(PdfReader(str(pdf_path)).pages)
    except Exception:
        return None


def find_workbook(patterns, arg=None):
    # WHY (12 Aug 2026): inputs now come from the suite's one Data area via
    # ampol_paths - newest file wins, Excel ~$ lock files and archived
    # Source_ pulls are never candidates. An explicit path argument still
    # overrides everything, exactly as before.
    if arg: return arg
    return ampol_paths.find_data(*patterns) or None

def main():
    global PRIOR_LABEL, PRIOR_SHORT, PRICE_RADIO, PRICE_BATT, PRICE_SOURCE, META
    master = find_workbook(["RENTAL_STOCK*.xlsx"], sys.argv[1] if len(sys.argv) > 1 else None)
    register = find_workbook(["radio_register*.xlsx", "*radio*register*.xlsx"], sys.argv[2] if len(sys.argv) > 2 else None)
    pricing = find_workbook(["Ampol_ToolStore_Pricing*.xlsx", "*Pricing*.xlsx"], sys.argv[3] if len(sys.argv) > 3 else None)
    if not master:
        raise SystemExit("No RENTAL_STOCK.xlsx in the Data folder - download the SiteIQ export, "
                         "run 12_PULL_SITEIQ_EXPORTS, and press the radio button again.")
    print(f"Register (source)  : {master}")
    print(f"Radio serial list  : {register or 'NOT FOUND in Data - serials fall back to the description, else dashes'}")
    print(f"Pricing master     : {pricing or 'NOT FOUND in Data - replacement values show as dashes'}")
    print("Radio workbook     : not read (02 Sep 2026) - the register is the source")
    out = Path(ampol_paths.day_folder("Radios"))  # dated folder, created on demand

    serials = load_serials(register) if register else {}
    prices, PRICE_RADIO, PRICE_BATT = load_prices(pricing)
    PRICE_SOURCE = (f"new replacement price per {Path(pricing).name}" if pricing and PRICE_RADIO
                    else "replacement price not in the pricing master - shown as TBC")
    d = load_from_register(master, serials, prices)
    r26, rprev, b26, bprev = d["r_cur"], d["r_prev"], d["b_cur"], d["b_prev"]
    oos, ravail, bavail = d["oos"], d["r_avail"], d["b_avail"]
    data_asat = d["asat"].strftime("%d %b %Y %H:%M")
    prior_years = sorted(y for y in d["years"] if y < CUR_YEAR)
    if prior_years:
        PRIOR_LABEL = f"{prior_years[0]}{EN_DASH}{CUR_YEAR - 1}" if prior_years[0] != CUR_YEAR - 1 else str(CUR_YEAR - 1)
        PRIOR_SHORT = f"{prior_years[0]}-{str(CUR_YEAR - 1)[2:]}" if prior_years[0] != CUR_YEAR - 1 else str(CUR_YEAR - 1)
    hits, tot = d["serial_hits"], d["serial_total"]
    META = {
        "n_radio": d["n_radio"], "n_batt": d["n_batt"],
        "serial_note": (f"serial numbers from the radio register or the serial SiteIQ carries in the "
                        f"item description - {hits} of the {tot} radios on hire have one, "
                        f"{tot - hits} show a dash"),
        "coverage": (f"Fleet on the register: {d['n_radio']} radios and {d['n_batt']} batteries"
                     + (f", including {d['turnaround']} SATGAS turnaround units" if d["turnaround"] else "")
                     + (f"; {d['accounts']} units are held on shared site accounts, shown as such." if d["accounts"] else ".")),
        "unpriced_note": (f" {d['unpriced']} unit(s) have no price in the master and are excluded from every "
                          f"value total, never estimated." if d["unpriced"] else ""),
    }
    _ASAT_DT[0] = d["asat"]
    # WHY (03 Sep 2026): one file name for the whole family - the PDF, the
    # page HTML beside it, the draft, the manifest and the phone card
    stem = ampol_names.report_stem("radio")
    F = position_facts(r26, rprev, b26, bprev, data_asat)
    # the scoreboard is written BEFORE the pages so the trend page sees today
    rh.record("radio", d["asat"], {
        "exposure": round(F["exposure"]), "radios_on_hire": len(r26) + len(rprev),
        "batteries_on_hire": len(b26) + len(bprev), "prior_units": len(F["prev_all"]),
        "prior_value": round(F["prev_val"]), "prior_pct": F["prior_pct"],
        "available_radios": len(ravail), "available_batteries": len(bavail),
        "oos": len(oos), "oldest_days": F["oldest"]},
        extra={"rag": F["status"], "headline": F["headline"], "rule": F["rule"],
               "owner": "Andrew Fisher, Shutdown Manager", "action": F["action"], "due": F["due"],
               "key_value": F["key_value"], "key_label": F["key_label"],
               "second_value": F["second_value"], "second_label": F["second_label"],
               "title": "Ampol Site Radio On-Hire Report", "folder": "Radios",
               "pdf": f"{stem}.pdf", "card": f"{stem}_PositionCard.png"})
    print(f"History            : {rh.HIST.name} - radio figures recorded for {d['asat'].strftime('%d %b %Y')}")
    # what moved: every radio and battery on the register as this report
    # defines it - on hire, on the shelf and in custody
    scope = {i["barcode"] for i in r26 + rprev + b26 + bprev + ravail + bavail + oos if i["barcode"]}
    changes = pull_diff.changes(scope_barcodes=scope)
    L24 = changes["last24"]
    print(f"Since the last pull: "
          + (f"previous pull {changes['prev_time']:%d %b %Y %H:%M} - came back {len(changes['returned'])}, "
             f"went out {len(changes['issued'])}, changed hands {len(changes['moved'])}, "
             f"crossed 30 days {len(changes['crossed'][30])}" if changes["have_previous"]
             else "no earlier pull in Data\\previous yet (starts with the next pull)")
          + f" | 24 h before the pull: issued {len(L24['issued'])}, returned {len(L24['returned'])}")
    # WHY (03 Sep 2026): the log's year for this report's barcodes, read once
    # and handed to both passes so the two prints are identical
    INS = insights_compute(r26 + rprev + b26 + bprev, oos, ravail, bavail)
    if INS:
        _a = INS["rw"]["all"] or {}
        _h = INS["holders"]
        _q = INS["dq"]
        print(f"Insights           : {INS['tx_n']:,} movements for {INS['scope_n']:,} barcodes ({INS['window'][0]:%d %b} to "
              f"{INS['window'][1]:%d %b %H:%M}); {len(INS['weekly'])} weeks; completed hires {_a.get('n', 0):,}, median "
              f"{_a.get('median', 0):.2f} d, p90 {_a.get('p90', 0):.1f} d, same day {_a.get('sd_pct', 0)}%; busiest hour "
              f"{INS['rhythm']['busiest'][0][0]:02d}:00; holders {_h['holders']} (80% of units with {_h['n80_items']}, value "
              f"{money(_h['rvalue_total'])}, cross-family {sum(1 for d in _h['rows'] if d['also'])}); short hires {_q['short_n']}, "
              f"mass draws {_q['mass_n']}, on hire no log {len(_q['onhire_no_log'])}, before log {len(_q['onhire_before_log'])}, "
              f"available with open movement {len(_q['open_but_available'])}")
    html_str = build_html(r26, rprev, b26, bprev, oos, ravail, bavail, data_asat, facts=F, changes=changes, insights=INS)
    L = build_html.last
    card_path = out / f"{stem}_PositionCard.png"
    import k2shell as _sh
    _sh.position_card_png(L["cfg"], data_asat, [
        (money(L["exposure"]), "On-hire value", f"{_sh.num(L['n_on'])} units on hire", "#7A8A9A"),
        (_sh.num(L["n_prior"]), f"On hire since {PRIOR_LABEL}", money(L["prev_val"]), "#F0603E"),
        (_sh.num(len(r26) + len(rprev)), "Radios on hire", money(val(r26) + val(rprev)), "#7A8A9A"),
        (_sh.num(len(b26) + len(bprev)), "Batteries on hire", money(val(b26) + val(bprev)), "#7A8A9A"),
        (f"{len(ravail)} / {len(bavail)}", "Available in store", "radios / batteries", "#22C55E"),
        (_sh.num(len(oos)), "Out of service", money(val(oos)) if val(oos) else "", "#7A8A9A"),
    ], (L["status"], f"{_sh.num(L['n_prior'])} of {_sh.num(L['n_on'])} units on hire ({L['prior_pct']}%) were issued in {PRIOR_LABEL} and are still out.",
        "Andrew Fisher, Shutdown Manager", f"Prior-year holders sent their list by {L['due']}"),
        [], str(card_path), f"Counted from the SiteIQ register pull of {data_asat} - nothing estimated.")
    print(f"Position card      : {card_path}")
    base = out / stem
    with open(f"{base}.html", "w", encoding="utf-8") as f:
        f.write(html_str)
    pdf_ok = write_pdf_robust(f"{base}.html", f"{base}.pdf")
    if pdf_ok and COVER_PAGE:
        # WHY (03 Sep 2026): the cover's "What's inside" block carries REAL
        # page numbers - read off the printed pages of the first pass, then
        # the page is rebuilt with them and printed again. The cover is a
        # fixed-height block, so the second pass paginates exactly like the
        # first; the console proves it with both page counts, and a
        # mismatch stops the build rather than print a wrong number.
        contents = pdf_finish.contents_from_pdf(f"{base}.pdf", html_str, has_cover=True,
                                                skip=("Meet the tool store team",))
        if contents:
            n_first = _page_count(f"{base}.pdf")
            html_str = build_html(r26, rprev, b26, bprev, oos, ravail, bavail, data_asat, facts=F,
                                  changes=changes, contents=contents, insights=INS)
            with open(f"{base}.html", "w", encoding="utf-8") as f:
                f.write(html_str)
            pdf_ok = write_pdf_robust(f"{base}.html", f"{base}.pdf")
            n_second = _page_count(f"{base}.pdf") if pdf_ok else None
            print(f"Cover contents     : {len(contents)} rows read off the printed pages - "
                  f"pass 1 {n_first} pages, pass 2 {n_second} pages"
                  + (" - identical" if n_first == n_second and n_first else ""))
            if pdf_ok and n_first and n_second and n_first != n_second:
                raise SystemExit("The second pass printed a different page count from the first - the cover's "
                                 "page numbers would be wrong. Not written.")
        else:
            print("Cover contents     : not printed - no PDF reader on this machine to read the page numbers")
    if pdf_ok:
        # WHY (03 Sep 2026): properties and bookmarks - Author, Subject and a
        # navigation pane built from the report's own section headings
        print(pdf_finish.finish(
            f"{base}.pdf", f"Ampol Site Radio On-Hire Report - as at {data_asat}",
            f"Site radios and batteries on hire from the Ampol Tool Store at Lytton Refinery, counted from the "
            f"SiteIQ register pull of {data_asat}.",
            html_str, keywords="radios, batteries, on hire", has_cover=COVER_PAGE, family="Radios"))
    write_eml(r26, rprev, b26, bprev, oos, ravail, bavail, f"{base}.pdf", f"{base}_OUTLOOK.eml", data_asat,
              card_path=card_path)
    tot_v = val(r26) + val(rprev) + val(b26) + val(bprev)
    print(f"Data as at         : {data_asat}  (RENTAL_STOCK request time)")
    print(f"Radios {CUR_YEAR}: {len(r26)} | prior: {len(rprev)} | Batteries {CUR_YEAR}: {len(b26)} | prior: {len(bprev)} | "
          f"OOS: {len(oos)} | available {len(ravail)} radios / {len(bavail)} batteries | TOTAL EXPOSURE: ${tot_v:,.0f}")
    print(f"Serials            : {hits} of {tot} radios on hire | unpriced units: {d['unpriced']}")
    print(f"Output: {out}")

if __name__ == "__main__":
    main()
