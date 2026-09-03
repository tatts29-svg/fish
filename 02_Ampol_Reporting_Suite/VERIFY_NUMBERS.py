#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VERIFY_NUMBERS - the truth table. Button 13.
Author: Andrew Fisher | POWERED BY SITEIQ

WHAT THIS IS
  A second, deliberately simple count of the key figures on today's
  pages, taken straight from the raw SiteIQ exports in Data\ with none
  of the report code involved - then checked against the pages that
  were built today. If a page says 1,500 and this script counts 1,500,
  the number is proven twice. If they differ, the run fails red and
  names the figure.

  It also sweeps every built page for the site's former name (the site
  is Ampol; the old word must not appear in any printed text) and
  prints how many raw SiteIQ lines still carry it, so the reader knows
  the rename is a display rule, disclosed, not a change to the data.

WHY (02 Sep 2026)
  "The data is wrong" is easy to say and hard to answer without proof.
  This is the proof: every figure below can be checked by hand in the
  exports, and the table is written to Reports\<date>\VERIFY_NUMBERS.txt
  so it travels with the day's reports.
"""
import glob
import os
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import ampol_paths
import ampol_names

import openpyxl

BASE = Path(__file__).resolve().parent
DATESTR = datetime.now().strftime("%Y-%m-%d")
REPORTS = BASE / "Reports" / DATESTR


def find(pattern):
    p = ampol_paths.find_data(pattern)
    if not p:
        sys.exit(f"ERROR: no {pattern} in the suite's Data folder.")
    return p


def rows_of(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c is not None else "" for c in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in it:
        if not any(c not in (None, "") for c in r):
            continue
        out.append({h: r[i] for h, i in ix.items() if i < len(r)})
    wb.close()
    return out


def s(v):
    return str(v if v is not None else "").strip()


def ddmmyyyy(v):
    if isinstance(v, datetime):
        return v
    try:
        return datetime.strptime(s(v)[:10], "%d/%m/%Y")
    except ValueError:
        return None


def page_text(folder_glob):
    """All text of the built HTML pages under Reports/<date>/<family>."""
    txt = []
    for f in glob.glob(str(REPORTS / folder_glob)):
        raw = open(f, encoding="utf-8", errors="ignore").read()
        txt.append(re.sub(r"<[^>]+>", " ", raw))
    return " ".join(txt)


def fmt(n):
    return f"{n:,}" if isinstance(n, int) else str(n)


# ---------------------------------------------------------------- the counts
def count_everything():
    reg = rows_of(find("RENTAL_STOCK*.xlsx"), "RENTAL_STOCK")
    st = rows_of(find("STOCKTAKE*.xlsx"), "STOCKTAKE")
    tx_path = find("TRANSACTIONS*.xlsx")
    tx = rows_of(tx_path, "CUSTOMER_CONTRACTOR_EQUIP")
    facts = {}
    old = Counter()

    # ---- register basics ----------------------------------------------
    by_status = Counter(s(r["ITEM_STATUS"]) for r in reg)
    facts["register: rows"] = len(reg)
    facts["register: on hire"] = by_status["On Hire"]
    facts["register: available for hire"] = by_status["Available for Hire"]
    old["register descriptions"] = sum(1 for r in reg if ampol_names.carries_former_name(r["ITEM_DESCRIPTION"]))
    old["register company accounts"] = sum(1 for r in reg if ampol_names.carries_former_name(r["COMPANY_NAME"]))

    # ---- gas monitors (same words as gasmon_engine, written out again) --
    gas_re = re.compile(r"x-am|gas monitor", re.I)
    not_gas = re.compile(r"charger|probe|pump|calibration gas|dock|cradle", re.I)
    gas = [r for r in reg if gas_re.search(s(r["ITEM_DESCRIPTION"])) and not not_gas.search(s(r["ITEM_DESCRIPTION"]))]
    facts["gas: fleet on register"] = len(gas)
    facts["gas: available in store"] = sum(1 for r in gas if s(r["ITEM_STATUS"]) == "Available for Hire")
    facts["gas: on hire"] = sum(1 for r in gas if s(r["ITEM_STATUS"]) == "On Hire")

    # ---- gas monitors, pull against pull (03 Sep 2026) -------------------
    # The newest earlier RENTAL_STOCK export in Data\previous whose request
    # time is before the current one, read here with the same gas words:
    # came back = on hire then, not on hire now; went out = on hire now,
    # not on hire then (or not on the register then). The gas report's
    # "since the last pull" page prints both.
    prev_reg, prev_time = None, None
    try:
        cur_req = None
        _wb = openpyxl.load_workbook(find("RENTAL_STOCK*.xlsx"), read_only=True, data_only=True)
        for r in _wb["REFERENCE_INFO"].iter_rows(min_row=2, max_row=2, values_only=True):
            for cell in r:
                try:
                    cur_req = datetime.strptime(s(cell), "%d/%m/%Y %I:%M %p")
                    break
                except ValueError:
                    pass
        _wb.close()
        cands = []
        for f in [p for d in ampol_paths.previous_dirs() for p in glob.glob(os.path.join(d, "*RENTAL_STOCK*.xlsx"))]:
            _w = openpyxl.load_workbook(f, read_only=True, data_only=True)
            t0 = None
            if "REFERENCE_INFO" in _w.sheetnames:
                for r in _w["REFERENCE_INFO"].iter_rows(min_row=2, max_row=2, values_only=True):
                    for cell in r:
                        try:
                            t0 = datetime.strptime(s(cell), "%d/%m/%Y %I:%M %p")
                            break
                        except ValueError:
                            pass
            _w.close()
            if t0 and cur_req and t0 < cur_req:
                cands.append((t0, f))
        if cands:
            prev_time, pf = max(cands)
            prev_reg = rows_of(pf, "RENTAL_STOCK")
    except Exception:
        prev_reg = None
    if prev_reg is not None:
        pg = {s(r["ITEM_BARCODE"]).upper(): s(r["ITEM_STATUS"]) for r in prev_reg
              if s(r["ITEM_BARCODE"]) and gas_re.search(s(r["ITEM_DESCRIPTION"])) and not not_gas.search(s(r["ITEM_DESCRIPTION"]))}
        cg = {s(r["ITEM_BARCODE"]).upper(): s(r["ITEM_STATUS"]) for r in gas if s(r["ITEM_BARCODE"])}
        facts["gas: came back since the last pull"] = sum(1 for b, st0 in pg.items() if st0 == "On Hire" and b in cg and cg[b] != "On Hire")
        facts["gas: went out since the last pull"] = sum(1 for b, st1 in cg.items() if st1 == "On Hire" and pg.get(b) != "On Hire")

    # ---- radios --------------------------------------------------------
    radio = [r for r in reg if re.search(r"motorola radio", s(r["ITEM_DESCRIPTION"]), re.I)
             and not re.search(r"charger|batter", s(r["ITEM_DESCRIPTION"]), re.I)]
    batt = [r for r in reg if re.search(r"radio batter", s(r["ITEM_DESCRIPTION"]), re.I)]
    facts["radio: radios on register"] = len(radio)
    facts["radio: batteries on register"] = len(batt)
    facts["radio: radios available"] = sum(1 for r in radio if s(r["ITEM_STATUS"]) == "Available for Hire")

    # ---- transactions --------------------------------------------------
    facts["transactions: rows this year"] = len(tx)
    same = sum(1 for r in tx if s(r["TRAN_END_DATE"]) and s(r["TRAN_END_DATE"]) == s(r["TRAN_START_DATE"]))
    facts["transactions: same-day %"] = round(same / len(tx) * 100) if tx else 0

    # ---- tooling master (workbook definition, written out again) -------
    exc = re.compile(r"RADIO|GAS MONITOR|GAS DETECTOR|MULTI GAS|MULTIGAS|DRAGER|LANYARD|STEEL COIL CLAMP", re.I)
    custody = {"T&I - TOOL STORE", "ALL-AROUND - REPAIRS", "BULK - YARD",
               "LOADING BAY - OUT OF SERVICE", "OUT OF - CALIBRATION",
               "RIGGING & 240V - OUT OF TAG DATE"}
    master, q = 0, Counter()
    legacy = 0
    for r in reg:
        if s(r["ITEM_STATUS"]) != "On Hire":
            continue
        d = ddmmyyyy(r["ON_HIRE_DATE"])
        if d and d.year < 2026:
            legacy += 1
            continue
        if exc.search(s(r["ITEM_DESCRIPTION"])) or s(r["HIRER_NAME"]).upper() in custody:
            continue
        master += 1
        if d:
            q[(d.month - 1) // 3 + 1] += 1
    facts["tooling: on hire (2026 tooling)"] = master
    facts["tooling: Q1 / Q2 / Q3"] = f"{q[1]} / {q[2]} / {q[3]}"
    facts["tooling: pre-2026 rows excluded"] = legacy
    facts["tooling: available (tooling)"] = sum(1 for r in reg if s(r["ITEM_STATUS"]) == "Available for Hire"
                                              and not exc.search(s(r["ITEM_DESCRIPTION"])))
    facts["tooling: rows on the Repairs account"] = sum(1 for r in reg if s(r["COMPANY_NAME"]) == "Repairs")

    # ---- what the transaction log says, store-wide (03 Sep 2026) --------
    # Counted here from the raw rows with none of txn_insights involved:
    # the pages print these four figures on the executive summary and the
    # utilisation report, and they must agree.
    import datetime as _dt
    from datetime import date as _date
    pull_stamp = None
    try:
        _wb = openpyxl.load_workbook(find("RENTAL_STOCK*.xlsx"), read_only=True, data_only=True)
        if "REFERENCE_INFO" in _wb.sheetnames:
            _rows = list(_wb["REFERENCE_INFO"].iter_rows(min_row=1, max_row=2, values_only=True))
            _hdr = [str(h or "").upper() for h in _rows[0]]
            for _i, _h in enumerate(_hdr):
                if "REQUESTED_DATE" in _h and _i < len(_rows[1]) and _rows[1][_i]:
                    pull_stamp = _dt.datetime.strptime(str(_rows[1][_i]).strip(), "%d/%m/%Y %I:%M %p")
        _wb.close()
    except Exception:
        pull_stamp = None
    pull_day = pull_stamp.date() if pull_stamp else _date.today()
    qn = (pull_day.month - 1) // 3
    qend = _date(pull_day.year, qn * 3 + 3, [31, 30, 30, 31][qn])
    crossing = 0
    for r in reg:
        if s(r["ITEM_STATUS"]) != "On Hire" or not s(r["ITEM_BARCODE"]):
            continue
        try:
            d0 = _dt.datetime.strptime(s(r["ON_HIRE_DATE"]), "%d/%m/%Y").date()
        except ValueError:
            continue
        if 0 <= (pull_day - d0).days < 90 <= (qend - d0).days:
            crossing += 1
    facts["store: on hire crossing 90 days by quarter close"] = crossing
    moved = {s(r["LATEST_BARCODE"]).upper() for r in tx if s(r["LATEST_BARCODE"])}
    facts["store: available items never moved this year"] = sum(
        1 for r in reg if s(r["ITEM_STATUS"]) == "Available for Hire" and s(r["ITEM_BARCODE"])
        and s(r["ITEM_BARCODE"]).upper() not in moved)
    short = 0
    hours = Counter()
    for r in tx:
        try:
            st0 = _dt.datetime.strptime(f"{s(r['TRAN_START_DATE'])} {s(r['TRAN_START_TIME'])}", "%d/%m/%Y %H:%M:%S")
        except ValueError:
            continue
        hours[st0.hour] += 1
        if s(r["TRAN_END_DATE"]):
            try:
                en0 = _dt.datetime.strptime(f"{s(r['TRAN_END_DATE'])} {s(r['TRAN_END_TIME'])}", "%d/%m/%Y %H:%M:%S")
            except ValueError:
                continue
            if 0 <= (en0 - st0).total_seconds() < 6 * 60:
                short += 1
    facts["transactions: closed inside 6 minutes"] = short
    busiest = max(hours.items(), key=lambda kv: (kv[1], -kv[0]))[0] if hours else None
    facts["transactions: busiest counter hour"] = f"{busiest:02d}:00" if busiest is not None else "-"

    # ---- stocktake -----------------------------------------------------
    regset = {s(r["ITEM_BARCODE"]).upper() for r in reg if s(r["ITEM_BARCODE"])}
    onhire_set = {s(r["ITEM_BARCODE"]).upper() for r in reg if s(r["ITEM_STATUS"]) == "On Hire"}
    countable, departed, instore, onhire = 0, 0, 0, 0
    for r in st:
        if not s(r["TOOL_STORE"]):
            continue
        bc = s(r["LATEST_BARCODE"]).upper()
        status = s(r["SIGHTED_STATUS"]).upper()
        action = s(r["LAST_SIGHTED_ACTION"]).upper()
        if status == "PENDING BRANCH RECEIPT" or (action == "DEPARTURE" and bc not in regset):
            departed += 1
            continue
        countable += 1
        if bc in onhire_set:
            onhire += 1
        else:
            instore += 1
    facts["stocktake: countable items"] = countable
    facts["stocktake: in store / on hire"] = f"{instore:,} / {onhire:,}"
    facts["stocktake: departed lines excluded"] = departed

    # ---- rigging register ------------------------------------------------
    rig_path = ampol_paths.find_data("Rigging*Register*.xlsx", "*Rigging*.xlsx")
    if rig_path:
        rig = rows_of(rig_path, "Rigging register Master")
        bcs = [s(r.get("ITEM_BARCODE")).upper() for r in rig]
        distinct = {b for b in bcs if b}
        found = {b for b in distinct if b in regset}
        facts["rigging: register rows / distinct barcodes"] = f"{len(rig)} / {len(distinct)}"
        facts["rigging: found in SiteIQ / not found"] = f"{len(found)} / {len(distinct) - len(found)}"
        facts["rigging: on hire (all accounts)"] = sum(1 for b in found if b in onhire_set)

    # ---- calibration register --------------------------------------------
    cal_path = ampol_paths.find_data("*Calibration*Register*.xlsx")
    if cal_path:
        cal = [r for r in rows_of(cal_path, "Register Entry") if s(r.get("New Asset No"))]
        asat = datetime.now().date()
        pulled = None
        wb = openpyxl.load_workbook(find("RENTAL_STOCK*.xlsx"), read_only=True, data_only=True)
        if "REFERENCE_INFO" in wb.sheetnames:
            for r in wb["REFERENCE_INFO"].iter_rows(min_row=2, max_row=2, values_only=True):
                for cell in r:
                    try:
                        pulled = datetime.strptime(s(cell), "%d/%m/%Y %I:%M %p")
                        break
                    except ValueError:
                        pass
        wb.close()
        if pulled:
            asat = pulled.date()
        b = Counter()
        for r in cal:
            d = r.get("Calibration Due")
            d = d if isinstance(d, datetime) else ddmmyyyy(d)
            if d is None:
                b["no date"] += 1
            else:
                left = (d.date() - asat).days
                b["overdue" if left < 0 else "due 30" if left <= 30 else "in cal"] += 1
        facts["calibration: assets on register"] = len(cal)
        facts["calibration: in cal / due 30 / overdue / no date"] = (
            f"{b['in cal']} / {b['due 30']} / {b['overdue']} / {b['no date']}")
    return facts, old


# ---------------------------------------------------------------- the checks
CHECKS = [
    # (fact key, page folder glob, how the page prints it)
    ("gas: fleet on register", "Gas_Monitors/*.html", None),
    ("gas: came back since the last pull", "Gas_Monitors/*.html", None),
    ("gas: went out since the last pull", "Gas_Monitors/*.html", None),
    ("gas: available in store", "Gas_Monitors/*.html", None),
    ("radio: radios available", "Radios/*.html", None),
    ("tooling: on hire (2026 tooling)", "Tooling/*Executive_Summary*.html", None),
    ("tooling: available (tooling)", "Tooling/*Executive_Summary*.html", None),
    ("tooling: rows on the Repairs account", "Tooling/*Executive_Summary*.html", None),
    ("transactions: rows this year", "Tooling/*Executive_Summary*.html", None),
    ("store: on hire crossing 90 days by quarter close", "Tooling/*Executive_Summary*.html", None),
    ("store: available items never moved this year", "Tooling/*Utilisation*.html", None),
    ("transactions: closed inside 6 minutes", "Tooling/*Executive_Summary*.html", None),
    ("transactions: busiest counter hour", "Tooling/*Executive_Summary*.html", "text"),
    ("stocktake: countable items", "Stocktake/*.html", None),
    ("stocktake: departed lines excluded", "Stocktake/*.html", None),
    ("calibration: assets on register", "Calibrations/*.html", None),
    ("rigging: register rows / distinct barcodes", "Rigging/*.html", "pair"),
]


def main():
    print("=" * 70)
    print(" COATES | VERIFY NUMBERS - the truth table")
    print(" Ampol Tool Store (Lytton Refinery) | POWERED BY SITEIQ")
    print("=" * 70)
    if not REPORTS.exists():
        sys.exit(f"ERROR: no Reports\\{DATESTR} folder - build the reports first (button 00).")
    facts, old = count_everything()
    lines = []
    lines.append(f"Counted from Data\\ exports at {datetime.now().strftime('%d %b %Y %H:%M')} - "
                 "none of the report code involved.")
    lines.append("")
    lines.append(f"{'FIGURE':46s} {'THIS COUNT':>14s}   ON TODAY'S PAGE")
    lines.append("-" * 70)
    fails = 0
    for key, val in facts.items():
        check = next((c for c in CHECKS if c[0] == key), None)
        verdict = ""
        if check:
            text = page_text(check[1])
            if not text:
                verdict = "page not built today"
            elif check[2] == "text":
                ok = str(val) in text
                verdict = "MATCHES" if ok else "NOT FOUND ON PAGE"
                fails += 0 if ok else 1
            elif check[2] == "pair":
                a, b = [x.strip() for x in str(val).split("/")]
                ok = fmt(int(a)) in text and fmt(int(b)) in text
                verdict = "MATCHES" if ok else "NOT FOUND ON PAGE"
                fails += 0 if ok else 1
            else:
                ok = fmt(val) in text
                verdict = "MATCHES" if ok else "NOT FOUND ON PAGE"
                fails += 0 if ok else 1
        lines.append(f"{key:46s} {fmt(val):>14s}   {verdict}")
    lines.append("")
    lines.append("Former site name (display rule, disclosed on every data page):")
    for k, v in old.items():
        lines.append(f"  raw SiteIQ lines still carrying it - {k}: {v:,}")
    pages = glob.glob(str(REPORTS / "*" / "*.html")) + glob.glob(str(REPORTS / "*" / "*.eml"))
    hits = []
    for f in pages:
        raw = open(f, encoding="utf-8", errors="ignore").read()
        n = len(re.findall(ampol_names.FORMER_SITE_NAME, re.sub(r"<[^>]+>", " ", raw), re.I))
        if n:
            hits.append((os.path.relpath(f, REPORTS), n))
    # the PDFs themselves, when a PDF reader is on the machine (some pages
    # are printed from a temporary file and have no HTML beside them)
    pdfs = glob.glob(str(REPORTS / "*" / "*.pdf"))
    try:
        import pymupdf
        for f in pdfs:
            text = " ".join(pg.get_text() for pg in pymupdf.open(f))
            n = len(re.findall(ampol_names.FORMER_SITE_NAME, text, re.I))
            if n:
                hits.append((os.path.relpath(f, REPORTS), n))
        pages += pdfs
    except ImportError:
        pass
    if hits:
        fails += 1
        lines.append("  PRINTED on today's pages (must be zero):")
        for f, n in hits:
            lines.append(f"    {f}: {n}")
    else:
        lines.append(f"  printed on today's pages: 0 (checked {len(pages)} files)")
    # WHY (03 Sep 2026): one dash style on every page - the long dash is a
    # typing habit that reads differently from family to family. The
    # sweep counts it in the printed text (tags stripped); must be zero.
    dash_hits = []
    for f in glob.glob(str(REPORTS / "*" / "*.html")):
        raw = open(f, encoding="utf-8", errors="ignore").read()
        n = len(re.findall("\u2014|&mdash;|&#8212;", re.sub(r"<style.*?</style>", " ", raw, flags=re.S)))
        if n:
            dash_hits.append((os.path.relpath(f, REPORTS), n))
    # WHY (03 Sep 2026): values must not silently vanish - the stocktake
    # worklist lost every price the day a reader looked for a workbook by
    # its old name. A page that prices its rows must carry dollar figures.
    lines.append("")
    lines.append("Values present (a priced page must carry dollar figures):")
    for rel, need in (("Stocktake/*Count_Worklist*.html", 50), ("Tooling/*Executive_Summary*.html", 20),
                      ("Radios/*.html", 20)):
        n = sum(len(re.findall(r"\$[\d,]+", open(f, encoding="utf-8").read()))
                for f in glob.glob(str(REPORTS / rel)))
        ok = n >= need
        if not ok:
            fails += 1
        lines.append(f"  {rel:40} {n:>6} dollar figures  {'MATCHES' if ok else 'FAIL - values missing'}")
    lines.append("")
    lines.append("One dash style (the long dash must not print):")
    if dash_hits:
        fails += 1
        for f, n in dash_hits:
            lines.append(f"    {f}: {n} long dash(es)")
    else:
        lines.append("  long dashes on today's pages: 0")
    lines.append("")
    if fails:
        lines.append(f"FAIL - {fails} figure(s) on today's pages do not match this count. Do not send.")
    else:
        lines.append("PASS - every checked figure on today's pages matches an independent count.")
    out = "\n".join(lines)
    print(out)
    (REPORTS / "VERIFY_NUMBERS.txt").write_text(out + "\n", encoding="utf-8")
    print(f"\nWritten: {REPORTS / 'VERIFY_NUMBERS.txt'}")
    sys.exit(1 if fails else 0)


if __name__ == "__main__":
    main()
