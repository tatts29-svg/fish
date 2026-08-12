#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=====================================================================
COATES - GAS MONITOR TRANSACTIONS ANALYTICS
Mines the SiteIQ TRANSACTIONS export for flow, rhythm and behaviour
=====================================================================
Author: Andrew Fisher
The Coates Way - Operational Excellence - POWERED BY SITEIQ

WHAT THIS IS
  The V18 workbook engine answers "where is every monitor RIGHT NOW".
  This module answers the questions the register cannot: when gear
  moves, how hard the morning peak hits, whether same-day returns are
  trending the right way, and WHICH PEOPLE keep gear overnight.

  It reads the SiteIQ TRANSACTIONS export (TRANSACTIONS.xlsx, sheet
  CUSTOMER_CONTRACTOR_EQUIP) - every issue and return with a
  timestamp - and filters to gas monitors by description ("X-am" /
  "gas monitor") plus the asset list in Gas_Monitor_Serial_Numbers.

DATA RULES
  - Dates are parsed explicitly as DD/MM/YYYY. Australian exports read
    with a US default produce plausible, wrong reports.
  - Internal Dräger service/custody accounts (Failed Bump Test, Out of
    Calibration, FCCU, Out of Service...) are EXCLUDED from crew
    behaviour stats and the people league - they are workflows, not
    people. They stay IN the concurrency curve, because a unit in
    custody is still a unit off the shelf.
  - The export window opens 01 May: anything issued before then never
    appears here, so "open transactions" UNDERCOUNTS the register's
    on-hire figure. This module is for FLOW analytics; the workbook
    register remains the authority for right-now counts. The
    reconciliation is printed, never hidden.
=====================================================================
"""

import os
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl


def _parse_dt(dv, tv):
    """DD/MM/YYYY + HH:MM:SS, parsed explicitly - never the US default."""
    if not dv:
        return None
    ds = str(dv).strip()[:10]
    ts = str(tv).strip()[:8] if tv else "00:00:00"
    for f in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(ds + " " + ts, f)
        except ValueError:
            continue
    return None


def _is_internal(who, co):
    w = (who or "").lower()
    c = (co or "").lower()
    return "dräger" in w or "drager" in w or c in ("dräger", "drager")


def load_serial_assets(path):
    """Asset numbers from Gas_Monitor_Serial_Numbers.xlsx (all sheets)."""
    if not path or not os.path.exists(path):
        return set()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    assets = set()
    for s in wb.sheetnames:
        for r in wb[s].iter_rows(min_row=2, values_only=True):
            if r and r[0]:
                assets.add(str(r[0]).strip().upper())
    wb.close()
    return assets


def load_transactions(path, serial_assets=None):
    """Gas monitor transactions -> list of dicts + the report window."""
    serial_assets = serial_assets or set()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    window = ""
    if "REFERENCE_INFO" in wb.sheetnames:
        for r in wb["REFERENCE_INFO"].iter_rows(min_row=2, max_row=2,
                                                values_only=True):
            window = str(r[0] or "")

    ws = wb["CUSTOMER_CONTRACTOR_EQUIP"]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        desc = (str(r[2] or "") + " " + str(r[4] or "")).lower()
        bc = str(r[5] or "").strip().upper()
        if ("x-am" not in desc and "gas monitor" not in desc
                and bc not in serial_assets):
            continue
        st = _parse_dt(r[9], r[10])
        if st is None:
            continue
        out.append({
            "co": str(r[0] or "").strip(),
            "who": str(r[1] or "").strip(),
            "bc": bc,
            "st": st,
            "en": _parse_dt(r[11], r[12]),
            "internal": _is_internal(str(r[1] or ""), str(r[0] or "")),
        })
    wb.close()
    return out, window


def compute_tx(tx, now=None):
    """All flow/rhythm/behaviour aggregates off the transaction list."""
    now = now or max([t["st"] for t in tx] + [t["en"] for t in tx if t["en"]])
    crew = [t for t in tx if not t["internal"]]
    crew_closed = [t for t in crew if t["en"]]
    m = {"window_start": min(t["st"] for t in tx), "window_end": now,
         "n_all": len(tx), "n_crew": len(crew),
         "n_internal": len(tx) - len(crew),
         "n_crew_closed": len(crew_closed),
         "n_crew_open": len(crew) - len(crew_closed)}

    # ---- same-day rate, overall and weekly trend (crew) ---------------
    sd = sum(1 for t in crew_closed if t["en"].date() == t["st"].date())
    m["same_day"] = sd
    m["same_day_pct"] = round(sd / len(crew_closed) * 100, 1) if crew_closed else 0
    weekly = defaultdict(lambda: [0, 0])
    for t in crew_closed:
        iso = t["st"].isocalendar()
        weekly[(iso[0], iso[1])][0] += 1
        if t["en"].date() == t["st"].date():
            weekly[(iso[0], iso[1])][1] += 1
    m["weekly_sd"] = []
    for k in sorted(weekly):
        n, s = weekly[k]
        # Monday of that ISO week, for the axis label
        mon = datetime.fromisocalendar(k[0], k[1], 1)
        m["weekly_sd"].append({"week": k[1], "label": mon.strftime("%d %b"),
                               "n": n, "sd": s,
                               "pct": round(s / n * 100, 1) if n else 0})
    cur_iso = now.isocalendar()
    m["current_week_partial"] = (cur_iso[0], cur_iso[1]) in weekly

    # ---- duration buckets (crew, closed) ------------------------------
    buckets = Counter()
    for t in crew_closed:
        h = (t["en"] - t["st"]).total_seconds() / 3600
        buckets["Same shift (under 12h)" if h < 12 else
                "Overnight (12-24h)" if h < 24 else
                "1-3 days" if h < 72 else
                "3-7 days" if h < 168 else
                "Over a week"] += 1
    order = ["Same shift (under 12h)", "Overnight (12-24h)", "1-3 days",
             "3-7 days", "Over a week"]
    m["dur_buckets"] = [(k, buckets.get(k, 0)) for k in order]

    # ---- hourly rhythm (crew) -----------------------------------------
    m["hour_issues"] = Counter(t["st"].hour for t in crew)
    m["hour_returns"] = Counter(t["en"].hour for t in crew_closed)
    imins = sorted(t["st"].hour * 60 + t["st"].minute for t in crew)
    rmins = sorted(t["en"].hour * 60 + t["en"].minute for t in crew_closed)

    def hm(mm):
        return f"{mm // 60:02d}:{mm % 60:02d}"
    m["median_issue"] = hm(imins[len(imins) // 2]) if imins else "-"
    m["median_return"] = hm(rmins[len(rmins) // 2]) if rmins else "-"
    early = sum(1 for t in crew if t["st"].hour < 6)
    m["pct_before_6"] = round(early / len(crew) * 100) if crew else 0

    hh = Counter((t["st"].date(), t["st"].hour) for t in crew)
    (rd, rh), rn = hh.most_common(1)[0]
    m["record_hour"] = {"date": rd, "hour": rh, "n": rn,
                        "every_s": 3600 // rn if rn else 0}

    # ---- daily volumes and weekday profile (crew) ---------------------
    m["daily_issues"] = Counter(t["st"].date() for t in crew)
    m["weekday_issues"] = Counter(t["st"].strftime("%a") for t in crew)

    # ---- concurrency: peak out at once per day (ALL tx incl custody) --
    events = []
    for t in tx:
        events.append((t["st"], 1))
        events.append((t["en"] if t["en"] else now, -1))
    events.sort(key=lambda e: (e[0], -e[1]))
    cur = 0
    day_peak = {}
    for ts, d in events:
        cur += d
        dt = ts.date()
        if dt not in day_peak or cur > day_peak[dt][0]:
            day_peak[dt] = (cur, ts)
    m["day_peaks"] = [{"date": d, "peak": day_peak[d][0], "at": day_peak[d][1]}
                      for d in sorted(day_peak)]
    rec = max(m["day_peaks"], key=lambda p: p["peak"])
    m["record_peak"] = rec

    # ---- average intraday net-flow curve, last 10 working days --------
    # (crew + custody: cumulative same-day issues minus returns, so it is
    # exactly how much deeper into the shelf the day digs as it runs)
    # exclude the current (partial) day - a half-finished day would drag
    # the average curve down and understate the pressure
    wdays = [d for d in sorted({t["st"].date() for t in tx})
             if d.strftime("%a") not in ("Sat", "Sun")
             and d < now.date()][-10:]
    m["curve_days"] = wdays
    half_hours = [x / 2 for x in range(6, 41)]     # 03:00 .. 20:00
    curves = defaultdict(list)
    for d in wdays:
        ev = []
        for t in tx:
            if t["st"].date() == d:
                ev.append((t["st"].hour + t["st"].minute / 60, 1))
            if t["en"] and t["en"].date() == d:
                ev.append((t["en"].hour + t["en"].minute / 60, -1))
        ev.sort()
        i, cur = 0, 0
        for hh2 in half_hours:
            while i < len(ev) and ev[i][0] <= hh2:
                cur += ev[i][1]
                i += 1
            curves[hh2].append(cur)
    m["net_curve"] = [(hh2, round(sum(v) / len(v), 1))
                      for hh2, v in sorted(curves.items())] if wdays else []
    m["net_curve_worst"] = [(hh2, max(v)) for hh2, v in sorted(curves.items())] \
        if wdays else []
    m["net_plateau"] = max((v for _, v in m["net_curve"]), default=0)
    m["net_plateau_worst"] = max((v for _, v in m["net_curve_worst"]), default=0)

    # ---- people league (crew only) ------------------------------------
    ppl = defaultdict(lambda: {"n": 0, "sd": 0, "ov": 0, "open": 0,
                               "co": Counter(), "maxd": 0})
    for t in crew:
        p = ppl[t["who"]]
        p["n"] += 1
        p["co"][t["co"]] += 1
        if t["en"]:
            if t["en"].date() == t["st"].date():
                p["sd"] += 1
            else:
                p["ov"] += 1
            p["maxd"] = max(p["maxd"], (t["en"] - t["st"]).days)
        else:
            p["open"] += 1
            p["maxd"] = max(p["maxd"], (now - t["st"]).days)
    league = []
    for name, p in ppl.items():
        closed_n = p["sd"] + p["ov"]
        league.append({
            "name": name, "co": p["co"].most_common(1)[0][0],
            "draws": p["n"], "sd": p["sd"], "ov": p["ov"], "open": p["open"],
            "not_sd": p["ov"] + p["open"], "maxd": p["maxd"],
            "sd_pct": round(p["sd"] / closed_n * 100) if closed_n else 0})
    league.sort(key=lambda x: -x["not_sd"])
    m["league"] = league
    m["longest_kept"] = max(crew_closed, key=lambda t: t["en"] - t["st"]) \
        if crew_closed else None
    return m
