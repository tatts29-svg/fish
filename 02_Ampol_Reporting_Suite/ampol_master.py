#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ampol_master - the one editable workbook: descriptions, pricing, serials.
Author: Andrew Fisher | POWERED BY SITEIQ

WHY (03 Sep 2026, Andrew): the four small workbooks Andrew maintained
(Tooling_Description_Mapping, New_Descriptions___Andrew,
Ampol_ToolStore_Pricing, Gas_Monitor_Serial_Numbers, radio_register)
become ONE file, Data\\Editable\\Ampol_Master.xlsx, with a tab each:

    Descriptions   ITEM_DESCRIPTION | ITEM_BARCODE | Corrected Description
                   | Correction Notes | Changed?
    Pricing        ITEM_DESCRIPTION | Avg Buy Price (New) | 3 Years Old 65%
                   of cost | 5 Years Old 50% of Cost | 10+ Years Old 30% of
                   Cost | Source to Buy
    Gas serials    ASSET NUMBER GAS MONITOR | SERIAL NUMBER | DESCRIPTION
    Radio serials  Barcode | Serial Number | Description | Status
    Read me        what each tab feeds and what must not change

The headers are the legacy files' headers, unchanged, so every reader
keeps its column names. A reader asks locate(kind, legacy patterns) and
gets (path, sheet): the master's tab when the master exists, else the
legacy file (sheet None = the reader's own sheet rule). Nothing breaks
on a laptop that has not run button 16 yet.
"""
import os
import re
from collections import OrderedDict, defaultdict

import openpyxl

import ampol_paths

MASTER_NAME = "Ampol_Master.xlsx"
MASTER_PATTERNS = ("Ampol_Master*.xlsx",)
SHEETS = {"descriptions": "Descriptions", "pricing": "Pricing",
          "gas_serials": "Gas serials", "radio_serials": "Radio serials"}
HEADERS = {
    "descriptions": ["ITEM_DESCRIPTION", "ITEM_BARCODE", "Corrected Description", "Correction Notes", "Changed?"],
    "pricing": ["ITEM_DESCRIPTION", "Avg Buy Price (New)", "3 Years Old 65% of cost",
                "5 Years Old 50% of Cost", "10+ Years Old 30% of Cost", "Source to Buy"],
    "gas_serials": ["ASSET NUMBER GAS MONITOR", "SERIAL NUMBER", "DESCRIPTION"],
    "radio_serials": ["Barcode", "Serial Number", "Description", "Status"],
}
_PATH = None


def path():
    """The master workbook, or '' when the laptop still runs the legacy files."""
    global _PATH
    if _PATH is None:
        _PATH = ampol_paths.find_data(*MASTER_PATTERNS) or ""
    return _PATH


def locate(kind, *legacy_patterns):
    """(path, sheet) for one input. The master's tab when the master exists
    and carries that tab; else the newest legacy file matching the
    patterns with sheet None (the reader applies its own sheet rule);
    else ('', None)."""
    p = path()
    if p:
        try:
            wb = openpyxl.load_workbook(p, read_only=True)
            has = SHEETS[kind] in wb.sheetnames
            wb.close()
        except Exception:
            has = False
        if has:
            return p, SHEETS[kind]
    legacy = ampol_paths.find_data(*legacy_patterns) if legacy_patterns else ""
    return (legacy or ""), None


def describe(kind, p, sheet):
    """One console line: where an input was read from."""
    if not p:
        return "NOT FOUND"
    where = os.path.relpath(p, ampol_paths.suite_dir())
    return f"{where}  [{sheet}]" if sheet else where


# ---------------------------------------------------------------------------
# Building the master from the legacy files (button 16)
# ---------------------------------------------------------------------------
def _rows(path, sheet=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c is not None else "" for c in next(it)]
    out = [list(r) for r in it if r and any(c not in (None, "") for c in r)]
    wb.close()
    return hdr, out


def _norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().upper()


def build(out_path, mapping=None, new_desc=None, pricing=None, gas=None, radio=None):
    """Write the master from whichever legacy files exist. Values only,
    never a formula; every row of every source kept; nothing invented.
    Returns {tab: rows} for the console."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    counts = OrderedDict()
    try:
        from openpyxl.styles import Font, PatternFill
        hf, hfill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F2A37")
    except Exception:
        hf = hfill = None

    def sheet(name, headers, rows, widths):
        ws = wb.create_sheet(name)
        ws.append(headers)
        if hf:
            for c in ws[1]:
                c.font, c.fill = hf, hfill
        for r in rows:
            ws.append(list(r)[:len(headers)] + [""] * max(0, len(headers) - len(r)))
        for i, w in enumerate(widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w
        ws.freeze_panes = "A2"
        if rows:
            ws.auto_filter.ref = ws.dimensions
        counts[name] = len(rows)

    # Descriptions: every mapping row (barcode-keyed), with the notes and
    # Changed? flag from New_Descriptions joined by barcode; rows only in
    # New_Descriptions are appended. Where both name a barcode with a
    # different corrected text, New_Descriptions (Andrew's file) wins and
    # the mapping's text goes into Correction Notes - the console says how
    # many, so nothing is silent.
    if mapping or new_desc:
        H = HEADERS["descriptions"]
        rows = OrderedDict()
        notes = {}
        if mapping:
            mh, mr = _rows(mapping, next((s for s in ("Use this", "In use", "CorrectedDescriptionsTable")), None))
            mi = {h: i for i, h in enumerate(mh)}
            ci = next((mi[h] for h in ("Corrected Description", "CorrectedDescriptionsTable.Corrected Description") if h in mi), None)
            for r in mr:
                bc = _norm(r[mi["ITEM_BARCODE"]]) if "ITEM_BARCODE" in mi else ""
                if not bc:
                    continue
                rows[bc] = [r[mi.get("ITEM_DESCRIPTION", 0)], r[mi["ITEM_BARCODE"]],
                            r[ci] if ci is not None else "", "", ""]
        conflicts = 0
        if new_desc:
            nh, nr = _rows(new_desc)
            ni = {h: i for i, h in enumerate(nh)}
            for r in nr:
                bc = _norm(r[ni["ITEM_BARCODE"]]) if "ITEM_BARCODE" in ni else ""
                if not bc:
                    continue
                new = r[ni.get("Corrected Description", 2)]
                note = r[ni["Correction Notes"]] if "Correction Notes" in ni else ""
                chg = r[ni["Changed?"]] if "Changed?" in ni else ""
                if bc in rows:
                    old = rows[bc][2]
                    if _norm(old) != _norm(new) and _norm(new):
                        conflicts += 1
                        note = f"{note or ''} [mapping had: {old}]".strip()
                        rows[bc][2] = new
                    rows[bc][3], rows[bc][4] = note or "", chg or ""
                else:
                    rows[bc] = [r[ni.get("ITEM_DESCRIPTION", 0)], r[ni["ITEM_BARCODE"]], new, note or "", chg or ""]
        sheet(SHEETS["descriptions"], H, list(rows.values()), [52, 18, 44, 40, 10])
        counts["descriptions conflicts (Andrew's text kept)"] = conflicts
    if pricing:
        ph, pr = _rows(pricing, "RENTAL_STOCK")
        H = HEADERS["pricing"]
        pi = {h: i for i, h in enumerate(ph)}
        order = [pi.get(h) for h in H]
        rows = [[(r[i] if i is not None and i < len(r) else "") for i in order] for r in pr]
        sheet(SHEETS["pricing"], H, rows, [52, 20, 22, 22, 24, 40])
    if gas:
        H = HEADERS["gas_serials"]
        seen, rows = set(), []
        wb0 = openpyxl.load_workbook(gas, read_only=True, data_only=True)
        for s in wb0.sheetnames:
            for r in wb0[s].iter_rows(min_row=2, values_only=True):
                if r and r[0] and len(r) > 1 and r[1] not in (None, ""):
                    k = _norm(r[0])
                    if k in seen:
                        continue
                    seen.add(k)
                    rows.append([r[0], r[1], r[2] if len(r) > 2 else ""])
        wb0.close()
        sheet(SHEETS["gas_serials"], H, rows, [26, 16, 40])
    if radio:
        rh, rr = _rows(radio, "Radio Register")
        H = HEADERS["radio_serials"]
        ri = {h: i for i, h in enumerate(rh)}
        order = [ri.get(h) for h in H]
        rows = [[(r[i] if i is not None and i < len(r) else "") for i in order] for r in rr]
        sheet(SHEETS["radio_serials"], H, rows, [16, 16, 40, 14])
    rm = wb.create_sheet("Read me", 0)
    for line in [
        "AMPOL MASTER - descriptions, pricing and serial numbers in one editable workbook",
        "Author: Andrew Fisher | POWERED BY SITEIQ",
        "",
        "Descriptions  - barcode -> corrected description. Feeds the tooling reports (button 04) and the",
        "                stocktake count worklist (05). A row with Changed? = No is skipped by the stocktake.",
        "Pricing       - description -> Avg Buy Price (New) and Source to Buy. Feeds every value figure",
        "                (tooling, radio, stocktake, executive summary). Numbers only in the price column.",
        "Gas serials   - barcode -> serial for gas monitors. Printed in brackets after the monitor's name.",
        "Radio serials - barcode -> serial for radios. Printed in brackets after the radio's name.",
        "",
        "Do not rename the tabs or change a header. Add rows freely. Keep row 1 as the header.",
        "Never type a price as text; leave an unknown price blank. Docs\\EXCEL_FILES_WE_READ.txt has",
        "the full rules and Docs\\PRICING_FILE_DO_NOT_CHANGE.txt the one-page note for a reviewer.",
        "The SiteIQ exports live in Data\\SiteIQ and are never edited; this file lives in Data\\Editable.",
    ]:
        rm.append([line])
    rm.column_dimensions["A"].width = 110
    if hf:
        rm["A1"].font = Font(bold=True, size=13)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return counts


if __name__ == "__main__":
    p = path()
    print(f"Master workbook : {p or 'not found - the legacy files are read until button 16 builds it'}")
    for k in SHEETS:
        pp, sh = locate(k, "*.xlsx")
        print(f"  {k:14} -> {describe(k, pp, sh)}")
