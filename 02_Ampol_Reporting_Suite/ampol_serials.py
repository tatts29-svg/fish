#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ampol_serials - barcode -> serial number, for the names as shown.
Author: Andrew Fisher | POWERED BY SITEIQ

WHY (03 Sep 2026, Andrew): a gas monitor or a radio prints under one
name with its serial in brackets - "Dräger X-am 5000 Gas Monitor
(ARSN-0637)", "Motorola Radio (122TYX0140)". The serial comes from the
two serial workbooks in Data\ (Gas_Monitor_Serial_Numbers.xlsx, every
sheet, columns A and B; radio_register.xlsx, sheet "Radio Register",
columns A and B), read once per run. Display only - matching and
counting never use a serial. A barcode with no serial on either list
prints without brackets; nothing is guessed.
"""
import openpyxl
import ampol_paths

_CACHE = None


def _read(path, sheets=None):
    out = {}
    if not path:
        return out
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return out
    try:
        for s in (sheets or wb.sheetnames):
            if s not in wb.sheetnames:
                continue
            for r in wb[s].iter_rows(min_row=2, values_only=True):
                if r and r[0] and len(r) > 1 and r[1] is not None and str(r[1]).strip():
                    out.setdefault(str(r[0]).strip().upper(), str(r[1]).strip())
    finally:
        wb.close()
    return out


def load():
    """{BARCODE: serial} from both serial workbooks; the gas list first."""
    global _CACHE
    if _CACHE is None:
        m = _read(ampol_paths.find_data("Gas_Monitor_Serial*.xlsx", "*serial*.xlsx"))
        for k, v in _read(ampol_paths.find_data("radio_register*.xlsx", "*radio*register*.xlsx"),
                          ["Radio Register"]).items():
            m.setdefault(k, v)
        _CACHE = m
    return _CACHE


def serial_for(barcode):
    """The serial on the lists for this barcode, or ''."""
    if not barcode:
        return ""
    return load().get(str(barcode).strip().upper(), "")


if __name__ == "__main__":
    m = load()
    print(f"{len(m):,} barcodes carry a serial on the two lists")
    for k in list(m)[:5]:
        print(f"  {k} -> {m[k]}")
