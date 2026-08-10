#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | CHECK THIS LAPTOP - will the on-hire report run here?
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  This changes NOTHING. It looks at the computer and the folder and
#  answers one question in plain words: will the report run here, and if
#  not, which single thing has to be fixed.
#
#  It also writes LAPTOP_CHECK.txt beside itself. When somebody says
#  "it doesn't work on my laptop", that file is the whole answer - which
#  Python, which folder, which exports, what's locked, what's missing -
#  so nobody has to play twenty questions over the phone.
#
#  Deliberately plain: standard library only, no openpyxl needed to run
#  (it reports on openpyxl rather than depending on it), and every
#  message says what to DO, not what went wrong.
# =====================================================================

import os
import platform
import shutil
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
_LIB = os.path.join(HERE, "_lib")
if os.path.isdir(_LIB) and _LIB not in sys.path:
    sys.path.insert(0, _LIB)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

NEEDED = ("RENTAL_STOCK", "SALES_STOCK", "TRANSACTIONS")
OPTIONAL = ("STOCKTAKE",)

OK, WARN, STOP = "OK", "WARN", "STOP"
lines = []          # everything we print also goes in the txt file
results = []        # (status, title, detail, fix)


def out(text=""):
    lines.append(text)
    try:
        print(text)
    except (UnicodeEncodeError, UnicodeError):
        flat = text.replace("·", "-").replace("—", "-").replace("–", "-")
        print(flat.encode("ascii", "replace").decode("ascii"))


def check(status, title, detail="", fix=""):
    results.append((status, title, detail, fix))
    tag = {OK: "[ OK ]", WARN: "[WARN]", STOP: "[STOP]"}[status]
    out("  {} {}".format(tag, title))
    for bit in (detail or "").split("\n"):
        if bit:
            out("         {}".format(bit))
    if fix:
        for bit in fix.split("\n"):
            out("         FIX: {}".format(bit) if bit == fix.split("\n")[0]
                else "              {}".format(bit))


def human_age(path):
    """'today', 'yesterday', or 'N days ago' - and the date."""
    import datetime as dt
    stamp = dt.datetime.fromtimestamp(os.path.getmtime(path))
    days = (dt.datetime.now().date() - stamp.date()).days
    if days <= 0:
        when = "today"
    elif days == 1:
        when = "yesterday"
    else:
        when = "{} days ago".format(days)
    return "{} ({})".format(stamp.strftime("%d %b %Y %H:%M"), when), days


def cloud_only(path):
    """True if this is a OneDrive file that hasn't actually come down.

    It shows in Explorer with a name, a size and a date, so it looks
    completely normal - and then Excel and Python both fail on it with
    "cannot access the file". Worth naming, because the fix is one
    right-click and nobody guesses it.
    """
    if os.name != "nt":
        return False
    try:
        import ctypes
        attrs = ctypes.windll.kernel32.GetFileAttributesW(str(path))
        if attrs == -1:
            return False
        OFFLINE = 0x1000
        RECALL_ON_OPEN = 0x40000
        RECALL_ON_DATA_ACCESS = 0x400000
        return bool(attrs & (OFFLINE | RECALL_ON_OPEN | RECALL_ON_DATA_ACCESS))
    except Exception:
        return False


def newest(folder, stem):
    hits = []
    for name in os.listdir(folder) if os.path.isdir(folder) else []:
        if not name.lower().endswith((".xlsx", ".xlsm")):
            continue
        if name.startswith("~$"):
            continue
        if name.upper().startswith(stem):
            hits.append(os.path.join(folder, name))
    return max(hits, key=os.path.getmtime) if hits else None


def project_of(path):
    """The SiteIQ job an export came from, if openpyxl is available."""
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        if "REFERENCE_INFO" not in wb.sheetnames:
            return ""
        rows = list(wb["REFERENCE_INFO"].iter_rows(min_row=1, max_row=2,
                                                   values_only=True))
        if len(rows) < 2:
            return ""
        head = [str(c or "").strip().upper() for c in rows[0]]
        if "PROJECT SCHEDULE" not in head:
            return ""
        return str(rows[1][head.index("PROJECT SCHEDULE")] or "").strip()
    except Exception:
        return None
    finally:
        try:
            wb.close()
        except Exception:
            pass


def main():
    out("=" * 68)
    out(" COATES | WILL THE ON-HIRE REPORT RUN ON THIS LAPTOP?")
    out(" Author: Andrew Fisher | POWERED BY SITEIQ")
    out("=" * 68)
    out()
    out(" Computer : {}".format(platform.node() or "(no name)"))
    out(" Windows  : {} {}".format(platform.system(), platform.release()))
    out(" Folder   : {}".format(HERE))
    out()
    out(" ---- THE ENGINE " + "-" * 51)

    # ---- 1. Python -------------------------------------------------
    ver = ".".join(str(n) for n in sys.version_info[:3])
    where = sys.executable or "(unknown)"
    private = os.path.normcase(HERE) in os.path.normcase(where)
    if sys.version_info >= (3, 8):
        check(OK, "Python {} is here and new enough".format(ver),
              "Using: {}".format(where) +
              ("\nThis is the kit's own private copy - nothing was "
               "installed on the computer." if private else ""))
    else:
        check(STOP, "Python {} is too old".format(ver),
              "Using: {}\nThe kit needs 3.8 or newer.".format(where),
              "Double-click 3_GET_PYTHON_NO_ADMIN.bat")

    # ---- 2. the Excel library --------------------------------------
    if not os.path.isdir(_LIB):
        check(STOP, "The _lib folder is missing",
              "That folder is the Excel library. It ships inside the kit "
              "so nothing has to be installed.",
              "Right-click the zip, Extract All, and work in the folder "
              "that comes out.")
    else:
        try:
            import openpyxl
            src = os.path.dirname(os.path.dirname(
                os.path.abspath(openpyxl.__file__)))
            from_kit = os.path.normcase(src) == os.path.normcase(_LIB)
            check(OK, "Excel library found - openpyxl {}".format(
                openpyxl.__version__),
                ("Loaded from the kit's own _lib folder. No install, no "
                 "internet, no pip." if from_kit else
                 "Loaded from {} - the computer's own copy, which is "
                 "fine.".format(src)))
        except ImportError as e:
            check(STOP, "The Excel library won't load",
                  "_lib is there but Python can't import it ({}).".format(e),
                  "Extract the zip again - a file in _lib didn't make it.")

    # ---- 3. is the kit whole? --------------------------------------
    missing = [f for f in ("build_onhire_workbook.py", "JOB.txt", "_RUN.bat")
               if not os.path.isfile(os.path.join(HERE, f))]
    if missing:
        check(STOP, "Part of the kit is missing",
              "Not in this folder: " + ", ".join(missing),
              "Right-click the zip, Extract All, and use that folder.")
    else:
        check(OK, "All the kit's files are here")

    out()
    out(" ---- WHERE IT'S SITTING " + "-" * 43)

    # ---- 4. the folder itself --------------------------------------
    low = HERE.lower()
    if os.sep + "temp" + os.sep in low or low.endswith(os.sep + "temp"):
        check(STOP, "This folder is inside Windows' Temp folder",
              "That means the kit was run from inside the zip. Windows "
              "unpacks one file at a time into Temp and throws it away.",
              "Right-click the zip, Extract All, choose Desktop, and use "
              "the folder that appears.")
    elif "onedrive" in low:
        check(WARN, "This folder is inside OneDrive",
              "It will work, but OneDrive sometimes locks a file mid-sync "
              "and the report then can't write the workbook.",
              "If you ever get a 'can't write' message, move the folder to "
              "the Desktop and carry on.")
    else:
        check(OK, "The folder is in a sensible place")

    if len(HERE) > 200:
        check(WARN, "The path to this folder is very long ({} characters)"
              .format(len(HERE)),
              "Windows starts refusing long file names past about 260.",
              "Move the folder nearer the top - C:\\Coates\\ or the "
              "Desktop.")

    # ---- 5. can we write here? ------------------------------------
    probe = os.path.join(HERE, "_write_test.tmp")
    try:
        with open(probe, "w", encoding="utf-8") as f:
            f.write("test")
        os.remove(probe)
        check(OK, "This folder can be written to")
    except OSError as e:
        check(STOP, "Can't write into this folder",
              str(e),
              "Copy the whole folder to your Desktop and run it from "
              "there. Program Files and most network drives are "
              "read-only.")

    try:
        free = shutil.disk_usage(HERE).free / (1024.0 ** 3)
        if free < 0.2:
            check(STOP, "The drive is full - {:.2f} GB free".format(free),
                  "The workbook needs a few MB to write.",
                  "Clear some space and run it again.")
        else:
            check(OK, "Disk space is fine - {:.1f} GB free".format(free))
    except Exception:
        pass

    out()
    out(" ---- TODAY'S SITEIQ EXPORTS " + "-" * 39)

    # ---- 6. the exports -------------------------------------------
    data = os.path.join(HERE, "Data_SiteIQ")
    if not os.path.isdir(data):
        check(STOP, "There's no Data_SiteIQ folder",
              "That's where the three SiteIQ exports go.",
              "Extract the zip again - the folder didn't come across.")
    else:
        locked = [n for n in os.listdir(data) if n.startswith("~$")]
        found, jobs = {}, {}
        for stem in NEEDED + OPTIONAL:
            p = newest(data, stem)
            found[stem] = p
        gone = [s for s in NEEDED if not found[s]]
        if gone:
            check(STOP, "Missing export{}: {}".format(
                "" if len(gone) == 1 else "s", ", ".join(gone)),
                "Looked in: {}".format(data),
                "Pull them out of SiteIQ and save them into Data_SiteIQ. "
                "Don't rename them - SiteIQ's own long file names are "
                "fine.")
        for stem in NEEDED + OPTIONAL:
            p = found[stem]
            if not p:
                if stem in OPTIONAL:
                    check(OK, "{} - not here, and that's fine".format(stem),
                          "It only adds one line to the Cover tab.")
                continue
            stamp, days = human_age(p)
            note = "{}\n{}".format(os.path.basename(p), stamp)
            if cloud_only(p):
                check(STOP, "{} is in OneDrive but hasn't downloaded"
                      .format(stem), note,
                      "It looks like it's there, but it's still in the "
                      "cloud. Right-click it, choose 'Always keep on this "
                      "device', wait for the green tick, then try again.")
                continue
            if os.path.getsize(p) == 0:
                check(STOP, "{} is an empty file (0 bytes)".format(stem),
                      note,
                      "The download didn't finish. Pull it out of SiteIQ "
                      "again and save it over the top.")
                continue
            if days >= 3:
                check(WARN, "{} is {} days old".format(stem, days), note,
                      "Pull a fresh one out of SiteIQ, or the report will "
                      "show old numbers with today's date on them.")
            else:
                check(OK, "{} is here".format(stem), note)
            job = project_of(p)
            if job:
                jobs[stem] = job

        if locked:
            check(WARN, "Something in Data_SiteIQ is open in Excel",
                  "Lock files present: " + ", ".join(locked),
                  "Close Excel. The report can usually still read them, "
                  "but it's cleaner shut.")

        if len(set(jobs.values())) > 1:
            detail = "\n".join("{}: {}".format(k, v)
                               for k, v in sorted(jobs.items()))
            check(STOP, "These exports are from MORE THAN ONE JOB",
                  detail,
                  "Take the odd one out of Data_SiteIQ. One job's exports "
                  "at a time - mixing two is the one mistake that would "
                  "put wrong numbers in front of a customer.")
        elif jobs:
            check(OK, "All the exports are from the same job",
                  "Job: {}".format(sorted(set(jobs.values()))[0]))

    # ---- 7. is the workbook we're about to write open? ------------
    stale = [n for n in os.listdir(HERE)
             if n.startswith("~$") and n.lower().endswith((".xlsx", ".xlsm"))]
    if stale:
        check(WARN, "A workbook in this folder is open in Excel",
              "Lock files: " + ", ".join(stale),
              "Close it before you press 2_MAKE_THE_REPORT.bat, or the "
              "refresh can't overwrite it.")

    out()
    out(" ---- OPENING WHAT IT MAKES " + "-" * 40)

    # ---- 8. Outlook and Excel, for opening the output -------------
    if os.name == "nt":
        try:
            import winreg
            for ext, what, why in (
                    (".eml", "the email draft",
                     "Without it the .eml won't open. The .html version "
                     "next to it opens in any browser instead."),
                    (".xlsx", "the workbook",
                     "You only need Excel to READ the workbook. The "
                     "report is built without it.")):
                try:
                    with winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, ext) as k:
                        winreg.QueryValueEx(k, "")
                    check(OK, "This computer knows how to open {}".format(
                        what))
                except OSError:
                    check(WARN, "Nothing is set up to open {} ({})".format(
                        what, ext), why)
        except ImportError:
            pass
    else:
        check(OK, "Not Windows - skipping the Outlook and Excel checks",
              "The report itself builds fine on any operating system.")

    # ---- the verdict ----------------------------------------------
    stops = [r for r in results if r[0] == STOP]
    warns = [r for r in results if r[0] == WARN]
    out()
    out("=" * 68)
    if stops:
        out(" NOT READY YET - {} thing{} to fix".format(
            len(stops), "" if len(stops) == 1 else "s"))
        out("=" * 68)
        out()
        for n, (_st, title, _d, fix) in enumerate(stops, 1):
            out(" {}. {}".format(n, title))
            if fix:
                for bit in fix.split("\n"):
                    out("    -> {}".format(bit))
        out()
        out(" Fix those and press 1_CHECK_THIS_LAPTOP.bat again.")
    else:
        out(" READY - this laptop will build the report")
        out("=" * 68)
        out()
        out(" Next: double-click  2_MAKE_THE_REPORT.bat")
        if warns:
            out()
            out(" {} thing{} worth knowing but nothing that stops it:"
                .format(len(warns), "" if len(warns) == 1 else "s"))
            for _st, title, _d, _f in warns:
                out("   - {}".format(title))
    out()
    out(" A copy of all of this has been saved as LAPTOP_CHECK.txt.")
    out(" Still stuck? Email Andrew that file - it answers everything.")
    out("=" * 68)

    try:
        with open(os.path.join(HERE, "LAPTOP_CHECK.txt"), "w",
                  encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
    except OSError:
        #  If we can't write the file we've already said why, above.
        pass

    return 1 if stops else 0


if __name__ == "__main__":
    sys.exit(main())
