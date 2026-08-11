#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | ON-HIRE WORKBOOK - every tab, built from today's exports
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  The shutdown on-hire workbook, rebuilt from the SiteIQ exports for
#  whichever job this computer is running (57_SWITCH_JOB.bat).
#
#  Twelve tabs, and every one of them works:
#
#    Cover                    what this is, what fed it, the headlines
#    Company Summary          per company: tooling, consumables, total
#    Detailed Onhire          every item out right now, who has it
#    Tooling Transactions     every tooling issue and return
#    Tooling Utilisation      per item type: turns, rating, what to do
#    Coates Tooling           the Coates-owned fleet on site
#    Consumable Transactions  every consumable issued
#    Consumables Available    what's left on the shelf
#    Coates Stock             the fleet with its on-hire dates
#    Consumable Utilisation   per line: usage, rating, what to do
#    Coates Labour            the shift roster - CARRIED OVER, not touched
#    Cost Breakdown           the day-by-day cost - CARRIED OVER, not touched
#
#  Two of those tabs are Andrew's own typing - the labour roster and
#  the cost breakdown. They are read out of the existing workbook and
#  written back exactly as they were, formulas and all. Nothing typed
#  by hand is ever recalculated, overwritten or "tidied".
#
#  The existing workbook is never written to. It carries macros, a
#  logo and live query connections that Python cannot round-trip
#  safely, so it stays exactly as it is and a clean dated workbook is
#  built alongside it.
#
#  It also writes the report as an EMAIL - the whole thing in the body,
#  Coates styling, the workbook on the paperclip, and NO COSTS. Not the
#  labour, not the cost breakdown, not a rate anywhere. That is the
#  operational picture: what is out, who has it, what is working and
#  what is not. The money stays in the workbook. It opens as an Outlook
#  DRAFT and waits - nothing sends by itself.
#
#  Output:  <job reports>\<today>\<Job> On-Hire Workbook - <date>.xlsx
#      and  <job folder>\<Job>_Onhire_Workbook_LATEST.xlsx
#      and  <job reports>\<today>\<Job> On-Hire Report - <date>.eml
# =====================================================================

import datetime as dt
import glob
import os
import shutil
import sys
from copy import copy

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

#  The Excel library travels WITH the kit, in _lib. Nothing to install,
#  no internet, no pip, no admin rights, no company proxy to argue with.
#  That is the whole reason this folder runs on a laptop it has never
#  been on before: put _lib in front of everything else on the path and
#  openpyxl is simply there.
_LIB = os.path.join(HERE, "_lib")
if os.path.isdir(_LIB) and _LIB not in sys.path:
    sys.path.insert(0, _LIB)


#  A Windows console is often still on an old code page, and printing one
#  fancy character to it (the dot between Rio Tinto and Weipa, a dash, a
#  curly quote) kills the whole run with UnicodeEncodeError - a wall of
#  red text, no report, and nothing that tells you why. Two guards:
#  ask the console for UTF-8, and if it won't have it, fall back to
#  plain ASCII rather than falling over.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

_ASCII = {"·": "-", "–": "-", "—": "-", "’": "'",
          "‘": "'", "“": '"', "”": '"', "…": "...",
          "°": " deg", "×": "x", "→": "->", " ": " "}


#  Everything said on screen is kept, and written out at the end as
#  LAST_RUN_LOG.txt. A black window that has been closed is gone; a file
#  can be emailed. "Send me the log" beats "what did it say exactly?".
_LOG = []


def say(text=""):
    """print() that cannot take the run down. Anything the console
    can't draw gets swapped for the plain-keyboard version."""
    _LOG.append(text)
    try:
        print(text)
    except (UnicodeEncodeError, UnicodeError):
        flat = text
        for fancy, plain in _ASCII.items():
            flat = flat.replace(fancy, plain)
        print(flat.encode("ascii", "replace").decode("ascii"))


def write_log():
    """Best effort, always. If we can't write it, that is not worth
    failing a good report over."""
    try:
        with open(os.path.join(HERE, "LAST_RUN_LOG.txt"), "w",
                  encoding="utf-8") as f:
            f.write("COATES | ON-HIRE REPORT - what the last run said\n")
            f.write("Run at {}\n".format(
                dt.datetime.now().strftime("%d %b %Y %H:%M")))
            f.write("Folder: {}\n".format(HERE))
            f.write("Python: {} at {}\n".format(
                ".".join(str(n) for n in sys.version_info[:3]),
                sys.executable))
            f.write("-" * 66 + "\n")
            f.write("\n".join(_LOG) + "\n")
    except OSError:
        pass

#  This script runs in two homes and must work in both:
#
#    1. inside the full suite, where site_config knows which job the
#       computer is on and report_paths knows where reports go, and
#    2. on its own, in a folder with nothing but this file, the button
#       and the exports - the kit that goes on the other computers.
#
#  Standing alone it works the job out from the exports themselves.
#  Every SiteIQ export carries the PROJECT SCHEDULE it came from on its
#  REFERENCE_INFO sheet, so the kit names itself off the data and there
#  is nothing to set up. Drop the exports in, press the button, done.
try:
    import report_paths
    import site_config
    STANDALONE = False
except ImportError:
    report_paths = site_config = None
    STANDALONE = True

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    say()
    say("=" * 66)
    say(" THE EXCEL LIBRARY IS MISSING FROM THIS FOLDER")
    say("=" * 66)
    say()
    say(" This kit carries its own copy of it, in a folder called _lib.")
    say(" It isn't here, which means the zip was only half unpacked.")
    say()
    say(" Looked in: {}".format(_LIB))
    say()
    say(" FIX, and it is always this: right-click the zip, choose")
    say(" 'Extract All', and work in the folder that comes out. Opening")
    say(" the zip and double-clicking a button INSIDE it only unpacks")
    say(" that one file, so _lib never arrives.")
    say()
    sys.exit(1)

# ---- the Coates look -------------------------------------------------
ORANGE = "FFF26222"
ORANGE_ALT = "FFFA4600"        # the shade the existing workbook uses
NEAR_BLACK = "FF1D1D1B"
WHITE = "FFFFFFFF"
GREY = "FF8B9099"
BAND_ROW_H = 26

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=WHITE)
HDR_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
BODY_FONT = Font(name="Calibri", size=11)
TITLE_FILL = PatternFill("solid", fgColor=NEAR_BLACK)
HDR_FILL = PatternFill("solid", fgColor=ORANGE_ALT)
THIN = Side(style="thin", color="FFD9D9D9")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

#  The header row is boxed in its own colour so the heading block reads
#  as one solid bar rather than a row of separate cells.
HDR_SIDE = Side(style="thin", color=ORANGE_ALT)
HDR_BORDER = Border(left=HDR_SIDE, right=HDR_SIDE, top=HDR_SIDE,
                    bottom=HDR_SIDE)

#  Every second row gets the faintest warm tint. Enough to follow a long
#  row across twelve columns, not enough to notice you are being helped.
BAND_FILL = PatternFill("solid", fgColor="FFFDF6F2")

DATE_FMT = "dd/mm/yyyy"
TIME_FMT = "hh:mm"
PCT_FMT = "0.0%"
MONEY_FMT = "$#,##0.00"
INT_FMT = "#,##0"

#  Hand-typed tabs. Read out of the old workbook, written back as-is.
CARRIED_OVER = ("Coates Labour", "Cost Breakdown")

#  Windows won't have these in a file name, and SiteIQ project names
#  routinely carry a slash or a colon ("31/07/2026 05:00"). Left alone
#  they go straight into the workbook name and the save fails with
#  something nobody can act on.
BAD_IN_NAMES = '\\/:*?"<>|'


class KitProblem(Exception):
    """Something the person at the keyboard can actually fix.

    Raised instead of letting a traceback out. main() catches it, prints
    the message as-is and stops - so every foreseeable failure reads like
    an instruction rather than a crash.
    """


def _safe_name(text, fallback="Shutdown"):
    """A file name Windows will accept, out of whatever the job is
    called. Keeps the words, drops the punctuation that isn't allowed."""
    out = "".join(" " if ch in BAD_IN_NAMES else ch for ch in _s(text))
    out = " ".join(out.split()).strip(" .")
    return out or fallback


def latest_workbook_name(label):
    """The always-newest copy. One place, because the builder writes it
    and carry_over reads it, and if those two ever disagree the typed
    tabs get silently wiped."""
    return "{}_Onhire_Workbook_LATEST.xlsx".format(
        _safe_name(label).replace(" ", "_"))


# =====================================================================
#  Standing on its own - the kit that goes on the other computers
# =====================================================================
class LooseSite(object):
    """The job, worked out from the exports sitting in the folder.

    No config, no setup, nothing to switch. Whatever exports are here
    decide what the workbook is called and who it is for. Drop Weipa's
    in and it is a Weipa workbook; drop Gladstone's in and it is a
    Gladstone one. JOB.txt can override any of the names if the
    SiteIQ project string isn't what you'd put in front of a customer.
    """

    def __init__(self, base):
        self.base = base
        self._d = self._read_job_txt(base)
        if not self._d.get("project_schedule"):
            self._d["project_schedule"] = self._sniff(base)
        stamp = self._d["project_schedule"]
        self._d.setdefault("job", stamp or "Shutdown")
        self._d.setdefault("customer", "")
        self._d.setdefault("location", "")
        self._d.setdefault("short", (self._d.get("customer")
                                     or self._d["job"]).strip())
        self._d.setdefault("author", "Andrew Fisher")

    @staticmethod
    def _read_job_txt(base):
        """Optional. One 'name: value' per line - job, customer,
        location, short. Anything missing is worked out instead."""
        out = {}
        p = os.path.join(base, "JOB.txt")
        try:
            with open(p, "rb") as f:
                raw = f.read()
        except OSError:
            return out
        #  Read as bytes and decode carefully. Notepad saves as ANSI
        #  unless you tell it otherwise, so the moment somebody types a
        #  name with an accent or a fancy dash in it - Boral's "Peppertree
        #  – Marulan", pasted out of an email - a strict UTF-8 read throws
        #  UnicodeDecodeError. That is a ValueError, not an OSError, so it
        #  used to escape and take the whole run down on the first thing
        #  anybody edits. Never again: UTF-8, then Windows ANSI, then give
        #  up gracefully and work the job out from the exports instead.
        text = None
        for codec in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                text = raw.decode(codec)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        if text is None:
            return out
        for line in text.splitlines():
            line = line.strip()
            if not line or line.startswith("#") or ":" not in line:
                continue
            k, v = line.split(":", 1)
            if v.strip():
                out[k.strip().lower().replace(" ", "_")] = v.strip()
        return out

    @staticmethod
    def _sniff(base):
        """The PROJECT SCHEDULE stamped on whichever export is here."""
        for d in (os.path.join(base, "Data_SiteIQ"), base):
            for stem in ("RENTAL_STOCK", "TRANSACTIONS", "SALES_STOCK",
                         "STOCKTAKE"):
                for p in sorted(glob.glob(os.path.join(
                        glob.escape(d), stem + "*.xlsx"))):
                    if os.path.basename(p).startswith("~$"):
                        continue
                    got = read_project_schedule(p)
                    if got:
                        return got
        return ""

    def __getattr__(self, name):
        if name.startswith("_"):
            raise AttributeError(name)
        return self._d.get(name, "")

    @property
    def header_line(self):
        bits = [b for b in (self._d.get("customer"), self._d.get("job"),
                            self._d.get("location")) if b]
        return " · ".join(bits) or "Shutdown tool store"

    @property
    def exports(self):
        return ["RENTAL_STOCK", "SALES_STOCK", "TRANSACTIONS", "STOCKTAKE"]

    @property
    def reports_dirname(self):
        return "Reports"

    def data_dirs(self, base=None):
        base = base or self.base
        return [os.path.join(base, "Data_SiteIQ"), base]

    def workbook(self, base=None):
        """Any .xlsm sitting in the kit - that's where the hand-typed
        Labour and Cost tabs are read from."""
        base = base or self.base
        hits = [p for p in glob.glob(os.path.join(glob.escape(base),
                                                  "*.xlsm"))
                if not os.path.basename(p).startswith("~$")]
        if hits:
            return max(hits, key=os.path.getmtime)
        #  No .xlsm? Then the only place the hand-typed Labour and Cost
        #  tabs can be is the workbook this kit built last time. Read
        #  them back out of it, or today's rebuild wipes anything typed
        #  since yesterday - which is the opposite of what the tab says
        #  on it. (In the full suite Andrew keeps his .xlsm here and this
        #  never comes up; in the standalone kit there isn't one.)
        latest = os.path.join(base, latest_workbook_name(
            self._d.get("short") or self._d.get("customer") or "Shutdown"))
        return latest if os.path.isfile(latest) else None


def read_project_schedule(path):
    """The SiteIQ project an export came from, off its REFERENCE_INFO
    sheet. Empty if it doesn't carry one - plenty of older pulls don't."""
    if site_config is not None:
        return site_config.export_project_schedule(path)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return ""
    try:
        if "REFERENCE_INFO" not in wb.sheetnames:
            return ""
        rows = list(wb["REFERENCE_INFO"].iter_rows(min_row=1, max_row=2,
                                                   values_only=True))
        if len(rows) < 2:
            return ""
        head = [str(c or "").strip().upper() for c in rows[0]]
        if "PROJECT SCHEDULE" not in head:
            return ""
        return str(rows[1][head.index("PROJECT SCHEDULE")] or "").strip()
    except Exception:
        return ""
    finally:
        try:
            wb.close()
        except Exception:
            pass


def the_site(base):
    """The job, however this copy is running."""
    if site_config is not None:
        return site_config.site(base=base)
    return LooseSite(base)


def find_export(base, pattern, s):
    """Newest file matching the pattern, across the folders this job
    keeps its exports in."""
    if report_paths is not None:
        return report_paths.find_export(base, pattern)
    hits = []
    for d in s.data_dirs(base):
        hits += [p for p in glob.glob(os.path.join(glob.escape(d), pattern))
                 if not os.path.basename(p).startswith("~$")]
    return max(hits, key=os.path.getmtime) if hits else None


def out_dirs(base, when, s):
    if report_paths is not None:
        return report_paths.out_dirs(base, when)
    day = os.path.join(base, s.reports_dirname, when.strftime("%Y-%m-%d"))
    os.makedirs(day, exist_ok=True)
    return {"day": day}


def note_sources(folder, entries):
    if report_paths is not None:
        return report_paths.note_sources(folder, entries)
    lines = ["DATA SOURCES FOR THIS WORKBOOK",
             "Generated " + dt.datetime.now().strftime("%d %b %Y - %H:%M"),
             "-" * 60]
    for p in entries:
        stamp = dt.datetime.fromtimestamp(os.path.getmtime(p))
        lines.append("{:<52} refreshed {}".format(
            os.path.basename(p), stamp.strftime("%d %b %Y - %H:%M")))
    out = os.path.join(folder, "DATA_SOURCES.txt")
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    return out


def wrong_job(path, base, s):
    """(ok, message). Standing alone there is no 'live job' to check
    against, so the test is that the exports AGREE WITH EACH OTHER -
    which catches the real mistake, refreshing off a mix of two jobs."""
    if site_config is not None:
        return site_config.belongs_to_live_job(path, base)
    want = (s.project_schedule or "").strip()
    got = read_project_schedule(path)
    if not want or not got or got.lower() == want.lower():
        return True, ""
    return False, ("{} is a {} export but the others in this folder are {}. "
                   "One job's exports at a time - take the odd one out and "
                   "run it again.".format(os.path.basename(path), got, want))


# =====================================================================
#  Reading the exports
# =====================================================================
def _rows(path, sheet, must_have=False):
    """A sheet as a list of dicts, keyed on its header row. Blank rows
    dropped - SiteIQ pads its exports out with them.

    must_have says this sheet is the whole point of the file. Without it
    a missing sheet is indistinguishable from an empty one, and the kit
    builds a complete, confident, all-zero report and puts an email draft
    in front of you ready to send to the customer. That is the worst
    thing this script could possibly do, so it stops instead.
    """
    if not path or not os.path.isfile(path):
        return []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        #  Half-downloaded from SiteIQ, saved as .xls or CSV with an
        #  .xlsx name, quarantined by antivirus, or a OneDrive placeholder
        #  that has not actually come down yet. All of them look like a
        #  file sitting there in Explorer.
        raise KitProblem(
            "This export can't be opened:\n"
            "         {}\n\n"
            "         {}\n\n"
            "         It is almost always a part-finished download. Pull "
            "it out of\n"
            "         SiteIQ again, save it over the top, and run the "
            "button again.".format(path, e))
    try:
        if sheet not in wb.sheetnames:
            if must_have:
                raise KitProblem(
                    "{} hasn't got a '{}' sheet in it.\n\n"
                    "         The sheets it does have: {}\n\n"
                    "         That usually means it is the wrong export, "
                    "or it was saved\n"
                    "         out of Excel rather than pulled from SiteIQ. "
                    "Pull a fresh\n"
                    "         one and don't open or re-save it on the way "
                    "through.\n\n"
                    "         Stopping here on purpose: with that sheet "
                    "missing the report\n"
                    "         would come out full of zeros and look "
                    "perfectly fine.".format(
                        os.path.basename(path), sheet,
                        ", ".join(wb.sheetnames) or "(none)"))
            return []
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        try:
            head = next(it)
        except StopIteration:
            return []
        hdr = [str(c or "").strip() for c in head]
        out = []
        for r in it:
            if not any(c not in (None, "") for c in r):
                continue
            out.append(dict(zip(hdr, r)))
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _s(v):
    """A cell as clean text."""
    return "" if v is None else str(v).strip()


def _n(v):
    """A cell as a number. SiteIQ ships numbers as text more often than
    not, so never trust the type."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except ValueError:
        return 0.0


def _date(v):
    """A cell as a date. SiteIQ mixes real dates with dd/mm/yyyy text
    inside the same column."""
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    t = _s(v)
    if not t:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S",
                "%d/%m/%Y %I:%M %p", "%Y-%m-%d %H:%M:%S"):
        try:
            return dt.datetime.strptime(t, fmt).date()
        except ValueError:
            continue
    return None


# =====================================================================
#  Usage ratings - the rules, written down
# =====================================================================
#  Tooling is rated on TURNS: how many times each unit went out.
#  One item that went out five times is working; five items that went
#  out once each are sitting there. Utilisation on its own can't tell
#  those apart, which is why the old sheet had a line rated "Low Use"
#  on 36 transactions and another rated "Good Use" on one.
def tooling_rating(transactions, total_qty, utilisation):
    turns = (transactions / total_qty) if total_qty else 0.0
    if turns <= 0:
        return "No Use", "Review / Reduce"
    if turns < 1:
        return "Low Use", "Keep Stock"
    if turns < 2:
        if utilisation >= 0.8:
            return "Good Use", "Monitor / Increase"
        return "Good Use", "Keep Stock"
    return "High Demand", "Increase Stock"


#  Consumables are rated on how much of the position has gone out.
def consumable_rating(usage):
    if usage <= 0:
        return "No Use", "Review / Reduce"
    if usage < 0.4:
        return "Low Use", "Keep Stock"
    if usage < 0.8:
        return "Good Use", "Keep Stock"
    return "High Demand", "Increase Stock"


# =====================================================================
#  Sheet writing
# =====================================================================
def band(ws, title, ncols):
    """The orange-on-black title strip across the top of every tab."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ncols, 1))
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = BAND_ROW_H


def sheet(wb, name, title, columns, rows, widths=None, formats=None):
    """One data tab, boxed up properly: title strip, header row, the
    rows, frozen panes, a filter, and print settings that put it on a
    page the right way round.

    `formats` maps column index (0-based) to a number format, and is
    also what decides which columns get right-aligned - a number that
    reads as a number belongs on the right, and a column of them lines
    up on the decimal instead of wandering.
    """
    ws = wb.create_sheet(name)
    band(ws, title, len(columns))

    for i, col in enumerate(columns, start=1):
        c = ws.cell(row=2, column=i, value=col)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.border = HDR_BORDER
        c.alignment = Alignment(vertical="center", horizontal="center",
                                wrap_text=True)
    ws.row_dimensions[2].height = 30

    #  Which columns are numeric, so they can be right-aligned. Dates and
    #  times read better centred than shoved to either edge.
    fmts = formats or {}
    right = {i for i, f in fmts.items()
             if f in (INT_FMT, MONEY_FMT, PCT_FMT, "#,##0.00")}
    middle = {i for i, f in fmts.items() if f in (DATE_FMT, TIME_FMT)}

    for r, row in enumerate(rows, start=3):
        stripe = BAND_FILL if (r % 2) else None
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = BODY_FONT
            c.border = CELL_BORDER
            if stripe is not None:
                c.fill = stripe
            if (i - 1) in right:
                c.alignment = Alignment(horizontal="right",
                                        vertical="center")
            elif (i - 1) in middle:
                c.alignment = Alignment(horizontal="center",
                                        vertical="center")
            else:
                c.alignment = Alignment(vertical="center", wrap_text=False)
            fmt = fmts.get(i - 1)
            if fmt:
                c.number_format = fmt

    for i, col in enumerate(columns, start=1):
        letter = get_column_letter(i)
        if widths and (i - 1) in widths:
            ws.column_dimensions[letter].width = widths[i - 1]
        else:
            longest = max([len(str(col))]
                          + [len(_s(r[i - 1])) for r in rows[:400]] or [10])
            ws.column_dimensions[letter].width = min(max(longest + 2, 11), 60)

    ws.freeze_panes = "A3"
    if rows:
        ws.auto_filter.ref = "A2:{}{}".format(
            get_column_letter(len(columns)), len(rows) + 2)
    #  An empty tab still says so out loud rather than looking broken.
    if not rows:
        c = ws.cell(row=3, column=1,
                    value="Nothing to show for this pull - no rows in the "
                          "export matched this tab.")
        c.font = Font(name="Calibri", size=11, italic=True, color=GREY)

    tidy_print(ws, len(columns))
    ws.sheet_properties.tabColor = ORANGE
    return ws


def tidy_print(ws, ncols, landscape=True):
    """Make a tab print like a report instead of like a spreadsheet
    somebody hit Ctrl+P on. Sideways, squeezed to one page wide, with
    the title strip and the headings repeated at the top of every page
    and the page number down the bottom."""
    ws.page_setup.orientation = ("landscape" if landscape else "portrait")
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:2"
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    ws.oddFooter.left.text = "COATES  |  POWERED BY SITEIQ"
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = "808080"
    ws.oddFooter.right.text = "Page &P of &N"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.color = "808080"


# =====================================================================
#  The tabs
# =====================================================================
def build(base=HERE, when=None):
    s = the_site(base)
    when = when or dt.date.today()
    asat = dt.datetime.now()

    def export(stem):
        return find_export(base, stem + "*.xlsx", s)

    p_rental = export("RENTAL_STOCK")
    p_sales = export("SALES_STOCK")
    p_trans = export("TRANSACTIONS")
    p_stock = export("STOCKTAKE")
    sources = [p for p in (p_rental, p_sales, p_trans, p_stock) if p]

    missing = [n for n, p in (("RENTAL_STOCK", p_rental),
                              ("SALES_STOCK", p_sales),
                              ("TRANSACTIONS", p_trans)) if not p]
    if missing:
        say("  STOP | These exports are not in {}:".format(
            os.path.relpath(s.data_dirs(base)[0], base)))
        for m in missing:
            say("       - " + m)
        say("       Pull them out of SiteIQ, save them there, run again.")
        return 1

    #  Right files, right job?
    for p in sources:
        ok, why = wrong_job(p, base, s)
        if not ok:
            say("  STOP | " + why)
            return 1

    rental = _rows(p_rental, "RENTAL_STOCK", must_have=True)
    sales = _rows(p_sales, "SALES_STOCK", must_have=True)
    stocktake = _rows(p_stock, "STOCKTAKE") if p_stock else []

    #  SiteIQ splits transactions across two sheets and which one a job
    #  fills depends on whether charging is switched on. Weipa's land in
    #  TRANSACTION_WITHOUT_CHARGES, K2's in TRANSACTION_CHARGES. Read
    #  both and de-duplicate on the transaction number, so the tabs fill
    #  either way instead of coming out empty on one job.
    tw = []
    seen = set()
    for sheet_name in ("TRANSACTION_CHARGES", "TRANSACTION_WITHOUT_CHARGES"):
        for r in _rows(p_trans, sheet_name):
            tid = _s(r.get("TRANSACTION_ID"))
            if tid and tid in seen:
                continue
            if tid:
                seen.add(tid)
            tw.append(r)

    #  Neither transaction sheet had anything in it. Either the export is
    #  the wrong one, or it was pulled for a date range with no movements.
    #  Either way, every transaction-driven tab and half the email would
    #  come out empty, so say so rather than send a hollow report.
    if not tw:
        raise KitProblem(
            "There are no transactions in {}.\n\n"
            "         Looked on both sheets SiteIQ uses - "
            "TRANSACTION_CHARGES and\n"
            "         TRANSACTION_WITHOUT_CHARGES - and both came back "
            "empty.\n\n"
            "         Usually the export was pulled for the wrong date "
            "range. Pull\n"
            "         it again covering the shutdown, save it over the "
            "top, and run\n"
            "         the button again.".format(os.path.basename(p_trans)))

    title = "Coates    |  {}  |  SHUTDOWN TOOLSTORE ON-HIRE REPORT" \
            "        Powered by SITEIQ".format(s.short or s.customer)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---- the pieces every tab is cut from -----------------------------
    onhire = [r for r in rental if _s(r.get("ITEM_STATUS")) == "On Hire"]
    coates_owned = [r for r in rental
                    if "COATES" in _s(r.get("OWNER")).upper()]
    tooling_tx = [r for r in tw
                  if _s(r.get("PRODUCT_CATEGORY")).lower() != "consumable"]
    consumable_tx = [r for r in sales if _n(r.get("SALES_QUANTITY")) > 0]
    consumable_stock = [r for r in sales if not _s(r.get("COMPANY_NAME"))]

    tabs = []                      # (name, rows) for the cover index

    # ---- 1. Company Summary -------------------------------------------
    companies = {}
    for r in tw:
        co = _s(r.get("EMPLOYER_NAME")) or "(not named)"
        kind = ("CONSUMABLES"
                if _s(r.get("PRODUCT_CATEGORY")).lower() == "consumable"
                else "TOOLING")
        d = companies.setdefault(co, {
            "TOOLING": {"tx": 0, "qty": 0.0, "types": set(), "open": 0},
            "CONSUMABLES": {"tx": 0, "qty": 0.0, "types": set(), "open": 0}})
        a = d[kind]
        a["tx"] += 1
        a["qty"] += _n(r.get("QUANTITY"))
        a["types"].add(_s(r.get("SKU/ITEM DESCRIPTION")))
        if kind == "TOOLING" and not _s(r.get("TRAN_END_DATE")):
            a["open"] += 1

    live_onhire = {}
    for r in onhire:
        live_onhire[_s(r.get("COMPANY_NAME"))] = \
            live_onhire.get(_s(r.get("COMPANY_NAME")), 0) + 1

    summary_rows = []
    for co in sorted(companies):
        d = companies[co]
        tot = {"tx": 0, "qty": 0.0, "types": 0, "onh": 0, "open": 0}
        for kind in ("TOOLING", "CONSUMABLES"):
            a = d[kind]
            #  Consumables are sold, not returned - nothing is ever
            #  "still out", so those two columns are zero by definition.
            onh = live_onhire.get(co, 0) if kind == "TOOLING" else 0
            opn = a["open"] if kind == "TOOLING" else 0
            summary_rows.append([co, kind, a["tx"], int(a["qty"]),
                                 len([t for t in a["types"] if t]), onh, opn])
            tot["tx"] += a["tx"]
            tot["qty"] += a["qty"]
            tot["types"] += len([t for t in a["types"] if t])
            tot["onh"] += onh
            tot["open"] += opn
        summary_rows.append([co, "TOTAL", tot["tx"], int(tot["qty"]),
                             tot["types"], tot["onh"], tot["open"]])

    sheet(wb, "Company Summary", title,
          ["Company", "Type", "Transactions", "Qty", "Item Types",
           "Currently Onhire Qty", "Items Not Returned"],
          summary_rows,
          widths={0: 26, 1: 15, 2: 14, 3: 10, 4: 13, 5: 21, 6: 20},
          formats={2: INT_FMT, 3: INT_FMT, 4: INT_FMT, 5: INT_FMT, 6: INT_FMT})
    tabs.append(("Company Summary", len(summary_rows),
                 "Per company: what they took, what they still have"))

    # ---- 2. Detailed Onhire -------------------------------------------
    det = sorted(([_s(r.get("COMPANY_NAME")), _s(r.get("HIRER_NAME")),
                   _s(r.get("ITEM_NUMBER")), _s(r.get("ITEM_DESCRIPTION")),
                   _s(r.get("ITEM_STATUS")), _date(r.get("ON_HIRE_DATE")),
                   _s(r.get("ON_HIRE_TIME"))] for r in onhire),
                 key=lambda x: (x[0], x[1], x[3]))
    sheet(wb, "Detailed Onhire", title,
          ["COMPANY_NAME", "HIRER_NAME", "ITEM_NUMBER", "ITEM_DESCRIPTION",
           "ITEM_STATUS", "ON_HIRE_DATE", "ON_HIRE_TIME"],
          det, widths={0: 24, 1: 22, 2: 26, 3: 50, 4: 18, 5: 14, 6: 14},
          formats={5: DATE_FMT})
    tabs.append(("Detailed Onhire", len(det),
                 "Every item out right now and who has it"))

    # ---- 3. Tooling Transactions --------------------------------------
    tt = sorted(([_s(r.get("EMPLOYER_NAME")), _s(r.get("HIRER_NAME")),
                  _s(r.get("SKU/ITEM_NUMBER")),
                  _s(r.get("SKU/ITEM DESCRIPTION")),
                  _date(r.get("TRAN_START_DATE")), _s(r.get("TRAN_START_TIME")),
                  _date(r.get("TRAN_END_DATE")), _s(r.get("TRAN_END_TIME")),
                  _s(r.get("TRANSACTION_ID"))] for r in tooling_tx),
                 key=lambda x: (x[4] or dt.date.min, x[0]))
    sheet(wb, "Tooling Transactions", title,
          ["EMPLOYER_NAME", "HIRER_NAME", "SKU/ITEM_NUMBER",
           "SKU/ITEM DESCRIPTION", "TRAN_START_DATE", "TRAN_START_TIME",
           "TRAN_END_DATE", "TRAN_END_TIME", "TRANSACTION_ID"],
          tt, widths={0: 24, 1: 22, 2: 26, 3: 50, 4: 15, 5: 14, 6: 15,
                      7: 14, 8: 16},
          formats={4: DATE_FMT, 6: DATE_FMT})
    tabs.append(("Tooling Transactions", len(tt),
                 "Every tooling issue and return this period"))

    # ---- 4. Tooling Utilisation ---------------------------------------
    fleet = {}
    for r in rental:
        d = fleet.setdefault(_s(r.get("ITEM_DESCRIPTION")),
                             {"total": 0, "onhire": 0, "avail": 0})
        d["total"] += 1
        st = _s(r.get("ITEM_STATUS"))
        if st == "On Hire":
            d["onhire"] += 1
        elif st == "Available for Hire":
            d["avail"] += 1

    used = {}
    for r in tooling_tx:
        d = used.setdefault(_s(r.get("SKU/ITEM DESCRIPTION")),
                            {"tx": 0, "qty": 0.0, "co": set(), "hi": set(),
                             "first": None, "last": None})
        d["tx"] += 1
        d["qty"] += _n(r.get("QUANTITY"))
        if _s(r.get("EMPLOYER_NAME")):
            d["co"].add(_s(r.get("EMPLOYER_NAME")))
        if _s(r.get("HIRER_NAME")):
            d["hi"].add(_s(r.get("HIRER_NAME")))
        day = _date(r.get("TRAN_START_DATE"))
        if day:
            d["first"] = day if d["first"] is None else min(d["first"], day)
            d["last"] = day if d["last"] is None else max(d["last"], day)

    tu = []
    for desc in sorted(fleet):
        f = fleet[desc]
        u = used.get(desc, {"tx": 0, "qty": 0.0, "co": set(), "hi": set(),
                            "first": None, "last": None})
        util = (f["onhire"] / f["total"]) if f["total"] else 0.0
        rating, rec = tooling_rating(u["tx"], f["total"], util)
        tu.append([desc, f["total"], f["onhire"], f["avail"], util,
                   u["tx"], int(u["qty"]), len(u["co"]), len(u["hi"]),
                   u["first"], u["last"], rating, rec])
    sheet(wb, "Tooling Utilisation", title,
          ["ITEM_DESCRIPTION", "Total Qty", "Qty On Hire", "Qty Available",
           "Current Utilisation %", "Total Transactions", "Total Qty Issued",
           "No. of Companies Using Item", "No. of Hirers Using Item",
           "First Used Date", "Last Used Date", "Usage Rating",
           "Recommendation"],
          tu, widths={0: 55, 1: 11, 2: 12, 3: 13, 4: 15, 5: 14, 6: 14,
                      7: 16, 8: 16, 9: 14, 10: 14, 11: 14, 12: 20},
          formats={1: INT_FMT, 2: INT_FMT, 3: INT_FMT, 4: PCT_FMT,
                   5: INT_FMT, 6: INT_FMT, 7: INT_FMT, 8: INT_FMT,
                   9: DATE_FMT, 10: DATE_FMT})
    tabs.append(("Tooling Utilisation", len(tu),
                 "Per item type: turns, rating, what to do about it"))

    # ---- 5. Coates Tooling --------------------------------------------
    ct = sorted(([_s(r.get("ITEM_NUMBER")), _s(r.get("ITEM_DESCRIPTION")),
                  _s(r.get("ITEM_STATUS"))] for r in coates_owned),
                key=lambda x: (x[1], x[0]))
    sheet(wb, "Coates Tooling", title,
          ["ITEM_NUMBER", "ITEM_DESCRIPTION", "ITEM_STATUS"],
          ct, widths={0: 26, 1: 60, 2: 20})
    tabs.append(("Coates Tooling", len(ct),
                 "The Coates-owned fleet on this site"))

    # ---- 6. Consumable Transactions -----------------------------------
    ctx = sorted(([_s(r.get("COMPANY_NAME")), _s(r.get("HIRER")),
                   _s(r.get("SKU_DESCRIPTION")), _n(r.get("SALES_QUANTITY")),
                   _date(r.get("SALES_DATE")), _s(r.get("SALES_TIME"))]
                  for r in consumable_tx),
                 key=lambda x: (x[0], x[2]))
    sheet(wb, "Consumable Transactions", title,
          ["COMPANY_NAME", "HIRER", "SKU_DESCRIPTION", "SALES_QUANTITY",
           "SALES_DATE", "SALES_TIME"],
          ctx, widths={0: 24, 1: 24, 2: 50, 3: 16, 4: 14, 5: 14},
          formats={3: "#,##0.00", 4: DATE_FMT})
    tabs.append(("Consumable Transactions", len(ctx),
                 "Every consumable issued this period"))

    # ---- 7. Consumables Available -------------------------------------
    ca = sorted(([_s(r.get("SKU_NUMBER")), _s(r.get("SKU_DESCRIPTION")),
                  int(_n(r.get("AVAILABLE_QUANTITY")))]
                 for r in consumable_stock), key=lambda x: x[1])
    sheet(wb, "Consumables Available", title,
          ["SKU_NUMBER", "SKU_DESCRIPTION", "AVAILABLE_QUANTITY"],
          ca, widths={0: 20, 1: 50, 2: 22}, formats={2: INT_FMT})
    tabs.append(("Consumables Available", len(ca),
                 "What's left on the shelf"))

    # ---- 8. Coates Stock ----------------------------------------------
    cs = sorted(([_s(r.get("ITEM_NUMBER")), _s(r.get("ITEM_DESCRIPTION")),
                  _s(r.get("ITEM_STATUS")), _date(r.get("ON_HIRE_DATE")),
                  _s(r.get("ON_HIRE_TIME"))] for r in coates_owned),
                key=lambda x: (x[1], x[0]))
    sheet(wb, "Coates Stock", title,
          ["ITEM_NUMBER", "ITEM_DESCRIPTION", "ITEM_STATUS", "ON_HIRE_DATE",
           "ON_HIRE_TIME"],
          cs, widths={0: 26, 1: 60, 2: 20, 3: 15, 4: 14},
          formats={3: DATE_FMT})
    tabs.append(("Coates Stock", len(cs),
                 "The fleet with its on-hire dates"))

    # ---- 9. Consumable Utilisation ------------------------------------
    shelf, sold = {}, {}
    for r in sales:
        desc = _s(r.get("SKU_DESCRIPTION"))
        if not _s(r.get("COMPANY_NAME")):
            shelf[desc] = shelf.get(desc, 0.0) + _n(r.get("AVAILABLE_QUANTITY"))
        q = _n(r.get("SALES_QUANTITY"))
        if q > 0:
            d = sold.setdefault(desc, {"qty": 0.0, "co": set(), "hi": set(),
                                       "first": None, "last": None})
            d["qty"] += q
            if _s(r.get("COMPANY_NAME")):
                d["co"].add(_s(r.get("COMPANY_NAME")))
            if _s(r.get("HIRER")):
                d["hi"].add(_s(r.get("HIRER")))
            day = _date(r.get("SALES_DATE"))
            if day:
                d["first"] = day if d["first"] is None else min(d["first"], day)
                d["last"] = day if d["last"] is None else max(d["last"], day)

    cu = []
    for desc in sorted(set(list(shelf) + list(sold))):
        avail = shelf.get(desc, 0.0)
        d = sold.get(desc, {"qty": 0.0, "co": set(), "hi": set(),
                            "first": None, "last": None})
        position = avail + d["qty"]
        usage = (d["qty"] / position) if position else 0.0
        rating, rec = consumable_rating(usage)
        cu.append([desc, int(avail), int(d["qty"]), int(position), usage,
                   len(d["co"]), len(d["hi"]), d["first"], d["last"],
                   rating, rec])
    sheet(wb, "Consumable Utilisation", title,
          ["SKU_DESCRIPTION", "Qty Available", "Total Sales Qty",
           "Total Stock Position", "Usage %", "No. of Companies Using Item",
           "No. of Hirers Using Item", "First Sold Date", "Last Sold Date",
           "Usage Rating", "Recommendation"],
          cu, widths={0: 50, 1: 14, 2: 16, 3: 19, 4: 12, 5: 16, 6: 16,
                      7: 14, 8: 14, 9: 14, 10: 20},
          formats={1: INT_FMT, 2: INT_FMT, 3: INT_FMT, 4: PCT_FMT,
                   5: INT_FMT, 6: INT_FMT, 7: DATE_FMT, 8: DATE_FMT})
    tabs.append(("Consumable Utilisation", len(cu),
                 "Per line: usage, rating, what to do about it"))

    # ---- 10/11. the two hand-typed tabs, carried over verbatim --------
    carried, labour_total, cost_total = carry_over(wb, s, base, title)
    for name, n in carried:
        tabs.append((name, n, "Typed by hand - carried over untouched"))

    # ---- 12. Cover ----------------------------------------------------
    cover(wb, s, when, asat, tabs, sources, base,
          headline={
              "Items on hire now": len(onhire),
              "Companies with gear out": len([c for c in live_onhire
                                              if live_onhire[c]]),
              "Coates items on site": len(coates_owned),
              "Tooling transactions": len(tooling_tx),
              "Consumable issues": len(ctx),
              "Consumable lines on shelf": len(ca),
              "Items counted at last stocktake": len(stocktake),
              "Labour cost to date": labour_total,
              "Cost breakdown total": cost_total,
          })
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(wb["Cover"])))

    # ---- save ---------------------------------------------------------
    try:
        dirs = out_dirs(base, when, s)
    except OSError as e:
        raise KitProblem(
            "Couldn't make today's folder under Reports.\n\n"
            "         {}\n\n"
            "         The folder is read-only, or it's a OneDrive one that "
            "isn't\n"
            "         signed in. Copy the whole kit to your Desktop and run "
            "it\n"
            "         from there.".format(e))
    label = _safe_name(s.short or s.customer or "Shutdown")
    stem = "{} On-Hire Workbook - {}".format(label, when.strftime("%d %b %Y"))
    out = os.path.join(dirs["day"], stem + ".xlsx")
    try:
        wb.save(out)
    except PermissionError:
        #  Nine times in ten this is yesterday's workbook still open in
        #  Excel. Say that, instead of thirty lines of red traceback.
        say()
        say("  STOP | Windows won't let me write the workbook:")
        say("         {}".format(out))
        say()
        say("         That file is open somewhere - almost always in Excel,")
        say("         sometimes only in the Preview pane of Explorer.")
        say()
        say("         Close it and press the button again. Nothing is lost.")
        say()
        return 1
    except OSError as e:
        say()
        say("  STOP | Couldn't write the workbook: {}".format(e))
        say("         {}".format(out))
        say()
        say("         Usual causes: the folder is read-only, the drive is")
        say("         full, or the kit is sitting somewhere you don't have")
        say("         permission to write - a network drive, or Program")
        say("         Files. Copy the whole folder to your Desktop and try")
        say("         again.")
        say()
        return 1

    #  The one to open. Same name every day, always the newest - so the
    #  shortcut on the desktop never goes stale.
    latest_dir = os.path.join(base, os.path.dirname(s.workbook_glob or ""))
    latest = os.path.join(latest_dir, latest_workbook_name(label))
    try:
        shutil.copyfile(out, latest)
    except OSError as e:
        say("  WARN | Couldn't refresh the LATEST copy ({}). The dated "
              "one above is fine - it's probably open in Excel.".format(e))
        latest = None

    #  From here on the workbook is safely on disk. Nothing below is worth
    #  throwing that away for, so each step warns and carries on instead
    #  of taking the whole run down with it.
    try:
        note_sources(dirs["day"], sources)
    except OSError:
        pass                    # the sources note is a nicety, not the report

    # ---- the email -----------------------------------------------------
    parts = email_parts(s, summary_rows, tu, cu, det, ctx, ca, coates_owned,
                        tooling_tx, live_onhire, sources, when)
    try:
        html_path, eml_path = write_email(base, s, when, asat, parts,
                                          latest or out, dirs["day"])
    except OSError as e:
        say("  WARN | The workbook saved fine, but the email couldn't be "
            "written ({}).".format(e))
        say("         Attach the workbook to a normal email this once, and "
            "run")
        say("         1_CHECK_THIS_LAPTOP.bat to see what is locking the "
            "folder.")
        html_path = eml_path = None

    say("=" * 66)
    say(" COATES | ON-HIRE WORKBOOK - {}".format(s.header_line))
    say("=" * 66)
    for name, n, _why in tabs:
        say("  {:<26} {:>6} rows".format(name, n))
    say("-" * 66)
    say("  Workbook  {}".format(os.path.relpath(out, base)))
    if latest:
        say("  Open this {}".format(os.path.relpath(latest, base)))
    if eml_path:
        say("  Email     {}".format(os.path.relpath(eml_path, base)))
        say("            double-click it - it opens as a DRAFT. Nothing sends")
        say("            until you press Send yourself.")
    if html_path:
        say("  Web page  {}".format(os.path.relpath(html_path, base)))
    say("=" * 66)
    return 0


def email_parts(s, summary_rows, tu, cu, det, ctx, ca, coates_owned,
                tooling_tx, live_onhire, sources, when):
    """Everything the email needs - as NUMBERS as well as text.

    The email draws a little bar beside each figure, and a bar needs the
    value AND the biggest value in its column, so this hands back both
    and lets the HTML do the formatting.

    It also works out the exceptions - the lines being asked for faster
    than we can turn them around, the consumables running down, the gear
    that has been out too long - because those are the only bits anybody
    has to act on. Everything in here is operational. No rates, no
    costs, no dollar signs: the money stays in the workbook.
    """
    LONG_OUT_DAYS = 21      # out longer than this and it wants a phone call
    CONS_TIGHT = 0.80       # this far down the position and we order

    n_companies = len([c for c, n in live_onhire.items() if n])
    coates_out = len([r for r in coates_owned
                      if _s(r.get("ITEM_STATUS")) == "On Hire"])

    #  ---- the exceptions. These drive the summary and the actions. ----
    hot = [r for r in tu if r[11] == "High Demand"]
    tight = [r for r in cu if r[4] >= CONS_TIGHT]
    dated = [r for r in det if r[5]]
    dated.sort(key=lambda r: r[5])
    long_out = [r for r in dated if (when - r[5]).days >= LONG_OUT_DAYS]
    idle_tool = [r for r in tu if r[11] == "No Use" and r[1] >= 2]
    idle_units = sum(r[1] for r in idle_tool)

    #  How long the oldest thing on site has been out, all up and per
    #  company. The workbook carries a "not returned" count as well, off
    #  the transaction end dates - but an item still out IS an item not
    #  returned, so in practice the two agree and printing both put the
    #  same column in the email twice. The live on-hire position wins,
    #  and the reconciliation stays in the workbook where it belongs.
    oldest_days = max([(when - r[5]).days for r in dated] or [0])
    oldest_item = dated[0][3] if dated else ""
    co_oldest = {}
    for r in dated:
        days = (when - r[5]).days
        if days > co_oldest.get(r[0], -1):
            co_oldest[r[0]] = days

    tiles = [
        ("Items on hire now", "{:,}".format(len(det)),
         "with {:,} companies".format(n_companies)),
        ("Longest out (days)", "{:,}".format(oldest_days),
         "oldest: {}".format(oldest_item) if oldest_item
         else "nothing dated on hire"),
        ("Coates items on site", "{:,}".format(len(coates_owned)),
         "{:,} of them out right now".format(coates_out)),
        ("Tooling issues & returns", "{:,}".format(len(tooling_tx)),
         "across {:,} item types".format(
             len({_s(r.get("SKU/ITEM DESCRIPTION")) for r in tooling_tx}))),
        ("Consumables issued", "{:,}".format(len(ctx)),
         "across {:,} lines".format(len({r[2] for r in ctx}))),
        ("Consumable lines on shelf", "{:,}".format(len(ca)),
         "{:,} at {:.0f}% or more".format(len(tight), CONS_TIGHT * 100)),
    ]

    #  The three or four things worth knowing before the tables. Every
    #  number here is counted off the exports - nothing is estimated.
    flags = [
        ("red" if hot else "green", "Order more",
         "{:,} tooling {}".format(len(hot),
                                  "line" if len(hot) == 1 else "lines")),
        ("red" if tight else "green", "Consumables tight",
         "{:,} {}".format(len(tight),
                          "line" if len(tight) == 1 else "lines")),
        ("amber" if long_out else "green",
         "Out over {} days".format(LONG_OUT_DAYS),
         "{:,} {}".format(len(long_out),
                          "item" if len(long_out) == 1 else "items")),
        ("slate", "Not moving at all",
         "{:,} units".format(idle_units)),
    ]

    #  The position, in the words you'd use on the phone.
    call = ["{:,} items are on hire to {:,} companies, and the oldest has "
            "been out {:,} days.".format(len(det), n_companies, oldest_days)]
    if hot or tight:
        bits = []
        if hot:
            bits.append("{:,} tooling {} being asked for faster than we can "
                        "turn {} around".format(
                            len(hot), "line is" if len(hot) == 1 else "lines are",
                            "it" if len(hot) == 1 else "them"))
        if tight:
            bits.append("{:,} consumable {} past {:.0f}% of the position"
                        .format(len(tight),
                                "line is" if len(tight) == 1 else "lines are",
                                CONS_TIGHT * 100))
        call.append("Worth knowing: " + " and ".join(bits) + ".")
    if long_out:
        call.append("{:,} {} been out longer than {} days.".format(
            len(long_out), "item has" if len(long_out) == 1 else "items have",
            LONG_OUT_DAYS))
    if idle_tool:
        call.append("The other way, {:,} units across {:,} lines haven't gone "
                    "out once - shelf space and hire we're carrying for "
                    "nothing.".format(idle_units, len(idle_tool)))

    #  What I need from you. Only the ones that actually apply, and each
    #  one names its own number so nobody has to go hunting for it.
    actions = []
    if hot:
        actions.append(("Top these up", "{:,} tooling {} marked Increase "
                        "Stock. Say the word and I'll get them on order."
                        .format(len(hot),
                                "line" if len(hot) == 1 else "lines")))
    if tight:
        actions.append(("Order consumables", "{:,} {} at {:.0f}% or more of "
                        "the position. Cheaper to order now than to run out "
                        "mid-shift.".format(
                            len(tight), "line" if len(tight) == 1 else "lines",
                            CONS_TIGHT * 100)))
    if long_out:
        actions.append(("Chase these back", "{:,} {} out longer than {} "
                        "days. Worth a call before it turns into a demob "
                        "problem.".format(
                            len(long_out),
                            "item" if len(long_out) == 1 else "items",
                            LONG_OUT_DAYS)))
    if idle_tool:
        actions.append(("Or send it home", "{:,} units across {:,} lines "
                        "haven't moved at all. Nod and I'll off-hire them."
                        .format(idle_units, len(idle_tool))))
    if not actions:
        actions.append(("Nothing needed", "No exceptions today - the store "
                        "is tracking. Full detail is in the workbook."))

    #  ---- the tables. Numbers stay numbers so the bars can be drawn. --
    by_co = {}
    for row in summary_rows:
        by_co.setdefault(row[0], {})[row[1]] = row
    companies = []
    for co in sorted(by_co, key=lambda c: -(by_co[c].get("TOTAL",
                                                         [0] * 7)[2])):
        t = by_co[co].get("TOOLING", [co, "", 0, 0, 0, 0, 0])
        c = by_co[co].get("CONSUMABLES", [co, "", 0, 0, 0, 0, 0])
        if not (t[2] or c[2] or t[5]):
            continue
        companies.append([co, t[5], t[2], c[2], co_oldest.get(co, 0)])

    #  Working hardest - the gear earning its keep, best turns first.
    hard = [r for r in tu if r[11] in ("High Demand", "Good Use")]
    hard.sort(key=lambda r: (-(r[5] / r[1] if r[1] else 0), -r[5]))
    hardest = [[r[0], r[1], r[2], r[5], r[11], r[12]] for r in hard[:12]]

    #  Not moving - biggest holdings that have not gone out at all.
    idle_tool.sort(key=lambda r: -r[1])
    idle_rows = [[r[0], r[1], r[2], r[12]] for r in idle_tool[:12]]

    #  Consumables worth watching - the ones running down, tightest first.
    watch = [r for r in cu if r[9] in ("High Demand", "Good Use")]
    watch.sort(key=lambda r: -r[4])
    cons = [[r[0], r[1], r[2], r[4], r[10]] for r in watch[:12]]

    #  Out the longest - oldest on-hire dates still open. The item NUMBER
    #  rides along, because four lines all saying "Rio Tinto DropMat Kit"
    #  are four different kits and somebody has to go and find them.
    longest = [[r[3], r[2], r[0], r[1], r[5], (when - r[5]).days]
               for r in dated[:12]]

    #  ---- everything on hire, company by company, person by person ----
    #  Alphabetical both ways, because this is the list people scroll
    #  through looking for their own name. Same gear description twice
    #  in one name is counted, not repeated - "4 x DropMat Kit" reads
    #  better than four identical lines and takes a quarter of the room.
    roll = {}
    for r in det:
        co = _s(r[0]) or "(no company named)"
        who = _s(r[1]) or "(nobody named on it)"
        p = roll.setdefault(co, {}).setdefault(
            who, {"items": {}, "oldest": None, "n": 0})
        desc = _s(r[3]) or "(no description)"
        p["items"][desc] = p["items"].get(desc, 0) + 1
        p["n"] += 1
        if r[5] and (p["oldest"] is None or r[5] < p["oldest"]):
            p["oldest"] = r[5]

    on_hire_roll = []
    for co in sorted(roll, key=lambda c: c.upper()):
        people = []
        for who in sorted(roll[co], key=lambda w: w.upper()):
            d = roll[co][who]
            people.append({
                "who": who,
                "n": d["n"],
                "since": d["oldest"],
                "days": (when - d["oldest"]).days if d["oldest"] else None,
                #  Most-held first, then alphabetical.
                "gear": sorted(d["items"].items(),
                               key=lambda kv: (-kv[1], kv[0].upper())),
            })
        on_hire_roll.append({
            "company": co,
            "items": sum(p["n"] for p in people),
            "lines": len({g for p in people for g, _q in p["gear"]}),
            "people": people,
        })

    stamp = ", ".join(sorted({
        dt.datetime.fromtimestamp(os.path.getmtime(p)).strftime("%d %b %Y")
        for p in sources})) or when.strftime("%d %b %Y")

    return {"tiles": tiles, "flags": flags, "call": call, "actions": actions,
            "companies": companies, "hardest": hardest, "idle": idle_rows,
            "consumables": cons, "longest": longest, "source_stamp": stamp,
            "roll": on_hire_roll,
            "long_out_days": LONG_OUT_DAYS, "cons_tight": CONS_TIGHT,
            "counts": {"onhire": len(det), "companies": n_companies,
                       "oldest_days": oldest_days, "hot": len(hot),
                       "tight": len(tight), "long_out": len(long_out),
                       "idle_units": idle_units}}


def carry_over(wb, s, base, title):
    """Copy Andrew's hand-typed tabs out of the existing workbook,
    values, formulas, dates and times exactly as he left them. If the
    old workbook isn't there the tabs are still created, headed and
    ready to type into - an empty tab that works beats a missing one."""
    src = s.workbook(base)
    carried, labour_total, cost_total = [], 0.0, 0.0

    old = None
    if src and os.path.isfile(src):
        try:
            old = openpyxl.load_workbook(src, data_only=False)
        except Exception as e:
            say("  WARN | Could not read {} ({}). The two hand-typed "
                  "tabs come through blank.".format(os.path.basename(src), e))

    #  The typed values, read a second time with formulas resolved, so
    #  the cover can show a total without re-implementing his sums.
    vals = None
    if src and os.path.isfile(src):
        try:
            vals = openpyxl.load_workbook(src, data_only=True)
        except Exception:
            vals = None

    for name in CARRIED_OVER:
        ws = wb.create_sheet(name)
        n = 0
        if old is not None and name in old.sheetnames:
            o = old[name]
            #  From row 2. Row 1 is the title band, which band() writes
            #  fresh below - copying it forward only carried the previous
            #  job's name into the new workbook for a moment, and counted
            #  itself as a row of typing that nobody typed.
            for row in o.iter_rows(min_row=2):
                filled = False
                for c in row:
                    if c.value is None:
                        continue
                    filled = True
                    nc = ws.cell(row=c.row, column=c.column, value=c.value)
                    #  Carry the LOOK across as well as the value. These
                    #  tabs are laid out by hand - boxed, filled, headed -
                    #  and copying only the text used to strip all of that
                    #  out on every refresh, so a tab that was tidy on
                    #  Monday was bare by Tuesday. copy() because a style
                    #  object belongs to one workbook and openpyxl will
                    #  not let it be shared between two.
                    nc.font = copy(c.font)
                    nc.fill = copy(c.fill)
                    nc.border = copy(c.border)
                    nc.alignment = copy(c.alignment)
                    if c.number_format:
                        nc.number_format = c.number_format
                if filled:
                    n += 1
            #  Row heights and the frozen header, so a long roster still
            #  scrolls with its headings in view.
            for k, dim in o.row_dimensions.items():
                if dim.height:
                    ws.row_dimensions[k].height = dim.height
            if o.freeze_panes:
                ws.freeze_panes = o.freeze_panes
            for rng in list(o.merged_cells.ranges):
                if rng.min_row <= 1:
                    #  Row 1 belongs to the title band, added below. Two
                    #  merged ranges that overlap make the file invalid
                    #  and Excel offers to "repair" it on open.
                    continue
                try:
                    ws.merge_cells(str(rng))
                except ValueError:
                    pass
            for k, dim in o.column_dimensions.items():
                if dim.width:
                    ws.column_dimensions[k].width = dim.width
        if n == 0:
            band(ws, title, 9)
            ws.cell(row=2, column=1,
                    value="Nothing typed in this tab yet - it is yours to "
                          "fill in and the rebuild never overwrites it.")
        else:
            #  Re-brand the title strip so it matches the other tabs.
            band(ws, title, max(5, ws.max_column))
        tidy_print(ws, max(5, ws.max_column))
        ws.sheet_properties.tabColor = NEAR_BLACK      # typed by hand
        carried.append((name, n))

    #  Totals for the cover, straight off the values copy.
    if vals is not None:
        if "Coates Labour" in vals.sheetnames:
            #  Shift lines only. The sheet carries its own grand total on
            #  a row with no date in it - counting that as well doubles
            #  the labour bill, which is exactly what it did first go.
            o = vals["Coates Labour"]
            for row in o.iter_rows(min_row=6, min_col=1, max_col=9,
                                   values_only=True):
                if _date(row[0]) is None:
                    continue
                labour_total += _n(row[8])
        if "Cost Breakdown" in vals.sheetnames:
            o = vals["Cost Breakdown"]
            for row in o.iter_rows(min_row=3, max_row=19, min_col=3,
                                   max_col=5, values_only=True):
                for v in row:
                    cost_total += _n(v)
        try:
            vals.close()
        except Exception:
            pass
    if old is not None:
        try:
            old.close()
        except Exception:
            pass
    return carried, labour_total, cost_total


def cover(wb, s, when, asat, tabs, sources, base, headline):
    """The tab that was blank. Says what this workbook is, what fed it,
    the headline numbers, and how the two ratings are worked out - so
    nobody has to take the recommendations on trust."""
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    band(ws, "Coates    |  {}  |  SHUTDOWN TOOLSTORE ON-HIRE REPORT"
             "        Powered by SITEIQ".format(s.short or s.customer), 6)

    def put(row, col, value, bold=False, size=11, color=None, fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Calibri", size=size, bold=bold,
                      color=color or "FF000000")
        if fmt:
            c.number_format = fmt
        return c

    def rule(row, text):
        c = put(row, 1, text, bold=True, size=12, color=WHITE)
        c.fill = PatternFill("solid", fgColor=ORANGE_ALT)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        return row + 1

    r = 3
    put(r, 1, s.customer, bold=True, size=20); r += 1
    put(r, 1, "{} · {}".format(s.job, s.location), size=13,
        color="FF595959"); r += 2
    put(r, 1, "As at"); put(r, 2, asat.strftime("%d %b %Y  %H:%M"), bold=True); r += 1
    put(r, 1, "Report date"); put(r, 2, when.strftime("%d %b %Y"), bold=True); r += 1
    put(r, 1, "Prepared by"); put(r, 2, s.author or "Andrew Fisher", bold=True); r += 1
    put(r, 1, "SiteIQ project"); put(r, 2, s.project_schedule or "-", bold=True); r += 2

    r = rule(r, "THE HEADLINES")
    for k, v in headline.items():
        put(r, 1, k)
        money = "cost" in k.lower()
        put(r, 2, round(v, 2) if money else int(v), bold=True,
            fmt=MONEY_FMT if money else INT_FMT)
        r += 1
    r += 1

    r = rule(r, "WHAT'S IN THIS WORKBOOK")
    put(r, 1, "Tab", bold=True); put(r, 2, "Rows", bold=True)
    put(r, 3, "What it tells you", bold=True); r += 1
    for name, n, why in tabs:
        put(r, 1, name)
        put(r, 2, n, fmt=INT_FMT)
        put(r, 3, why, color="FF595959")
        r += 1
    r += 1

    r = rule(r, "WHERE THE NUMBERS CAME FROM")
    for p in sources:
        put(r, 1, os.path.basename(p))
        put(r, 2, dt.datetime.fromtimestamp(os.path.getmtime(p))
            .strftime("%d %b %Y  %H:%M"), color="FF595959")
        r += 1
    r += 1

    r = rule(r, "HOW THE RATINGS ARE WORKED OUT")
    for line in [
        "Tooling is rated on TURNS - transactions divided by how many of",
        "that item are on site. One item that went out five times is",
        "working; five items that went out once each are sitting there.",
        "   no turns          No Use        Review / Reduce",
        "   under 1 turn      Low Use       Keep Stock",
        "   1 to 2 turns      Good Use      Keep Stock  (Monitor / Increase",
        "                                   if 80% or more are out right now)",
        "   2 turns or more   High Demand   Increase Stock",
        "",
        "Consumables are rated on how much of the position has gone out -",
        "sales divided by sales plus what's left on the shelf.",
        "   nothing sold      No Use        Review / Reduce",
        "   under 40%         Low Use       Keep Stock",
        "   40% to 80%        Good Use      Keep Stock",
        "   over 80%          High Demand   Increase Stock",
    ]:
        put(r, 1, line, color="FF404040")
        r += 1
    r += 1

    r = rule(r, "WHAT IS TYPED BY HAND")
    for line in [
        "Coates Labour and Cost Breakdown are yours. The rebuild reads",
        "them out of the previous workbook and writes them straight back -",
        "it never recalculates them and never overwrites your typing.",
        "Everything else on every other tab is built from the SiteIQ",
        "exports listed above, so a fresh pull refreshes the lot.",
    ]:
        put(r, 1, line, color="FF404040")
        r += 1

    for col, w in ((1, 46), (2, 24), (3, 52), (4, 12), (5, 12), (6, 12)):
        ws.column_dimensions[get_column_letter(col)].width = w
    #  The cover is the page people print and put on the wall, so it
    #  goes portrait while the data tabs go landscape.
    tidy_print(ws, 6, landscape=False)
    ws.print_title_rows = "1:1"
    ws.sheet_properties.tabColor = NEAR_BLACK
    return ws


# =====================================================================
#  THE EMAIL - the whole report in the body, and it has to look the part
# =====================================================================
#  Outlook on Windows does not draw email with a browser engine. It
#  draws it with WORD. That one fact decides every choice below:
#
#    * Tables do the layout. No flexbox, no grid, no floats, no divs
#      holding a column up.
#    * Every style is written inline, on the element it belongs to. A
#      stylesheet in the head gets thrown away.
#    * bgcolor="" goes on alongside background: - Word trusts the
#      attribute more than the style. Belt and braces on every panel,
#      which also stops Outlook's dark mode inverting half the report.
#    * The body is a fixed pixel width centred inside an Outlook-only
#      wrapper, because Word ignores max-width and would otherwise
#      stretch the tables across a 34-inch monitor. Everything else
#      gets the fluid version and reads fine on a phone.
#    * Border-spacing is ignored by Word, so gaps between the tiles are
#      real spacer cells, not spacing.
#    * The bars and the status chips are nested tables with a
#      background colour. NO IMAGES ANYWHERE: no logo file to go
#      missing, no red X where a picture should be, nothing to
#      download on site, no "click to display images" bar across the
#      top. It looks the same on the store laptop, on a phone at the
#      gate, and in Gmail.
#
#  NO COSTS anywhere in it. Not the labour, not the cost breakdown, not
#  a rate and not a dollar sign. This is the operational picture and it
#  can go to anyone. The money stays in the workbook.
# =====================================================================

#  Coates orange and near-black are the brand. The rest is the
#  supporting greyscale, kept deliberately narrow so the orange means
#  something when it does appear.
E_ORANGE = "#F26222"
E_INK = "#1D1D1B"
E_PAGE = "#EDEFF2"
E_CARD = "#FFFFFF"
E_LINE = "#DCE1E7"
E_MUTED = "#5B6472"
E_ZEBRA = "#F7F8FA"
E_TRACK = "#E4E8ED"
E_FONT = "Segoe UI,Calibri,Arial,sans-serif"
E_WIDTH = 760

#  Red / amber / green / slate, as (chip background, text, bar). Muted
#  on purpose - a report that shouts in six colours reads like a
#  warning label. These are the only colours that carry a meaning.
E_RAG = {
    "red": ("#FBE7E6", "#A32015", "#D0342C"),
    "amber": ("#FDF1DC", "#8A5300", "#E08A24"),
    "green": ("#E8F3E9", "#1E6B32", "#409A55"),
    "slate": ("#E9ECF0", "#4A5361", "#98A0AA"),
    "orange": ("#FDEAE0", "#A6410F", E_ORANGE),
}

#  Ratings and recommendations carry their own colour, so the same word
#  always looks the same wherever it turns up in the report.
E_TONE = {
    "High Demand": "red", "Good Use": "green",
    "Low Use": "amber", "No Use": "slate",
    "Increase Stock": "red", "Monitor / Increase": "amber",
    "Keep Stock": "green", "Review / Reduce": "slate",
}

E_TXT = ("font-family:{f};font-size:11pt;color:{i};line-height:1.55"
         .format(f=E_FONT, i=E_INK))
E_SMALL = ("font-family:{f};font-size:9pt;color:{m};line-height:1.5"
           .format(f=E_FONT, m=E_MUTED))


def _esc(v):
    return (_s(v).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def _fmt(v):
    """Only used for display - the thousands comma."""
    try:
        return "{:,}".format(int(v))
    except (TypeError, ValueError):
        return _esc(v)


def _spacer(px, colspan=1, bg=""):
    """A row of empty space that Word actually honours. A td with no
    content collapses, so it gets a non-breaking space and a font size
    of nothing to sit on."""
    return ("<tr><td colspan='{c}'{b} style=\"height:{p}px;font-size:1px;"
            "line-height:{p}px\">&nbsp;</td></tr>".format(
                c=colspan, p=px,
                b=" bgcolor=\"{}\"".format(bg) if bg else ""))


def _rule(px, colour, colspan=1):
    return ("<tr><td colspan='{c}' bgcolor=\"{k}\" style=\"background:{k};"
            "height:{p}px;font-size:1px;line-height:{p}px\">&nbsp;</td></tr>"
            .format(c=colspan, k=colour, p=px))


def _pill(text, tone):
    """A status chip. Word squares off the corners and that is fine -
    the colour is doing the work, not the radius."""
    bg, fg, _bar = E_RAG.get(tone, E_RAG["slate"])
    return ("<table cellpadding='0' cellspacing='0' border='0' "
            "style='border-collapse:collapse'><tr>"
            "<td bgcolor=\"{bg}\" style=\"background:{bg};border-radius:3px;"
            "padding:3px 7px;font-family:{f};font-size:8.5pt;"
            "font-weight:700;color:{fg};white-space:nowrap\">{t}</td>"
            "</tr></table>".format(bg=bg, fg=fg, f=E_FONT, t=_esc(text)))


def _bar(value, biggest, tone="orange"):
    """A bar the width of its cell, drawn out of two table cells: the
    filled part and the track behind it. This is the whole reason the
    report reads at a glance instead of being a wall of digits."""
    _bg, _fg, fill = E_RAG.get(tone, E_RAG["orange"])
    try:
        pct = (float(value) / float(biggest) * 100.0) if biggest else 0.0
    except (TypeError, ValueError, ZeroDivisionError):
        pct = 0.0
    pct = max(0.0, min(100.0, pct))
    if pct <= 0:
        return ("<table cellpadding='0' cellspacing='0' border='0' "
                "width='100%' style='border-collapse:collapse;width:100%'>"
                "<tr><td bgcolor=\"{t}\" style=\"background:{t};height:5px;"
                "font-size:1px;line-height:5px\">&nbsp;</td></tr></table>"
                .format(t=E_TRACK))
    left = int(round(pct))
    left = max(3, min(100, left))
    cells = ("<td width='{l}%' bgcolor=\"{f}\" style=\"background:{f};"
             "height:5px;font-size:1px;line-height:5px\">&nbsp;</td>"
             .format(l=left, f=fill))
    if left < 100:
        cells += ("<td width='{r}%' bgcolor=\"{t}\" style=\"background:{t};"
                  "height:5px;font-size:1px;line-height:5px\">&nbsp;</td>"
                  .format(r=100 - left, t=E_TRACK))
    return ("<table cellpadding='0' cellspacing='0' border='0' width='100%' "
            "style='border-collapse:collapse;width:100%'><tr>{c}</tr></table>"
            .format(c=cells))


def _metric(text, value, biggest, tone="orange"):
    """The number, and under it the bar that says how big it is next to
    everything else in the column. A whole column of these is a chart."""
    return ("<div style=\"font-family:{f};font-size:11pt;font-weight:700;"
            "color:{i};text-align:right;line-height:1.2\">{t}</div>"
            "<div style=\"font-size:1px;line-height:4px\">&nbsp;</div>{b}"
            .format(f=E_FONT, i=E_INK, t=_esc(text),
                    b=_bar(value, biggest, tone)))


def _lead(text, sub=""):
    """First column of a table row: the thing the row is about."""
    out = ("<div style=\"font-family:{f};font-size:10.5pt;color:{i};"
           "font-weight:600;line-height:1.35\">{t}</div>".format(
               f=E_FONT, i=E_INK, t=_esc(text)))
    if sub:
        out += ("<div style=\"font-family:{f};font-size:9pt;color:{m};"
                "padding-top:2px\">{s}</div>".format(f=E_FONT, m=E_MUTED,
                                                     s=_esc(sub)))
    return out


def _plain(text, align="left", muted=False):
    return ("<div style=\"font-family:{f};font-size:10.5pt;color:{c};"
            "text-align:{a};line-height:1.35\">{t}</div>".format(
                f=E_FONT, c=E_MUTED if muted else E_INK, a=align,
                t=_esc(text)))


def _section(number, title, subtitle):
    """An orange numbered square against a dark bar. Gives the report a
    spine you can scroll to, and tells you what the table is FOR before
    you start reading numbers out of it."""
    return (
        "<table cellpadding='0' cellspacing='0' border='0' width='100%' "
        "style='border-collapse:collapse;width:100%;margin:0'>"
        "<tr>"
        "<td width='38' bgcolor=\"{o}\" align='center' valign='middle' "
        "style=\"width:38px;background:{o};font-family:{f};font-size:15pt;"
        "font-weight:800;color:#ffffff\">{n}</td>"
        "<td bgcolor=\"{k}\" style=\"background:{k};padding:9px 14px\">"
        "<div style=\"font-family:{f};font-size:12pt;font-weight:700;"
        "color:#ffffff;line-height:1.25\">{t}</div>"
        "<div style=\"font-family:{f};font-size:9pt;color:#B9C0C9;"
        "padding-top:2px\">{s}</div></td>"
        "</tr></table>".format(o=E_ORANGE, k=E_INK, f=E_FONT, n=number,
                               t=_esc(title), s=_esc(subtitle)))


def _company_band(name, items, people, lines):
    """The strip that starts each company's block in the on-hire roll.
    Orange, so you can scroll and land on the company you want."""
    return (
        "<table cellpadding='0' cellspacing='0' border='0' width='100%' "
        "style='border-collapse:collapse;width:100%'><tr>"
        "<td bgcolor=\"{o}\" style=\"background:{o};padding:8px 12px;"
        "font-family:{f};font-size:11.5pt;font-weight:700;color:#ffffff\">"
        "{n}</td>"
        "<td bgcolor=\"{o}\" align='right' style=\"background:{o};"
        "padding:8px 12px;font-family:{f};font-size:9.5pt;color:#FFE4D6;"
        "white-space:nowrap\">{i} item{s} &nbsp;&middot;&nbsp; {l} line{ls}"
        " &nbsp;&middot;&nbsp; {p} {pw}</td>"
        "</tr></table>".format(
            o=E_ORANGE, f=E_FONT, n=_esc(name), i=_fmt(items),
            s="" if items == 1 else "s", l=_fmt(lines),
            ls="" if lines == 1 else "s", p=_fmt(people),
            pw="person" if people == 1 else "people"))


def _gear_list(gear):
    """What one person has out. Counted, so four of the same kit is one
    line that says four, not four lines that look like a glitch."""
    if not gear:
        return ""
    #  One styled block, not one per line. With 112 items on hire the
    #  per-line version added 40KB of repeated styling to the email, and
    #  Gmail quietly clips anything over about 100KB - so the bottom of
    #  the report would have disappeared behind a "view entire message"
    #  link on half the recipients.
    lines = "<br>".join(
        "<b style=\"color:{i}\">{q} &times;</b> {d}".format(
            i=E_INK, q=qty, d=_esc(desc)) for desc, qty in gear)
    return ("<div style=\"font-family:{f};font-size:9.5pt;color:{m};"
            "line-height:1.55\">{l}</div>".format(
                f=E_FONT, m=E_MUTED, l=lines))


def _table(cols, rows, widths=None, aligns=None, note="", valign=None):
    """One data table. Dark header, an orange rule under it, zebra rows,
    and every cell carrying its own styling because Word will not read a
    stylesheet. Cells arrive as finished HTML from the caller - that is
    what lets a cell hold a bar or a status chip instead of text."""
    if not rows:
        return ("<table cellpadding='0' cellspacing='0' border='0' "
                "width='100%' style='border-collapse:collapse;width:100%'>"
                "<tr><td bgcolor=\"{c}\" style=\"background:{c};border:1px "
                "solid {l};border-top:none;padding:14px;font-family:{f};"
                "font-size:10.5pt;color:{m}\">Nothing to report here today."
                "</td></tr></table>".format(c=E_CARD, l=E_LINE, f=E_FONT,
                                            m=E_MUTED))
    widths = widths or {}
    aligns = aligns or {}
    ncols = len(cols)
    out = ["<table cellpadding='0' cellspacing='0' border='0' width='100%' "
           "style='border-collapse:collapse;width:100%'>"]

    out.append("<tr>")
    for i, c in enumerate(cols):
        w = (" width='{}%'".format(widths[i])) if i in widths else ""
        ws = ("width:{}%;".format(widths[i])) if i in widths else ""
        out.append(
            "<th{w} bgcolor=\"{k}\" style=\"{ws}background:{k};padding:8px "
            "10px;font-family:{f};font-size:8.5pt;font-weight:700;"
            "color:#ffffff;letter-spacing:.5px;text-transform:uppercase;"
            "text-align:{a};vertical-align:bottom\">{c}</th>".format(
                w=w, ws=ws, k=E_INK, f=E_FONT,
                a=aligns.get(i, "left"), c=_esc(c)))
    out.append("</tr>")
    #  A real 3px orange row rather than a border, because Word draws a
    #  background reliably and a border on a th sometimes not at all.
    out.append(_rule(3, E_ORANGE, colspan=ncols))

    for n, row in enumerate(rows):
        bg = E_CARD if n % 2 == 0 else E_ZEBRA
        out.append("<tr>")
        for i, cell in enumerate(row):
            out.append(
                "<td bgcolor=\"{b}\" style=\"background:{b};padding:7px 10px;"
                "border-bottom:1px solid {l};vertical-align:{v}\">{c}</td>"
                .format(b=bg, l=E_LINE, c=cell,
                        v=valign or ("middle" if i else "top")))
        out.append("</tr>")
    out.append("</table>")

    if note:
        out.append(
            "<table cellpadding='0' cellspacing='0' border='0' width='100%' "
            "style='border-collapse:collapse;width:100%'><tr>"
            "<td bgcolor=\"{c}\" style=\"background:{c};padding:7px 10px 0;"
            "font-family:{f};font-size:9pt;color:{m};line-height:1.45\">"
            "{n}</td></tr></table>".format(c=E_PAGE, f=E_FONT, m=E_MUTED,
                                           n=_esc(note)))
    return "".join(out)


def _masthead(s, when, asat):
    """The top of the report. Built out of type and two brand colours -
    no logo file, because a logo file is one more thing that can arrive
    as a red X on somebody else's laptop."""
    return (
        "<table cellpadding='0' cellspacing='0' border='0' width='100%' "
        "style='border-collapse:collapse;width:100%'>"
        + _rule(5, E_ORANGE)
        + "<tr><td bgcolor=\"{k}\" style=\"background:{k};padding:16px 18px\">"
          "<table cellpadding='0' cellspacing='0' border='0' width='100%' "
          "style='border-collapse:collapse;width:100%'><tr>"
          "<td valign='top' style=\"font-family:{f}\">"
          "<div style=\"font-family:{f};font-size:20pt;font-weight:800;"
          "color:{o};letter-spacing:3px;line-height:1.1\">COATES</div>"
          "<div style=\"font-family:{f};font-size:8.5pt;font-weight:700;"
          "color:#9AA2AC;letter-spacing:1.6px;text-transform:uppercase;"
          "padding-top:3px\">Shutdown Tool Store</div></td>"
          "<td valign='top' align='right' style=\"font-family:{f}\">"
          "<div style=\"font-family:{f};font-size:15pt;font-weight:700;"
          "color:#ffffff;line-height:1.2\">On-Hire Report</div>"
          "<div style=\"font-family:{f};font-size:10pt;color:#B9C0C9;"
          "padding-top:3px\">{d}</div></td>"
          "</tr></table></td></tr>".format(k=E_INK, f=E_FONT, o=E_ORANGE,
                                           d=when.strftime("%d %b %Y"))
        + "<tr><td bgcolor=\"{o}\" style=\"background:{o};padding:9px 18px\">"
          "<table cellpadding='0' cellspacing='0' border='0' width='100%' "
          "style='border-collapse:collapse;width:100%'><tr>"
          "<td style=\"font-family:{f};font-size:10.5pt;font-weight:700;"
          "color:#ffffff\">{hdr}</td>"
          "<td align='right' style=\"font-family:{f};font-size:9.5pt;"
          "color:#FFE4D6;white-space:nowrap\">As at {asat}</td>"
          "</tr></table></td></tr>".format(
              o=E_ORANGE, f=E_FONT, hdr=_esc(s.header_line),
              asat=asat.strftime("%d %b %Y  %H:%M"))
        + "<tr><td bgcolor=\"#2A2F36\" style=\"background:#2A2F36;"
          "padding:6px 18px;font-family:{f};font-size:8.5pt;color:#9AA2AC;"
          "letter-spacing:.8px\">POWERED BY <span style=\"color:{o};"
          "font-weight:800\">SITEIQ</span> &nbsp;&nbsp;|&nbsp;&nbsp; "
          "Author: {who} &nbsp;&nbsp;|&nbsp;&nbsp; Operational report - "
          "no costs</td></tr>".format(f=E_FONT, o=E_ORANGE,
                                      who=_esc(s.author or "Andrew Fisher"))
        + "</table>")


def _call_card(parts):
    """The position, up front, in the words you'd use on the phone -
    then the flags. Nobody should have to read six tables to find out
    whether today is a good day or a bad one."""
    out = ["<table cellpadding='0' cellspacing='0' border='0' width='100%' "
           "style='border-collapse:collapse;width:100%'>"
           "<tr><td bgcolor=\"{c}\" style=\"background:{c};border:1px solid "
           "{l};border-left:4px solid {o};padding:14px 16px\">".format(
               c=E_CARD, l=E_LINE, o=E_ORANGE)]
    out.append("<div style=\"font-family:{f};font-size:8.5pt;font-weight:700;"
               "color:{o};letter-spacing:1.4px;text-transform:uppercase;"
               "padding-bottom:6px\">Where we sit today</div>".format(
                   f=E_FONT, o=E_ORANGE))
    for i, line in enumerate(parts["call"]):
        size = "12pt" if i == 0 else "10.5pt"
        weight = "600" if i == 0 else "400"
        colour = E_INK if i == 0 else E_MUTED
        out.append("<div style=\"font-family:{f};font-size:{z};"
                   "font-weight:{w};color:{c};line-height:1.5;"
                   "padding-bottom:4px\">{t}</div>".format(
                       f=E_FONT, z=size, w=weight, c=colour, t=_esc(line)))
    out.append("</td></tr></table>")

    #  The flags. Four little chips, spacer cells between them because
    #  Word will not honour border-spacing.
    flags = parts["flags"]
    out.append(_spacer(10))
    out.append("<table cellpadding='0' cellspacing='0' border='0' "
               "width='100%' style='border-collapse:collapse;width:100%'>"
               "<tr>")
    #  Even quarters, so four flags of different word-lengths still line
    #  up as four matching cards instead of four different sizes.
    each = int(100 / max(1, len(flags)))
    for i, (tone, label, value) in enumerate(flags):
        bg, fg, bar = E_RAG.get(tone, E_RAG["slate"])
        if i:
            out.append("<td width='8' style='width:8px;font-size:1px'>"
                       "&nbsp;</td>")
        out.append(
            "<td width='{pc}%' bgcolor=\"{c}\" style=\"width:{pc}%;"
            "background:{c};border:1px solid {l};"
            "border-top:3px solid {bar};padding:9px 11px\">"
            "<div style=\"font-family:{f};font-size:8.5pt;font-weight:700;"
            "color:{m};letter-spacing:.6px;text-transform:uppercase\">{lb}"
            "</div>"
            "<div style=\"font-family:{f};font-size:11pt;font-weight:700;"
            "color:{fg};padding-top:3px\">{v}</div></td>".format(
                pc=each, c=E_CARD, l=E_LINE, bar=bar, f=E_FONT, m=E_MUTED,
                fg=fg, lb=_esc(label), v=_esc(value)))
    out.append("</tr></table>")
    return "".join(out)


def _tiles(tiles):
    """The headline numbers, three across, two rows. White cards with an
    orange top rule - the one place in the report where the numbers are
    allowed to be big."""
    out = ["<table cellpadding='0' cellspacing='0' border='0' width='100%' "
           "style='border-collapse:collapse;width:100%'>"]
    for start in range(0, len(tiles), 3):
        chunk = tiles[start:start + 3]
        if start:
            out.append(_spacer(9, colspan=5))
        out.append("<tr>")
        for i, (label, value, sub) in enumerate(chunk):
            if i:
                out.append("<td width='9' style='width:9px;font-size:1px'>"
                           "&nbsp;</td>")
            out.append(
                "<td width='33%' bgcolor=\"{c}\" valign='top' "
                "style=\"width:33%;background:{c};border:1px solid {l};"
                "border-top:3px solid {o};padding:12px 14px\">"
                "<div style=\"font-family:{f};font-size:23pt;font-weight:800;"
                "color:{k};line-height:1.05\">{v}</div>"
                "<div style=\"font-family:{f};font-size:8.5pt;font-weight:700;"
                "color:{k};letter-spacing:.6px;text-transform:uppercase;"
                "padding-top:6px\">{lb}</div>"
                "<div style=\"font-family:{f};font-size:9pt;color:{m};"
                "padding-top:2px\">{s}</div></td>".format(
                    c=E_CARD, l=E_LINE, o=E_ORANGE, f=E_FONT, k=E_INK,
                    m=E_MUTED, v=_esc(value), lb=_esc(label), s=_esc(sub)))
        for _ in range(3 - len(chunk)):
            out.append("<td width='9' style='width:9px;font-size:1px'>"
                       "&nbsp;</td>")
            out.append("<td width='33%' bgcolor=\"{p}\" style=\"width:33%;"
                       "background:{p}\">&nbsp;</td>".format(p=E_PAGE))
        out.append("</tr>")
    out.append("</table>")
    return "".join(out)


def _actions_card(parts):
    """What I need from you. Only the things that actually apply, each
    one carrying its own number so nobody has to go and count."""
    out = ["<table cellpadding='0' cellspacing='0' border='0' width='100%' "
           "style='border-collapse:collapse;width:100%'>"]
    out.append("<tr><td bgcolor=\"{k}\" style=\"background:{k};"
               "padding:10px 14px;font-family:{f};font-size:11.5pt;"
               "font-weight:700;color:#ffffff\">What I need from you</td>"
               "</tr>".format(k=E_INK, f=E_FONT))
    out.append(_rule(3, E_ORANGE))
    for n, (head, body) in enumerate(parts["actions"], 1):
        bg = E_CARD if n % 2 else E_ZEBRA
        out.append(
            "<tr><td bgcolor=\"{b}\" style=\"background:{b};padding:10px 14px;"
            "border-bottom:1px solid {l}\">"
            "<table cellpadding='0' cellspacing='0' border='0' width='100%' "
            "style='border-collapse:collapse;width:100%'><tr>"
            "<td width='26' valign='top' style=\"width:26px;font-family:{f};"
            "font-size:11pt;font-weight:800;color:{o}\">{n}</td>"
            "<td valign='top'>"
            "<div style=\"font-family:{f};font-size:10.5pt;font-weight:700;"
            "color:{k}\">{h}</div>"
            "<div style=\"font-family:{f};font-size:10pt;color:{m};"
            "padding-top:2px;line-height:1.45\">{t}</div>"
            "</td></tr></table></td></tr>".format(
                b=bg, l=E_LINE, f=E_FONT, o=E_ORANGE, k=E_INK, m=E_MUTED,
                n=n, h=_esc(head), t=_esc(body)))
    out.append("</table>")
    return "".join(out)


def email_html(s, when, asat, parts):
    """The whole report as an email body. Costs deliberately absent."""
    c = parts["counts"]
    pre = ("{} items on hire with {} companies, oldest out {} days. "
           "{} tooling lines to top up, {} consumable lines tight.".format(
               _fmt(c["onhire"]), _fmt(c["companies"]), _fmt(c["oldest_days"]),
               _fmt(c["hot"]), _fmt(c["tight"])))

    body = []

    #  Inbox preview line. Hidden in the email itself - this is the text
    #  that shows under the subject in the list, and without it Outlook
    #  shows "Morning," which tells nobody anything.
    body.append("<div style=\"display:none;font-size:1px;color:{p};"
                "max-height:0;max-width:0;overflow:hidden;mso-hide:all\">"
                "{t}</div>".format(p=E_PAGE, t=_esc(pre)))

    body.append(_masthead(s, when, asat))
    body.append(_spacer(14))

    body.append("<table cellpadding='0' cellspacing='0' border='0' "
                "width='100%' style='border-collapse:collapse;width:100%'>"
                "<tr><td style=\"{x};padding:0 0 12px\">"
                "<div style=\"{x};padding-bottom:10px\">Morning,</div>"
                "<div style=\"{x}\">Here is where the tool store sits as at "
                "{d}. The position is below, then the detail behind it. The "
                "full workbook is attached &mdash; twelve tabs, every line "
                "behind these numbers.</div>"
                "</td></tr></table>".format(x=E_TXT,
                                            d=when.strftime("%d %B %Y")))

    body.append(_call_card(parts))
    body.append(_spacer(16))

    body.append(_section("1", "The position",
                         "The six numbers that describe the store today"))
    body.append(_spacer(9))
    body.append(_tiles(parts["tiles"]))
    body.append(_spacer(18))

    #  ---- 2. who is holding gear --------------------------------------
    rows = parts["companies"]
    limit = parts["long_out_days"]
    mx = max([r[1] for r in rows] or [0])
    mx2 = max([r[2] for r in rows] or [0])
    mx3 = max([r[3] for r in rows] or [0])
    mx4 = max([r[4] for r in rows] or [0])
    body.append(_section("2", "Who is holding gear",
                         "Every company on site, biggest holding first"))
    body.append(_table(
        ["Company", "Tooling out", "Tooling issues", "Consumables",
         "Longest out"],
        [[_lead(r[0]),
          _metric(_fmt(r[1]), r[1], mx, "orange"),
          _metric(_fmt(r[2]), r[2], mx2, "slate"),
          _metric(_fmt(r[3]), r[3], mx3, "slate"),
          _metric("{} d".format(_fmt(r[4])), r[4], mx4,
                  "red" if r[4] >= limit else
                  ("amber" if r[4] >= limit * 0.6 else "slate"))]
         for r in rows],
        widths={0: 32, 1: 17, 2: 17, 3: 17, 4: 17},
        aligns={1: "right", 2: "right", 3: "right", 4: "right"},
        note="Tooling out is what is on hire in their name right now - which "
             "is also what has not come back. Longest out is the oldest item "
             "still in their name, in days."))
    body.append(_spacer(18))

    #  ---- 3. working hardest ------------------------------------------
    rows = parts["hardest"]
    mx = max([r[3] for r in rows] or [0])
    body.append(_section("3", "Working hardest",
                         "The gear earning its keep, most turns first"))
    body.append(_table(
        ["Item", "On site", "Out now", "Times issued", "Rating", "Do"],
        [[_lead(r[0]),
          _plain(_fmt(r[1]), "right"),
          _plain(_fmt(r[2]), "right", muted=not r[2]),
          _metric(_fmt(r[3]), r[3], mx, E_TONE.get(r[4], "orange")),
          _pill(r[4], E_TONE.get(r[4], "slate")),
          _plain(r[5], "left", muted=True)] for r in rows],
        widths={0: 33, 1: 9, 2: 9, 3: 16, 4: 15, 5: 18},
        aligns={1: "right", 2: "right", 3: "right"},
        note="Rated on turns - times issued against how many we hold. "
             "Anything on Increase Stock is going out faster than we can "
             "turn it around, and the shift feels that before we do."))
    body.append(_spacer(18))

    #  ---- 4. not moving -----------------------------------------------
    rows = parts["idle"]
    mx = max([r[1] for r in rows] or [0])
    body.append(_section("4", "Not moving",
                         "Biggest holdings that have not gone out once"))
    body.append(_table(
        ["Item", "On site", "Out now", "Do"],
        [[_lead(r[0]),
          _metric(_fmt(r[1]), r[1], mx, "slate"),
          _plain(_fmt(r[2]), "right", muted=not r[2]),
          _pill(r[3], E_TONE.get(r[3], "slate"))] for r in rows],
        widths={0: 52, 1: 16, 2: 12, 3: 20},
        aligns={1: "right", 2: "right"},
        note="Every one of these is shelf space and hire we are carrying "
             "for nothing. Say the word on any line and it goes home."))
    body.append(_spacer(18))

    #  ---- 5. consumables to watch -------------------------------------
    rows = parts["consumables"]
    tight = parts["cons_tight"]
    body.append(_section("5", "Consumables to watch",
                         "The lines running down, tightest first"))
    body.append(_table(
        ["Line", "On shelf", "Gone out", "Used", "Do"],
        [[_lead(r[0]),
          _plain(_fmt(r[1]), "right", muted=not r[1]),
          _plain(_fmt(r[2]), "right"),
          _metric("{:.0f}%".format(round(r[3] * 100)), r[3] * 100, 100,
                  "red" if r[3] >= tight else
                  ("amber" if r[3] >= 0.6 else "green")),
          _pill(r[4], E_TONE.get(r[4], "slate"))] for r in rows],
        widths={0: 40, 1: 12, 2: 12, 3: 16, 4: 20},
        aligns={1: "right", 2: "right", 3: "right"},
        note="Used is what has gone out against the whole position. Past "
             "{:.0f}% we order, because running a line dry mid-shift costs "
             "more than the stock does.".format(tight * 100)))
    body.append(_spacer(18))

    #  ---- 6. out the longest ------------------------------------------
    rows = parts["longest"]
    mx = max([r[5] for r in rows] or [0])
    limit = parts["long_out_days"]
    body.append(_section("6", "Out the longest",
                         "Oldest gear still out, worth a phone call"))
    body.append(_table(
        ["Item", "Company", "Who has it", "On hire since", "Days"],
        [[_lead(r[0], r[1]),
          _plain(r[2]),
          _plain(r[3]),
          _plain(r[4].strftime("%d %b %Y"), "left", muted=True),
          _metric(_fmt(r[5]), r[5], mx,
                  "red" if r[5] >= limit else
                  ("amber" if r[5] >= limit * 0.6 else "slate"))]
         for r in rows],
        widths={0: 30, 1: 20, 2: 20, 3: 16, 4: 14},
        aligns={4: "right"},
        note="Anything past {} days is flagged red. Most of the time it is "
             "still on the job and fine - but it is the list that stops a "
             "demob turning into a search party.".format(limit)))
    body.append(_spacer(18))

    #  ---- 7. everything on hire, company by company --------------------
    limit = parts["long_out_days"]
    body.append(_section("7", "Everything on hire, company by company",
                         "Alphabetical - every item out, and who has it"))
    for block in parts["roll"]:
        body.append(_spacer(10))
        body.append(_company_band(block["company"], block["items"],
                                  len(block["people"]), block["lines"]))
        rows = []
        for p in block["people"]:
            if p["days"] is None:
                story = "{} item{} out. No on-hire date recorded.".format(
                    _fmt(p["n"]), "" if p["n"] == 1 else "s")
            else:
                story = ("{} item{} out, the oldest since {} - {} day{}."
                         .format(_fmt(p["n"]), "" if p["n"] == 1 else "s",
                                 p["since"].strftime("%d %b %Y"), p["days"],
                                 "" if p["days"] == 1 else "s"))
            #  No separate Items or Days column: the story line already
            #  says "3 items out, the oldest since 05 Aug - 5 days", and
            #  a table that prints the same number twice reads like a
            #  mistake. The chip stays, because that is the bit you scan
            #  for rather than read.
            rows.append([
                _lead(p["who"], story) +
                "<div style=\"font-size:1px;line-height:6px\">&nbsp;</div>" +
                _gear_list(p["gear"]),
                _pill("{} d".format(p["days"]) if p["days"] is not None
                      else "no date",
                      "red" if (p["days"] or 0) >= limit else
                      ("amber" if (p["days"] or 0) >= limit * 0.6
                       else "slate")),
            ])
        body.append(_table(["Who has it, and what they've got",
                            "Longest out"],
                           rows, widths={0: 82, 1: 18},
                           aligns={1: "left"}, valign="top"))
    body.append(_spacer(6))
    body.append("<table cellpadding='0' cellspacing='0' border='0' "
                "width='100%' style='border-collapse:collapse;width:100%'>"
                "<tr><td bgcolor=\"{p}\" style=\"background:{p};"
                "font-family:{f};font-size:9pt;color:{m};line-height:1.45\">"
                "That is every item on hire as at the time above - {n} of "
                "them. If a name here has finished up, tell me and I'll "
                "chase the gear back before demob.</td></tr></table>".format(
                    p=E_PAGE, f=E_FONT, m=E_MUTED, n=_fmt(c["onhire"])))
    body.append(_spacer(18))

    body.append(_actions_card(parts))
    body.append(_spacer(16))

    body.append("<table cellpadding='0' cellspacing='0' border='0' "
                "width='100%' style='border-collapse:collapse;width:100%'>"
                "<tr><td style=\"{x}\">"
                "<div style=\"{x};padding-bottom:12px\">Anything here you "
                "want chased, tell me and I'll get on it.</div>"
                "<div style=\"{x}\">{who}<br>"
                "<span style=\"font-family:{f};font-size:10pt;color:{m}\">"
                "Shutdown Tool Store &nbsp;|&nbsp; Coates</span></div>"
                "</td></tr></table>".format(
                    x=E_TXT, f=E_FONT, m=E_MUTED,
                    who=_esc(s.author or "Andrew Fisher")))
    body.append(_spacer(16))

    #  ---- footer -------------------------------------------------------
    body.append(
        "<table cellpadding='0' cellspacing='0' border='0' width='100%' "
        "style='border-collapse:collapse;width:100%'>"
        + _rule(3, E_ORANGE) +
        "<tr><td bgcolor=\"{k}\" style=\"background:{k};padding:12px 16px;"
        "font-family:{f};font-size:9pt;color:#9AA2AC;line-height:1.6\">"
        "Built from the SiteIQ exports of {src}. Figures are the tool "
        "store's own record as at {asat}, not a live feed.<br>"
        "Costs are in the attached workbook, not in this email.<br>"
        "<span style=\"color:#ffffff;font-weight:700\">POWERED BY "
        "<span style=\"color:{o}\">SITEIQ</span></span>"
        " &nbsp;|&nbsp; Author: {who}"
        "</td></tr></table>".format(
            k=E_INK, f=E_FONT, o=E_ORANGE, src=_esc(parts["source_stamp"]),
            asat=asat.strftime("%d %b %Y %H:%M"),
            who=_esc(s.author or "Andrew Fisher")))

    inner = "".join(body)

    #  Word ignores max-width, so Outlook gets a fixed-width table of its
    #  own and everybody else gets the fluid one. Without this the report
    #  stretches the full width of a big monitor and looks like a
    #  spreadsheet somebody pasted in.
    return (
        "<div style=\"background:{p};margin:0;padding:0\">"
        "<table cellpadding='0' cellspacing='0' border='0' width='100%' "
        "bgcolor=\"{p}\" style=\"border-collapse:collapse;width:100%;"
        "background:{p}\"><tr><td align='center' style='padding:16px 10px'>"
        "<!--[if mso]><table cellpadding='0' cellspacing='0' border='0' "
        "width='{w}' align='center'><tr><td width='{w}'><![endif]-->"
        "<table cellpadding='0' cellspacing='0' border='0' width='100%' "
        "style=\"border-collapse:collapse;width:100%;max-width:{w}px;"
        "margin:0 auto;text-align:left\"><tr><td>{inner}</td></tr></table>"
        "<!--[if mso]></td></tr></table><![endif]-->"
        "</td></tr></table></div>".format(p=E_PAGE, w=E_WIDTH, inner=inner))


def email_text(s, when, parts):
    """The plain-text half, for anyone reading on a locked-down phone."""
    c = parts["counts"]
    lines = ["COATES | ON-HIRE REPORT",
             s.header_line,
             "As at {}".format(when.strftime("%d %B %Y")),
             "POWERED BY SITEIQ | Author: {}".format(
                 s.author or "Andrew Fisher"),
             "",
             "Morning,",
             ""]
    lines += [t for t in parts["call"]]
    lines += ["", "THE POSITION", ""]
    for label, value, sub in parts["tiles"]:
        lines.append("  {:<28} {:>7}   {}".format(label, value, sub))
    lines += ["", "EVERYTHING ON HIRE - COMPANY BY COMPANY", ""]
    for block in parts["roll"]:
        lines.append("  {}  ({} items, {} people)".format(
            block["company"], block["items"], len(block["people"])))
        for p in block["people"]:
            when_txt = ("oldest {} ({} days)".format(
                p["since"].strftime("%d %b %Y"), p["days"])
                if p["days"] is not None else "no on-hire date")
            lines.append("    {} - {} items, {}".format(
                p["who"], p["n"], when_txt))
            for desc, qty in p["gear"]:
                lines.append("        {} x {}".format(qty, desc))
        lines.append("")

    lines += ["", "WHAT I NEED FROM YOU", ""]
    for n, (head, bodytext) in enumerate(parts["actions"], 1):
        lines.append("  {}. {} - {}".format(n, head, bodytext))
    lines += ["",
              "The detail is in the attached workbook - twelve tabs.",
              "Costs are in the workbook, not in this email.",
              "",
              "Anything here you want chased, tell me and I'll get on it.",
              "",
              s.author or "Andrew Fisher",
              "Shutdown Tool Store | Coates",
              ""]
    return "\n".join(lines)


def write_email(base, s, when, asat, parts, workbook_path, out_dir):
    """The report as an Outlook draft, with the workbook on the
    paperclip. NOTHING SENDS - it opens as a draft and waits for you."""
    import email.message
    import email.policy

    label = _safe_name(s.short or s.customer or "Shutdown")
    stem = "{} On-Hire Report - {}".format(label, when.strftime("%d %b %Y"))
    body = email_html(s, when, asat, parts)

    #  The same report as a web page, for anybody without Outlook and for
    #  printing. Same HTML, so what you check here is what they get.
    html_path = os.path.join(out_dir, stem + ".html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write("<!DOCTYPE html><html lang='en'><head><meta charset='utf-8'>"
                "<meta name='viewport' content='width=device-width,"
                "initial-scale=1'><title>{}</title></head>"
                "<body style=\"margin:0;padding:0;background:{}\">{}</body>"
                "</html>".format(_esc(stem), E_PAGE, body))

    msg = email.message.EmailMessage(policy=email.policy.SMTP)
    msg["Subject"] = "{} | On-Hire Report | {}".format(
        s.customer or label, when.strftime("%d %B %Y"))
    #  X-Unsent is what makes Outlook open it as a draft you press Send
    #  on yourself, rather than something already gone.
    msg["X-Unsent"] = "1"
    msg.set_content(email_text(s, when, parts))
    msg.add_alternative(body, subtype="html")

    if workbook_path and os.path.isfile(workbook_path):
        try:
            with open(workbook_path, "rb") as f:
                msg.add_attachment(
                    f.read(), maintype="application",
                    subtype=("vnd.openxmlformats-officedocument"
                             ".spreadsheetml.sheet"),
                    filename=os.path.basename(workbook_path))
        except OSError as e:
            say("  WARN | Couldn't attach the workbook to the email ({}). "
                "The email is still there - drag the workbook onto it "
                "before you send.".format(e))

    eml_path = os.path.join(out_dir, stem + ".eml")
    with open(eml_path, "wb") as f:
        f.write(msg.as_bytes())
    return html_path, eml_path


def main(argv):
    when = None
    if len(argv) > 1:
        try:
            when = dt.datetime.strptime(argv[1], "%Y-%m-%d").date()
        except ValueError:
            say("Date must look like 2026-08-10. Leave it off for today.")
            write_log()
            return 1
    try:
        rc = build(HERE, when)
    except KitProblem as problem:
        #  Something the person at the keyboard can fix. The message was
        #  written to be read, so print it and stop - no traceback, no
        #  half a report, and no email draft sitting there looking ready.
        say()
        say("  STOP | {}".format(problem))
        say()
        write_log()
        return 1
    except Exception:
        #  Last line of defence. A raw traceback in a black window at
        #  05:00 tells nobody anything, and the window gets closed before
        #  anyone can read it. Say something useful, keep the detail in
        #  the log, and hand back a failure.
        import traceback
        say()
        say("=" * 66)
        say(" SOMETHING UNEXPECTED WENT WRONG")
        say("=" * 66)
        say()
        say(" The report did NOT finish, and nothing has been sent.")
        say()
        say(" This is not something you can fix from here. Do this:")
        say()
        say("   1. Press 1_CHECK_THIS_LAPTOP.bat")
        say("   2. Email Andrew both files this folder now contains:")
        say("        LAPTOP_CHECK.txt")
        say("        LAST_RUN_LOG.txt")
        say()
        say(" The technical detail is at the bottom of LAST_RUN_LOG.txt.")
        say("=" * 66)
        _LOG.append("")
        _LOG.append("---- technical detail, for Andrew ----")
        _LOG.append(traceback.format_exc())
        write_log()
        return 1
    write_log()
    return rc


if __name__ == "__main__":
    sys.exit(main(sys.argv))
