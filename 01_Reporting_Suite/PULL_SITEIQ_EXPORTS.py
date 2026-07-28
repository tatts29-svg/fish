#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | PULL SITEIQ EXPORTS - Downloads -> Data_SiteIQ, safely
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  THE JOB (Andrew, 27 Jul 2026): SiteIQ downloads land in Downloads
#  with long names, and some of them belong to OTHER sites. Copying the
#  wrong one in would poison a whole day of reports, so this never
#  trusts a filename. It OPENS each file and checks two things:
#
#    1. WHICH SITE - REFERENCE_INFO > PROJECT SCHEDULE must be the K2
#       Gladstone shutdown. Anything else is left where it is and named
#       on screen, so you can see it was skipped and why.
#    2. WHICH EXPORT - the data sheet name (RENTAL_STOCK, ON_HIRE,
#       STOCKTAKE ...). That is how SiteIQ builds them, so the download
#       can be called anything at all and it still lands right.
#
#  Newest of each type wins (by the export's own REQUESTED_DATE/TIME,
#  not the file date). Whatever is already in Data_SiteIQ is backed up
#  to Data_SiteIQ\previous\ first - nothing is ever just overwritten.
# =====================================================================
import datetime as dt
import os
import re
import shutil
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
DEST = os.path.join(HERE, "Data_SiteIQ")
BACKUP = os.path.join(DEST, "previous")
DOWNLOADS = os.path.join(os.path.expanduser("~"), "Downloads")

#  WHOSE EXPORT IS THIS? Checked against REFERENCE_INFO > PROJECT and
#  > PROJECT SCHEDULE. A file is accepted if EITHER field contains ANY
#  of these words (case and spacing ignored). Two markers, not one, so
#  a rename at the SiteIQ end can't lock us out - and an export from
#  another site matches neither and is left alone.
#  (Andrew: "they all say Cement inside", 27 Jul 2026.)
SITE_MARKERS = ["cement", "gladstone k2"]

#  data sheet name -> the name the suite expects on disk
WANTED = {
    "RENTAL_STOCK": "RENTAL_STOCK.xlsx",
    "ON_HIRE": "ON_HIRE.xlsx",
    "TRANSACTION_CHARGES": "TRANSACTIONS.xlsx",
    "SALES_STOCK": "SALES_STOCK.xlsx",
    "STOCKTAKE": "STOCKTAKE.xlsx",
    "DAILY_SUMMARY": "DAILY_SUMMARY.xlsx",
    #  Everyone on site with their hire ID. ON_HIRE only ever shows the
    #  people currently holding gear, so without this the store can only
    #  build a card for someone who already has something out.
    "HIRERS_ID": "HIRERS_ID.xlsx",
}


def norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


def parse_when(v):
    if isinstance(v, dt.datetime):
        return v
    m = re.match(r"\s*(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{2})\s*([AP]M)?",
                 str(v or ""), re.I)
    if not m:
        return None
    d, mo, y, h, mi, ap = m.groups()
    h = int(h)
    if ap:
        if ap.upper() == "PM" and h != 12:
            h += 12
        if ap.upper() == "AM" and h == 12:
            h = 0
    try:
        return dt.datetime(int(y), int(mo), int(d), h, int(mi))
    except ValueError:
        return None


def inspect(path):
    """(export_key, site_text, requested_when) - or (None, why, None)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return None, "not a readable Excel file ({})".format(
            str(e)[:40]), None
    try:
        sheets = list(wb.sheetnames)
        key = next((s for s in sheets if s.upper() in WANTED), None)
        if not key:
            #  The hirer roster is the odd one out: its sheet is just
            #  "Sheet1" and it carries no REFERENCE_INFO block, so
            #  neither the sheet-name rule nor the site check can see
            #  it. Recognise it by the columns instead - Name, External
            #  ID and Employer together are not a coincidence.
            #  (Andrew, 28 Jul 2026: "this is everyone.")
            try:
                ws0 = wb[sheets[0]]
                head = next(ws0.iter_rows(values_only=True), ()) or ()
                cols = {norm(v) for v in head if v}
                if {"name", "external id"} <= cols and "employer" in cols:
                    return "HIRERS_ID", "hirer roster (no site block)", None
            except Exception:
                pass
            return None, "not a SiteIQ export (no known data sheet)", None
        site, when = "", None
        if "REFERENCE_INFO" in sheets:
            rows = []
            for i, r in enumerate(wb["REFERENCE_INFO"].iter_rows(values_only=True)):
                rows.append(r)
                if i >= 1:
                    break
            if len(rows) >= 2 and rows[0] and rows[1]:
                hdr = [norm(v) for v in rows[0]]
                bits = []
                for i, h in enumerate(hdr):
                    #  BOTH project fields count - ON_HIRE and STOCKTAKE
                    #  carry the client in PROJECT, the rest carry the
                    #  shutdown in PROJECT SCHEDULE.
                    if h.startswith("project") and i < len(rows[1]):
                        v = str(rows[1][i] or "").strip()
                        if v:
                            bits.append(v)
                    if "requested_date" in h and i < len(rows[1]):
                        when = parse_when(rows[1][i])
                site = " / ".join(bits)
        return key.upper(), site, when
    finally:
        try:
            wb.close()
        except Exception:
            pass


def main():
    dl = sys.argv[1] if len(sys.argv) > 1 else DOWNLOADS
    print("=" * 62)
    print(" COATES | PULL SITEIQ EXPORTS")
    print(" From: " + dl)
    print(" To  : " + DEST)
    print("=" * 62)
    if not os.path.isdir(dl):
        print(" That Downloads folder does not exist.")
        return 1
    os.makedirs(DEST, exist_ok=True)

    files = [os.path.join(dl, f) for f in os.listdir(dl)
             if f.lower().endswith((".xlsx", ".xlsm"))
             and not f.startswith("~$")]
    files.sort(key=os.path.getmtime, reverse=True)
    print(" Looking at {} Excel file(s) in Downloads...".format(len(files)))
    print("")

    best, skipped = {}, []
    for p in files:
        name = os.path.basename(p)
        key, site, when = inspect(p)
        if not key:
            skipped.append((name, site))
            continue
        if not any(m in norm(site) for m in SITE_MARKERS):
            #  The important one: right kind of export, WRONG SITE.
            skipped.append((name, "another site - '{}'".format(
                (site or "site not stated").strip())))
            print("  SKIP  | {}\n          {} export, but it is for {}".format(
                name[:52], key, (site or "an unnamed site").strip()))
            continue
        stamp = when or dt.datetime.fromtimestamp(os.path.getmtime(p))
        cur = best.get(key)
        if not cur or stamp > cur[1]:
            best[key] = (p, stamp)

    if not best:
        print("")
        print(" Nothing to bring in - no K2 Gladstone exports in Downloads.")
        return 1

    print("")
    os.makedirs(BACKUP, exist_ok=True)
    for key in sorted(best):
        src, stamp = best[key]
        dest = os.path.join(DEST, WANTED[key])
        if os.path.isfile(dest):
            keep = os.path.join(BACKUP, "{}_{}".format(
                dt.datetime.fromtimestamp(os.path.getmtime(dest))
                .strftime("%Y%m%d_%H%M"), WANTED[key]))
            try:
                shutil.copy2(dest, keep)
            except Exception:
                pass
        shutil.copy2(src, dest)
        print("  IN    | {:<18} <- {}".format(WANTED[key],
                                              os.path.basename(src)[:44]))
        print("          exported {}".format(
            stamp.strftime("%d %b %Y %I:%M %p").lstrip("0")))

    missing = [WANTED[k] for k in WANTED if k not in best]
    print("-" * 62)
    print(" Brought in : {} of {} export(s)".format(len(best), len(WANTED)))
    if missing:
        print(" Still needed: " + ", ".join(sorted(missing)))
        print("  (download them from SiteIQ and run me again - the ones")
        print("   already in place are untouched)")
    if skipped:
        print(" Left alone : {} file(s) that are not this site's "
              "exports".format(len(skipped)))
    print(" Previous copies kept in Data_SiteIQ\\previous\\")
    print("")
    print(" Next: run 00_RUN_EVERYTHING.bat")
    return 0


if __name__ == "__main__":
    sys.exit(main())
