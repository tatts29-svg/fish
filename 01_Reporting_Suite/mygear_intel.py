#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | UTILISATION INTELLIGENCE - the engine
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 2 Aug 2026): spread utilisation across equivalent
#  assets, right-size the next shutdown's fleet, and catch incomplete
#  issues before gear disappears.
#
#  This module ONLY computes. The page that draws it is
#  build_utilisation_intel.py. Same split as everywhere else in the
#  suite, so the counter and the manager can never be told different
#  stories by two copies of the same sum.
#
#  IT DOES NOT REPLACE k2_utilisation.py. That module owns the
#  site-wide "what has moved" question and the three-signal definition
#  of used, and this one imports that definition rather than writing a
#  second one. Two numbers both called utilisation is exactly the trap
#  this suite already paid for once.
#
#  ---------------------------------------------------------------
#  THE TWO UTILISATIONS, AND WHY ONE NUMBER LIES
#  ---------------------------------------------------------------
#  COMMERCIAL   - the asset was on charge or on hire to anybody,
#                 including the site holding account.
#  CLIENT-ISSUED - the asset was in the hands of a NAMED hirer at a
#                 company. The holding account is not a person.
#
#  On the 29 Jul pull, 347 charge lines worth $46,558 sat against
#  "Site Plant Equipment - ." - 24% of everything charged, with no
#  downstream allocation recorded anywhere. Report one number and that
#  gear looks fully worked. Report both and the question becomes
#  answerable.
#
#  ---------------------------------------------------------------
#  WHAT A SCAN PROVES, AND WHAT IT DOES NOT
#  ---------------------------------------------------------------
#  A transaction proves an asset was ISSUED. It does not prove anybody
#  picked it up and pulled the trigger. Every label in this module and
#  on the page says "issued" or "deployed" for that reason. Nothing
#  here claims an asset was physically used.
#
#  ---------------------------------------------------------------
#  WHAT THE EXPORTS CANNOT TELL US - stated, never guessed
#  ---------------------------------------------------------------
#  Audited against the live exports on 2 Aug 2026. NONE of these
#  appear in any SiteIQ export this suite receives:
#      * calibration dates, bump tests, inspection or test-and-tag
#        status for an individual asset
#      * reserved / quarantined / out-of-service / under-repair flags
#        (ITEM_STATUS carries five values and none of them are these)
#      * deliberate contingency (ITEM_ON_STANDBY_HIRE is "No" on every
#        row on this job, so the field cannot carry the meaning)
#      * which battery pairs with which radio
#      * whether an accessory or kit is complete
#      * the date an asset arrived on site, or any period it spent out
#        of service
#
#  So this module NEVER says an asset is safe, serviceable, calibrated
#  or complete, and the page says plainly that it cannot see those
#  things. LIMITS below is carried onto the page - it is not decoration
#  and it is not to be quietly dropped to make the screen tidier.
# =====================================================================
import collections
import datetime as dt
import math
import os
import re as _RE

try:
    import k2_utilisation as _K2
    SHUT_START = _K2.SHUT_START
    _sheet = _K2._sheet
    _num = _K2._num
    _date = _K2._date
    _txt = _K2._txt
except Exception:                                    # pragma: no cover
    _K2 = None
    SHUT_START = dt.date(2026, 7, 13)

#  The operational day runs 06:00 to 06:00. A radio booked out at 02:00
#  belongs to the night shift that started the evening before, not to a
#  new day - counting it as a new day would put a phantom peak in every
#  night-shift variant.
DAY_START_HOUR = 6

#  Fleet sizing adds a contingency buffer to PEAK CONCURRENT demand,
#  never to the average. Average utilisation is how a store runs out of
#  what it just sent back. Change this one number to change every
#  recommendation on the page.
CONTINGENCY_PCT = 0.10
CONTINGENCY_MIN = 1

#  SiteIQ's own holding account. Not a person, not a workfront - gear
#  parked against the site rather than issued to anybody.
#
#  THE FULL NAME IS "Site Plant Equipment ." (Andrew, 2 Aug 2026), and
#  SiteIQ stores it two ways: RENTAL_STOCK and ON_HIRE write it as
#  "Site Plant Equipment - ." because they use a First - Last layout,
#  while the transaction sheets flatten it to "Site Plant Equipment .".
#  Both are the same account and both must match. Swept the exports:
#  those are the only two spellings, 506 rows and 361 rows, and the
#  name appears in HIRER_NAME only - never in EMPLOYER_NAME.
#
#  MATCHED EXACTLY, NOT BY PREFIX. Everything in this engine hangs off
#  this one name: match too loosely and a genuine hirer called
#  something like "Site Plant Equipment Hire Pty Ltd" gets counted as
#  the holding account, and all of their days stop being client-issued;
#  match too tightly and the pool itself reads as a customer and
#  inflates the very figure this is here to separate. So it is an exact
#  match on the canonical name, and anything that comes CLOSE without
#  matching is reported rather than silently decided either way.
HOLDING_ACCOUNT = 'site plant equipment .'
HOLDING_NEAR = 'site plant equipment'

#  THE ADMIN ACCOUNT. (Andrew, 2 Aug 2026: "anything in admin is all
#  admin costs".) Stored as "Admin Costs Admin Costs" - 4 lines worth
#  $119,125 on the 29 Jul pull: LABOUR, TRANSPORT-EX, TRAVEL EXPENSES
#  and MISC. Real money, and none of it belongs to a single asset, so
#  it is counted and shown on its own rather than folded into what the
#  fleet earned.
#
#  MATCHED ON THE ACCOUNT, NOT ON THE CATEGORY. Both rules pick the
#  same four lines today, but they are not the same question: a service
#  charge raised against a REAL hirer belongs to that hirer's machine,
#  not to admin, and a category test would quietly send it to admin
#  anyway. The account name is what Andrew actually means. Where the
#  two disagree, the page says so instead of picking a winner.
ADMIN_ACCOUNT = 'admin costs admin costs'
ADMIN_NEAR = 'admin costs'


def _is_admin(name):
    return _norm_hirer(name) == ADMIN_ACCOUNT


def is_near_admin(name):
    n = _norm_hirer(name)
    return n.startswith(ADMIN_NEAR) and n != ADMIN_ACCOUNT


def _norm_hirer(name):
    """SiteIQ's First - Last layout and its flattened twin, as one."""
    s = ' '.join(str(name or '').split())
    return s.replace(' - ', ' ').strip().lower()

#  What "ready to hire" means in ITEM_STATUS terms. Everything else is
#  either out, not here yet, or on its way off site.
READY_STATUS = 'Available for Hire'
OUT_STATUS = 'On Hire'

#  GEAR THAT HAS LEFT. (Andrew, 2 Aug 2026: departed gear shows as
#  pending baseplan, or "could show as pending branch receipt".) It is
#  not part of the fleet any more, so it must not be counted as
#  mobilised - doing that inflates the surplus and would have this page
#  recommending a cut to a fleet that has already been cut.
DEPARTED_STATUS = ('Pending Baseplan', 'Failed Baseplan',
                   'Pending Branch Receipt')

LIMITS = [
    "Issued is not used. A transaction proves an asset was booked out, "
    "not that anybody operated it.",
    "The page cannot see tags, calibration, bump tests, servicing or "
    "whether a kit is complete - none of it is in any SiteIQ export. "
    "Check the item before you issue it.",
    "The page cannot see reserved, quarantined or out-of-service gear "
    "either, so an asset held back on purpose will look idle. If you "
    "keep passing over a tool the page recommends, say why - that is "
    "how the hidden problem gets found.",
    "Utilisation is measured against the days the TRANSACTIONS ACTUALLY "
    "COVER, not the length of the shutdown. A day the export does not "
    "reach is not a day the gear sat idle - it is a day we cannot see.",
    "Arrival dates are not exported, so an asset that landed part way "
    "through the supplied period still reads against all of it.",
    "Revenue is summed from the charge lines. The Daily Summary is a "
    "job total by category and cannot be split per asset, so a lagging "
    "Daily Summary shows here as PENDING, never as zero.",
]


# ------------------------------------------------------------------ #
#  fallbacks, only used if k2_utilisation could not be imported
# ------------------------------------------------------------------ #
if _K2 is None:                                      # pragma: no cover
    def _num(v):
        try:
            return float(str(v).strip())
        except (TypeError, ValueError, AttributeError):
            return 0.0

    def _date(v):
        if isinstance(v, dt.datetime):
            return v.date()
        if isinstance(v, dt.date):
            return v
        s = ('' if v is None else str(v)).strip()
        for f in ('%d/%m/%Y %I:%M %p', '%d/%m/%Y %H:%M', '%d/%m/%Y',
                  '%Y-%m-%d'):
            try:
                return dt.datetime.strptime(s, f).date()
            except ValueError:
                continue
        return None

    def _txt(d, k):
        v = d.get(k)
        return '' if v is None else str(v).strip()

    def _sheet(path, name):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if name not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        try:
            hdr = [str(c) if c is not None else '' for c in next(it)]
        except StopIteration:
            wb.close()
            return []
        out = [dict(zip(hdr, r)) for r in it
               if r and any(c not in (None, '') for c in r)]
        wb.close()
        return out


# ------------------------------------------------------------------ #
#  small helpers
# ------------------------------------------------------------------ #
def _time(v):
    """HH:MM:SS out of SiteIQ, or None. Never guesses a time."""
    s = ('' if v is None else str(v)).strip()
    if not s:
        return None
    if isinstance(v, dt.time):
        return v
    if isinstance(v, dt.datetime):
        return v.time()
    for f in ('%H:%M:%S', '%H:%M', '%I:%M:%S %p', '%I:%M %p'):
        try:
            return dt.datetime.strptime(s, f).time()
        except ValueError:
            continue
    return None


def _stamp(datev, timev):
    """One datetime out of SiteIQ's separate date and time columns."""
    d = _date(datev)
    if not d:
        return None
    t = _time(timev)
    return dt.datetime.combine(d, t or dt.time(0, 0))


def opday(when):
    """The operational day a moment belongs to - 06:00 to 06:00."""
    if when is None:
        return None
    if isinstance(when, dt.date) and not isinstance(when, dt.datetime):
        return when
    return (when.date() - dt.timedelta(days=1)
            if when.hour < DAY_START_HOUR else when.date())


def _is_holding(name):
    return _norm_hirer(name) == HOLDING_ACCOUNT


def is_near_holding(name):
    """Looks like the holding account but is not it. Never guessed."""
    n = _norm_hirer(name)
    return n.startswith(HOLDING_NEAR) and n != HOLDING_ACCOUNT


def _variant(row):
    """SiteIQ spells the column with a trailing space in some exports."""
    return (_txt(row, 'PRODUCT_VARIANT')
            or _txt(row, 'PRODUCT_VARIANT ')
            or '')


#  A UNIT TAG IS NOT A MODEL. (Andrew, 2 Aug 2026: "radios and gas
#  monitors some welders are found in transactions customers contractor
#  equip" - he was right, and this is the second half of using it.)
#
#  CUSTOMER_CONTRACTOR_EQUIP names each radio with its own serial -
#  "Motorola DP4801e Two-Way Radio 871TNK7668" - so pooling on the raw
#  string gives 72 pools of one, which can rank nothing. The serial is
#  stripped so identical handsets pool. Welders carry a plant tag in
#  brackets, "(HGA014)", and go the same way.
#
#  THE DANGER IS OVER-STRIPPING, and it is a real one: "Spanner - Combo
#  (46mm)" through to (27mm) are different spanners, and merging them
#  would put six sizes in one pool and recommend the wrong one. So a
#  trailing token only counts as a serial if it has at least 3 digits
#  AND 2 letters AND is not a unit of measure. Checked across every
#  variant name in all three sheets: 4 merges, all of them genuine
#  (72 radios, 7 welders), and zero collisions among the stock codes,
#  the charged names, or anything carrying a size.
_UNIT_TOKEN = _RE.compile(
    r'^\d+(\.\d+)?\s*(V|A|AH|T|M|MM|CM|KG|KVA|LM|W|KW|HP|L|MTR|IN|FT|G)$',
    _RE.I)
_PAREN_TAIL = _RE.compile(r'\s*\(([A-Za-z0-9]{4,})\)\s*$')
_BARE_TAIL = _RE.compile(r'\s+([A-Za-z0-9]{7,})$')


def _is_serial(tok):
    digits = sum(c.isdigit() for c in tok)
    alpha = sum(c.isalpha() for c in tok)
    return digits >= 3 and alpha >= 2 and not _UNIT_TOKEN.match(tok)


def norm_variant(v):
    """The model, with the individual unit's own tag taken off."""
    s = (v or '').strip()
    m = _PAREN_TAIL.search(s)
    if m and _is_serial(m.group(1)):
        s = s[:m.start()].strip()
    m = _BARE_TAIL.search(s)
    if m and _is_serial(m.group(1)):
        s = s[:m.start()].strip()
    return s


def _pct(a, b):
    return (100.0 * a / b) if b else 0.0


# ------------------------------------------------------------------ #
#  THE READ
# ------------------------------------------------------------------ #
def read(rental_path, txn_path, onhire_path=None, today=None):
    """Everything the intelligence page needs, or None with no stock."""
    if not rental_path or not os.path.isfile(rental_path):
        return None
    today = today or dt.date.today()
    now = dt.datetime.combine(today, dt.time(23, 59))
    stock = _sheet(rental_path, 'RENTAL_STOCK')
    if not stock:
        return None

    # ---------------- the asset register, keyed on ITEM_NUMBER -------
    #  Audited 2 Aug 2026: 5337 stock rows, 5337 distinct item numbers,
    #  no duplicates - so an item number IS one physical asset here and
    #  asset-level rotation is a real question, not an inferred one.
    assets = {}
    dup_items = 0
    for r in stock:
        item = _txt(r, 'ITEM_NUMBER')
        if not item:
            continue
        if item in assets:
            dup_items += 1
            continue
        assets[item] = {
            'item': item,
            'bc': _txt(r, 'ITEM_BARCODE'),
            'desc': _txt(r, 'ITEM_DESCRIPTION'),
            'variant': _variant(r),
            'family': _txt(r, 'PRODUCT_FAMILY') or 'Other',
            'product': _txt(r, 'PRODUCT'),
            'store': _txt(r, 'TOOLSTORE'),
            'unit': _txt(r, 'STORAGE_UNIT') or 'Unassigned',
            #  how the project splits the cost - k2_utilisation owns the
            #  list of plant storage units, so there is one answer
            'plant': bool(_K2 and _K2._is_plant(_txt(r, 'STORAGE_UNIT'))),
            'status': _txt(r, 'ITEM_STATUS'),
            'holder': _txt(r, 'HIRER_NAME'),
            'holderCo': _txt(r, 'COMPANY_NAME'),
            'onHireDate': None,
            'cycles': 0, 'openLines': 0,
            'clientDays': 0.0, 'commercialDays': 0.0, 'holdingDays': 0.0,
            'revenue': 0.0, 'pendingRevenue': False,
            'lastOut': None, 'firstOut': None,
            'issued': False, 'clientIssued': False, 'holdingOnly': False,
            'spans': [],
        }
    by_bc = {a['bc']: a for a in assets.values() if a['bc']}

    # ---------------- current state, straight off ON_HIRE ------------
    #  ON_HIRE is the truth about RIGHT NOW; the transaction sheets run
    #  a day behind billing. Audited: all 860 on-hire item numbers
    #  reconcile to the stock register.
    onhire_rows = []
    if onhire_path and os.path.isfile(onhire_path):
        onhire_rows = _sheet(onhire_path, 'ON_HIRE')
    oh_unmatched = 0
    oh_backfilled = 0
    for r in onhire_rows:
        item = _txt(r, 'ITEM_NUMBER')
        a = assets.get(item)
        if not a:
            oh_unmatched += 1
            continue
        a['onHireDate'] = _date(r.get('START_DATE'))
        a['holder'] = _txt(r, 'HIRER_NAME') or a['holder']
        a['holderCo'] = _txt(r, 'COMPANY') or a['holderCo']
        if not a['variant']:
            v = _variant(r)
            if v:
                a['variant'] = v
                oh_backfilled += 1

    # ---------------- activity, from BOTH transaction sheets ---------
    #  Charged lines AND moved-without-charge lines. Reading only the
    #  charge sheet is what reported the radios at 0% - see the header
    #  of k2_utilisation.py. That lesson applies here unchanged.
    spans = collections.defaultdict(list)   # variant -> (start, end) list
    seen_txn = set()
    tx_rows = 0
    grouped_variants = set()
    #  VARIANT NAMES. Stock spells a variant as a code (RUBCHUTE1M);
    #  the transaction sheets spell the SAME variant as words ("Rubbish
    #  Chute 1M"). Audited 2 Aug 2026: 227 transaction spellings, 791
    #  stock spellings, ZERO in common - so the two can never be joined
    #  on the string, and anything that tries silently produces nothing.
    #  They join through the item number, and that also teaches us the
    #  readable name for each code, which is what the page shows.
    variant_name = {}
    #  MONEY THAT IS NOT AN ASSET. LABOUR, TRANSPORT, TRAVEL EXPENSES
    #  and MISC are booked against "Admin Costs" with no item behind
    #  them - $119,124 of the 29 Jul pull. They are real money and they
    #  are NOT this fleet's earnings, so they are counted separately and
    #  shown, never folded in and never silently dropped.
    money = {'asset': 0.0, 'offRegister': 0.0, 'service': 0.0}
    off_register = {}
    unmatched_lines = 0
    near_names = collections.Counter()
    near_admin = collections.Counter()
    money_split = collections.Counter()

    cap = None

    def _activity(rows, charged):
        nonlocal tx_rows, unmatched_lines
        for r in rows:
            tx_rows += 1
            item = _txt(r, 'SKU/ITEM_NUMBER') or _txt(r, 'ITEM_NUMBER')
            bc = _txt(r, 'LATEST_BARCODE')
            hirer_raw = _txt(r, 'HIRER_NAME')
            a = assets.get(item) or by_bc.get(bc)
            qty = _num(r.get('QUANTITY')) or 1.0
            total = _num(r.get('TOTAL_CHARGE ($)')) if charged else 0.0
            if qty > 1:
                #  A line for 6 barriers is quantity-days, not one asset
                #  going out six times. Recorded so the page can refuse
                #  to pretend it knows which barrier moved.
                grouped_variants.add(a['variant'] if a else _variant(r))
            if not a:
                unmatched_lines += 1
                cat = _txt(r, 'PRODUCT_CATEGORY')
                admin = _is_admin(hirer_raw)
                if admin != (cat == 'Service'):
                    #  the account and the category disagree about what
                    #  this line is. Recorded, never quietly resolved.
                    money_split[(hirer_raw or '(no hirer)', cat)] += total
                #  ADMIN MEANS THE ADMIN ACCOUNT. Andrew's rule, and a
                #  service charge raised against a real hirer is not
                #  admin just because it is a service - it belongs to
                #  that hirer's work, not to the job's overheads. So
                #  only the account decides, and the divergence above
                #  is what gets a human to look.
                if admin:
                    money['service'] += total
                else:
                    #  plant hired outside the store's rental register -
                    #  real gear, real money, just not this fleet
                    money['offRegister'] += total
                    o = off_register.setdefault(item, {
                        'item': item, 'desc': _txt(r, 'SKU/ITEM DESCRIPTION'),
                        'revenue': 0.0, 'lines': 0})
                    o['revenue'] += total
                    o['lines'] += 1
                continue
            money['asset'] += total
            vn = norm_variant(_variant(r))
            if vn and a['variant'] and a['variant'] not in variant_name:
                variant_name[a['variant']] = vn
            start = _stamp(r.get('TRAN_START_DATE'), r.get('TRAN_START_TIME'))
            end = _stamp(r.get('TRAN_END_DATE'), r.get('TRAN_END_TIME'))
            hirer = _txt(r, 'HIRER_NAME')
            holding = _is_holding(hirer)
            if is_near_holding(hirer):
                near_names[hirer] += 1
            if is_near_admin(hirer):
                near_admin[hirer] += 1

            a['issued'] = True
            if holding:
                a['holdingOnly'] = a['holdingOnly'] or not a['clientIssued']
            else:
                a['clientIssued'] = True
                a['holdingOnly'] = False

            #  ONE HIRE = ONE ISSUE-TO-RETURN CYCLE, not one scan. A
            #  line with no end date is still out and is not a
            #  completed cycle yet.
            tid = r.get('TRANSACTION_ID')
            key = (a['item'], tid)
            if end and tid is not None and key not in seen_txn:
                seen_txn.add(key)
                a['cycles'] += 1
            elif not end:
                a['openLines'] += 1

            #  MONEY BEFORE DATES, deliberately. Further down there is a
            #  `continue` that throws away a span whose start is after
            #  its end - nonsense dates, rightly not turned into usage.
            #  While this block sat below it, that same `continue` also
            #  threw away the CHARGE on those lines: $192.04 counted in
            #  the store total but in neither the plant book nor the
            #  tooling book, so the two never summed back to the whole.
            #  A date this suite cannot read is a reason to distrust the
            #  span. It is not a reason to lose the money.
            if charged:
                shifts = _num(r.get('SHIFTS'))
                if total:
                    a['revenue'] += total
                elif shifts:
                    #  Charged shifts recorded but the money has not
                    #  landed yet. PENDING, never $0 - a nought here
                    #  reads as "this earned nothing", which is a
                    #  different and wrong statement.
                    a['pendingRevenue'] = True

            if start:
                #  a span cannot run past the data. Beyond src_to is NO
                #  DATA, not more usage and not idleness.
                if cap and end is None:
                    end = cap
                elif cap and end and end > cap:
                    end = cap
                if floor and start < floor:
                    start = floor
                if end and start > end:
                    continue
                a['firstOut'] = min(a['firstOut'] or start, start)
                a['lastOut'] = max(a['lastOut'] or start, end or start)
                #  COLLECTED, NOT SUMMED. The same physical hire can
                #  appear in both transaction sheets, so adding span
                #  lengths as they arrive double-counts the days - 36
                #  assets came out billed for more days than the export
                #  even covers, which is impossible on its face. The
                #  intervals are merged per asset below, exactly the way
                #  peak concurrent already had to be fixed.
                a['spans'].append((start, end or now, holding))
                #  keyed on the ASSET's variant, never the transaction's
                #  spelling of it - see variant_name above
                if a['variant']:
                    spans[a['variant']].append(
                        (start, end or now, holding, a['item']))

    #  ---- BACKFILL THE MISSING VARIANTS, BEFORE ANYTHING IS POOLED --
    #  (Andrew, 2 Aug 2026: "radios and gas monitors some welders are
    #  found in transactions customers contractor equip".) He is right,
    #  and it matters: RENTAL_STOCK leaves PRODUCT_VARIANT blank on 809
    #  assets - every radio, every gas monitor and 9 welders among them -
    #  but CUSTOMER_CONTRACTOR_EQUIP names the variant on the lines that
    #  moved them. Read it and 293 assets get a pool to be compared in;
    #  skip it and the store's most-chased gear can never be ranked.
    #  The charge sheet adds none of these, checked - it is the movement
    #  sheet that carries them, which is the same reason it exists.
    tc_rows = ec_rows = []
    backfilled = 0
    #  THE SOURCE WINDOW. (Andrew's Flame Off sheet, 2 Aug 2026.) His
    #  rule, and it is a better one than mine was: "NO DATA = outside
    #  supplied transaction period". A day the export does not cover is
    #  not a day the asset sat idle - it is a day we cannot see, and
    #  measuring against it marks gear down for our own reporting gap.
    #
    #  So utilisation is measured against the days the TRANSACTIONS
    #  ACTUALLY COVER, not against the length of the shutdown. His sheet
    #  reads 21 source days off a 13/07-02/08 pull and divides by that;
    #  this engine was dividing by days-since-flame-off, which is a
    #  different and more flattering number every morning.
    src_from = src_to = None
    #  MEASURE FROM MOBILISATION. (Andrew, 2 Aug 2026: "majority of gear
    #  startee going onhire on the 20/07/2026".) Days before the gear
    #  started going out are not idle days - there was nothing to issue
    #  yet - so counting them marks every asset down for the job not
    #  having started. Same principle as his NO DATA rule, at the other
    #  end of the window.
    mob = None
    try:
        import shutdown_day as _SD
        mob = getattr(_SD, 'MOBILISED', None)
    except Exception:
        mob = None
    if txn_path and os.path.isfile(txn_path):
        tc_rows = _sheet(txn_path, 'TRANSACTION_CHARGES')
        ec_rows = _sheet(txn_path, 'CUSTOMER_CONTRACTOR_EQUIP')
        _ds = [_date(r.get(k)) for r in tc_rows + ec_rows
               for k in ('TRAN_START_DATE', 'TRAN_END_DATE')]
        _ds = [x for x in _ds if x]
        if _ds:
            src_from, src_to = min(_ds), max(_ds)
        for r in ec_rows + tc_rows:
            item = _txt(r, 'SKU/ITEM_NUMBER') or _txt(r, 'ITEM_NUMBER')
            a = assets.get(item) or by_bc.get(_txt(r, 'LATEST_BARCODE'))
            if a and not a['variant']:
                v = _variant(r)
                if v:
                    a['variant'] = v
                    backfilled += 1

    #  Every variant, from wherever it came, through the same normaliser
    #  so a model pools as a model - see norm_variant.
    merged_names = collections.defaultdict(set)
    for a in assets.values():
        if a['variant']:
            n = norm_variant(a['variant'])
            if n != a['variant']:
                merged_names[n].add(a['variant'])
            a['variant'] = n

    #  the window everything is measured in: from the later of the
    #  export's first day and mobilisation, to the export's last day
    win_from = src_from
    if src_from and mob and mob > src_from:
        win_from = mob
    win_to = src_to
    if win_to:
        cap = dt.datetime.combine(win_to, dt.time(23, 59, 59))
    floor = (dt.datetime.combine(win_from, dt.time(0, 0))
             if win_from else None)
    if tc_rows or ec_rows:
        _activity(tc_rows, True)
        _activity(ec_rows, False)

    # ---------------- per asset: merge the intervals, then measure ---
    def _merged_days(ivs):
        """Days covered by these intervals, counting overlap once."""
        if not ivs:
            return 0.0
        ivs = sorted(ivs)
        total = 0.0
        cs, ce = ivs[0]
        for st, en in ivs[1:]:
            if st <= ce:
                ce = max(ce, en)
            else:
                total += (ce - cs).total_seconds()
                cs, ce = st, en
        total += (ce - cs).total_seconds()
        return max(0.0, total / 86400.0)

    def _merge(ivs):
        """The same merge, but handing back the intervals themselves.

        _merged_days answers "how many days" and throws the shape away.
        A timeline needs the shape - which days, not how many - and it
        has to be the SAME merge or the chart and the utilisation figure
        will disagree about the same tool on the same day.
        """
        if not ivs:
            return []
        ivs = sorted(ivs)
        out = [list(ivs[0])]
        for st, en in ivs[1:]:
            if st <= out[-1][1]:
                out[-1][1] = max(out[-1][1], en)
            else:
                out.append([st, en])
        return [tuple(x) for x in out]

    for a in assets.values():
        sp = a.pop('spans', [])
        a['commercialDays'] = _merged_days([(s0, e0) for s0, e0, _h in sp])
        a['clientDays'] = _merged_days([(s0, e0) for s0, e0, h in sp if not h])
        #  what was on charge but never in a named hirer's hands
        a['holdingDays'] = max(0.0, a['commercialDays'] - a['clientDays'])
        #  KEPT FOR THE TIMELINE, as dates. Merged, so one asset can
        #  only ever count once on a given day - the same rule that had
        #  to be put on peak concurrent when it read 8 on a fleet of 4.
        a['outSpans'] = [(s0.date(), e0.date()) for s0, e0
                         in _merge([(s0, e0) for s0, e0, _h in sp])]
        a['clientSpans'] = [(s0.date(), e0.date()) for s0, e0
                            in _merge([(s0, e0) for s0, e0, h in sp
                                       if not h])]

    # ---------------- per asset: idle, and the third signal ----------
    for a in assets.values():
        if a['status'] == OUT_STATUS:
            a['issued'] = True
            d = a['onHireDate']
            if d:
                a['lastOut'] = max(a['lastOut'] or now,
                                   dt.datetime.combine(d, dt.time(6, 0)))
            if not _is_holding(a['holder']):
                a['clientIssued'] = True
                a['holdingOnly'] = False
            elif not a['clientIssued']:
                a['holdingOnly'] = True
        a['idleDays'] = (None if a['status'] == OUT_STATUS
                         else ((today - a['lastOut'].date()).days
                               if a['lastOut'] else None))
        a['neverIssued'] = not a['issued']

    days_in = max(1, (today - SHUT_START).days)
    #  the denominator every utilisation figure on this page divides by
    source_days = ((win_to - win_from).days + 1) if (win_from and win_to) else 0
    supplied_days = ((src_to - src_from).days + 1) if (src_from and src_to) else 0

    # ---------------- per variant ------------------------------------
    #  An asset with no PRODUCT_VARIANT cannot be compared with anything,
    #  so it can never be recommended and never sizes a fleet. 670 of
    #  5337 on the 29 Jul pull. They are counted and named as a gap
    #  rather than quietly pooled into a bucket that means nothing.
    variants = {}
    novariant = [a for a in assets.values() if not a['variant']]
    for a in assets.values():
        v = a['variant']
        if not v:
            continue
        s = variants.get(v)
        if s is None:
            s = variants[v] = {
                'variant': v,
                'name': variant_name.get(v, '') or a['desc'],
                'desc': a['desc'], 'family': a['family'],
                'product': a['product'], 'store': a['store'],
                'units': collections.Counter(),
                'assets': 0, 'ready': 0, 'out': 0, 'awaiting': 0,
                'departed': 0,
                'otherStatus': 0, 'issuedOnce': 0, 'neverIssued': 0,
                'clientIssuedOnce': 0, 'holdingOnly': 0,
                'cycles': 0, 'revenue': 0.0, 'pendingRevenue': False,
                'clientDays': 0.0, 'commercialDays': 0.0,
                'holdingDays': 0.0, 'items': [],
            }
        s['assets'] += 1
        s['units'][a['unit']] += 1
        s['items'].append(a['item'])
        if a['status'] == READY_STATUS:
            s['ready'] += 1
        elif a['status'] == OUT_STATUS:
            s['out'] += 1
        elif a['status'] == 'Awaiting Arrival':
            s['awaiting'] += 1
        elif a['status'] in DEPARTED_STATUS:
            s['departed'] += 1
        else:
            s['otherStatus'] += 1
        s['issuedOnce'] += 1 if a['issued'] else 0
        s['neverIssued'] += 0 if a['issued'] else 1
        s['clientIssuedOnce'] += 1 if a['clientIssued'] else 0
        s['holdingOnly'] += 1 if a['holdingOnly'] else 0
        s['cycles'] += a['cycles']
        s['revenue'] += a['revenue']
        s['pendingRevenue'] = s['pendingRevenue'] or a['pendingRevenue']
        s['clientDays'] += a['clientDays']
        s['commercialDays'] += a['commercialDays']
        s['holdingDays'] += a['holdingDays']

    #  PEAK CONCURRENT - the only honest basis for sizing a fleet.
    #  Sweep the issue intervals; the high-water mark is what the job
    #  actually needed at once. Average utilisation would size this
    #  fleet at half what the busiest shift required.
    for v, s in variants.items():
        #  PER ASSET FIRST, THEN SWEEP. One physical hire can appear in
        #  BOTH transaction sheets - charged in one, movement in the
        #  other - and a naive sweep of raw lines counts it twice. That
        #  put a peak of 8 on a fleet of 4, which is not a tight fleet,
        #  it is arithmetic that cannot be true. So each asset's own
        #  intervals are merged into the time it was out, and the sweep
        #  counts ASSETS - one asset can never contribute more than 1.
        per_item = collections.defaultdict(list)
        for start, end, holding, item in spans.get(v, ()):
            per_item[item].append((start, end, holding))
        ev = []
        for item, ivs in per_item.items():
            ivs.sort()
            cs, ce, ch = ivs[0]
            for st, en, hd in ivs[1:]:
                if st <= ce:                      # overlaps or touches
                    ce = max(ce, en)
                    ch = ch and hd               # client contact wins
                else:
                    ev.append((cs, 1, ch))
                    ev.append((ce, -1, ch))
                    cs, ce, ch = st, en, hd
            ev.append((cs, 1, ch))
            ev.append((ce, -1, ch))
        ev.sort(key=lambda x: (x[0], -x[1]))
        cur = peak = 0
        cur_c = peak_c = 0
        peak_days = set()
        for when, delta, holding in ev:
            cur += delta
            if not holding:
                cur_c += delta
            if cur > peak:
                peak = cur
                peak_days = {opday(when)}
            elif cur == peak and peak:
                peak_days.add(opday(when))
            peak_c = max(peak_c, cur_c)
        s['peak'] = peak
        s['peakClient'] = peak_c
        s['peakDays'] = len([d for d in peak_days if d])
        s['grouped'] = v in grouped_variants

        base = max(s['peak'], s['out'])
        buffer_n = max(CONTINGENCY_MIN, int(math.ceil(base * CONTINGENCY_PCT))) \
            if base else 0
        s['contingency'] = buffer_n
        s['recommended'] = base + buffer_n if base else 0
        #  what is actually standing on site: not the ones still
        #  coming, and not the ones that have gone
        mobilised = s['assets'] - s['awaiting'] - s['departed']
        s['mobilised'] = mobilised
        s['surplus'] = max(0, mobilised - s['recommended'])
        s['short'] = max(0, s['recommended'] - mobilised)
        s['clientPct'] = _pct(s['clientDays'],
                              s['commercialDays']) if s['commercialDays'] else 0.0
        s['everIssuedPct'] = _pct(s['issuedOnce'], s['assets'])

        #  CONFIDENCE. Says how much weight the recommendation carries,
        #  because a fleet call made on four days of data is a guess
        #  wearing a number.
        if s['grouped']:
            s['confidence'] = 'low'
            s['why'] = ('lines for more than one item at a time - this is '
                        'quantity-days, not asset rotation')
        elif days_in < 7 or not s['peak']:
            s['confidence'] = 'low'
            s['why'] = '{} day(s) of shutdown so far'.format(days_in)
        elif days_in >= 14 and s['peakDays'] >= 2:
            s['confidence'] = 'high'
            s['why'] = 'peak reached on {} separate days'.format(s['peakDays'])
        else:
            s['confidence'] = 'medium'
            s['why'] = 'peak reached once, {} days in'.format(days_in)

        #  THE CALL, in the words a decision gets made in.
        #
        #  BRING MORE needs EVIDENCE OF UNMET DEMAND, not just a fleet
        #  that happens to be fully out. The first cut called it on any
        #  variant with everything on hire, which fired on 154 variants -
        #  mostly one-off specialist items where "all of it is out" means
        #  the single one we own is in use, not that the job wanted two.
        #  So: the whole fleet has to have been out AT ONCE, and it has
        #  to be a fleet rather than a one-off. Anything smaller is
        #  reported as an observation, not a recommendation.
        fully = s['peak'] >= mobilised and mobilised > 0
        if not s['issuedOnce']:
            #  Says what was OBSERVED, not what to do. "Reduce to nought"
            #  off a fleet nobody has touched is a recommendation this
            #  page has not earned - it might be contingency, and the
            #  exports cannot show contingency.
            s['call'] = 'NONE ISSUED YET'
        elif s['short'] and fully and mobilised >= 3:
            s['call'] = 'BRING MORE'
        elif fully:
            s['call'] = 'FULLY DEPLOYED'
        elif s['surplus'] >= max(2, int(mobilised * 0.25)):
            s['call'] = 'REDUCE NEXT TIME'
        elif s['neverIssued'] and s['neverIssued'] >= max(2, int(
                s['assets'] * 0.5)) and s['assets'] > 2:
            s['call'] = 'ROTATE STOCK'
        elif s['surplus']:
            s['call'] = 'TRIM'
        else:
            s['call'] = 'RIGHT-SIZED'

    variant_rows = sorted(variants.values(),
                          key=lambda x: (-x['assets'], x['variant']))
    for s in variant_rows:
        s['units'] = s['units'].most_common(3)
        s.pop('items', None)

    # ---------------- totals -----------------------------------------
    tot = {
        'assets': len(assets),
        'ready': sum(1 for a in assets.values() if a['status'] == READY_STATUS),
        'out': sum(1 for a in assets.values() if a['status'] == OUT_STATUS),
        'awaiting': sum(1 for a in assets.values()
                        if a['status'] == 'Awaiting Arrival'),
        'departed': sum(1 for a in assets.values()
                        if a['status'] in DEPARTED_STATUS),
        'issuedOnce': sum(1 for a in assets.values() if a['issued']),
        'neverIssued': sum(1 for a in assets.values() if not a['issued']),
        'clientIssuedOnce': sum(1 for a in assets.values() if a['clientIssued']),
        'holdingOnly': sum(1 for a in assets.values() if a['holdingOnly']),
        'cycles': sum(a['cycles'] for a in assets.values()),
        'revenue': money['asset'],
        'revenueOffRegister': money['offRegister'],
        'revenueService': money['service'],
        'revenueAll': money['asset'] + money['offRegister'] + money['service'],
        'noVariant': len(novariant),
        'clientDays': sum(a['clientDays'] for a in assets.values()),
        'commercialDays': sum(a['commercialDays'] for a in assets.values()),
        'holdingDays': sum(a['holdingDays'] for a in assets.values()),
        'variants': len(variant_rows),
    }
    #  PLANT and TOOLING as separate books, because the project pays
    #  for them separately. Tooling comes off the shutdown and Andrew's
    #  toolstores; plant comes from wherever it can be got and is
    #  managed on site - one average across the two answers nobody.
    for lab, want in (('plant', True), ('tooling', False)):
        grp = [a for a in assets.values() if a['plant'] is want]
        cd = sum(a['clientDays'] for a in grp)
        md = sum(a['commercialDays'] for a in grp)
        tot[lab] = {
            'assets': len(grp),
            'out': sum(1 for a in grp if a['status'] == OUT_STATUS),
            'ready': sum(1 for a in grp if a['status'] == READY_STATUS),
            'neverIssued': sum(1 for a in grp if not a['issued']),
            'holdingOnly': sum(1 for a in grp if a['holdingOnly']),
            'clientDays': cd, 'commercialDays': md,
            'revenue': sum(a['revenue'] for a in grp),
            'clientPct': _pct(cd, md),
        }
    #  THE TWO BOOKS MUST ADD BACK TO THE ONE TOTAL. Every asset is
    #  either plant or it is tooling, so plant + tooling has to equal
    #  the store's charged revenue to the cent. When it did not, the
    #  difference was money silently dropped by a data-quality skip -
    #  which read as if the job had spent less. This carries the gap so
    #  a page can refuse to quote a split that does not reconcile,
    #  rather than showing two numbers that quietly do not meet.
    tot['splitGap'] = round(
        tot['revenue'] - tot['plant']['revenue'] - tot['tooling']['revenue'],
        2)
    tot['clientPct'] = _pct(tot['clientDays'], tot['commercialDays'])
    tot['everIssuedPct'] = _pct(tot['issuedOnce'], tot['assets'])
    tot['neverPct'] = _pct(tot['neverIssued'], tot['assets'])

    return {
        'asof': today.isoformat(),
        'shutStart': SHUT_START.isoformat(),
        'daysIn': days_in,
        'sourceFrom': win_from.isoformat() if win_from else '',
        'sourceTo': win_to.isoformat() if win_to else '',
        'sourceDays': source_days,
        'suppliedFrom': src_from.isoformat() if src_from else '',
        'suppliedTo': src_to.isoformat() if src_to else '',
        'suppliedDays': supplied_days,
        'mobilised': mob.isoformat() if mob else '',
        'dayStartHour': DAY_START_HOUR,
        'contingencyPct': CONTINGENCY_PCT,
        'totals': tot,
        'variants': variant_rows,
        'assets': assets,
        'limits': list(LIMITS),
        'offRegister': sorted(off_register.values(),
                              key=lambda x: -x['revenue']),
        'join': {
            'stockRows': len(stock),
            'assets': len(assets),
            'duplicateItemNumbers': dup_items,
            'onHireRows': len(onhire_rows),
            'onHireUnmatched': oh_unmatched,
            'txnRows': tx_rows,
            'txnUnmatched': unmatched_lines,
            'groupedVariants': len(grouped_variants),
            'noVariantAssets': len(novariant),
            'variantNamesLearned': len(variant_name),
            'variantsBackfilled': backfilled + oh_backfilled,
            'variantsFromTransactions': backfilled,
            'variantsFromOnHire': oh_backfilled,
            'plantUnits': list(getattr(_K2, 'PLANT_UNITS', ())),
            'unknownUnits': sorted({a['unit'] for a in assets.values()
                                    if _K2 and a['unit']
                                    and not _K2.is_known_unit(a['unit'])}),
            'nearHoldingNames': near_names.most_common(),
            'nearAdminNames': near_admin.most_common(),
            'accountVsCategory': [(h, c, round(v, 2))
                                  for (h, c), v in money_split.most_common()],
            'variantsMerged': sum(len(v) for v in merged_names.values()),
            'variantMergeGroups': len(merged_names),
        },
    }


# ------------------------------------------------------------------ #
#  NEXT TOOL UP - which one of these do I hand over?
# ------------------------------------------------------------------ #
#  Eligibility first, ranking second. An asset that is not ready to
#  hire is not a candidate at any score.
#
#  The page CANNOT check tags, calibration or completeness - see LIMITS
#  - so eligibility here means "SiteIQ says it is on the shelf and free",
#  and the recommendation says so in those words. It never says safe.
def rank(data, variant, store=None, limit=5):
    """The assets to issue next in this variant, best first."""
    if not data:
        return []
    out = []
    for a in data['assets'].values():
        if (a['variant'] or '') != variant:
            continue
        if a['status'] != READY_STATUS:
            continue
        if store and a['store'] and a['store'] != store:
            continue
        out.append(a)
    #  never issued, then least client time, then fewest completed
    #  cycles, then longest since it last went out, then lowest revenue.
    #  Revenue is LAST and is shown but never leads - the Daily Summary
    #  lags, and a lagging number must not pick the tool.
    out.sort(key=lambda a: (
        0 if a['neverIssued'] else 1,
        round(a['clientDays'], 3),
        a['cycles'],
        -(a['idleDays'] if a['idleDays'] is not None else 10 ** 6),
        round(a['revenue'], 2),
        a['item'],
    ))
    return out[:limit]


# ------------------------------------------------------------------ #
#  READY FLEETS - a radio without a battery is not a radio
# ------------------------------------------------------------------ #
#  SiteIQ has no compatibility table, so "compatible" here means what
#  the DESCRIPTIONS actually say and nothing more:
#      Motorola DP4801e Two-Way Radio  <->  Motorola NNTN8129 IMPRES
#  That is a real statement off the register (72 radios, 104 IMPRES
#  batteries on the 29 Jul pull), not an invented pairing. The Milwaukee
#  18V tool batteries are deliberately NOT in it - they charge tools,
#  not handsets, and folding them in would triple the battery count and
#  make the ready fleet look healthy when it is not.
#
#  Radio Cover is excluded from the radio count for the same reason: it
#  is an accessory, and counting 37 covers as 37 radios would be a lie
#  with a straight face.
KITS = [
    {'key': 'radio', 'name': 'Two-way radios',
     'item': _RE.compile(r'two-?way\s+radio', _RE.I),
     'itemNot': _RE.compile(r'cover|charger|antenna', _RE.I),
     'mate': _RE.compile(r'impres\s+battery|nntn', _RE.I),
     'mateName': 'IMPRES battery',
     'unit': 'a radio and a battery'},
    {'key': 'gas', 'name': 'Gas monitors',
     'item': _RE.compile(r'gas\s+(?:monitor|detector)|gasalert|microclip'
                         r'|altair|multi-?gas', _RE.I),
     'itemNot': _RE.compile(r'charger|cradle|dock', _RE.I),
     'mate': None, 'mateName': '', 'unit': 'a monitor'},
]


def _kind(kit, desc):
    if kit['item'].search(desc) and not kit['itemNot'].search(desc):
        return 'item'
    if kit['mate'] is not None and kit['mate'].search(desc):
        return 'mate'
    return None


def ready_fleets(data):
    """Complete usable sets, not raw counts, plus who holds a half one."""
    if not data:
        return []
    out = []
    for kit in KITS:
        items = {'ready': 0, 'out': 0, 'total': 0}
        mates = {'ready': 0, 'out': 0, 'total': 0}
        holders = collections.defaultdict(lambda: {'item': 0, 'mate': 0,
                                                   'co': ''})
        for a in data['assets'].values():
            k = _kind(kit, a['desc'])
            if not k:
                continue
            bucket = items if k == 'item' else mates
            bucket['total'] += 1
            if a['status'] == READY_STATUS:
                bucket['ready'] += 1
            elif a['status'] == OUT_STATUS:
                bucket['out'] += 1
                who = a['holder'] or '(not recorded)'
                holders[who][k] += 1
                holders[who]['co'] = a['holderCo']
        #  THE READY FLEET is the smaller of the two halves. 78 radios
        #  on the shelf with 12 batteries is a fleet of 12.
        if kit['mate'] is not None:
            ready = min(items['ready'], mates['ready'])
            limiter = ('batteries' if mates['ready'] < items['ready']
                       else ('handsets' if items['ready'] < mates['ready']
                             else 'balanced'))
        else:
            ready = items['ready']
            limiter = ''
        pairs = []
        for who, h in holders.items():
            if _is_holding(who):
                continue
            if kit['mate'] is None:
                continue
            if h['item'] and not h['mate']:
                pairs.append({'who': who, 'co': h['co'], 'kind': 'no-mate',
                              'n': h['item']})
            elif h['mate'] and not h['item']:
                pairs.append({'who': who, 'co': h['co'], 'kind': 'no-item',
                              'n': h['mate']})
        pairs.sort(key=lambda x: (-x['n'], x['who']))
        out.append({
            'key': kit['key'], 'name': kit['name'], 'unit': kit['unit'],
            'items': items, 'mates': mates, 'mateName': kit['mateName'],
            'readySets': ready, 'limiter': limiter, 'pairs': pairs,
        })
    return out


# ------------------------------------------------------------------ #
#  ISSUE INTEGRITY - flags that are jobs, not red lights
# ------------------------------------------------------------------ #
#  Only checks the exports can actually answer. There is deliberately
#  NO calibration, tag or service check here - see LIMITS. A flag this
#  page cannot evidence is a flag it does not raise.
def integrity(data, fleets=None, rotate_min=6):
    """Things worth walking somewhere about, worst first."""
    if not data:
        return []
    jobs = []
    fleets = fleets if fleets is not None else ready_fleets(data)

    for f in fleets:
        for p in f['pairs']:
            if p['kind'] == 'no-mate':
                jobs.append({
                    'sev': 'act', 'job': 'COMPLETE THE KIT',
                    'what': '{} holds {} {} and no {} is recorded'.format(
                        p['who'], p['n'], f['name'].lower(), f['mateName']),
                    'who': p['who'], 'co': p['co'],
                    'do': 'Check whether one went out unscanned, or get one '
                          'to them.'})
            else:
                jobs.append({
                    'sev': 'look', 'job': 'INVESTIGATE',
                    'what': '{} holds {} {} with no handset'.format(
                        p['who'], p['n'], f['mateName']),
                    'who': p['who'], 'co': p['co'],
                    'do': 'Spare or charging stock is fine - if it is not, '
                          'the handset is unaccounted for.'})

    #  ROTATE STOCK. One tool flogged while its equivalents sit.
    for s in data['variants']:
        if s['assets'] < 3 or s['grouped'] or not s['cycles']:
            continue
        if s['neverIssued'] >= rotate_min and s['cycles'] >= rotate_min:
            jobs.append({
                'sev': 'rotate', 'job': 'ROTATE STOCK',
                'what': '{} - {} hire cycles across the pool while {} of {} '
                        'have never gone out'.format(
                            s['name'] or s['variant'], s['cycles'],
                            s['neverIssued'], s['assets']),
                'who': '', 'co': '',
                'do': 'Issue the untouched ones next - open Next Tool Up on '
                      'this variant.'})

    #  CHARGED, NOTHING ISSUED DOWNSTREAM. The holding-account money.
    hold = [a for a in data['assets'].values() if a['holdingOnly']]
    if hold:
        rev = sum(a['revenue'] for a in hold)
        jobs.append({
            'sev': 'money', 'job': 'REVIEW THE CHARGE',
            'what': '{} asset(s) have only ever been booked to the site '
                    'holding account{}'.format(
                        len(hold),
                        ' - ${:,.0f} charged'.format(rev) if rev else ''),
            'who': '', 'co': '',
            'do': 'Charged with no downstream allocation recorded. Either '
                  'the allocation was never scanned or the charge needs a '
                  'look.'})

    #  ON HIRE AND ON THE SHELF AT THE SAME TIME. Cannot both be true.
    both = [a for a in data['assets'].values()
            if a['status'] == READY_STATUS and a['onHireDate']]
    if both:
        jobs.append({
            'sev': 'act', 'job': 'CORRECT THE LOCATION',
            'what': '{} asset(s) read Available for Hire in the stock '
                    'register and On Hire in the on-hire export'.format(
                        len(both)),
            'who': '', 'co': '',
            'do': 'One of the two is stale. Check the item, then fix it in '
                  'SiteIQ - the page only reports what it was given.'})

    order = {'act': 0, 'money': 1, 'rotate': 2, 'look': 3}
    jobs.sort(key=lambda j: order.get(j['sev'], 9))
    return jobs


def variant_of(data, needle):
    """Find variants by item number, barcode, variant code or words."""
    if not data:
        return []
    q = str(needle or '').strip().lower()
    if not q:
        return []
    hits = collections.OrderedDict()
    for a in data['assets'].values():
        if not a['variant']:
            continue
        if (q == a['item'].lower() or q == (a['bc'] or '').lower()
                or q in a['variant'].lower() or q in a['desc'].lower()):
            hits.setdefault(a['variant'], 0)
            hits[a['variant']] += 1
    return list(hits.items())
