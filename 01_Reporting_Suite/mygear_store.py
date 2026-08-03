#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | MY GEAR - WHAT'S IN THE STORE
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 29 Jul 2026): "a what's available in the tool store -
#  consumables and rental - a really defined list of what is available
#  for hire. saves them waiting and then being told no. so easy to look
#  for something a baby can do it and find it."
#
#  The honest shape of it:
#
#  AVAILABLE means SiteIQ's own word. RENTAL_STOCK carries ITEM_STATUS,
#  and "Available for Hire" is SiteIQ saying the item is on the shelf -
#  we never infer it from a missing hirer name, and we never estimate.
#
#  A ZERO IS AS USEFUL AS A NUMBER. "None right now" saves the same walk
#  as "yes, four of them" - so out-of-stock lines are shown, not hidden.
#
#  IT IS A MORNING SNAPSHOT, NOT A LIVE FEED. The page is built once a
#  day from that morning's export. If a bloke reads "4 available" at ten
#  o'clock and three went at eight, we have made his day worse and he
#  will never trust the page again. So every line carries the time it
#  was true.
#
#  AND HE IS STANDING RIGHT THERE. The page is only reachable on the
#  store's own Wi-Fi - a bloke reads it at the front of the store, not
#  from the other side of site. So "ring us and we'll put one aside" is
#  daft advice: the counter is two metres away. The wording sends him
#  to the window, which is where he already is.
#  (Andrew, 29 Jul 2026: "people can only view this info from out the
#  front. so saying ring the store is wrong.")
#
#  BROWSE THE WAY THE STORE IS LAID OUT. SiteIQ's STORAGE_UNIT is where
#  the thing physically lives - Tooling, Electrical, Rigging, Welding,
#  Air, Radios. Those are the aisles, so those are the buttons. Nobody
#  has to learn a menu; they walk the store with their thumb.
# =====================================================================
import os
import re

#  The permanent do-not-show list. Optional on purpose: a suite missing
#  the module still builds and simply hides nothing.
try:
    import hidden_stock as _HID
except Exception:
    _HID = None

#  Quantities come out of SALES_STOCK as TEXT ("10", "0"), not numbers.
#  Read as numbers they silently become nothing, and every consumable
#  reads "none left" - which is the exact wrong answer, and a confident
#  one. (Found 29 Jul 2026.)
def _qty(v):
    try:
        return int(float(str(v).strip().replace(',', '')))
    except (TypeError, ValueError):
        return None


AVAILABLE = 'available for hire'

#  THE AISLES - named the way a bloke asks for the thing.
#
#  SiteIQ's own taxonomy is no use for browsing: 1,888 of the 2,230
#  items in the Tooling unit are all filed as "Industrial Tools &
#  Equipment", so its categories collapse into one giant bucket. The
#  real signal is in the item names, and those are already written in
#  crew language - Socket, Spanner, Grinder, Chain Block, Extension
#  Lead. So the categories are mined from the names, first rule wins.
#
#  ORDER MATTERS. "Welding Lead" has to reach WELDING before "Lead"
#  reaches LEADS & POWER, and "Tool Lanyard" has to reach its own
#  category before "Tool" reaches HAND TOOLS. Read the list as a
#  sequence of questions, not a set. (29 Jul 2026)
CATS = [
    ('Sockets',        'M12 2a10 10 0 1 0 0 20 10 10 0 0 0 0-20zm0 5a5 5 0 1 1 0 10 5 5 0 0 1 0-10z',
     ('socket', 'ratchet ring', 'ratchet spanner', 'torque bar',
      'crowsfeet', 'crows feet', 'drive adaptor', 'extension bar',
      'universal joint', 'wobble')),
    ('Spanners',       'M14.7 6.3a4 4 0 0 0-5.4 5.4L3 18l3 3 6.3-6.3a4 4 0 0 0 5.4-5.4l-2.6 2.6-2.4-2.4z',
     ('spanner', 'wrench', 'allen key', 'hex key', 'podger')),
    ('Tool lanyards',  'M12 3v7M9 10h6l-1 11h-4z',
     ('lanyard', 'tool tether', 'tether')),
    ('Lifting & rigging', 'M12 3v6M12 9l-5 8h10l-5-8zM5 21h14',
     ('sling', 'shackle', 'bow ', 'chain block', 'lever block', 'cumalong',
      'come along', 'eye bolt', 'eyebolt', 'turnbuckle', 'hoist', 'winch',
      'strop', 'safety anchor', 'load bind', 'round ,', 'plate clamp',
      'beam clamp', 'girder', 'spreader', 'lifting', 'rope', 'webbing')),
    ('Power tools',    'M10 4h6l4 4-4 4h-6l-2-4zM8 8H3M8 8l-4 8h6',
     ('impact wrench', 'rattle gun', 'drill', 'saw', 'nut runner',
      'multi tool', 'sander', 'planer', 'router', 'nibbler', 'shear',
      'power tool', 'jigsaw', 'recip', 'annular', 'magnetic base',
      'mag drill', 'plasma cutter', 'tap and die', 'reamer')),
    ('Grinding & cutting', 'M6 18a5 5 0 1 0 0-10 5 5 0 0 0 0 10zM11 13h10M17 9l4 4-4 4',
     ('grinder', 'bi metal', 'hole saw', 'cut off', 'cutoff', 'abrasive',
      'disc', 'burr', 'chop saw', 'band saw')),
    ('Hand tools',     'M4 20l8-8M14 6l4 4M12 4l8 8-4 4-8-8z',
     ('plier', 'hammer', 'screwdriver', 'drift pin', 'chisel', 'file ',
      'punch', 'crowbar', 'pry bar', 'knife', 'snip', 'vice', 'clamp',
      'mallet', 'saw horse', 'hacksaw', 'brush', 'scraper', 'shovel',
      'broom', 'bucket', 'wheelbarrow', 'wheel barrow', 'screw driver',
      'podge', 'stilsen', 'stillson', 'bolt cutter', 'crow bar',
      'wedge', 'pick up tool', 'grease gun', 'oil can', 'funnel',
      'tin snip', 'rivet', 'stapler', 'sledge')),
    ('Leads & power',  'M9 3v5M15 3v5M7 8h10v3a5 5 0 0 1-5 5 5 5 0 0 1-5-5V8zM12 16v5',
     ('extension lead', 'lead ext', 'distribution board', 'earth leakage',
      'generator', 'transformer', 'power board', 'rcd', 'switchboard',
      'cable', 'inverter', 'adaptor', 'adapter', 'ceeform', 'clipsal',
      'plug', 'socket outlet')),
    ('Air & hoses',    'M4 12h10a3 3 0 1 0-3-3M4 17h8a2.5 2.5 0 1 1-2.5 2.5',
     ('air hose', 'hose whip', 'air ', 'compressor', 'fitting', 'coupling',
      'regulator', 'blow gun', 'hose', 'airline', 'air line',
      'y piece', 'manifold')),
    ('Welding',        'M3 15h7l3-9 3 6h5M6 21h12',
     ('weld', 'argon', 'electrode', 'oxy', 'acetylene', 'gouging',
      'purge', 'tig', 'mig')),
    ('Lighting',       'M9 21h6M10 17h4a5 5 0 1 0-4 0zM12 3v2',
     ('light', 'lamp', 'flood', 'torch')),
    ('Fans & ventilation', 'M12 12a4 4 0 0 0 4-4 4 4 0 0 0-8 0 4 4 0 0 0 4 4zm0 0a4 4 0 0 1 4 4 4 4 0 0 1-8 0 4 4 0 0 1 4-4z',
     ('fan', 'ducting', 'blower', 'ventilat', 'extract')),
    ('Hydraulics',     'M4 10h9l3-4h4v12h-4l-3-4H4zM7 10v4',
     ('hydraulic', 'porta power', 'jack', 'torque wrench', 'tensioner',
      'hi torque', 'ram ')),
    ('Radios',         'RECT:8,7,8,14,2|M12 7V2M12 2l3 2M10.5 11h3M10.5 14h3',
     ('motorola', 'radio', 'impres', 'aerial', 'antenna', 'earpiece')),
    ('Gas monitors',   'RECT:7,5,10,16,2.5|M9.5 2.5h5',
     ('honeywell bw', 'gas det', 'gas mon', 'flex4', 'clip4', 'calibration gas')),
    ('Batteries & chargers', 'RECT:6,4,12,17,2|M10 2h4M13 9l-3 4h4l-3 4',
     ('battery', 'batt ', 'charger')),
    ('Measuring',      'M2 9h20v6H2zM6 9v3M10 9v4M14 9v3M18 9v4',
     ('tape measure', 'measur', 'level', 'square', 'caliper', 'gauge',
      'thermometer', 'multimeter', 'laser', 'dial indicator', 'divider',
      'steel rule', 'straight edge', 'stamp', 'marker', 'scriber',
      'protractor', 'feeler')),
    ('Safety gear',    'M12 3l7 3v6c0 4-3 7-7 9-4-2-7-5-7-9V6z',
     ('safety', 'harness', 'fall arrest', 'barrier', 'bunting', 'sign',
      'first aid', 'eyewash', 'eye wash', 'diphoterine', 'spill',
      'extinguisher', 'respirator', 'ear muff', 'glove', 'overall',
      'goggle', 'visor', 'helmet', 'vest', 'wipes', 'cloth', 'rag',
      'cleaner', 'degreaser', 'absorbent')),
    ('Ladders & access', 'M7 3v18M17 3v18M7 7h10M7 12h10M7 17h10',
     ('ladder', 'platform', 'trestle', 'scaffold', 'step ', 'staging')),
    ('Pumps & cleaning', 'M5 20h14M8 20v-6a4 4 0 0 1 8 0v6M12 4v6',
     ('pump', 'vacuum', 'vac ', 'pressure clean', 'hoover', 'blaster')),
]

#  Whatever no rule claims. Kept as a real category rather than hidden -
#  a bloke can still browse it, and a fat "Other" is the signal that a
#  rule is missing.
OTHER = ('Other gear', 'M5 8h14v11H5zM9 8V5h6v3')


def _cat_of(name):
    n = ' ' + (name or '').lower() + ' '
    for cat, _p, keys in CATS:
        for k in keys:
            if k in n:
                return cat
    return OTHER[0]


#  Gear that must never be offered. SiteIQ keeps retired assets on the
#  register with the reason written into the description; a catalogue
#  that offers a bloke an OBSOLETE socket has cost him the walk it was
#  meant to save. (29 Jul 2026)
#  Each marker must be unambiguous. A bare 'SCRAP' would quietly hide
#  every SCRAPER in the store, and a catalogue that silently drops real
#  gear is worse than one that shows a retired socket.
NEVER_OFFER = ('OBSOLETE', 'DO NOT USE', 'DO NOT HIRE', 'SCRAPPED',
               'WRITTEN OFF', 'QUARANTINE', 'FAULTY', 'OUT OF SERVICE')


def _offerable(desc):
    u = (desc or '').upper()
    return not any(m in u for m in NEVER_OFFER)


def _icon(spec, size=22):
    body = ''
    for part in spec.split('|'):
        if part.startswith('RECT:'):
            x, y, w, h, r = part[5:].split(',')
            body += ("<rect x='{}' y='{}' width='{}' height='{}' rx='{}'/>"
                     .format(x, y, w, h, r))
        else:
            body += "<path d='{}'/>".format(part)
    return ("<svg viewBox='0 0 24 24' width='{s}' height='{s}' fill='none' "
            "stroke='currentColor' stroke-width='1.8' stroke-linecap='round' "
            "stroke-linejoin='round'>{b}</svg>").format(s=size, b=body)


def _unit_of(raw):
    """Where the thing physically lives - the aisle you walk to.

    This is no longer how the catalogue is BROWSED (the crew-language
    categories above do that); it is the direction you give a bloke
    once he has found it. So it only has to be tidy and honest.
    """
    s = (raw or '').strip()
    if not s:
        return 'Ask at the counter'
    low = s.lower()
    if 'hi torque' in low or 'hydraulic' in low:
        return 'Hydraulics'
    if 'cement' in low or 'site plant' in low or 'own equipment' in low:
        #  the client's own gear the store looks after - named honestly
        #  so nobody mistakes it for Coates hire stock
        return 'Cement own gear'
    if 'consumable' in low:
        return 'Consumables'
    #  SiteIQ writes them plainly already (Tooling, Electrical, Rigging,
    #  Welding, Air, Radios, Laydown, Gas Monitors) - keep its word
    return s.split('-')[0].strip().title()


def _tidy(desc, master=None, item_number=''):
    """The plain-English name if the master file has one, else SiteIQ's.

    The master is keyed on ITEM_NUMBER, so the number has to be handed
    over - without it every lookup misses and the crews get SiteIQ's
    raw wording instead of the 4,400-odd names Andrew has renamed."""
    d = (desc or '').strip()
    if master is not None:
        try:
            d = master.disp(item_number, d) or d
        except Exception:
            pass
    #  SiteIQ writes consumables in SHOUTING CAPS with the size trailing
    #  after a dash. Title-case reads better on a phone; the detail is
    #  kept, never trimmed away.
    if d.isupper() and len(d) > 6:
        d = d.title().replace(' - ', ' - ')
    return re.sub(r'\s{2,}', ' ', d).strip()


# ---------------------------------------------------------------------
#  READ - what is actually on the shelf this morning
# ---------------------------------------------------------------------
def read_availability(rental_path, sales_path, master=None):
    """Returns {'hire': [...], 'cons': [...], 'stats': {...}}.

    Every line: name, how many, which aisle. Nothing inferred."""
    import openpyxl
    hire, cons = {}, {}
    #  how many lines the do-not-show list took out this run - reported,
    #  never silent
    _hid_n = [0]
    #  and how many the register itself marks unofferable. Five assets
    #  on the 3 Aug pull read "Available for Hire" in SiteIQ while
    #  carrying **OBSOLETE** in the description, so the shelf count
    #  came out 4,208 against a register saying 4,213 and said nothing
    #  about the difference. Correct to exclude them; wrong to be quiet
    #  about it - a gap nobody explains is the one somebody re-derives
    #  by hand at the worst moment (3 Aug 2026).
    _obs_n = [0]
    #  the photo key: PRODUCT_VARIANT for hire gear, SKU number for
    #  consumables - one thumbnail in Gear_Lookup\thumbs covers every
    #  item behind the code (Andrew, 30 Jul 2026: "thumbnails for
    #  everything ... for product variants")
    vkey = {}

    if rental_path and os.path.isfile(rental_path):
        wb = openpyxl.load_workbook(rental_path, read_only=True, data_only=True)
        ws = wb['RENTAL_STOCK'] if 'RENTAL_STOCK' in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        hdr = [str(c or '').strip() for c in next(rows)]
        ix = {h: i for i, h in enumerate(hdr) if h}
        need = {'ITEM_STATUS', 'ITEM_DESCRIPTION', 'STORAGE_UNIT'}
        if need <= set(ix):
            #  BOTH SIDES OF THE SHELF (Andrew, 3 Aug 2026: "its
            #  missing heaps of gear"). 27 of the 30 gas monitors were
            #  out with crews, so the catalogue showed 3 and a worker
            #  concluded the store barely owns any. Gear that is OUT is
            #  still gear we HAVE - it is counted separately and the
            #  page says "all out with crews" instead of hiding it.
            out_n = {}
            for r in rows:
                if not r:
                    continue
                status = str(r[ix['ITEM_STATUS']] or '').strip().lower()
                if status not in (AVAILABLE, 'on hire'):
                    continue          # SiteIQ's word, never our guess
                raw = r[ix['ITEM_DESCRIPTION']]
                if not _offerable(raw):
                    _obs_n[0] += 1
                    continue          # retired, faulty or written off
                name = _tidy(raw, master,
                             r[ix['ITEM_NUMBER']] if 'ITEM_NUMBER' in ix else '')
                if not name or not _offerable(name):
                    continue
                unit = _unit_of(r[ix['STORAGE_UNIT']])
                k = (name, _cat_of(name), unit)
                if status == 'on hire':
                    out_n[k] = out_n.get(k, 0) + 1
                    hire.setdefault(k, 0)
                else:
                    hire[k] = hire.get(k, 0) + 1
                if k not in vkey and 'PRODUCT_VARIANT' in ix:
                    _vk = str(r[ix['PRODUCT_VARIANT']] or '').strip().upper()
                    if not _vk:
                        #  radios / gas monitors have no variant - the
                        #  MODEL photo key is derived off the name
                        import mygear_thumbs as _T2
                        _vk = _T2.derived_keys(name)[1]
                    vkey[k] = _vk
        wb.close()

    if sales_path and os.path.isfile(sales_path):
        wb = openpyxl.load_workbook(sales_path, read_only=True, data_only=True)
        ws = wb['SALES_STOCK'] if 'SALES_STOCK' in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        hdr = [str(c or '').strip() for c in next(rows)]
        ix = {h: i for i, h in enumerate(hdr) if h}
        if {'SKU_DESCRIPTION', 'AVAILABLE_QUANTITY'} <= set(ix):
            hcol = ix.get('HIRER')
            for r in rows:
                if not r:
                    continue
                #  a line against a person's name is gear they have taken,
                #  not the shelf - the store's own line has no hirer
                if hcol is not None and str(r[hcol] or '').strip():
                    continue
                if not _offerable(r[ix['SKU_DESCRIPTION']]):
                    continue
                #  THE DO-NOT-SHOW LIST (Andrew, 2 Aug 2026: "can we
                #  permently remove these and ensure these dont show up
                #  again"). Applied before the line is counted, so a
                #  hidden item cannot pad a shelf count either.
                if _HID is not None and 'SKU_NUMBER' in ix:
                    if _HID.hidden(r[ix['SKU_NUMBER']],
                                   r[ix['STORAGE_UNIT']] if 'STORAGE_UNIT' in ix else ''):
                        _hid_n[0] += 1
                        continue
                name = _tidy(r[ix['SKU_DESCRIPTION']], master)
                q = _qty(r[ix['AVAILABLE_QUANTITY']])
                if not name or q is None:
                    continue
                unit = _unit_of(r[ix['STORAGE_UNIT']]) if 'STORAGE_UNIT' in ix \
                    else 'Consumables'
                #  duplicate SKU records fold here the same way the
                #  consumables watch folds them: SiteIQ carries dead
                #  twin records at zero stock beside the live line, and
                #  a crew reading "NONE right now" under an item the
                #  shelf holds plenty of is worse than no list at all
                #  (Andrew, 30 Jul 2026: "remove the 6 duplicates these
                #  will be ones with 0 stock")
                #  "(alt)" is the master list's marker for a twin record
                #  it had to rename - fold it with its live line too
                _fk = re.sub(r'\s*\(alt\)\s*$', '', name, flags=re.I).upper()
                cons.setdefault((_fk, _cat_of(name)), []).append(
                    {'n': name, 'u': unit, 'q': q,
                     'sku': str(r[ix['SKU_NUMBER']] or '').strip().upper()
                            if 'SKU_NUMBER' in ix else ''})
        wb.close()

    def _fl(name):
        #  compliance letters by description - same authority as the
        #  worker card (equipment_compliance); blank when unbound
        try:
            import equipment_compliance as _EC
            f = _EC.flags(None, name)
        except Exception:
            return ''
        return (('E' if f.get('electrical') else '')
                + ('R' if f.get('rigging') else '')
                + ('L' if f.get('logbook') else ''))

    #  the kit list (Andrew, 1 Aug 2026): "this item goes out and comes
    #  back with" - the same lists as the tool store stickers. Lives in
    #  kits.txt (protected, written at the counter); the shipped
    #  kits_EXAMPLE.txt shows the format:  match words | contents
    def _kits():
        rules = []
        for fn in ('kits.txt', 'kits_EXAMPLE.txt'):
            p = os.path.join(os.path.dirname(os.path.abspath(__file__)), fn)
            if not os.path.isfile(p):
                continue
            with open(p, encoding='utf-8', errors='replace') as fh:
                for line in fh:
                    line = line.strip()
                    if not line or line.startswith('#') or '|' not in line:
                        continue
                    w, c = line.split('|', 1)
                    words = tuple(x.upper() for x in w.split() if x)
                    if words and c.strip():
                        rules.append((words, c.strip()))
            break
        return rules
    KITS = _kits()

    #  SDS links (Andrew, 1 Aug 2026): "my qr codes is the direct link
    #  to the sds" - the A3 quick-scan board, living on every chemical's
    #  card. sds.txt (protected) beats the shipped sds_EXAMPLE.txt.
    def _sds_rules():
        rules = []
        for fn in ('sds.txt', 'sds_EXAMPLE.txt'):
            p = os.path.join(os.path.dirname(os.path.abspath(__file__)), fn)
            if not os.path.isfile(p):
                continue
            with open(p, encoding='utf-8', errors='replace') as fh:
                for line in fh:
                    line = line.strip()
                    if not line or line.startswith('#') or '|' not in line:
                        continue
                    w, u = line.split('|', 1)
                    ws = tuple(x.upper() for x in w.split() if x)
                    if ws and u.strip().startswith('http'):
                        rules.append((ws, u.strip()))
            break
        return rules
    SDS = _sds_rules()

    def _sds_of(name):
        up = ' ' + name.upper() + ' '
        for ws, u in SDS:
            if all(w in up for w in ws):
                return u
        return ''

    def _kit_of(name):
        up = ' ' + name.upper() + ' '
        for words, c in KITS:
            if all(w in up for w in words):
                return c
        return ''

    def pack(d, kind):
        out = []
        for (name, cat, unit), n in d.items():
            #  c = the category you look under, u = the aisle you walk to
            it = {'n': name, 'c': cat, 'u': unit,
                  'q': int(n), 'k': kind,
                  'v': vkey.get((name, cat, unit), ''),
                  'fl': _fl(name)}
            #  how many of this line are out with crews - carried only
            #  when the reader would otherwise be lied to by silence
            _o = out_n.get((name, cat, unit), 0) if kind == 'h' else 0
            if _o:
                it['o'] = int(_o)
            _kt = _kit_of(name)
            if _kt:
                it['kt'] = _kt
            _sd = _sds_of(name)
            if _sd:
                it['sd'] = _sd
            out.append(it)
        out.sort(key=lambda x: x['n'].lower())
        return out

    H = pack(hire, 'h')
    #  one line per consumable: quantities summed across its records,
    #  aisle and photo/scan key taken from the LIVE record (most stock)
    C = []
    for (_fk, cat), recs in cons.items():
        live = max(recs, key=lambda x: x['q'])
        _ci = {'n': live['n'], 'c': cat, 'u': live['u'],
               'q': int(sum(x['q'] for x in recs)), 'k': 'c',
               'v': live['sku'], 'fl': ''}
        _sd = _sds_of(live['n'])
        if _sd:
            _ci['sd'] = _sd
        C.append(_ci)
    C.sort(key=lambda x: x['n'].lower())
    stats = {
        'hireItems': sum(x['q'] for x in H),
        'hireLines': len(H),
        'hireSingles': sum(1 for x in H if x['q'] == 1),
        'consLines': len([x for x in C if x['q'] > 0]),
        'consUnits': sum(x['q'] for x in C if x['q'] > 0),
        'consOut': len([x for x in C if x['q'] == 0]),
        'consLow': len([x for x in C if 0 < x['q'] <= 5]),
    }
    stats['hidden'] = _hid_n[0]
    stats['obsolete'] = _obs_n[0]
    return {'hire': H, 'cons': C, 'stats': stats}


# ---------------------------------------------------------------------
#  THE SCREEN
# ---------------------------------------------------------------------
CSS = """
/* ---- WHAT'S IN THE STORE ---------------------------------------- */
.stwrap{padding:2px 0 26px}
.stnote{background:#171B22;border:1px solid #2A313C;border-left:4px solid var(--org);
  border-radius:12px;padding:12px 14px;margin:0 0 14px;font-size:13px;
  line-height:1.6;color:#C7CED8}
.stnote b{color:#fff}
.stsearch{position:sticky;top:0;z-index:5;background:#0A0E14;padding:8px 0 10px;
  margin:0 0 4px}
.stsearch input{width:100%;background:#171B22;border:1.5px solid #2A313C;
  color:#fff;font-family:inherit;font-size:17px;padding:15px 15px 15px 44px;
  border-radius:13px;-webkit-appearance:none}
.stsearch input:focus{outline:none;border-color:var(--org)}
.stsearch{position:relative}
.stsearch svg{position:absolute;left:14px;top:23px;width:19px;height:19px;
  color:#7B8798;pointer-events:none}
/* THE CONTAINER LOOK (Andrew, 31 Jul 2026: "like you're inside the
   tooling container... like a Coates catalogue") - corrugated-steel
   banner, rack-ticket bay labels */
.ctop{background:repeating-linear-gradient(90deg,#141922 0 16px,#1A212C 16px 32px);
 border:1px solid #2A313C;border-left:5px solid var(--org);border-radius:12px;
 padding:12px 16px;margin:2px 0 12px}
.ctop b{display:block;color:#EAF0F7;font-size:13px;letter-spacing:2.5px;
 font-weight:900}
.ctop span{display:block;color:#8A97A8;font-size:11.5px;margin-top:3px}
.stcats{display:grid;grid-template-columns:repeat(2,1fr);gap:10px;margin:4px 0 14px}
.stcat{background:#151A22;border:1px solid #2A313C;border-radius:12px;
  padding:12px 6px 10px;text-align:center;color:var(--org);cursor:pointer;
  font-family:inherit;min-height:74px}
.stcat.on{border-color:var(--org);background:#1E1710}
.stcat .cimg{width:100%;height:104px;border-radius:10px;object-fit:cover;
 display:block;margin:0 0 7px}
.stcat b{display:block;background:var(--org);color:#fff;border-radius:6px;
 padding:3px 6px;margin:2px 4px 0;font-size:11.5px;font-weight:800;
  letter-spacing:.4px;margin-top:5px}
.stcat span{display:block;color:#8A97A8;font-size:11px;font-weight:700;
  margin-top:2px}
.stfam{display:flex;align-items:center;gap:9px;font-size:11.5px;font-weight:800;
 letter-spacing:1px;text-transform:uppercase;color:#F26222;
 padding:0;margin:14px 0 7px}
.stfam b{background:var(--org);color:#fff;border-radius:6px;padding:3px 9px;
 font-weight:800}
.stfam span{color:#8A97A8;font-weight:700;letter-spacing:.4px;text-transform:none;
 flex:none}
.stfam::after{content:'';flex:1;height:1px;background:#2A313C;min-width:14px}
/*  THE SEARCH RESULT ROW (Andrew, 3 Aug 2026 - the last place the
    squeeze was still living). A 112px picture and a 74px count column
    beside the words left about 150px of a 390px phone for the text,
    so a serial-named gas monitor came down the screen six lines high
    and its alternatives list read as a wall of half-names.
    On a phone: picture and count on the top row, the words underneath
    at full width. From 560px up it goes back to one line.  */
.strow{display:grid;grid-template-columns:112px 1fr;gap:10px 12px;
  align-items:center;background:#151A22;
  border:1px solid #2A313C;border-left:4px solid #2BB673;border-radius:12px;
  padding:12px 13px;margin-bottom:8px;min-height:60px}
.strow .stn{grid-column:1/-1}
@media(min-width:560px){
  .strow{display:flex;align-items:center;gap:12px}
  .strow .stn{grid-column:auto}
}
.strow.few{border-left-color:#F5A623}
.strow.none{border-left-color:#E23B2E;opacity:.72}
/* THE SHOWROOM GRID - browsing a category faces you with the gear,
   big photo cards two across, count badge on the corner (31 Jul 2026:
   "get these this big and bigger... make this elite") */
.stgrid{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.stgrid .stfam{grid-column:1/-1}
/* STAGE TWO - THE CATALOGUE SPREAD (Andrew, 31 Jul 2026: "like a
   Coates catalogue") - open a bay and it reads like the catalogue:
   the bay-front hero over the corrugated steel, shelf tickets, and
   the lead item of every shelf running full-width like the
   catalogue's feature shot */
.cspread{position:relative;grid-column:1/-1;border:1px solid #2A313C;
 border-radius:14px;overflow:hidden;min-height:118px;
 background:repeating-linear-gradient(90deg,#141922 0 16px,#1A212C 16px 32px)}
.cspread img{position:absolute;top:0;left:0;width:100%;height:100%;
 object-fit:cover;opacity:.42}
.cspread .ct{position:relative;padding:32px 14px 13px;
 background:linear-gradient(180deg,rgba(10,14,20,0) 0%,rgba(10,14,20,.94) 74%)}
.cspread .ct b{display:inline-block;background:var(--org);color:#fff;
 border-radius:7px;padding:4px 12px;font-size:15px;font-weight:900;
 letter-spacing:2px}
.cspread .ct span{display:block;color:#C7CED8;font-size:11.5px;
 font-weight:700;margin-top:6px}
/*  THE PICTURE IS A PICTURE, NOT A BILLBOARD.
    (Andrew, 4 Aug 2026: "in the section whats in the store you have
    enlarged the pics so so so much. needs to look like how we use
    things in our area.")

    On a phone this card stacked - a 132px-tall photograph running the
    FULL WIDTH of the screen with the words underneath. One item filled
    half a phone, so walking an aisle of forty lines was forty screens
    of photography. That is not how anything else in here reads: the
    toolbox talk, the guide rows and the stores board are all a small
    square picture on the left with the words beside it.

    So this is that shape too, at every width. 96px square on a phone,
    132px from 560px up where there is room. The reason it was stacked
    in the first place was that a 132px photo AND a QR tile left about
    120px of a 390px phone for the text - so the QR comes out of the
    picture and sits in the body as a chip instead (.stqr.s2, which was
    already here for the list rows). Roughly 265px for the words on a
    390px phone, and the photo still reads at a glance.  */
.stcard.wide{grid-column:1/-1;flex-direction:row;align-items:flex-start;
 gap:11px;padding:11px}
/*  A SQUARE, not a strip. align-self:stretch made the picture as tall
    as the whole card - 96 wide by 228 tall on a card with a few chips
    on it - so a spanner lying flat came out cropped to a sliver. It is
    a fixed square at the top left, exactly like the toolbox talk and
    the guide rows.  */
.stcard.wide .im{width:96px;height:96px;flex:none;align-self:flex-start;
 aspect-ratio:auto;border-radius:12px;overflow:hidden}
.stcard.wide .bd{flex:1;min-width:0;padding:0}
.stcard.wide .bd b{font-size:14px;overflow-wrap:break-word}
@media(min-width:560px){
  .stcard.wide .im{width:132px;height:132px}
  .stcard.wide .bd b{font-size:15px}
}
/*  the alternatives list reads as a list, not a paragraph  */
.stalt span{display:block;margin-top:3px}
.stcard{background:#151A22;border:1px solid #28323F;border-radius:16px;
 overflow:hidden;display:flex;flex-direction:column}
.stcard.none{border-color:#5a2622}
/* the aisle IS the address - wear it like one (Andrew, 1 Aug 2026) */
.loc{font-style:normal;color:#FFB347;font-weight:800;letter-spacing:.4px;
  text-transform:uppercase;font-size:10px}
.loc:before{content:"📍 ";font-size:9px}
/*  THE SCAN CHIP, SIZED. Moving it out of the photograph and into the
    body made it a static element, and a static SVG with no width takes
    the whole line - a white block right across the card, bigger than
    the product photo next to it. It is a 46px chip. (4 Aug 2026.)  */
.stqr.s2{position:static;display:inline-block;margin-top:7px;
 width:46px;height:46px;box-shadow:none}
.stqr.s2 svg{width:100%;height:100%;display:block}
.stqr{position:absolute;bottom:8px;right:8px;background:#fff;border-radius:7px;
  padding:3px;line-height:0;cursor:pointer;box-shadow:0 2px 8px rgba(0,0,0,.45)}
#stqrbig{display:none;position:fixed;inset:0;z-index:99;background:rgba(6,9,14,.96);
  flex-direction:column;align-items:center;justify-content:center;gap:10px;
  text-align:center;padding:24px}
#stqrbig .qrs{background:#fff;border-radius:16px;padding:16px;line-height:0}
#stqrbig b{color:#EAF0F7;font-size:16px;max-width:320px}
#stqrbig span{color:#8A97A8;font-size:12px;letter-spacing:1px}
#stqrbig i{color:#F26222;font-style:normal;font-weight:900;letter-spacing:2px;
  font-size:12px}
#stqrbig u{color:#5A6472;font-size:10.5px;text-decoration:none}
.stsds{display:block;margin-top:5px;font-size:10px;font-weight:800;
  letter-spacing:.5px;color:#F0B429;text-decoration:none;
  border:1px solid #4a3a10;border-radius:8px;padding:5px 8px;text-align:center}
.stkit{margin-top:5px;font-size:10.5px;color:#8A97A8;line-height:1.5}
.stkit b{color:#FFB347}
.strider{margin-top:4px;font-size:10px;font-weight:800;letter-spacing:.5px;
  color:#F0B429;text-transform:uppercase}
.stt{display:block;margin-top:4px;font-size:10.5px;font-weight:800;
  letter-spacing:.6px}
.stt i{display:inline-block;width:7px;height:7px;border-radius:50%;
  margin-right:5px;vertical-align:1px}
.stt.g{color:#35D68A}.stt.g i{background:#35D68A}
.stt.a{color:#F0B429}.stt.a i{background:#F0B429}
.stt.r{color:#FF5A4D}.stt.r i{background:#FF5A4D}
.stask{margin-top:4px;font-size:10px;color:#8A97A8;font-style:italic}
.stsdsoff{margin-top:5px;font-size:10px;line-height:1.5;color:#F0B429;background:#191307;border:1px solid #4a3a10;border-radius:8px;padding:6px 8px}
.stsdsoff b{color:#FFD27A}
.stalt{margin-top:5px;font-size:10.5px;color:#8A97A8;line-height:1.55}
.stalt b{color:#35D68A;letter-spacing:.3px}
.stalt span{color:#C3CDDA;font-weight:700}
.stcard .im{position:relative;aspect-ratio:1/1;background:#1B2330;
 display:flex;align-items:center;justify-content:center;overflow:hidden}
.stcard .im img{width:100%;height:100%;object-fit:cover;display:block}
.stcard .gmono{color:#8A97A8;font-weight:900;font-size:30px;letter-spacing:1px;
 display:flex;flex-direction:column;align-items:center;justify-content:center;
 width:100%;height:100%}
.stcard .gmono u{display:block;font-size:7.5px;font-weight:800;letter-spacing:.8px;
 color:#6B7789;text-decoration:none;margin-top:4px;text-align:center}
.stcard .qb{position:absolute;top:8px;right:8px;border-radius:9px;
 padding:3px 10px;font-size:14px;font-weight:900;color:#08210c;background:#35D68A}
.stcard .qb.a{background:#F0B429;color:#2a1e05}
.stcard .qb.r{background:#FF5A4D;color:#fff}
.stcard .bd{padding:9px 11px 11px}
.stcard .bd b{display:block;font-size:13px;font-weight:700;color:#EAF0F7;line-height:1.35}
.stcard .bd span{display:block;font-size:10.5px;color:#8A97A8;margin-top:3px;line-height:1.5}
.vc{font-family:Consolas,Menlo,monospace;font-style:normal;font-size:9.5px;color:#6E7A8A}
.cchips{margin-top:5px}
.cchip{display:inline-block;color:#fff;border-radius:6px;padding:2px 8px;
 font-size:9.5px;font-weight:800;letter-spacing:.4px;margin:0 5px 3px 0}
.cchip.lbk{background:#3A2E08;color:#F5C032;border:1px solid #6b551b}
.strow .sth{flex:none;width:112px;height:112px;border-radius:14px;overflow:hidden;
 background:#20262e;display:flex;align-items:center;justify-content:center}
.strow .sth img{width:100%;height:100%;object-fit:cover;display:block}
.strow .sth.mono{color:#8A97A8;font-weight:900;font-size:23px;letter-spacing:.5px}
.strow .stn{flex:1;min-width:0}
.strow .stn b{display:block;font-size:14.5px;font-weight:700;color:#EAF0F7;
  line-height:1.35}
.strow .stn span{display:block;font-size:11.5px;color:#8A97A8;margin-top:3px;
  font-weight:700;letter-spacing:.4px;text-transform:uppercase}
.strow .stq{flex:none;text-align:right;min-width:74px}
.strow .stq b{display:block;font-size:21px;font-weight:900;color:#2BB673;
  line-height:1}
.strow.few .stq b{color:#F5A623}
.strow.none .stq b{color:#E23B2E;font-size:14px;padding-top:4px}
.strow .stq span{display:block;font-size:9.5px;color:#7B8798;font-weight:800;
  letter-spacing:.9px;text-transform:uppercase;margin-top:3px}
.stmore{display:block;width:100%;background:transparent;border:1px solid #2A313C;
  color:#A9B3C0;font-family:inherit;font-weight:800;font-size:13px;
  letter-spacing:.6px;text-transform:uppercase;padding:14px;border-radius:12px;
  margin-top:6px;min-height:48px}
.stcount{font-size:11.5px;color:#8A97A8;font-weight:800;letter-spacing:.8px;
  text-transform:uppercase;margin:2px 2px 9px}
.stnil{text-align:center;color:#8A97A8;font-size:14px;padding:26px 10px;
  line-height:1.7}
.stnil b{display:block;color:#fff;font-size:16px;margin-bottom:6px}
@media (prefers-reduced-motion:no-preference){
  .strow{animation:strise .26s ease both}
  @keyframes strise{from{opacity:0;transform:translateY(6px)}to{opacity:1;transform:none}}
}
"""

JS = """
/* ---- WHAT'S IN THE STORE: search and aisles ----------------------
   Deliberately dumb to use: one box, big buttons, three colours.
   Typing beats browsing for the bloke who knows what he wants;
   the aisles are there for the one who doesn't. */
var STORE_SHOWN = 60;
function stNorm(s){ return (s||'').toLowerCase().replace(/[^a-z0-9 ]+/g,' '); }
/* WORDS PEOPLE ACTUALLY SAY vs words the register writes (Andrew,
   3 Aug 2026: "monitors don't show... same with the crowfeet" - said
   a thousand times because typing "crowfeet" can never substring-match
   "Crowsfeet", it is one letter short, and the register calls a gas
   monitor a Multi-Gas Detector). Each spoken word expands to the
   register's spellings; a line matches if ANY expansion lands. */
var ST_SYN={
  /*  the workers' catalogue calls them CROWFOOT (the master's word);
      the register calls them CROWSFEET. Every spelling a bloke might
      type lands on both, in both directions - the one-way map was why
      typing the register's own spelling found nothing.  */
  'crowfeet':['crowsfeet','crows feet','crowfoot','crowsfoot'],
  'crowsfeet':['crowfoot','crows feet','crowfeet'],
  'crowsfoot':['crowfoot','crowsfeet'],
  'crowfoot':['crowsfeet','crowfeet'],
  'crows':['crowfoot','crowsfeet'],
  'feet':['crowfoot','crowsfeet'],
  'monitor':['detector','monitor'],
  'monitors':['detector','monitors'],
  'detector':['detector','monitor'],
  'radio':['radio','two way'],
  'battery':['battery','batteries'],
  'extension':['extension','extn'],
  'lead':['lead','extn lead']
};
function stWordHits(hay,hayFlat,w){
  if(hay.indexOf(w)>=0) return true;
  /* the no-spaces pass: "crows feet" typed as one word still lands */
  if(hayFlat.indexOf(w.replace(/ /g,''))>=0) return true;
  var alts=ST_SYN[w]||[];
  for(var i=0;i<alts.length;i++){
    if(hay.indexOf(alts[i])>=0) return true;
    if(hayFlat.indexOf(alts[i].replace(/ /g,''))>=0) return true;
  }
  return false;
}
/* The FAMILY a line belongs to - the name with its trailing size cut
   off, then the first segment: "1/2in Drive Crowsfeet-27MM" and its 30
   siblings are one family, "Allen Key Metric - 12mm" one, "Hydraulic
   Torque Wrench - Cassette - 32mm" one. Sockets alone is 283 lines in
   one aisle; family seams turn that scroll into a dozen signposted
   shelves. Search is untouched - typing already beats browsing.
   (Andrew, 29 Jul 2026: "can we go deeper into types of tooling
   groups when searching") */
function stFam(n){
  var f=String(n||'')
    .replace(/[-\\u2013]\\s*[\\d,\\/. ]+\\s*(mm|in|inch|")?\\s*$/i,'')
    .split(' - ')[0].replace(/[-\\s]+$/,'').trim();
  return f||String(n||'');
}
/* NONE never dead-ends (Andrew, 1 Aug 2026): when the shelf is empty,
   the card offers the nearest sizes of the same family that ARE on the
   shelf - every "no" comes with a "but here's your answer". */
function stAlt(it){
  var fam = stFam(it.n), out = [], i, s;
  for(i=0;i<STORE.length;i++){ s = STORE[i];
    if(s.n === it.n || !s.q || s.q <= 0) continue;
    if(stFam(s.n) !== fam) continue;
    out.push(s); if(out.length >= 3) break;
  }
  var ask='<div class="stask">or ask at the window &mdash; two metres away.</div>';
  if(!out.length) return ask;
  /*  one alternative per line - joined with dots they wrapped into a
      paragraph of half-names on a phone (3 Aug 2026)  */
  return '<div class="stalt"><b>But on the shelf:</b>'
    + out.map(function(s){return '<span>' + s.n + ' &times;' + s.q
        + '</span>'}).join('') + '</div>' + ask;
}
/* tap the corner QR: full-screen "show this at the window" (Andrew,
   1 Aug 2026) - the counter's scanner reads it straight off the phone */
function stQR(ev,el){
  ev.stopPropagation();
  var c=el.getAttribute('data-c'), n=el.getAttribute('data-n');
  var o=document.getElementById('stqrbig');
  if(!o){o=document.createElement('div');o.id='stqrbig';document.body.appendChild(o);}
  o.innerHTML='<div class="qrs">'+qr(c,240)+'</div><b>'+n+'</b><span>'+c+'</span>'
    +'<i>SHOW THIS AT THE WINDOW</i><u>tap anywhere to close</u>';
  o.style.display='flex';
  o.onclick=function(){o.style.display='none';};
}
/* The one thing on this page that genuinely needs the internet: the SDS
   itself lives on the manufacturer's site. Everything else is already in
   the phone. So if a bloke taps it out of range, say so plainly instead
   of handing him a browser error page - and point him at the A3 board on
   the wall, which never needs a signal. (1 Aug 2026) */
function sdsGo(a){
  if(navigator && navigator.onLine === false){
    var w=a.nextSibling;
    if(!w || w.className!=='stsdsoff'){
      w=document.createElement('div'); w.className='stsdsoff';
      a.parentNode.insertBefore(w, a.nextSibling);
    }
    w.innerHTML='<b>Connection unavailable.</b> The safety data sheet lives '
      +'on the maker&rsquo;s website. Use the SDS board on the store wall, '
      +'or try again in Wi-Fi range.';
    return false;
  }
  return true;
}
function stKit(it){
  var h='';
  if(it.sd) h+='<a class="stsds" href="'+it.sd+'" target="_blank" rel="noopener"'
    +' onclick="return sdsGo(this)">&#9888; SAFETY DATA SHEET &mdash; tap to open</a>';
  if(it.kt) h+='<div class="stkit"><b>Goes out &amp; comes back with:</b> '+it.kt+'</div>';
  if(it.fl && it.fl.indexOf('R')>=0)
    h+='<div class="strider">Rated lifting &mdash; check it before you rig it</div>';
  return h;
}
function stRender(reset){
  var box = document.getElementById('st-list'); if(!box) return;
  if(reset) STORE_SHOWN = 60;
  var term = stNorm((document.getElementById('st-q')||{}).value);
  var words = term.split(/\\s+/).filter(function(w){return w.length>1;});
  var cat = window.ST_CAT || 'All';
  var hits = STORE.filter(function(it){
    if(cat !== 'All' && it.c !== cat) return false;
    if(!words.length) return true;
    /* search across the name, the category and the aisle - a bloke who
       types "rigging" or "sockets" should get somewhere too */
    var hay = stNorm(it.n + ' ' + it.c + ' ' + it.u);
    var hayFlat = hay.replace(/ /g,'');
    for(var i=0;i<words.length;i++)
      if(!stWordHits(hay,hayFlat,words[i])) return false;
    return true;
  });
  var head = document.getElementById('st-count');
  if(head){
    /*  AND HOW MANY OF THEM HAVE NO PHOTO YET.
        (Andrew, 4 Aug 2026: "this is missing so so many pics in this
        area please fix.")
        The pictures are not lost - they were never taken. 1,073 lines
        on the shelf and 35 photographs between them. A screen full of
        two-letter monograms with nothing saying why reads as broken,
        and a bloke stops trusting the ones that ARE photographs.
        So it says the number, the same way every other cut list in
        here says what it left out. Drop a photo named after the code
        into Photos\ and the next 04 build closes the gap - the wanted
        list is 56_PHOTO_HUNT.  */
    var noPic = 0;
    for(var z=0; z<hits.length; z++){
      var hv = hits[z].v;
      if(!hv || (typeof hasThumb==='function' && !hasThumb(hv))) noPic++;
    }
    head.textContent = hits.length
      ? (hits.length + (hits.length===1?' thing':' things')
         + (cat==='All'?'':' in ' + cat)
         + (noPic ? ' \u00b7 ' + noPic + ' with no photo yet' : ''))
      : '';
  }
  if(!hits.length){
    box.innerHTML = '<div class="stnil"><b>Nothing matches that</b>'
      + 'Try a shorter word &mdash; <b style="display:inline">grinder</b>, '
      + '<b style="display:inline">batt</b>, <b style="display:inline">hose</b>. '
      + 'Or ask at the counter, the store carries more than the shelf shows.</div>';
    return;
  }
  var slice = hits.slice(0, STORE_SHOWN), html = '';
  /* BROWSING is a showroom, SEARCHING is a list. Walk into a category
     and the gear faces you as big photo cards - the picture IS the
     catalogue (Andrew, 31 Jul 2026: "get these this big and bigger
     again for their sub folder locations, make this elite"). Type a
     search and it drops back to the quick list. */
  var grid = !words.length;
  /* family seams only when BROWSING a big aisle - a search result or a
     small aisle stays a flat list, seams there are just noise */
  var seams = !words.length && cat !== 'All' && hits.length > 25;
  var lastFam = null, feat = false;
  if(grid) html += '<div class="stgrid">';
  /* STAGE TWO - the bay-front spread: walk into a bay and it opens
     like a catalogue page - the bay's rack photo big over the
     corrugated steel, and how much the bay holds */
  if(grid && cat !== 'All'){
    var fams = {}, famT = 0;
    for(var j2=0;j2<hits.length;j2++) fams[stFam(hits[j2].n)] = 1;
    for(var k2 in fams) famT++;
    html += '<div class="cspread">'
      + (window.ST_FACE ? '<img src="' + window.ST_FACE + '" alt="">' : '')
      + '<div class="ct"><b>' + cat.toUpperCase() + '</b>'
      + '<span>' + hits.length + (hits.length===1?' line':' lines')
      + ' in this bay' + (famT>1 ? ' &middot; ' + famT + ' shelves' : '')
      + '</span></div></div>';
  }
  for(var i=0;i<slice.length;i++){
    var it = slice[i];
    if(seams){
      var fam = stFam(it.n);
      if(fam !== lastFam){
        var famN = 0;
        for(var j=0;j<hits.length;j++) if(stFam(hits[j].n)===fam) famN++;
        html += '<div class="stfam"><b>' + fam + '</b>'
          + '<span>' + famN + (famN===1?' size':' sizes / kinds') + '</span></div>';
        lastFam = fam;
        /* the shelf's lead item runs full-width - the catalogue's
           feature shot - then the size run follows as cards */
        feat = true;
      }
    }
    var cls = it.q === 0 ? (it.o ? 'few' : 'none') : (it.q <= 3 ? 'few' : '');
    var big = it.q === 0 ? (it.o ? it.o : 'NONE') : it.q;
    var lab = it.q === 0 ? (it.o ? 'out with crews' : 'right now') : (it.k === 'c' ? 'on the shelf' : 'ready to hire');
    if(grid){
      /* ONE CARD SHAPE. (Andrew, 3 Aug 2026, pointing at the landscape
         row: "this is now the size we go".)
         This used to run only the LEAD item of each family full-width
         as a feature shot, and everything after it as a stacked grid
         card - so one screen held two different card shapes, the
         stacked ones left a dead column beside them, and a tall photo
         and a flat one came out completely different sizes. That is
         the "all odd sizes" he is describing. Every card is the wide
         row now: square photo left at a fixed 132px, everything else
         right. */
      var wide = true; feat = false;
      html += '<div class="stcard ' + (wide?'wide ':'') + cls + '">'
        + '<div class="im">'
        + ((it.v && (typeof hasThumb!=='function' || hasThumb(it.v)))
            ? '<img src="thumbs/' + encodeURIComponent(tsafe(it.v))
            + '.jpg" loading="lazy" alt="" data-m="' + thMono(it.n)
            + '" onerror="thxg(this)">'
          /*  no photograph yet, and it says so - a bare monogram looks
              like a picture that failed to load  */
          : '<span class="gmono">' + thMono(it.n)
            + '<u>NO PHOTO YET</u></span>')
        + '<span class="qb ' + (it.q===0?'r':(it.q<=3?'a':'g')) + '">'
        + big + '</span>'
        + '</div>'
        + '<div class="bd"><b>' + it.n + '</b>'
        + '<span><i class="loc">' + it.u + '</i>'
        + (it.v ? ' &middot; <i class="vc">' + it.v + '</i>' : '') + '</span>'
        + '<span class="stt ' + (it.q===0?'r':(it.q<=3?'a':'g')) + '"><i></i>'
        + (it.q===0 ? (it.o ? 'WE HAVE ' + it.o + ' &mdash; ALL OUT WITH CREWS' : 'NONE RIGHT NOW')
           : (it.q<=3 ? 'RUNNING LOW &mdash; ' + it.q + (it.o ? ' free &middot; ' + it.o + ' out' : '')
              : (it.k==='c' ? 'ON THE SHELF &mdash; ' : 'READY TO HIRE &mdash; ') + it.q
                + (it.o ? ' &middot; ' + it.o + ' out with crews' : '')))
        + '</span>'
        + stChips(it.fl) + stKit(it) + (it.q===0 ? stAlt(it) : '')
        /*  the scan code, as a chip under the words rather than a white
            square sitting on top of the product photograph  */
        + (window.qr ? '<span class="stqr s2" data-c="' + (it.v||it.n)
           + '" data-n="' + it.n + '" onclick="stQR(event,this)">'
           + qr(it.v||it.n,40) + '</span>' : '')
        + '</div></div>';
    } else {
      html += '<div class="strow ' + cls + '" style="animation-delay:'
        + Math.min(i*14,280) + 'ms">'
        + thTile(it.v, it.n)
        + '<div class="stn"><b>' + it.n + '</b><span><i class="loc">' + it.u + '</i>'
        + (it.v ? ' &middot; <i class="vc">' + it.v + '</i>' : '') + '</span>'
        + '<span class="stt ' + (it.q===0?'r':(it.q<=3?'a':'g')) + '"><i></i>'
        + (it.q===0 ? (it.o ? 'WE HAVE ' + it.o + ' &mdash; ALL OUT WITH CREWS' : 'NONE RIGHT NOW')
           : (it.q<=3 ? 'RUNNING LOW &mdash; ' + it.q + (it.o ? ' free &middot; ' + it.o + ' out' : '')
              : (it.k==='c' ? 'ON THE SHELF &mdash; ' : 'READY TO HIRE &mdash; ') + it.q
                + (it.o ? ' &middot; ' + it.o + ' out with crews' : '')))
        + '</span>'
        + stChips(it.fl) + stKit(it) + (it.q===0 ? stAlt(it) : '') + '</div>'
        + '<div class="stq"><b>' + big + '</b><span>' + lab + '</span>'
        + (window.qr ? '<span class="stqr s2" data-c="' + (it.v||it.n) + '" data-n="' + it.n
           + '" onclick="stQR(event,this)">' + qr(it.v||it.n,34) + '</span>' : '')
        + '</div></div>';
    }
  }
  if(grid) html += '</div>';
  box.innerHTML = html;
  var more = document.getElementById('st-more');
  if(more){
    if(hits.length > STORE_SHOWN){
      more.style.display = 'block';
      more.textContent = 'Show ' + Math.min(60, hits.length - STORE_SHOWN)
        + ' more of ' + hits.length;
    } else { more.style.display = 'none'; }
  }
}
/* the picture tile: the variant's thumbnail out of Gear_Lookup/thumbs,
   or a two-letter monogram until its photo is collected (56_PHOTO_HUNT
   is the wanted list). Lazy-loaded so a thousand rows stay quick. */
function thMono(n){
  /*  LETTERS, NOT THE SIZE. It took the first character of the first
      two words, so every spanner on the shelf wore "11" - 1 1/16",
      1 1/2", 1 1/4" all came out identical, which is worse than no
      monogram at all because it looks like information.
      The size is already on the line underneath; this wants the WORDS.
      "1 1/16 Combination Spanner" -> CS. (4 Aug 2026.)  */
  var all=String(n||'').split(/[^A-Za-z0-9]+/).filter(function(x){return x});
  /*  a word has to START with letters - "25mm" is a size, not a word,
      and "25mm Air Hose" was coming out as 2A instead of AH  */
  var w=all.filter(function(x){ return /^[A-Za-z]{2,}/.test(x); });
  if(!w.length) w=all;
  return ((w[0]||'?').charAt(0)+(w[1]||w[0]||'').charAt(0)).toUpperCase();
}
function tsafe(v){return String(v).replace(/[/:*?"<>|]/g,'_')}
function thTile(v,n){
  /* ASK THE MANIFEST, NOT THE NETWORK. Emitting an <img> for a photo
     that does not exist leaves an empty hole on the phone until the
     404 comes back - and below the fold, lazy loading means it never
     even asks, so the hole stays. See hasThumb in BUILD_MY_GEAR. */
  if(!v || (typeof hasThumb==='function' && !hasThumb(v)))
    return '<span class="sth mono" title="No photo collected yet">'
      +thMono(n)+'</span>';
  return '<span class="sth"><img src="thumbs/'+encodeURIComponent(tsafe(v))
    +'.jpg" loading="lazy" alt="" data-m="'+thMono(n)
    +'" onerror="thx(this)"></span>';
}
function thx(img){
  var s=img.parentNode;
  s.className='sth mono';
  s.textContent=img.getAttribute('data-m')||'?';
}
/* grid-card photo fallback: the big two-letter tile */
function thxg(img){
  var d=img.parentNode;
  var q=d.querySelector('.qb'), s=d.querySelector('.stqr');
  d.innerHTML='<span class="gmono">'+(img.getAttribute('data-m')||'?')+'</span>';
  if(q)d.appendChild(q);
  if(s)d.appendChild(s);
}
/* compliance chips for the crew: the tag colour word and the log book.
   __TAGC__/__TAGX__ are stamped by the build from the compliance
   master's windows - Jul/Aug prints BLUE. */
function stChips(fl){
  if(!fl) return '';
  var out='';
  if((fl.indexOf('E')>=0||fl.indexOf('R')>=0)&&'__TAGC__')
    out+='<span class="cchip" style="background:__TAGX__">TAG __TAGC__</span>';
  if(fl.indexOf('L')>=0)
    out+='<span class="cchip lbk">&#128221; LOG BOOK</span>';
  return out?'<div class="cchips">'+out+'</div>':'';
}
function stCat(name, el){
  window.ST_CAT = name;
  /* the tapped bay's rack photo becomes the spread's hero shot */
  window.ST_FACE = '';
  if(el && name !== 'All'){
    var im = el.querySelector('.cimg');
    if(im && im.style.display !== 'none') window.ST_FACE = im.getAttribute('src');
  }
  var all = document.querySelectorAll('.stcat');
  for(var i=0;i<all.length;i++) all[i].className = 'stcat';
  if(el) el.className = 'stcat on';
  stRender(true);
  /* stepping into a bay lands you IN the bay - at its spread, not
     back at the top of the bay wall. EVERYTHING goes back to the top. */
  var b = document.getElementById('gsbody'); if(b){
    var L = document.getElementById('st-count');
    if(name !== 'All' && L){
      b.scrollTop = Math.max(0, L.getBoundingClientRect().top
        - b.getBoundingClientRect().top + b.scrollTop - 78);
    } else { b.scrollTop = 0; }
  }
}
function stMore(){ STORE_SHOWN += 60; stRender(false); }
"""


def pane(data, asof):
    """The whole screen, ready to drop in as a guide pane."""
    st = data['stats']
    counts = {}
    for it in data['hire'] + data['cons']:
        counts[it['c']] = counts.get(it['c'], 0) + 1
    total = len(data['hire']) + len(data['cons'])

    #  category cards wear a REAL photo when one of their lines has one
    #  - picked at build time against the thumbs folder, so nothing 404s
    #  (Andrew, 31 Jul 2026: "product categories - images for these too")
    import mygear_thumbs as _T4
    _tdir = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         'Gear_Lookup', 'thumbs')
    catv = {}
    for it in data['hire'] + data['cons']:
        _c0 = it['c']
        if _c0 in catv:
            continue
        _v0 = (it.get('v') or '').strip()
        if _v0 and os.path.isfile(
                os.path.join(_tdir, _T4.safe_name(_v0) + '.jpg')):
            catv[_c0] = _T4.safe_name(_v0)

    def _cface(name, path):
        #  a DEDICATED category render wins (Andrew's pack, 31 Jul 2026:
        #  CATEGORY_SOCKETS.jpg and friends); else the first pictured
        #  item in the aisle; else the line icon
        _ck = 'CATEGORY_' + re.sub(r'[^A-Z0-9]+', '_',
                                   name.upper()).strip('_')
        if os.path.isfile(os.path.join(_tdir, _ck + '.jpg')):
            return ("<img class='cimg' src='thumbs/{v}.jpg' alt='' "
                    "onerror=\"this.style.display='none'\">").format(v=_ck)
        if name in catv:
            return ("<img class='cimg' src='thumbs/{v}.jpg' alt='' "
                    "onerror=\"this.style.display='none'\">").format(
                        v=catv[name])
        return _icon(path)

    cats = ("<button class='stcat on' type='button' onclick=\"stCat('All',this)\">"
            + _cface('Everything', 'M4 6h16M4 12h16M4 18h16')
            + "<b>EVERYTHING</b><span>{}</span></button>".format(total))
    #  biggest first: the aisles a crew actually reaches for, at the top
    for name, path, _keys in sorted(CATS, key=lambda c: -counts.get(c[0], 0)):
        if not counts.get(name):
            continue
        cats += ("<button class='stcat' type='button' onclick=\"stCat('{n}',this)\">"
                 "{i}<b>{u}</b><span>{c}</span></button>".format(
                     n=name, u=name.upper(), i=_cface(name, path),
                     c=counts[name]))
    if counts.get(OTHER[0]):
        cats += ("<button class='stcat' type='button' onclick=\"stCat('{n}',this)\">"
                 "{i}<b>{u}</b><span>{c}</span></button>".format(
                     n=OTHER[0], u=OTHER[0].upper(),
                     i=_cface(OTHER[0], OTHER[1]),
                     c=counts[OTHER[0]]))

    mag = ("<svg viewBox='0 0 24 24' fill='none' stroke='currentColor' "
           "stroke-width='2' stroke-linecap='round'><circle cx='11' cy='11' "
           "r='7'/><path d='M20 20l-3.5-3.5'/></svg>")

    return (
        "<div class='stwrap'>"
        "<div class='stnote'>Everything the store had on the shelf "
        "<b>as at {asof}</b>. It is this morning's count, not a live one "
        "&mdash; if it's the last one, or you can't see it here, "
        "<b>ask at the window</b>. We'll have more out the back.</div>"
        "<div class='stsearch'>{mag}<input id='st-q' type='search' "
        "inputmode='search' autocomplete='off' "
        "placeholder='Type what you need &mdash; grinder, batt, hose' "
        "oninput='stRender(true)' aria-label='Search the store'></div>"
        "<div class='ctop'><b>STEP INTO THE TOOL STORE</b>"
        "<span>Walk the bays below &mdash; every picture is the real gear "
        "on these shelves. Tap a bay, or ask the search like you'd ask at "
        "the window.</span></div>"
        "<div class='stcats'>{cats}</div>"
        "<div class='stcount' id='st-count'></div>"
        "<div id='st-list'></div>"
        "<button class='stmore' id='st-more' type='button' onclick='stMore()' "
        "style='display:none'>Show more</button>"
        "</div>"
    ).format(asof=asof or 'this morning', cats=cats, mag=mag)
