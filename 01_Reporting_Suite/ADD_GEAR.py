#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | ADD GEAR - one new item, every file it belongs in
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  New gear turns up on site - a sub-hire, a swap, an extra machine -
#  and three files need to know about it before the reports can tell
#  its story properly:
#
#    1. K2_MASTER_EQUIPMENT_PRICING.xlsx  the display name, the Plant
#       ID, the replacement cost and the compliance flags (logbook,
#       tag, return daily). This one drives EVERY report.
#    2. ImportSample_Contracted Rates and Prices.xlsx  the daily hire
#       rate, so cost and utilisation figures are right.
#    3. Plant_ID_Register.csv  the old asset -> Plant ID list the
#       dashboard still falls back to.
#
#  I ask, I check for clashes, I back up each file, then I write.
#  Nothing is overwritten without telling you first.
#
#  IMPORTANT - THE ONE THING I CANNOT DO
#  On-hire comes from SiteIQ. This makes the reports describe the item
#  correctly, but it only APPEARS on a report once it is in SiteIQ and
#  shows up in the RENTAL_STOCK / ON_HIRE export. Book it in at the
#  counter as normal.
# =====================================================================
import csv
import datetime as dt
import os
import shutil
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
MASTER = os.path.join(HERE, "K2_MASTER_EQUIPMENT_PRICING.xlsx")
RATES = os.path.join(HERE, "ImportSample_Contracted Rates and Prices.xlsx")
PIDCSV = os.path.join(HERE, "Plant_ID_Register.csv")
BK = os.path.join(HERE, "Updates", "gear_backups")


def ask(label, default=""):
    d = " [{}]".format(default) if default else ""
    try:
        v = input(" {}{}: ".format(label, d)).strip()
    except EOFError:
        v = ""
    return v or default


def yn(label, default="y"):
    v = ask(label + " (y/n)", default).lower()
    return v.startswith("y")


def backup(path):
    if not os.path.isfile(path):
        return
    os.makedirs(BK, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy2(path, os.path.join(
        BK, "{}_{}".format(stamp, os.path.basename(path))))


def main():
    print("=" * 62)
    print(" COATES | ADD GEAR")
    print("=" * 62)
    import openpyxl

    item = ask("Item / SKU barcode number (e.g. FK505)").upper()
    if not item:
        print(" Nothing entered - stopping.")
        return 1
    desc = ask("Description as it should read on reports")
    variant = ask("Product variant ID (for the rate)", item).upper()
    pid = ask("Plant ID (blank if none)")
    rate = ask("Daily hire rate in dollars (blank if none)")
    repl = ask("Replacement cost in dollars (blank = TBC)")
    unit = ask("Storage unit", "Cement Site Plant")
    cat = ask("Equipment category", "Plant - Forklifts")
    logbook = "Y" if yn("Needs daily pre-start / logbook?", "y") else ""
    elec = "Y" if yn("Needs an electrical test tag?", "n") else ""
    rig = "Y" if yn("Needs a rigging tag?", "n") else ""
    ret = ask("Return requirement (blank = none)", "")
    note = ask("Note for your own records (e.g. SUB-HIRE, start 27/07/2026)")

    # ---- clash checks BEFORE we touch anything -------------------------
    wb = openpyxl.load_workbook(MASTER)
    ws = wb["MASTER"]
    hdr = [str(c.value or "").strip() for c in ws[1]]
    col = {h: i for i, h in enumerate(hdr)}
    existing_row = None
    pid_owner = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col["ITEM_NUMBER"] + 1).value
        if str(v or "").strip().upper() == item:
            existing_row = r
        p = ws.cell(row=r, column=col["PLANT_ID"] + 1).value
        if pid and str(p or "").strip() == pid:
            pid_owner = (str(v or ""),
                         str(ws.cell(row=r,
                                     column=col["ITEM_DESCRIPTION"] + 1).value or ""),
                         r)
    print("")
    if existing_row:
        print(" NOTE: {} is already in the master (row {}).".format(
            item, existing_row))
        if not yn(" Update that row instead of adding a new one?", "y"):
            print(" Stopping - nothing changed.")
            return 1
    release_row, release_when = None, ""
    if pid_owner and (not existing_row or pid_owner[0].upper() != item):
        print(" CLASH: Plant ID {} is already on {} - {}".format(
            pid, pid_owner[0], pid_owner[1][:40]))
        print("        Two items with one Plant ID makes the reports lie -")
        print("        a lookup can return either one.")
        if yn(" Has {} been off hired?".format(pid_owner[0]), "y"):
            #  A clean handover: the old machine gives the number up and
            #  the new one takes it. Recorded both ways so the swap can
            #  always be explained. (Andrew's forklift, 27 Jul 2026.)
            release_when = ask(" When was it off hired?",
                               dt.date.today().strftime("%d/%m/%Y"))
            release_row = pid_owner[2]
            print("        Plant ID {} will be released from {} and given"
                  " to {}.".format(pid, pid_owner[0], item))
        elif not yn(" Take the Plant ID anyway?", "n"):
            pid = ""
            print("        Left blank - sort it out in the master later.")

    if not yn("\n Write these changes now?", "y"):
        print(" Nothing changed.")
        return 1

    # ---- 1. the master --------------------------------------------------
    backup(MASTER)
    row = existing_row or (ws.max_row + 1)
    def put(name, val):
        if name in col and val != "":
            ws.cell(row=row, column=col[name] + 1).value = val
    put("STORAGE_UNIT", unit)
    put("ITEM_NUMBER", item)
    put("PLANT_ID", pid)
    put("ITEM_DESCRIPTION", desc)
    put("PRODUCT_VARIANT", variant)
    put("NEW_DESCRIPTION", desc)
    if repl:
        put("REPLACEMENT_COST_AUD", float(repl.replace("$", "").replace(",", "")))
        put("REPLACEMENT_PRICE_SOURCE", "Entered by Andrew Fisher {}".format(
            dt.date.today().strftime("%d %b %Y")))
    put("EQUIPMENT_CATEGORY", cat)
    put("ELECTRICAL_TAG", elec)
    put("RIGGING_TAG", rig)
    put("LOGBOOK_REQUIRED", logbook)
    put("RETURN_REQUIREMENT", ret)
    if release_row:
        ws.cell(row=release_row, column=col["PLANT_ID"] + 1).value = None
        print("  MASTER  | Plant ID {} released from {} (off hired {})".format(
            pid, pid_owner[0], release_when))
    wb.save(MASTER)
    wb.close()
    print("  MASTER  | {} {} at row {}".format(
        item, "updated" if existing_row else "added", row))

    # ---- 2. the rate ----------------------------------------------------
    if rate:
        backup(RATES)
        wb = openpyxl.load_workbook(RATES)
        ws = wb["Coates Equipment"]
        hit = None
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=1).value or "").strip().upper() == variant:
                hit = r
                break
        r = hit or (ws.max_row + 1)
        ws.cell(row=r, column=1).value = variant
        ws.cell(row=r, column=2).value = float(rate.replace("$", "").replace(",", ""))
        wb.save(RATES)
        wb.close()
        print("  RATES   | {} = ${}/day ({})".format(
            variant, rate, "updated" if hit else "added"))

    # ---- 3. the plant id register --------------------------------------
    if pid:
        backup(PIDCSV)
        rows, seen = [], False
        if os.path.isfile(PIDCSV):
            with open(PIDCSV, newline="", encoding="utf-8-sig") as f:
                for i, rr in enumerate(csv.reader(f)):
                    if i and rr and rr[0].strip().upper() == item:
                        rr = [item, pid]
                        seen = True
                    if (release_row and i and rr
                            and rr[0].strip().upper() == pid_owner[0].upper()):
                        continue          # the old machine gives the number up
                    rows.append(rr)
        if not rows:
            rows = [["Asset Number", "Plant ID"]]
        if not seen:
            rows.append([item, pid])
        with open(PIDCSV, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)
        print("  PLANT ID| {} -> Plant ID {}".format(item, pid))

    if note or release_row:
        with open(os.path.join(HERE, "GEAR_ADDED_LOG.txt"), "a",
                  encoding="utf-8") as f:
            if release_row:
                f.write("{} | Plant ID {} handover: {} off hired {} -> {}\n"
                        .format(dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
                                pid, pid_owner[0], release_when, item))
            if note:
                f.write("{} | {} | {} | {}\n".format(
                    dt.datetime.now().strftime("%Y-%m-%d %H:%M"), item,
                    desc, note))
        print("  LOG     | noted in GEAR_ADDED_LOG.txt")

    print("-" * 62)
    print(" Done. Backups of every file changed are in")
    print(" Updates\\gear_backups\\")
    print("")
    print(" ONE MORE THING: this tells the REPORTS how to describe the")
    print(" item. It shows up on a report once SiteIQ has it - book it")
    print(" in at the counter, then pull a fresh RENTAL_STOCK / ON_HIRE")
    print(" export and run the morning as normal.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
