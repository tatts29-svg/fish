#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | MYBRANCH FLEET LISTING - the branch's view of the same gear
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 2 Aug 2026): "is this handy to have or no" - the
#  MyBranch "Fleet Listing by Availability Status" export.
#
#  It is, and it fixed a live fault the day it arrived. On the 2 Aug
#  pull, 44 assets that the store register calls "Available for Hire"
#  are RESERVED at branch level. Fleet Details was ranking them top of
#  the list and telling a counter hand to go and get one. Andrew's own
#  rule for that screen was to rank every asset "without hiding
#  out-of-service or reserved gear inside the percentage" - and until
#  this file turned up there was no way to know which ones were
#  reserved, because SiteIQ's RENTAL_STOCK does not carry it.
#
#  ---------------------------------------------------------------
#  TWO LEVELS OF HIRE, AND THEY ARE NOT A CONTRADICTION
#  ---------------------------------------------------------------
#  635 assets read "On Hire" here and "Available for Hire" in the store
#  register. Both are right. The BRANCH has hired the gear to the
#  project; the STORE has not yet issued it to a crew. That is the same
#  commercial-versus-client-issued split this suite already draws, seen
#  from the other end - so branch "On Hire" is NEVER treated as a
#  reason to exclude an asset from the shelf. It only means the branch
#  has it out to Cement Australia.
#
#  ---------------------------------------------------------------
#  IT COVERS PART OF THE STORE, AND THE PART IS NAMED
#  ---------------------------------------------------------------
#  976 of 5,337 K2 assets appear in it (18.3%), 915 of them on GLST.
#  The rest have no branch record at all. An asset this file has never
#  heard of is NOT assumed to be fine and NOT assumed to be reserved -
#  it is simply left as the store register has it, and the coverage is
#  printed so nobody reads a clean screen as a clean fleet.
#
#  Drop the export in Data_MyBranch\. No file, no change to anything.
# =====================================================================
import glob
import os

HERE = os.path.dirname(os.path.abspath(__file__))
FOLDER = os.path.join(HERE, 'Data_MyBranch')

#  Availability values that mean the asset must not be recommended.
#  Read off the export itself, not invented - every one of these is a
#  real value on the 2 Aug pull. The word against each is what the
#  screen prints, so a store hand is told WHY, not just "no".
NOT_ISSUABLE = {
    'reserved': 'reserved for someone else',
    'reserved, in service': 'reserved, and in service',
    'inspection pending': 'inspection pending',
    'in service': 'in service',
    'off site for repair': 'off site for repair',
    'in transfer': 'in transfer between branches',
    'wait for config job': 'waiting on a config job',
    'off hired': 'off hired at the branch',
}

#  Branch-level hire. NOT an exclusion - see the header. Named here so
#  nobody later adds it to the set above by mistake.
BRANCH_HIRE = ('on hire', 'on hire, in service')

_CACHE = None


def _norm(s):
    return ' '.join(str(s or '').split()).lower()


def newest(folder=None):
    hits = [p for p in glob.glob(os.path.join(folder or FOLDER, '*.xlsx'))
            if not os.path.basename(p).startswith('~')]
    return max(hits, key=os.path.getmtime) if hits else None


def load(folder=None):
    """{plant number: record}. Empty when there is no export, which is
    a perfectly good state - the screens just carry on without it."""
    global _CACHE
    if _CACHE is not None and folder is None:
        return _CACHE
    out = {}
    path = newest(folder)
    if path:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            hdr = [str(c).strip() if c is not None else ''
                   for c in next(rows)]
            ix = {h: i for i, h in enumerate(hdr) if h}

            def cell(r, name):
                i = ix.get(name)
                if i is None or i >= len(r) or r[i] is None:
                    return ''
                return str(r[i]).strip()

            for r in rows:
                if not r:
                    continue
                pn = cell(r, 'Plant Number')
                if not pn:
                    continue
                av = cell(r, 'Availability')
                out[pn] = {
                    'branch': cell(r, 'Branch'),
                    'availability': av,
                    'status': cell(r, 'Availability Status'),
                    'category': cell(r, 'Category'),
                    'type': cell(r, 'Type'),
                    'model': cell(r, 'Model'),
                    'days': cell(r, 'Days in Status'),
                    'cost': cell(r, 'Original Cost'),
                    'wdv': cell(r, 'WDV'),
                    'why': NOT_ISSUABLE.get(_norm(av), ''),
                }
            wb.close()
        except Exception:
            #  A broken export changes nothing. Losing an exclusion is a
            #  nuisance; a store with no screen on the wall is worse.
            out = {}
    if folder is None:
        _CACHE = out
    return out


def record(item):
    return load().get(str(item or '').strip())


def blocked(item):
    """('reserved for someone else', '10') or ('', '').

    Returns the reason this asset must not be recommended, and how many
    days it has been in that state. Empty means either the branch is
    happy with it or the branch has never heard of it - and those two
    are different, which is what `covers` is for.
    """
    rec = record(item)
    if not rec or not rec['why']:
        return '', ''
    return rec['why'], rec['days']


def covers(items):
    """How many of these assets the branch export actually knows about.

    A screen that quietly treats 'no branch record' as 'nothing wrong'
    is lying by omission on four assets in five, so every page using
    this prints the coverage next to the exclusions.
    """
    r = load()
    seen = sum(1 for i in items if str(i) in r)
    return {'known': seen, 'total': len(items), 'file': newest() or ''}


def available_elsewhere(model_words, exclude_branch='', limit=8):
    """Who else in the state has one on the shelf right now.

    The reason to keep the whole export rather than just the K2 rows:
    when the job needs another twenty of something, this is the only
    thing on the machine that can answer where they are.
    """
    words = [w for w in _norm(model_words).split() if w]
    if not words:
        return []
    counts = {}
    for pn, rec in load().items():
        if _norm(rec['availability']) != 'available':
            continue
        if exclude_branch and rec['branch'] == exclude_branch:
            continue
        hay = _norm(rec['model'] + ' ' + rec['type'])
        if all(w in hay for w in words):
            counts[rec['branch']] = counts.get(rec['branch'], 0) + 1
    out = sorted(counts.items(), key=lambda kv: -kv[1])
    return out[:limit]
