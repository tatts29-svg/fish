#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | GONE FROM SITE - what departed, and on what day
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 5 Aug 2026): "if someone searched for asset that has
#  been departed what will show. can we advise its been departed and on
#  what day."
#
#  WHAT SHOWED BEFORE. Nothing - and worse than nothing.
#
#  An asset that leaves site drops out of RENTAL_STOCK completely.
#  There is no "departed" row in that export, so the moment the board
#  is rebuilt the gear simply is not there any more. Search its number
#  and you get "Nothing matches that" - the identical answer you get
#  for a typo. A bloke cannot tell the difference between "that is not
#  a real asset number", "it is hidden", and "it went back to the
#  branch on Tuesday".
#
#  And in the window between the gear leaving and the next build, the
#  board says something worse than nothing: on 5 Aug the store board
#  was still offering a Skid Steer Loader 2.6t as "WE HAVE 1 - ALL OUT
#  WITH CREWS". It had departed at 06:56 the morning before.
#
#  WHERE THE TRUTH LIVES. Not in a diff of yesterday's register against
#  today's - that would only ever be a guess, and it could not tell a
#  departure from a partial export. SiteIQ already records this
#  properly, in STOCKTAKE:
#
#      SIGHTED_STATUS         "Pending Branch Receipt"
#      LAST_SIGHTED_ACTION    "Departure"
#      LAST_SIGHTED_DATE_TIME 04/08/2026 06:56
#      LAST_SIGHTED_BY        Thomas Maher
#
#  232 lines carried that status on 5 Aug 2026, 231 of them Departures.
#  So this module reads, it does not infer. Every date on a screen is
#  the date SiteIQ stamped, not a date this suite worked out.
#
#  THE HONESTY RULES
#    - Only a row that says BOTH "Pending Branch Receipt" AND an action
#      of "Departure" is called departed. Anything else is left alone.
#    - A hidden line is never reported as departed. It is not gone, it
#      is held off the board on purpose, and saying otherwise would be
#      a different lie to the one hiding it silently would be.
#    - Still on the register AND pending branch receipt: it depends
#      WHEN. The two exports are pulled hours apart - on 5 Aug the
#      register was requested 07:48 and the stocktake ran to 10:02, so
#      gear that left at 09:22 is still on a register taken before it
#      went. A departure AFTER the register pull is believed; one
#      BEFORE it, on gear the register still lists, means it came back
#      and nothing is said.
#    - No STOCKTAKE, no claims. An empty index says nothing rather than
#      guessing from an absence.
#    - No money. This runs on the store Wi-Fi like everything else in
#      My Gear and carries no rate, no total, no dollar.
# =====================================================================
from __future__ import print_function

import datetime as dt
import glob
import io
import os
import re

HERE = os.path.dirname(os.path.abspath(__file__))

GONE_STATUS = 'pending branch receipt'
GONE_ACTION = 'departure'


def _txt(v):
    return '' if v is None else str(v).strip()


def _key(s):
    """Match the way a person types, not the way a database stores.
    Case and stray spaces only - never fuzzy, because telling somebody
    the wrong machine went home is worse than telling them nothing."""
    return re.sub(r'\s+', ' ', _txt(s)).strip().upper()


def _when(v):
    """SiteIQ stamps 'dd/mm/yyyy hh:mm'. Anything it cannot parse is
    returned as-is rather than dropped - a date I do not recognise is
    still the date on the record, and inventing a tidier one would be
    the one thing this file is for stopping."""
    if isinstance(v, dt.datetime):
        return v
    s = _txt(v)
    #  SiteIQ stamps these with AM/PM - '03/08/2026 12:42 PM'. Missing
    #  that shape did not crash anything, it just quietly left every
    #  date as text, so the whole list stopped sorting by date and
    #  nothing looked wrong. Order the patterns most-specific first.
    for f in ('%d/%m/%Y %I:%M:%S %p', '%d/%m/%Y %I:%M %p',
              '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y',
              '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return dt.datetime.strptime(s, f)
        except ValueError:
            continue
    return s or None


def _stocktake_path():
    cands = [p for p in
             glob.glob(os.path.join(HERE, 'Data_SiteIQ', 'STOCKTAKE*.xlsx'))
             + glob.glob(os.path.join(HERE, 'STOCKTAKE*.xlsx'))
             if not os.path.basename(p).startswith('~$')]
    return max(cands, key=os.path.getmtime) if cands else None


def pulled_at(path):
    """When SiteIQ was asked for this export - it stamps it on the
    REFERENCE_INFO sheet. Two exports taken hours apart disagree about
    the same asset, and knowing which one is fresher is the difference
    between an honest answer and a confident wrong one."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if 'REFERENCE_INFO' not in wb.sheetnames:
            wb.close()
            return None
        rows = list(wb['REFERENCE_INFO'].iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2:
            return None
        hdr = [_txt(c).upper() for c in rows[0]]
        for i, h in enumerate(hdr):
            if 'REQUESTED_DATE' in h or 'TO_DATE' in h:
                v = _when(rows[1][i] if i < len(rows[1]) else None)
                if isinstance(v, dt.datetime):
                    return v
    except Exception:
        return None
    return None


_CACHE = None


def load(path=None, on_register=None, include_hidden=False):
    """Everything that has gone back to the branch.

    on_register: item numbers / barcodes STILL in RENTAL_STOCK. A row
    in that set is dropped ONLY if it departed before the register was
    pulled - see the note at the check itself. Pass None to skip it.

    include_hidden: HIDING IS A DISPLAY DECISION, DEPARTING IS A FACT.
    The store search must never mention a held line, so it takes the
    default and hidden gear is filtered out. Anything working out an
    asset's LIFECYCLE - plant_life.py - wants the truth regardless of
    whether the line is on the board, or a hidden item that genuinely
    went home gets quietly recorded as still being here.

    Returns {KEY: record} where KEY is both the SKU and the barcode, so
    a bloke can type either one off the sticker."""
    global _CACHE
    if (_CACHE is not None and path is None and on_register is None
            and not include_hidden):
        return _CACHE

    out = {}
    reg_pulled = None
    if on_register:
        c = [x for x in
             glob.glob(os.path.join(HERE, 'Data_SiteIQ', 'RENTAL_STOCK*.xlsx'))
             + glob.glob(os.path.join(HERE, 'RENTAL_STOCK*.xlsx'))
             if not os.path.basename(x).startswith('~$')]
        if c:
            reg_pulled = pulled_at(max(c, key=os.path.getmtime))
    p = path or _stocktake_path()
    if not p or not os.path.isfile(p):
        return out
    try:
        import openpyxl
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb['STOCKTAKE'] if 'STOCKTAKE' in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        hdr = [_txt(c) for c in next(rows)]
        ix = {h: i for i, h in enumerate(hdr) if h}
        need = ('SIGHTED_STATUS', 'LAST_SIGHTED_ACTION', 'DESCRIPTION')
        if any(n not in ix for n in need):
            wb.close()
            return out

        def cell(r, name):
            i = ix.get(name)
            return r[i] if i is not None and i < len(r) else None

        try:
            import hidden_stock
        except Exception:
            hidden_stock = None

        for r in rows:
            if not r:
                continue
            if _txt(cell(r, 'SIGHTED_STATUS')).lower() != GONE_STATUS:
                continue
            if _txt(cell(r, 'LAST_SIGHTED_ACTION')).lower() != GONE_ACTION:
                continue
            sku = _txt(cell(r, 'SKU_NUMBER'))
            bar = _txt(cell(r, 'LATEST_BARCODE'))
            unit = _txt(cell(r, 'STORAGE_UNIT'))
            #  A held line is not a departed line. It is on the shelf
            #  and kept off the board on purpose.
            if (not include_hidden) and hidden_stock and (
                    hidden_stock.hidden(sku, unit)
                    or hidden_stock.hidden(bar, unit)):
                continue
            #  STILL ON THE REGISTER? Then it depends WHEN it left.
            #
            #  "The register wins" was my first rule and it is wrong.
            #  The two exports are pulled at different times - on
            #  5 Aug the register was requested at 07:48 and the
            #  stocktake ran to 10:02. A hose that departed at 09:22
            #  is therefore still sitting on a register that was taken
            #  two hours before it left, and dropping it would have
            #  told a storeman "nothing matches" about gear that had
            #  gone out the gate that morning.
            #
            #  So: a departure AFTER the register was pulled means the
            #  register simply has not caught up - believe it. A
            #  departure BEFORE the register was pulled, on gear the
            #  register still lists, means it came back - believe the
            #  register and say nothing.
            when = _when(cell(r, 'LAST_SIGHTED_DATE_TIME'))
            if on_register and (_key(sku) in on_register
                                or _key(bar) in on_register):
                if not (reg_pulled and isinstance(when, dt.datetime)
                        and when > reg_pulled):
                    continue
            rec = {
                'sku': sku,
                'barcode': bar,
                'desc': _txt(cell(r, 'DESCRIPTION')),
                'unit': unit,
                'when': when,
                'who': _txt(cell(r, 'LAST_SIGHTED_BY')),
                'family': _txt(cell(r, 'PRODUCT_FAMILY')),
                'product': _txt(cell(r, 'PRODUCT')),
            }
            for k in (sku, bar):
                if k:
                    out[_key(k)] = rec
        wb.close()
    except Exception:
        #  A broken stocktake tells the board nothing. It must never be
        #  able to take the store search down.
        return {}
    if path is None and on_register is None and not include_hidden:
        _CACHE = out
    return out


def on_register_keys(rental_path=None):
    """Item numbers and barcodes still in RENTAL_STOCK - the live set a
    departure has to be checked against."""
    keys = set()
    p = rental_path
    if not p:
        c = [x for x in
             glob.glob(os.path.join(HERE, 'Data_SiteIQ', 'RENTAL_STOCK*.xlsx'))
             + glob.glob(os.path.join(HERE, 'RENTAL_STOCK*.xlsx'))
             if not os.path.basename(x).startswith('~$')]
        p = max(c, key=os.path.getmtime) if c else None
    if not p or not os.path.isfile(p):
        return keys
    try:
        import openpyxl
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = (wb['RENTAL_STOCK'] if 'RENTAL_STOCK' in wb.sheetnames
              else wb.active)
        rows = ws.iter_rows(values_only=True)
        hdr = [_txt(c) for c in next(rows)]
        ix = {h: i for i, h in enumerate(hdr) if h}
        for r in rows:
            if not r:
                continue
            for col in ('ITEM_NUMBER', 'ITEM_BARCODE'):
                i = ix.get(col)
                if i is not None and i < len(r):
                    k = _key(r[i])
                    if k:
                        keys.add(k)
        wb.close()
    except Exception:
        return set()
    return keys


def say_when(rec):
    """The date, the way a person says it. '4 Aug at 6:56am'."""
    w = rec.get('when')
    if isinstance(w, dt.datetime):
        h = w.strftime('%I:%M%p').lstrip('0').lower()
        return w.strftime('%-d %b') if os.name != 'nt' else w.strftime('%d %b').lstrip('0'), h
    return (_txt(w), '')


def payload(limit_desc=True):
    """What the built screens carry: a small map the page can look up
    without shipping the whole stocktake to a phone."""
    reg = on_register_keys()
    idx = load(on_register=reg)
    #  one entry per record, not per key, plus a by-description index so
    #  a NAME search ("skid steer") finds it as well as a number
    seen, items = {}, []
    for k, rec in idx.items():
        rid = id(rec)
        if rid in seen:
            continue
        seen[rid] = len(items)
        w = rec.get('when')
        items.append({
            'k': rec['sku'] or rec['barcode'],
            'b': rec['barcode'],
            'd': rec['desc'],
            'u': rec['unit'],
            'w': (w.strftime('%d/%m/%Y %H:%M')
                  if isinstance(w, dt.datetime) else _txt(w)),
            'y': rec['who'],
        })
    keys = {k: seen[id(rec)] for k, rec in idx.items()}
    return {'items': items, 'keys': keys}


def main():
    reg = on_register_keys()
    idx = load(on_register=reg)
    seen = set()
    recs = []
    for rec in idx.values():
        if id(rec) in seen:
            continue
        seen.add(id(rec))
        recs.append(rec)
    recs.sort(key=lambda r: (r['when'] if isinstance(r['when'], dt.datetime)
                             else dt.datetime.min), reverse=True)
    print('=' * 70)
    print(' COATES | GONE FROM SITE')
    print('=' * 70)
    print(' Read from : ' + (os.path.basename(_stocktake_path() or '-')))
    print(' Still on the register : {:,} item/barcode key(s)'.format(len(reg)))
    print(' Departed  : {} line(s) pending branch receipt'.format(len(recs)))
    print('')
    if not recs:
        print(' Nothing has departed, or the stocktake does not say so.')
        print(' Nothing is guessed here - no stocktake means no claims.')
        return 0
    print(' {:<22} {:<38} {:<17} {}'.format(
        'Code', 'What it is', 'Departed', 'Signed off by'))
    print(' ' + '-' * 96)
    for r in recs[:40]:
        w = r['when']
        ws = (w.strftime('%d %b %Y %H:%M') if isinstance(w, dt.datetime)
              else _txt(w))
        print(' {:<22} {:<38} {:<17} {}'.format(
            (r['sku'] or r['barcode'])[:22], r['desc'][:38], ws, r['who'][:22]))
    if len(recs) > 40:
        print('')
        print(' ... and {} more. All {} are on the board - this list is'
              .format(len(recs) - 40, len(recs)))
        print(' cut at 40 for the screen, not in the data.')
    return 0


if __name__ == '__main__':
    import sys
    sys.exit(main())
