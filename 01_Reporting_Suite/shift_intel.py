#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | SHIFT & COUNTER INTEL - the engine
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 3 Aug 2026): "I really want to know what each shift is
#  doing ... graphs on when they are busy at the counter and not busy.
#  days 6am to 6pm, night 6pm to 6am ... full transparency on what
#  happens. To the point the times between scans."
#
#  This reads the two feeds that carry a CLOCK and turns them into what
#  each shift actually did:
#
#    TRANSACTIONS   every issue and return, to the second
#    STOCKTAKE      every sighting, to the minute, with who counted it
#
#  ---------------------------------------------------------------
#  THE 5-SECOND RULE, AND WHY THE FIRST ANSWER WAS WRONG
#  ---------------------------------------------------------------
#  Andrew's rule: "scan look then scan look should never ever be less
#  than 5 seconds". Straight off the data, the fastest minute in the
#  whole shut belongs to Andrew himself - 70 scans in sixty seconds.
#  A tool that flagged that would have been worse than useless.
#
#  Those 70 were seventy identical Rubbish Chutes in one storage unit,
#  action blank, not a stocktake at all - a stack being moved, counted
#  as a stack, exactly as anyone would. Same for 50 Barriers and 43
#  identical air fittings.
#
#  So the rule is applied to what it was written about - a bloke
#  walking a shelf, looking at a DIFFERENT thing each time:
#     * only rows whose action is Stocktake
#     * only DISTINCT descriptions inside the minute (a stack of
#       identical stock is one look, however many units)
#     * 13+ distinct in a minute = under 5 seconds a look
#
#  On the 3 Aug pull that leaves 7 minutes out of 452, and only one
#  SUSTAINED run - four consecutive minutes. That distinction is the
#  whole value: one fast minute is a shelf of obvious gear, four in a
#  row is a pattern. Isolated minutes are reported as noise and named
#  as such, so nobody gets a conversation they did not earn.
# =====================================================================
import datetime as dt
import re
from collections import Counter, defaultdict

#  Andrew's shift boundary, 3 Aug 2026: day 06:00-18:00, night after.
DAY_FROM, DAY_TO = 6, 18

#  A look at a different thing, per his rule. 60/5 = 12, so the 13th
#  distinct item inside one minute is the first that cannot have had
#  five seconds.
LOOKS_PER_MIN = 12


def _ts(v):
    """SiteIQ writes its clocks as text, in two shapes:
       '31/07/2026'                  (a date column)
       '31/07/2026 03:39 PM'         (stocktake, minute precision)
    plus a separate '10:39:11' time column on transactions."""
    if isinstance(v, dt.datetime):
        return v
    if isinstance(v, dt.date):
        return dt.datetime(v.year, v.month, v.day)
    s = str(v or '').strip()
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})'
                 r'(?:\s+(\d{1,2}):(\d{2})(?::(\d{2}))?\s*(AM|PM)?)?', s, re.I)
    if not m:
        return None
    d, mo, y, h, mi, se, ap = m.groups()
    h = int(h or 0)
    ap = (ap or '').upper()
    if ap == 'PM' and h != 12:
        h += 12
    if ap == 'AM' and h == 12:
        h = 0
    try:
        return dt.datetime(int(y), int(mo), int(d), h, int(mi or 0),
                           int(se or 0))
    except ValueError:
        return None


def _clock(v):
    """'10:39:11' -> (10, 39, 11)."""
    m = re.match(r'(\d{1,2}):(\d{2})(?::(\d{2}))?', str(v or '').strip())
    if not m:
        return None
    return int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)


def shift_of(when):
    """DAY or NIGHT, and the shift's own date.

    A night shift starting 18:00 Monday runs into Tuesday morning, and
    both halves belong to Monday night - otherwise every night's work
    is split across two rows and neither reads right.
    """
    if DAY_FROM <= when.hour < DAY_TO:
        return 'DAY', when.date()
    if when.hour < DAY_FROM:
        return 'NIGHT', (when - dt.timedelta(days=1)).date()
    return 'NIGHT', when.date()


def read_counter(txn_path, days=21):
    """Every issue and return that carries a clock, newest first."""
    import openpyxl
    out = []
    wb = openpyxl.load_workbook(txn_path, read_only=True, data_only=True)
    for sn in ('TRANSACTION_CHARGES', 'TRANSACTION_WITHOUT_CHARGES',
               'CUSTOMER_CONTRACTOR_EQUIP'):
        if sn not in wb.sheetnames:
            continue
        rows = wb[sn].iter_rows(values_only=True)
        try:
            hdr = [str(c or '').strip() for c in next(rows)]
        except StopIteration:
            continue
        ix = {h: i for i, h in enumerate(hdr)}

        def col(*names):
            for n in names:
                if n in ix:
                    return ix[n]
            return None

        c_item = col('SKU/ITEM_NUMBER', 'ITEM_NUMBER')
        c_desc = col('SKU/ITEM DESCRIPTION', 'ITEM_DESCRIPTION')
        c_who = col('HIRER_NAME')
        c_co = col('EMPLOYER_NAME', 'COMPANY_NAME')
        c_fam = col('PRODUCT_FAMILY')
        pairs = ((col('TRAN_START_DATE'), col('TRAN_START_TIME'), 'OUT'),
                 (col('TRAN_END_DATE'), col('TRAN_END_TIME'), 'BACK'))
        for r in rows:
            if not r:
                continue
            for cd, ct, way in pairs:
                if cd is None or cd >= len(r):
                    continue
                d = _ts(r[cd])
                if not d:
                    continue
                hms = _clock(r[ct]) if (ct is not None and ct < len(r)) \
                    else None
                if hms:
                    d = d.replace(hour=hms[0], minute=hms[1], second=hms[2])
                out.append({
                    'at': d, 'way': way,
                    'i': str(r[c_item] or '').strip() if c_item is not None
                         else '',
                    'n': str(r[c_desc] or '').strip()[:52] if c_desc is not None
                         else '',
                    'w': str(r[c_who] or '').strip()[:28] if c_who is not None
                         else '',
                    'co': str(r[c_co] or '').strip()[:30] if c_co is not None
                          else '',
                    'f': str(r[c_fam] or '').strip()[:26] if c_fam is not None
                         else '',
                })
    wb.close()
    if out:
        newest = max(x['at'] for x in out)
        cut = newest - dt.timedelta(days=days)
        out = [x for x in out if x['at'] >= cut]
    out.sort(key=lambda x: x['at'], reverse=True)
    return out


def read_sightings(stocktake_path):
    """Every stocktake sighting with a clock and a name."""
    import openpyxl
    out = []
    wb = openpyxl.load_workbook(stocktake_path, read_only=True,
                                data_only=True)
    ws = wb['STOCKTAKE'] if 'STOCKTAKE' in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c or '').strip() for c in next(rows)]
    ix = {h: i for i, h in enumerate(hdr)}
    need = ('LAST_SIGHTED_DATE_TIME', 'LAST_SIGHTED_BY')
    if not all(n in ix for n in need):
        wb.close()
        return out
    for r in rows:
        if not r:
            continue
        t = _ts(r[ix['LAST_SIGHTED_DATE_TIME']])
        w = str(r[ix['LAST_SIGHTED_BY']] or '').strip()
        if not (t and w):
            continue
        out.append({
            'at': t, 'w': w,
            'act': str(r[ix.get('LAST_SIGHTED_ACTION', -1)] or '').strip(),
            'u': str(r[ix.get('STORAGE_UNIT', -1)] or '').strip(),
            'd': str(r[ix.get('DESCRIPTION', -1)] or '').strip(),
        })
    wb.close()
    return out


def hourly(events, key='at'):
    """Counter load by hour of the day, 0-23."""
    h = [0] * 24
    for e in events:
        h[e[key].hour] += 1
    return h


def by_shift(counter, sight):
    """What each shift did, split by the areas Andrew names: gear out,
    gear back, and stocktake counting."""
    out = {}

    def slot(when):
        s, d = shift_of(when)
        return out.setdefault((d, s), {
            'date': d, 'shift': s, 'out': 0, 'back': 0, 'count': 0,
            'units': Counter(), 'people': Counter(), 'fam': Counter()})

    for e in counter:
        b = slot(e['at'])
        b['out' if e['way'] == 'OUT' else 'back'] += 1
        if e['f']:
            b['fam'][e['f']] += 1
    for s in sight:
        if s['act'] != 'Stocktake':
            continue
        b = slot(s['at'])
        b['count'] += 1
        if s['u']:
            b['units'][s['u']] += 1
        b['people'][s['w']] += 1
    return sorted(out.values(), key=lambda x: (x['date'], x['shift']),
                  reverse=True)


def scan_pace(sight):
    """Andrew's 5-second rule, applied to what it was written about.

    Returns runs of consecutive fast minutes. A run of one is noise and
    is reported as such; a run of several is the pattern worth a word.
    """
    per = defaultdict(set)
    for s in sight:
        if s['act'] != 'Stocktake':
            continue           # a stack being moved is not a shelf walk
        per[(s['w'], s['at'].replace(second=0))].add(s['d'])

    fast = sorted(k for k, v in per.items() if len(v) > LOOKS_PER_MIN)
    runs, cur = [], []
    for k in fast:
        if cur and k[0] == cur[-1][0] and \
                (k[1] - cur[-1][1]).total_seconds() <= 60:
            cur.append(k)
        else:
            if cur:
                runs.append(cur)
            cur = [k]
    if cur:
        runs.append(cur)

    out = []
    for run in runs:
        looks = sum(len(per[k]) for k in run)
        mins = len(run)
        out.append({
            'who': run[0][0],
            'from': run[0][1],
            'to': run[-1][1] + dt.timedelta(minutes=1),
            'mins': mins,
            'looks': looks,
            'gap': round(60.0 * mins / looks, 1),
            'sustained': mins >= 3,
        })
    out.sort(key=lambda r: (-r['mins'], r['gap']))
    return {
        'runs': out,
        'minutes': len(per),
        'fastMinutes': len(fast),
        'counters': len({k[0] for k in per}),
        'looks': sum(len(v) for v in per.values()),
    }


def queue(counter):
    """The rush at the window, counted in PEOPLE not rows.

    Counting rows makes a stack of seventy chutes look like seventy
    blokes queueing. Distinct hirers served inside one minute is the
    honest measure of how many were standing there.
    """
    per = defaultdict(set)
    for e in counter:
        if e['w']:
            per[e['at'].replace(second=0)].add(e['w'])
    sizes = {k: len(v) for k, v in per.items()}
    by_hour = defaultdict(int)
    for k, n in sizes.items():
        if n >= 3:
            by_hour[k.hour] += 1
    worst = sorted(sizes.items(), key=lambda x: -x[1])[:8]
    return {
        'minutes': len(sizes),
        'busy': sum(1 for n in sizes.values() if n >= 3),
        'peak': max(sizes.values()) if sizes else 0,
        'worst': worst,
        'byHour': [by_hour.get(h, 0) for h in range(24)],
    }


def durations(txn_path):
    """How long gear actually stays out, off paired start/end times."""
    import openpyxl
    hours = []
    wb = openpyxl.load_workbook(txn_path, read_only=True, data_only=True)
    for sn in ('TRANSACTION_CHARGES', 'CUSTOMER_CONTRACTOR_EQUIP'):
        if sn not in wb.sheetnames:
            continue
        rows = wb[sn].iter_rows(values_only=True)
        try:
            hdr = [str(c or '').strip() for c in next(rows)]
        except StopIteration:
            continue
        ix = {h: i for i, h in enumerate(hdr)}
        if 'TRAN_START_DATE' not in ix or 'TRAN_END_DATE' not in ix:
            continue
        for r in rows:
            if not r:
                continue
            a, b = _ts(r[ix['TRAN_START_DATE']]), _ts(r[ix['TRAN_END_DATE']])
            if not (a and b):
                continue
            ta = _clock(r[ix['TRAN_START_TIME']]) if 'TRAN_START_TIME' in ix \
                else None
            tb = _clock(r[ix['TRAN_END_TIME']]) if 'TRAN_END_TIME' in ix \
                else None
            if ta:
                a = a.replace(hour=ta[0], minute=ta[1], second=ta[2])
            if tb:
                b = b.replace(hour=tb[0], minute=tb[1], second=tb[2])
            h = (b - a).total_seconds() / 3600.0
            if 0 <= h < 24 * 45:
                hours.append(h)
    wb.close()
    if not hours:
        return None
    hours.sort()
    #  the buckets a store hand thinks in, not equal-width maths
    buckets = [('Under 2h', 0, 2), ('2-6h', 2, 6), ('6-12h', 6, 12),
               ('12-24h', 12, 24), ('1-3 days', 24, 72),
               ('3-7 days', 72, 168), ('Over a week', 168, 10 ** 6)]
    counts = []
    for label, lo, hi in buckets:
        counts.append((label, sum(1 for h in hours if lo <= h < hi)))
    return {
        'n': len(hours),
        'median': hours[len(hours) // 2],
        'shift': sum(1 for h in hours if h <= 12),
        'week': sum(1 for h in hours if h > 168),
        'buckets': counts,
    }


def aisle_coverage(rental_path, sight):
    """Which aisles get walked, and which have never been laid eyes on."""
    import openpyxl
    units = Counter()
    wb = openpyxl.load_workbook(rental_path, read_only=True, data_only=True)
    ws = wb['RENTAL_STOCK'] if 'RENTAL_STOCK' in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c or '').strip() for c in next(rows)]
    ix = {h: i for i, h in enumerate(hdr)}
    if 'STORAGE_UNIT' in ix:
        for r in rows:
            if not r:
                continue
            u = str(r[ix['STORAGE_UNIT']] or '').strip()
            if u:
                units[u] += 1
    wb.close()
    counted = Counter()
    last = {}
    for x in sight:
        if x['act'] != 'Stocktake' or not x['u']:
            continue
        counted[x['u']] += 1
        if x['u'] not in last or x['at'] > last[x['u']]:
            last[x['u']] = x['at']
    rows_out = []
    for u, n in units.most_common():
        rows_out.append({'unit': u, 'assets': n, 'counted': counted.get(u, 0),
                         'last': last.get(u)})
    return {
        'rows': rows_out,
        'aisles': len(units),
        'walked': sum(1 for r in rows_out if r['counted']),
        'unseen': sum(r['assets'] for r in rows_out if not r['counted']),
    }
