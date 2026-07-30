#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | K2 COST TRACKING SNAPSHOT
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Reads the Daily Tracking tab (forecast vs actual) and the live plant
#  position, and tells the cost story. Dark Coates dashboard framed on
#  white A4 sheets so it carries cleanly into PDF / email. Read-only -
#  the workbook is never changed.
#
#  LAYOUT v2 - 24 Jul 2026 (one story per page)
#    - Each page carries ONE story: the position, cost streams, plant
#      utilisation, the landing, people & assurance, the team. If a
#      section has nothing to say (no charges, no roster) it collapses
#      to a line or drops off - it never pads a page.
#    - Numbers live once: the KPI tiles carry the figures, the story
#      carries the why. Only the headline landing/plan repeat, because
#      they ARE the call.
#    - Bigger type, lifted contrast - readable at 05:30 on a phone.
#    - Outputs land in Reports\<YYYY-MM-DD>\ (one folder per day) via
#      report_paths.py. Page PNGs are kept there too, next to the
#      HTML / PDF / EML, so the whole day lives together.
# =====================================================================

import os
import re
import glob
import html
import json
import datetime as dt
import openpyxl

import report_paths
import recipients

HERE = os.path.dirname(os.path.abspath(__file__))


def _gfind(pattern):
    r"""SiteIQ export candidates: Data_SiteIQ\ first, suite root second -
    saving over the top works in either place."""
    hits = []
    for d in (os.path.join(HERE, "Data_SiteIQ"), HERE):
        hits += [p for p in glob.glob(os.path.join(d, pattern))
                 if not os.path.basename(p).startswith("~$")]
    return hits
ORG = "#F26222"
GOOD = "#0ca30c"
AMBER = "#fab219"
BAD = "#d03b3b"
FCAST = "#F26222"


def money0(v):
    try:
        v = float(v)
    except Exception:
        return str(v)
    return ("-$" + "{:,.0f}".format(abs(v))) if v < 0 else "$" + "{:,.0f}".format(v)


def find_workbook():
    c = (glob.glob(os.path.join(HERE, "Cement_Australia_Report*K2*.xlsm"))
         or glob.glob(os.path.join(HERE, "*.xlsm")))
    if not c:
        raise SystemExit("  ! No K2 workbook (.xlsm) found next to this script.")
    return max(c, key=os.path.getmtime)


# ---- cost story: Daily Tracking trued up to SiteIQ's invoiced truth ------
# THE ALIGNMENT RULE (24 Jul 2026, per A. Fisher - the client sees TRUE):
#   * DAILY_SUMMARY is SiteIQ's invoiced truth, and it's retrospective -
#     each fresh export can revise prior days. Plant & tooling hire is
#     therefore shown as SiteIQ charges it: ONE hire line, trued to the
#     invoiced figure for every day SiteIQ has charged. Those days are
#     LOCKED at the invoice.
#   * Days already run that SiteIQ hasn't invoiced yet (usually just
#     today) carry the tracked live figure, clearly marked PROVISIONAL -
#     it can move a little until the invoice lands, then it locks.
#   * Radio / gas / personnel / accommodation are contract arrangements
#     that don't ride in SiteIQ's HIRE column - they stay sourced from
#     the K2 workbook (their own truth).
#   * Nothing is ever scaled, split or estimated to force a match - the
#     true-up page shows tracked v invoiced day by day, drift and all.

STREAMS = [("Transport", 20, 21),
           ("Personnel / Labour", 14, 15), ("Accommodation", 17, 18),
           ("Radio", 8, 9), ("Gas Monitor", 11, 12),
           ("Damage Recovery", 23, 24)]

# Streams recognised at their forecast value for days that have already passed:
# the cost is incurred when it happens (e.g. mobilisation transport on 20 Jul,
# gear delivered to site) but billed in arrears, so a $0 actual understates the
# true spend. Recognise it at the contracted forecast for elapsed days. Per A. Fisher.
RECOGNISE_AT_FORECAST = {"Transport"}


def read_daily_summary():
    """SiteIQ's invoiced truth: {date: {'hire': $, 'transport': $, ...}}
    from the newest DAILY_SUMMARY export, plus the export window end.
    Missing file = empty dict - the snapshot then says plainly that hire
    is tracked-only this run, it never pretends."""
    hits = [p for p in _gfind("DAILY_SUMMARY*.xlsx")
            if not os.path.basename(p).startswith("~$")]
    if not hits:
        return {}, None, None
    path = max(hits, key=os.path.getmtime)
    days = {}
    asat = None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ref = list(wb["REFERENCE_INFO"].iter_rows(values_only=True))
            asat = str(ref[1][2])          # DATE (TO) - the export window end
        except Exception:
            pass
        ws = wb["DAILY_SUMMARY"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = {str(v).strip().upper(): i for i, v in enumerate(rows[0]) if v}
        for r in rows[1:]:
            label = str(r[0] or "")
            m = re.match(r"TOTAL FOR:\s*(\d{2})/(\d{2})/(\d{4})", label)
            if not m:
                continue
            d = dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))

            def col(name):
                i = hdr.get(name)
                v = r[i] if i is not None else 0
                try:
                    return float(v)
                except (TypeError, ValueError):
                    return 0.0
            days[d] = {"hire": col("HIRE"), "transport": col("TRANSPORT"),
                       "labour": col("LABOUR"), "damage": col("DAMAGE"),
                       "service": col("SERVICE"), "total": col("TOTAL"),
                       "consumable": col("CONSUMABLE"), "fuel": col("FUEL"),
                       "disposal": col("DISPOSAL"), "other": col("OTHER"),
                       "waiver": col("WAIVER"), "stamp": col("STAMP DUTY")}
        wb.close()
    except Exception:
        return {}, None, None
    return days, path, asat


SUBHIRE_RATE = 72.59      # Vantage welders, per unit per day - the client
                          # charge confirmed by A. Fisher 24 Jul 2026 (register
                          # rate). On site = charged, same as site plant.
                          # CLIENT-SAFE NAME: the reports say "Welders" only -
                          # never the arrangement behind them.


def read_subhire_welders():
    """The separately-invoiced welders (barcode SUB*): SiteIQ moves them but
    never charges them, so they ride outside DAILY_SUMMARY - like radios.
    Each unit charges from its first evidenced day on site (earliest SiteIQ
    movement or on-hire date - never an assumed arrival)."""
    arrivals = {}

    def note(bc, d):
        if bc and d and (bc not in arrivals or d < arrivals[bc]):
            arrivals[bc] = d

    def au_date(v):
        s = str(v or "").strip().split()[0]
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return dt.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    # movement evidence (historical, incl. pool days)
    tx = _gfind("TRANSACTIONS*.xlsx")
    if tx:
        try:
            twb = openpyxl.load_workbook(max(tx, key=os.path.getmtime),
                                         read_only=True, data_only=True)
            if "CUSTOMER_CONTRACTOR_EQUIP" in twb.sheetnames:
                rows = list(twb["CUSTOMER_CONTRACTOR_EQUIP"].iter_rows(values_only=True))
                hdr = {str(v).strip(): i for i, v in enumerate(rows[0]) if v}
                for r in rows[1:]:
                    bc = str(r[hdr.get("LATEST_BARCODE", -1)] or "").upper()
                    if bc.startswith("SUB"):
                        note(bc, au_date(r[hdr.get("TRAN_START_DATE", -1)]))
            twb.close()
        except Exception:
            pass
    # current position evidence
    rs = _gfind("RENTAL_STOCK*.xlsx")
    if rs:
        try:
            rwb = openpyxl.load_workbook(max(rs, key=os.path.getmtime),
                                         read_only=True, data_only=True)
            if "RENTAL_STOCK" in rwb.sheetnames:
                rows = list(rwb["RENTAL_STOCK"].iter_rows(values_only=True))
                hdr = {str(v).strip(): i for i, v in enumerate(rows[0]) if v}
                for r in rows[1:]:
                    bc = str(r[hdr.get("ITEM_BARCODE", -1)] or "").upper()
                    if bc.startswith("SUB"):
                        note(bc, au_date(r[hdr.get("ON_HIRE_DATE", -1)]))
            rwb.close()
        except Exception:
            pass
    return {"arrivals": arrivals, "rate": SUBHIRE_RATE}


#  Seed for MISC_CLAIM_NAMES.txt - the four lines Andrew named on
#  30 Jul 2026, each checked to the cent (qty x unit price = the line).
_MISC_NAMES_SEED = """\
# MISC claim lines - what they actually are.
# SiteIQ books ad-hoc consumable claims as "MISC" with no description
# at all, so the cost snapshot fills the words in from this file.
# One line each:   amount = what it is
# Add a new line whenever a new MISC claim appears, save, run 02.
1090.50 = Ledlenser MH10 600lm Rechargeable Headlamp - 5 x $218.10
511.20 = Paramount Safety Bollard Stem - 24 x $21.30
389.76 = Paramount Bollard Bases 6KG - 24 x $16.24
306.00 = Squids 3155 Elastic Tool Lanyard and Clamp - 20 x $15.30
"""


def read_misc_names():
    """SiteIQ books ad-hoc consumable claims as bare "MISC" lines - the
    export genuinely carries no description (SKU/ITEM_NUMBER says MISC,
    the description column is blank). This file supplies the real words,
    keyed by the line's dollar amount. Created with Andrew's four named
    lines on first run; he adds a line in Notepad whenever a new MISC
    claim appears. Lives beside the suite like every register - never
    in git, never in an update zip."""
    path = os.path.join(HERE, "MISC_CLAIM_NAMES.txt")
    if not os.path.isfile(path):
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(_MISC_NAMES_SEED)
        except OSError:
            return {}
    names = {}
    try:
        with open(path, encoding="utf-8-sig") as f:
            for ln in f:
                ln = ln.strip()
                if not ln or ln.startswith("#") or "=" not in ln:
                    continue
                a, d = ln.split("=", 1)
                try:
                    amt = round(float(a.replace("$", "").replace(",", "")), 2)
                except ValueError:
                    continue
                if d.strip():
                    names[amt] = d.strip()
    except OSError:
        pass
    return names


def read_service_invoiced():
    """Progress claims booked into SiteIQ as PRODUCT_CATEGORY 'Service'
    (labour, accommodation and the like). These are LUMP claims covering a
    stretch of days - they land in DAILY_SUMMARY under SERVICE and in
    TRANSACTION_CHARGES with the dollars in the SERVICE ($) column. The
    cost story stays on the daily accrual (so days read evenly); these
    claims are reconciled openly on the true-up page - never double
    counted, never treated as one day's cost. Per A. Fisher 24 Jul 2026."""
    out = {"labour": 0.0, "accom": 0.0, "other": [], "total": 0.0,
           "rows": [], "unnamed": 0}
    misc_names = read_misc_names()
    tx = _gfind("TRANSACTIONS*.xlsx")
    if not tx:
        return out
    try:
        twb = openpyxl.load_workbook(max(tx, key=os.path.getmtime),
                                     read_only=True, data_only=True)
        if "TRANSACTION_CHARGES" not in twb.sheetnames:
            twb.close()
            return out
        rows = list(twb["TRANSACTION_CHARGES"].iter_rows(values_only=True))
        twb.close()
    except Exception:
        return out
    if not rows:
        return out
    hdr = {str(v).strip().upper(): i for i, v in enumerate(rows[0]) if v}

    def cell(r, name):
        i = hdr.get(name)
        return r[i] if i is not None and i < len(r) else None

    for r in rows[1:]:
        cat = str(cell(r, "PRODUCT_CATEGORY") or "").strip().lower()
        if cat != "service":
            continue
        amt = cell(r, "SERVICE ($)")
        try:
            amt = float(amt)
        except (TypeError, ValueError):
            amt = 0.0
        if not amt:
            try:
                amt = float(cell(r, "TOTAL_CHARGE ($)") or 0)
            except (TypeError, ValueError):
                amt = 0.0
        if not amt:
            continue
        desc = str(cell(r, "SKU/ITEM DESCRIPTION") or cell(r, "SKU/ITEM_NUMBER") or "").strip()
        if not desc or desc.upper() in ("MISC", "MISCELLANEOUS"):
            #  SiteIQ sent no words at all - fill them in from Andrew's
            #  names file (keyed by the line's exact dollar amount)
            nm = misc_names.get(round(amt, 2))
            if nm:
                desc = nm
            else:
                desc = "MISC"
                out["unnamed"] += 1
        du = desc.upper()
        when = str(cell(r, "SHIFT_CHARGE_DATE FROM") or cell(r, "TRAN_START_DATE") or "")
        out["rows"].append((when, desc, amt))
        out["total"] += amt
        if "LABOUR" in du or "LABOR" in du or "PERSONNEL" in du:
            out["labour"] += amt
        elif "ACCOM" in du:
            out["accom"] += amt
        else:
            out["other"].append((desc, amt))
    return out


def read_hire_split(wb):
    """Split SiteIQ's ONE invoiced hire line into Plant and Tooling - by
    ITEM, from SiteIQ's own charge lines (TRANSACTION_CHARGES), classified
    against the site's own plant identity: the workbook's Plant Equipment
    register, Plant_ID_Register.csv, and everything RENTAL_STOCK holds in
    the plant storage units (site plant, rubbish chutes, barriers - the
    same definition the prefill and plant dashboard use). Every charged
    line lands in exactly one bucket, so plant + tooling always sums to
    the invoiced total to the cent - a partition, never an estimate.
    Sub-hired welders (SUB*) are excluded outright: they are invoiced
    separately and carry their own stream. Returns None if TRANSACTIONS
    is missing - the caller then keeps the combined line, honestly."""
    def norm(v):
        return str(v).strip().upper() if v is not None else ""
    plant_keys = set()
    # the workbook's Plant Equipment register (asset numbers)
    try:
        for r in list(wb["Plant Equipment"].iter_rows(values_only=True))[1:]:
            a = norm(r[4]) if len(r) > 4 else ""
            if a and not a.startswith("VARIANCE"):
                plant_keys.add(a)
    except Exception:
        pass
    # the Plant ID register (Asset Number -> Plant ID)
    try:
        import csv
        with open(os.path.join(HERE, "Plant_ID_Register.csv"), newline="") as f:
            for row in csv.DictReader(f):
                a = norm(row.get("Asset Number"))
                if a:
                    plant_keys.add(a)
    except Exception:
        pass
    # RENTAL_STOCK plant storage units (current position)
    rs = _gfind("RENTAL_STOCK*.xlsx")
    if rs:
        try:
            rwb = openpyxl.load_workbook(max(rs, key=os.path.getmtime),
                                         read_only=True, data_only=True)
            rows = list(rwb["RENTAL_STOCK"].iter_rows(values_only=True))
            rwb.close()
            H = {norm(v): i for i, v in enumerate(rows[0]) if v}
            PLANT_SU = {"CEMENT SITE PLANT", "CEMENT RUBBISH CHUTES",
                        "CEMENT BARRIERS"}
            for r in rows[1:]:
                if norm(r[H.get("STORAGE_UNIT", -1)]) in PLANT_SU:
                    for col in ("ITEM_BARCODE", "SKU_NUMBER", "ITEM_NUMBER"):
                        if col in H:
                            k = norm(r[H[col]])
                            if k and not k.startswith("SUB"):
                                plant_keys.add(k)
        except Exception:
            pass
    plant_keys.discard("")
    if not plant_keys:
        return None

    def is_plant(bc):
        if bc in plant_keys:
            return True
        return any(bc.startswith(k + "-") or k.startswith(bc + "-")
                   for k in plant_keys)

    tx = _gfind("TRANSACTIONS*.xlsx")
    if not tx:
        return None
    try:
        twb = openpyxl.load_workbook(max(tx, key=os.path.getmtime),
                                     read_only=True, data_only=True)
        if "TRANSACTION_CHARGES" not in twb.sheetnames:
            twb.close()
            return None
        rows = list(twb["TRANSACTION_CHARGES"].iter_rows(values_only=True))
        twb.close()
    except Exception:
        return None
    if not rows:
        return None
    H = {norm(v): i for i, v in enumerate(rows[0]) if v}
    if "HIRE_CHARGE ($)" not in H or "LATEST_BARCODE" not in H:
        return None
    plant = tool = 0.0
    for r in rows[1:]:
        try:
            hire = float(r[H["HIRE_CHARGE ($)"]] or 0)
        except (TypeError, ValueError):
            hire = 0.0
        if not hire:
            continue
        bc = norm(r[H["LATEST_BARCODE"]])
        if bc.startswith("SUB"):
            continue          # welders: separate invoice, separate stream
        if is_plant(bc):
            plant += hire
        else:
            tool += hire
    return {"plant": plant, "tool": tool, "total": plant + tool,
            "keys": len(plant_keys)}


def read_hire_daily(wb):
    """The invoiced hire, PER DAY, split three ways: Plant, Tooling and
    Gear. (Andrew, 30 Jul 2026: "we need to know the whole plant gear
    and tooling how much for each for each day.")

    Every charged line is spread evenly across its own SHIFT_CHARGE
    date span and dropped into exactly one bucket, so the three columns
    sum to SiteIQ's daily invoice - PROVEN to tie to the cent against
    DAILY_SUMMARY on every day the TRANSACTIONS pull covers. A
    partition of the invoice, never an estimate.

    The buckets, by the store's own identity:
      Plant   - the plant registers + the plant storage units (same
                definition as read_hire_split, the prefill and the
                plant dashboard - ONE definition across the suite)
      Tooling - barcodes living in the tool aisles (Tooling,
                Hydraulics- Hi Torque)
      Gear    - everything else: electrical, rigging, welding, air,
                lighting and the rest of the store

    ONE TIMING TRAP, named where it bites: SiteIQ posts a day's charge
    lines the NEXT morning around 09:30. A TRANSACTIONS export pulled
    at 06:30 will carry $0 against a day DAILY_SUMMARY has already
    invoiced - the caller must show a dash and say "re-pull after
    09:30", never a silent zero.
    """
    def norm(v):
        return str(v).strip().upper() if v is not None else ""

    #  the plant identity - same three sources as read_hire_split
    plant_keys = set()
    try:
        for r in list(wb["Plant Equipment"].iter_rows(values_only=True))[1:]:
            a = norm(r[4]) if len(r) > 4 else ""
            if a and not a.startswith("VARIANCE"):
                plant_keys.add(a)
    except Exception:
        pass
    try:
        import csv
        with open(os.path.join(HERE, "Plant_ID_Register.csv"), newline="") as f:
            for row in csv.DictReader(f):
                a = norm(row.get("Asset Number"))
                if a:
                    plant_keys.add(a)
    except Exception:
        pass
    unit_of = {}
    rs = _gfind("RENTAL_STOCK*.xlsx")
    if rs:
        try:
            rwb = openpyxl.load_workbook(max(rs, key=os.path.getmtime),
                                         read_only=True, data_only=True)
            rows = list(rwb["RENTAL_STOCK"].iter_rows(values_only=True))
            rwb.close()
            H = {norm(v): i for i, v in enumerate(rows[0]) if v}
            PLANT_SU = {"CEMENT SITE PLANT", "CEMENT RUBBISH CHUTES",
                        "CEMENT BARRIERS"}
            for r in rows[1:]:
                su = norm(r[H.get("STORAGE_UNIT", -1)])
                bc = norm(r[H.get("ITEM_BARCODE", -1)])
                if bc:
                    unit_of[bc] = su
                if su in PLANT_SU and bc and not bc.startswith("SUB"):
                    plant_keys.add(bc)
        except Exception:
            pass
    plant_keys.discard("")

    def is_plant(bc):
        if bc in plant_keys:
            return True
        return any(bc.startswith(k + "-") or k.startswith(bc + "-")
                   for k in plant_keys)

    TOOL_SU = {"TOOLING", "HYDRAULICS- HI TORQUE"}

    tx = _gfind("TRANSACTIONS*.xlsx")
    if not tx:
        return None
    try:
        twb = openpyxl.load_workbook(max(tx, key=os.path.getmtime),
                                     read_only=True, data_only=True)
        if "TRANSACTION_CHARGES" not in twb.sheetnames:
            twb.close()
            return None
        rows = list(twb["TRANSACTION_CHARGES"].iter_rows(values_only=True))
        twb.close()
    except Exception:
        return None
    if not rows:
        return None
    H = {norm(v): i for i, v in enumerate(rows[0]) if v}
    for need in ("HIRE_CHARGE ($)", "LATEST_BARCODE",
                 "SHIFT_CHARGE_DATE FROM"):
        if need not in H:
            return None

    def pdate(v):
        s = str(v or "").strip().split()[0] if v else ""
        try:
            return dt.datetime.strptime(s, "%d/%m/%Y").date()
        except ValueError:
            return None

    daily = {}
    for r in rows[1:]:
        try:
            hire = float(r[H["HIRE_CHARGE ($)"]] or 0)
        except (TypeError, ValueError):
            hire = 0.0
        if not hire:
            continue
        bc = norm(r[H["LATEST_BARCODE"]])
        if bc.startswith("SUB"):
            continue          # welders: separate invoice, separate stream
        f = pdate(r[H["SHIFT_CHARGE_DATE FROM"]])
        t = pdate(r[H.get("SHIFT_CHARGE_DATE_TO", -1)]) or f
        if not f:
            continue
        if is_plant(bc):
            bucket = "plant"
        elif unit_of.get(bc, "") in TOOL_SU:
            bucket = "tool"
        else:
            bucket = "gear"
        ndays = max(1, (t - f).days + 1)
        per = hire / ndays
        for i in range(ndays):
            d = f + dt.timedelta(days=i)
            slot = daily.setdefault(d, {"plant": 0.0, "tool": 0.0,
                                        "gear": 0.0})
            slot[bucket] += per
    return daily or None


def cost_story(wb, ds=None, sub=None, hs=None):
    ws = wb["Daily Tracking"]
    ds = ds or {}
    sub = sub or {"arrivals": {}, "rate": SUBHIRE_RATE}

    def num(r, c):
        v = ws.cell(r, c).value
        return v if isinstance(v, (int, float)) else 0

    rows = [r for r in range(2, 60) if isinstance(ws.cell(r, 1).value, (dt.datetime, dt.date))]
    active = [r for r in rows if num(r, 27) > 0]
    d0 = ws.cell(active[0], 1).value
    d1 = ws.cell(active[-1], 1).value
    today = dt.date.today()

    def rowdate(r):
        v = ws.cell(r, 1).value
        return v.date() if isinstance(v, dt.datetime) else v

    # ---- plant & tooling hire: ONE line, as SiteIQ invoices it ----------
    # (split back into Plant and Tooling for the stream table when the
    #  charge lines are here to do it truthfully - see read_hire_split)
    hire_fc = hire_act = hire_full = 0.0
    p_fc = p_full = t_fc = t_full = 0.0       # plant/tooling plan pieces
    prov_p = prov_t = 0.0                     # provisional split (workbook)
    invoiced_total = provisional_total = 0.0
    provisional_days = 0
    recon = []            # the true-up table, day by day
    last_inv = max(ds) if ds else None
    for r in rows:
        d = rowdate(r)
        fc_day = num(r, 2) + num(r, 5)                # plant fc + tooling fc
        tracked = num(r, 3) + num(r, 6)               # tracked actuals (live)
        # blank is not zero: a day whose tracking cells were never filled
        # in must say "not logged", not "$0 tracked" - $0 reads like
        # nothing happened and drags a phantom drift into the net figure
        # (29 Jul 2026 - the 27th)
        tracked_logged = (isinstance(ws.cell(r, 3).value, (int, float))
                          or isinstance(ws.cell(r, 6).value, (int, float)))
        hire_full += fc_day
        p_full += num(r, 2)
        t_full += num(r, 5)
        is_active = num(r, 27) > 0
        if is_active:
            hire_fc += fc_day
            p_fc += num(r, 2)
            t_fc += num(r, 5)
        # a day LOCKS only when its hire is genuinely in: either the hire
        # figure is present, or the day sits fully behind the export's
        # last day. A row that exists only because a SERVICE progress
        # claim posted (hire still $0 on an unfinished day) stays
        # provisional - it never locks hire at zero. (24 Jul 2026)
        ds_locked = d in ds and (ds[d]["hire"] > 0 or (ds and d < max(ds)))
        if ds_locked:                                 # invoiced - locked
            act_day = ds[d]["hire"]
            invoiced_total += act_day
            recon.append({"d": d,
                          "tracked": tracked if tracked_logged else None,
                          "invoiced": act_day, "status": "Invoiced"})
        elif is_active and d <= today and tracked > 0:
            act_day = tracked                         # run, not yet invoiced
            provisional_total += act_day
            provisional_days += 1
            prov_p += num(r, 3)
            prov_t += num(r, 6)
            recon.append({"d": d, "tracked": tracked, "invoiced": None,
                          "status": "Provisional"})
        else:
            act_day = 0.0
            if is_active and d <= today:
                recon.append({"d": d, "tracked": tracked, "invoiced": None,
                              "status": "No charges"})
        if is_active:
            hire_act += act_day
    # invoiced days that fall outside the workbook's active window still
    # count - SiteIQ charged them (e.g. early gear before the shut)
    for d, v in ds.items():
        if d not in [rowdate(r) for r in active] and v["hire"]:
            pass    # already included above via the full row walk when the
                    # date exists in Daily Tracking; a date SiteIQ charged
                    # that Daily Tracking doesn't carry at all would need a
                    # row added in the workbook - flagged on the true-up page
    ds_only = sorted(d for d in ds
                     if d not in {rowdate(r) for r in rows} and ds[d]["hire"])
    # "Trued to SiteIQ" has to mean every invoiced day is IN the table and
    # the totals - a day SiteIQ charged that Daily Tracking doesn't carry
    # yet still gets a real row, flagged so the workbook row gets added
    # (29 Jul 2026 - the missing 28th)
    for d in ds_only:
        invoiced_total += ds[d]["hire"]
        hire_act += ds[d]["hire"]
        recon.append({"d": d, "tracked": None, "invoiced": ds[d]["hire"],
                      "status": "Invoiced", "noworkbook": True})
    recon.sort(key=lambda x: x["d"])

    # Split the hire line into Plant and Tooling when the charge lines can
    # do it truthfully: the split must land on SiteIQ's invoiced total to
    # within a dollar, or we keep the combined line and say so - the report
    # never apportions. Provisional (not-yet-invoiced) days add their
    # workbook-tracked plant/tooling actuals to each side.
    if invoiced_total == 0:
        inv_p = inv_t = 0.0          # nothing invoiced yet - workbook split
        split_ok = True
    elif hs is not None and abs(hs["total"] - invoiced_total) < 1.0:
        inv_p, inv_t = hs["plant"], hs["tool"]
        split_ok = True
    else:
        split_ok = False             # charge lines absent or out of step with
                                     # DAILY_SUMMARY - keep the combined line
    if split_ok:
        p_act = inv_p + prov_p
        t_act = inv_t + prov_t
        streams = [
            {"name": "Plant hire", "fc": p_fc, "act": p_act,
             "var": p_act - p_fc, "full": p_full},
            {"name": "Tooling hire", "fc": t_fc, "act": t_act,
             "var": t_act - t_fc, "full": t_full},
        ]
    else:
        streams = [{"name": "Plant & tooling hire", "fc": hire_fc,
                    "act": hire_act, "var": hire_act - hire_fc,
                    "full": hire_full}]
    for name, fc, ac in STREAMS:
        f = sum(num(r, fc) for r in active)
        a = sum(num(r, ac) for r in active)
        # recognise incurred-but-unbilled cost at forecast for elapsed days
        if name in RECOGNISE_AT_FORECAST and a < f:
            a = f
        full = sum(num(r, fc) for r in rows)
        streams.append({"name": name, "fc": f, "act": a, "var": a - f, "full": full})

    # ---- welders: on site = charged, invoiced separately (like radios) --
    # Day value = units on site that day x the confirmed rate. Days run
    # count as actual (and plan, radio-style - they were brought on as
    # added scope, so plan moves with them); days ahead hold the current
    # on-site count at the rate.
    def sub_day(d):
        return sum(1 for a in sub["arrivals"].values() if a and a <= d) * sub["rate"]
    if sub["arrivals"]:
        w_act = sum(sub_day(rowdate(r)) for r in active
                    if rowdate(r) <= today)
        w_full = sum(sub_day(rowdate(r)) for r in rows)
        streams.append({"name": "Welders", "fc": w_act, "act": w_act,
                        "var": 0.0, "full": w_full})
    F = sum(s["fc"] for s in streams)
    A = sum(s["act"] for s in streams)
    FULL = sum(s["full"] for s in streams)

    # ---- month view: days run at trued/tracked value, days ahead at fc --
    months = {}
    for r in rows:
        d = rowdate(r)
        fc_total = num(r, 26)
        if d in ds and (ds[d]["hire"] > 0 or d < max(ds)):
            hire_day = ds[d]["hire"]           # locked at the invoice
        elif num(r, 27) > 0 and d <= today:
            hire_day = num(r, 3) + num(r, 6)   # run day (0 is a real zero)
        else:
            hire_day = None                    # not run yet - forecast
        others_act = (num(r, 9) + num(r, 12) + num(r, 15) + num(r, 18)
                      + max(num(r, 21), num(r, 20)) + num(r, 24))
        welders_day = (sum(1 for a in sub["arrivals"].values() if a and a <= d)
                       * sub["rate"]) if sub["arrivals"] else 0.0
        key = d.strftime("%B %Y")
        m = months.setdefault(key, {"run": 0.0, "ahead": 0.0, "order": d.replace(day=1)})
        if hire_day is not None:
            m["run"] += hire_day + others_act + welders_day
        else:
            m["ahead"] += fc_total + welders_day
    # invoiced days missing from Daily Tracking still belong to their month
    for d in ds_only:
        welders_day = (sum(1 for a in sub["arrivals"].values() if a and a <= d)
                       * sub["rate"]) if sub["arrivals"] else 0.0
        m = months.setdefault(d.strftime("%B %Y"),
                              {"run": 0.0, "ahead": 0.0, "order": d.replace(day=1)})
        m["run"] += ds[d]["hire"] + welders_day
    month_rows = sorted(months.items(), key=lambda kv: kv[1]["order"])

    return {"streams": streams, "F": F, "A": A, "V": A - F, "FULL": FULL,
            "still": FULL - A, "days": len(active),
            "d0": d0, "d1": d1,
            "recon": recon, "invoiced": invoiced_total,
            "provisional": provisional_total, "prov_days": provisional_days,
            "last_inv": last_inv, "ds_only": ds_only,
            "months": month_rows}


CHARGE_LINES = ["Damage Recovery", "Consumables Sales", "Fuel", "Transport"]


def read_charges():
    """Pull the Charge Reporter's Costing Feed (Damage/Consumables/Fuel/Transport)
    so logged charges show up in the report. Empty-safe."""
    feed = os.path.join(HERE, "Coates_K2_Charge_Reporter", "CHARGE_REGISTER.xlsx")
    lines = {k: {"count": 0, "total": 0.0, "firm": 0.0} for k in CHARGE_LINES}
    if not os.path.isfile(feed):
        return lines, 0.0, 0.0
    try:
        wb = openpyxl.load_workbook(feed, read_only=True, data_only=True)
        if "Costing Feed" not in wb.sheetnames:
            wb.close()
            return lines, 0.0, 0.0
        ws = wb["Costing Feed"]
        rows = ws.iter_rows(values_only=True)
        hdr = [str(v) if v is not None else "" for v in next(rows)]
        ix = {h: i for i, h in enumerate(hdr)}
        for r in rows:
            line = str(r[ix.get("Cost Line", 3)] or "").strip()
            amt = r[ix.get("Amount", 6)]
            status = str(r[ix.get("Status", 7)] or "").strip().lower()
            if line not in lines or not isinstance(amt, (int, float)):
                continue
            lines[line]["count"] += 1
            lines[line]["total"] += amt
            if status in ("approved", "invoiced"):
                lines[line]["firm"] += amt
        wb.close()
    except Exception:
        return lines, 0.0, 0.0
    total = sum(v["total"] for v in lines.values())
    firm = sum(v["firm"] for v in lines.values())
    return lines, total, firm


SHUTDOWN_START = dt.date(2026, 7, 20)
# people not counted toward the per-person Take 5 tally (management / liaison)
SAFETY_EXCLUDE = {"andrew fisher"}


def read_roster(wb, asof):
    """Tool-store man-hours (day/night) and auto safety tally from the Shutdown
    Costing roster, from 20 Jul. To-date = up to the as-at day; plan = full."""
    if "Shutdown Costing" not in wb.sheetnames:
        return None
    ws = wb["Shutdown Costing"]
    asof = asof.date() if isinstance(asof, dt.datetime) else asof
    td = {"Day": 0.0, "Night": 0.0}
    pl = {"Day": 0.0, "Night": 0.0}
    persondays = set(); shifts = set(); days = set()
    for r in range(16, ws.max_row + 1):
        d = ws.cell(r, 1).value
        if not isinstance(d, (dt.datetime, dt.date)):
            continue
        dd = d.date() if isinstance(d, dt.datetime) else d
        if dd < SHUTDOWN_START:
            continue
        shift = str(ws.cell(r, 3).value or "").strip().title()
        shift = shift if shift in ("Day", "Night") else "Day"
        name = str(ws.cell(r, 5).value or "").strip()
        bill = ws.cell(r, 10).value
        bill = bill if isinstance(bill, (int, float)) else 0
        if not name:
            continue
        pl[shift] += bill
        if dd <= asof:
            td[shift] += bill
            shifts.add((dd, shift)); days.add(dd)
            if name.lower() not in SAFETY_EXCLUDE:
                persondays.add((dd, name))   # Take 5 count excludes management/liaison
    return {
        "td_day": td["Day"], "td_night": td["Night"], "td_total": td["Day"] + td["Night"],
        "pl_day": pl["Day"], "pl_night": pl["Night"], "pl_total": pl["Day"] + pl["Night"],
        "persondays": len(persondays), "shifts": len(shifts), "days": len(days),
        # auto safety engagement: 3 Take 5s/person/day, 1 conversation/shift, 1 pre-start/day
        "take5": 3 * len(persondays), "convos": len(shifts), "prestarts": len(days),
    }


TEAM_DAY = [("Thomas Maher", "Dayshift Lead"),
            ("Helen O'Grady", "Dayshift Tool Store Controller"),
            ("Lucas Armstrong", "Day Shift Mechanic")]
TEAM_NIGHT = [("Cody Mitchell", "Nightshift Tool Store Controller"),
              ("Adam Boyce", "Night Shift Tool Store Controller")]


def _person(name, role, boss=False):
    parts = name.replace("O'", "O").split()
    ini = (parts[0][:1] + (parts[-1][:1] if len(parts) > 1 else "")).upper()
    return ("<div class='tperson{b}'><div class='tavatar'>{i}</div>"
            "<div><div class='tn'>{n}</div><div class='tr'>{r}</div></div></div>").format(
        b=" boss" if boss else "", i=ini, n=html.escape(name), r=html.escape(role))


def team_html():
    day = "".join(_person(n, r) for n, r in TEAM_DAY)
    night = "".join(_person(n, r) for n, r in TEAM_NIGHT)
    chips = ("<div class='twedo'>"
             "<span><b>&#10003;</b> The right gear to the right crew</span>"
             "<span><b>&#10003;</b> Looked after, maintained &amp; ready</span>"
             "<span><b>&#10003;</b> Checked and right &mdash; every time</span></div>")
    return (
        "<div class='card'>"
        "<h2>Meet the tool store team</h2>"
        "<div class='cap'>The crew running your K2 tool store &mdash; getting the right gear to the right "
        "crew, keeping it looked after and ready, and making sure everything that leaves the store is right. "
        "Day and night, so the shutdown keeps moving. Something not on hand or not right? Tell us and we'll sort it.</div>"
        + chips +
        "<div class='team-top'>" + _person("Andrew Fisher", "Shutdown Manager", boss=True) + "</div>"
        "<div class='tconn'></div>"
        "<div class='tgroups'>"
        "<div class='tgroup'><h4>Day shift</h4>" + day + "</div>"
        "<div class='tgroup'><h4>Night shift</h4>" + night + "</div>"
        "</div></div>")


def predicted_breakdown(cost):
    """Composition of the full-shutdown forecast (hire lines) — sums to FULL."""
    fby = {s["name"]: s["full"] for s in cost["streams"]}
    equip = sum(fby.get(k, 0) for k in ("Plant hire", "Tooling hire",
                                        "Plant & tooling hire", "Radio",
                                        "Gas Monitor", "Welders"))
    labour = sum(fby.get(k, 0) for k in ("Personnel / Labour", "Accommodation"))
    transport = fby.get("Transport", 0)
    damage = fby.get("Damage Recovery", 0)
    rows = [("Plant, tooling, radio &amp; gas hire", equip),
            ("Labour &amp; accommodation", labour),
            ("Transport", transport)]
    if damage > 0:
        rows.append(("Damage recovery (forecast)", damage))
    total = equip + labour + transport + damage
    return rows, total


def rag(s):
    """Return (colour, status text) for a cost stream, to-date."""
    f, a, v = s["fc"], s["act"], s["var"]
    if f <= 0 and a <= 0:
        return None
    if abs(v) < 1:
        return (GOOD, "On plan")
    if a <= 0 and f > 0:
        return (AMBER, "Not billed on days run")
    if a / f < 0.25:
        return (AMBER, "Well under on days run &mdash; locked")
    if v < 0 and abs(v) / f < 0.15:
        return (GOOD, "Tracking close (slightly under)")
    if v < 0:
        return (AMBER, "Under on days run &mdash; locked")
    return (GOOD, "On plan")


# ---- layout v2 styles ----------------------------------------------------
# Bigger type and lifted greys throughout - the old sizes read fine on a
# desktop but went muddy on phones and print. Same palette and design
# language as v1; more contrast, more air, one story per page.
PANEL_CSS = """
@page{size:A4;margin:0}
*{box-sizing:border-box;-webkit-print-color-adjust:exact;print-color-adjust:exact}
html,body{margin:0;background:#fff}
body{font-family:'Segoe UI',system-ui,Arial,sans-serif;color:#E6E9EE}
/* A4 geometry in px (1px = 1/96in, exact in Edge/Chrome headless AND print;
   mm heights render short in some headless modes, so we never use them here).
   Sheet 794x1123px = 210x297mm; panel fills the sheet inside 34px margins. */
.sheet{width:794px;height:1123px;overflow:hidden;margin:0 auto;padding:34px;background:#fff}
.sheet + .sheet{page-break-before:always}
.panel{background:#0E1116;border-radius:16px;border-top:9px solid #F26222;padding:26px 30px 20px;box-shadow:0 2px 12px rgba(0,0,0,.15);height:1055px;overflow:hidden;display:flex;flex-direction:column;position:relative}
.grow{flex:1 1 auto;min-height:8px}
@media print{.panel{box-shadow:none}}
.hd{display:flex;justify-content:space-between;align-items:flex-start;gap:24px}
.hd .brand{font-size:30px;font-weight:900;color:#F26222;letter-spacing:-.5px}
.hd .brand span{color:#fff;font-size:15px;font-weight:600;margin-left:8px}
.hd .site{color:#A9B1BD;font-size:12px;letter-spacing:.5px;margin-top:10px;text-transform:uppercase}
.hd h1{color:#fff;font-size:24px;margin:9px 0 3px}
.hd .rp{color:#A9B1BD;font-size:13px}
.hd .siteiq{text-align:right;color:#F26222;font-weight:800;font-size:12px;letter-spacing:1px}
.hd .meta{text-align:right;color:#A9B1BD;font-size:10.5px;line-height:1.7;margin-top:8px;text-transform:uppercase}
.hd .meta b{color:#fff;font-size:12.5px}
.hd2{display:flex;justify-content:space-between;align-items:flex-start;gap:24px}
.hd2 .brand{font-size:22px;font-weight:900;color:#F26222;letter-spacing:-.5px}
.hd2 .brand span{color:#fff;font-size:12px;font-weight:600;margin-left:7px}
.hd2 h1{color:#fff;font-size:19px;margin:6px 0 2px}
.hd2 .rp{color:#A9B1BD;font-size:12.5px}
.hd2 .siteiq{text-align:right;color:#F26222;font-weight:800;font-size:11px;letter-spacing:1px}
.hd2 .meta{text-align:right;color:#A9B1BD;font-size:10.5px;margin-top:6px;text-transform:uppercase;line-height:1.7}.hd2 .meta b{color:#fff;font-size:11.5px}
.rule{height:2px;background:linear-gradient(90deg,#F26222,transparent);margin:16px 0 20px}
.banner{display:flex;gap:16px;align-items:flex-start;background:#131a15;border:1px solid #24422a;border-left:5px solid #0ca30c;border-radius:12px;padding:18px 20px;margin-bottom:20px}
.banner.watch{background:#1c1710;border-color:#4a3a14;border-left-color:#fab219}
.pill{flex:none;background:#0ca30c;color:#08210c;font-weight:900;font-size:12px;letter-spacing:.5px;padding:6px 13px;border-radius:16px;text-transform:uppercase;margin-top:2px}
.pill.watch{background:#fab219;color:#2a1e05}
.banner p{margin:0;font-size:14.5px;line-height:1.65;color:#D7DBE2}.banner b{color:#fff}
.kpis{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:20px}
.kpi{background:#171B22;border:1px solid #2A313C;border-left:4px solid #F26222;border-radius:11px;padding:16px 14px;min-width:0}
.kpi.g{border-left-color:#0ca30c}.kpi.b{border-left-color:#F26222}
.kpi .v{font-size:20px;font-weight:900;color:#fff;line-height:1.1;letter-spacing:-.3px}
.kpi .l{font-size:11px;color:#A9B1BD;text-transform:uppercase;letter-spacing:.5px;margin-top:8px;font-weight:700}
.kpi .s{font-size:11.5px;color:#949CA8;margin-top:4px;line-height:1.4}
.storyc{background:#171B22;border:1px solid #2A313C;border-left:4px solid #F26222;border-radius:12px;padding:20px 24px;margin-top:4px}
.storyc h2{color:#fff;font-size:16px;margin:0 0 10px}
.storyc p{color:#D7DBE2;font-size:14px;line-height:1.75;margin:0 0 12px}.storyc p:last-child{margin-bottom:0}.storyc b{color:#fff}.storyc .hl{color:#F26222;font-weight:700}
.card{background:#171B22;border:1px solid #2A313C;border-radius:12px;padding:18px 22px}
.card h2{color:#fff;font-size:17px;margin:0 0 5px}
.card .cap{color:#9BA3AF;font-size:12.5px;margin-bottom:12px;line-height:1.5}
.cols{display:grid;grid-template-columns:1fr 1fr;gap:16px}
table{width:100%;border-collapse:collapse;font-size:13px}
th{color:#9BA3AF;text-align:left;font-size:10.5px;text-transform:uppercase;letter-spacing:.6px;padding:0 8px 9px;border-bottom:1px solid #2A313C}
td{padding:10px 8px;border-bottom:1px solid #20262e;vertical-align:top;color:#D7DBE2}
td.n{text-align:right;font-variant-numeric:tabular-nums;color:#fff}
tr:last-child td{border-bottom:0}
table.tight{font-size:12px}
table.tight td{padding:9px 6px}
table.tight th{padding:0 6px 8px}
table.tight td:first-child,table.tight td:last-child{white-space:nowrap}
.dot{display:inline-block;width:9px;height:9px;border-radius:50%;margin-right:8px;vertical-align:0}
.footnote{color:#A9B1BD;font-size:11.5px;line-height:1.55;margin-top:10px;border-top:1px dashed #2A313C;padding-top:9px}
.footnote b{color:#F26222}
.util{font-size:52px;font-weight:900;color:#fff;line-height:1;margin:10px 0 4px;letter-spacing:-1px}.util span{font-size:18px;color:#A9B1BD;font-weight:600;letter-spacing:0}
.bar{height:15px;background:#20262e;border-radius:8px;overflow:hidden;margin:16px 0 8px}
.bar > i{display:block;height:100%;background:linear-gradient(90deg,#F26222,#f7883f);border-radius:8px}
.barends{display:flex;justify-content:space-between;color:#A9B1BD;font-size:12.5px}
.callout{background:#0E1116;border:1px solid #2A313C;border-left:4px solid #fab219;border-radius:10px;padding:14px 18px;margin-top:18px;font-size:15px;color:#D7DBE2}.callout b{color:#fff;font-size:17px}
.uln{font-size:13.5px;color:#D7DBE2;line-height:1.7;margin-top:14px}.uln b{color:#fff}
.bigmoney{font-size:44px;font-weight:900;color:#fff;line-height:1.05;letter-spacing:-1px}
.bigsub{color:#A9B1BD;font-size:11px;text-transform:uppercase;letter-spacing:.6px;margin-top:8px;font-weight:700}
.ceiling{margin-top:12px;color:#A9B1BD;font-size:12.5px}.ceiling b{color:#D7DBE2}
.strip{display:flex;align-items:flex-start;gap:12px;margin-top:16px;padding:14px 18px}
.strip p{margin:0;font-size:13px;line-height:1.6;color:#D7DBE2}.strip b{color:#fff}
.foot{position:absolute;left:30px;right:30px;bottom:14px;display:flex;justify-content:space-between;gap:20px;color:#9BA3AF;font-size:10px;line-height:1.55;border-top:1px solid #2A313C;padding-top:8px}
.foot b{color:#A9B1BD}.foot span:first-child{flex:1}
.team-top{display:flex;justify-content:center;margin:8px 0 0}
.tperson{display:flex;align-items:center;gap:10px;background:#171B22;border:1px solid #2A313C;border-radius:9px;padding:8px 13px}
.tperson.boss{border-left:4px solid #F26222;padding-right:22px}
.tavatar{width:34px;height:34px;border-radius:50%;background:#F26222;color:#fff;font-weight:800;font-size:13px;display:flex;align-items:center;justify-content:center;flex:none;letter-spacing:.5px}
.tperson .tn{color:#fff;font-weight:700;font-size:13.5px}
.tperson .tr{color:#9BA3AF;font-size:10.5px;text-transform:uppercase;letter-spacing:.4px;margin-top:2px}
.tconn{height:10px;border-left:2px solid #2A313C;width:0;margin:2px auto}
.tgroups{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:10px}
.tgroup{background:#141922;border:1px solid #2A313C;border-radius:10px;padding:11px 13px}
.tgroup h4{color:#F26222;font-size:11px;text-transform:uppercase;letter-spacing:1px;margin:0 0 8px}
.tgroup .tperson{margin-bottom:7px;background:#0E1116}
.tgroup .tperson:last-child{margin-bottom:0}
.twedo{display:flex;gap:9px;flex-wrap:wrap;margin:10px 0 4px}
.twedo span{background:#171B22;border:1px solid #2A313C;border-left:3px solid #F26222;border-radius:8px;padding:7px 12px;font-size:12px;color:#D7DBE2}
.twedo b{color:#F26222}
"""

# ---------------------------------------------------------------------
#  EXECUTIVE LIGHT THEME - 27 Jul 2026. The snapshot was the original
#  dark design the page engine copied; when the whole suite went
#  executive white it kept the black look. This override brings it into
#  the family - same switch as everything else (report_pages.EXEC_LIGHT),
#  flip that off and the dark snapshot comes straight back.
# ---------------------------------------------------------------------
try:
    import report_pages as _RP
    _LIGHT = bool(getattr(_RP, "EXEC_LIGHT", False))
except Exception:
    _LIGHT = False

LIGHT_CSS = """
body{color:#20242B}
.panel{background:#fff;border:6px solid #F26222;border-top:9px solid #F26222;box-shadow:none}
.hd .brand span{color:#14181F}.hd .site{color:#6B7380}
.hd h1{color:#14181F}.hd .rp{color:#5B6472}
.hd .meta{color:#6B7380}.hd .meta b{color:#14181F}
.hd2 .brand span{color:#14181F}.hd2 h1{color:#14181F}.hd2 .rp{color:#5B6472}
.hd2 .meta{color:#6B7380}.hd2 .meta b{color:#14181F}
.rule{background:linear-gradient(90deg,#F26222,#EFFF3D,transparent)}
.banner{background:#E9F6EC;border-color:#BFE3C8}
.banner.watch{background:#FFF4D6;border-color:#F0DCA0}
.banner p{color:#33393F}.banner b{color:#14181F}
.kpi{background:#F6F7F9;border-color:#E3E6EB}
.kpi .v{color:#14181F}.kpi .l{color:#5B6472}.kpi .s{color:#6B7380}
.storyc{background:#FFF3EC;border-color:#F5DECD}
.storyc h2{color:#14181F}.storyc p{color:#33393F}.storyc b{color:#14181F}
.card{background:#F6F7F9;border-color:#E3E6EB}
.card h2{color:#14181F}.card .cap{color:#6B7380}
th{color:#5B6472;border-bottom-color:#E3E6EB}
td{color:#262C35;border-bottom-color:#EDF0F3}
td.n{color:#14181F}
.footnote{color:#5B6472;border-top-color:#E3E6EB}
.util{color:#14181F}.util span{color:#5B6472}
.bar{background:#EDF0F4}
.bar > i{background:linear-gradient(90deg,#F26222,#EFFF3D)}
.barends{color:#5B6472}
.callout{background:#FFFBEA;border-color:#F0DCA0;color:#33393F}
.callout b{color:#14181F}
.uln{color:#33393F}.uln b{color:#14181F}
.bigmoney{color:#14181F}.bigsub{color:#5B6472}
.ceiling{color:#5B6472}.ceiling b{color:#33393F}
.strip p{color:#33393F}.strip b{color:#14181F}
.foot{color:#6B7380;border-top-color:#E3E6EB}.foot b{color:#5B6472}
.tperson{background:#F6F7F9;border-color:#E3E6EB}
.tperson .tn{color:#14181F}.tperson .tr{color:#5B6472}
.tconn{border-left-color:#E3E6EB}
.tgroup{background:#F9FAFB;border-color:#E3E6EB}
.tgroup .tperson{background:#fff}
.twedo span{background:#F6F7F9;border-color:#E3E6EB;color:#33393F}
""" if _LIGHT else ""

#  Inline styles always beat the LIGHT_CSS overrides above, so anything
#  coloured inline must pick its ink by theme - #fff ink on the white
#  executive page was Andrew's "some words you just cant see"
#  (29 Jul 2026). White-on-ORANGE (the plant/tooling bar labels) is
#  readable on both themes and stays #fff on purpose.
INK = "#14181F" if _LIGHT else "#fff"        # strong ink (totals, headings)
MUTED = "#5B6472" if _LIGHT else "#A9B1BD"   # secondary ink (notes, labels)
FAINT = "#6B7380" if _LIGHT else "#9BA3AF"   # small caps section labels
TRACK = "#EDF0F4" if _LIGHT else "#20262e"   # bar track behind the fill



def build_html(cost, ov, charges_data, roster, wbname, svc=None):
    """Assemble the report as one story per A4 sheet. Sections with nothing to
    say collapse to a line (charges at zero) or drop off (no roster) - the
    page count adapts. Returns (full_html, [per-page html], n_pages)."""
    now = dt.datetime.now()
    asat = now.strftime("%d %b %Y - %H:%M")
    period = "{} – {}".format(cost["d0"].strftime("%d %b"), cost["d1"].strftime("%d %b %Y"))
    F, A, V, FULL = cost["F"], cost["A"], cost["V"], cost["FULL"]
    under = -V                 # dollars under the to-date plan (positive)
    tracking = FULL + V        # forecast to completion less the under-run tracked so far

    # under-run drivers (streams with actual < forecast to date)
    unders = sorted([s for s in cost["streams"] if s["var"] < -1], key=lambda s: s["var"])
    top2 = unders[:2]
    driver_names = " and ".join(s["name"].split(" / ")[0].lower() for s in top2) or "timing"

    # ---- page 1 · the position ------------------------------------------
    # The banner carries the call (landing vs plan). The tiles carry the
    # figures. The story carries the why - in words, not repeated dollars.
    banner = (
        "Cost is tracking <b>under the plan and under control</b> &mdash; and trued to SiteIQ's "
        "invoicing every morning. Days SiteIQ has invoiced are locked at the invoiced figure, "
        "{prov} and days ahead are held at forecast &mdash; tracking to land at <b>{TR}</b>, under "
        "the <b>{FULL}</b> plan, and not over it."
    ).format(TR=money0(tracking), FULL=money0(FULL),
             prov=("today runs provisional until its invoice lands,"
                   if cost["prov_days"] else "every day run is invoiced,"))

    kpis = "".join([
        "<div class='kpi'><div class='v'>{}</div><div class='l'>Forecast to date</div><div class='s'>The plan · {}</div></div>".format(money0(F), period),
        "<div class='kpi g'><div class='v'>{}</div><div class='l'>Actual to date</div><div class='s'>per Daily Tracking</div></div>".format(money0(A)),
        "<div class='kpi'><div class='v'>{}</div><div class='l'>Variance to date</div><div class='s'>Under plan so far</div></div>".format(money0(V)),
        "<div class='kpi b'><div class='v'>{}</div><div class='l'>Forecast to completion</div><div class='s'>The plan</div></div>".format(money0(FULL)),
        "<div class='kpi g'><div class='v'>{}</div><div class='l'>Tracking to land</div><div class='s'>Actuals locked + plan ahead</div></div>".format(money0(tracking)),
    ])

    story = (
        "<div class='storyc'><h2>The story so far</h2>"
        "<p>{days} active days in, the K2 tool store is <b>under the plan and under control</b> &mdash; and the "
        "numbers are <b>trued to SiteIQ's invoicing</b>: <span class='hl'>{inv}</span> of plant &amp; tooling hire "
        "is invoiced and locked{prov}. The true-up page shows tracked v invoiced, day by day, drift and all.</p>"
        "<p>The variance is an under-run, not an overrun. The gap is <span class='hl'>timing, not scope</span> "
        "&mdash; mobilisation transport is recognised at forecast, and the under-run sits in {drivers} against "
        "forecast on the days already run. Nothing has been dropped or deferred &mdash; the crews have had the "
        "gear they need, when they needed it.</p>"
        "<p>With the days still to come held at forecast, the landing stays under the plan &mdash; and we won't bill "
        "over it. This snapshot refreshes each week alongside the plant position.</p></div>"
    ).format(days=cost["days"], drivers=driver_names, inv=money0(cost["invoiced"]),
             prov=("; today's <b>{}</b> is provisional until its invoice lands".format(
                       money0(cost["provisional"])) if cost["provisional"] else ""))

    page1 = (
        "<div class='hd'><div>"
        "<div class='brand'>COATES<span>Equipped for anything</span></div>"
        "<div class='site'>Cement Australia · K2 Shutdown 2026 · Gladstone</div>"
        "<h1>Cost Tracking Snapshot</h1>"
        "<div class='rp'>Reporting period {period} · {days} active days</div>"
        "</div><div><div class='siteiq'>POWERED BY SITEIQ</div>"
        "<div class='meta'>As at<br><b>{asat}</b><br><br>Author<br><b>Andrew Fisher</b><br><br>PAGEMARK</div></div></div>"
        "<div class='rule'></div>"
        "<div class='banner'><span class='pill'>On plan</span><p>{banner}</p></div>"
        #  THE STREAM STAMP (Andrew, 28 Jul 2026): every dollar on this
        #  snapshot is the SiteIQ stream. The separate monthly invoice is
        #  its own page and its own money - said plainly on page one so
        #  the two can never be read as one figure.
        "<div class='banner watch'><span class='pill watch'>Cost stream</span>"
        "<p><b>SiteIQ invoicing only.</b> Every figure on this snapshot "
        "bills through SiteIQ. The separate monthly invoice &mdash; "
        "radios, gas monitors, sub-hired welders and forklift, "
        "compressor, crib gear and its own transport &mdash; is tracked "
        "on its own page and <b>none of its dollars appear here</b>. "
        "The two streams never share a figure.</p></div>"
        "<div class='kpis'>{kpis}</div>"
    ).format(period=period, days=cost["days"], asat=asat, banner=banner, kpis=kpis) + story

    # ---- page 2 · cost streams ------------------------------------------
    body = []
    for s in cost["streams"]:
        r = rag(s)
        if not r:
            continue
        colour, status = r
        # variance shown as rounded(actual) - rounded(forecast) so the three
        # columns always reconcile to the eye - never out by $1 on a client page
        var_disp = round(s["act"]) - round(s["fc"])
        body.append(
            ("<tr><td>{name}</td><td class='n' style='color:" + MUTED + "'>{full}</td>"
             "<td class='n'>{fc}</td><td class='n'>{act}</td>"
             "<td class='n' style='color:{vc}'>{var}</td>"
             "<td><span class='dot' style='background:{c}'></span>{st}</td></tr>").format(
                name=html.escape(s["name"]), full=money0(s["full"]),
                fc=money0(s["fc"]), act=money0(s["act"]),
                var=money0(var_disp), vc=(MUTED if abs(var_disp) < 1 else INK),
                c=colour, st=status))
    drivers_cap = " and ".join(s["name"].split(" / ")[0] for s in top2) or "Timing"
    hire_is_split = any(s["name"] == "Plant hire" for s in cost["streams"])
    if hire_is_split:
        hire_note = (
            "SiteIQ invoices plant and tooling as one hire line, trued daily to the "
            "DAILY_SUMMARY export (invoiced days locked, today provisional); it is split here item by item "
            "from SiteIQ's own charge lines &mdash; plant register assets to Plant, tool store gear to "
            "Tooling &mdash; and the two always sum exactly to the invoiced figure. The combined line is "
            "reconciled day by day on the true-up page.")
    else:
        hire_note = (
            "Plant &amp; tooling hire is shown the way SiteIQ invoices it &mdash; one hire "
            "line, trued daily to the DAILY_SUMMARY export (invoiced days locked, today provisional).")
    streams_page = (
        "<div class='card'><h2>Why we're under &mdash; by cost stream</h2>"
        "<div class='cap'>The plan column is what we thought each stream would cost for the whole shut "
        "&mdash; the original numbers. Forecast vs actual then cover the days already run. Amber means "
        "under on days run &mdash; locked in our favour, not at risk.</div>"
        "<table class='tight'><thead><tr><th>Cost stream</th><th style='text-align:right'>Plan &middot; full shut</th>"
        "<th style='text-align:right'>Forecast</th>"
        "<th style='text-align:right'>Actual</th><th style='text-align:right'>Variance</th><th>Status</th></tr></thead>"
        "<tbody>{body}</tbody></table>"
        "<div class='footnote'>{hire_note} Transport is recognised at forecast &mdash; "
        "mobilisation incurred, billed in arrears. Welders, radio and gas monitor hire are tracked here "
        "for the full cost picture and invoiced separately from SiteIQ's hire line &mdash; the Welders "
        "stream carries its own plan and is never counted in Plant hire. Personnel / labour and "
        "accommodation are shown at the daily accrual and invoiced in progress claims &mdash; reconciled "
        "on the true-up page. Days still to come are costed at forecast.</div></div>"
    ).format(body="".join(body), drivers=drivers_cap, hire_note=hire_note)

    # ---- page · the true-up: tracked v invoiced, day by day --------------
    # Quiet stretches (3+ consecutive no-charge days) collapse into one row,
    # and the day table deals itself onto extra pages at 18 rows a sheet -
    # so as the shut runs on, the story grows MORE pages, never a squashed
    # or clipped one. (A. Fisher's rule: it fits, or it goes to the next page.)
    st_dot = {"Invoiced": GOOD, "Provisional": AMBER, "No charges": "#5b6472"}

    #  the per-day plant/tooling/gear partition (Andrew, 30 Jul 2026:
    #  "how much for each for each day") - shown ONLY when it ties to
    #  the day's invoice, dashed with a reason when the TRANSACTIONS
    #  pull predates the day's charge lines (they post ~09:30 next
    #  morning; a 06:30 pull carries $0 against an invoiced day)
    hire_daily = cost.get("hire_daily") or {}
    stale_split_days = []

    def rec_row(d_lab, tracked, invoiced, status, flag="", day=None):
        # tracked None = the workbook never logged that day - a dash, not
        # a phantom $0 (there is no drift to show when there is no tracking)
        no_track = tracked is None
        drift_disp = (money0(round(tracked) - round(invoiced))
                      if (invoiced is not None and not no_track) else "&mdash;")
        quiet = no_track or invoiced is None or abs(tracked - invoiced) < 1
        # drift you can read on EITHER theme - #fff on the light page was
        # Andrew's "some words you just cant see" (29 Jul 2026)
        loud = INK
        sp = hire_daily.get(day) if day else None
        sp_tot = (sp["plant"] + sp["tool"] + sp["gear"]) if sp else 0.0
        ok = (sp is not None and
              (invoiced is None or abs(sp_tot - invoiced) < 1))
        if sp and invoiced is not None and not ok:
            stale_split_days.append(d_lab)
        if sp is None and invoiced is not None and status == "Invoiced":
            stale_split_days.append(d_lab)
        p_c = money0(sp["plant"]) if ok else "&mdash;"
        t_c = money0(sp["tool"]) if ok else "&mdash;"
        g_c = money0(sp["gear"]) if ok else "&mdash;"
        return ("<tr><td style='white-space:nowrap'>{d}</td><td class='n'>{t}</td><td class='n'>{i}</td>"
                "<td class='n' style='color:{mut}'>{p}</td>"
                "<td class='n' style='color:{mut}'>{tl}</td>"
                "<td class='n' style='color:{mut}'>{g}</td>"
                "<td class='n' style='color:{dc}'>{dr}</td>"
                "<td style='white-space:nowrap'><span class='dot' style='background:{c}'></span>{st}</td></tr>").format(
            d=d_lab, t="&mdash;" if no_track else money0(tracked),
            i=money0(invoiced) if invoiced is not None else "&mdash;",
            p=p_c, tl=t_c, g=g_c, mut=MUTED,
            dr=drift_disp, dc="#A9B1BD" if quiet else loud, c=st_dot[status],
            #  eight columns share this page now - the long status wording
            #  ran off the card edge (caught 30 Jul 2026 in the raster).
            #  The coloured dot carries "invoiced"; the word just says the
            #  consequence.
            st=("Locked" if status == "Invoiced" else status) + flag)

    rec_rows = []
    net_drift = 0.0
    i0 = 0
    recon = cost["recon"]
    while i0 < len(recon):
        x = recon[i0]
        # a run of quiet days (nothing tracked, nothing invoiced) -> one row
        j = i0
        while (j < len(recon) and recon[j]["status"] == "No charges"
               and recon[j]["tracked"] == 0):
            j += 1
        if j - i0 >= 3:
            rec_rows.append(rec_row(
                "{} &ndash; {} &middot; {} days".format(
                    recon[i0]["d"].strftime("%d %b"),
                    recon[j - 1]["d"].strftime("%d %b"), j - i0),
                0.0, None, "No charges"))
            i0 = j
            continue
        flag = ""
        if x.get("noworkbook"):
            flag = " (no row)"
        elif x["tracked"] is None:
            flag = " (no log)"
        rec_rows.append(rec_row(x["d"].strftime("%d %b"), x["tracked"],
                                x["invoiced"], x["status"], flag,
                                day=x["d"]))
        i0 += 1
    # net drift must count every invoiced day, collapsed rows included -
    # but only days that were actually tracked can drift (a never-logged
    # day has nothing to measure against)
    net_drift = sum((x["tracked"] - x["invoiced"]) for x in recon
                    if x["invoiced"] is not None and x["tracked"] is not None)
    ds_note = ""
    if cost["ds_only"]:
        ds_note = (" SiteIQ shows charges on {} &mdash; date(s) not yet in the "
                   "workbook's Daily Tracking. They are in this table and in "
                   "every total (trued to SiteIQ); add the workbook row(s) and "
                   "the tracked column fills in.".format(
                       ", ".join(d.strftime("%d %b") for d in cost["ds_only"])))
    _tu_head = ("<table style='font-size:11.5px'><thead><tr><th>Day</th><th style='text-align:right'>Tracked</th>"
                "<th style='text-align:right'>SiteIQ invoiced</th>"
                "<th style='text-align:right'>Plant</th>"
                "<th style='text-align:right'>Tooling</th>"
                "<th style='text-align:right'>Gear</th>"
                "<th style='text-align:right'>Drift</th>"
                "<th>Status</th></tr></thead><tbody>")
    split_note = (" Plant / Tooling / Gear is the invoiced figure split by "
                  "SiteIQ's own charge lines &mdash; plant by the site's plant "
                  "registers and plant storage units, tooling by the tool "
                  "aisles, gear is the rest of the store; the three always "
                  "sum to the day's invoice. The tooling column runs small "
                  "because the hand-tool aisles go out largely free-issue "
                  "&mdash; the charged store dollars sit in rigging, "
                  "electrical and welding gear. That is the real shape of "
                  "the hire, not missing data.")
    if stale_split_days:
        split_note += (" A dash in the split means the TRANSACTIONS pull "
                       "predates that day's charge lines (SiteIQ posts them "
                       "about 09:30 the next morning) &mdash; re-run 28 after "
                       "09:30 and rebuild, and {} fill(s) in.".format(
                           ", ".join(stale_split_days)))
    _no_track_days = [x["d"].strftime("%d %b") for x in recon
                      if x["invoiced"] is not None and x["tracked"] is None]
    if _no_track_days:
        split_note += (" Tracking was not logged on {} &mdash; the invoice "
                       "carries those days in every total; run "
                       "01_RUN_PREFILL_DAILY each morning so the drift check "
                       "stays alive.".format(", ".join(_no_track_days)))
    _tu_foot = ("<div class='footnote'>Drift is the live morning position v the final invoiced figure &mdash; gear "
                "on- and off-hired during the day, waivers and corrections. Net drift across invoiced days: "
                "<b>{nd}</b>. The report always carries the invoiced figure, so the drift never reaches the "
                "totals.{extra}{sp}</div>").format(nd=money0(round(net_drift)),
                                                   extra=ds_note,
                                                   sp=split_note)
    #  Page budget. The panel is 1055px with overflow hidden - anything
    #  past the edge is simply gone. The footnote above grew its split
    #  and timing notes until rows + footnote pushed past that edge
    #  ("cant see the bottom", A.F. 30 Jul 2026), so the LAST page now
    #  gives the footnote its room first: estimate its rendered lines
    #  from the text length (~105 chars a line at 11.5px), budget what's
    #  left for table rows (~44px each after ~330px of header, rule and
    #  card chrome), and balance the rows across the pages that need.
    TRUEUP_ROWS_PER_PAGE = 18
    _foot_chars = len(re.sub(r"<[^>]+>", "", _tu_foot))
    _foot_px = 20 + 18 * max(2, (_foot_chars + 104) // 105)
    _last_max = max(4, min(TRUEUP_ROWS_PER_PAGE, (1055 - 330 - _foot_px) // 44))
    _n = len(rec_rows)
    _pages = (1 if _n <= _last_max else
              1 + (_n - _last_max + TRUEUP_ROWS_PER_PAGE - 1) // TRUEUP_ROWS_PER_PAGE)
    _sizes = [_n // _pages + (1 if k < _n % _pages else 0) for k in range(_pages)]
    _k = 0
    while _sizes[-1] > _last_max and _k < len(_sizes) - 1:
        if _sizes[_k] < TRUEUP_ROWS_PER_PAGE:
            _sizes[_k] += 1
            _sizes[-1] -= 1
        else:
            _k += 1
    row_chunks, _at = [], 0
    for _s in _sizes:
        row_chunks.append(rec_rows[_at:_at + _s])
        _at += _s
    row_chunks = row_chunks or [[]]
    trueup_pages = []
    for ci, chunk in enumerate(row_chunks):
        first, last_c = (ci == 0), (ci == len(row_chunks) - 1)
        h2 = ("The true-up &mdash; tracked v SiteIQ invoiced, day by day"
              + ("" if first else " (continued)"))
        cap = ("Our live tracking against what SiteIQ actually invoiced (DAILY_SUMMARY, the latest "
               "export). Invoiced days are locked at the invoice; the tracked column shows how close the live view "
               "ran. Welders, radio and gas monitor hire are invoiced separately, so they sit outside this "
               "reconciliation by design. This is how the report stays true, every single day."
               if first else "Continuing the day-by-day reconciliation.")
        trueup_pages.append(
            "<div class='card'><h2>" + h2 + "</h2><div class='cap'>" + cap + "</div>"
            + _tu_head + "".join(chunk) + "</tbody></table>"
            + (_tu_foot if last_c else "") + "</div>")

    # -- labour & accommodation progress claims (SERVICE in SiteIQ) -------
    # A story of its own, so it gets its OWN page (never squeezed under the
    # true-up table): the position summary, then the claim lines exactly as
    # SiteIQ booked them - an invoice register, shown to the cent.
    svc = svc or {"labour": 0.0, "accom": 0.0, "other": [], "total": 0.0, "rows": []}
    claims_page = ""
    claims_lines_page = ""
    if svc["total"] > 0:
        acc_lab = next((s2["act"] for s2 in cost["streams"]
                        if s2["name"] == "Personnel / Labour"), 0.0)
        acc_acc = next((s2["act"] for s2 in cost["streams"]
                        if s2["name"] == "Accommodation"), 0.0)

        def claim_row(label, accrued, claimed):
            gap = round(claimed) - round(accrued)
            note = ("claimed ahead of accrual" if gap > 0 else
                    "accrual ahead of claims" if gap < 0 else "square")
            return ("<tr><td>{l}</td><td class='n'>{a}</td><td class='n'>{c}</td>"
                    "<td class='n'>{g}</td><td>{n}</td></tr>").format(
                l=label, a=money0(accrued), c=money0(claimed),
                g=money0(gap), n=note)
        crows = (claim_row("Personnel / Labour", acc_lab, svc["labour"])
                 + claim_row("Accommodation", acc_acc, svc["accom"]))
        for desc, amt in svc["other"]:
            crows += ("<tr><td>{d}</td><td class='n'>&mdash;</td>"
                      "<td class='n'>{a}</td><td class='n'>&mdash;</td>"
                      "<td>other service line</td></tr>").format(
                          d=html.escape(desc[:70]), a=money0(amt))

        def claim_when(w):
            s = str(w or "").strip().split()[0] if str(w or "").strip() else ""
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    return dt.datetime.strptime(s, fmt).strftime("%d %b %Y")
                except ValueError:
                    continue
            return html.escape(s) or "&mdash;"
        lines = list(svc.get("rows") or [])
        MAX_LINES = 12
        dropped, dropped_amt = 0, 0.0
        if len(lines) > MAX_LINES:                 # keep the LATEST claims
            dropped = len(lines) - (MAX_LINES - 1)
            dropped_amt = sum(a for _w, _de, a in lines[:dropped])
            lines = lines[dropped:]
        lrows = ""
        if dropped:
            lrows += ("<tr><td>&mdash;</td><td>{n} earlier claim line(s)</td>"
                      "<td class='n'>${a:,.2f}</td></tr>").format(n=dropped, a=dropped_amt)
        for w, de, a in lines:
            lrows += ("<tr><td style='white-space:nowrap'>{w}</td><td>{d}</td>"
                      "<td class='n'>${a:,.2f}</td></tr>").format(
                          w=claim_when(w), d=html.escape(de[:70]), a=a)
        lrows += ("<tr><td></td><td style='color:" + INK + ";font-weight:700'>Total claimed</td>"
                  "<td class='n' style='color:" + INK + ";font-weight:700'>${a:,.2f}</td></tr>"
                  ).format(a=svc["total"])
        claims_page = (
            "<div class='card'>"
            "<h2>Labour &amp; accommodation &mdash; invoiced in progress claims</h2>"
            "<div class='cap'>These streams are invoiced in lump progress claims through SiteIQ "
            "(they land under SERVICE), covering a stretch of days at a time. The cost story runs "
            "on the daily accrual so every day reads evenly &mdash; the claims and the accrual meet "
            "in the middle and even out across the shut. Neither is double counted.</div>"
            "<table><thead><tr><th>Stream</th><th style='text-align:right'>Accrued to date</th>"
            "<th style='text-align:right'>Claimed to date</th><th style='text-align:right'>Gap</th>"
            "<th>Position</th></tr></thead><tbody>" + crows + "</tbody></table>"
            "<div class='footnote'>A claim ahead of accrual is simply billing for days the accrual "
            "hasn't reached yet &mdash; it does not change the landing, and the total claimed will "
            "finish level with the plan, not over it."
            + ((" {n} line(s) still read MISC &mdash; SiteIQ sent no description and "
                "MISC_CLAIM_NAMES.txt doesn't name that amount yet. Open the file in "
                "Notepad, add a line like &ldquo;1090.50 = what it is&rdquo;, save, "
                "run 02.").format(n=svc.get("unnamed", 0))
               if svc.get("unnamed") else "")
            + "</div></div>")
        #  The invoice register rides on its OWN sheet - two cards on one
        #  1055px panel clipped this one at the edge once the claim lines
        #  grew ("cant see the bottom", A.F. 30 Jul 2026).
        claims_lines_page = (
            "<div class='card'>"
            "<h2>The claims as SiteIQ booked them</h2>"
            "<div class='cap'>Straight from the TRANSACTIONS export (PRODUCT_CATEGORY &ldquo;Service&rdquo;), "
            "to the cent &mdash; the same lines that appear on the invoice.</div>"
            "<table><thead><tr><th>Booked</th><th>Claim line</th>"
            "<th style='text-align:right'>Amount</th></tr></thead><tbody>" + lrows +
            "</tbody></table></div>")

    # ---- page 3 · plant utilisation -------------------------------------
    util = ov["util"]
    util_page = (
        "<div class='card'>"
        "<div><span class='pill watch'>Idle plant · client insight</span></div>"
        "<div style='color:" + INK + ";font-size:19px;font-weight:800;margin-top:14px'>Plant utilisation</div>"
        "<div class='cap'>Site plant on charge, as at {d}</div>"
        "<div class='util'>{util}%<span> deployed</span></div>"
        "<div class='bar'><i style='width:{util}%'></i></div>"
        "<div class='kpis' style='grid-template-columns:repeat(3,1fr);margin:18px 0 0'>"
        "<div class='kpi g'><div class='v'>{dep}</div><div class='l'>Out with crews</div><div class='s'>Working &amp; earning</div></div>"
        "<div class='kpi' style='border-left-color:#fab219'><div class='v'>{park}</div><div class='l'>On site, not yet out</div><div class='s'>Charged, used or not</div></div>"
        "<div class='kpi'><div class='v'>{tot}</div><div class='l'>Total plant on charge</div><div class='s'>On the ground at K2</div></div></div>"
        "<div class='callout'><b>{idle}/day</b> of plant is on site but not out with a crew ({share}%).</div>"
        "<div class='uln'>While gear is on site it's charged, used or not &mdash; so this is a live view of "
        "what's on the ground, not a saving. It's shared so you can see exactly what you're carrying, and use "
        "it to plan gear delivery and future shutdowns around real demand.</div></div>"
    ).format(d=cost["d1"].strftime("%d %b"), util=util, dep=ov["deployed"], park=ov["parked"],
             tot=ov["total"], idle=money0(ov["id_daily"]), share=ov["idle_share"])

    # ---- page 4 · where it lands ----------------------------------------
    charges, ch_total, ch_firm = charges_data
    ch_count = sum(charges[ln]["count"] for ln in CHARGE_LINES)

    pred_rows, pred_total = predicted_breakdown(cost)   # composition sums to the plan (FULL)
    pmax = max([a for _l, a in pred_rows] + [1.0])
    pbars = "".join(
        ("<div style='display:flex;align-items:center;gap:10px;margin:8px 0'>"
         "<div style='width:230px;color:" + MUTED + ";font-size:12.5px'>{l}</div>"
         "<div style='flex:1;background:" + TRACK + ";border-radius:6px;overflow:hidden'>"
         "<div style='height:14px;width:{w:.1f}%;background:#F26222;border-radius:6px'></div></div>"
         "<div style='width:90px;text-align:right;color:" + INK + ";font-size:12.5px'>{v}</div></div>").format(
            l=l, w=max(2.0, a / pmax * 100), v=money0(a)) for l, a in pred_rows)

    charges_strip = ""
    if ch_count == 0:
        charges_strip = (
            "<div class='card strip'><span class='pill watch' style='margin-top:1px'>Charges</span>"
            "<p><b>Client charges &mdash; none logged yet.</b> Damage, consumables sold, fuel and transport "
            "are raised in the Charge Reporter and added here as they're approved.</p></div>")

    landing_page = (
        "<div class='card' style='border-left:4px solid #F26222'>"
        "<h2>Predicted full-shutdown cost</h2>"
        "<div class='cap'>Where K2 is tracking to land &mdash; days run locked at actual, days ahead at forecast.</div>"
        "<div class='bigmoney'>{TR}</div>"
        "<div class='bigsub'>Predicted landing · locked + plan ahead</div>"
        "<div class='ceiling'>The plan / forecast to completion is <b>{FULL}</b> &mdash; that's the ceiling, "
        "and we won't bill over it.</div>"
        "<div style='margin-top:20px;color:" + FAINT + ";font-size:11px;text-transform:uppercase;letter-spacing:.6px;"
        "font-weight:700'>What the plan is made of</div>"
        "<div style='margin-top:6px'>{bars}</div>"
        "<div style='margin-top:20px;color:" + FAINT + ";font-size:11px;text-transform:uppercase;letter-spacing:.6px;"
        "font-weight:700'>The month view &mdash; where the landing falls</div>"
        "<div style='margin-top:6px'>{months}</div>"
        "<div class='footnote'>Month figures are the days already run at their trued value plus the days "
        "ahead at forecast &mdash; they sum exactly to the landing. The under-run on invoiced days is "
        "banked &mdash; it won't climb back up. Client charges (damage, consumables, fuel, transport) sit "
        "outside the hire plan and are added as they're logged.</div></div>"
    ).format(TR=money0(tracking), FULL=money0(FULL), bars=pbars,
             months="".join(
                 ("<div style='display:flex;align-items:center;gap:10px;margin:8px 0'>"
                  "<div style='width:230px;color:" + MUTED + ";font-size:12.5px'>{m}</div>"
                  "<div style='flex:1;background:" + TRACK + ";border-radius:6px;overflow:hidden'>"
                  "<div style='height:14px;width:{w:.1f}%;background:linear-gradient(90deg,#F26222,#f7883f);"
                  "border-radius:6px'></div></div>"
                  "<div style='width:150px;text-align:right;color:" + INK + ";font-size:12.5px'>{v}"
                  "<span style='color:" + FAINT + ";font-size:10.5px'> ({run} run)</span></div></div>").format(
                     m=k, v=money0(v["run"] + v["ahead"]), run=money0(v["run"]),
                     w=max(2.0, (v["run"] + v["ahead"]) /
                           max(1.0, max(mv["run"] + mv["ahead"] for _k, mv in cost["months"])) * 100))
                 for k, v in cost["months"])) + charges_strip

    # ---- page 5 · charges (own page only once there's something to say) --
    charges_page = None
    if ch_count > 0:
        ch_rows = "".join(
            "<tr><td>{ln}</td><td class='n'>{c}</td><td class='n'>{t}</td><td class='n'>{f}</td></tr>".format(
                ln=html.escape(ln), c=charges[ln]["count"], t=money0(charges[ln]["total"]),
                f=money0(charges[ln]["firm"])) for ln in CHARGE_LINES if charges[ln]["count"] > 0)
        charges_page = (
            "<div class='card'><h2>Charges captured this shutdown</h2>"
            "<div class='cap'>Logged in the Charge Reporter &mdash; Damage, Consumables sold, Fuel, Transport. "
            "Firm = Approved or Invoiced.</div>"
            "<table><thead><tr><th>Cost line</th><th style='text-align:right'>Count</th>"
            "<th style='text-align:right'>Total</th><th style='text-align:right'>Firm</th></tr></thead><tbody>"
            + ch_rows +
            "<tr><td><b>Total</b></td><td class='n'><b>{c}</b></td><td class='n'><b>{t}</b></td>"
            "<td class='n'><b>{f}</b></td></tr>".format(c=ch_count, t=money0(ch_total), f=money0(ch_firm))
            + "</tbody></table>"
            "<div class='footnote'><b>{f}</b> firm and <b>{p}</b> provisional across {c} charge(s). Firm "
            "charges roll into the costing under their line.</div></div>".format(
                f=money0(ch_firm), p=money0(ch_total - ch_firm), c=ch_count))

    # ---- page 6 · people & assurance (drops off if no roster) ------------
    people_page = None
    if roster:
        mh = ("<div class='card'><h2>Tool store man-hours</h2>"
              "<div class='cap'>From 20 Jul &middot; day &amp; night shifts</div>"
              "<div class='kpis' style='grid-template-columns:repeat(3,1fr);margin-bottom:0'>"
              "<div class='kpi'><div class='v'>{tt:.0f}h</div><div class='l'>Total</div></div>"
              "<div class='kpi'><div class='v'>{dd:.0f}h</div><div class='l'>Day</div></div>"
              "<div class='kpi'><div class='v'>{nn:.0f}h</div><div class='l'>Night</div></div></div>"
              "<div class='footnote'>Planned to completion: <b>{pt:.0f}h</b> ({pd:.0f} day / {pn:.0f} night). "
              "Night shift ramps up as the shutdown peaks.</div></div>").format(
                  tt=roster["td_total"], dd=roster["td_day"], nn=roster["td_night"],
                  pt=roster["pl_total"], pd=roster["pl_day"], pn=roster["pl_night"])
        sf = ("<div class='card'><h2>Safety engagement &mdash; daily</h2>"
              "<div class='cap'>Auto-tallied from the roster, from 20 Jul ({days} days)</div>"
              "<table><thead><tr><th>Activity</th><th style='text-align:right'>Standard</th>"
              "<th style='text-align:right'>Logged</th></tr></thead><tbody>"
              "<tr><td>Take 5s</td><td class='n'>3 / person / day</td><td class='n'>{t5}</td></tr>"
              "<tr><td>Safety conversations</td><td class='n'>1 / shift</td><td class='n'>{cv}</td></tr>"
              "<tr><td>Pre-start meetings</td><td class='n'>1 / day</td><td class='n'>{ps}</td></tr>"
              "</tbody></table>"
              "<div class='footnote'>Captured automatically across {days} day(s) and {sh} shift(s) &mdash; "
              "{pp} person-days on the tools. Every shift starts with a conversation; each day opens with a "
              "pre-start.</div></div>").format(
                  days=roster["days"], t5=roster["take5"], cv=roster["convos"],
                  ps=roster["prestarts"], sh=roster["shifts"], pp=roster["persondays"])
        people_page = "<div class='cols'>" + mh + sf + "</div>"

    # ---- last page · the team (footer lives on the last page) ------------
    team_page = team_html()

    footer = ("<div class='foot'><span>Data source: <b>SiteIQ</b> · Extract {asat} · Contracted daily rates; "
              "unpriced items excluded, never estimated · Point-in-time snapshot, shared weekly.</span>"
              "<span style='white-space:nowrap'>Author: <b>Andrew Fisher</b> · POWERED BY SITEIQ</span></div>"
              ).format(asat=asat)

    # assemble - one story per sheet; empty sections have already dropped off
    sections = [("__first__", page1),
                ("The cost detail · forecast vs actual by stream", streams_page)]
    for tp in trueup_pages:
        sections.append(("The true-up · tracked v SiteIQ invoiced", tp))
    if claims_page:
        sections.append(("Progress claims · labour &amp; accommodation", claims_page))
    if claims_lines_page:
        sections.append(("Progress claims · as SiteIQ booked them", claims_lines_page))
    sections += [("Plant on the ground · utilisation &amp; idle plant", util_page),
                 ("The landing · predicted cost &amp; the month view", landing_page)]
    if charges_page:
        sections.append(("Client charges · captured this shutdown", charges_page))
    if people_page:
        sections.append(("People &amp; assurance · man-hours and safety", people_page))
    sections.append(("The team · who's running your tool store", team_page))

    total_pages = len(sections)

    def _hdr(sub, pg):
        return ("<div class='hd2'><div>"
                "<div class='brand'>COATES<span>Equipped for anything</span></div>"
                "<h1>Cost Tracking Snapshot</h1><div class='rp'>{sub}</div>"
                "</div><div><div class='siteiq'>POWERED BY SITEIQ</div>"
                "<div class='meta'>As at <b>{asat}</b><br>Page {pg} of {tp}</div></div></div>"
                ).format(sub=sub, asat=asat, pg=pg, tp=total_pages)

    sheets = []
    for i, (sub, inner) in enumerate(sections, 1):
        last = (i == total_pages)
        if sub == "__first__":
            content = inner.replace("PAGEMARK", "Page 1 of {}".format(total_pages))
        else:
            content = _hdr(sub, i) + "<div class='rule'></div>" + inner
        sheets.append("<section class='sheet'><div class='panel'>" + content +
                      "<div class='grow'></div>" + (footer if last else "") + "</div></section>")

    # Concatenate (never .format) around PANEL_CSS — the CSS has its own { } braces
    # (e.g. @page{size:A4}) that .format would try to read as fields.
    _head = ("<!DOCTYPE html><html><head><meta charset='utf-8'>"
             "<meta name='viewport' content='width=device-width,initial-scale=1'>"
             "<title>Coates K2 - Cost Tracking Snapshot</title><style>" + PANEL_CSS + LIGHT_CSS + "</style></head><body>")
    _tail = "</body></html>"
    full = _head + "".join(sheets) + _tail
    pages = [_head + s + _tail for s in sheets]
    return full, pages, total_pages


# =====================================================================
#  COATES BILLING FORECAST (INTERNAL) - what SiteIQ will invoice
#  Plant & tooling hire only. Radio, gas monitors and the welders are
#  invoiced separately and NEVER appear in these figures; labour,
#  accommodation and transport are entered or claimed directly, so they
#  land in the month they're put in. THE MONTH-END RULE (A. Fisher,
#  24 Jul 2026): SiteIQ does not invoice the last day of a month within
#  that month - invoice July on 31 Jul and the 31st's hire is not in it;
#  it registers into August's billing.
# =====================================================================
def coates_billing(wb, ds=None, hs=None):
    """The SiteIQ billing forecast by BILLING month: every day's plant &
    tooling hire value (invoiced days locked at DAILY_SUMMARY, today at
    tracked, days ahead at forecast - identical rules to the cost story)
    assigned to the month SiteIQ will actually invoice it in: its own
    month, except the month's LAST day, which registers into the next
    month. Returns month rows, the roll-over days, and the plant/tooling
    forecast totals (invoiced split + workbook split ahead)."""
    import calendar
    ws = wb["Daily Tracking"]
    ds = ds or {}

    def num(r, c):
        v = ws.cell(r, c).value
        return v if isinstance(v, (int, float)) else 0

    rows = [r for r in range(2, 60) if isinstance(ws.cell(r, 1).value, (dt.datetime, dt.date))]
    today = dt.date.today()

    def rowdate(r):
        v = ws.cell(r, 1).value
        return v.date() if isinstance(v, dt.datetime) else v

    months = {}
    msplit = {}
    rolls = []
    run_total = 0.0
    prov_p = prov_t = fut_p = fut_t = 0.0
    # everything else Coates bills through SiteIQ, by CALENDAR month - no
    # month-end rule for these: they land in the month they're entered or
    # claimed (A. Fisher, 24 Jul 2026). Radio / gas / welders stay out.
    others = {}
    OTHERS = (("labour", 14, 15), ("accom", 17, 18),
              ("transport", 20, 21), ("damage", 23, 24))
    for r in rows:
        d = rowdate(r)
        p_fc, t_fc = num(r, 2), num(r, 5)
        tracked = num(r, 3) + num(r, 6)
        is_active = num(r, 27) > 0
        ds_locked = d in ds and (ds[d]["hire"] > 0 or (ds and d < max(ds)))
        if ds_locked:
            v, st = ds[d]["hire"], "run"
        elif is_active and d <= today and tracked > 0:
            v, st = tracked, "run"
            prov_p += num(r, 3)
            prov_t += num(r, 6)
        else:
            v, st = p_fc + t_fc, "ahead"
            fut_p += p_fc
            fut_t += t_fc
        if st == "run":
            run_total += v
        last_day = calendar.monthrange(d.year, d.month)[1]
        if d.day == last_day:            # the month-end rule: rolls forward
            bm = (d.replace(day=1) + dt.timedelta(days=32)).replace(day=1)
            if v:
                rolls.append({"d": d, "v": v,
                              "to": bm.strftime("%B %Y")})
        else:
            bm = d.replace(day=1)
        m = months.setdefault(bm, {"run": 0.0, "ahead": 0.0,
                                   "rolled_in": 0.0, "order": bm})
        m[st] += v
        if d.day == last_day and v:
            m["rolled_in"] += v
        # plant v tooling composition for THIS billing month: locked and
        # provisional days at the workbook's tracked split, days ahead at
        # the plan split. The month's BILL stays the SiteIQ total above -
        # the composition explains it, the drift line reconciles it.
        ms = msplit.setdefault(bm, {"plant": 0.0, "tool": 0.0,
                                    "trk_locked": 0.0, "inv_locked": 0.0})
        if ds_locked:
            ms["plant"] += num(r, 3)
            ms["tool"] += num(r, 6)
            ms["trk_locked"] += tracked
            ms["inv_locked"] += v
        elif st == "run":
            ms["plant"] += num(r, 3)
            ms["tool"] += num(r, 6)
        else:
            ms["plant"] += p_fc
            ms["tool"] += t_fc
        # -- everything else that pops up, calendar month ------------------
        cm = d.replace(day=1)
        om = others.setdefault(cm, {k: 0.0 for k, _f, _a in OTHERS})
        run_day = is_active and d <= today
        for key, fcol, acol in OTHERS:
            if key == "transport":
                # recognised at forecast - incurred, billed in arrears
                dv = max(num(r, acol), num(r, fcol)) if run_day else num(r, fcol)
            else:
                dv = num(r, acol) if run_day else num(r, fcol)
            om[key] += dv
    # fuel, consumables and any other SiteIQ charge that pops up: no plan
    # column exists for these, so the ONLY figures shown are what SiteIQ
    # has actually invoiced (DAILY_SUMMARY) - added as raised, never
    # forecast. Calendar month of the charged day.
    fuelcons, misc = {}, {}
    for d, v in ds.items():
        cm = d.replace(day=1)
        fuelcons[cm] = fuelcons.get(cm, 0.0) + v.get("fuel", 0.0) + v.get("consumable", 0.0)
        misc[cm] = misc.get(cm, 0.0) + (v.get("disposal", 0.0) + v.get("other", 0.0)
                                        + v.get("waiver", 0.0) + v.get("stamp", 0.0))
    month_rows = sorted(months.items(), key=lambda kv: kv[1]["order"])
    # plant / tooling forecast totals - invoiced split where truthful
    invoiced_total = run_total - (prov_p + prov_t)
    if invoiced_total < 1:
        inv_p = inv_t = 0.0
        split_ok = True
    elif hs is not None and abs(hs["total"] - invoiced_total) < 1.0:
        inv_p, inv_t = hs["plant"], hs["tool"]
        split_ok = True
    else:
        split_ok = False
        inv_p = inv_t = 0.0
    total = sum(m["run"] + m["ahead"] for _k, m in month_rows)
    other_rows = sorted(others.items())
    others_total = {k: sum(om[k] for _c, om in other_rows)
                    for k, _f, _a in OTHERS}
    grand = (total + sum(others_total.values())
             + sum(fuelcons.values()) + sum(misc.values()))
    return {"months": month_rows, "rolls": rolls, "total": total,
            "run": run_total, "split_ok": split_ok,
            "plant": inv_p + prov_p + fut_p if split_ok else None,
            "tool": inv_t + prov_t + fut_t if split_ok else None,
            "others": other_rows, "others_total": others_total,
            "fuelcons": fuelcons, "misc": misc, "grand": grand,
            "msplit": msplit}


def billing_doc(bill, asat, wbname, ds_name, svc=None, charges_data=None):
    """The INTERNAL billing forecast, two framed pages: page 1 is the
    all-in picture - every stream Coates bills through SiteIQ, month by
    month; page 2 is the SiteIQ hire line in detail (the month-end rule
    at work, and the plant/tooling split)."""
    svc = svc or {"labour": 0.0, "accom": 0.0, "other": [], "total": 0.0}
    hire_m = dict(bill["months"])
    other_m = dict(bill["others"])
    mlist = sorted(set(list(hire_m.keys()) + list(other_m.keys())
                       + list(bill["fuelcons"].keys()) + list(bill["misc"].keys())))

    def hire_val(cm):
        m = hire_m.get(cm)
        return (m["run"] + m["ahead"]) if m else 0.0

    def oth(cm, key):
        return other_m.get(cm, {}).get(key, 0.0)

    # ---- page 1 · the all-in picture ------------------------------------
    def month_all_in(cm):
        return (hire_val(cm) + sum(oth(cm, k) for k in
                                   ("labour", "accom", "transport", "damage"))
                + bill["fuelcons"].get(cm, 0.0) + bill["misc"].get(cm, 0.0))

    tiles = ""
    for cm in mlist[:2]:
        tiles += ("<div class='kpi'><div class='v'>{v}</div><div class='l'>{l} billing &middot; all in</div>"
                  "<div class='s'>every SiteIQ-billed stream</div></div>").format(
            v=money0(month_all_in(cm)), l=cm.strftime("%B"))
    tiles += ("<div class='kpi g'><div class='v'>{v}</div><div class='l'>Total &middot; all in</div>"
              "<div class='s'>whole shut, all SiteIQ billing</div></div>").format(v=money0(bill["grand"]))

    mh = "".join("<th style='text-align:right'>{}</th>".format(cm.strftime("%b %Y"))
                 for cm in mlist)

    def allin_row(name, vals, total, note, bold=False):
        style = " style='color:" + INK + ";font-weight:700'" if bold else ""
        cells = "".join("<td class='n'{s}>{v}</td>".format(s=style, v=money0(v))
                        for v in vals)
        return ("<tr><td{s}>{n}</td>{c}<td class='n'{s2}>{t}</td>"
                "<td style='color:" + MUTED + ";white-space:normal'>{note}</td></tr>").format(
            s=style, n=name, c=cells, t=money0(total),
            s2=" style='color:" + INK + ";font-weight:700'", note=note)

    roll_note = "; ".join("{} &rarr; {}".format(x["d"].strftime("%d %b"),
                                                x["to"].split()[0])
                          for x in bill["rolls"]) or ""
    hire_note = "month-end rule" + (": " + roll_note if roll_note else "")
    lab_note = ("progress claims &middot; {} claimed to date".format(money0(svc["labour"]))
                if svc["labour"] > 0 else "progress claims as entered")
    acc_note = ("progress claims &middot; {} claimed to date".format(money0(svc["accom"]))
                if svc["accom"] > 0 else "progress claims as entered")
    rows_1 = (
        allin_row("Plant &amp; tooling hire", [hire_val(cm) for cm in mlist],
                  bill["total"], hire_note)
        + allin_row("Personnel / Labour", [oth(cm, "labour") for cm in mlist],
                    bill["others_total"]["labour"], lab_note)
        + allin_row("Accommodation", [oth(cm, "accom") for cm in mlist],
                    bill["others_total"]["accom"], acc_note)
        + allin_row("Transport", [oth(cm, "transport") for cm in mlist],
                    bill["others_total"]["transport"], "billed in arrears as entered")
        + allin_row("Damage recovery", [oth(cm, "damage") for cm in mlist],
                    bill["others_total"]["damage"], "as raised &amp; approved")
        + allin_row("Fuel &amp; consumables", [bill["fuelcons"].get(cm, 0.0) for cm in mlist],
                    sum(bill["fuelcons"].values()), "as raised &middot; invoiced figures only"))
    if sum(bill["misc"].values()) > 0:
        rows_1 += allin_row("Other SiteIQ charges", [bill["misc"].get(cm, 0.0) for cm in mlist],
                            sum(bill["misc"].values()),
                            "disposal, waivers &amp; the like &middot; as invoiced")
    rows_1 += allin_row("Total", [month_all_in(cm) for cm in mlist],
                        bill["grand"], "", bold=True)

    ch_note = ""
    if charges_data:
        charges, ch_total, ch_firm = charges_data
        ch_count = sum(charges[ln]["count"] for ln in CHARGE_LINES)
        if ch_count:
            ch_note = (" Charge Reporter logged to date: {} across {} charge(s) "
                       "({} firm) &mdash; firm charges roll into their lines as "
                       "they bill.").format(money0(ch_total), ch_count, money0(ch_firm))

    page1 = (
        "<div class='banner'><span class='pill watch'>Month-end rule</span>"
        "<p>SiteIQ does <b>not</b> invoice the last day of a month within that month. Invoice July on "
        "31 Jul and the 31st's plant &amp; tooling hire is <b>not in it</b> &mdash; it registers into "
        "August's billing. Only the hire line works this way &mdash; everything else lands in the "
        "month it's entered or claimed.</p></div>"
        "<div class='kpis' style='grid-template-columns:repeat(3,1fr)'>{tiles}</div>"
        "<div class='card'><h2>Everything Coates bills &mdash; month by month</h2>"
        "<div class='cap'>Hire at invoiced/tracked/forecast (page 2 has the detail); labour and "
        "accommodation at the daily plan, invoiced in progress claims; transport recognised at "
        "forecast, billed in arrears; damage, fuel and consumables added as raised &mdash; never "
        "forecast.</div>"
        "<table class='tight'><thead><tr><th>Stream</th>{mh}"
        "<th style='text-align:right'>Total</th><th>Bills</th></tr></thead>"
        "<tbody>{rows}</tbody></table>"
        "<div class='footnote'>Radio, gas monitor and welder hire are invoiced separately and are "
        "<b>not</b> in these figures; no sub-hired gear is included.{ch}</div></div>"
    ).format(tiles=tiles, mh=mh, rows=rows_1, ch=ch_note)

    # ---- page 2 · the SiteIQ hire line in detail ------------------------
    mr = bill["months"]
    body_rows = ""
    for k, m in mr:
        note = "&mdash;"
        # a day rolled OUT of this calendar month
        rolled_out = [x for x in bill["rolls"]
                      if x["d"].strftime("%B %Y") == k.strftime("%B %Y")]
        rolled_in = [x for x in bill["rolls"] if x["to"] == k.strftime("%B %Y")]
        bits = []
        for x in rolled_in:
            bits.append("{} ({}) rolls IN from {}".format(
                x["d"].strftime("%d %b"), money0(x["v"]), x["d"].strftime("%B")))
        for x in rolled_out:
            bits.append("{} ({}) registers into {}".format(
                x["d"].strftime("%d %b"), money0(x["v"]), x["to"].split()[0]))
        if bits:
            note = "; ".join(bits)
        body_rows += ("<tr><td>{m}</td><td class='n'>{run}</td><td class='n'>{ah}</td>"
                      "<td class='n' style='color:" + INK + ";font-weight:700'>{tot}</td>"
                      "<td style='color:" + MUTED + "'>{note}</td></tr>").format(
            m=k.strftime("%B %Y"), run=money0(m["run"]), ah=money0(m["ahead"]),
            tot=money0(m["run"] + m["ahead"]), note=note)
    if bill["split_ok"]:
        split_card = (
            "<div class='card' style='margin-top:16px'><h2>By stream &mdash; whole shut</h2>"
            "<div class='cap'>Invoiced to date split item-by-item from SiteIQ's charge lines; "
            "days ahead at the workbook plan.</div>"
            "<table><thead><tr><th>Stream</th><th style='text-align:right'>Forecast billing</th></tr></thead>"
            "<tbody><tr><td>Plant hire</td><td class='n'>{p}</td></tr>"
            "<tr><td>Tooling hire</td><td class='n'>{t}</td></tr></tbody></table>"
            "<div class='footnote'>Together {tot} &mdash; the total above. Locked days never move; "
            "the forecast never estimates.</div></div>"
        ).format(p=money0(bill["plant"]), t=money0(bill["tool"]), tot=money0(bill["total"]))
    else:
        split_card = (
            "<div class='card' style='margin-top:16px'><h2>By stream &mdash; whole shut</h2>"
            "<div class='cap'>The plant/tooling split is unavailable this run (charge lines missing or "
            "out of step with DAILY_SUMMARY) &mdash; the combined SiteIQ hire figure above is the "
            "billing truth; the split returns with the next matching export. Never apportioned.</div></div>")
    # ---- the month's split: plant v tooling, drawn ----------------------
    msp = bill.get("msplit") or {}
    split_rows = ""
    split_bars = ""
    drift_bits = []
    for cm in sorted(msp):
        e = msp[cm]
        comp = e["plant"] + e["tool"]
        mtot = hire_m.get(cm)
        mtot_v = (mtot["run"] + mtot["ahead"]) if mtot else comp
        split_rows += ("<tr><td>{m}</td><td class='n'>{p}</td>"
                       "<td class='n'>{t}</td>"
                       "<td class='n' style='color:" + INK + ";font-weight:700'>{tot}</td></tr>").format(
            m=cm.strftime("%B %Y"), p=money0(e["plant"]), t=money0(e["tool"]),
            tot=money0(mtot_v))
        if comp > 0:
            pw = e["plant"] / comp * 100
            split_bars += (
                "<div style='display:flex;align-items:center;gap:10px;margin:7px 0'>"
                "<div style='width:110px;color:" + MUTED + ";font-size:12px'>{m}</div>"
                "<div style='flex:1;display:flex;border-radius:6px;overflow:hidden;height:15px'>"
                "<div style='width:{pw:.1f}%;background:#F26222'></div>"
                "<div style='width:{tw:.1f}%;background:#fab219'></div></div>"
                "<div style='width:190px;text-align:right;color:" + INK + ";font-size:11.5px'>"
                "{pp:.0f}% plant &middot; {tp:.0f}% tooling</div></div>").format(
                m=cm.strftime("%b %Y"), pw=pw, tw=100 - pw,
                pp=pw, tp=100 - pw)
        if e["inv_locked"]:
            drift_bits.append("{m}: SiteIQ invoiced {i} v tracked {t} on locked "
                              "days (drift {d})".format(
                                  m=cm.strftime("%b"),
                                  i=money0(e["inv_locked"]),
                                  t=money0(e["trk_locked"]),
                                  d=money0(round(e["inv_locked"] - e["trk_locked"]))))
    split_month_card = ""
    if split_rows:
        split_month_card = (
            "<div class='card' style='margin-top:16px'>"
            "<h2>The month's split &mdash; plant v tooling</h2>"
            "<div class='cap'>What each billing month is made of: locked and "
            "run days at the workbook's tracked split, days ahead at the "
            "plan split. The month's bill is the SiteIQ total &mdash; the "
            "composition here is how you explain it.</div>"
            "<table class='tight'><thead><tr><th>Billing month</th>"
            "<th style='text-align:right'>Plant</th>"
            "<th style='text-align:right'>Tooling</th>"
            "<th style='text-align:right'>Month bill</th></tr></thead>"
            "<tbody>" + split_rows + "</tbody></table>"
            "<div style='margin-top:10px'>" + split_bars + "</div>"
            "<div class='footnote'>"
            "<span style='display:inline-block;width:9px;height:9px;background:#F26222;"
            "border-radius:2px;margin-right:5px'></span>Plant "
            "<span style='display:inline-block;width:9px;height:9px;background:#fab219;"
            "border-radius:2px;margin:0 5px 0 14px'></span>Tooling &mdash; "
            + ("; ".join(drift_bits) + ". The invoiced figure is always the "
               "bill; the tracked split explains its make-up." if drift_bits
               else "no locked-day drift yet.")
            + "</div></div>")

    page2 = (
        "<div class='card'><h2>The SiteIQ hire line &mdash; billing month by month</h2>"
        "<div class='cap'>Days SiteIQ has invoiced are locked at the invoice; today runs at tracked; "
        "days ahead at the workbook plan &mdash; each day assigned to the month SiteIQ will actually "
        "bill it in.</div>"
        "<table><thead><tr><th>Billing month</th><th style='text-align:right'>Invoiced + tracked</th>"
        "<th style='text-align:right'>At forecast</th><th style='text-align:right'>Total</th>"
        "<th>Month-end days</th></tr></thead><tbody>{rows}</tbody></table>"
        "<div class='footnote'>Plant &amp; tooling hire only &mdash; the SiteIQ hire line. "
        "Sources: {ds} (invoiced days) &middot; {wb} (plan).</div></div>"
        "{msc}"
    ).format(rows=body_rows, ds=html.escape(ds_name), wb=html.escape(wbname),
             msc=split_month_card)
    page3 = split_card
    if bill.get("split_ok") and (bill.get("plant") or 0) + (bill.get("tool") or 0) > 0:
        _pt = bill["plant"] + bill["tool"]
        _pw = bill["plant"] / _pt * 100
        page3 += (
            "<div class='card' style='margin-top:16px'>"
            "<h2>The whole shut, one bar</h2>"
            "<div class='cap'>Every SiteIQ hire dollar, plant v tooling.</div>"
            "<div style='display:flex;border-radius:8px;overflow:hidden;height:34px'>"
            "<div style='width:{pw:.1f}%;background:#F26222;display:flex;align-items:center;"
            "justify-content:center;color:#fff;font-weight:800;font-size:13px'>"
            "PLANT {p} &middot; {pp:.0f}%</div>"
            "<div style='width:{tw:.1f}%;background:#fab219;display:flex;align-items:center;"
            "justify-content:center;color:#1a1a1a;font-weight:800;font-size:13px'>"
            "TOOLING {t} &middot; {tp:.0f}%</div></div>"
            "<div class='footnote'>Easy to explain: roughly {pp:.0f} cents of "
            "every SiteIQ hire dollar this shut is plant, {tp:.0f} is tooling "
            "&mdash; invoiced to date split item-by-item from SiteIQ's own "
            "charge lines, days ahead at the workbook plan.</div></div>").format(
            pw=_pw, tw=100 - _pw, p=money0(bill["plant"]),
            t=money0(bill["tool"]), pp=_pw, tp=100 - _pw)

    # ---- assemble: three framed sheets, numbered, INTERNAL header --------
    sections = [("Everything Coates bills &middot; all in, month by month", page1),
                ("The SiteIQ hire line &middot; month by month, split", page2),
                ("Plant v tooling &middot; the whole shut", page3)]
    footer = ("<div class='foot'><span>COATES INTERNAL &middot; billing forecast &middot; "
              "what SiteIQ bills, month by month.</span>"
              "<span style='white-space:nowrap'>Author: <b>Andrew Fisher</b> · POWERED BY SITEIQ</span></div>")

    def _hdr(sub, pg):
        return ("<div class='hd2'><div>"
                "<div class='brand'>COATES<span>Equipped for anything</span></div>"
                "<h1>Coates Billing Forecast</h1><div class='rp'>{sub}</div>"
                "</div><div><div class='siteiq'>COATES INTERNAL</div>"
                "<div class='meta'>As at <b>{asat}</b><br>Not for client distribution<br>"
                "Page {pg} of {tp}</div></div></div>").format(
            sub=sub, asat=asat, pg=pg, tp=len(sections))

    sheets = []
    for i, (sub, inner) in enumerate(sections, 1):
        last = (i == len(sections))
        sheets.append("<section class='sheet'><div class='panel'>"
                      + _hdr(sub, i) + "<div class='rule'></div>" + inner
                      + "<div class='grow'></div>" + (footer if last else "")
                      + "</div></section>")
    _head = ("<!DOCTYPE html><html><head><meta charset='utf-8'>"
             "<meta name='viewport' content='width=device-width,initial-scale=1'>"
             "<title>Coates K2 - Billing Forecast (INTERNAL)</title><style>" + PANEL_CSS + LIGHT_CSS + "</style></head><body>")
    _tail = "</body></html>"
    full = _head + "".join(sheets) + _tail
    pages = [_head + s + _tail for s in sheets]
    return full, pages


def make_pdf(html_path, pdf_path):
    """PDF via weasyprint, else headless Edge/Chrome (as on a Coates PC)."""
    try:
        #  Probe quietly: on a machine without weasyprint's native
        #  libraries the import prints a scary multi-line banner before
        #  failing, and the Edge fallback below was already making the
        #  PDF perfectly well. (Andrew's morning run, 27 Jul 2026.)
        import contextlib
        import io as _io
        with contextlib.redirect_stdout(_io.StringIO()), \
                contextlib.redirect_stderr(_io.StringIO()):
            from weasyprint import HTML
            HTML(filename=html_path).write_pdf(pdf_path)
        return os.path.isfile(pdf_path) and os.path.getsize(pdf_path) > 0
    except Exception:
        pass
    import subprocess
    edge = _find_browser()
    if not edge:
        return False
    prof = os.path.join(_tmp(), "coates_snap_pdf")
    url = "file:///" + os.path.abspath(html_path).replace("\\", "/")
    try:
        subprocess.run([edge, "--headless", "--disable-gpu", "--no-first-run"]
                       + _browser_extra() +
                       ["--user-data-dir=" + prof, "--no-pdf-header-footer",
                        "--print-to-pdf=" + os.path.abspath(pdf_path), url],
                       capture_output=True, timeout=180)
    except Exception:
        return False
    return os.path.isfile(pdf_path) and os.path.getsize(pdf_path) > 0


def _find_browser():
    """Locate a headless-capable Edge/Chrome as shipped on a Coates PC.
    One list for the whole kit - see browser_engine.py."""
    try:
        import browser_engine
        return browser_engine.find_browser()
    except Exception:
        pass
    import shutil
    return next((p for p in [
        os.path.expandvars(r"%ProgramFiles(x86)%\Microsoft\Edge\Application\msedge.exe"),
        os.path.expandvars(r"%ProgramFiles%\Microsoft\Edge\Application\msedge.exe"),
        shutil.which("msedge"), shutil.which("microsoft-edge"),
        shutil.which("google-chrome"), shutil.which("chrome"),
        shutil.which("chromium"), shutil.which("chromium-browser"),
    ] if p and os.path.isfile(p)), None)


def _tmp():
    """Browser profiles belong in the temp folder, not in the kit. TEMP
    is a Windows variable and was undefined everywhere else."""
    import tempfile
    return os.environ.get("TEMP") or tempfile.gettempdir()


def _reserved(browser):
    try:
        import email_images
        return email_images.reserved_height(browser)
    except Exception:
        return 120


def _trim(png):
    try:
        import email_images
        email_images._crop(png, 794 * 2, 1123 * 2)
    except Exception:
        pass


def _browser_extra():
    try:
        import browser_engine
        return browser_engine.extra_args()
    except Exception:
        return [] if os.name == "nt" else ["--no-sandbox"]


def render_page_images(page_htmls, outdir, stamp, prefix="Coates_K2_Cost_Snapshot"):
    """Render each standalone page HTML to a PNG — one A4 sheet each, drawn by the
    same headless Edge/Chrome that builds the PDF, so the email matches it page for
    page. The PNGs are KEPT in the day folder (handy for Teams / WhatsApp / decks).
    Returns the list of PNG paths, or [] if a browser isn't available."""
    browser = _find_browser()
    if not browser:
        return []
    import subprocess
    prof = os.path.join(_tmp(), "coates_snap_img")
    pngs = []
    for i, ph in enumerate(page_htmls, 1):
        tmp_html = os.path.join(outdir, "_render_page{}.html".format(i))
        with open(tmp_html, "w", encoding="utf-8") as f:
            f.write(ph)
        png = os.path.join(outdir, "{}_{}_Page{}.png".format(prefix, stamp, i))
        url = "file:///" + os.path.abspath(tmp_html).replace("\\", "/")
        try:
            subprocess.run([browser, "--headless", "--disable-gpu", "--no-first-run",
                            "--hide-scrollbars", "--force-device-scale-factor=2"]
                           + _browser_extra() + [
                            "--user-data-dir=" + prof,
                            #  Taller than the sheet by whatever the browser
                            #  reserves for itself - otherwise the foot of
                            #  every page, team strip and all, is cut off.
                            #  See email_images.reserved_height (25 Jul 2026).
                            "--window-size=794,{}".format(1123 + _reserved(browser)),
                            "--screenshot=" + os.path.abspath(png), url],
                           capture_output=True, timeout=120)
            _trim(png)
        except Exception:
            pass
        try:
            os.remove(tmp_html)
        except Exception:
            pass
        if os.path.isfile(png) and os.path.getsize(png) > 0:
            pngs.append(png)
    return pngs if len(pngs) == len(page_htmls) else []


def image_email(cost, png_paths, to="", cc=""):
    """The snapshot pages pasted INTO the email body, in order - built by
    the shared email_images module (inline disposition, no filenames, no
    attachments, CRLF), so it renders in the compose body, not on the
    attachment strip. Pre-addressed from the recipients book (tag: COST)."""
    import email_images
    subject = "K2 Shutdown - Cost Tracking Snapshot (as at {})".format(
        cost["d1"].strftime("%d %b"))
    return email_images.build_image_eml(subject, png_paths, to=to, cc=cc, baked=False)


def _framed(inner, width=680):
    """One Coates border around an email body, shared with the rest of
    the kit (email_images.frame_html)."""
    try:
        import email_images
        return email_images.frame_html(inner, width=width)
    except Exception:
        return ("<html><body style='margin:0;padding:0;background:#ffffff;'>"
                + inner + "</body></html>")


def email_body(cost, ov, charges_data, roster, asat):
    """Fallback dark Coates-style email body (Outlook-safe tables) for machines
    with no browser to draw the page images. Same one-story-per-section flow:
    the call, the tiles, the why, streams, plant, landing, charges, people, team."""
    F, A, V, FULL = cost["F"], cost["A"], cost["V"], cost["FULL"]
    tracking = FULL + V
    period = "{} - {}".format(cost["d0"].strftime("%d %b"), cost["d1"].strftime("%d %b %Y"))
    FN = "font-family:Calibri,Arial,sans-serif;"
    CARD, LINE, ORGC, TXT, MUT, W = "#171B22", "#2A313C", "#F26222", "#D7DBE2", "#A9B1BD", "#ffffff"

    def sect(title, inner, sub=""):
        h = ("<tr><td style='padding:18px 22px 0;'><div style='{f}font-size:15px;font-weight:bold;color:{w};"
             "border-bottom:2px solid {o};padding-bottom:5px;'>{t}</div>").format(f=FN, w=W, o=ORGC, t=title)
        if sub:
            h += "<div style='{f}font-size:11px;color:{m};margin-top:3px;'>{s}</div>".format(f=FN, m=MUT, s=sub)
        return h + "</td></tr><tr><td style='padding:10px 22px 0;'>" + inner + "</td></tr>"

    kdefs = [("Forecast to date", money0(F)), ("Actual to date", money0(A)), ("Variance", money0(V)),
             ("Forecast to completion", money0(FULL)), ("Tracking to land", money0(tracking))]
    kcells = "".join(
        "<td width='20%' bgcolor='{c}' style='{f}border:1px solid {l};border-left:3px solid {o};padding:10px 8px;vertical-align:top;'>"
        "<div style='font-size:16px;font-weight:bold;color:{w};'>{v}</div>"
        "<div style='font-size:9px;color:{m};text-transform:uppercase;margin-top:4px;'>{lab}</div></td>".format(
            f=FN, c=CARD, l=LINE, o=ORGC, w=W, m=MUT, v=v, lab=lab) for lab, v in kdefs)
    kpis = "<table width='100%' cellpadding='0' cellspacing='3'><tr>" + kcells + "</tr></table>"

    unders = sorted([s for s in cost["streams"] if s["var"] < -1], key=lambda s: s["var"])
    drivers = " and ".join(s["name"].split(" / ")[0].lower() for s in unders[:2]) or "timing"
    story = ("<div style='{f}font-size:13px;color:{t};line-height:1.65;'>"
             "{days} active days in, the K2 tool store is <b style='color:{w};'>under the plan and under control</b> "
             "&mdash; and trued to SiteIQ's invoicing daily: days SiteIQ has invoiced are <b style='color:{w};'>locked "
             "at the invoiced figure</b>. The gap is timing, not scope: mobilisation transport is recognised at "
             "forecast, and the under-run sits in {drivers} against forecast on days already run. Nothing has been "
             "dropped or deferred.</div>"
             ).format(f=FN, t=TXT, w=W, days=cost["days"], drivers=drivers)

    sr = ""
    j = 0
    for s in cost["streams"]:
        r = rag(s)
        if not r:
            continue
        colour, status = r
        bg = CARD if j % 2 == 0 else "#12161d"
        j += 1
        sr += ("<tr><td bgcolor='{bg}' style='{f}font-size:12px;color:{t};padding:6px 9px;border-bottom:1px solid {l};'>{n}</td>"
               "<td bgcolor='{bg}' style='{f}font-size:12px;color:{w};padding:6px 9px;border-bottom:1px solid {l};text-align:right;'>{fc}</td>"
               "<td bgcolor='{bg}' style='{f}font-size:12px;color:{w};padding:6px 9px;border-bottom:1px solid {l};text-align:right;'>{ac}</td>"
               "<td bgcolor='{bg}' style='{f}font-size:12px;color:{t};padding:6px 9px;border-bottom:1px solid {l};'><span style='color:{c};'>&#9679;</span> {st}</td></tr>").format(
                   bg=bg, f=FN, t=TXT, w=W, l=LINE, n=html.escape(s["name"]), fc=money0(s["fc"]), ac=money0(s["act"]), c=colour, st=status)
    stream_tbl = ("<table width='100%' cellpadding='0' cellspacing='0' style='border-collapse:collapse;'>"
                  "<tr><td bgcolor='#0c0f14' style='{f}font-size:9px;color:{m};text-transform:uppercase;padding:6px 9px;'>Cost stream</td>"
                  "<td bgcolor='#0c0f14' style='{f}font-size:9px;color:{m};text-transform:uppercase;padding:6px 9px;text-align:right;'>Forecast</td>"
                  "<td bgcolor='#0c0f14' style='{f}font-size:9px;color:{m};text-transform:uppercase;padding:6px 9px;text-align:right;'>Actual</td>"
                  "<td bgcolor='#0c0f14' style='{f}font-size:9px;color:{m};text-transform:uppercase;padding:6px 9px;'>Status</td></tr>{r}</table>").format(f=FN, m=MUT, r=sr)

    plant = ("<div style='{f}font-size:13px;color:{t};line-height:1.6;'><b style='color:{w};'>{util}% of site plant is out with crews</b>; "
             "<b style='color:{w};'>{idle}/day</b> is on site but not yet out ({share}%). While gear is on site it's charged, used or not &mdash; "
             "a live view of what's on the ground, shared so you can plan gear delivery and future shutdowns around real demand.</div>"
             ).format(f=FN, t=TXT, w=W, util=ov["util"], idle=money0(ov["id_daily"]), share=ov["idle_share"])

    predicted = ("<div style='{f}'><span style='font-size:26px;font-weight:bold;color:{w};'>{TR}</span> "
                 "<span style='font-size:12px;color:{m};'>predicted landing &middot; the {FULL} plan is the ceiling</span></div>"
                 "<div style='{f}font-size:12.5px;color:{t};line-height:1.55;margin-top:6px;'>Days already run are locked at actual; days "
                 "still to come are costed at forecast. It won't go over the plan.</div>").format(
                     f=FN, w=W, m=MUT, t=TXT, TR=money0(tracking), FULL=money0(FULL))

    charges, ch_total, ch_firm = charges_data
    ch_count = sum(charges[ln]["count"] for ln in CHARGE_LINES)
    if ch_count == 0:
        charges_blk = ("<div style='{f}font-size:12.5px;color:{t};line-height:1.55;'><b style='color:{w};'>None logged yet.</b> "
                       "Damage, consumables sold, fuel and transport are raised in the Charge Reporter and added as they're approved.</div>"
                       ).format(f=FN, t=TXT, w=W)
    else:
        cr = "".join(
            "<tr><td bgcolor='{bg}' style='{f}font-size:12px;color:{t};padding:6px 9px;border-bottom:1px solid {l};'>{ln}</td>"
            "<td bgcolor='{bg}' style='{f}font-size:12px;color:{w};padding:6px 9px;border-bottom:1px solid {l};text-align:right;'>{c}</td>"
            "<td bgcolor='{bg}' style='{f}font-size:12px;color:{w};padding:6px 9px;border-bottom:1px solid {l};text-align:right;'>{tot}</td></tr>".format(
                bg=(CARD if i % 2 == 0 else "#12161d"), f=FN, t=TXT, w=W, l=LINE, ln=ln, c=charges[ln]["count"], tot=money0(charges[ln]["total"]))
            for i, ln in enumerate(CHARGE_LINES) if charges[ln]["count"] > 0)
        charges_blk = ("<table width='100%' cellpadding='0' cellspacing='0' style='border-collapse:collapse;'>"
                       "<tr><td bgcolor='#0c0f14' style='{f}font-size:9px;color:{m};text-transform:uppercase;padding:6px 9px;'>Cost line</td>"
                       "<td bgcolor='#0c0f14' style='{f}font-size:9px;color:{m};text-transform:uppercase;padding:6px 9px;text-align:right;'>Count</td>"
                       "<td bgcolor='#0c0f14' style='{f}font-size:9px;color:{m};text-transform:uppercase;padding:6px 9px;text-align:right;'>Total</td></tr>{r}</table>"
                       "<div style='{f}font-size:10.5px;color:{m};margin-top:5px;'>{fi} firm; firm charges flow into the costing as they're approved.</div>"
                       ).format(f=FN, m=MUT, r=cr, fi=money0(ch_firm))

    people_blk = ""
    if roster:
        people_blk = ("<div style='{f}font-size:13px;color:{t};line-height:1.65;'><b style='color:{w};'>{td:,.0f}h</b> worked to date "
                      "(<b style='color:{w};'>{day:,.0f}h</b> day / <b style='color:{w};'>{night:,.0f}h</b> night), planned to completion "
                      "<b style='color:{o};'>{pl:,.0f}h</b>. Safety: <b style='color:{w};'>{t5}</b> Take 5s, <b style='color:{w};'>{cv}</b> "
                      "safety conversations, <b style='color:{w};'>{ps}</b> pre-starts across {d} day(s). Every shift starts with a "
                      "conversation; each day opens with a pre-start.</div>").format(
                          f=FN, t=TXT, w=W, o=ORGC, td=roster["td_total"], day=roster["td_day"], night=roster["td_night"],
                          pl=roster["pl_total"], t5=roster["take5"], cv=roster["convos"], ps=roster["prestarts"], d=roster["days"])

    team_rows = ""
    for grp, members in (("Day shift", TEAM_DAY), ("Night shift", TEAM_NIGHT)):
        team_rows += "<div style='{f}font-size:10px;color:{o};text-transform:uppercase;font-weight:bold;margin:9px 0 3px;'>{g}</div>".format(f=FN, o=ORGC, g=grp)
        for n, role in members:
            team_rows += "<div style='{f}font-size:12.5px;color:{t};margin:2px 0;'><b style='color:{w};'>{n}</b> &middot; {r}</div>".format(f=FN, t=TXT, w=W, n=html.escape(n), r=html.escape(role))
    team = ("<div style='{f}font-size:12.5px;color:{t};line-height:1.55;margin-bottom:4px;'>The crew running your K2 tool store &mdash; right gear to the right crew, "
            "looked after and ready, checked and right every time. Led by <b style='color:{w};'>Andrew Fisher</b>, Shutdown Manager.</div>{rows}").format(f=FN, t=TXT, w=W, rows=team_rows)

    #  One Coates border around the whole thing - the same frame every
    #  other report email uses. (A. Fisher, 25 Jul 2026)
    return _framed(
        "<table role='presentation' width='680' cellpadding='0' cellspacing='0' align='center' bgcolor='#0E1116' style='border-top:5px solid {o};'>"
        "<tr><td style='padding:18px 22px 6px;'><table role='presentation' width='100%'><tr>"
        "<td><span style='{f}font-size:24px;font-weight:900;color:{o};'>COATES</span> <span style='{f}font-size:12px;color:{w};'>Equipped for anything</span><br>"
        "<span style='{f}font-size:20px;font-weight:bold;color:{w};'>Cost Tracking Snapshot</span><br>"
        "<span style='{f}font-size:10.5px;color:{m};'>Cement Australia &middot; K2 Shutdown 2026 &middot; Gladstone &nbsp;|&nbsp; Reporting period {period} &middot; {days} active days</span></td>"
        "<td align='right' valign='top' style='{f}font-size:10px;font-weight:bold;color:{o};'>POWERED BY SITEIQ<br>"
        "<span style='color:{m};font-weight:normal;'>As at {asat}<br>Author Andrew Fisher</span></td></tr></table></td></tr>"
        "<tr><td style='padding:10px 22px 0;'><table width='100%' cellpadding='0' cellspacing='0'><tr>"
        "<td bgcolor='#0f2a14' style='border-left:4px solid #0ca30c;padding:13px 15px;{f}font-size:13px;color:{t};line-height:1.6;'>"
        "<span style='background:#0ca30c;color:#062b0a;font-weight:bold;font-size:10.5px;padding:3px 10px;'>ON PLAN</span> &nbsp;{banner}</td></tr></table></td></tr>"
        "<tr><td style='padding:12px 22px 0;'>{kpis}</td></tr>"
        "{s_story}{s_stream}{s_plant}{s_pred}{s_charges}{s_people}{s_team}"
        "<tr><td style='padding:18px 22px 16px;'><div style='border-top:1px solid {l};padding-top:10px;{f}font-size:9.5px;color:{m};line-height:1.55;'>"
        "Figures from the K2 workbook Daily Tracking (read-only) at contracted daily rates; unpriced items excluded, never estimated. A client day runs 6:00am to 6:00am. "
        "Point-in-time snapshot, refreshed and shared weekly.<br><span style='color:{o};font-weight:bold;'>Care Deeply &middot; Customer Focused &middot; Be Our Best &middot; One Team &middot; Competitive Spirit</span><br>"
        "Author: Andrew Fisher &nbsp;|&nbsp; POWERED BY <span style='color:{o};font-weight:bold;'>SITEIQ</span></div></td></tr>"
        "</table>"
    ).format(
        f=FN, o=ORGC, w=W, m=MUT, t=TXT, l=LINE, period=period, days=cost["days"], asat=asat,
        banner=("Cost is tracking <b style='color:{w};'>under the plan and under control</b> &mdash; tracking to land at "
                "<b style='color:{w};'>{TR}</b>, under the <b style='color:{w};'>{FULL}</b> plan, and not over it.").format(
                    w=W, TR=money0(tracking), FULL=money0(FULL)),
        kpis=kpis,
        s_story=sect("The story so far", story),
        s_stream=sect("Why we're under &mdash; by cost stream", stream_tbl, "Forecast vs actual to date."),
        s_plant=sect("Plant utilisation", plant, "Idle plant &middot; client insight"),
        s_pred=sect("Predicted full-shutdown cost", predicted),
        s_charges=sect("Client charges", charges_blk),
        s_people=sect("People &amp; assurance", people_blk) if people_blk else "",
        s_team=sect("Meet the tool store team", team))


def main():
    wbpath = find_workbook()
    wb = openpyxl.load_workbook(wbpath, data_only=True)
    ds, ds_path, ds_asat = read_daily_summary()
    if ds:
        print(" SiteIQ invoiced truth: {} ({} day(s), window to {})".format(
            os.path.basename(ds_path), len(ds), ds_asat or "?"))
    else:
        print("  ! No DAILY_SUMMARY export found - hire shows TRACKED figures"
              " only this run (not trued to invoicing). Drop the export in"
              " and run again for the trued view.")
    sub = read_subhire_welders()
    if sub["arrivals"]:
        print(" Welders (invoiced separately): {} unit(s) on site @ {}/day".format(
            len(sub["arrivals"]), money0(sub["rate"])))
    # sub-hired units must never sit in the Plant Equipment register - the
    # register drives the plant plan, and welders carry their own stream
    try:
        _reg_subs = [str(r[4]).strip() for r in
                     list(wb["Plant Equipment"].iter_rows(values_only=True))[1:]
                     if len(r) > 4 and r[4] and str(r[4]).strip().upper().startswith("SUB")]
        if _reg_subs:
            print("  ! {} SUB item(s) found in the Plant Equipment register "
                  "({}). Sub-hired welders are tracked in their own Welders "
                  "stream - remove them from the register so the Plant plan "
                  "stays clean.".format(len(_reg_subs), ", ".join(_reg_subs[:4])))
    except Exception:
        pass
    hs = read_hire_split(wb)
    if hs:
        print(" Hire split (from SiteIQ charge lines): plant {} | tooling {}".format(
            money0(hs["plant"]), money0(hs["tool"])))
    cost = cost_story(wb, ds, sub, hs)
    #  the per-day plant/tooling/gear partition for the true-up page
    #  (Andrew, 30 Jul 2026: "how much for each for each day")
    cost["hire_daily"] = read_hire_daily(wb) or {}
    if cost["hire_daily"]:
        _hd = cost["hire_daily"]
        print(" Daily split: {} day(s) of charge lines | plant {} | "
              "tooling {} | gear {}".format(
                  len(_hd),
                  money0(sum(v["plant"] for v in _hd.values())),
                  money0(sum(v["tool"] for v in _hd.values())),
                  money0(sum(v["gear"] for v in _hd.values()))))
        _missed = [d for d in (ds or {}) if d not in _hd]
        if _missed:
            #  say WHEN the pull happened, off the export's own
            #  REFERENCE_INFO - "your file was too early" lands harder
            #  with the clock time on it (Andrew, 30 Jul 2026: a 06:26
            #  TRANSACTIONS against a 09:33 DAILY_SUMMARY left 29 Jul
            #  blank and looked like missing data)
            _pulled = ""
            try:
                _txf = _gfind("TRANSACTIONS*.xlsx")
                if _txf:
                    _twb = openpyxl.load_workbook(
                        max(_txf, key=os.path.getmtime),
                        read_only=True, data_only=True)
                    if "REFERENCE_INFO" in _twb.sheetnames:
                        _rr = list(_twb["REFERENCE_INFO"]
                                   .iter_rows(values_only=True))
                        if len(_rr) > 1 and _rr[1]:
                            _pulled = str(_rr[1][-1] or "").strip()
                    _twb.close()
            except Exception:
                _pulled = ""
            print("  ! The split is blank for {days}. Your TRANSACTIONS "
                  "export was pulled{at} - BEFORE SiteIQ posted those "
                  "days' charge lines (they post about 09:30 the next "
                  "morning). Download TRANSACTIONS again now, run 28, "
                  "then re-run 02 - the blanks fill in.".format(
                      days=", ".join(d.strftime("%d %b")
                                     for d in sorted(_missed)),
                      at=(" at " + _pulled) if _pulled else " too early"))
    _nolog = [d for d in (ds or {})
              if cost.get("recon") and any(
                  x["d"] == d and x["invoiced"] is not None
                  and x["tracked"] is None for x in cost["recon"])]
    if _nolog:
        print("  ! Tracking not logged on {} - the invoice carries those "
              "days; run 01_RUN_PREFILL_DAILY each morning so the drift "
              "check stays alive.".format(", ".join(
                  d.strftime("%d %b") for d in sorted(_nolog))))
    if not any(s["name"] == "Plant hire" for s in cost["streams"]):
        print("  ! Plant/Tooling split unavailable this run (charge lines "
              "missing or out of step with DAILY_SUMMARY) - showing the "
              "combined hire line, never an apportioned split.")
    bill = coates_billing(wb, ds, hs)     # INTERNAL SiteIQ billing forecast
    roster = read_roster(wb, cost["d1"])
    wb.close()
    # live plant position (reuses the proven plant-dashboard logic; importing it
    # also rebuilds the plant dashboard HTML into today's Reports folder)
    try:
        import build_plant_dashboard as bpd
        ov = bpd.DATA["ov"]
    except Exception as exc:
        print("  ! Could not read the plant position ({}); showing cost story only.".format(exc))
        ov = {"util": 0, "deployed": 0, "parked": 0, "total": 0, "id_daily": 0, "idle_share": 0}
    charges_data = read_charges()
    svc = read_service_invoiced()
    if svc["total"] > 0:
        print(" Progress claims (SERVICE): labour {} | accommodation {}{}".format(
            money0(svc["labour"]), money0(svc["accom"]),
            " | other {}".format(money0(sum(a for _d, a in svc["other"])))
            if svc["other"] else ""))
    page, page_htmls, n_pages = build_html(cost, ov, charges_data, roster,
                                           os.path.basename(wbpath), svc=svc)

    # one folder per day, organised: Pages\ PDF\ Emails\ (A. Fisher)
    dirs = report_paths.out_dirs(HERE)
    outdir = dirs["day"]
    stamp = dt.datetime.now().strftime("%Y-%m-%d")
    out = os.path.join(dirs["pages"], "Coates_K2_Cost_Tracking_Snapshot_{}.html".format(stamp))
    with open(out, "w", encoding="utf-8") as f:
        f.write(page)
    pdf_out = os.path.join(dirs["pdf"], os.path.basename(out).replace(".html", ".pdf"))
    if make_pdf(out, pdf_out):
        print(" PDF: {}".format(pdf_out))
    # Email — the same pages as the PDF, one under the other, as inline images
    # so it opens looking identical in Outlook. Falls back to a reflowed
    # Coates-style email only if no browser is on the machine to draw the pages.
    eml_out = os.path.join(dirs["emails"], "Coates_K2_Cost_Tracking_Snapshot_{}.eml".format(stamp))
    made_email = False
    r_to, r_cc = recipients.resolve(HERE, "", "COST")
    #  THE CEMENT MASTER EMAIL (Andrew, 28 Jul 2026): the snapshot stays
    #  pasted in the body, and the full daily pack rides the paperclip -
    #  exec summary, safety, returns, site plant, consumables, Cement's
    #  own on-hire report and the all-companies summary. The PDFs are
    #  built later in the same 00 run (company reports are step 6, this
    #  is step 5); the draft is created at step 9, after everything
    #  exists - MAKE_OUTLOOK_DRAFTS skips and names anything missing.
    _pack = [os.path.join(dirs["pdf"], n.format(stamp)) for n in (
        "Coates_K2_Executive_Summary_{}.pdf",
        "Coates_K2_Safety_Assurance_{}.pdf",
        "Coates_K2_Returns_Performance_{}.pdf",
        "Coates_SitePlant_Report_{}.pdf",
        "Coates_K2_Consumables_Report_{}.pdf",
        "Coates_OnHire_Report_CEMENT_AUSTRALIA_HOLDINGS_PTY_LTD_{}.pdf",
        "Coates_OnHire_Summary_ALL_COMPANIES_{}.pdf",
    )]
    pngs = render_page_images(page_htmls, dirs["emails"], stamp)
    if pngs:
        try:
            import email_images
            m = image_email(cost, pngs, to=r_to, cc=r_cc)
            email_images.save_eml(m, eml_out)
            cids = ["pg{}".format(i) for i in range(1, len(pngs) + 1)]
            email_images.write_manifest(eml_out, str(m["Subject"]),
                                        email_images.images_body(cids, baked=False), pngs,
                                        report="COST",
                                        attachments=_pack)
            print(" Email ({}-page): {}".format(n_pages, eml_out))
            made_email = True
        except Exception as exc:
            print("  ! page-image email skipped ({})".format(exc))
    if not made_email:
        # fallback: reflowed dark Coates-style email capturing the whole snapshot
        asat_full = dt.datetime.now().strftime("%d %b %Y - %H:%M")
        eb = email_body(cost, ov, charges_data, roster, asat_full)
        try:
            from email.message import EmailMessage
            m = EmailMessage()
            m["Subject"] = "K2 Shutdown - Cost Tracking Snapshot (as at {})".format(cost["d1"].strftime("%d %b"))
            m["From"] = "Andrew Fisher"
            m["To"] = r_to
            if r_cc:
                m["Cc"] = r_cc
            m.set_content("This report is best viewed in HTML. The Cost Tracking Snapshot is in the message body.")
            m.add_alternative(eb, subtype="html")
            with open(eml_out, "w", encoding="utf-8") as f:
                f.write(m.as_string())
            print(" Email (reflow): {}".format(eml_out))
        except Exception as exc:
            print("  ! email .eml skipped ({})".format(exc))

    # ---- COATES INTERNAL: the SiteIQ billing forecast --------------------
    # Its own file, its own email - it never rides inside the client report.
    try:
        bdoc, bpages = billing_doc(
            bill, dt.datetime.now().strftime("%d %b %Y - %H:%M"),
            os.path.basename(wbpath),
            os.path.basename(ds_path) if ds_path else "DAILY_SUMMARY (not found)",
            svc=svc, charges_data=charges_data)
        bstem = "Coates_K2_Billing_Forecast_INTERNAL_{}".format(stamp)
        bhtml = os.path.join(dirs["pages"], bstem + ".html")
        with open(bhtml, "w", encoding="utf-8") as f:
            f.write(bdoc)
        bpdf = os.path.join(dirs["pdf"], bstem + ".pdf")
        if make_pdf(bhtml, bpdf):
            print(" Billing forecast PDF: {}".format(bpdf))
        b_to, b_cc = recipients.resolve(HERE, "", "BILLING")
        bpngs = render_page_images(bpages, dirs["emails"], stamp,
                                   prefix="Coates_K2_Billing_Forecast_INTERNAL")
        if bpngs:
            import email_images
            bsubj = "K2 Shutdown - Coates Billing Forecast (INTERNAL) (as at {})".format(
                cost["d1"].strftime("%d %b"))
            bm = email_images.build_image_eml(bsubj, bpngs, to=b_to, cc=b_cc, baked=False)
            beml = os.path.join(dirs["emails"], bstem + ".eml")
            email_images.save_eml(bm, beml)
            bcids = ["pg{}".format(i) for i in range(1, len(bpngs) + 1)]
            email_images.write_manifest(beml, bsubj,
                                        email_images.images_body(bcids, baked=False), bpngs,
                                        report="BILLING")
            print(" Billing forecast email (INTERNAL, {}-page): {}".format(len(bpngs), beml))
        print(" COATES billing forecast - all SiteIQ-billed streams:")
        _hm = dict(bill["months"])
        _om = dict(bill["others"])
        for _cm in sorted(set(list(_hm.keys()) + list(_om.keys()))):
            _hv = (_hm[_cm]["run"] + _hm[_cm]["ahead"]) if _cm in _hm else 0.0
            _ov = sum(_om.get(_cm, {}).values())
            _fx = bill["fuelcons"].get(_cm, 0.0) + bill["misc"].get(_cm, 0.0)
            print("   {} {}  (hire {} + other streams {})".format(
                _cm.strftime("%B %Y"), money0(_hv + _ov + _fx),
                money0(_hv), money0(_ov + _fx)))
        print("   Total (all in) {}".format(money0(bill["grand"])))
        for x in bill["rolls"]:
            print("   NOTE: {} ({}) does not invoice in {} - registers into {}".format(
                x["d"].strftime("%d %b"), money0(x["v"]),
                x["d"].strftime("%B"), x["to"].split()[0]))
    except Exception as exc:
        print("  ! billing forecast skipped ({})".format(exc))

    # provenance - which files fed this day's reports, and how fresh they were
    report_paths.note_sources(outdir, [
        wbpath,
        report_paths.find_export(HERE, "DAILY_SUMMARY*.xlsx") or os.path.join(HERE, "DAILY_SUMMARY.xlsx"),
        os.path.join(HERE, "Coates_K2_Charge_Reporter", "CHARGE_REGISTER.xlsx"),
        report_paths.find_export(HERE, "ON_HIRE*.xlsx") or os.path.join(HERE, "ON_HIRE.xlsx"),
        report_paths.find_export(HERE, "RENTAL_STOCK*.xlsx") or os.path.join(HERE, "RENTAL_STOCK.xlsx"),
        os.path.join(HERE, "ImportSample_Contracted Rates and Prices.xlsx"),
        os.path.join(HERE, "Plant_ID_Register.csv"),
    ])

    print("=" * 62)
    print(" COATES | K2 COST TRACKING SNAPSHOT  ({} pages, one story each)".format(n_pages))
    print(" Forecast to date {}  |  Actual {}  |  Variance {}".format(
        money0(cost["F"]), money0(cost["A"]), money0(cost["V"])))
    print(" Plan to completion {}  |  tracking to land {}".format(
        money0(cost["FULL"]), money0(cost["FULL"] + cost["V"])))
    print(" Hire trued to SiteIQ: invoiced {} locked | provisional {} ({} day(s))".format(
        money0(cost["invoiced"]), money0(cost["provisional"]), cost["prov_days"]))
    mtot = sum(v["run"] + v["ahead"] for _k, v in cost["months"])
    print(" Month view: " + "  |  ".join(
        "{} {}".format(k, money0(v["run"] + v["ahead"])) for k, v in cost["months"])
        + "  (sums to {})".format(money0(mtot)))
    print(" Today's folder: {}".format(outdir))
    # finish with REAL Outlook drafts - pages in the compose body, addressed,
    # sitting in Drafts ready to send (covers today's dashboard email too)
    import email_images
    email_images.make_native_drafts(HERE)


if __name__ == "__main__":
    main()
