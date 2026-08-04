#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | SUBMIT CHARGES - the 28 Jul 2026 batch
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 28 Jul 2026): "can i get you to submit 4 things for me,
#  fill in the form for me and submit please."
#
#  This puts them through the Charge Reporter's OWN submit() - the same
#  code path as typing them into the form by hand. That matters: the
#  charge IDs come off the register's own counter, the advice HTML/PDF
#  and the Outlook draft are built the usual way, and both the category
#  sheet AND the Costing Feed get written. Editing the spreadsheet
#  directly would have skipped all of it and quietly broken the feed the
#  cost snapshot reads.
#
#  It runs against YOUR register on YOUR machine, so nothing of yours is
#  overwritten by a copy of mine.
#
#  Nothing is emailed. Each charge produces an Outlook DRAFT you read
#  and send yourself, exactly like a form submission.
#
#  Run it twice and it will say so rather than double-charging.
# =====================================================================
import datetime as dt
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
KIT = os.path.join(HERE, "Coates_K2_Charge_Reporter")

#  ---- the batch -----------------------------------------------------
#  Requested by Ben Vandenbroek, approved 28/07/2026 (both).
DATE = "2026-07-28"
COMPANY = "Cement Australia"

CHARGES = [
    {
        "category": "transport",
        "company": COMPANY, "date": DATE, "time": "07:00",
        "raised_by": "Andrew Fisher", "owner": "Andrew Fisher",
        "site_ref": "K2 Shutdown 2026",
        "move_from": "Brisbane",
        "move_to": "Gladstone",
        "load_desc": "Conveyor",
        "direction": "One way",
        "charge_total": "1134.00",
        "status": "Approved",
        "story": "Transport charge for the conveyor, Brisbane to "
                 "Gladstone, for the K2 shutdown. Approved 28/07/2026.",
    },
    {
        "category": "consumables",
        "company": COMPANY, "date": DATE, "time": "09:00",
        "raised_by": "Ben Vandenbroek", "owner": "Andrew Fisher",
        "site_ref": "Ben Vandenbroek",
        "description": "Paramount Safety Bollard Stem",
        "quantity": "24", "unit": "Each",
        "unit_price": "21.30", "charge_total": "511.20",
        "ordered_by": "Ben Vandenbroek", "po_ref": "Ben Vandenbroek",
        "status": "Approved",
        "story": "Requested by Ben Vandenbroek and approved 28/07/2026. "
                 "24 x Paramount Safety Bollard Stem at $21.30 each.",
    },
    {
        "category": "consumables",
        "company": COMPANY, "date": DATE, "time": "09:00",
        "raised_by": "Ben Vandenbroek", "owner": "Andrew Fisher",
        "site_ref": "Ben Vandenbroek",
        "description": "Paramount Bollard Bases 6KG",
        #  $16.24, not the $16.22 first quoted - SiteIQ is the billing
        #  system of record and it is charging $389.76. Two cents a unit,
        #  48c over the line, and it would have sat on the billing
        #  reconciliation as an exception every single run until one side
        #  moved. (Confirmed by Andrew, 28 Jul 2026.)
        "quantity": "24", "unit": "Each",
        "unit_price": "16.24", "charge_total": "389.76",
        "ordered_by": "Ben Vandenbroek", "po_ref": "Ben Vandenbroek",
        "status": "Approved",
        "story": "Requested by Ben Vandenbroek and approved 28/07/2026. "
                 "24 x Paramount Bollard Bases 6KG at $16.24 each.",
    },
    {
        "category": "consumables",
        "company": COMPANY, "date": DATE, "time": "09:00",
        "raised_by": "Ben Vandenbroek", "owner": "Andrew Fisher",
        "site_ref": "Ben Vandenbroek",
        "description": "Squids 3155 Elastic Tool Lanyard and Clamp",
        "quantity": "20", "unit": "Each",
        "unit_price": "15.30", "charge_total": "306.00",
        "ordered_by": "Ben Vandenbroek", "po_ref": "Ben Vandenbroek",
        "status": "Approved",
        "story": "Requested by Ben Vandenbroek and approved 28/07/2026. "
                 "20 x Squids 3155 Elastic Tool Lanyard and Clamp at "
                 "$15.30 each.",
    },
]

#  ---- NOT TO BE CHARGED, FOR NOW --------------------------------------
#  Andrew, 4 Aug 2026, on the welder: "don't charge for it at the moment."
#
#  A decision not to charge is worth as much as a charge, and it is the
#  one that goes missing - it lives in somebody's head until they are on
#  leave and it gets billed by accident. So it is written down here, in
#  the file that raises charges, and this script PRINTS it every single
#  run. Nothing on this list is silent.
#
#  Take the line out when it is time to charge, and nothing here stops
#  you - this is a "not yet", not a "never".
#
#  (what it is, why it is held, who said so and when)
ON_HOLD = [
    ("The welder", "Not to be charged at the moment.",
     "Andrew Fisher, 4 Aug 2026"),
]


#  Prices that moved after the charge was first raised. Keyed by what the
#  register already holds, so a re-run corrects the row in place instead
#  of raising a second charge for the same goods.
SUPERSEDED = {
    "paramount bollard bases 6kg": [389.28],
}


def already_there(C, c):
    """Has this charge been raised before?

    Returns (charge_id, what_to_do): "skip" if the register already holds
    it at this exact figure, "amend" if it holds the SAME GOODS at a
    price we now know is superseded.

    Matching on amount alone was a trap. Correct the bollard price from
    $389.28 to $389.76 and the old row no longer matches, so a re-run
    would raise a SECOND charge for the same 24 bases - the customer
    billed twice, and nothing in the output would say so. Same goods on
    the same day is the same charge, whatever the price says."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(C.REGISTER_PATH, read_only=True,
                                    data_only=True)
    except Exception:
        return None, None
    key = "consumables" if c["category"] == "consumables" else "transport"
    sheet = C.sheet_name(key)
    if sheet not in wb.sheetnames:
        return None, None
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, None
    hdr = {str(v).strip(): i for i, v in enumerate(rows[0]) if v}

    def cell(r, name):
        i = hdr.get(name)
        return r[i] if i is not None and i < len(r) else None

    want_desc = (c.get("description") or c.get("load_desc") or "").strip().lower()
    want_amt = round(float(c["charge_total"]), 2)
    for r in rows[1:]:
        if not cell(r, "Charge ID"):
            continue
        d = cell(r, "Date")
        d = d.date() if isinstance(d, dt.datetime) else d
        if str(d) != c["date"]:
            continue
        got = (str(cell(r, "Consumables supplied")
                   or cell(r, "Load / what was moved") or "").strip().lower())
        try:
            amt = round(float(cell(r, "Total charge") or 0), 2)
        except (TypeError, ValueError):
            amt = None
        if got != want_desc:
            continue
        cid = str(cell(r, "Charge ID"))
        if amt == want_amt:
            return cid, "skip"
        if amt in SUPERSEDED.get(want_desc, []):
            return cid, "amend"
        #  same goods, same day, a price we were never told about -
        #  don't guess. Flag it and let Andrew decide.
        return cid, "conflict:{}".format(amt)
    return None, None


def amend(C, charge_id, c):
    """Correct an existing charge's price in place - the category sheet
    row AND its Costing Feed line, which is what the cost snapshot
    reads. Miss the feed and the two disagree silently."""
    import openpyxl
    wb = openpyxl.load_workbook(C.REGISTER_PATH)
    key = "consumables" if c["category"] == "consumables" else "transport"
    total = float(c["charge_total"])
    stamp = dt.datetime.now()
    touched = []

    def row_of(ws, want_id):
        hdr = {str(x.value).strip(): x.column for x in ws[1] if x.value}
        for r in range(2, ws.max_row + 1):
            i = hdr.get("Charge ID")
            if i and str(ws.cell(row=r, column=i).value or "").strip() == want_id:
                return r, hdr
        return None, hdr

    ws = wb[C.sheet_name(key)]
    r, hdr = row_of(ws, charge_id)
    if r:
        for col, val in (("Unit price", float(c["unit_price"]) if
                          c.get("unit_price") else None),
                         ("Total charge", total),
                         ("Amount to Costing", total),
                         ("Last Updated", stamp)):
            i = hdr.get(col)
            if i and val is not None:
                ws.cell(row=r, column=i).value = val
        touched.append(C.sheet_name(key))

    if "Costing Feed" in wb.sheetnames:
        fs = wb["Costing Feed"]
        r, hdr = row_of(fs, charge_id)
        if r:
            i = hdr.get("Amount")
            if i:
                fs.cell(row=r, column=i).value = total
            touched.append("Costing Feed")

    wb.save(C.REGISTER_PATH)
    return touched


def main():
    print("=" * 70)
    print(" COATES | SUBMIT CHARGES - Cement Australia K2, 28 Jul 2026")
    print("=" * 70)
    if not os.path.isdir(KIT):
        print(" Coates_K2_Charge_Reporter folder isn't next to me.")
        return 1
    sys.path.insert(0, KIT)
    try:
        import charge_reporter as C
    except Exception as e:
        print(" Couldn't load the Charge Reporter: {}".format(e))
        return 1
    C.ensure_dirs()
    C.ensure_register()

    total = sum(float(c["charge_total"]) for c in CHARGES)
    print("")
    print(" About to submit {} charge(s), {} total:".format(
        len(CHARGES), "${:,.2f}".format(total)))
    for c in CHARGES:
        what = c.get("description") or "{} - {} to {}".format(
            c["load_desc"], c["move_from"], c["move_to"])
        qty = "{} x ${}".format(c["quantity"], c["unit_price"]) \
            if c.get("quantity") else ""
        print("   {:<12} {:<46} {:>12}  {}".format(
            c["category"], what[:46],
            "${:,.2f}".format(float(c["charge_total"])), qty))
    print("")

    #  Before the YES, not after it. What is deliberately NOT being
    #  charged belongs in front of you while you are deciding whether
    #  to submit - and it has to survive you cancelling.
    if ON_HOLD:
        print(" ON HOLD - deliberately NOT charged:")
        for what, why, who in ON_HOLD:
            print("   {:<28} {}".format(what, why))
            print("   {:<28} ({})".format("", who))
        print(" Take the line out of ON_HOLD in SUBMIT_CHARGES.py when it")
        print(" is time to charge for it.")
        print("")

    if "--yes" not in sys.argv:
        try:
            if input(" Type YES to submit: ").strip().upper() != "YES":
                print(" Cancelled. Nothing submitted.")
                return 0
        except EOFError:
            print(" Cancelled. Nothing submitted.")
            return 0

    print("")
    done, skipped, fixed, clash = [], [], [], []
    for c in CHARGES:
        what = c.get("description") or c.get("load_desc")
        cid, action = already_there(C, c)
        if action == "skip":
            print("  SKIP   {} - already on the register as {}".format(
                what[:44], cid))
            skipped.append(cid)
            continue
        if action == "amend":
            touched = amend(C, cid, c)
            print("  AMEND  {} - {} corrected to {}".format(
                what[:44], cid, "${:,.2f}".format(float(c["charge_total"]))))
            print("         updated: {}".format(", ".join(touched) or "nothing"))
            fixed.append(cid)
            continue
        if action and action.startswith("conflict"):
            was = action.split(":", 1)[1]
            print("  STOP   {} - {} is already on the register at ${}, and "
                  "I was told ${}.".format(what[:44], cid, was,
                                           c["charge_total"]))
            print("         Not touching it. Tell me which is right.")
            clash.append(cid)
            continue
        try:
            res = C.submit(dict(c))
        except Exception as e:
            print("  FAIL  {} - {}".format(c.get("description")
                                           or c.get("load_desc"), e))
            continue
        done.append(res)
        print("  DONE  {:<10} {:<46} {}".format(
            res["id"], (c.get("description") or c.get("load_desc"))[:46],
            "${:,.2f}".format(float(c["charge_total"]))))
        for k in ("pdf", "eml"):
            if res.get(k):
                print("        {}".format(res[k]))

    print("")
    print("-" * 70)
    if done:
        amt = sum(float(c["charge_total"]) for c, r in zip(CHARGES, CHARGES)
                  if True)
        print(" Submitted : {} charge(s)".format(len(done)))
        print(" Register  : Coates_K2_Charge_Reporter\\CHARGE_REGISTER.xlsx")
        print("             (category sheet + Costing Feed both updated)")
        print(" Documents : Coates_K2_Charge_Reporter\\Output_Charge_Reports\\")
        print("")
        print(" Each one has an Outlook draft waiting. NOTHING HAS BEEN SENT -")
        print(" open the draft, read it, send it yourself.")
        print("")
        print(" Your cost snapshot will pick these up next time it runs.")
    if fixed:
        print(" Corrected : {} charge(s) repriced in place - {}".format(
            len(fixed), ", ".join(fixed)))
        print("             Both the category sheet and the Costing Feed.")
        print("             No second charge was raised for the same goods.")
    if skipped:
        print(" Skipped   : {} already on the register - not charged "
              "twice.".format(len(skipped)))
    if clash:
        print(" !! STOPPED: {} charge(s) are on the register at a price I "
              "wasn't given.".format(len(clash)))
        print("             Nothing was changed for those. Check them.")
    if ON_HOLD:
        print(" On hold   : {} thing(s) deliberately not charged - listed "
              "above.".format(len(ON_HOLD)))
    if not done and not skipped and not fixed:
        print(" Nothing submitted.")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
