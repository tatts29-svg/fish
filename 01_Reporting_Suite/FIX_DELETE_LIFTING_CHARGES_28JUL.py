#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | ONE-OFF FIX - remove the 28 Jul lifting-consumable charges
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 30 Jul 2026): "can we delete these records please i
#  have deleted off siteiq" - the seventeen lifting-consumable charge
#  lines raised 28 Jul 14:00 (CON-20260728-004 through -020) were
#  removed in SiteIQ, so the register must stop carrying them or every
#  report keeps counting $2,785 of charges that no longer exist.
#
#  What this does, in order:
#    1. Backs up CHARGE_REGISTER.xlsx to Updates\ before touching it.
#    2. Deletes rows whose Charge ID is on the list below - from the
#       "Consumables sold" sheet AND the "Costing Feed" sheet. Exact
#       ID match only; nothing else is touched.
#    3. Moves those charges' generated files (report HTML/PDF, drafts)
#       out of Output_Charge_Reports into the same backup folder -
#       moved, never destroyed, like everything in this suite.
#
#  Safe to run twice: a second run finds nothing and says so.
#  This is a one-off tool. Once it has run clean, delete it or let
#  51_TIDY sweep it - it retires itself from the version manifest by
#  not being part of any update after today.
# =====================================================================
import datetime as dt
import glob
import os
import shutil
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
REG = os.path.join(HERE, 'Coates_K2_Charge_Reporter', 'CHARGE_REGISTER.xlsx')
OUT = os.path.join(HERE, 'Coates_K2_Charge_Reporter', 'Output_Charge_Reports')

#  the exact records Andrew deleted in SiteIQ - nothing more, ever
DELETE_IDS = {'CON-20260728-%03d' % n for n in range(4, 21)}


def main():
    print('=' * 62)
    print(' COATES | ONE-OFF FIX - remove the 28 Jul lifting charges')
    print(' {} Charge IDs: CON-20260728-004 .. CON-20260728-020'.format(
        len(DELETE_IDS)))
    print('=' * 62)
    if not os.path.isfile(REG):
        print(' CHARGE_REGISTER.xlsx not found - nothing to do here.')
        print(' Looked in: ' + REG)
        return 1
    try:
        import openpyxl
    except ImportError:
        print(' openpyxl is missing - run 21_CHECK_MY_SETUP first.')
        return 1

    #  ---- 1. backup, before anything ----
    stamp = dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    bdir = os.path.join(HERE, 'Updates',
                        'pre_delete_lifting_28jul_' + stamp)
    os.makedirs(bdir, exist_ok=True)
    shutil.copy2(REG, os.path.join(bdir, 'CHARGE_REGISTER.xlsx'))
    print(' Backup   : ' + os.path.relpath(bdir, HERE))

    #  ---- 2. delete the rows ----
    wb = openpyxl.load_workbook(REG)
    removed = {}
    for sheet, id_col in (('Consumables sold', 1), ('Costing Feed', 2)):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        #  bottom-up so row numbers stay honest while deleting
        gone = 0
        for r in range(ws.max_row, 1, -1):
            v = str(ws.cell(r, id_col).value or '').strip()
            if v in DELETE_IDS:
                ws.delete_rows(r)
                gone += 1
        removed[sheet] = gone
    total = sum(removed.values())
    if total:
        wb.save(REG)
        for sheet, gone in removed.items():
            print(' {:<18} {} row(s) removed'.format(sheet, gone))
    else:
        print(' Nothing to remove - these charges are already gone.')
        print(' (Safe: this tool has either run before, or the register')
        print('  never carried them on this machine.)')

    #  ---- 3. move the generated artefacts aside ----
    moved = 0
    if os.path.isdir(OUT):
        for cid in sorted(DELETE_IDS):
            for f in glob.glob(os.path.join(OUT, '*' + cid + '*')):
                shutil.move(f, os.path.join(bdir, os.path.basename(f)))
                moved += 1
    if moved:
        print(' Artefacts : {} generated file(s) moved to the backup '
              'folder'.format(moved))

    print('')
    if total or moved:
        print(' Done. Now run 00_RUN_EVERYTHING (or 02) so the reports')
        print(' rebuild without these charges.')
    print(' The originals are in the backup folder if this ever needs')
    print(' to be undone.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
