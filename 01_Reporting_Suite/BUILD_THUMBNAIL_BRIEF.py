#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | THE THUMBNAIL BRIEF - every code, and the picture it needs
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Andrew, 5 Aug 2026: "can you give me the full list in excel of
#  [every variant]. then next to it full elite high quality upgrade
#  description of what the thumbnail image should look like. please
#  ensure we have a colomn to refence how this matches with the image
#  against the description too."
#
#  One workbook: every code on the register, the road its picture
#  travels (real photo / house render / supplier re-dress), the full
#  art direction for the shot, the checklist that proves the finished
#  image IS this item and not its neighbour, and a QC column to sign
#  each one off. Re-run it any morning and it rebuilds from the newest
#  exports - the register moves, the brief moves with it.
#
#  Run it:  py BUILD_THUMBNAIL_BRIEF.py
#  Output:  Coates_K2_Thumbnail_Upgrade_Brief.xlsx (next to this file)
# =====================================================================
import os
import re
import datetime as dt

import mygear_thumbs as MT

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'Coates_K2_Thumbnail_Upgrade_Brief.xlsx')

#  THE HOUSE STYLE - one paragraph, word for word the same on every
#  row, because a brief that drifts per item is how a catalogue stops
#  matching itself. This is the look Andrew signed off on the showcase
#  pages (5 Aug): the flogging-spanner card.
STYLE = ("HOUSE STYLE (identical for every picture): studio product "
         "photograph on a near-black charcoal floor and backdrop "
         "(#14181F). One warm orange rim light raking in low from the "
         "left, a soft white top fill, and a gentle pool of orange "
         "glow on the floor under the item. Deep soft shadows, "
         "pin-sharp focus, high detail. The item is centred, filling "
         "about 75% of a square 800x800 frame. Nothing else in shot: "
         "no hands, no people, no props, no added text or watermark, "
         "no price, no other brand's logo - only markings the real "
         "item carries.")

#  Camera angle by the SHAPE of the thing, not its name.
ANG_FLAT = ("Laid flat on the floor at a 30-45 degree diagonal, camera "
            "high at about 40 degrees so the whole length is in frame.")
ANG_STAND = ("Standing upright on the floor, camera at item height, "
             "three-quarter view so two faces show.")
ANG_COIL = ("Coiled in a neat single loop, connectors/ends crossed to "
            "the front so both fittings are visible.")
ANG_HERO = ("Three-quarter hero angle from slightly above, the working "
            "end toward the camera.")


def _sizes(name):
    """Every size-ish marking in the register name - these are the
    markings the finished image has to carry to BE this code."""
    #  Ordered longest-form first, and overlap-aware below: "1 1/16in"
    #  must swallow its own "1/16in", not stand beside it.
    pats = [
        r'\d+\s+\d+/\d+\s*(?:in(?:ch)?\b|")',  # 1 1/16in
        r'\d+/\d+\s*(?:in(?:ch)?\b|")',        # 9/16"
        r'\d+(?:\.\d+)?\s*(?:in(?:ch)?\b|")',  # 1in / 8"
        r'\d+(?:\.\d+)?\s*mm',                # 32mm
        r'\d+(?:\.\d+)?\s*(?:t|tonne)\b',     # 1.5t
        r'\d+(?:\.\d+)?\s*kva',               # 20kVA
        r'\d+(?:\.\d+)?\s*(?:kg|lb)\b',       # 45kg
        r'\d+\s*v\b',                         # 18V
        r'\d+\s*a\b',                         # 500A
        r'\d+\s*m\s*drop',                    # 10M Drop
        r'\d+(?:\.\d+)?\s*m\b',               # 250M
        r'\d+\s*(?:ft|lm)\b',                 # 19ft / 4400lm
    ]
    found, spans, low = [], [], name.lower()
    for p in pats:
        for m in re.finditer(p, low):
            if any(m.start() < e and m.end() > s0 for s0, e in spans):
                continue
            spans.append((m.start(), m.end()))
            t = ' '.join(m.group(0).split())
            if t not in found:
                found.append(t)
    return found


#  (match-words, angle, scene line, item-proof checklist)
#  First match wins - order matters: "socket set" before "socket".
RULES = [
    (('socket set',), ANG_FLAT,
     'The full set open in its blow-mould case, sockets in size order.',
     'Open case with every socket seated; drive size correct; '
     'sizes ascend left to right'),
    (('socket',), ANG_STAND,
     'One single socket standing on its drive end, hex opening toward '
     'the camera, top face angled to catch the rim light.',
     'ONE socket only; six-point black impact finish; correct drive '
     'square; hex opening visible'),
    (('spanner - flogging', 'spanner flogging', 'slogging'), ANG_FLAT,
     'One flogging spanner, ring end to the upper left, the stubby '
     'striking anvil catching the orange light.',
     'Ring end PLUS square striking anvil (no open jaw); '
     'one spanner only'),
    (('ratchet ring',), ANG_FLAT,
     'One ratchet ring spanner, ratchet head to the front.',
     'Ratchet mechanism visible in the ring end; both stated sizes '
     'apply to its two ends'),
    (('spanner - combo', 'spanner combo', 'combination spanner'),
     ANG_FLAT,
     'One combination spanner, open jaw to the lower left, ring end '
     'upper right.',
     'Ring end one side AND open jaw the other; one spanner only'),
    (('podger', 'podge bar'), ANG_FLAT,
     'One podger bar, the tapered alignment spike filling the lower '
     'half of the frame.',
     'Long tapered spike is unmistakable; spanner/open end at the '
     'other tip'),
    (('spanner - open', 'open end', 'open - end'), ANG_FLAT,
     'One open-end spanner, jaws toward the camera.',
     'Open jaws BOTH ends or as named; no ring end unless named'),
    (('shifter', 'adjustable wrench'), ANG_FLAT,
     'One adjustable shifter, jaws part-open toward the camera, '
     'adjuster worm catching the light.',
     'Adjustable jaw and worm screw visible; stated length looks '
     'proportionate'),
    (('spanner',), ANG_FLAT,
     'One spanner laid diagonally, working ends both visible.',
     'Spanner type matches the name exactly; one spanner only'),
    (('stilsen', 'stillson', 'pipe wrench'), ANG_FLAT,
     'One pipe wrench, serrated jaws toward the camera, part open.',
     'Serrated pipe jaws; adjuster nut visible; aluminium body if '
     'the name says so'),
    (('chain block',), ANG_STAND,
     'The chain block hanging from its top hook against the dark, '
     'load chain falling to a neat pool of chain on the floor.',
     'Hand-chain wheel visible; BOTH hooks in frame; air motor '
     'visible if the name says Air'),
    (('lever hoist', 'cumalong', 'come along'), ANG_HERO,
     'The lever hoist at three-quarter angle, lever arm toward the '
     'camera, load chain draped forward.',
     'Ratchet lever prominent; both hooks visible; rated tonnage tag '
     'legible'),
    (('tirfor', 'winch',), ANG_HERO,
     'The winch body at three-quarter angle, wire rope entering and '
     'leaving, operating handle socket visible.',
     'Wire-rope winch body (not chain); tonnage marking legible'),
    (('lifting bag', 'bull bag',), ANG_STAND,
     'The lifting bag upright and filled square, lifting loops up.',
     'All lifting loops visible; rated load tag legible'),
    (('crane', 'cage',), ANG_HERO,
     'The cage at three-quarter angle so depth reads, lifting points '
     'to the top.',
     'All four lifting points visible; open structure reads as a cage'),
    (('eye bolt', 'eyebolt',), ANG_STAND,
     'One eye bolt standing on its thread collar, eye to the camera.',
     'Forged eye and thread both visible; WLL stamp toward camera'),
    (('shackle',), ANG_STAND,
     'One shackle standing on its bow, pin toward the camera.',
     'Pin and bow both visible; rated stamp toward camera'),
    (('sling', 'ratchet strap', 'tie down'), ANG_COIL,
     'Neatly coiled, both end fittings crossed to the front.',
     'Both end fittings visible; rated tag legible and flat'),
    (('hydraulic cylinder', 'flat jack'), ANG_STAND,
     'The cylinder upright on the floor, saddle up, coupler toward '
     'the camera.',
     'Ram saddle and hydraulic coupler both visible; tonnage marking '
     'legible'),
    (('hydraulic pump',), ANG_HERO,
     'The pump at three-quarter angle, gauge face angled to camera, '
     'hose coiled once beside it.',
     'Gauge, reservoir and coupler all visible'),
    (('torque wrench - cassette', 'hydraulic torque'), ANG_HERO,
     'The hydraulic torque wrench head at three-quarter angle, drive '
     'or cassette opening toward the camera, twin couplers up.',
     'Cassette/link size legible; twin swivel couplers visible'),
    (('torque wrench',), ANG_FLAT,
     'The torque wrench laid diagonally, scale window toward the '
     'camera.',
     'Adjustment scale visible; drive size correct'),
    (('breaker', 'demolition'), ANG_HERO,
     'The breaker leaning working-tip down, handles toward camera.',
     'Chisel/point fitted; power type (electric/air) reads correctly'),
    (('drill magnetic', 'magnetic base drill', 'mag base'), ANG_HERO,
     'The mag-base drill upright on its magnet, motor and feed '
     'handles toward the camera.',
     'Magnet base, column and cutter visible; power type correct'),
    (('drill',), ANG_HERO,
     'The drill standing on its battery or rested on its side handle, '
     'chuck toward the camera.',
     'Chuck toward camera; battery fitted if cordless'),
    (('grinder',), ANG_HERO,
     'The grinder at three-quarter angle, guard and disc toward the '
     'camera.',
     'Guard fitted; disc size reads right; power type correct'),
    (('holesaw', 'hole saw'), ANG_STAND,
     'The holesaw standing cutting-teeth up, arbor beside it.',
     'Tooth ring sharp in the light; diameter stamp legible'),
    (('bolt cutter',), ANG_FLAT,
     'Bolt cutters part open, jaws toward the camera.',
     'Compound jaws visible; stated length proportionate'),
    (('g clamp', 'g-clamp'), ANG_STAND,
     'One G clamp standing upright, screw part-wound.',
     'Frame and screw both visible; throat size proportionate'),
    (('pipe clamp',), ANG_HERO,
     'The pipe clamp at three-quarter angle, chain or yoke open.',
     'Clamping range hardware visible; stainless finish if named'),
    (('welder',), ANG_HERO,
     'The welder set at three-quarter angle on the dark floor, '
     'control panel angled to the camera, leads coiled once in '
     'front.',
     'Control panel legible; amperage class reads right; diesel '
     'units show their frame/tank'),
    (('welding lead', 'welding hose'), ANG_COIL,
     'Neatly coiled, both end fittings crossed to the front.',
     'End fittings visible; lead/hose type reads correctly'),
    (('electrode oven',), ANG_STAND,
     'The oven upright, door ajar just enough to read as an oven.',
     'Door, latch and capacity read correctly'),
    (('electrode', 'welding rod'), ANG_FLAT,
     'The sealed packet laid diagonal with three rods fanned in '
     'front.',
     'Packet label legible; rod type/size readable'),
    (('generator',), ANG_HERO,
     'The generator set at three-quarter angle, control panel toward '
     'the camera.',
     'kVA class reads right; frame and panel visible'),
    (('distribution board', 'distro'), ANG_STAND,
     'The board upright, outlet face square to the camera.',
     'Every outlet and breaker visible; rating legible'),
    (('lead - extension', 'extension lead', 'lead 3 phase', 'lead ',),
     ANG_COIL,
     'The lead in a neat coil, both plugs crossed to the front.',
     'Both plug ends visible; phase/amp rating reads correctly'),
    (('lighting tower', 'light tower'), ANG_HERO,
     'The tower on its trailer, mast up, all heads lit warm against '
     'the dark.',
     'Mast extended; light heads ON; trailer visible'),
    (('light', 'lamp'), ANG_HERO,
     'The light standing lit, beam catching haze faintly above.',
     'Lamp visibly ON; mount/stand type matches the name'),
    (('exhaust fan', 'ventilation fan', ' fan'), ANG_HERO,
     'The fan face-on at slight angle so the cage depth reads, duct '
     'collar visible if it has one.',
     'Blade cage sharp; diameter reads right; power type correct'),
    (('ducting',), ANG_COIL,
     'The flexible duct in a short S-curve, one open end toward the '
     'camera.',
     'Diameter reads right; spiral flex structure visible'),
    (('radio',), ANG_STAND,
     'The handheld radio upright, screen and keys toward the camera, '
     'antenna up.',
     'Screen face legible; antenna fitted; belt clip visible'),
    (('charger',), ANG_STAND,
     'The charger face-on at slight angle, bay(s) empty and visible.',
     'Charge bays and indicator lights visible'),
    (('battery',), ANG_STAND,
     'The battery pack standing upright, contacts toward the camera, '
     'charge-gauge face visible.',
     'Voltage class legible; contact end visible'),
    (('gas monitor', 'gas detector'), ANG_STAND,
     'The monitor upright, screen lit toward the camera.',
     'Screen visibly lit; sensor ports visible; clip visible'),
    (('harness',), ANG_STAND,
     'The harness hung open on an invisible support so the strap '
     'layout reads like worn.',
     'Dorsal D-ring visible; leg loops hang correctly'),
    (('rope',), ANG_COIL,
     'The rope in a neat coil, one tail crossed to the front.',
     'Diameter reads right; fibre type reads correctly'),
    (('ladder', 'trestle'), ANG_HERO,
     'The ladder part-open at three-quarter angle.',
     'Rung count proportionate to stated size; feet visible'),
    (('scissor lift', 'boom lift', 'ewp'), ANG_HERO,
     'The machine at three-quarter angle, platform slightly raised.',
     'Platform, rails and controls visible; electric class reads '
     'right'),
    (('pump',), ANG_HERO,
     'The pump at three-quarter angle, inlet and outlet toward the '
     'camera.',
     'Ports visible; power type correct'),
    (('file - ', 'hacksaw', 'screwdriver', 'plier', 'multigrip',
      'vice grip', 'vise grip', 'allen key', 'hex key', 'punch',
      'chisel', 'trowel', 'knife'), ANG_FLAT,
     'The tool laid diagonally on the dark floor, working end toward '
     'the camera.',
     'Tool type unmistakable; working end sharp in the light'),
    (('vernier', 'caliper', 'dial indicator', 'micrometer', 'scriber',
      'base magnetic', 'steel rule', 'tape measure', 'tape - measure',
      'spirit level'), ANG_FLAT,
     'The instrument laid diagonal on the dark floor, scale or dial '
     'toward the camera.',
     'Scale/dial legible; measuring faces visible'),
    (('hammer',), ANG_FLAT,
     'One hammer laid diagonal, head to the upper left catching the '
     'light.',
     'Head material reads right (copper glows warm); handle type '
     'matches the name'),
    (('crow bar', 'crowbar', 'pinch bar', 'wrecking'), ANG_FLAT,
     'The bar laid full-length diagonal, working tip toward the '
     'camera.',
     'Working tip shape correct; stated length proportionate'),
]

#  Consumables: the picture is the PRODUCT AS HANDED OVER THE COUNTER.
CONS_SCENE = ('The item exactly as handed over the counter - packet, '
              'tin, pair or spool - upright, label face square to the '
              'camera.')
CONS_CHECK = ('Label brand/size legible; quantity form matches how the '
              'store issues it (pair/pack/roll/tin)')


def classify(name):
    low = ' ' + ' '.join(str(name).lower().split()) + ' '
    for keys, ang, scene, check in RULES:
        if any(k in low for k in keys):
            return ang, scene, check
    return (ANG_HERO,
            'The item at a three-quarter hero angle on the dark floor, '
            'working side toward the camera.',
            'Item type unmistakably matches the register name')


def road_for(code, entry):
    """Which road this picture travels - matched to what it is FOR."""
    if entry.get('k') == 'cons':
        return ('3 - SUPPLIER RE-DRESS',
                'Supplier pack-shot through the house dresser: '
                'background off, dark floor on, orange glow added.')
    low = ' ' + entry['n'].lower() + ' '
    real = ('welder', 'generator', 'light', 'tower', 'scissor', 'boom',
            'radio', 'gas monitor', 'gas detector', 'chain block',
            'lever hoist', 'cumalong', 'tirfor', 'winch', 'hydraulic',
            'torque wrench', 'drill magnetic', 'breaker', 'compressor',
            'pump', 'fan', 'charger', 'battery', 'harness',
            'distribution board', 'electrode oven', 'ladder',
            'trestle', 'sling', 'shackle', 'lifting', 'crane')
    if any(k in low for k in real):
        return ('1 - REAL PHOTO',
                'Our actual unit at the 30-second counter booth as it '
                'crosses on demob: black mat, one lamp, scan the '
                'barcode and the phone names the file by this code.')
    return ('2 - HOUSE RENDER',
            'Generated in the locked house style, one image per '
            'description family; spot-check against the shelf before '
            'it ships.')


def brief_for(entry):
    #  A consumable's picture is the product AS HANDED OVER THE
    #  COUNTER - packet, tin, pair or spool - not a hero pose.
    if entry.get('k') == 'cons':
        ang = ('Camera straight on at label height, the item turned '
               'a few degrees so it still has depth.')
        scene, check = CONS_SCENE, CONS_CHECK
    else:
        ang, scene, check = classify(entry['n'])
    sizes = _sizes(entry['n'])
    marks = ('THE MARKINGS THAT MUST BE READABLE: ' + ' | '.join(sizes)
             + '.') if sizes else ''
    item = ('THE ITEM: {n}. {scene} {ang} {marks}').format(
        n=entry['n'].strip(), scene=scene, ang=ang, marks=marks).strip()
    checklist = check
    if sizes:
        checklist += ' ; stamped/labelled sizes readable: ' + \
            ', '.join(sizes)
    checklist += (' ; house style holds (dark floor, orange rim glow, '
                  'item fills ~75%, no text/people/price)')
    return item + ' ' + STYLE, checklist


def build():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, \
        Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    reg = MT.variant_register(HERE)
    tdir = os.path.join(HERE, 'Gear_Lookup', 'thumbs')
    have = set()
    if os.path.isdir(tdir):
        codes = {MT.safe_name(str(c).strip().upper()): c for c in reg}
        for f in os.listdir(tdir):
            if f.lower().endswith('.jpg') and f[:-4].upper() in codes:
                have.add(codes[f[:-4].upper()])

    rows = []
    for code in sorted(reg):
        e = reg[code]
        road, road_note = road_for(code, e)
        brief, checklist = brief_for(e)
        rows.append((road, e.get('f') or '', code, e['n'],
                     'Consumable' if e.get('k') == 'cons' else 'Hire',
                     e.get('q') or 0,
                     'YES' if code in have else '',
                     brief, checklist, '', road_note))
    rows.sort(key=lambda r: (r[0], r[1], r[3].upper()))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Thumbnail Brief'
    ORANGE, DARK = 'F26222', '1D1D1B'
    thin = Border(bottom=Side(style='thin', color='DDDDDD'))

    ws.merge_cells('A1:K1')
    c = ws.cell(1, 1, 'COATES K2 | THE THUMBNAIL UPGRADE BRIEF - every '
                'code on the register, the picture it needs, and the '
                'proof it got it')
    c.font = Font(bold=True, size=13, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=DARK)
    ws.merge_cells('A2:K2')
    c = ws.cell(2, 1, 'Built {} from the newest SiteIQ exports - '
                '{:,} codes. Re-run BUILD_THUMBNAIL_BRIEF.py any '
                'morning and this rebuilds itself. Work a ROAD at a '
                'time; sign each picture off in the IMAGE MATCHES '
                'column only after every line of the checklist '
                'passes.'.format(
                    dt.date.today().strftime('%d %b %Y'), len(rows)))
    c.font = Font(size=9, color='555555')
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[2].height = 30

    heads = ['Road', 'SiteIQ Family', 'Variant Code', 'Register Name',
             'Kind', 'Qty', 'Has Photo Today',
             'THE THUMBNAIL BRIEF - full art direction for this image',
             'MATCH CHECKLIST - what the finished image must show to '
             'BE this code', 'IMAGE MATCHES (Y / N / REDO)',
             'How this road works']
    for i, h in enumerate(heads, 1):
        c = ws.cell(4, i, h)
        c.font = Font(bold=True, size=9, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=ORANGE)
        c.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[4].height = 30

    for r, row in enumerate(rows, 5):
        for i, v in enumerate(row, 1):
            c = ws.cell(r, i, v)
            c.font = Font(size=9)
            c.border = thin
            c.alignment = Alignment(wrap_text=(i in (4, 8, 9, 11)),
                                    vertical='top')
        if row[6] == 'YES':
            ws.cell(r, 7).font = Font(size=9, bold=True,
                                      color='1F7A44')

    widths = [16, 22, 22, 38, 11, 6, 9, 72, 52, 13, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = 'A4:K{}'.format(4 + len(rows))
    dv = DataValidation(type='list', formula1='"Y,N,REDO"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add('J5:J{}'.format(4 + len(rows)))

    #  Sheet 2: the house style and the roads, once, in full - so the
    #  workbook can be handed to a photographer or an image tool with
    #  no other document beside it.
    g = wb.create_sheet('House Style & Roads')
    g.column_dimensions['A'].width = 110
    lines = [
        ('COATES K2 | THE HOUSE STYLE - ONE LOOK FOR EVERY PICTURE',
         True),
        ('', False), (STYLE, False), ('', False),
        ('THE THREE ROADS', True),
        ('ROAD 1 - REAL PHOTO: plant, gas monitors, radios, welders, '
         'hydraulics, lifting gear and everything else that carries a '
         'serial or a tag. Our actual unit, photographed at the '
         '30-second counter booth as it crosses on demob returns - '
         'black mat, one lamp on a floor mark, scan the barcode and '
         'the phone names the file by the code.', False),
        ('ROAD 2 - HOUSE RENDER: commodity tooling (sockets, spanners, '
         'hand tools). One generated image per description family in '
         'the locked house style - a 46mm and a 50mm socket differ by '
         'an engraving, and the card already prints the size. '
         'Spot-check every render against the shelf before it ships.',
         False),
        ('ROAD 3 - SUPPLIER RE-DRESS: consumables. The supplier '
         'pack-shot through the house dresser - background off, dark '
         'floor on, orange glow added. No photography at all.', False),
        ('', False),
        ('THE FILE LAW', True),
        ('800x800 JPG, 80KB or less, named thumbs\\<VARIANT CODE>.jpg '
         '(Windows-unsafe characters swapped to _ exactly as the '
         'pages already do). The moment a file lands, My Gear, the '
         'category grid, the reports and the print-outs all pick it '
         'up on the next build - no other change.', False),
        ('', False),
        ('THE RULES THAT RIDE ALONG', True),
        ('A catalogue picture is never the asset\'s condition photo - '
         'damage evidence stays its own photo, filed against the '
         'asset. The tag colour band on every card keeps coming from '
         'the register, never from the picture. No picture carries a '
         'price, a rate, or another company\'s name. A rendered '
         'picture must match the type and size it claims - the MATCH '
         'CHECKLIST column is the sign-off, and the IMAGE MATCHES '
         'column is who signed it.', False),
    ]
    for i, (t, bold) in enumerate(lines, 1):
        c = g.cell(i, 1, t)
        c.font = Font(bold=bold, size=10 if bold else 9,
                      color=(ORANGE if bold else '222222'))
        c.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(OUT)
    return len(rows), sum(1 for r in rows if r[6] == 'YES'), OUT


if __name__ == '__main__':
    n, have, path = build()
    print('=' * 66)
    print(' COATES | THUMBNAIL UPGRADE BRIEF')
    print('=' * 66)
    print(' {:,} codes briefed | {} already have a photo'.format(
        n, have))
    print(' Written: ' + path)
