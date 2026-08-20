#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | HVAC & POWER CATALOG PAGE - what we can get, one page
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Renders HVAC_POWER_EQUIPMENT_CATALOG.xlsx (the register - see
#  equipment_catalog.py) as ONE printable branded page:
#  HVAC_POWER_CATALOG.html, beside the suite like HOW_IT_WORKS.html.
#
#  What the page answers at the counter or on the phone to the BDM:
#    * what HVAC and power gear the fleet catalog carries
#    * the exact MODEL code to order it by (= SiteIQ PRODUCT_VARIANT,
#      = Product Variant ID on the contracted-rates import)
#    * the pricing group it bills under (a billing family, NOT a rate -
#      no dollars on this page, so it can sit on the counter)
#    * which of it is already on K2 - read live from the newest
#      RENTAL_STOCK export, stamped with that export's own as-at
#
#  Manners: the register missing is a dead button, so that failure
#  SHOUTS in a band and exits 1. The rental stock missing only costs
#  the "on K2 now" column - the page still builds and says in plain
#  English which export was absent. No JavaScript on the page; print
#  is A4 through the browser. Safe to run twice - it just rebuilds.
# =====================================================================

import datetime as dt
import os
import re
import sys

import equipment_catalog

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "HVAC_POWER_CATALOG.html")


def esc(v):
    return (str(v).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;"))


def au_date(d):
    """11 Jul 2026 style, 24-hour when there's a time."""
    if d.hour or d.minute:
        return d.strftime("%d %b %Y %H:%M").lstrip("0")
    return d.strftime("%d %b %Y").lstrip("0")


# ---------------------------------------------------------------------
#  Rental stock - live "on K2 now" counts by PRODUCT_VARIANT
# ---------------------------------------------------------------------

def find_newest(pattern):
    hits = []
    for d in (os.path.join(HERE, "Data_SiteIQ"), HERE):
        if not os.path.isdir(d):
            continue
        for fn in os.listdir(d):
            if fn.startswith("~$"):
                continue
            if re.match(pattern, fn, re.I):
                p = os.path.join(d, fn)
                if os.path.isfile(p):
                    hits.append(p)
    return max(hits, key=os.path.getmtime) if hits else None


def _parse_au(s):
    """SiteIQ's own stamp - '29/07/2026 05:42 AM', day first, always."""
    s = str(s).strip()
    for fmt in ("%d/%m/%Y %I:%M %p", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def load_rental_stock():
    """(counts, asat_line, note). counts maps PRODUCT_VARIANT ->
    {status: n}. Absent or unreadable export -> empty counts and a
    plain-English note; the page builds regardless."""
    path = find_newest(r"RENTAL_STOCK.*\.xlsx$")
    if not path:
        return {}, "", ("The RENTAL_STOCK export wasn't in Data_SiteIQ\\ "
                        "or the suite folder, so the ON K2 NOW column "
                        "shows dashes. Pull it (button 28) and run 57 "
                        "again for live counts.")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        asat = None
        if "REFERENCE_INFO" in wb.sheetnames:
            info = list(wb["REFERENCE_INFO"].iter_rows(values_only=True))
            if len(info) >= 2:
                hdr = [str(c or "").strip().upper() for c in info[0]]
                for i, h in enumerate(hdr):
                    if "REQUESTED_DATE" in h and i < len(info[1]):
                        asat = _parse_au(info[1][i])

        ws = (wb["RENTAL_STOCK"] if "RENTAL_STOCK" in wb.sheetnames
              else wb[wb.sheetnames[0]])
        rows = ws.iter_rows(values_only=True)
        hdr = [str(c or "").strip().upper() for c in next(rows)]
        try:
            i_v = hdr.index("PRODUCT_VARIANT")
            i_s = hdr.index("ITEM_STATUS")
        except ValueError:
            wb.close()
            return {}, "", ("The RENTAL_STOCK export ({}) had no "
                            "PRODUCT_VARIANT column, so the ON K2 NOW "
                            "column shows dashes."
                            .format(os.path.basename(path)))
        counts = {}
        for row in rows:
            v = str(row[i_v]).strip().upper() if len(row) > i_v and row[i_v] else ""
            s = str(row[i_s]).strip() if len(row) > i_s and row[i_s] else ""
            if v:
                counts.setdefault(v, {})
                counts[v][s] = counts[v].get(s, 0) + 1
        wb.close()
        if asat:
            asat_line = "{} export, data as at {}".format(
                os.path.basename(path), au_date(asat))
        else:
            asat_line = "{} export, file time {}".format(
                os.path.basename(path),
                au_date(dt.datetime.fromtimestamp(os.path.getmtime(path))))
        return counts, asat_line, ""
    except Exception as e:
        return {}, "", ("Couldn't read the RENTAL_STOCK export ({}), so "
                        "the ON K2 NOW column shows dashes.".format(e))


_STATUS_WORDS = (
    ("On Hire", "on hire"),
    ("Available for Hire", "in the store"),
    ("Awaiting Arrival", "arriving"),
)


def onsite_phrase(status_counts):
    """{'On Hire': 2} -> '2 on hire'. Anything unexpected is counted
    honestly rather than dropped."""
    if not status_counts:
        return ""
    bits = []
    left = dict(status_counts)
    for raw, word in _STATUS_WORDS:
        if left.get(raw):
            bits.append("{} {}".format(left.pop(raw), word))
    other = sum(left.values())
    if other:
        bits.append("{} other".format(other))
    return " + ".join(bits)


# ---------------------------------------------------------------------
#  The page
# ---------------------------------------------------------------------

STYLE = """
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Calibri,'Segoe UI',Arial,sans-serif;background:#fff;color:#20242B;
     font-size:11.5pt;line-height:1.5;max-width:200mm;margin:0 auto;padding:10mm 8mm}
.frame{border:6px solid #F26222;border-top:9px solid #F26222;border-radius:18px;padding:22px 26px}
.band{background:#14181F;color:#fff;padding:14px 18px;border-radius:10px;margin-bottom:14px}
.band .k{color:#F26222;font-weight:800;letter-spacing:2px;font-size:9pt;text-transform:uppercase}
.band h1{font-size:17pt;margin-top:2px}
.band p{color:#C9CFD8;font-size:10pt;margin-top:4px}
h2{font-size:12pt;color:#14181F;background:#F6F7F9;border:1px solid #E3E6EB;
   border-left:5px solid #F26222;border-radius:9px;padding:6px 12px;margin:16px 0 6px}
h2 .tc{color:#5B6472;font-weight:400;font-size:9.5pt}
.note{background:#FFF3EC;border-left:4px solid #F26222;border-radius:0 8px 8px 0;
      padding:10px 14px;font-size:10.5pt;margin:10px 0}
.note b.o{color:#F26222}
table{border-collapse:collapse;width:100%;margin:6px 0 4px;font-size:10pt}
td,th{border:1px solid #E3E6EB;padding:5px 9px;text-align:left;vertical-align:top}
th{background:#14181F;color:#fff;font-size:8.5pt;letter-spacing:1px;text-transform:uppercase;white-space:nowrap}
td.m{font-family:Consolas,'Courier New',monospace;font-weight:700;white-space:nowrap}
td.g b{white-space:nowrap}
td.now{white-space:nowrap}
td.now b{color:#F26222}
tr.share td{background:#FDF6EF}
.small{color:#5B6472;font-size:9pt;margin:2px 0 8px}
.ft{color:#5B6472;font-size:9pt;border-top:1px solid #E3E6EB;margin-top:16px;padding-top:8px}
b.o{color:#F26222}
.tw{overflow-x:auto}
@media (max-width:480px){body{font-size:10.5pt;padding:4mm}
  .frame{padding:12px 10px;border-width:4px;border-top-width:6px}}
@media print{body{padding:0}.frame{border-radius:0}.tw{overflow-x:visible}
  h2{page-break-after:avoid}tr{page-break-inside:avoid}
  thead{display:table-header-group}
  .ft,.note{page-break-inside:avoid}}
@page{size:A4;margin:10mm}
"""


def type_groups(cat):
    """[(category, type_desc, type_code, [recs sorted small->large])] in
    the register's own order of first appearance."""
    order = []
    groups = {}
    for r in cat.models():
        key = (r["category"], r["type_desc"], r["type_code"])
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(r)

    def size_key(rec):
        c = equipment_catalog.capacity(rec["model"])
        return (0, c[0], rec["model"]) if c else (1, 0, rec["model"])

    return [(k[0], k[1], k[2], sorted(groups[k], key=size_key))
            for k in order]


def build():
    print("=" * 70)
    print(" COATES | HVAC & POWER CATALOG PAGE")
    print("=" * 70)

    cat = equipment_catalog.load()
    if not cat.loaded:
        print("!" * 70)
        print("! THE CATALOG REGISTER IS MISSING - PAGE NOT BUILT")
        print("! Looked for HVAC_POWER_EQUIPMENT_CATALOG*.xlsx beside the")
        print("! suite and in Data_SiteIQ\\. Put the register back (or")
        print("! restore it from Updates\\backups) and run 57 again.")
        print("!" * 70)
        return 1

    stock, stock_asat, stock_note = load_rental_stock()
    if stock_asat:
        print("  Rental stock     : {}".format(stock_asat))
    else:
        print("  Rental stock     : (not read - on-site column will show "
              "dashes)")

    groups = type_groups(cat)
    n_models = len(cat.models())
    cat_counts = []
    for code in cat.categories():
        inside = ["{} {}".format(len(rs), tdesc)
                  for c, tdesc, tcode, rs in groups if c == code]
        n = sum(len(rs) for c, _t, _tc, rs in groups if c == code)
        cat_counts.append("{} {} ({})".format(n, code, ", ".join(inside)))

    # What's already on the job, catalog models only
    onsite_bits = []
    n_onsite = 0
    for _c, _t, _tc, recs in groups:
        for r in recs:
            ph = onsite_phrase(stock.get(r["model"], {}))
            if ph:
                n_onsite += sum(stock.get(r["model"], {}).values())
                onsite_bits.append("{} ({})".format(r["desc"], ph))

    # Catalog-family variants on site that this extract does NOT cover
    fam = re.compile(r"^(ACPAC|CHILL|AHU\d|GENERATOR)")
    outside = []
    for v in sorted(stock):
        if fam.match(v) and not cat.rec(v):
            outside.append("{} ×{}".format(v, sum(stock[v].values())))

    now = dt.datetime.now()
    h = []
    h.append("<!DOCTYPE html><html><head><meta charset='utf-8'>"
             "<meta name='viewport' "
             "content='width=device-width,initial-scale=1'>"
             "<title>HVAC &amp; Power Catalog - K2</title>"
             "<style>" + STYLE + "</style></head><body>")
    h.append("<div class='frame'>")
    h.append("<div class='band'><div class='k'>COATES &middot; HVAC &amp; "
             "POWER EQUIPMENT CATALOG</div>"
             "<h1>What we can get, what it's called, what it bills under</h1>"
             "<p>Cement Australia K2 Shutdown 2026 &middot; Gladstone "
             "&middot; Author: Andrew Fisher &middot; POWERED BY SITEIQ</p>"
             "</div>")

    # The position, first
    pos = ("<div class='note'><b class='o'>The position.</b> {n} active "
           "models in the fleet catalog: {cats}.").format(
               n=n_models, cats=esc("; ".join(cat_counts)))
    if stock:
        if onsite_bits:
            pos += (" Already on K2: {}.".format(esc("; ".join(onsite_bits))))
        else:
            pos += " None of it is on K2 yet."
    if stock_note:
        pos += " {}".format(esc(stock_note))
    pos += "</div>"
    h.append(pos)

    h.append("<p class='small'>Order by the <b>MODEL</b> code - it's the "
             "same code SiteIQ calls PRODUCT_VARIANT in every export and "
             "the contracted-rates import calls Product Variant ID. The "
             "pricing group is the billing family a model charges under - "
             "a code, not a rate; there are no dollars on this page. "
             "Listed smallest to largest.</p>")

    for code in cat.categories():
        for c, tdesc, tcode, recs in groups:
            if c != code:
                continue
            h.append(("<h2>{cat} &middot; {t} <span class='tc'>type {tc} "
                      "&middot; {n} models</span></h2>").format(
                          cat=esc(c), t=esc(tdesc), tc=esc(tcode),
                          n=len(recs)))
            shared = cat.shared_groups()
            rows_html = []
            rows_html.append("<div class='tw'><table>"
                             "<thead><tr><th>Model</th><th>What it is</th>"
                             "<th>On K2 now</th>"
                             "<th>Pricing group</th></tr></thead><tbody>")
            for r in recs:
                ph = onsite_phrase(stock.get(r["model"], {}))
                now_cell = "<b>{}</b>".format(esc(ph)) if ph else "&ndash;"
                cls = " class='share'" if r["group"] in shared else ""
                rows_html.append(
                    ("<tr{cls}><td class='m'>{m}</td><td>{d}</td>"
                     "<td class='now'>{now}</td>"
                     "<td class='g'><b>{g}</b> &middot; {gd}</td>"
                     "</tr>").format(
                        cls=cls, m=esc(r["model"]), d=esc(r["desc"]),
                        g=esc(r["group"]), gd=esc(r["group_desc"]),
                        now=now_cell))
            rows_html.append("</tbody></table></div>")
            h.append("".join(rows_html))
            for gcode, ms in sorted(cat.shared_groups().items()):
                if any(r["group"] == gcode for r in recs):
                    h.append(("<p class='small'>Pricing group {g} carries "
                              "{n} models ({ms}) - one billing family, "
                              "exactly as the extract arrived.</p>").format(
                                  g=esc(gcode), n=len(ms),
                                  ms=esc(", ".join(sorted(ms)))))

    if outside:
        h.append(("<div class='note'><b class='o'>Also on site.</b> The "
                  "rental stock carries catalog-family gear this extract "
                  "doesn't cover (other type codes - mostly the smaller "
                  "gensets): {}.</div>").format(esc(", ".join(outside))))

    ft = []
    ft.append("Catalog register: {} (file time {}).".format(
        esc(os.path.basename(cat.path)), au_date(cat.mtime)))
    if stock_asat:
        ft.append("On-site counts: {}.".format(esc(stock_asat)))
    ft.append("Built {} &middot; 57_HVAC_POWER_CATALOG.bat rebuilds this "
              "page &middot; add or fix a model in the register and run "
              "it again.".format(au_date(now)))
    ft.append("Author: Andrew Fisher &middot; POWERED BY SITEIQ")
    h.append("<div class='ft'>{}</div>".format("<br>".join(ft)))
    h.append("</div></body></html>")

    with open(OUT, "w", encoding="utf-8") as f:
        f.write("".join(h))

    print("  Catalog page     : {} models across {} type groups".format(
        n_models, len(groups)))
    if stock:
        print("  On K2 now        : {} catalog item(s){}".format(
            n_onsite, "" if n_onsite else " - none of the catalog is on "
            "the job yet"))
    print("")
    print("  Written : {}".format(os.path.basename(OUT)))
    print("  Print   : open it and Ctrl+P - it lays out for A4.")
    return 0


if __name__ == "__main__":
    sys.exit(build())
