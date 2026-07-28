#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | K2 EXCEL ALIGNMENT CHECK - the proof, cell by cell
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Read-only. Opens the K2 workbook and recomputes every headline figure
#  from the workbook's OWN tables with the same rules the report suite
#  uses, then compares against what the workbook cells actually show.
#  Run it after 13_ALIGN_EXCEL_DATA (00_RUN_EVERYTHING does both):
#
#      ALL LINED UP  -> the Excel and the reports are telling one story
#      difference    -> the exact cell, the live figure and the shown
#                       figure are printed - nothing drifts silently
#
#  Also checks the master file (K2_MASTER_EQUIPMENT_PRICING.xlsx) is
#  flowing: on-hire tooling prices in the Excel v the master, by item.
# =====================================================================

import os
import sys
import glob
import datetime as dt
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
TOL = 0.011


def find_newest(pattern):
    hits = []
    for d in (os.path.join(HERE, "Data_SiteIQ"), HERE):
        hits += [p for p in glob.glob(os.path.join(d, pattern))
                 if not os.path.basename(p).startswith("~$")]
    return max(hits, key=os.path.getmtime) if hits else None


def num(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def to_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, (int, float)):
        try:
            return (dt.datetime(1899, 12, 30) + dt.timedelta(days=float(v))).date()
        except Exception:
            return None
    return None


class Book(object):
    def __init__(self, path):
        self.path = path
        self.wb = openpyxl.load_workbook(path, data_only=True)  # READ ONLY
        self.tables = {}
        for wsname in self.wb.sheetnames:
            ws = self.wb[wsname]
            try:
                for tname, ref in (ws.tables or {}).items():
                    self.tables[tname] = (wsname, ref.ref if hasattr(ref, "ref") else str(ref))
            except Exception:
                pass

    def table(self, name):
        if name not in self.tables:
            return []
        wsname, ref = self.tables[name]
        ws = self.wb[wsname]
        min_c, min_r, max_c, max_r = openpyxl.utils.range_boundaries(ref)
        hdr = [("" if ws.cell(row=min_r, column=c).value is None
                else str(ws.cell(row=min_r, column=c).value).strip())
               for c in range(min_c, max_c + 1)]
        rows = []
        for r in range(min_r + 1, max_r + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(min_c, max_c + 1)]
            if any(v not in (None, "") for v in vals):
                rows.append(dict(zip(hdr, vals)))
        return rows

    def cell(self, sheet, addr):
        try:
            return self.wb[sheet][addr].value
        except Exception:
            return None


def days_on(oh, today):
    d = to_date(oh)
    return (today - d).days + 1 if d else 0


def main():
    today = dt.date.today()
    wb_path = (glob.glob(os.path.join(HERE, "Cement_Australia_Report*K2*.xlsm"))
               or [None])
    wb_path = max(wb_path, key=lambda p: os.path.getmtime(p) if p else 0)
    if not wb_path:
        print("  ! No K2 workbook found beside this script.")
        sys.exit(1)
    bk = Book(wb_path)
    print("  K2 EXCEL ALIGNMENT CHECK - {}".format(os.path.basename(wb_path)))
    print("  " + "-" * 66)

    rows = []

    def chk(label, live, shown):
        try:
            ok = abs(float(live) - float(shown)) < TOL
        except Exception:
            ok = str(live) == str(shown)
        rows.append((label, live, shown, ok))

    # ---- registers, recomputed off the workbook's own tables -------------
    tool = bk.table("Tooling_Onhire__4")
    milw = bk.table("RENTAL_STOCK__3")
    rad_all = bk.table("Radios_Onhire")
    handsets = [r for r in rad_all
                if str(r.get("Asset Type ID") or "") != "RADIO_BATTERY"]
    batts = [r for r in rad_all
             if str(r.get("Asset Type ID") or "") == "RADIO_BATTERY"]
    gas = bk.table("Gas_Monitors_Onhire")

    KD = lambda a: bk.cell("K2 Dashboard", a)
    DD = lambda a: bk.cell("Dashboard Data", a)

    # discipline grid (rows 17-21: tooling, milwaukee, radios, batteries, gas)
    for rix, (name, reg, replcol) in zip(
            (17, 18, 19, 20, 21),
            [("General Tooling", tool, "Replacement Cost"),
             ("Milwaukee", milw, "Replacement Cost"),
             ("Radios", handsets, "Replacement Cost"),
             ("Radio Batteries", batts, None),
             ("Gas Monitors", gas, "Replacement Fee")]):
        chk("{} items out".format(name), len(reg), KD("B{}".format(rix)))
        expo = sum(num(r.get(replcol)) for r in reg) if replcol else 0.0
        chk("{} exposure".format(name), expo, KD("C{}".format(rix)))
        ds = [days_on(r.get("On Hire Date"), today) for r in reg]
        chk("{} overdue 2d".format(name), sum(1 for d in ds if d == 2),
            KD("D{}".format(rix)))
        chk("{} high 3d".format(name), sum(1 for d in ds if d == 3),
            KD("E{}".format(rix)))
        chk("{} critical 4d+".format(name), sum(1 for d in ds if d >= 4),
            KD("F{}".format(rix)))

    # replacement exposure tile (tooling + milwaukee + radios + gas)
    expo_all = (sum(num(r.get("Replacement Cost")) for r in tool)
                + sum(num(r.get("Replacement Cost")) for r in milw)
                + sum(num(r.get("Replacement Cost")) for r in handsets)
                + sum(num(r.get("Replacement Fee")) for r in gas))
    chk("Replacement exposure tile", expo_all, DD("H11"))

    # critical returns tile
    crit = sum(1 for reg in (tool, milw, handsets, gas)
               for r in reg if days_on(r.get("On Hire Date"), today) >= 4)
    chk("Critical returns 4d+ tile", crit, DD("H10"))

    # ---- financial control ------------------------------------------------
    est = sum(num(r.get("Est Total")) for r in bk.table("K2EstimateTable"))
    reg_fc = sum(num(r.get("Forecast Hire Value"))
                 for r in bk.table("AssetRegisterTable2219"))
    # the tracking table's range carries its TOTAL row - real dates only
    track = [r for r in bk.table("DailyTrackingTable")
             if to_date(r.get("Date")) is not None]
    radio_fc = sum(num(r.get("Radio Hire F/C")) for r in track)
    gas_fc = sum(num(r.get("Gas Monitor Hire F/C")) for r in track)
    lab_fc = sum(num(r.get("Personnel / Labour F/C")) for r in track)
    acc_fc = sum(num(r.get("Accommodation F/C")) for r in track)
    tra_fc = sum(num(r.get("Transport F/C")) for r in track)
    dmg_fc = sum(num(r.get("Damage Recovery F/C")) for r in track)
    reconciled = round(reg_fc + est + radio_fc + gas_fc + lab_fc + acc_fc
                       + tra_fc + dmg_fc, 2)
    tracker = round(sum(num(r.get("Daily Forecast Total")) for r in track), 2)
    chk("Planned plant forecast", reg_fc, DD("H14"))
    chk("Detailed tooling forecast", est, DD("H24"))
    chk("Reconciled forecast", reconciled, DD("H31"))
    chk("Daily Tracking forecast", tracker, DD("H15"))
    chk("Model gap", round(reconciled - tracker, 2), DD("H32"))

    # ---- dashboard radio rows (post-align these are live values) ---------
    chk("Radios issued (dashboard row)", len(handsets), DD("K10"))
    chk("Radio batteries issued (dashboard row)", len(batts), DD("K11"))
    bulk = bk.table("BulkEquipmentTable")
    batt_fleet = sum(num(b.get("Qty On Site")) for b in bulk
                     if str(b.get("Equipment Type") or "").startswith("Radio Battery"))
    chk("Radio battery fleet", batt_fleet, DD("N11"))
    radio_daily = sum(num(b.get("Daily Fleet Charge")) for b in bulk
                      if str(b.get("Equipment Type") or "").startswith("Two-Way"))
    chk("Radios daily value (fleet charge)", radio_daily, DD("L10"))

    # ---- refresh stamp + error count -------------------------------------
    stamp = to_date(DD("H1"))
    fresh = bool(stamp and stamp >= today - dt.timedelta(days=1))
    rows.append(("Refresh stamp current",
                 stamp.strftime("%d %b") if stamp else "-",
                 "today" if fresh else "STALE", fresh))
    live_err = num(bk.cell("Controls & Exceptions", "D13"))
    chk("Query error count (config v live detector)",
        live_err, bk.cell("Control Config", "B20"))

    # ---- accountability: live score v table rows -------------------------
    comps = {}
    for reg, replcol in [(tool, "Replacement Cost"), (milw, "Replacement Cost"),
                         (handsets, "Replacement Cost"), (batts, None),
                         (gas, "Replacement Fee")]:
        for r in reg:
            comp = str(r.get("Company") or "").strip() or "(no company)"
            c = comps.setdefault(comp, {"days": [], "expo": 0.0})
            c["days"].append(days_on(r.get("On Hire Date"), today))
            if replcol:
                c["expo"] += num(r.get(replcol))
    ca_rows = {str(r.get("Company") or "").strip(): r
               for r in bk.table("CompanyAccountabilityTable")}
    chk("Accountability companies", len(comps), len(ca_rows))
    for comp, c in sorted(comps.items()):
        t = ca_rows.get(comp)
        if t is None:
            rows.append(("Accountability row: " + comp, "live", "MISSING", False))
            continue
        chk("{} items".format(comp), len(c["days"]), num(t.get("Total Items")))
        chk("{} exposure".format(comp), c["expo"],
            num(t.get("Replacement Exposure")))

    # ---- master file flow into the Excel ---------------------------------
    mpath = find_newest("K2_MASTER_EQUIPMENT_PRICING*.xlsx")
    spath = find_newest("RENTAL_STOCK*.xlsx")
    if mpath:
        mwb = openpyxl.load_workbook(mpath, read_only=True, data_only=True)
        mws = mwb[mwb.sheetnames[0]]
        it = mws.iter_rows(values_only=True)
        mh = [str(c or "").strip().upper() for c in next(it)]
        gi = {n: (mh.index(n) if n in mh else None) for n in
              ("ITEM_NUMBER", "REPLACEMENT_COST_AUD", "NEW_DESCRIPTION")}
        master = {}
        for r in it:
            k = str(r[gi["ITEM_NUMBER"]] or "").strip() if gi["ITEM_NUMBER"] is not None else ""
            if k:
                master[k] = {
                    "price": r[gi["REPLACEMENT_COST_AUD"]] if gi["REPLACEMENT_COST_AUD"] is not None else None,
                    "new": str(r[gi["NEW_DESCRIPTION"]] or "").strip() if gi["NEW_DESCRIPTION"] is not None else "",
                }
        mwb.close()
        bc2item = {}
        if spath:
            swb = openpyxl.load_workbook(spath, read_only=True, data_only=True)
            sws = swb["RENTAL_STOCK"] if "RENTAL_STOCK" in swb.sheetnames else swb[swb.sheetnames[0]]
            sit = sws.iter_rows(values_only=True)
            sh = [str(c or "").strip().upper() for c in next(sit)]
            if "ITEM_BARCODE" in sh and "ITEM_NUMBER" in sh:
                bi, ii = sh.index("ITEM_BARCODE"), sh.index("ITEM_NUMBER")
                for r in sit:
                    b = str(r[bi] or "").strip().upper()
                    i = str(r[ii] or "").strip()
                    if b and i and b not in bc2item:
                        bc2item[b] = i
            swb.close()
        n_check = n_match = 0
        for r in tool:
            bc = str(r.get("Barcode") or "").strip()
            item = bc2item.get(bc.upper()) or (bc if bc in master else None)
            if not item or item not in master:
                continue
            mp = master[item]["price"]
            if mp in (None, ""):
                continue
            n_check += 1
            if abs(num(r.get("Replacement Cost")) - float(mp)) < TOL:
                n_match += 1
        rows.append(("Master prices flowing into tooling tab",
                     "{} of {} rows priced from master".format(n_match, n_check),
                     "all" if n_check == 0 or n_match == n_check else "partial",
                     n_check == 0 or n_match == n_check))

    # ---- report ----------------------------------------------------------
    good = [r for r in rows if r[3]]
    bad = [r for r in rows if not r[3]]
    print("  {} of {} checks lined up.".format(len(good), len(rows)))
    if bad:
        print("  NOT ALIGNED - live figure v what the Excel shows:")
        for label, live, shown, _ok in bad:
            fl = ("{:,.2f}".format(live) if isinstance(live, (int, float)) else live)
            fs = ("{:,.2f}".format(shown) if isinstance(shown, (int, float))
                  else ("(blank)" if shown is None else shown))
            print("    - {:44s} live {}  v  shown {}".format(str(label)[:44], fl, fs))
        print("  Run 13_ALIGN_EXCEL_DATA (after a save) to true these up.")
        sys.exit(1)
    print("  ALL LINED UP - the Excel and the reports are telling one story.")
    sys.exit(0)


if __name__ == "__main__":
    main()
