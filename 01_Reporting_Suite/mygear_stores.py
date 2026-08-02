#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | STORES TEAM PAGE - the counter's own view
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 29 Jul 2026): "can Coates have a direct sign in - opens
#  up stores staff page. all the product groups, how many available,
#  how many on hire, company and name, where abouts, what storage unit
#  ... a section of stock take and a percentage of what has been stock
#  taked and what has not ... on hire list longer than 3 days, a
#  section you look through to ensure its not in the store, done by
#  storage unit so your being directed where to go."
#
#  This is a DIFFERENT JOB from the crew page. A crew member asks "what
#  have I got and what can I grab". The stores team asks "what is
#  missing, who has it, and where do I walk to find it". So this page
#  is built around the hunt, not the browse.
#
#  ON THE SIGN-IN, HONESTLY. The page is a static file on the store's
#  own Wi-Fi - there is no server to check a password against. So the
#  payload is ENCRYPTED with the passcode, using the same scheme that
#  already protects every crew member's card: without the code the file
#  is meaningless noise, and the code is never written into the page.
#  That is real protection against someone browsing the Wi-Fi, and it
#  is NOT enterprise sign-on - a determined person with the file and
#  time could work at it. For a list of who is holding which grinder on
#  a shutdown, already printed on the client's daily reports, that is
#  the proportionate answer. Say so plainly rather than call it secure.
#
#  THE SITE PLANT SPLIT. 509 items are out longer than three days, but
#  294 of them are barriers, rubbish chutes and site plant that live on
#  site by design - they are not lost, and putting them in a hunt list
#  turns it into 60% noise that nobody reads twice. Tools and plant are
#  therefore counted apart: the chase list is the gear that should
#  plausibly be back on a shelf.
# =====================================================================
import base64
import datetime as dt
import io
import json
import os
import re

#  The permanent do-not-show list (Andrew, 2 Aug 2026: "can we permently
#  remove these"). Optional, so a suite without the module still builds.
try:
    import hidden_stock as _HID
except Exception:
    _HID = None

#  Days from Flame Off. Optional the same way - a suite without the
#  module simply shows no day tag.
try:
    import shutdown_day as _SD
except Exception:
    _SD = None


def _day_tag(d=None):
    """The header chip, whole. A named day wears the orange; an ordinary
    one stays quiet steel so the named ones actually stand out. Returns
    an empty string if the module is missing, so the header just loses
    the chip rather than the page losing the header."""
    if _SD is None:
        return ''
    try:
        n = _SD.day(d)
        m = _SD.milestone(d)
        bits = '<b>DAY {}</b>'.format(n)
        if m:
            bits += ' &middot; ' + m
        else:
            nx = _SD.next_milestone(d)
            #  only shout about what is close enough to plan around
            if nx and nx['away'] <= 7:
                bits += ' &middot; {} in {} day{}'.format(
                    nx['name'], nx['away'], '' if nx['away'] == 1 else 's')
        return '<div class="fday{}">{}</div>'.format(' big' if m else '', bits)
    except Exception:
        return ''

M32 = 0xFFFFFFFF


def _imul(a, b):
    return (a * b) & M32


def _xmur3(s):
    h = 1779033703 ^ len(s)
    for ch in s:
        h = _imul(h ^ ord(ch), 3432918353)
        h = ((h << 13) | (h >> 19)) & M32

    def call():
        nonlocal h
        h = _imul(h ^ (h >> 16), 2246822507)
        h = _imul(h ^ (h >> 13), 3266489909)
        h ^= h >> 16
        return h & M32
    return call


def _mulberry32(a):
    #  Copied verbatim from BUILD_MY_GEAR's proven pair - the Python and
    #  the JavaScript have to produce the identical stream or the page
    #  will not open, and "close enough" fails silently.
    a &= M32

    def call():
        nonlocal a
        a = (a + 0x6D2B79F5) & M32
        t = _imul(a ^ (a >> 15), (a | 1) & M32)
        t = ((t + _imul(t ^ (t >> 7), (t | 61) & M32)) & M32) ^ t
        t &= M32
        return (t ^ (t >> 14)) & M32
    return call


def enc(code, s):
    """Same scheme as the crew cards - XOR against a seeded stream.
    r() >> 24 is the integer form of the page's Math.floor(rnd()*256)."""
    r = _mulberry32(_xmur3(code + '|CoatesK2stores2026')())
    return base64.b64encode(
        bytes((ord(c) ^ (r() >> 24)) & 0xFF for c in s)).decode('ascii')


def tag(code):
    return format(_xmur3(code + '|CoatesK2storestag2026')(), 'x')


def au_date(v):
    """SiteIQ writes these as TEXT ("28/07/2026"), not dates - the third
    field family in this export set to do it. Read as a date they come
    back as nothing, and every ageing figure silently reads zero, which
    is a confident wrong answer. (29 Jul 2026)"""
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    m = re.match(r'\s*(\d{1,2})/(\d{1,2})/(\d{4})', str(v or ''))
    if not m:
        return None
    try:
        return dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None


def au_dt(v):
    """Date AND time. STOCKTAKE writes "25/07/2026 08:47 AM", so a count
    can be put on the shift that actually did it."""
    sv = str(v or '').strip()
    m = re.match(r'\s*(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{2})\s*([AaPp][Mm])?', sv)
    if not m:
        d = au_date(v)
        return (d, None) if d else (None, None)
    try:
        day = dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None, None
    h, mi = int(m.group(4)), int(m.group(5))
    ap = (m.group(6) or '').lower()
    if ap == 'pm' and h != 12:
        h += 12
    elif ap == 'am' and h == 12:
        h = 0
    return day, (h, mi)


#  Storage units that live on site by design - out for days is normal,
#  not a thing to chase round the store.
PLANT_UNITS = ('site plant', 'barrier', 'chute', 'laydown')


def _is_plant(unit):
    u = (unit or '').lower()
    return any(k in u for k in PLANT_UNITS)


def _hm(v):
    m = re.match(r'\s*(\d{1,2}):(\d{2})', str(v or ''))
    return (int(m.group(1)), int(m.group(2))) if m else None


def _shift_of(d, hm):
    """Which shift-day a movement belongs to.

    Nights run 18:00 to 06:00, so they straddle midnight: 02:00 on the
    14th is still the NIGHT OF THE 13th and has to score for the crew
    that was actually standing there. Andrew's own example, 29 Jul 2026:
    "they do 18:00 to 6:00 so 13/07/2026 into 14/07/2026".
    """
    if not d or not hm:
        return None, None
    h = hm[0]
    if h >= 18:
        return d, 'N'
    if h < 6:
        return d - dt.timedelta(days=1), 'N'
    return d, 'D'


def _night_ran(hm):
    """Was a night crew really on?

    Andrew, 29 Jul 2026: "if there is no transactions after 18:30
    through to 5:00 generally means no nightshift." The SHIFT runs
    18:00-06:00, but the DETECTION window is deliberately tighter - a
    day-shift straggler scanning at 18:05, or an early bird at 05:40,
    must not conjure a night crew that was never there.
    """
    if not hm:
        return False
    h, m = hm
    return (h > 18 or (h == 18 and m >= 30)) or h < 5


def _battle(txn_path, stocktake_path=None):
    """Day versus night, every shift-day of the shut.

    An issue is counted at its start, a return at its end - so a tool
    taken on days and dropped back at night scores one for each crew,
    which is exactly right: both did a piece of work.

    FAIRNESS. Nights did not start until part-way through the shut, and
    scoring the empty nights as day-shift wins would rig the ladder and
    the night crew would dismiss the whole board in one look. So a
    shift-day only counts as a contest once BOTH crews are running; the
    earlier days are shown, and plainly marked as days-only.
    """
    import openpyxl
    if not os.path.isfile(txn_path):
        return None
    wb = openpyxl.load_workbook(txn_path, read_only=True, data_only=True)
    if 'TRANSACTION_CHARGES' not in wb.sheetnames:
        wb.close()
        return None
    rows = list(wb['TRANSACTION_CHARGES'].iter_rows(values_only=True))
    wb.close()
    if not rows:
        return None
    hdr = [str(c or '').strip() for c in rows[0]]
    ix = {h: i for i, h in enumerate(hdr) if h}
    need = ('TRAN_START_DATE', 'TRAN_START_TIME')
    if not all(k in ix for k in need):
        return None

    tally = {}

    nights_on = {}
    counters = {'D': {}, 'N': {}}   # name -> stocktake sightings, by shift

    def slot(d):
        return tally.setdefault(d, {'D': {'i': 0, 'r': 0, 's': 0},
                                    'N': {'i': 0, 'r': 0, 's': 0}})
    for r in rows[1:]:
        if not r or not any(c not in (None, '') for c in r):
            continue
        _t = _hm(r[ix['TRAN_START_TIME']])
        d, sh = _shift_of(au_date(r[ix['TRAN_START_DATE']]), _t)
        if d:
            slot(d)[sh]['i'] += 1
            if _night_ran(_t):
                nights_on[d] = True
        if 'TRAN_END_DATE' in ix and 'TRAN_END_TIME' in ix:
            _t2 = _hm(r[ix['TRAN_END_TIME']])
            d2, sh2 = _shift_of(au_date(r[ix['TRAN_END_DATE']]), _t2)
            if d2:
                slot(d2)[sh2]['r'] += 1
                if _night_ran(_t2):
                    nights_on[d2] = True

    #  stocktake counts, scored onto the same shift-days
    if stocktake_path and os.path.isfile(stocktake_path):
        wb2 = openpyxl.load_workbook(stocktake_path, read_only=True,
                                     data_only=True)
        ws2 = wb2['STOCKTAKE'] if 'STOCKTAKE' in wb2.sheetnames else wb2.active
        srows = list(ws2.iter_rows(values_only=True))
        wb2.close()
        if srows:
            sh_ = [str(c or '').strip() for c in srows[0]]
            si = {h: i for i, h in enumerate(sh_) if h}
            if 'LAST_SIGHTED_DATE_TIME' in si:
                #  WHO is doing the counting, by shift - the power bar
                #  means nothing if nobody is named (Andrew, 30 Jul
                #  2026: "who what shift is doing what")
                for r in srows[1:]:
                    if not r:
                        continue
                    d0, hm = au_dt(r[si['LAST_SIGHTED_DATE_TIME']])
                    if not d0:
                        continue
                    dd, ss = _shift_of(d0, hm or (12, 0))
                    if dd:
                        slot(dd)[ss]['s'] += 1
                        if _night_ran(hm):
                            nights_on[dd] = True
                        _by = (str(r[si['LAST_SIGHTED_BY']] or '').strip()
                               if 'LAST_SIGHTED_BY' in si else '')
                        if _by:
                            counters[ss][_by] = counters[ss].get(_by, 0) + 1

    days = sorted(tally)
    if not days:
        return None
    out, dw, nw, tie = [], 0, 0, 0
    for d in days:
        v = tally[d]
        D = v['D']['i'] + v['D']['r']
        N = v['N']['i'] + v['N']['r']
        #  a day is only a contest if a night crew actually worked it
        contest = bool(nights_on.get(d))
        w = ''
        if contest:
            if D > N:
                w = 'D'
                dw += 1
            elif N > D:
                w = 'N'
                nw += 1
            else:
                w = 'T'
                tie += 1
        out.append({'d': d.strftime('%a %d %b'), 'iso': d.isoformat(),
                    'di': v['D']['i'], 'dr': v['D']['r'],
                    'ni': v['N']['i'], 'nr': v['N']['r'],
                    'ds': v['D']['s'], 'ns': v['N']['s'],
                    'D': D, 'N': N, 'w': w, 'c': contest})
    return {
        'days': out,
        'dayTotal': sum(x['D'] for x in out),
        'nightTotal': sum(x['N'] for x in out),
        'dayWins': dw, 'nightWins': nw, 'ties': tie,
        'dayStock': sum(x['ds'] for x in out),
        'nightStock': sum(x['ns'] for x in out),
        'nightDays': sum(1 for x in out if x['c']),
        'whoD': sorted(([w, n] for w, n in counters['D'].items()),
                       key=lambda x: -x[1])[:6],
        'whoN': sorted(([w, n] for w, n in counters['N'].items()),
                       key=lambda x: -x[1])[:6],
    }


def _money(v):
    try:
        return float(str(v).strip().replace('$', '').replace(',', ''))
    except (TypeError, ValueError):
        return None


def _pricing(onhire_path, master=None):
    """What the gear on hire is costing per day, for the manager view.

    The point is a manager glancing down a column and spotting the line
    that is wrong - so the odd ones are surfaced, not buried: anything on
    hire at a zero rate (earning nothing), and the dearest lines.
    Everything here is SiteIQ's own SHIFT_RATE; nothing is calculated up
    or estimated. (Andrew, 29 Jul 2026: "see visually if anything does
    not look correct or wrong")
    """
    import openpyxl
    if not onhire_path or not os.path.isfile(onhire_path):
        return None
    import mygear_store as MS
    wb = openpyxl.load_workbook(onhire_path, read_only=True, data_only=True)
    ws = wb['ON_HIRE'] if 'ON_HIRE' in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return None
    hdr = [str(c or '').strip() for c in rows[0]]
    ix = {h: i for i, h in enumerate(hdr) if h}
    if 'SHIFT_RATE' not in ix:
        return None

    def g(r, k):
        return str(r[ix[k]] or '').strip() if k in ix else ''

    cats, firms, items, zero = {}, {}, {}, []
    total = 0.0
    for r in rows[1:]:
        if not r or not any(c not in (None, '') for c in r):
            continue
        rate = _money(r[ix['SHIFT_RATE']])
        desc = g(r, 'ITEM_DESCRIPTION')
        if not desc:
            continue
        name = MS._tidy(desc, master, g(r, 'ITEM_NUMBER'))
        co = g(r, 'COMPANY') or 'Not named'
        cat = MS._cat_of(name)
        if rate is None or rate <= 0:
            zero.append({'n': name, 'co': co, 'w': g(r, 'HIRER_NAME')})
            rate = 0.0
        total += rate
        cats[cat] = cats.get(cat, 0.0) + rate
        firms[co] = firms.get(co, 0.0) + rate
        e = items.setdefault(name, {'n': name, 'q': 0, 'r': rate, 'v': 0.0})
        e['q'] += 1
        e['v'] += rate
        e['r'] = max(e['r'], rate)

    top = sorted(items.values(), key=lambda x: -x['v'])[:60]
    return {
        'perDay': round(total, 2),
        'cats': sorted(([k, round(v, 2)] for k, v in cats.items()),
                       key=lambda t: -t[1]),
        'firms': sorted(([k, round(v, 2)] for k, v in firms.items()),
                        key=lambda t: -t[1]),
        'top': [{'n': x['n'], 'q': x['q'], 'r': round(x['r'], 2),
                 'v': round(x['v'], 2)} for x in top],
        'zero': zero[:80],
        'zeroN': len(zero),
        'week': round(total * 7, 2),
    }


#  The phone-keypad alias. The crew page's ID box brings up the
#  number-only keypad on a phone - right for 900 workers typing digit
#  IDs, and a wall for a stores code made of letters (caught 29 Jul
#  2026: "no option to enter letters"). Every letter code therefore
#  gets a numeric twin spelled on the phone keypad - NOIS is 6647 -
#  so the same door opens from a numeric keypad. The alias never
#  appears in the page; it works exactly like the manager code does,
#  by decrypting the real code out of a key blob.
_KEYPAD = {c: d for d, letters in
           (('2', 'ABC'), ('3', 'DEF'), ('4', 'GHI'), ('5', 'JKL'),
            ('6', 'MNO'), ('7', 'PQRS'), ('8', 'TUV'), ('9', 'WXYZ'))
           for c in letters}


def keypad_alias(code):
    """The numeric twin of a letter code, or '' when there isn't one
    (already all digits, too short, or carries characters with no
    keypad home)."""
    code = (code or '').upper().strip()
    if not code or code.isdigit() or len(code) < 3:
        return ''
    out = []
    for ch in code:
        if ch.isdigit():
            out.append(ch)
        elif ch in _KEYPAD:
            out.append(_KEYPAD[ch])
        else:
            return ''
    return ''.join(out)


def _plant_id(item):
    """The orange Plant ID for an asset, via the compliance module the
    rest of the suite already uses. BUILD_MY_GEAR binds the master file
    before calling read(), so the lookup is free; run standalone with
    nothing bound, it degrades to blank rather than breaking."""
    try:
        import equipment_compliance as _EC
        return _EC.plant_id(item) or ''
    except Exception:
        return ''


def _comp_fl(item, name):
    """Compact compliance letters for a row: E electrical tag, R rigging
    / height tag (both BLUE for Jul-Aug per the master), L logbook -
    generators, forklifts, booms and the like that need the daily
    pre-start written up (Andrew, 31 Jul 2026: "have like a log book so
    they know it's a piece of equipment that needs the log book filled
    in"). Same authority as the worker card: equipment_compliance."""
    try:
        import equipment_compliance as _EC
        f = _EC.flags(item, name)
    except Exception:
        return ''
    s = ''
    if f.get('electrical'):
        s += 'E'
    if f.get('rigging'):
        s += 'R'
    if f.get('logbook'):
        s += 'L'
    return s


def _tag_now():
    try:
        import equipment_compliance as _EC
        return {'c': _EC.tag_colour()[0] or '', 'x': _EC.tag_hex()}
    except Exception:
        return {'c': '', 'x': '#8A97A8'}


def read(rental_path, stocktake_path, master=None, today=None,
         txn_path=None, sales_path=None, base=None):
    """Everything the counter needs, from the two registers.

    sales_path/base bring in the consumables shelf. They are optional so
    a site with no consumables export still gets a board rather than an
    error - the pane simply does not appear.
    """
    import openpyxl
    today = today or dt.date.today()

    def sheet(path, name):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[name] if name in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c or '').strip() for c in rows[0]]
        ix = {h: i for i, h in enumerate(hdr) if h}
        body = [r for r in rows[1:] if r and any(c is not None for c in r)]
        wb.close()
        return ix, body

    import mygear_store as MS          # the crew catalogue's category brain
    ix, rs = sheet(rental_path, 'RENTAL_STOCK')

    def g(r, k, d=''):
        return str(r[ix[k]] or '').strip() if k in ix else d

    groups, chase_t, chase_p, idle = {}, [], [], []
    find_av = {}          # the finder's shelf index: name\x1funit -> [items]
    arrivals, plant = [], {'out': [], 'idle': [], 'free': []}
    roster = []                                   # every on-hire item
    #  Same four rules as the daily hit list in the company report
    #  builder (menu H) - one standard across the suite, not two.
    hits = {'radio': [], 'gas': [], 'bat': [], 'tool': []}
    #  For the FRESH LOOK loader (Andrew, 29 Jul 2026: "just using a
    #  once off raw data for on the spot info"): a phone-loaded export
    #  carries SiteIQ's raw wording, so the page carries the rename map
    #  - every item number whose display name Andrew has cleaned up in
    #  the master - and the Plant IDs, so an on-the-spot print reads
    #  exactly like the morning build.
    renames = {}
    pid_map = {}
    #  item number -> product variant, so the stocktake pane can show
    #  the same photos as everywhere else (Andrew, 31 Jul 2026)
    var_by_item = {}
    #  ...and the model-photo fallback for serial-keyed gear (radios,
    #  gas monitors - SiteIQ gives them NO variant, so the photo key is
    #  derived: serial first, model second)
    alt_by_item = {}
    import mygear_thumbs as _TH
    _hidden_n = 0
    for r in rs:
        #  the do-not-show list, applied before anything is counted
        if _HID is not None and _HID.hidden(g(r, 'SKU_NUMBER'),
                                            g(r, 'STORAGE_UNIT')):
            _hidden_n += 1
            continue
        status = g(r, 'ITEM_STATUS')
        #  Anything not yet on the shelf and not out - ordered, in
        #  transit, or stuck in Baseplan. The counter needs to know it is
        #  coming rather than wonder where it went.
        #  (Andrew, 29 Jul 2026: "flag anything that is in arrival status")
        if status not in ('Available for Hire', 'On Hire'):
            if status:
                arrivals.append({
                    'n': MS._tidy(g(r, 'ITEM_DESCRIPTION'), master,
                                  g(r, 'ITEM_NUMBER')),
                    'u': g(r, 'STORAGE_UNIT') or 'Unfiled',
                    's': status})
            continue
        raw = g(r, 'ITEM_DESCRIPTION')

        #  --- the full on-hire roster and the hit list ---
        #  Collected BEFORE the offerable filter on purpose: an item
        #  marked DO NOT HIRE that is somehow out with a crew is exactly
        #  the item the counter must be able to chase and print. The
        #  crew-facing catalogue hides it; the counter's roster cannot.
        if status == 'On Hire':
            d0 = au_date(g(r, 'ON_HIRE_DATE'))
            dy0 = (today - d0).days if d0 else None
            who0 = g(r, 'HIRER_NAME') or 'Not named'
            co0 = g(r, 'COMPANY_NAME') or 'Not named'
            unit0 = g(r, 'STORAGE_UNIT') or 'Unfiled'
            itm0 = g(r, 'ITEM_NUMBER')
            nm0 = MS._tidy(raw, master, itm0) if raw else 'Unnamed item'
            #  the item number and the orange Plant ID travel with every
            #  row - "bring number 41 back" is how a machine is actually
            #  asked for over the counter (Andrew, 29 Jul 2026)
            pid0 = _plant_id(itm0)
            roster.append({'n': nm0, 'u': unit0, 'w': who0, 'co': co0,
                           'd': dy0, 'i': itm0, 'p': pid0,
                           'fl': _comp_fl(itm0, nm0)})
            #  The hit list (Andrew, 29 Jul 2026): "who has not brought
            #  back radios after one day. and milwaukee batteries after
            #  1 day and milwaukee tooling after 3 days."
            #  "After N days" is read conservatively - flagged once the
            #  item is PAST its allowance, so a radio issued yesterday
            #  for a shift that ended this morning is not accused an
            #  hour before the bloke walks it back in.
            if dy0 is not None and 'site plant' not in who0.lower():
                up0 = raw.upper()
                hit = {'n': nm0, 'w': who0, 'co': co0, 'd': dy0,
                       'u': unit0, 'i': itm0, 'p': pid0}
                if unit0 == 'Radios' and dy0 > 1:
                    hits['radio'].append(hit)
                elif unit0 == 'Gas Monitors' and dy0 > 1:
                    hits['gas'].append(hit)
                elif 'MILWAUKEE' in up0:
                    if (('BATT' in up0 or 'CHARGER' in up0)
                            and dy0 > 1):
                        hits['bat'].append(hit)
                    elif ('BATT' not in up0 and 'CHARGER' not in up0
                          and dy0 > 3):
                        hits['tool'].append(hit)

        if not raw or not MS._offerable(raw):
            continue
        name = MS._tidy(raw, master, g(r, 'ITEM_NUMBER'))
        _itm = g(r, 'ITEM_NUMBER')
        if _itm:
            if name != raw:
                renames[_itm] = name
            _pd = _plant_id(_itm)
            if _pd:
                pid_map[_itm] = _pd
        cat = MS._cat_of(name)
        unit = g(r, 'STORAGE_UNIT') or 'Unfiled'
        key = (cat, name)
        e = groups.setdefault(key, {'c': cat, 'n': name, 'av': 0, 'oh': 0,
                                    'u': {}, 'who': [], 'v': '', 'fl': ''})
        if not e['fl']:
            e['fl'] = _comp_fl(_itm, name)
        _vrow = str(g(r, 'PRODUCT_VARIANT') or '').strip().upper()
        if _vrow:
            if _itm:
                var_by_item[_itm] = _vrow
        else:
            #  no variant from SiteIQ (radios, gas monitors): derive -
            #  serial-named photo covers the one unit, model-named
            #  photo covers the fleet
            _dser, _dmod = _TH.derived_keys(raw)
            _vrow = _dmod
            if _itm:
                if _dser:
                    var_by_item[_itm] = _dser
                    if _dmod:
                        alt_by_item[_itm] = _dmod
                elif _dmod:
                    var_by_item[_itm] = _dmod
        if not e['v']:
            #  the photo key - one thumbnail per variant covers every
            #  item behind it (Andrew, 30 Jul 2026)
            e['v'] = _vrow
        #  plant rows carry their SiteIQ family, item number and Plant
        #  ID so the Plant tab can read by CATEGORY - "find welder,
        #  see what's available, what's idle, who's got the rest"
        #  (Andrew, 30 Jul 2026)
        _fam = str(g(r, 'PRODUCT_FAMILY') or '').strip().title() or 'Other Plant'
        _pidv = pid_map.get(_itm, '')
        if status == 'Available for Hire':
            e['av'] += 1
            e['u'][unit] = e['u'].get(unit, 0) + 1
            #  THE FINDER's shelf index (Andrew, 31 Jul 2026: "where's
            #  grinder 1219644?"). Item numbers grouped under
            #  name+unit so names are stored once, not 4,000 times.
            if _itm:
                find_av.setdefault(name + '\x1f' + unit, []).append(_itm)
            if _is_plant(unit):
                plant['free'].append({'n': name, 'u': unit, 'f': _fam,
                                      'i': _itm, 'p': _pidv,
                                      'fl': _comp_fl(_itm, name)})
        else:
            e['oh'] += 1
            d = au_date(g(r, 'ON_HIRE_DATE'))
            days = (today - d).days if d else None
            who = g(r, 'HIRER_NAME') or 'Not named'
            co = g(r, 'COMPANY_NAME') or 'Not named'
            e['who'].append({'w': who, 'co': co, 'd': days, 'u': unit})
            if days is not None and days > 3:
                #  item number rides along so the aisle chase list can
                #  print and scan like everything else (30 Jul 2026)
                row = {'n': name, 'u': unit, 'w': who, 'co': co, 'd': days,
                       'i': _itm}
                (chase_p if _is_plant(unit) else chase_t).append(row)
            if 'site plant' in who.lower():
                idle.append({'n': name, 'u': unit, 'd': days, 'i': _itm})
                if _is_plant(unit):
                    plant['idle'].append({'n': name, 'u': unit, 'd': days,
                                          'f': _fam, 'i': _itm, 'p': _pidv,
                                          'fl': _comp_fl(_itm, name)})
            elif _is_plant(unit):
                plant['out'].append({'n': name, 'u': unit, 'w': who,
                                     'co': co, 'd': days,
                                     'f': _fam, 'i': _itm, 'p': _pidv,
                                     'fl': _comp_fl(_itm, name)})

    #  every list row gets its photo keys - "every time a tool or
    #  consumable is mentioned a thumbnail is to be shown" (Andrew,
    #  31 Jul 2026). Serial key first, model fallback in 'va'.
    def _attach_v(_row):
        _iv = _row.get('i', '')
        _row['v'] = var_by_item.get(_iv, '')
        _av = alt_by_item.get(_iv, '')
        if _av:
            _row['va'] = _av
    for _row in roster:
        _attach_v(_row)
    for _hl in hits.values():
        for _row in _hl:
            _attach_v(_row)
    for _row in chase_t + chase_p + idle:
        _attach_v(_row)
    for _row in (plant.get('out', []) + plant.get('idle', [])
                 + plant.get('free', [])):
        if 'v' not in _row:
            _attach_v(_row)

    #  HIRE HISTORY off the charge feed: how many times each item went
    #  out this shut, and who had it last (Andrew, 31 Jul 2026)
    hire_hist = {}
    try:
        if txn_path and os.path.isfile(txn_path):
            import openpyxl as _px
            _wbt = _px.load_workbook(txn_path, read_only=True,
                                     data_only=True)
            if 'TRANSACTION_CHARGES' in _wbt.sheetnames:
                _tr = _wbt['TRANSACTION_CHARGES'].iter_rows(
                    values_only=True)
                _hd = [str(c or '').strip() for c in next(_tr)]
                _tix = {h: i for i, h in enumerate(_hd)}
                if 'ITEM_NUMBER' in _tix:
                    _hn = _tix.get('HIRER_NAME')
                    _sdx = _tix.get('TRAN_START_DATE')
                    for _r2 in _tr:
                        if not _r2:
                            continue
                        _iv = str(_r2[_tix['ITEM_NUMBER']] or '').strip()
                        if not _iv:
                            continue
                        _d2 = au_date(_r2[_sdx]) if _sdx is not None else None
                        _e2 = hire_hist.setdefault(_iv, [0, '', None])
                        _e2[0] += 1
                        _w2 = (str(_r2[_hn] or '').strip()
                               if _hn is not None else '')
                        if _d2 and (_e2[2] is None or _d2 >= _e2[2]):
                            _e2[2] = _d2
                            if _w2:
                                _e2[1] = _w2[:24]
    except Exception:
        hire_hist = {}

    #  stocktake - how much of the store has actually been laid eyes on
    stock = {'total': 0, 'w1': 0, 'w3': 0, 'w7': 0, 'stale': []}
    #  the pin-point rebuild (Andrew, 31 Jul 2026: "pin point where they
    #  need to be doing the stock... click into that storage area...
    #  seen last 30/14/7/3/1 days all pickable... a percentage... and
    #  highlight things of concern that keep not being found"):
    #    _unit_rows  - per storage area, lines folded by (bucket, name,
    #                  status) with a quantity - the screen shows types
    #                  and counts, never a wall of asset numbers
    #    _unit_walks - the DATES a team counted anything in that area,
    #                  so a stale item can say "missed on N walks"
    _unit_rows = {}
    _unit_walks = {}
    #  sightings per (shift-day, D/N) per area - the raw material of the
    #  completion record (Andrew, 31 Jul 2026: "a record of what they
    #  actually did on the previous night... how do i know what they
    #  have done i dont")
    _sight = {}
    if stocktake_path and os.path.isfile(stocktake_path):
        six, sk = sheet(stocktake_path, 'STOCKTAKE')

        def sg(r, k):
            return str(r[six[k]] or '').strip() if k in six else ''
        def _is_plant_stock(unit_s, name_s):
            #  plant and the bulk yard gear (chutes, hoppers, frames,
            #  barriers) are OUT of the daily stock count - they are
            #  audited on the Plant tab's audit sheet instead (Andrew,
            #  31 Jul 2026: "remove these off our stock checks... as
            #  well as plant gear out of the daily stock count checks")
            uu, nn = unit_s.upper(), name_s.upper()
            return (_is_plant(unit_s) or 'CHUTE' in uu or 'BARRIER' in uu
                    or 'CHUTE' in nn or 'HOPPER' in nn or 'BARRIER' in nn)
        for r in sk:
            d = au_date(sg(r, 'LAST_SIGHTED_DATE_TIME'))
            if not d:
                continue
            if _is_plant_stock(sg(r, 'STORAGE_UNIT'), sg(r, 'DESCRIPTION')):
                continue
            #  hidden lines never reach the stocktake either - if they
            #  did they would drag the counted-in-7-days percentage
            #  down forever for stock the store does not care about
            if _HID is not None and _HID.hidden(
                    sg(r, 'SKU_NUMBER'), sg(r, 'STORAGE_UNIT')):
                continue
            _st_raw = sg(r, 'SIGHTED_STATUS')
            #  ON HIRE IS NOT A STOCK COUNT (Andrew, 2 Aug 2026: "ensure
            #  anything in here is not items onhire that guys are trying
            #  to stock take").
            #
            #  It is not in the store. It is in a bloke's ute or up a
            #  scaffold. Putting it on a counting sheet asks the night
            #  shift to walk an aisle looking for something that was
            #  never going to be on the shelf, and then marks the store
            #  down when they cannot find it. Every one of those lines
            #  was a guaranteed fail dragging the score with it.
            #
            #  Out of the list, out of the total, out of the percentage.
            #  Where that gear actually is stays answerable at the
            #  counter - it is on the roster, the chase list and the
            #  hit list, which is where a missing item belongs.
            if _st_raw == 'On Hire':
                stock['onhire_skipped'] = stock.get('onhire_skipped', 0) + 1
                continue
            age = (today - d).days
            stock['total'] += 1
            _u_st = sg(r, 'STORAGE_UNIT') or 'Unfiled'
            _nm_st = sg(r, 'DESCRIPTION')[:60]
            _it_st = sg(r, 'ITEM_OR_CONSUMABLE') or sg(r, 'SKU_NUMBER')
            _sfl = ('A' if _st_raw in ('Available for Hire',
                                       'In Stock', 'Stock Low') else 'X')
            _bk = (1 if age <= 1 else 3 if age <= 3 else 7 if age <= 7
                   else 14 if age <= 14 else 30 if age <= 30 else 99)
            _fold = _unit_rows.setdefault(_u_st, {})
            _fr = _fold.setdefault((_bk, _nm_st, _sfl),
                                   {'q': 0, 'a': 0, 'v': '', 'ii': []})
            _fr['q'] += 1
            _fr['a'] = max(_fr['a'], age)
            if not _fr['v']:
                #  group tile prefers the MODEL photo; serial photos
                #  belong to single items
                _fr['v'] = (alt_by_item.get(_it_st)
                            or var_by_item.get(_it_st, ''))
            if len(_fr['ii']) < 40 and _it_st:
                #  grouped item numbers, not scattered - the counter
                #  ticks these off under one picture
                _fr['ii'].append([_it_st, age])
            _unit_walks.setdefault(_u_st, set()).add(d)
            _d0, _hm0 = au_dt(r[six['LAST_SIGHTED_DATE_TIME']])
            if _d0:
                _sd, _sh = _shift_of(_d0, _hm0 or (12, 0))
                if _sd and 0 <= (today - _sd).days <= 14:
                    _se = _sight.setdefault((_sd.isoformat(), _sh), {})
                    _se[_u_st] = _se.get(_u_st, 0) + 1
            if age <= 1:
                stock['w1'] += 1
            if age <= 3:
                stock['w3'] += 1
            if age <= 7:
                stock['w7'] += 1
            else:
                #  the not-found sheet needs the item number and the
                #  hire status: an AVAILABLE uncounted item should be
                #  findable on its shelf; an ON-HIRE one is out with a
                #  crew and must NOT be hunted for. (Andrew, 29 Jul
                #  2026: "print out of items per storage unit of items
                #  missing in stock take... only looking for things
                #  that are available... option here for just onhire
                #  so they know areas to look and items no to look
                #  for")
                _st = sg(r, 'SIGHTED_STATUS')
                _hh = hire_hist.get(_it_st, [0, '', None])
                stock['stale'].append({
                    'n': _nm_st,
                    'u': _u_st,
                    'd': age, 'by': sg(r, 'LAST_SIGHTED_BY')[:26],
                    'i': _it_st,
                    'v': var_by_item.get(_it_st, '') or _fr['v'],
                    'va': alt_by_item.get(_it_st, ''),
                    'hc': _hh[0], 'lh': _hh[1],
                    's': _sfl})
    stock['stale'].sort(key=lambda x: -x['d'])

    #  the area scoreboard: worst-counted first, so "where do we walk
    #  today" is answered by reading top to bottom
    _areas = []
    for _u_st, _fold in _unit_rows.items():
        _tot_u = sum(x['q'] for x in _fold.values())
        _s7 = sum(x['q'] for (_b, _n, _s), x in _fold.items() if _b <= 7)
        _bks = {1: 0, 3: 0, 7: 0, 14: 0, 30: 0, 99: 0}
        for (_b, _n, _s), x in _fold.items():
            _bks[_b] += x['q']
        _areas.append({
            'u': _u_st, 't': _tot_u, 's7': _s7,
            'p': int(round(100.0 * _s7 / _tot_u)) if _tot_u else 100,
            'b': [_bks[1], _bks[3], _bks[7], _bks[14], _bks[30], _bks[99]]})
    _areas.sort(key=lambda a: (a['p'], -a['t']))
    stock['areas'] = _areas
    stock['rows'] = [
        {'u': _u_st, 'b': _k[0], 'n': _k[1], 's': _k[2],
         'q': _x['q'], 'a': _x['a'], 'v': _x['v'], 'ii': _x['ii']}
        for _u_st, _fold in _unit_rows.items()
        for _k, _x in _fold.items()]

    #  THE GHOST LIST - the real worry: an area the team HAS walked,
    #  more than once, and the item still has not turned up. Every walk
    #  date in the area after the item's own last sighting counts as a
    #  miss. Two misses = it is not "waiting its turn" any more.
    _ghosts = []
    for _x in stock['stale']:
        if _x['s'] != 'A':
            continue
        _ld = today - dt.timedelta(days=_x['d'])
        _m = sum(1 for _w in _unit_walks.get(_x['u'], ()) if _w > _ld)
        if _m >= 2:
            _g = dict(_x)
            _g['m'] = min(_m, 9)
            _ghosts.append(_g)
    _ghosts.sort(key=lambda g: (-g['m'], -g['d']))
    stock['ghosts'] = _ghosts[:40]

    #  TODAY'S COUNTING ORDERS (Andrew, 30 Jul 2026: "daily dayshift
    #  what to check like a category or categories... for both shifts
    #  so they know exactly what they need to do. split needs to be 30%
    #  days and 70% nights - that part we dont wanna show on here").
    #
    #  The workload is the findable backlog: AVAILABLE items not laid
    #  eyes on in 7 days, per storage unit. On-hire items are excluded
    #  - they count when they come back, not on a walk. Nights carry
    #  the bigger share because nights are quieter at the counter; the
    #  weighting is deliberate policy and deliberately NOT printed.
    #  The pick rotates with the date so the same crew does not own the
    #  same aisle forever, and both shifts see the same orders all day.
    _backlog = {}
    for x in stock['stale']:
        if x['s'] == 'A':
            _backlog[x['u']] = _backlog.get(x['u'], 0) + 1
    orders = {'d': [], 'n': [], 'dN': 0, 'nN': 0}
    _units = sorted(_backlog.items(), key=lambda kv: (-kv[1], kv[0]))
    if _units:
        _rot = today.toordinal() % len(_units)
        _units = _units[_rot:] + _units[:_rot]
        _total = sum(n for _u, n in _units)
        _NIGHT_SHARE = 0.7          # policy, never rendered
        for _u, _n in _units:
            if (orders['nN'] < _total * _NIGHT_SHARE
                    or not orders['n']):
                orders['n'].append({'u': _u, 'n': _n})
                orders['nN'] += _n
            else:
                orders['d'].append({'u': _u, 'n': _n})
                orders['dN'] += _n
        #  a shift with nothing to do reads as a mistake - if days got
        #  nothing and there are at least two aisles, hand them the
        #  lightest night aisle
        if not orders['d'] and len(orders['n']) > 1:
            _mv = min(orders['n'], key=lambda o: o['n'])
            orders['n'].remove(_mv)
            orders['nN'] -= _mv['n']
            orders['d'].append(_mv)
            orders['dN'] += _mv['n']
        orders['d'].sort(key=lambda o: -o['n'])
        orders['n'].sort(key=lambda o: -o['n'])

    #  clear direction, no overlap: every area card carries WHOSE aisle
    #  it is today, so day and night never walk the same patch
    _own = {}
    for _o in orders['d']:
        _own[_o['u']] = 'd'
    for _o in orders['n']:
        _own[_o['u']] = 'n'
    for _a in stock.get('areas', []):
        _a['own'] = _own.get(_a['u'], '')

    #  THE COMPLETION RECORD - writes itself, every build. Today's
    #  orders are remembered in stocktake_log.json; the sightings each
    #  shift actually made (off the scanner timestamps) are written
    #  against each day. Nobody fills in a form; the record simply
    #  exists next morning. Local data file - never in git or a zip.
    stock['log'] = []
    try:
        import json as _json
        _lp = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           'stocktake_log.json')
        try:
            with open(_lp, 'r', encoding='utf-8') as _f:
                _slog = _json.load(_f)
        except (OSError, ValueError):
            _slog = {}
        _tiso = today.isoformat()
        _e0 = _slog.setdefault(_tiso, {})
        if 'orders' not in _e0:
            _e0['orders'] = {
                'd': [[_o['u'], _o['n']] for _o in orders['d']],
                'n': [[_o['u'], _o['n']] for _o in orders['n']]}
        for (_di, _sh), _cnts in _sight.items():
            _e = _slog.setdefault(_di, {})
            _e.setdefault('done', {})[_sh.lower()] = _cnts
        _cut = (today - dt.timedelta(days=30)).isoformat()
        for _k in [k for k in list(_slog) if k < _cut]:
            del _slog[_k]
        with open(_lp, 'w', encoding='utf-8') as _f:
            _json.dump(_slog, _f)
        for _di in sorted([k for k in _slog if k <= _tiso],
                          reverse=True)[:7]:
            _e = _slog.get(_di, {})
            _dd = dt.date.fromisoformat(_di)
            stock['log'].append({
                'dt': _dd.strftime('%a %d %b'),
                'today': _di == _tiso,
                'o': _e.get('orders', {'d': [], 'n': []}),
                'c': _e.get('done', {})})
    except Exception as _le:
        stock['log'] = []
        print('  Stocktake record: skipped this build ({})'.format(_le))

    battle = _battle(txn_path, stocktake_path) if txn_path else None

    G = sorted(groups.values(), key=lambda e: (e['c'], e['n'].lower()))
    chase_t.sort(key=lambda x: (-x['d'], x['u']))
    chase_p.sort(key=lambda x: -x['d'])

    def by_unit(rows):
        out = {}
        for x in rows:
            out[x['u']] = out.get(x['u'], 0) + 1
        return sorted(out.items(), key=lambda t: -t[1])

    arrivals.sort(key=lambda x: (x['s'], x['n'].lower()))

    #  --- consumables (Andrew, 29 Jul 2026) ---
    #  The counter gets the same shelf figures as the manager's
    #  utilisation report, from the same engine, so the two can never
    #  drift apart and tell a crew two different stories. Only the
    #  lines that need a decision travel into the page - the full
    #  71-line register belongs in the report, not on a tablet.
    cons = None
    if sales_path:
        try:
            import k2_consumables
            cd = k2_consumables.read(sales_path, stocktake_path,
                                     base or '.', today=today)
        except Exception:
            cd = None                     # never stop the board over this
        if cd:
            def slim(rows, n):
                return [{'n': x['desc'], 'a': x['avail'], 'u': x['used'],
                         'b': round(x['burn'], 1),
                         'c': (int(x['cover']) if x['cover'] is not None
                               else None),
                         'k': x['sku'],
                         'ct': x['counted'], 'v': x['varAdj'],
                         'tw': x['twinHolds'],
                         'w': x.get('why', '')}
                        for x in rows[:n]]
            cons = {
                'folded': cd.get('folded', 0),
                #  the full register travels too - the stock check &
                #  reorder sheet is a walk of EVERY line with a pen,
                #  not a summary (Andrew, 29 Jul 2026: "updated stock
                #  check. qty we have. stock last stock check info.
                #  then a reorder column so we can put down a number")
                'all': [{'n': x['desc'], 'k': x['sku'],
                         'a': int(x['avail']), 'u': int(x['used']),
                         'ct': (None if x['counted'] is None
                                else int(x['counted'])),
                         'co': x['countedOn'], 'by': x['countedBy'],
                         #  every movement this line has had, so the
                         #  counter can tap a consumable and see who
                         #  took what and when (2 Aug 2026)
                         'tx': x.get('tx') or [],
                         'un': x.get('unit', '')}
                        for x in sorted(cd['items'],
                                        key=lambda i: i['desc'].upper())],
                'avail': int(cd['avail']), 'used': int(cd['used']),
                'skus': cd['skus'], 'moves': cd['moves'],
                'end': cd['end'], 'daysLeft': cd['daysLeft'],
                'checkedPct': int(cd['checkedPct'] + .5),
                'matchPct': int(cd['matchPct'] + .5),
                'checked': cd['checked'],
                'order': slim(cd['order'], 40),
                'watch': slim(cd['watch'], 20),
                'records': slim(cd['records'], 20),
                'off': slim(cd['off'], 30),
                'dead': len(cd['dead']),
                'explained': cd['explained'],
                'countDays': cd['countDays'],
            }

    #  ONE ordering rule for the whole board (Andrew, 29 Jul 2026):
    #  companies A-Z, hirers inside a company A-Z, and a person's items
    #  longest-held first with A-Z breaking ties. The same list always
    #  reads in the same order, so the counter finds a name by eye
    #  instead of by scanning.
    for k in hits:
        hits[k].sort(key=lambda x: (-x['d'], x['w'].upper(), x['n'].upper()))
    roster.sort(key=lambda x: (x['co'].upper(), x['w'].upper(),
                               -(x['d'] if x['d'] is not None else -1),
                               x['n'].upper()))
    for e in groups.values():
        e['who'].sort(key=lambda w: (-(w['d'] if w['d'] is not None else -1),
                                     w['w'].upper()))

    return {
        'battle': battle,
        'arrivals': arrivals,
        'cons': cons,
        'hasCons': bool(cons),
        'hits': hits,
        'hitN': sum(len(v) for v in hits.values()),
        'roster': roster,
        'ren': renames,
        'pids': pid_map,
        'plant': plant,
        'hasPlant': bool(plant['out'] or plant['idle'] or plant['free']),
        'groups': G,
        #  THE FINDER: every available item by number, names stored once
        'find': {'av': find_av},
        #  the CURRENT test-tag colour, from the compliance master's
        #  windows (Jul-Aug = BLUE) - the page prints the word it is
        #  told, so a new quarter needs no page change
        'tag': _tag_now(),
        #  hire history per item off the charge feed - times out this
        #  shut + who had it last, shown wherever an item is named
        #  (Andrew, 31 Jul 2026)
        'hist': {k: [v[0], v[1]] for k, v in hire_hist.items() if v[0]},
        'chase': {'tools': chase_t, 'plant': chase_p,
                  'toolUnits': by_unit(chase_t),
                  'plantUnits': by_unit(chase_p)},
        'stock': stock,
        'orders': orders,
        'idle': idle,
        'tiles': {
            'avail': sum(e['av'] for e in G),
            'onhire': sum(e['oh'] for e in G),
            'lines': len(G),
            'chase': len(chase_t),
            'plantOut': len(chase_p),
            'idle': len(idle),
            'stockPct': int(100.0 * stock['w7'] / stock['total'] + 0.5)
                        if stock['total'] else 0,
            'stale': len(stock['stale']),
            'arrivals': len(arrivals),
            'plantOn': len(plant['out']),
            'plantIdle': len(plant['idle']),
            'plantFree': len(plant['free']),
        },
    }


# ---------------------------------------------------------------------
#  THE PAGE
# ---------------------------------------------------------------------
#  The in-browser .xlsx reader for the FRESH LOOK loader. A RAW string
#  on purpose: it is full of regex backslashes, and Python quietly
#  eating one of them has already broken this page's script block twice
#  today. Proven against a real phone-downloaded export: 1,017 rows,
#  19,323 cells, zero mismatches against openpyxl. It is a faithful
#  port of zlib's reference inflate (puff) plus a minimal zip walk and
#  just enough XML for shared strings and one sheet.
#  QR LITE, PORTED TO JS - the same encoder as qr_lite.py, line for
#  line, so a QR printed off this board and a QR on a poster come from
#  ONE proven implementation. Ported (not rewritten) 30 Jul 2026 for
#  Andrew's "add a QR Code for all Item Numbers ... this allows us to
#  scan these sheets"; verified against qr_lite.qr_matrix byte for byte
#  in the headless browser before shipping. Byte mode, level M,
#  versions 1-10, all eight masks scored per the standard. No external
#  library, nothing fetched - the sheets stay offline like everything
#  else. Raw string: keep backslashes out of the JS.
_QR_JS = r"""
var QRL=(function(){
var EXP=new Array(512),LOG=new Array(256);
(function(){var x=1,i;for(i=0;i<255;i++){EXP[i]=x;LOG[x]=i;x<<=1;if(x&256)x^=285;}
for(i=255;i<512;i++)EXP[i]=EXP[i-255];})();
function mul(a,b){return(a===0||b===0)?0:EXP[LOG[a]+LOG[b]];}
function rsGen(n){var g=[1],i,j;
 for(i=0;i<n;i++){var g2=[],k;for(k=0;k<g.length+1;k++)g2.push(0);
  for(j=0;j<g.length;j++){g2[j]^=g[j];g2[j+1]^=mul(g[j],EXP[i]);}g=g2;}
 return g;}
function rsEnc(data,n){var g=rsGen(n),res=data.slice(),i,j;
 for(i=0;i<n;i++)res.push(0);
 for(i=0;i<data.length;i++){var c=res[i];
  if(c){for(j=1;j<g.length;j++)res[i+j]^=mul(g[j],c);}}
 return res.slice(data.length);}
var M={1:[26,10,1],2:[44,16,1],3:[70,26,1],4:[100,18,2],5:[134,24,2],
 6:[172,16,4],7:[196,18,4],8:[242,22,4],9:[292,22,5],10:[346,26,5]};
var ALIGN={1:[],2:[6,18],3:[6,22],4:[6,26],5:[6,30],6:[6,34],
 7:[6,22,38],8:[6,24,42],9:[6,26,46],10:[6,28,50]};
var FORMAT=[0x5412,0x5125,0x5E7C,0x5B4B,0x45F9,0x40CE,0x4F97,0x4AA0];
var VBITS={7:0x07C94,8:0x085BC,9:0x09A99,10:0x0A4D3};
function cap(v){return M[v][0]-M[v][1]*M[v][2];}
function pickVer(n){var v,cci,need;
 for(v=1;v<=10;v++){cci=v<10?8:16;
  need=Math.floor((4+cci+n*8+7)/8);
  if(need<=cap(v))return v;}
 return 0;}
function bitstream(data,ver){var cci=ver<10?8:16,bits=[],i,j;
 function put(val,n){for(var k=n-1;k>=0;k--)bits.push((val>>k)&1);}
 put(4,4);put(data.length,cci);
 for(i=0;i<data.length;i++)put(data[i],8);
 var capb=cap(ver)*8;
 put(0,Math.min(4,capb-bits.length));
 while(bits.length%8)bits.push(0);
 var words=[];
 for(i=0;i<bits.length;i+=8){var w=0;for(j=0;j<8;j++)w=(w<<1)|bits[i+j];words.push(w);}
 var pad=[236,17],k2=0;
 while(words.length<cap(ver)){words.push(pad[k2%2]);k2++;}
 return words;}
function interleave(words,ver){var eccN=M[ver][1],nb=M[ver][2];
 var dcount=cap(ver),shrt=Math.floor(dcount/nb),extra=dcount%nb;
 var blocks=[],ecs=[],pos=0,i,b;
 for(b=0;b<nb;b++){var size=shrt+((b>=nb-extra)?1:0);
  var blk=words.slice(pos,pos+size);pos+=size;
  blocks.push(blk);ecs.push(rsEnc(blk,eccN));}
 var out=[],mx=0;
 for(b=0;b<blocks.length;b++)if(blocks[b].length>mx)mx=blocks[b].length;
 for(i=0;i<mx;i++)for(b=0;b<blocks.length;b++)
  if(i<blocks[b].length)out.push(blocks[b][i]);
 for(i=0;i<eccN;i++)for(b=0;b<ecs.length;b++)out.push(ecs[b][i]);
 return out;}
function matrix(ver,words){var size=ver*4+17,m=[],r,c,i,j;
 for(r=0;r<size;r++){var row=[];for(c=0;c<size;c++)row.push(null);m.push(row);}
 function finder(r0,c0){for(var i2=-1;i2<8;i2++)for(var j2=-1;j2<8;j2++){
  var rr=r0+i2,cc=c0+j2;
  if(rr>=0&&rr<size&&cc>=0&&cc<size){
   var on=((i2>=0&&i2<=6)&&(j2===0||j2===6))
    ||((j2>=0&&j2<=6)&&(i2===0||i2===6))
    ||(i2>=2&&i2<=4&&j2>=2&&j2<=4);
   m[rr][cc]=on?1:0;}}}
 finder(0,0);finder(0,size-7);finder(size-7,0);
 for(i=8;i<size-8;i++){var v=(i%2===0)?1:0;m[6][i]=v;m[i][6]=v;}
 var al=ALIGN[ver],a,b2;
 for(a=0;a<al.length;a++)for(b2=0;b2<al.length;b2++){
  r=al[a];c=al[b2];
  if(m[r][c]!==null)continue;
  for(i=-2;i<3;i++)for(j=-2;j<3;j++)
   m[r+i][c+j]=(Math.max(Math.abs(i),Math.abs(j))!==1)?1:0;}
 m[size-8][8]=1;
 for(i=0;i<9;i++){if(m[8][i]===null)m[8][i]=0;if(m[i][8]===null)m[i][8]=0;}
 for(i=0;i<8;i++){if(m[8][size-1-i]===null)m[8][size-1-i]=0;
  if(m[size-1-i][8]===null)m[size-1-i][8]=0;}
 if(ver>=7){var vb=VBITS[ver];
  for(i=0;i<18;i++){var bb=(vb>>i)&1;
   m[Math.floor(i/3)][size-11+i%3]=bb;
   m[size-11+i%3][Math.floor(i/3)]=bb;}}
 var bits=[];
 for(i=0;i<words.length;i++)for(j=7;j>=0;j--)bits.push((words[i]>>j)&1);
 var idx=0,up=true,col=size-1;
 while(col>0){
  if(col===6)col-=1;
  var rows=[];
  if(up){for(r=size-1;r>=0;r--)rows.push(r);}
  else{for(r=0;r<size;r++)rows.push(r);}
  for(var ri=0;ri<rows.length;ri++){r=rows[ri];
   for(var ci=0;ci<2;ci++){c=(ci===0)?col:col-1;
    if(m[r][c]===null){m[r][c]=(idx<bits.length)?bits[idx]:0;idx++;}}}
  up=!up;col-=2;}
 return {m:m,size:size};}
function reservedMap(size,ver){var res=[],r,c,i;
 for(r=0;r<size;r++){var row=[];for(c=0;c<size;c++)row.push(false);res.push(row);}
 function block(r0,c0,h,w){for(var i2=0;i2<h;i2++)for(var j2=0;j2<w;j2++){
  var rr=r0+i2,cc=c0+j2;
  if(rr>=0&&rr<size&&cc>=0&&cc<size)res[rr][cc]=true;}}
 block(0,0,9,9);block(0,size-8,9,8);block(size-8,0,8,9);
 for(i=0;i<size;i++){res[6][i]=true;res[i][6]=true;}
 var al=ALIGN[ver],a,b2;
 for(a=0;a<al.length;a++)for(b2=0;b2<al.length;b2++){
  r=al[a];c=al[b2];
  if((r<9&&c<9)||(r<9&&c>size-10)||(r>size-10&&c<9))continue;
  block(r-2,c-2,5,5);}
 if(ver>=7){block(0,size-11,6,3);block(size-11,0,3,6);}
 return res;}
function applyMask(m,size,mask,ver){var g=[],r,c,i;
 for(r=0;r<size;r++)g.push(m[r].slice());
 var res=reservedMap(size,ver);
 for(r=0;r<size;r++)for(c=0;c<size;c++){
  if(res[r][c])continue;
  var cond;
  switch(mask){
   case 0:cond=((r+c)%2===0);break;
   case 1:cond=(r%2===0);break;
   case 2:cond=(c%3===0);break;
   case 3:cond=((r+c)%3===0);break;
   case 4:cond=((Math.floor(r/2)+Math.floor(c/3))%2===0);break;
   case 5:cond=(((r*c)%2+(r*c)%3)===0);break;
   case 6:cond=((((r*c)%2+(r*c)%3)%2)===0);break;
   default:cond=((((r+c)%2+(r*c)%3)%2)===0);}
  if(cond)g[r][c]^=1;}
 var fmt=FORMAT[mask];
 for(i=0;i<15;i++){var b=(fmt>>i)&1;
  if(i<6)g[i][8]=b;
  else if(i===6)g[7][8]=b;
  else if(i===7)g[8][8]=b;
  else if(i===8)g[8][7]=b;
  else g[8][14-i]=b;
  if(i<8)g[8][size-1-i]=b;
  else g[size-15+i][8]=b;}
 g[size-8][8]=1;
 return g;}
function penalty(g,size){var p=0,r,c,i,k;
 var lines=[];
 for(r=0;r<size;r++)lines.push(g[r]);
 for(c=0;c<size;c++){var col=[];for(r=0;r<size;r++)col.push(g[r][c]);lines.push(col);}
 for(i=0;i<lines.length;i++){var run=0,prev=-1,line=lines[i];
  for(k=0;k<line.length;k++){var v=line[k];
   if(v===prev){run+=1;}
   else{if(run>=5)p+=3+(run-5);run=1;prev=v;}}
  if(run>=5)p+=3+(run-5);}
 for(r=0;r<size-1;r++)for(c=0;c<size-1;c++){
  var s=g[r][c]+g[r][c+1]+g[r+1][c]+g[r+1][c+1];
  if(s===0||s===4)p+=3;}
 var dark=0;
 for(r=0;r<size;r++)for(c=0;c<size;c++)dark+=g[r][c];
 p+=10*Math.floor(Math.abs(Math.floor(dark*100/(size*size))-50)/5);
 return p;}
function toBytes(s){var out=[],i,c;
 for(i=0;i<s.length;i++){c=s.charCodeAt(i);
  if(c<128)out.push(c);
  else if(c<2048){out.push(192|(c>>6));out.push(128|(c&63));}
  else{out.push(224|(c>>12));out.push(128|((c>>6)&63));out.push(128|(c&63));}}
 return out;}
var MEMO={};
function qrMatrix(text){
 if(MEMO[text])return MEMO[text];
 var data=toBytes(String(text));
 var ver=pickVer(data.length);
 if(!ver)return null;
 var words=interleave(bitstream(data,ver),ver);
 var base=matrix(ver,words);
 var best=null,bestp=-1;
 for(var mask=0;mask<8;mask++){
  var g=applyMask(base.m,base.size,mask,ver);
  var p=penalty(g,base.size);
  if(bestp<0||p<bestp){best=g;bestp=p;}}
 MEMO[text]=best;
 return best;}
function qrSvg(text,px){
 var m=qrMatrix(text);
 if(!m)return '';
 var n=m.length,quiet=2,total=n+quiet*2,rects='',r,c;
 for(r=0;r<n;r++){c=0;
  while(c<n){
   if(m[r][c]){var run=1;
    while(c+run<n&&m[r][c+run])run++;
    rects+='<rect x="'+(c+quiet)+'" y="'+(r+quiet)
      +'" width="'+run+'" height="1"/>';
    c+=run;}
   else c++;}}
 return '<svg class="qrs" width="'+px+'" height="'+px+'" viewBox="0 0 '
  +total+' '+total+'" shape-rendering="crispEdges" '
  +'xmlns="http://www.w3.org/2000/svg">'
  +'<rect width="'+total+'" height="'+total+'" fill="#ffffff"/>'
  +'<g fill="#101317">'+rects+'</g></svg>';}
return {matrix:qrMatrix,svg:qrSvg};
})();
/* the one call every list uses: the QR IS the item number - scan it
   with the store scanner or a phone and you get the number itself,
   exactly as typed. Empty when a row has no number: never a QR of
   nothing. */
function qr(t,px){ if(t==null||t===''){return '';} return QRL.svg(String(t),px||46); }
"""

_READER_JS = r"""
function _inflate(src){
  var pos=0, bitbuf=0, bitcnt=0, out=[];
  var MAXBITS=15;
  function bits(n){
    var val=bitbuf;
    while(bitcnt<n){ val|=src[pos++]<<bitcnt; bitcnt+=8; }
    bitbuf=val>>>n; bitcnt-=n;
    return val&((1<<n)-1);
  }
  function huft(lens,n){
    var counts=[],syms=[],offs=[],i;
    for(i=0;i<=MAXBITS;i++) counts[i]=0;
    for(i=0;i<n;i++) counts[lens[i]]++;
    if(counts[0]===n) return {c:counts,s:syms};
    offs[1]=0;
    for(i=1;i<MAXBITS;i++) offs[i+1]=offs[i]+counts[i];
    for(i=0;i<n;i++) if(lens[i]) syms[offs[lens[i]]++]=i;
    return {c:counts,s:syms};
  }
  function decode(h){
    var code=0,first=0,index=0,len=1;
    while(len<=MAXBITS){
      code|=bits(1);
      var count=h.c[len];
      if(code-first<count) return h.s[index+(code-first)];
      index+=count; first+=count;
      first<<=1; code<<=1; len++;
    }
    throw new Error('bad code');
  }
  var LENS=[3,4,5,6,7,8,9,10,11,13,15,17,19,23,27,31,35,43,51,59,67,83,99,115,131,163,195,227,258];
  var LEXT=[0,0,0,0,0,0,0,0,1,1,1,1,2,2,2,2,3,3,3,3,4,4,4,4,5,5,5,5,0];
  var DISTS=[1,2,3,4,5,7,9,13,17,25,33,49,65,97,129,193,257,385,513,769,1025,1537,2049,3073,4097,6145,8193,12289,16385,24577];
  var DEXT=[0,0,0,0,1,1,2,2,3,3,4,4,5,5,6,6,7,7,8,8,9,9,10,10,11,11,12,12,13,13];
  function codes(lc,dc){
    while(true){
      var sym=decode(lc);
      if(sym<256){ out.push(sym); }
      else if(sym===256){ return; }
      else{
        sym-=257;
        var len=LENS[sym]+bits(LEXT[sym]);
        var dsym=decode(dc);
        var dist=DISTS[dsym]+bits(DEXT[dsym]);
        var from=out.length-dist;
        for(var i=0;i<len;i++) out.push(out[from+i]);
      }
    }
  }
  var fixedL=null, fixedD=null;
  function fixed(){
    if(!fixedL){
      var lens=[],i;
      for(i=0;i<144;i++) lens[i]=8;
      for(;i<256;i++) lens[i]=9;
      for(;i<280;i++) lens[i]=7;
      for(;i<288;i++) lens[i]=8;
      fixedL=huft(lens,288);
      var dl=[]; for(i=0;i<30;i++) dl[i]=5;
      fixedD=huft(dl,30);
    }
    codes(fixedL,fixedD);
  }
  var ORDER=[16,17,18,0,8,7,9,6,10,5,11,4,12,3,13,2,14,1,15];
  function dynamic(){
    var nlen=bits(5)+257, ndist=bits(5)+1, ncode=bits(4)+4;
    var lens=[],i;
    for(i=0;i<19;i++) lens[i]=0;
    for(i=0;i<ncode;i++) lens[ORDER[i]]=bits(3);
    var lench=huft(lens,19);
    var all=[]; i=0;
    while(i<nlen+ndist){
      var sym=decode(lench);
      if(sym<16){ all[i++]=sym; }
      else{
        var rep,val=0;
        if(sym===16){ val=all[i-1]; rep=3+bits(2); }
        else if(sym===17){ rep=3+bits(3); }
        else{ rep=11+bits(7); }
        while(rep--) all[i++]=val;
      }
    }
    codes(huft(all.slice(0,nlen),nlen),
          huft(all.slice(nlen),ndist));
  }
  function stored(){
    bitbuf=0; bitcnt=0;
    var len=src[pos]|(src[pos+1]<<8); pos+=4;
    for(var i=0;i<len;i++) out.push(src[pos++]);
  }
  var last;
  do{
    last=bits(1);
    var type=bits(2);
    if(type===0) stored();
    else if(type===1) fixed();
    else if(type===2) dynamic();
    else throw new Error('bad block type');
  }while(!last);
  return new Uint8Array(out);
}
function _zipEntries(buf){
  var v=new DataView(buf.buffer||buf), n=v.byteLength, i;
  for(i=n-22;i>=0;i--) if(v.getUint32(i,true)===0x06054b50) break;
  if(i<0) throw new Error('not a zip');
  var count=v.getUint16(i+10,true), off=v.getUint32(i+16,true);
  var entries={}, p=off;
  for(var e=0;e<count;e++){
    if(v.getUint32(p,true)!==0x02014b50) break;
    var method=v.getUint16(p+10,true);
    var csize=v.getUint32(p+20,true);
    var nlen=v.getUint16(p+28,true), xlen=v.getUint16(p+30,true),
        clen=v.getUint16(p+32,true);
    var lho=v.getUint32(p+42,true);
    var name='';
    for(var c=0;c<nlen;c++) name+=String.fromCharCode(buf[p+46+c]);
    entries[name]={method:method,csize:csize,lho:lho};
    p+=46+nlen+xlen+clen;
  }
  return entries;
}
function _zipRead(buf,entries,name){
  var e=entries[name];
  if(!e) return null;
  var v=new DataView(buf.buffer||buf);
  var p=e.lho;
  if(v.getUint32(p,true)!==0x04034b50) throw new Error('bad local header');
  var nlen=v.getUint16(p+26,true), xlen=v.getUint16(p+28,true);
  var data=buf.subarray(p+30+nlen+xlen, p+30+nlen+xlen+e.csize);
  var raw=e.method===8?_inflate(data):data;
  if(typeof TextDecoder!=='undefined') return new TextDecoder('utf-8').decode(raw);
  var s='';
  for(var i=0;i<raw.length;i++) s+=String.fromCharCode(raw[i]);
  try{ return decodeURIComponent(escape(s)); }catch(err){ return s; }
}
function _unent(s){
  return String(s)
    .replace(/&#x([0-9a-fA-F]+);/g,function(_,h){return String.fromCharCode(parseInt(h,16))})
    .replace(/&#(\d+);/g,function(_,d){return String.fromCharCode(parseInt(d,10))})
    .replace(/&lt;/g,'<').replace(/&gt;/g,'>')
    .replace(/&quot;/g,'"').replace(/&apos;/g,"'").replace(/&amp;/g,'&');
}
function _sst(xml){
  var out=[], re=/<si[ >][\s\S]*?<\/si>|<si\/>/g, m;
  while((m=re.exec(xml))){
    var t='', tr=/<t[^>]*>([\s\S]*?)<\/t>/g, tm;
    while((tm=tr.exec(m[0]))) t+=_unent(tm[1]);
    out.push(t);
  }
  return out;
}
function _colIdx(ref){
  var n=0;
  for(var i=0;i<ref.length;i++){
    var c=ref.charCodeAt(i);
    if(c>=65&&c<=90) n=n*26+(c-64); else break;
  }
  return n-1;
}
function _sheetRows(xml,sst){
  var rows=[], rre=/<row[^>]*>([\s\S]*?)<\/row>/g, rm;
  while((rm=rre.exec(xml))){
    var cells=[], cre=/<c ([^>]*?)(?:\/>|>([\s\S]*?)<\/c>)/g, cm;
    while((cm=cre.exec(rm[1]))){
      var attrs=cm[1], body=cm[2]||'';
      var ref=/r="([A-Z]+)\d+"/.exec(attrs);
      var typ=/t="([^"]+)"/.exec(attrs);
      var val='';
      var vm=/<v>([\s\S]*?)<\/v>/.exec(body);
      if(vm) val=_unent(vm[1]);
      var im=/<is>[\s\S]*?<\/is>/.exec(body);
      if(im){ var t='',tr=/<t[^>]*>([\s\S]*?)<\/t>/g,tm;
        while((tm=tr.exec(im[0]))) t+=_unent(tm[1]); val=t; }
      if(typ&&typ[1]==='s') val=sst[parseInt(val,10)]||'';
      cells[ref?_colIdx(ref[1]):cells.length]=val;
    }
    rows.push(cells);
  }
  return rows;
}
function readXlsxSheet(bytes,wantSheet){
  var buf=bytes instanceof Uint8Array?bytes:new Uint8Array(bytes);
  var entries=_zipEntries(buf);
  var wb=_zipRead(buf,entries,'xl/workbook.xml')||'';
  var rels=_zipRead(buf,entries,'xl/_rels/workbook.xml.rels')||'';
  var relmap={}, rr=/<Relationship [^>]*Id="([^"]+)"[^>]*Target="([^"]+)"[^>]*\/>/g, rm2;
  while((rm2=rr.exec(rels))) relmap[rm2[1]]=rm2[2];
  var sheetFile=null, sr=/<sheet [^>]*name="([^"]+)"[^>]*r:id="([^"]+)"/g, sm;
  while((sm=sr.exec(wb))){
    if(!wantSheet||_unent(sm[1]).toUpperCase()===String(wantSheet).toUpperCase()){
      sheetFile='xl/'+relmap[sm[2]].replace(/^\//,'').replace(/^xl\//,'');
      if(wantSheet) break;
    }
  }
  if(!sheetFile) return null;
  var sst=_sst(_zipRead(buf,entries,'xl/sharedStrings.xml')||'');
  var rows=_sheetRows(_zipRead(buf,entries,sheetFile)||'',sst);
  if(!rows.length) return {header:[],rows:[]};
  var header=rows[0].map(function(h){return String(h||'').trim()});
  var out=[];
  for(var i=1;i<rows.length;i++){
    var o={},any=false;
    for(var j=0;j<header.length;j++){
      if(!header[j]) continue;
      var v=rows[i][j]==null?'':String(rows[i][j]).trim();
      o[header[j]]=v;
      if(v) any=true;
    }
    if(any) out.push(o);
  }
  return {header:header,rows:out};
}
"""

PAGE = """<!DOCTYPE html><!-- MY GEAR HQ · the Coates stores board · designed and built by Andrew Fisher --><html lang="en-AU"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
<meta name="theme-color" content="#F26222">
<title>MY GEAR HQ · Coates Stores Team</title><style>__FONTCSS__
:root{--org:#F26222;--ink:#0A0E14;--pnl:#151A22;--pnl2:#1C232D;--line:#2A3340;
 --txt:#E9EEF5;--dim:#98A4B4;--gd:#2BB673;--am:#F5A623;--rd:#E23B2E;--neon:#EFFF3D}
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
body{background:var(--ink);color:var(--txt);font-family:-apple-system,
 BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif;font-size:16px;line-height:1.5}
.wrap{max-width:820px;margin:0 auto;padding:0 13px 90px}
/* ---- the MY GEAR HQ masthead (the brand, 1 Aug 2026). Forged
   carbon: the weave, the brushed-steel MY, the forged-orange GEAR,
   the ember line. The counter's page stopped being a tab and became
   the headquarters. ---- */
header{background:#0a0b0d;padding:20px 15px 13px;margin:0 -13px 14px;
 position:relative;overflow:hidden;text-align:center}
header:before{content:"";position:absolute;inset:0;opacity:.5;
 background:repeating-linear-gradient(45deg,#101216 0 5px,#16181d 5px 10px),
 repeating-linear-gradient(-45deg,rgba(255,255,255,.02) 0 5px,transparent 5px 10px)}
header:after{content:"";position:absolute;inset:0;
 background:radial-gradient(ellipse 90% 130% at 50% 28%,transparent 40%,rgba(0,0,0,.72) 100%)}
.hqmast{position:relative;z-index:2;font-family:'Archivo Black','Arial Black',Arial,sans-serif;
 font-size:52px;line-height:1;letter-spacing:2px;display:flex;justify-content:center;
 align-items:baseline;gap:11px}
.hqmy{background:linear-gradient(180deg,#fff 0%,#d9dde3 45%,#9aa2ad 58%,#e8ebef 80%,#b6bcc5 100%);
 -webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent;
 filter:drop-shadow(0 2px 4px rgba(0,0,0,.8))}
.hqgear{background:linear-gradient(180deg,#ffb877 0%,#F26222 40%,#b93c0a 58%,#ff8a3d 80%,#cc4a10 100%);
 -webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent;
 filter:drop-shadow(0 0 12px rgba(242,98,34,.5)) drop-shadow(0 2px 5px rgba(0,0,0,.8))}
.hqtag{font-family:'Archivo Black','Arial Black',Arial,sans-serif;font-size:19px;
 align-self:flex-start;background:linear-gradient(180deg,#fff,#aeb5bf);
 -webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent;
 filter:drop-shadow(0 0 8px rgba(242,98,34,.4))}
.hqline{position:relative;z-index:2;width:84%;margin:11px auto 9px;height:3px;border-radius:2px;
 background:linear-gradient(90deg,transparent,#7a2f08 12%,#F26222 34%,#ffd9ae 50%,#F26222 66%,#7a2f08 88%,transparent);
 box-shadow:0 0 14px 3px rgba(242,98,34,.6),0 0 44px 9px rgba(242,98,34,.3)}
.sub{position:relative;z-index:2;font-size:10.5px;letter-spacing:2px;color:#98A4B4;
 font-weight:700;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.sub b{color:#F26222;font-weight:800}
/* gate */
#gate{padding:46px 6px;text-align:center}
#gate h2{font-size:19px;font-weight:900;margin-bottom:6px}
#gate p{color:var(--dim);font-size:13.5px;margin-bottom:20px;line-height:1.65}
#gate input{width:100%;max-width:300px;background:var(--pnl);border:1.5px solid var(--line);
 color:#fff;font-family:inherit;font-size:19px;text-align:center;letter-spacing:5px;
 padding:15px;border-radius:13px}
#gate input:focus{outline:none;border-color:var(--org)}
#gate button{display:block;width:100%;max-width:300px;margin:11px auto 0;
 background:linear-gradient(135deg,var(--org),#C44C28);color:#fff;border:0;
 font-family:inherit;font-weight:900;font-size:15px;letter-spacing:1.4px;
 padding:15px;border-radius:13px;min-height:50px}
#gerr{color:var(--rd);font-size:13.5px;margin-top:12px;font-weight:700;min-height:20px}
/* board */
.tiles{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:14px}
/* min-width:0 - a grid item defaults to min-width:auto, so a long label
   like FREE ON THE GROUND could push the whole row wider than a 360px
   phone. This lets the tile shrink and the label wrap instead. */
.tile{background:var(--pnl);border:1px solid var(--line);border-radius:12px;
 padding:12px 10px;text-align:center;min-width:0}
.tile b{display:block;font-size:24px;font-weight:900;color:var(--neon);line-height:1.1;
 font-variant-numeric:tabular-nums}
.tile.g b{color:var(--gd)}.tile.a b{color:var(--am)}.tile.r b{color:var(--rd)}
.tile span{display:block;font-size:9.5px;color:var(--dim);font-weight:800;
 letter-spacing:.8px;text-transform:uppercase;margin-top:5px;line-height:1.3}
/* the instrument bridge: the six numbers are CONTROLS, not decoration
   (Andrew's pack - "the overhead figures are also controls"). A tile
   that goes somewhere says so with a chevron; one that has nowhere to
   go stays a plain tile rather than pretending. */
button.tile{width:100%;font:inherit;cursor:pointer;position:relative;
 display:block;border:1px solid var(--line)}
button.tile:after{content:"\\203A";position:absolute;top:7px;right:9px;
 font-size:13px;line-height:1;color:#4E5A6B;font-weight:800}
button.tile:active{background:var(--pnl2)}

/* ---- THE SIX BAYS (Andrew's Store Street pack, 2 Aug 2026) -------
   His six physical destinations, built the way I said I would only
   build them: WITHOUT costing the counter a tap. Each bay is a place
   in the store, but every screen that used to be one tap from home is
   STILL one tap - the bay carries its own doors on its face instead of
   hiding them one level down. Twelve destinations, six bays, one tap.
------------------------------------------------------------------ */
.bays{padding:2px 0 22px}
.youare{display:flex;align-items:center;gap:11px;margin:2px 0 14px;
 padding:11px 14px;border:1px solid var(--line);border-radius:14px;
 background:linear-gradient(160deg,#121A27,#0C121C)}
.youare i{flex:none;width:26px;height:26px;border-radius:50%;
 border:1px solid var(--org);display:flex;align-items:center;
 justify-content:center;color:var(--org)}
.youare i svg{width:14px;height:14px;fill:none;stroke:currentColor;
 stroke-width:1.8;stroke-linecap:round}
.youare small{display:block;font-size:8.5px;font-weight:800;letter-spacing:2.2px;
 color:var(--dim)}
.youare b{display:block;margin-top:3px;font-size:12.5px;letter-spacing:1.4px;
 color:#F5F7FB}
.bay{position:relative;margin-bottom:12px;padding:14px 15px 12px;
 border:1px solid var(--line);border-left:3px solid #2A3547;
 border-radius:0 16px 16px 0;background:linear-gradient(160deg,#121A27,#0C121C)}
.bay.org{border-left-color:var(--org)}
.bay.amb{border-left-color:var(--am)}
.bay.red{border-left-color:var(--rd)}
.bhd{display:flex;gap:12px;align-items:flex-start}
.bic{flex:none;width:42px;height:42px;border-radius:13px;display:flex;
 align-items:center;justify-content:center;background:rgba(242,98,34,.14)}
.bic svg{width:21px;height:21px;fill:none;stroke:var(--org);stroke-width:1.7;
 stroke-linecap:round;stroke-linejoin:round}
/* the photographed bay. Bigger than the line drawing was, because it is
   a picture of the place and not a glyph - it has to be worth looking
   at. Thin steel edge, no orange wash over the top of it. */
.bic.art{width:74px;height:74px;border-radius:15px;overflow:hidden;
 background:#0B111A;border:1px solid #2A3547;padding:0}
.bic.art img{width:100%;height:100%;object-fit:cover;display:block}
.bay.amb .bic{background:rgba(240,180,41,.14)}.bay.amb .bic svg{stroke:var(--am)}
.bay.red .bic{background:rgba(255,90,77,.14)}.bay.red .bic svg{stroke:var(--rd)}
.bmeta{min-width:0;flex:1}
.bmeta small{display:block;font-size:8.5px;font-weight:800;letter-spacing:2.2px;
 color:var(--org)}
.bmeta b{display:block;margin-top:4px;font-size:16.5px;font-weight:800;
 color:#F5F7FB;letter-spacing:-.2px}
.bmeta span{display:block;margin-top:3px;font-size:11.5px;color:var(--dim);
 line-height:1.45}
.bst{display:flex;align-items:center;gap:7px;margin-top:10px;font-size:8.5px;
 font-weight:800;letter-spacing:1.4px;color:var(--dim)}
.bst i{width:7px;height:7px;border-radius:50%;background:#4E5A6B;flex:none}
.bst.g i{background:var(--gd);box-shadow:0 0 8px rgba(53,214,138,.7)}
.bst.a i{background:var(--am);box-shadow:0 0 8px rgba(240,180,41,.7)}
.bst.r i{background:var(--rd);box-shadow:0 0 8px rgba(255,90,77,.7)}
/* the bay's own doors - every one of them still a single tap */
.bdoors{display:flex;flex-wrap:wrap;gap:7px;margin-top:11px}
.bdoors button{display:inline-flex;align-items:center;gap:7px;
 padding:9px 12px;border:1px solid var(--line);border-radius:11px;
 background:#0B111A;color:#DCE3EC;font:800 12px/1 inherit;cursor:pointer}
.bdoors button:active{background:var(--pnl2)}
.bdoors button em{font-style:normal;color:var(--org);font-weight:800}
.bdoors button .bn{background:var(--pnl2);color:#C7CED8;border-radius:7px;
 padding:2px 7px;font-size:11px}
.bdoors button.hot{border-color:var(--rd)}
.bdoors button.hot .bn{background:var(--rd);color:#fff}
.bdoors button.wide{flex:1 1 100%;justify-content:space-between;
 border-color:var(--org);background:#1E1710;font-size:13.5px;padding:12px 14px}
/* tabs */
/* Ten tabs do not fit one row. On a phone they scroll sideways, which is
   the natural thing to do with a thumb - but on a laptop nobody swipes a
   tab strip, so the last tabs (Money among them) sat off the right edge
   where a manager would never find them. Wrap once there is width for it.
   (Caught 29 Jul 2026 in a 1100px screenshot - OUR STANDA... cut in half.) */
/* THE STORE MENU - the board opens on a menu you can read, not a wall
   of fifteen tabs (Andrew, 31 Jul 2026: "easily understood and move
   yourself around easily... find my gear should be easily found").
   One tap in, the MENU button brings you straight back. */
.crumb{display:flex;align-items:center;gap:11px;position:sticky;top:0;z-index:9;
 background:var(--ink);padding:8px 0 10px}
.crumb button{background:var(--org);border:0;color:#fff;font-family:inherit;
 font-weight:900;font-size:12px;letter-spacing:1.2px;padding:11px 15px;
 border-radius:11px;min-height:44px}
.crumb b{color:var(--dim);font-size:13px;font-weight:800;letter-spacing:.8px;
 text-transform:uppercase}
/* the old grouped menu's styling came out with the menu itself -
   every rule in it (.home .hgroup .hbtn) had no markup left to
   style once the bays landed */
.pane{display:none}.pane.on{display:block}
/* the print hub wizard - one step at a time */
.prstep{font-size:11px;font-weight:900;letter-spacing:1.4px;color:var(--org);
 margin:15px 2px 6px;text-transform:uppercase}
.prchips{display:flex;flex-wrap:wrap;gap:8px;margin:4px 0 6px}
.chip{background:var(--pnl);border:1.5px solid var(--line);color:var(--dim);
 font-family:inherit;font-weight:800;font-size:12.5px;padding:11px 14px;
 border-radius:11px;min-height:44px;cursor:pointer}
.chip.on{background:var(--org);border-color:var(--org);color:#fff}
.prq{width:100%;background:var(--pnl);border:1.5px solid var(--line);color:#fff;
 font-family:inherit;font-size:16px;padding:13px 14px;border-radius:12px;
 -webkit-appearance:none;margin:2px 0 8px}
.prq:focus{outline:none;border-color:var(--org)}
/* rows */
.srch{width:100%;background:var(--pnl);border:1.5px solid var(--line);color:#fff;
 font-family:inherit;font-size:16px;padding:13px 15px;border-radius:12px;margin-bottom:12px}
.srch:focus{outline:none;border-color:var(--org)}
.grp{background:var(--pnl);border:1px solid var(--line);border-radius:12px;
 margin-bottom:9px;overflow:hidden}
.grp>button{width:100%;background:none;border:0;color:var(--txt);font-family:inherit;
 text-align:left;padding:13px 14px;display:flex;align-items:center;gap:12px;min-height:58px}
.grp .gn{flex:1;min-width:0}
.grp .gn b{display:block;font-size:14.5px;font-weight:700}
.grp .gn span{display:block;font-size:11px;color:var(--dim);font-weight:700;
 letter-spacing:.4px;text-transform:uppercase;margin-top:2px}
.grp .gq{flex:none;text-align:right}
.grp .gq b{display:block;font-size:17px;font-weight:900;color:var(--gd);line-height:1}
.grp .gq span{display:block;font-size:9.5px;color:var(--dim);font-weight:800;
 letter-spacing:.7px;text-transform:uppercase;margin-top:2px}
.bar{display:flex;height:8px;border-radius:99px;overflow:hidden;background:#0E1319;
 margin:0 14px 12px}
.bar i{display:block;height:100%}
.bar .ba{background:var(--gd)}.bar .bo{background:var(--org)}
.kids{display:none;padding:0 12px 12px}
.kids.on{display:block}
.kfam{font-size:10.5px;font-weight:800;letter-spacing:1px;text-transform:uppercase;
 color:var(--org);border-left:3px solid var(--org);padding:3px 0 3px 8px;
 margin:12px 0 6px}
.kid{background:var(--pnl2);border:1px solid var(--line);border-radius:10px;
 padding:10px 12px;margin-bottom:7px}
.kid .kt{display:flex;gap:10px;align-items:baseline}
.kid .kt b{flex:1;font-size:13.5px;font-weight:700}
.kid .kt em{font-style:normal;font-size:12px;font-weight:800;color:var(--gd)}
.kid .kt em.o{color:var(--org)}
.kid .kw{font-size:11.5px;color:var(--dim);margin-top:5px;line-height:1.6}
.kid .kw b{color:#C7CED8;font-weight:700}
/* compliance chips + the variant code (the photo key, worth knowing) */
.cchips{margin-top:6px}
.cchip{display:inline-block;color:#fff;border-radius:6px;padding:2.5px 9px;
 font-size:10px;font-weight:800;letter-spacing:.4px;margin:0 6px 4px 0}
.cchip.lbk{background:#3A2E08;color:#F5C032;border:1px solid #6b551b}
.vcode{font-family:Consolas,Menlo,monospace;font-size:10.5px;color:#8A97A8}
/* the finder's answer cards - shelf green, out orange, hunt red */
.fcard{display:flex;gap:12px;align-items:center;background:var(--pnl);
 border:1px solid var(--line);border-left:5px solid var(--gd);
 border-radius:14px;padding:14px;padding-right:84px;margin-bottom:10px;
 position:relative}
.fcard.fo{border-left-color:var(--org)}
.fcard.fw{border-left-color:var(--rd)}
.fcard .fbody{flex:1;min-width:0}
.fhead{font-size:11px;font-weight:900;letter-spacing:1.2px}
.fcard.fa .fhead{color:var(--gd)}
.fcard.fo .fhead{color:var(--org)}
.fcard.fw .fhead{color:var(--rd)}
.fname{font-size:15px;font-weight:800;color:#fff;margin:3px 0}
/* the catalogue picture tile - photo when collected, monogram until */
.kid.kidth{display:flex;gap:11px;align-items:flex-start}
.kid.kidth .kbody{flex:1;min-width:0}
/* stocktake v2 - area cards, freshness chips, ghost list (31 Jul 2026) */
.ucard{display:block;width:100%;text-align:left;background:var(--pnl);
 border:1px solid var(--line);border-radius:14px;padding:12px 14px;
 margin:8px 0;cursor:pointer;font-family:inherit;color:var(--txt)}
.ucard:active{border-color:var(--org)}
.urow{display:flex;justify-content:space-between;align-items:center;gap:10px}
.urow .un{min-width:0}
.urow .un b{font-size:15px;display:block}
.urow .un span{font-size:12px;color:var(--dim)}
.upct{font-size:21px;font-weight:900;flex:none}
.ubar{height:8px;background:#20262e;border-radius:5px;overflow:hidden;margin-top:9px}
.ubar i{display:block;height:100%}
.uwarn{color:var(--rd);font-size:11.5px;font-weight:700;margin-top:6px}
.stchips{display:flex;flex-wrap:wrap;gap:8px;margin:12px 0}
.chipk{background:var(--pnl);border:1.5px solid var(--line);border-radius:999px;
 padding:8px 13px;font-family:inherit;color:var(--txt);font-size:12.5px;
 font-weight:700;cursor:pointer}
.chipk b{margin-left:5px}
.chipk.on{border-color:var(--org);background:rgba(242,98,34,.16)}
.chipk:disabled{opacity:.35;cursor:default}
.ghp{display:inline-block;background:var(--rd);color:#fff;border-radius:999px;
 font-size:10.5px;font-weight:800;padding:2px 9px;letter-spacing:.5px}
/* the call cards - folded how-to at the top of every tab (31 Jul 2026) */
.hgrp button .gn b{color:var(--org)}
.hcard{padding:4px 2px 8px}
.hcard .hs{margin:10px 0}
.hcard .hs>b{display:block;font-size:11px;letter-spacing:2px;color:var(--dim);
 margin-bottom:4px}
.hcard .hs span{font-size:14px;line-height:1.6;color:var(--txt)}
.hcard .hs ol{margin:0 0 0 20px;font-size:14px;line-height:1.65;color:var(--txt)}
.hcard .hs ol li{margin-bottom:5px}
/* 132px. ONE PICTURE SIZE ACROSS THE WHOLE SUITE. (Andrew,
   3 Aug 2026: "needs to be fixed in stores and workers".)
   The store board ran 132, the stores board 112 and the worker
   page 100 - three sizes for the same photo of the same tool,
   so the gear changed size as a bloke moved between screens.
   132 is the one Andrew pointed at. If this ever moves, it moves
   on all four: mygear_store, mygear_stores, BUILD_MY_GEAR and
   build_fleet_detail. */
.kth2{flex:none;width:132px;height:132px;border-radius:14px;overflow:hidden;
 background:#20262e;display:flex;align-items:center;justify-content:center}
.kth2 img{width:100%;height:100%;object-fit:cover;display:block}
.kth2.mono{color:#8A97A8;font-weight:900;font-size:23px;letter-spacing:.5px}
/* the on-screen scan sticker: white QR tile pinned to the row's right,
   room reserved with padding so text never runs underneath it */
.kid.hasqr{position:relative;padding-right:80px}
.kqr{position:absolute;right:9px;top:50%;transform:translateY(-50%);line-height:0}
.kqr svg{border-radius:5px;display:block}
.kw.kwq{display:flex;align-items:center;gap:10px}
.kw.kwq span{flex:1;min-width:0}
.kw.kwq svg{flex:none;border-radius:4px}
/* .stmore was being used on the plant toggle but never defined here, so it
   rendered as a grey system button in the middle of a Coates page.
   (Caught 29 Jul 2026 in a screenshot.) */
.stmore{background:var(--pnl2);border:1.5px solid var(--org);color:var(--org);
 font-family:inherit;font-weight:800;font-size:12px;letter-spacing:.5px;
 padding:11px 16px;border-radius:999px;min-height:44px;text-transform:uppercase;
 cursor:pointer}
.stmore:hover{background:var(--org);color:#fff}
.kw.cut{color:var(--dim);font-style:italic;padding:9px 2px 2px}
/* ENTRANCE MOTION - one second of life when the board unlocks, then
   everything sits still. Working screens hold still; only the unlock
   moves, and the hit-list pulse below stays the single looping thing
   on the page so it keeps its meaning. Bars inside a hidden pane run
   their entrance the first time that pane is opened - a display:none
   element does not start animating until it is shown.
   (Andrew, 29 Jul 2026: "proceed".) */
@media (prefers-reduced-motion: no-preference){
 .bar i,.bb i{transform-origin:left;animation:growx .7s ease-out both}
 .brow .bb i{animation-duration:.55s}
 .score .sc b{animation:bpunch .55s cubic-bezier(.2,1.6,.4,1) both}
 .score .sc:last-child b{animation-delay:.12s}
 .tiles .tile{animation:tfup .4s ease both}
 .tiles .tile:nth-child(2){animation-delay:.05s}
 .tiles .tile:nth-child(3){animation-delay:.1s}
 .tiles .tile:nth-child(4){animation-delay:.15s}
 .tiles .tile:nth-child(5){animation-delay:.2s}
 .tiles .tile:nth-child(6){animation-delay:.25s}
}
@keyframes growx{from{transform:scaleX(0)}to{transform:scaleX(1)}}
@keyframes bpunch{0%{opacity:0;transform:scale(.4)}
 70%{transform:scale(1.12)}100%{opacity:1;transform:none}}
@keyframes tfup{from{opacity:0;transform:translateY(7px)}
 to{opacity:1;transform:none}}
@media print{.bar i,.bb i,.score .sc b,.tiles .tile{animation:none!important}}
/* the hit-list menu button glows - it is the one that means walk somewhere */
/* print & send */
.prpick{display:flex;gap:9px;flex-wrap:wrap;margin-bottom:6px}
.prbtns{display:flex;gap:9px;flex-wrap:wrap;margin:4px 0 13px}
a.stmore{display:inline-block;text-decoration:none;line-height:1.55}
select.srch{appearance:none;-webkit-appearance:none}
/* THE PRINT SHEET. Hidden on screen; when the html element carries
   class "pr" and the print dialog opens, the board disappears and only
   this white Coates A4 page goes to the printer. Print colours are
   forced so the orange rule and the late-day marks survive "save ink"
   defaults. */
#prsheet{display:none}
@media print{
 html.pr header,html.pr #gate,html.pr #app,html.pr .wrap{display:none!important}
 html.pr body{background:#fff}
 html.pr #prsheet{display:block;background:#fff;color:#14181F;
  font-family:"Segoe UI",Roboto,Helvetica,Arial,sans-serif;font-size:12.5px;
  -webkit-print-color-adjust:exact;print-color-adjust:exact}
 #prsheet .phead{display:flex;justify-content:space-between;align-items:flex-start;
  border-top:5px solid #F26222;padding:12px 2px 10px;margin-bottom:6px}
 #prsheet .pbrand{font-size:21px;font-weight:800;letter-spacing:.5px;color:#14181F}
 #prsheet .pbrand b{color:#F26222}
 #prsheet .pbrand span{display:block;font-size:9.5px;letter-spacing:2px;
  color:#8A94A2;font-weight:700;margin-top:2px}
 #prsheet .pmeta{text-align:right;font-size:10.5px;color:#5B6472;line-height:1.6}
 #prsheet .ptitle{font-size:18px;font-weight:800;margin:6px 2px 2px}
 #prsheet .psub{font-size:11px;color:#5B6472;margin:0 2px 12px}
 #prsheet .pwho{font-weight:800;font-size:12.5px;margin:14px 2px 4px;
  padding-left:8px;border-left:3px solid #F26222;page-break-after:avoid}
 #prsheet .pwho span{color:#8A94A2;font-weight:700;font-size:10.5px}
 #prsheet .ptab{width:100%;border-collapse:collapse;page-break-inside:auto}
 #prsheet .ptab tr{page-break-inside:avoid}
 #prsheet .ptab th{text-align:left;font-size:9px;text-transform:uppercase;
  letter-spacing:.8px;color:#8A94A2;padding:4px 8px;border-bottom:1px solid #D5DBE3}
 #prsheet .ptab td{padding:5px 8px;border-bottom:1px solid #EDF0F4;vertical-align:top}
 #prsheet .ptab .pn{text-align:right;white-space:nowrap;font-variant-numeric:tabular-nums}
 #prsheet .ptab .pn.late{color:#C1440E;font-weight:800}
 /* write-in furniture: ruled boxes, tick squares and signature lines
    for the clipboard sheets - drawn, not typed, because the whole
    point of these pages is a pen */
 #prsheet .pbox{display:inline-block;width:17mm;height:6.5mm;
  border:1.2px solid #14181F;border-radius:2px;vertical-align:middle}
 #prsheet .ptick{display:inline-block;width:5.5mm;height:5.5mm;
  border:1.2px solid #14181F;border-radius:2px;vertical-align:middle}
 #prsheet .pline{display:inline-block;width:52mm;
  border-bottom:1.2px solid #14181F;height:5.5mm;vertical-align:baseline}
 #prsheet .pline.short{width:26mm}
 #prsheet .pchk td{padding:6px 8px}
 #prsheet .pintr{background:#FDECE7;border:1.5px solid #F26222;color:#C1440E;
  border-radius:8px;padding:7px 12px;margin:0 0 12px;font-size:10.5px;
  font-weight:800;letter-spacing:.8px}
 #prsheet .ppid{display:inline-block;background:#F26222;color:#fff;
  border-radius:8px;padding:1px 7px;font-size:9px;font-weight:800;
  margin-left:5px;vertical-align:1px}
 /* the scan column: the QR IS the item number - sized so a store
    scanner or a phone reads it off paper first go (12.5mm with its
    own quiet zone), crisp SVG so it prints razor sharp */
 #prsheet .pqr{width:14mm;padding:3px 4px}
 #prsheet .pqr svg{width:12.5mm;height:12.5mm;display:block}
 #prsheet .hcv{border:2.5px solid #F26222;border-radius:14px;
  padding:14px 18px;margin-bottom:9mm;page-break-inside:avoid}
 #prsheet .hct{font-size:19px;font-weight:900;color:#F26222;margin-bottom:6px}
 #prsheet .hcs{font-size:12px;line-height:1.62;margin:7px 0;color:#14181F}
 #prsheet .hcs>b{letter-spacing:1.5px;font-size:9.5px;color:#5B6472;display:block}
 #prsheet .hcs ol{margin:3px 0 0 18px}
 #prsheet .hcs ol li{margin-bottom:3px}
 #prsheet .pfoot{margin-top:16px;padding-top:9px;border-top:1px solid #D5DBE3;
  text-align:center;font-size:9.5px;color:#8A94A2;line-height:1.7}
}
@page{size:A4;margin:12mm}
.where{font-size:11px;color:var(--neon);font-weight:800;letter-spacing:.5px;
 text-transform:uppercase;margin-top:5px}
.uhead{font-size:12px;font-weight:900;letter-spacing:1.2px;text-transform:uppercase;
 color:var(--org);margin:16px 0 8px;padding-left:9px;border-left:3px solid var(--org)}
.note{background:var(--pnl);border:1px solid var(--line);border-left:4px solid var(--org);
 border-radius:12px;padding:12px 14px;font-size:13px;color:#C7CED8;line-height:1.65;
 margin-bottom:14px}
.note b{color:#fff}
.ring{display:flex;align-items:center;gap:16px;background:var(--pnl);
 border:1px solid var(--line);border-radius:14px;padding:16px;margin-bottom:14px}
.ring .rv{font-size:40px;font-weight:900;color:var(--gd);line-height:1}
.ring .rt{flex:1;font-size:13px;color:var(--dim);line-height:1.6}
.ring .rt b{display:block;color:#fff;font-size:15px;font-weight:800;margin-bottom:3px}
.foot{text-align:center;color:#6E7A8A;font-size:11px;padding:24px 8px;line-height:1.8}
/* battle */
.score{display:flex;align-items:stretch;gap:10px;margin-bottom:14px}
.sc{flex:1;background:var(--pnl);border:1px solid var(--line);border-radius:14px;
 padding:14px 10px;text-align:center}
.sc.day{border-color:#F5A623}.sc.night{border-color:#2E9BF0}
.sc b{display:block;font-size:34px;font-weight:900;line-height:1;
 font-variant-numeric:tabular-nums}
.sc.day b{color:#F5A623}.sc.night b{color:#2E9BF0}
.sc span{display:block;font-size:10px;color:var(--dim);font-weight:800;
 letter-spacing:1px;text-transform:uppercase;margin-top:6px}
.sc em{display:block;font-style:normal;font-size:11.5px;color:#C7CED8;
 font-weight:700;margin-top:4px}
.vs{display:flex;align-items:center;font-size:13px;font-weight:900;color:var(--dim)}
.brow{background:var(--pnl);border:1px solid var(--line);border-radius:12px;
 padding:11px 13px;margin-bottom:8px}
.brow .bd{display:flex;justify-content:space-between;align-items:baseline;
 font-size:12.5px;font-weight:800;margin-bottom:7px}
.brow .bd .win{font-size:10.5px;font-weight:900;letter-spacing:.8px;
 text-transform:uppercase;padding:3px 8px;border-radius:99px}
.win.d{background:#F5A623;color:#2a1e05}.win.n{background:#2E9BF0;color:#04223b}
.win.t{background:#3A4553;color:#C7CED8}.win.x{background:none;color:#6E7A8A}
.bb{display:flex;height:22px;border-radius:7px;overflow:hidden;background:#0E1319}
.bb i{display:flex;align-items:center;justify-content:center;font-size:10.5px;
 font-weight:900;color:#151A22;min-width:0}
.bb .bd2{background:linear-gradient(90deg,#F5A623,#e0940f)}
.bb .bn{background:linear-gradient(90deg,#2E9BF0,#1c7fcc);color:#fff}
.bnums{display:flex;justify-content:space-between;font-size:10.5px;color:var(--dim);
 font-weight:700;margin-top:5px}
/* standards */
.std{background:var(--pnl);border:1px solid var(--line);border-left:4px solid var(--org);
 border-radius:12px;padding:14px 15px;margin-bottom:10px}
.std h3{font-size:14.5px;font-weight:900;margin-bottom:7px;letter-spacing:-.2px}
.std p{font-size:13.5px;color:#C7CED8;line-height:1.7}
.std p.ref{font-size:12px;color:var(--dim);margin-top:10px;
 border-left:2px solid var(--org);padding-left:10px}
.std p.ref b{color:#C7CED8}
.std p b{color:#fff}
.sop{background:var(--pnl2);border:1px solid var(--line);border-radius:10px;
 padding:11px 13px;margin-top:9px}
.sop .n{display:inline-block;background:var(--org);color:#fff;font-size:11px;
 font-weight:900;padding:2px 9px;border-radius:99px;margin-bottom:6px}
.sop>b{display:block;font-size:13.5px;margin-bottom:4px}
/*  only the SOP's own title is a block - a nested bold inside the
    body must stay inline or the procedure reads as a column of
    fragments instead of a sentence  */
.sop span b{display:inline;color:#fff}
.sop span{font-size:12.5px;color:var(--dim);line-height:1.65}
.rule{display:flex;gap:10px;align-items:flex-start;padding:9px 0;
 border-top:1px solid var(--line);font-size:13px;color:#C7CED8;line-height:1.6}
.rule:first-of-type{border-top:0}
.rule i{flex:none;width:7px;height:7px;border-radius:50%;background:var(--org);
 margin-top:7px}
/*  THE MASTER SEARCH (2 Aug 2026)  */
.wcard{display:block;width:100%;text-align:left;margin-bottom:8px;
 padding:12px 14px;border:1px solid var(--line);border-radius:14px;
 background:var(--pnl);color:#DCE3EC;font:inherit;cursor:pointer}
.wcard:active{background:var(--pnl2)}
.wtop{display:flex;align-items:center;justify-content:space-between;gap:10px}
.wtop b{font-size:14.5px;font-weight:800;color:#F5F7FB;min-width:0}
.wtop em{flex:none;font-style:normal;font-weight:800;font-size:12px;
 color:var(--org);background:#1E1710;border:1px solid var(--org);
 border-radius:9px;padding:3px 9px}
.wsub{margin-top:4px;font-size:11.5px;color:var(--dim)}
.wbs{display:flex;flex-wrap:wrap;gap:6px;margin-top:8px}
.wb{font-size:9px;font-weight:800;letter-spacing:1px;padding:4px 9px;
 border-radius:999px;border:1px solid var(--line);color:var(--dim)}
.wb.a{border-color:var(--am);color:var(--am)}
.wb.r{border-color:var(--rd);color:var(--rd)}
.wback{margin:2px 0 12px;padding:9px 14px;border:1px solid var(--line);
 border-radius:999px;background:var(--pnl);color:#C7CED8;font:inherit;
 font-size:12px;font-weight:800;letter-spacing:.6px;cursor:pointer}
.wprof{padding:15px 16px;border:1px solid var(--line);
 border-left:3px solid var(--org);border-radius:0 16px 16px 0;
 background:linear-gradient(160deg,#121A27,#0C121C)}
.wpn{font-size:20px;font-weight:800;color:#F5F7FB;letter-spacing:-.3px;
 line-height:1.2}
.wpc{margin-top:4px;font-size:12px;color:var(--dim)}
.wnums{display:grid;grid-template-columns:repeat(3,1fr);gap:7px;margin-top:13px}
.wnums span{padding:10px 5px;text-align:center;border:1px solid var(--line);
 border-radius:11px;background:#0B111A;font-size:8px;font-weight:800;
 letter-spacing:.7px;color:var(--dim)}
.wnums b{display:block;margin-bottom:3px;font-size:19px;color:#F5F7FB}
.wnums .g b{color:var(--gd)}.wnums .r b{color:var(--rd)}
.wage{display:flex;height:10px;margin-top:12px;border-radius:6px;
 overflow:hidden;border:1px solid var(--line);background:var(--pnl)}
.wage i{display:block}
.wage .g{background:var(--gd)}.wage .a{background:var(--am)}
.wage .r{background:var(--rd)}
.wlg{display:flex;flex-wrap:wrap;gap:11px;margin-top:8px;font-size:10.5px;
 color:var(--dim)}
.wlg i{display:inline-block;width:8px;height:8px;border-radius:50%;
 margin-right:5px;background:#4E5A6B}
.wlg .g{background:var(--gd)}.wlg .a{background:var(--am)}
.wlg .r{background:var(--rd)}
.wold{margin-top:10px;font-size:12px;color:var(--dim)}
.wold b{color:#F5F7FB}
.wunits{display:flex;flex-wrap:wrap;gap:6px;margin-bottom:10px}
.wunits span{font-size:11px;color:var(--dim);border:1px solid var(--line);
 border-radius:9px;padding:6px 10px;background:var(--pnl)}
.wunits b{color:#F5F7FB}
/*  a name you can walk through  */
.wlink{color:var(--org);font-weight:800;cursor:pointer;
 border-bottom:1px dotted rgba(242,98,34,.6)}
/*  a consumable you can open  */
.kid.tapme{cursor:pointer}
.kid.tapme:active{background:var(--pnl2)}
.txn{color:var(--org)}
.txh{padding:14px 15px;border:1px solid var(--line);
 border-left:3px solid var(--org);border-radius:0 16px 16px 0;
 background:linear-gradient(160deg,#121A27,#0C121C);margin-bottom:14px}
.txn1{font-size:17px;font-weight:800;color:#F5F7FB;line-height:1.25}
.txn2{margin-top:4px;font-size:11.5px;color:var(--dim)}
.txr{display:flex;align-items:center;gap:10px;padding:10px 12px;
 border:1px solid var(--line);border-radius:12px;background:var(--pnl);
 margin-bottom:7px}
.txd{flex:none;width:88px;font-size:11px;color:var(--dim);
 font-variant-numeric:tabular-nums}
.txw{flex:1;min-width:0;font-size:12.5px;color:#DCE3EC}
.txw span{display:block;font-size:10.5px;color:var(--dim);margin-top:2px}
.txq{flex:none;font-weight:800;font-size:13px;color:var(--am);
 font-variant-numeric:tabular-nums}
.txq.back{color:var(--gd)}
/*  the light sheet  */
.shs{position:fixed;inset:0;z-index:96;display:none}
.shs.on{display:block}
.shbg{position:absolute;inset:0;background:rgba(3,6,11,.72);
 -webkit-backdrop-filter:blur(3px);backdrop-filter:blur(3px)}
.shcard{position:absolute;left:0;right:0;bottom:0;max-height:88%;
 display:flex;flex-direction:column;border-top:1px solid var(--line);
 border-radius:20px 20px 0 0;background:#0A0F16;
 box-shadow:0 -18px 44px rgba(0,0,0,.6);
 animation:shup .32s cubic-bezier(.16,1,.3,1) both}
@keyframes shup{from{transform:translateY(14%);opacity:0}
 to{transform:none;opacity:1}}
.shhd{display:flex;align-items:center;justify-content:space-between;gap:10px;
 padding:14px 16px;border-bottom:1px solid var(--line)}
.shhd b{font-size:11px;font-weight:800;letter-spacing:2.2px;color:var(--org);
 text-transform:uppercase}
.shhd button{padding:9px 15px;border:1px solid var(--line);border-radius:999px;
 background:var(--pnl);color:#C7CED8;font:inherit;font-size:12px;
 font-weight:800;cursor:pointer}
.shbd{overflow-y:auto;-webkit-overflow-scrolling:touch;
 padding:14px 16px calc(20px + env(safe-area-inset-bottom))}
@media (min-width:900px){.shcard{left:50%;right:auto;bottom:0;
 width:min(760px,94vw);transform:translateX(-50%);
 border-radius:20px 20px 0 0}
 @keyframes shup{from{transform:translateX(-50%) translateY(14%);opacity:0}
  to{transform:translateX(-50%);opacity:1}}}
/*  DAYS FROM FLAME OFF (Andrew, 2 Aug 2026). Nobody on a shutdown
    thinks in calendar dates - they think "we're on day 9". A named day
    wears the orange; an ordinary one stays quiet steel so the named
    ones actually stand out. */
.fday{margin-top:9px;display:inline-flex;align-items:center;gap:8px;
 padding:7px 13px;border:1px solid var(--line);border-radius:999px;
 background:var(--pnl);color:var(--dim);font-size:10px;font-weight:800;
 letter-spacing:1.8px;text-transform:uppercase}
.fday.big{border-color:var(--org);color:#FFC7A6;background:#1E1710}
.fday b{color:#F5F7FB}
.fday.big b{color:var(--org)}

/* ---- STORE STREET (Andrew's pack, 2 Aug 2026) --------------------
   His own words in the README: "This replaces the roller-door idea
   with a believable command-centre arrival." So the shutter is gone
   from this page and the counter walks in instead.

   The code is accepted FIRST - nothing here is a security step. Then
   the room powers on in stages, exactly as he timed it:

     0.00s  the access console says it is bringing Store Street online
     0.42s  six ceiling circuits energise in sequence, near to far
     0.45s  the orange floor route lights from the threshold to the
            counter, one section at a time
     0-5.9s the sharp room plate scales slowly - a walking pace, not a
            zoom. Nothing is blurred.
     5.90s  the board takes the screen

   Two plates, not one squeezed: a portrait street for a phone and a
   landscape one for the counter laptop. 117 KB and 130 KB - and it is
   a <picture> with a media source, NOT two images with one hidden. A
   display:none image still downloads; the rig proved both plates were
   arriving on the laptop before this was changed.
------------------------------------------------------------------ */
#sdoor{position:fixed;inset:0;z-index:90;display:none;background:#03060B;
 overflow:hidden}
#sdoor.on{display:block}
.sd-plate,.sd-black,.sd-vig,.sd-circ,.sd-floor{position:absolute;inset:0}
.sd-plate{overflow:hidden}
.sd-plate picture{position:absolute;inset:0;display:block}
.sd-plate img{width:100%;height:100%;object-fit:cover;object-position:center;
 transform:scale(1.015);transform-origin:50% 58%;user-select:none;
 filter:saturate(.82) brightness(.46) contrast(1.06)}
#sdoor.go .sd-plate img{animation:sd-walk 5.9s cubic-bezier(.2,.67,.26,1) both}
.sd-black{background:rgba(0,2,6,.64)}
#sdoor.go .sd-black{animation:sd-power 5.4s ease-out both}
.sd-vig{pointer-events:none;background:
 radial-gradient(circle at 50% 50%,transparent 32%,rgba(0,0,0,.48) 100%),
 linear-gradient(90deg,rgba(0,0,0,.46),transparent 30%,transparent 70%,
  rgba(0,0,0,.46))}
/* six ceiling circuits, waking near to far */
.sd-circ,.sd-floor{pointer-events:none}
.sd-circ i{position:absolute;top:0;left:calc(18% + var(--n,0) * 11%);width:4px;
 height:27%;opacity:0;transform:rotate(15deg);
 background:linear-gradient(180deg,#fff,rgba(255,255,255,0));
 box-shadow:0 0 22px rgba(255,255,255,.8),0 0 52px rgba(255,255,255,.34)}
.sd-circ i:nth-child(1){--n:0}.sd-circ i:nth-child(2){--n:1}
.sd-circ i:nth-child(3){--n:2}.sd-circ i:nth-child(4){--n:3}
.sd-circ i:nth-child(5){--n:4}.sd-circ i:nth-child(6){--n:5}
#sdoor.go .sd-circ i{animation:sd-circuit 2.2s calc(var(--n) * .42s) ease-out both}
/* the orange floor route, lighting from the threshold to the counter */
.sd-floor{left:34%;right:34%;top:auto;bottom:0;height:58%;perspective:600px}
.sd-floor i{position:absolute;left:50%;bottom:calc(6% + var(--n,0) * 14%);
 width:calc(15rem - var(--n,0) * 1.65rem);max-width:86vw;height:3px;opacity:0;
 transform:translateX(-50%) rotateX(64deg);
 background:linear-gradient(90deg,transparent,#F26222,#FFD2BD,#F26222,transparent);
 box-shadow:0 0 16px rgba(244,90,27,.92)}
.sd-floor i:nth-child(1){--n:0}.sd-floor i:nth-child(2){--n:1}
.sd-floor i:nth-child(3){--n:2}.sd-floor i:nth-child(4){--n:3}
.sd-floor i:nth-child(5){--n:4}.sd-floor i:nth-child(6){--n:5}
#sdoor.go .sd-floor i{animation:sd-fl 1.3s calc(.45s + var(--n) * .48s) ease-out both}
/* the access console - in the board's own materials, not the pack's */
.sd-con{position:absolute;z-index:6;left:16px;right:16px;bottom:5%;
 padding:20px 20px 18px;border:1px solid var(--line);border-left:3px solid var(--org);
 border-radius:0 18px 18px 0;background:linear-gradient(160deg,#121A27,#0C121C);
 box-shadow:0 22px 44px rgba(0,0,0,.62)}
.sd-con .ok{display:flex;align-items:center;gap:8px;font:800 10px/1 inherit;
 letter-spacing:2.2px;color:#35D68A}
.sd-con .ok i{width:8px;height:8px;border-radius:50%;background:#35D68A;
 box-shadow:0 0 10px rgba(53,214,138,.8)}
.sd-con .eyeb{margin-top:14px;font-size:9.5px;font-weight:800;letter-spacing:2.4px;
 color:#F26222}
.sd-con b{display:block;margin-top:7px;font-size:22px;font-weight:800;
 color:#F5F7FB;letter-spacing:-.4px;line-height:1.15}
.sd-con span{display:block;margin-top:7px;font-size:12.5px;color:#8794A6;
 line-height:1.55}
.sd-bar{position:relative;height:3px;margin-top:16px;border-radius:3px;
 background:#1A2331;overflow:hidden}
.sd-bar:after{content:"";position:absolute;inset:0;transform-origin:left;
 transform:scaleX(0);background:linear-gradient(90deg,#F26222,#FFA24D)}
#sdoor.go .sd-bar:after{animation:sd-prog 5.9s linear both}
.sd-steps{margin-top:9px;font-size:9px;font-weight:800;letter-spacing:2.2px;
 color:#5A6472}
/* on the counter laptop the console sits to the left at eye height, the
   way he drew it - stretched across a 1400px street it read as a banner */
@media (min-width:900px){
 .sd-con{left:clamp(34px,7vw,124px);right:auto;bottom:auto;top:50%;
  width:min(520px,44vw);transform:translateY(-50%);padding:30px 32px 26px}
 .sd-con b{font-size:30px}
 .sd-con span{font-size:14px}}
.sd-skip{position:absolute;z-index:12;right:14px;
 top:14px;top:calc(14px + env(safe-area-inset-top));
 background:rgba(8,13,20,.78);border:1px solid #2A3547;color:#DCE3EC;
 font:inherit;font-size:12px;font-weight:800;letter-spacing:1.4px;
 border-radius:999px;padding:12px 18px;cursor:pointer;
 -webkit-backdrop-filter:blur(6px);backdrop-filter:blur(6px)}
#sdoor.out{animation:sd-handoff .45s ease both}
/* the board rises as the room lets go - a hand-off, not a cut */
#app.sdgrow{animation:sd-arrive .5s cubic-bezier(.16,1,.3,1) both}
@keyframes sd-walk{
 0%{transform:scale(1.015) translateY(0);
  filter:saturate(.78) brightness(.44) contrast(1.06)}
 18%{transform:scale(1.025) translateY(0);
  filter:saturate(.84) brightness(.54) contrast(1.05)}
 68%{transform:scale(1.09) translateY(.8%);
  filter:saturate(.9) brightness(.75) contrast(1.04)}
 100%{transform:scale(1.145) translateY(1.4%);
  filter:saturate(.92) brightness(.84) contrast(1.03)}}
@keyframes sd-power{0%{opacity:1}28%{opacity:.78}68%{opacity:.35}
 100%{opacity:.08}}
@keyframes sd-circuit{0%{opacity:0}14%{opacity:1}22%{opacity:.26}
 31%,100%{opacity:.85}}
@keyframes sd-fl{0%{opacity:0;transform:translateX(-50%) rotateX(64deg) scaleX(.2)}
 28%,100%{opacity:1;transform:translateX(-50%) rotateX(64deg) scaleX(1)}}
@keyframes sd-prog{from{transform:scaleX(0)}to{transform:scaleX(1)}}
@keyframes sd-handoff{from{opacity:1}to{opacity:0}}
@keyframes sd-arrive{from{opacity:0;transform:translateY(12px)}
 to{opacity:1;transform:none}}
/* THE DOCK. The board is long and the counter reads it one-handed, so
   the five places they go all shift live at the bottom of the screen.
   Nothing was taken off the home screen to pay for it - every tile and
   every section is still exactly where it was. */
#sdock{position:fixed;left:0;right:0;bottom:0;z-index:60;display:none;
 grid-template-columns:repeat(5,1fr);gap:2px;
 padding:6px 6px calc(6px + env(safe-area-inset-bottom));
 border-top:1px solid var(--line);background:rgba(9,14,21,.94);
 -webkit-backdrop-filter:blur(10px);backdrop-filter:blur(10px)}
#sdock.on{display:grid}
/* on the counter laptop the dock lines up with the board above it
   rather than sprawling across 1400px of empty screen */
@media (min-width:900px){#sdock{max-width:820px;margin:0 auto;
 border-left:1px solid var(--line);border-right:1px solid var(--line);
 border-radius:14px 14px 0 0}}
#sdock button{display:flex;flex-direction:column;align-items:center;gap:4px;
 padding:8px 2px 6px;border:0;border-radius:12px;background:transparent;
 color:#8794A6;font:800 8.5px/1 inherit;letter-spacing:1.2px;cursor:pointer}
#sdock button svg{width:19px;height:19px;fill:none;stroke:currentColor;
 stroke-width:1.8;stroke-linecap:round;stroke-linejoin:round}
#sdock button:active{background:#111A27}
#sdock button.on{color:#F26222}
body.hasdock{padding-bottom:64px}
@media (prefers-reduced-motion: reduce){#sdoor{display:none!important}
 #app.sdgrow{animation:none!important}}
</style></head><body>
<div class="wrap">
<header>
<div class="hqmast"><span class="hqmy">MY</span><span class="hqgear">GEAR</span><span class="hqtag">HQ</span></div>
<div class="hqline"></div>
<div class="sub"><b>COATES</b> STORES TEAM &middot; K2 &middot; __ASOF__</div>
__DAYTAG__
</header>

<div id="gate">
  <h2>Coates stores staff</h2>
  <p>Enter the store code to open the board.<br>
  This page is for the team behind the counter.</p>
  <input id="code" type="password" inputmode="text" autocomplete="off"
         autocapitalize="none" autocorrect="off" spellcheck="false"
         placeholder="CODE" aria-label="Store code">
  <button onclick="unlock()" type="button">OPEN THE BOARD</button>
  <div id="gerr"></div>
  <div style="margin-top:26px;font-size:9.5px;letter-spacing:2px;color:#5A6472;
   font-weight:700">MY GEAR HQ &middot; POWERED BY SITEIQ &middot;
   DESIGNED &amp; BUILT BY ANDREW FISHER</div>
</div>

<div id="app" style="display:none"></div>
</div>
<div id="sdoor">
 <div class="sd-plate">
  <picture>
   <source media="(min-width:900px)" srcset="art/store-street-desktop.webp">
   <img src="art/store-street-mobile.webp" alt="" onerror="sdBail()">
  </picture>
 </div>
 <div class="sd-black"></div>
 <div class="sd-circ"><i></i><i></i><i></i><i></i><i></i><i></i></div>
 <div class="sd-floor"><i></i><i></i><i></i><i></i><i></i><i></i></div>
 <div class="sd-vig"></div>
 <div class="sd-con">
  <div class="ok"><i></i>CODE ACCEPTED</div>
  <div class="eyeb">COATES STORES COMMAND CENTRE</div>
  <b>Bringing Store Street online</b>
  <span>Lighting the route, waking each work bay and linking this
   morning&rsquo;s SiteIQ snapshot.</span>
  <div class="sd-bar"></div>
  <div class="sd-steps">WALKWAY &middot; BAYS &middot; INSTRUMENTS &middot; DATA</div>
 </div>
 <button type="button" class="sd-skip" onclick="sdSkip()">SKIP WALK-IN</button>
</div>
<div id="sdock">
 <button type="button" onclick="dock('home')" data-d="home"><svg viewBox="0 0 24 24"><path d="m4 11 8-7 8 7M6 10v10h12V10"/></svg>HOME</button>
 <button type="button" onclick="dock('find')" data-d="find"><svg viewBox="0 0 24 24"><circle cx="10.5" cy="10.5" r="5.7"/><path d="m15 15 4.5 4.5"/></svg>FIND</button>
 <button type="button" onclick="dock('chase')" data-d="chase"><svg viewBox="0 0 24 24"><path d="M7.5 14.5h9l-1-7h-7zM6 17h12M9 4.5l-1-2M15 4.5l1-2"/></svg>HUNT</button>
 <button type="button" onclick="dock('groups')" data-d="groups"><svg viewBox="0 0 24 24"><path d="M4 4v16M20 4v16M4 8h16M4 14h16M7 5.5h3M13 5.5h4M7 11h5"/></svg>FLOOR</button>
 <button type="button" onclick="dock('print')" data-d="print"><svg viewBox="0 0 24 24"><path d="M7 9V4h10v5M7 17H4V9h16v8h-3M7 14h10v6H7z"/></svg>PRINT</button>
</div>
<div id="prsheet"></div>
<script>
var PAYLOAD="__PAYLOAD__",TAG="__TAG__",ASOF="__ASOF__";
var ATAG="__ATAG__",AKEY="__AKEY__";
var MPAY="__MPAYLOAD__",MTAG="__MTAG__",MKEY="__MKEY__";
var MGR=null,SHOW_PLANT=true;
function xmur3(s){for(var i=0,h=1779033703^s.length;i<s.length;i++){
 h=Math.imul(h^s.charCodeAt(i),3432918353);h=h<<13|h>>>19}
 return function(){h=Math.imul(h^h>>>16,2246822507);h=Math.imul(h^h>>>13,3266489909);
 h^=h>>>16;return h>>>0}}
function mulberry32(a){return function(){a|=0;a=a+0x6D2B79F5|0;
 var t=Math.imul(a^a>>>15,1|a);t=t+Math.imul(t^t>>>7,61|t)^t;
 return((t^t>>>14)>>>0)/4294967296}}
function tagOf(c){return(xmur3(c+'|CoatesK2storestag2026')()>>>0).toString(16)}
function dec(c,b64){var rnd=mulberry32(xmur3(c+'|CoatesK2stores2026')());
 var raw=atob(b64),o='';for(var i=0;i<raw.length;i++){
 o+=String.fromCharCode(raw.charCodeAt(i)^Math.floor(rnd()*256))}return o}
/* The manager layer is a SECOND payload under its own code. Money is not
   the counter's business, so it is not merely hidden on the stores view -
   it is not decryptable with the stores code at all. */
function mtagOf(c){return(xmur3(c+'|CoatesK2mgrtag2026')()>>>0).toString(16)}
function mdec(c,b64){var rnd=mulberry32(xmur3(c+'|CoatesK2mgr2026')());
 var raw=atob(b64),o='';for(var i=0;i<raw.length;i++){
 o+=String.fromCharCode(raw.charCodeAt(i)^Math.floor(rnd()*256))}return o}
//__READER__//
//__QRJS__//
var D=null;
function unlock(){
  /* The stores code is a word - upper-casing it means a bloke on a wet
     tablet at 5am does not fail the gate over a lower-case letter. The
     MANAGER code is a password with deliberate mixed case, so upper-casing
     it destroyed it - the lower-case half arrived upper-case and the tag
     never matched. Keep the raw string for the manager check.
     (Caught 29 Jul 2026, in the browser probe - the manager path had been
     "verified" in Python, where the upper-casing does not happen.) */
  var raw=(document.getElementById('code').value||'').trim();
  var c=raw.toUpperCase();
  if(!raw){return}
  if(MTAG){
    var mc=(mtagOf(raw)===MTAG)?raw:((mtagOf(c)===MTAG)?c:null);
    if(mc){ try{ MGR=JSON.parse(mdec(mc,MPAY)); }catch(e){ MGR=null; }
            if(MGR){ c=mc; } }
  }
  /* the phone-keypad alias: the numeric twin decrypts the real code
     out of its blob, then walks through the same gate as everyone */
  if(!MGR && ATAG && AKEY && tagOf(c)===ATAG){
    try{ c=dec(c,AKEY).toUpperCase(); }catch(e){}
  }
  if(tagOf(c)!==TAG && !MGR){document.getElementById('gerr').textContent=
    'That code does not open this board. Ask Andrew.';return}
  if(MGR && tagOf(c)!==TAG){
    /* The manager code opens the board as well - it is a superset, not a
       second door to remember. It gets there by decrypting the STORES
       code out of a tiny key blob: writing the stores code into the page
       so the manager could reach the board would have handed it to
       anyone who opened View Source, which is the whole gate gone.
       (Caught 29 Jul 2026, in the build.) */
    try{ D=JSON.parse(dec(mdec(c,MKEY),PAYLOAD)); }catch(e){}
    if(D){ openBoard(); return; }
  }
  try{ D=JSON.parse(dec(c,PAYLOAD)); }catch(e){
    document.getElementById('gerr').textContent='Could not open the board.';return}
  openBoard();
}

/* ================= THE STORES COMMAND CENTRE ENTRY ===============
   Andrew's third pack, 2 Aug 2026. Both unlock paths - stores code and
   manager code - now come through here, so there is one way in and one
   place the door can be turned off.

   ORDER MATTERS. safeRender() runs FIRST, before a frame of animation.
   What grows into view during the crossing is the real board carrying
   today's real numbers, not a mock-up of one. If anything at all goes
   wrong with the door the board is already built and already on screen.
================================================================= */
function openBoard(){
  document.getElementById('gate').style.display='none';
  var a=document.getElementById('app');
  a.style.display='block';
  safeRender();                       // the board is live before the door moves
  if(sdCan()) sdPlay(); else { window.scrollTo(0,0); dockOn(); }
}
var SD_DEAD=false, SD_BUSY=false, SD_T=[];
function sdCan(){
  try{
    if(SD_DEAD||SD_BUSY) return false;
    if(window.matchMedia&&matchMedia('(prefers-reduced-motion: reduce)').matches)
      return false;
  }catch(e){}
  return !!document.getElementById('sdoor');
}
function sdBail(){ SD_DEAD=true; sdEnd(); }
function sdSkip(){ sdEnd(); }
function sdEnd(){
  var d=document.getElementById('sdoor'), a=document.getElementById('app');
  for(var i=0;i<SD_T.length;i++) clearTimeout(SD_T[i]);
  SD_T=[]; SD_BUSY=false;
  if(d) d.className='';
  if(a) a.className=a.className.replace(/\bsdgrow\b/,'').trim();
  document.body.style.overflow='';
  window.scrollTo(0,0);
  dockOn();
}
function sdPlay(){
  var d=document.getElementById('sdoor'), a=document.getElementById('app');
  if(!d){ window.scrollTo(0,0); dockOn(); return; }
  SD_BUSY=true;
  document.body.style.overflow='hidden';
  d.className='on';
  void d.offsetWidth;                 // restart the animation cleanly
  d.className='on go';                // circuits, floor route and the walk
  /* the hand-off: the room lets go over 450ms while the board rises
     under it, so it reads as arriving rather than as a cut */
  SD_T.push(setTimeout(function(){
    d.className='on go out';
    if(a) a.className=(a.className+' sdgrow').trim();
  }, 5900));
  SD_T.push(setTimeout(sdEnd, 5900+450));
}
document.addEventListener('keydown',function(e){
  if(SD_BUSY&&(e.key==='Escape'||e.keyCode===27)) sdSkip();
});
/* ---- THE DOCK ----------------------------------------------------
   The board is long and the counter reads it one-handed while somebody
   is standing at the window. The five places they actually go all shift
   now sit at the bottom of the screen.

   It is ADDITIVE. Nothing came off the home screen to pay for it -
   every tile, every section and all fifteen panes are exactly where
   they were, so nobody has to relearn the board.
------------------------------------------------------------------- */
function dockOn(){
  var k=document.getElementById('sdock');
  if(!k) return;
  k.className='on';
  document.body.classList.add('hasdock');
  dockMark('home');
}
function dockMark(k){
  var bs=document.querySelectorAll('#sdock button');
  for(var i=0;i<bs.length;i++)
    bs[i].className=(bs[i].getAttribute('data-d')===k)?'on':'';
}
function dock(k){
  if(k==='home'){ home(); } else { nav(k); }
  dockMark(k);
}
function safeRender(){
  /* NEVER a blank screen: if drawing the board throws for any reason,
     put the gate straight back up with the code still typed in - one
     Enter from in, no refreshing (Andrew, 31 Jul 2026: "cant refresh
     to see things its silly") */
  try{ render(); }
  catch(e){
    document.getElementById('app').style.display='none';
    document.getElementById('gate').style.display='';
    var ge=document.getElementById('gerr');
    if(ge)ge.textContent='The board hiccupped opening - press Enter to go again.';
  }
}
document.getElementById('code').addEventListener('keydown',function(e){
  if(e.key==='Enter') unlock();});
/* Came through the override on the crew page? Open straight up. The code
   travels in sessionStorage, never the address bar, and is cleared the
   moment it is used so a shared tablet does not stay unlocked. */
(function(){ try{
  var h=sessionStorage.getItem('k2stores');
  if(h){ sessionStorage.removeItem('k2stores');
         /* wait for EVERY script block to parse before unlocking -
            render() reaches helpers defined further down the page, and
            firing early left a hidden gate and an empty app: the blank
            page Andrew hit coming through the crew-page shortcut
            (31 Jul 2026) */
         setTimeout(function(){
           try{ document.getElementById('code').value=h; unlock(); }
           catch(e2){}
           /* belt and braces: whatever happened, never leave a blank
              page - if the board is not up, the gate must be */
           var ap=document.getElementById('app');
           if(!ap||ap.style.display!=='block'||!ap.innerHTML){
             document.getElementById('gate').style.display='';
           }
         },50); }
}catch(e){} })();

function esc(s){return String(s==null?'':s).replace(/&/g,'&amp;')
  .replace(/</g,'&lt;').replace(/>/g,'&gt;')}
/* THE INSTRUMENT BRIDGE. Andrew's pack: "the overhead figures are also
   controls." So a tile that has somewhere sensible to send you becomes
   a button and wears a chevron; one that does not stays a plain tile
   rather than looking tappable and doing nothing. */
function tile(v,l,cls,go){
  if(!go) return '<div class="tile '+(cls||'')+'"><b>'+v+'</b><span>'+l+'</span></div>';
  return '<button type="button" class="tile '+(cls||'')+'" onclick="nav(\\''+go+'\\')">'
   +'<b>'+v+'</b><span>'+l+'</span></button>';
}
/* Long lists are cut short so the page stays quick on a tablet. A cut list
   that says nothing is a lie by omission - a storeman scrolling 80 of 253
   idle machines would swear the other 173 are not on site. Every cut says
   so. (29 Jul 2026.) */
function more(shown,total,word){
  return total>shown?'<div class="kw cut">Showing the first '+shown+' of '
    +total.toLocaleString()+' '+(word||'lines')+' &mdash; the rest are in SiteIQ.</div>':'';
}

/* THE SIX BAYS (Andrew's Store Street pack, 2 Aug 2026).
   The board used to open on fifteen tabs in a row, then on a grouped
   menu, and now on his six physical bays.

   THE ONE RULE I BUILT IT UNDER. I told him I would not swap twelve
   doors for six, because Stocktake going from one tap to two is a real
   cost at a counter with a bloke waiting at the window. So the bays
   carry their own doors ON THEIR FACE. Six places in the store, every
   destination still one tap. Nothing was buried to make it pretty.

   MENU on the crumb bar brings you straight back from anywhere, and
   the dock along the bottom carries the five most-used. */
function bayIcon(k){
  var P={
   find:'<circle cx="10.5" cy="10.5" r="5.7"/><path d="m15 15 4.5 4.5M8 10.5h5M10.5 8v5"/>',
   hunt:'<path d="M7.5 14.5h9l-1-7h-7zM6 17h12M9 4.5l-1-2M15 4.5l1-2M5.5 7 3.5 6M18.5 7l2-1"/>',
   floor:'<path d="M4 4v16M20 4v16M4 8h16M4 14h16M7 5.5h3M13 5.5h4M7 11h5M15 11h2M7 17h4"/>',
   print:'<path d="M7 9V4h10v5M7 17H4V9h16v8h-3M7 14h10v6H7zM16.5 11.5h.01"/>',
   plant:'<path d="M4 17h13M6 17l2-7h6l2 7M9 10V6h4l3 4M16 10h3v7M7 20a2 2 0 1 0 0-4M17 20a2 2 0 1 0 0-4"/>',
   ctrl:'<path d="M12 3 5 6v5c0 4.5 2.8 8 7 10 4.2-2 7-5.5 7-10V6l-7-3Zm-3.5 9 2.2 2.2 4.8-5"/>'};
  return '<svg viewBox="0 0 24 24">'+(P[k]||P.find)+'</svg>';
}
/* one door on a bay's face. n = the live count on it, hot = it is red */
function bayDoor(k,label,n,cls){
  return '<button type="button" onclick="nav(\\''+k+'\\')" class="'+(cls||'')+'">'
   +label+(n!=null&&n!==''?'<span class="bn">'+n+'</span>':'')
   +'<em>&rsaquo;</em></button>';
}
/* Andrew's own bay photographs, 2 Aug 2026. Each bay gets the picture
   of the place it is, at 320px and about 13 KB apiece - the whole set
   is 80 KB. If the Art folder ever goes missing the line drawing takes
   over on the spot, so the board never shows a broken square. */
var BAYART={find:'01-find-it-counter',hunt:'02-the-hunt',
 floor:'03-store-floor',print:'04-print-works',plant:'05-plant-desk',
 ctrl:'06-store-control'};
function bayArt(ic){
  var f=BAYART[ic];
  if(!f) return '<div class="bic">'+bayIcon(ic)+'</div>';
  return '<div class="bic art"><img src="art/bays/'+f+'.webp" alt="" '
   +'loading="lazy" decoding="async" '
   +'onerror="this.parentNode.className=\\'bic\\';'
   +'this.parentNode.innerHTML=bayIcon(\\''+ic+'\\')"></div>';
}
function bay(n,cls,ic,title,sub,stCls,stTx,doors){
  return '<section class="bay '+(cls||'')+'"><div class="bhd">'
   +bayArt(ic)+'<div class="bmeta">'
   +'<small>BAY 0'+n+'</small><b>'+title+'</b><span>'+sub+'</span>'
   +'</div></div>'
   +'<div class="bst '+(stCls||'')+'"><i></i>'+stTx+'</div>'
   +'<div class="bdoors">'+doors+'</div></section>';
}
function homeMenu(){
  var t=D.tiles, h='<div class="bays">';
  h+='<div class="youare"><i><svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="3"/>'
   +'<path d="M12 3v3M12 18v3M3 12h3M18 12h3"/></svg></i>'
   +'<div><small>YOU ARE HERE</small><b>STORE STREET</b></div></div>';

  /* BAY 01 - the counter's most-asked question */
  h+=bay(1,'org','find','Find It Counter',
    'Scan or type anything &mdash; the answer is the shelf, the person, '
    +'or the hunt.','g','SEARCH READY',
    bayDoor('find','Open the counter search',null,'wide')
    +bayDoor('who','Any company or name'));

  /* BAY 02 - the hunt. Red when there is a hit list, amber when there
     is only a chase list, green when the store is square. */
  var hcls=D.hitN?'red':(t.chase?'amb':''),
      hst=D.hitN?'r':(t.chase?'a':'g'),
      htx=D.hitN?(D.hitN+' ON THE HIT LIST')
         :(t.chase?(t.chase+' TO CHASE'):'NOTHING OVERDUE');
  h+=bay(2,hcls,'hunt','The Hunt',
    'Out past its return rule &mdash; who has it and for how long.',
    hst,htx,
    bayDoor('chase','Chase up',t.chase)
    +(D.hitN?bayDoor('hits','Hit list',D.hitN,'hot'):'')
    +(t.arrivals?bayDoor('arr','Arriving',t.arrivals):''));

  /* BAY 03 - the store floor */
  h+=bay(3,'org','floor','Store Floor',
    'Aisles, lines, counts and consumables &mdash; the shelf itself.',
    (t.stale?'a':'g'),
    t.lines.toLocaleString()+' LINES &middot; '+t.stockPct+'% COUNTED',
    bayDoor('groups','Product groups',t.lines.toLocaleString())
    +bayDoor('aisle','Walk an aisle')
    +bayDoor('stock','Stocktake',t.stockPct+'%')
    +bayDoor('battle','Day v Night')
    +(D.hasCons?bayDoor('cons','Consumables',D.cons.order.length||null):''));

  /* BAY 04 - print works */
  h+=bay(4,'org','print','Print Works',
    'Every report from one place: person, company, whole site, radios, '
    +'gas, hit list.','g','PRINT READY',
    bayDoor('print','Open the print hub',null,'wide'));

  /* BAY 05 - plant desk. Only built when the export carried plant. */
  if(D.hasPlant||t.idle){
    h+=bay(5,(t.idle?'amb':''),'plant','Plant Desk',
      'The machines &mdash; free to hire, out with companies, and what '
      +'is parked doing nothing.',
      (t.idle?'a':'g'),
      (t.idle?t.idle+' IDLE ON HIRE':'NOTHING PARKED'),
      (D.hasPlant?bayDoor('plant','Plant status'):'')
      +bayDoor('idle','Idle plant',t.idle));
  }

  /* BAY 06 - store control. The manager's money screen only appears
     for a manager code, same as it always has. */
  h+=bay(6,'','ctrl','Store Control',
    'A fresh read off a raw export, and the rules of this store.',
    'g','STORE CONTROL',
    bayDoor('fresh','Fresh look')
    +bayDoor('std','Our standards')
    +(MGR?bayDoor('mgr','Money'):''));

  return h+'</div>';
}
function render(){
  var t=D.tiles;
  var h='<div class="crumb" id="crumb" style="display:none">'
   +'<button type="button" onclick="home()">&#8962; MENU</button>'
   +'<b id="crumb-t"></b></div>'
   +'<div class="pane on" id="p-home"><div class="tiles">'
   +tile(t.avail.toLocaleString(),'On the shelf','g','groups')
   +tile(t.onhire.toLocaleString(),'Out with crews','','chase')
   +tile(t.lines.toLocaleString(),'Different things','','groups')
   +tile(t.chase,'Chase up','a','chase')
   +tile(t.stockPct+'%','Counted in 7 days','g','stock')
   +tile(t.stale,'Not counted','r','stock')
   +'</div>'
   +homeMenu()
   +'</div>'
   +'<div class="pane" id="p-find">'+helpBar('find')+paneFind()+'</div>'
   +'<div class="pane" id="p-who">'+helpBar('who')+paneWho()+'</div>'
   +'<div class="pane" id="p-groups">'+helpBar('groups')+paneGroups()+'</div>'
   +'<div class="pane" id="p-chase">'+helpBar('chase')+paneChase()+'</div>'
   +(D.hitN?'<div class="pane" id="p-hits">'+helpBar('hits')+paneHits()+'</div>':'')
   +'<div class="pane" id="p-print">'+helpBar('print')+panePrint()+'</div>'
   +'<div class="pane" id="p-fresh">'+helpBar('fresh')+paneFresh()+'</div>'
   +'<div class="pane" id="p-stock">'+helpBar('stock')+paneStock()+'</div>'
   +'<div class="pane" id="p-aisle">'+helpBar('aisle')+paneAisle()+'</div>'
   +'<div class="pane" id="p-battle">'+helpBar('battle')+paneBattle()+'</div>'
   +(D.hasCons?'<div class="pane" id="p-cons">'+helpBar('cons')+paneCons()+'</div>':'')
   +'<div class="pane" id="p-std">'+helpBar('std')+paneStd()+'</div>'
   +(t.arrivals?'<div class="pane" id="p-arr">'+helpBar('arr')+paneArr()+'</div>':'')
   +(D.hasPlant?'<div class="pane" id="p-plant">'+helpBar('plant')+panePlant()+'</div>':'')
   +(MGR?'<div class="pane" id="p-mgr">'+paneMgr()+'</div>':'')
   +'<div class="pane" id="p-idle">'+helpBar('idle')+paneIdle()+'</div>'
   +'<div class="foot">Built from this morning\\'s SiteIQ exports &middot; '
   +'read-only &middot; POWERED BY SITEIQ<br>Author: Andrew Fisher</div>';
  document.getElementById('app').innerHTML=h;
  countTiles();
}
/* The unlock count-up. Numbers run from zero to their real value in
   three-quarters of a second, once, then never move again. The REAL
   value is parsed off the page and written back at the end, so a
   glitch mid-animation can never leave a wrong number standing - the
   worst failure mode is the right number appearing instantly. */
function countTiles(){
  try{
    if(window.matchMedia&&matchMedia('(prefers-reduced-motion: reduce)').matches)return;
  }catch(e){}
  var els=[].slice.call(document.querySelectorAll('#app .tiles .tile b'));
  els.forEach(function(el){
    var raw=(el.textContent||'').trim();
    var m=raw.match(/^([0-9,]+)(%?)$/);
    if(!m)return;
    var target=parseInt(m[1].replace(/,/g,''),10);
    if(isNaN(target)||target===0)return;
    var pct=m[2],t0=null,DUR=750,done=false;
    function fin(){ if(!done){done=true; el.textContent=raw;} }
    function step(ts){
      if(done)return;
      if(t0===null)t0=ts;
      var k=Math.min(1,(ts-t0)/DUR);
      k=1-Math.pow(1-k,3);
      el.textContent=Math.round(target*k).toLocaleString()+pct;
      if(k<1)requestAnimationFrame(step); else fin();
    }
    requestAnimationFrame(step);
    /* the guarantee: whatever happens to the animation frames - a
       background tab, a throttled webview, a headless browser - the
       real number is standing 900ms after unlock, full stop */
    setTimeout(fin,DUR+150);
  });
}
/* one pane at a time: nav(key) steps in, home() steps back out. The
   crumb bar names where you are and MENU is always one tap away. */
function nav(k){
  var p=document.getElementById('p-'+k);
  if(!p){ home(); return; }
  var panes=document.querySelectorAll('.pane');
  for(var j=0;j<panes.length;j++) panes[j].className='pane';
  p.className='pane on';
  var names={mgr:'Money',print:'Print hub',who:'Master search'};
  document.getElementById('crumb').style.display='flex';
  document.getElementById('crumb-t').textContent=
    names[k]||(HOWTO[k]?HOWTO[k].t:k);
  window.scrollTo(0,0);
  /* the master search opens with the biggest holders already on screen -
     an empty box tells a bloke nothing about what is in here */
  if(k==='who'&&typeof whoFind==='function'){ try{ whoFind(); }catch(e){} }
  /* keep the dock honest - a pane it does not carry lights nothing
     rather than leaving the last one lit and lying about where you are */
  if(typeof dockMark==='function') dockMark(k);
}
function home(){
  var panes=document.querySelectorAll('.pane');
  for(var j=0;j<panes.length;j++) panes[j].className='pane';
  document.getElementById('p-home').className='pane on';
  document.getElementById('crumb').style.display='none';
  window.scrollTo(0,0);
  if(typeof dockMark==='function') dockMark('home');
}
/* THE FINDER - the counter's most-asked question, answered in one box
   (Andrew, 31 Jul 2026: "where's grinder 1219644?"). Scan or type an
   item number, Plant ID or name; the answer is the shelf, the person,
   or the hunt list. The index is built once, on first search, from
   the three truths already in the payload: the shelf (find.av), the
   roster (out with crews) and the stocktake stale list (missing). */
var FIND_IDX=null;
function findIdx(){
  if(FIND_IDX) return FIND_IDX;
  var ix={}, SEP=String.fromCharCode(31);
  var av=(D.find&&D.find.av)||{};
  Object.keys(av).forEach(function(k){
    var p=k.split(SEP), n=p[0], u=p[1]||'';
    av[k].forEach(function(it){ ix[String(it).toUpperCase()]={s:'A',n:n,u:u,k:k}; });
  });
  D.roster.forEach(function(x){
    if(x.i) ix[String(x.i).toUpperCase()]={s:'O',n:x.n,u:x.u,w:x.w,co:x.co,d:x.d,p:x.p};
  });
  (D.stock.stale||[]).forEach(function(x){
    if(!x.i) return;
    var key=String(x.i).toUpperCase(), e=ix[key];
    if(e&&e.s==='A'){ e.st=x.d; e.by=x.by; }
    else if(!e) ix[key]={s:'M',n:x.n,u:x.u,d2:x.d,by:x.by,hs:x.s};
  });
  var nv={}, fln={};
  D.groups.forEach(function(g){
    if(g.v) nv[g.n.toUpperCase()]=g.v;
    if(g.fl) fln[g.n.toUpperCase()]=g.fl;
  });
  FIND_IDX={ix:ix,nv:nv,fln:fln,av:av};
  return FIND_IDX;
}
function paneFind(){
  return '<div class="note"><b>Where is it?</b> Scan the barcode with the '
   +'hand scanner, or type the item number off a sheet or sticker. The '
   +'answer is the shelf, the person holding it, or the hunt list. Plant '
   +'IDs and names work too.</div>'
   +'<input class="srch" id="fq" placeholder="Scan or type an item number, Plant ID or name" '
   +'autocomplete="off" autocapitalize="characters" oninput="findGo()" '
   +'onkeydown="if(event.keyCode===13)findGo()">'
   +'<div id="fout"><div class="kw" style="padding:12px 2px">Waiting for a '
   +'number&hellip; the hand scanner types it and presses Enter for you.</div></div>';
}
function findGo(){
  var q=(document.getElementById('fq').value||'').trim().toUpperCase();
  var out=document.getElementById('fout');
  if(q.length<3){out.innerHTML='<div class="kw" style="padding:12px 2px">'
    +'Keep typing&hellip; three characters gets it looking.</div>';return}
  var F=findIdx(), hits=[], i;
  if(F.ix[q]) hits.push([q,F.ix[q]]);
  if(!hits.length){
    var ks=Object.keys(F.ix);
    for(i=0;i<ks.length&&hits.length<12;i++){
      if(ks[i].indexOf(q)>=0) hits.push([ks[i],F.ix[ks[i]]]);
    }
  }
  if(!hits.length){
    D.roster.forEach(function(x){
      if(hits.length<12&&x.p&&String(x.p).toUpperCase()===q)
        hits.push([x.i||'',{s:'O',n:x.n,u:x.u,w:x.w,co:x.co,d:x.d,p:x.p}]);
    });
  }
  if(!hits.length){
    var ks2=Object.keys(F.ix);
    for(i=0;i<ks2.length&&hits.length<12;i++){
      var e2=F.ix[ks2[i]];
      if(e2.n&&e2.n.toUpperCase().indexOf(q)>=0) hits.push([ks2[i],e2]);
    }
  }
  if(!hits.length){
    out.innerHTML='<div class="note" style="border-left-color:var(--rd)">'
      +'<b>Nothing on the register matches.</b> Check the number &mdash; or '
      +'it may have arrived after this morning&rsquo;s build. A fresh SiteIQ '
      +'export in the Fresh look tab covers the gap.</div>';
    return;
  }
  out.innerHTML=hits.map(function(h){return findCard(h[0],h[1],F)}).join('')
   +(hits.length>=12?'<div class="kw cut">Showing the first 12 &mdash; keep '
     +'typing to narrow it.</div>':'');
}
function histLine(i){
  var h2=(D.hist||{})[i];
  if(!h2)return '';
  return ' &middot; hired '+h2[0]+'&times; this shut'
    +(h2[1]?' &middot; last: <b>'+esc(h2[1])+'</b>':'');
}
function findCard(it,e,F){
  var v=F.nv[(e.n||'').toUpperCase()]||'';
  var th=v?'<span class="kth2"><img src="thumbs/'+encodeURIComponent(tsafe(v))
    +'.jpg" loading="lazy" alt="" data-m="'+thMono(e.n)+'" onerror="thx(this)"></span>'
    :'<span class="kth2 mono">'+thMono(e.n)+'</span>';
  var head,body,cls;
  if(e.s==='O'){
    cls='fo'; head='OUT WITH A CREW';
    /*  the name is tappable straight off the finder card (Andrew,
        2 Aug 2026: "if you see things onhire to someone you also should
        be able to clkick on the name and it maybe take you to that
        person to find out more info")  */
    body='<b>'+(typeof whoLink==='function'?whoLink(e.w||'?',e.co||'')
                :esc(e.w||'?'))+'</b> &middot; '+esc(e.co||'')
      +(e.d!=null?' &middot; <b>'+e.d+(e.d===1?' day':' days')+'</b> out':'')
      +'<br>Lives in '+esc(e.u||'?')+histLine(it);
  } else if(e.s==='A'){
    cls='fa'; head='ON THE SHELF';
    var mates=(F.av[e.k||'']||[]).length;
    body='Aisle: <b>'+esc(e.u||'?')+'</b>'
      +(mates>1?' &middot; '+mates+' of these available':'');
    /* who has the rest (Andrew, 1 Aug 2026): the shelf answer also
       says where the others went - the chase starts from the tool */
    var outm=[], nrest=0, j2, x2;
    for(j2=0;j2<D.roster.length;j2++){ x2=D.roster[j2];
      if(x2.n===e.n){ nrest++; if(outm.length<3) outm.push(x2); } }
    if(nrest) body+='<br>Out with crews: <b>'+nrest+'</b> &middot; '
      +outm.map(function(x3){return (typeof whoLink==='function'
          ?whoLink(x3.w||'?',x3.co||''):esc(x3.w||'?'))+' ('
        +(x3.d!=null?x3.d+'d':'?')+')'}).join(' &middot; ')
      +(nrest>3?' &middot; +'+(nrest-3)+' more in CHASE UP':'');
    if(e.st!=null){cls='fw';head='SHOULD BE ON THE SHELF';
      body+='<br><b>Not sighted in '+e.st+'d</b>'
        +(e.by?' &middot; last seen by '+esc(e.by):'')
        +' &mdash; confirm it is really there.';}
  } else {
    cls='fw';
    head=(e.hs==='O')?'ON HIRE - NOT COUNTED'
        :'NOT SEEN IN '+(e.d2!=null?e.d2+'d':'A WHILE');
    body='Lives in <b>'+esc(e.u||'?')+'</b>'+histLine(it)
      +(e.by?' &middot; last sighted by <b>'+esc(e.by)+'</b>':'')
      +'<br>On this aisle&rsquo;s hunt list &mdash; Stocktake tab.';
  }
  return '<div class="fcard '+cls+'">'+th
   +'<div class="fbody"><div class="fhead">'+head+'</div>'
   +'<div class="fname">'+esc(e.n||'Unnamed item')+'</div>'
   +'<div class="kw">Item '+esc(it)
   +(e.p?' &middot; <b style="color:var(--org)">Plant ID '+esc(e.p)+'</b>':'')
   +(v?' &middot; <span class="vcode">'+esc(v)+'</span>':'')+'</div>'
   +'<div class="kw">'+body+'</div>'+compChips(e.fl||(F.fln?F.fln[(e.n||'').toUpperCase()]:''))+'</div>'
   +'<div class="kqr">'+qr(it,56)+'</div></div>';
}
function paneGroups(){
  var cats={};
  D.groups.forEach(function(g){ (cats[g.c]=cats[g.c]||[]).push(g); });
  var h='<input class="srch" id="gq" placeholder="Search everything the store holds" '
       +'oninput="filterGroups()">' + '<div id="glist">';
  Object.keys(cats).sort().forEach(function(c){
    var list=cats[c], av=0, oh=0;
    list.forEach(function(g){av+=g.av;oh+=g.oh});
    var tot=av+oh||1;
    h+='<div class="grp" data-cat="'+esc(c)+'">'
      +'<button type="button" onclick="tog(this)">'
      +'<div class="gn"><b>'+esc(c)+'</b><span>'+list.length+' different things</span></div>'
      +'<div class="gq"><b>'+av+'</b><span>on shelf</span></div>'
      +'<div class="gq"><b style="color:var(--org)">'+oh+'</b><span>out</span></div>'
      +'</button>'
      +'<div class="bar"><i class="ba" style="width:'+(100*av/tot)+'%"></i>'
      +'<i class="bo" style="width:'+(100*oh/tot)+'%"></i></div>'
      +'<div class="kids">'+kidsWithSeams(list)+'</div></div>';
  });
  return h+'</div>';
}
/* Family seams inside a big category - same rule as the crew
   catalogue, so the counter and the crew read the same shelves.
   Sockets is 283 lines; a seam every family turns the scroll into
   signposted shelving. Small categories stay flat. */
function famOf(n){
  var f=String(n||'')
    .replace(/[-\\u2013]\\s*[\\d,\\/. ]+\\s*(mm|in|inch|")?\\s*$/i,'')
    .split(' - ')[0].replace(/[-\\s]+$/,'').trim();
  return f||String(n||'');
}
function kidsWithSeams(list){
  if(list.length<=25) return list.map(kid).join('');
  var out='',last=null;
  list.forEach(function(g){
    var f=famOf(g.n);
    if(f!==last){out+='<div class="kfam">'+esc(f)+'</div>';last=f;}
    out+=kid(g);
  });
  return out;
}
function kid(g){
  var units=Object.keys(g.u||{}).map(function(u){return u+' ('+g.u[u]+')'}).join(' &middot; ');
  var who='';
  if(g.who && g.who.length){
    who='<div class="kw">'+g.who.slice(0,12).map(function(w){
      return '<b>'+esc(w.w)+'</b> &middot; '+esc(w.co)
        +(w.d==null?'':' &middot; '+w.d+(w.d===1?' day':' days')+' out');
    }).join('<br>')+(g.who.length>12?'<br>+ '+(g.who.length-12)+' more':'')+'</div>';
  }
  return '<div class="kid kidth">'+thTile(g.v,g.n)
    +'<div class="kbody"><div class="kt"><b>'+esc(g.n)+'</b>'
    +'<em>'+g.av+' free</em>'+(g.oh?'<em class="o">'+g.oh+' out</em>':'')+'</div>'
    +(units?'<div class="where">'+esc(units)
      +(g.v?' &middot; <span class="vcode">'+esc(g.v)+'</span>':'')+'</div>':'')
    +compChips(g.fl)+who+'</div></div>';
}
/* the picture tile: the variant's thumbnail (Gear_Lookup/thumbs), or
   a two-letter monogram until its photo is collected - 56_PHOTO_HUNT
   is the wanted list. Lazy-loaded; a missing file falls back clean. */
/* compliance chips: the tag colour word comes from the payload (the
   master's windows - Jul/Aug is BLUE), LOG BOOK marks the machines
   that need the daily pre-start written up. One helper, every list. */
function compChips(fl){
  if(!fl) return '';
  var t=D.tag||{c:'',x:'#8A97A8'}, out='';
  if(fl.indexOf('E')>=0&&t.c)
    out+='<span class="cchip" style="background:'+t.x+'">TAG '+esc(t.c)+' &middot; ELECTRICAL</span>';
  if(fl.indexOf('R')>=0&&t.c)
    out+='<span class="cchip" style="background:'+t.x+'">TAG '+esc(t.c)+' &middot; RIGGING / HEIGHT</span>';
  if(fl.indexOf('L')>=0)
    out+='<span class="cchip lbk">&#128221; LOG BOOK &middot; DAILY PRE-START</span>';
  return out?'<div class="cchips">'+out+'</div>':'';
}
function thMono(n){
  var w=String(n||'').split(/[^A-Za-z0-9]+/).filter(function(x){return x});
  return ((w[0]||'?').charAt(0)+(w[1]||w[0]||'').charAt(0)).toUpperCase();
}
function tsafe(v){return String(v).replace(/[/:*?"<>|]/g,'_')}
function thTile(v,n,a){
  if(!v) return '<span class="kth2 mono">'+thMono(n)+'</span>';
  /* If the build told us which photos exist, believe it rather than
     firing a request to find out. Neither the serial shot nor the
     model shot present = draw the monogram now. Guarded, so a page
     built without a manifest behaves exactly as it always did. */
  if(typeof hasThumb==='function' && !hasThumb(v) && !(a && hasThumb(a)))
    return '<span class="kth2 mono">'+thMono(n)+'</span>';
  return '<span class="kth2"><img src="thumbs/'+encodeURIComponent(tsafe(v))
    +'.jpg" loading="lazy" alt="" data-m="'+thMono(n)
    +(a&&a!==v?'" data-a="'+esc(a):'')
    +'" onerror="thx(this)"></span>';
}
function thx(img){
  /* photo chain: serial photo -> model photo (data-a) -> monogram */
  var a=img.getAttribute('data-a');
  if(a){img.removeAttribute('data-a');
    img.src='thumbs/'+encodeURIComponent(tsafe(a))+'.jpg';return;}
  var s=img.parentNode;
  s.className=s.className.replace(' mono','')+' mono';
  s.textContent=img.getAttribute('data-m')||'?';
}
function tog(b){
  var k=b.parentNode.querySelector('.kids');
  k.className = k.className.indexOf('on')>=0 ? 'kids' : 'kids on';
}
function filterGroups(){
  var q=(document.getElementById('gq').value||'').toLowerCase().trim();
  var grps=document.querySelectorAll('#glist .grp');
  for(var i=0;i<grps.length;i++){
    var g=grps[i], txt=g.textContent.toLowerCase();
    var hit=!q||txt.indexOf(q)>=0;
    g.style.display=hit?'':'none';
    if(q&&hit) g.querySelector('.kids').className='kids on';
  }
}
function paneChase(){
  var c=D.chase;
  var h='<div class="note"><b>Gear out longer than three days.</b> Walk the aisle, '
   +'check the shelf, then chase the name. Site plant, barriers and chutes are '
   +'listed separately below &mdash; they live on site by design and are not lost.</div>';
  h+='<div class="uhead">Tools to chase &mdash; '+c.tools.length+'</div>';
  var byU={};
  c.tools.forEach(function(x){(byU[x.u]=byU[x.u]||[]).push(x)});
  c.toolUnits.forEach(function(p){
    var u=p[0], list=byU[u]||[];
    h+='<div class="grp"><button type="button" onclick="tog(this)">'
      +'<div class="gn"><b>'+esc(u)+'</b><span>go and look here</span></div>'
      +'<div class="gq"><b style="color:var(--am)">'+list.length+'</b><span>items</span></div>'
      +'</button><div class="kids">'
      +list.map(function(x){
        return '<div class="kid kidth">'+thTile(x.v,x.n,x.va)
          +'<div class="kbody"><div class="kt"><b>'+esc(x.n)+'</b>'
          +'<em class="o">'+x.d+' days</em></div>'
          +'<div class="kw"><b>'+wl(x)+'</b> &middot; '+esc(x.co)
          +(x.i?' &middot; Item '+esc(x.i):'')+histLine(x.i)+'</div></div></div>';
      }).join('')+'</div></div>';
  });
  h+='<div class="uhead">Site plant, barriers &amp; chutes &mdash; '+c.plant.length
    +' (normal)</div><div class="note">Out for days because that is where they '
    +'live. Kept out of the chase list so it stays worth reading.</div>';
  return h;
}
function paneStock(){
  /* STOCKTAKE v2 (Andrew, 31 Jul 2026): no walls of asset numbers on
     screen. Pin-point WHERE to count (area cards, worst first, each
     with a percentage), tap in for tappable freshness buckets, and
     lead with the ghost list - items the team has walked past more
     than once and still not found. Numbers live on the printed sheet
     where the scanner needs them, not on the phone. */
  var s=D.stock, t=D.tiles;
  var h='<div class="ring"><div class="rv">'+t.stockPct+'%</div>'
   +'<div class="rt"><b>of the store counted in the last 7 days</b>'
   +s.w1.toLocaleString()+' sighted in the last day &middot; '
   +s.total.toLocaleString()+' lines on the register &middot; plant, chutes, '
   +'hoppers and barriers are audited on the Plant tab, not counted here'
   +'</div></div>';
  var aIx={}; (s.areas||[]).forEach(function(a,i){aIx[a.u]=i;});
  /* the ghost list first - the thing actually worth worrying about */
  var gh=s.ghosts||[];
  if(gh.length){
    h+='<div class="uhead">Walked past &mdash; still not found</div>'
     +'<div class="note" style="border-left-color:var(--rd)"><b>'+gh.length
     +' item'+(gh.length===1?'':'s')+' live in aisles the team has already '
     +'counted &mdash; more than once &mdash; and still have not turned '
     +'up.</b> Stop re-hunting them on every lap: one proper look, then flag '
     +'it to Andrew so the loss conversation starts now, not at demob.</div>'
     +gh.slice(0,12).map(function(x){
       return '<div class="kid kidth">'+thTile(x.v,x.n,x.va)
        +'<div class="kbody"><div class="kt"><b>'+esc(x.n)+'</b>'
        +'<em class="o">'+x.d+'d</em></div>'
        +'<div class="kw">Lives in <b>'+esc(x.u)+'</b>'
        +(x.i?' &middot; Item '+esc(x.i):'')
        +(x.hc?' &middot; hired '+x.hc+'&times; this shut':'')
        +(x.lh?' &middot; last had it: <b>'+esc(x.lh)+'</b>':'')+'</div>'
        +'<div class="kw" style="margin-top:4px"><span class="ghp">MISSED ON '
        +x.m+' WALK'+(x.m===1?'':'S')+'</span></div></div></div>';
     }).join('')+more(12,gh.length,'items');
  }
  /* today's counting orders - one aisle chip per shift, no item walls.
     The 70/30 share between the shifts is set in the build and is not
     printed anywhere on this pane - deliberately. */
  var od=D.orders||{d:[],n:[]};
  if((od.d.length+od.n.length)>0){
    h+='<div class="uhead">Today&rsquo;s counting orders</div>'
     +'<div class="note"><b>Tap your aisle, count the oldest bucket first, '
     +'tick the paper sheet as you walk.</b> The board picks the aisles so '
     +'nobody has to decide at 18:05 what &ldquo;do some stocktake&rdquo; '
     +'means.</div>';
    [['DAY SHIFT &middot; 06:00&ndash;18:00',od.d],
     ['NIGHT SHIFT &middot; 18:00&ndash;06:00',od.n]].forEach(function(rw){
      h+='<div class="kw" style="padding:6px 2px;font-weight:800;color:var(--dim)">'
       +rw[0]+'</div><div class="prpick">'
       +(rw[1].length?rw[1].map(function(o){
          return '<button class="stmore" type="button" onclick="stArea('
           +(aIx[o.u]!=null?aIx[o.u]:0)+')">'+esc(o.u)
           +' &middot; '+o.n+' to sight</button>';
        }).join(''):'<span class="kw">Nothing assigned &mdash; the other '
          +'shift carries today&rsquo;s backlog.</span>')
       +'</div>';
    });
  }
  /* THE RECORD - what each shift actually counted, written by the
     scanner timestamps. The answer to "how do I know what they did". */
  var lg=s.log||[];
  if(lg.length){
    h+='<div class="uhead">The record &mdash; what each shift actually did</div>'
     +'<div class="note"><b>This writes itself from the scanner timestamps '
     +'&mdash; nobody fills in a form.</b> Every day: what was assigned, '
     +'what actually got sighted, what got missed. An aisle walked by both '
     +'shifts shows up too &mdash; that&rsquo;s wasted legs.</div>';
    lg.forEach(function(dy){
      var cd=dy.c.d||{}, cn=dy.c.n||{};
      var totD=0, totN=0, u2;
      for(u2 in cd) totD+=cd[u2];
      for(u2 in cn) totN+=cn[u2];
      h+='<div class="grp"><button type="button" onclick="tog(this)">'
       +'<div class="gn"><b>'+dy.dt+(dy.today?' &middot; today so far':'')+'</b>'
       +'<span>assigned v actually counted, both shifts</span></div>'
       +'<div class="gq"><b>'+(totD+totN)+'</b><span>sighted</span></div>'
       +'</button><div class="kids'+(dy.today?' on':'')+'">'
       +stLogShift('DAY SHIFT',dy.o.d||[],cd,cn)
       +stLogShift('NIGHT SHIFT',dy.o.n||[],cn,cd)
       +'</div></div>';
    });
  }
  /* the area scoreboard - pin-point where counting is needed: worst
     areas first, a percentage on every card */
  h+='<div class="uhead">Where to walk &mdash; every storage area</div>';
  h+=(s.areas||[]).map(function(a,i){
    var c=a.p>=90?'var(--gd)':a.p>=70?'var(--am)':'var(--rd)';
    return '<button type="button" class="ucard" onclick="stArea('+i+')">'
     +'<div class="urow"><div class="un"><b>'+esc(a.u)
     +(a.own==='d'?' <span class="ghp" style="background:var(--am)">DAYS TODAY</span>'
      :a.own==='n'?' <span class="ghp" style="background:var(--org)">NIGHTS TODAY</span>':'')
     +'</b><span>'+a.s7+' of '+a.t+' sighted this week</span></div>'
     +'<div class="upct" style="color:'+c+'">'+a.p+'%</div></div>'
     +'<div class="ubar"><i style="width:'+a.p+'%;background:'+c+'"></i></div>'
     +(a.b[5]?'<div class="uwarn">'+a.b[5]+' not seen in 30+ days</div>':'')
     +'</button>';
  }).join('');
  return h;
}
/* one shift's line in the record: each assigned aisle scored CLEARED /
   PARTIAL / NOT TOUCHED, plus anything counted off-list, plus a flag
   when the other shift walked the same aisle the same day */
function stLogShift(lab,assigned,counted,other){
  var tot=0,u; for(u in counted) tot+=counted[u];
  var h='<div class="kw" style="padding:8px 2px 4px;font-weight:800;color:var(--dim)">'
   +lab+' &middot; '+tot+' line'+(tot===1?'':'s')+' sighted</div>';
  if(!assigned.length&&!tot){
    return h+'<div class="kw" style="padding:0 2px 8px">Nothing assigned, '
     +'nothing counted.</div>';
  }
  var seen={};
  assigned.forEach(function(o){
    var u3=o[0], n3=o[1], c3=counted[u3]||0; seen[u3]=1;
    var pill,pc;
    if(c3>=n3){pill='CLEARED';pc='var(--gd)';}
    else if(c3>0){pill='PARTIAL &middot; '+c3+' of '+n3;pc='var(--am)';}
    else{pill='NOT TOUCHED';pc='var(--rd)';}
    h+='<div class="kid"><div class="kt"><b>'+esc(u3)+'</b>'
     +'<em><span class="ghp" style="background:'+pc+'">'+pill+'</span></em></div>'
     +'<div class="kw">'+n3+' assigned &middot; '+c3+' sighted'
     +(other[u3]?' &middot; <b style="color:var(--rd)">also walked by the '
       +'other shift &mdash; overlap</b>':'')
     +'</div></div>';
  });
  var extra=[];
  for(u in counted){ if(!seen[u]) extra.push(esc(u)+' ('+counted[u]+')'); }
  if(extra.length){
    h+='<div class="kw" style="padding:4px 2px 8px">Also counted off-list: '
     +extra.join(' &middot; ')+'</div>';
  }
  return h;
}
/* ONE AREA, ZOOMED IN - freshness buckets you can tap: seen today /
   3 / 7 / 14 / 30 days / longer. Pictures and counts, never a wall of
   asset numbers - those live on the printed sheet with the scan codes. */
var _BK=[1,3,7,14,30,99];
var _BKL=['Seen today','2&ndash;3 days','4&ndash;7 days',
          '8&ndash;14 days','15&ndash;30 days','30+ days'];
function stArea(i,b){
  var s=D.stock, a=s.areas[i];
  if(!a) return;
  if(b==null){
    /* land on the oldest bucket that holds WALKABLE gear - a bucket
       of nothing but on-hire lines is not a job */
    var av0={};
    (s.rows||[]).forEach(function(x){
      if(x.u===a.u&&x.s!=='O') av0[x.b]=true;});
    for(var j=5;j>=0;j--){ if(av0[_BK[j]]){ b=_BK[j]; break; } }
    if(b==null){ for(var j2=5;j2>=0;j2--){ if(a.b[j2]>0){ b=_BK[j2]; break; } } }
    if(b==null) b=7;
  }
  var c=a.p>=90?'var(--gd)':a.p>=70?'var(--am)':'var(--rd)';
  var h='<div class="prbtns" style="margin-top:2px">'
   +'<button class="stmore" type="button" onclick="stBack()">&larr; All areas</button>'
   +'</div>'
   +'<div class="ring"><div class="rv" style="color:'+c+'">'+a.p+'%</div>'
   +'<div class="rt"><b>'+esc(a.u)
   +(a.own==='d'?' <span class="ghp" style="background:var(--am)">DAYS TODAY</span>'
    :a.own==='n'?' <span class="ghp" style="background:var(--org)">NIGHTS TODAY</span>':'')
   +'</b>'+a.s7+' of '+a.t
   +' lines sighted in the last 7 days</div></div>'
   +'<div class="stchips">'
   +_BK.map(function(bk,j){
     var n=a.b[j];
     var cc=j<2?'var(--gd)':j<4?'var(--am)':'var(--rd)';
     return '<button type="button" class="chipk'+(bk===b?' on':'')+'" '
      +(n?'onclick="stArea('+i+','+bk+')"':'disabled')+'>'
      +_BKL[j]+'<b style="color:'+cc+'">'+n+'</b></button>';
   }).join('')+'</div>';
  var rows=(s.rows||[]).filter(function(x){return x.u===a.u&&x.b===b;});
  var av=rows.filter(function(x){return x.s!=='O';});
  var oh=rows.filter(function(x){return x.s==='O';});
  function srt(l){ l.sort(function(x,y){ if(y.q!==x.q) return y.q-x.q;
    return x.n.toUpperCase()<y.n.toUpperCase()?-1:1; }); return l; }
  var old=b>7;
  function line(x){
    /* grouped, not all over the shop: one picture, the variant code,
       then the item numbers as tidy chips under it */
    var chips='';
    if(x.ii&&x.ii.length&&old){
      chips='<div class="kw" style="margin-top:5px;line-height:2">'
       +x.ii.slice(0,14).map(function(p){
         return '<span class="vcode">'+esc(p[0])+' &middot; '+p[1]+'d</span>';
       }).join(' ')
       +(x.ii.length>14?' <b>+'+(x.ii.length-14)+' more on the sheet</b>':'')
       +'</div>';
    }
    return '<div class="kid kidth">'+thTile(x.v,x.n)
     +'<div class="kbody"><div class="kt"><b>'+esc(x.n)+'</b>'
     +'<em'+(old?' class="o"':'')+'>&times;'+x.q+'</em></div>'
     +'<div class="kw">'+(x.v?'<span class="vcode">'+esc(x.v)+'</span> &middot; ':'')
     +(old?'oldest sighting <b>'+x.a+' days</b> ago'
       :'sighted inside this window')+'</div>'+chips+'</div></div>';
  }
  if(av.length){
    h+='<div class="uhead">'+(old?'Go and lay eyes on these':'Counted &mdash; ticked off')
     +' &middot; '+av.reduce(function(t2,x){return t2+x.q;},0)+'</div>'
     +srt(av).slice(0,40).map(line).join('')
     +more(40,av.length,'lines');
  }
  if(oh.length){
    h+='<div class="uhead">Out with crews &mdash; do not hunt these</div>'
     +'<div class="kw" style="padding:0 2px 6px">They count when they come '
     +'back through the window, not on a walk.</div>'
     +srt(oh).slice(0,25).map(line).join('')
     +more(25,oh.length,'lines');
  }
  var nA=s.stale.filter(function(x){return x.u===a.u&&x.s==='A';}).length;
  var nO=s.stale.filter(function(x){return x.u===a.u&&x.s==='O';}).length;
  if(nA||nO){
    h+='<div class="prbtns">'
     +(nA?'<button class="stmore" type="button" onclick="stPrint(\\''+esc(a.u)
       +'\\',\\'A\\')">&#128424; Not found sheet ('+nA+')</button>':'')
     +(nO?'<button class="stmore" type="button" onclick="stPrint(\\''+esc(a.u)
       +'\\',\\'O\\')">&#128424; On hire &mdash; do not hunt ('+nO+')</button>':'')
     +'</div>';
  }
  document.getElementById('p-stock').innerHTML=h;
}
function stBack(){document.getElementById('p-stock').innerHTML=helpBar('stock')+paneStock();}
/* THE CALL CARDS (Andrew, 31 Jul 2026: "how-to cards... dummy it down
   so anyone understands... call cards on all areas so it helps people
   with no knowledge"). One card per tab: what it is, how to use it,
   what good looks like. They sit folded at the top of every tab and
   print as a laminate deck for the counter. */
/* ==================================================================
   THE MASTER SEARCH (Andrew, 2 Aug 2026: "a master serach file that
   alow the team to seravh any company any hirer and find out all
   infomation as much infomation as possible about that company or
   hirer").

   Everything the board knows about a name, in one place: what they
   hold, how old it is, what is overdue, what is on the hit list, which
   aisles it came from, and the whole list.

   It is built off the SAME roster every other pane uses, so it can
   never disagree with the chase list or a printed sheet. One truth,
   read six ways.
================================================================== */
var WHO_IX=null, WHO_LAST=null;
function whoKey(w,co){ return String(w)+'\u001F'+String(co); }
function whoIndex(){
  if(WHO_IX) return WHO_IX;
  var co={}, pp={};
  (D.roster||[]).forEach(function(x){
    var c=x.co||'Not named', w=x.w||'Not named';
    (co[c]=co[c]||{n:c,rows:[],people:{}}).rows.push(x);
    co[c].people[w]=(co[c].people[w]||0)+1;
    var k=whoKey(w,c);
    (pp[k]=pp[k]||{n:w,co:c,rows:[]}).rows.push(x);
  });
  /* the chase list and the hit list folded onto the same names, so a
     profile shows the trouble as well as the count */
  var mark=function(list,key){
    (list||[]).forEach(function(x){
      var c=x.co||'Not named', w=x.w||'Not named', k=whoKey(w,c);
      if(co[c]) co[c][key]=(co[c][key]||0)+1;
      if(pp[k]) pp[k][key]=(pp[k][key]||0)+1;
    });
  };
  mark((D.chase&&D.chase.tools)||[], 'chase');
  var H=D.hits||{};
  ['radio','gas','bat','tool'].forEach(function(g){ mark(H[g]||[], 'hit'); });
  WHO_IX={co:co, pp:pp};
  return WHO_IX;
}
function whoAges(rows){
  var a={g:0,a:0,r:0,u:0,old:null};
  rows.forEach(function(x){
    var d=(x.d==null?null:parseInt(x.d));
    if(d==null||isNaN(d)){a.u++;return}
    if(d<=2)a.g++; else if(d<=4)a.a++; else a.r++;
    if(!a.old||d>a.old.d) a.old={d:d,n:x.n};
  });
  return a;
}
/* A SAFELY QUOTED ARGUMENT FOR AN INLINE onclick.
   Names carry apostrophes - O'Brien - and one of those ends the JS
   string early and breaks the whole button.
   Backslash-escaping is the obvious fix and it is the WRONG one here:
   this JavaScript is written out of a Python string, so every
   backslash gets halved on the way through and the escaping arrives
   mangled. (It did, and node caught it.)
   encodeURIComponent has no backslashes at all, turns the apostrophe
   into %27 and the quote into %22, and is safe inside an HTML
   attribute as well as a JS string. Every reader decodes it back. */
/*  encodeURIComponent does NOT escape the apostrophe - it is an
    unreserved mark, so O'BRIEN came through intact and ended the JS
    string early, exactly the fault this function exists to prevent.
    Caught on the rig with a deliberate O'Brien. Escape it by hand.  */
function jq(s){ return "'"+encodeURIComponent(String(s==null?'':s))
  .split("'").join('%27')+"'"; }
function jd(s){ try{ return decodeURIComponent(String(s==null?'':s)); }
                catch(e){ return String(s==null?'':s); } }
function whoBadges(o){
  var b='';
  if(o.hit) b+='<span class="wb r">'+o.hit+' on the hit list</span>';
  if(o.chase) b+='<span class="wb a">'+o.chase+' to chase</span>';
  return b;
}
function whoCoCard(o){
  var np=0; for(var k in o.people) np++;
  return '<button class="wcard" type="button" onclick="whoOpen(\\'co\\','
   +jq(o.n)+')"><div class="wtop"><b>'+esc(o.n)+'</b><em>'+o.rows.length
   +'</em></div><div class="wsub">'+np+' name'+(np===1?'':'s')
   +' &middot; '+o.rows.length+' item'+(o.rows.length===1?'':'s')
   +' on hire</div>'+(whoBadges(o)?'<div class="wbs">'+whoBadges(o)+'</div>':'')
   +'</button>';
}
function whoPpCard(o){
  return '<button class="wcard" type="button" onclick="whoOpen(\\'pp\\','
   +jq(o.n)+','+jq(o.co)+')"><div class="wtop"><b>'+esc(o.n)+'</b><em>'
   +o.rows.length+'</em></div><div class="wsub">'+esc(o.co)+'</div>'
   +(whoBadges(o)?'<div class="wbs">'+whoBadges(o)+'</div>':'')+'</button>';
}
function whoTop(){
  /* nothing typed yet - show the biggest holders, because that is who
     the counter is usually chasing anyway */
  var ix=whoIndex(), rc=[];
  for(var c in ix.co) rc.push(ix.co[c]);
  rc.sort(function(a,b){return b.rows.length-a.rows.length});
  if(!rc.length) return '';
  return '<div class="uhead">Holding the most right now</div>'
    +rc.slice(0,6).map(whoCoCard).join('');
}
function whoFind(){
  var q=(document.getElementById('wq')||{value:''}).value
        .toUpperCase().replace(/^\s+|\s+$/g,'');
  var out=document.getElementById('wout'); if(!out) return;
  if(q.length<2){
    out.innerHTML='<div class="kw" style="padding:12px 2px">Type two '
      +'letters of a company or a person&rsquo;s name. Every name the '
      +'store has gear out to is in here.</div>'+whoTop();
    return;
  }
  var ix=whoIndex(), rc=[], rp=[];
  for(var c in ix.co) if(c.toUpperCase().indexOf(q)>=0) rc.push(ix.co[c]);
  for(var k in ix.pp) if(ix.pp[k].n.toUpperCase().indexOf(q)>=0) rp.push(ix.pp[k]);
  rc.sort(function(a,b){return b.rows.length-a.rows.length});
  rp.sort(function(a,b){return b.rows.length-a.rows.length});
  if(!rc.length&&!rp.length){
    out.innerHTML='<div class="kw cut" style="padding:12px 2px">Nothing '
      +'matches &ldquo;'+esc(q)+'&rdquo;. Only names holding gear right '
      +'now are in here &mdash; somebody who has brought everything back '
      +'will not show.</div>'; return;
  }
  var h='';
  if(rc.length){
    h+='<div class="uhead">Companies &mdash; '+rc.length+'</div>'
      +rc.slice(0,30).map(whoCoCard).join('');
  }
  if(rp.length){
    h+='<div class="uhead">People &mdash; '+rp.length+'</div>'
      +rp.slice(0,40).map(whoPpCard).join('');
  }
  out.innerHTML=h;
}
function whoOpen(kind,a,b){
  a=jd(a); b=jd(b);
  var ix=whoIndex(), o=(kind==='co')?ix.co[a]:ix.pp[whoKey(a,b)];
  var out=document.getElementById('wout'); if(!out) return;
  if(!o){
    /* Not on the roster - so they are holding no hire gear. That is a
       true answer but a dead end, and a bloke who just tapped a name
       deserves better than a shrug. Show what we DO have on them: the
       consumables they have taken. Caught on the rig by tapping a name
       out of a movements sheet who had never held a tool. */
    out.innerHTML='<button class="wback" type="button" onclick="whoFind()">'
      +'&lsaquo; Back to the search</button>'
      +'<div class="wprof"><div class="wph"><div class="wpn">'+esc(a)+'</div>'
      +'<div class="wpc">'+esc(b||'')+'</div></div>'
      +'<div class="wold"><b>No hire gear in this name right now.</b> '
      +'Nothing to chase &mdash; everything is back.</div></div>'
      +whoCons(a);
    window.scrollTo(0,0); return; }
  WHO_LAST={kind:kind,a:a,b:b};
  var rows=o.rows.slice().sort(function(x,y){
    return (y.d||0)-(x.d||0) || String(x.n).localeCompare(String(y.n)); });
  var ag=whoAges(rows), kinds={}, units={};
  rows.forEach(function(x){ kinds[x.n]=1; units[x.u]=(units[x.u]||0)+1; });
  var nk=0; for(var z in kinds) nk++;
  var ul=[]; for(var u in units) ul.push([u,units[u]]);
  ul.sort(function(x,y){return y[1]-x[1]});
  var np=0; if(kind==='co'){ for(var k0 in o.people) np++; }

  var h='<button class="wback" type="button" onclick="whoFind()">'
   +'&lsaquo; Back to the search</button>'
   +'<div class="wprof"><div class="wph"><div class="wpn">'+esc(o.n)+'</div>'
   +'<div class="wpc">'+esc(kind==='co'
       ? (np+' name'+(np===1?'':'s')+' with gear out')
       : o.co)+'</div></div>'
   +'<div class="wnums">'
   +'<span><b>'+rows.length+'</b>ITEMS OUT</span>'
   +'<span><b>'+nk+'</b>DIFFERENT THINGS</span>'
   +'<span class="'+(ag.r?'r':'g')+'"><b>'+ag.r+'</b>5+ DAYS</span>'
   +'</div>'
   +'<div class="wage"><i class="g" style="flex:'+(ag.g||0.001)+'"></i>'
   +'<i class="a" style="flex:'+(ag.a||0.001)+'"></i>'
   +'<i class="r" style="flex:'+(ag.r||0.001)+'"></i></div>'
   +'<div class="wlg"><span><i class="g"></i>0&ndash;2d '+ag.g+'</span>'
   +'<span><i class="a"></i>3&ndash;4d '+ag.a+'</span>'
   +'<span><i class="r"></i>5+d '+ag.r+'</span>'
   +(ag.u?'<span><i></i>no date '+ag.u+'</span>':'')+'</div>'
   +(ag.old?'<div class="wold">Longest out: <b>'+esc(ag.old.n)
     +'</b> &middot; '+ag.old.d+' days</div>':'')
   +(whoBadges(o)?'<div class="wbs">'+whoBadges(o)+'</div>':'')
   +'</div>';

  if(ul.length>1){
    h+='<div class="uhead">Across '+ul.length+' aisles</div><div class="wunits">'
      +ul.slice(0,12).map(function(x){return '<span>'+esc(x[0])
        +' <b>'+x[1]+'</b></span>'}).join('')+'</div>';
  }
  if(kind==='co'){
    var ppl=[]; for(var w in o.people) ppl.push([w,o.people[w]]);
    ppl.sort(function(x,y){return y[1]-x[1]});
    h+='<div class="uhead">Who has it &mdash; '+ppl.length+'</div>'
     +ppl.map(function(x){
        var po=ix.pp[whoKey(x[0],o.n)]||{rows:[]};
        return '<button class="wcard" type="button" onclick="whoOpen(\\'pp\\','
          +jq(x[0])+','+jq(o.n)+')"><div class="wtop"><b>'+esc(x[0])
          +'</b><em>'+x[1]+'</em></div>'
          +(whoBadges(po)?'<div class="wbs">'+whoBadges(po)+'</div>':'')
          +'</button>'; }).join('');
  }
  h+='<div class="uhead">Every item &mdash; longest out first</div>'
   +rows.slice(0,300).map(function(x){
      return '<div class="kid kidth">'+thTile(x.v,x.n,x.va)
       +'<div class="kbody"><div class="kt"><b>'+esc(x.n)+'</b>'
       +'<em class="'+((x.d||0)>4?'o':'')+'">'
       +(x.d==null?'&mdash;':x.d+'d')+'</em></div>'
       +'<div class="kw">'+esc(x.u)
       +(x.i?' &middot; Item '+esc(x.i):'')
       +(x.p?' &middot; <b style="color:var(--org)">Plant ID '+esc(x.p)+'</b>':'')
       +(kind==='co'?' &middot; '+whoLink(x.w,x.co):'')
       +'</div></div></div>'; }).join('')
   +more(300,rows.length,'items');

  if(kind==='pp') h+=whoCons(o.n);
  h+='<div class="prbtns" style="margin-top:14px">'
   +'<button class="stmore" type="button" onclick="whoPrint()">'
   +'&#128424; Print this list</button></div>';
  out.innerHTML=h;
  window.scrollTo(0,0);
}
/* A hirer's name, wherever it appears, is a door to their profile
   (Andrew: "you also should be able to clkick on the name and it maybe
   take you to that person to find out more info"). */
/* the same door from any list row that carries a hirer */
function wl(x){ return (x&&x.w)?whoLink(x.w,x.co||''):''; }
/* every consumable movement in a person's name, gathered off the sales
   feed. Used both on a full profile and on the "holding nothing" one,
   because a bloke who has taken forty pairs of gloves and no tools is
   still somebody the store knows about. */
function whoCons(w){
  var c=D.cons; if(!c||!c.all) return '';
  var rows=[];
  c.all.forEach(function(x){
    (x.tx||[]).forEach(function(m){
      if(m.w===w) rows.push({n:x.n,k:x.k,d:m.d,q:m.q,r:m.r});
    });
  });
  if(!rows.length) return '';
  rows.sort(function(p1,p2){return String(p2.d).localeCompare(String(p1.d))});
  var took=0; rows.forEach(function(m){took+=(m.q||0)});
  return '<div class="uhead">Consumables taken &mdash; '+rows.length
   +' movement'+(rows.length===1?'':'s')+', '+Math.round(took)+' item'
   +(Math.round(took)===1?'':'s')+'</div>'
   +rows.slice(0,120).map(function(m){
     return '<div class="txr"><div class="txd">'+esc(m.d||'&mdash;')+'</div>'
      +'<div class="txw">'+esc(m.n)+'<span>SKU '+esc(m.k)+'</span></div>'
      +'<div class="txq'+((m.r||0)>0?' back':'')+'">'
      +((m.q||0)?'-'+Math.round(m.q):'')
      +((m.r||0)?' +'+Math.round(m.r):'')+'</div></div>';
   }).join('')+more(120,rows.length,'movements');
}
function whoLink(w,co){
  if(!w) return '';
  return '<span class="wlink" onclick="whoJump(event,'+jq(w)+','+jq(co)+')">'
    +esc(w)+'</span>';
}
function whoJump(ev,w,co){
  if(ev&&ev.stopPropagation) ev.stopPropagation();
  w=jd(w); co=jd(co);
  /* a name tapped inside the movements sheet has to close it on the
     way out, or the profile opens underneath and nothing appears to
     happen (caught on the rig) */
  if(typeof shClose==='function') shClose();
  nav('who');
  var i=document.getElementById('wq'); if(i) i.value=w;
  whoOpen('pp',w,co);
}
function whoPrint(){
  if(!WHO_LAST) return;
  nav('print');
  if(WHO_LAST.kind==='co'){ PRW.kind='co'; PRW.co=WHO_LAST.a; }
  else { PRW.kind='pp'; PRW.pp=whoKey(WHO_LAST.a,WHO_LAST.b); }
  prShow();
}
function paneWho(){
  return '<input class="srch" id="wq" placeholder="Any company, any name"'
   +' oninput="whoFind()" autocomplete="off" autocapitalize="characters"'
   +' spellcheck="false">'
   +'<div id="wout"></div>';
}
var HOWTO={
 who:{t:'Master search',
  w:'Everything the board knows about a company or a person, on one '
   +'screen. Somebody rings and asks &ldquo;what have my blokes got?&rdquo; '
   +'&mdash; this is the answer, without opening five tabs.',
  h:['Type two letters of a company or a name.',
     'Tap a company to see every one of its people, or tap a name to go '
     +'straight to that bloke.',
     'Anywhere you see a name in orange &mdash; on this screen, in the '
     +'finder, on a chase list &mdash; tap it and it brings you here.',
     'Print this list sends it straight to the print hub, already loaded.'],
  g:'The profile gives you the count, how old it all is, the aisles it '
   +'came out of, the longest-held item, and anything on the chase or hit '
   +'list. Only names holding gear RIGHT NOW appear &mdash; somebody who '
   +'has brought everything back will not show, and that is the answer '
   +'too.'},
 find:{t:'Find it',
  w:'The where-is-it tool. Somebody asks &ldquo;have we got a 2-tonne '
   +'lever block?&rdquo; or &ldquo;where&rsquo;s item 1312687?&rdquo; &mdash; '
   +'this answers it in seconds.',
  h:['Point the hand scanner at any barcode &mdash; a sheet, a sticker or '
     +'the gear itself. It types the number for you.',
     'No scanner? Type the item number, a Plant ID, or just a word like '
     +'&ldquo;welder&rdquo;.',
     'Read the card that comes back &mdash; that is the answer.'],
  g:'GREEN card = it&rsquo;s on the shelf &mdash; the card names the aisle '
   +'and how many. ORANGE = out with a crew &mdash; who, which company, how '
   +'many days. So you never promise gear we don&rsquo;t have, and you never '
   +'hunt for gear that&rsquo;s legitimately out.'},
 groups:{t:'Product groups',
  w:'The whole store sorted by type, like supermarket aisles &mdash; '
   +'batteries, leads, rigging, hand tools and so on.',
  h:['Tap a group to open it and see every line inside.',
     'Green number = on the shelf right now. Orange = out with crews.',
     'Or type in the search box and jump straight to the thing.'],
  g:'Use it whenever someone asks &ldquo;have we got&hellip;&rdquo;. The '
   +'green/orange bar shows in one look whether a group is running dry.'},
 chase:{t:'Chase up',
  w:'Gear that&rsquo;s been out longer than its return rule. Radios and gas '
   +'monitors come back daily; Milwaukee tooling gets three days.',
  h:['Open a group and read the lines: who has it, their company, how many '
     +'days.',
     'One name at a time &mdash; ring them or walk to their crew.',
     'When it comes through the window, it drops off this list by itself '
     +'tomorrow morning.'],
  g:'A shrinking number. If your group grows day after day, say so at '
   +'handover &mdash; don&rsquo;t let it become demob&rsquo;s problem.'},
 hits:{t:'Hit list',
  w:'Today&rsquo;s worst overdues boiled down to one short list &mdash; the '
   +'stuff that actually matters this shift.',
  h:['Print the walk-around sheet (button is right there).',
     'One lap of site &mdash; every stop is on the sheet with a scan code.',
     'Tick each line as gear lands back, hand the sheet in at handover.'],
  g:'An empty hit list by end of shift. One lap, done properly, clears it.'},
 print:{t:'Print hub',
  w:'The one-stop print shop. Every report the board can make, picked '
   +'from one dropdown and walked through step by step &mdash; branded '
   +'COATES paper or an Outlook email.',
  h:['STEP 1: pick the report from the dropdown &mdash; one person, one '
     +'company, whole site, radios, gas, hit list, or the call-card deck.',
     'STEP 2: pick who it&rsquo;s for &mdash; type a name, or go company '
     +'first then the person.',
     'STEP 3: pick how it reads &mdash; grouped by worker (names A to Z) '
     +'or one big table; products A to Z or longest out first.',
     'Look at the preview, then tap Print (store printer or Save as PDF) '
     +'or Email via Outlook. Nothing prints before you&rsquo;ve seen it.'],
  g:'A company asks &ldquo;what do we have on hire?&rdquo; and walks away '
   +'with the answer on paper, same visit. That&rsquo;s the standard.'},
 fresh:{t:'Fresh look',
  w:'What moved since yesterday morning &mdash; everything that came back '
   +'and everything that went out.',
  h:['Open it at the start of your shift.',
     'Scan the returns &mdash; anything back early or late tells a story.',
     'Use the print button if you want it on paper for the handover.'],
  g:'You start the shift already knowing what happened &mdash; nobody has '
   +'to tell you.'},
 stock:{t:'Stocktake',
  w:'Proof we&rsquo;ve laid eyes on everything we own. Every storage area '
   +'has a percentage &mdash; 100% means every line in it was sighted in '
   +'the last 7 days. The store&rsquo;s overall score is the big ring.',
  h:['Find your aisle under TODAY&rsquo;S COUNTING ORDERS and tap it &mdash; '
     +'the board picks the aisles, you don&rsquo;t have to choose.',
     'Work the OLDEST bucket first (30+ days, then 15&ndash;30, and so on). '
     +'The buckets are the row of pills &mdash; tap one to see what&rsquo;s '
     +'in it, with pictures.',
     'Print the NOT-FOUND SHEET for the aisle. Walk it, scan or tick every '
     +'line you lay eyes on, sign it, hand it in.',
     'Gear marked &ldquo;out with crews&rdquo; is NOT hunted &mdash; it '
     +'counts itself when it comes back through the window.',
     'Stay on YOUR shift&rsquo;s aisles &mdash; every area card wears a '
     +'DAYS TODAY or NIGHTS TODAY badge. THE RECORD underneath shows what '
     +'each shift actually counted, straight off the scanner timestamps '
     +'&mdash; cleared, partial, or not touched. It writes itself.'],
  g:'Your aisles sitting green (90%+) at handover. And the WALKED PAST '
   +'list shrinking: if something isn&rsquo;t there after two proper looks, '
   +'flag it to Andrew the same day. Reporting a missing item early is a '
   +'win for honesty &mdash; sitting on it is the only fail.'},
 aisle:{t:'Walk an aisle',
  w:'One aisle, everything about it, on one screen &mdash; built for '
   +'standing in the aisle with a phone in your hand.',
  h:['Pick the aisle you&rsquo;re standing in.',
     'It shows what should be on the shelf, what&rsquo;s out, what to '
     +'chase and what hasn&rsquo;t been counted.',
     'Sort one aisle completely before moving to the next.'],
  g:'The aisle you just walked has nothing left to ask you.'},
 battle:{t:'Day v Night',
  w:'The shift-against-shift scorecard &mdash; who counted more, who '
   +'matched more, night by night. Bragging rights only, but the numbers '
   +'are real.',
  h:['Have a look at the start of shift.',
     'The bar chart shows each day&rsquo;s scores side by side.',
     'Win the night. That&rsquo;s it, that&rsquo;s the instruction.'],
  g:'Both shifts pushing the stocktake percentage up because neither wants '
   +'to lose. Rivalry does what memos can&rsquo;t.'},
 cons:{t:'Consumables',
  w:'The consumables shelf &mdash; gloves, discs, tape, batteries-in-'
   +'packets. Pictures, live counts, what&rsquo;s gone out, and what needs '
   +'ordering before the next delivery cutoff.',
  h:['The tiles up top are the story: on the shelf, issued so far, and TO '
     +'ORDER in red.',
     'Scroll the A&ndash;Z shelf &mdash; every line has its picture and '
     +'scan code.',
     'Print the stock check &amp; reorder sheet when you do the count.'],
  g:'Nothing hits zero unannounced. The 13 in red today is tomorrow&rsquo;s '
   +'order, placed in time.'},
 std:{t:'Our standards',
  w:'The rules of this store &mdash; what we promise every person at the '
   +'window, and what we expect back.',
  h:['Read it once properly on your first day.',
     'When a call is unclear mid-shift, this is the tie-breaker.',
     'Hold each other to it &mdash; politely.'],
  g:'Be nice, be polite &mdash; we&rsquo;re here to help. Every visitor '
   +'leaves with the right gear and a straight answer.'},
 arr:{t:'Arriving',
  w:'What&rsquo;s already on order and inbound to the store.',
  h:['Check here BEFORE you re-order anything.',
     'Each line shows what&rsquo;s coming and how many.',
     'If a crew is waiting on it, you can tell them it&rsquo;s inbound.'],
  g:'No double-orders, no &ldquo;didn&rsquo;t know it was coming&rdquo;.'},
 plant:{t:'Plant',
  w:'The machines &mdash; welders, generators, forklifts, towers &mdash; '
   +'sorted by type, split into FREE to hire, IDLE on the ground, and OUT '
   +'with a company.',
  h:['Tap a category to see its machines &mdash; Plant IDs ride on every '
     +'line.',
     'Machines with a &#128221; LOG BOOK chip need their daily pre-start '
     +'filled in &mdash; remind the crew taking one.',
     'Print the IDLE PLANT AUDIT SHEET to physically sight machines, or '
     +'the DEMOB CHECKLIST when we&rsquo;re packing the job up.'],
  g:'Every machine accounted for on paper anyone can follow: back &rarr; '
   +'off-hired &rarr; picked up.'},
 idle:{t:'Idle plant',
  w:'Machines that are on hire but parked doing nothing &mdash; costing '
   +'without earning.',
  h:['Scan the list &mdash; anything sitting long is a question.',
     'Ask the holder: still needed, or can it come back?',
     'Tell Andrew what can come off &mdash; he runs the off-hire.'],
  g:'A short list. Machines either work here or go home.'}};
function helpBar(k){
  var c=HOWTO[k]; if(!c) return '';
  return '<div class="grp hgrp"><button type="button" onclick="tog(this)">'
   +'<div class="gn"><b>&#10067; New here? How &ldquo;'+c.t
   +'&rdquo; works</b><span>plain English &middot; 20 seconds &middot; tap to open</span></div>'
   +'</button><div class="kids"><div class="hcard">'
   +'<div class="hs"><b>WHAT IT IS</b><span>'+c.w+'</span></div>'
   +'<div class="hs"><b>HOW TO USE IT</b><ol>'
   +c.h.map(function(x){return '<li>'+x+'</li>';}).join('')+'</ol></div>'
   +'<div class="hs"><b>WHAT GOOD LOOKS LIKE</b><span>'+c.g+'</span></div>'
   +'<div class="prbtns" style="margin-top:10px"><button class="stmore" '
   +'type="button" onclick="howtoPrint()">&#128424; Print the full deck '
   +'&mdash; laminate for the counter</button></div>'
   +'</div></div></div>';
}
/* the printable deck - one call card per tab, big type, laminate-ready */
function howtoPrint(){
  var order=['find','groups','chase','hits','stock','aisle','cons','plant',
             'idle','fresh','arr','print','battle','std'];
  var body=order.map(function(k){
    var c=HOWTO[k]; if(!c) return '';
    return '<div class="hcv">'
     +'<div class="hct">'+c.t+'</div>'
     +'<div class="hcs"><b>WHAT IT IS</b> '+c.w+'</div>'
     +'<div class="hcs"><b>HOW TO USE IT</b><ol>'
     +c.h.map(function(x){return '<li>'+x+'</li>';}).join('')+'</ol></div>'
     +'<div class="hcs"><b>WHAT GOOD LOOKS LIKE</b> '+c.g+'</div>'
     +'</div>';
  }).join('');
  /* prFrame titles are esc()'d - literal dash, never &mdash; here */
  prFrame('The stores board — call cards',
   'one card per tab &middot; plain English &middot; print, cut, laminate, '
    +'keep at the counter &mdash; built for a first-day start with zero '
    +'knowledge',
   ASOF, body, false);
}
/* THE NOT-FOUND SHEET - one aisle, one piece of paper, a tick box per
   line. kind A = available: should be on the shelf, go and lay eyes
   on it. kind O = on hire: out with a crew, printed so the counter
   knows what NOT to hunt for. */
function stPrint(unit,kind){
  var rows=D.stock.stale.filter(function(x){
    return x.u===unit && x.s===kind;});
  if(!rows.length) return;
  rows=rows.slice().sort(function(a,b){
    if(a.d!==b.d) return b.d-a.d;
    return a.n.toUpperCase()<b.n.toUpperCase()?-1:1;});
  var isA=kind==='A';
  var body='<table class="ptab pchk">'
   +'<tr><th>Item</th><th>Item no</th><th>Scan</th><th class="pn">Days uncounted</th>'
   +'<th>Last sighted by</th>'
   +(isA?'<th class="pn">FOUND</th><th>Where / note</th>':'')
   +'</tr>'
   +rows.map(function(x){
     return '<tr><td>'+esc(x.n)+'</td><td>'+esc(x.i||'—')+'</td>'
       +'<td class="pqr">'+qr(x.i)+'</td>'
       +'<td class="pn">'+x.d+'</td><td>'+esc(x.by||'—')+'</td>'
       +(isA?'<td class="pn"><span class="ptick"></span></td>'
            +'<td><span class="pline"></span></td>':'')
       +'</tr>';
   }).join('')+'</table>'
   +(isA?'<div class="pwho" style="margin-top:14px">Walked by '
     +'<span class="pline"></span> &nbsp; Date <span class="pline short"></span></div>':'');
  prFrame(
    isA?('Stocktake — '+unit+' — not yet found')
       :('Stocktake — '+unit+' — on hire, do not hunt'),
    rows.length+' item'+(rows.length===1?'':'s')+' &middot; '
     +(isA?'should be on the shelf in this aisle &mdash; tick each one as you lay eyes on it'
          :'out with crews right now &mdash; they count when they come back, not on the walk'),
    ASOF, body, false);
}
/* WALK AN AISLE - one screen per aisle, for a bloke standing in it.
   Everything about that aisle in one place: what should be on the shelf,
   what is out, what to chase, what has not been counted. He is holding
   the phone in the aisle, so the aisle is the unit of thought - not the
   product group. (Andrew, 29 Jul 2026: "when we walk the aisle, be good
   to click on it and it gives you the detailed list of what you're
   looking for and in which aisle") */
function paneAisle(){
  var A={};
  function slot(u){ return A[u]=A[u]||{shelf:[],out:[],chase:[],stale:[]}; }
  D.groups.forEach(function(g){
    Object.keys(g.u||{}).forEach(function(u){
      slot(u).shelf.push({n:g.n,q:g.u[u]});
    });
    (g.who||[]).forEach(function(w){ slot(w.u).out.push({n:g.n,w:w.w,co:w.co,d:w.d}); });
  });
  D.chase.tools.forEach(function(x){ slot(x.u).chase.push(x); });
  D.chase.plant.forEach(function(x){ slot(x.u).chase.push(x); });
  (D.stock.stale||[]).forEach(function(x){ slot(x.u).stale.push(x); });

  var h='<div class="note"><b>Pick the aisle you are standing in.</b> '
   +'Everything about it in one place &mdash; what should be on the shelf, '
   +'what is out, what to chase, and what has not been counted.</div>';
  Object.keys(A).sort(function(a,b){
    return (A[b].shelf.length+A[b].chase.length)-(A[a].shelf.length+A[a].chase.length);
  }).forEach(function(u){
    var a=A[u];
    var onShelf=a.shelf.reduce(function(s,x){return s+x.q},0);
    h+='<div class="grp"><button type="button" onclick="tog(this)">'
      +'<div class="gn"><b>'+esc(u)+'</b><span>'+a.shelf.length+' lines &middot; '
      +a.out.length+' out</span></div>'
      +'<div class="gq"><b>'+onShelf+'</b><span>on shelf</span></div>'
      +(a.chase.length?'<div class="gq"><b style="color:var(--am)">'+a.chase.length
        +'</b><span>chase</span></div>':'')
      +(a.stale.length?'<div class="gq"><b style="color:var(--rd)">'+a.stale.length
        +'</b><span>uncounted</span></div>':'')
      +'</button><div class="kids">';
    if(a.chase.length){
      h+='<div class="uhead">Look for these first &mdash; out over 3 days</div>';
      h+=a.chase.slice(0,40).map(function(x){
        return '<div class="kid'+(x.i?' hasqr':'')+'">'
          +(x.i?'<div class="kqr">'+qr(x.i,56)+'</div>':'')
          +'<div class="kt"><b>'+esc(x.n)+'</b>'
          +'<em class="o">'+x.d+' days</em></div>'
          +'<div class="kw"><b>'+wl(x)+'</b> &middot; '+esc(x.co)
          +(x.i?' &middot; Item '+esc(x.i):'')+'</div></div>';
      }).join('')+more(40,a.chase.length,'items');
    }
    if(a.stale.length){
      h+='<div class="uhead">Not counted in over a week</div>';
      h+=a.stale.slice(0,40).map(function(x){
        return '<div class="kid"><div class="kt"><b>'+esc(x.n)+'</b>'
          +'<em class="o">'+x.d+' days</em></div></div>';
      }).join('')+more(40,a.stale.length,'assets');
    }
    h+='<div class="uhead">Should be on this shelf</div>';
    h+=a.shelf.sort(function(x,y){return y.q-x.q}).slice(0,80).map(function(x){
      return '<div class="kid"><div class="kt"><b>'+esc(x.n)+'</b>'
        +'<em>'+x.q+'</em></div></div>';
    }).join('');
    h+=more(80,a.shelf.length,'lines');
    h+='</div></div>';
  });
  return h;
}
/* DAY v NIGHT - the power bar. Issues counted at the start, returns at
   the end, so a tool taken on days and dropped back at night scores one
   for each crew: both did a piece of work. */
function paneBattle(){
  var b=D.battle;
  if(!b) return '<div class="note">No transaction export this morning, so '
    +'the ladder could not be built.</div>';
  var h='<div class="score">'
   +'<div class="sc day"><b>'+b.dayWins+'</b><span>Day shift</span>'
   +'<em>'+b.dayTotal.toLocaleString()+' movements</em></div>'
   +'<div class="vs">v</div>'
   +'<div class="sc night"><b>'+b.nightWins+'</b><span>Night shift</span>'
   +'<em>'+b.nightTotal.toLocaleString()+' movements</em></div></div>';
  h+='<div class="note"><b>Every issue and every return, shift by shift, '
   +'since the shut started.</b> Nights run 18:00 to 06:00, so a movement at '
   +'two in the morning counts for the night before &mdash; the crew that was '
   +'actually standing there.'
   +(b.nightsFrom?' Nights started '+b.nightsFrom+', so only the days both '
     +'crews were running are scored.':'')+'</div>';
  b.days.slice().reverse().forEach(function(x){
    var tot=x.D+x.N||1;
    var w=x.w==='D'?'<span class="win d">Day</span>'
        :x.w==='N'?'<span class="win n">Night</span>'
        :x.w==='T'?'<span class="win t">Tie</span>'
        :'<span class="win x">days only</span>';
    h+='<div class="brow"><div class="bd"><span>'+x.d+'</span>'+w+'</div>'
      +'<div class="bb">'
      +(x.D?'<i class="bd2" style="width:'+(100*x.D/tot)+'%">'+(x.D>tot*0.12?x.D:'')+'</i>':'')
      +(x.N?'<i class="bn" style="width:'+(100*x.N/tot)+'%">'+(x.N>tot*0.12?x.N:'')+'</i>':'')
      +'</div>'
      +'<div class="bnums"><span>DAY '+x.di+' out &middot; '+x.dr+' back</span>'
      +'<span>NIGHT '+x.ni+' out &middot; '+x.nr+' back</span></div></div>';
  });
  /* THE STOCK COUNT BATTLE (Andrew, 30 Jul 2026: "how many stock
     counts are being done between nightshift and dayshift... a power
     bar who what shift is doing what"). Same shift windows as the
     movements above - day 06:00-18:00, night 18:00-06:00 rolling into
     the next morning - and every count scored by its OWN sighting
     timestamp, so a 02:00 count belongs to the crew that stood there. */
  var sd=b.dayStock||0, sn=b.nightStock||0;
  if(sd+sn>0){
    h+='<div class="uhead">Stock counts &mdash; shift by shift</div>'
     +'<div class="score">'
     +'<div class="sc day"><b>'+sd.toLocaleString()+'</b><span>Day shift</span>'
     +'<em>items sighted</em></div>'
     +'<div class="vs">v</div>'
     +'<div class="sc night"><b>'+sn.toLocaleString()+'</b><span>Night shift</span>'
     +'<em>items sighted</em></div></div>';
    b.days.slice().reverse().forEach(function(x){
      if(!(x.ds+x.ns)) return;
      var tot=x.ds+x.ns;
      h+='<div class="brow"><div class="bd"><span>'+x.d+'</span>'
        +(x.ds>x.ns?'<span class="win d">Day</span>'
          :x.ns>x.ds?'<span class="win n">Night</span>'
          :'<span class="win t">Tie</span>')+'</div>'
        +'<div class="bb">'
        +(x.ds?'<i class="bd2" style="width:'+(100*x.ds/tot)+'%">'+(x.ds>tot*0.12?x.ds:'')+'</i>':'')
        +(x.ns?'<i class="bn" style="width:'+(100*x.ns/tot)+'%">'+(x.ns>tot*0.12?x.ns:'')+'</i>':'')
        +'</div>'
        +'<div class="bnums"><span>DAY '+x.ds+' sighted</span>'
        +'<span>NIGHT '+x.ns+' sighted</span></div></div>';
    });
    /* the names behind the bars - a scoreboard nobody is on is a
       scoreboard nobody reads */
    var wD=b.whoD||[], wN=b.whoN||[];
    if(wD.length||wN.length){
      h+='<div class="uhead">Who is doing the counting</div>';
      [['Day shift',wD],['Night shift',wN]].forEach(function(row){
        if(!row[1].length) return;
        h+='<div class="grp"><button type="button" onclick="tog(this)">'
          +'<div class="gn"><b>'+row[0]+'</b><span>top counters, whole shut</span></div>'
          +'<div class="gq"><b>'+row[1].length+'</b><span>named</span></div>'
          +'</button><div class="kids on">'
          +row[1].map(function(w){
            return '<div class="kid"><div class="kt"><b>'+esc(w[0])+'</b>'
              +'<em>'+w[1].toLocaleString()+' sighted</em></div></div>';
          }).join('')+'</div></div>';
      });
    }
    h+='<div class="note">Today&rsquo;s counting orders for each shift live on '
      +'the <b>Stocktake</b> tab &mdash; each crew has its aisles named, '
      +'with the not-found sheet ready to print.</div>';
  }
  return h;
}
/* OUR STANDARDS - quoted from SWMS-CTS-001 Rev 4, not paraphrased. The
   store's own procedure is the authority; this screen is a reminder of
   it at the counter, never a replacement for signing on to it. */
/* HIT LIST - the three overdue rules, holder first.
   (Andrew, 29 Jul 2026: "who has not brought back radios after one day
   and milwaukee batteries after 1 day and milwaukee tooling after 3
   days.") Named by person because that is how you chase gear - you do
   not ring an item description. */
function paneHits(){
  var H=D.hits;
  var rules=[
    ['radio','Radios','allowed a day &mdash; these have been kept longer',H.radio],
    ['gas','Gas monitors','allowed a day &mdash; these have been kept longer',H.gas||[]],
    ['bat','Milwaukee batteries','allowed a day &mdash; these have been kept longer',H.bat],
    ['tool','Milwaukee tooling','allowed three days &mdash; these have been kept longer',H.tool]];
  var h='<div class="note"><b>The hit list.</b> Gear with a hard return '
   +'rule that is past it. Radios, gas monitors and Milwaukee batteries '
   +'come back daily; Milwaukee tooling gets three days. Same four rules '
   +'as the daily hit list in the report pack. Walk this list before '
   +'smoko and most of it walks back in.'
   +'<br><button class="stmore" style="margin-top:10px" type="button" '
   +'onclick="prSet(\\'hits\\');prGo()">&#128424; Print the walk-around sheet</button>'
   +'</div>';
  rules.forEach(function(rl){
    if(!rl[3].length){
      h+='<div class="grp"><button type="button" onclick="tog(this)">'
        +'<div class="gn"><b>'+rl[1]+'</b><span>'+rl[2]+'</span></div>'
        +'<div class="gq"><b style="color:var(--gd)">0</b><span>clear</span>'
        +'</div></button><div class="kids"><div class="kw" style="padding:8px 2px">'
        +'Nothing overdue. Whoever is running this rule, keep going.</div>'
        +'</div></div>';
      return;
    }
    h+='<div class="grp"><button type="button" onclick="tog(this)">'
      +'<div class="gn"><b>'+rl[1]+'</b><span>'+rl[2]+'</span></div>'
      +'<div class="gq"><b style="color:var(--rd)">'+rl[3].length+'</b>'
      +'<span>to chase</span></div></button><div class="kids">'
      +rl[3].map(function(x){
        return '<div class="kid kidth'+(x.i?' hasqr':'')+'">'+thTile(x.v,x.n,x.va)
          +'<div class="kbody">'
          +(x.i?'<div class="kqr">'+qr(x.i,56)+'</div>':'')
          +'<div class="kt"><b>'+wl(x)+'</b>'
          +'<em class="o">'+x.d+' day'+(x.d===1?'':'s')+'</em></div>'
          +'<div class="kw">'+esc(x.co)+' &middot; '+esc(x.n)
          +(x.i?' &middot; Item '+esc(x.i):'')
          +(x.p?' &middot; <b style="color:var(--org)">Plant ID '+esc(x.p)+'</b>':'')
          +'</div></div></div>';
      }).join('')+'</div></div>';
  });
  return h;
}

/* PRINT & SEND - clean Coates printouts of anything on hire, straight
   from the counter. (Andrew, 29 Jul 2026: "section for printing onhire
   reports via company and person... radio print out of all onhire and
   how long for. same with gas monitors. this can be sent via phone to
   outlook email or wifi printer.")

   Print goes through the phone or laptop's own print dialog - that is
   where the Wi-Fi printer lives. Email opens the phone's Outlook with
   the report already written into the body; a static offline page
   cannot attach a file to an email, so the body IS the report, and the
   print button is there when paper is wanted. */
/* THE PRINT HUB WIZARD (Andrew, 31 Jul 2026: "one stop shop for all
   your printing needs... a drop down box to advise what report we
   want... then once picked there is more options to do with that kind
   of report and so on, until you're comfortable enough to print").

   One dropdown carries EVERY report this board can print. Picking one
   opens the next step - who it's for - then how it should read
   (workers A to Z, products A to Z, or longest out first), then the
   preview, and only then the print button. Nothing prints before
   you've seen it. Workers print their own gear from their My Gear
   page - this hub is the stores team's. */
function panePrint(){
  return '<div class="note"><b>Every print, one place, step by step.</b> '
   +'Pick the report, pick who it&rsquo;s for, pick how it reads, look '
   +'at the preview, then print. Print uses this device&rsquo;s print '
   +'menu &mdash; pick the store Wi-Fi printer there, or Save as PDF. '
   +'Email opens Outlook with the report written in.</div>'
   +'<div class="prstep">Step 1 &mdash; what do you want to print?</div>'
   +'<select id="prkind" class="srch" onchange="prWiz(1)">'
   +'<option value="">Pick a report&hellip;</option>'
   +'<option value="pp">One person &mdash; their personal on-hire report</option>'
   +'<option value="co">One company &mdash; everything they have on hire</option>'
   +'<option value="all">Whole site &mdash; everything on hire ('+D.roster.length+' items)</option>'
   +(D.hitN?'<option value="hits">Hit list &mdash; overdue returns walk-around ('+D.hitN+')</option>':'')
   +'<option value="radios">Radios on hire</option>'
   +'<option value="gas">Gas monitors on hire</option>'
   +'<option value="deck">How-to call cards &mdash; the counter deck</option>'
   +'</select>'
   +'<div id="prsteps"></div><div id="prout"></div>'
   +'<div class="uhead">Prints that live with their job</div>'
   +'<div class="prpick">'
   +'<button class="stmore" type="button" onclick="nav(\\'stock\\')">Not-found sheet &mdash; in Stocktake</button>'
   +'<button class="stmore" type="button" onclick="nav(\\'aisle\\')">Aisle sheets &mdash; in Walk an aisle</button>'
   +(D.hasCons?'<button class="stmore" type="button" onclick="nav(\\'cons\\')">Stock check &amp; reorder &mdash; in Consumables</button>':'')
   +(D.hasPlant?'<button class="stmore" type="button" onclick="nav(\\'plant\\')">Idle plant audit &amp; demob &mdash; in Plant</button>':'')
   +'</div>';
}
var PRCUR=null;
/* the wizard's memory: what report, who for, and how it should read */
var PRW={kind:'',co:'',pp:'',pco:'',q:'',layout:'grp',order:'days'};
var PRPPL=[];
function prWiz(reset){
  var sel=document.getElementById('prkind');
  var k=sel?sel.value:PRW.kind;
  if(reset){
    PRW.kind=k; PRW.co=''; PRW.pp=''; PRW.pco=''; PRW.q='';
    /* sensible starting shape per report - still changeable below */
    PRW.layout=(k==='all')?'flat':'grp';
    PRW.order='days';
  }
  var s=document.getElementById('prsteps'), h='';
  if(!k){ s.innerHTML=''; prShow(); return; }
  if(k==='deck'){
    s.innerHTML='<div class="prstep">Step 2 &mdash; it&rsquo;s one deck, '
     +'ready to go</div>'
     +'<div class="prbtns"><button class="stmore" type="button" '
     +'onclick="howtoPrint()">&#128424; Print the call-card deck</button></div>';
    document.getElementById('prout').innerHTML=''; PRCUR=null; return;
  }
  if(k==='pp') h+=prStepPerson();
  if(k==='co') h+=prStepCompany();
  h+=prStepOpts(k);
  s.innerHTML=h;
  if(k==='pp') prPplFilter();
  prShow();
}
function prCoOpts(sel){
  var cos={};
  D.roster.forEach(function(x){ cos[x.co]=(cos[x.co]||0)+1; });
  return Object.keys(cos).sort().map(function(c){
    return '<option value="'+esc(c)+'"'+(c===sel?' selected':'')+'>'
      +esc(c)+' ('+cos[c]+' items)</option>';}).join('');
}
function prStepCompany(){
  return '<div class="prstep">Step 2 &mdash; which company?</div>'
   +'<select id="prco" class="srch" '
   +'onchange="PRW.co=this.value;prWiz()">'
   +'<option value="">Pick a company&hellip;</option>'+prCoOpts(PRW.co)
   +'</select>';
}
function prStepPerson(){
  return '<div class="prstep">Step 2 &mdash; who is it for?</div>'
   +'<input class="prq" id="prq" type="search" autocomplete="off" '
   +'placeholder="Type a name&hellip;" value="'+esc(PRW.q)+'" '
   +'oninput="PRW.q=this.value;prPplFilter()">'
   +'<div class="kw" style="padding:0 2px 6px">&hellip;or go company '
   +'first, then the person:</div>'
   +'<select id="prco2" class="srch" '
   +'onchange="PRW.pco=this.value;prPplFilter()">'
   +'<option value="">Any company</option>'+prCoOpts(PRW.pco)+'</select>'
   +'<div id="prppl"></div>';
}
function prPplFilter(){
  var box=document.getElementById('prppl'); if(!box) return;
  var q=(PRW.q||'').toLowerCase(), co=PRW.pco||'';
  var pps={};
  D.roster.forEach(function(x){
    if(co&&x.co!==co) return;
    var k=x.w+'\\u001F'+x.co; pps[k]=(pps[k]||0)+1;
  });
  PRPPL=Object.keys(pps).sort().filter(function(k){
    return !q||k.split('\\u001F')[0].toLowerCase().indexOf(q)>=0;});
  if(!PRPPL.length){
    box.innerHTML='<div class="kw" style="padding:8px 2px">No one matches '
     +'that&hellip; shorter word? They may also have nothing on hire.</div>';
    return;
  }
  var shown=PRPPL.slice(0,30);
  box.innerHTML='<div class="prpick">'+shown.map(function(k,i){
    var p=k.split('\\u001F');
    return '<button class="chip'+(k===PRW.pp?' on':'')+'" type="button" '
     +'onclick="prPickPp('+i+')">'+esc(p[0])+' &middot; '+esc(p[1])
     +' ('+pps[k]+')</button>';
  }).join('')+'</div>'
  +(PRPPL.length>shown.length?'<div class="kw" style="padding:2px">'
    +(PRPPL.length-shown.length)+' more &mdash; keep typing to narrow it.</div>':'');
}
function prPickPp(i){
  PRW.pp=PRPPL[i]||'';
  prWiz();
}
function prChip(f,v,n){
  return '<button type="button" class="chip'+(PRW[f]===v?' on':'')
   +'" onclick="PRW.'+f+'=\\''+v+'\\';prWiz()">'+n+'</button>';
}
function prStepOpts(k){
  var h='', step=(k==='co'||k==='pp')?'Step 3':'Step 2';
  if(k==='hits') return '';
  if(k==='co'&&!PRW.co) return '';
  if(k==='pp'&&!PRW.pp) return '';
  if(k==='co'||k==='all'){
    h+='<div class="prstep">'+step+' &mdash; how should it read?</div>'
     +'<div class="prchips">'
     +prChip('layout','grp','Grouped by worker &mdash; names A to Z')
     +prChip('layout','flat','One big table &mdash; no grouping')
     +'</div>';
  } else {
    h+='<div class="prstep">'+step+' &mdash; what order?</div>';
  }
  h+='<div class="prchips">'
   +prChip('order','days','Longest out first')
   +prChip('order','az','Products A to Z')
   +'</div>';
  return h;
}
/* items in the order the wizard asked for - products A to Z, or days
   on hire highest to lowest (ties break the other way) */
function prOrd(list){
  var az=(PRW.order==='az');
  return list.slice().sort(function(a,b){
    var an=a.n.toUpperCase(), bn=b.n.toUpperCase();
    var da=(a.d==null?-1:a.d), db=(b.d==null?-1:b.d);
    if(az){ if(an!==bn) return an<bn?-1:1; return db-da; }
    if(da!==db) return db-da;
    return an<bn?-1:(an>bn?1:0);
  });
}
function prRows(kind){
  var v;
  if(kind==='hits'){
    /* the walk-around sheet: every overdue item under its rule, so one
       lap of the site clears the lot (Andrew, 29 Jul 2026: "add the
       hit list print") */
    var rows=[];
    [['Radios - allowed a day',D.hits.radio],
     ['Gas monitors - allowed a day',D.hits.gas||[]],
     ['Milwaukee batteries - allowed a day',D.hits.bat],
     ['Milwaukee tooling - allowed three days',D.hits.tool]]
     .forEach(function(rl){ rl[1].forEach(function(x){
       rows.push({n:x.n,u:rl[0],w:x.w,co:x.co,d:x.d,i:x.i,p:x.p}); }); });
    return {t:'Hit list — overdue returns', r:rows,
            sub:'past their return rule — the rule sits beside each item; one visit per name clears the lot'};
  }
  if(kind==='all'){
    /* the whole book in one sheet: every item on hire across every
       company, with who and company on each row (Andrew, 31 Jul 2026:
       "Whole Company Onhire report with all details, alphabetical
       order, with company as well") - the wizard picks the layout */
    return {t:'Whole site — everything on hire',
      r:D.roster.slice()};
  }
  if(kind==='radios') return {t:'Radios on hire',
    r:D.roster.filter(function(x){return x.u==='Radios'})};
  if(kind==='gas') return {t:'Gas monitors on hire',
    r:D.roster.filter(function(x){return x.u==='Gas Monitors'})};
  if(kind==='co'){v=PRW.co;
    return v?{t:v+' — gear on hire',
      r:D.roster.filter(function(x){return x.co===v})}:null;}
  if(kind==='pp'){v=PRW.pp;
    if(!v) return null;
    var p=v.split('\\u001F');
    return {t:p[0]+' ('+p[1]+') — personal on-hire report',
      r:D.roster.filter(function(x){return x.w===p[0]&&x.co===p[1]})};}
  return null;
}
/* kept for the Hit list pane's one-tap "print the walk-around sheet" -
   and any old habit of calling prSet directly */
function prSet(kind){ PRW.kind=kind; prShow(); }
function prShow(){
  var out=document.getElementById('prout'); if(!out) return;
  if(!PRW.kind){ out.innerHTML=''; PRCUR=null; return; }
  var got=prRows(PRW.kind);
  if(!got||!got.r.length){
    out.innerHTML=got?'<div class="kw" style="padding:10px 2px">Nothing on hire there right now.</div>':'';
    PRCUR=null; return;
  }
  got.r=prOrd(got.r);
  PRCUR=got; PRCUR.kind=PRW.kind;
  var h='<div class="prstep">Last look &mdash; then print</div>'
   +'<div class="uhead">'+esc(got.t)+' &mdash; '+got.r.length+' item'
   +(got.r.length===1?'':'s')+'</div>'
   +'<div class="prbtns">'
   +'<button class="stmore" type="button" onclick="prGo()">&#128424; Print / save PDF</button>'
   +'<a class="stmore" id="prmail" href="#">&#9993; Email via Outlook</a>'
   +'</div>'
   +got.r.slice(0,200).map(function(x){
     return '<div class="kid kidth">'+thTile(x.v,x.n,x.va)
       +'<div class="kbody"><div class="kt"><b>'+esc(x.n)+'</b>'
       +'<em class="'+((x.d||0)>4?'o':'')+'">'+(x.d==null?'&mdash;':x.d+'d')+'</em></div>'
       +'<div class="kw">'+wl(x)+' &middot; '+esc(x.co)+' &middot; '+esc(x.u)
       +(x.i?' &middot; Item '+esc(x.i):'')
       +(x.p?' &middot; <b style="color:var(--org)">Plant ID '+esc(x.p)+'</b>':'')
       +'</div></div></div>';
   }).join('')+more(200,got.r.length,'items');
  out.innerHTML=h;
  var m=document.getElementById('prmail');
  m.href=prMailto(got);
}
function prMailto(got){
  /* NL is built with fromCharCode so this survives living inside a
     Python template - a literal backslash-n here has already been
     eaten once by Python and broken the whole script block.
     (Caught 29 Jul 2026 by node --check on the built page.) */
  var NL=String.fromCharCode(10), RULE=Array(41).join('-');
  var b='COATES | K2 TOOL STORE - ON HIRE'+NL+got.t.replace(/—/g,'-')
   +NL+'As at '+ASOF+NL+RULE+NL;
  var lines=got.r.map(function(x){
    return (x.d==null?'-':x.d+'d')+'  '+x.n
      +(x.i?' (Item '+x.i+')':'')+(x.p?' [Plant ID '+x.p+']':'')
      +'  |  '+x.w+' - '+x.co;});
  var body=b, cut=0;
  for(var i=0;i<lines.length;i++){
    if(body.length+lines[i].length>1700){cut=lines.length-i;break;}
    body+=lines[i]+NL;
  }
  if(cut)body+='...plus '+cut+' more - see the printed report.'+NL;
  body+=RULE+NL+got.r.length+' items on hire.'+NL
   +'POWERED BY SITEIQ | Author: Andrew Fisher'+NL;
  return 'mailto:?subject='+encodeURIComponent('Coates K2 on hire - '+got.t)
   +'&body='+encodeURIComponent(body);
}
function prGo(flat){
  if(!PRCUR)return;
  var el=document.getElementById('prsheet');
  var ordTxt=(PRW.order==='az')?'products A to Z':'longest out first';
  if(PRCUR.flat||flat
     ||(PRW.layout==='flat'&&(PRCUR.kind==='co'||PRCUR.kind==='all'))){
    /* the one-big-table sheet reads like a dictionary: every item on
       one flat table, who and company on the line - in the order the
       wizard asked for */
    var flat=prOrd(PRCUR.r);
    var fbody='<table class="ptab">'
      +'<tr><th>Item</th><th>Item no</th><th>Scan</th><th>Who has it</th>'
      +'<th>Company</th><th>Aisle</th><th class="pn">Days on hire</th></tr>'
      +flat.map(function(x){
        return '<tr><td>'+esc(x.n)
          +(x.p?' <span class="ppid">Plant ID '+esc(x.p)+'</span>':'')
          +'</td><td>'+esc(x.i||'—')+'</td>'
          +'<td class="pqr">'+qr(x.i)+'</td>'
          +'<td>'+esc(x.w)+'</td><td>'+esc(x.co)+'</td><td>'+esc(x.u)+'</td>'
          +'<td class="pn'+((x.d||0)>4?' late':'')+'">'
          +(x.d==null?'—':x.d)+'</td></tr>';
      }).join('')+'</table>';
    prFrame(PRCUR.t+' — one table',
      PRCUR.r.length+' item'+(PRCUR.r.length===1?'':'s')+' &middot; '
        +(PRCUR.sub||ordTxt+', with who has it and their company on every line'),
      PRCUR.asof||ASOF, fbody, PRCUR.interim);
    return;
  }
  /* The grouped sheet: companies A-Z, hirers inside a company A-Z,
     items inside a name in the order the wizard asked for - longest
     out first, or products A to Z. The group key leads with the
     company so the sort and the page read the same way. */
  var byWho={};
  PRCUR.r.forEach(function(x){var k=x.co+'\\u001F'+x.w;
    (byWho[k]=byWho[k]||[]).push(x);});
  var body='';
  Object.keys(byWho).sort(function(a,b){
    return a.toUpperCase()<b.toUpperCase()?-1:1;
  }).forEach(function(k){
    var list=prOrd(byWho[k]);
    var p=k.split('\\u001F');
    body+='<div class="pwho">'+esc(p[1])+' <span>'+esc(p[0])+' &middot; '
      +list.length+' item'+(list.length===1?'':'s')
      +'</span></div><table class="ptab">'
      +'<tr><th>Item</th><th>Item no</th><th>Scan</th><th>Aisle</th>'
      +'<th class="pn">Days on hire</th></tr>'
      +list.map(function(x){
        return '<tr><td>'+esc(x.n)
          +(x.p?' <span class="ppid">Plant ID '+esc(x.p)+'</span>':'')
          +'</td><td>'+esc(x.i||'—')+'</td>'
          +'<td class="pqr">'+qr(x.i)+'</td><td>'+esc(x.u)+'</td>'
          +'<td class="pn'+((x.d||0)>4?' late':'')+'">'
          +(x.d==null?'—':x.d)+'</td></tr>';
      }).join('')+'</table>';
  });
  prFrame(PRCUR.t,
    PRCUR.r.length+' item'+(PRCUR.r.length===1?'':'s')+' &middot; '
      +(PRCUR.sub||'names A to Z, items '+ordTxt
        +' &middot; anything over 4 days is marked'),
    PRCUR.asof||ASOF, body, PRCUR.interim);
}
/* One frame for EVERY printed sheet - brand rule, title, as-at,
   footer - so a stock check, a hit list and a company report all
   come off the printer as the same family of document. */
function prFrame(title,sub,asof,body,interim){
  var el=document.getElementById('prsheet');
  el.innerHTML='<div class="phead"><div class="pbrand">COATES<b>.</b>'
   +'<span>POWERED BY SITEIQ</span></div>'
   +'<div class="pmeta">Cement Australia K2 Shutdown 2026 &middot; Gladstone'
   +'<br>As at '+esc(asof||ASOF)+'</div></div>'
   +'<div class="ptitle">'+esc(title)+'</div>'
   +'<div class="psub">'+sub+'</div>'
   +(interim?'<div class="pintr">INTERIM &mdash; phone-loaded fresh look. '
     +'The next morning build is the record.</div>':'')
   +body
   +'<div class="pfoot">Built from this morning&rsquo;s SiteIQ exports '
   +'&middot; read-only &middot; POWERED BY SITEIQ &middot; '
   +'Author: Andrew Fisher</div>';
  document.documentElement.className='pr';
  window.print();
}
window.addEventListener('afterprint',function(){
  document.documentElement.className='';
});

/* FRESH LOOK - load a raw SiteIQ export off THIS phone and read it on
   the spot. (Andrew, 29 Jul 2026: "what about just using a once off
   raw data for on the spot info print off if needed.")

   The deal, stated on the pane so nobody mistakes it: this view lives
   on this phone only, until the next proper build replaces it. It
   never touches the morning truth - the payload stays exactly what
   the laptop built. The raw export arrives with SiteIQ's shouting
   descriptions; the rename map rides in the payload, so the fresh
   view reads with the same clean names as everything else. SHIFT_RATE
   is in the raw file and is deliberately never rendered - money stays
   behind the manager code, whatever file gets loaded. */
var FR=null;
function paneFresh(){
  return '<div class="note"><b>A fresh look, straight off this phone.</b> '
   +'Download the On Hire report from SiteIQ onto this phone, pick it '
   +'below, and read it here &mdash; who holds what, as at the minute you '
   +'pulled it. Lives on this phone only; the next morning build is still '
   +'the record.</div>'
   +'<input type="file" id="frfile" accept=".xlsx" '
   +'style="display:none" onchange="frPick(this)">'
   +'<button class="stmore" type="button" '
   +'onclick="document.getElementById(\\'frfile\\').click()">'
   +'&#128194; Pick the ON_HIRE export</button>'
   +'<div id="frout"></div>';
}
function frPick(inp){
  var f=inp.files&&inp.files[0];
  if(!f) return;
  var out=document.getElementById('frout');
  out.innerHTML='<div class="kw" style="padding:10px 2px">Reading '
    +esc(f.name)+'&hellip;</div>';
  var rd=new FileReader();
  rd.onload=function(){ frParse(new Uint8Array(rd.result), f.name); };
  rd.onerror=function(){ out.innerHTML='<div class="kw" style="padding:10px 2px">'
    +'Could not read that file. Try picking it again.</div>'; };
  rd.readAsArrayBuffer(f);
}
function frTidy(d){
  d=String(d||'').replace(/\\s{2,}/g,' ').trim();
  if(d===d.toUpperCase() && d.length>6){
    d=d.toLowerCase().replace(/(^|[\\s(\\/-])([a-z])/g,
      function(_,a,b){return a+b.toUpperCase()});
  }
  return d;
}
function frParse(bytes,fname){
  var out=document.getElementById('frout');
  var sheet=null, ref=null;
  try{
    sheet=readXlsxSheet(bytes,'ON_HIRE');
    ref=readXlsxSheet(bytes,'REFERENCE_INFO');
  }catch(e){
    out.innerHTML='<div class="kw" style="padding:10px 2px">That file did '
      +'not open as an Excel export ('+esc(String(e))+'). Download it '
      +'fresh from SiteIQ and try again.</div>';
    return;
  }
  if(!sheet||!sheet.rows.length){
    out.innerHTML='<div class="kw" style="padding:10px 2px">No ON_HIRE '
      +'sheet in that file. This loader reads the On Hire report &mdash; '
      +'download that one from SiteIQ and pick it again.</div>';
    return;
  }
  var asof='';
  if(ref&&ref.rows.length){
    /* the sheet has REQUESTED_BY (a name) and REQUESTED_DATE/TIME -
       match on DATE or the pulled-at time comes out "Andrew Fisher"
       (caught 29 Jul 2026, first probe run) */
    var rk=Object.keys(ref.rows[0]);
    for(var i=0;i<rk.length;i++){
      var K=rk[i].toUpperCase();
      if(K.indexOf('REQUESTED')>=0&&K.indexOf('DATE')>=0){
        asof=ref.rows[0][rk[i]];break;}
    }
  }
  var now=new Date();
  function days(ds){
    var m=/^(\\d{1,2})\\/(\\d{1,2})\\/(\\d{4})/.exec(ds||'');
    if(!m) return null;
    var d=new Date(+m[3],+m[2]-1,+m[1]);
    return Math.max(0,Math.round((new Date(now.getFullYear(),now.getMonth(),
      now.getDate())-d)/86400000));
  }
  var rows=sheet.rows.map(function(r){
    var itm=r.ITEM_NUMBER||'';
    var nm=(D.ren&&D.ren[itm])||frTidy(r.ITEM_DESCRIPTION);
    return {n:nm, w:r.HIRER_NAME||'Not named', co:r.COMPANY||'Not named',
            d:days(r.START_DATE), i:itm, p:(D.pids&&D.pids[itm])||'',
            u:(r.PRODUCT_FAMILY||'').replace(/\\s{2,}/g,' ')};
  });
  rows.sort(function(a,b){
    var ka=[a.co.toUpperCase(),a.w.toUpperCase(),-(a.d==null?-1:a.d),a.n.toUpperCase()];
    var kb=[b.co.toUpperCase(),b.w.toUpperCase(),-(b.d==null?-1:b.d),b.n.toUpperCase()];
    for(var j=0;j<4;j++){ if(ka[j]<kb[j])return -1; if(ka[j]>kb[j])return 1; }
    return 0;
  });
  FR={rows:rows, asof:asof, fname:fname,
      loaded:now.getHours()+':'+('0'+now.getMinutes()).slice(-2)};
  frShow('all');
}
function frRows(kind){
  if(kind==='radios') return FR.rows.filter(function(x){
    return /radio/i.test(x.n)||/radio/i.test(x.u);});
  if(kind==='gas') return FR.rows.filter(function(x){
    return /gas monitor|bw flex/i.test(x.n);});
  if(kind==='co'){var v=document.getElementById('frco').value;
    return v?FR.rows.filter(function(x){return x.co===v;}):[];}
  if(kind==='pp'){var v2=document.getElementById('frpp').value;
    if(!v2) return [];
    var p=v2.split('\\u001F');
    return FR.rows.filter(function(x){return x.w===p[0]&&x.co===p[1];});}
  return FR.rows;
}
function frShow(kind){
  var out=document.getElementById('frout');
  if(!FR){out.innerHTML='';return;}
  var rows=frRows(kind);
  var cos={},pps={};
  FR.rows.forEach(function(x){
    cos[x.co]=(cos[x.co]||0)+1;
    var k=x.w+'\\u001F'+x.co; pps[k]=(pps[k]||0)+1;
  });
  var coOpts=Object.keys(cos).sort().map(function(c){
    return '<option value="'+esc(c)+'">'+esc(c)+' ('+cos[c]+')</option>';}).join('');
  var ppOpts=Object.keys(pps).sort().map(function(k){
    var p=k.split('\\u001F');
    return '<option value="'+esc(k)+'">'+esc(p[0])+' &middot; '+esc(p[1])
      +' ('+pps[k]+')</option>';}).join('');
  var title={all:'Everything on hire',radios:'Radios on hire',
             gas:'Gas monitors on hire'}[kind];
  if(kind==='co') title=(document.getElementById('frco')||{}).value;
  if(kind==='pp'){var pv=(document.getElementById('frpp')||{}).value||'';
    title=pv.split('\\u001F')[0];}
  PRCUR={t:'FRESH LOOK — '+(title||'on hire'),
         sub:'INTERIM — export pulled '+esc(FR.asof||('today '+FR.loaded))
           +', loaded on a phone at '+FR.loaded
           +'. The next morning build is the record.',
         asof:FR.asof||'', r:rows, interim:true};
  var h='<div class="note"><b>'+FR.rows.length+' items on hire</b> in '
    +esc(FR.fname)+(FR.asof?' &middot; pulled '+esc(FR.asof):'')
    +' &middot; loaded '+FR.loaded+'. This view lives on this phone only.</div>'
    +'<div class="prpick">'
    +'<button class="stmore" type="button" onclick="frShow(\\'all\\')">Everything</button>'
    +'<button class="stmore" type="button" onclick="frShow(\\'radios\\')">Radios</button>'
    +'<button class="stmore" type="button" onclick="frShow(\\'gas\\')">Gas monitors</button>'
    +'</div>'
    +'<div class="uhead">One company</div>'
    +'<select id="frco" class="srch" onchange="frShow(\\'co\\')">'
    +'<option value="">Pick a company&hellip;</option>'+coOpts+'</select>'
    +'<div class="uhead">One person</div>'
    +'<select id="frpp" class="srch" onchange="frShow(\\'pp\\')">'
    +'<option value="">Pick a person&hellip;</option>'+ppOpts+'</select>';
  if(rows.length){
    h+='<div class="uhead">'+esc(title||'')+' &mdash; '+rows.length+' item'
      +(rows.length===1?'':'s')+'</div>'
      +'<div class="prbtns">'
      +'<button class="stmore" type="button" onclick="prGo()">&#128424; Print / save PDF</button>'
      +'<a class="stmore" href="'+prMailto(PRCUR)+'">&#9993; Email via Outlook</a>'
      +'</div>'
      +rows.slice(0,200).map(function(x){
        return '<div class="kid"><div class="kt"><b>'+esc(x.n)+'</b>'
          +'<em class="'+((x.d||0)>4?'o':'')+'">'+(x.d==null?'&mdash;':x.d+'d')+'</em></div>'
          +'<div class="kw">'+wl(x)+' &middot; '+esc(x.co)
          +(x.i?' &middot; Item '+esc(x.i):'')
          +(x.p?' &middot; <b style="color:var(--org)">Plant ID '+esc(x.p)+'</b>':'')
          +'</div></div>';
      }).join('')+more(200,rows.length,'items');
  }else{
    h+='<div class="kw" style="padding:10px 2px">Nothing matches that pick.</div>';
  }
  out.innerHTML=h;
}

/* CONSUMABLE STOCK CHECK & REORDER SHEET - the clipboard walk.
   (Andrew, 29 Jul 2026: "i want a consumable print out like updated
   stock check. qty we have stock last stock check info then a reorder
   column so we can put down a number we need.")
   Every line, A-Z, with an empty ruled box to write the count and the
   reorder number in pen. Paper is the interface here on purpose - a
   clipboard walk beats a phone walk in a dusty aisle. */
function consPrint(){
  var c=D.cons; if(!c||!c.all) return;
  var body='<table class="ptab pchk">'
   +'<tr><th>Item</th><th>SKU</th><th>Scan</th><th class="pn">On shelf</th>'
   +'<th class="pn">Used so far</th><th>Last counted</th>'
   +'<th class="pn">Counted now</th><th class="pn">REORDER</th></tr>'
   +c.all.map(function(x){
     var last=(x.ct==null?'never':(x.ct+' on '
       +(x.co?x.co.split('-').reverse().join('/'):'?')
       +(x.by?' by '+esc(x.by):'')));
     return '<tr><td>'+esc(x.n)+'</td><td>'+esc(x.k)+'</td>'
       +'<td class="pqr">'+qr(x.k)+'</td>'
       +'<td class="pn">'+x.a+'</td><td class="pn">'+x.u+'</td>'
       +'<td>'+last+'</td>'
       +'<td class="pn"><span class="pbox"></span></td>'
       +'<td class="pn"><span class="pbox"></span></td></tr>';
   }).join('')+'</table>'
   +'<div class="pwho" style="margin-top:14px">Counted by '
   +'<span class="pline"></span> &nbsp; Date <span class="pline short"></span>'
   +' &nbsp; Sign <span class="pline short"></span></div>';
  prFrame('Consumables — stock check & reorder sheet',
    c.all.length+' lines A&ndash;Z &middot; write the count and the reorder '
     +'number in the boxes &middot; shelf figures as at the morning build',
    ASOF, body, false);
}

/* CONSUMABLES - the shelf, the count, and what has to be ordered.
   (Andrew, 29 Jul 2026: "consumables stock. how many available. how many
   used. have we stock checked was it right was it ok. percentage of stock
   take. do we need to order.")

   The order of the blocks is deliberate and is NOT the order of the
   questions. The false alarms come first, because SiteIQ flags ten lines
   as Stock Low and not one of them is an empty shelf - and a storeman who
   reads "stock low" at the counter will say "we are out of those" to the
   next bloke who asks. Kill that before anything else. */
/* ==================================================================
   EVERY MOVEMENT A CONSUMABLE HAS HAD (Andrew, 2 Aug 2026: "same with
   consumable we sjhould be able to click on it and come up with all
   transactions that product has had").

   Tap a line on the shelf and this opens its whole history off the
   sales feed - who took it, which company, how many, and when, newest
   first. Every name in it is a door to that person's profile, so the
   counter can go from "who took the last of the gloves" to "what else
   has that bloke got" in two taps.
================================================================== */
/* A LIGHT SHEET. The board had no overlay of its own - every pane was
   full-screen and swapped. A movement history does not deserve a pane
   (you are not navigating to it, you are glancing at it and coming
   straight back), so this is a sheet that slides up over whatever you
   were reading and leaves it exactly where it was.
   Escape closes it, the browser back button closes it, and tapping the
   dark area closes it. Three ways out, because one is never enough on
   a phone with gloves on. */
function shSheet(title,html){
  var el=document.getElementById('shsheet');
  if(!el){
    el=document.createElement('div'); el.id='shsheet'; el.className='shs';
    el.innerHTML='<div class="shbg" onclick="shClose()"></div>'
      +'<div class="shcard"><div class="shhd"><b id="shttl"></b>'
      +'<button type="button" onclick="shClose()">Close</button></div>'
      +'<div class="shbd" id="shbd"></div></div>';
    document.body.appendChild(el);
  }
  document.getElementById('shttl').textContent=title||'';
  var bd=document.getElementById('shbd');
  bd.innerHTML=html||''; bd.scrollTop=0;
  el.className='shs on';
  document.body.style.overflow='hidden';
  try{ history.pushState({sh:1},''); }catch(e){}
}
function shClose(){
  var el=document.getElementById('shsheet');
  if(el) el.className='shs';
  document.body.style.overflow='';
}
window.addEventListener('popstate',function(){ shClose(); });
document.addEventListener('keydown',function(e){
  var el=document.getElementById('shsheet');
  if(el&&el.className.indexOf('on')>=0&&(e.key==='Escape'||e.keyCode===27))
    shClose();
});
function consTx(i){
  var c=D.cons, x=c&&c.all&&c.all[i]; if(!x) return;
  var tx=x.tx||[];
  var took=0, back=0, who={};
  tx.forEach(function(m){ took+=(m.q||0); back+=(m.r||0);
    if(m.w) who[m.w]=(who[m.w]||0)+(m.q||0); });
  var wl2=[]; for(var w in who) wl2.push([w,who[w]]);
  wl2.sort(function(a,b){return b[1]-a[1]});
  var h='<div class="txh"><div class="txn1">'+esc(x.n)+'</div>'
   +'<div class="txn2">SKU '+esc(x.k)+(x.un?' &middot; '+esc(x.un):'')+'</div>'
   +'<div class="wnums" style="margin-top:12px">'
   +'<span><b>'+x.a+'</b>ON SHELF</span>'
   +'<span><b>'+Math.round(took)+'</b>TAKEN</span>'
   +'<span><b>'+Math.round(back)+'</b>BROUGHT BACK</span></div></div>';
  if(wl2.length){
    h+='<div class="uhead">Who has taken them</div><div class="wunits">'
      +wl2.slice(0,12).map(function(a){
         return '<span>'+whoLink(a[0],'')+' <b>'+Math.round(a[1])+'</b></span>';
       }).join('')+'</div>';
  }
  h+='<div class="uhead">Every movement &mdash; newest first ('+tx.length+')</div>';
  h+=tx.length?tx.map(function(m){
      return '<div class="txr"><div class="txd">'+esc(m.d||'&mdash;')+'</div>'
       +'<div class="txw">'+(m.w?whoLink(m.w,m.co||''):'<i>not named</i>')
       +(m.co?'<span>'+esc(m.co)+'</span>':'')+'</div>'
       +'<div class="txq'+((m.r||0)>0?' back':'')+'">'
       +((m.q||0)?'-'+Math.round(m.q):'')
       +((m.r||0)?' +'+Math.round(m.r):'')+'</div></div>';
    }).join('')
   :'<div class="kw">Nothing has moved on this line yet.</div>';
  shSheet('Movements', h);
}
function paneCons(){
  var c=D.cons;
  var h='<div class="note"><b>The consumables shelf.</b> What we hold, what '
   +'has gone out the window, and what has to be on order before '
   +c.end.split('-').reverse().join('/')+' &mdash; '+c.daysLeft+' days away.'
   +'<br><button class="stmore" style="margin-top:10px" type="button" '
   +'onclick="consPrint()">&#128424; Print the stock check &amp; reorder sheet</button>'
   +'</div>'
   +'<div class="tiles">'
   +tile(c.avail.toLocaleString(),'On the shelf','g')
   +tile(c.used.toLocaleString(),'Issued so far')
   +tile(c.skus,'Different lines')
   +tile(c.checkedPct+'%','Stocktake done','g')
   +tile(c.matchPct+'%','Counts that matched',c.matchPct>=95?'g':'a')
   +tile(c.order.length,'To order',c.order.length?'r':'g')
   +'</div>';

  /* THE SHELF, SHOWCASED - every line A-Z with its picture, SKU,
     live shelf count and a scan code, OPEN by default: this is the
     front window of the consumables area (Andrew, 30 Jul 2026: "get
     these consumable picture and their barcodes in the stores team
     section under consumables too please lets show case this area").
     The QR is the SKU number, so the counter scans the screen or the
     printed sheet the same way. Twins are already folded - one item,
     one line, the true total. */
  if(c.all&&c.all.length){
    h+='<div class="grp"><button type="button" onclick="tog(this)">'
      +'<div class="gn"><b>The shelf &mdash; every line A&ndash;Z</b>'
      +'<span>picture &middot; SKU &middot; on shelf &middot; scan code</span></div>'
      +'<div class="gq"><b>'+c.all.length+'</b><span>lines</span></div>'
      +'</button><div class="kids on">'
      +c.all.map(function(x,xi){
        var nt=(x.tx||[]).length;
        return '<div class="kid kidth hasqr'+(nt?' tapme':'')+'"'
          +(nt?' onclick="consTx('+xi+')"':'')+'>'
          +thTile(x.k,x.n)
          +'<div class="kbody"><div class="kt"><b>'+esc(x.n)+'</b>'
          +'<em'+(x.a?'':' class="o"')+'>'+x.a+' on shelf</em></div>'
          +'<div class="kw">SKU '+esc(x.k)+' &middot; '+x.u+' issued so far'
          +(nt?' &middot; <b class="txn">'+nt+' movement'+(nt===1?'':'s')
              +' &rsaquo;</b>':'')+'</div></div>'
          +'<div class="kqr">'+qr(x.k,56)+'</div></div>';
      }).join('')+'</div></div>';
  }
  if(c.folded){
    h+='<div class="note">'+c.folded+' duplicate SKU record'
      +(c.folded===1?'':'s')+' folded into their live line &mdash; SiteIQ '
      +'will not merge them, so this board does. One item, one line, the '
      +'true total, and no false stock-low flags.</div>';
  }
  if(c.records.length){
    h+='<div class="note"><b>Do not tell anyone we are out of these.</b> '
      +'The system has them at zero, the count found gear on the shelf. '
      +'Post the count in SiteIQ and the flag clears.</div>'
      +'<div class="grp"><button type="button" onclick="tog(this)">'
      +'<div class="gn"><b>Says low, but the count found stock</b>'
      +'<span>check the shelf before you say no</span></div>'
      +'<div class="gq"><b style="color:var(--am)">'+c.records.length+'</b>'
      +'<span>lines</span></div></button><div class="kids">'
      +c.records.map(function(x){
        return '<div class="kid"><div class="kt"><b>'+esc(x.n)+'</b>'
          +'<em>'+(x.ct||0)+' counted</em></div><div class="kw">'
          +'System says none, the count found stock &middot; '+esc(x.k)
          +'</div></div>';
      }).join('')+'</div></div>';
  }

  h+='<div class="uhead">Do we need to order?</div>';
  if(!c.order.length){
    h+='<div class="note">Nothing needs ordering. Every line that is moving '
      +'has enough on the shelf to reach the finish date at the rate it is '
      +'going.</div>';
  }else{
    h+='<div class="note">Worked out from what actually went out the window '
      +'&mdash; the rate each line is going since it first moved, against the '
      +c.daysLeft+' days left. The minimum and reorder fields in SiteIQ are '
      +'all set to zero on this job, so there is no trigger in there to '
      +'read.</div>'
      +'<div class="grp"><button type="button" onclick="tog(this)">'
      +'<div class="gn"><b>Order these</b><span>runs out before the finish '
      +'date</span></div><div class="gq"><b style="color:var(--rd)">'
      +c.order.length+'</b><span>lines</span></div></button>'
      +'<div class="kids">'
      +c.order.map(function(x){
        var need=Math.max(0,Math.round(x.b*c.daysLeft-x.a));
        return '<div class="kid kidth hasqr">'+thTile(x.k,x.n)
          +'<div class="kbody"><div class="kt"><b>'+esc(x.n)+'</b>'
          /* cover is floored to whole days, so anything under 24 hours
             arrives as 0. "0 days left" reads like a rounding error;
             "under a day" reads like the warning it is. */
          +'<em class="o">'+(x.c==null?'&mdash;':(x.c===0?'under a day'
             :x.c+(x.c===1?' day left':' days left')))
          +'</em></div><div class="kw">'+x.a+' on the shelf &middot; '
          +'going out '+x.b+'/day &middot; <b>order about '+need+' more</b>'
          +'</div></div>'
          +'<div class="kqr">'+qr(x.k,56)+'</div></div>';
      }).join('')+'</div></div>';
  }
  if(c.watch.length){
    h+='<div class="grp"><button type="button" onclick="tog(this)">'
      +'<div class="gn"><b>Worth watching</b><span>tight, not short</span>'
      +'</div><div class="gq"><b style="color:var(--am)">'+c.watch.length
      +'</b><span>lines</span></div></button><div class="kids">'
      +c.watch.map(function(x){
        return '<div class="kid"><div class="kt"><b>'+esc(x.n)+'</b>'
          +'<em>'+x.c+' days left</em></div><div class="kw">'+x.a
          +' on the shelf &middot; going out '+x.b+'/day</div></div>';
      }).join('')+'</div></div>';
  }

  h+='<div class="uhead">Was the count right?</div>'
    +'<div class="note"><b>'+c.checked+' of '+c.skus+' lines counted &mdash; '
    +c.checkedPct+'%.</b> Anything sold after a count is taken off before the '
    +'line is called wrong, so a count from last week is judged fairly. '
    +(c.off.length
       ? 'That leaves <b>'+c.off.length+'</b> real gaps.'
       : 'Every count lines up.')
    +'</div>';
  if(c.off.length){
    h+='<div class="grp"><button type="button" onclick="tog(this)">'
      +'<div class="gn"><b>Counts that did not line up</b>'
      +'<span>system against the shelf</span></div>'
      +'<div class="gq"><b style="color:var(--am)">'+c.off.length+'</b>'
      +'<span>lines</span></div></button><div class="kids">'
      +c.off.map(function(x){
        var v=x.v||0;
        return '<div class="kid"><div class="kt"><b>'+esc(x.n)+'</b>'
          +'<em class="'+(v<0?'o':'')+'">'+(v>0?'+':'')+Math.round(v)+'</em>'
          +'</div><div class="kw">system '+x.a+' &middot; counted '+(x.ct||0)
          +' &middot; '+(v>0?'more on the shelf than the system knows'
                            :'less on the shelf than the system says')
          +'</div></div>';
      }).join('')+'</div></div>';
  }
  return h;
}
function paneStd(){
  return '<div class="note"><b>SWMS-CTS-001 Rev 4</b> &middot; Tool Store '
   +'Operation, Cement Australia K2 2026 &middot; issued 15 Jul 2026. '
   +'This is the reminder at the counter. The SWMS itself is the document '
   +'you sign on to.</div>'
   +'<div class="std"><h3>The one that never bends</h3>'
   +'<p><b>Nothing issues or returns without scanning.</b> No exceptions, '
   +'no doing it later, no verbal hand-outs.</p></div>'
   +'<div class="std"><h3>The two SOPs</h3>'
   +'<div class="sop"><span class="n">SOP 1</span><b>Issue</b>'
   +'<span>Verify the hirer and their company. Inspect tools and leads; '
   +'confirm electrical tags, rigging inspection status, torque and '
   +'hydraulic compliance, and gas monitor bump and charge status. '
   +'<b style="color:#fff">Do not issue non-compliant gear.</b> '
   +'Scan every issue in SiteIQ.</span></div>'
   +'<div class="sop"><span class="n">SOP 2</span><b>Return &mdash; two stages, '
   +'two scans</b>'
   +'<span><b style="color:#fff">Stage 1:</b> receive, identify and '
   +'<b style="color:#fff">scan immediately</b>; place in the controlled '
   +'returns bay. Confirm condition with the returning person and record '
   +'faults.<br><b style="color:#fff">Stage 2:</b> inspect, clean, test, '
   +'charge and confirm compliance, then <b style="color:#fff">scan it '
   +'available</b>.<br>Never bypass either stage. Task gloves and P2 where '
   +'contamination or dust requires it.</span></div></div>'
   +'<div class="std"><h3>Why two scans</h3>'
   +'<p>The first scan says <b>it is back and it is ours again</b>. The '
   +'second says <b>it is fit to go out</b>. One scan cannot say both, and '
   +'the gap between them is where a faulty tool would otherwise walk '
   +'straight back onto a job.</p>'
   +'<p class="ref">Same words as the two A3 posters at the counter &mdash; '
   +'<b>Issue: Double Scan</b> and <b>Return: Double Scan</b>. If the poster '
   +'and this page ever disagree, the SWMS wins.</p></div>'
   +'<div class="std"><h3>Faults &mdash; every time</h3>'
   +'<p>Out of Service tag, photograph, written report, SiteIQ status and '
   +'physical quarantine. Record the asset, the fault, who reported it and '
   +'their company and contact, the location, date and time, the action and '
   +'who was notified.</p></div>'
   +'<div class="std"><h3>Handover</h3>'
   +'<p>Communicate open risks, faults, deliveries, equipment status, client '
   +'actions and owners to the incoming shift.</p></div>'
   +'<div class="std"><h3>The Coates Way at the counter</h3>'
   +'<div class="rule"><i></i><span><b>Care Deeply</b> &mdash; check in on '
   +'your workmates. A smile and "how are you?" matters. Concerns and '
   +'fatigue get raised, not carried.</span></div>'
   +'<div class="rule"><i></i><span><b>Stop Work Authority</b> &mdash; '
   +'anyone can stop the job. If it is not safe, stop, make the call, '
   +'ask.</span></div>'
   +'<div class="rule"><i></i><span><b>Best Service &amp; Value</b> &mdash; '
   +'gear that leaves this window bumped, charged and understood is a crew '
   +'that walks straight to the job.</span></div></div>';
}
function paneArr(){
  var h='<div class="note"><b>On its way, or stuck.</b> Not on a shelf and '
   +'not out with a crew &mdash; ordered, in transit, or held up in Baseplan. '
   +'Worth knowing before you tell someone we have not got one.</div>';
  var by={};
  D.arrivals.forEach(function(x){(by[x.s]=by[x.s]||[]).push(x)});
  Object.keys(by).sort().forEach(function(st){
    var list=by[st];
    h+='<div class="grp"><button type="button" onclick="tog(this)">'
      +'<div class="gn"><b>'+esc(st)+'</b><span>status in SiteIQ</span></div>'
      +'<div class="gq"><b style="color:var(--am)">'+list.length+'</b>'
      +'<span>items</span></div></button><div class="kids">'
      +list.map(function(x){
        return '<div class="kid"><div class="kt"><b>'+esc(x.n)+'</b></div>'
          +'<div class="kw">bound for '+esc(x.u)+'</div></div>';
      }).join('')+'</div></div>';
  });
  return h;
}
/* PLANT - shown only where there is plant, because not every store has
   it. (Andrew, 29 Jul 2026: "some sites will be different so maybe a
   toggle on and off for when we have plant onsite") */
function panePlant(){
  /* BY CATEGORY (Andrew, 30 Jul 2026: "with the plant gear lets
     categorise it better so if i want to find welder in there it
     shows me what is available and what in idle and who onhire and
     company"). One group per SiteIQ product family: free and idle
     first (the answer to "have we got one?"), then who holds the
     rest, named with company, days, Plant ID and a scan code. */
  var p=D.plant, t=D.tiles;
  var h='<div class="note"><b>Plant on this site, by category.</b> Open a '
   +'category: what is free right now, what sits idle on charge, and who '
   +'holds the rest. Turn it off if this store is tools only.'
   +'<div class="prbtns" style="margin-top:10px">'
   +'<button class="stmore" type="button" onclick="plantAudit()">&#128424; '
   +'Idle plant audit sheet ('+(t.plantIdle+t.plantFree)+')</button>'
   +'<button class="stmore" type="button" onclick="plantDemob()">&#128424; '
   +'Plant demob checklist ('+(t.plantOn+t.plantIdle+t.plantFree)+')</button>'
   +'<button class="stmore" type="button" onclick="togglePlant()">Hide plant '
   +'from this board</button></div></div>'
   +'<div class="tiles">'
   +tile(t.plantOn,'Out with crews')
   +tile(t.plantIdle,'Idle on charge','a')
   +tile(t.plantFree,'Free on the ground','g')
   +'</div>';
  var cats={};
  function slot(f){return cats[f]=cats[f]||{free:[],idle:[],out:[]};}
  (p.free||[]).forEach(function(x){if(!isBulk(x))slot(x.f||'Other Plant').free.push(x);});
  (p.idle||[]).forEach(function(x){if(!isBulk(x))slot(x.f||'Other Plant').idle.push(x);});
  (p.out||[]).forEach(function(x){slot(x.f||'Other Plant').out.push(x);});
  /* bulk yard gear reads as counts, one line per kind - nobody scrolls
     86 identical chutes (Andrew, 31 Jul 2026) */
  var bulkS=bulkAgg([['idle',p.idle||[]],['free',p.free||[]]]);
  if(bulkS.length){
    var bqS=bulkS.reduce(function(s,b){return s+b.q},0);
    h+='<div class="grp"><button type="button" onclick="tog(this)">'
      +'<div class="gn"><b>Bulk gear &mdash; chutes, hoppers &amp; barriers</b>'
      +'<span>counted stock, not serial-chased &middot; audit sheet covers it</span></div>'
      +'<div class="gq"><b>'+bqS+'</b><span>pieces</span></div>'
      +'</button><div class="kids">'
      +bulkS.map(function(b){
        return '<div class="kw kwq"><span>&bull; '+esc(b.n)+' &middot; '
          +esc(b.u)+'</span><b style="color:var(--am)">'+b.q+' on the ground</b></div>';
      }).join('')+'</div></div>';
  }
  function byDays(a,b){var da=(a.d==null?-1:a.d),db=(b.d==null?-1:b.d);
    if(da!==db)return db-da;
    return a.n.toUpperCase()<b.n.toUpperCase()?-1:1;}
  function byName(a,b){return a.n.toUpperCase()<b.n.toUpperCase()?-1:1;}
  function prow(x,em,emCls){
    return '<div class="kid kidth'+(x.i?' hasqr':'')+'">'+thTile(x.v,x.n,x.va)
      +'<div class="kbody">'
      +(x.i?'<div class="kqr">'+qr(x.i,56)+'</div>':'')
      +'<div class="kt"><b>'+esc(x.n)
      +(x.p?' <span style="color:var(--org);font-weight:800">&middot; Plant ID '
        +esc(x.p)+'</span>':'')
      +'</b><em'+(emCls?' class="'+emCls+'"':'')+'>'+em+'</em></div>'
      +'<div class="kw">'+(x.i?'Item '+esc(x.i)+' &middot; ':'')+esc(x.u)
      +(x.w?' &middot; <b>'+esc(x.w)+'</b>'+(x.co?' &middot; '+esc(x.co):''):'')
      +histLine(x.i)
      +'</div>'+compChips(x.fl)+'</div></div>';
  }
  Object.keys(cats).sort(function(a,b){
    var A=cats[a],B=cats[b];
    return (B.free.length+B.idle.length+B.out.length)
          -(A.free.length+A.idle.length+A.out.length);
  }).forEach(function(f){
    var c=cats[f], total=c.free.length+c.idle.length+c.out.length;
    h+='<div class="grp"><button type="button" onclick="tog(this)">'
      +'<div class="gn"><b>'+esc(f)+'</b><span>'+total+' machine'
      +(total===1?'':'s')+' on site</span></div>'
      +'<div class="gq"><b style="color:var(--gd)">'+c.free.length+'</b><span>free</span></div>'
      +'<div class="gq"><b style="color:var(--am)">'+c.idle.length+'</b><span>idle</span></div>'
      +'<div class="gq"><b style="color:var(--org)">'+c.out.length+'</b><span>out</span></div>'
      +'</button><div class="kids">';
    if(c.free.length){
      h+='<div class="uhead">Free on the ground &mdash; take one</div>'
        +c.free.slice().sort(byName).slice(0,40).map(function(x){
          return prow(x,'free','');}).join('')
        +more(40,c.free.length,'free');
    }
    if(c.idle.length){
      h+='<div class="uhead">Idle &mdash; on charge, nobody using it</div>'
        +c.idle.slice().sort(byDays).slice(0,40).map(function(x){
          return prow(x,(x.d!=null?x.d+'d idle':'idle'),'o');}).join('')
        +more(40,c.idle.length,'idle');
    }
    if(c.out.length){
      h+='<div class="uhead">Out with crews &mdash; who holds what</div>'
        +c.out.slice().sort(byDays).slice(0,60).map(function(x){
          return prow(x,(x.d!=null?x.d+'d out':'out'),'o');}).join('')
        +more(60,c.out.length,'machines');
    }
    h+='</div></div>';
  });
  return h;
}
function togglePlant(){
  SHOW_PLANT=!SHOW_PLANT;
  D.hasPlant=SHOW_PLANT;
  render();
}
/* THE IDLE PLANT AUDIT SHEET (Andrew, 31 Jul 2026: "someone can print
   and go check it off to ensure it matches up"). Every machine the
   register says is ON THE GROUND - idle on charge or free - one line
   each with a SIGHTED tick and a condition note. Plant ID printed
   where the register has one; a WRITE-IN box where it does not, so
   the walk that checks the plant also completes the ID register. */
/* BULK PLANT is counted, never ticked line by line (Andrew, 31 Jul
   2026: "the Rubbish chutes and the Hopers and the frames as well as
   both type of barriers can be just counts please"). One line per
   KIND with the expected number and a write-in COUNTED box - that is
   how a yard actually checks 86 chutes. */
function isBulk(x){
  var u=(x.u||'').toUpperCase(), n=(x.n||'').toUpperCase();
  return u.indexOf('CHUTE')>=0 || u.indexOf('BARRIER')>=0
      || n.indexOf('CHUTE')>=0 || n.indexOf('HOPPER')>=0
      || n.indexOf('BARRIER')>=0
      || (n.indexOf('FRAME')>=0 && (u.indexOf('CHUTE')>=0||n.indexOf('CHUTE')>=0));
}
function bulkAgg(lists){
  var agg={};
  lists.forEach(function(pair){
    pair[1].forEach(function(x){
      if(!isBulk(x)) return;
      var k=x.n+' | '+(x.u||'');
      var e=agg[k]=agg[k]||{n:x.n,u:x.u||'',q:0};
      e.q++;
    });
  });
  return Object.keys(agg).sort().map(function(k){return agg[k];});
}
function plantByCat(lists){
  var cats={};
  lists.forEach(function(pair){
    pair[1].forEach(function(x){
      var f=x.f||'Other Plant';
      (cats[f]=cats[f]||[]).push({x:x,tag:pair[0]});
    });
  });
  return cats;
}
function pidCell(x){
  return x.p?'<span class="ppid">'+esc(x.p)+'</span>'
            :'<span class="pbox" title="write the Plant ID on"></span>';
}
function byDaysName(a,b){
  var da=(a.x.d==null?-1:a.x.d), db=(b.x.d==null?-1:b.x.d);
  if(da!==db) return db-da;
  return a.x.n.toUpperCase()<b.x.n.toUpperCase()?-1:1;
}
function plantAudit(){
  var p=D.plant;
  var ground=[['idle',(p.idle||[]).filter(function(x){return !isBulk(x)})],
              ['free',(p.free||[]).filter(function(x){return !isBulk(x)})]];
  var bulk=bulkAgg([['idle',p.idle||[]],['free',p.free||[]]]);
  var cats=plantByCat(ground);
  var total=0, noid=0, body='';
  if(bulk.length){
    var bq=bulk.reduce(function(s,b){return s+b.q},0);
    body+='<div class="pintr">COUNT THESE &mdash; CHUTES, HOPPERS, FRAMES '
      +'&amp; BARRIERS &mdash; '+bq+' PIECES, COUNTED NOT TICKED</div>'
      +'<table class="ptab pchk">'
      +'<tr><th>Kind</th><th>Where</th><th class="pn">Register says</th>'
      +'<th class="pn">Counted</th><th class="pn">Matches</th><th>Note</th></tr>'
      +bulk.map(function(b){
        return '<tr><td>'+esc(b.n)+'</td><td>'+esc(b.u)+'</td>'
          +'<td class="pn"><b>'+b.q+'</b></td>'
          +'<td class="pn"><span class="pbox"></span></td>'
          +'<td class="pn"><span class="ptick"></span></td>'
          +'<td><span class="pline"></span></td></tr>';
      }).join('')+'</table>';
  }
  Object.keys(cats).sort(function(a,b){return cats[b].length-cats[a].length;})
   .forEach(function(f){
    var list=cats[f].slice().sort(byDaysName);
    total+=list.length;
    body+='<div class="pwho">'+esc(f)+' <span>'+list.length+' machine'
      +(list.length===1?'':'s')+' on the ground</span></div>'
      +'<table class="ptab pchk">'
      +'<tr><th>Plant ID</th><th>Machine</th><th>Scan</th><th>Where</th>'
      +'<th class="pn">Idle</th><th class="pn">SIGHTED</th><th>Condition / note</th></tr>'
      +list.map(function(e){var x=e.x;
        if(!x.p)noid++;
        return '<tr><td>'+pidCell(x)+'</td>'
          +'<td>'+esc(x.n)+'<br><span style="color:#8A94A2">Item '+esc(x.i||'&mdash;')+'</span></td>'
          +'<td class="pqr">'+qr(x.i)+'</td>'
          +'<td>'+esc(x.u||'')+(e.tag==='free'?' &middot; free':'')+'</td>'
          +'<td class="pn">'+(x.d!=null?x.d+'d':'&mdash;')+'</td>'
          +'<td class="pn"><span class="ptick"></span></td>'
          +'<td><span class="pline"></span></td></tr>';
      }).join('')+'</table>';
  });
  body+='<div class="pwho" style="margin-top:14px">Walked by '
   +'<span class="pline"></span> &nbsp; Date <span class="pline short"></span>'
   +' &nbsp; Sign <span class="pline short"></span></div>';
  var bq2=bulk.reduce(function(s,b){return s+b.q},0);
  prFrame('Plant — idle & free audit sheet',
    total+' machine'+(total===1?'':'s')+' ticked one by one'
     +(bq2?' &middot; '+bq2+' pieces of bulk gear on count lines':'')
     +' &middot; lay eyes on every line, note anything wrong'
     +(noid?' &middot; '+noid+' line'+(noid===1?'':'s')+' missing a Plant ID '
       +'&mdash; write it in the box and it goes on the register':''),
    ASOF, body, false);
}
/* THE PLANT DEMOB CHECKLIST - the whole plant fleet in demob order:
   first what is still out (who to chase), then the ground gear to
   off-hire and have collected. One paper trail from full site to
   clean site. */
function plantDemob(){
  var p=D.plant;
  var out=(p.out||[]).slice().sort(function(a,b){
    var ca=(a.co||'').toUpperCase(), cb=(b.co||'').toUpperCase();
    if(ca!==cb) return ca<cb?-1:1;
    var da=(a.d==null?-1:a.d), db=(b.d==null?-1:b.d);
    if(da!==db) return db-da;
    return a.n.toUpperCase()<b.n.toUpperCase()?-1:1;
  });
  var body='';
  if(out.length){
    body+='<div class="pintr">STILL OUT WITH CREWS &mdash; '+out.length
      +' MACHINE'+(out.length===1?'':'S')+' TO GET BACK FIRST</div>'
      +'<table class="ptab pchk">'
      +'<tr><th>Plant ID</th><th>Machine</th><th>Scan</th>'
      +'<th>Who has it</th><th class="pn">Days</th>'
      +'<th class="pn">BACK</th><th>Note</th></tr>'
      +out.map(function(x){
        return '<tr><td>'+pidCell(x)+'</td>'
          +'<td>'+esc(x.n)+'<br><span style="color:#8A94A2">Item '+esc(x.i||'&mdash;')+'</span></td>'
          +'<td class="pqr">'+qr(x.i)+'</td>'
          +'<td><b>'+esc(x.w||'?')+'</b><br><span style="color:#8A94A2">'+esc(x.co||'')+'</span></td>'
          +'<td class="pn">'+(x.d!=null?x.d:'&mdash;')+'</td>'
          +'<td class="pn"><span class="ptick"></span></td>'
          +'<td><span class="pline"></span></td></tr>';
      }).join('')+'</table>';
  }
  var bulkD=bulkAgg([['idle',p.idle||[]],['free',p.free||[]]]);
  var ground=0;
  if(bulkD.length){
    var bqD=bulkD.reduce(function(s,b){return s+b.q},0);
    ground+=bqD;
    body+='<div class="pwho">Bulk gear &mdash; chutes, hoppers, frames &amp; '
      +'barriers <span>'+bqD+' pieces &middot; count them back, off-hire by '
      +'the line</span></div>'
      +'<table class="ptab pchk">'
      +'<tr><th>Kind</th><th>Where</th><th class="pn">Register says</th>'
      +'<th class="pn">Counted back</th><th class="pn">OFF-HIRED</th><th>Note</th></tr>'
      +bulkD.map(function(b){
        return '<tr><td>'+esc(b.n)+'</td><td>'+esc(b.u)+'</td>'
          +'<td class="pn"><b>'+b.q+'</b></td>'
          +'<td class="pn"><span class="pbox"></span></td>'
          +'<td class="pn"><span class="ptick"></span></td>'
          +'<td><span class="pline"></span></td></tr>';
      }).join('')+'</table>';
  }
  var cats=plantByCat([['idle',(p.idle||[]).filter(function(x){return !isBulk(x)})],
                       ['free',(p.free||[]).filter(function(x){return !isBulk(x)})]]);
  Object.keys(cats).sort(function(a,b){return cats[b].length-cats[a].length;})
   .forEach(function(f){
    var list=cats[f].slice().sort(byDaysName);
    ground+=list.length;
    body+='<div class="pwho">'+esc(f)+' <span>'+list.length+' on the ground '
      +'&mdash; off-hire, then collection</span></div>'
      +'<table class="ptab pchk">'
      +'<tr><th>Plant ID</th><th>Machine</th><th>Scan</th><th>Where</th>'
      +'<th class="pn">OFF-HIRED</th><th class="pn">PICKED UP</th><th>Note</th></tr>'
      +list.map(function(e){var x=e.x;
        return '<tr><td>'+pidCell(x)+'</td>'
          +'<td>'+esc(x.n)+'<br><span style="color:#8A94A2">Item '+esc(x.i||'&mdash;')+'</span></td>'
          +'<td class="pqr">'+qr(x.i)+'</td>'
          +'<td>'+esc(x.u||'')+'</td>'
          +'<td class="pn"><span class="ptick"></span></td>'
          +'<td class="pn"><span class="ptick"></span></td>'
          +'<td><span class="pline"></span></td></tr>';
      }).join('')+'</table>';
  });
  body+='<div class="pwho" style="margin-top:14px">Demob led by '
   +'<span class="pline"></span> &nbsp; Date <span class="pline short"></span>'
   +' &nbsp; Sign <span class="pline short"></span></div>';
  var fin=(D.cons&&D.cons.end)
    ?' &middot; finish date '+D.cons.end.split('-').reverse().join('/')
      +(D.cons.daysLeft!=null?' &mdash; '+D.cons.daysLeft+' days away':'')
    :'';
  prFrame('Plant demob checklist',
    (out.length+ground)+' machines on the books: '+out.length+' still out, '
     +ground+' on the ground'+fin
     +' &middot; back &rarr; off-hired &rarr; picked up, ticked in that order',
    ASOF, body, false);
}
/* MONEY - the manager layer. A day rate against every line so a wrong
   one shows itself. */
function paneMgr(){
  var m=MGR;
  if(!m||!m.perDay) return '<div class="note">No rate data in the on-hire export.</div>';
  var h='<div class="ring"><div class="rv" style="font-size:30px">$'
   +m.perDay.toLocaleString(undefined,{minimumFractionDigits:2})+'</div>'
   +'<div class="rt"><b>on hire, per day</b>'
   +'$'+m.week.toLocaleString(undefined,{maximumFractionDigits:0})
   +' a week at today&rsquo;s rates. Every figure is SiteIQ&rsquo;s own '
   +'SHIFT_RATE &mdash; nothing here is estimated.</div></div>';
  if(m.zeroN){
    h+='<div class="note" style="border-left-color:var(--rd)"><b>'+m.zeroN
     +' items are on hire at a zero rate.</b> Earning nothing while they sit '
     +'with a crew. Worth a look &mdash; it is usually a rate that never got '
     +'set, not gear that is meant to be free.</div>'
     +'<div class="grp"><button type="button" onclick="tog(this)">'
     +'<div class="gn"><b>Zero-rate lines</b><span>check these first</span></div>'
     +'<div class="gq"><b style="color:var(--rd)">'+m.zeroN+'</b><span>items</span></div>'
     +'</button><div class="kids">'
     +m.zero.map(function(x){
       return '<div class="kid"><div class="kt"><b>'+esc(x.n)+'</b></div>'
         +'<div class="kw">'+esc(x.co)+(x.w?' &middot; '+esc(x.w):'')+'</div></div>';
     }).join('')+'</div></div>';
  }
  function money(v){return '$'+v.toLocaleString(undefined,{maximumFractionDigits:0})}
  function barlist(title,rows,sub){
    if(!rows.length) return '';
    var max=rows[0][1]||1;
    var o='<div class="uhead">'+title+'</div>';
    rows.slice(0,14).forEach(function(r){
      o+='<div class="brow"><div class="bd"><span>'+esc(r[0])+'</span>'
        +'<span>'+money(r[1])+sub+'</span></div>'
        +'<div class="bb"><i class="bd2" style="width:'+(100*r[1]/max)+'%"></i></div></div>';
    });
    return o+more(14,rows.length,'groups');
  }
  h+=barlist('By product group',m.cats,'/day');
  h+=barlist('By company',m.firms,'/day');
  h+='<div class="uhead">Dearest lines &mdash; per day</div>';
  m.top.slice(0,30).forEach(function(x){
    h+='<div class="kid"><div class="kt"><b>'+esc(x.n)+'</b>'
      +'<em>'+money(x.v)+'/day</em></div>'
      +'<div class="kw">'+x.q+' out &middot; '+money(x.r)+' each</div></div>';
  });
  h+=more(30,m.top.length,'lines');
  return h;
}
function paneIdle(){
  var h='<div class="note"><b>Held by the site plant pool.</b> On charge, on site, '
   +'not out with a crew &mdash; the cost-saving watch list.</div>';
  var byU={};
  D.idle.forEach(function(x){(byU[x.u]=byU[x.u]||[]).push(x)});
  Object.keys(byU).sort(function(a,b){return byU[b].length-byU[a].length})
   .forEach(function(u){
    var list=byU[u];
    h+='<div class="grp"><button type="button" onclick="tog(this)">'
      +'<div class="gn"><b>'+esc(u)+'</b><span>idle but on charge</span></div>'
      +'<div class="gq"><b>'+list.length+'</b><span>items</span></div>'
      +'</button><div class="kids">'
      +list.slice(0,60).map(function(x){
        return '<div class="kid kidth">'+thTile(x.v,x.n,x.va)
          +'<div class="kbody"><div class="kt"><b>'+esc(x.n)+'</b>'
          +(x.d==null?'':'<em class="o">'+x.d+' days</em>')+'</div>'
          +'<div class="kw">'+(x.i?'Item '+esc(x.i):'')
          +histLine(x.i)+'</div></div></div>';
      }).join('')+more(60,list.length,'machines')+'</div></div>';
  });
  return h;
}
</script></body></html>"""


def menc(code, s):
    r = _mulberry32(_xmur3(code + '|CoatesK2mgr2026')())
    return base64.b64encode(
        bytes((ord(c) ^ (r() >> 24)) & 0xFF for c in s)).decode('ascii')


def mtag(code):
    return format(_xmur3(code + '|CoatesK2mgrtag2026')(), 'x')


def build(data, code, asof, pricing=None, mgr_code=None):
    """The finished, gated page. Neither code appears in the file.

    Two payloads under two codes: the stores board, and the money. A
    manager code opens both - it is a superset, not a second door to
    remember - but the stores code cannot reach the money at all,
    because it is separately encrypted rather than merely hidden.
    """
    blob = json.dumps(data, separators=(',', ':'), ensure_ascii=True)
    _alias = keypad_alias(code)
    import mygear_font
    page = (PAGE.replace('__FONTCSS__', mygear_font.FONT_CSS)
                .replace('//__READER__//', _READER_JS)
                .replace('//__QRJS__//', _QR_JS)
                .replace('__PAYLOAD__', enc(code, blob))
                .replace('__TAG__', tag(code))
                .replace('__ATAG__', tag(_alias) if _alias else '')
                .replace('__AKEY__', enc(_alias, code) if _alias else '')
                .replace('__ASOF__', asof or 'this morning')
                .replace('__DAYTAG__', _day_tag()))
    if pricing and mgr_code:
        mblob = json.dumps(pricing, separators=(',', ':'), ensure_ascii=True)
        page = (page.replace('__MPAYLOAD__', menc(mgr_code, mblob))
                    .replace('__MTAG__', mtag(mgr_code))
                    #  the stores code, encrypted under the manager code -
                    #  never the code itself
                    .replace('__MKEY__', menc(mgr_code, code)))
    else:
        page = (page.replace('__MPAYLOAD__', '').replace('__MTAG__', '')
                    .replace('__MKEY__', ''))
    return page
