#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | THE OFFLINE DAY - keep trading when the internet dies
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | MY GEAR HQ
#
#  WHY (Andrew, 1 Aug 2026): "siteiq runs off the cloud how could we
#  move this offline if needed... i wanna go offline how can we"
#
#  THE HONEST SHAPE OF IT
#  ----------------------
#  Most of this suite is ALREADY offline. The pages are files on the
#  laptop, the store Wi-Fi serves them without touching the internet,
#  and every report builds from exports already sitting in
#  Data_SiteIQ. Nothing in the store needs a signal to keep working.
#
#  Exactly ONE thing needs the internet: pulling a fresh SiteIQ export
#  each morning (01/PULL) and submitting charges back. So "going
#  offline" is not a rebuild - it is a plan for the hours when the
#  uplink is down:
#
#    1. Know how old the data is, out loud, before trusting it.
#    2. Print the counter's paper so gear can still go out and come
#       back with a name against it.
#    3. Capture every movement on paper in the same shape SiteIQ
#       wants it, so catching up afterwards is typing, not detective
#       work.
#
#  That is what this button prints. It never invents a number: every
#  figure on the pack carries the age of the export it came from.
# =====================================================================
import datetime as dt
import html
import os
import sys

import report_paths as RP

HERE = os.path.dirname(os.path.abspath(__file__))
BLANK_ROWS = 34           # rows on each paper log page (fills A4)
BLANK_PAGES = 2           # how many of each log to print


def _sheet(path, name):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if name not in wb.sheetnames:
            return []
        rows = list(wb[name].iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []
    hdr = [(str(h).strip() if h is not None else '') for h in rows[0]]
    out = []
    for r in rows[1:]:
        d = {}
        for i, h in enumerate(hdr):
            if h:
                d[h] = r[i] if i < len(r) else None
        out.append(d)
    return out


def _t(d, k):
    v = d.get(k)
    return '' if v is None else str(v).strip()


def _date(v):
    if not v:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()[:10]
    for f in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return dt.datetime.strptime(s, f).date()
        except ValueError:
            pass
    return None


def _esc(s):
    return html.escape(str(s or ''))


# ---------------------------------------------------------------------
#  the paper
# ---------------------------------------------------------------------
CSS = """
*{box-sizing:border-box;margin:0;padding:0}
body{background:#5A5F66;font-family:Arial,Helvetica,sans-serif;color:#1D1D1B}
.sheet{width:794px;min-height:1123px;background:#fff;margin:14px auto;
 padding:0 0 26px;position:relative;page-break-after:always;
 border:3px solid #F26222;border-radius:4px;overflow:hidden}
.sheet:last-child{page-break-after:auto}
.band{background:#1D1D1B;color:#fff;padding:16px 26px;display:flex;
 justify-content:space-between;align-items:flex-start}
.k{color:#F26222;font-size:9.5pt;font-weight:bold;letter-spacing:2px}
.t{font-size:19pt;font-weight:bold;margin-top:4px}
.j{color:#B9BEC6;font-size:9pt;margin-top:4px}
.r{text-align:right;font-size:8.5pt;color:#B9BEC6}
.r b{color:#F26222;letter-spacing:1px;display:block;font-size:9.5pt}
.meta{padding:9px 26px;font-size:8.5pt;color:#555;border-bottom:1px solid #eee}
.meta b{color:#1D1D1B}
.body{padding:16px 26px 0}
h2{font-size:12pt;margin:14px 0 6px;color:#1D1D1B}
h2:first-child{margin-top:0}
.note{background:#FFF6F0;border-left:4px solid #F26222;padding:10px 12px;
 font-size:9.5pt;line-height:1.55;margin-bottom:12px}
.warn{background:#FDECEA;border-left:4px solid #C0392B}
.good{background:#EAF7EF;border-left:4px solid #2BB673}
table{width:100%;border-collapse:collapse;font-size:8.5pt}
th{background:#1D1D1B;color:#fff;text-align:left;padding:5px 6px;
 font-size:7.5pt;letter-spacing:1px}
td{padding:4px 6px;border-bottom:1px solid #e6e6e6;vertical-align:top}
tr:nth-child(even) td{background:#FAFAFA}
.co{background:#F26222;color:#fff;font-weight:bold;font-size:9pt;
 padding:5px 8px;margin:12px 0 0}
.pn{font-weight:bold;font-size:9pt;padding:6px 6px 2px;border-bottom:2px solid #1D1D1B}
.log td{height:23px;border-bottom:1px solid #c9c9c9}
.log td.w{border-left:1px solid #ececec}
.foot{position:absolute;bottom:8px;left:26px;right:26px;
 border-top:2px solid #F26222;padding-top:7px;display:flex;
 justify-content:space-between;font-size:7.5pt;color:#777}
.foot b{color:#F26222;letter-spacing:1px}
.big{font-size:26pt;font-weight:bold;color:#C0392B}
.stat{display:flex;gap:10px;margin:10px 0 14px}
.stat div{flex:1;border:1px solid #ddd;border-radius:6px;padding:9px 10px;
 text-align:center}
.stat b{display:block;font-size:17pt;color:#F26222}
.stat span{font-size:7.5pt;letter-spacing:1px;color:#666}
ol{margin:0 0 10px 18px;font-size:9.5pt;line-height:1.7}
ul{margin:0 0 10px 18px;font-size:9.5pt;line-height:1.6}
@media print{body{background:#fff}
 .sheet{margin:0;border:0;border-radius:0;box-shadow:none}}
@page{size:A4;margin:8mm}
"""


def sheet(kicker, title, asof, body, page, pages):
    return ("<div class='sheet'><div class='band'><div>"
            "<div class='k'>COATES &middot; " + kicker + "</div>"
            "<div class='t'>" + title + "</div>"
            "<div class='j'>Cement Australia K2 Shutdown 2026 &middot; "
            "Gladstone</div></div>"
            "<div class='r'><b>OFFLINE PACK</b>MY GEAR HQ<br>"
            "Andrew Fisher</div></div>"
            "<div class='meta'>Printed <b>"
            + dt.datetime.now().strftime('%d %b %Y - %H:%M')
            + "</b> &nbsp;|&nbsp; Data as at <b>" + asof + "</b>"
            "</div><div class='body'>" + body + "</div>"
            "<div class='foot'><span><b>MY GEAR HQ</b> &middot; "
            "the offline day</span><span>Page " + str(page) + " of "
            + str(pages) + "</span></div></div>")


def log_table(kind):
    """Blank rows in the shape SiteIQ wants back - so catching up is
    typing, not detective work."""
    if kind == 'out':
        cols = ("DATE", "TIME", "ID NUMBER", "NAME", "COMPANY",
                "ITEM NUMBER", "WHAT IT IS", "STAFF")
        widths = ("8%", "6%", "11%", "17%", "17%", "12%", "21%", "8%")
    else:
        cols = ("DATE", "TIME", "ITEM NUMBER", "WHAT IT IS", "BACK FROM",
                "CONDITION", "STAFF")
        widths = ("8%", "6%", "13%", "26%", "20%", "17%", "10%")
    h = "<table class='log'><tr>"
    for c, w in zip(cols, widths):
        h += "<th style='width:" + w + "'>" + c + "</th>"
    h += "</tr>"
    for _ in range(BLANK_ROWS):
        h += "<tr>" + "".join("<td class='w'></td>" for _ in cols) + "</tr>"
    return h + "</table>"


#  A4 sheet budget in pixels of body. The sheets are a FIXED height on
#  purpose - they are printed - so anything taller than the sheet would
#  be silently clipped. With 600+ items on hire that is most of the
#  pack, so the long sections are packed into pages here rather than
#  trusting the browser to break them. (Found 1 Aug 2026 - the sim only
#  had 8 items on hire and hid it.)
BODY_PX = 845
NOTE_PX = 78
CO_PX = 40
PERSON_PX = 26
ROW_PX = 19
THEAD_PX = 22


def _pack(blocks, first_note):
    """blocks: list of (cost, html, group_label_or_None). Returns a list
    of page bodies. A group whose header lands mid-page carries on with
    a (continued) header on the next one - nobody should have to guess
    whose gear they are looking at."""
    pages, cur, used, group = [], [], first_note and NOTE_PX or 0, None
    budget = BODY_PX

    def flush():
        if cur:
            pages.append("".join(cur))

    for cost, html_, grp in blocks:
        need = cost
        head = ""
        if grp is not None and grp != group:
            head = ""          # the block carries its own header
        if used + need > budget and cur:
            flush()
            cur, used = [], 0
            if grp is not None:
                cur.append("<div class='co'>" + _esc(grp)
                           + " &mdash; continued</div>")
                used += CO_PX
        cur.append(head + html_)
        used += need
        if grp is not None:
            group = grp
    flush()
    return pages or [""]


def build(base=None):
    base = base or HERE
    rental = RP.find_export(base, '*RENTAL_STOCK*.xlsx')
    if not rental:
        print("  No RENTAL_STOCK export found in Data_SiteIQ.")
        print("  The offline pack is built from the LAST export you")
        print("  pulled - without one there is nothing honest to print.")
        return 1

    age = dt.datetime.fromtimestamp(os.path.getmtime(rental))
    asof = age.strftime('%d %b %Y - %H:%M')
    hours = (dt.datetime.now() - age).total_seconds() / 3600.0
    stock = _sheet(rental, 'RENTAL_STOCK')
    if not stock:
        print("  That export has no RENTAL_STOCK sheet in it.")
        return 1

    today = dt.date.today()
    out, shelf = [], {}
    for r in stock:
        status = _t(r, 'ITEM_STATUS')
        desc = _t(r, 'ITEM_DESCRIPTION')
        unit = _t(r, 'STORAGE_UNIT') or 'Unassigned'
        if status == 'On Hire':
            d = _date(r.get('ON_HIRE_DATE'))
            out.append({
                'co': _t(r, 'COMPANY') or 'Not named',
                'who': _t(r, 'HIRER_NAME') or 'Not named',
                'no': _t(r, 'ITEM_NUMBER'),
                'd': desc, 'u': unit,
                'days': (today - d).days if d else None,
            })
        elif status.lower().startswith('available'):
            shelf.setdefault(unit, {}).setdefault(desc, 0)
            shelf[unit][desc] += 1

    #  who has what, the way the counter reads it: company, then person
    by_co = {}
    for x in out:
        by_co.setdefault(x['co'], {}).setdefault(x['who'], []).append(x)

    pages = []

    # ---- 1. the situation ------------------------------------------
    fresh = ("good" if hours <= 26 else ("" if hours <= 48 else "warn"))
    say = ("This morning's export &mdash; trust it." if hours <= 26 else
           ("Yesterday's export. Anything moved since is NOT on these "
            "pages." if hours <= 48 else
            "This data is more than two days old. Treat every line as a "
            "starting point, not the truth."))
    b = ("<div class='note " + fresh + "'><b>The data on this pack is "
         + ("%.0f" % hours) + " hours old.</b><br>" + say + "</div>")
    b += ("<div class='stat'><div><b>" + str(len(out))
          + "</b><span>ITEMS OUT</span></div><div><b>" + str(len(by_co))
          + "</b><span>COMPANIES</span></div><div><b>"
          + str(sum(sum(v.values()) for v in shelf.values()))
          + "</b><span>ON THE SHELF</span></div></div>")
    b += ("<h2>What still works with no internet</h2>"
          "<ul>"
          "<li><b>Everything in the store.</b> My Gear and the stores "
          "board are files on this laptop, served over the store's own "
          "Wi-Fi. Workers' phones do not touch the internet to read "
          "them.</li>"
          "<li><b>Every report button.</b> They build from the exports "
          "already in Data_SiteIQ &mdash; no signal needed.</li>"
          "<li><b>The label printer.</b> The Zebra prints over the "
          "store network.</li>"
          "<li><b>This pack.</b> Printed paper needs nothing at all.</li>"
          "</ul>"
          "<h2>What does not</h2>"
          "<ul>"
          "<li><b>Pulling a fresh SiteIQ export</b> (button 01) &mdash; "
          "that is the one job that needs the uplink.</li>"
          "<li><b>Submitting charges back to SiteIQ</b> &mdash; they "
          "queue and go when the line returns.</li>"
          "<li><b>Emailing the packs</b> &mdash; drafts still build, "
          "they just sit in Outlook until you have a signal.</li>"
          "</ul>"
          "<h2>So the plan is simple</h2>"
          "<ol>"
          "<li>Keep trading. Use the ISSUE and RETURN logs in this pack "
          "&mdash; every movement gets a name against it on paper.</li>"
          "<li>Tell the crews the phone page is a snapshot from <b>"
          + asof + "</b> and anything since is on paper at the "
          "window.</li>"
          "<li>When the line comes back: pull a fresh export, type the "
          "paper in, then run the day as normal. The catch-up page at "
          "the back walks it through.</li>"
          "</ol>"
          "__CONTENTS__")
    pages.append(("THE OFFLINE DAY", "When the uplink drops", b))

    # ---- 2. who has what -------------------------------------------
    note = ("<div class='note'>Every item on hire as at <b>" + asof
            + "</b>, by company then by person. This is the sheet that "
            "lets you hand gear back to the right name with no "
            "screen.</div>")
    blocks = []
    for co in sorted(by_co, key=lambda s: s.upper()):
        people = by_co[co]
        n = sum(len(v) for v in people.values())
        blocks.append((CO_PX,
                       "<div class='co'>" + _esc(co) + " &mdash; " + str(n)
                       + (" item" if n == 1 else " items") + "</div>", co))
        for who in sorted(people, key=lambda s: s.upper()):
            items = sorted(people[who], key=lambda x: -(x['days'] or 0))
            h = ("<div class='pn'>" + _esc(who) + " &middot; "
                 + str(len(items)) + "</div><table>")
            for it in items:
                h += ("<tr><td style='width:14%'>" + _esc(it['no'])
                      + "</td><td>" + _esc(it['d'])
                      + "</td><td style='width:20%'>" + _esc(it['u'])
                      + "</td><td style='width:10%'>"
                      + (str(it['days']) + "d" if it['days'] is not None
                         else "&mdash;") + "</td></tr>")
            h += "</table>"
            blocks.append((PERSON_PX + ROW_PX * len(items), h, co))
    for i, body in enumerate(_pack(blocks, True)):
        pages.append(("WHO HAS WHAT", "On hire right now",
                      (note if i == 0 else "") + body))

    # ---- 3. what is on the shelf -----------------------------------
    note = ("<div class='note'>What SiteIQ said was available at <b>"
            + asof + "</b>. Cross a line off as it goes out on paper "
            "&mdash; that is your live count until the export "
            "refreshes.</div>")
    THEAD = ("<tr><th>WHAT IT IS</th><th style='width:12%'>ON SHELF</th>"
             "<th style='width:16%'>GONE OUT</th></tr>")
    blocks = []
    for unit in sorted(shelf, key=lambda s: s.upper()):
        items = shelf[unit]
        blocks.append((CO_PX, "<div class='co'>" + _esc(unit) + " &mdash; "
                       + str(sum(items.values())) + " on the shelf</div>",
                       unit))
        names = sorted(items, key=lambda s: s.upper())
        #  a long aisle is broken into chunks so one shelf can span
        #  pages without the table being cut in half
        CH = 30
        for c in range(0, len(names), CH):
            part = names[c:c + CH]
            h = "<table>" + THEAD
            for d in part:
                h += ("<tr><td>" + _esc(d) + "</td><td><b>"
                      + str(items[d]) + "</b></td><td></td></tr>")
            h += "</table>"
            blocks.append((THEAD_PX + ROW_PX * len(part), h, unit))
    for i, body in enumerate(_pack(blocks, True)):
        pages.append(("ON THE SHELF", "What we had, and what has gone",
                      (note if i == 0 else "") + body))

    # ---- 4/5. the paper logs ---------------------------------------
    for i in range(BLANK_PAGES):
        b = ("<div class='note'>Fill a line for <b>every item that "
             "leaves</b>. ID number and name are what make it stick "
             "&mdash; a line with no name cannot be charged and cannot "
             "be chased.</div>" + log_table('out'))
        pages.append(("ISSUE LOG", "Gear going out &mdash; on paper", b))
    for i in range(BLANK_PAGES):
        b = ("<div class='note'>Fill a line for <b>every item that comes "
             "back</b>. Note the condition while the bloke is still "
             "standing there &mdash; damage found later is an argument, "
             "damage found now is a fact.</div>" + log_table('in'))
        pages.append(("RETURN LOG", "Gear coming back &mdash; on paper", b))

    # ---- 6. the catch-up -------------------------------------------
    b = ("<div class='note good'><b>The line is back. Here is the order "
         "to do it in.</b> Do not skip step 1 &mdash; typing paper into "
         "stale data is how two records disagree forever.</div>"
         "<ol>"
         "<li><b>Pull a fresh export</b> &mdash; button 01. Everything "
         "below works off the new file.</li>"
         "<li><b>Type the ISSUE log into SiteIQ</b>, oldest first. Use "
         "the date and time off the paper, not the time you are typing "
         "&mdash; the hire days have to be right or the invoice will "
         "not be.</li>"
         "<li><b>Type the RETURN log in</b>, same rule, oldest first.</li>"
         "<li><b>Log any damage found</b> on return while the paper is "
         "in front of you.</li>"
         "<li><b>Submit the consumable charges</b> &mdash; they queue "
         "while the line is down and go through once it is back.</li>"
         "<li><b>Pull ANOTHER fresh export</b> and run "
         "<b>00_RUN_EVERYTHING</b>. The day's reports, the phone page "
         "and the board all rebuild off the corrected data.</li>"
         "<li><b>Check button 40</b> on both laptops &mdash; same "
         "version, same numbers.</li>"
         "</ol>"
         "<h2>Keep the paper</h2>"
         "<div class='note'>Staple this pack shut when it is typed in "
         "and keep it with the day's reports. If a company ever queries "
         "a hire day from an offline stretch, the signed paper is the "
         "answer &mdash; and it is the only record of those hours that "
         "was written at the window.</div>")
    pages.append(("CATCH-UP", "When the line comes back", b))

    # ---- 7. emergency ----------------------------------------------
    try:
        import mygear_guides as MG
        em = MG.EMERGENCY
        groups = MG.CONTACT_GROUPS
    except Exception:
        em, groups = {"internal": "2222", "external": "07 4970 2222",
                      "channel": "1"}, []
    b = ("<div class='note warn'><div class='big'>EMERGENCY &mdash; "
         + _esc(em.get('internal', '2222')) + " internal &middot; "
         + _esc(em.get('external', '')) + "</div>"
         "UHF channel <b>" + _esc(em.get('channel', '1')) + "</b>. Say "
         "<b>&ldquo;Emergency, Emergency, Emergency&rdquo;</b>, then who "
         "you are, where you are, and what has happened.<br>"
         "<b>None of this needs the internet.</b> Radios and phones do "
         "not care that SiteIQ is down.</div>")
    #  the board runs well past one A4 - pack it like the rest
    blocks = []
    for name, colour, people in groups:
        blocks.append((CO_PX, "<div class='co' style='background:" + colour
                       + "'>" + name.replace('&amp;', '&') + "</div>",
                       name))
        h = "<table>"
        for p in people:
            h += ("<tr><td style='width:26%'><b>" + _esc(p[0])
                  + "</b></td><td>" + p[1] + "</td>"
                  "<td style='width:20%'><b>" + _esc(p[2]) + "</b></td></tr>")
        h += "</table>"
        blocks.append((22 * len(people), h, name))
    cpages = _pack(blocks, True)
    for i, body in enumerate(cpages):
        pages.append(("CONTACTS", "Every number, on paper",
                      (b if i == 0 else "") + body))

    # ---- the contents, now that we know how long each part ran -----
    marks, seen = [], set()
    for i, (k, _t2, _b) in enumerate(pages, start=1):
        if k not in seen:
            seen.add(k)
            marks.append((k, i))
    con = ("<h2>What is in this pack</h2><table>"
           "<tr><th>SECTION</th><th style='width:22%'>PAGES</th></tr>")
    for j, (k, start) in enumerate(marks):
        end = marks[j + 1][1] - 1 if j + 1 < len(marks) else len(pages)
        con += ("<tr><td>" + k.title() + "</td><td><b>"
                + (str(start) if start == end
                   else str(start) + " &ndash; " + str(end))
                + "</b></td></tr>")
    con += ("</table><div class='note'>Short of paper? <b>Print the "
            "situation page, both logs, the catch-up and the contacts"
            "</b> &mdash; that is enough to keep trading. The "
            "who-has-what and shelf pages are the bulk, and they are "
            "the ones worth having when a company argues about a hire "
            "day.</div>")
    pages[0] = (pages[0][0], pages[0][1],
                pages[0][2].replace("__CONTENTS__", con))

    # ---- write ------------------------------------------------------
    n = len(pages)
    doc = ("<!DOCTYPE html><html lang='en-AU'><head><meta charset='utf-8'>"
           "<title>Coates K2 - The Offline Day</title><style>" + CSS
           + "</style></head><body>")
    for i, (k, t, body) in enumerate(pages, start=1):
        doc += sheet(k, t, asof, body, i, n)
    doc += "</body></html>"

    dirs = RP.out_dirs(base)
    path = os.path.join(dirs['pages'],
                        'Coates_K2_Offline_Day_%s.html' % today.isoformat())
    with open(path, 'w', encoding='utf-8') as f:
        f.write(doc)

    print("  Data as at : %s  (%.0f hours old)" % (asof, hours))
    print("  On hire    : %d item(s) across %d compan%s"
          % (len(out), len(by_co), 'y' if len(by_co) == 1 else 'ies'))
    print("  On shelf   : %d item(s)"
          % sum(sum(v.values()) for v in shelf.values()))
    print("  Pages      : %d" % n)
    print("  Written    : %s" % os.path.relpath(path, base))
    print("")
    print("  Print it and put it in the drawer BEFORE you need it.")
    return 0


def main():
    print("=" * 66)
    print(" COATES | THE OFFLINE DAY - keep trading when the line drops")
    print("=" * 66)
    return build()


if __name__ == '__main__':
    try:
        sys.exit(main())
    except (KeyboardInterrupt, SystemExit):
        raise
    except BaseException as e:
        print("  Something went wrong: %r" % (e,))
        sys.exit(1)
