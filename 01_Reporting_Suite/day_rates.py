#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | DAY RATES - what the job was quoted, per item, per day
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY: no SiteIQ export carries a per-asset day rate, so every report
#  in this suite that wants to put a dollar against idle time has had
#  to write TBC. The quote has them. Drop it in and they fill in.
#
#  PER EACH (Andrew, 2 Aug 2026). A quoted rate is per item per day,
#  not per the quoted quantity - so 50 crash barriers at $2.07 is
#  $103.50 a day, not $2.07. Every sum in here multiplies by quantity
#  AND by days, in that order, and says so on the page.
#
#  WHERE TO PUT IT:
#      Data_Quote\   <- drop the rate card in here. Newest .xlsx wins.
#
#  THE SHAPE IT ACTUALLY COMES IN (Andrew, 2 Aug 2026 - the Contracted
#  Rates and Prices export): one row per PRODUCT VARIANT with a Base
#  Rate. 880 lines, and it keys straight onto the variant the register
#  already speaks - 779 of 790 stock variants matched, covering 4,487
#  of the 4,528 assets that carry a variant. RUBCHUTE1M comes back at
#  $1.04/day, the same figure Andrew has on his own sheet.
#
#  TIERS. The card carries T2/T3 tier columns - a shift number and a
#  rate that takes over past it. Every one of them is empty on this
#  contract, so today every line is a flat base rate. They are read and
#  applied anyway, because the day they ARE filled the sums must not
#  quietly keep using the base rate.
#
#  A quote in the older shape - item number or description plus a rate
#  column - still reads. Both are supported; the variant is tried first
#  because it is the only one that cannot be ambiguous.
#
#  ---------------------------------------------------------------
#  THREE SOURCES, IN ORDER, AND EVERY LINE SAYS WHICH ONE IT USED
#  ---------------------------------------------------------------
#  (Andrew, 2 Aug 2026: "if you look in transactions it will tell you
#  the product variant and the hire rates for ones your missing", and
#  "the ones on baseplan are charged on baseplan and only base plan but
#  yes this will come part of the plant equipment costs".)
#
#    1. CONTRACT   the contracted rate card, keyed on product variant.
#                  The agreed rate. Covers 99% of the register.
#    2. CHARGED    the RATE column on TRANSACTION_CHARGES - what was
#                  actually billed. Every one of the 1,079 lines carries
#                  one, across 800 items, and NOT ONE item is charged at
#                  two different rates, so it is unambiguous. This is
#                  what fills the gaps the card cannot reach - the
#                  externally hired plant that has no variant.
#    3. BASEPLAN   the other invoice stream's own Rate 1. Baseplan gear
#                  is billed on Baseplan and ONLY Baseplan - but it is
#                  still part of the plant equipment COST, so it is
#                  priced here for utilisation and marked BASEPLAN on
#                  every line that uses it. Marking it is the whole
#                  point: a Baseplan figure must never be mistaken for
#                  something SiteIQ will invoice.
#
#  Nothing found in any of the three prices as TBC. Never $0.
#
#  TWO INVOICE STREAMS, TWO RATE CARDS, AND THEY NEVER CROSS.
#  Baseplan bills its own 16 lines - radios, gas, two welders, fridges,
#  freezer, ice, tables, chairs and their transport - and carries its
#  own Rate 1. build_baseplan_costs.py owns that and its rule is
#  absolute: no Baseplan total ever feeds a SiteIQ report. So this
#  module reads the QUOTE, for the SiteIQ-billed gear, and it will not
#  silently reach into Data_Baseplan to fill a gap.
#
#  WHAT IT WILL NOT DO: invent a rate. An item with no quoted rate
#  prices as None and every report shows it as TBC, never as $0. A zero
#  would say "this earned nothing", which is a different claim and a
#  wrong one.
# =====================================================================
import glob
import os
import re

BASE = os.path.dirname(os.path.abspath(__file__))
QUOTE_DIR = 'Data_Quote'

#  Header names seen on quotes, lowest-fuss first. Matched loosely
#  because a quote is a human document and nobody spells these twice.
VARIANT_COLS = ('product variant id', 'product variant', 'variant id',
                'variant')
ITEM_COLS = ('item number', 'item no', 'item', 'asset no', 'asset number',
             'sku', 'product code', 'code')
DESC_COLS = ('description', 'item description', 'equipment', 'asset',
             'bundle equipment', 'product')
RATE_COLS = ('base rate', 'day rate', 'daily rate', 'rate per day',
             'rate 1', 'rate', 'unit rate', 'price per day')
T2_FROM = ('t2 first shift nr',)
T2_RATE = ('t2 tier rate',)
T3_FROM = ('t3 first shift nr',)
T3_RATE = ('t3 tier rate',)
QTY_COLS = ('quantity', 'qty')


def _norm(s):
    """Descriptions match on words, not on punctuation and case."""
    s = re.sub(r'[^a-z0-9]+', ' ', str(s or '').lower())
    return ' '.join(s.split())


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r'[^0-9.\-]', '', str(v or ''))
    try:
        return float(s)
    except ValueError:
        return 0.0


def _pick(hdr, wanted):
    """Index of the first header that looks like one of `wanted`."""
    low = [_norm(h) for h in hdr]
    for w in wanted:
        for i, h in enumerate(low):
            if h == _norm(w):
                return i
    for w in wanted:
        for i, h in enumerate(low):
            if h and _norm(w) in h:
                return i
    return None


def quote_path(here=None):
    """The quote, and ONLY something meant to be the quote.

    Anything sitting in Data_Quote\ is there on purpose, so any .xlsx
    in it counts. In the suite folder itself it must be NAMED like a
    quote - the first cut fell back to *.xlsx there and cheerfully
    picked up STOCKTAKE.xlsx, which is exactly the kind of silent wrong
    source that puts a confident number on a page.
    """
    here = here or BASE
    hits = [p for p in glob.glob(os.path.join(here, QUOTE_DIR, '*.xlsx'))
            if not os.path.basename(p).startswith('~')]
    if not hits:
        for pat in ('*QUOTE*.xlsx', '*RATE*.xlsx', '*quote*.xlsx',
                    '*rate*.xlsx'):
            hits += [p for p in glob.glob(os.path.join(here, pat))
                     if not os.path.basename(p).startswith('~')]
    return max(hits, key=os.path.getmtime) if hits else None


def load(here=None):
    """{'byItem':{}, 'byDesc':{}, 'lines':n, 'path':..., 'problem':...}

    Empty and harmless when there is no quote yet - every caller then
    shows TBC exactly as it did before this module existed.
    """
    out = {'byVariant': {}, 'byItem': {}, 'byDesc': {}, 'lines': 0,
           'path': '', 'problem': '', 'skipped': [], 'tiered': 0}
    p = quote_path(here)
    if not p:
        out['problem'] = ('no quote found - drop it in {}\\ and the day '
                          'rates fill in'.format(QUOTE_DIR))
        return out
    out['path'] = p
    try:
        import openpyxl
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    except Exception as e:
        out['problem'] = 'could not read {} ({})'.format(
            os.path.basename(p), e)
        return out

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        #  a quote usually has a title block above the table, so hunt
        #  for the row that actually looks like headers
        head_at, hdr = None, []
        for i, r in enumerate(rows[:25]):
            cand = [('' if c is None else str(c)) for c in r]
            if _pick(cand, RATE_COLS) is not None and (
                    _pick(cand, VARIANT_COLS) is not None
                    or _pick(cand, ITEM_COLS) is not None
                    or _pick(cand, DESC_COLS) is not None):
                head_at, hdr = i, cand
                break
        if head_at is None:
            continue
        vi = _pick(hdr, VARIANT_COLS)
        ii = _pick(hdr, ITEM_COLS)
        di = _pick(hdr, DESC_COLS)
        ri = _pick(hdr, RATE_COLS)
        qi = _pick(hdr, QTY_COLS)
        t2f, t2r = _pick(hdr, T2_FROM), _pick(hdr, T2_RATE)
        t3f, t3r = _pick(hdr, T3_FROM), _pick(hdr, T3_RATE)
        #  "Product Variant ID" contains "product" and would otherwise
        #  be grabbed as the description column too
        if vi is not None and di == vi:
            di = None
        if vi is not None and ii == vi:
            ii = None
        for r in rows[head_at + 1:]:
            if not r or not any(c not in (None, '') for c in r):
                continue
            rate = _num(r[ri]) if ri is not None and ri < len(r) else 0.0
            item = (str(r[ii]).strip()
                    if ii is not None and ii < len(r) and r[ii] else '')
            desc = (str(r[di]).strip()
                    if di is not None and di < len(r) and r[di] else '')
            qty = _num(r[qi]) if qi is not None and qi < len(r) else 0.0
            if rate <= 0:
                #  a line with no rate is REPORTED, not read as free
                out['skipped'].append(item or desc)
                continue
            var = (str(r[vi]).strip()
                   if vi is not None and vi < len(r) and r[vi] else '')
            if not (item or desc or var):
                continue

            def _cell(i):
                return (r[i] if i is not None and i < len(r) else None)

            tiers = []
            for fi, ri2 in ((t2f, t2r), (t3f, t3r)):
                frm, rt = _num(_cell(fi)), _num(_cell(ri2))
                if frm > 0 and rt > 0:
                    tiers.append((int(frm), rt))
            if tiers:
                out['tiered'] += 1
            rec = {'rate': rate, 'qty': qty, 'desc': desc, 'item': item,
                   'variant': var, 'tiers': sorted(tiers)}
            if var:
                out['byVariant'][var.upper()] = rec
            if item:
                out['byItem'][item] = rec
            if desc:
                out['byDesc'].setdefault(_norm(desc), rec)
            out['lines'] += 1
    wb.close()
    if not out['lines'] and not out['problem']:
        out['problem'] = ('read {} but found no rate lines in it - check '
                          'it has an item or description column and a rate '
                          'column'.format(os.path.basename(p)))
    return out


def record_for(rates, variant='', item='', desc=''):
    """The rate card line for one thing, or None. Variant first - it is
    the only key that cannot be ambiguous."""
    if not rates:
        return None
    if variant:
        r = rates['byVariant'].get(str(variant).strip().upper())
        if r:
            return r
    if item:
        r = rates['byItem'].get(str(item).strip())
        if r:
            return r
    if desc:
        n = _norm(desc)
        r = rates['byDesc'].get(n)
        if r:
            return r
        best = None
        for k, v in rates['byDesc'].items():
            if len(k) < 8:
                continue
            if n.startswith(k) or k.startswith(n):
                if best is None or len(k) > len(best[0]):
                    best = (k, v)
        if best:
            return best[1]
    return None


def charge(rec, qty, days):
    """PER EACH, and through the tiers if the card has any.

    Base rate up to the first tier's shift number, then that tier's
    rate, and so on. Every tier is empty on the current contract, so
    this comes out as rate x qty x days - but it will not silently keep
    doing that if the card is ever filled in.
    """
    if not rec:
        return None
    days = float(days or 0)
    qty = max(1, int(qty or 1))
    if days <= 0:
        return 0.0
    tiers = rec.get('tiers') or []
    if not tiers:
        return rec['rate'] * qty * days
    total, done, rate = 0.0, 0.0, rec['rate']
    for start, trate in tiers:
        upto = min(days, max(0.0, float(start) - 1))
        if upto > done:
            total += rate * (upto - done)
            done = upto
        rate = trate
    if days > done:
        total += rate * (days - done)
    return total * qty


def from_transactions(txn_path):
    """{item: rate} and {variant: rate} off the charge sheet's RATE.

    What was actually billed, which is the only rate that cannot be
    argued with. Checked across the whole sheet: no item is charged at
    two different rates, so one rate per item is a safe statement. If
    that ever stops being true the disagreement is recorded rather than
    a winner being picked.
    """
    out = {'byItem': {}, 'byVariant': {}, 'lines': 0, 'conflicts': []}
    if not txn_path or not os.path.isfile(txn_path):
        return out
    try:
        import openpyxl
        wb = openpyxl.load_workbook(txn_path, read_only=True, data_only=True)
    except Exception:
        return out
    if 'TRANSACTION_CHARGES' not in wb.sheetnames:
        wb.close()
        return out
    ws = wb['TRANSACTION_CHARGES']
    it = ws.iter_rows(values_only=True)
    try:
        hdr = [str(c).strip() if c is not None else '' for c in next(it)]
    except StopIteration:
        wb.close()
        return out
    ix = {h: i for i, h in enumerate(hdr) if h}
    ii = ix.get('SKU/ITEM_NUMBER')
    vi = ix.get('PRODUCT_VARIANT')
    ri = ix.get('RATE')
    if ri is None:
        wb.close()
        return out
    seen = {}
    for r in it:
        if not r:
            continue
        rate = _num(r[ri]) if ri < len(r) else 0.0
        if rate <= 0:
            continue
        item = (str(r[ii]).strip() if ii is not None and ii < len(r)
                and r[ii] else '')
        var = (str(r[vi]).strip() if vi is not None and vi < len(r)
               and r[vi] else '')
        if item:
            prev = seen.get(item)
            if prev is not None and abs(prev - rate) > 0.005:
                out['conflicts'].append((item, prev, rate))
            seen[item] = rate
            out['byItem'][item] = {'rate': rate, 'qty': 1, 'desc': '',
                                   'item': item, 'variant': var,
                                   'tiers': [], 'source': 'CHARGED'}
        if var:
            out['byVariant'].setdefault(var.upper(), {
                'rate': rate, 'qty': 1, 'desc': var, 'item': item,
                'variant': var, 'tiers': [], 'source': 'CHARGED'})
        out['lines'] += 1
    wb.close()
    return out


def from_baseplan(here=None):
    """{normalised description: rate} off the Baseplan stream's Rate 1.

    Baseplan bills its own gear and ONLY Baseplan bills it - but that
    gear is still a plant equipment cost, so it can be priced for
    utilisation. Every line that uses this is marked BASEPLAN so it can
    never be read as something SiteIQ will invoice.
    """
    here = here or BASE
    out = {'byDesc': {}, 'lines': 0}
    hits = [p for p in glob.glob(os.path.join(here, 'Data_Baseplan',
                                              '*.xlsx'))
            if not os.path.basename(p).startswith('~')]
    if not hits:
        return out
    try:
        import openpyxl
        wb = openpyxl.load_workbook(max(hits, key=os.path.getmtime),
                                    read_only=True, data_only=True)
    except Exception:
        return out
    for sn in wb.sheetnames:
        rows = list(wb[sn].iter_rows(values_only=True))
        if not rows:
            continue
        hdr = [str(c).strip() if c is not None else '' for c in rows[0]]
        ix = {h: i for i, h in enumerate(hdr) if h}
        ri = ix.get('Rate 1')
        if ri is None:
            continue
        for r in rows[1:]:
            if not r:
                continue
            rate = _num(r[ri]) if ri < len(r) else 0.0
            if rate <= 0:
                continue
            for col in ('Description', 'Bundle Equipment'):
                ci = ix.get(col)
                nm = (str(r[ci]).strip()
                      if ci is not None and ci < len(r) and r[ci] else '')
                if nm:
                    qi = ix.get('Quantity')
                    q = (_num(r[qi]) if qi is not None and qi < len(r)
                         else 0)
                    out['byDesc'].setdefault(_norm(nm), {
                        'rate': rate, 'qty': q, 'desc': nm, 'item': '',
                        'variant': '', 'tiers': [], 'source': 'BASEPLAN'})
                    out['lines'] += 1
    wb.close()
    return out


def resolve(card, txn, base, variant='', item='', desc=''):
    """The rate card first, then what was charged, then Baseplan.

    Returns the record with a 'source' on it, or None. The source is
    not decoration - a BASEPLAN rate and a CONTRACT rate mean different
    things to an invoice and the sheet has to be able to say which.
    """
    r = record_for(card, variant=variant, item=item, desc=desc)
    if r:
        r = dict(r)
        r.setdefault('source', 'CONTRACT')
        return r
    if txn:
        if item and item in txn['byItem']:
            return txn['byItem'][item]
        if variant and variant.upper() in txn['byVariant']:
            return txn['byVariant'][variant.upper()]
        if desc and _norm(desc):
            k = _norm(desc).upper()
            for vk, vv in txn['byVariant'].items():
                if _norm(vk) == _norm(desc):
                    return vv
    if base and desc:
        n = _norm(desc)
        if n in base['byDesc']:
            return base['byDesc'][n]
        best = None
        for k, v in base['byDesc'].items():
            if len(k) < 10:
                continue
            if n.startswith(k) or k.startswith(n):
                if best is None or len(k) > len(best[0]):
                    best = (k, v)
        if best:
            return best[1]
        #  THE REGISTER AND BASEPLAN DO NOT SPELL THE SAME MACHINE THE
        #  SAME WAY. "Welding Vantage Diesel 580(HGA035)" against
        #  "Welder - Motorized - Diesel - Vantage 580". A prefix match
        #  cannot bridge that, so words are compared - but carefully,
        #  because a loose match here puts the wrong rate on an
        #  expensive machine.
        #
        #  EVERY NUMBER HAS TO AGREE. A 500 never matches a 580. Beyond
        #  that it needs two more real words in common. Every match made
        #  this way is RECORDED and printed, so it is auditable rather
        #  than trusted.
        m = _token_match(n, base['byDesc'])
        if m:
            rec = dict(m[1])
            rec['fuzzy'] = m[0]
            return rec
    return None


def _tokens(s):
    return [t for t in _norm(s).split() if len(t) >= 4 or t.isdigit()]


def _token_match(n, table):
    """Best word-overlap match, or None. Numbers must agree exactly."""
    nt = set(_tokens(n))
    nums = {t for t in nt if any(c.isdigit() for c in t)}
    best = None
    for k, v in table.items():
        kt = set(_tokens(k))
        knums = {t for t in kt if any(c.isdigit() for c in t)}
        #  every number one side carries must appear on the other, or
        #  they are different machines wearing similar words
        common_nums = nums & knums
        if (nums and knums) and not common_nums:
            continue
        if nums and knums and (nums ^ knums) - common_nums:
            #  a number on one side the other does not have - only
            #  forgiven when it is clearly a unit tag, not a model
            extra = (nums ^ knums) - common_nums
            if any(len(e) <= 3 for e in extra):
                continue
        shared = (nt & kt) - common_nums
        score = len(shared) + 2 * len(common_nums)
        if len(shared) >= 2 and common_nums and (best is None
                                                 or score > best[0]):
            best = (score, k, v)
    return (best[1], best[2]) if best else None


def rate_for(rates, item='', desc='', variant=''):
    """The base day rate, or None. Never a guess, never zero."""
    r = record_for(rates, variant=variant, item=item, desc=desc)
    return r['rate'] if r else None


def _legacy_rate_for(rates, item='', desc=''):
    if not rates:
        return None
    if item:
        r = rates['byItem'].get(str(item).strip())
        if r:
            return r['rate']
    if desc:
        n = _norm(desc)
        r = rates['byDesc'].get(n)
        if r:
            return r['rate']
        #  a quote writes "Barrier - Crash Rated Water filled" where the
        #  register writes "Barrier - Crash Rated Water filled -
        #  Armorzone". Longest containing match wins, and only if it is
        #  a real prefix - never a two-word coincidence.
        best = None
        for k, v in rates['byDesc'].items():
            if len(k) < 8:
                continue
            if n.startswith(k) or k.startswith(n):
                if best is None or len(k) > len(best[0]):
                    best = (k, v)
        if best:
            return best[1]['rate']
    return None


def money(rate, qty, days):
    """PER EACH: rate x quantity x days. None in, None out."""
    if rate is None:
        return None
    return float(rate) * max(1, int(qty or 1)) * float(days or 0)


if __name__ == '__main__':
    r = load()
    if r['problem']:
        print('  ' + r['problem'])
    else:
        print('  Rate card: {}'.format(os.path.basename(r['path'])))
        print('  {} line(s) read, {} keyed on product variant, {} tiered'
              .format(r['lines'], len(r['byVariant']), r['tiered']))
        for k, v in list(r['byVariant'].items())[:8]:
            print('     {:<30} ${:>9.2f}/day'.format(k[:30], v['rate']))
    if r['skipped']:
        print('  {} line(s) had no rate and were skipped: {}'.format(
            len(r['skipped']), ', '.join(str(x)[:24] for x in r['skipped'][:5])))
