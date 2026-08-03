#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | BUILD DAILY EMAIL PACKS
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Turns today's built reports into the day's email packs: one folder
#  per company, dated, with the report that goes in the body, the PDFs
#  that go on the paperclip, a ready-addressed draft and the record of
#  what was sent to whom.
#
#      K2 DAILY REPORTING
#      |-- 00 MASTER DAILY REPORTS \ 2026-07-25 \
#      |-- 01 COMPANY REPORTS \ DGH Engineering \ 2026-07-25 \
#      |       DGH Engineering - On-Hire Report - 2026-07-25.html
#      |       Daily Safety & Compliance Report - 2026-07-25.pdf
#      |       Email Draft - 2026-07-25.eml
#      |       Email Record - 2026-07-25.txt
#      |-- 02 EMAIL CONTROL \    the day's status board
#      \-- 03 CONTACTS & SETTINGS \
#
#  THE WORKBOOK IS THE ROUTING SOURCE
#  ----------------------------------
#  Every address comes out of K2_Daily_Email_Report_Allocation.xlsx.
#  Not one is written into this script or any other. Change the
#  workbook, run again, and the packs follow it.
#
#  THE COMPANY NAME IS THE CONTROL KEY
#  -----------------------------------
#  A DGH report can never be sent on the Cleanaway recipient list. The
#  company name ties the data, the folder and the recipient list
#  together, every pack is checked before it is called ready, and one
#  of those checks reads the report itself looking for another
#  contractor's name.
#
#  NOTHING SENDS FROM HERE. Every pack is a draft. The drafts are
#  plain .eml files - double-click one and it opens in Outlook,
#  addressed, with the report in the body. You press Send.
#
#  Usage:
#      py BUILD_DAILY_EMAIL_PACKS.py
#      py BUILD_DAILY_EMAIL_PACKS.py --date 2026-07-24
#      py BUILD_DAILY_EMAIL_PACKS.py --dry-run
# =====================================================================

from __future__ import print_function

import os
import re
import sys
import glob
import json
import shutil
import argparse
import datetime as dt

import k2_daily_packs as K

HERE = os.path.dirname(os.path.abspath(__file__))
BRAND = "#F26222"

# ---------------------------------------------------------------------
#  Console that survives a Windows code page
# ---------------------------------------------------------------------
def say(msg=""):
    try:
        print(msg)
    except Exception:
        print(msg.encode("ascii", "replace").decode("ascii"))


def esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;"))


# ---------------------------------------------------------------------
#  Finding today's built reports
# ---------------------------------------------------------------------
def report_dirs(date_tag, base=None):
    day = os.path.join(base or os.path.join(HERE, "Reports"), date_tag)
    return {"pdf": os.path.join(day, "PDF"),
            "pages": os.path.join(day, "Pages"),
            "emails": os.path.join(day, "Emails"),
            "day": day}


def built(dirs, stem, date_tag, ext):
    """THE report for this stem and this day - by its exact name.

    This used to search for anything STARTING with the stem, which is
    how a report belonging to someone else could be picked up: the
    register spells Veolia Refractory just 'Veolia', so a pattern of
    'Coates_OnHire_Report_VEOLIA*' would happily match a
    'VEOLIA_WATER_TECHNOLOGIES' report and file it in Veolia
    Refractory's pack. One exact filename, or nothing. (25 Jul 2026)"""
    name = "{}_{}{}".format(stem, date_tag, ext)
    for d in dirs:
        path = os.path.join(d, name)
        if os.path.isfile(path):
            return path
    return None


def is_stale(pdf, html):
    """A PDF older than the report it was made from is yesterday's
    picture of today's numbers. Named, never filed quietly."""
    if not pdf or not html:
        return False
    try:
        return os.path.getmtime(pdf) < os.path.getmtime(html) - 2
    except OSError:
        return False


# ---------------------------------------------------------------------
#  The email itself
# ---------------------------------------------------------------------
def fingerprint(subject, images, attachments, to, cc, body=""):
    """One short string standing for everything that IS this email: who
    it goes to, what it says, and the CONTENT of every page picture and
    attachment - not just their names.

    Without this, a second run of the day would leave the first run's
    draft sitting there: the addresses, the subject and the file names
    are all identical, so nothing looked changed, while the pictures in
    the body were the earlier report. (25 Jul 2026)"""
    import hashlib
    h = hashlib.md5()
    h.update(subject.encode("utf-8", "replace"))
    h.update((body or "").encode("utf-8", "replace"))
    for e in sorted(x.lower() for x in list(to) + list(cc)):
        h.update(e.encode("utf-8", "replace"))
    for _cid, path in images:
        h.update(_digest(path).encode("ascii"))
    for path in attachments:
        h.update(os.path.basename(path).encode("utf-8", "replace"))
        h.update(_digest(path).encode("ascii"))
    return h.hexdigest()


def _digest(path):
    import hashlib
    h = hashlib.md5()
    try:
        with open(path, "rb") as fh:
            for chunk in iter(lambda: fh.read(1 << 16), b""):
                h.update(chunk)
    except Exception:
        return "missing"
    return h.hexdigest()


def build_eml(subject, body_html, images, attachments, to, cc,
              sender="Andrew Fisher"):
    """The report in the body AND the PDFs on the paperclip.

    Built as the standard nesting Outlook expects:

        multipart/mixed
          multipart/related
            multipart/alternative -> text + the HTML body
            image/png (inline, cid, no filename)  x N
          application/pdf (attachment, with filename) x N

    The inline page images carry no filename and are marked inline, so
    Outlook draws them in the body and leaves the paperclip for the
    PDFs that belong there.
    """
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.image import MIMEImage
    from email.mime.application import MIMEApplication

    root = MIMEMultipart("mixed")
    root["Subject"] = subject
    root["From"] = sender
    #  COMMAS, not semicolons. A semicolon is what you type into Outlook's
    #  To box; inside the file itself the standard separator is a comma,
    #  and a strict mail server reads a semicolon list as ONE malformed
    #  address - so only the first person gets it. Found and fixed
    #  25 Jul 2026; read_back() below proves every address survives.
    root["To"] = ", ".join(to)
    if cc:
        root["Cc"] = ", ".join(cc)
    root["X-Unsent"] = "1"          # opens in Outlook as an unsent draft
    #  What this draft is made of, so a later run can tell at a glance
    #  whether it is still the right email. Outlook ignores it.
    root["X-Coates-Pack"] = fingerprint(subject, images, attachments,
                                        to, cc, body_html)

    related = MIMEMultipart("related")
    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(
        "This report is best viewed in HTML - the report is in the "
        "message body.", "plain", "utf-8"))
    alt.attach(MIMEText(body_html, "html", "utf-8"))
    related.attach(alt)
    for cid, png in images:
        with open(png, "rb") as fh:
            img = MIMEImage(fh.read(), "png")
        img.add_header("Content-ID", "<{}>".format(cid))
        img.add_header("Content-Disposition", "inline")
        related.attach(img)
    root.attach(related)

    for path in attachments:
        if not os.path.isfile(path):
            continue
        with open(path, "rb") as fh:
            data = fh.read()
        sub = "pdf" if path.lower().endswith(".pdf") else "octet-stream"
        part = MIMEApplication(data, sub)
        part.add_header("Content-Disposition", "attachment",
                        filename=os.path.basename(path))
        root.attach(part)
    return root


def eml_bytes(msg):
    """CRLF line endings, as the standard requires and as Outlook's
    parser wants. The proven report emails are written this way; the
    packs were going out with bare newlines until 25 Jul 2026."""
    import io
    import email.policy
    from email.generator import BytesGenerator
    buf = io.BytesIO()
    BytesGenerator(buf, mangle_from_=False,
                   policy=email.policy.compat32.clone(linesep="\r\n")
                   ).flatten(msg)
    return buf.getvalue()


def read_back(eml_path):
    """Open the draft we just wrote the way a mail client will, and
    report what it actually finds. Writing the file is not proof it is
    right - reading it back is."""
    import email
    import email.policy
    with open(eml_path, "rb") as fh:
        msg = email.message_from_binary_file(fh, policy=email.policy.default)

    def addrs(field):
        out = []
        for a in (msg[field].addresses if msg[field] else ()):
            if a.addr_spec:
                out.append(a.addr_spec.lower())
        return out

    inline, files = [], []
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        cid = (part.get("Content-ID") or "").strip("<>")
        if part.get_content_disposition() == "inline" and cid:
            inline.append(cid)
        elif part.get_content_disposition() == "attachment":
            files.append(part.get_filename() or "")
    html = ""
    for part in msg.walk():
        if part.get_content_type() == "text/html":
            html = part.get_content()
            break
    return {"to": addrs("To"), "cc": addrs("Cc"),
            "stamp": msg["X-Coates-Pack"] or "",
            "subject": msg["Subject"] or "", "inline": inline,
            "attachments": files,
            "cids_used": sorted(set(re.findall(r"cid:([A-Za-z0-9_.-]+)", html)))}


def plain_body(greeting, lead, note_line, lines):
    """Used when there are no page pictures to paste - the report goes on
    the paperclip instead and the email says so plainly, with the same
    greeting and sign-off as every other pack."""
    import email_images
    items = "".join(
        "<li style='margin:0 0 6px'>{}</li>".format(esc(x)) for x in lines)
    inner = (
        "<table role='presentation' width='100%' cellpadding='0' "
        "cellspacing='0' style='width:100%;background:#0E1116'>"
        "<tr><td style='padding:18px 22px;color:#fff'>"
        "<div style='font:700 11px/1.2 Segoe UI,Arial;letter-spacing:.16em;"
        "color:{b}'>COATES | CEMENT AUSTRALIA K2 SHUTDOWN 2026</div>"
        "<div style='font:400 14px/1.6 Segoe UI,Arial;padding-top:10px;"
        "color:#D7DBE2'>{n}</div>"
        "<ul style='margin:10px 0 0;padding-left:20px;color:#D7DBE2;"
        "font:400 13px/1.7 Segoe UI,Arial'>{i}</ul></td></tr></table>"
    ).format(b=BRAND, n=esc(note_line), i=items)
    return email_images.letter(inner, greeting=greeting, lead=lead,
                               signoff=SIGNOFF, width=760)

def images_for(report_html, emails_dir):
    """The report's own page images, in order, as built by the report
    kit. Returns ([(cid, path), ...], [problems]) - an empty list if the
    kit didn't make them on this machine, which is not a fault; a
    problem means the pictures are there but don't add up to the whole
    report, and the pack must not go out on them."""
    if not report_html:
        return [], []
    stem = os.path.splitext(os.path.basename(report_html))[0]
    hits = glob.glob(os.path.join(emails_dir, stem + "_Email*.png"))

    def order(p):
        m = re.search(r"_Email(\d+)\.png$", p)
        return int(m.group(1)) if m else 0

    hits = [h for h in sorted(hits, key=order) if order(h) > 0]

    #  The pictures ARE the email. If a page is missing from the middle,
    #  the email still reads as a complete report - the pages are simply
    #  renumbered and the gap disappears. So the numbering is checked for
    #  gaps, and the count against what the report kit wrote down when it
    #  photographed the report. (25 Jul 2026)
    numbers = [order(h) for h in hits]
    gaps = numbers != list(range(1, len(numbers) + 1))
    want = 0
    side = os.path.join(emails_dir, stem + "_Email.pages")
    if os.path.isfile(side):
        try:
            with open(side, encoding="utf-8") as fh:
                want = int((fh.read().split() or ["0"])[0])
        except Exception:
            want = 0
    problems = []
    if gaps:
        problems.append("the page pictures are not a clean run 1..{} "
                        "({} found)".format(len(numbers),
                                            ", ".join(map(str, numbers[:8]))))
    elif want and len(hits) != want:
        #  TOO LONG TO PASTE IS NOT THE SAME AS BROKEN. The photographer
        #  stops at MAX_SLICES pages, so a 41-page report yields exactly
        #  40 pictures - a clean run 1..40, nothing damaged, just a
        #  report longer than an email body can hold. Counting that as a
        #  fault held Cement Australia's own daily pack: 6 attachments
        #  ready, addresses right, and the whole thing unsendable
        #  because the report grew by one page (3 Aug 2026).
        #
        #  No pictures, no error: the caller already has the path for
        #  that, and it attaches the complete PDF and says so in the
        #  body. Same report, all 41 pages, and it goes.
        try:
            import email_images as _EI
            _cap = _EI.MAX_SLICES
        except Exception:
            _cap = 40
        if len(hits) == _cap and want > _cap:
            return [], []
        problems.append("{} page picture(s) for a {}-page report".format(
            len(hits), want))
    if problems:
        return [], problems

    #  The report kit photographs every page of the report (it asks the
    #  document how many there are), so what is on disk IS the whole
    #  report. This used to re-photograph into the shared Reports folder
    #  when it looked short, which could delete the very pages the other
    #  emails were built from. It doesn't touch them now. (25 Jul 2026)
    return [("pg{}".format(i), h) for i, h in enumerate(hits, 1)], []


#  What the email says either side of the report. Andrew's words from
#  the mock-up, 25 Jul 2026 - short, plain, and it reads like a person
#  sent it rather than a system.
SIGNOFF = ("Regards,", "Your Coates Tool Store Team")
SENDER = "Coates Tool Store"


def images_stack(cids):
    """The stacked page pictures, exactly as the shared module builds
    them - one definition, one width (794, the pictures' own), and the
    orange border already baked into the pictures themselves."""
    import email_images
    return email_images.images_stack(cids)


# ---------------------------------------------------------------------
#  The day's control sheet
# ---------------------------------------------------------------------
def control_html(packs, date_tag, book, generated, fixed_cc):
    disp = dt.datetime.strptime(date_tag, "%Y-%m-%d").strftime("%d %b %Y")
    tone = {K.ST_DRAFT_READY: ("#12805C", "#E4F5EE"),
            K.ST_CHECKED: ("#12805C", "#E4F5EE"),
            K.ST_GENERATED: ("#8A6100", "#FFF3D6"),
            K.ST_NOT_GENERATED: ("#5A6472", "#EEF1F4"),
            K.ST_FAILED: ("#B3261E", "#FDECEA"),
            K.ST_SENT: ("#12805C", "#E4F5EE")}
    rows = []
    for p in packs:
        fg, bg = tone.get(p["status"], ("#5A6472", "#EEF1F4"))
        checks = p.get("checks") or []
        passed = sum(1 for _, ok, _ in checks if ok)
        nogear = p["status"] == K.ST_NOT_GENERATED
        if nogear:
            #  Nothing on hire is a fact, not a fault - it reads calmly.
            detail = "<span style='color:#9AA6B2'>{}</span>".format(
                esc(p["errors"][0]) if p["errors"] else
                "Nothing on hire today - folder and record made, no email.")
            score = "&ndash;"
        elif p["errors"]:
            detail = "<br>".join("<span style='color:#FF8A80'>&#10007; {}</span>"
                                 .format(esc(e)) for e in p["errors"])
            score = "{}/{}".format(passed, len(checks))
        else:
            detail = ("<span style='color:#4CD4A0'>All {} checks passed"
                      "</span>".format(len(checks)))
            score = "{}/{}".format(passed, len(checks))
        rows.append(
            "<tr><td><b>{co}</b><div class='sub'>{kind}</div></td>"
            "<td><span class='pill' style='color:{fg};background:{bg}'>{st}</span></td>"
            "<td class='num'>{pc}</td>"
            "<td>{body}</td><td class='num'>{na}</td>"
            "<td class='num'>{nt} / {nc}</td>"
            "<td class='num'>{mb}</td><td>{detail}</td></tr>".format(
                co=esc(p["company"]), kind=esc(p["kind"]), fg=fg, bg=bg,
                st=esc(p["status"]), pc=score,
                body=esc(os.path.basename(p["body_file"] or "") or "-"),
                na=len(p["attachments"]), nt=len(p["to"]), nc=len(p["cc"]),
                mb="{:.1f}".format(p.get("size_mb") or 0),
                detail=detail))
    counts = {}
    for p in packs:
        counts[p["status"]] = counts.get(p["status"], 0) + 1
    caption = {K.ST_DRAFT_READY: "checked and addressed - open and send",
               K.ST_FAILED: "held back - see the detail column",
               K.ST_NOT_GENERATED: "nothing on hire today - no email",
               K.ST_SENT: "gone", K.ST_CHECKED: "checked",
               K.ST_GENERATED: "built, not yet checked"}
    tiles = "".join(
        "<div class='tile'><div class='n'>{}</div><div class='k'>{}</div>"
        "<div class='c'>{}</div></div>".format(
            v, esc(k), esc(caption.get(k, "")))
        for k, v in sorted(counts.items()))
    return ("<!DOCTYPE html><html><head><meta charset='utf-8'>"
            "<meta name='viewport' content='width=device-width,initial-scale=1'>"
            "<title>K2 Daily Email Control - {disp}</title><style>"
            "body{{margin:0;background:#0E1116;color:#E9EDF2;"
            "font:400 14px/1.5 Segoe UI,Arial,sans-serif}}"
            ".wrap{{max-width:1180px;margin:0 auto;padding:26px 18px 60px}}"
            ".kick{{font:700 11px/1.2 Segoe UI;letter-spacing:.18em;color:{b}}}"
            "h1{{font:700 30px/1.25 Segoe UI;margin:8px 0 4px}}"
            ".meta{{color:#9AA6B2;font-size:13px;margin-bottom:20px}}"
            ".tiles{{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:22px}}"
            ".tile{{background:#171B21;border:1px solid #242A32;border-radius:12px;"
            "padding:14px 18px;min-width:130px}}"
            ".tile .n{{font:700 26px/1 Segoe UI;color:#fff}}"
            ".tile .k{{font-size:12px;color:#E9EDF2;padding-top:4px;font-weight:600}}"
            ".tile .c{{font-size:11px;color:#9AA6B2;padding-top:2px;max-width:190px}}"
            "table{{width:100%;border-collapse:collapse;background:#171B21;"
            "border:1px solid #242A32;border-radius:12px;overflow:hidden}}"
            "th{{text-align:left;font:700 11px/1.2 Segoe UI;letter-spacing:.1em;"
            "text-transform:uppercase;color:#9AA6B2;padding:12px 14px;"
            "background:#12161B;border-bottom:1px solid #242A32}}"
            "td{{padding:12px 14px;border-bottom:1px solid #1F242B;"
            "vertical-align:top;font-size:13px}}"
            "td.num{{text-align:right;white-space:nowrap}}"
            ".sub{{color:#9AA6B2;font-size:12px;padding-top:2px}}"
            ".pill{{display:inline-block;padding:3px 10px;border-radius:999px;"
            "font:700 11px/1.4 Segoe UI;white-space:nowrap}}"
            ".foot{{color:#9AA6B2;font-size:12px;margin-top:20px;"
            "border-top:1px solid #242A32;padding-top:14px}}"
            "</style></head><body><div class='wrap'>"
            "<div class='kick'>COATES | CEMENT AUSTRALIA K2 SHUTDOWN 2026 - GLADSTONE</div>"
            "<h1>Daily Email Control</h1>"
            "<div class='meta'>Reporting day {disp} &nbsp;|&nbsp; built {gen} "
            "&nbsp;|&nbsp; routing from {book}</div>"
            "<div class='tiles'>{tiles}</div>"
            "<table><tr><th>Company</th><th>Status</th><th>Checks</th>"
            "<th>In the email body</th><th>Attach</th><th>To / Cc</th>"
            "<th>MB</th><th>Detail</th></tr>{rows}</table>"
            "<div class='foot'>Nothing here has been sent. Each pack is a "
            "draft in its company folder - open the .eml, read it, press "
            "Send.<br>Fixed oversight Cc on every external pack: {fixed}<br>"
            "Andrew Fisher | Shutdown Manager | Coates | POWERED BY SITEIQ"
            "</div></div></body></html>").format(
                b=BRAND, disp=disp, gen=generated.strftime("%d %b %Y %H:%M"),
                book=esc(os.path.basename(book)), tiles=tiles,
                rows="".join(rows), fixed=esc("; ".join(fixed_cc) or "none"))


def control_txt(packs, date_tag, book, generated):
    L = []
    w = L.append
    w("COATES | K2 SHUTDOWN 2026 - DAILY EMAIL CONTROL")
    w("Reporting day {}   built {}".format(
        date_tag, generated.strftime("%d %b %Y %H:%M")))
    w("Routing from {}".format(os.path.basename(book)))
    w("=" * 74)
    for p in packs:
        w("{:<22} {:<26} To {:>2}  Cc {:>2}  Attach {:>1}".format(
            p["company"][:22], p["status"], len(p["to"]), len(p["cc"]),
            len(p["attachments"])))
        for e in p["errors"]:
            w("    ! " + e)
    w("=" * 74)
    w("Nothing has been sent. Open the .eml in each company folder and")
    w("press Send when you are happy with it.")
    return "\n".join(L) + "\n"


def keep_current(path, text, scrub=()):
    """A status board has to show today as it stands NOW.

    Records are never overwritten - that's rule 7 and it protects the
    'Sent by' line you write in by hand. But the control sheet is the
    page you open to decide what to send, and a board still showing
    yesterday's green while the console says HELD is worse than no
    board. So: same story, leave it; changed story, the earlier board is
    kept beside it and the file you open is the current one."""
    import re as _re

    def norm(s):
        for pat in scrub:
            s = _re.sub(pat, "", s)
        return s

    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as fh:
                if norm(fh.read()) == norm(text):
                    return path, "unchanged"
        except Exception:
            pass
        stem, ext = os.path.splitext(path)
        moved = False
        for n in range(1, 500):
            old = "{} (earlier {}){}".format(stem, n, ext)
            if not os.path.exists(old):
                try:
                    os.rename(path, old)
                    moved = True
                except OSError:
                    pass
                break
        if not moved:
            #  Couldn't put the earlier one safely aside, so the current
            #  one goes beside it instead. Nothing is ever written over.
            for n in range(1, 500):
                alt = "{} (current {}){}".format(stem, n, ext)
                if not os.path.exists(alt):
                    with open(alt, "w", encoding="utf-8") as fh:
                        fh.write(text)
                    return alt, "written beside the earlier one"
            return path, "could not be written - the earlier one is kept"
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(text)
    return path, "written"


def hold_or_release(pack, folder, date_tag):
    """A .eml in a company folder means ONE thing: checked, and ready to
    send. Anything else keeps everything it built - so you can see what
    went wrong - but not as a file you can double-click by mistake.
    "If any check fails, the email stays in Draft - Check Required":
    this is that rule, made physical."""
    live = os.path.join(folder, "Email Draft - {}.eml".format(date_tag))
    held = os.path.join(folder, "HOLD - Email Draft - {}.eml.txt".format(date_tag))
    ready = pack["status"] == K.ST_DRAFT_READY

    if ready:
        #  It came right. A HOLD file left over from an earlier run would
        #  sit there looking like today's problem - move it out of the way
        #  rather than delete it, so the trail survives.
        if os.path.isfile(held):
            for n in range(1, 60):
                keep = os.path.join(folder,
                                    "HOLD - Email Draft - {} (earlier {})"
                                    ".eml.txt".format(date_tag, n))
                if not os.path.exists(keep):
                    try:
                        os.rename(held, keep)
                    except OSError:
                        pass
                    break
        return ""

    if not os.path.isfile(live):
        return ""

    #  Keep any earlier held draft rather than writing over it - rule 7
    #  applies to these too.
    if os.path.isfile(held):
        for n in range(1, 60):
            keep = os.path.join(folder, "HOLD - Email Draft - {} (earlier {})"
                                        ".eml.txt".format(date_tag, n))
            if not os.path.exists(keep):
                try:
                    os.rename(held, keep)
                except OSError:
                    pass
                break
    try:
        os.rename(live, held)
        pack["draft"] = held
        return "held - the draft is not sendable until this is fixed"
    except OSError as exc:
        #  Couldn't rename it. A sendable draft for a failed pack is the
        #  one thing we must not leave behind, so it goes - the record
        #  beside it says exactly what was in it.
        try:
            os.remove(live)
            pack["errors"].append(
                "the draft could not be renamed ({}) so it was removed - "
                "the record beside this file has its details".format(exc))
        except OSError:
            pack["errors"].append(
                "WARNING: a sendable draft is still in this folder for a "
                "pack that FAILED its checks. Delete it by hand.")
        return "held"


def failed_pack(c, book, root, date_tag, fixed_cc, generated, why):
    """A pack that fell over. It still gets a folder, a record and a red
    line on the board - a company must never simply vanish from the
    day."""
    company = c["company"]
    folder = os.path.join(root, K.F_COMPANY, K._slug(company), date_tag)
    if not os.path.isdir(folder):
        os.makedirs(folder)
    pack = {"company": company, "kind": c.get("kind") or "External contractor",
            "data_company": None, "folder": folder, "date_tag": date_tag,
            "book": book["path"], "to": list(c["to"]), "cc": list(c["cc"]),
            "fixed_cc": fixed_cc, "body_file": None, "attachments": [],
            "errors": ["this pack could not be built: {}".format(why),
                       "nothing has been sent - fix it and run again"],
            "status": K.ST_FAILED, "size_mb": 0.0, "stale": [], "checks": [],
            "readback": [], "mapped": True, "reissued": [],
            "report_attached": "", "superseded": "",
            "subject": "{} - {} - NOT BUILT".format(company, date_tag)}
    try:
        hold_or_release(pack, folder, date_tag)
    except Exception:
        pass
    try:
        K.write_record(pack, date_tag, generated)
    except Exception:
        pass
    return pack


def write_register(root, packs, date_tag, generated, book):
    """One row per company per day, appended - the trail across the whole
    shutdown, never overwritten."""
    path = os.path.join(root, K.F_CONTROL, "Email Control Register.xlsx")
    try:
        import openpyxl
    except Exception:
        return None
    head = ["Date", "Company", "Package", "Status", "Checks passed",
            "Checks total", "To", "Cc", "Body file", "Attachments",
            "Size MB", "Built", "Issues"]
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path)
        ws = wb["Register"] if "Register" in wb.sheetnames else wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Register"
        ws.append(head)
        for c in ws[1]:
            c.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            c.fill = openpyxl.styles.PatternFill("solid", fgColor="0E1116")
        ws.freeze_panes = "A2"
        for col, wid in zip("ABCDEFGHIJKLM",
                            (12, 26, 20, 24, 13, 13, 40, 46, 44, 40, 9, 18, 60)):
            ws.column_dimensions[col].width = wid
    #  Rule 7: a day already in the register is left exactly as it is.
    #  A ledger, not a snapshot: the same company on the same day appears
    #  again if its status changed, so the trail shows a pack that was
    #  held and then fixed. An unchanged status writes nothing.
    seen = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[0] and r[1]:
            seen.add((str(r[0]).strip()[:10], str(r[1]).strip(),
                      str(r[3] or "").strip()))
    added = 0
    for p in packs:
        if (date_tag, p["company"], p["status"]) in seen:
            continue
        checks = p.get("checks") or []
        ws.append([date_tag, p["company"], p["kind"], p["status"],
                   sum(1 for _, ok, _ in checks if ok), len(checks),
                   "; ".join(p["to"]), "; ".join(p["cc"]),
                   os.path.basename(p["body_file"] or ""),
                   "; ".join(os.path.basename(a) for a in p["attachments"]),
                   round(p.get("size_mb") or 0, 2),
                   generated.strftime("%Y-%m-%d %H:%M"),
                   " | ".join(p["errors"])])
        added += 1
    #  Nothing new to say? Then don't touch the file at all - re-saving a
    #  workbook rewrites every byte in it, and a register that changes on
    #  a run that added nothing is a register you can't trust.
    if added or not os.path.isfile(path):
        wb.save(path)
    return path, added


# ---------------------------------------------------------------------
#  The run
# ---------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(
        description="Build the day's company email packs from the "
                    "reports already generated.")
    ap.add_argument("--date", help="reporting day, YYYY-MM-DD "
                                   "(default: today)")
    ap.add_argument("--root", help="where K2 DAILY REPORTING lives "
                                   "(default: beside this script)")
    ap.add_argument("--dry-run", action="store_true", dest="dry",
                    help="show what would be built, write nothing")
    ap.add_argument("--book", help="a specific routing workbook (default: "
                                   "the one beside this script)")
    ap.add_argument("--reports", help="where the built reports live "
                                      "(default: Reports beside this script)")
    a = ap.parse_args()

    generated = dt.datetime.now()
    date_tag = a.date or generated.strftime("%Y-%m-%d")
    try:
        disp = dt.datetime.strptime(date_tag, "%Y-%m-%d").strftime("%d %b %Y")
    except ValueError:
        say("That date doesn't look like YYYY-MM-DD: {}".format(date_tag))
        return 2
    root = a.root or os.path.join(HERE, K.ROOT_NAME)

    say("")
    say("=" * 68)
    say(" COATES | K2 DAILY EMAIL PACKS - {}".format(disp))
    say("=" * 68)

    book = K.load_allocation(HERE, path=a.book)
    if not book:
        say("")
        say(" No routing workbook found.")
        say(" Put K2_Daily_Email_Report_Allocation.xlsx beside this script")
        say(" and run again. Nothing is addressed without it.")
        return 1
    if not book["companies"]:
        say("")
        say(" The routing workbook opened, but no companies could be read")
        say(" from its Company Routing sheet.")
        say("")
        say(" The sheet needs a header row containing 'Company', with the")
        say(" companies listed underneath it. Nothing has been built -")
        say(" fix the sheet and run again.")
        return 1
    fixed_cc = book["fixed_cc"]
    say(" Routing     : {}".format(os.path.basename(book["path"])))
    say(" Companies   : {}".format(len(book["companies"])))
    say(" Fixed Cc    : {}".format(len(fixed_cc)))
    if not fixed_cc:
        say("   ! The four fixed oversight contacts were not found on the")
        say("     Summary sheet - every external pack will be held.")

    dirs = report_dirs(date_tag, a.reports)
    if not os.path.isdir(dirs["day"]):
        say("")
        say(" No reports built for {} yet.".format(disp))
        say(" Run 00_RUN_EVERYTHING.bat first, then run this again.")
        return 1
    say(" Reports from: {}".format(os.path.join("Reports", date_tag)))

    #  ---- the master reports -----------------------------------------
    say("")
    say(" MASTER REPORTS")
    masters = {}
    stale_all = []
    for label, stem in K.MASTERS:
        html = built([dirs["pages"]], stem, date_tag, ".html")
        pdf = built([dirs["pdf"]], stem, date_tag, ".pdf")
        if is_stale(pdf, html):
            stale_all.append(label)
        masters[label] = {"html": html, "pdf": pdf, "stem": stem}
        mark = "PDF + HTML" if (pdf and html) else (
            "HTML only - no PDF" if html else "NOT BUILT")
        if is_stale(pdf, html):
            mark += "  (PDF older than the report - rebuild)"
        say("   {:<34} {}".format(label[:34], mark))

    if a.dry:
        say("")
        say(" Dry run - nothing written.")
        return 0

    #  ---- the folders -------------------------------------------------
    made = K.ensure_tree(root, [c["company"] for c in book["companies"]],
                         date_tag)
    say("")
    say(" FOLDERS")
    say("   {}".format(root))
    say("   {} new folder(s) created".format(len(made)))

    #  file the masters
    mdir = os.path.join(root, K.F_MASTER, date_tag)
    filed_master = {}
    for label, info in masters.items():
        name = K.FILED_NAME.get(label, label + " - {d}").format(d=date_tag)
        for src, ext in ((info["pdf"], ".pdf"), (info["html"], ".html")):
            if not src:
                continue
            dest, note, reissued = K._place(src, mdir, name + ext)
            filed_master.setdefault(label, {})[ext] = dest
            if reissued:
                say("   {} was re-issued today - filed as {}".format(
                    label, os.path.basename(dest)))

    #  the routing workbook, kept with the day it routed
    sdir = os.path.join(root, K.F_SETTINGS)
    K._place(book["path"], sdir, os.path.basename(book["path"]))
    snap = os.path.join(sdir, "Routing in use - {}.txt".format(date_tag))
    if not os.path.exists(snap):
        lines = ["COATES | K2 - ROUTING IN USE ON {}".format(date_tag),
                 "Taken from {}".format(os.path.basename(book["path"])), "=" * 70,
                 "FIXED OVERSIGHT Cc ON EVERY EXTERNAL PACK",
                 "  " + ("; ".join(fixed_cc) or "(none found)"), ""]
        for c in book["companies"]:
            lines.append("{}  [{}]".format(c["company"], c["kind"]))
            lines.append("  To: " + ("; ".join(c["to"]) or "(nobody)"))
            lines.append("  Cc: " + ("; ".join(c["cc"]) or "(nobody)"))
            lines.append("")
        lines.append("Change the workbook, not this file - this is only the")
        lines.append("record of what the packs used on the day.")
        with open(snap, "w", encoding="utf-8") as fh:
            fh.write("\n".join(lines) + "\n")

    #  ---- a pack per company ------------------------------------------
    say("")
    say(" COMPANY PACKS")
    packs = []
    for c in book["companies"]:
        try:
            packs.append(build_pack(c, book, root, date_tag, disp, dirs,
                                    filed_master, masters, fixed_cc,
                                    generated))
        except Exception as exc:
            #  Whatever went wrong with one company, the other eleven
            #  still get their pack, and this one is held and named.
            import traceback
            say("   {:<24} HELD - {}".format(c["company"][:24], exc))
            packs.append(failed_pack(c, book, root, date_tag, fixed_cc,
                                     generated,
                                     traceback.format_exc().strip()
                                     .splitlines()[-1]))

    #  ---- the control sheet -------------------------------------------
    cdir = os.path.join(root, K.F_CONTROL)
    ch = os.path.join(cdir, "Daily Email Control - {}.html".format(date_tag))
    ct = os.path.join(cdir, "Daily Email Control - {}.txt".format(date_tag))
    #  The build time is the only thing that legitimately differs between
    #  two runs of the same day - scrub it before deciding whether the
    #  day's story has actually changed.
    when = r"\d{2} \w{3} \d{4} \d{2}:\d{2}"
    #  The control sheet is a STATUS BOARD, not a record - it has to show
    #  today as it stands right now. If a re-run changes the story, the
    #  earlier board is kept beside it and the file you open is the
    #  current one. (A record is different: see write_once.)
    for path, text in ((ch, control_html(packs, date_tag, book["path"],
                                         generated, fixed_cc)),
                       (ct, control_txt(packs, date_tag, book["path"],
                                        generated))):
        keep_current(path, text, scrub=("built " + when,))
    reg = write_register(root, packs, date_tag, generated, book["path"])

    #  ---- what happened ------------------------------------------------
    ready = [p for p in packs if p["status"] == K.ST_DRAFT_READY]
    held = [p for p in packs if p["status"] == K.ST_FAILED]
    nogear = [p for p in packs if p["status"] == K.ST_NOT_GENERATED]
    say("")
    say("=" * 68)
    say(" {} ready to send   |   {} held for checking   |   {} no gear on hire"
        .format(len(ready), len(held), len(nogear)))
    say("=" * 68)
    if held:
        say("")
        say(" HELD - fix these, then run again:")
        for p in held:
            say("   {}".format(p["company"]))
            for e in p["errors"]:
                say("      - {}".format(e))
    if nogear:
        say("")
        say(" NO GEAR ON HIRE - folder and record made, no email:")
        say("   " + ", ".join(p["company"] for p in nogear))
        #  ...and prove that claim against the live export before
        #  anybody believes it. "No gear on hire" is the one line here
        #  that closes a contractor's day without sending them
        #  anything, so it is the one that has to be right. It was
        #  wrong for four of them on 3 Aug - thirty items between
        #  Sync Lift, Universal Cranes, Tasman Rope Access and Industec
        #  - because the map behind it was a snapshot of an older week.
        try:
            _oh = sorted(
                [p for p in glob.glob(os.path.join(HERE, "Data_SiteIQ",
                                                   "ON_HIRE*.xlsx"))
                 + glob.glob(os.path.join(HERE, "ON_HIRE*.xlsx"))
                 if not os.path.basename(p).startswith("~$")],
                key=os.path.getmtime)
            _stale = K.check_stale_aliases(_oh[-1]) if _oh else []
        except Exception:
            _stale = []
        if _stale:
            say("")
            say(" *** THAT IS WRONG - the company map is out of date:")
            for _wb, _siq, _n in _stale:
                say("     {} shows nothing, but SiteIQ has {} item(s) "
                    "under \"{}\"".format(_wb, _n, _siq))
            say("     Those contractors got no email and no chase list.")
            say("     Fix ALIASES in k2_daily_packs.py and run again.")
    if stale_all:
        say("")
        say(" REBUILD THESE - the PDF is older than the report it came from:")
        for s in stale_all:
            say("   " + s)
    say("")
    say(" Control sheet: {}".format(os.path.relpath(ch, HERE)))
    if reg:
        say(" Register     : {} ({} row(s) added)".format(
            os.path.relpath(reg[0], HERE), reg[1]))
    say("")
    say(" Nothing has been sent. Open the .eml in a company folder,")
    say(" read it, press Send.")
    say("")
    return 0


def build_pack(c, book, root, date_tag, disp, dirs, filed_master, masters,
               fixed_cc, generated):
    """One company: file the reports, write the draft, run the checks,
    set the status, leave the record."""
    company = c["company"]
    kind = c["kind"] or "External contractor"
    external = kind.lower().startswith("external")
    data_company, mapped = K.resolve_company(company)
    folder = os.path.join(root, K.F_COMPANY, K._slug(company), date_tag)

    pack = {"company": company, "kind": kind, "data_company": data_company,
            "folder": folder, "date_tag": date_tag, "book": book["path"],
            "to": list(c["to"]), "cc": list(c["cc"]), "fixed_cc": fixed_cc,
            "body_file": None, "attachments": [], "errors": [],
            "status": K.ST_NOT_GENERATED, "size_mb": 0.0, "stale": [],
            "checks": [], "readback": [], "mapped": mapped,
            "reissued": [], "report_attached": "", "superseded": ""}

    #  --- who is allowed on this email, from the contact register ------
    #  Coates and Cement Australia are on every pack by design (we print
    #  the report; they own the site). Everyone else must be listed on
    #  the Contacts sheet against THIS company - which is what stops an
    #  address pasted into the wrong routing row from ever going out.
    reg = K.contact_index(book)
    home = {"coates", "cement australia"}
    #  ONLY the contact register decides who is allowed on an email. The
    #  fixed Cc list is what MUST be there, never what MAY be there - an
    #  address pasted into every Cc column would otherwise make itself
    #  legitimate just by being everywhere. (25 Jul 2026)
    allowed = set()
    for key in list(home) + [K._key(company)]:
        allowed |= reg.get(key, set())
    everyone = [e.lower() for e in pack["to"] + pack["cc"]]
    pack["unknown_addresses"] = sorted(set(e for e in everyone
                                           if e not in allowed))
    if not external:
        inside = K.domains_of(reg, ["Cement Australia", "Coates"])
        pack["outsiders"] = sorted(set(
            e for e in everyone if e.split("@")[-1] not in inside))

    #  --- the report that goes in the body ----------------------------
    if external:
        #  Subject exactly as A. Fisher set it out (25 Jul 2026):
        #      DGH Engineering | Daily On-Hire Report | 23 July 2026
        pack["subject"] = "{} | Daily On-Hire Report | {}".format(
            company, dt.datetime.strptime(date_tag, "%Y-%m-%d")
            .strftime("%d %B %Y").lstrip("0"))
        stem = None
        if data_company:
            #  THE ACTIVITY & ACCOUNTABILITY REPORT. This is the company
            #  report now - the two-level one, company position first then
            #  a page per person. The old flat on-hire report still builds
            #  for the store's own use; it is not what a contractor gets.
            #  (A. Fisher, 25 Jul 2026 - the whole point of the rebuild.)
            stem = "Coates_K2_Activity_{}".format(
                re.sub(r"[^A-Za-z0-9]+", "_", data_company).strip("_").upper())
        greeting = "Hi {} Team,".format(company)
        lead = ["Below is your current company on-hire position for the K2 "
                "shutdown.<br>This is a read-only SiteIQ snapshot."]
        src_html = built([dirs["pages"]], stem, date_tag, ".html") if stem else None
        filed_body = "{} - On-Hire Report - {}.html".format(
            K._slug(company), date_tag)
    else:
        pack["subject"] = "Cement Australia | Executive Daily Summary | {}".format(
            dt.datetime.strptime(date_tag, "%Y-%m-%d")
            .strftime("%d %B %Y").lstrip("0"))
        greeting = "Hi team,"
        lead = ["Below is the K2 shutdown executive daily summary.<br>"
                "The supporting reports are attached."]
        src_html = masters.get("Executive Daily Summary", {}).get("html")
        filed_body = "Executive Daily Summary - {}.html".format(date_tag)

    note = ""
    if src_html:
        dest, note, reissued = K._place(src_html, folder, filed_body)
        pack["body_file"] = dest
        if reissued:
            pack["reissued"].append(os.path.basename(dest))
        pdf_src = None
        if external and data_company:
            pdf_src = built([dirs["pdf"]], stem, date_tag, ".pdf")
        elif not external:
            pdf_src = masters.get("Executive Daily Summary", {}).get("pdf")
        if is_stale(pdf_src, src_html):
            pack["stale"].append(os.path.basename(pdf_src))

    #  --- what goes on the paperclip -----------------------------------
    #  Nothing on hire means nothing to attach either - the folder and
    #  the record are made so the day is complete, and that is all.
    wanted = ([K.SAFETY_PDF] if external else list(K.CEMENT_ATTACHMENTS))
    if not data_company:
        wanted = []
    for label in wanted:
        src = masters.get(label, {}).get("pdf")
        if not src:
            pack["errors"].append(
                "{} has no PDF - it was not built for {}".format(label, disp))
            continue
        if is_stale(src, masters.get(label, {}).get("html")):
            pack["stale"].append(os.path.basename(src))
        name = K.FILED_NAME.get(label, label + " - {d}").format(d=date_tag) + ".pdf"
        dest, _n, reissued = K._place(src, folder, name)
        if reissued:
            pack["reissued"].append(os.path.basename(dest))
        pack["attachments"].append(dest)

    #  --- the draft -----------------------------------------------------
    eml = os.path.join(folder, "Email Draft - {}.eml".format(date_tag))
    if pack["body_file"]:
        imgs, img_problems = images_for(src_html, dirs["emails"])
        pack["errors"] += img_problems

        #  Will the whole thing actually go? Base64 adds about a third.
        #  A report too big to email is attached as its PDF instead, and
        #  the email says so - it is never quietly cut short.
        #  +0.5 MB for MIME headers and the HTML body - the same margin
        #  email_images.email_weight_mb carries (a 33-page email once
        #  estimated 9.9 MB and landed 10.2, just over the line).
        weight = (sum(os.path.getsize(p) for _c, p in imgs)
                  + sum(os.path.getsize(x) for x in pack["attachments"]
                        if os.path.isfile(x))) * 1.37 / 1048576.0 + 0.5
        if imgs and weight > K.MAX_EMAIL_MB:
            report_pdf = None
            if external and data_company:
                report_pdf = built([dirs["pdf"]], stem, date_tag, ".pdf")
            else:
                report_pdf = masters.get("Executive Daily Summary",
                                         {}).get("pdf")
            if report_pdf:
                name = "{} - {}.pdf".format(
                    os.path.splitext(os.path.basename(pack["body_file"]))[0]
                    .rsplit(" - ", 1)[0], date_tag)
                dest, _n, _r = K._place(report_pdf, folder, name)
                pack["attachments"].insert(0, dest)
                pack["report_attached"] = os.path.basename(dest)
                say("     {}: report is {:.0f} MB of pages - attached as a "
                    "PDF instead of pasted in".format(company, weight))
                imgs = []

        if imgs:
            #  The pictures carry their own baked orange border - the
            #  letter adds the greeting, sign-off and white space only.
            import email_images
            body = email_images.letter(
                images_stack([cid for cid, _p in imgs]),
                greeting=greeting, lead=lead, signoff=SIGNOFF,
                draw_border=False)
        else:
            #  No page pictures on this machine, so the report cannot be
            #  in the body. It then has to be ON the email - an email
            #  that SAYS the report is attached and doesn't attach it is
            #  worse than no email at all. (25 Jul 2026)
            imgs = []
            if not pack["report_attached"]:
                fallback = built([dirs["pdf"]], stem, date_tag, ".pdf") \
                    if (external and data_company) else \
                    masters.get("Executive Daily Summary", {}).get("pdf")
                fname = "{} - {}{}".format(
                    os.path.splitext(os.path.basename(pack["body_file"]))[0]
                    .rsplit(" - ", 1)[0], date_tag,
                    ".pdf" if fallback else ".html")
                dest, _n, _r = K._place(fallback or pack["body_file"],
                                        folder, fname)
                if dest not in pack["attachments"]:
                    pack["attachments"].insert(0, dest)
                pack["report_attached"] = os.path.basename(dest)
            body = plain_body(
                greeting, lead,
                "The full report for {} is attached to this email.".format(disp),
                [os.path.basename(x) for x in pack["attachments"]])

        #  A draft already here that is NOT the email we would write now
        #  must not stay clickable - that is how a corrected routing gets
        #  sent with the old addresses. The old one is kept, renamed so
        #  it cannot be opened and sent by accident.
        if os.path.exists(eml):
            want = fingerprint(pack["subject"], imgs, pack["attachments"],
                               pack["to"], pack["cc"], body)
            same = False
            try:
                had = read_back(eml)
                same = (had.get("stamp") == want
                        and sorted(had["to"]) ==
                        sorted(e.lower() for e in pack["to"])
                        and sorted(had["cc"]) ==
                        sorted(e.lower() for e in pack["cc"]))
            except Exception:
                same = False
            if same:
                note = "already filed - left untouched"
                pack["size_mb"] = os.path.getsize(eml) / 1048576.0
            else:
                for n in range(1, 60):
                    old = "{} (superseded {}).eml.txt".format(eml[:-4], n)
                    if not os.path.exists(old):
                        os.rename(eml, old)
                        pack["superseded"] = os.path.basename(old)
                        break
        if not os.path.exists(eml):
            msg = build_eml(pack["subject"], body, imgs, pack["attachments"],
                            pack["to"], pack["cc"], sender=SENDER)
            with open(eml, "wb") as fh:
                fh.write(eml_bytes(msg))
            pack["size_mb"] = os.path.getsize(eml) / 1048576.0
            if pack["superseded"]:
                note = "rewritten - the earlier draft is kept as {}".format(
                    pack["superseded"])
        #  When the body is the report's page images there is nothing to
        #  keep beside it - the report itself is already in this folder
        #  and reads far better than a page of dead image links. Only the
        #  written-out body (no images on this machine) is worth filing.
        if not imgs:
            K.write_once(os.path.join(
                folder, "Email Body - {}.html".format(date_tag)), body,
                ignore=())
        pack["draft"] = eml

        #  --- read the draft back, as a mail client would --------------
        try:
            got = read_back(eml)
        except Exception as exc:
            _rb_err = "the draft could not be read back: {}".format(exc)
            pack["readback"].append(_rb_err)
            got = None
        rb = list(pack["readback"])
        if got:
            want_to = sorted(e.lower() for e in pack["to"])
            want_cc = sorted(e.lower() for e in pack["cc"])
            if sorted(got["to"]) != want_to:
                rb.append(
                    "the draft's To reads back as {} address(es), not the {} "
                    "in the workbook".format(len(got["to"]), len(want_to)))
            if sorted(got["cc"]) != want_cc:
                rb.append(
                    "the draft's Cc reads back as {} address(es), not the {} "
                    "in the workbook".format(len(got["cc"]), len(want_cc)))
            if sorted(got["cids_used"]) != sorted(got["inline"]):
                rb.append(
                    "the report images and the email body disagree "
                    "({} in the body, {} attached inline)".format(
                        len(got["cids_used"]), len(got["inline"])))
            want_files = sorted(os.path.basename(x)
                                for x in pack["attachments"])
            if got["subject"] != pack["subject"]:
                rb.append("the draft's subject reads back as '{}'".format(
                    got["subject"]))
            if sorted(got["attachments"]) != want_files:
                rb.append(
                    "the draft's attachments read back as {}, not {}".format(
                        ", ".join(got["attachments"]) or "none",
                        ", ".join(want_files) or "none"))
        pack["readback"] = rb

    #  A report re-issued during the day is normal - a fresh SiteIQ pull
    #  and a second run. The pack is simply rebuilt from the new report.
    #  It is only a problem if you have ALREADY SENT today's, because
    #  then sending again is your decision, not the script's.
    if pack["reissued"] and not K.record_signed(folder, date_tag):
        note = "report re-issued today - the pack was rebuilt from it"
        pack["reissued"] = []

    #  --- the checks, then the status ------------------------------------
    pack["checks"] = K.run_checks(pack, date_tag)
    pack["errors"] += K.failures(pack["checks"])

    if not mapped:
        #  Nobody has said what this company is called in the register.
        #  That is a person's decision, not something to guess at - and
        #  it must never look like a quiet day.
        pack["status"] = K.ST_FAILED
        pack["errors"] = [
            "'{}' is not on the company map, so no report could be matched "
            "to it. Either the workbook spelling changed or this is a new "
            "company - add it to ALIASES in k2_daily_packs.py with the name "
            "the on-hire register uses.".format(company)]
    elif not data_company:
        pack["status"] = K.ST_NOT_GENERATED
        #  Nothing on hire is not a failure - it is a fact. The folder and
        #  the record are still made so the day is complete, and the
        #  reason is stated once instead of as a wall of failed checks.
        pack["errors"] = ["{} holds nothing on hire today - no report to "
                          "send. The folder and this record are made so "
                          "the day is complete.".format(company)]
        keep = ("Every address on the contact register for this company",
                "Four fixed oversight contacts in Cc",
                "At least one confirmed address in To")
        pack["errors"] += [e for e in K.failures(pack["checks"])
                           if e.split(" (")[0] in keep]
    elif not pack["body_file"]:
        pack["status"] = K.ST_FAILED
    elif pack["errors"]:
        pack["status"] = K.ST_FAILED
    else:
        pack["status"] = K.ST_DRAFT_READY

    #  A .eml in a company folder means ONE thing: checked, and ready to
    #  send. A held pack keeps everything it built so you can see what
    #  went wrong, but not as a file you can double-click and send by
    #  mistake. "If any check fails, the email stays in Draft - Check
    #  Required" - this is that rule, made physical.
    note = hold_or_release(pack, folder, date_tag) or note

    K.write_record(pack, date_tag, generated)

    flag = {K.ST_DRAFT_READY: "READY", K.ST_FAILED: "HELD",
            K.ST_NOT_GENERATED: "no gear on hire"}.get(pack["status"],
                                                       pack["status"])
    if not pack["mapped"]:
        flag = "NOT ON THE MAP"
    say("   {:<24} {:<16} {:>2} To {:>2} Cc {:>2} attach {:>5} MB  {}".format(
        company[:24], flag, len(pack["to"]), len(pack["cc"]),
        len(pack["attachments"]), "{:.1f}".format(pack["size_mb"]),
        note or ""))
    return pack


if __name__ == "__main__":
    sys.exit(main())
