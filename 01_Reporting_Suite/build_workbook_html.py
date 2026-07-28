#!/usr/bin/env python3
# =====================================================================
#  COATES | K2 WORKBOOK -> HTML MIRROR
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Turns every tab of the K2 workbook into its own clean, dark
#  Coates-branded web page, all tied together with a Home hub and a
#  side menu. Read-only: it NEVER changes your workbook - it just
#  reads the latest saved copy and rebuilds the pages.
#
#  Usage:  double-click 11_RUN_WORKBOOK_HTML.bat
#          (or: python build_workbook_html.py [path-to-workbook.xlsm])
#  Output: K2_Workbook_Pages\  (open index.html)
# =====================================================================

import sys, os, re, html, glob
from datetime import datetime, date, time
import openpyxl
from openpyxl.utils import get_column_letter

# ---------- brand ----------
ORG   = "#F26222"   # Coates orange
BG    = "#0E1116"
CARD  = "#171B22"
LINE  = "#2A313C"
INK   = "#E6E9EE"
MUT   = "#8B9099"
HEADBG= "#1D1D1B"

# validated data-viz series colours (colour-blind safe on the dark surface)
FCAST = "#3987e5"   # Forecast series (blue)
ACT   = "#d95926"   # Actual series (deep orange, Coates family)
GOOD  = "#0ca30c"
BAD   = "#d03b3b"
AMBER = "#fab219"

SITE  = "Cement Australia · K2 Shutdown 2026 · Gladstone"

# ---------- how the tabs are grouped in the menu / hub ----------
# (stripped sheet name -> group). Order here = order in the menu.
GROUPS = [
    ("On-Hire", ["Gas Monitors Onhire", "Radios Onhire", "Tooling Onhire.",
                 "Plant Onhire", "Milwaukee Onhire"]),
    ("Plant & Site", ["Plant Onsite", "Live Utilisation", "Plant Equipment",
                      "Plant Variance"]),
    ("Costing & Spend", ["Shutdown Costing", "Company Accountability", "Company Spend"]),
    ("Daily & Dashboards", ["Daily Tracking", "K2 Dashboard"]),
    ("Tooling & Consumables", ["Tooling to Supply", "Consumables Stock",
                               "Consumables Utilisation", "Consumables Movement"]),
    ("Stocktake", ["Stocktake Log", "Stocktake Coverage", "RENTAL_STOCK (2)"]),
    ("Compliance & Safety", ["Coates Way + Safety", "Compliance + Assurance"]),
    ("Data & Config", ["Dashboard Data", "Controls & Exceptions",
                       "Control Config", "Control Data"]),
]

# pure backend/config tabs: clean table only (auto KPIs/charts would be noise)
NO_VIZ = {"Control Config", "Control Data", "Controls & Exceptions"}

# featured live reports (linked from the hub if they exist next door)
FEATURED = [
    ("Plant On Site Dashboard", "Coates_K2_Plant_Dashboard_LATEST.html"),
    ("Site Plant Report",       "Coates_SitePlant_Report_LATEST.html"),
    ("Plant Audit List",        "Coates_K2_Plant_Audit_List_LATEST.html"),
    ("Executive Summary",       "Coates_K2_Executive_Summary_LATEST.html"),
]


def slug(s):
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return "tab_" + (s or "sheet")


def group_of(name):
    n = name.strip()
    for gname, members in GROUPS:
        if n in members:
            return gname
    return "Other"


def fmt_value(cell):
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, datetime):
        if v.hour or v.minute or v.second:
            return v.strftime("%d %b %Y %H:%M")
        return v.strftime("%d %b %Y")
    if isinstance(v, date):
        return v.strftime("%d %b %Y")
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, (int, float)):
        fmt = cell.number_format or "General"
        try:
            if "%" in fmt:
                dp = 2 if "0.00%" in fmt else (1 if "0.0%" in fmt else 0)
                return ("{:,.%df}%%" % dp).format(v * 100)
            if any(sym in fmt for sym in ("$", "€", "£", "[$")):
                return ("-$" + "{:,.2f}".format(abs(v))) if v < 0 else ("$" + "{:,.2f}".format(v))
            if "#,##0" in fmt or "* #" in fmt:
                return "{:,.2f}".format(v) if ".00" in fmt else \
                       ("{:,.0f}".format(v) if float(v).is_integer() else "{:,.2f}".format(v))
            if isinstance(v, int) or float(v).is_integer():
                return str(int(v))
            return ("%.4f" % v).rstrip("0").rstrip(".")
        except Exception:
            return str(v)
    return str(v)


def used_bounds(ws):
    """Real used range by scanning for non-empty cells."""
    maxr = maxc = 0
    for row in ws.iter_rows():
        for c in row:
            if c.value not in (None, ""):
                if c.row > maxr:
                    maxr = c.row
                if c.column > maxc:
                    maxc = c.column
    return maxr, maxc


def render_table(ws):
    maxr, maxc = used_bounds(ws)
    if maxr == 0:
        return '<p class="empty">This tab has no data.</p>', 0, 0

    # merged cells
    covered = set()
    spans = {}
    for rng in ws.merged_cells.ranges:
        r0, c0, r1, c1 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        if r0 > maxr or c0 > maxc:
            continue
        spans[(r0, c0)] = (min(r1, maxr) - r0 + 1, min(c1, maxc) - c0 + 1)
        for r in range(r0, min(r1, maxr) + 1):
            for c in range(c0, min(c1, maxc) + 1):
                if (r, c) != (r0, c0):
                    covered.add((r, c))

    # ITEM_NUMBER is the identifier - drop any "Barcode" column from the mirror
    # (scan every row: some tabs have a second barcode header in a side-panel).
    _bc = ("barcode", "item barcode", "item_barcode", "latest barcode", "barcode no")
    skip = set()
    for c in range(1, maxc + 1):
        for r in range(1, maxr + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() in _bc:
                skip.add(c)
                break

    out = ['<div class="tblwrap"><table class="grid">']
    for r in range(1, maxr + 1):
        tag = "th" if r == 1 else "td"
        out.append('<tr class="{}">'.format("hdr" if r == 1 else "bd"))
        for c in range(1, maxc + 1):
            if c in skip or (r, c) in covered:
                continue
            cell = ws.cell(row=r, column=c)
            val = html.escape(fmt_value(cell))
            attr = ""
            if (r, c) in spans:
                rs, cs = spans[(r, c)]
                cs -= sum(1 for cc in range(c, c + cs) if cc in skip)  # drop skipped cols
                if rs > 1:
                    attr += ' rowspan="%d"' % rs
                if cs > 1:
                    attr += ' colspan="%d"' % cs
            # right-align numbers
            cls = ""
            if tag == "td" and isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cls = ' class="n"'
            out.append("<{t}{a}{c}>{v}</{t}>".format(t=tag, a=attr, c=cls, v=val))
        out.append("</tr>")
    out.append("</table></div>")
    return "".join(out), maxr, maxc


CSS = """
*{box-sizing:border-box}
html,body{margin:0;padding:0}
body{background:__BG__;color:__INK__;font-family:'Segoe UI',Arial,Helvetica,sans-serif;font-size:14px}
a{color:inherit;text-decoration:none}
/* sidebar */
.side{position:fixed;top:0;left:0;width:238px;height:100vh;overflow-y:auto;background:#0B0E12;
  border-right:1px solid __LINE__;padding:0 0 40px 0;z-index:40}
.side .sbrand{padding:16px 16px 12px 16px;border-bottom:1px solid __LINE__;position:sticky;top:0;background:#0B0E12}
.side .sbrand b{color:#fff;font-size:18px;font-weight:800;letter-spacing:.5px}
.side .sbrand i{color:__ORG__;font-style:normal}
.side .sbrand .eq{display:block;color:__MUT__;font-size:11px;margin-top:2px}
.side a.home{display:block;margin:10px 12px;padding:9px 12px;background:__ORG__;color:#fff;border-radius:7px;font-weight:700;font-size:13px}
.side .grp{color:__MUT__;font-size:10.5px;letter-spacing:1px;text-transform:uppercase;padding:12px 16px 5px 16px;font-weight:700}
.side a.lnk{display:block;padding:7px 16px 7px 16px;color:#c4c9d2;font-size:12.5px;border-left:3px solid transparent}
.side a.lnk:hover{background:#141922;color:#fff}
.side a.lnk.act{background:#141922;color:#fff;border-left:3px solid __ORG__;font-weight:600}
/* main */
.main{margin-left:238px;padding:0}
header.hd{background:__HEADBG__;border-bottom:3px solid __ORG__;padding:16px 26px}
header.hd .r1{display:flex;justify-content:space-between;align-items:flex-start;gap:12px;flex-wrap:wrap}
header.hd .kick{color:__ORG__;font-weight:700;font-size:11px;letter-spacing:1.5px;text-transform:uppercase}
header.hd h1{margin:3px 0 2px 0;font-size:23px;color:#fff}
header.hd .site{color:__MUT__;font-size:12.5px}
header.hd .siteiq{text-align:right;color:#fff;font-weight:700;font-size:12px}
header.hd .siteiq .o{color:__ORG__}
header.hd .meta{color:__MUT__;font-size:11.5px;text-align:right;margin-top:3px}
.body{padding:22px 26px 60px 26px}
.card{background:__CARD__;border:1px solid __LINE__;border-radius:12px;padding:6px;overflow:hidden}
.tblwrap{overflow-x:auto;max-width:100%}
table.grid{border-collapse:collapse;width:100%;font-size:12.5px}
table.grid td,table.grid th{border:1px solid #232a34;padding:5px 9px;text-align:left;vertical-align:top;white-space:nowrap}
table.grid tr.hdr th{position:sticky;top:0;background:#20262F;color:#fff;font-weight:700;border-bottom:2px solid __ORG__;z-index:2}
table.grid tr.bd:nth-child(even) td{background:#141922}
table.grid tr.bd:nth-child(odd) td{background:#12161d}
table.grid td.n{text-align:right;font-variant-numeric:tabular-nums}
.empty{color:__MUT__;padding:26px;font-style:italic}
.badge{display:inline-block;background:rgba(242,98,34,.15);color:__ORG__;border:1px solid rgba(242,98,34,.42);
  border-radius:20px;padding:3px 12px;font-size:11px;font-weight:700;letter-spacing:.6px;text-transform:uppercase}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(178px,1fr));gap:12px;margin:12px 0 16px 0}
.kpi{background:__CARD__;border:1px solid __LINE__;border-left:3px solid __ORG__;border-radius:11px;padding:13px 15px}
.kpi .v{color:#fff;font-size:23px;font-weight:800;line-height:1.12;font-variant-numeric:tabular-nums}
.kpi .l{color:__MUT__;font-size:10.5px;margin-top:5px;text-transform:uppercase;letter-spacing:.6px}
.viz{background:__CARD__;border:1px solid __LINE__;border-radius:12px;padding:15px 17px;margin-bottom:16px;overflow-x:auto}
.viz .vnote{color:__MUT__;font-size:11px;margin-top:8px}
.chartimg{width:100%;max-width:980px;border:1px solid __LINE__;border-radius:10px;margin:0 0 16px 0;display:block}
.note{color:__MUT__;font-size:12px;margin:4px 0 14px 0}
footer.ft{margin-left:238px;padding:14px 26px 30px 26px;color:__MUT__;font-size:11px;border-top:1px solid __LINE__}
/* hub */
.hub{padding:22px 26px 60px 26px}
.hero{background:linear-gradient(120deg,#171B22,#10141a);border:1px solid __LINE__;border-radius:14px;padding:22px 24px;margin-bottom:20px}
.hero h2{margin:0 0 4px 0;color:#fff;font-size:20px}
.hero p{margin:0;color:__MUT__;font-size:13px}
.grpttl{color:#fff;font-size:14px;font-weight:700;margin:22px 0 10px 0;padding-bottom:6px;border-bottom:1px solid __LINE__}
.grpttl span{color:__MUT__;font-weight:400;font-size:12px;margin-left:8px}
.tiles{display:grid;grid-template-columns:repeat(auto-fill,minmax(214px,1fr));gap:12px}
.tile{background:__CARD__;border:1px solid __LINE__;border-left:3px solid __ORG__;border-radius:10px;padding:13px 14px;transition:transform .08s,border-color .12s}
.tile:hover{transform:translateY(-2px);border-color:__ORG__}
.tile b{color:#fff;font-size:13.5px;display:block}
.tile .dim{color:__MUT__;font-size:11px;margin-top:5px}
.feat{background:__CARD__;border:1px solid __LINE__;border-radius:10px;padding:13px 14px}
.feat a{display:inline-block;margin:5px 8px 5px 0;padding:8px 13px;background:#141922;border:1px solid __LINE__;border-radius:7px;color:#e6e9ee;font-size:12.5px;font-weight:600}
.feat a:hover{border-color:__ORG__;color:#fff}
.hamb{display:none}
@media(max-width:820px){
  .side{transform:translateX(-100%);transition:transform .2s}
  body.nav .side{transform:none}
  .main,footer.ft{margin-left:0}
  .hamb{display:inline-block;position:fixed;top:12px;left:12px;z-index:60;background:__ORG__;color:#fff;
    border:none;border-radius:8px;padding:8px 12px;font-size:16px;font-weight:800;cursor:pointer}
}
""".replace("__BG__", BG).replace("__CARD__", CARD).replace("__LINE__", LINE)\
   .replace("__INK__", INK).replace("__MUT__", MUT).replace("__ORG__", ORG)\
   .replace("__HEADBG__", HEADBG)


def build_sidebar(pages, active_slug):
    # pages: list of dicts {name, slug, group}
    parts = ['<button class="hamb" onclick="document.body.classList.toggle(\'nav\')">☰</button>']
    parts.append('<nav class="side">')
    parts.append('<div class="sbrand"><b>COATES<i>.</i></b><span class="eq">Equipped for anything</span></div>')
    parts.append('<a class="home" href="index.html">■  Home hub</a>')
    for gname, _members in GROUPS + [("Other", [])]:
        grp_pages = [p for p in pages if p["group"] == gname]
        if not grp_pages:
            continue
        parts.append('<div class="grp">%s</div>' % html.escape(gname))
        for p in grp_pages:
            act = " act" if p["slug"] == active_slug else ""
            parts.append('<a class="lnk%s" href="%s.html">%s</a>' %
                         (act, p["slug"], html.escape(p["name"].strip())))
    parts.append("</nav>")
    return "".join(parts)


def page_shell(title, sidebar, header, body_html, footer, extra=""):
    return (
        "<!doctype html><html lang='en'><head><meta charset='utf-8'>"
        "<meta name='viewport' content='width=device-width,initial-scale=1'>"
        "<title>" + html.escape(title) + " · Coates K2</title>"
        "<style>" + CSS + "</style></head><body>"
        + sidebar +
        "<div class='main'>" + header + "<div class='body'>" + extra + body_html + "</div></div>"
        + footer +
        "</body></html>"
    )


def header_html(title, asat):
    return (
        "<header class='hd'><div class='r1'>"
        "<div><div class='kick'>Coates · K2 Workbook</div>"
        "<h1>" + html.escape(title) + "</h1>"
        "<div class='site'>" + SITE + "</div></div>"
        "<div><div class='siteiq'>POWERED BY <span class='o'>SITEIQ</span></div>"
        "<div class='meta'>As at " + asat + "<br>Author: <b style='color:#cfd3da'>Andrew Fisher</b></div>"
        "</div></div></header>"
    )


def footer_html(wbname, sheet, gen):
    return ("<footer class='ft'>Source: <b>" + html.escape(wbname) + "</b>"
            + ("" if sheet is None else " · tab: <b>" + html.escape(sheet) + "</b>")
            + " · Generated " + gen +
            " · Point-in-time snapshot of the last saved workbook. Figures are read-only; the workbook is the source of truth."
            "</footer>")


def svg_line_chart(cats, series, title, w=980, h=360):
    """Pure-Python inline SVG line chart (no external libraries).
    series = list of (name, values, colour, dashed_bool)."""
    vals = [v for _n, vv, _c, _d in series for v in vv if isinstance(v, (int, float))]
    if len(cats) < 2 or not vals:
        return ""
    pl, pr, pt, pb = 66, 18, 54, 68
    pw, ph = w - pl - pr, h - pt - pb
    vmax = max(vals + [0.0]); vmin = min(vals + [0.0])
    if vmax == vmin:
        vmax = vmin + 1
    n = len(cats)
    def X(i): return pl + pw * i / (n - 1)
    def Y(v): return pt + ph * (1 - (v - vmin) / (vmax - vmin))
    o = ["<svg viewBox='0 0 %d %d' width='100%%' style='max-width:%dpx;display:block' "
         "xmlns='http://www.w3.org/2000/svg' font-family='Segoe UI,Arial,sans-serif'>" % (w, h, w)]
    o.append("<text x='%d' y='27' fill='#fff' font-size='15' font-weight='700'>%s</text>"
             % (pl, html.escape(title)))
    for k in range(6):
        v = vmin + (vmax - vmin) * k / 5; yy = Y(v)
        o.append("<line x1='%d' y1='%.1f' x2='%.1f' y2='%.1f' stroke='%s' stroke-width='0.7' opacity='0.5'/>"
                 % (pl, yy, pl + pw, yy, LINE))
        o.append("<text x='%d' y='%.1f' fill='%s' font-size='9' text-anchor='end'>$%s</text>"
                 % (pl - 8, yy + 3, MUT, "{:,.0f}".format(v)))
    step = max(1, n // 12)
    for i in range(0, n, step):
        xx = X(i); yb = pt + ph + 16
        o.append("<text x='%.1f' y='%.1f' fill='%s' font-size='9' text-anchor='end' "
                 "transform='rotate(-45 %.1f %.1f)'>%s</text>"
                 % (xx, yb, MUT, xx, yb, html.escape(cats[i])))
    for name, vv, colour, dashed in series:
        dash = " stroke-dasharray='6 4'" if dashed else ""
        pts = " ".join("%.1f,%.1f" % (X(i), Y(v)) for i, v in enumerate(vv)
                       if isinstance(v, (int, float)))
        if pts:
            o.append("<polyline points='%s' fill='none' stroke='%s' stroke-width='2.4'%s stroke-linejoin='round'/>"
                     % (pts, colour, dash))
        if not dashed:
            for i, v in enumerate(vv):
                if isinstance(v, (int, float)):
                    o.append("<circle cx='%.1f' cy='%.1f' r='2.6' fill='%s'/>" % (X(i), Y(v), colour))
    lx = pl + 8
    for j, (name, vv, colour, dashed) in enumerate(series):
        yy = pt + 2 + j * 16
        dash = " stroke-dasharray='6 4'" if dashed else ""
        o.append("<line x1='%d' y1='%d' x2='%d' y2='%d' stroke='%s' stroke-width='3'%s/>"
                 % (lx, yy, lx + 22, yy, colour, dash))
        o.append("<text x='%d' y='%d' fill='%s' font-size='10'>%s</text>"
                 % (lx + 28, yy + 3, INK, html.escape(str(name))))
    o.append("</svg>")
    return "".join(o)


def dashboard_chart_svg(wb):
    """The one workbook chart: Daily Forecast vs Actual, off 'Dashboard Data'."""
    if "Dashboard Data" not in wb.sheetnames:
        return ""
    ws = wb["Dashboard Data"]
    h1 = ws.cell(1, 2).value or "Forecast"
    h2 = ws.cell(1, 3).value or "Actual"
    cats, s1, s2 = [], [], []
    for r in range(2, 37):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        if isinstance(a, (datetime, date)):
            a = a.strftime("%d %b")
        cats.append(str(a))
        b = ws.cell(r, 2).value; c = ws.cell(r, 3).value
        s1.append(b if isinstance(b, (int, float)) else None)
        s2.append(c if isinstance(c, (int, float)) else None)
    return svg_line_chart(cats, [(str(h1), s1, MUT, True), (str(h2), s2, ORG, False)],
                          "Daily Forecast vs Actual ($)")


def money0(v):
    try:
        v = float(v)
    except Exception:
        return str(v)
    return ("-$" + "{:,.0f}".format(abs(v))) if v < 0 else "$" + "{:,.0f}".format(v)


def _hint(name):
    n = str(name).lower()
    keys = ("value", "cost", "total", "amount", "exposure", "spend", "qty",
            "quantit", "hire", "daily", "price", "fee", "charge", "revenue",
            "$", "sum", "period")
    return 1 if any(k in n for k in keys) else 0


def _no_sum(name):
    """Columns that must NOT be summed into a 'Total' KPI/chart: per-unit rates,
    averages, ratios, percentages - adding them up is meaningless."""
    n = str(name).lower()
    bad = ("rate", "avg", "average", " per ", "per hirer", "per unit", "each",
           "ratio", "%", "util", "unit price", "unit cost", "/day", "/ day")
    return any(b in n for b in bad)


def _row_stats(ws, r, maxc):
    ne = txt = 0
    for c in range(1, maxc + 1):
        v = ws.cell(r, c).value
        if v in (None, ""):
            continue
        ne += 1
        if isinstance(v, str):
            txt += 1
    return ne, txt


def find_header(ws, maxr, maxc):
    """First row (within the top 5) that looks like a real column header with
    data beneath it. Skips leading title-banner rows. Returns row index or None."""
    for r in range(1, min(5, maxr) + 1):
        ne, txt = _row_stats(ws, r, maxc)
        if ne >= 3 and txt >= ne * 0.7:
            ne2 = _row_stats(ws, r + 1, maxc)[0] if r + 1 <= maxr else 0
            if ne2 >= 2:
                return r
    return None


def find_main_cols(ws, hr, maxc):
    """Width of the main table: contiguous non-empty header cells from column A,
    stopping at the first gap of 2+ empty header cells (which precedes any
    right-hand side-panel)."""
    last = 0; gap = 0
    for c in range(1, maxc + 1):
        if ws.cell(hr, c).value not in (None, ""):
            last = c; gap = 0
        else:
            gap += 1
            if gap >= 2 and last > 0:
                break
    return last


def find_block_end(ws, hr, main_maxc, maxr):
    """Last row of the first contiguous data block under the header (stops at the
    first fully-blank row across the main columns, so trailing summary blocks are
    left out of the aggregation)."""
    for r in range(hr + 1, maxr + 1):
        if all(ws.cell(r, c).value in (None, "") for c in range(1, main_maxc + 1)):
            return r - 1
    return maxr


def profile_columns(ws, hr, block_end, main_maxc):
    start = hr + 1
    profs = []
    for c in range(1, main_maxc + 1):
        hv = ws.cell(hr, c).value
        name = str(hv).strip() if hv not in (None, "") else ("Col " + get_column_letter(c))
        nums = []; dates = 0; texts = 0; nonempty = 0; currency = False; distinct = set()
        for r in range(start, block_end + 1):
            v = ws.cell(r, c).value
            if v in (None, ""):
                continue
            nonempty += 1; distinct.add(str(v))
            if isinstance(v, bool):
                texts += 1
            elif isinstance(v, (int, float)):
                nums.append(v)
                if not currency:
                    fmt = ws.cell(r, c).number_format or ""
                    if any(s in fmt for s in ("$", "€", "£", "[$")):
                        currency = True
            elif isinstance(v, (datetime, date)):
                dates += 1
            else:
                texts += 1
        if nonempty == 0:
            kind = "empty"
        elif len(nums) >= nonempty * 0.6:
            kind = "num"
        elif dates >= nonempty * 0.6:
            kind = "date"
        else:
            kind = "text"
        profs.append({"idx": c, "name": name, "kind": kind, "currency": currency,
                      "nonempty": nonempty, "total": sum(nums) if nums else 0.0,
                      "ncount": len(nums), "distinct": len(distinct)})
    return profs


def pick_kpis(profs, nrows):
    tiles = [("Records", "{:,}".format(nrows))]
    valcols = [p for p in profs if p["kind"] == "num" and (p["currency"] or _hint(p["name"]))
               and not _no_sum(p["name"])]
    valcols.sort(key=lambda p: (p["currency"], _hint(p["name"]), abs(p["total"])), reverse=True)
    for p in valcols[:3]:
        v = money0(p["total"]) if p["currency"] else "{:,.0f}".format(p["total"])
        tiles.append(("Total · " + p["name"], v))
    if len(tiles) < 4:
        for p in profs:
            if p["kind"] == "text" and 1 < p["distinct"] < max(2, nrows):
                tiles.append(("Distinct · " + p["name"], "{:,}".format(p["distinct"])))
                break
    return tiles[:4]


def pick_chart(profs, ws, hr, block_end):
    start = hr + 1
    datecols = [p for p in profs if p["kind"] == "date"]
    numcols = [p for p in profs if p["kind"] == "num" and p["ncount"] >= 3
               and (p["currency"] or _hint(p["name"])) and not _no_sum(p["name"])]
    numcols.sort(key=lambda p: (p["currency"], _hint(p["name"]), abs(p["total"])), reverse=True)
    textcols = [p for p in profs if p["kind"] == "text" and 2 <= p["distinct"] <= 40]
    if not numcols:
        return ""
    vc = numcols[0]
    # category present -> aggregated top-N bars (best for item/list tabs)
    if textcols:
        lc = textcols[0]
        agg = {}
        for r in range(start, block_end + 1):
            k = ws.cell(r, lc["idx"]).value
            v = ws.cell(r, vc["idx"]).value
            if k in (None, "") or not isinstance(v, (int, float)):
                continue
            agg[str(k)] = agg.get(str(k), 0) + v
        pairs = sorted(agg.items(), key=lambda kv: abs(kv[1]), reverse=True)[:12]
        if len(pairs) >= 2:
            return svg_bar_h(pairs, "Top " + str(len(pairs)) + " · " + vc["name"] +
                             " by " + lc["name"], currency=vc["currency"])
    # otherwise a date column -> time-series line (Daily Tracking, Dashboard Data, etc.)
    if datecols:
        dc = datecols[0]
        pts = []
        for r in range(start, block_end + 1):
            d = ws.cell(r, dc["idx"]).value
            v = ws.cell(r, vc["idx"]).value
            if d in (None, "") or not isinstance(d, (datetime, date)):
                continue
            pts.append((d, v if isinstance(v, (int, float)) else None))
        pts.sort(key=lambda t: t[0])  # sort by date so the line reads left-to-right
        cats = [d.strftime("%d %b") for d, _v in pts]
        vals = [v for _d, v in pts]
        if sum(1 for v in vals if isinstance(v, (int, float))) >= 3:
            return svg_line_chart(cats, [(vc["name"], vals, ORG, False)],
                                  vc["name"] + " over " + dc["name"])
    return ""


def svg_bar_h(pairs, title, currency=False, w=980):
    n = len(pairs)
    top, rowh = 46, 28
    h = top + n * rowh + 12
    maxv = max([abs(v) for _k, v in pairs] + [1.0])
    labw, gap, valw = 220, 12, 120
    bx = labw + gap
    bw = w - bx - valw
    o = ["<svg viewBox='0 0 %d %d' width='100%%' style='max-width:%dpx;display:block' "
         "xmlns='http://www.w3.org/2000/svg' font-family='Segoe UI,Arial,sans-serif'>" % (w, h, w)]
    o.append("<text x='0' y='26' fill='#fff' font-size='15' font-weight='700'>%s</text>" % html.escape(title))
    for i, (k, v) in enumerate(pairs):
        y = top + i * rowh
        lab = html.escape((k[:30] + "…") if len(k) > 31 else k)
        o.append("<text x='0' y='%d' fill='%s' font-size='11.5'>%s</text>" % (y + 13, INK, lab))
        wv = max(2.0, bw * abs(v) / maxv)
        o.append("<rect x='%d' y='%d' width='%.1f' height='16' rx='3' fill='%s' opacity='0.92'/>" % (bx, y + 2, wv, ORG))
        vs = money0(v) if currency else "{:,.0f}".format(v)
        o.append("<text x='%.1f' y='%d' fill='%s' font-size='11'>%s</text>" % (bx + wv + 8, y + 14, MUT, vs))
    o.append("</svg>")
    return "".join(o)


def kpi_tiles(tiles):
    chips = "".join("<div class='kpi'><div class='v'>%s</div><div class='l'>%s</div></div>"
                    % (html.escape(str(v)), html.escape(str(l))) for l, v in tiles)
    return "<div class='kpis'>" + chips + "</div>"


def svg_grouped_bars(streams, title, w=980):
    """Two bars per row (Forecast vs Actual). streams = [(name, fc, act), ...]."""
    streams = [(n, f, a) for n, f, a in streams if (f or a)]
    n = len(streams)
    if not n:
        return ""
    top, grp = 60, 48
    h = top + n * grp + 6
    maxv = max([max(f, a) for _n, f, a in streams] + [1.0])
    labw, gap, valw = 200, 14, 116
    bx = labw + gap
    bw = w - bx - valw
    o = ["<svg viewBox='0 0 %d %d' width='100%%' style='max-width:%dpx;display:block' "
         "xmlns='http://www.w3.org/2000/svg' font-family='system-ui,Segoe UI,Arial,sans-serif'>" % (w, h, w)]
    o.append("<text x='0' y='20' fill='#fff' font-size='16' font-weight='700'>%s</text>" % html.escape(title))
    o.append("<rect x='0' y='32' width='12' height='12' rx='2' fill='%s'/><text x='18' y='42' fill='%s' font-size='11'>Forecast</text>" % (FCAST, INK))
    o.append("<rect x='94' y='32' width='12' height='12' rx='2' fill='%s'/><text x='112' y='42' fill='%s' font-size='11'>Actual</text>" % (ACT, INK))
    for i, (nm, f, a) in enumerate(streams):
        y = top + i * grp
        o.append("<text x='0' y='%d' fill='%s' font-size='12.5' font-weight='600'>%s</text>" % (y + 15, INK, html.escape(nm)))
        fw = max(2.0, bw * f / maxv); aw = max(2.0, bw * a / maxv)
        o.append("<rect x='%d' y='%d' width='%.1f' height='13' rx='4' fill='%s'/>" % (bx, y + 3, fw, FCAST))
        o.append("<text x='%.1f' y='%d' fill='%s' font-size='10.5'>%s</text>" % (bx + fw + 7, y + 13, MUT, money0(f)))
        o.append("<rect x='%d' y='%d' width='%.1f' height='13' rx='4' fill='%s'/>" % (bx, y + 20, aw, ACT))
        o.append("<text x='%.1f' y='%d' fill='%s' font-size='10.5'>%s</text>" % (bx + aw + 7, y + 30, MUT, money0(a)))
    o.append("</svg>")
    return "".join(o)


def daily_tracking_dashboard(wb):
    """Curated Forecast vs Actual dashboard for the Daily Tracking tab."""
    ws = wb["Daily Tracking"]

    def num(r, c):
        v = ws.cell(r, c).value
        return v if isinstance(v, (int, float)) else 0

    streams = [("Plant Equipment", 2, 3), ("Tooling", 5, 6), ("Radio", 8, 9),
               ("Gas Monitor", 11, 12), ("Personnel / Labour", 14, 15),
               ("Accommodation", 17, 18), ("Transport", 20, 21),
               ("Damage Recovery", 23, 24)]
    rows = [r for r in range(2, 60) if isinstance(ws.cell(r, 1).value, (datetime, date))]
    active = [r for r in rows if num(r, 27) > 0]  # billed days (Daily Actual Total > 0)
    if not active:
        return ""

    todate = [(nm, sum(num(r, fc) for r in active), sum(num(r, ac) for r in active))
              for nm, fc, ac in streams]
    fulljob = [(nm, sum(num(r, fc) for r in rows)) for nm, fc, _ac in streams]
    F = sum(t[1] for t in todate); A = sum(t[2] for t in todate); V = A - F
    FULL = sum(t[1] for t in fulljob)
    pct = (V / F * 100) if F else 0

    hero = ("<div class='kpis'>"
            "<div class='kpi'><div class='v'>%s</div><div class='l'>Forecast · to date</div></div>"
            "<div class='kpi'><div class='v' style='color:#8fbaf0'>%s</div><div class='l'>Actual · to date</div></div>"
            "<div class='kpi' style='border-left-color:%s'><div class='v'>%s</div>"
            "<div class='l'>Actual − Forecast (%s%.0f%%)</div></div>"
            "<div class='kpi' style='border-left-color:%s'><div class='v'>%s</div>"
            "<div class='l'>Forecast · to completion</div></div>"
            "</div>") % (money0(F), money0(A), AMBER, money0(V),
                         ("+" if V >= 0 else ""), pct, FCAST, money0(FULL))

    gb = svg_grouped_bars(todate, "Forecast vs Actual by cost stream — to date")
    cats = [ws.cell(r, 1).value.strftime("%d %b") for r in active]
    trend = svg_line_chart(cats,
                           [("Forecast", [num(r, 26) for r in active], FCAST, False),
                            ("Actual", [num(r, 27) for r in active], ACT, False)],
                           "Daily Forecast vs Actual ($) — billed days")
    comp = svg_bar_h(sorted([(n, v) for n, v in fulljob if v > 0],
                            key=lambda t: t[1], reverse=True),
                     "Forecast to completion by cost stream — full job", currency=True)

    out = hero
    out += ("<div class='viz'>" + gb +
            "<div class='vnote'>Summed over the active days (12–22 Jul). Forecast is the plan; Actual is "
            "per Daily Tracking — Plant and Tooling are genuine SiteIQ hire, while Labour, Accommodation "
            "and Radio are accrued to forecast pending billing. Transport is forecast but not yet billed.</div></div>")
    out += ("<div class='viz'>" + trend +
            "<div class='vnote'>The 20 Jul forecast spike is the big plant mobilisation; actual "
            "bills through as gear is picked up and processed.</div></div>")
    out += ("<div class='viz'>" + comp +
            "<div class='vnote'>What the full $" + "{:,.0f}".format(FULL) +
            " job forecast is made of. This is forecast to completion, not billed to date.</div></div>")
    return out


def main():
    # locate workbook
    here = os.path.dirname(os.path.abspath(__file__))
    if len(sys.argv) > 1 and os.path.isfile(sys.argv[1]):
        wbpath = sys.argv[1]
    else:
        cands = glob.glob(os.path.join(here, "Cement_Australia_Report*K2*.xlsm")) or \
                glob.glob(os.path.join(here, "*.xlsm"))
        if not cands:
            print("  ! No workbook (.xlsm) found next to this script.")
            return 1
        wbpath = max(cands, key=os.path.getmtime)

    wbname = os.path.basename(wbpath)
    outdir = os.path.join(os.path.dirname(wbpath), "K2_Workbook_Pages")
    os.makedirs(outdir, exist_ok=True)
    gen = datetime.now().strftime("%d %b %Y %H:%M")
    asat = datetime.fromtimestamp(os.path.getmtime(wbpath)).strftime("%d %b %Y %H:%M")

    print("  Reading (read-only): %s" % wbname)
    wb = openpyxl.load_workbook(wbpath, data_only=True)

    # build the page list in workbook order
    pages = []
    for ws in wb.worksheets:
        nm = ws.title
        pages.append({"name": nm, "slug": slug(nm), "group": group_of(nm), "ws": ws})

    # render each tab
    for p in pages:
        ws = p["ws"]
        sidebar = build_sidebar(pages, p["slug"])
        header = header_html(ws.title.strip(), asat)
        # visual layer: section badge + KPI tiles + a chart where the data suits it
        maxr, maxc = used_bounds(ws)
        hr = find_header(ws, maxr, maxc) if maxr else None
        if hr is not None:
            main_maxc = find_main_cols(ws, hr, maxc)
            block_end = find_block_end(ws, hr, main_maxc, maxr) if main_maxc else hr
            ndata = block_end - hr
            profs = profile_columns(ws, hr, block_end, main_maxc) if main_maxc else []
        else:
            main_maxc = 0; block_end = 0; ndata = 0; profs = []
        header_ok = hr is not None and ndata >= 3 and ws.title.strip() not in NO_VIZ
        viz = "<div style='margin-bottom:2px'><span class='badge'>" + html.escape(p["group"]) + "</span></div>"
        if ws.title.strip() == "Daily Tracking":
            # curated Forecast vs Actual dashboard
            viz += daily_tracking_dashboard(wb)
        else:
            if header_ok and profs:
                viz += kpi_tiles(pick_kpis(profs, ndata))
            if ws.title.strip() == "K2 Dashboard":
                svg = dashboard_chart_svg(wb)
                vnote = "Rebuilt from Dashboard Data on every run · the live chart stays in the .xlsm."
            elif header_ok:
                svg = pick_chart(profs, ws, hr, block_end) if profs else ""
                vnote = "Auto-built from this tab's data · read-only snapshot."
            else:
                svg = ""
            if svg:
                viz += "<div class='viz'>" + svg + "<div class='vnote'>" + vnote + "</div></div>"
        extra = ""
        table_html, nr, nc = render_table(ws)
        body = viz + "<div class='card'>" + table_html + "</div>"
        footer = footer_html(wbname, ws.title.strip(), gen)
        htmlpage = page_shell(ws.title.strip(), sidebar, header, body, footer, extra)
        with open(os.path.join(outdir, p["slug"] + ".html"), "w", encoding="utf-8") as f:
            f.write(htmlpage)
        p["rows"], p["cols"] = nr, nc

    # build the home hub
    sidebar = build_sidebar(pages, active_slug="__home__")
    header = header_html("Workbook — all tabs", asat)
    hub = ["<div class='hub'>"]
    hub.append("<div class='hero'><h2>K2 Workbook — every tab, on the web</h2>"
               "<p>All %d tabs of your Cement Australia K2 report, one page each, linked together. "
               "Read-only mirror of the last saved workbook — nothing here changes the file.</p></div>"
               % len(pages))

    # featured live reports (only if they exist next door in Output_Company_Reports)
    rep_dir = os.path.join(os.path.dirname(wbpath), "Output_Company_Reports")
    feats = [(lbl, fn) for (lbl, fn) in FEATURED if os.path.isfile(os.path.join(rep_dir, fn))]
    if feats:
        hub.append("<div class='grpttl'>Featured live reports</div><div class='feat'>")
        for lbl, fn in feats:
            hub.append("<a href='../Output_Company_Reports/%s'>▶ %s</a>" % (fn, html.escape(lbl)))
        hub.append("</div>")

    for gname, _m in GROUPS + [("Other", [])]:
        grp_pages = [p for p in pages if p["group"] == gname]
        if not grp_pages:
            continue
        hub.append("<div class='grpttl'>%s <span>%d tabs</span></div><div class='tiles'>" %
                   (html.escape(gname), len(grp_pages)))
        for p in grp_pages:
            dim = ("%d rows × %d cols" % (p["rows"], p["cols"])) if p.get("rows") else "no data"
            hub.append("<a class='tile' href='%s.html'><b>%s</b><div class='dim'>%s</div></a>" %
                       (p["slug"], html.escape(p["name"].strip()), dim))
        hub.append("</div>")
    hub.append("</div>")

    footer = footer_html(wbname, None, gen)
    with open(os.path.join(outdir, "index.html"), "w", encoding="utf-8") as f:
        f.write(page_shell("Home hub", sidebar, header, "".join(hub), footer))

    print("  Wrote %d tab pages + Home hub to: %s" % (len(pages), outdir))
    print("  Open: %s" % os.path.join(outdir, "index.html"))
    return 0


if __name__ == "__main__":
    sys.exit(main())
