#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | K2 ALIGN PREP - feeds ALIGN_EXCEL_DATA.ps1 without COM
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Reads the master equipment file + the RENTAL_STOCK export with
#  openpyxl (no Excel involved - corporate Excel dialogs, sensitivity
#  labels and wedged hidden instances can't touch it) and writes ONE
#  plain CSV beside this script for the align script to Import-Csv:
#
#      _align_master.csv:  barcode,item,new_desc,price,source
#
#  One row per barcode we can map to a master item (plus item-number
#  rows for barcodes that ARE item numbers). Deleted and rebuilt every
#  run. 13_ALIGN_EXCEL_DATA.bat runs this first automatically.
# =====================================================================

import csv
import glob
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "_align_master.csv")


def find_newest(pattern):
    hits = []
    for d in (os.path.join(HERE, "Data_SiteIQ"), HERE):
        hits += [p for p in glob.glob(os.path.join(d, pattern))
                 if not os.path.basename(p).startswith("~$")]
    return max(hits, key=os.path.getmtime) if hits else None


def main():
    try:
        import openpyxl
    except ImportError:
        print("  ! openpyxl missing - run: py -m pip install openpyxl")
        sys.exit(1)
    try:
        if os.path.isfile(OUT):
            os.remove(OUT)
    except Exception:
        pass

    mpath = find_newest("K2_MASTER_EQUIPMENT_PRICING*.xlsx")
    if not mpath:
        print("  Align prep: master file not found - price/name true-up "
              "will be skipped (everything else still aligns).")
        sys.exit(0)
    master = {}
    wb = openpyxl.load_workbook(mpath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c or "").strip().upper() for c in next(rows)]

    def col(n):
        return hdr.index(n) if n in hdr else None

    ci, cn, cp, cs = (col("ITEM_NUMBER"), col("NEW_DESCRIPTION"),
                      col("REPLACEMENT_COST_AUD"),
                      col("REPLACEMENT_PRICE_SOURCE"))
    for r in rows:
        item = str(r[ci] or "").strip() if ci is not None else ""
        if not item:
            continue
        price = r[cp] if (cp is not None and cp < len(r)) else None
        master[item] = {
            "new": str(r[cn] or "").strip() if cn is not None else "",
            "price": ("" if price in (None, "") else str(price)),
            "src": str(r[cs] or "").strip() if cs is not None else "",
        }
    wb.close()

    bc2item = {}
    spath = find_newest("RENTAL_STOCK*.xlsx")
    if spath:
        wb = openpyxl.load_workbook(spath, read_only=True, data_only=True)
        ws = (wb["RENTAL_STOCK"] if "RENTAL_STOCK" in wb.sheetnames
              else wb[wb.sheetnames[0]])
        rows = ws.iter_rows(values_only=True)
        hdr = [str(c or "").strip().upper() for c in next(rows)]
        if "ITEM_BARCODE" in hdr and "ITEM_NUMBER" in hdr:
            bi, ii = hdr.index("ITEM_BARCODE"), hdr.index("ITEM_NUMBER")
            for r in rows:
                bc = str(r[bi] or "").strip()
                it = str(r[ii] or "").strip()
                if bc and it and bc.upper() not in bc2item:
                    bc2item[bc.upper()] = it
        wb.close()

    n = 0
    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["barcode", "item", "new_desc", "price", "source"])
        # every mapped barcode
        for bc, item in bc2item.items():
            rec = master.get(item)
            if rec:
                w.writerow([bc, item, rec["new"], rec["price"], rec["src"]])
                n += 1
        # item numbers as their own keys (some Excel-tab barcodes ARE items)
        for item, rec in master.items():
            if item.upper() not in bc2item:
                w.writerow([item.upper(), item, rec["new"], rec["price"],
                            rec["src"]])
                n += 1
    print("  Align prep: {} master rows staged for the Excel true-up "
          "({} barcodes mapped).".format(n, len(bc2item)))


if __name__ == "__main__":
    main()
