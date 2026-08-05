# ==========================================================================
#  COATES | K2 PLANT ON SITE DASHBOARD
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Builds a single self-contained HTML dashboard from the SiteIQ exports in
#  this folder:
#    - Plant On Hire  (site plant out with a crew) by category, with Plant ID
#    - Site Plant Idle (parked in the Site Plant Equipment pool, still charged)
#    - Rubbish Chutes & Barriers by type (quantity only)
#
#  Reads (auto-found in this folder, newest wins):
#    RENTAL_STOCK*.xlsx            - the live on-hire picture (required)
#    Cement_Australia_Report*K2*.xlsm - Plant Equipment register, for categories
#    Plant_ID_Register.csv         - your Asset Number -> Plant ID list (editable)
#
#  Output: Reports\<YYYY-MM-DD>\Coates_K2_Plant_Dashboard_<date>.html
#  (one folder per day - see report_paths.py). Open it in any browser or
#  email it - no internet needed.
#
#  25 Jul 2026 - COMPLIANCE UNDER EVERY PLANT LINE (A. Fisher)
#  Every plant line now carries its own compliance requirement under the
#  item name - current tag colour, daily pre-start and logbook, back
#  daily - read straight off K2_MASTER_EQUIPMENT_PRICING.xlsx through
#  equipment_compliance.py. The reason is simple: a supervisor looking at
#  this dashboard should be able to see what has to be checked before the
#  gear is used without ringing the store, and the Coates team can check
#  the gear over knowing what it's meant to have. The search box reads the
#  same words, so typing "logbook" or "blue" finds the gear it applies to.
#  A compliance tally sits with the site totals so the exposure is one
#  number, not a hunt through the tabs. Nothing is guessed here - no flag
#  in the master means nothing shows on the line.
#
#  Usage: double-click 07_RUN_PLANT_DASHBOARD.bat
# ==========================================================================
import openpyxl, csv, os, sys, re, glob, datetime, json, html as _html

import report_paths
import master_equipment
import equipment_compliance as EC
# Master display names (item-number keyed). Classification (cat_for)
# stays on the export's ORIGINAL description so categories never move
# because of a rename; only the shown name improves.
MASTER = master_equipment.load(os.path.dirname(os.path.abspath(__file__)),
                               quiet=True)
# Point the compliance engine at the same master - one file drives the
# names, the costs and the compliance flags, so a fix in the spreadsheet
# fixes every report on the next run. Safe with no master: no badges.
EC.bind(MASTER)

HERE = os.path.dirname(os.path.abspath(__file__))

def norm(v): return '' if v is None else str(v).strip()

def find_newest(pattern, what, required=True):
    hits = []
    for d in (os.path.join(HERE, 'Data_SiteIQ'), HERE):
        hits += [f for f in glob.glob(os.path.join(d, pattern))
                 if not os.path.basename(f).startswith('~$')]
    if not hits:
        if required:
            print("\nPROBLEM: couldn't find %s." % what)
            print("  Looked for: %s" % pattern)
            print("  In: %s" % HERE)
            print("Drop that export into this folder and run again.\n")
            sys.exit(1)
        return None
    return max(hits, key=os.path.getmtime)

# ---- 1. locate inputs ----------------------------------------------------
stock_path = find_newest('RENTAL_STOCK*.xlsx', 'the RENTAL_STOCK export')
wb_path    = find_newest('Cement_Australia_Report*K2*.xlsm', 'the K2 workbook (for categories)', required=False)
pid_path   = os.path.join(HERE, 'Plant_ID_Register.csv')

# ---- 2. plant categories from the register (variant -> category) ---------
var2cat = {}
if wb_path:
    try:
        wb = openpyxl.load_workbook(wb_path, data_only=True)
        reg = wb['Plant Equipment']
        for tn in reg.tables:
            if tn == 'AssetRegisterTable2219':
                cells = reg[reg.tables[tn].ref]
                hdr = [norm(c.value) for c in cells[0]]; ci = {h:i for i,h in enumerate(hdr)}
                for row in cells[1:]:
                    v = norm(row[ci['Product Variant']].value).upper()
                    cat = norm(row[ci['Category']].value)
                    if v and cat and v not in var2cat: var2cat[v] = cat
    except Exception as e:
        print("NOTE: couldn't read categories from the workbook (%s) - using descriptions instead." % e)

# Meet the team - printed on every page of the dashboard, the same crew
# strip the report packs carry. (A. Fisher, 25 Jul 2026)
TEAM_STRIP = (
    "<div class='tl'>Your Coates tool store team &middot; "
    "<span style='color:#8B9099;letter-spacing:1px'>MY GEAR HQ</span></div>"
    "<div class='tn'><b>Andrew Fisher</b> Shutdown Manager QLD &amp; NT"
    " &nbsp;|&nbsp; <span style='color:#F26222;font-weight:700'>DAY</span>"
    " <b>Thomas Maher</b> Dayshift Lead &middot; <b>Helen O'Grady</b>"
    " Dayshift Tool Store Controller &middot; <b>Lucas Armstrong</b>"
    " Day Shift Mechanic &nbsp;|&nbsp;"
    " <span style='color:#F26222;font-weight:700'>NIGHT</span>"
    " <b>Cody Mitchell</b> Nightshift Tool Store Controller &middot;"
    " <b>Adam Boyce</b> Night Shift Tool Store Controller</div>")


def cat_for(variant, family, product, desc):
    v = variant.upper()
    if v and v in var2cat: return var2cat[v]
    if v:  # only prefix-match a REAL (non-empty) variant
        for k in var2cat:
            if len(k) >= 8 and (v.startswith(k) or k.startswith(v)): return var2cat[k]
    f = (desc + ' ' + family + ' ' + product).upper()
    if 'WELD' in f: return 'Welders'                 # welders first (names contain "AIR")
    if 'SCISSOR' in f: return 'Scissor Lifts'
    if 'BOOM' in f or 'KNUCKLE' in f: return 'Booms'
    if 'FORK' in f or 'TELEHANDLER' in f or 'PALLET' in f or 'MATERIALS HANDLING' in f: return 'Forklifts & Pallet Trucks'
    if 'GENERATOR' in f: return 'Generators'
    if 'LIGHT' in f: return 'Lighting'
    if 'COMPRESSOR' in f or 'HOSE' in f: return 'Compressors & Hoses'
    if 'LEAD' in f or 'DISTRIBUTION' in f or 'BOARD' in f or 'HVAC' in f or 'CABLE' in f or 'ELECTR' in f: return 'Distribution Boards & Leads'
    if 'EXCAVATOR' in f or 'SKID STEER' in f or 'BUCKET' in f: return 'Earthmoving'
    return 'Miscellaneous'

# ---- 2b. hire rates: contracted first, then register (prefix-tolerant) ---
def _num(x):
    try: return float(re.sub(r'[$,]', '', str(x)))
    except Exception: return None
rates = {}; regrate = {}
rates_path = find_newest('ImportSample*Contracted*Rates*and*Prices*.xlsx', 'the Contracted Rates file', required=False)
if rates_path:
    try:
        wrt = openpyxl.load_workbook(rates_path, data_only=True, read_only=True)
        for row in wrt['Coates Equipment'].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                v = norm(row[0]).upper(); rt = _num(row[1]) if len(row) > 1 else None
                if v and rt and rt > 0 and v not in rates: rates[v] = rt
    except Exception as e:
        print("NOTE: couldn't read the Contracted Rates file (%s) - costs may be blank." % e)
if wb_path:
    try:
        for tn in reg.tables:
            if tn == 'AssetRegisterTable2219':
                cells = reg[reg.tables[tn].ref]; hdr = [norm(c.value) for c in cells[0]]; ci = {h:i for i,h in enumerate(hdr)}
                for row in cells[1:]:
                    v = norm(row[ci['Product Variant']].value).upper(); rt = _num(row[ci['Daily Rate']].value)
                    if v and rt and rt > 0 and v not in regrate: regrate[v] = rt
    except Exception:
        pass

def rate_for(variant, product=''):
    for src in (rates, regrate):
        for key in (variant.upper(), product.upper()):
            if key and key in src: return src[key]
    for src in (regrate, rates):
        for key in (variant.upper(), product.upper()):
            if not key: continue
            for k in src:
                if len(k) >= 8 and (key.startswith(k) or k.startswith(key)): return src[k]
    return None

def parse_dt(s):
    s = norm(s).split()[0] if norm(s) else ''
    for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
        try: return datetime.datetime.strptime(s, fmt).date()
        except Exception: pass
    return None

# ---- 3. Plant ID register (Asset Number -> Plant ID) ---------------------
pidmap = {}
if os.path.exists(pid_path):
    with open(pid_path, newline='') as f:
        for row in csv.DictReader(f):
            a = norm(row.get('Asset Number')); i = norm(row.get('Plant ID'))
            if a and i: pidmap[a] = i
else:
    print("NOTE: Plant_ID_Register.csv not found - Plant ID column will show '-'.")

def pid_for(asset):
    #  The MASTER file is the suite standard for Plant IDs (same source
    #  as the Site Plant report's pills) - the old CSV register only knew
    #  a handful. Master first, CSV as the fallback. (Andrew, 26 Jul
    #  2026: "make sure plant IDs are on ALL plant".)
    mid = EC.plant_id(asset)
    if mid: return mid
    if asset in pidmap: return pidmap[asset]
    for k in pidmap:                       # SKU may be a prefix of the full barcode
        if k.startswith(asset + '-') or asset.startswith(k): return pidmap[k]
    return ''

# ---- 4. read RENTAL_STOCK, classify and cost -----------------------------
from collections import defaultdict
rs = openpyxl.load_workbook(stock_path, data_only=True, read_only=True)
ws = rs['RENTAL_STOCK']; rows = list(ws.iter_rows(values_only=True))
H = {norm(h): i for i, h in enumerate(rows[0])}
def g(r, k): return norm(r[H[k]]) if k in H else ''

# as-at reference (the export's pull time) drives days-on-hire
asat = ''
try:
    ref = list(rs['REFERENCE_INFO'].iter_rows(values_only=True)); asat = norm(ref[1][2])
except Exception:
    pass
if not asat:
    asat = datetime.datetime.fromtimestamp(os.path.getmtime(stock_path)).strftime('%d/%m/%Y %H:%M')
asat_date = parse_dt(asat) or datetime.date.today()

def days_on(onh):
    d = parse_dt(onh)
    return ((asat_date - d).days + 1) if (d and asat_date >= d) else None

onhire = defaultdict(list); idle = defaultdict(list)
chutes = defaultdict(lambda: {'qty': 0, 'daily': 0.0, 'todate': 0.0})
barriers = defaultdict(lambda: {'qty': 0, 'daily': 0.0, 'todate': 0.0})
for r in rows[1:]:
    su = g(r, 'STORAGE_UNIT').upper()
    var = g(r, 'PRODUCT_VARIANT'); prod = g(r, 'PRODUCT'); desc = g(r, 'ITEM_DESCRIPTION')
    if su == 'CEMENT SITE PLANT':
        if g(r, 'ITEM_STATUS').upper() == 'AVAILABLE FOR HIRE': continue   # off charge - excluded
        sku = g(r, 'SKU_NUMBER')                      # internal key for Plant ID lookup
        cat = cat_for(var, g(r, 'PRODUCT_FAMILY'), prod, desc)
        rt = rate_for(var, prod); dys = days_on(g(r, 'ON_HIRE_DATE'))
        # 'cb' = the compliance chips that sit UNDER the item name (already
        # escaped HTML from equipment_compliance - never re-escape it in the
        # page). 'ct' is the same requirement in plain words so the search
        # box finds it: type "logbook" or "blue" and you get the gear it
        # applies to. Both are "" when the master says this item has nothing
        # to declare, so a clean line stays clean.
        rec = {'asset': g(r, 'ITEM_NUMBER') or sku or '-',
               'desc': MASTER.disp(g(r, 'ITEM_NUMBER'), desc),
               'cb': EC.badges_html(g(r, 'ITEM_NUMBER'), desc, wrap=False),
               'ct': EC.badges_text(g(r, 'ITEM_NUMBER'), desc),
               'onhire': g(r, 'ON_HIRE_DATE'), 'pid': pid_for(sku),
               'rate': round(rt, 2) if rt else None, 'days': dys,
               'todate': round(rt * dys, 2) if (rt and dys) else None}
        hi = g(r, 'HIRER_NAME'); co = g(r, 'COMPANY_NAME')
        if 'SITE PLANT EQUIPMENT' in hi.upper():
            idle[cat].append(rec)
        elif co or hi:
            rec['hirer'] = hi; rec['company'] = co; onhire[cat].append(rec)
    elif su in ('CEMENT RUBBISH CHUTES', 'CEMENT BARRIERS'):
        bucket = chutes if su == 'CEMENT RUBBISH CHUTES' else barriers
        rt = rate_for(var, prod); dys = days_on(g(r, 'ON_HIRE_DATE'))
        b = bucket[MASTER.disp(g(r, 'ITEM_NUMBER'), desc)]; b['qty'] += 1
        if rt: b['daily'] += rt
        if rt and dys: b['todate'] += rt * dys

def order(d): return {k: sorted(v, key=lambda x: (x['desc'], x['asset'])) for k, v in sorted(d.items(), key=lambda kv: -len(kv[1]))}
def qty2(d): return {k: {'qty': v['qty'], 'daily': round(v['daily'], 2), 'todate': round(v['todate'], 2)}
                     for k, v in sorted(d.items(), key=lambda kv: -kv[1]['qty'])}
DATA = {'asat': asat, 'onhire': order(onhire), 'idle': order(idle),
        'chutes': qty2(chutes), 'barriers': qty2(barriers)}

def _tot(sec):
    daily = todate = 0.0; unp = 0
    for cat in DATA[sec]:
        for it in DATA[sec][cat]:
            if it['rate']: daily += it['rate']
            else: unp += 1
            if it['todate']: todate += it['todate']
    return round(daily, 2), round(todate, 2), unp
oh_d, oh_t, oh_u = _tot('onhire'); id_d, id_t, id_u = _tot('idle')
ch_d = round(sum(v['daily'] for v in DATA['chutes'].values()), 2); ch_t = round(sum(v['todate'] for v in DATA['chutes'].values()), 2)
ba_d = round(sum(v['daily'] for v in DATA['barriers'].values()), 2); ba_t = round(sum(v['todate'] for v in DATA['barriers'].values()), 2)
DATA['tot'] = {'oh_daily': oh_d, 'oh_todate': oh_t, 'oh_unp': oh_u,
               'id_daily': id_d, 'id_todate': id_t, 'id_unp': id_u,
               'ch_daily': ch_d, 'ch_todate': ch_t, 'ba_daily': ba_d, 'ba_todate': ba_t,
               'all_daily': round(oh_d + id_d + ch_d + ba_d, 2),
               'all_todate': round(oh_t + id_t + ch_t + ba_t, 2)}

n_onhire = sum(len(v) for v in DATA['onhire'].values())
n_idle   = sum(len(v) for v in DATA['idle'].values())
n_chute  = sum(v['qty'] for v in DATA['chutes'].values()); n_barr = sum(v['qty'] for v in DATA['barriers'].values())
n_plant  = n_onhire + n_idle

# ---- overview / showcase metrics -----------------------------------------
plant_daily = round(oh_d + id_d, 2)
cat_daily = {}
for _sec in ('onhire', 'idle'):
    for _cat, _items in DATA[_sec].items():
        cat_daily[_cat] = round(cat_daily.get(_cat, 0) + sum(it['rate'] for it in _items if it['rate']), 2)
cat_daily = dict(sorted(cat_daily.items(), key=lambda kv: -kv[1]))
_idle_all = [dict(it, cat=_cat) for _cat, _items in DATA['idle'].items() for it in _items if it['rate']]
top_idle  = sorted(_idle_all, key=lambda x: (-(x['rate'] or 0), -(x['todate'] or 0)))[:8]
long_idle = sorted([x for x in _idle_all if x['days']], key=lambda x: -x['days'])[:6]
DATA['ov'] = {'deployed': n_onhire, 'parked': n_idle, 'total': n_plant,
              'util': round(n_onhire / n_plant * 100, 1) if n_plant else 0,
              'oh_daily': oh_d, 'id_daily': id_d, 'plant_daily': plant_daily,
              'idle_share': round(id_d / plant_daily * 100) if plant_daily else 0,
              'cat_daily': cat_daily, 'top_idle': top_idle, 'long_idle': long_idle}

# ---- compliance exposure -------------------------------------------------
# The same flags that print under each line, counted up. Only plant that is
# actually OUT WITH A CREW is counted - gear parked in the Coates idle pool
# isn't in anyone's hands, so counting it here would overstate what the
# supervisors have to action today. The tag colour name comes from the one
# place it lives (equipment_compliance.TAG_PERIODS), so when the inspection
# colour rolls over this label follows on the next run.
_oh_all = [it for _items in DATA['onhire'].values() for it in _items]
DATA['comp'] = EC.summarise(_oh_all, item_key='asset', desc_key='desc')
DATA['comp']['tagname'] = (EC.tag_colour()[0] or 'Tag check').title()

# ---- 5. render the dashboard --------------------------------------------
HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>K2 Plant On Site Dashboard — Cement Australia</title>
<style>
:root{--org:#F26222;--org2:#C44C28;--bg:#0E1116;--card:#171B22;--card2:#1E232C;--line:#2A313C;
--tx:#EAEDF1;--mut:#95A0AF;--grn:#35B37E;--amb:#E8A32C;--red:#E5544B;--chip:#232935;}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--tx);font-family:"Segoe UI",Roboto,Helvetica,Arial,sans-serif;
font-size:14px;line-height:1.45;-webkit-font-smoothing:antialiased}
.wrap{max-width:1180px;margin:0 auto;padding:0 18px 60px}
header{border-top:5px solid var(--org);background:linear-gradient(180deg,#141922,#0E1116);padding:22px 0 18px;margin-bottom:20px}
.hd{max-width:1180px;margin:0 auto;padding:0 18px;display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:12px}
.brand .logo{font-weight:800;font-size:23px;letter-spacing:.5px}
.brand .logo b{color:var(--org)}
.brand .tag{color:var(--mut);font-size:12px;letter-spacing:2px;text-transform:uppercase;margin-top:2px}
.brand h1{font-size:19px;margin-top:12px;font-weight:700}
.brand .sub{color:var(--mut);font-size:13px;margin-top:3px}
.meta{text-align:right;font-size:12px;color:var(--mut)}
.meta .siteiq{color:var(--org);font-weight:700;letter-spacing:1px;font-size:11px}
.meta .who{color:var(--tx);font-weight:600}
.kpis{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-bottom:22px}
@media(max-width:820px){.kpis{grid-template-columns:repeat(2,1fr)}}
.kpi{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:15px 16px;position:relative;overflow:hidden}
.kpi:before{content:"";position:absolute;left:0;top:0;bottom:0;width:4px;background:var(--org)}
.kpi.amber:before{background:var(--amb)}
.kpi .n{font-size:30px;font-weight:800;letter-spacing:-.5px}
.kpi .l{color:var(--mut);font-size:11.5px;text-transform:uppercase;letter-spacing:.6px;margin-top:3px}
.kpi .h{color:var(--mut);font-size:11px;margin-top:6px}
.tabs{display:flex;gap:4px;border-bottom:1px solid var(--line);margin-bottom:18px;flex-wrap:wrap}
.tab{background:none;border:none;color:var(--mut);font-size:14px;font-weight:600;padding:11px 16px;cursor:pointer;border-bottom:3px solid transparent;font-family:inherit}
.tab:hover{color:var(--tx)}
.tab.on{color:var(--tx);border-bottom-color:var(--org)}
.tab .b{background:var(--chip);color:var(--mut);border-radius:20px;padding:1px 8px;font-size:11px;margin-left:6px;font-weight:700}
.tab.on .b{background:var(--org);color:#fff}
.search{width:100%;background:var(--card);border:1px solid var(--line);border-radius:9px;color:var(--tx);
padding:10px 13px;font-size:13.5px;font-family:inherit;margin-bottom:16px}
.search:focus{outline:none;border-color:var(--org)}
.note{background:rgba(232,163,44,.09);border:1px solid rgba(232,163,44,.35);border-radius:9px;padding:10px 14px;
font-size:12.5px;color:#F0D9A8;margin-bottom:16px;display:flex;gap:9px;align-items:flex-start}
.note b{color:var(--amb)}
.cat{margin-bottom:14px;border:1px solid var(--line);border-radius:11px;overflow:hidden;background:var(--card)}
.cat>summary{list-style:none;cursor:pointer;padding:12px 15px;display:flex;justify-content:space-between;align-items:center;
background:var(--card2);font-weight:700;font-size:14.5px}
.cat>summary::-webkit-details-marker{display:none}
.cat>summary .cn{display:flex;align-items:center;gap:10px}
.cat>summary .dot{width:9px;height:9px;border-radius:50%;background:var(--org)}
.cat>summary .cc{background:var(--chip);color:var(--mut);border-radius:20px;padding:2px 11px;font-size:12px;font-weight:700}
table{width:100%;border-collapse:separate;border-spacing:0 6px}
th{text-align:left;color:var(--mut);font-size:11px;text-transform:uppercase;letter-spacing:.5px;padding:4px 15px 2px;font-weight:600}
td{padding:11px 15px;font-size:13.5px;background:#191E27;border-top:1px solid var(--line);border-bottom:1px solid var(--line);vertical-align:middle}
.row td:first-child{border-left:1px solid var(--line);border-top-left-radius:9px;border-bottom-left-radius:9px}
.row td:last-child{border-right:1px solid var(--line);border-top-right-radius:9px;border-bottom-right-radius:9px}
.row:hover td{background:rgba(242,98,34,.07);border-color:rgba(242,98,34,.45)}
.asset{font-family:"Consolas",ui-monospace,monospace;background:var(--chip);color:var(--org);border-radius:6px;
padding:2px 8px;font-size:12.5px;font-weight:600;white-space:nowrap}
.pid{display:inline-block;min-width:34px;text-align:center;background:var(--org);color:#fff;border-radius:6px;
padding:2px 7px;font-weight:800;font-size:13.5px}
.pidnone{color:var(--mut);font-weight:400}
/* .cb = the compliance chips under an item name. The chips themselves are
   inline-styled so they survive print, PDF and an email body - this class
   only spaces them off the description. */
.cb{margin-top:3px;line-height:1.6}
.mut{color:var(--mut)}
.qtygrid{display:grid;grid-template-columns:1fr 1fr;gap:16px}
@media(max-width:720px){.qtygrid{grid-template-columns:1fr}}
.panel{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:18px}
.panel h3{font-size:15px;margin-bottom:3px}
.panel .pt{color:var(--mut);font-size:12px;margin-bottom:15px}
.qrow{display:flex;align-items:center;gap:12px;margin-bottom:13px}
.qrow .ql{width:220px;font-size:13.5px}
.qrow .qbar{flex:1;height:12px;background:var(--chip);border-radius:7px;overflow:hidden}
.qrow .qfill{height:100%;background:linear-gradient(90deg,var(--org2),var(--org));border-radius:7px}
.qrow .qn{width:40px;text-align:right;font-weight:800;font-size:16px}
.qrow .qm{width:92px;text-align:right;color:var(--org);font-weight:700;font-size:13px}
.ptot{border-top:1px solid var(--line);margin-top:6px;padding-top:12px;display:flex;align-items:center;gap:10px;font-weight:700}
.ptot .sep2{flex:1}
.ptot .qn{color:var(--org);font-size:18px;width:40px;text-align:right}
.ptot .qm{width:92px;text-align:right;color:var(--org);font-size:15px}
.ptot2{margin-top:8px;color:var(--mut);font-size:12.5px;text-align:right}
.ptot2 b{color:var(--tx)}
td.tocost{text-align:right;color:var(--org);font-weight:600}
.csum{display:flex;align-items:center;gap:12px}
.cmoney{color:var(--org);font-weight:700;font-size:13px}
.tally{display:flex;gap:26px;flex-wrap:wrap;align-items:center;background:var(--card);border:1px solid var(--line);border-radius:11px;padding:13px 18px;margin-bottom:16px}
.tally .ti{display:flex;flex-direction:column}
.tally .tl{color:var(--mut);font-size:10.5px;text-transform:uppercase;letter-spacing:.5px}
.tally .tv{font-size:23px;font-weight:800;margin-top:1px}
.tally .tv.org{color:var(--org)} .tally .tv.amb{color:var(--amb)}
.tally .sep{flex:1;min-width:8px}
.tally .note2{color:var(--mut);font-size:11px;max-width:270px;line-height:1.4}
.ovhero{display:grid;grid-template-columns:1fr 1.25fr;gap:16px;margin-bottom:16px}
@media(max-width:820px){.ovhero{grid-template-columns:1fr}}
.ovcard{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:18px}
.ovcard.wide{margin-bottom:16px}
.ovh{font-size:14.5px;font-weight:700;margin-bottom:14px}
.ovh .ovmut{color:var(--mut);font-weight:400;font-size:12px}
.ovdonut{display:flex;align-items:center;gap:22px}
.ovleg{font-size:13.5px;line-height:2}
.ovleg .ovmut{color:var(--mut);font-size:12px;margin-top:3px}
.dotc{display:inline-block;width:9px;height:9px;border-radius:50%;margin-right:8px;vertical-align:middle}
.ovbig{font-size:40px;font-weight:800;color:var(--amb);line-height:1}
.ovbigsub{font-size:15px;font-weight:600;color:var(--mut)}
.ovsub{margin:8px 0 16px;color:var(--tx);font-size:13.5px}
.ovsub .amb{color:var(--amb)}
.splitbar{display:flex;height:30px;border-radius:8px;overflow:hidden;background:var(--chip)}
.splitbar .seg{display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:700;color:#fff;white-space:nowrap}
.splitbar .seg.org{background:var(--org)}
.splitbar .seg.amb{background:var(--amb);color:#3a2b06}
.splitleg{display:flex;gap:22px;margin-top:11px;font-size:12.5px;color:var(--mut)}
.ovtbl{width:100%;border-collapse:separate;border-spacing:0 6px}
.ovtbl th{text-align:left;color:var(--mut);font-size:11px;text-transform:uppercase;letter-spacing:.5px;padding:2px 12px 4px;font-weight:600}
.ovtbl td{background:#191E27;border-top:1px solid var(--line);border-bottom:1px solid var(--line);padding:10px 12px;font-size:13px}
.ovtbl td:first-child{border-left:1px solid var(--line);border-radius:8px 0 0 8px}
.ovtbl td:last-child{border-right:1px solid var(--line);border-radius:0 8px 8px 0}
.ovtbl tr:hover td{background:rgba(242,98,34,.06)}
footer{margin-top:34px;padding-top:16px;border-top:1px solid var(--line);color:var(--mut);font-size:11.5px;line-height:1.7}
footer b{color:var(--tx)}
.hide{display:none}
.empty{padding:24px 15px;color:var(--mut);text-align:center;font-size:13px}
@media print{body{background:#fff;color:#000}.tabs,.search{display:none}.tabview{display:block!important}
.cat,.kpi,.panel{border-color:#ccc}.tab{display:none}}
</style></head>
<body>
<header><div class="hd">
<div class="brand">
<div class="logo"><b>COATES</b> &nbsp;<span style="color:var(--mut);font-weight:400;font-size:13px">Equipped for anything</span></div>
<div class="tag">Cement Australia · K2 Shutdown 2026 · Gladstone</div>
<h1>Plant On Site Dashboard</h1>
<div class="sub">Site plant, barriers &amp; rubbish chutes — on hire, idle and quantities</div>
</div>
<div class="meta">
<div class="siteiq">POWERED BY SITEIQ</div>
<div style="margin-top:8px">As at<br><span class="who">__ASAT__</span></div>
<div style="margin-top:8px">Author<br><span class="who">Andrew Fisher</span></div>
</div>
</div></header>
<div class="wrap">

<div class="kpis">
<div class="kpi"><div class="n">__NPLANT__</div><div class="l">Site Plant On Site</div><div class="h">on charge</div></div>
<div class="kpi"><div class="n">__NONHIRE__</div><div class="l">On Hire to Crews</div><div class="h">in use</div></div>
<div class="kpi amber"><div class="n">__NIDLE__</div><div class="l">Idle Pool</div><div class="h">parked · still charged</div></div>
<div class="kpi"><div class="n">__NCHUTE__</div><div class="l">Rubbish Chutes</div><div class="h">on site</div></div>
<div class="kpi"><div class="n">__NBARR__</div><div class="l">Barriers</div><div class="h">on site</div></div>
</div>

<div id="site-tally"></div>
<div id="comp-tally"></div>

<div class="tabs">
<button class="tab on" data-t="ov">Overview</button>
<button class="tab" data-t="onhire">Plant On Hire <span class="b">__NONHIRE__</span></button>
<button class="tab" data-t="idle">Site Plant Idle <span class="b">__NIDLE__</span></button>
<button class="tab" data-t="cb">Chutes &amp; Barriers</button>
</div>

<input class="search" id="q" placeholder="Search asset number or item…">

<div class="tabview" id="v-ov"><div id="ov-body"></div></div>

<div class="tabview hide" id="v-onhire">
<div id="onhire-tally"></div>
<div class="pt mut" style="margin-bottom:14px;color:var(--mut);font-size:12.5px">Site plant out with a crew, grouped by category. Daily $ is the contracted hire rate; To Date $ is that rate × days on hire. Category headers show the daily run-rate.</div>
<div id="onhire-body"></div>
</div>

<div class="tabview hide" id="v-idle">
<div class="note"><b>⚠ Cost-saving watch.</b><div>These units are parked in the Site Plant Equipment (Coates) pool — available but <b>still on charge</b>. The daily figure is what idle plant costs per day; the to-date figure is the spend so far — that's the saving on the table.</div></div>
<div id="idle-tally"></div>
<div id="idle-body"></div>
</div>

<div class="tabview hide" id="v-cb">
<div class="qtygrid">
<div class="panel"><h3>Rubbish Chutes</h3><div class="pt">By type · quantity on site</div><div id="chutes"></div></div>
<div class="panel"><h3>Barriers</h3><div class="pt">By type · quantity on site</div><div id="barriers"></div></div>
</div>
</div>

<footer>
<b>Data source:</b> SiteIQ RENTAL_STOCK export · <b>Extract:</b> __ASAT__ · Snapshot in time — refresh the export for a live picture.<br>
<b>Definitions:</b> On Hire = out with a named crew/company. Idle Pool = on hire to "Site Plant Equipment" (Coates), available but still charged. "Available for Hire" plant (off charge) is excluded.<br>
<b>Plant ID</b> is your site register number, matched to each unit by asset number. "—" = not on the register.<br>
<b>Compliance chips</b> under an item name come from the K2 master equipment file — a current inspection tag, a daily pre-start and logbook entry, or a same-day return. No chip means the master has nothing recorded against that item; it is never guessed here.<br>
<b>Costs:</b> Daily $ = contracted daily hire rate (Coates Equipment card, register fallback). To Date $ = daily rate × days on hire (inclusive of the on-hire day). Items with no rate on file show "—" and are excluded from every total — never estimated.<br>
Care Deeply · Customer Focused · Be Our Best · One Team · Competitive Spirit &nbsp;|&nbsp; Author: Andrew Fisher · POWERED BY SITEIQ
</footer>
</div>

<script>
var DATA=__DATA__;
function esc(s){return (s||'').replace(/[&<>]/g,function(c){return{'&':'&amp;','<':'&lt;','>':'&gt;'}[c]})}
function money(n){ if(n==null) return '<span class="pidnone">—</span>'; return '$'+Number(n).toLocaleString('en-AU',{minimumFractionDigits:2,maximumFractionDigits:2}); }
function money0(n){ return '$'+Math.round(Number(n)||0).toLocaleString('en-AU'); }
function renderPlant(host, groups, withHirer){
  var el=document.getElementById(host), html='';
  Object.keys(groups).forEach(function(cat){
    var items=groups[cat], cd=0;
    items.forEach(function(it){ if(it.rate) cd+=it.rate; });
    html+='<details class="cat" open><summary><span class="cn"><span class="dot"></span>'+esc(cat)+'</span>'
        +'<span class="csum"><span class="cmoney">'+money0(cd)+'/day</span><span class="cc">'+items.length+'</span></span></summary>';
    html+='<table><thead><tr><th style="width:66px">Plant ID</th><th style="width:126px">Asset No.</th><th>Item</th>'
        +(withHirer?'<th style="width:180px">Hirer / Company</th>':'')
        +'<th style="width:96px">On Hire</th><th style="width:90px;text-align:right">Daily $</th><th style="width:104px;text-align:right">To Date $</th></tr></thead><tbody>';
    items.forEach(function(it){
      html+='<tr class="row" data-s="'+esc(('plant '+(it.pid||'')+' '+it.asset+' '+it.desc+' '+(it.ct||'')+' '+(it.hirer||'')+' '+(it.company||'')).toLowerCase())+'">';
      html+='<td>'+(it.pid!==''&&it.pid!=null?'<span class="pid">'+esc(String(it.pid))+'</span>':'<span class="pidnone">—</span>')+'</td>';
      // it.cb is already-escaped chip HTML from equipment_compliance - it
      // must NOT go through esc() or the badges print as raw tags.
      html+='<td><span class="asset">'+esc(it.asset)+'</span></td><td>'+esc(it.desc)+(it.cb?'<div class="cb">'+it.cb+'</div>':'')+'</td>';
      if(withHirer) html+='<td>'+esc(it.hirer||'—')+(it.company?'<br><span class="mut" style="font-size:11.5px">'+esc(it.company)+'</span>':'')+'</td>';
      html+='<td class="mut">'+esc(it.onhire||'—')+(it.days?'<br><span style="font-size:11px">'+it.days+' day'+(it.days>1?'s':'')+'</span>':'')+'</td>';
      html+='<td style="text-align:right">'+(it.rate!=null?'<b>'+money(it.rate)+'</b>':'<span class="pidnone">—</span>')+'</td>';
      html+='<td class="tocost">'+(it.todate!=null?money(it.todate):'<span class="pidnone">—</span>')+'</td></tr>';
    });
    html+='</tbody></table></details>';
  });
  el.innerHTML=html||'<div class="empty">No items.</div>';
}
function renderQty(host, obj){
  var el=document.getElementById(host), totq=0, totd=0, tott=0, max=0;
  Object.keys(obj).forEach(function(k){ totq+=obj[k].qty; totd+=obj[k].daily; tott+=obj[k].todate; if(obj[k].qty>max)max=obj[k].qty; });
  var html='';
  Object.keys(obj).forEach(function(k){
    var o=obj[k], pct=max?Math.round(o.qty/max*100):0;
    html+='<div class="qrow"><div class="ql">'+esc(k)+'</div><div class="qbar"><div class="qfill" style="width:'+pct+'%"></div></div>'
        +'<div class="qn">'+o.qty+'</div><div class="qm">'+money0(o.daily)+'/day</div></div>';
  });
  html+='<div class="ptot"><span>Total</span><span class="sep2"></span><span class="qn">'+totq+'</span><span class="qm">'+money0(totd)+'/day</span></div>';
  html+='<div class="ptot2">Cost to date: <b>'+money0(tott)+'</b></div>';
  el.innerHTML=html;
}
function tally(host, daily, todate, unp, kind){
  document.getElementById(host).innerHTML='<div class="tally">'
    +'<div class="ti"><span class="tl">Daily hire</span><span class="tv '+(kind==='idle'?'amb':'org')+'">'+money0(daily)+'<span style="font-size:13px;font-weight:600">/day</span></span></div>'
    +'<div class="ti"><span class="tl">'+(kind==='idle'?'Idle cost to date':'Cost to date')+' (since on-hire)</span><span class="tv">'+money0(todate)+'</span></div>'
    +'<div class="sep"></div>'
    +(unp?'<div class="note2">'+unp+' item'+(unp>1?'s':'')+' with no rate on file — excluded from totals</div>':'')
    +'</div>';
}
document.getElementById('site-tally').innerHTML='<div class="tally">'
  +'<div class="ti"><span class="tl">Whole site — daily hire</span><span class="tv org">'+money0(DATA.tot.all_daily)+'<span style="font-size:13px;font-weight:600">/day</span></span></div>'
  +'<div class="ti"><span class="tl">Whole site — cost to date</span><span class="tv">'+money0(DATA.tot.all_todate)+'</span></div>'
  +'<div class="sep"></div><div class="note2">Site plant, chutes &amp; barriers at contracted daily rates · unpriced items excluded, never estimated</div></div>';
// Compliance exposure - the same requirement that prints under each line,
// counted for the plant that's out with crews. A cell only appears when
// there is something to report, and the whole strip stays away when the
// master flags nothing - the dashboard never shows an empty obligation.
(function(){
  var c=DATA.comp||{};
  if(!c.any) return;
  function ti(label,val,kind){ return val?('<div class="ti"><span class="tl">'+label+'</span>'
    +'<span class="tv'+(kind?' '+kind:'')+'">'+val+'</span></div>'):''; }
  document.getElementById('comp-tally').innerHTML='<div class="tally">'
    +ti(esc(c.tagname)+' tag · electrical',c.electrical,'org')
    +ti(esc(c.tagname)+' tag · rigging',c.rigging,'org')
    +ti('Pre-start &amp; logbook',c.logbook,'amb')
    +ti('Back daily',c.ret,'amb')
    +'<div class="ti"><span class="tl">Lines with a check</span><span class="tv">'+c.any
    +'<span style="font-size:13px;font-weight:600"> of '+c.total+'</span></span></div>'
    +'<div class="sep"></div><div class="note2">Plant out with crews that needs a current tag, '
    +'a logbook entry or a same-day return · straight off the master, never guessed</div></div>';
})();
tally('onhire-tally',DATA.tot.oh_daily,DATA.tot.oh_todate,DATA.tot.oh_unp,'onhire');
tally('idle-tally',DATA.tot.id_daily,DATA.tot.id_todate,DATA.tot.id_unp,'idle');
function renderOverview(){
  var o=DATA.ov, C=314.159, dash=(o.util/100*C).toFixed(1);
  var ohPct=o.plant_daily?Math.round(o.oh_daily/o.plant_daily*100):0, idPct=100-ohPct, h='';
  h+='<div class="ovhero">';
  h+='<div class="ovcard"><div class="ovh">Fleet deployment</div><div class="ovdonut">'
    +'<svg viewBox="0 0 120 120" width="150" height="150"><circle cx="60" cy="60" r="50" fill="none" stroke="#232935" stroke-width="13"/>'
    +'<circle cx="60" cy="60" r="50" fill="none" stroke="#F26222" stroke-width="13" stroke-linecap="round" transform="rotate(-90 60 60)" stroke-dasharray="'+dash+' '+C.toFixed(1)+'"/>'
    +'<text x="60" y="58" text-anchor="middle" font-size="27" font-weight="800" fill="#EAEDF1">'+o.util+'%</text>'
    +'<text x="60" y="77" text-anchor="middle" font-size="8.5" letter-spacing="1" fill="#95A0AF">DEPLOYED</text></svg>'
    +'<div class="ovleg"><div><span class="dotc" style="background:#F26222"></span>'+o.deployed+' out with crews</div>'
    +'<div><span class="dotc" style="background:#3A424E"></span>'+o.parked+' idle in the pool</div>'
    +'<div class="ovmut">'+o.total+' units on charge</div></div></div></div>';
  h+='<div class="ovcard"><div class="ovh">Daily plant spend — where it goes</div>'
    +'<div class="ovbig">'+money0(o.id_daily)+'<span class="ovbigsub"> /day idle</span></div>'
    +'<div class="ovsub"><b class="amb">'+o.idle_share+'%</b> of daily plant hire is sitting idle</div>'
    +'<div class="splitbar"><div class="seg org" style="width:'+ohPct+'%">'+(ohPct>=14?'On hire':'')+'</div><div class="seg amb" style="width:'+idPct+'%">'+(idPct>=14?'Idle':'')+'</div></div>'
    +'<div class="splitleg"><span><span class="dotc" style="background:#F26222"></span>On hire '+money0(o.oh_daily)+'/day</span><span><span class="dotc" style="background:#E8A32C"></span>Idle '+money0(o.id_daily)+'/day</span></div></div>';
  h+='</div>';
  var cats=o.cat_daily, mx=0; Object.keys(cats).forEach(function(k){if(cats[k]>mx)mx=cats[k]});
  h+='<div class="ovcard wide"><div class="ovh">Daily hire by category <span class="ovmut">— all site plant on charge</span></div>';
  Object.keys(cats).forEach(function(k){var pct=mx?Math.round(cats[k]/mx*100):0;
    h+='<div class="qrow"><div class="ql" style="width:200px">'+esc(k)+'</div><div class="qbar"><div class="qfill" style="width:'+pct+'%"></div></div><div class="qm" style="width:96px">'+money0(cats[k])+'/day</div></div>';});
  h+='</div>';
  h+='<div class="ovcard wide"><div class="ovh">Top idle cost targets <span class="ovmut">— biggest daily saving if off-hired or redeployed</span></div>';
  h+='<table class="ovtbl"><thead><tr><th style="width:64px">Plant ID</th><th style="width:118px">Asset</th><th>Item</th><th style="width:150px">Category</th><th style="width:82px;text-align:right">Daily $</th><th style="width:64px;text-align:right">Days</th><th style="width:100px;text-align:right">To Date $</th></tr></thead><tbody>';
  o.top_idle.forEach(function(it){h+='<tr class="row"><td>'+(it.pid?'<span class="pid">'+esc(String(it.pid))+'</span>':'<span class="pidnone">—</span>')+'</td>'
    +'<td><span class="asset">'+esc(it.asset)+'</span></td><td>'+esc(it.desc)+(it.cb?'<div class="cb">'+it.cb+'</div>':'')+'</td><td class="mut">'+esc(it.cat)+'</td>'
    +'<td style="text-align:right"><b>'+money(it.rate)+'</b></td><td style="text-align:right" class="mut">'+(it.days||'—')+'</td><td class="tocost">'+money(it.todate)+'</td></tr>';});
  h+='</tbody></table></div>';
  document.getElementById('ov-body').innerHTML=h;
}
renderOverview();
renderPlant('onhire-body',DATA.onhire,true);
renderPlant('idle-body',DATA.idle,false);
renderQty('chutes',DATA.chutes);
renderQty('barriers',DATA.barriers);
document.getElementById('q').style.display='none';
// tabs
document.querySelectorAll('.tab').forEach(function(t){t.onclick=function(){
  document.querySelectorAll('.tab').forEach(function(x){x.classList.remove('on')});t.classList.add('on');
  document.querySelectorAll('.tabview').forEach(function(v){v.classList.add('hide')});
  document.getElementById('v-'+t.dataset.t).classList.remove('hide');
  document.getElementById('q').style.display=(t.dataset.t==='onhire'||t.dataset.t==='idle')?'':'none';
}});
// search
document.getElementById('q').oninput=function(){
  var s=this.value.toLowerCase().trim();
  document.querySelectorAll('.row').forEach(function(r){
    r.style.display=(!s||r.dataset.s.indexOf(s)>=0)?'':'none';
  });
  document.querySelectorAll('.cat').forEach(function(c){
    var vis=c.querySelectorAll('.row:not([style*="none"])').length;
    c.style.display=(s&&vis===0)?'none':'';
  });
};
</script>
</body></html>'''

html = HTML_TEMPLATE
html = html.replace('__ASAT__', asat)
html = html.replace('__NPLANT__', str(n_plant)).replace('__NONHIRE__', str(n_onhire)).replace('__NIDLE__', str(n_idle))
html = html.replace('__NCHUTE__', str(n_chute)).replace('__NBARR__', str(n_barr))
html = html.replace('__DATA__', json.dumps(DATA))

_dirs = report_paths.out_dirs(HERE)      # organised day folder
outdir = _dirs['day']
stamp = datetime.datetime.now().strftime('%Y-%m-%d')
outpath = os.path.join(_dirs['pages'], 'Coates_K2_Plant_Dashboard_%s.html' % stamp)
with open(outpath, 'w', encoding='utf-8') as f:
    f.write(html)
print("==============================================================")
print(" COATES | K2 PLANT ON SITE DASHBOARD")
print(" Author: Andrew Fisher | POWERED BY SITEIQ")
print("==============================================================")
print(" Data as at : %s" % asat)
print(" Plant on site : %d  (on hire %d, idle %d)" % (n_plant, n_onhire, n_idle))
print(" Chutes : %d   Barriers : %d" % (n_chute, n_barr))
print(" Plant IDs matched : %d" % sum(1 for s in ('onhire','idle') for c in DATA[s] for it in DATA[s][c] if it['pid']))
print("")
print(" Saved: %s" % outpath)
print(" Double-click it to open, or attach it to an email for the team.")

# ---- email version: the dashboard pasted into the body as images ----------
# Same look as the Cost Tracking Snapshot email - the dashboard captured
# in order as clear full-width images, with the interactive HTML attached
# for anyone who wants to click around. Skipped quietly (with a note) if
# no browser is on this machine.
try:
    import json as _json
    import email_images
    import report_pages
    # Build the PAGED email edition: the dashboard's overview dealt onto
    # framed sheets (wider 1000x1414 pages) - real pages, never sliced.
    # The interactive original stays the attachment.
    _pages_path = os.path.join(_dirs['pages'],
                               'Coates_K2_Plant_Dashboard_%s_EmailPages.html' % stamp)
    _hdrn = ("<div class='bandc'><div>"
             "<div class='k'>COATES &middot; PLANT ON SITE DASHBOARD</div>"
             "<h1>Site plant, chutes &amp; barriers</h1></div>"
             "<div class='m'><span class='siq'>POWERED BY SITEIQ</span><br>"
             "As at <b>%s</b><br><span class='pgnum'></span></div></div>") % asat
    _wrap = (
        "<style>" + report_pages.page_css(1000, 1414, 1346) +
        "body{background:#fff}.tabs,.search{display:none!important}"
        ".ppanel .hd{padding:4px 2px 0}"
        "</style>"
        "<script>(function(){"
        "var stage=document.createElement('div');stage.id='stage';"
        "var sheets=document.createElement('div');sheets.id='sheets';"
        "var hd=document.querySelector('header .hd');"
        "window.PAGES_HDR1=hd?\"<div class='hd'>\"+hd.innerHTML+\"</div>\":'';"
        "window.PAGES_HDRN=" + _json.dumps(_hdrn) + ";"
        "var foot=document.querySelector('footer');"
        "window.PAGES_FOOT=foot?'<footer>'+foot.innerHTML+'</footer>':'';"
        # Meet the team on every printed dashboard page too, same as the
        # report packs. (A. Fisher, 25 Jul 2026)
        "window.PAGES_TEAM=" + _json.dumps(TEAM_STRIP) + ";"
        "if(foot&&foot.parentNode)foot.parentNode.removeChild(foot);"
        "var wrap=document.querySelector('.wrap');var kids=[];"
        "if(wrap){while(wrap.firstElementChild)kids.push(wrap.removeChild(wrap.firstElementChild));}"
        "for(var i=0;i<kids.length;i++){var k=kids[i];var cn=''+(k.className||'');"
        "if(k.id==='v-ov'){var ov=k.querySelector('#ov-body')||k;"
        "while(ov.firstElementChild)stage.appendChild(ov.removeChild(ov.firstElementChild));}"
        "else if(cn.indexOf('tabview')>=0||cn.indexOf('tabs')>=0||cn.indexOf('search')>=0){}"
        "else stage.appendChild(k);}"
        "var header=document.querySelector('header');"
        "if(header&&header.parentNode)header.parentNode.removeChild(header);"
        "if(wrap&&wrap.parentNode)wrap.parentNode.removeChild(wrap);"
        "document.body.appendChild(stage);document.body.appendChild(sheets);"
        "})();</script>"
        "<script>" + report_pages.PAGINATE_JS + "</script>")
    with open(_pages_path, 'w', encoding='utf-8') as f:
        f.write(html.replace('</body>', _wrap + '</body>', 1))
    _eml_imgs = email_images.capture_report_images(
        _pages_path, _dirs['emails'], 'Coates_K2_Plant_Dashboard_%s' % stamp,
        width=1000, page_h=1414)
    if _eml_imgs:
        import recipients
        _to, _cc = recipients.resolve(HERE, '', 'PLANT')
        _subject = 'Coates K2 - Plant On Site Dashboard - %s' % (
            datetime.datetime.now().strftime('%d %b %Y'))
        _msg = email_images.build_image_eml(_subject, _eml_imgs,
                                            to=_to, cc=_cc)
        _eml_path = os.path.join(_dirs['emails'],
                                 'Coates_K2_Plant_Dashboard_%s.eml' % stamp)
        email_images.save_eml(_msg, _eml_path)
        _cids = ['pg%d' % i for i in range(1, len(_eml_imgs) + 1)]
        email_images.write_manifest(_eml_path, _subject,
                                    email_images.images_body(_cids, width=1000),
                                    _eml_imgs, report='PLANT')
        print(" Email: %s  (dashboard in the body, %d images)" % (
            _eml_path, len(_eml_imgs)))
        if __name__ == '__main__':
            # run directly (not imported by the snapshot): finish with the
            # native Outlook draft - dashboard pages in the compose body
            email_images.make_native_drafts(HERE)
    else:
        print(" Email skipped - no Edge/Chrome found to draw the dashboard "
              "images (the HTML above still attaches fine).")
except Exception as _exc:
    print(" Email skipped (%s) - the HTML above still attaches fine." % _exc)
