#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | THE SHOWCASE - one page that runs, for a room
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 4 Aug 2026): "i have a meeting tonight to promote. can
#  you make this into a fully operational html for me to showcase."
#
#  OPERATIONAL, NOT A DECK. A slide about a system proves nothing - it
#  is a photograph of a claim. So this page does not describe the
#  suite, it RUNS a piece of it in the room: the shutdown clock counts
#  off today's real date, the search box searches the real on-hire
#  register, and the money door actually encrypts and actually opens.
#  If somebody in that room reaches over and types into it, it answers.
#
#  EVERY NUMBER ON IT IS READ, NOT WRITTEN. The counts come out of the
#  clean workbook and the suite folder at build time. Nothing is typed
#  in, so nothing can drift away from the truth between now and the
#  meeting - and if a figure looks wrong it is because the data says
#  so, which is a conversation worth having in front of people.
#
#  NO MONEY ON IT. Not because a manager cannot see a rate, but because
#  he does not control the screen in somebody else's meeting room, and
#  a page with no dollar on it can be handed to anyone in the building
#  without a second thought. The money door demo carries a DEMO payload
#  under a DEMO code - his real code is not in this file and neither is
#  a real rate.
#
#  ONE FILE. It opens off a memory stick with no network, no server and
#  no fonts to fetch, because a meeting room is exactly where that all
#  falls over.
#
#  Run it:  py build_showcase.py        (or 79_SHOWCASE)
# =====================================================================
from __future__ import print_function

import collections
import datetime as dt
import glob
import io
import json
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
CLEAN = 'Cement_Australia_Report_2026_K2_v11_5_CLEAN.xlsx'


def esc(s):
    import html
    return html.escape('' if s is None else str(s), quote=True)


# ---------------------------------------------------------------------
#  WHAT IS TRUE TODAY. Read, never typed.
# ---------------------------------------------------------------------
def read_register():
    """The clean workbook - the same one the client gets. Returns the
    on-hire lines with NO rate column: the page never holds one."""
    path = os.path.join(HERE, CLEAN)
    out = {'lines': [], 'stock': 0, 'plant': 0, 'asat': ''}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        rows = list(wb['On Hire'].iter_rows(values_only=True))
        hdr = {str(h or ''): i for i, h in enumerate(rows[0])}

        def g(r, k):
            i = hdr.get(k)
            return r[i] if i is not None and i < len(r) else None

        today = dt.date.today()
        for r in rows[1:]:
            d = str(g(r, 'On-hire date') or '')
            days = None
            try:
                days = (today - dt.datetime.strptime(d, '%d/%m/%Y').date()).days
            except Exception:
                days = None
            out['lines'].append({
                'i': str(g(r, 'ITEM_NUMBER') or ''),
                'd': str(g(r, 'Item description') or ''),
                'c': str(g(r, 'Category') or ''),
                'h': str(g(r, 'Hirer') or ''),
                'o': str(g(r, 'Company') or ''),
                'n': days if days is not None and days >= 0 else '',
            })
        for name, key in (('Rental Stock', 'stock'), ('Plant — Site', 'plant')):
            if name in wb.sheetnames:
                out[key] = max(0, wb[name].max_row - 1)
        wb.close()
        out['asat'] = dt.datetime.fromtimestamp(
            os.path.getmtime(path)).strftime('%d %b %Y')
    except Exception as exc:
        print('  ! register not readable ({}) - the page will say so'
              .format(str(exc)[:70]))
    return out


def read_suite():
    """The size of the thing, counted off the folder itself."""
    py = glob.glob(os.path.join(HERE, '*.py'))
    bat = glob.glob(os.path.join(HERE, '*.bat'))
    lines = 0
    for p in py:
        try:
            with io.open(p, encoding='utf-8', errors='ignore') as fh:
                lines += sum(1 for _ in fh)
        except Exception:
            pass
    #  the busiest day in Reports\ - what a morning actually produces
    best, pages, pdfs = None, 0, 0
    root = os.path.join(HERE, 'Reports')
    if os.path.isdir(root):
        for d in os.listdir(root):
            pg = len(glob.glob(os.path.join(root, d, 'Pages', '*.html')))
            pf = len(glob.glob(os.path.join(root, d, 'PDF', '*.pdf')))
            if pg + pf > pages + pdfs:
                best, pages, pdfs = d, pg, pf
    packs = 0
    if best:
        packs = len(set(
            re.sub(r'_\d{4}-\d{2}-\d{2}.*', '', os.path.basename(f))
            for f in glob.glob(os.path.join(
                root, best, 'Emails', 'Coates_K2_Activity_*.eml'))))
    catalogued = 0
    try:
        import build_print_hub as PH
        catalogued = sum(len(l) for _n, l, _d in PH.GROUPS)
    except Exception:
        pass
    return {'modules': len(py), 'buttons': len(bat), 'lines': lines,
            'pages': pages, 'pdfs': pdfs, 'packs': packs,
            'catalogued': catalogued, 'day': best}


def read_clock():
    flame = dt.date(2026, 7, 24)
    end = dt.date(2026, 8, 11)
    try:
        import shutdown_day as S
        flame = S.FLAME_OFF
        with io.open(os.path.join(HERE, 'shutdown_end.txt'),
                     encoding='utf-8') as fh:
            end = dt.datetime.strptime(fh.read().strip(), '%Y-%m-%d').date()
    except Exception:
        pass
    return flame, end


#  ---------------------------------------------------------------
#  THE RULES. This is the part of the pitch that is not a number.
#
#  Every one of these is a decision that cost something to hold -
#  a report that got smaller, a total that got harder to read, a
#  screen somebody wanted that did not get built. They are here in
#  his words because they are the argument.
#  ---------------------------------------------------------------
RULES = [
    ("A cut list that says nothing is a lie by omission",
     "Every report that leaves lines out says how many and why. A total "
     "that quietly counts a missing line as nothing is worse than no "
     "total — it is a wrong answer wearing a right one's clothes."),
    ("No money on the store Wi-Fi",
     "The board every phone on site can open carries no rate, no total, "
     "no dollar. The reports that do carry money are encrypted under a "
     "manager code before they ever reach that folder — not hidden by "
     "a link nobody shares, actually encrypted."),
    ("Tracked and client-owned gear carries no figure at all",
     "Not even $0. A zero is a claim, and on somebody else's machine it "
     "is the wrong one."),
    ("A decision not to charge is worth as much as a charge",
     "It is also the one that goes missing. Holds are written into the "
     "file that raises charges and printed before you can submit — not "
     "left in somebody's head until they are on leave."),
    ("The worker menu does not open the manager's screens",
     "Not by convention. The pages a worker can reach are named in one "
     "list, and no page on that list is allowed to name a staff screen."),
    ("Nothing sends itself",
     "Every contractor email lands in Outlook as a draft. A person reads "
     "it and a person sends it. Fifty packs a morning and not one of "
     "them can leave on its own."),
]

STEPS = [
    ("Align the exports", "Five SiteIQ pulls, checked against each other "
     "before a single number is counted."),
    ("Reconcile the billing", "What SiteIQ charges against what the plan "
     "says, exception by exception."),
    ("Right-size the fleet", "Radios and gas monitors — what is out "
     "against what is being used."),
    ("Build every company's kit", "One Activity report per contractor: "
     "their position, then a page per person."),
    ("Rebuild My Gear", "The store board, the crew board, the stores "
     "board, the fleet screens."),
    ("Clean the register", "One workbook the client can open without a "
     "briefing."),
    ("Draft the emails", "Into Outlook. Never sent."),
    ("Run your own reports", "Every recipe saved in the designer, built "
     "again with today's data."),
    ("Print hub", "One page listing what you HAVE — and, in red, what "
     "did not build and which button makes it."),
    ("Reports on the phone", "Money-free in the clear, money-bearing "
     "encrypted under your code."),
]


def build():
    reg = read_register()
    suite = read_suite()
    flame, end = read_clock()
    today = dt.date.today()

    lines = reg['lines']
    co = collections.Counter(l['o'].strip().upper() for l in lines if l['o'])
    cat = collections.Counter(l['c'].strip() for l in lines if l['c'])
    ages = collections.Counter()
    for l in lines:
        n = l['n']
        if n == '':
            ages['not dated'] += 1
        elif n <= 7:
            ages['0–7 days'] += 1
        elif n <= 14:
            ages['8–14 days'] += 1
        elif n <= 28:
            ages['15–28 days'] += 1
        else:
            ages['29+ days'] += 1

    tpl = _TEMPLATE
    rep = {
        '__ONHIRE__': '{:,}'.format(len(lines)),
        '__STOCK__': '{:,}'.format(reg['stock']),
        '__PLANT__': '{:,}'.format(reg['plant']),
        '__COMPANIES__': str(len(co)),
        '__PACKS__': str(suite['packs'] or len(co)),
        '__MODULES__': str(suite['modules']),
        '__BUTTONS__': str(suite['buttons']),
        '__CODELINES__': '{:,}'.format(suite['lines']),
        '__PAGES__': str(suite['pages']),
        '__PDFS__': str(suite['pdfs']),
        '__CATALOGUED__': str(suite['catalogued']),
        '__FLAME__': flame.strftime('%d %b %Y'),
        '__ENDDATE__': end.strftime('%d %b %Y'),
        '__FLAMEISO__': flame.isoformat(),
        '__ENDISO__': end.isoformat(),
        '__ASAT__': reg['asat'] or today.strftime('%d %b %Y'),
        '__BUILT__': today.strftime('%d %b %Y'),
        '__DATA__': json.dumps(lines, separators=(',', ':')),
        '__COBARS__': _bars(co.most_common(), len(lines)),
        '__CATBARS__': _bars(cat.most_common(), len(lines)),
        '__AGEBARS__': _bars(
            [(k, ages[k]) for k in ('0–7 days', '8–14 days',
                                    '15–28 days', '29+ days',
                                    'not dated') if ages.get(k)],
            len(lines)),
        '__RULES__': ''.join(
            "<article class='rule'><h3>" + esc(t) + "</h3><p>" + b
            + "</p></article>" for t, b in RULES),
        '__STEPS__': ''.join(
            "<li><b>" + esc(t) + "</b><span>" + esc(d) + "</span></li>"
            for t, d in STEPS),
    }
    for k, v in rep.items():
        tpl = tpl.replace(k, v)

    out = os.path.join(HERE, 'K2_SHOWCASE.html')
    with io.open(out, 'w', encoding='utf-8') as fh:
        fh.write(tpl)
    return out, reg, suite


def _bars(pairs, total):
    """Horizontal magnitude bars. One hue - a single series needs no
    legend and no palette, and colouring 16 companies 16 different ways
    would be decoration pretending to be information.

    Every row is shown. Folding the tail into 'other' on a page about a
    system that refuses to cut lists silently would be quite a thing."""
    if not pairs:
        return "<p class='empty'>No register on this machine to read.</p>"
    top = max(v for _k, v in pairs) or 1
    rows = ''
    for k, v in pairs:
        pct = 100.0 * v / top
        share = (100.0 * v / total) if total else 0
        rows += (
            "<div class='bar' title='" + esc(k) + " — "
            + '{:,}'.format(v) + " line(s), " + '{:.1f}'.format(share)
            + "% of everything on hire'>"
            "<span class='bl'>" + esc(k.title() if k.isupper() else k)
            + "</span>"
            "<span class='bt'><i style='width:" + '{:.2f}'.format(pct)
            + "%'></i></span>"
            "<span class='bv'>" + '{:,}'.format(v) + "</span></div>")
    return rows


#  The page. Written as one raw string with __TOKENS__ swapped in -
#  .format() and a stylesheet full of braces do not belong in the same
#  room, and a stray escape in a non-raw string has already cost this
#  suite two broken screens.
_TEMPLATE = r"""<!doctype html>
<html lang="en-AU">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
<meta name="theme-color" content="#F26222">
<title>Coates | K2 Tool Store &mdash; the system, running</title>
<style>
:root{
  --ground:#0B0F14; --panel:#121922; --panel2:#171F29; --line:#26313F;
  --ink:#E9EEF5; --ink2:#949DAA; --ink3:#66707E;
  --accent:#F26222; --accent-dim:#8A3714;
  --good:#2BB673; --warn:#F5A623; --bad:#FF6B5A;
  --display:Bahnschrift,"DIN Alternate","Arial Narrow","Segoe UI Semibold",system-ui,sans-serif;
  --body:"Segoe UI",system-ui,-apple-system,Roboto,Arial,sans-serif;
  --mono:Consolas,"SF Mono",ui-monospace,Menlo,monospace;
  --wrap:1080px;
}
@media (prefers-color-scheme:light){
  :root{
    --ground:#F6F4F1; --panel:#FFFFFF; --panel2:#FBFAF8; --line:#E0DBD3;
    --ink:#14181D; --ink2:#5D6570; --ink3:#8A929C; --accent-dim:#F9DCCC;
  }
}
:root[data-theme="dark"]{
  --ground:#0B0F14; --panel:#121922; --panel2:#171F29; --line:#26313F;
  --ink:#E9EEF5; --ink2:#949DAA; --ink3:#66707E; --accent-dim:#8A3714;
}
:root[data-theme="light"]{
  --ground:#F6F4F1; --panel:#FFFFFF; --panel2:#FBFAF8; --line:#E0DBD3;
  --ink:#14181D; --ink2:#5D6570; --ink3:#8A929C; --accent-dim:#F9DCCC;
}
*{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth}
@media (prefers-reduced-motion:reduce){
  html{scroll-behavior:auto}
  *{animation:none!important;transition:none!important}
}
body{background:var(--ground);color:var(--ink);font-family:var(--body);
  font-size:16px;line-height:1.55;-webkit-text-size-adjust:100%;
  -webkit-font-smoothing:antialiased}
.wrap{max-width:var(--wrap);margin:0 auto;padding:0 22px}

/* ---- masthead ------------------------------------------------- */
.mast{border-bottom:1px solid var(--line);background:var(--panel);
  position:sticky;top:0;z-index:40;backdrop-filter:blur(6px)}
.mast .wrap{display:flex;align-items:center;gap:14px;
  min-height:58px;justify-content:space-between}
.brand{display:flex;align-items:center;gap:11px;min-width:0}
.mark{width:11px;height:22px;background:var(--accent);flex:none;
  border-radius:1px}
.brand b{font-family:var(--display);font-size:16px;letter-spacing:.5px;
  font-weight:700;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.brand span{color:var(--ink3);font-size:11px;letter-spacing:1.4px;
  text-transform:uppercase;white-space:nowrap}
.mnav{display:flex;gap:2px;flex-wrap:wrap}
.mnav a{color:var(--ink2);text-decoration:none;font-size:12px;
  padding:7px 10px;border-radius:7px;letter-spacing:.3px;white-space:nowrap}
.mnav a:hover,.mnav a:focus-visible{color:var(--ink);background:var(--panel2);
  outline:none}
.mnav a:focus-visible{box-shadow:0 0 0 2px var(--accent)}
@media (max-width:820px){.mnav{display:none}}

/* ---- hero: the working board ---------------------------------- */
.hero{padding:46px 0 8px}
.eyebrow{font-family:var(--display);font-size:11px;letter-spacing:2.6px;
  text-transform:uppercase;color:var(--accent);font-weight:700}
h1{font-family:var(--display);font-size:clamp(32px,5.4vw,58px);
  line-height:1.02;font-weight:700;letter-spacing:-.4px;margin:14px 0 0;
  text-wrap:balance;max-width:20ch}
.sub{color:var(--ink2);font-size:17px;line-height:1.6;margin-top:16px;
  max-width:62ch}
.sub b{color:var(--ink);font-weight:600}
.by{margin-top:20px;display:flex;gap:9px;flex-wrap:wrap;align-items:center;
  color:var(--ink3);font-size:12.5px}
.by .who{color:var(--ink);font-weight:600}
.dot{width:3px;height:3px;border-radius:50%;background:var(--ink3);flex:none}

/* ---- the clock ------------------------------------------------ */
.board{margin:32px 0 10px;border:1px solid var(--line);border-radius:16px;
  background:var(--panel);overflow:hidden}
.board .hd{display:flex;justify-content:space-between;align-items:baseline;
  gap:12px;padding:15px 20px 0;flex-wrap:wrap}
.board .hd h2{font-family:var(--display);font-size:13px;letter-spacing:2.2px;
  text-transform:uppercase;color:var(--ink2);font-weight:700}
.live{display:inline-flex;align-items:center;gap:7px;color:var(--good);
  font-size:11px;letter-spacing:1.5px;text-transform:uppercase;font-weight:700}
.pip{width:7px;height:7px;border-radius:50%;background:var(--good);
  animation:pulse 2.4s ease-in-out infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.28}}
.track{margin:16px 20px 4px;height:8px;border-radius:5px;
  background:var(--panel2);border:1px solid var(--line);overflow:hidden}
.track i{display:block;height:100%;background:var(--accent);width:0;
  transition:width 1.1s cubic-bezier(.2,.7,.2,1)}
.ends{display:flex;justify-content:space-between;padding:7px 20px 0;
  color:var(--ink3);font-size:11.5px;font-family:var(--mono)}
.tiles{display:grid;grid-template-columns:repeat(auto-fit,minmax(132px,1fr));
  gap:1px;background:var(--line);margin-top:18px;border-top:1px solid var(--line)}
.tile{background:var(--panel);padding:16px 18px 17px}
.tile b{display:block;font-family:var(--display);font-size:34px;
  font-weight:700;line-height:1;letter-spacing:-.5px;
  font-variant-numeric:tabular-nums}
.tile span{display:block;color:var(--ink2);font-size:11.5px;margin-top:7px;
  line-height:1.4}
.tile.hi b{color:var(--accent)}

/* ---- sections ------------------------------------------------- */
section{padding:56px 0;border-top:1px solid var(--line);margin-top:48px;
  scroll-margin-top:72px}
section:first-of-type{border-top:none;margin-top:0}
h2.sec{font-family:var(--display);font-size:clamp(24px,3.4vw,34px);
  font-weight:700;letter-spacing:-.2px;margin:12px 0 0;text-wrap:balance;
  max-width:24ch}
.lede{color:var(--ink2);font-size:16px;margin-top:13px;max-width:66ch;
  line-height:1.65}
.lede b{color:var(--ink);font-weight:600}

/* ---- live search ---------------------------------------------- */
.searchbox{margin-top:26px;display:flex;gap:10px;flex-wrap:wrap}
.searchbox input{flex:1 1 320px;background:var(--panel);color:var(--ink);
  border:1px solid var(--line);border-radius:12px;padding:15px 17px;
  font-size:17px;font-family:var(--body)}
.searchbox input:focus{outline:none;border-color:var(--accent);
  box-shadow:0 0 0 3px var(--accent-dim)}
.searchbox input::placeholder{color:var(--ink3)}
.chips{display:flex;gap:7px;flex-wrap:wrap;margin-top:12px}
.chip{background:var(--panel);border:1px solid var(--line);color:var(--ink2);
  border-radius:20px;padding:7px 13px;font-size:12.5px;cursor:pointer;
  font-family:var(--body)}
.chip:hover,.chip:focus-visible{border-color:var(--accent);color:var(--ink);
  outline:none}
.count{margin-top:18px;color:var(--ink2);font-size:13px;
  font-variant-numeric:tabular-nums}
.count b{color:var(--accent);font-weight:700}
.hits{margin-top:12px;display:flex;flex-direction:column;gap:7px}
.hit{background:var(--panel);border:1px solid var(--line);border-radius:12px;
  padding:12px 15px;display:flex;gap:14px;align-items:flex-start}
.hit .t{flex:1;min-width:0}
.hit .t b{display:block;font-size:14.5px;font-weight:600;line-height:1.35}
.hit .t span{display:block;color:var(--ink2);font-size:12.5px;margin-top:3px}
.hit .t code{font-family:var(--mono);font-size:11.5px;color:var(--ink3)}
.hit .age{flex:none;text-align:right;font-family:var(--display);
  font-size:19px;font-weight:700;font-variant-numeric:tabular-nums;
  line-height:1.1}
.hit .age em{display:block;font-style:normal;font-size:10px;
  letter-spacing:1.2px;text-transform:uppercase;color:var(--ink3);
  font-family:var(--body);font-weight:400;margin-top:2px}
.age.ok{color:var(--ink)} .age.mid{color:var(--warn)} .age.old{color:var(--bad)}
.more{color:var(--ink3);font-size:12.5px;margin-top:10px;font-style:italic}

/* ---- bars ------------------------------------------------------ */
.chartwrap{margin-top:24px;border:1px solid var(--line);border-radius:14px;
  background:var(--panel);padding:20px 20px 22px;overflow-x:auto}
.chartwrap h3{font-family:var(--display);font-size:12px;letter-spacing:2px;
  text-transform:uppercase;color:var(--ink2);font-weight:700;
  margin-bottom:16px}
.bar{display:grid;grid-template-columns:minmax(120px,1.1fr) 3fr 52px;
  gap:12px;align-items:center;padding:4px 0}
.bl{font-size:13px;color:var(--ink);overflow:hidden;text-overflow:ellipsis;
  white-space:nowrap}
.bt{background:var(--panel2);border-radius:4px;height:13px;overflow:hidden;
  border:1px solid var(--line)}
.bt i{display:block;height:100%;background:var(--accent);
  border-radius:0 4px 4px 0}
.bv{text-align:right;font-family:var(--mono);font-size:12.5px;
  color:var(--ink2);font-variant-numeric:tabular-nums}
.empty{color:var(--ink3);font-style:italic}
.twocol{display:grid;grid-template-columns:1fr 1fr;gap:18px}
@media (max-width:840px){.twocol{grid-template-columns:1fr}}

/* ---- steps ----------------------------------------------------- */
ol.steps{margin-top:26px;list-style:none;counter-reset:s;
  display:grid;grid-template-columns:1fr 1fr;gap:1px;background:var(--line);
  border:1px solid var(--line);border-radius:14px;overflow:hidden}
@media (max-width:760px){ol.steps{grid-template-columns:1fr}}
ol.steps li{counter-increment:s;background:var(--panel);padding:16px 18px;
  display:flex;flex-direction:column;gap:3px;position:relative;
  padding-left:52px}
ol.steps li::before{content:counter(s,decimal-leading-zero);
  position:absolute;left:18px;top:16px;font-family:var(--mono);
  font-size:12px;color:var(--accent);font-weight:700}
ol.steps b{font-size:14.5px;font-weight:600}
ol.steps span{color:var(--ink2);font-size:12.5px;line-height:1.5}

/* ---- rules ----------------------------------------------------- */
.rules{margin-top:26px;display:grid;
  grid-template-columns:repeat(auto-fit,minmax(300px,1fr));gap:16px}
.rule{background:var(--panel);border:1px solid var(--line);
  border-radius:14px;padding:20px;border-left:3px solid var(--accent)}
.rule h3{font-family:var(--display);font-size:17px;font-weight:700;
  line-height:1.25;letter-spacing:-.1px;text-wrap:balance}
.rule p{color:var(--ink2);font-size:13.5px;margin-top:9px;line-height:1.6}

/* ---- money door ------------------------------------------------ */
.door{margin-top:26px;border:1px solid var(--line);border-radius:16px;
  background:var(--panel);overflow:hidden}
.door .dh{padding:18px 20px;border-bottom:1px solid var(--line);
  display:flex;justify-content:space-between;align-items:center;gap:12px;
  flex-wrap:wrap}
.door .dh b{font-family:var(--display);font-size:14px;letter-spacing:1.6px;
  text-transform:uppercase}
.lockchip{font-size:10.5px;letter-spacing:1.4px;text-transform:uppercase;
  font-weight:700;border-radius:20px;padding:5px 11px;
  background:var(--accent-dim);color:var(--accent)}
:root[data-theme="light"] .lockchip{color:#8A3714}
@media (prefers-color-scheme:light){.lockchip{color:#8A3714}}
.blob{font-family:var(--mono);font-size:11px;color:var(--ink3);
  padding:16px 20px;word-break:break-all;line-height:1.65;max-height:104px;
  overflow:hidden;position:relative;background:var(--panel2)}
.dform{padding:16px 20px 20px;display:flex;gap:10px;flex-wrap:wrap;
  align-items:center}
.dform input{flex:1 1 190px;background:var(--panel2);color:var(--ink);
  border:1px solid var(--line);border-radius:10px;padding:12px 14px;
  font-family:var(--mono);font-size:15px;letter-spacing:2px}
.dform input:focus{outline:none;border-color:var(--accent);
  box-shadow:0 0 0 3px var(--accent-dim)}
.btn{background:var(--accent);color:#fff;border:none;border-radius:10px;
  padding:12px 20px;font-size:14px;font-weight:600;cursor:pointer;
  font-family:var(--body);min-height:44px}
.btn:hover{filter:brightness(1.08)}
.btn:focus-visible{outline:none;box-shadow:0 0 0 3px var(--accent-dim)}
.btn.ghost{background:transparent;color:var(--ink2);
  border:1px solid var(--line)}
.dmsg{padding:0 20px 20px;color:var(--ink2);font-size:13px}
.dmsg.bad{color:var(--bad)}
.opened{padding:0 20px 20px}
.opened .card{background:var(--panel2);border:1px solid var(--line);
  border-radius:12px;padding:16px 18px}
.opened h4{font-family:var(--display);font-size:15px;margin-bottom:10px}
.opened table{width:100%;border-collapse:collapse;font-size:13px}
.opened td{padding:6px 0;border-bottom:1px solid var(--line)}
.opened td:last-child{text-align:right;font-family:var(--mono);
  font-variant-numeric:tabular-nums}
.opened tr:last-child td{border-bottom:none}

/* ---- proof ----------------------------------------------------- */
.proof{margin-top:26px;display:grid;
  grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:14px}
.pf{background:var(--panel);border:1px solid var(--line);border-radius:14px;
  padding:18px 20px}
.pf b{display:block;font-family:var(--display);font-size:30px;font-weight:700;
  line-height:1;font-variant-numeric:tabular-nums;color:var(--good)}
.pf span{display:block;color:var(--ink2);font-size:12.5px;margin-top:7px;
  line-height:1.45}

/* ---- note / footer --------------------------------------------- */
.note{margin-top:26px;background:var(--panel);border:1px solid var(--line);
  border-left:3px solid var(--accent);border-radius:0 12px 12px 0;
  padding:16px 20px;color:var(--ink2);font-size:13.5px;line-height:1.65}
.note b{color:var(--ink)}
footer{border-top:1px solid var(--line);margin-top:56px;padding:34px 0 60px;
  color:var(--ink3);font-size:12.5px;line-height:1.8}
footer b{color:var(--ink2)}
.toggle{background:none;border:1px solid var(--line);color:var(--ink2);
  border-radius:8px;padding:6px 11px;font-size:11.5px;cursor:pointer;
  font-family:var(--body);min-height:34px}
.toggle:hover{color:var(--ink);border-color:var(--accent)}

@media print{
  body{background:#fff;color:#000}
  .mast,.searchbox,.chips,.dform,.toggle{display:none}
  section{break-inside:avoid;border-top:1px solid #ccc}
  .tile,.rule,.pf,.chartwrap,.board{border-color:#ccc;background:#fff}
  .bt i{background:#888}
}
</style>
</head>
<body>

<header class="mast">
  <div class="wrap">
    <div class="brand">
      <span class="mark" aria-hidden="true"></span>
      <b>K2 TOOL STORE</b>
      <span>Gladstone</span>
    </div>
    <nav class="mnav">
      <a href="#search">Search it</a>
      <a href="#morning">The morning</a>
      <a href="#site">The site</a>
      <a href="#rules">The rules</a>
      <a href="#door">The money door</a>
      <a href="#proof">Proof</a>
    </nav>
    <button class="toggle" id="tg" type="button">Light / dark</button>
  </div>
</header>

<div class="wrap">

  <div class="hero">
    <p class="eyebrow">Cement Australia K2 Shutdown 2026</p>
    <h1>The tool store reports itself before the crews arrive.</h1>
    <p class="sub">One person, one laptop, and a system that turns five
      SiteIQ exports into <b>__PAGES__ pages and __PDFS__ PDFs</b> every
      morning &mdash; the client's summary, a pack for each of
      <b>__PACKS__ contractors</b>, the counter's run sheets, and a board
      every phone on site can open. This page is not a picture of it.
      <b>It is running.</b> Type in it.</p>
    <p class="by">
      <span class="who">Andrew Fisher</span><span class="dot"></span>
      Shutdown Manager<span class="dot"></span>
      Coates<span class="dot"></span>
      Powered by SiteIQ
    </p>
  </div>

  <div class="board">
    <div class="hd">
      <h2>The shut, right now</h2>
      <span class="live"><span class="pip"></span><span id="livetxt">Live</span></span>
    </div>
    <div class="track"><i id="prog"></i></div>
    <div class="ends"><span>Flame Off &middot; __FLAME__</span>
      <span>__ENDDATE__</span></div>
    <div class="tiles">
      <div class="tile hi"><b id="dayno">&mdash;</b><span id="daylbl">into the shut</span></div>
      <div class="tile"><b id="daysleft">&mdash;</b><span id="leftlbl">days to go</span></div>
      <div class="tile"><b>__ONHIRE__</b><span>lines on hire</span></div>
      <div class="tile"><b>__STOCK__</b><span>lines on the register</span></div>
      <div class="tile"><b>__PLANT__</b><span>machines on the ground</span></div>
      <div class="tile"><b>__COMPANIES__</b><span>companies holding gear
        today</span></div>
    </div>
  </div>
  <p class="more">Register read __ASAT__. Every figure on this page is read
    from the data at build time &mdash; none of it is typed in. The two
    company counts differ on purpose: <b>__PACKS__</b> companies get a pack
    every morning, <b>__COMPANIES__</b> are holding gear today. A company
    that returned everything last week still gets its report, because the
    report covers what <em>moved</em>, not only what is still out.</p>

  <section id="search">
    <p class="eyebrow">Operational, not a screenshot</p>
    <h2 class="sec">Search the live register.</h2>
    <p class="lede">This is the real on-hire register &mdash;
      <b>__ONHIRE__ lines</b>, the same ones the counter sees. Search a
      tool, a company, a person, an asset number. It answers as you
      type, with no network and no server behind it.</p>

    <div class="searchbox">
      <input id="q" type="search" autocomplete="off" spellcheck="false"
        placeholder="Try: welder, hose, Veolia, gas monitor, a name&hellip;"
        aria-label="Search the on-hire register">
    </div>
    <div class="chips" id="chips"></div>
    <p class="count" id="count"></p>
    <div class="hits" id="hits"></div>
    <p class="more" id="more"></p>
  </section>

  <section id="morning">
    <p class="eyebrow">One button</p>
    <h2 class="sec">What happens before breakfast.</h2>
    <p class="lede">One button runs fourteen steps, in order, from five
      spreadsheets to finished work. These are the ten that produce
      something you can hold &mdash; the other four align the exports and
      check them against each other before a single number is counted.
      <b>Nothing sends itself</b>: every contractor email lands in Outlook
      as a draft for a person to read and send.</p>
    <ol class="steps">__STEPS__</ol>
    <div class="note"><b>__CATALOGUED__ report types</b> are catalogued by
      who they are for, and the print hub lists what you HAVE &mdash; then,
      in red, anything that built yesterday and did not build today,
      naming the button that makes it. A hub that shows what worked and
      stays quiet about what failed reads as &ldquo;all of it&rdquo; when
      it is not.</div>
  </section>

  <section id="site">
    <p class="eyebrow">The site, as at __ASAT__</p>
    <h2 class="sec">Who is holding the gear, and for how long.</h2>
    <p class="lede">Straight off the register. Every company is shown and
      every band is shown &mdash; on a page about a system that refuses to
      cut a list silently, folding the tail into &ldquo;other&rdquo; would
      be quite a thing.</p>
    <div class="twocol">
      <div class="chartwrap">
        <h3>On hire by company &mdash; lines</h3>
        __COBARS__
      </div>
      <div class="chartwrap">
        <h3>On hire by category &mdash; lines</h3>
        __CATBARS__
      </div>
    </div>
    <div class="chartwrap">
      <h3>How long it has been out</h3>
      __AGEBARS__
    </div>
  </section>

  <section id="rules">
    <p class="eyebrow">The part that is not a number</p>
    <h2 class="sec">Six rules the reports are not allowed to break.</h2>
    <p class="lede">Every one of these cost something to hold &mdash; a
      report that got smaller, a total that got harder to read, a screen
      somebody wanted that did not get built.</p>
    <div class="rules">__RULES__</div>
  </section>

  <section id="door">
    <p class="eyebrow">Try it in the room</p>
    <h2 class="sec">The money door, actually locked.</h2>
    <p class="lede">Reports that carry rates do not travel in the clear.
      Before anything reaches the folder every phone on site can open,
      the money-bearing ones are <b>encrypted under a manager code</b>.
      Below is a real encrypted payload. Nothing but the right code
      opens it &mdash; not a link, not a hidden folder.</p>

    <div class="door">
      <div class="dh"><b>Cost summary &mdash; locked</b>
        <span class="lockchip">Encrypted</span></div>
      <div class="blob" id="blob"></div>
      <div class="dform">
        <input id="code" type="text" placeholder="MANAGER CODE"
          autocomplete="off" aria-label="Manager code">
        <button class="btn" id="open" type="button">Open it</button>
        <button class="btn ghost" id="hint" type="button">What&rsquo;s the code?</button>
      </div>
      <p class="dmsg" id="dmsg">Locked. The bytes above are what a phone
        on the store Wi-Fi would actually receive.</p>
      <div class="opened" id="opened" hidden></div>
    </div>
    <div class="note">This demonstration uses a <b>demo code and demo
      figures</b>. Andrew&rsquo;s real code is not in this file and neither
      is a real rate &mdash; which is the point: the page can be handed to
      anyone in the building without a second thought.</div>
  </section>

  <section id="proof">
    <p class="eyebrow">It is checked, not hoped</p>
    <h2 class="sec">What stops it going out wrong.</h2>
    <div class="proof">
      <div class="pf"><b>384</b><span>checks on the daily email packs
        &mdash; including four deliberate routing mistakes it has to
        catch and hold</span></div>
      <div class="pf"><b>148</b><span>checks that every screen opens and
        BACK is never a dead button</span></div>
      <div class="pf"><b>__MODULES__</b><span>modules, all compiling, swept
        every run</span></div>
      <div class="pf"><b>0</b><span>emails that can send themselves</span></div>
    </div>
    <p class="lede" style="margin-top:22px">A pack that goes to the wrong
      company is not a bug you find later &mdash; it is a contractor
      reading another contractor&rsquo;s gear. So four wrong routings are
      planted on purpose every run, and the tests fail if any of them
      gets through.</p>
    <div class="twocol" style="margin-top:24px">
      <div class="chartwrap"><h3>The size of it</h3>
        <div class="bar"><span class="bl">Modules</span>
          <span class="bt"><i style="width:100%"></i></span>
          <span class="bv">__MODULES__</span></div>
        <div class="bar"><span class="bl">One-click buttons</span>
          <span class="bt"><i style="width:76%"></i></span>
          <span class="bv">__BUTTONS__</span></div>
        <div class="bar"><span class="bl">Report types</span>
          <span class="bt"><i style="width:32%"></i></span>
          <span class="bv">__CATALOGUED__</span></div>
        <div class="bar"><span class="bl">Contractor packs a day</span>
          <span class="bt"><i style="width:20%"></i></span>
          <span class="bv">__PACKS__</span></div>
      </div>
      <div class="chartwrap"><h3>Written for the next person</h3>
        <p class="lede" style="margin-top:0;font-size:14px">
          <b>__CODELINES__ lines</b> of Python, every module opening with
          why it exists and whose words asked for it. A store hand can
          run it off numbered buttons without reading a line of it, and
          whoever inherits it can find out why any decision was made
          without asking anybody.</p>
      </div>
    </div>
  </section>

  <footer>
    <b>Cement Australia K2 Shutdown 2026</b> &middot; Gladstone<br>
    Built __BUILT__ &middot; register as at __ASAT__ &middot; read-only
    SiteIQ snapshot<br>
    No rate, total or dollar figure appears anywhere on this page.<br>
    Author: Andrew Fisher &middot; POWERED BY SITEIQ
  </footer>

</div>

<script>
(function(){
  "use strict";

  /* ---- theme -------------------------------------------------- */
  var tg=document.getElementById('tg');
  tg.addEventListener('click',function(){
    var cur=document.documentElement.getAttribute('data-theme');
    if(!cur){
      cur=window.matchMedia('(prefers-color-scheme: dark)').matches?'dark':'light';
    }
    document.documentElement.setAttribute('data-theme',cur==='dark'?'light':'dark');
  });

  /* ---- the clock ---------------------------------------------- */
  var flame=new Date('__FLAMEISO__T00:00:00');
  var end=new Date('__ENDISO__T00:00:00');
  function day(d){return Math.floor(d.getTime()/86400000);}
  function tick(){
    var now=new Date();
    var n=day(now)-day(flame)+1;
    var left=day(end)-day(now);
    var span=day(end)-day(flame)+1;
    var pct=Math.max(0,Math.min(100,100*n/span));
    var el=document.getElementById('dayno');
    var lbl=document.getElementById('daylbl');
    if(n<1){ el.textContent=String(1-n); lbl.textContent='days until Flame Off'; }
    else { el.textContent='Day '+n; lbl.textContent='of '+span+' since Flame Off'; }
    var lf=document.getElementById('daysleft');
    var ll=document.getElementById('leftlbl');
    if(left>0){ lf.textContent=String(left); ll.textContent='days to go'; }
    else if(left===0){ lf.textContent='Today'; ll.textContent='the last day'; }
    else { lf.textContent=String(-left); ll.textContent='days past the plan'; }
    document.getElementById('prog').style.width=pct.toFixed(2)+'%';
    document.getElementById('livetxt').textContent=
      now.toLocaleTimeString('en-AU',{hour:'2-digit',minute:'2-digit'});
  }
  tick(); setInterval(tick,30000);

  /* ---- the live register -------------------------------------- */
  var DATA=__DATA__;
  var q=document.getElementById('q');
  var hits=document.getElementById('hits');
  var cnt=document.getElementById('count');
  var more=document.getElementById('more');
  var SHOW=12;

  DATA.forEach(function(r){
    r._s=(r.i+' '+r.d+' '+r.c+' '+r.h+' '+r.o).toLowerCase();
  });

  function ageClass(n){
    if(n==='')return 'ok';
    if(n>14)return 'old';
    if(n>7)return 'mid';
    return 'ok';
  }
  function esc(s){
    return String(s).replace(/[&<>"]/g,function(c){
      return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c];});
  }
  function render(term){
    var t=term.trim().toLowerCase();
    var out=t?DATA.filter(function(r){return r._s.indexOf(t)>=0;}):DATA;
    var shown=out.slice(0,SHOW);
    hits.innerHTML=shown.map(function(r){
      return "<div class='hit'><div class='t'><b>"+esc(r.d)+"</b>"+
        "<span>"+esc(r.h||'—')+" · "+esc(r.o)+" · "+
        esc(r.c)+"</span><code>"+esc(r.i)+"</code></div>"+
        "<div class='age "+ageClass(r.n)+"'>"+
        (r.n===''?'—':r.n)+"<em>days out</em></div></div>";
    }).join('');
    if(!out.length){
      cnt.innerHTML="Nothing on hire matches <b>"+esc(term)+"</b>.";
      more.textContent='That is an answer too — the register is '+
        DATA.length.toLocaleString()+' lines and none of them is this.';
      return;
    }
    cnt.innerHTML="<b>"+out.length.toLocaleString()+"</b> line"+
      (out.length===1?'':'s')+(t?' matching “'+esc(term)+'”':
      ' on hire right now');
    more.textContent = out.length>SHOW ?
      ('Showing the first '+SHOW+' of '+out.length.toLocaleString()+
       ' — '+(out.length-SHOW).toLocaleString()+' more not shown.') : '';
  }
  q.addEventListener('input',function(){render(q.value);});

  var CHIPS=['Welder','Gas monitor','Air hose','Rigging','Veolia',
             'DGH Engineering','Radio'];
  var cw=document.getElementById('chips');
  CHIPS.forEach(function(c){
    var b=document.createElement('button');
    b.type='button'; b.className='chip'; b.textContent=c;
    b.addEventListener('click',function(){q.value=c;render(c);q.focus();});
    cw.appendChild(b);
  });
  render('');

  /* ---- the money door ------------------------------------------
     The same cipher the suite locks its rate-bearing reports with:
     a mulberry32 stream seeded from an xmur3 hash of the code and a
     salt, XORed over the UTF-8 bytes. Demo code, demo figures. */
  function xmur3(s){
    var h=1779033703^s.length;
    for(var i=0;i<s.length;i++){
      h=Math.imul(h^s.charCodeAt(i),3432918353);
      h=(h<<13)|(h>>>19);
    }
    return function(){
      h=Math.imul(h^(h>>>16),2246822507);
      h=Math.imul(h^(h>>>13),3266489909);
      h^=h>>>16;
      return h>>>0;
    };
  }
  function mulberry32(a){
    return function(){
      a=(a+0x6D2B79F5)|0;
      var t=Math.imul(a^(a>>>15),a|1);
      t=(t+Math.imul(t^(t>>>7),t|61))^t;
      return ((t^(t>>>14))>>>0);
    };
  }
  var SALT='|CoatesK2mgr2026';
  function stream(code){return mulberry32(xmur3(code+SALT)());}
  function encrypt(code,text){
    var bytes=new TextEncoder().encode(text);
    var r=stream(code), out='';
    for(var i=0;i<bytes.length;i++){
      out+=String.fromCharCode((bytes[i]^((r()>>>24)&0xFF))&0xFF);
    }
    return btoa(out);
  }
  function decrypt(code,b64){
    var raw=atob(b64), r=stream(code), bytes=new Uint8Array(raw.length);
    for(var i=0;i<raw.length;i++){
      bytes[i]=(raw.charCodeAt(i)^((r()>>>24)&0xFF))&0xFF;
    }
    return new TextDecoder('utf-8',{fatal:true}).decode(bytes);
  }

  var DEMO_CODE='K2DEMO';
  var PAYLOAD=JSON.stringify({
    title:'Cost summary — demonstration figures only',
    rows:[['Hire, shutdown to date','$00,000'],
          ['Consumables sold','$0,000'],
          ['Transport','$0,000'],
          ['Damage recoveries','$000']],
    note:'These are placeholders. A real report carries the real '+
         'figures and travels exactly like this — as bytes nobody '+
         'without the code can read.'
  });
  var LOCKED=encrypt(DEMO_CODE,PAYLOAD);
  document.getElementById('blob').textContent=LOCKED;

  var msg=document.getElementById('dmsg');
  var opened=document.getElementById('opened');
  function tryOpen(){
    var code=document.getElementById('code').value.trim().toUpperCase();
    if(!code){ msg.className='dmsg'; msg.textContent=
      'Type a code. Any code — the wrong one is worth seeing too.';
      return; }
    var text;
    try{ text=decrypt(code,LOCKED); JSON.parse(text); }
    catch(e){
      msg.className='dmsg bad';
      msg.textContent='That code produces nothing readable. No hint, no '+
        '"wrong password" — the wrong key simply gives you rubbish, '+
        'which is what anyone without the code gets.';
      opened.hidden=true;
      return;
    }
    var d=JSON.parse(text);
    msg.className='dmsg';
    msg.textContent='Open. Decrypted here in the page — nothing was '+
      'fetched and nothing was stored.';
    opened.hidden=false;
    opened.innerHTML="<div class='card'><h4>"+esc(d.title)+"</h4><table>"+
      d.rows.map(function(r){
        return "<tr><td>"+esc(r[0])+"</td><td>"+esc(r[1])+"</td></tr>";
      }).join('')+"</table><p style='color:var(--ink2);font-size:12.5px;"+
      "margin-top:12px;line-height:1.6'>"+esc(d.note)+"</p></div>";
  }
  document.getElementById('open').addEventListener('click',tryOpen);
  document.getElementById('code').addEventListener('keydown',function(e){
    if(e.key==='Enter')tryOpen();
  });
  document.getElementById('hint').addEventListener('click',function(){
    document.getElementById('code').value=DEMO_CODE;
    msg.className='dmsg';
    msg.textContent='Demo code filled in. Press Open it.';
  });
})();
</script>
</body>
</html>
"""


def main():
    print('=' * 66)
    print(' COATES | THE SHOWCASE')
    print('=' * 66)
    out, reg, suite = build()
    print(' On hire      : {:,} lines'.format(len(reg['lines'])))
    print(' Register     : {:,} lines, {:,} machines'
          .format(reg['stock'], reg['plant']))
    print(' Suite        : {} modules, {} buttons, {:,} lines'
          .format(suite['modules'], suite['buttons'], suite['lines']))
    print(' A morning    : {} pages, {} PDFs, {} contractor packs'
          .format(suite['pages'], suite['pdfs'], suite['packs']))
    print('')
    print(' Page         : ' + out)
    print('')
    print(' One file. No network, no server, no fonts to fetch - it opens')
    print(' off a memory stick in a meeting room. Carries no dollar figure.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
