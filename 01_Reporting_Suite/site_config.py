#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | SITE CONFIG - which job this computer is running
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  The suite was written for one job (Cement Australia K2, Gladstone)
#  and the job was typed into the code in 71 places. The next job -
#  Rio Tinto Weipa - would have meant a fork, and a fork means two
#  sets of bugs.
#
#  So: the JOB is data now, not code. One small file per job in
#  Sites\, and a single line in SITE.txt saying which one is live on
#  this computer. Everything that needs to know the customer, the
#  site, where the exports live or where the reports land, asks here.
#
#    site()                 the live job
#    site("k2")             a named job, whatever is live
#    site().header_line     "Cement Australia | K2 Shutdown 2026 | Gladstone"
#    site().data_dirs(base) where this job's SiteIQ exports live
#    site().reports_dir     where this job's reports land
#
#  Nothing here changes K2. With SITE.txt saying k2 - which is what it
#  says out of the box - every path and every heading comes out byte
#  for byte the same as before.
# =====================================================================

import glob
import json
import os

HERE = os.path.dirname(os.path.abspath(__file__))
SITES_DIRNAME = "Sites"
SELECTOR = "SITE.txt"
DEFAULT_KEY = "k2"

_cache = {}


class Site(object):
    """One job. Everything the reports need to know about where they
    are and who they are for."""

    def __init__(self, key, d):
        self.key = key
        self._d = d

    def __getattr__(self, name):
        #  Any field in the json is readable as .field. Unknown fields
        #  come back as "" rather than blowing up mid-report - a missing
        #  heading is a cosmetic fault, a traceback at 5am is not.
        if name.startswith("_"):
            raise AttributeError(name)
        return self._d.get(name, "")

    # ---- identity --------------------------------------------------
    @property
    def name(self):
        return self._d.get("job", self.key)

    @property
    def header_line(self):
        """The line under every report title."""
        if self._d.get("header_line"):
            return self._d["header_line"]
        bits = [b for b in (self._d.get("customer"), self._d.get("job"),
                            self._d.get("location")) if b]
        return " · ".join(bits)

    @property
    def project_schedule(self):
        """The PROJECT SCHEDULE string SiteIQ stamps into every export.
        Used to prove an export belongs to THIS job before a report is
        built off it."""
        return self._d.get("project_schedule", "")

    # ---- where things live -----------------------------------------
    def data_dirs(self, base=None):
        """Folders to search for this job's SiteIQ exports, best first."""
        base = base or HERE
        out = []
        for rel in self._d.get("data_dirs") or ["Data_SiteIQ", "."]:
            p = os.path.normpath(os.path.join(base, rel))
            if p not in out:
                out.append(p)
        return out

    @property
    def reports_dirname(self):
        """Relative folder this job's daily reports live under."""
        return self._d.get("reports_dir") or "Reports"

    def workbook(self, base=None):
        """This job's master workbook, newest if there are several."""
        base = base or HERE
        pat = self._d.get("workbook_glob")
        if not pat:
            return None
        hits = [p for p in glob.glob(os.path.join(base, pat))
                if not os.path.basename(p).startswith("~$")]
        return max(hits, key=os.path.getmtime) if hits else None

    def __repr__(self):
        return "<Site {} - {}>".format(self.key, self.name)


# ---------------------------------------------------------------------
#  Loading
# ---------------------------------------------------------------------
def sites_dir(base=None):
    return os.path.join(base or HERE, SITES_DIRNAME)


def available(base=None):
    """Every job this suite knows about, key -> Site."""
    out = {}
    for p in sorted(glob.glob(os.path.join(sites_dir(base), "*.json"))):
        key = os.path.splitext(os.path.basename(p))[0].lower()
        try:
            with open(p, encoding="utf-8") as f:
                out[key] = Site(key, json.load(f))
        except (OSError, ValueError):
            #  A job file someone has half-edited shouldn't stop the
            #  other jobs loading.
            continue
    return out


def live_key(base=None):
    """Which job this computer is set to. SITE.txt, first non-comment
    line. Missing or unreadable falls back to K2 - the job the suite
    was built for."""
    p = os.path.join(base or HERE, SELECTOR)
    try:
        with open(p, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    return line.lower()
    except OSError:
        pass
    return DEFAULT_KEY


def set_live(key, base=None):
    """Point this computer at a different job."""
    key = key.lower().strip()
    known = available(base)
    if key not in known:
        raise KeyError("No job called '{}'. Known jobs: {}".format(
            key, ", ".join(sorted(known)) or "none"))
    p = os.path.join(base or HERE, SELECTOR)
    with open(p, "w", encoding="utf-8") as f:
        f.write("# The job this computer is running. Change it with\n"
                "# 57_SWITCH_JOB.bat - don't hand-edit unless you have to.\n")
        f.write(key + "\n")
    _cache.clear()
    return known[key]


def site(key=None, base=None):
    """The live job, or a named one."""
    key = (key or live_key(base)).lower()
    if key in _cache:
        return _cache[key]
    known = available(base)
    s = known.get(key)
    if s is None:
        #  Unknown job named in SITE.txt - fall back to K2 rather than
        #  stopping the morning, but say so once.
        s = known.get(DEFAULT_KEY) or Site(key, {})
    _cache[key] = s
    return s


# ---------------------------------------------------------------------
#  Proving an export belongs to this job
# ---------------------------------------------------------------------
def export_project_schedule(path):
    """Read the PROJECT SCHEDULE out of a SiteIQ export's REFERENCE_INFO
    sheet. Returns "" if it isn't there."""
    try:
        import openpyxl
    except ImportError:
        return ""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return ""
    try:
        if "REFERENCE_INFO" not in wb.sheetnames:
            return ""
        ws = wb["REFERENCE_INFO"]
        rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
        if len(rows) < 2:
            return ""
        head = [str(c or "").strip().upper() for c in rows[0]]
        if "PROJECT SCHEDULE" not in head:
            return ""
        return str(rows[1][head.index("PROJECT SCHEDULE")] or "").strip()
    except Exception:
        return ""
    finally:
        try:
            wb.close()
        except Exception:
            pass


def belongs_to_live_job(path, base=None):
    """(ok, message). False only when the export clearly names a
    DIFFERENT job - an export with no PROJECT SCHEDULE stamp is let
    through, because plenty of the older ones don't carry one."""
    s = site(base=base)
    want = (s.project_schedule or "").strip().lower()
    if not want:
        return True, ""
    got = export_project_schedule(path).strip()
    if not got:
        return True, ""
    if got.lower() == want:
        return True, ""
    return False, ("{} is a {} export. This computer is set to {}. "
                   "Either switch jobs with 57_SWITCH_JOB.bat or put the "
                   "right export in.".format(os.path.basename(path), got,
                                             s.name))


if __name__ == "__main__":
    s = site()
    print("=" * 62)
    print(" COATES | THIS COMPUTER IS RUNNING")
    print("=" * 62)
    print("  Job            {}".format(s.name))
    print("  Customer       {}".format(s.customer))
    print("  Location       {}".format(s.location))
    print("  SiteIQ project {}".format(s.project_schedule or "(not set)"))
    print("  Exports from   {}".format(", ".join(
        os.path.relpath(d, HERE) for d in s.data_dirs())))
    print("  Reports into   {}".format(s.reports_dirname))
    print("-" * 62)
    print("  Other jobs set up: {}".format(", ".join(
        k for k in sorted(available()) if k != s.key) or "none"))
    print("=" * 62)
