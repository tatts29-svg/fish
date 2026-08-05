#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | EQUIPMENT ACTIVITY & ACCOUNTABILITY - the model
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Two levels, one calculation.
#
#    LEVEL 1  the company's whole position - for their management
#    LEVEL 2  every hirer, on their own page - the evidence behind it
#
#  THE RECONCILIATION RULE
#  -----------------------
#  A company figure is NEVER computed on its own. Every company card is
#  literally sum(hirer[card] for hirer in company). There is no second
#  calculation, so there is nothing to drift from. reconcile() proves it
#  on every run and shouts if it ever stops being true.
#
#  WHERE THE NUMBERS COME FROM
#  ---------------------------
#    Issued / returned / same day / recovered
#        TRANSACTIONS.xlsx -> TRAN_START_DATE and TRAN_END_DATE, per asset
#        per hirer. Two sheets, no overlap in TRANSACTION_ID (checked
#        25 Jul 2026), so both are read and de-duped on the ID anyway.
#    Still on hire
#        The on-hire register, via the models already built.
#    Consumables
#        TRANSACTIONS.xlsx -> CUSTOMER_CONTRACTOR_EQUIP where
#        PRODUCT_CATEGORY = "Consumable". Per A. Fisher, 25 Jul 2026:
#        "sales stock you will find in transactions". LATEST_BARCODE is
#        the material number. There is NO price in any export, so no
#        dollar figure is shown - units and product types only.
#    Damage
#        CHARGE_REGISTER.xlsx -> "Damage - Breakdown".
#    Compliance
#        The four master-file columns, via equipment_compliance.
#
#  WHAT CONSUMABLES CAN NEVER TOUCH
#  --------------------------------
#  Still on hire, not returned same day, overdue, recovered, outstanding.
#  They are split at load time by PRODUCT_CATEGORY and never re-join the
#  hire list, so it is structurally impossible - not a rule someone has
#  to remember.
# =====================================================================

import os
import re
import glob
import datetime as dt
from collections import defaultdict

import equipment_compliance as EC

_HERE = os.path.dirname(os.path.abspath(__file__))


# ---------------------------------------------------------------------
#  helpers
# ---------------------------------------------------------------------
def _txt(v):
    return "" if v is None else str(v).strip()


def _date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = _txt(v)
    if not s:
        return None
    for f in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d %b %Y", "%d/%m/%y"):
        try:
            return dt.datetime.strptime(s[:10], f).date()
        except ValueError:
            pass
    return None


def _num(v):
    try:
        return float(_txt(v).replace(",", "") or 0)
    except ValueError:
        return 0.0


def _find(pattern):
    hits = [p for p in glob.glob(os.path.join(_HERE, pattern))
            if not os.path.basename(p).startswith("~$")]
    return max(hits, key=os.path.getmtime) if hits else None


# ---------------------------------------------------------------------
#  1. LOAD - hire movements and consumable movements, kept apart
# ---------------------------------------------------------------------
def load_movements(path=None):
    """Everything that moved, split at the door.

    Returns (hire, consumables, period). `hire` and `consumables` never
    share a row - PRODUCT_CATEGORY decides once, here, and that is the
    only place the two can be confused."""
    path = path or _find("TRANSACTIONS*.xlsx")
    hire, cons, period = [], [], (None, None)
    if not path:
        return hire, cons, period
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return hire, cons, period

    # SiteIQ declares its own window - honour it rather than assume one
    try:
        ref = list(wb["REFERENCE_INFO"].iter_rows(values_only=True))
        bits = _txt(ref[1][0]).split(" - ")
        if len(bits) == 2:
            period = (_date(bits[0]), _date(bits[1]))
    except Exception:
        pass

    seen = set()
    for sheet in ("TRANSACTION_CHARGES", "CUSTOMER_CONTRACTOR_EQUIP",
                  "TRANSACTION_WITHOUT_CHARGES"):
        if sheet not in wb.sheetnames:
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        hdr = [_txt(c).upper() for c in rows[0]]

        def col(*names):
            for n in names:
                if n in hdr:
                    return hdr.index(n)
            return None

        c_emp = col("EMPLOYER_NAME")
        c_hir = col("HIRER_NAME")
        c_desc = col("SKU/ITEM DESCRIPTION", "SKU_DESCRIPTION")
        c_item = col("SKU/ITEM_NUMBER", "SKU_NUMBER")
        c_bc = col("LATEST_BARCODE")
        c_cat = col("PRODUCT_CATEGORY")
        c_qty = col("QUANTITY")
        c_s = col("TRAN_START_DATE")
        c_st = col("TRAN_START_TIME")
        c_e = col("TRAN_END_DATE")
        c_id = col("TRANSACTION_ID")
        if c_hir is None or c_s is None:
            continue

        for r in rows[1:]:
            def g(i):
                return _txt(r[i]) if i is not None and i < len(r) else ""

            hirer = g(c_hir)
            if not hirer:
                continue
            tid = g(c_id)
            # de-dupe across sheets on the transaction id; fall back to the
            # asset + stamps when an export ever omits it
            key = tid or "|".join([g(c_bc), g(c_s), g(c_e), hirer])
            if key in seen:
                continue
            seen.add(key)

            rec = {"company": g(c_emp), "hirer": hirer,
                   "desc": g(c_desc), "item": g(c_item),
                   "barcode": g(c_bc), "tx_id": tid,
                   "qty": _num(r[c_qty]) if c_qty is not None else 1.0,
                   "start": _date(r[c_s]) if c_s is not None else None,
                   "start_time": g(c_st),
                   "end": _date(r[c_e]) if c_e is not None else None}

            if g(c_cat).upper() == "CONSUMABLE":
                cons.append(rec)          # <- never returns to the hire list
            else:
                hire.append(rec)
    wb.close()
    return hire, cons, period


def load_damage(path=None):
    """The damage register - every field the report needs already exists
    on the 'Damage - Breakdown' sheet. Neutral wording throughout: what
    was reported and what its status is, never who is at fault."""
    path = path or os.path.join(_HERE, "Coates_K2_Charge_Reporter",
                                "CHARGE_REGISTER.xlsx")
    out = []
    if not os.path.exists(path):
        return out
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Damage - Breakdown" not in wb.sheetnames:
            return out
        rows = list(wb["Damage - Breakdown"].iter_rows(values_only=True))
        hdr = {_txt(c): i for i, c in enumerate(rows[0])}

        def g(r, name):
            i = hdr.get(name)
            return _txt(r[i]) if i is not None and i < len(r) else ""

        for r in rows[1:]:
            if not g(r, "Charge ID"):
                continue
            out.append({
                "cid": g(r, "Charge ID"),
                "company": g(r, "Client / company"),
                "person": g(r, "Person using / returning item"),
                "asset": g(r, "Asset no. (item number)"),
                "plant_id": g(r, "Plant ID"),
                "desc": g(r, "Item description"),
                "when": _date(g(r, "Date")),
                "location": g(r, "Location"),
                "found": g(r, "Found during"),
                "event": g(r, "Event type"),
                "op_status": g(r, "Operational status"),
                "oos": g(r, "Out-of-service tag"),
                "offhired": g(r, "Off-hired"),
                "exposure": _num(g(r, "Replacement-cost exposure")),
                "repair": _num(g(r, "Repair / quoted cost")),
                "status": g(r, "Cost status"),
                "owner": g(r, "Coates owner"),
            })
        wb.close()
    except Exception:
        pass
    return out


# ---------------------------------------------------------------------
#  2. CLASSIFY - one item, one bucket, used at both levels
# ---------------------------------------------------------------------
_RADIO = re.compile(r"two-way radio|\bradio\b|dp4801", re.I)
_GAS = re.compile(r"gas monitor|gas detector|multi-gas|bw flex|draeger|"
                  r"drager|x-am|gasalert|microclip|altair|ventis", re.I)
_MILW = re.compile(r"milwaukee", re.I)

MIX_ORDER = ["Radios", "Gas monitors", "Milwaukee", "Rigging", "Plant",
             "Electrical", "General tooling", "Other"]


def classify(desc, asset=None, storage_unit=""):
    """The equipment-mix bucket. Compliance flags come first because they
    are data (the master file), not a guess at a description."""
    d = "{} {}".format(desc or "", storage_unit or "")
    if _GAS.search(d) or storage_unit.upper() == "GAS MONITORS":
        return "Gas monitors"
    if _RADIO.search(d) or storage_unit.upper() == "RADIOS":
        return "Radios"
    if _MILW.search(d):
        return "Milwaukee"
    f = EC.flags(asset, desc)
    if f["rigging"]:
        return "Rigging"
    if f["logbook"]:
        return "Plant"
    if f["electrical"]:
        return "Electrical"
    if storage_unit.upper() == "CEMENT SITE PLANT":
        return "Plant"
    return "General tooling"


def _blank_cards():
    return {"issued": 0, "returned": 0, "same": 0, "not_same": 0,
            "recovered": 0, "still": 0}


# ---------------------------------------------------------------------
#  3. BUILD - hirers first, company as the sum of them
# ---------------------------------------------------------------------

#  ---------------------------------------------------------------------
#  THE SHUTDOWN CURVE
#
#  One line per company: how much of our gear they have been holding,
#  every day from Flame Off to the planned end, with today marked and
#  the run-down they need to be on to finish clear.
#
#  ANCHORED, NOT ACCUMULATED. Running the movements up from zero drifts
#  - the transaction feed and the on-hire register do not reconcile to
#  the item, and by day thirteen DGH's accumulated total was 140 against
#  a real 233. A curve that ends on the wrong number is worse than no
#  curve, because the last point is the one somebody reads.
#
#  So the SHAPE comes from the movements, which is what they are good
#  for, and the curve is scaled so its last point is exactly what the
#  register says they are holding today. The daily in/out bars are raw
#  counts and are not scaled at all.
#
#  Returns [] when there is nothing to draw - no movements, or nothing
#  on hire - and the report simply leaves the chart out.
#  ---------------------------------------------------------------------
def shut_window():
    """(Flame Off, planned end). Pulled out of _curve on 5 Aug 2026 so
    the four figures under the chart can be worked out even when there
    is no chart - Cleanaway's movements net to zero by today, so the
    curve cannot be drawn, but "8 on hire, 13 days to go" is still true
    and still worth saying."""
    try:
        import shutdown_day as SD
        flame = SD.FLAME_OFF
    except Exception:
        flame = dt.date(2026, 7, 24)
    try:
        import io as _io, os as _os
        p = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)),
                          'shutdown_end.txt')
        end = dt.datetime.strptime(
            _io.open(p, encoding='utf-8').read().strip(), '%Y-%m-%d').date()
    except Exception:
        end = flame + dt.timedelta(days=18)
    return flame, end


def _curve(moves, still_now, today):
    flame, end = shut_window()
    if not moves or end <= flame:
        return []

    out = defaultdict(int)
    back = defaultdict(int)
    for m in moves:
        if m.get("start"):
            out[m["start"]] += 1
        if m.get("end"):
            back[m["end"]] += 1

    days = [flame + dt.timedelta(days=i)
            for i in range((end - flame).days + 1)]
    held, raw = 0, []
    for d in days:
        held += out.get(d, 0) - back.get(d, 0)
        raw.append(max(0, held))

    past = [i for i, d in enumerate(days) if d <= today]
    if not past:
        return []
    last = raw[past[-1]] or 0
    scale = (float(still_now) / last) if last else 0.0

    rows = []
    for i, d in enumerate(days):
        rows.append({
            "d": d,
            "future": d > today,
            "held": (raw[i] * scale) if (scale and d <= today) else None,
            "out": out.get(d, 0),
            "back": back.get(d, 0),
        })
    return rows


def build(models, today=None, window=None, hirer_id_for=None,
          exclude_company=None, movements=None, damage=None):
    """models: the existing {company: model} dict from the on-hire report.

    Returns {"period": (start, end), "companies": [company_model, ...]}
    with every company card equal to the sum of its hirers'."""
    today = today or dt.date.today()
    hire, cons, period = movements or load_movements()
    damage = damage if damage is not None else load_damage()
    w0, w1 = window or period
    if not w0:
        starts = [m["start"] for m in hire if m["start"]]
        w0 = min(starts) if starts else today
    if not w1:
        w1 = today

    def in_win(d):
        return bool(d) and w0 <= d <= w1

    def norm_co(s):
        return re.sub(r"[^A-Z0-9]", "", _txt(s).upper())

    skip = {norm_co(c) for c in (exclude_company or [])}

    # ---- index the on-hire register by (company, PERSON KEY) -----------
    #  The two sources spell people differently - the on-hire register has
    #  "Bradley - Logiudice", the transaction export has "Bradley
    #  Logiudice". Keyed on the raw name they become two people, one with
    #  all the gear and no movements, one with all the movements and no
    #  gear. _pkey() strips it back to letters so they merge, and the
    #  tidied register spelling is what shows on the page.
    #  (A. Fisher, 25 Jul 2026 - found in build)
    on_hire = defaultdict(list)
    display_of = {}
    names = {}

    def keep_name(k, raw):
        n = _clean_name(raw)
        if n and (k not in names or len(n) > len(names[k])):
            names[k] = n

    for m in models.values():
        if norm_co(m["display"]) in skip:
            continue
        for r in m["rows"]:
            k = _pkey(r.get("hirer"))
            on_hire[(m["display"], k)].append(r)
            keep_name(k, r.get("hirer"))
        display_of[norm_co(m["company"])] = m["display"]
        display_of[norm_co(m["display"])] = m["display"]

    def disp(raw):
        return display_of.get(norm_co(raw), _txt(raw) or "(not recorded)")

    # ---- index movements by (company, PERSON KEY) ---------------------
    mv = defaultdict(list)
    for m in hire:
        if norm_co(m["company"]) in skip:
            continue
        if not (in_win(m["start"]) or in_win(m["end"])):
            continue
        k = _pkey(m["hirer"])
        mv[(disp(m["company"]), k)].append(m)
        keep_name(k, m["hirer"])

    cv = defaultdict(list)
    for c in cons:
        if norm_co(c["company"]) in skip:
            continue
        if c["start"] and not in_win(c["start"]):
            continue
        k = _pkey(c["hirer"])
        cv[(disp(c["company"]), k)].append(c)
        keep_name(k, c["hirer"])

    dmg = defaultdict(list)
    model_keys = {norm_co(c) for c in models}
    for d in damage:
        k = norm_co(d["company"])
        if k and k not in model_keys:
            #  The Charge Reporter logs short client names ("Cement
            #  Australia"); the hire register carries the full entity
            #  ("Cement Australia Holdings Pty Ltd"). A damage against
            #  the short name used to fall off this report entirely.
            #  One unambiguous prefix match bridges it; anything
            #  ambiguous stays put rather than guessing.
            near = [m for m in model_keys
                    if m.startswith(k) or k.startswith(m)]
            if len(near) == 1:
                k = near[0]
        dmg[k].append(d)

    # ---- every person we know anything about --------------------------
    keys = set(on_hire) | set(mv) | set(cv)
    by_company = defaultdict(list)
    for co, k in keys:
        by_company[co].append(k)

    companies = []
    for co in sorted(by_company, key=str.upper):
        hirers = []
        for k in sorted(set(by_company[co])):
            hirers.append(_hirer(co, names.get(k, k), on_hire.get((co, k), []),
                                 mv.get((co, k), []),
                                 cv.get((co, k), []),
                                 dmg.get(norm_co(co), []),
                                 today, w0, in_win, hirer_id_for))
        hirers.sort(key=lambda h: (-h["cards"]["still"],
                                   -h["cards"]["issued"], h["name"].upper()))
        for i, h in enumerate(hirers, 1):
            h["n"] = i
            h["of"] = len(hirers)
        c_model = _company(co, hirers, dmg.get(norm_co(co), []), today)
        #  THE SHUTDOWN CURVE - what this company took, gave back, and is
        #  still holding, day by day. (Andrew, 5 Aug 2026, on seeing the
        #  mockup: "love option a".)
        c_model["curve"] = _curve(
            [m for k in set(by_company[co]) for m in mv.get((co, k), [])],
            c_model["cards"]["still"], today)
        companies.append(c_model)

    companies.sort(key=lambda c: (-c["cards"]["still"], c["display"].upper()))
    return {"period": (w0, w1), "companies": companies,
            "today": today, "damage": damage}


def _hirer(company, name, rows, moves, cons, dmg_all, today, w0, in_win,
           hirer_id_for):
    """One person. Everything on their page is computed here, once."""
    c = _blank_cards()
    c["still"] = len(rows)
    for m in moves:
        if in_win(m["start"]):
            c["issued"] += 1
        if in_win(m["end"]):
            c["returned"] += 1
            if m["start"] == m["end"]:
                c["same"] += 1
            elif m["start"] and m["start"] < w0:
                # previously outstanding equipment, brought back this period
                c["recovered"] += 1

    #  NOT RETURNED SAME DAY - only gear that was actually due back that
    #  day. The master's RETURN_REQUIREMENT column is what says so: gas
    #  monitors, radios and battery gear. Authorised long-term hire is
    #  never counted late, exactly as A. Fisher specified.
    daily = [r for r in rows
             if EC.flags(r.get("asset"), r.get("description"))["ret"]]
    c["not_same"] = sum(1 for r in daily
                        if r.get("days") is not None and r["days"] >= 1)

    mix = {k: 0 for k in MIX_ORDER}
    comp = {"rigging": 0, "electrical": 0, "logbook": 0, "ret": 0}
    exposure = 0.0
    for r in rows:
        mix[classify(r.get("description"), r.get("asset"),
                     _txt(r.get("storage_unit")))] += 1
        f = EC.flags(r.get("asset"), r.get("description"))
        for k in ("rigging", "electrical", "logbook"):
            if f[k]:
                comp[k] += 1
        if f["ret"]:
            comp["ret"] += 1
        if r.get("repl"):
            exposure += r["repl"]

    lines = sorted(cons, key=lambda x: (x["start"] or dt.date.min,
                                        x["desc"]))
    key = _pkey(name)
    mine = [d for d in dmg_all if _pkey(d["person"]) == key]

    #  HOW MANY TIMES THEY CAME OVER. Every item is scanned separately, so
    #  the row count is items, not trips. A trip to the counter is one
    #  date and time - Bradley's 120 items on 20 Jul were 15 separate
    #  visits. Both numbers are real and both are worth showing.
    #  (A. Fisher, 25 Jul 2026)
    stamps = set()
    dates = set()
    for m in list(moves) + list(lines):
        d = m.get("start")
        if not d:
            continue
        dates.add(d)
        stamps.add((d, _txt(m.get("start_time"))))
    visits = len(stamps)
    days_over = len(dates)

    return {
        "name": name or "(not recorded)",
        "hirer_id": _hid(rows) or (hirer_id_for(name) if hirer_id_for else ""),
        "company": company,
        "cards": c,
        "mix": mix,
        "compliance": comp,
        "exposure": exposure,
        "oldest": max([r["days"] for r in rows
                       if r.get("days") is not None] or [0]),
        "rows": sorted(rows, key=lambda r: -(r.get("days") or 0)),
        "visits": visits,
        "days_over": days_over,
        "movements": len(moves) + len(lines),
        "cons": {"txns": len(lines),
                 "units": sum(x["qty"] for x in lines),
                 "types": len({x["barcode"] or x["desc"] for x in lines}),
                 "lines": lines},
        "damage": mine,
        "actions": _actions(c, comp, mine),
    }


def _company(display, hirers, dmg, today):
    """The company page. Every card is the sum of the hirer pages - there
    is deliberately no independent calculation here to disagree with."""
    c = _blank_cards()
    mix = {k: 0 for k in MIX_ORDER}
    comp = {"rigging": 0, "electrical": 0, "logbook": 0, "ret": 0}
    cons = {"txns": 0, "units": 0.0, "types": set(), "by_hirer": [],
            "top": {}}
    exposure = 0.0
    age = {"1-7": 0, "8-30": 0, "30+": 0}
    visits = movements = 0
    using_store = 0
    for h in hirers:
        visits += h["visits"]
        movements += h["movements"]
        if h["movements"]:
            using_store += 1
        for k in c:
            c[k] += h["cards"][k]
        for k in mix:
            mix[k] += h["mix"][k]
        for k in comp:
            comp[k] += h["compliance"][k]
        exposure += h["exposure"]
        cons["txns"] += h["cons"]["txns"]
        cons["units"] += h["cons"]["units"]
        for line in h["cons"]["lines"]:
            cons["types"].add(line["barcode"] or line["desc"])
            t = cons["top"].setdefault(line["desc"],
                                       {"desc": line["desc"], "qty": 0.0,
                                        "material": line["barcode"]})
            t["qty"] += line["qty"]
        if h["cons"]["units"]:
            cons["by_hirer"].append({"name": h["name"],
                                     "units": h["cons"]["units"],
                                     "txns": h["cons"]["txns"]})
        for r in h["rows"]:
            d = r.get("days")
            if d is None:
                continue
            age["1-7" if d <= 7 else "8-30" if d <= 30 else "30+"] += 1

    cons["types"] = len(cons["types"])
    cons["top"] = sorted(cons["top"].values(), key=lambda t: -t["qty"])
    cons["by_hirer"].sort(key=lambda x: -x["units"])

    d_open = [d for d in dmg if d["status"].lower() not in
              ("closed", "invoiced", "written off", "declined")]
    damage = {
        "reported": len(dmg),
        "open": len(d_open),
        "closed": len(dmg) - len(d_open),
        "oos": sum(1 for d in dmg if d["oos"].lower().startswith("y")),
        "offhired": sum(1 for d in dmg if d["offhired"].lower().startswith("y")),
        "exposure": sum(d["repair"] or d["exposure"] for d in dmg),
        "items": dmg,
    }
    return {
        "display": display, "hirers": hirers, "cards": c, "mix": mix,
        "compliance": comp, "exposure": exposure, "age": age,
        "cons": cons, "damage": damage,
        "oldest": max([h["oldest"] for h in hirers] or [0]),
        "n_hirers": len(hirers),
        "visits": visits, "movements": movements,
        "using_store": using_store,
        "actions": [a for h in hirers for a in h["actions"]],
    }


def _actions(cards, comp, damage):
    """Genuine exceptions only - never a list of everything."""
    out = []
    if cards["not_same"]:
        out.append(("HIGH", "{} item{} due back today not yet returned".format(
            cards["not_same"], "" if cards["not_same"] == 1 else "s"),
            "Return to the tool store this shift"))
    for d in damage:
        if d["status"].lower() not in ("closed", "invoiced"):
            out.append(("HIGH", "Damage reported - {} ({})".format(
                d["desc"], d["cid"]), "Awaiting inspection"))
    if comp["logbook"]:
        out.append(("MEDIUM", "{} plant item{} on hire".format(
            comp["logbook"], "" if comp["logbook"] == 1 else "s"),
            "Daily pre-start and logbook entry each shift"))
    return out


def _pkey(name):
    """Everything that is just a person's name, stripped to letters, so
    "Bradley - Logiudice" and "Bradley Logiudice" are the same person."""
    return re.sub(r"[^a-z]", "", _txt(name).lower().split("(")[0])


def _clean_name(raw):
    """Display spelling: drop the stray hyphens the register carries and
    any trailing "(Company)" the damage sheet adds."""
    n = re.sub(r"\s*\(.*?\)\s*", " ", _txt(raw))
    n = re.sub(r"\s*-\s*", " ", n)
    return re.sub(r"\s+", " ", n).strip()


def _hid(rows):
    for r in rows:
        if r.get("hirer_id"):
            return _txt(r["hirer_id"])
    return ""


# ---------------------------------------------------------------------
#  4. RECONCILE - prove it, every run
# ---------------------------------------------------------------------
def reconcile(model, verbose=True):
    """Company totals must equal the sum of their hirer pages. Because the
    company is BUILT as that sum this cannot fail - which is the point.
    It runs anyway, so that if anyone ever adds an independent company
    calculation the next run says so instead of quietly shipping a report
    that does not add up."""
    bad = []
    for co in model["companies"]:
        for k, v in co["cards"].items():
            s = sum(h["cards"][k] for h in co["hirers"])
            if s != v:
                bad.append((co["display"], k, v, s))
        for k, v in co["compliance"].items():
            s = sum(h["compliance"][k] for h in co["hirers"])
            if s != v:
                bad.append((co["display"], "compliance." + k, v, s))
    if verbose:
        if bad:
            print("  RECONCILE   : FAILED - company totals do not match "
                  "their hirer pages")
            for co, k, v, s in bad:
                print("     {} {}: company {} vs hirers {}".format(co, k, v, s))
        else:
            n = sum(len(c["hirers"]) for c in model["companies"])
            print("  Reconcile   : PASS - {} companies, {} hirer pages, "
                  "every total ties".format(len(model["companies"]), n))
    return bad
