#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | ON-HIRE WORKBOOK - every tab, built from today's exports
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  The shutdown on-hire workbook, rebuilt from the SiteIQ exports for
#  whichever job this computer is running (57_SWITCH_JOB.bat).
#
#  Twelve tabs, and every one of them works:
#
#    Cover                    what this is, what fed it, the headlines
#    Company Summary          per company: tooling, consumables, total
#    Detailed Onhire          every item out right now, who has it
#    Tooling Transactions     every tooling issue and return
#    Tooling Utilisation      per item type: turns, rating, what to do
#    Coates Tooling           the Coates-owned fleet on site
#    Consumable Transactions  every consumable issued
#    Consumables Available    what's left on the shelf
#    Coates Stock             the fleet with its on-hire dates
#    Consumable Utilisation   per line: usage, rating, what to do
#    Coates Labour            the shift roster - CARRIED OVER, not touched
#    Cost Breakdown           the day-by-day cost - CARRIED OVER, not touched
#
#  Two of those tabs are Andrew's own typing - the labour roster and
#  the cost breakdown. They are read out of the existing workbook and
#  written back exactly as they were, formulas and all. Nothing typed
#  by hand is ever recalculated, overwritten or "tidied".
#
#  The existing workbook is never written to. It carries macros, a
#  logo and live query connections that Python cannot round-trip
#  safely, so it stays exactly as it is and a clean dated workbook is
#  built alongside it.
#
#  It also writes the report as an EMAIL - the whole thing in the body,
#  Coates styling, the workbook on the paperclip, and NO COSTS. Not the
#  labour, not the cost breakdown, not a rate anywhere. That is the
#  operational picture: what is out, who has it, what is working and
#  what is not. The money stays in the workbook. It opens as an Outlook
#  DRAFT and waits - nothing sends by itself.
#
#  Output:  <job reports>\<today>\<Job> On-Hire Workbook - <date>.xlsx
#      and  <job folder>\<Job>_Onhire_Workbook_LATEST.xlsx
#      and  <job reports>\<today>\<Job> On-Hire Report - <date>.eml
# =====================================================================

import datetime as dt
import glob
import os
import shutil
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

#  This script runs in two homes and must work in both:
#
#    1. inside the full suite, where site_config knows which job the
#       computer is on and report_paths knows where reports go, and
#    2. on its own, in a folder with nothing but this file, the button
#       and the exports - the kit that goes on the other computers.
#
#  Standing alone it works the job out from the exports themselves.
#  Every SiteIQ export carries the PROJECT SCHEDULE it came from on its
#  REFERENCE_INFO sheet, so the kit names itself off the data and there
#  is nothing to set up. Drop the exports in, press the button, done.
try:
    import report_paths
    import site_config
    STANDALONE = False
except ImportError:
    report_paths = site_config = None
    STANDALONE = True

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is not installed. Run any numbered button once and it "
          "will put it on for you, or: pip install openpyxl")
    sys.exit(1)

# ---- the Coates look -------------------------------------------------
ORANGE = "FFF26222"
ORANGE_ALT = "FFFA4600"        # the shade the existing workbook uses
NEAR_BLACK = "FF1D1D1B"
WHITE = "FFFFFFFF"
GREY = "FF8B9099"
BAND_ROW_H = 26

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=WHITE)
HDR_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
BODY_FONT = Font(name="Calibri", size=11)
TITLE_FILL = PatternFill("solid", fgColor=NEAR_BLACK)
HDR_FILL = PatternFill("solid", fgColor=ORANGE_ALT)
THIN = Side(style="thin", color="FFD9D9D9")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATE_FMT = "dd/mm/yyyy"
TIME_FMT = "hh:mm"
PCT_FMT = "0.0%"
MONEY_FMT = "$#,##0.00"
INT_FMT = "#,##0"

#  Hand-typed tabs. Read out of the old workbook, written back as-is.
CARRIED_OVER = ("Coates Labour", "Cost Breakdown")


# =====================================================================
#  Standing on its own - the kit that goes on the other computers
# =====================================================================
class LooseSite(object):
    """The job, worked out from the exports sitting in the folder.

    No config, no setup, nothing to switch. Whatever exports are here
    decide what the workbook is called and who it is for. Drop Weipa's
    in and it is a Weipa workbook; drop Gladstone's in and it is a
    Gladstone one. JOB.txt can override any of the names if the
    SiteIQ project string isn't what you'd put in front of a customer.
    """

    def __init__(self, base):
        self.base = base
        self._d = self._read_job_txt(base)
        if not self._d.get("project_schedule"):
            self._d["project_schedule"] = self._sniff(base)
        stamp = self._d["project_schedule"]
        self._d.setdefault("job", stamp or "Shutdown")
        self._d.setdefault("customer", "")
        self._d.setdefault("location", "")
        self._d.setdefault("short", (self._d.get("customer")
                                     or self._d["job"]).strip())
        self._d.setdefault("author", "Andrew Fisher")

    @staticmethod
    def _read_job_txt(base):
        """Optional. One 'name: value' per line - job, customer,
        location, short. Anything missing is worked out instead."""
        out = {}
        p = os.path.join(base, "JOB.txt")
        try:
            with open(p, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#") or ":" not in line:
                        continue
                    k, v = line.split(":", 1)
                    if v.strip():
                        out[k.strip().lower().replace(" ", "_")] = v.strip()
        except OSError:
            pass
        return out

    @staticmethod
    def _sniff(base):
        """The PROJECT SCHEDULE stamped on whichever export is here."""
        for d in (os.path.join(base, "Data_SiteIQ"), base):
            for stem in ("RENTAL_STOCK", "TRANSACTIONS", "SALES_STOCK",
                         "STOCKTAKE"):
                for p in sorted(glob.glob(os.path.join(d, stem + "*.xlsx"))):
                    if os.path.basename(p).startswith("~$"):
                        continue
                    got = read_project_schedule(p)
                    if got:
                        return got
        return ""

    def __getattr__(self, name):
        if name.startswith("_"):
            raise AttributeError(name)
        return self._d.get(name, "")

    @property
    def header_line(self):
        bits = [b for b in (self._d.get("customer"), self._d.get("job"),
                            self._d.get("location")) if b]
        return " · ".join(bits) or "Shutdown tool store"

    @property
    def exports(self):
        return ["RENTAL_STOCK", "SALES_STOCK", "TRANSACTIONS", "STOCKTAKE"]

    @property
    def reports_dirname(self):
        return "Reports"

    def data_dirs(self, base=None):
        base = base or self.base
        return [os.path.join(base, "Data_SiteIQ"), base]

    def workbook(self, base=None):
        """Any .xlsm sitting in the kit - that's where the hand-typed
        Labour and Cost tabs are read from."""
        base = base or self.base
        hits = [p for p in glob.glob(os.path.join(base, "*.xlsm"))
                if not os.path.basename(p).startswith("~$")]
        return max(hits, key=os.path.getmtime) if hits else None


def read_project_schedule(path):
    """The SiteIQ project an export came from, off its REFERENCE_INFO
    sheet. Empty if it doesn't carry one - plenty of older pulls don't."""
    if site_config is not None:
        return site_config.export_project_schedule(path)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return ""
    try:
        if "REFERENCE_INFO" not in wb.sheetnames:
            return ""
        rows = list(wb["REFERENCE_INFO"].iter_rows(min_row=1, max_row=2,
                                                   values_only=True))
        if len(rows) < 2:
            return ""
        head = [str(c or "").strip().upper() for c in rows[0]]
        if "PROJECT SCHEDULE" not in head:
            return ""
        return str(rows[1][head.index("PROJECT SCHEDULE")] or "").strip()
    except Exception:
        return ""
    finally:
        try:
            wb.close()
        except Exception:
            pass


def the_site(base):
    """The job, however this copy is running."""
    if site_config is not None:
        return site_config.site(base=base)
    return LooseSite(base)


def find_export(base, pattern, s):
    """Newest file matching the pattern, across the folders this job
    keeps its exports in."""
    if report_paths is not None:
        return report_paths.find_export(base, pattern)
    hits = []
    for d in s.data_dirs(base):
        hits += [p for p in glob.glob(os.path.join(d, pattern))
                 if not os.path.basename(p).startswith("~$")]
    return max(hits, key=os.path.getmtime) if hits else None


def out_dirs(base, when, s):
    if report_paths is not None:
        return report_paths.out_dirs(base, when)
    day = os.path.join(base, s.reports_dirname, when.strftime("%Y-%m-%d"))
    os.makedirs(day, exist_ok=True)
    return {"day": day}


def note_sources(folder, entries):
    if report_paths is not None:
        return report_paths.note_sources(folder, entries)
    lines = ["DATA SOURCES FOR THIS WORKBOOK",
             "Generated " + dt.datetime.now().strftime("%d %b %Y - %H:%M"),
             "-" * 60]
    for p in entries:
        stamp = dt.datetime.fromtimestamp(os.path.getmtime(p))
        lines.append("{:<52} refreshed {}".format(
            os.path.basename(p), stamp.strftime("%d %b %Y - %H:%M")))
    out = os.path.join(folder, "DATA_SOURCES.txt")
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    return out


def wrong_job(path, base, s):
    """(ok, message). Standing alone there is no 'live job' to check
    against, so the test is that the exports AGREE WITH EACH OTHER -
    which catches the real mistake, refreshing off a mix of two jobs."""
    if site_config is not None:
        return site_config.belongs_to_live_job(path, base)
    want = (s.project_schedule or "").strip()
    got = read_project_schedule(path)
    if not want or not got or got.lower() == want.lower():
        return True, ""
    return False, ("{} is a {} export but the others in this folder are {}. "
                   "One job's exports at a time - take the odd one out and "
                   "run it again.".format(os.path.basename(path), got, want))


# =====================================================================
#  Reading the exports
# =====================================================================
def _rows(path, sheet):
    """A sheet as a list of dicts, keyed on its header row. Blank rows
    dropped - SiteIQ pads its exports out with them."""
    if not path or not os.path.isfile(path):
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return []
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        try:
            head = next(it)
        except StopIteration:
            return []
        hdr = [str(c or "").strip() for c in head]
        out = []
        for r in it:
            if not any(c not in (None, "") for c in r):
                continue
            out.append(dict(zip(hdr, r)))
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _s(v):
    """A cell as clean text."""
    return "" if v is None else str(v).strip()


def _n(v):
    """A cell as a number. SiteIQ ships numbers as text more often than
    not, so never trust the type."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except ValueError:
        return 0.0


def _date(v):
    """A cell as a date. SiteIQ mixes real dates with dd/mm/yyyy text
    inside the same column."""
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    t = _s(v)
    if not t:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S",
                "%d/%m/%Y %I:%M %p", "%Y-%m-%d %H:%M:%S"):
        try:
            return dt.datetime.strptime(t, fmt).date()
        except ValueError:
            continue
    return None


# =====================================================================
#  Usage ratings - the rules, written down
# =====================================================================
#  Tooling is rated on TURNS: how many times each unit went out.
#  One item that went out five times is working; five items that went
#  out once each are sitting there. Utilisation on its own can't tell
#  those apart, which is why the old sheet had a line rated "Low Use"
#  on 36 transactions and another rated "Good Use" on one.
def tooling_rating(transactions, total_qty, utilisation):
    turns = (transactions / total_qty) if total_qty else 0.0
    if turns <= 0:
        return "No Use", "Review / Reduce"
    if turns < 1:
        return "Low Use", "Keep Stock"
    if turns < 2:
        if utilisation >= 0.8:
            return "Good Use", "Monitor / Increase"
        return "Good Use", "Keep Stock"
    return "High Demand", "Increase Stock"


#  Consumables are rated on how much of the position has gone out.
def consumable_rating(usage):
    if usage <= 0:
        return "No Use", "Review / Reduce"
    if usage < 0.4:
        return "Low Use", "Keep Stock"
    if usage < 0.8:
        return "Good Use", "Keep Stock"
    return "High Demand", "Increase Stock"


# =====================================================================
#  Sheet writing
# =====================================================================
def band(ws, title, ncols):
    """The orange-on-black title strip across the top of every tab."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ncols, 1))
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = BAND_ROW_H


def sheet(wb, name, title, columns, rows, widths=None, formats=None):
    """One data tab: title strip, header row, the rows, frozen panes and
    a filter. `formats` maps column index (0-based) to a number format."""
    ws = wb.create_sheet(name)
    band(ws, title, len(columns))
    for i, col in enumerate(columns, start=1):
        c = ws.cell(row=2, column=i, value=col)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    for r, row in enumerate(rows, start=3):
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = BODY_FONT
            c.border = CELL_BORDER
            fmt = (formats or {}).get(i - 1)
            if fmt:
                c.number_format = fmt

    for i, col in enumerate(columns, start=1):
        letter = get_column_letter(i)
        if widths and (i - 1) in widths:
            ws.column_dimensions[letter].width = widths[i - 1]
        else:
            longest = max([len(str(col))]
                          + [len(_s(r[i - 1])) for r in rows[:400]] or [10])
            ws.column_dimensions[letter].width = min(max(longest + 2, 11), 60)

    ws.freeze_panes = "A3"
    if rows:
        ws.auto_filter.ref = "A2:{}{}".format(
            get_column_letter(len(columns)), len(rows) + 2)
    #  An empty tab still says so out loud rather than looking broken.
    if not rows:
        c = ws.cell(row=3, column=1,
                    value="Nothing to show for this pull - no rows in the "
                          "export matched this tab.")
        c.font = Font(name="Calibri", size=11, italic=True, color=GREY)
    return ws


# =====================================================================
#  The tabs
# =====================================================================
def build(base=HERE, when=None):
    s = the_site(base)
    when = when or dt.date.today()
    asat = dt.datetime.now()

    def export(stem):
        return find_export(base, stem + "*.xlsx", s)

    p_rental = export("RENTAL_STOCK")
    p_sales = export("SALES_STOCK")
    p_trans = export("TRANSACTIONS")
    p_stock = export("STOCKTAKE")
    sources = [p for p in (p_rental, p_sales, p_trans, p_stock) if p]

    missing = [n for n, p in (("RENTAL_STOCK", p_rental),
                              ("SALES_STOCK", p_sales),
                              ("TRANSACTIONS", p_trans)) if not p]
    if missing:
        print("  STOP | These exports are not in {}:".format(
            os.path.relpath(s.data_dirs(base)[0], base)))
        for m in missing:
            print("       - " + m)
        print("       Pull them out of SiteIQ, save them there, run again.")
        return 1

    #  Right files, right job?
    for p in sources:
        ok, why = wrong_job(p, base, s)
        if not ok:
            print("  STOP | " + why)
            return 1

    rental = _rows(p_rental, "RENTAL_STOCK")
    sales = _rows(p_sales, "SALES_STOCK")
    stocktake = _rows(p_stock, "STOCKTAKE") if p_stock else []

    #  SiteIQ splits transactions across two sheets and which one a job
    #  fills depends on whether charging is switched on. Weipa's land in
    #  TRANSACTION_WITHOUT_CHARGES, K2's in TRANSACTION_CHARGES. Read
    #  both and de-duplicate on the transaction number, so the tabs fill
    #  either way instead of coming out empty on one job.
    tw = []
    seen = set()
    for sheet_name in ("TRANSACTION_CHARGES", "TRANSACTION_WITHOUT_CHARGES"):
        for r in _rows(p_trans, sheet_name):
            tid = _s(r.get("TRANSACTION_ID"))
            if tid and tid in seen:
                continue
            if tid:
                seen.add(tid)
            tw.append(r)

    title = "Coates    |  {}  |  SHUTDOWN TOOLSTORE ON-HIRE REPORT" \
            "        Powered by SITEIQ".format(s.short or s.customer)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---- the pieces every tab is cut from -----------------------------
    onhire = [r for r in rental if _s(r.get("ITEM_STATUS")) == "On Hire"]
    coates_owned = [r for r in rental
                    if _s(r.get("OWNER")).upper() == "COATES"]
    tooling_tx = [r for r in tw
                  if _s(r.get("PRODUCT_CATEGORY")).lower() != "consumable"]
    consumable_tx = [r for r in sales if _n(r.get("SALES_QUANTITY")) > 0]
    consumable_stock = [r for r in sales if not _s(r.get("COMPANY_NAME"))]

    tabs = []                      # (name, rows) for the cover index

    # ---- 1. Company Summary -------------------------------------------
    companies = {}
    for r in tw:
        co = _s(r.get("EMPLOYER_NAME")) or "(not named)"
        kind = ("CONSUMABLES"
                if _s(r.get("PRODUCT_CATEGORY")).lower() == "consumable"
                else "TOOLING")
        d = companies.setdefault(co, {
            "TOOLING": {"tx": 0, "qty": 0.0, "types": set(), "open": 0},
            "CONSUMABLES": {"tx": 0, "qty": 0.0, "types": set(), "open": 0}})
        a = d[kind]
        a["tx"] += 1
        a["qty"] += _n(r.get("QUANTITY"))
        a["types"].add(_s(r.get("SKU/ITEM DESCRIPTION")))
        if kind == "TOOLING" and not _s(r.get("TRAN_END_DATE")):
            a["open"] += 1

    live_onhire = {}
    for r in onhire:
        live_onhire[_s(r.get("COMPANY_NAME"))] = \
            live_onhire.get(_s(r.get("COMPANY_NAME")), 0) + 1

    summary_rows = []
    for co in sorted(companies):
        d = companies[co]
        tot = {"tx": 0, "qty": 0.0, "types": 0, "onh": 0, "open": 0}
        for kind in ("TOOLING", "CONSUMABLES"):
            a = d[kind]
            #  Consumables are sold, not returned - nothing is ever
            #  "still out", so those two columns are zero by definition.
            onh = live_onhire.get(co, 0) if kind == "TOOLING" else 0
            opn = a["open"] if kind == "TOOLING" else 0
            summary_rows.append([co, kind, a["tx"], int(a["qty"]),
                                 len([t for t in a["types"] if t]), onh, opn])
            tot["tx"] += a["tx"]
            tot["qty"] += a["qty"]
            tot["types"] += len([t for t in a["types"] if t])
            tot["onh"] += onh
            tot["open"] += opn
        summary_rows.append([co, "TOTAL", tot["tx"], int(tot["qty"]),
                             tot["types"], tot["onh"], tot["open"]])

    sheet(wb, "Company Summary", title,
          ["Company", "Type", "Transactions", "Qty", "Item Types",
           "Currently Onhire Qty", "Items Not Returned"],
          summary_rows,
          widths={0: 26, 1: 15, 2: 14, 3: 10, 4: 13, 5: 21, 6: 20},
          formats={2: INT_FMT, 3: INT_FMT, 4: INT_FMT, 5: INT_FMT, 6: INT_FMT})
    tabs.append(("Company Summary", len(summary_rows),
                 "Per company: what they took, what they still have"))

    # ---- 2. Detailed Onhire -------------------------------------------
    det = sorted(([_s(r.get("COMPANY_NAME")), _s(r.get("HIRER_NAME")),
                   _s(r.get("ITEM_NUMBER")), _s(r.get("ITEM_DESCRIPTION")),
                   _s(r.get("ITEM_STATUS")), _date(r.get("ON_HIRE_DATE")),
                   _s(r.get("ON_HIRE_TIME"))] for r in onhire),
                 key=lambda x: (x[0], x[1], x[3]))
    sheet(wb, "Detailed Onhire", title,
          ["COMPANY_NAME", "HIRER_NAME", "ITEM_NUMBER", "ITEM_DESCRIPTION",
           "ITEM_STATUS", "ON_HIRE_DATE", "ON_HIRE_TIME"],
          det, widths={0: 24, 1: 22, 2: 26, 3: 50, 4: 18, 5: 14, 6: 14},
          formats={5: DATE_FMT})
    tabs.append(("Detailed Onhire", len(det),
                 "Every item out right now and who has it"))

    # ---- 3. Tooling Transactions --------------------------------------
    tt = sorted(([_s(r.get("EMPLOYER_NAME")), _s(r.get("HIRER_NAME")),
                  _s(r.get("SKU/ITEM_NUMBER")),
                  _s(r.get("SKU/ITEM DESCRIPTION")),
                  _date(r.get("TRAN_START_DATE")), _s(r.get("TRAN_START_TIME")),
                  _date(r.get("TRAN_END_DATE")), _s(r.get("TRAN_END_TIME")),
                  _s(r.get("TRANSACTION_ID"))] for r in tooling_tx),
                 key=lambda x: (x[4] or dt.date.min, x[0]))
    sheet(wb, "Tooling Transactions", title,
          ["EMPLOYER_NAME", "HIRER_NAME", "SKU/ITEM_NUMBER",
           "SKU/ITEM DESCRIPTION", "TRAN_START_DATE", "TRAN_START_TIME",
           "TRAN_END_DATE", "TRAN_END_TIME", "TRANSACTION_ID"],
          tt, widths={0: 24, 1: 22, 2: 26, 3: 50, 4: 15, 5: 14, 6: 15,
                      7: 14, 8: 16},
          formats={4: DATE_FMT, 6: DATE_FMT})
    tabs.append(("Tooling Transactions", len(tt),
                 "Every tooling issue and return this period"))

    # ---- 4. Tooling Utilisation ---------------------------------------
    fleet = {}
    for r in rental:
        d = fleet.setdefault(_s(r.get("ITEM_DESCRIPTION")),
                             {"total": 0, "onhire": 0, "avail": 0})
        d["total"] += 1
        st = _s(r.get("ITEM_STATUS"))
        if st == "On Hire":
            d["onhire"] += 1
        elif st == "Available for Hire":
            d["avail"] += 1

    used = {}
    for r in tooling_tx:
        d = used.setdefault(_s(r.get("SKU/ITEM DESCRIPTION")),
                            {"tx": 0, "qty": 0.0, "co": set(), "hi": set(),
                             "first": None, "last": None})
        d["tx"] += 1
        d["qty"] += _n(r.get("QUANTITY"))
        if _s(r.get("EMPLOYER_NAME")):
            d["co"].add(_s(r.get("EMPLOYER_NAME")))
        if _s(r.get("HIRER_NAME")):
            d["hi"].add(_s(r.get("HIRER_NAME")))
        day = _date(r.get("TRAN_START_DATE"))
        if day:
            d["first"] = day if d["first"] is None else min(d["first"], day)
            d["last"] = day if d["last"] is None else max(d["last"], day)

    tu = []
    for desc in sorted(fleet):
        f = fleet[desc]
        u = used.get(desc, {"tx": 0, "qty": 0.0, "co": set(), "hi": set(),
                            "first": None, "last": None})
        util = (f["onhire"] / f["total"]) if f["total"] else 0.0
        rating, rec = tooling_rating(u["tx"], f["total"], util)
        tu.append([desc, f["total"], f["onhire"], f["avail"], util,
                   u["tx"], int(u["qty"]), len(u["co"]), len(u["hi"]),
                   u["first"], u["last"], rating, rec])
    sheet(wb, "Tooling Utilisation", title,
          ["ITEM_DESCRIPTION", "Total Qty", "Qty On Hire", "Qty Available",
           "Current Utilisation %", "Total Transactions", "Total Qty Issued",
           "No. of Companies Using Item", "No. of Hirers Using Item",
           "First Used Date", "Last Used Date", "Usage Rating",
           "Recommendation"],
          tu, widths={0: 55, 1: 11, 2: 12, 3: 13, 4: 15, 5: 14, 6: 14,
                      7: 16, 8: 16, 9: 14, 10: 14, 11: 14, 12: 20},
          formats={1: INT_FMT, 2: INT_FMT, 3: INT_FMT, 4: PCT_FMT,
                   5: INT_FMT, 6: INT_FMT, 7: INT_FMT, 8: INT_FMT,
                   9: DATE_FMT, 10: DATE_FMT})
    tabs.append(("Tooling Utilisation", len(tu),
                 "Per item type: turns, rating, what to do about it"))

    # ---- 5. Coates Tooling --------------------------------------------
    ct = sorted(([_s(r.get("ITEM_NUMBER")), _s(r.get("ITEM_DESCRIPTION")),
                  _s(r.get("ITEM_STATUS"))] for r in coates_owned),
                key=lambda x: (x[1], x[0]))
    sheet(wb, "Coates Tooling", title,
          ["ITEM_NUMBER", "ITEM_DESCRIPTION", "ITEM_STATUS"],
          ct, widths={0: 26, 1: 60, 2: 20})
    tabs.append(("Coates Tooling", len(ct),
                 "The Coates-owned fleet on this site"))

    # ---- 6. Consumable Transactions -----------------------------------
    ctx = sorted(([_s(r.get("COMPANY_NAME")), _s(r.get("HIRER")),
                   _s(r.get("SKU_DESCRIPTION")), _n(r.get("SALES_QUANTITY")),
                   _date(r.get("SALES_DATE")), _s(r.get("SALES_TIME"))]
                  for r in consumable_tx),
                 key=lambda x: (x[0], x[2]))
    sheet(wb, "Consumable Transactions", title,
          ["COMPANY_NAME", "HIRER", "SKU_DESCRIPTION", "SALES_QUANTITY",
           "SALES_DATE", "SALES_TIME"],
          ctx, widths={0: 24, 1: 24, 2: 50, 3: 16, 4: 14, 5: 14},
          formats={3: "#,##0.00", 4: DATE_FMT})
    tabs.append(("Consumable Transactions", len(ctx),
                 "Every consumable issued this period"))

    # ---- 7. Consumables Available -------------------------------------
    ca = sorted(([_s(r.get("SKU_NUMBER")), _s(r.get("SKU_DESCRIPTION")),
                  int(_n(r.get("AVAILABLE_QUANTITY")))]
                 for r in consumable_stock), key=lambda x: x[1])
    sheet(wb, "Consumables Available", title,
          ["SKU_NUMBER", "SKU_DESCRIPTION", "AVAILABLE_QUANTITY"],
          ca, widths={0: 20, 1: 50, 2: 22}, formats={2: INT_FMT})
    tabs.append(("Consumables Available", len(ca),
                 "What's left on the shelf"))

    # ---- 8. Coates Stock ----------------------------------------------
    cs = sorted(([_s(r.get("ITEM_NUMBER")), _s(r.get("ITEM_DESCRIPTION")),
                  _s(r.get("ITEM_STATUS")), _date(r.get("ON_HIRE_DATE")),
                  _s(r.get("ON_HIRE_TIME"))] for r in coates_owned),
                key=lambda x: (x[1], x[0]))
    sheet(wb, "Coates Stock", title,
          ["ITEM_NUMBER", "ITEM_DESCRIPTION", "ITEM_STATUS", "ON_HIRE_DATE",
           "ON_HIRE_TIME"],
          cs, widths={0: 26, 1: 60, 2: 20, 3: 15, 4: 14},
          formats={3: DATE_FMT})
    tabs.append(("Coates Stock", len(cs),
                 "The fleet with its on-hire dates"))

    # ---- 9. Consumable Utilisation ------------------------------------
    shelf, sold = {}, {}
    for r in sales:
        desc = _s(r.get("SKU_DESCRIPTION"))
        if not _s(r.get("COMPANY_NAME")):
            shelf[desc] = shelf.get(desc, 0.0) + _n(r.get("AVAILABLE_QUANTITY"))
        q = _n(r.get("SALES_QUANTITY"))
        if q > 0:
            d = sold.setdefault(desc, {"qty": 0.0, "co": set(), "hi": set(),
                                       "first": None, "last": None})
            d["qty"] += q
            if _s(r.get("COMPANY_NAME")):
                d["co"].add(_s(r.get("COMPANY_NAME")))
            if _s(r.get("HIRER")):
                d["hi"].add(_s(r.get("HIRER")))
            day = _date(r.get("SALES_DATE"))
            if day:
                d["first"] = day if d["first"] is None else min(d["first"], day)
                d["last"] = day if d["last"] is None else max(d["last"], day)

    cu = []
    for desc in sorted(set(list(shelf) + list(sold))):
        avail = shelf.get(desc, 0.0)
        d = sold.get(desc, {"qty": 0.0, "co": set(), "hi": set(),
                            "first": None, "last": None})
        position = avail + d["qty"]
        usage = (d["qty"] / position) if position else 0.0
        rating, rec = consumable_rating(usage)
        cu.append([desc, int(avail), int(d["qty"]), int(position), usage,
                   len(d["co"]), len(d["hi"]), d["first"], d["last"],
                   rating, rec])
    sheet(wb, "Consumable Utilisation", title,
          ["SKU_DESCRIPTION", "Qty Available", "Total Sales Qty",
           "Total Stock Position", "Usage %", "No. of Companies Using Item",
           "No. of Hirers Using Item", "First Sold Date", "Last Sold Date",
           "Usage Rating", "Recommendation"],
          cu, widths={0: 50, 1: 14, 2: 16, 3: 19, 4: 12, 5: 16, 6: 16,
                      7: 14, 8: 14, 9: 14, 10: 20},
          formats={1: INT_FMT, 2: INT_FMT, 3: INT_FMT, 4: PCT_FMT,
                   5: INT_FMT, 6: INT_FMT, 7: DATE_FMT, 8: DATE_FMT})
    tabs.append(("Consumable Utilisation", len(cu),
                 "Per line: usage, rating, what to do about it"))

    # ---- 10/11. the two hand-typed tabs, carried over verbatim --------
    carried, labour_total, cost_total = carry_over(wb, s, base, title)
    for name, n in carried:
        tabs.append((name, n, "Typed by hand - carried over untouched"))

    # ---- 12. Cover ----------------------------------------------------
    cover(wb, s, when, asat, tabs, sources, base,
          headline={
              "Items on hire now": len(onhire),
              "Companies with gear out": len([c for c in live_onhire
                                              if live_onhire[c]]),
              "Coates items on site": len(coates_owned),
              "Tooling transactions": len(tooling_tx),
              "Consumable issues": len(ctx),
              "Consumable lines on shelf": len(ca),
              "Items counted at last stocktake": len(stocktake),
              "Labour cost to date": labour_total,
              "Cost breakdown total": cost_total,
          })
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(wb["Cover"])))

    # ---- save ---------------------------------------------------------
    dirs = out_dirs(base, when, s)
    label = (s.short or s.customer or "Shutdown").strip()
    stem = "{} On-Hire Workbook - {}".format(label, when.strftime("%d %b %Y"))
    out = os.path.join(dirs["day"], stem + ".xlsx")
    wb.save(out)

    #  The one to open. Same name every day, always the newest - so the
    #  shortcut on the desktop never goes stale.
    latest_dir = os.path.join(base, os.path.dirname(s.workbook_glob or ""))
    latest = os.path.join(latest_dir, "{}_Onhire_Workbook_LATEST.xlsx".format(
        label.replace(" ", "_")))
    try:
        shutil.copyfile(out, latest)
    except OSError as e:
        print("  WARN | Couldn't refresh the LATEST copy ({}). The dated "
              "one above is fine - it's probably open in Excel.".format(e))
        latest = None

    note_sources(dirs["day"], sources)

    # ---- the email -----------------------------------------------------
    parts = email_parts(s, summary_rows, tu, cu, det, ctx, ca, coates_owned,
                        tooling_tx, live_onhire, sources, when)
    html_path, eml_path = write_email(base, s, when, asat, parts,
                                      latest or out, dirs["day"])

    print("=" * 66)
    print(" COATES | ON-HIRE WORKBOOK - {}".format(s.header_line))
    print("=" * 66)
    for name, n, _why in tabs:
        print("  {:<26} {:>6} rows".format(name, n))
    print("-" * 66)
    print("  Workbook  {}".format(os.path.relpath(out, base)))
    if latest:
        print("  Open this {}".format(os.path.relpath(latest, base)))
    print("  Email     {}".format(os.path.relpath(eml_path, base)))
    print("            double-click it - it opens as a DRAFT. Nothing sends")
    print("            until you press Send yourself.")
    print("  Web page  {}".format(os.path.relpath(html_path, base)))
    print("=" * 66)
    return 0


def email_parts(s, summary_rows, tu, cu, det, ctx, ca, coates_owned,
                tooling_tx, live_onhire, sources, when):
    """The handful of tables that go in the email body. Everything here
    is operational - what is out, what is working, what is not. No
    rates, no costs, no dollar signs."""
    def pct(v):
        return "{:.0f}%".format(round(v * 100))

    tiles = [
        ("Items on hire now", "{:,}".format(len(det))),
        ("Companies holding gear", "{:,}".format(
            len([c for c, n in live_onhire.items() if n]))),
        ("Coates items on site", "{:,}".format(len(coates_owned))),
        ("Tooling issues & returns", "{:,}".format(len(tooling_tx))),
        ("Consumables issued", "{:,}".format(len(ctx))),
        ("Consumable lines on shelf", "{:,}".format(len(ca))),
    ]

    #  Companies, busiest first. Only the TOTAL lines, so the table
    #  reads at a glance rather than three rows per outfit.
    by_co = {}
    for row in summary_rows:
        co, kind = row[0], row[1]
        d = by_co.setdefault(co, {})
        d[kind] = row
    companies = []
    for co in sorted(by_co, key=lambda c: -(by_co[c].get("TOTAL", [0] * 7)[2])):
        t = by_co[co].get("TOOLING", [co, "", 0, 0, 0, 0, 0])
        c = by_co[co].get("CONSUMABLES", [co, "", 0, 0, 0, 0, 0])
        if not (t[2] or c[2] or t[5]):
            continue
        companies.append([co, "{:,}".format(t[5]), "{:,}".format(t[6]),
                          "{:,}".format(t[2]), "{:,}".format(c[2])])

    #  Working hardest - the gear earning its keep.
    hard = [r for r in tu if r[11] in ("High Demand", "Good Use")]
    hard.sort(key=lambda r: (-(r[5] / r[1] if r[1] else 0), -r[5]))
    hardest = [[r[0], "{:,}".format(r[1]), "{:,}".format(r[2]),
                "{:,}".format(r[5]), r[11], r[12]] for r in hard[:12]]

    #  Not moving - biggest holdings that have not gone out at all.
    idle = [r for r in tu if r[11] == "No Use" and r[1] >= 2]
    idle.sort(key=lambda r: -r[1])
    idle_rows = [[r[0], "{:,}".format(r[1]), "{:,}".format(r[2]),
                  "{:,}".format(r[5]), r[12]] for r in idle[:12]]

    #  Consumables worth watching - the ones running down.
    watch = [r for r in cu if r[9] in ("High Demand", "Good Use")]
    watch.sort(key=lambda r: -r[4])
    cons = [[r[0], "{:,}".format(r[1]), "{:,}".format(r[2]), pct(r[4]),
             r[10]] for r in watch[:12]]

    #  Out the longest - oldest on-hire dates still open.
    dated = [r for r in det if r[5]]
    dated.sort(key=lambda r: r[5])
    longest = [[r[3], r[0], r[1], r[5].strftime("%d %b %Y"),
                "{:,}".format((when - r[5]).days)] for r in dated[:12]]

    stamp = ", ".join(sorted({
        dt.datetime.fromtimestamp(os.path.getmtime(p)).strftime("%d %b %Y")
        for p in sources})) or when.strftime("%d %b %Y")

    return {"tiles": tiles, "companies": companies, "hardest": hardest,
            "idle": idle_rows, "consumables": cons, "longest": longest,
            "source_stamp": stamp}


def carry_over(wb, s, base, title):
    """Copy Andrew's hand-typed tabs out of the existing workbook,
    values, formulas, dates and times exactly as he left them. If the
    old workbook isn't there the tabs are still created, headed and
    ready to type into - an empty tab that works beats a missing one."""
    src = s.workbook(base)
    carried, labour_total, cost_total = [], 0.0, 0.0

    old = None
    if src and os.path.isfile(src):
        try:
            old = openpyxl.load_workbook(src, data_only=False)
        except Exception as e:
            print("  WARN | Could not read {} ({}). The two hand-typed "
                  "tabs come through blank.".format(os.path.basename(src), e))

    #  The typed values, read a second time with formulas resolved, so
    #  the cover can show a total without re-implementing his sums.
    vals = None
    if src and os.path.isfile(src):
        try:
            vals = openpyxl.load_workbook(src, data_only=True)
        except Exception:
            vals = None

    for name in CARRIED_OVER:
        ws = wb.create_sheet(name)
        n = 0
        if old is not None and name in old.sheetnames:
            o = old[name]
            for row in o.iter_rows():
                filled = False
                for c in row:
                    if c.value is None:
                        continue
                    filled = True
                    nc = ws.cell(row=c.row, column=c.column, value=c.value)
                    nc.font = Font(name=c.font.name or "Calibri",
                                   size=c.font.sz or 11, bold=c.font.b)
                    if c.number_format:
                        nc.number_format = c.number_format
                if filled:
                    n += 1
            for rng in list(o.merged_cells.ranges):
                try:
                    ws.merge_cells(str(rng))
                except ValueError:
                    pass
            for k, dim in o.column_dimensions.items():
                if dim.width:
                    ws.column_dimensions[k].width = dim.width
        if n == 0:
            band(ws, title, 9)
            ws.cell(row=2, column=1,
                    value="Nothing typed in this tab yet - it is yours to "
                          "fill in and the rebuild never overwrites it.")
        else:
            #  Re-brand the title strip so it matches the other tabs.
            band(ws, title, max(5, ws.max_column))
        carried.append((name, n))

    #  Totals for the cover, straight off the values copy.
    if vals is not None:
        if "Coates Labour" in vals.sheetnames:
            #  Shift lines only. The sheet carries its own grand total on
            #  a row with no date in it - counting that as well doubles
            #  the labour bill, which is exactly what it did first go.
            o = vals["Coates Labour"]
            for row in o.iter_rows(min_row=6, min_col=1, max_col=9,
                                   values_only=True):
                if _date(row[0]) is None:
                    continue
                labour_total += _n(row[8])
        if "Cost Breakdown" in vals.sheetnames:
            o = vals["Cost Breakdown"]
            for row in o.iter_rows(min_row=3, max_row=19, min_col=3,
                                   max_col=5, values_only=True):
                for v in row:
                    cost_total += _n(v)
        try:
            vals.close()
        except Exception:
            pass
    if old is not None:
        try:
            old.close()
        except Exception:
            pass
    return carried, labour_total, cost_total


def cover(wb, s, when, asat, tabs, sources, base, headline):
    """The tab that was blank. Says what this workbook is, what fed it,
    the headline numbers, and how the two ratings are worked out - so
    nobody has to take the recommendations on trust."""
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    band(ws, "Coates    |  {}  |  SHUTDOWN TOOLSTORE ON-HIRE REPORT"
             "        Powered by SITEIQ".format(s.short or s.customer), 6)

    def put(row, col, value, bold=False, size=11, color=None, fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Calibri", size=size, bold=bold,
                      color=color or "FF000000")
        if fmt:
            c.number_format = fmt
        return c

    def rule(row, text):
        c = put(row, 1, text, bold=True, size=12, color=WHITE)
        c.fill = PatternFill("solid", fgColor=ORANGE_ALT)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        return row + 1

    r = 3
    put(r, 1, s.customer, bold=True, size=20); r += 1
    put(r, 1, "{} · {}".format(s.job, s.location), size=13,
        color="FF595959"); r += 2
    put(r, 1, "As at"); put(r, 2, asat.strftime("%d %b %Y  %H:%M"), bold=True); r += 1
    put(r, 1, "Report date"); put(r, 2, when.strftime("%d %b %Y"), bold=True); r += 1
    put(r, 1, "Prepared by"); put(r, 2, s.author or "Andrew Fisher", bold=True); r += 1
    put(r, 1, "SiteIQ project"); put(r, 2, s.project_schedule or "-", bold=True); r += 2

    r = rule(r, "THE HEADLINES")
    for k, v in headline.items():
        put(r, 1, k)
        money = "cost" in k.lower()
        put(r, 2, round(v, 2) if money else int(v), bold=True,
            fmt=MONEY_FMT if money else INT_FMT)
        r += 1
    r += 1

    r = rule(r, "WHAT'S IN THIS WORKBOOK")
    put(r, 1, "Tab", bold=True); put(r, 2, "Rows", bold=True)
    put(r, 3, "What it tells you", bold=True); r += 1
    for name, n, why in tabs:
        put(r, 1, name)
        put(r, 2, n, fmt=INT_FMT)
        put(r, 3, why, color="FF595959")
        r += 1
    r += 1

    r = rule(r, "WHERE THE NUMBERS CAME FROM")
    for p in sources:
        put(r, 1, os.path.basename(p))
        put(r, 2, dt.datetime.fromtimestamp(os.path.getmtime(p))
            .strftime("%d %b %Y  %H:%M"), color="FF595959")
        r += 1
    r += 1

    r = rule(r, "HOW THE RATINGS ARE WORKED OUT")
    for line in [
        "Tooling is rated on TURNS - transactions divided by how many of",
        "that item are on site. One item that went out five times is",
        "working; five items that went out once each are sitting there.",
        "   no turns          No Use        Review / Reduce",
        "   under 1 turn      Low Use       Keep Stock",
        "   1 to 2 turns      Good Use      Keep Stock  (Monitor / Increase",
        "                                   if 80% or more are out right now)",
        "   2 turns or more   High Demand   Increase Stock",
        "",
        "Consumables are rated on how much of the position has gone out -",
        "sales divided by sales plus what's left on the shelf.",
        "   nothing sold      No Use        Review / Reduce",
        "   under 40%         Low Use       Keep Stock",
        "   40% to 80%        Good Use      Keep Stock",
        "   over 80%          High Demand   Increase Stock",
    ]:
        put(r, 1, line, color="FF404040")
        r += 1
    r += 1

    r = rule(r, "WHAT IS TYPED BY HAND")
    for line in [
        "Coates Labour and Cost Breakdown are yours. The rebuild reads",
        "them out of the previous workbook and writes them straight back -",
        "it never recalculates them and never overwrites your typing.",
        "Everything else on every other tab is built from the SiteIQ",
        "exports listed above, so a fresh pull refreshes the lot.",
    ]:
        put(r, 1, line, color="FF404040")
        r += 1

    for col, w in ((1, 46), (2, 24), (3, 52), (4, 12), (5, 12), (6, 12)):
        ws.column_dimensions[get_column_letter(col)].width = w
    return ws


# =====================================================================
#  THE EMAIL - the whole report, in the body, no costs
# =====================================================================
#  Built as plain HTML tables with the styling written into every cell.
#  Outlook renders that reliably; a stylesheet at the top of the
#  document it quietly throws away. No pictures, no attachments beyond
#  the workbook, nothing fetched from the internet - so it looks the
#  same on a phone at the gate as it does on the store laptop.
#
#  NO COSTS. Not the labour, not the cost breakdown, not a rate or a
#  dollar sign anywhere. This is the operational picture - what is
#  out, who has it, what is working and what is not. The money stays
#  in the workbook.
# =====================================================================
E_TXT = ("font-family:Segoe UI,Calibri,Arial,sans-serif;font-size:11.5pt;"
         "color:#1B1F24;line-height:1.55")
E_TD = ("padding:7px 10px;border:1px solid #E3E6EB;font-family:Segoe UI,"
        "Calibri,Arial,sans-serif;font-size:10.5pt;color:#1B1F24;"
        "vertical-align:top")
E_TH = ("padding:7px 10px;border:1px solid #14181F;background:#14181F;"
        "color:#ffffff;font-family:Segoe UI,Calibri,Arial,sans-serif;"
        "font-size:9.5pt;letter-spacing:.6px;text-transform:uppercase;"
        "text-align:left")


def _esc(v):
    return (_s(v).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;"))


def _etable(cols, rows, aligns=None, note=""):
    """One table, styled cell by cell so Outlook keeps it."""
    if not rows:
        return ("<p style=\"{s};color:#5B6472;margin:0 0 16px\">"
                "Nothing to report here today.</p>".format(s=E_TXT))
    aligns = aligns or {}
    out = ["<table cellpadding='0' cellspacing='0' border='0' "
           "style='border-collapse:collapse;width:100%;margin:0 0 6px'>",
           "<tr>"]
    for i, c in enumerate(cols):
        out.append("<th style=\"{s};text-align:{a}\">{c}</th>".format(
            s=E_TH, a=aligns.get(i, "left"), c=_esc(c)))
    out.append("</tr>")
    for n, row in enumerate(rows):
        bg = "#ffffff" if n % 2 == 0 else "#F6F7F9"
        out.append("<tr>")
        for i, v in enumerate(row):
            out.append("<td style=\"{s};background:{b};text-align:{a}\">"
                       "{v}</td>".format(s=E_TD, b=bg,
                                         a=aligns.get(i, "left"),
                                         v=_esc(v)))
        out.append("</tr>")
    out.append("</table>")
    if note:
        out.append("<p style=\"{s};font-size:9.5pt;color:#5B6472;"
                   "margin:0 0 16px\">{n}</p>".format(s=E_TXT, n=_esc(note)))
    return "".join(out)


def _eheading(text):
    return ("<table cellpadding='0' cellspacing='0' border='0' "
            "style='border-collapse:collapse;width:100%;margin:22px 0 8px'>"
            "<tr><td style=\"padding:7px 12px;background:#F6F7F9;"
            "border:1px solid #E3E6EB;border-left:5px solid #F26222;"
            "font-family:Segoe UI,Calibri,Arial,sans-serif;font-size:12pt;"
            "font-weight:700;color:#14181F\">{t}</td></tr></table>"
            .format(t=_esc(text)))


def _etiles(pairs):
    """The headline numbers, three across. A table, because that is the
    only layout every mail client agrees on."""
    out = ["<table cellpadding='0' cellspacing='0' border='0' "
           "style='border-collapse:separate;border-spacing:8px 8px;"
           "width:100%;margin:4px 0 6px'>"]
    for i in range(0, len(pairs), 3):
        out.append("<tr>")
        for label, value in pairs[i:i + 3]:
            out.append(
                "<td width='33%' style=\"background:#F6F7F9;border:1px solid "
                "#E3E6EB;border-left:4px solid #F26222;padding:11px 13px\">"
                "<div style=\"font-family:Segoe UI,Calibri,Arial,sans-serif;"
                "font-size:21pt;font-weight:800;color:#14181F;"
                "line-height:1.1\">{v}</div>"
                "<div style=\"font-family:Segoe UI,Calibri,Arial,sans-serif;"
                "font-size:9pt;color:#5B6472;text-transform:uppercase;"
                "letter-spacing:.6px;padding-top:4px\">{l}</div></td>"
                .format(v=_esc(value), l=_esc(label)))
        for _ in range(3 - len(pairs[i:i + 3])):
            out.append("<td width='33%'></td>")
        out.append("</tr>")
    out.append("</table>")
    return "".join(out)


def email_html(s, when, asat, parts):
    """The whole report as an email body. Costs deliberately absent."""
    h = ["<div style=\"background:#ffffff;padding:0;margin:0\">",
         "<table cellpadding='0' cellspacing='0' border='0' "
         "style='border-collapse:collapse;width:100%;max-width:820px;"
         "margin:0 auto'>",
         "<tr><td style='padding:0 0 4px'>"]

    #  The band across the top - the Coates look, no image needed.
    h.append(
        "<table cellpadding='0' cellspacing='0' border='0' "
        "style='border-collapse:collapse;width:100%;background:#14181F;"
        "border-top:6px solid #F26222;border-radius:0'>"
        "<tr><td style='padding:15px 18px'>"
        "<div style=\"font-family:Segoe UI,Calibri,Arial,sans-serif;"
        "font-size:9pt;font-weight:800;letter-spacing:2px;color:#F26222;"
        "text-transform:uppercase\">COATES &middot; SHUTDOWN TOOL STORE</div>"
        "<div style=\"font-family:Segoe UI,Calibri,Arial,sans-serif;"
        "font-size:17pt;font-weight:700;color:#ffffff;padding-top:3px\">"
        "On-Hire Report</div>"
        "<div style=\"font-family:Segoe UI,Calibri,Arial,sans-serif;"
        "font-size:10pt;color:#C9CFD8;padding-top:3px\">{hdr}</div>"
        "<div style=\"font-family:Segoe UI,Calibri,Arial,sans-serif;"
        "font-size:9.5pt;color:#8B9099;padding-top:8px\">As at {asat}"
        " &nbsp;&middot;&nbsp; POWERED BY <span style=\"color:#F26222;"
        "font-weight:800\">SITEIQ</span> &nbsp;&middot;&nbsp; Author: "
        "{who}</div>"
        "</td></tr></table>".format(hdr=_esc(s.header_line),
                                    asat=asat.strftime("%d %b %Y  %H:%M"),
                                    who=_esc(s.author or "Andrew Fisher")))

    h.append("<div style=\"{s};padding:16px 2px 0\">".format(s=E_TXT))
    h.append("<p style=\"{s};margin:0 0 14px\">Morning,</p>".format(s=E_TXT))
    h.append("<p style=\"{s};margin:0 0 14px\">Here is where the tool store "
             "sits as at {d}. The full workbook is attached &mdash; twelve "
             "tabs, every line behind these numbers.</p>".format(
                 s=E_TXT, d=when.strftime("%d %B %Y")))

    h.append(_eheading("The position"))
    h.append(_etiles(parts["tiles"]))

    h.append(_eheading("Who is holding gear"))
    h.append(_etable(
        ["Company", "Tooling out", "Not returned", "Tooling issues",
         "Consumable issues"],
        parts["companies"],
        aligns={1: "right", 2: "right", 3: "right", 4: "right"},
        note="Tooling out is what is on hire in their name right now."))

    h.append(_eheading("Working hardest"))
    h.append(_etable(
        ["Item", "On site", "Out now", "Times issued", "Rating", "Do"],
        parts["hardest"],
        aligns={1: "right", 2: "right", 3: "right"},
        note="Rated on turns - times issued against how many we hold. "
             "Anything on Increase Stock is being asked for faster than "
             "we can turn it around."))

    h.append(_eheading("Not moving"))
    h.append(_etable(
        ["Item", "On site", "Out now", "Times issued", "Do"],
        parts["idle"],
        aligns={1: "right", 2: "right", 3: "right"},
        note="Biggest holdings that have not gone out at all. Every one of "
             "these is shelf space and hire we are carrying for nothing."))

    h.append(_eheading("Consumables to watch"))
    h.append(_etable(
        ["Line", "On shelf", "Gone out", "Used", "Do"],
        parts["consumables"],
        aligns={1: "right", 2: "right", 3: "right"},
        note="Used is what has gone out against the whole position. Over "
             "80% and we order before it bites."))

    h.append(_eheading("Out the longest"))
    h.append(_etable(
        ["Item", "Company", "Who has it", "On hire since", "Days"],
        parts["longest"], aligns={4: "right"},
        note="Worth a phone call before it turns into a demob problem."))

    h.append("<p style=\"{s};margin:18px 0 0\">Anything here you want "
             "chased, tell me and I'll get on it.</p>".format(s=E_TXT))
    h.append("<p style=\"{s};margin:16px 0 0\">{who}<br>"
             "<span style=\"color:#5B6472\">Coates &middot; Shutdown Tool "
             "Store</span></p>".format(s=E_TXT,
                                       who=_esc(s.author or "Andrew Fisher")))
    h.append("<p style=\"{s};font-size:9pt;color:#8B9099;margin:18px 0 0;"
             "border-top:1px solid #E3E6EB;padding-top:8px\">Built from the "
             "SiteIQ exports of {src}. Figures are the tool store's record "
             "as at the time above. Costs are in the workbook, not in this "
             "email.</p>".format(s=E_TXT, src=_esc(parts["source_stamp"])))
    h.append("</div></td></tr></table></div>")
    return "".join(h)


def email_text(s, when, parts):
    """The plain-text half, for anyone reading on a locked-down phone."""
    lines = ["COATES | ON-HIRE REPORT",
             s.header_line,
             "As at {}".format(when.strftime("%d %B %Y")),
             "",
             "Morning,",
             "",
             "Where the tool store sits today. Full workbook attached.",
             ""]
    for label, value in parts["tiles"]:
        lines.append("  {:<32} {}".format(label, value))
    lines += ["", "The detail is in the attached workbook - twelve tabs.",
              "Costs are in the workbook, not in this email.", "",
              s.author or "Andrew Fisher",
              "Coates - Shutdown Tool Store", ""]
    return "\n".join(lines)


def write_email(base, s, when, asat, parts, workbook_path, out_dir):
    """The report as an Outlook draft, with the workbook on the
    paperclip. NOTHING SENDS - it opens as a draft and waits for you."""
    import email.message
    import email.policy

    label = (s.short or s.customer or "Shutdown").strip()
    stem = "{} On-Hire Report - {}".format(label, when.strftime("%d %b %Y"))
    body = email_html(s, when, asat, parts)

    html_path = os.path.join(out_dir, stem + ".html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write("<!DOCTYPE html><html><head><meta charset='utf-8'>"
                "<title>{}</title></head><body style='margin:0;padding:14px;"
                "background:#ffffff'>{}</body></html>".format(_esc(stem),
                                                              body))

    msg = email.message.EmailMessage(policy=email.policy.SMTP)
    msg["Subject"] = "{} | On-Hire Report | {}".format(
        s.customer or label, when.strftime("%d %B %Y"))
    #  X-Unsent is what makes Outlook open it as a draft you press Send
    #  on yourself, rather than something already gone.
    msg["X-Unsent"] = "1"
    msg.set_content(email_text(s, when, parts))
    msg.add_alternative(body, subtype="html")

    if workbook_path and os.path.isfile(workbook_path):
        with open(workbook_path, "rb") as f:
            msg.add_attachment(
                f.read(), maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=os.path.basename(workbook_path))

    eml_path = os.path.join(out_dir, stem + ".eml")
    with open(eml_path, "wb") as f:
        f.write(msg.as_bytes())
    return html_path, eml_path


def main(argv):
    when = None
    if len(argv) > 1:
        try:
            when = dt.datetime.strptime(argv[1], "%Y-%m-%d").date()
        except ValueError:
            print("Date must look like 2026-08-10. Leave it off for today.")
            return 1
    return build(HERE, when)


if __name__ == "__main__":
    sys.exit(main(sys.argv))
