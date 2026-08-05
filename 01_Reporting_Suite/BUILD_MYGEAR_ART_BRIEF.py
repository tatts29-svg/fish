#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | MY GEAR ART BRIEF - every named PLACE, and its picture
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Andrew, 5 Aug 2026: "every page. every scene. everything that can
#  be clicked into needs to have a image. if their is a catergory that
#  you click into therr should be a image. if there is a location your
#  in. there should be a image. if you enter a area. there should be
#  cinematic experience. nothing should be a plain no visual image.
#  if it has a name there should be visual image."
#
#  The tools and consumables have their own workbook
#  (BUILD_THUMBNAIL_BRIEF.py). This one is the PLACES: the doors, the
#  scenes, the category tiles, the aisles, the bays, the gates, the
#  street - every named surface of My Gear, what picture it carries
#  today, exactly how a picture attaches to it, and the art direction
#  for the one it should carry.
#
#  Run it:  py BUILD_MYGEAR_ART_BRIEF.py
#  Output:  Coates_K2_MyGear_Art_Brief.xlsx (next to this file)
# =====================================================================
import os
import re
import glob
import datetime as dt

HERE = os.path.dirname(os.path.abspath(__file__))
ART = os.path.join(HERE, 'Gear_Lookup', 'art')
OUT = os.path.join(HERE, 'Coates_K2_MyGear_Art_Brief.xlsx')

#  THE PLACE STYLE - the world every scene lives in. One paragraph,
#  word for word on every row, same discipline as the item catalogue.
STYLE = ("PLACE STYLE (identical for every scene): a cinematic film "
         "still from inside the K2 tool store world. Near-black "
         "charcoal environment (#0A0E14 deepening to #14181F), warm "
         "orange rim light and wayfinding glow (#F26222) pooling "
         "exactly where the eye should go, a few cool white practical "
         "lights far back, faint haze for depth, pin-sharp foreground. "
         "Shot like a film, not a diagram. No people, no readable "
         "third-party branding, no added text or watermark, no "
         "prices. Scenes and banners land 1600x900; square tiles land "
         "800x800; JPG quality 80-85.")


def slug(s):
    return re.sub(r'[^a-z0-9]+', '-', str(s).lower()).strip('-')


def _have():
    """Every art file that exists today, so HAS ART reports the truth
    of this machine, not a memory of it."""
    out = {}
    for p in glob.glob(os.path.join(ART, '*.*')):
        out[os.path.basename(p).lower()] = 'art/' + os.path.basename(p)
    for sub in ('bays', 'icons', 'reports'):
        for p in glob.glob(os.path.join(ART, sub, '*.*')):
            out[sub + '/' + os.path.basename(p).lower()] = \
                'art/' + sub + '/' + os.path.basename(p)
    return out


def rows():
    have = _have()

    def r(name, kind, where, target, drawn_by, hook_live, brief,
          have_key=None):
        f = have.get((have_key or '').lower(), '')
        attach = ('FILE: Gear_Lookup\\{t}  |  DRAWN BY: {d}  |  HOOK: '
                  '{h}').format(
            t=target.replace('/', '\\'), d=drawn_by,
            h='live today - drop the file in and it shows on the next '
              'build' if hook_live else
              'to be built - the slot is designed, the page does not '
              'ask for this file yet')
        return [name, kind, where, ('YES - ' + f) if f else '',
                attach, brief + ' ' + STYLE, '']

    out = []

    # ---- 1. THE JOURNEY IN --------------------------------------------
    out += [
        r('The roller door opening', 'Journey scene',
          'The very first thing on screen when My Gear opens',
          'art/roller-shutter.webp + art/tool-store-interior.webp',
          'index.html opening animation (dwOpen)', True,
          'Two frames of one moment: the Coates roller shutter '
          'two-thirds up, and behind it the tool store interior '
          'revealed - racking lanes running away into the dark, every '
          'lane edge-lit orange, the floor holding a long warm '
          'reflection. The feeling: the store is opening FOR YOU.',
          'roller-shutter.webp'),
        r('The front door backdrop', 'Journey scene',
          'Behind the landing page, every visit',
          'art/store.jpg', "index.html body background", True,
          'The store facade at night from a worker\'s eye height, '
          'shutter down but glowing at the seams, the My Gear QR '
          'poster lit in its window, wet concrete in front carrying '
          'the orange spill.', 'store.jpg'),
        r('The card gate (SCAN YOUR CARD)', 'Journey scene',
          'The ID step between the door and your gear',
          'art/scene-card-gate.jpg', 'index.html card-scan panel',
          False,
          'A worker\'s ID card mid-swipe over a scanner pad, macro '
          'depth of field, the pad\'s status ring glowing orange '
          'about to go green - the half-second before the store knows '
          'you.'),
        r('Your personal gear bay', 'Journey scene',
          'The reveal when your own list opens',
          'art/personal-gear-bay.webp', 'index.html YOUR GEAR panel '
          '(wg reveal)', True,
          'One bay of racking that is unmistakably YOURS: a name '
          'plate catching the light, your gear staged neat on three '
          'shelves, one tagged tool front and centre.',
          'personal-gear-bay.webp'),
        r('My Gear HQ door', 'Journey scene',
          'The OPEN MY GEAR HQ button on the front door',
          'art/scene-hq-door.jpg', 'index.html HQ button', False,
          'A steel personnel door ajar onto a small lit control room '
          '- screens on the wall showing the store\'s numbers as '
          'coloured glows, keys still in the lock.'),
    ]

    # ---- 2. THE FRONT DOOR FURNITURE ----------------------------------
    tiles = [
        ('THE STORE tile', 'store.jpg',
         'Racking lanes from the counter\'s point of view, one lane '
         'lit brighter than the rest - come in and look.'),
        ('RADIO CHANNELS tile', 'radio.jpg',
         'A handheld radio upright in its charger pocket, channel '
         'display lit, the row of chargers falling away dark behind.'),
        ('GAS MONITORS tile', 'gas.jpg',
         'A gas monitor face-on with its screen lit teal-green '
         'against the dark, docking station glow behind.'),
        ('CONTACTS tile', 'contacts.jpg',
         'The counter phone under a single warm light, the laminated '
         'contact card beside it - the shot says "someone answers".'),
        ('TRADE TABLES tile', 'tables.jpg',
         'A torque chart clipboard hanging on the racking end, '
         'orange edge light down its side.'),
    ]
    for name, key, brief in tiles:
        out.append(r(name, 'Front-door tile',
                     'Front door - the navigation cards',
                     'art/' + key, 'index.html _fd_tile("' +
                     key.split('.')[0] + '")', True, brief, key))
    stats = [
        ('USERS stat tile', 'st_users.jpg',
         'A queue of gloved hands passing cards over the counter '
         'scanner, motion-blurred, one card sharp.'),
        ('ON HIRE stat tile', 'st_onhire.jpg',
         'A wall of tagged gear checked out - empty shadow-board '
         'silhouettes with orange outlines where tools should be.'),
        ('READY TO HIRE stat tile', 'st_ready.jpg',
         'A full shelf face-on, every hook loaded, tags all facing '
         'the same way - abundance, squared away.'),
    ]
    for name, key, brief in stats:
        out.append(r(name, 'Stat tile', 'Front door - the instrument '
                     'strip', 'art/' + key,
                     'index.html _stat("' + key.split('.')[0] + '")',
                     True, brief, key))
    out += [
        r('Toolbox talk banner (store days)', 'Scene banner',
          'Front door - today\'s toolbox talk, general topics',
          'art/store.jpg', 'index.html toolbox_talk() art picker',
          True, 'Covered by THE STORE tile image - same file, doing '
          'quiet duty behind the talk.', 'store.jpg'),
        r('"Nothing matches that" empty state', 'State screen',
          'Any search that finds nothing',
          'art/state-no-match.jpg', 'index.html #gnone empty state',
          False,
          'A single empty hook on a shadow board, its painted '
          'outline lit by one soft orange spot - not a sad picture, '
          'a "not here, ask us" picture.'),
        r('Departed-asset answer', 'State screen',
          'Searching an asset that has left site',
          'art/state-departed.jpg', 'index.html goNow()/stGone() '
          'departed panel', False,
          'Tail lights of a truck leaving the gate at dusk, the '
          'load silhouetted, gatehouse light streaking the wet '
          'road - gone, and gone PROPERLY, with a record.'),
    ]

    # ---- 3. THE CATEGORY TILES ----------------------------------------
    cats = [
        ('Everything', 'the whole store in one frame - drill, chain '
         'block, spanner fan, tape measure staged as a family '
         'portrait'),
        ('Sockets', 'a ratchet laid between two arcs of impact '
         'sockets in size order'),
        ('Spanners', 'a fan of combination spanners opening like a '
         'hand of cards'),
        ('Hand Tools', 'hammer, pliers and shifter crossed on the '
         'dark floor'),
        ('Lifting & Rigging', 'a lever hoist hanging with its chain '
         'pooled, shackles and a sling coiled at its foot'),
        ('Power Tools', 'drill, grinder and jigsaw in a low '
         'three-quarter group, batteries clicked in'),
        ('Measuring', 'tape, vernier and steel rule laid in a '
         'draughtsman\'s diagonal'),
        ('Safety Gear', 'white hard hat, clear glasses and hi-vis '
         'gloves under one clean spot'),
        ('Hydraulics', 'a hand pump with hose coiled once to a '
         'cylinder, couplers catching the light'),
        ('Leads & Power', 'an orange lead coiled beside a black '
         'distro box, outlets toward camera'),
        ('Air & Hoses', 'an air hose coil with a blow gun and '
         'claw couplings crossed in front'),
        ('Welding', 'a welding helmet three-quarter with electrode '
         'holder and leads draped forward'),
        ('Lighting', 'a tripod work light lit warm against the '
         'dark, throwing its own pool'),
        ('Radios', 'one radio in a six-bay charger, every pocket '
         'LED lit'),
        ('Batteries & Chargers', 'battery packs on a twin charger, '
         'charge bars glowing'),
        ('Gas Monitors', 'four monitors docked in a line, screens '
         'lit, one lifted forward'),
        ('Grinding & Cutting', 'a grinder with a stack of discs '
         'fanned beside it'),
        ('Tool Lanyards', 'two coiled lanyards with karabiners '
         'crossed to the front'),
        ('Fans & Ventilation', 'a drum fan face-on with flexible '
         'duct curving away into the dark'),
        ('Ladders & Access', 'a platform ladder part-open at '
         'three-quarter, feet sharp in the light'),
        ('Pumps & Cleaning', 'a submersible pump with its layflat '
         'hose coiled, wet floor reflections'),
    ]
    for name, scene in cats:
        out.append(r('Category: ' + name, 'Category tile',
                     'The what\'s-in-the-store grid - tap to open the '
                     'category', 'art/cat-' + slug(name) + '.jpg',
                     'the category grid tile background (concept page '
                     'today; same law when the grid goes live on the '
                     'worker catalogue)', False,
                     'A family portrait for the category: ' + scene +
                     '. Square tile, the name plate area kept clear '
                     'at the bottom third.'))

    # ---- 4. THE AISLES - where a thing LIVES --------------------------
    units = []
    try:
        import openpyxl
        p = max(glob.glob(os.path.join(HERE, 'Data_SiteIQ',
                                       'RENTAL_STOCK*.xlsx')),
                key=os.path.getmtime)
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        rws = list(wb['RENTAL_STOCK'].iter_rows(values_only=True))
        wb.close()
        ix = {str(v or '').strip(): i for i, v in enumerate(rws[0])}
        seen = {}
        for x in rws[1:]:
            v = str(x[ix.get('STORAGE_UNIT', -1)] or '').strip()
            if v:
                seen[v] = seen.get(v, 0) + 1
        units = sorted(seen.items(), key=lambda kv: -kv[1])
    except Exception:
        pass
    for unit, n in units:
        out.append(r('Aisle: ' + unit, 'Aisle & location',
                     'The location chip on every item card, and the '
                     'aisle header when browsing by location '
                     '({:,} lines live here)'.format(n),
                     'art/aisle-' + slug(unit) + '.jpg',
                     'item card location chip + aisle section header',
                     False,
                     'THIS aisle as a place: its own racking lane '
                     'shot from the entrance, the gear that lives '
                     'here recognisable on the shelves, an aisle '
                     'name plate catching the orange wayfinding '
                     'glow, lane running away into the dark.'))

    # ---- 5. THE STORES SIDE -------------------------------------------
    bays = [
        ('Find-It Counter', '01-find-it-counter.webp',
         'the counter bay - scanner, screen and the hand-over slab '
         'under one warm light'),
        ('The Hunt', '02-the-hunt.webp',
         'a torch beam cutting down a dark racking lane - the '
         'looking-for-it bay'),
        ('Store Floor', '03-store-floor.webp',
         'the widest shot of the floor - lanes, ladders and the '
         'counter glow far back'),
        ('Print Works', '04-print-works.webp',
         'the printer mid-page under a desk lamp, labels and a '
         'guillotine beside it'),
        ('Plant Desk', '05-plant-desk.webp',
         'the plant board - keys on hooks, prestart books squared '
         'up, one clipboard lit'),
        ('Store Control', '06-store-control.webp',
         'the control desk - twin screens glowing over a tidy '
         'keyboard, radio in its cradle'),
    ]
    for name, f, scene in bays:
        out.append(r('Stores bay: ' + name, 'Stores bay',
                     'The stores team board - the six working bays',
                     'art/bays/' + f, 'stores.html bay card', True,
                     'Already shot to the standard: ' + scene + '.',
                     'bays/' + f))
    out += [
        r('The stores gate (team code)', 'Gate scene',
          'The code step into the stores board',
          'art/scene-stores-gate.jpg', 'stores.html code gate', False,
          'A keypad on a steel door edge, one thumb\'s reach away, '
          'its keys back-lit - STAFF written in worn stencil above.'),
        r('The manager gate', 'Gate scene',
          'The manager code step into the money pages',
          'art/scene-manager-gate.jpg',
          'phone_reports manager gate panel', False,
          'A heavier door, a brass-edged key switch beside the '
          'keypad, a thin line of warm light under the door - the '
          'room the numbers live in.'),
    ]

    # ---- 6. CREW & FLEET ----------------------------------------------
    out += [
        r('Crew scorecard hero', 'Page hero',
          'Top of every personal crew page',
          'art/scene-crew-hero.jpg', 'crew.html header band', False,
          'A workbench at shift\'s end - gloves down, tagged tools '
          'lined up square for return, one card propped against a '
          'coffee cup. The picture says: your record, kept properly.'),
        r('Fleet details hero', 'Page hero',
          'Top of the fleet details page (button 68)',
          'art/scene-fleet-hero.jpg', 'fleet.html header band', False,
          'A long rank of identical machines in the laydown at '
          'dusk, plant numbers stencilled and catching the light, '
          'perspective compressing them into one family.'),
        r('Alternatives band', 'Section banner',
          'The "every asset in this fleet" section on fleet pages',
          'art/scene-alternatives.jpg', 'fleet.html alternatives '
          'section', False,
          'Three of the same machine side by side, one pulled '
          'half a metre forward into the light - the "or take this '
          'one" shot.'),
    ]
    icons = sorted(os.path.basename(p) for p in
                   glob.glob(os.path.join(ART, 'icons', '*.webp')))
    out.append(r('Crew scorecard icon set ({} badges)'.format(
        len(icons)), 'Icon set',
        'The badges on every crew scorecard',
        'art/icons/<badge>.webp', 'crew.html badge renderer', True,
        'Already drawn to one style: ' + ', '.join(
            i.replace('.webp', '') for i in icons) + '. Any NEW badge '
        'joins in the same style: single object, orange-lit, dark '
        'ground, square.', 'icons/' + icons[0] if icons else None))

    # ---- 7. STORE STREET - the reports shelf --------------------------
    out.append(r('Store Street', 'Journey scene',
                 'The reports shelf everyone lands on - "where do you '
                 'want to go"',
                 'art/scene-store-street.jpg',
                 'reports/index.html header', False,
                 'A narrow lane of small lit shop fronts at night - '
                 'each doorway a different warm glow, wet ground '
                 'stitching them together. The street the reports '
                 'live on.'))
    fronts = [
        ('My Gear', 'a lit doorway with a wall of tagged tools '
         'glimpsed inside'),
        ('The stores team board', 'a staff door ajar on the six-bay '
         'board glowing inside'),
        ('Fleet Details', 'a window onto the machine rank at dusk'),
        ('Utilisation', 'a window of gauges - needles all lit, one '
         'deep in the orange'),
        ('Returns performance', 'a doorway with a conveyor of '
         'returned gear rolling in'),
        ('Stocktake scorecard', 'a window of tally boards, one '
         'column lit gold'),
        ('Stocktake run sheet', 'a clipboard on a nail under a '
         'doorway lamp'),
        ('Consumables watch', 'shelves of boxes through a hatch, '
         'one row down to two'),
        ('Consumable requests', 'a request chit spiked on a counter '
         'nail, lamplight'),
        ('Offline day pack', 'a storm shutter half down, a battery '
         'lantern lit inside'),
        ('Every asset ranked', 'a podium of three machines under '
         'three spots'),
        ('Coates OnHire Summary ALL COMPANIES', 'a wide window onto '
         'the whole yard at night, every bay glowing'),
        ('Coates SitePlant Report', 'a window onto the big plant '
         'line, beacons turning'),
        ('How My Gear works', 'a small bright doorway with the QR '
         'poster lit inside it'),
    ]
    for name, scene in fronts:
        out.append(r('Street front: ' + name, 'Street front tile',
                     'Store Street - the tile that opens this report',
                     'art/reports/' + slug(name) + '.jpg',
                     'reports/index.html report tile', False,
                     'This report\'s shop front on Store Street: ' +
                     scene + '. Small tile; the name plate area kept '
                     'clear.'))

    # ---- 8. THE PRINT PIECES ------------------------------------------
    out += [
        r('QR window poster', 'Print piece',
          'The poster on the store window that starts everything',
          'art/poster-qr-hero.jpg', 'QR_Window_Poster.html hero slot',
          False,
          'A phone mid-scan of the window QR at night, the store '
          'glowing behind the glass - the moment of entry, printed '
          'A3.'),
        r('How My Gear works guide', 'Print piece',
          'The handout beside the counter',
          'art/poster-how-hero.jpg', 'MyGear_How_It_Works.html hero '
          'slot', False,
          'Three hands three phones: scanning the window, reading a '
          'gear card, showing the counter - one strip, one story.'),
    ]
    return out


def build():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, \
        Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    data = rows()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'My Gear Art Brief'
    ORANGE, DARK = 'F26222', '1D1D1B'
    thin = Border(bottom=Side(style='thin', color='DDDDDD'))

    ws.merge_cells('A1:G1')
    c = ws.cell(1, 1, 'COATES K2 | MY GEAR ART BRIEF - every named '
                'place, the picture it carries, and the one it should')
    c.font = Font(bold=True, size=13, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=DARK)
    ws.merge_cells('A2:G2')
    c = ws.cell(2, 1, 'Built {} by walking the live pages and the art '
                'folder on this machine - {} named surfaces. The rule '
                'this workbook serves: IF IT HAS A NAME, IT HAS A '
                'PICTURE. Tools and consumables live in their own '
                'workbook (Coates_K2_Thumbnail_Upgrade_Brief.xlsx); '
                'this is everything else.'.format(
                    dt.date.today().strftime('%d %b %Y'), len(data)))
    c.font = Font(size=9, color='555555')
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[2].height = 28

    heads = ['Part of My Gear', 'Kind', 'Where you meet it',
             'Has art today',
             'HOW THE IMAGE ATTACHES - file, page and hook',
             'THE IMAGE BRIEF - how it should look',
             'IMAGE MATCHES (Y / N / REDO)']
    for i, h in enumerate(heads, 1):
        c = ws.cell(4, i, h)
        c.font = Font(bold=True, size=9, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=ORANGE)
        c.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[4].height = 28

    for rn, row in enumerate(data, 5):
        for i, v in enumerate(row, 1):
            c = ws.cell(rn, i, v)
            c.font = Font(size=9)
            c.border = thin
            c.alignment = Alignment(wrap_text=(i in (3, 5, 6)),
                                    vertical='top')
        if row[3]:
            ws.cell(rn, 4).font = Font(size=9, bold=True,
                                       color='1F7A44')

    widths = [34, 15, 34, 30, 46, 68, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = 'A4:G{}'.format(4 + len(data))
    dv = DataValidation(type='list', formula1='"Y,N,REDO"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add('G5:G{}'.format(4 + len(data)))

    g = wb.create_sheet('Place Style & The Law')
    g.column_dimensions['A'].width = 110
    lines = [
        ('COATES K2 | THE PLACE STYLE - ONE WORLD FOR EVERY SCENE',
         True), ('', False), (STYLE, False), ('', False),
        ('HOW A PICTURE ATTACHES', True),
        ('Everything non-item lives in Gear_Lookup\\art\\ and is '
         'named for its job: scene-*.jpg for journey scenes, cat-*.jpg '
         'for category tiles, aisle-*.jpg for locations, bays\\*.webp '
         'for the stores bays, icons\\*.webp for scorecard badges, '
         'reports\\*.jpg for Store Street fronts, poster-*.jpg for '
         'the print pieces. Where the HOOK column says LIVE, dropping '
         'the file in is the whole job - the page picks it up on the '
         'next build. Where it says TO BE BUILT, the page does not '
         'ask for that file yet; the name is reserved so art can be '
         'made now and wired later without renaming a thing.', False),
        ('', False),
        ('THE RULES THAT RIDE ALONG', True),
        ('Every image degrades gracefully - pages must keep their '
         'monogram/plain fallback for the day a file is missing, '
         'exactly as the item thumbs do. Nothing in any scene shows '
         'a price, a person\'s face, or another company\'s branding. '
         'Scenes are the STORE\'S OWN world - the more they look '
         'like our actual store, the better they land; the six '
         'stores bays are the standard to match.', False),
    ]
    for i, (t, bold) in enumerate(lines, 1):
        c = g.cell(i, 1, t)
        c.font = Font(bold=bold, size=10 if bold else 9,
                      color=(ORANGE if bold else '222222'))
        c.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(OUT)
    return len(data), sum(1 for x in data if x[3]), OUT


if __name__ == '__main__':
    n, have, path = build()
    print('=' * 66)
    print(' COATES | MY GEAR ART BRIEF')
    print('=' * 66)
    print(' {} named surfaces | {} already carry art | {} to make'
          .format(n, have, n - have))
    print(' Written: ' + path)
