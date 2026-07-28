#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | ADD NEW CONTACTS - John Pickels' list, 28 Jul 2026
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY: John Pickels (Maintenance Planner, Cement Australia) sent
#  through contact addresses for four contractors, and Andrew asked for
#  five people - David, John, Ben, Thomas and Cody - to be CC'd on
#  every company report email.
#
#  What this does, once, safely:
#    1. Adds the four contractor contacts to the recipients book -
#       switched OFF (Include = No), the same rule as every import:
#       nothing starts emailing a client because a row appeared.
#       Turn each on with 34_SWITCH_ON.bat when you're ready.
#    2. Adds John Pickels to the FIXED CC block of the routing
#       workbook's Summary sheet - the same block that already carries
#       Thomas, Cody, Ben and David. Every company email (packs AND
#       Outlook drafts) reads its CC list from there.
#
#  Both workbooks are backed up first. Safe to run twice - a contact
#  that is already in the book is skipped, never doubled.
# =====================================================================
import os
import re
import shutil
import sys
import time

HERE = os.path.dirname(os.path.abspath(__file__))
RECIP = os.path.join(HERE, "Coates_Report_Recipients.xlsx")
ROUTE = os.path.join(HERE, "K2_Daily_Email_Report_Allocation.xlsx")
BK = os.path.join(HERE, "Updates", "contact_backups")

#  The four contractor contacts, exactly as John sent them. Company
#  spellings checked against the live rental data 28 Jul 2026 - all on
#  hire except Bosch, whose row waits until they hold gear.
NEW_ROWS = [
    ("INDUSTEC FILTRATION", "Admin (Industec)",
     "admin@industec.com.au", "To", "No", "COMPANY",
     "From John Pickels 28 Jul 2026 - shared Industec/Filtercare inbox"),
    ("FILTERCARE", "Admin (Industec & Filtercare)",
     "admin@industec.com.au", "To", "No", "COMPANY",
     "From John Pickels 28 Jul 2026 - shared Industec/Filtercare inbox"),
    ("INDUSTRIAL RESCUE & EMERGENCY TRAINING", "Chris Short",
     "chris.short@iretcq.com", "To", "No", "COMPANY",
     "From John Pickels 28 Jul 2026"),
    ("BOSCH", "Gerhardus Van Der Westhuizen",
     "gerhardus.vanderwesthuizen@boschrexroth.com.au", "To", "No",
     "COMPANY",
     "From John Pickels 28 Jul 2026 - not in SiteIQ yet; confirm the "
     "company spelling when Bosch holds gear"),
    ("CUTRITE", "Matt Campbell", "matt@cutrite.net.au", "To", "No",
     "COMPANY", "From John Pickels 28 Jul 2026"),
]

JOHN = ("John Pickels", "John.Pickels@cemaust.com.au")

#  The Cement master email (Cost Tracking Snapshot) - the To and CC
#  exactly as Andrew's own working draft of 28 Jul 2026. These are
#  people he already emails every day, so they land switched ON with
#  the COST tag: the snapshot draft comes out fully addressed.
COST_ROWS = [
    ("ALL", "David Moore", "david.moore@cemaust.com.au",
     "To", "Yes", "COST", "Cement master email 28 Jul 2026"),
    ("ALL", "Benjamin Vandenbroek", "benjamin.vandenbroek@cemaust.com.au",
     "To", "Yes", "COST", "Cement master email 28 Jul 2026"),
    ("ALL", "John Pickels", "John.Pickels@cemaust.com.au",
     "To", "Yes", "COST", "Cement master email 28 Jul 2026"),
    ("ALL", "Carmen Groat", "carmen.groat@cemaust.com.au",
     "CC", "Yes", "COST", "Cement master email 28 Jul 2026"),
    ("ALL", "Hazel Barba", "hazel.barba@cemaust.com.au",
     "CC", "Yes", "COST", "Cement master email 28 Jul 2026"),
    ("ALL", "Grant Holdsworth", "grant.holdsworth@cemaust.com.au",
     "CC", "Yes", "COST", "Cement master email 28 Jul 2026"),
    ("ALL", "Daniel Osborne", "daniel.osborne@cemaust.com.au",
     "CC", "Yes", "COST", "Cement master email 28 Jul 2026"),
    ("ALL", "Aaron Watts", "aaron.watts@cemaust.com.au",
     "CC", "Yes", "COST", "Cement master email 28 Jul 2026"),
    ("ALL", "Adam McInnes", "Adam.McInnes@cemaust.com.au",
     "CC", "Yes", "COST", "Cement master email 28 Jul 2026"),
    ("ALL", "Andrew Fisher (personal)", "atfish3r@hotmail.com",
     "CC", "Yes", "COST", "Cement master email 28 Jul 2026"),
    ("ALL", "Thomas Maher", "thomas.maher@coates.com.au",
     "CC", "Yes", "COST", "Cement master email 28 Jul 2026"),
    ("ALL", "Cody Mitchell", "cody.mitchell@coates.com.au",
     "CC", "Yes", "COST", "Cement master email 28 Jul 2026"),
]


def backup(path, tag):
    os.makedirs(BK, exist_ok=True)
    dst = os.path.join(BK, "{}_{}_{}".format(
        time.strftime("%Y%m%d_%H%M%S"), tag, os.path.basename(path)))
    shutil.copy2(path, dst)
    return dst


def main():
    print("=" * 70)
    print(" COATES | ADD NEW CONTACTS - John Pickels' list, 28 Jul 2026")
    print("=" * 70)
    try:
        import openpyxl
    except ImportError:
        print(" openpyxl isn't installed - run:  py -m pip install openpyxl")
        return 1

    # ---- 1. the contractor contacts -> recipients book -----------------
    if not os.path.isfile(RECIP):
        print(" Can't find Coates_Report_Recipients.xlsx here - run me "
              "from the suite folder.")
        return 1
    wb = openpyxl.load_workbook(RECIP)
    ws = wb["Recipients"] if "Recipients" in wb.sheetnames else wb.active
    hdr_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10,
                                         values_only=True), 1):
        cells = [str(c or "").strip().lower() for c in row]
        if "company" in cells and "email" in cells:
            hdr_row = i
            break
    if not hdr_row:
        print(" Couldn't find the header row in the Recipients sheet - "
              "nothing changed. Tell Claude.")
        return 1
    have = set()          # (company, email) pairs already in the book
    cost_have = set()     # emails already wired to the COST email
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        vals = [str(c or "").strip() for c in row]
        if len(vals) >= 3 and "@" in vals[2]:
            have.add((vals[0].upper(), vals[2].lower()))
            tags = (vals[5].upper() if len(vals) > 5 else "")
            if "COST" in tags or tags.strip() == "ALL":
                cost_have.add(vals[2].lower())
    added = skipped = 0
    bpath = None

    def _add(row):
        nonlocal bpath, added
        if bpath is None:
            bpath = backup(RECIP, "pre_john_pickels")
        ws.append(list(row))
        added += 1

    for r in NEW_ROWS:
        key = (r[0].upper(), r[2].lower())
        if key in have:
            skipped += 1
            continue
        _add(r)
        have.add(key)
    #  The COST list dedupes on the email alone - if someone already
    #  gets the snapshot (their row says COST or ALL), no second row.
    for r in COST_ROWS:
        if r[2].lower() in cost_have:
            skipped += 1
            continue
        _add(r)
        cost_have.add(r[2].lower())
    if added:
        wb.save(RECIP)
    print("")
    print(" Recipients book: {} contact(s) added, {} already there."
          .format(added, skipped))
    if added:
        print("   John's four contractor contacts land switched OFF - turn")
        print("   each company on with 34_SWITCH_ON.bat when you're ready.")
        print("   The Cement snapshot To/CC land switched ON with the COST")
        print("   tag - the snapshot draft comes out fully addressed.")
        print("   Backup: " + os.path.basename(bpath))

    # ---- 2. John Pickels -> the fixed CC block -------------------------
    if not os.path.isfile(ROUTE):
        print(" Routing workbook not found - John Pickels NOT added to "
              "the CC block. Run me again when it's here.")
        return 1
    wb2 = openpyxl.load_workbook(ROUTE)
    if "Summary" not in wb2.sheetnames:
        print(" No Summary sheet in the routing workbook - CC unchanged.")
        return 1
    ws2 = wb2["Summary"]
    grid = [[("" if c is None else str(c).strip()) for c in r]
            for r in ws2.iter_rows(values_only=True)]
    already = any(JOHN[1].lower() in " ".join(r).lower() for r in grid)
    if already:
        print(" CC block: John Pickels is already there. Nothing to do.")
    else:
        head_i = last_email_i = None
        for i, r in enumerate(grid, 1):
            head = " ".join(r).upper()
            if head_i is None and "FIXED" in head and "CC" in head:
                head_i = i
                continue
            if head_i is not None:
                if re.search(r"[\w.+-]+@[\w.-]+\.\w+", " ".join(r)):
                    last_email_i = i
                elif last_email_i is not None:
                    break
        if head_i is None or last_email_i is None:
            print(" Couldn't find the FIXED CC block on the Summary sheet -")
            print(" CC unchanged. Add him by hand under the same heading:")
            print("   {} <{}>".format(*JOHN))
        else:
            backup(ROUTE, "pre_john_pickels")
            ws2.insert_rows(last_email_i + 1)
            ws2.cell(row=last_email_i + 1, column=1).value = (
                "{} <{}>".format(*JOHN))
            wb2.save(ROUTE)
            print(" CC block: John Pickels added - every company email now")
            print(" CCs David, John, Ben, Thomas and Cody automatically.")

    print("")
    print(" Next run of the reports picks all of this up. Check with")
    print(" 18_CHECK_DAILY_EMAIL_PACKS.bat before you send.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
