#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | BUSINESS UTILISATION - TU and ROC, Coates eyes only
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 31 Jul 2026): "the invoice breakdown thats for the
#  client. now i want this broken down further for Coates as a
#  business - a full utilisation on costs, TU, not sure if we can do
#  ROC... but more so TU."
#
#  TWO NUMBERS, HONESTLY MADE:
#
#  TU - TIME UTILISATION, per storage unit and for the site. The
#  rental question: of the days this fleet has stood on site, how many
#  did it spend in a worker's hands? Two flavours, both shown:
#    * TU BILLED   = billed shift-days (TRANSACTION_CHARGES.SHIFTS)
#                    over fleet-days on site. The money meter.
#    * TU OCCUPIED = days accrued by everything on hire right now
#                    (ON_HIRE_DATE to today) over fleet-days. Catches
#                    the free-issue gear - radios live here, because
#                    they bill on the separate monthly invoice, not
#                    through the charge lines.
#  Fleet-days = items held x days since the shut started (13 Jul).
#  Arrival dates aren't in the exports, so gear that landed mid-shut
#  reads slightly LOW, never flattered. Stated, not hidden.
#
#  ROC - RETURN ON CAPITAL, per storage unit. Revenue earned against
#  the replacement value of the fleet deployed. Revenue = billed
#  shift-days x the item's own SiteIQ SHIFT_RATE (unit average where
#  an item has no current rate line). Capital = the master's
#  REPLACEMENT_COST_AUD. Both estimates are labelled with their
#  coverage - X of Y items priced - and the period figure is shown
#  next to the annualised one so nobody quotes a 19-day number as a
#  year. It is an ESTIMATE and says so on the page.
#
#  COATES INTERNAL. This page never joins the client packs - money
#  and utilisation of OUR fleet are our business.
# =====================================================================
import collections
import datetime as dt
import os
import sys

import report_paths as RP
import k2_utilisation as KU
import master_equipment


def read_fleet_costs():
    """Real dollars from the MyBranch fleet file (the 60 Fleet folder):
    plant number -> (original cost, WDV). K2 asset numbers ARE the
    network plant numbers (proven 1 Aug 2026), so plant-tracked gear
    prices from what Coates actually paid, not the master's estimate.
    No file, no fuss - ROC falls back to the estimates, labelled."""
    try:
        import build_fleet_finder as FF
        src = FF.find_export()
        if not src:
            return {}
        return {r['p']: (float(r['oc'] or 0), float(r['wv'] or 0))
                for r in FF.read_fleet(src) if r.get('p')}
    except (KeyboardInterrupt, SystemExit):
        raise
    except BaseException as e:
        print('  Fleet file skipped (%r) - using estimates.' % (e,))
        return {}

HERE = os.path.dirname(os.path.abspath(__file__))
ORG, INK, DIM = '#F26222', '#1D1D1B', '#8A94A2'
GOOD, AMBER, BAD = '#0ca30c', '#fab219', '#d03b3b'


def money(v):
    return '${:,.0f}'.format(v)


def read_rates(onhire_path):
    """item_number -> SHIFT_RATE off the current ON_HIRE export."""
    import openpyxl
    rates = {}
    if not onhire_path or not os.path.isfile(onhire_path):
        return rates
    try:
        wb = openpyxl.load_workbook(onhire_path, read_only=True,
                                    data_only=True)
        ws = wb['ON_HIRE'] if 'ON_HIRE' in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        ix = {str(c or '').strip(): i for i, c in enumerate(rows[0])}
        if 'SHIFT_RATE' not in ix or 'ITEM_NUMBER' not in ix:
            return rates
        for r in rows[1:]:
            it = str(r[ix['ITEM_NUMBER']] or '').strip()
            rt = KU._num(r[ix['SHIFT_RATE']])
            if it and rt > 0:
                rates[it] = rt
    except Exception:
        pass
    return rates


_STOP = {'AND', 'THE', 'FOR', 'WITH', 'OF', 'TWO', 'WAY', 'PER', 'NEW'}


def _toks(s):
    import re as _re
    return set(t for t in _re.split(r'[^A-Z0-9]+', str(s or '').upper())
               if len(t) >= 3 and t not in _STOP)


def read_monthly_rates():
    """The monthly-stream day rates off the newest Baseplan charges
    export (the 58 folder) - the radios, gas monitors and welders earn
    THERE, not through the SiteIQ lines, so without this their FU and
    revenue would read zero while they earn every day. Returns a list
    of (token-set, rate) to word-match item descriptions against."""
    import glob as _g
    import openpyxl
    hits = [p for p in _g.glob(os.path.join(HERE, 'Baseplan', '*.xlsx'))
            if not os.path.basename(p).startswith('~$')]
    for d in (os.path.join(HERE, 'Data_SiteIQ'), HERE):
        hits += [p for p in _g.glob(os.path.join(d, 'BASEPLAN_CHARGES*.xlsx'))
                 if not os.path.basename(p).startswith('~$')]
    if not hits:
        return []
    out = []
    try:
        wb = openpyxl.load_workbook(max(hits, key=os.path.getmtime),
                                    read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        ix = {str(c or '').strip(): i for i, c in enumerate(rows[0])}
        if 'Description' not in ix or 'Rate 1' not in ix:
            return []
        for r in rows[1:]:
            rate = KU._num(r[ix['Rate 1']])
            tk = _toks(r[ix['Description']])
            if rate > 0 and tk:
                out.append((tk, rate))
    except Exception:
        return []
    return out


def build(rental_path, txn_path, onhire_path, master, today=None):
    today = today or dt.date.today()
    days_in = max(1, (today - KU.SHUT_START).days)
    stock = KU._sheet(rental_path, 'RENTAL_STOCK')
    if not stock:
        return None

    shifts = collections.defaultdict(float)
    if txn_path and os.path.isfile(txn_path):
        for r in KU._sheet(txn_path, 'TRANSACTION_CHARGES'):
            b = KU._txt(r, 'LATEST_BARCODE')
            if b:
                shifts[b] += KU._num(r.get('SHIFTS'))
    rates = read_rates(onhire_path)
    mrates = read_monthly_rates()

    def kw_rate(desc):
        """Best word-match against the Baseplan lines, or 0."""
        tk = _toks(desc)
        best, score = 0.0, 0
        for btk, rate in mrates:
            got = len(tk & btk)
            if got > score:
                best, score = rate, got
        return best if score else 0.0

    #  the variant fold: PRODUCT_VARIANT where SiteIQ gives one, the
    #  derived model key where it doesn't (radios and gas monitors fold
    #  serial by serial into their fleet) - same rule as the photos,
    #  so this page and My Gear agree on what "a variant" is
    import mygear_thumbs as TH
    FLEET = read_fleet_costs()
    if FLEET:
        print('  Fleet file: real cost/WDV for {:,} plant numbers.'.format(len(FLEET)))

    def _blank():
        return {'n': 0, 'billed': 0.0, 'occ': 0.0, 'out': 0,
                'plant': False, 'repl': 0.0, 'priced': 0, 'rev': 0.0,
                'rated': 0, 'rate_sum': 0.0, 'pend': [],
                'pot': 0.0, 'unrated': 0, 'monthly': 0,
                'realp': 0, 'wdv': 0.0, 'wdv0': 0}
    U = collections.defaultdict(_blank)
    V = collections.defaultdict(_blank)
    vname = {}
    for r in stock:
        unit = KU._txt(r, 'STORAGE_UNIT') or 'Unassigned'
        desc = KU._txt(r, 'ITEM_DESCRIPTION')
        var = KU._txt(r, 'PRODUCT_VARIANT').upper()
        if not var:
            _ser, var = TH.derived_keys(desc)
            var = var or (desc.upper()[:40] or 'UNCODED')
            #  the fleet name reads clean - no one serial's tail on it
            import re as _re
            desc = _re.sub(r'\s*-?\s*Serial\s+\S+\s*$', '', desc,
                           flags=_re.I) or desc
        vk = (unit, var)
        if vk not in vname:
            vname[vk] = desc or var
        bc = KU._txt(r, 'ITEM_BARCODE')
        it = KU._txt(r, 'ITEM_NUMBER')
        sh = shifts.get(bc, 0.0)
        on_hire = KU._txt(r, 'ITEM_STATUS') == 'On Hire'
        od = 0
        if on_hire:
            d = KU._date(r.get('ON_HIRE_DATE'))
            if d:
                od = min(days_in, max(0, (today - d).days))
        p = master.price(it) if master else None
        flc = None
        for cand in (it, KU._txt(r, 'ASSET_NUMBER'), bc):
            if cand and cand in FLEET:
                flc = FLEET[cand]
                break
        if flc and flc[0] > 0:
            p = flc[0]
        rt = rates.get(it)
        #  no SiteIQ rate line = maybe the monthly stream: radios, gas,
        #  welders earn on the Baseplan invoice, occupancy IS their
        #  billing, so their revenue = days out x the invoice rate
        mrt = 0.0 if rt else kw_rate(desc)
        for agg in (U[unit], V[vk]):
            agg['n'] += 1
            agg['plant'] = KU._is_plant(unit)
            agg['billed'] += sh
            if on_hire:
                agg['out'] += 1
                agg['occ'] += od
            if p:
                agg['repl'] += p
                agg['priced'] += 1
            if flc and flc[0] > 0:
                agg['realp'] += 1
                agg['wdv'] += flc[1]
                if flc[1] <= 0 and sh > 0:
                    agg['wdv0'] += 1
            if rt:
                agg['rated'] += 1
                agg['rate_sum'] += rt
                agg['rev'] += sh * rt
                agg['pot'] += rt * days_in
            elif mrt:
                agg['monthly'] += 1
                agg['rev'] += od * mrt
                agg['billed'] += od      # monthly gear: days out ARE billed days
                agg['pot'] += mrt * days_in
            else:
                agg['unrated'] += 1
                if sh:
                    agg['pend'].append(sh)   # billed days awaiting a rate

    def _row(name, u, extra=None):
        fleet_days = u['n'] * days_in
        #  billed days with no current rate line price at the group's
        #  own average - estimate, and counted as such. The same average
        #  stands in for the unrated items' POTENTIAL, so FU's
        #  denominator covers the whole fleet, not just the rated bit.
        avg = (u['rate_sum'] / u['rated']) if u['rated'] else 0.0
        rev = u['rev'] + sum(u['pend']) * avg
        pot = u['pot'] + u['unrated'] * days_in * avg
        out = {
            'name': name, 'n': u['n'], 'out': u['out'],
            'plant': u['plant'],
            'fleetDays': fleet_days,
            'billed': u['billed'],
            'occ': u['occ'],
            'tuB': 100.0 * u['billed'] / fleet_days if fleet_days else 0.0,
            'tuO': 100.0 * u['occ'] / fleet_days if fleet_days else 0.0,
            'repl': u['repl'], 'priced': u['priced'],
            'realp': u['realp'], 'wdv': u['wdv'], 'wdv0': u['wdv0'],
            'rev': rev, 'pot': pot, 'monthly': u['monthly'],
            'fu': 100.0 * rev / pot if pot else None,
            'roc': 100.0 * rev / u['repl'] if u['repl'] else None,
        }
        if extra:
            out.update(extra)
        return out

    rows = []
    for name, u in U.items():
        r = _row(name, u)
        #  the variant drill-down under this unit, biggest fleet first
        vrows = [_row(vname[(un, vc)], V[(un, vc)], {'code': vc})
                 for (un, vc) in V if un == name]
        vrows.sort(key=lambda x: (-x['fleetDays'], x['name'].upper()))
        r['vars'] = vrows
        rows.append(r)
    rows.sort(key=lambda x: -x['fleetDays'])

    tot = {'n': sum(r['n'] for r in rows),
           'fleetDays': sum(r['fleetDays'] for r in rows),
           'billed': sum(r['billed'] for r in rows),
           'occ': sum(r['occ'] for r in rows),
           'repl': sum(r['repl'] for r in rows),
           'priced': sum(r['priced'] for r in rows),
           'realp': sum(r['realp'] for r in rows),
           'wdv': sum(r['wdv'] for r in rows),
           'wdv0': sum(r['wdv0'] for r in rows),
           'rev': sum(r['rev'] for r in rows),
           'pot': sum(r['pot'] for r in rows),
           'monthly': sum(r['monthly'] for r in rows)}
    tot['tuB'] = 100.0 * tot['billed'] / tot['fleetDays'] if tot['fleetDays'] else 0.0
    tot['tuO'] = 100.0 * tot['occ'] / tot['fleetDays'] if tot['fleetDays'] else 0.0
    tot['roc'] = 100.0 * tot['rev'] / tot['repl'] if tot['repl'] else None
    tot['fu'] = 100.0 * tot['rev'] / tot['pot'] if tot['pot'] else None
    return {'rows': rows, 'tot': tot, 'daysIn': days_in, 'today': today,
            'hasMonthly': bool(mrates)}


MENU = {'1': {'tu', 'fu', 'roc'}, '2': {'tu'}, '3': {'fu'}, '4': {'roc'},
        '5': {'tu', 'fu'}, '6': {'tu', 'roc'}, '7': {'fu', 'roc'}}


def parse_show(text):
    """'2', 'TU FU', 'roc' ... -> the set of metrics to show. Empty or
    unreadable = the lot, so Enter-and-go never breaks."""
    t = str(text or '').strip().upper()
    if not t:
        return {'tu', 'fu', 'roc'}
    if t in MENU:
        return MENU[t]
    got = set(m for m in ('TU', 'FU', 'ROC') if m in t.replace(',', ' ').split())
    return set(x.lower() for x in got) or {'tu', 'fu', 'roc'}


def ask_show(argv):
    """The picker (Andrew, 31 Jul 2026: 'options to pick what to show
    before the script starts'). Arguments win - 59 can be scripted -
    otherwise a plain menu, Enter for everything."""
    if argv:
        return parse_show(' '.join(argv))
    print('')
    print(' What should this run show?  (Enter = the lot)')
    print('   1  TU + FU + ROC - the full scorecard')
    print('   2  TU only  - the time story')
    print('   3  FU only  - money earned v money possible')
    print('   4  ROC only - return on capital')
    print('   5  TU + FU')
    print('   6  TU + ROC')
    print('   7  FU + ROC')
    try:
        return parse_show(input(' Pick a number, or type e.g. TU FU : '))
    except (EOFError, KeyboardInterrupt):
        return {'tu', 'fu', 'roc'}


def html(d, show=None):
    show = show or {'tu', 'fu', 'roc'}
    tu, fu, roc_on = 'tu' in show, 'fu' in show, 'roc' in show
    tot, rows, days_in = d['tot'], d['rows'], d['daysIn']
    ann = 365.0 / days_in

    def bar(p, col):
        w = min(100.0, p)
        return ("<div style='background:#eee;border-radius:6px;height:10px;"
                "width:120px;display:inline-block;vertical-align:middle'>"
                "<div style='background:{c};height:10px;border-radius:6px;"
                "width:{w:.0f}%'></div></div> {p:.0f}%").format(
                    c=col, w=w, p=p)

    def tu_col(p):
        return GOOD if p >= 50 else AMBER if p >= 25 else BAD

    h = ["""<!DOCTYPE html><html><head><meta charset='utf-8'>
<title>Coates K2 - Business Utilisation</title><style>
body{font-family:Segoe UI,Arial,sans-serif;color:#1D1D1B;margin:26px;
 font-size:13px}
h1{font-size:22px;margin:0}
h2{font-size:15px;border-left:4px solid #F26222;padding-left:9px;
 margin:24px 0 8px}
.brand b{color:#F26222;font-size:26px}
.brand span{color:#8A94A2;font-size:11px;letter-spacing:2px;display:block}
.meta{color:#8A94A2;font-size:11px;margin:4px 0 14px}
.intern{background:#1D1D1B;color:#fff;display:inline-block;padding:5px 12px;
 border-radius:7px;font-weight:700;font-size:11px;letter-spacing:1.5px}
.story{background:#FFF4EE;border-left:4px solid #F26222;border-radius:8px;
 padding:12px 14px;margin:12px 0;line-height:1.65}
table{width:100%;border-collapse:collapse;font-size:12.5px}
th{text-align:left;color:#8A94A2;font-size:10.5px;letter-spacing:.8px;
 text-transform:uppercase;padding:6px 8px;border-bottom:2px solid #F26222}
td{padding:7px 8px;border-bottom:1px solid #eee;vertical-align:middle}
tr.tot td{font-weight:700;border-top:2px solid #1D1D1B}
.r{text-align:right}
.note{color:#5a6470;font-size:11.5px;line-height:1.7;margin:8px 0}
.foot{color:#8A94A2;font-size:10.5px;text-align:center;margin-top:26px}
@media print{body{margin:10mm}}
</style></head><body>
<div class='brand'><b>COATES</b><span>POWERED BY SITEIQ</span></div>
<h1>Business utilisation &mdash; __PICKED__</h1>
<div class='meta'>Cement Australia K2 Shutdown 2026 &middot; Gladstone
 &middot; day __DAYS__ of the shut &middot; as at __ASOF__
 &middot; Author: Andrew Fisher</div>
<div class='intern'>COATES INTERNAL &mdash; NOT FOR CLIENT DISTRIBUTION</div>
"""
             #  token swap, never .format - the CSS above is full of
             #  { } braces (the suite's own documented trap)
             .replace('__DAYS__', str(days_in))
             .replace('__ASOF__', d['today'].strftime('%d %b %Y'))
             .replace('__PICKED__', ' &amp; '.join(
                 m for m, on in (('TU', tu), ('FU', fu), ('ROC', roc_on))
                 if on))]

    pos = ["<div class='story'><b>The position:</b> the fleet has stood "
           "{fd:,} fleet-days on site".format(fd=tot['fleetDays'])]
    if tu:
        pos.append(" and spent {occ:,.0f} of them in workers' hands "
                   "&mdash; site TU {tuO:.0f}% occupied, {tuB:.0f}% "
                   "billed".format(occ=tot['occ'], tuO=tot['tuO'],
                                   tuB=tot['tuB']))
    if fu:
        pos.append(". Estimated revenue to date {rev} of a possible {pot} "
                   "&mdash; <b>FU {v}</b>".format(
                       rev=money(tot['rev']), pot=money(tot['pot']),
                       v=('{:.0f}%'.format(tot['fu'])
                          if tot['fu'] is not None else 'n/a')))
    if roc_on:
        pos.append(". Against {repl} of capital deployed ({pr} of {n} "
                   "items priced): ROC {p} for the period, {a} annualised"
                   .format(repl=money(tot['repl']), pr=tot['priced'],
                           n=tot['n'],
                           p=('{:.1f}%'.format(tot['roc'])
                              if tot['roc'] is not None else 'n/a'),
                           a=('{:.0f}%'.format(tot['roc'] * ann)
                              if tot['roc'] is not None else 'n/a')))
    pos.append(".</div>")
    h.append(''.join(pos).replace('. .', '.'))

    if tu:
        h.append("<h2>TU &mdash; time utilisation by storage unit</h2>"
                 "<div class='note'>TU OCCUPIED counts every day an item "
                 "has stood on hire. TU BILLED counts shift-days that "
                 "charged through the SiteIQ lines, plus the "
                 "monthly-stream gear (radios, gas, welders) whose every "
                 "day out bills on the Baseplan invoice. Fleet-days = "
                 "items held &times; {d} days; gear that arrived mid-shut "
                 "reads slightly low, never flattered.</div>"
                 .format(d=days_in))
        h.append("<table><tr><th>Storage unit</th><th class='r'>Items</th>"
                 "<th class='r'>Out now</th><th class='r'>Fleet-days</th>"
                 "<th class='r'>Occupied days</th><th>TU occupied</th>"
                 "<th class='r'>Billed days</th><th>TU billed</th></tr>")
        for r in rows:
            h.append(("<tr><td><b>{name}</b>{pl}</td><td class='r'>{n}</td>"
                      "<td class='r'>{out}</td><td class='r'>{fd:,}</td>"
                      "<td class='r'>{occ:,.0f}</td><td>{tb1}</td>"
                      "<td class='r'>{bd:,.0f}</td><td>{tb2}</td></tr>")
                     .format(
                name=r['name'],
                pl=(" <span style='color:#8A94A2;font-size:10px'>PLANT"
                    "</span>" if r['plant'] else ''),
                n=r['n'], out=r['out'], fd=r['fleetDays'], occ=r['occ'],
                tb1=bar(r['tuO'], tu_col(r['tuO'])),
                bd=r['billed'], tb2=bar(r['tuB'], tu_col(r['tuB']))))
        h.append(("<tr class='tot'><td>WHOLE SITE</td><td class='r'>{n}"
                  "</td><td></td><td class='r'>{fd:,}</td>"
                  "<td class='r'>{occ:,.0f}</td>"
                  "<td>{b1}</td><td class='r'>{bd:,.0f}</td><td>{b2}</td>"
                  "</tr></table>").format(
            n=tot['n'], fd=tot['fleetDays'], occ=tot['occ'],
            b1=bar(tot['tuO'], tu_col(tot['tuO'])),
            bd=tot['billed'], b2=bar(tot['tuB'], tu_col(tot['tuB']))))

    if fu or roc_on:
        h.append("<h2>The money view &mdash; "
             + ('FU &amp; ROC' if fu and roc_on else ('FU' if fu else 'ROC'))
             + " by storage unit (estimate)</h2>"
             "<div class='note'>FU (financial utilisation) = revenue "
             "earned &divide; revenue if every item hired every day at "
             "its rate. Revenue: billed shift-days &times; the item's "
             "SiteIQ SHIFT_RATE; the monthly-stream gear (radios, gas, "
             "welders) earns its Baseplan invoice rate for every day "
             "out, read straight off the 58 folder"
             + ("" if d.get('hasMonthly') else
                " &mdash; <b>no Baseplan export found, so the monthly "
                "stream reads zero this run</b>")
             + ". Items with no rate anywhere ride at the unit average. "
             "Capital = the master's replacement cost; units missing "
             "prices read LOW on capital, HIGH on ROC."
             + ((" <b>Real dollars:</b> {rp} unit(s) priced from the "
                "MyBranch fleet file (what Coates actually paid); "
                "book value (WDV) {wd} - and <b>{w0} written-down "
                "unit(s) (WDV $0) earned hire this period - pure "
                "return on dead capital.</b>").format(
                 rp=tot.get('realp', 0), wd=money(tot.get('wdv', 0)),
                 w0=tot.get('wdv0', 0)) if tot.get('realp') else '')
             + "</div>")
        def _mcells(r, is_tot=False):
            cs = ["<td class='r'>" + money(r['rev']) + "</td>"]
            if fu:
                cs.append("<td class='r'>" + money(r['pot']) + "</td>")
                cs.append("<td>" + (bar(r['fu'], tu_col(r['fu']))
                                    if r['fu'] is not None else '&mdash;')
                          + "</td>")
            if roc_on:
                roc = r['roc']
                cs.append("<td class='r'>" + money(r['repl']) + "</td>")
                cs.append("<td class='r'>{}/{}</td>".format(r['priced'],
                                                            r['n']))
                cs.append("<td class='r'>" + ('{:.1f}%'.format(roc)
                          if roc is not None else '&mdash;') + "</td>")
                cs.append("<td class='r'><b>" + ('{:.0f}%'.format(roc * ann)
                          if roc is not None else '&mdash;') + "</b></td>")
            return ''.join(cs)

        head = ["<table><tr><th>Storage unit</th>",
                "<th class='r'>Revenue est.</th>"]
        if fu:
            head.append("<th class='r'>Potential</th><th>FU</th>")
        if roc_on:
            head.append("<th class='r'>Capital</th><th class='r'>Priced"
                        "</th><th class='r'>ROC period</th>"
                        "<th class='r'>ROC annualised</th>")
        head.append("</tr>")
        h.append(''.join(head))
        rocrows = sorted([r for r in rows
                          if r['repl'] or r['rev'] or r['pot']],
                         key=lambda x: -(x['rev']))
        for r in rocrows:
            h.append("<tr><td><b>" + r['name'] + "</b></td>"
                     + _mcells(r) + "</tr>")
        h.append("<tr class='tot'><td>WHOLE SITE</td>"
                 + _mcells(tot, True) + "</tr></table>")

    #  ---- the drill-down: every product variant, unit by unit --------
    picked = ' &amp; '.join(m for m, on in
                            (('TU', tu), ('FU', fu), ('ROC', roc_on)) if on)
    h.append("<h2>The drill-down &mdash; " + picked
             + " by product variant</h2>"
             "<div class='note'>Every variant in every unit &mdash; the "
             "radios and gas monitors fold serial by serial into their "
             "fleet, same rule as My Gear. This is where the fleet-cut "
             "conversation gets specific: a variant at 0% since day one "
             "is a demob candidate by name, not by feel. Sorted biggest "
             "fleet first inside each unit.</div>")
    for r in rows:
        vs = r.get('vars') or []
        if not vs:
            continue
        h.append("<h3 style='font-size:13px;margin:16px 0 4px;color:#1D1D1B'>"
                 + r['name']
                 + " <span style='color:#8A94A2;font-weight:400'>&middot; "
                 + str(len(vs)) + " variant" + ('' if len(vs) == 1 else 's')
                 + "</span></h3>")
        vh = ["<table><tr><th>Product variant</th><th class='r'>Items</th>"
              "<th class='r'>Out</th>"]
        if tu:
            vh.append("<th>TU occupied</th><th>TU billed</th>")
        if fu or roc_on:
            vh.append("<th class='r'>Revenue est.</th>")
        if fu:
            vh.append("<th class='r'>FU</th>")
        if roc_on:
            vh.append("<th class='r'>Capital</th><th class='r'>ROC ann."
                      "</th>")
        vh.append("</tr>")
        h.append(''.join(vh))
        for v in vs:
            roc = v['roc']
            cs = [("<tr><td><b>{nm}</b><br><span style='color:#8A94A2;"
                   "font-family:Consolas,monospace;font-size:10px'>{cd}"
                   "</span></td><td class='r'>{n}</td>"
                   "<td class='r'>{out}</td>").format(
                nm=v['name'], cd=v.get('code', ''), n=v['n'], out=v['out'])]
            if tu:
                cs.append("<td>" + bar(v['tuO'], tu_col(v['tuO']))
                          + "</td><td>" + bar(v['tuB'], tu_col(v['tuB']))
                          + "</td>")
            if fu or roc_on:
                cs.append("<td class='r'>"
                          + (money(v['rev']) if v['rev'] else '&mdash;')
                          + "</td>")
            if fu:
                cs.append("<td class='r'><b>"
                          + ('{:.0f}%'.format(v['fu'])
                             if v['fu'] is not None else '&mdash;')
                          + "</b></td>")
            if roc_on:
                cs.append("<td class='r'>"
                          + (money(v['repl']) if v['repl'] else '&mdash;')
                          + "</td><td class='r'><b>"
                          + ('{:.0f}%'.format(roc * ann)
                             if roc is not None else '&mdash;')
                          + "</b></td>")
            cs.append("</tr>")
            h.append(''.join(cs))
        h.append("</table>")

    h.append("<div class='foot'>Built from this morning's SiteIQ exports "
             "&middot; COATES INTERNAL &middot; POWERED BY SITEIQ &middot; "
             "Author: Andrew Fisher</div></body></html>")
    return ''.join(h)


def main():
    print('=' * 66)
    print(' COATES | BUSINESS UTILISATION - TU, FU & ROC, Coates eyes only')
    print('=' * 66)
    show = ask_show(sys.argv[1:])
    print(' Showing  : ' + ' + '.join(
        m for m, k in (('TU', 'tu'), ('FU', 'fu'), ('ROC', 'roc'))
        if k in show))
    rental = RP.find_export(HERE, 'RENTAL_STOCK*.xlsx')
    txn = RP.find_export(HERE, 'TRANSACTIONS*.xlsx')
    onhire = RP.find_export(HERE, 'ON_HIRE*.xlsx')
    if not rental:
        print(' No RENTAL_STOCK export - run the morning downloads first.')
        return 1
    master = None
    try:
        master = master_equipment.load(HERE, quiet=True)
    except Exception:
        pass
    d = build(rental, txn, onhire, master)
    if not d:
        return 1
    out_dir = RP.day_folder(HERE)
    out = os.path.join(out_dir, 'Coates_K2_Business_Utilisation_{}.html'
                       .format(d['today'].isoformat()))
    with open(out, 'w', encoding='utf-8') as f:
        f.write(html(d, show))
    t = d['tot']
    print(' Site TU  : {:.0f}% occupied | {:.0f}% billed  (day {} of the shut)'
          .format(t['tuO'], t['tuB'], d['daysIn']))
    if t['fu'] is not None:
        print(' Site FU  : {:.0f}%  ({} earned of {} possible)'.format(
            t['fu'], money(t['rev']), money(t['pot'])))
    if t['roc'] is not None:
        print(' Site ROC : {:.1f}% period | {:.0f}% annualised  '
              '({} of {} items priced)'.format(
                  t['roc'], t['roc'] * 365.0 / d['daysIn'],
                  t['priced'], t['n']))
    print('')
    print(' Written to : ' + out)
    print(' COATES INTERNAL - keep it out of the client packs.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
