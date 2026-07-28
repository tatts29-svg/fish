#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | UPDATE CONTACTS - your Outlook list into the address book
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 27 Jul 2026): "we need to update all my contacts for
#  emails" - and he keeps the real list in Outlook, exported as a CSV.
#  Retyping 65 people into a spreadsheet is how addresses go stale and
#  how a supervisor quietly stops getting his company's report.
#
#  THE PART THAT MATTERS MOST - the company name.
#  recipients.resolve() matches a contact to a report by an EXACT
#  company-name compare. Get the name wrong and nothing errors: the
#  report builds, the email sends, and the person simply is not on it.
#
#  So where does the right name come from? NOT the folders under
#  01 COMPANY REPORTS - those are yesterday's output and SiteIQ has
#  been seen to change its spelling between exports (VEOLIA REFRACTORY
#  -> VEOLIA, SYNCLIFT -> SYNC LIFT ENGINEERING, DKE GROUP -> DARK
#  KNIGHT ENGINEERING). A first cut of this script trusted those
#  folders and wrote names that were already out of date.
#
#  The only honest source is the live rental data put through the very
#  functions the report itself uses:
#
#       load_rental() -> item["company"]        (already canonical)
#       display_company(canonical)              -> the string resolve()
#                                                  will be handed
#
#  That is what this script reads. If it cannot read it, it STOPS and
#  says so. It does not fall back to a weaker source and hand back
#  confident-looking wrong answers - that is exactly how the first
#  version went wrong on the new laptop.
#
#  WHAT IT WILL NOT DO: switch anyone on. Every new contact lands with
#  Include = No. Turning 65 people into live recipients of client
#  reports is Andrew's call, made with his eyes open - never a side
#  effect of an import.
# =====================================================================
import csv
import glob
import os
import re
import shutil
import sys
import time

HERE = os.path.dirname(os.path.abspath(__file__))
BOOK = os.path.join(HERE, "Coates_Report_Recipients.xlsx")
BK = os.path.join(HERE, "Updates", "contact_backups")

#  Coates' own people belong to every report, not to a client company.
INTERNAL = {"COATES"}

#  Tags a brand-new client contact gets. COMPANY = their own on-hire
#  report, DEMOB = the gear-return push. Deliberately narrow: nobody is
#  put on ALL by an import.
DEFAULT_TAGS = "COMPANY,DEMOB"


def norm(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def find_csv():
    """The Outlook export, wherever Andrew happened to save it."""
    spots = [HERE,
             os.path.join(HERE, "K2 DAILY REPORTING", "03 CONTACTS & SETTINGS"),
             os.path.join(os.path.dirname(HERE), "K2 DAILY REPORTING",
                          "03 CONTACTS & SETTINGS"),
             os.path.join(os.path.expanduser("~"), "Downloads"),
             os.path.join(os.path.expanduser("~"), "Desktop")]
    hits = []
    for s in spots:
        for f in glob.glob(os.path.join(s, "*.csv")):
            if os.path.basename(f).startswith("~"):
                continue
            try:
                with open(f, encoding="utf-8-sig", errors="replace") as fh:
                    head = fh.readline().lower()
            except Exception:
                continue
            if "e-mail address" in head or "email address" in head:
                hits.append(f)
    return max(hits, key=os.path.getmtime) if hits else None


def live_companies():
    """The company names the reports will ACTUALLY be built under, taken
    from today's rental data through the report's own naming functions.
    Returns (display_names, canonical_of, error) - error set means we
    could not read it, and the caller must stop rather than guess."""
    try:
        import build_company_onhire_report as B
    except Exception as e:
        return [], None, "couldn't load the report engine ({})".format(e)
    try:
        items = B.load_rental()[0]
    except Exception as e:
        return [], None, "couldn't read the rental export ({})".format(e)
    canon = sorted({i["company"] for i in items if i.get("company")})
    if not canon:
        return [], None, "the rental export has no companies in it"
    return sorted({B.display_company(c) for c in canon}), B, ""


def match_company(name, live, B):
    """An Outlook company name -> the string a report is built under.

    Step 1 runs it through the report's own canonical + display pipeline,
    so every alias Andrew has already taught the suite applies for free -
    'Cement Australia' becomes the Holdings entity, 'DKE Group' becomes
    Dark Knight Engineering.
    Step 2 falls back to ignoring spacing and punctuation, then to a
    prefix match, which is what turns 'DGH' into 'DGH Engineering' and
    'Veolia Refractory' into 'Veolia'.
    Anything still unmatched comes back as None and is put in front of
    Andrew - it is never quietly guessed."""
    disp = B.display_company(B.canonical_company(name))
    if disp in live:
        return disp, ("exact" if disp == name else "alias")
    n = norm(name)
    for c in live:
        if norm(c) == n:
            return c, "spelling"
    cands = [c for c in live if norm(c).startswith(n) or n.startswith(norm(c))]
    if len(cands) == 1:
        return cands[0], "matched"
    return None, "none"


def name_from_email(em):
    """Outlook rows with no first or last name still have to be called
    something, and "Aburgun" isn't it.

    Coates and its contractors address people the same way all through
    this list - initial then surname: jadams, kdollery, psheehan,
    jbryant, blynch. So a single run-on prefix is split back out. Short
    prefixes are shared mailboxes, not people (hse@, mat@, ca@), and get
    left upper-case rather than mangled into "H Se"."""
    p = em.split("@")[0]
    if "." in p or "_" in p:
        return " ".join(w.capitalize() for w in re.split(r"[._]+", p) if w)
    if len(p) <= 3:
        return p.upper()
    m = re.match(r"^([A-Za-z])([A-Za-z]{4,})$", p)
    if m:
        return "{} {}".format(m.group(1).upper(), m.group(2).capitalize())
    return p.capitalize()


def read_csv(path):
    people = []
    with open(path, encoding="utf-8-sig", errors="replace", newline="") as fh:
        for r in csv.DictReader(fh):
            r = {(k or "").strip(): (v or "").strip() for k, v in r.items()}
            em = r.get("E-mail Address") or r.get("Email Address") or ""
            if "@" not in em:
                continue
            nm = (r.get("First Name", "") + " " + r.get("Last Name", "")).strip()
            if not nm:
                nm = name_from_email(em)
            people.append({
                "name": nm,
                "company": r.get("Company", "").strip(),
                "title": r.get("Job Title", "").strip(),
                "email": em,
                "mobile": r.get("Mobile Phone", "").strip(),
            })
    return people


def _edit(a, b):
    """Plain edit distance - 'Wacher' and 'Wachter' are one keystroke
    apart and are the same bloke."""
    if a == b:
        return 0
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1,
                           prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def same_person(a, b):
    """Two book rows that are one human written twice.

    Deliberately hard to satisfy. The book legitimately holds one mailbox
    under two names - ca@highrisksolutions.com.au is Mat's AND Josh's -
    and holds one person twice on purpose, once for their company and
    once for the site-wide safety pack. Neither may be merged.

    So: same company, and either the same name with the email missing off
    one row, or the same address with names a keystroke apart."""
    ca, cb = norm(a["company"]), norm(b["company"])
    if ca != cb or not ca:
        return False
    na, nb = norm(a["name"]), norm(b["name"])
    ea, eb = a["email"].strip().lower(), b["email"].strip().lower()
    if "(" in a["name"] or "(" in b["name"]:
        return False                      # the safety-pack style variants
    if na and na == nb and (not ea or not eb):
        return True                       # typed in before the address was
    if ea and ea == eb:
        if na in nb or nb in na:
            return True
        ta, tb = na[-9:], nb[-9:]         # surname-ish tail
        return _edit(ta, tb) <= 2
    return False


def merge_into(keep, drop):
    """Fold the weaker row into the stronger one, never losing a Yes and
    never losing a report tag."""
    if drop["include"].lower() == "yes":
        keep["include"] = "Yes"
    if not keep["email"]:
        keep["email"] = drop["email"]
    if len(drop["name"]) > len(keep["name"]):
        keep["name"] = drop["name"]
    tags = [t.strip() for t in (keep["reports"] + "," + drop["reports"]).split(",")]
    seen = []
    for t in tags:
        if t and t.upper() not in [s.upper() for s in seen]:
            seen.append(t)
    if "ALL" in [s.upper() for s in seen]:
        seen = ["ALL"]
    keep["reports"] = ",".join(seen)
    if len(drop["notes"]) > len(keep["notes"]):
        keep["notes"] = drop["notes"]


def claim_row(person, rows, taken):
    """Which existing book row is this Outlook contact, if any?

    Email is the identity - but the book legitimately holds the SAME
    address twice (ca@highrisksolutions.com.au is Mat's and Josh's
    shared inbox; Kevin Dollery has a Walz row and a separate safety-pack
    row), so the name breaks the tie. And a row typed in ahead of the
    address - "Kellie Kay" with the email column still empty - has to be
    filled in, not duplicated beside itself.

    Keyed on email alone these merge people who are not the same person;
    keyed on email+name they duplicate people who are ("S Wacher" in the
    book, " Wachter" in Outlook, one mailbox between them). Hence: email
    first, name only to choose between several."""
    em = person["email"].lower()
    pn = norm(person["name"])
    same = [i for i in rows if rows[i]["email"].lower() == em and i not in taken]
    if same:
        for i in same:                       # exact same human, named the same
            if norm(rows[i]["name"]) == pn:
                return i
        for i in same:                       # not a parenthesised variant row
            if "(" not in rows[i]["name"]:
                return i
        return same[0]
    #  no email match - a placeholder row waiting for this person's address
    for i in rows:
        if i in taken or rows[i]["email"].strip():
            continue
        if norm(rows[i]["name"]) == pn and pn:
            return i
    return None


def main():
    print("=" * 68)
    print(" COATES | UPDATE CONTACTS - Outlook list into the address book")
    print("=" * 68)
    try:
        import openpyxl
        from copy import copy
    except ImportError:
        print(" openpyxl isn't installed - run:  py -m pip install openpyxl")
        return 1
    if not os.path.isfile(BOOK):
        print(" Coates_Report_Recipients.xlsx isn't in this folder.")
        return 1

    src = sys.argv[1] if len(sys.argv) > 1 else find_csv()
    if not src or not os.path.isfile(src):
        print(" No Outlook contacts CSV found.")
        print(" Export from Outlook (People > Manage > Export contacts) and")
        print(" save it in this folder or in Downloads, then run me again.")
        return 1
    print(" Contacts : " + os.path.basename(src))

    live, B, err = live_companies()
    if err:
        #  Stopping is the whole point. A silent fallback here is what put
        #  wrong company names in the book the first time round.
        print("")
        print(" STOPPED - I can't tell what the reports call each company.")
        print(" Reason: " + err)
        print("")
        print(" The right names come from today's rental data, not from")
        print(" yesterday's report folders. Run 28_PULL_SITEIQ_EXPORTS.bat")
        print(" then 03_RUN_COMPANY_REPORT.bat, and run me again.")
        print(" Nothing has been changed.")
        return 1
    print(" Book     : Coates_Report_Recipients.xlsx")
    print(" Companies: {} live, read from today's rental data".format(len(live)))
    print(" Read     : {} contact(s) from Outlook".format(len(read_csv(src))))
    print("")

    people = read_csv(src)

    # ---- each contact's company, as the reports will spell it ---------
    unmatched = {}
    renamed = set()
    for p in people:
        if p["company"].upper() in INTERNAL:
            p["book"], p["tags"] = "ALL", "ALL"
            continue
        hit, why = match_company(p["company"], live, B)
        p["book"] = hit or p["company"]
        p["tags"] = DEFAULT_TAGS
        if hit and why != "exact":
            renamed.add((p["company"], hit))
        if not hit:
            unmatched[p["company"]] = unmatched.get(p["company"], 0) + 1

    if renamed:
        print(" Outlook name -> the name the report is built under:")
        for a, b in sorted(renamed):
            print("   {:<24} -> {}".format(a, b))
        print("")

    # ---- read the book, keeping every decision Andrew has made --------
    wb = openpyxl.load_workbook(BOOK)
    ws = wb["Recipients"]
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [str(v).strip().lower() if v is not None else "" for v in row]
        if vals and vals[0] == "company" and "email" in vals:
            hdr = i
            break
    if not hdr:
        print(" Couldn't find the header row in the Recipients sheet.")
        return 1

    rows, order = {}, []
    for i in range(hdr + 1, ws.max_row + 1):
        c = [ws.cell(row=i, column=j).value for j in range(1, 8)]
        if not any(str(v or "").strip() for v in c):
            continue
        rows[i] = {"company": str(c[0] or "").strip(),
                   "name": str(c[1] or "").strip(),
                   "email": str(c[2] or "").strip(),
                   "addas": str(c[3] or "").strip() or "To",
                   "include": str(c[4] or "").strip() or "No",
                   "reports": str(c[5] or "").strip(),
                   "notes": str(c[6] or "").strip()}
        order.append(i)

    # ---- tidy up anyone written into the book twice -------------------
    #  An earlier version of this script keyed people on email AND name,
    #  so one man spelled two ways ("S Wacher" / " Wachter") landed as two
    #  rows. Those books are already out there, so clean them here rather
    #  than making Andrew start again.
    merged = []
    for a_i in list(order):
        if a_i not in rows:
            continue
        for b_i in list(order):
            if b_i == a_i or b_i not in rows or a_i not in rows:
                continue
            if same_person(rows[a_i], rows[b_i]):
                keep, drop = (a_i, b_i) if a_i < b_i else (b_i, a_i)
                merged.append((rows[keep]["name"], rows[drop]["name"],
                               rows[keep]["email"] or rows[drop]["email"]))
                merge_into(rows[keep], rows[drop])
                del rows[drop]
                order.remove(drop)
    if merged:
        print(" Same person written in twice - merged into one row:")
        for a, b, em in merged:
            print("   {} + {}  ({})".format(a, b, em or "no email"))
        print("")

    # ---- fold the Outlook list in ------------------------------------
    taken, added, updated = set(), [], []
    for p in people:
        i = claim_row(p, rows, taken)
        note = " | ".join(x for x in (p["title"], p["mobile"]) if x)
        if i is not None:
            taken.add(i)
            r = rows[i]
            was = r["company"]
            #  An ALL row (the safety pack) is site-wide on purpose - the
            #  import must not drag it back to one company.
            if was.upper() != "ALL":
                newc = p["book"]
                #  A company that has handed everything back drops off the
                #  live list, and Outlook's shorter spelling would then
                #  overwrite the fuller name the report was built under -
                #  "Industec Filtration" back to "Industec", every run,
                #  in both directions. Keep the more specific name.
                if (newc not in live and was
                        and norm(was).startswith(norm(newc)) and was != newc):
                    newc = was
                r["company"] = newc
            if not r["email"]:
                r["email"] = p["email"]
            #  "Aburgun" is this script's own earlier guess, not something
            #  Andrew typed - safe to improve. A name he wrote is left
            #  exactly as he wrote it.
            if norm(r["name"]) == norm(p["email"].split("@")[0]):
                r["name"] = p["name"]
            if note:
                r["notes"] = note
            if was != r["company"]:
                updated.append((was, r["company"], p["email"], r["include"]))
            continue
        nxt = (max(rows) if rows else hdr) + 1
        rows[nxt] = {"company": p["book"], "name": p["name"],
                     "email": p["email"], "addas": "To",
                     "include": "No",          # never switched on by import
                     "reports": p["tags"], "notes": note}
        order.append(nxt)
        taken.add(nxt)
        added.append(p)

    # ---- book rows Outlook didn't cover, still filed under a dead name -
    for i in order:
        r = rows[i]
        if i in taken or not r["company"] or r["company"].upper() == "ALL":
            continue
        if r["company"] in live:
            continue
        hit, _ = match_company(r["company"], live, B)
        if hit and hit != r["company"]:
            updated.append((r["company"], hit, r["email"], r["include"]))
            r["company"] = hit

    if updated:
        #  Grouped, not one line per person - 27 identical Cement
        #  Australia lines buries the one that matters.
        print(" Book rows renamed to match the reports:")
        grp = {}
        for old, new, em, inc in updated:
            g = grp.setdefault((old, new), {"n": 0, "live": []})
            g["n"] += 1
            #  Only a company that HAS gear on hire builds a report, so
            #  only that case can have been dropping a live recipient.
            #  The rest were simply waiting for their first hire.
            if inc.lower() == "yes" and new in live and em:
                g["live"].append(em)
        for (old, new), g in sorted(grp.items()):
            n = "" if g["n"] == 1 else "   ({} people)".format(g["n"])
            print("   {:<26} -> {}{}".format(old[:26], new, n))
            for em in g["live"]:
                print("      {:<44} <-- WAS BEING DROPPED".format(em))
        print("")

    # ---- write it back ------------------------------------------------
    os.makedirs(BK, exist_ok=True)
    shutil.copy2(BOOK, os.path.join(
        BK, "Coates_Report_Recipients_%s.xlsx" % time.strftime("%Y%m%d_%H%M%S")))

    styles = [copy(ws.cell(row=hdr + 1, column=j)._style) for j in range(1, 8)]
    last = max(ws.max_row, hdr + len(order))
    for i in range(hdr + 1, last + 1):
        for j in range(1, 8):
            ws.cell(row=i, column=j).value = None
    for n, i in enumerate(order):
        r = rows[i]
        for j, v in enumerate([r["company"], r["name"], r["email"], r["addas"],
                               r["include"], r["reports"], r["notes"]], 1):
            cell = ws.cell(row=hdr + 1 + n, column=j)
            cell.value = v
            cell._style = styles[j - 1]
    try:
        wb.save(BOOK)
    except PermissionError:
        print(" ! The book is open in Excel. Close it and run me again.")
        print(" ! Nothing has been changed.")
        return 1

    # ---- say plainly what happened ------------------------------------
    on = sum(1 for i in order if rows[i]["include"].lower() == "yes")
    print(" Added    : {} new contact(s), all set Include = No".format(len(added)))
    print(" Book now : {} row(s), {} switched on".format(len(order), on))
    print(" Backup   : Updates\\contact_backups\\")
    if added:
        print("")
        print(" NEW - open the book and put Yes against anyone who should")
        print(" actually receive their company's report:")
        bycomp = {}
        for p in added:
            bycomp.setdefault(p["book"], []).append(p)
        for comp in sorted(bycomp):
            tail = "" if comp in live or comp == "ALL" else "   (no gear on hire)"
            print("   {}{}".format(comp, tail))
            for p in sorted(bycomp[comp], key=lambda x: x["name"]):
                print("      {:<24} {}".format(p["name"][:24], p["email"]))
    if unmatched:
        print("")
        print(" I COULDN'T PLACE THESE - no company on hire matches the name")
        print(" Outlook has. They're in the book, but no report will reach")
        print(" them until the company name lines up. Tell me the right one")
        print(" and I'll bridge it, same as DKE Group -> Dark Knight:")
        for c in sorted(unmatched):
            print("   {:<24} {} contact(s)".format(c, unmatched[c]))
    print("")
    print(" Companies on hire right now:")
    for c in live:
        print("   " + c)
    print("")
    print(" Open Coates_Report_Recipients.xlsx, flip Include to Yes for the")
    print(" people you want on, save, then run your reports as normal.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
