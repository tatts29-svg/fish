# ==========================================================================
#  COATES | K2 CLEAN REGISTER  ->  Cement_Australia_Report_2026_K2_v11_5_CLEAN.xlsx
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  A clean, barcode-free workbook built straight off your current data.
#    - ITEM_NUMBER is the asset identifier everywhere. ITEM_BARCODE is never
#      read, shown or written - it appears nowhere.
#    - Plant ID is shown wherever one is allocated (matched by ITEM_NUMBER
#      against Plant_ID_Register.csv).
#    - Damage / Consumables / Transport charges are pulled from the Charge
#      Reporter register so that data sits right here in the sheet.
#
#  25 Jul 2026 - COMPLIANCE CHECK COLUMN (A. Fisher)
#  Every tab that names an item now carries a "Compliance check" column
#  sitting straight after the item description - current tag colour, daily
#  pre-start and logbook, back daily - read off
#  K2_MASTER_EQUIPMENT_PRICING.xlsx through equipment_compliance.py.
#  On a screen we put that requirement UNDER the description; a spreadsheet
#  can't do that, so it gets its own column instead - same words, same one
#  source. The point is that a bloke can filter or sort this register and
#  see exactly what has to be tagged, logged or brought back, and the
#  Coates team can check the gear over before it goes out again. No flag in
#  the master means a blank cell - nothing is guessed in here.
#
#  Never touches your live workbook. Re-run any time to refresh from the
#  latest RENTAL_STOCK / ON_HIRE / charge data.
# ==========================================================================
import os, csv, datetime as dt
import master_equipment
import equipment_compliance as EC
import openpyxl
# Master display names (item-number keyed) - the CLEAN register shows
# Andrew's corrected descriptions; empty-safe without the file.
MASTER = master_equipment.load(os.path.dirname(os.path.abspath(__file__)),
                               quiet=True)
# Same master feeds the compliance column, so one fix in the spreadsheet
# fixes the workbook and every other report on the next run. Safe with no
# master file: the column simply comes out blank.
EC.bind(MASTER)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(HERE, "Cement_Australia_Report_2026_K2_v11_5_CLEAN.xlsx")

# ---- Coates house style ---------------------------------------------------
ORANGE = "F26222"; INK = "0E1116"; SLATE = "1F2937"
HEADFILL = PatternFill("solid", fgColor=ORANGE)
BANDFILL = PatternFill("solid", fgColor="F5F6F8")
IDFILL   = PatternFill("solid", fgColor="FCE9DE")   # soft orange for Plant ID
# Compliance check reads as its own thing, the same way Plant ID does - a
# soft steel-blue panel so the eye finds the gear that has to be checked
# without hunting. Deliberately NOT the tag colour: the tag colour rolls
# over (see equipment_compliance.TAG_PERIODS) and a hard-coded fill here
# would go stale and start telling people the wrong thing. The words in the
# cell carry the colour; the fill just says "this one needs a check".
COMPFILL = PatternFill("solid", fgColor="EAF1FA")
FONT     = "Arial"
HEADFONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
CELLFONT = Font(name=FONT, size=10, color="1F2937")
IDFONT   = Font(name=FONT, size=10, bold=True, color=ORANGE)
COMPFONT = Font(name=FONT, size=10, bold=True, color="1F5FA8")
COMPCOL  = "Compliance check"       # header name the styling keys off
THIN     = Side(style="thin", color="E2E5EA")
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def norm(v):
    return "" if v is None else str(v).strip()


def load_plant_ids():
    """ITEM_NUMBER (Asset Number in the register) -> Plant ID. 74/74 match."""
    p = os.path.join(HERE, "Plant_ID_Register.csv")
    m = {}
    if os.path.isfile(p):
        with open(p, newline="") as f:
            for row in csv.DictReader(f):
                a = norm(row.get("Asset Number")); i = norm(row.get("Plant ID"))
                if a and i:
                    m[a] = i
    return m


def read_sheet(path, sheet=None):
    """Return (headers, list-of-dicts) from a spreadsheet, cached values."""
    if not os.path.isfile(path):
        return [], []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    hdr = [norm(h) for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        out.append({hdr[i]: r[i] for i in range(len(hdr)) if i < len(r)})
    return hdr, out


def as_at():
    """Data as-at from ON_HIRE / RENTAL_STOCK reference info, else today."""
    for f, sh in (("ON_HIRE.xlsx", "REFERENCE_INFO"), ("RENTAL_STOCK.xlsx", "REFERENCE_INFO")):
        f = next((c for c in (os.path.join("Data_SiteIQ", f), f)
                  if os.path.isfile(os.path.join(HERE, c))), f)
        p = os.path.join(HERE, f)
        if os.path.isfile(p):
            try:
                wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
                if sh in wb.sheetnames:
                    for row in wb[sh].iter_rows(values_only=True):
                        for c in row:
                            s = norm(c)
                            if "/2026" in s or "/2025" in s:
                                wb.close(); return s
                wb.close()
            except Exception:
                pass
    return dt.datetime.now().strftime("%d/%m/%Y %H:%M")


# ---- sheet writer ---------------------------------------------------------
def write_table(ws, columns, records, id_col="Plant ID"):
    """columns = [(header, key_or_fn, width), ...]; records = list of dicts."""
    # header
    for c, (head, _key, width) in enumerate(columns, 1):
        cell = ws.cell(1, c, head)
        cell.fill = HEADFILL; cell.font = HEADFONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 22
    # body
    for i, rec in enumerate(records):
        r = i + 2
        band = (i % 2 == 1)
        for c, (head, key, _w) in enumerate(columns, 1):
            val = key(rec) if callable(key) else rec.get(key, "")
            cell = ws.cell(r, c, val if val not in (None, "") else "")
            cell.font = CELLFONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if head == id_col and val not in (None, "", "—"):
                cell.font = IDFONT
                cell.fill = IDFILL
            elif head == COMPCOL and val not in (None, "", "—"):
                # Only a line that actually has a requirement gets the
                # treatment - a blank compliance cell bands like any other
                # so the zebra stripes still read straight down the page.
                cell.font = COMPFONT
                cell.fill = COMPFILL
            elif band:
                cell.fill = BANDFILL
    ws.freeze_panes = "A2"
    if records:
        ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(columns)), len(records) + 1)
    else:
        note = ws.cell(2, 1, "No charges logged yet — raise them in the Charge Reporter and they appear here.")
        note.font = Font(name=FONT, italic=True, color="8B9099", size=10)
    # tidy printing: landscape, fit to page width, repeat the header row
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:1"


def style_cover(ws, asat, counts):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 66
    band = ws.cell(2, 2, "COATES")
    band.font = Font(name=FONT, bold=True, size=26, color=ORANGE)
    ws.cell(2, 3, "Equipped for anything").font = Font(name=FONT, size=12, color="1F2937")
    ws.cell(4, 2, "Cement Australia K2 Shutdown 2026 — Gladstone").font = Font(name=FONT, bold=True, size=14, color=INK)
    ws.cell(5, 2, "Clean Equipment Register  ·  v11.5 CLEAN").font = Font(name=FONT, size=12, color="1F2937")
    ws.cell(6, 2, "Data as at %s   ·   POWERED BY SITEIQ   ·   Author: Andrew Fisher" % asat).font = Font(name=FONT, size=9, color="8B9099")

    y = 8
    lines = [
        ("What this is", True),
        ("A clean copy of your K2 equipment data. ITEM_NUMBER is the asset identifier on every", False),
        ("tab — nothing else stands in for it. Plant ID is shown wherever one is allocated. Damage,", False),
        ("Consumables and Transport charges are pulled in from the Charge Reporter. Your live", False),
        ("workbook is never touched — re-run 08_RUN_CLEAN_REPORT.bat any time to refresh.", False),
        ("", False),
        ("Tabs in this workbook", True),
    ]
    tabs = [
        ("On Hire", "Everything currently on hire — %d items, keyed on ITEM_NUMBER + Plant ID" % counts.get("onhire", 0)),
        ("Plant — Site", "Site plant on the ground — %d units, Plant ID front and centre" % counts.get("plant", 0)),
        ("Rental Stock", "Full stock list — %d items, clean (filter by Storage Unit / Status)" % counts.get("stock", 0)),
        ("Damage Charges", "Damage recoveries logged in the Charge Reporter (%d)" % counts.get("damage", 0)),
        ("Consumables Charges", "Consumables sold to the client (%d)" % counts.get("consumables", 0)),
        ("Transport Charges", "Transport charges logged (%d)" % counts.get("transport", 0)),
    ]
    for text, head in lines:
        cell = ws.cell(y, 2, text)
        cell.font = Font(name=FONT, bold=head, size=11 if head else 10,
                         color=ORANGE if head else "1F2937")
        y += 1
    for name, desc in tabs:
        ws.cell(y, 2, name).font = Font(name=FONT, bold=True, size=10, color=INK)
        ws.cell(y, 3, desc).font = Font(name=FONT, size=10, color="1F2937")
        y += 1
    y += 1
    ws.cell(y, 2, "Note").font = Font(name=FONT, bold=True, size=10, color=ORANGE)
    ws.cell(y + 1, 2, "Plant ID is matched to each unit by ITEM_NUMBER against Plant_ID_Register.csv.").font = Font(name=FONT, size=9, color="8B9099")
    ws.cell(y + 2, 2, "A blank Plant ID simply means no site number is allocated to that item.").font = Font(name=FONT, size=9, color="8B9099")
    ws.cell(y + 3, 2, "Compliance check sits beside the description — the tag colour, pre-start and logbook, or same-day").font = Font(name=FONT, size=9, color="8B9099")
    ws.cell(y + 4, 2, "return that item needs. It comes off K2_MASTER_EQUIPMENT_PRICING.xlsx; blank means nothing recorded.").font = Font(name=FONT, size=9, color="8B9099")
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    # print area follows the last note line written above - keep in step if
    # another note is added, otherwise the cover prints short.
    ws.print_area = "A1:C%d" % (y + 5)


def main():
    pid = load_plant_ids()

    def pid_of(itemno):
        return pid.get(norm(itemno), "")

    # ---- source data ------------------------------------------------------
    _rs_p = next((c for c in (os.path.join(HERE, "Data_SiteIQ", "RENTAL_STOCK.xlsx"),
              os.path.join(HERE, "RENTAL_STOCK.xlsx")) if os.path.isfile(c)),
             os.path.join(HERE, "RENTAL_STOCK.xlsx"))
    _, stock = read_sheet(_rs_p, "RENTAL_STOCK")
    _, onh   = read_sheet(os.path.join(HERE, "ON_HIRE.xlsx"), "ON_HIRE")

    # storage unit (category) by item number, from the master stock
    su_by_item = {norm(r.get("ITEM_NUMBER")): norm(r.get("STORAGE_UNIT")) for r in stock}

    asat = as_at()
    wb = openpyxl.Workbook()

    # ---- Cover (fill last once counts are known) --------------------------
    cover = wb.active; cover.title = "Cover"

    # ---- On Hire ----------------------------------------------------------
    ws = wb.create_sheet("On Hire")
    onh_recs = []
    for r in onh:
        itemno = norm(r.get("ITEM_NUMBER"))
        onh_recs.append({
            "pid": pid_of(itemno), "item": itemno,
            "desc": MASTER.disp(itemno, norm(r.get("ITEM_DESCRIPTION"))),
            "cat": su_by_item.get(itemno, "") or norm(r.get("PRODUCT_FAMILY")),
            "fam": norm(r.get("PRODUCT_FAMILY")),
            "hirer": norm(r.get("HIRER_NAME")), "company": norm(r.get("COMPANY")),
            "start": norm(r.get("START_DATE")), "rate": r.get("SHIFT_RATE"),
        })
    onh_recs.sort(key=lambda x: (x["cat"], x["desc"], x["item"]))
    write_table(ws, [
        ("Plant ID", "pid", 9), ("ITEM_NUMBER", "item", 20), ("Item description", "desc", 42),
        (COMPCOL, lambda x: EC.badges_plain(x["item"], x["desc"]), 34),
        ("Category", "cat", 20), ("Product family", "fam", 26), ("Hirer", "hirer", 20),
        ("Company", "company", 30), ("On-hire date", "start", 13),
        ("Shift rate ($)", lambda x: x["rate"] if isinstance(x["rate"], (int, float)) else "", 12),
    ], onh_recs)

    # ---- Plant — Site -----------------------------------------------------
    ws = wb.create_sheet("Plant — Site")
    onh_by_item = {norm(r.get("ITEM_NUMBER")): r for r in onh}
    plant_recs = []
    for r in stock:
        if norm(r.get("STORAGE_UNIT")).upper() != "CEMENT SITE PLANT":
            continue
        itemno = norm(r.get("ITEM_NUMBER"))
        oh = onh_by_item.get(itemno)
        plant_recs.append({
            "pid": pid_of(itemno), "item": itemno,
            "desc": MASTER.disp(itemno, norm(r.get("ITEM_DESCRIPTION"))),
            "prod": norm(r.get("PRODUCT")), "var": norm(r.get("PRODUCT_VARIANT")),
            "status": norm(r.get("ITEM_STATUS")),
            "onhire": norm(r.get("ON_HIRE_DATE")),
            "hirer": norm(oh.get("HIRER_NAME")) if oh else "",
        })

    def _pk(x):
        try:
            return (0, int(x["pid"]))
        except (ValueError, TypeError):
            return (1, 0)
    plant_recs.sort(key=lambda x: (_pk(x), x["item"]))
    write_table(ws, [
        ("Plant ID", "pid", 9), ("ITEM_NUMBER", "item", 20), ("Item description", "desc", 40),
        (COMPCOL, lambda x: EC.badges_plain(x["item"], x["desc"]), 34),
        ("Product", "prod", 26), ("Variant", "var", 26), ("Status", "status", 16),
        ("On-hire date", "onhire", 13), ("Hirer", "hirer", 20),
    ], plant_recs)

    # ---- Rental Stock (all) ----------------------------------------------
    ws = wb.create_sheet("Rental Stock")
    stock_recs = []
    for r in stock:
        itemno = norm(r.get("ITEM_NUMBER"))
        stock_recs.append({
            "pid": pid_of(itemno), "item": itemno,
            "desc": MASTER.disp(itemno, norm(r.get("ITEM_DESCRIPTION"))),
            "fam": norm(r.get("PRODUCT_FAMILY")), "prod": norm(r.get("PRODUCT")),
            "status": norm(r.get("ITEM_STATUS")), "toolstore": norm(r.get("TOOLSTORE")),
            "su": norm(r.get("STORAGE_UNIT")), "onhire": norm(r.get("ON_HIRE_DATE")),
        })
    stock_recs.sort(key=lambda x: (x["su"], x["desc"], x["item"]))
    write_table(ws, [
        ("Plant ID", "pid", 9), ("ITEM_NUMBER", "item", 20), ("Item description", "desc", 42),
        (COMPCOL, lambda x: EC.badges_plain(x["item"], x["desc"]), 34),
        ("Product family", "fam", 26), ("Product", "prod", 24), ("Status", "status", 16),
        ("Toolstore", "toolstore", 16), ("Storage unit", "su", 18), ("On-hire date", "onhire", 13),
    ], stock_recs)

    # ---- Charge tabs (from the Charge Reporter register) ------------------
    reg = os.path.join(HERE, "Coates_K2_Charge_Reporter", "CHARGE_REGISTER.xlsx")
    charge_counts = {}

    def charge_tab(title, sheet, columns, key):
        _, recs = read_sheet(reg, sheet)
        ws = wb.create_sheet(title)
        write_table(ws, columns, recs, id_col="Plant ID")
        charge_counts[key] = len(recs)

    charge_tab("Damage Charges", "Damage - Breakdown", [
        ("Charge ID", "Charge ID", 14), ("Date", "Date", 12), ("Company", "Client / company", 26),
        ("ITEM_NUMBER", "Asset no. (item number)", 20), ("Plant ID", "Plant ID", 9),
        ("Item description", "Item description", 34),
        # Charge rows come straight out of the Charge Reporter, so the keys
        # are its column names rather than the short ones used above.
        (COMPCOL, lambda x: EC.badges_plain(x.get("Asset no. (item number)"),
                                           x.get("Item description")), 34),
        ("What happened", "What happened / why - the story", 44),
        ("Amount to costing ($)", "Amount to Costing", 16), ("Cost status", "Cost status", 14),
        ("Cost line", "Cost Line", 16),
    ], "damage")
    charge_tab("Consumables Charges", "Consumables sold", [
        ("Charge ID", "Charge ID", 14), ("Date", "Date", 12), ("Company", "Client / company", 26),
        ("Consumables supplied", "Consumables supplied", 32), ("Qty", "Quantity", 8),
        ("Unit", "Unit", 10), ("Unit price ($)", "Unit price", 12), ("Total ($)", "Total charge", 12),
        ("Status", "Charge status", 14), ("Cost line", "Cost Line", 16),
    ], "consumables")
    charge_tab("Transport Charges", "Transport", [
        ("Charge ID", "Charge ID", 14), ("Date", "Date", 12), ("Company", "Client / company", 26),
        ("From", "From", 20), ("To", "To", 20), ("Load / moved", "Load / what was moved", 28),
        ("Truck / rego", "Truck / rego", 16), ("Total ($)", "Total charge", 12),
        ("Status", "Charge status", 14), ("Cost line", "Cost Line", 16),
    ], "transport")

    # ---- Cover ------------------------------------------------------------
    counts = {"onhire": len(onh_recs), "plant": len(plant_recs), "stock": len(stock_recs),
              "damage": charge_counts.get("damage", 0),
              "consumables": charge_counts.get("consumables", 0),
              "transport": charge_counts.get("transport", 0)}
    style_cover(cover, asat, counts)

    wb.save(OUT)
    print("=" * 64)
    print(" COATES | K2 CLEAN REGISTER (v11.5 CLEAN)")
    print(" ITEM_NUMBER is the identifier · Plant ID where allocated · no barcode")
    print(" On hire: %d   Site plant: %d   Rental stock: %d" % (counts["onhire"], counts["plant"], counts["stock"]))
    print(" Plant IDs allocated: %d" % sum(1 for r in stock_recs if r["pid"]))
    print(" Charges  damage:%d  consumables:%d  transport:%d" % (counts["damage"], counts["consumables"], counts["transport"]))
    print(" Saved: %s" % OUT)
    print("=" * 64)


if __name__ == "__main__":
    main()
