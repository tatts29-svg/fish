#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | MY GEAR THUMBNAILS - a picture for every product variant
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 30 Jul 2026): "having a thumbnail of what the item
#  looks like in my gear ... what if i got thumbnails for everything /
#  for product variants". The register speaks variant codes
#  (BATCORDMIL18V5AH is every Milwaukee 18V 5Ah battery on site), so
#  ONE photo per variant covers every serialised item behind it.
#
#  THE DEAL:
#    Photos\                 <- Andrew drops pictures in, named by the
#                               variant code (BATCORDMIL18V5AH.jpg) or
#                               consumable SKU (CEME9016-590674.jpg).
#                               Any size, any source - phone, website.
#    Gear_Lookup\thumbs\     <- this module shrinks each one to a fast
#                               ~120px thumbnail the pages lazy-load.
#
#  Run 56_PHOTO_HUNT for the wanted list (every code, its plain name,
#  and search links). The shrinking uses the same headless Edge/Chrome
#  the PDFs use - browser_engine finds it - so nothing gets installed.
#  Incremental: only new or changed photos are re-shrunk, so the
#  morning build stays quick. No browser on the machine = photos wait,
#  the build says so, nothing breaks.
# =====================================================================
import base64
import json
import os
import re
import subprocess
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
SIZE = 384           # thumbnail edge, px - showroom-grid sharp on phones

PHOTOS_README = """\
DROP GEAR PHOTOS IN HERE
========================
One picture per product variant, named by its code:

    BATCORDMIL18V5AH.jpg      <- covers EVERY Milwaukee 18V 5Ah battery
    CEME9016-590674.jpg       <- a consumable, by its SKU number

Run 56_PHOTO_HUNT.bat for the full wanted list - every code, what it
is in plain English, and a search button each. jpg or png, any size,
any source (phone photo of the shelf beats a studio shot). The next
04 build shrinks them and the pictures appear in My Gear.

CLOSE COUNTS. Underscores and spaces in the filename don't matter,
and extra words on the end are fine once the code part is there:

    MOTOROLA_DP4801E_TWO_WAY_RADIO_WITH_HANDPIECE.jpg
      -> covers the model key MOTOROLADP4801ETWOWAYRADIO,
         and through it EVERY radio, every serial number.
"""


_SER_RE = re.compile(r'^[0-9]{3}[A-Z0-9]{5,}$')


def derived_keys(desc):
    """Photo keys for gear SiteIQ gives NO product variant - the radios
    and gas monitors (31 Jul 2026).

    'Motorola DP4801e Two-Way Radio - Serial 871TNK7668'
        -> ('871TNK7668', 'MOTOROLADP4801ETWOWAYRADIO')
    'Multi-Gas Detector - Honeywell BW Flex - Serial GM206136'
        -> ('GM206136', 'MULTIGASDETECTORHONEYWELLBWFLEX')

    A photo named by the SERIAL covers that one unit; a photo named by
    the model code covers the fleet. The pages try serial first, then
    model, then the monogram.

    The register writes the serial behind the word SERIAL, so split on
    the word, not on guessing the serial's shape - gas serials like
    GM206136 start with letters and a shape-guess misses them, which
    left every monitor keyed apart from its own fleet and no photo able
    to cover them (caught 31 Jul 2026: 'none have pictures attached')."""
    d = re.sub(r'[^A-Z0-9 ]', '', str(desc or '').strip().upper())
    m = re.search(r'\bSERIAL\s+([A-Z0-9]+)\s*$', d)
    if m:
        return m.group(1), ''.join(d[:m.start()].split())[:40]
    toks = d.split()
    ser = ''
    if (toks and _SER_RE.match(toks[-1])
            and any(c.isdigit() for c in toks[-1])
            and any(c.isalpha() for c in toks[-1])):
        ser = toks[-1]
        toks = toks[:-1]
    return ser, ''.join(toks)[:40]


def safe_name(code):
    """The filename a code can actually be saved under. Windows bans
    \\ / : * ? " < > | in filenames, and 275 register codes carry one
    (SOCKET1/2DR11MM, TRAILERMINIEXCAVATOR<4.5T) - so their photos
    arrive with those swapped to _ (Andrew's 790-variant pack,
    31 Jul 2026). Every lookup between a code and a file goes
    through here; the pages do the same swap in tsafe()."""
    return re.sub(r'[\\/:*?"<>|]', '_', code)


def photos_dir(here=None):
    d = os.path.join(here or HERE, 'Photos')
    os.makedirs(d, exist_ok=True)
    rd = os.path.join(d, '_READ_ME_FIRST.txt')
    if not os.path.isfile(rd):
        try:
            with open(rd, 'w', encoding='utf-8') as f:
                f.write(PHOTOS_README)
        except OSError:
            pass
    return d


_MB_CACHE = {}


def _master_book(here):
    """The master equipment list, loaded once per folder. It is what
    renames a register description into the words the workers' page
    shows, so photo keys can be derived on both sides of that rename."""
    key = here or HERE
    if key not in _MB_CACHE:
        try:
            import master_equipment as _ME
            _MB_CACHE[key] = _ME.load(key, quiet=True)
        except Exception:
            _MB_CACHE[key] = None
    return _MB_CACHE[key]


def variant_register(here=None):
    """Every code a photo can be named after: rental PRODUCT_VARIANTs
    and consumable SKU numbers, with plain names, families and how many
    items ride behind each - straight off the newest exports."""
    import glob
    import openpyxl
    here = here or HERE
    out = {}

    def gfind(pat):
        hits = []
        for d in (os.path.join(here, 'Data_SiteIQ'), here):
            hits += [p for p in glob.glob(os.path.join(d, pat))
                     if not os.path.basename(p).startswith('~$')]
        return max(hits, key=os.path.getmtime) if hits else None

    rp = gfind('RENTAL_STOCK*.xlsx')
    if rp:
        wb = openpyxl.load_workbook(rp, read_only=True, data_only=True)
        rows = list(wb['RENTAL_STOCK'].iter_rows(values_only=True))
        wb.close()
        ix = {str(v or '').strip(): i for i, v in enumerate(rows[0])}
        for r in rows[1:]:
            code = str(r[ix.get('PRODUCT_VARIANT', -1)] or '').strip().upper()
            drv = 0
            alt2 = ''
            if not code:
                #  radios / gas monitors / crowsfeet: no variant in
                #  SiteIQ, so a MODEL key is derived from the name.
                #
                #  AND THE NAME DEPENDS WHO IS ASKING. The stores board
                #  derives from the RAW register description; the
                #  workers' catalogue derives from the name after the
                #  master equipment list has tidied it. Same asset, two
                #  codes: 12DRIVECROWSFEET27MM and CROWFOOT27MM12INDRIVE.
                #  Registering only the first is why Andrew's crowsfoot
                #  photo never appeared in the workers' catalogue no
                #  matter how many times it was pinned (3 Aug 2026).
                _raw = r[ix.get('ITEM_DESCRIPTION', -1)]
                _ser, code = derived_keys(_raw)
                drv = 1
                try:
                    import mygear_store as _MST
                    _tidied = _MST._tidy(
                        _raw, _master_book(here),
                        r[ix.get('ITEM_NUMBER', -1)]
                        if 'ITEM_NUMBER' in ix else '')
                    _s2, alt2 = derived_keys(_tidied)
                    if alt2 == code:
                        alt2 = ''
                except Exception:
                    alt2 = ''
                if not code:
                    code, alt2 = alt2, ''
                if not code:
                    continue
            e = out.setdefault(code, {'n': '', 'f': '', 'q': 0, 'k': 'hire',
                                      'drv': drv})
            e['q'] += 1
            if not e['n']:
                e['n'] = str(r[ix.get('ITEM_DESCRIPTION', -1)] or '').strip()
                e['f'] = str(r[ix.get('PRODUCT_FAMILY', -1)] or '').strip().title()
            if alt2:
                e2 = out.setdefault(alt2, {'n': e['n'], 'f': e['f'],
                                           'q': 0, 'k': 'hire', 'drv': 1})
                e2['q'] += 1
    sp = gfind('SALES_STOCK*.xlsx')
    if sp:
        wb = openpyxl.load_workbook(sp, read_only=True, data_only=True)
        ws = wb['SALES_STOCK'] if 'SALES_STOCK' in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        ix = {str(v or '').strip(): i for i, v in enumerate(rows[0])}
        for r in rows[1:]:
            code = str(r[ix.get('SKU_NUMBER', -1)] or '').strip().upper()
            if not code or code in out:
                continue
            desc = str(r[ix.get('SKU_DESCRIPTION', -1)] or '').strip()
            if not desc:
                continue
            try:
                q = int(float(r[ix.get('AVAILABLE_QUANTITY', -1)] or 0))
            except (TypeError, ValueError):
                q = 0
            out[code] = {'n': desc, 'f': 'Consumables', 'q': q, 'k': 'cons'}
    return out


def _photo_files(here=None):
    """Photos\\ and one level of subfolders - a zip of pictures dragged
    in whole (folder and all) still counts, nobody has to know to
    flatten it first (30 Jul 2026, Andrew's consumables pack arrived
    exactly like that)."""
    d = photos_dir(here)
    out = {}
    dirs = [d] + sorted(os.path.join(d, s) for s in os.listdir(d)
                        if os.path.isdir(os.path.join(d, s)))
    for dd in dirs:
        for fn in sorted(os.listdir(dd)):
            stem, ext = os.path.splitext(fn)
            if (ext.lower() in ('.jpg', '.jpeg', '.png')
                    and not fn.startswith('_')
                    and stem.strip().upper() not in out):
                out[stem.strip().upper()] = os.path.join(dd, fn)
    return out


def _norm(s):
    """A code with everything but the letters and digits stripped -
    what two names have to share to be the same thing."""
    return re.sub(r'[^A-Z0-9]', '', str(s or '').upper())


#  ANDREW'S OWN PHOTOS, PINNED BY HAND (3 Aug 2026: "just hardcode
#  these in be easier in the update"). His filenames on his machine,
#  spelled exactly as he saved them, mapped straight to their register
#  codes - no fuzzy rule between him and his own pictures. A FAMILY:
#  value claims every code that starts with the stem, so the one
#  crowsfoot photo covers all fourteen sizes without a rename.
#  Keys are matched on the normalised stem, so underscores don't count.
#  TWO VOCABULARIES, AND THE PINS MUST SPEAK BOTH (3 Aug 2026).
#  The stores board and Fleet Details use SiteIQ's register codes.
#  The workers' catalogue renames everything through the master
#  equipment list, so the SAME gas monitor is
#      HONEYWELLBWFLEXMULTIGASDETECTOR    on the board
#      MULTIGASDETECTORHONEYWELLBWFLEX    in the workers' catalogue
#  and the same crowsfoot is 12DRIVECROWSFEET27MM one side and
#  CROWFOOT27MM12INDRIVE the other. Every earlier fix I made was on
#  the register side alone, which is exactly why Andrew kept seeing
#  the monitors and the crowsfeet missing after each one. Both
#  spellings are pinned here now.
PINNED = {
    'HONEYWELLBWFLEX4GASMONITOR': [
        'HONEYWELLBWFLEXMULTIGASDETECTOR',        # register / board
        'MULTIGASDETECTORHONEYWELLBWFLEX',        # master / workers
        'FAMILY:MULTIGASDETECTORHONEYWELLBWFLEX'],
    'MOTOROLADP4801ETWOWAYRADIOWITHHANDPIECE': [
        'MOTOROLADP4801ETWOWAYRADIO',
        'FAMILY:MOTOROLADP4801ETWOWAYRADIO'],
    'MOTOROLANNTN8129IMPRESBATTERY': [
        'MOTOROLANNTN8129IMPRESBATTERY',
        'RADIOBATTERYMOTOROLANNTN8129IMPRES',
        'FAMILY:RADIOBATTERYMOTOROLANNTN8129IMPRES'],
    '12DRIVECROWSFEET43MM': [
        'FAMILY:12DRIVECROWSFEET',                # register spelling
        'FAMILY:CROWFOOT'],                       # master spelling
}


def alias_photos(photos, codes, loose=None):
    """Photos named ALMOST right still land (Andrew's radio and gas
    monitor shots, 31 Jul 2026: none matched - underscores, a trailing
    WITH_HANDPIECE, and a gas photo worded like the box while the
    register words it like the report).

    Three ways a photo is claimed for a code it wasn't exactly named as:
      1. Same letters and digits - underscores, spaces and dashes in
         the filename don't matter.
      2. The filename STARTS with the code and runs on - extra words
         like WITH_HANDPIECE are fine - but only for long codes
         (16+ characters), so a photo called SOCKET.jpg can never
         claim every socket in the register.
      3. For DERIVED codes only (loose = {code: plain name} - the
         radios and gas monitors, whose codes never appear on any
         sheet a filename could be copied from): the photo's words are
         matched against the item's plain name. HONEYWELL_BW_FLEX_4_
         GAS_MONITOR shares Honeywell, BW, Flex and Gas with
         'Multi-Gas Detector - Honeywell BW Flex' - that's a claim.
         Needs at least 3 shared words and 60% of the photo's words,
         so it stays too tight to misfire on the rest of the store.

    The register keys for radios and gas monitors are model codes;
    every serialised unit falls back to its model photo on the pages,
    so one claimed photo pictures the whole fleet."""
    if not photos:
        return photos
    loose = loose or {}
    normed, toks = {}, {}
    for stem, path in photos.items():
        normed.setdefault(_norm(stem), path)
        toks[path] = set(t for t in re.split(r'[^A-Z0-9]+', stem.upper()) if t)
    #  the hand-pinned claims first - they are the spoken word, and no
    #  automatic rule below may overrule or miss them
    for pstem, targets in PINNED.items():
        path = normed.get(pstem)
        if not path:
            continue
        for t in targets:
            if t.startswith('FAMILY:'):
                fam = _norm(t[7:])
                for code in codes:
                    c = safe_name(str(code or '').strip().upper())
                    if c and _norm(c).startswith(fam) and c not in photos:
                        photos[c] = path
            else:
                if t not in photos:
                    photos[t] = path
    for code in codes:
        code = safe_name(str(code or '').strip().upper())
        if not code or code in photos:
            continue
        nc = _norm(code)
        if not nc:
            continue
        hit = normed.get(nc)
        if hit is None and len(nc) >= 16:
            for ns in normed:
                if ns.startswith(nc):
                    hit = normed[ns]
                    break
        #  ONE PHOTO, WHOLE FAMILY (Andrew, 3 Aug 2026: "crows feet
        #  all look the same"). A photo named WITHOUT the size claims
        #  every size behind it: 12DRIVECROWSFEET.jpg covers
        #  1/2DRIVECROWSFEET27MM through 50MM, one picture for nine
        #  codes. The stem must be 12+ characters and the code must
        #  START with it, so a photo called SOCKET.jpg still cannot
        #  claim every socket in the register - and a photo that DOES
        #  name a size (43MM) claims only that size, exactly as now.
        if hit is None:
            for ns in normed:
                if len(ns) >= 12 and nc.startswith(ns):
                    hit = normed[ns]
                    break
        if hit is None and code in loose:
            name = set(t for t in re.split(
                r'[^A-Z0-9]+', str(loose[code] or '').upper()) if t)
            best, bestn = None, 0
            for path, ft in toks.items():
                got = len(ft & name)
                if got >= 3 and got >= 0.6 * len(ft) and got > bestn:
                    best, bestn = path, got
            hit = best
        if hit is not None:
            photos[code] = hit
    return photos


def refresh(here=None, quiet=False):
    """Shrink new/changed photos into Gear_Lookup\\thumbs. Returns
    (photos_found, thumbs_made_now, thumbs_ready_total)."""
    here = here or HERE
    photos = _photo_files(here)
    #  near-enough filenames claim their register code - the register
    #  read can fail quietly (no exports on this machine = no aliases,
    #  exact names still work)
    try:
        reg = variant_register(here)
        photos = alias_photos(photos, reg.keys(),
                              loose={c: v.get('n', '') for c, v in reg.items()
                                     if v.get('drv')})
    except Exception:
        pass
    tdir = os.path.join(here, 'Gear_Lookup', 'thumbs')
    os.makedirs(tdir, exist_ok=True)

    def tpath(code):
        return os.path.join(tdir, code + '.jpg')
    #  bumping SIZE regenerates the lot once: a marker file remembers
    #  the size the folder was built at (31 Jul 2026 - grid went big)
    marker = os.path.join(tdir, '_size.txt')
    try:
        old_size = open(marker).read().strip()
    except OSError:
        old_size = ''
    rebuild_all = old_size != str(SIZE)
    pend = [(c, p) for c, p in photos.items()
            if rebuild_all or not os.path.isfile(tpath(c))
            or os.path.getmtime(tpath(c)) < os.path.getmtime(p)]
    ready = sum(1 for c in photos if os.path.isfile(tpath(c)))
    if not pend:
        return (len(photos), 0, ready)
    try:
        with open(marker, 'w') as f:
            f.write(str(SIZE))
    except OSError:
        pass
    try:
        import browser_engine
        browser = browser_engine.find_browser()
        extra = browser_engine.extra_args()
    except Exception:
        browser, extra = None, []
    if not browser:
        if not quiet:
            print('  Thumbnails: {} new photo(s) waiting - no headless '
                  'browser on this machine to shrink them.'.format(len(pend)))
        return (len(photos), 0, ready)

    made = 0
    prof = os.path.join(tempfile.gettempdir(), 'coates_thumbs_prof')
    #  smaller batches and a longer clock: 50 x 1024px images could
    #  outrun the old 30s budget on a slow laptop, and every image in a
    #  timed-out batch was silently skipped - the page then shows a
    #  monogram while the hunt says DONE ("some pictures dont show",
    #  31 Jul 2026). Failures now get a second, smaller go, and
    #  whatever still won't shrink is NAMED at the end.
    def _shrink(batch):
        items = []
        for code, p in batch:
            url = 'file:///' + os.path.abspath(p).replace('\\', '/').replace(' ', '%20')
            items.append([code, url])
        page = ('<!DOCTYPE html><html><body><pre id=out></pre><script>'
                'var IT=' + json.dumps(items) + ';var R={},left=IT.length;'
                'function fin(){document.getElementById("out").textContent='
                '"@@"+JSON.stringify(R)+"@@";}'
                'IT.forEach(function(it){var im=new Image();'
                'im.onload=function(){try{'
                'var c=document.createElement("canvas");'
                'c.width=' + str(SIZE) + ';c.height=' + str(SIZE) + ';'
                'var x=c.getContext("2d");x.fillStyle="#fff";'
                'x.fillRect(0,0,c.width,c.height);'
                'var s=Math.max(c.width/im.width,c.height/im.height);'
                'var w=im.width*s,h=im.height*s;'
                'x.drawImage(im,(c.width-w)/2,(c.height-h)/2,w,h);'
                'R[it[0]]=c.toDataURL("image/jpeg",0.82);}catch(e){R[it[0]]="";}'
                'if(--left===0)fin();};'
                'im.onerror=function(){R[it[0]]="";if(--left===0)fin();};'
                'im.src=it[1];});'
                '</script></body></html>')
        hp = os.path.join(tempfile.gettempdir(), 'coates_thumb_batch.html')
        with open(hp, 'w', encoding='utf-8') as f:
            f.write(page)
        try:
            r = subprocess.run(
                [browser, '--headless', '--disable-gpu', '--no-first-run',
                 '--user-data-dir=' + prof, '--allow-file-access-from-files',
                 '--virtual-time-budget=90000', '--dump-dom'] + extra +
                ['file:///' + hp.replace('\\', '/')],
                capture_output=True, timeout=300)
            dom = r.stdout.decode('utf-8', 'replace')
        except Exception:
            dom = ''
        m = re.search(r'@@(\{.*?\})@@', dom, re.S)
        if not m:
            return 0
        try:
            res = json.loads(m.group(1))
        except ValueError:
            return 0
        n_ok = 0
        for code, dataurl in res.items():
            if not dataurl or ',' not in dataurl:
                continue
            try:
                raw = base64.b64decode(dataurl.split(',', 1)[1])
                with open(tpath(code), 'wb') as f:
                    f.write(raw)
                n_ok += 1
            except Exception:
                continue
        return n_ok

    for k in range(0, len(pend), 25):
        made += _shrink(pend[k:k + 25])
    #  second, gentler go for anything the first pass dropped
    failed = [(c, p) for c, p in pend if not os.path.isfile(tpath(c))]
    for k in range(0, len(failed), 5):
        made += _shrink(failed[k:k + 5])
    still = sorted(c for c, p in pend if not os.path.isfile(tpath(c)))
    if still and not quiet:
        print('  Thumbnails: {} photo(s) would NOT shrink - the page shows'
              ' their two-letter tile until they do:'.format(len(still)))
        for c in still[:10]:
            print('     ' + c)
        if len(still) > 10:
            print('     ...and {} more'.format(len(still) - 10))
        print('  Open each in Paint, Save As JPEG over itself, run 04 again.')
    ready = sum(1 for c in photos if os.path.isfile(tpath(c)))
    return (len(photos), made, ready)


if __name__ == '__main__':
    n, made, ready = refresh()
    print('Photos: {} | shrunk this run: {} | thumbnails ready: {}'.format(
        n, made, ready))
    sys.exit(0)
