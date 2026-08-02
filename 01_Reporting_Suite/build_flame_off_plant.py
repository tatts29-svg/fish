#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | FLAME OFF - SITE PLANT UTILISATION - INTERNAL ONLY
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Andrew built this by hand on 2 Aug 2026. This rebuilds it from the
#  exports so it comes out with the morning pull instead of an evening.
#  His workbook is the specification - the rule, the states, the
#  columns and the day grid are all his.
#
#  HIS RULE, VERBATIM, AND IT IS THE WHOLE FILE:
#      SITE PLANT = onsite and chargeable but not allocated.
#      USED       = on hire to another employer/person.
#      NO DATA    = outside supplied transaction period.
#
#  WHAT COUNTS AS SITE PLANT (Andrew, 2 Aug 2026: "what ever was in the
#  site plant equipment"). It is an ACCOUNT HISTORY, not a storage unit
#  and not a category: any asset that has been on the Site Plant
#  Equipment account at any point in the supplied period belongs on
#  this sheet, wherever it sits now. Checked against his own list - the
#  rule finds 137 of his 138 individual assets, and the one it misses
#  is a water blaster that had moved on by the pull this was tested
#  against. It finds 321 in total, of which 176 are the bulk lines he
#  groups: 70 chutes, 50 crash barriers, 40 crowd barriers, 8 frames,
#  8 hoppers. That is his grouped total exactly.
#
#  AND THINGS GET DEPARTED (his words). Gear leaves site, and that
#  turns up at stocktake - the asset then reads Pending Baseplan, or
#  simply Available. A departed asset is NOT idle site plant and must
#  not be counted as though it were sitting there earning nothing, so
#  it gets its own state.
#
#  THE MONEY. Day rates come off the QUOTE, read by day_rates.py - drop
#  it in Data_Quote\ and the columns fill in. PER EACH (Andrew, 2 Aug
#  2026): a quoted rate is per item per day, so 50 crash barriers at
#  $2.07 is $103.50 a day, not $2.07. Every figure here is
#  rate x quantity x days.
#
#  No quote, or a line the quote does not price, and the cell reads TBC.
#  Never $0 - a nought says "this earned nothing", which is a different
#  claim and a wrong one.
#
#  Run it:  py build_flame_off_plant.py   (or 66_RUN_FLAME_OFF_PLANT)
# =====================================================================
import collections
import datetime as dt
import glob
import os
import sys

import day_rates as DR
import mygear_intel as MI

BASE = os.path.dirname(os.path.abspath(__file__))

try:
    import shutdown_day as SD
except Exception:
    SD = None

#  The bulk lines. Andrew, 2 Aug 2026: "anything in rubbish chutes and
#  barriers also part of siteplant. but it was put seperate as it was
#  bulky." So they ARE site plant - they are in the 321 and in every
#  total on this sheet - they are simply shown as one line with a
#  quantity instead of 176 rows.
#
#  That is the right call for a second reason too: they go out by the
#  pallet, not by the asset. A crash barrier is not rotated, it is
#  dropped where it is wanted and collected at the end, so a per-asset
#  utilisation figure on one would be a number with nothing behind it.
GROUPS = [
    ('Crash Barriers', 'Barrier - Crash Rated'),
    ('Crowd Barriers', 'Barrier - Crowd Control'),
    ('Rubbish Chute Frames', 'Rubbish Chute Frame'),
    ('Rubbish Chute Hoppers', 'Rubbish Chute Top Hopper'),
    ('Rubbish Chutes', 'Rubbish Chute 1M'),
]

#  Statuses that mean the asset is no longer standing on site as plant.
#  Andrew, 2 Aug 2026: departed gear shows as pending baseplan, or
#  "could show as pending branch receipt". Branch receipt is the branch
#  waiting to book it back in - it has left site, so it is departed.
#
#  AVAILABLE FOR HIRE IS NOT DEPARTED. He settled it: "available for
#  hire means its generally onsite. not being charged until such time
#  it going onhire." So it is standing on site earning nothing, which
#  is a different fact from gone, and commercially the sharper of the
#  two - see the three states below.
DEPARTED_STATUS = ('Pending Baseplan', 'Failed Baseplan',
                   'Pending Branch Receipt')

#  Statuses this sheet has a rule for. Anything else SiteIQ starts
#  sending gets NAMED rather than quietly filed under "not seen" - a
#  status we do not understand is exactly the thing that should surface,
#  not the thing that should vanish.
KNOWN_STATUS = DEPARTED_STATUS + ('Available for Hire', 'On Hire',
                                  'Awaiting Arrival')


def _norm_status(s):
    """Case and spacing must not decide whether gear counts as gone."""
    return ' '.join(str(s or '').split()).lower()


DEPARTED_NORM = {_norm_status(x) for x in DEPARTED_STATUS}
KNOWN_NORM = {_norm_status(x) for x in KNOWN_STATUS}

S_NODATA = 'NO DATA'
S_PLANT = 'SITE PLANT'
S_OFF = 'OFF / NOT SEEN'
S_GONE = 'DEPARTED'


def _newest(pattern):
    hits = [q for q in glob.glob(os.path.join(BASE, 'Data_SiteIQ', pattern))
            if not os.path.basename(q).startswith('~')]
    hits += [q for q in glob.glob(os.path.join(BASE, pattern))
             if not os.path.basename(q).startswith('~')]
    return max(hits, key=os.path.getmtime) if hits else None


def _is_plant_account(name):
    #  ONE definition, in mygear_intel, so this sheet and the
    #  intelligence page can never disagree about what the account is.
    return MI._is_holding(name)


def _group_of(desc):
    d = (desc or '').lower()
    for label, needle in GROUPS:
        if needle.lower() in d:
            return label
    return None


def collect(today=None):
    """Every asset that has been on the Site Plant account, with its
    day-by-day state across the shutdown window."""
    today = today or dt.date.today()
    rental = _newest('RENTAL_STOCK*.xlsx')
    txn = _newest('TRANSACTIONS*.xlsx')
    onhire = _newest('ON_HIRE*.xlsx')
    if not rental or not txn:
        return None

    stock = {MI._txt(r, 'ITEM_NUMBER'): r
             for r in MI._sheet(rental, 'RENTAL_STOCK')
             if MI._txt(r, 'ITEM_NUMBER')}
    oh = MI._sheet(onhire, 'ON_HIRE') if onhire else []
    tc = MI._sheet(txn, 'TRANSACTION_CHARGES')
    ec = MI._sheet(txn, 'CUSTOMER_CONTRACTOR_EQUIP')

    #  the supplied window - outside it is NO DATA, never idleness
    ds = [MI._date(r.get(k)) for r in tc + ec
          for k in ('TRAN_START_DATE', 'TRAN_END_DATE')]
    ds = [d for d in ds if d]
    if not ds:
        return None
    src_from, src_to = min(ds), max(ds)

    #  ---- who has been on the account, from every source we have ----
    on_account = set()
    for r in stock.values():
        if _is_plant_account(MI._txt(r, 'HIRER_NAME')):
            on_account.add(MI._txt(r, 'ITEM_NUMBER'))
    for r in oh:
        if _is_plant_account(MI._txt(r, 'HIRER_NAME')):
            on_account.add(MI._txt(r, 'ITEM_NUMBER'))
    for r in tc + ec:
        if _is_plant_account(MI._txt(r, 'HIRER_NAME')):
            on_account.add(MI._txt(r, 'SKU/ITEM_NUMBER'))
    on_account.discard('')

    #  names that look like the account but are not - reported, because
    #  a spelling nobody told us about is exactly what would quietly
    #  drop assets off this sheet
    near = collections.Counter()
    for r in list(stock.values()) + oh:
        nm = MI._txt(r, 'HIRER_NAME')
        if MI.is_near_holding(nm):
            near[nm] += 1
    for r in tc + ec:
        nm = MI._txt(r, 'HIRER_NAME')
        if MI.is_near_holding(nm):
            near[nm] += 1

    #  ---- every span, with who held it -------------------------------
    spans = collections.defaultdict(list)
    desc_of = {}
    for r in tc + ec:
        item = MI._txt(r, 'SKU/ITEM_NUMBER')
        if item not in on_account:
            continue
        s = MI._date(r.get('TRAN_START_DATE'))
        e = MI._date(r.get('TRAN_END_DATE')) or src_to
        if not s:
            continue
        who = MI._txt(r, 'HIRER_NAME')
        emp = MI._txt(r, 'EMPLOYER_NAME') or who
        spans[item].append((s, e, _is_plant_account(who), emp))
        desc_of.setdefault(item, MI._txt(r, 'SKU/ITEM DESCRIPTION'))
    #  ON_HIRE is the truth about right now, and carries hires that
    #  started before the transaction period ever began
    for r in oh:
        item = MI._txt(r, 'ITEM_NUMBER')
        if item not in on_account:
            continue
        s = MI._date(r.get('START_DATE'))
        if not s:
            continue
        who = MI._txt(r, 'HIRER_NAME')
        emp = MI._txt(r, 'COMPANY') or who
        spans[item].append((s, src_to, _is_plant_account(who), emp))

    #  ---- the timeline ----------------------------------------------
    first = SD.date_of(SD.FIRST_DAY) if SD else src_from
    last = SD.date_of(SD.LAST_DAY) if SD else src_to
    days = []
    d = first
    while d <= last:
        days.append(d)
        d += dt.timedelta(days=1)

    lines = []
    grouped = collections.defaultdict(
        lambda: {'qty': 0, 'items': [], 'sample': ''})
    for item in sorted(on_account):
        row = stock.get(item)
        desc = (MI._txt(row, 'ITEM_DESCRIPTION') if row
                else desc_of.get(item, ''))
        g = _group_of(desc)
        if g:
            grouped[g]['qty'] += 1
            grouped[g]['items'].append(item)
            #  the REGISTER's wording, kept for pricing. The quote says
            #  "Barrier - Crash Rated Water filled - Armorzone"; the
            #  group label says "Crash Barriers". Matching the label
            #  against the quote priced 3 lines of 150 - it has to be
            #  matched on what the gear is actually called.
            if not grouped[g]['sample']:
                grouped[g]['sample'] = desc
            continue
        lines.append((item, desc, row, [item]))

    def states_for(items):
        """One row of day states for an asset (or a group of them)."""
        out = []
        status = ''
        for it in items:
            r = stock.get(it)
            if r is not None:
                status = MI._txt(r, 'ITEM_STATUS')
                break
        departed = _norm_status(status) in DEPARTED_NORM
        for day_d in days:
            if day_d < src_from or day_d > src_to:
                out.append(S_NODATA)
                continue
            used_by, on_plant = None, False
            for it in items:
                for s, e, holding, emp in spans.get(it, ()):
                    if s <= day_d <= e:
                        if holding:
                            on_plant = True
                        else:
                            used_by = emp or 'another hirer'
                            break
                if used_by:
                    break
            if used_by:
                out.append('USED: ' + used_by)
            elif on_plant:
                out.append(S_PLANT)
            elif departed:
                #  gear that has left site. NOT idle plant - it is not
                #  here. Andrew: "things get departed too. this happens
                #  in stock take."
                out.append(S_GONE)
            else:
                out.append(S_OFF)
        return out, status

    rows = []
    for item, desc, row, items in lines:
        st, status = states_for(items)
        rows.append({'key': item, 'desc': desc, 'priceDesc': desc,
                     'qty': 1, 'type': 'INDIVIDUAL',
                     'status': status, 'states': st})
    for label, g in grouped.items():
        st, status = states_for(g['items'])
        rows.append({'key': 'GROUP - ' + label, 'desc': label,
                     'priceDesc': g['sample'],
                     'qty': g['qty'], 'type': 'GROUPED',
                     'status': status, 'states': st})

    #  ---- the measures, off the states, so they can never disagree ---
    for r in rows:
        used = [i for i, s in enumerate(r['states']) if s.startswith('USED:')]
        r['daysUsed'] = len(used)
        r['daysPlant'] = sum(1 for s in r['states'] if s == S_PLANT)
        r['daysGone'] = sum(1 for s in r['states'] if s == S_GONE)
        r['firstUsed'] = days[used[0]] if used else None
        r['lastUsed'] = days[used[-1]] if used else None
        r['employers'] = sorted({s[6:] for s in r['states']
                                 if s.startswith('USED:')})
    source_days = (src_to - src_from).days + 1
    for r in rows:
        r['util'] = (r['daysUsed'] / source_days) if source_days else 0.0

    #  ---- price it, if the quote is in --------------------------------
    rates = DR.load(BASE)
    priced = unpriced = 0
    for r in rows:
        first = (r['key'] if r['type'] == 'INDIVIDUAL' else '')
        rate = DR.rate_for(rates, item=first,
                           desc=r.get('priceDesc') or r['desc'])
        r['rate'] = rate
        if rate is None:
            unpriced += 1
        else:
            priced += 1
        r['$used'] = DR.money(rate, r['qty'], r['daysUsed'])
        r['$plant'] = DR.money(rate, r['qty'], r['daysPlant'])
        #  days on site with no charge at all - the other money story
        idle = sum(1 for x in r['states'] if x == S_OFF)
        r['idleDays'] = idle
        r['$idle'] = DR.money(rate, r['qty'], idle)

    #  any status we have no rule for, named so it cannot hide
    unknown = collections.Counter()
    for item in on_account:
        r = stock.get(item)
        if r is None:
            continue
        st = MI._txt(r, 'ITEM_STATUS')
        if st and _norm_status(st) not in KNOWN_NORM:
            unknown[st] += 1

    rows.sort(key=lambda r: (r['type'] != 'GROUPED', r['desc'] or r['key']))
    return {
        'days': days, 'rows': rows, 'srcFrom': src_from, 'srcTo': src_to,
        'sourceDays': source_days, 'today': today,
        'totalAssets': sum(r['qty'] for r in rows),
        'individual': sum(1 for r in rows if r['type'] == 'INDIVIDUAL'),
        'groups': sum(1 for r in rows if r['type'] == 'GROUPED'),
        'used': sum(1 for r in rows if r['daysUsed'] > 0),
        'neverUsed': sum(r['qty'] for r in rows if r['daysUsed'] == 0),
        'departed': sum(r['qty'] for r in rows if r['daysGone'] > 0),
        #  AMBIGUOUS ON PURPOSE. Andrew: departed gear "will show as
        #  pending baseplan. or may show as available." Pending Baseplan
        #  is unarguable and gets called DEPARTED. Available is NOT -
        #  an asset reading Available is either back on the shelf or
        #  gone off site, and nothing in the export separates the two.
        #  It is counted and named as unresolved rather than guessed
        #  into whichever column makes the sheet look tidier.
        #  ON SITE AND EARNING NOTHING. Not departed, not allocated,
        #  not even on charge. Site Plant days are revenue Coates IS
        #  taking with nobody allocated against it; these are days the
        #  gear is here and taking nothing at all. Two different
        #  arguments, and lumping them together loses both.
        'onSiteIdle': sum(r['qty'] for r in rows
                          if r['status'] == MI.READY_STATUS),
        'unknownStatus': unknown.most_common(),
        'rates': rates, 'priced': priced, 'unpriced': unpriced,
        '$used': sum(r['$used'] or 0 for r in rows),
        '$plant': sum(r['$plant'] or 0 for r in rows),
        '$idle': sum(r['$idle'] or 0 for r in rows),
        'nearNames': near.most_common(),
    }


def write_xlsx(d, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Flame Off Analysis'
    ORANGE = 'F26222'
    org = Font(bold=True, color='FFFFFF', size=13)
    hdrf = Font(bold=True, color='FFFFFF', size=9)
    fill_org = PatternFill('solid', fgColor=ORANGE)
    fill_hdr = PatternFill('solid', fgColor='1D1D1B')
    fills = {
        S_PLANT: PatternFill('solid', fgColor='FFE9DC'),
        S_NODATA: PatternFill('solid', fgColor='EFEFEF'),
        S_OFF: PatternFill('solid', fgColor='FFFFFF'),
        S_GONE: PatternFill('solid', fgColor='E4E9F0'),
    }
    used_fill = PatternFill('solid', fgColor='C9E7D2')

    ws['A1'] = 'FLAME OFF - SITE PLANT UTILISATION'
    ws['A1'].font = org
    ws['A1'].fill = fill_org
    dayno = (' | Day 0 = {}'.format(SD.FLAME_OFF.strftime('%A %d/%m/%Y'))
             if SD else '')
    ws['A2'] = ('Built {}{} | Source coverage {} to {} ({} days)'.format(
        d['today'].strftime('%d/%m/%Y'), dayno,
        d['srcFrom'].strftime('%d/%m/%Y'), d['srcTo'].strftime('%d/%m/%Y'),
        d['sourceDays']))
    ws['A3'] = ('Rule: SITE PLANT = onsite and chargeable but not allocated. '
                'USED = on hire to another employer/person. '
                'NO DATA = outside supplied transaction period. '
                'DEPARTED = off site, showing '
                + ' or '.join(DEPARTED_STATUS) + '. '
                'AVAILABLE FOR HIRE = on site and NOT being charged until '
                'it goes back on hire - here, but earning nothing.')
    ws['A5'] = ('Day rates are PER EACH - rate x quantity x days. '
                + (('Priced from ' + os.path.basename(d['rates']['path'])
                    + ': ' + str(d['priced']) + ' line(s) priced, '
                    + str(d['unpriced']) + ' still TBC.')
                   if d['rates'].get('path') else
                   ('No quote loaded yet - ' + d['rates'].get('problem', ''))))
    ws['A4'] = ('Day rates are not in any SiteIQ export - the rate column '
                'reads TBC rather than a guess. With rates, SITE PLANT days '
                'price what is being charged with nobody allocated, and '
                'ON SITE NOT CHARGING prices what is standing here earning '
                'nothing at all.')

    labels = ['TOTAL UNIQUE ASSETS', 'INDIVIDUAL LINES', 'GROUPED LINES',
              'ASSETS / GROUPS USED', 'NEVER USED IN SOURCE', 'DEPARTED',
              'ON SITE, NOT CHARGING', 'SOURCE DAYS']
    vals = [d['totalAssets'], d['individual'], d['groups'], d['used'],
            d['neverUsed'], d['departed'], d['onSiteIdle'], d['sourceDays']]
    for i, (lab, v) in enumerate(zip(labels, vals)):
        c = ws.cell(row=6, column=1 + i * 2, value=lab)
        c.font = hdrf
        c.fill = fill_hdr
        ws.cell(row=7, column=1 + i * 2, value=v).font = Font(bold=True,
                                                              size=12)

    head = ['Asset No / Group', 'Asset Description', 'Day Rate', 'Qty',
            'Line Type', 'Status', 'First Used', 'Last Used', 'Days Used',
            'Days Site Plant', 'Days Departed', 'Utilisation %',
            '$ Used', '$ Site Plant', '$ On Site Not Charging', 'Used By']
    R = 9
    for i, h in enumerate(head, start=1):
        c = ws.cell(row=R, column=i, value=h)
        c.font = hdrf
        c.fill = fill_hdr
        c.alignment = Alignment(wrap_text=True, vertical='center')
    for j, day_d in enumerate(d['days']):
        n = SD.day(day_d) if SD else j
        c = ws.cell(row=R, column=len(head) + 1 + j,
                    value='{}\n{}'.format(n, day_d.strftime('%d/%m')))
        c.font = hdrf
        c.fill = fill_hdr
        c.alignment = Alignment(wrap_text=True, horizontal='center')

    for k, r in enumerate(d['rows']):
        row = R + 1 + k
        ws.cell(row=row, column=1, value=r['key'])
        ws.cell(row=row, column=2, value=r['desc'])
        ws.cell(row=row, column=3,
                value='TBC' if r['rate'] is None else r['rate'])
        ws.cell(row=row, column=4, value=r['qty'])
        ws.cell(row=row, column=5, value=r['type'])
        ws.cell(row=row, column=6, value=r['status'])
        ws.cell(row=row, column=7,
                value=r['firstUsed'].strftime('%d/%m/%Y') if r['firstUsed']
                else '')
        ws.cell(row=row, column=8,
                value=r['lastUsed'].strftime('%d/%m/%Y') if r['lastUsed']
                else '')
        ws.cell(row=row, column=9, value=r['daysUsed'])
        ws.cell(row=row, column=10, value=r['daysPlant'])
        ws.cell(row=row, column=11, value=r['daysGone'])
        c = ws.cell(row=row, column=12, value=r['util'])
        c.number_format = '0%'
        for off, key in ((13, '$used'), (14, '$plant'), (15, '$idle')):
            v = r[key]
            c = ws.cell(row=row, column=off,
                        value='TBC' if v is None else v)
            if v is not None:
                c.number_format = '"$"#,##0.00'
        ws.cell(row=row, column=16, value=', '.join(r['employers'][:4]))
        for j, s in enumerate(r['states']):
            c = ws.cell(row=row, column=len(head) + 1 + j, value=s)
            c.fill = used_fill if s.startswith('USED:') else fills.get(
                s, fills[S_OFF])
            c.font = Font(size=8)

    widths = [22, 40, 10, 6, 12, 16, 12, 12, 10, 13, 12, 11,
              13, 14, 20, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(i)].width = w
    for j in range(len(d['days'])):
        ws.column_dimensions[openpyxl.utils.get_column_letter(
            len(head) + 1 + j)].width = 15
    ws.freeze_panes = ws.cell(row=R + 1, column=3)
    wb.save(path)


def main():
    today = dt.date.today()
    d = collect(today)
    if not d:
        print('PROBLEM: need RENTAL_STOCK and TRANSACTIONS exports in '
              'Data_SiteIQ. Pull them from SiteIQ and run again.')
        return 1
    out_dir = os.path.join(BASE, 'Reports', today.isoformat(), 'Pages')
    if not os.path.isdir(out_dir):
        os.makedirs(out_dir)
    out = os.path.join(out_dir,
                       'Flame_Off_Site_Plant_Utilisation_{}.xlsx'.format(
                           today.strftime('%d%b%Y')))
    write_xlsx(d, out)

    print('=' * 62)
    print(' COATES | FLAME OFF - SITE PLANT UTILISATION')
    print('=' * 62)
    print('')
    print(' Scope        : whatever has been on the Site Plant Equipment')
    print('                account in the supplied period, wherever it')
    print('                sits now.')
    print(' Source       : {} to {} ({} days)'.format(
        d['srcFrom'].strftime('%d/%m/%Y'), d['srcTo'].strftime('%d/%m/%Y'),
        d['sourceDays']))
    print('')
    print(' Total assets : {:,}   ({} individual lines, {} grouped)'.format(
        d['totalAssets'], d['individual'], d['groups']))
    print(' Used         : {} line(s) went out to a named employer'.format(
        d['used']))
    print(' Never used   : {:,} asset(s) never left the account'.format(
        d['neverUsed']))
    print(' Departed     : {:,} asset(s) have left site ({})'.format(
        d['departed'], ' / '.join(DEPARTED_STATUS)))
    if d['nearNames']:
        print(' ' + '!' * 58)
        print(' Name(s) that look like the Site Plant Equipment account')
        print(' but do not match it exactly. Nothing of theirs is on this')
        print(' sheet - say if it should be:')
        for nm, n in d['nearNames'][:6]:
            print('     {:<38} {} row(s)'.format(nm[:38], n))
        print(' ' + '!' * 58)
    if d['unknownStatus']:
        print(' ' + '!' * 58)
        print(' SiteIQ sent status values this sheet has no rule for.')
        print(' They are being treated as "not seen", which may be wrong:')
        for st, n in d['unknownStatus']:
            print('     {:<34} {} asset(s)'.format(st[:34], n))
        print(' Add them to DEPARTED_STATUS or KNOWN_STATUS in')
        print(' build_flame_off_plant.py once you know which they are.')
        print(' ' + '!' * 58)
    if d['onSiteIdle']:
        print(' On site,     : {:,} asset(s) read Available for Hire - here,'
              .format(d['onSiteIdle']))
        print(' not charging   on the shelf, and not being charged at all')
        print('                until they go back on hire. Site Plant days')
        print('                earn with nobody allocated; these earn')
        print('                nothing. Different arguments, both worth')
        print('                having.')
    print('')
    r = d['rates']
    if r.get('path'):
        print(' Day rates    : {} - {} line(s) priced, {} still TBC'.format(
            os.path.basename(r['path']), d['priced'], d['unpriced']))
        print(' Per each     : rate x quantity x days.')
        print('   Used          ${:>12,.2f}'.format(d['$used']))
        print('   Site plant    ${:>12,.2f}   charged, nobody allocated'
              .format(d['$plant']))
        print('   Not charging  ${:>12,.2f}   on site, earning nothing'
              .format(d['$idle']))
        if r.get('skipped'):
            print(' {} quote line(s) carried no rate: {}'.format(
                len(r['skipped']),
                ', '.join(str(x)[:22] for x in r['skipped'][:4])))
    else:
        #  Make the folder and say what goes in it. The update zip is
        #  flat, so it cannot ship an empty directory - if this did not
        #  create it, the instruction would point at somewhere that does
        #  not exist.
        qd = os.path.join(BASE, DR.QUOTE_DIR)
        try:
            if not os.path.isdir(qd):
                os.makedirs(qd)
            rd = os.path.join(qd, 'READ_ME.txt')
            if not os.path.isfile(rd):
                with open(rd, 'w') as fh:
                    fh.write(
                        'DAY RATES - drop the job quote in this folder.\n'
                        '=============================================\n\n'
                        'Any .xlsx in here is read as the quote. Newest\n'
                        'wins. It needs two things:\n\n'
                        '  * an item number OR a description column\n'
                        '  * a day rate column\n\n'
                        'Rates are PER EACH - per item, per day. 50 crash\n'
                        'barriers at $2.07 is $103.50 a day.\n\n'
                        'A line with no rate is skipped and named, never\n'
                        'read as free. An item the quote does not price\n'
                        'shows TBC, never $0.\n\n'
                        'This is the SiteIQ stream only. Baseplan bills\n'
                        'its own 16 lines and carries its own rates - the\n'
                        'two never mix.\n')
        except Exception:
            pass
        print(' Day rates    : {}'.format(r.get('problem', 'not loaded')))
        print('                Every money column reads TBC until it is in.')
        print('                Folder ready: {}\\'.format(DR.QUOTE_DIR))
    print('')
    print(' Workbook     : ' + out)
    print(' COATES INTERNAL - do not send this to the client.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
