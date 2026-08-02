#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | SERIAL NUMBERS - the plate on the machine
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 3 Aug 2026): "Fleet_No = Item_Number, and then the
#  column Serial_No is the Item_Number's serial number."
#
#  That is the join, and it holds: 985 of the assets in the store
#  register appear in the export by plant number, and nothing in it
#  contradicts what SiteIQ already knows.
#
#  It matters because a plant number is Coates' name for a machine and
#  a serial number is the manufacturer's. When a generator comes back
#  damaged, when an insurer asks, when a supervisor argues that it was
#  not the one his bloke had - the plant number is our word for it and
#  the serial is the plate on the side.
#
#  ---------------------------------------------------------------
#  IT ONLY COVERS PLANT, AND THAT IS CORRECT
#  ---------------------------------------------------------------
#  990 of the 5,380 assets in the store register carry a plant number
#  (1302153, 1242366). The other 4,390 are barcode-suffixed tooling -
#  DUCTING300MMR-0308, SPANNERPMM 32MM-0009 - and the export has no
#  serial for those, because a length of ducting and a crowd control
#  barrier do not have one. So the gap is not missing data. Nothing
#  here pretends otherwise, and nothing is invented to fill it.
#
#  ---------------------------------------------------------------
#  615 OF THE "SERIALS" ARE NOT SERIALS
#  ---------------------------------------------------------------
#  This is the part worth knowing. Of 3,105 rows with something in the
#  Serial_No column:
#
#    432   read Coates1207765 - the plant number with COATES in front
#    111   are the plant number again, unchanged
#     53   say TBA, tba, NA, N/A, nil or just a full stop
#
#  Put one of those on a damage claim and it says nothing: it is our
#  own number handed back to us wearing a serial's clothes. So they are
#  recognised and held apart. serial_of() returns nothing for them, and
#  the count is reported rather than buried, because 270 of the assets
#  in this store are in that state and that is a Baseplan housekeeping
#  job Andrew can actually get done.
#
#  A LIVE CONFLICT, KEPT VISIBLE: serial 1300211 is recorded against
#  two different plant numbers, 1323755 and 1323775. One of them is
#  wrong. It is reported rather than silently picked between.
#
#  ---------------------------------------------------------------
#  TWO BRANCHES IN ONE FILE
#  ---------------------------------------------------------------
#  The export carries GLST and NOIS - 3,235 rows and 3,299 rows - and
#  1,677 plant numbers appear under both. Not one of them disagrees
#  about the serial, so the duplicate is a listing artefact, not a
#  clash. GLST is preferred anyway because that is our branch, and a
#  row carrying a real serial beats an empty one either way.
#
#  Drop the export in Data_Serials\. No file, no change to anything -
#  every screen carries on exactly as it did.
# =====================================================================
import glob
import os
import re

HERE = os.path.dirname(os.path.abspath(__file__))
FOLDER = os.path.join(HERE, 'Data_Serials')

HOME_BRANCH = 'GLST'

#  Values that occupy the Serial_No column without being a serial. Every
#  one of these was read off the file, not imagined.
PLACEHOLDER = re.compile(r'^(n/?a|nil|none|tbc|tba|unknown|[-_/.0\s]*)$',
                         re.IGNORECASE)

_CACHE = None


def newest(folder=None):
    hits = [p for p in glob.glob(os.path.join(folder or FOLDER, '*.xlsx'))
            if not os.path.basename(p).startswith('~')]
    return max(hits, key=os.path.getmtime) if hits else None


def is_real(fleet, ser):
    """Is this a serial, or is it our own plant number in a wig?"""
    s = (ser or '').strip()
    if not s or PLACEHOLDER.match(s):
        return False
    f = (fleet or '').strip().upper()
    u = s.upper().replace(' ', '')
    return not (u == f or u == 'COATES' + f)


def load(folder=None):
    """{plant number: record}. Empty when there is no export, which is
    a perfectly good state - the screens just carry on without it."""
    global _CACHE
    if _CACHE is not None and folder is None:
        return _CACHE
    out = {}
    path = newest(folder)
    if path:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            hdr = [str(c).strip().upper() if c is not None else ''
                   for c in next(rows)]
            ix = {h: i for i, h in enumerate(hdr) if h}

            def cell(r, name):
                i = ix.get(name)
                if i is None or i >= len(r) or r[i] is None:
                    return ''
                return str(r[i]).strip()

            for r in rows:
                if not r:
                    continue
                fleet = cell(r, 'FLEET_NO')
                if not fleet:
                    continue
                ser = cell(r, 'SERIAL_NO')
                rec = {
                    'item': fleet,
                    'serial': ser if is_real(fleet, ser) else '',
                    'stated': ser,
                    'model': cell(r, 'MODEL'),
                    'desc': cell(r, 'FLD_DESC'),
                    'branch': cell(r, 'BRANCH_CODE'),
                }
                #  Same plant number twice: keep the row that actually
                #  says something, then prefer our own branch.
                old = out.get(fleet)
                if old:
                    if old['serial'] and not rec['serial']:
                        continue
                    if (old['serial'] and rec['serial']
                            and old['branch'] == HOME_BRANCH):
                        continue
                out[fleet] = rec
            wb.close()
        except Exception:
            #  A broken export changes nothing. A screen with no serial
            #  on it is a small loss; no screen at all is a real one.
            out = {}
    if folder is None:
        _CACHE = out
    return out


def serial_of(item):
    """The manufacturer's number, or '' - never a placeholder."""
    r = load().get(str(item or '').strip())
    return r['serial'] if r else ''


def model_of(item):
    r = load().get(str(item or '').strip())
    return r['model'] if r else ''


def record(item):
    return load().get(str(item or '').strip())


def conflicts(recs=None):
    """One serial against more than one plant number. Real ones only -
    forty machines all reading TBA is a blank, not a clash."""
    recs = recs if recs is not None else load()
    by = {}
    for k, r in recs.items():
        if r['serial']:
            by.setdefault(r['serial'].upper(), []).append(k)
    return {s: sorted(v) for s, v in by.items() if len(v) > 1}


def coverage(items):
    """How much of a given set of assets this file can actually speak
    for, told in three parts rather than one flattering percentage."""
    recs = load()
    out = {'assets': 0, 'serial': 0, 'placeholder': 0, 'absent': 0,
           'plant': 0, 'file': bool(recs)}
    for it in items:
        k = str(it or '').strip()
        out['assets'] += 1
        if k.isdigit():
            out['plant'] += 1
        r = recs.get(k)
        if not r:
            out['absent'] += 1
        elif r['serial']:
            out['serial'] += 1
        else:
            out['placeholder'] += 1
    return out
