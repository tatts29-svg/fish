#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | CONSUMABLES - what is on the shelf, what went out,
#                          was the count right, and do we need to order
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 29 Jul 2026): "consumables stock. how many available.
#  how many used. have we stock checked was it right was it ok.
#  percentage of stock take. do we need to order. looking at the date
#  range ending around 11th august."
#
#  ---------------------------------------------------------------
#  THE ONE THING THAT WILL BITE ANYONE WHO EDITS THIS FILE
#  ---------------------------------------------------------------
#  SALES_STOCK.xlsx is TWO different kinds of row in one sheet, and
#  nothing in the file says so:
#
#    * STOCK LINES  - no SALES_DATE. These carry AVAILABLE_QUANTITY,
#                     MINIMUM_QUANTITY, REORDER_QUANTITY and the
#                     storage unit. This is the shelf. 72 rows today.
#    * SALE ROWS    - a SALES_DATE, a SALES_QUANTITY, a company and a
#                     hirer. AVAILABLE_QUANTITY on these rows is 0,
#                     not "none left". 197 rows today.
#
#  Sum AVAILABLE_QUANTITY over every row and today you get the right
#  answer BY ACCIDENT, because the sale rows happen to be zero. Sum
#  SALES_QUANTITY over every row and you also get away with it. Group
#  by SKU without splitting them and you will not: 39 SKUs have both
#  kinds of row, and the shelf figure quietly becomes whichever row
#  sorted last. Split the two first. Always.
#
#  Every quantity in this export is TEXT, not a number - "10", "0.00".
#  Dates are TEXT too, d/m/Y. Coerce, never assume.
#
#  ---------------------------------------------------------------
#  WHAT "DO WE NEED TO ORDER" CAN AND CANNOT MEAN HERE
#  ---------------------------------------------------------------
#  SiteIQ has MINIMUM_QUANTITY and REORDER_QUANTITY fields, which is
#  where a reorder trigger belongs. On this job every one of them is
#  zero - nobody has set them. So a reorder point cannot be read out
#  of SiteIQ, and this file does NOT pretend otherwise. It works the
#  answer out from what actually went out the window: burn rate since
#  the SKU first moved, against the days left to the finish date. The
#  page says which of the two it used, every time.
#
#  SiteIQ does carry its own "Stock Low" sighting status, and that IS
#  read. But it is reported next to the count, not instead of it -
#  see below, because today all ten of those flags are false alarms.
# =====================================================================
import collections
import datetime as dt
import io
import os

#  The finish date the ordering question is asked against. Andrew:
#  "looking at the date range ending around 11th august". Kept in a
#  file so it can move without a code change - shutdowns always move.
END_FILE = 'shutdown_end.txt'
DEFAULT_END = '2026-08-11'


def finish_date(base):
    """The date we have to keep the shelves stocked to."""
    p = os.path.join(base, END_FILE)
    raw = DEFAULT_END
    if os.path.isfile(p):
        try:
            with io.open(p, encoding='utf-8') as fh:
                raw = (fh.read().strip() or DEFAULT_END)
        except OSError:
            pass
    else:
        try:
            with io.open(p, 'w', encoding='utf-8') as fh:
                fh.write(DEFAULT_END + '\n')
        except OSError:
            pass
    for f in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return dt.datetime.strptime(raw.strip(), f).date()
        except ValueError:
            continue
    return dt.datetime.strptime(DEFAULT_END, '%Y-%m-%d').date()


def _num(v):
    """SiteIQ ships quantities as text. '10', '0.00', '' and None all
    turn up in the same column."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return 0.0


def _date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = ('' if v is None else str(v)).strip()
    if not s:
        return None
    for f in ('%d/%m/%Y %I:%M %p', '%d/%m/%Y %H:%M', '%d/%m/%Y', '%Y-%m-%d'):
        try:
            return dt.datetime.strptime(s, f).date()
        except ValueError:
            continue
    return None


def _sheet(path, name):
    """Rows of a sheet as dicts, headers stripped.

    REORDER_QUANTITY ships with a trailing space in its header. Strip
    every header or the lookup fails on a column that is plainly there.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[name] if name in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []
    hdr = [(str(h).strip() if h is not None else '') for h in rows[0]]
    out = []
    for r in rows[1:]:
        d = {}
        for i, h in enumerate(hdr):
            if h:
                d[h] = r[i] if i < len(r) else None
        out.append(d)
    return out


def _txt(d, k):
    v = d.get(k)
    return '' if v is None else str(v).strip()


def read(sales_path, stocktake_path, base, today=None):
    """The whole consumables picture, or None if there is no export.

    Returns a dict the page layers can render without doing any more
    thinking: totals, the count, the order list, and the flags.
    """
    if not sales_path or not os.path.isfile(sales_path):
        return None
    today = today or dt.date.today()
    end = finish_date(base)
    days_left = max(0, (end - today).days)

    rows = _sheet(sales_path, 'SALES_STOCK')

    #  --- split the sheet in two. See the header comment. ---
    shelf = {}
    sales = collections.defaultdict(list)
    for r in rows:
        sku = _txt(r, 'SKU_NUMBER')
        if not sku:
            continue
        if _txt(r, 'SALES_DATE'):
            sales[sku].append({
                'when': _date(r.get('SALES_DATE')),
                'qty': _num(r.get('SALES_QUANTITY')),
                'ret': _num(r.get('RETURN_QUANTITY')),
                'who': _txt(r, 'HIRER'),
                'co': _txt(r, 'COMPANY_NAME'),
            })
        else:
            s = shelf.setdefault(sku, {
                'sku': sku,
                'desc': _txt(r, 'SKU_DESCRIPTION') or sku,
                'barcode': _txt(r, 'SKU_BARCODE'),
                'unit': _txt(r, 'STORAGE_UNIT'),
                'avail': 0.0, 'min': 0.0, 'reorder': 0.0,
            })
            #  one SKU can hold a line in two storage units - add, do
            #  not overwrite, or the shelf silently loses a location
            s['avail'] += _num(r.get('AVAILABLE_QUANTITY'))
            s['min'] += _num(r.get('MINIMUM_QUANTITY'))
            s['reorder'] += _num(r.get('REORDER_QUANTITY'))

    if not shelf:
        return None

    #  --- the count ---
    counted = {}
    if stocktake_path and os.path.isfile(stocktake_path):
        for r in _sheet(stocktake_path, 'STOCKTAKE'):
            #  PRODUCT_CATEGORY is the flag. ITEM_OR_CONSUMABLE reads
            #  like it would be, but it holds an item number.
            if _txt(r, 'PRODUCT_CATEGORY') != 'Consumable':
                continue
            sku = _txt(r, 'SKU_NUMBER')
            if not sku:
                continue
            c = counted.setdefault(sku, {
                'qty': 0.0, 'when': _date(r.get('LAST_SIGHTED_DATE_TIME')),
                'by': _txt(r, 'LAST_SIGHTED_BY'),
                'status': _txt(r, 'SIGHTED_STATUS'),
                'unit': _txt(r, 'STORAGE_UNIT'),
            })
            c['qty'] += _num(r.get('SIGHTED_QUANTITY'))

    #  --- twins: the same item carrying two SKU records ---
    #  Six items on this job have two SKUs each: a live one holding the
    #  stock and a duplicate sitting on zero. The duplicate is what
    #  trips SiteIQ's "Stock Low". SiteIQ will not merge them (Andrew,
    #  29 Jul 2026: "the twins dont merge can we hide"), so this engine
    #  FOLDS them instead: the duplicate's shelf figures, counts and
    #  sales are absorbed into the live record and the duplicate line
    #  disappears from every page. One item, one line, no false alarm -
    #  and if SiteIQ ever grows a third record for the same item, it
    #  folds in with no code change.
    by_desc = collections.defaultdict(list)
    for sku, s in shelf.items():
        by_desc[s['desc'].strip().upper()].append(sku)
    folded = 0
    folded_skus = []
    for _d, skus in by_desc.items():
        if len(skus) < 2:
            continue
        #  the record with the most stock is the live one; ties break
        #  by SKU so two zero-stock twins still fold deterministically
        skus.sort(key=lambda k: (-shelf[k]['avail'], k))
        prime = skus[0]
        for dup in skus[1:]:
            shelf[prime]['avail'] += shelf[dup]['avail']
            shelf[prime]['min'] += shelf[dup]['min']
            shelf[prime]['reorder'] += shelf[dup]['reorder']
            sales[prime].extend(sales.pop(dup, []))
            cdup = counted.pop(dup, None)
            if cdup:
                cp = counted.setdefault(prime, {
                    'qty': 0.0, 'when': cdup['when'], 'by': cdup['by'],
                    'status': cdup['status'], 'unit': cdup['unit']})
                cp['qty'] += cdup['qty']
                #  the freshest count wins the date and the status -
                #  except "Stock Low", which never survives a fold that
                #  just proved the stock exists on the other record
                if cdup['when'] and (not cp['when']
                                     or cdup['when'] > cp['when']):
                    cp['when'] = cdup['when']
                    cp['by'] = cdup['by']
                if cp['status'] == 'Stock Low' and shelf[prime]['avail'] > 0:
                    cp['status'] = 'In Stock'
            del shelf[dup]
            folded += 1
            folded_skus.append((dup, prime))

    items = []
    for sku, s in sorted(shelf.items(), key=lambda kv: kv[1]['desc']):
        moves = sales.get(sku, [])
        used = sum(m['qty'] for m in moves)
        returned = sum(m['ret'] for m in moves)
        firsts = [m['when'] for m in moves if m['when']]
        first = min(firsts) if firsts else None
        #  Burn is measured from the day the item FIRST moved, not from
        #  the start of the shut. Something that only started going out
        #  on Monday burns at its Monday rate, not a rate watered down
        #  by the fortnight it sat there.
        span = max(1, (today - first).days + 1) if first else 0
        burn = (used / span) if (used and span) else 0.0
        cover = (s['avail'] / burn) if burn > 0 else None

        c = counted.get(sku)
        var = None
        var_adj = None
        sold_since = 0.0
        if c and c['when']:
            #  A count taken on the 28th SHOULD differ from today's
            #  stock if gear went out on the 29th. Take that off before
            #  calling anything a discrepancy.
            sold_since = sum(m['qty'] for m in moves
                             if m['when'] and m['when'] > c['when'])
            var = c['qty'] - s['avail']
            var_adj = (c['qty'] - sold_since) - s['avail']

        low = bool(c and c['status'] == 'Stock Low')

        items.append({
            'sku': sku, 'desc': s['desc'], 'unit': s['unit'],
            'barcode': s['barcode'],
            'avail': s['avail'], 'min': s['min'], 'reorder': s['reorder'],
            'used': used, 'returned': returned, 'moves': len(moves),
            'first': first.isoformat() if first else '',
            'span': span, 'burn': burn, 'cover': cover,
            'counted': (c['qty'] if c else None),
            'countedOn': (c['when'].isoformat() if (c and c['when']) else ''),
            'countedBy': (c['by'] if c else ''),
            'status': (c['status'] if c else ''),
            'low': low, 'var': var, 'varAdj': var_adj,
            'soldSince': sold_since,
            #  twins are folded before this loop ever runs; the fields
            #  stay so the pages need no schema change
            'twins': [], 'twinHolds': 0.0,
        })

    #  --- classify. Order of these tests matters: a false alarm must
    #  be caught BEFORE it lands on the order list. ---
    order, watch, dead, empty, records = [], [], [], [], []
    for it in items:
        #  1. flagged low but a count found stock the system has at
        #     zero - a records problem, not a shortage. (Twin-SKU false
        #     alarms no longer reach here; they are folded away above.)
        if it['low'] and (it['counted'] or 0) > 0:
            it['why'] = 'uncounted'
            records.append(it)
            continue
        #  2. will run out before the finish date at the current rate
        if it['cover'] is not None and it['cover'] < days_left:
            order.append(it)
            continue
        #  3. empty shelf and nothing moving - nothing to burn, but
        #     nothing there either
        if it['avail'] <= 0 and it['burn'] <= 0:
            empty.append(it)
            continue
        #  4. tight but not short
        if it['cover'] is not None and it['cover'] < days_left * 1.5:
            watch.append(it)
            continue
        #  5. sat on the shelf the whole job and never moved
        if it['burn'] <= 0 and it['avail'] > 0:
            dead.append(it)

    order.sort(key=lambda x: (x['cover'] if x['cover'] is not None else 9e9))
    watch.sort(key=lambda x: (x['cover'] if x['cover'] is not None else 9e9))
    dead.sort(key=lambda x: -x['avail'])
    records.sort(key=lambda x: -(x['counted'] or 0))

    #  --- the count, scored ---
    checked = [it for it in items if it['counted'] is not None]
    matched = [it for it in checked if abs(it['varAdj'] or 0) < 0.001]
    off = [it for it in checked if abs(it['varAdj'] or 0) >= 0.001]
    off.sort(key=lambda x: -abs(x['varAdj'] or 0))
    explained = [it for it in checked
                 if abs(it['var'] or 0) >= 0.001
                 and abs(it['varAdj'] or 0) < 0.001]
    count_days = sorted({it['countedOn'] for it in checked if it['countedOn']})

    total_avail = sum(it['avail'] for it in items)
    total_used = sum(it['used'] for it in items)

    return {
        'asof': today.isoformat(),
        'end': end.isoformat(),
        'daysLeft': days_left,
        'items': items,
        'skus': len(items),
        'avail': total_avail,
        'used': total_used,
        'returned': sum(it['returned'] for it in items),
        'moves': sum(it['moves'] for it in items),
        'checked': len(checked),
        'checkedPct': (100.0 * len(checked) / len(items)) if items else 0.0,
        'matched': len(matched),
        'matchPct': (100.0 * len(matched) / len(checked)) if checked else 0.0,
        'off': off,
        'explained': len(explained),
        'countDays': count_days,
        'order': order,
        'watch': watch,
        'dead': dead,
        'empty': empty,
        'records': records,
        'folded': folded,
        'foldedSkus': folded_skus,
        'lowFlags': sum(1 for it in items if it['low']),
        #  every reorder point in the export is zero - say so rather
        #  than let a reader assume the trigger came from SiteIQ
        'reorderSet': sum(1 for it in items
                          if it['min'] > 0 or it['reorder'] > 0),
    }
