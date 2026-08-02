# ==========================================================================
#  COATES | BUILD MY GEAR - "My Shutdown Story" scorecard builder
#  Cement Australia K2 Shutdown 2026 - Gladstone
#
#  Builds Gear_Lookup\index.html - the QR-at-the-window page where crew
#  enter their ID and see their complete shutdown scorecard:
#    - gear on hire now (with age + category colours)
#    - Returns Score (driven by ACTUAL same-day returns, transactions truth)
#    - store scorecard: visits, transactions, consumables, radios,
#      gas monitors, client gear, damage charges
#  Every number comes from the SiteIQ exports sitting next to this script:
#    ON_HIRE*.xlsx       (current on-hire; carries EXTERNAL_ID = card ID)
#    TRANSACTIONS*.xlsx  (transaction history; returns, damages, radios)
#  Nothing is invented: if an export has no damage rows, cards say so.
#
#  Run: 04_RUN_MY_GEAR.bat  (or: python BUILD_MY_GEAR.py)
#  Then serve with 05_START_GEAR_LOOKUP.bat as usual.
#
#  25 Jul 2026 - COMPLIANCE BADGES ON EVERY LINE (A. Fisher)
#  Each gear line now carries its own compliance requirement under the
#  description - tag colour, pre-start and logbook, back-daily - straight
#  off K2_MASTER_EQUIPMENT_PRICING.xlsx via equipment_compliance.py. The
#  reason is simple: the bloke reading his card at the window should not
#  have to guess whether the thing in his hand needs a current tag or a
#  logbook entry. It kills the "nobody told me" excuse and it lets the
#  Coates team check gear over before it goes back out. Nothing is
#  guessed here - no flag in the master means nothing shows on the card.
#  A short summary line sits under the person's gear list as well, so the
#  obligation reads as one job, not a scatter of little chips.
#
#  Author: Andrew Fisher | POWERED BY SITEIQ
# ==========================================================================
import base64, glob, html, io, json, os, re, sys
import datetime as dt
import master_equipment
# The site guides (contact board, radio, gas monitor) and the
# phone-first bits: save-to-phone, scan-your-card, ?id= links.
import mygear_font
import mygear_guides
import mygear_ui
import equipment_compliance as EC
# K2_MASTER_EQUIPMENT_PRICING.xlsx - item-number-keyed display names
# (renames) applied wherever gear is shown; empty-safe without the file.
#  Replacement costs on the printed A4 (Andrew, 26 Jul 2026: the elite
#  per-person document carries every bit of information we have) - same
#  matching engine as every company report, so the numbers can never
#  disagree between a person's page and their company's pack.
import build_company_onhire_report as _BC
MASTER = master_equipment.load(os.path.dirname(os.path.abspath(__file__)),
                               quiet=True)
# Point the compliance engine at the same master - one file drives the
# names, the costs and the compliance flags, so a fix in the spreadsheet
# fixes every report on the next run. Safe with no master: no badges.
EC.bind(MASTER)
try:
    _REPL, _repl_path = _BC.load_replacement()
except Exception:
    _REPL = {}


def _repl_cost(asset, desc):
    c = MASTER.price(asset)
    if c is None and _REPL:
        c = _REPL.get(str(desc or "").upper())
    return c

BASE = os.path.dirname(os.path.abspath(__file__))

# --------------------------- crypto (matches the page JS exactly) ---------
M32 = 0xFFFFFFFF
def _imul(a, b): return (a * b) & M32
def _xmur3(s):
    h = (1779033703 ^ len(s)) & M32
    for ch in s:
        h = _imul(h ^ ord(ch), 3432918353)
        h = ((h << 13) | (h >> 19)) & M32
    def call():
        nonlocal h
        h = _imul(h ^ (h >> 16), 2246822507)
        h = _imul(h ^ (h >> 13), 3266489909)
        h ^= h >> 16
        return h & M32
    return call
def _mulberry32(a):
    a &= M32
    def call():
        nonlocal a
        a = (a + 0x6D2B79F5) & M32
        t = _imul(a ^ (a >> 15), (a | 1) & M32)
        t = ((t + _imul(t ^ (t >> 7), (t | 61) & M32)) & M32) ^ t
        t &= M32
        return (t ^ (t >> 14)) & M32
    return call
def lk_tag(idno): return format(_xmur3(idno + '|CoatesK2tag2026')(), 'x')
def lk_enc(idno, s):
    r = _mulberry32(_xmur3(idno + '|CoatesK2gear2026')())
    return base64.b64encode(bytes((ord(c) ^ (r() >> 24)) & 0xFF for c in s)).decode()

# --------------------------- helpers --------------------------------------
def die(msg):
    print('PROBLEM: ' + msg)
    sys.exit(1)

def find_export(pattern, what, required=True):
    """Newest match wins - handles both the short kit names (ON_HIRE.xlsx)
    and the raw SiteIQ names (ON_HIRE_23_07_2026 08_50 AM 1.xlsx)."""
    _cand = (glob.glob(os.path.join(BASE, 'Data_SiteIQ', pattern))
             + glob.glob(os.path.join(BASE, pattern)))
    hits = sorted(_cand,
                  key=os.path.getmtime, reverse=True)
    if not hits:
        if required:
            die('No ' + what + ' export found. Pull it from SiteIQ, drop it in '
                'this folder (' + pattern + ') and run again.')
        print('  NOTE: no ' + what + ' export found (' + pattern + ') - '
              'building without it; descriptions fall back to ON_HIRE\'s.')
        return None
    return hits[0]

def parse_date(v):
    """Australian DD/MM/YYYY only - US parsing makes plausible wrong reports."""
    if v in (None, ''): return None
    if isinstance(v, dt.datetime): return v.date()
    if isinstance(v, dt.date): return v
    s = str(v).strip().split(' ')[0]
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m: return dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    return None


def _newest_file(pattern):
    """Newest match, Data_SiteIQ first then the suite folder."""
    import glob
    hits = [q for q in glob.glob(os.path.join(BASE, 'Data_SiteIQ', pattern))
            if not os.path.basename(q).startswith('~')]
    hits += [q for q in glob.glob(os.path.join(BASE, pattern))
             if not os.path.basename(q).startswith('~')]
    return max(hits, key=os.path.getmtime) if hits else None

def norm_name(s):
    # ON_HIRE uses "First - Last", transactions use "First Last"
    return re.sub(r'\s*-\s*', ' ', str(s or '').strip())

def fmt_d(d): return d.strftime('%d %b %Y').lstrip('0') if d else '-'

# --------------------------- categories & colours -------------------------
CAT_COLOURS = {'Electrical': '#FFA24D', 'Rigging': '#F26222',
               'Plant': '#C44C28', 'Tooling': '#8A97A8', 'Radios': '#FFD27A'}
def category(desc):
    d = str(desc or '').lower()
    if re.search(r'radio|battery|batte|antenna', d): return 'Radios'
    if re.search(r'extension lead|distribution board|lighting tower|generator'
                 r'|240v|415v|power lead|rcd|transformer', d): return 'Electrical'
    if re.search(r'chain block|shackle|sling|lever hoist|hoist|rigging', d): return 'Rigging'
    if re.search(r'forklift|skid steer|loader|welder|welding|excavat|telehandler'
                 r'|compressor|boom|scissor|dozer|roller|dumper', d): return 'Plant'
    return 'Tooling'
GAS_RE = re.compile(r'gas monitor|gas detector|gasalert|microclip|altair'
                    r'|quattro|ventis|draeger|drager|x-am|odalog', re.I)
RADIO_RE = re.compile(r'radio|battery|batte|antenna', re.I)

# --------------------------- read ON_HIRE ---------------------------------
def read_onhire(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    asof = ''
    if 'REFERENCE_INFO' in wb.sheetnames:
        rows = list(wb['REFERENCE_INFO'].iter_rows(values_only=True))
        if len(rows) > 1:
            hdr = [str(c) for c in rows[0]]
            for k, v in zip(hdr, rows[1]):
                if 'REQUESTED_DATE' in k and v:
                    d = parse_date(v)
                    tm = re.search(r'(\d{1,2}:\d{2})\s*([AP]M)?', str(v))
                    asof = (fmt_d(d) + (' ' + tm.group(0) if tm else '')) if d else str(v)
    ws = wb['ON_HIRE'] if 'ON_HIRE' in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c) for c in rows[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    #  HIRER_ID rides along so the card can show BOTH scannable numbers -
    #  the card number the crews quote, and SiteIQ's own hirer ID the
    #  store counter uses. (Andrew, 27 Jul 2026.)
    for need in ('EXTERNAL_ID', 'HIRER_NAME', 'COMPANY', 'ITEM_NUMBER',
                 'ITEM_DESCRIPTION', 'START_DATE'):
        if need not in ix:
            die('ON_HIRE export is missing the ' + need + ' column - the page '
                'cannot be built from it. Re-pull the standard ON_HIRE report.')
    people = {}
    for r in rows[1:]:
        if not r or r[ix['EXTERNAL_ID']] in (None, ''): continue
        # the site idle pool ("Site Plant Equipment") is not a person -
        # its barriers/chutes would otherwise become a 300-item "card"
        if norm_name(r[ix['HIRER_NAME']]).lower().startswith('site plant equipment'):
            continue
        idno = str(r[ix['EXTERNAL_ID']]).strip()
        p = people.setdefault(idno, {
            'name': norm_name(r[ix['HIRER_NAME']]),
            'company': str(r[ix['COMPANY']] or '').strip(),
            'hid': (str(r[ix['HIRER_ID']]).strip()
                    if 'HIRER_ID' in ix and r[ix['HIRER_ID']] not in (None, '')
                    else ''),
            'items': []})
        _itm = str(r[ix['ITEM_NUMBER']] or '').strip()
        p['items'].append({
            'item': _itm,
            'barcode': str(r[ix['ITEM_BARCODE']] or '').strip().upper()
                       if 'ITEM_BARCODE' in ix else '',
            'desc': MASTER.disp(_itm, str(r[ix['ITEM_DESCRIPTION']] or '').strip()),
            'start': parse_date(r[ix['START_DATE']])})
    return people, asof

# ------------------- read RENTAL_STOCK / SALES_STOCK ----------------------
# RENTAL_STOCK is the fleet register - the authoritative asset number and
# CURRENT description for every barcode (descriptions get corrected there,
# and everything joins on to replacement costs the same way).
#  item number -> PRODUCT_VARIANT code: the photo key, so a worker's own
#  gear list can show what each item LOOKS like (Andrew, 31 Jul 2026:
#  "next level for people ... to see what it looks like what they have
#  in their name"). Filled by read_rental as it walks the register.
VAR_OF_ITEM = {}
ALT_OF_ITEM = {}


def read_rental(path):
    import openpyxl
    by_bc = {}
    if not path: return by_bc
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['RENTAL_STOCK'] if 'RENTAL_STOCK' in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c) for c in next(rows)]
    ix = {h.strip(): i for i, h in enumerate(hdr)}
    if not {'ITEM_BARCODE', 'ITEM_NUMBER', 'ITEM_DESCRIPTION'} <= set(ix):
        print('  NOTE: RENTAL_STOCK export missing expected columns - '
              'descriptions fall back to ON_HIRE\'s.')
        return by_bc
    for r in rows:
        if not r: continue
        bc = str(r[ix['ITEM_BARCODE']] or '').strip().upper()
        _itm2 = str(r[ix['ITEM_NUMBER']] or '').strip()
        if _itm2 and 'PRODUCT_VARIANT' in ix and _itm2 not in VAR_OF_ITEM:
            _v = str(r[ix['PRODUCT_VARIANT']] or '').strip().upper()
            if _v:
                VAR_OF_ITEM[_itm2] = _v
            else:
                #  radios / gas monitors: serial photo first, model
                #  photo as the fallback (31 Jul 2026)
                import mygear_thumbs as _T3
                _ser3, _mod3 = _T3.derived_keys(
                    r[ix.get('ITEM_DESCRIPTION', 0)])
                if _ser3:
                    VAR_OF_ITEM[_itm2] = _ser3
                    ALT_OF_ITEM[_itm2] = _mod3
                elif _mod3:
                    VAR_OF_ITEM[_itm2] = _mod3
        if bc and bc not in by_bc:
            _itm = str(r[ix['ITEM_NUMBER']] or '').strip()
            by_bc[bc] = (_itm,
                         MASTER.disp(_itm, str(r[ix['ITEM_DESCRIPTION']] or '').strip()))
    return by_bc

# SALES_STOCK is the consumables catalogue - authoritative names for the
# consumable SKUs that flow through the transaction sheets.
def read_sales(path):
    import openpyxl
    names = {}
    if not path: return names
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['SALES_STOCK'] if 'SALES_STOCK' in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c) for c in next(rows)]
    ix = {h.strip(): i for i, h in enumerate(hdr)}
    if not {'SKU_DESCRIPTION'} <= set(ix): return names
    for r in rows:
        if not r: continue
        desc = str(r[ix['SKU_DESCRIPTION']] or '').strip()
        if not desc: continue
        for col in ('SKU_BARCODE', 'SKU_NUMBER'):
            if col in ix and r[ix[col]] not in (None, ''):
                names[str(r[ix[col]]).strip().upper()] = desc
    return names

# --------------------------- read TRANSACTIONS ----------------------------
def read_transactions(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    def sheet_rows(name):
        if name not in wb.sheetnames: return []
        rows = list(wb[name].iter_rows(values_only=True))
        if not rows: return []
        hdr = [str(c) for c in rows[0]]
        return [dict(zip(hdr, r)) for r in rows[1:]
                if r and any(c not in (None, '') for c in r)]
    return sheet_rows('TRANSACTION_CHARGES'), sheet_rows('CUSTOMER_CONTRACTOR_EQUIP')

# --------------------------- build ----------------------------------------
#  DAYS FROM FLAME OFF (Andrew, 2 Aug 2026: "days from Flame Off - this
#  starts at 0 on the 24/07/2026 ... 02/08/2026 being the Power Outage
#  day which is 9"). Optional import so a suite without the module just
#  loses the tag rather than the build.
try:
    import shutdown_day as _SDAY
except Exception:
    _SDAY = None


def _flame_tag(d=None):
    """DAY 9 - POWER OUTAGE, or DAY 9, or nothing at all."""
    if _SDAY is None:
        return ''
    try:
        n = _SDAY.day(d)
        m = _SDAY.milestone(d)
        return 'DAY {}{}'.format(n, (' - ' + m) if m else '')
    except Exception:
        return ''


def build():
    onhire_path = find_export('ON_HIRE*.xlsx', 'ON_HIRE')
    txn_path = find_export('TRANSACTIONS*.xlsx', 'TRANSACTIONS')
    rental_path = find_export('RENTAL_STOCK*.xlsx', 'RENTAL_STOCK', required=False)
    sales_path = find_export('SALES_STOCK*.xlsx', 'SALES_STOCK', required=False)
    for pth in (onhire_path, txn_path, rental_path, sales_path):
        if pth: print('Reading  ' + os.path.basename(pth))
    people, asof = read_onhire(onhire_path)
    tc, eq = read_transactions(txn_path)
    rental_bc = read_rental(rental_path)    # barcode -> (asset no, current desc)
    sales_names = read_sales(sales_path)    # sku/barcode -> consumable name
    #  What is actually ON THE SHELF this morning - the store catalogue
    #  (Andrew, 29 Jul 2026: "saves them waiting and then being told no")
    try:
        import mygear_store
        STOCK = mygear_store.read_availability(rental_path, sales_path, MASTER)
    except Exception as _e:
        print('  NOTE: store catalogue not built ({}) - the rest of the '
              'page is unaffected.'.format(_e))
        STOCK = {'hire': [], 'cons': [], 'stats': {}}
    # transactions run a day behind billing - the true coverage boundary is
    # the latest date actually present in the file; the card says so.
    txn_dates = [parse_date(r.get(k)) for r in tc + eq
                 for k in ('TRAN_START_DATE', 'TRAN_END_DATE')]
    txn_dates = [d for d in txn_dates if d]
    txn_to = fmt_d(max(txn_dates)) if txn_dates else ''
    if not people:
        die('ON_HIRE export has no rows with an EXTERNAL_ID - nothing to build.')
    # ---- roster memory: people who returned EVERYTHING drop off ON_HIRE,
    # but returning it all is the gold standard - they keep their card.
    # MY_GEAR_PEOPLE.csv remembers every id/name/company ever seen in an
    # ON_HIRE pull (ids still only ever come from the on-hire report).
    import csv
    cache_path = os.path.join(BASE, 'MY_GEAR_PEOPLE.csv')
    if os.path.exists(cache_path):
        with io.open(cache_path, encoding='utf-8', newline='') as f:
            for row in csv.reader(f):
                if len(row) >= 3 and row[0] and row[0] not in people:
                    people[row[0]] = {'name': row[1], 'company': row[2],
                                      'hid': '', 'items': []}
    # ---- the site roster: EVERYONE, with their hire ID ---------------
    #  (Andrew, 28 Jul 2026: "see this - we will need this on both
    #  computers, this is everyone.")
    #  ON_HIRE only ever lists people currently holding gear, so a bloke
    #  who has handed everything back - or never taken anything - had no
    #  card and no way to get one. HIRERS_ID.xlsx is the site's own list
    #  of every hirer and their External ID, so now everybody has a page,
    #  even if that page says "nothing in your name".
    try:
        import openpyxl as _rx
        _rp = _newest_file('HIRERS_ID*.xlsx')
        if _rp:
            _rw = _rx.load_workbook(_rp, read_only=True, data_only=True)
            _rr = list(_rw[_rw.sheetnames[0]].iter_rows(values_only=True))
            _rh = {}
            for _row in _rr[:5]:
                _c = [str(v).strip().lower() if v is not None else ''
                      for v in _row]
                if 'name' in _c and 'external id' in _c:
                    _rh = {v: i for i, v in enumerate(_c) if v}
                    break
            _new = 0
            if _rh:
                _ni, _ei = _rh['name'], _rh['external id']
                _ci = _rh.get('employer')
                for _row in _rr:
                    _id = (str(_row[_ei] or '').strip()
                           if _ei < len(_row) else '')
                    _nm = (str(_row[_ni] or '').strip()
                           if _ni < len(_row) else '')
                    if not _id or not _nm or _nm.lower() == 'name':
                        continue
                    if _id in people:
                        continue
                    people[_id] = {
                        'name': norm_name(_nm), 'hid': '', 'items': [],
                        'company': (str(_row[_ci] or '').strip()
                                    if _ci is not None and _ci < len(_row)
                                    else '')}
                    _new += 1
            print('  Site roster: {} more people given a card from {} '
                  '(everyone on site, not just those holding gear).'
                  .format(_new, os.path.basename(_rp)))
            _rw.close()
    except Exception as _e:
        print('  (site roster not read: {})'.format(_e))

    # ---- IDs typed in by hand (Andrew, 28 Jul 2026: "how about i get a
    # full full list, as this won't work doing it this way").
    # ON_HIRE is the ONLY export carrying an EXTERNAL_ID, so anyone who
    # has never held gear under their own ID could never get a card - not
    # even to be told they are holding nothing. The 'ID Cards' sheet of
    # Coates_Report_Recipients.xlsx existed for this and was never read.
    # It is now: put a number against a name there and they get a card.
    try:
        import openpyxl as _ox
        _book = os.path.join(BASE, 'Coates_Report_Recipients.xlsx')
        if os.path.exists(_book):
            _wb = _ox.load_workbook(_book, read_only=True, data_only=True)
            if 'ID Cards' in _wb.sheetnames:
                _rows = list(_wb['ID Cards'].iter_rows(values_only=True))
                _h = {}
                for _r in _rows[:8]:
                    _c = [str(v).strip() if v is not None else '' for v in _r]
                    if 'ID No' in _c and 'Hirer Name' in _c:
                        _h = {v: i for i, v in enumerate(_c) if v}
                        break
                _added = 0
                if _h:
                    for _r in _rows:
                        _id = str(_r[_h['ID No']] or '').strip() \
                            if _h['ID No'] < len(_r) else ''
                        _nm = str(_r[_h['Hirer Name']] or '').strip() \
                            if _h['Hirer Name'] < len(_r) else ''
                        if not _id or not _nm or _id == 'ID No':
                            continue
                        if 'DEMO' in ' '.join(str(v) for v in _r).upper():
                            continue
                        if _id not in people:
                            people[_id] = {'name': norm_name(_nm),
                                           'company': '', 'hid': '',
                                           'items': []}
                            _added += 1
                if _added:
                    print('  {} card(s) added from the ID Cards sheet - people '
                          'with no gear on hire still get a page.'.format(_added))
            _wb.close()
    except Exception as _e:
        print('  (ID Cards sheet not read: {})'.format(_e))

    with io.open(cache_path, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        for idno, p in people.items():
            w.writerow([idno, p['name'], p['company']])
    today = dt.date.today()
    by_name = {p['name']: idno for idno, p in people.items()}

    # per-person transaction truth (Coates rental = TRANSACTION_CHARGES)
    stats = {n: {'ret': 0, 'same': 0, 'dmg': [], 'txn_keys': set(),
                 'dates': set(), 'radio': [0, 0], 'gas': [0, 0],
                 'oth': [0, 0], 'cons': {}} for n in by_name}
    def key_of(row):
        item = str(row.get('SKU/ITEM_NUMBER') or row.get('ITEM_NUMBER') or '').strip()
        d = parse_date(row.get('TRAN_START_DATE'))
        return (item.upper(), d.isoformat() if d else '')
    skip_pool = lambda n: n.lower().startswith('site plant equipment')

    for row in tc:
        n = norm_name(row.get('HIRER_NAME'))
        if n not in stats or skip_pool(n): continue
        s = stats[n]
        s['txn_keys'].add(key_of(row))
        for k in ('TRAN_START_DATE', 'TRAN_END_DATE'):
            d = parse_date(row.get(k))
            if d: s['dates'].add(d)
        sd, ed = parse_date(row.get('TRAN_START_DATE')), parse_date(row.get('TRAN_END_DATE'))
        if ed:
            s['ret'] += 1
            if sd == ed: s['same'] += 1
        try: dmg = float(row.get('DAMAGE ($)') or 0)
        except (TypeError, ValueError): dmg = 0
        if dmg > 0:
            s['dmg'].append(str(row.get('SKU/ITEM DESCRIPTION') or 'item')[:40])
        desc = str(row.get('SKU/ITEM DESCRIPTION') or '')
        if GAS_RE.search(desc):
            s['gas'][0] += 1
            if ed: s['gas'][1] += 1

    for row in eq:
        n = norm_name(row.get('HIRER_NAME'))
        if n not in stats or skip_pool(n): continue
        s = stats[n]
        s['txn_keys'].add(key_of(row))
        for k in ('TRAN_START_DATE', 'TRAN_END_DATE'):
            d = parse_date(row.get(k))
            if d: s['dates'].add(d)
        ed = parse_date(row.get('TRAN_END_DATE'))
        desc = str(row.get('PRODUCT_VARIANT') or row.get('SKU/ITEM DESCRIPTION') or '')
        cat = str(row.get('PRODUCT_CATEGORY') or '')
        try: qty = int(float(row.get('QUANTITY') or 1))
        except (TypeError, ValueError): qty = 1
        if cat == 'Consumable':
            # SALES_STOCK carries the authoritative consumable names
            bc = str(row.get('LATEST_BARCODE') or '').strip().upper()
            sku = str(row.get('SKU/ITEM_NUMBER') or '').strip().upper()
            best = sales_names.get(bc) or sales_names.get(sku) or desc
            k = re.sub(r'\s*-\s*K2/?\d*$', '', best.strip()[:40]).strip(' -')
            s['cons'][k] = s['cons'].get(k, 0) + max(1, qty)
        elif GAS_RE.search(desc):
            s['gas'][0] += 1
            if ed: s['gas'][1] += 1
        elif RADIO_RE.search(desc):
            s['radio'][0] += 1
            if ed: s['radio'][1] += 1
        else:
            s['oth'][0] += 1
            if ed: s['oth'][1] += 1

    # on-hire items also count as store activity (and catch unbilled days)
    for idno, p in people.items():
        s = stats[p['name']]
        for it in p['items']:
            s['txn_keys'].add((it['item'].upper(), it['start'].isoformat() if it['start'] else ''))
            if it['start']: s['dates'].add(it['start'])
            if GAS_RE.search(it['desc']) and not parse_date(None):
                pass  # gas items currently out are counted via txn sheets when billed

    # ---- Charge Reporter damage register: damage recorded against a person
    # also shows on their card (e.g. gear billed outside SiteIQ's charge
    # lines). Same alert as SiteIQ damage charges; de-duplicated by item
    # description so a charge that later reaches SiteIQ never doubles up.
    # (A. Fisher, 24 Jul 2026)
    reg_path = os.path.join(BASE, 'Coates_K2_Charge_Reporter', 'CHARGE_REGISTER.xlsx')
    dmg_unattached = []   # (person, item) with no card in today's ON_HIRE
    if os.path.isfile(reg_path):
        try:
            import openpyxl
            rwb = openpyxl.load_workbook(reg_path, read_only=True, data_only=True)
            if 'Damage - Breakdown' in rwb.sheetnames:
                rrows = list(rwb['Damage - Breakdown'].iter_rows(values_only=True))
                rhdr = {str(v).strip(): i for i, v in enumerate(rrows[0]) if v}
                for rr in rrows[1:]:
                    def rg(col):
                        i = rhdr.get(col)
                        return rr[i] if i is not None and i < len(rr) else None
                    if not rg('Charge ID'):
                        continue
                    who = norm_name(re.sub(r'\s*\(.*?\)\s*', ' ', str(rg('Person using / returning item') or '')))
                    who = re.sub(r'\s+', ' ', who).strip()
                    d = str(rg('Item description') or 'item')[:40]
                    if who in stats:
                        if d not in stats[who]['dmg']:
                            stats[who]['dmg'].append(d)
                    else:
                        #  Damage against someone with no card today used
                        #  to vanish without a word - the register said
                        #  $600, the page said 0, and nobody knew why.
                        #  It still can't sit on a card that doesn't
                        #  exist, but now it is SAID, and counted in the
                        #  site-wide damage figure. (Found 26 Jul 2026 -
                        #  Mason Thomas's radio.)
                        dmg_unattached.append((who or 'unnamed', d))
            rwb.close()
        except Exception:
            pass    # unreadable register never breaks the gear page
    if dmg_unattached:
        print('  NOTE: {} damage charge(s) in the register belong to people '
              'with no card in today\'s ON_HIRE export: {}. They are counted '
              'in the site-wide damage figure and will attach the moment the '
              'person appears in an ON_HIRE pull.'.format(
                  len(dmg_unattached),
                  '; '.join('{} ({})'.format(w, d) for w, d in dmg_unattached)))

    # ---------------- score model (documented, transparent) ---------------
    # Returns Score = 100 x (0.75 x same-day rate + 0.25 x returned rate)
    #   same-day rate = same-day returns / returns
    #   returned rate = returns / total items had (had = out now + returned)
    # Reproduces the approved anchors: 1-of-1 same-day + 1 still out -> 88;
    # 1 return (not same-day) of 12 -> 2. No returns yet -> no score shown.
    order = list(people.keys())  # ON_HIRE first-appearance order (stable ties)
    computed = {}
    for idno in order:
        p = people[idno]; s = stats[p['name']]
        out_now = len(p['items']); had = out_now + s['ret']
        if s['ret'] > 0:
            sd_rate = s['same'] / s['ret']
            rt_rate = s['ret'] / had if had else 0
            score = int(100 * (0.75 * sd_rate + 0.25 * rt_rate) + 0.5)
        else:
            score = 0
        computed[idno] = {'out': out_now, 'had': had, 'ret': s['ret'],
                          'same': s['same'], 'score': score}
    total = len(order)
    site_sorted = sorted(order, key=lambda i: (-computed[i]['score'], order.index(i)))
    site_rank = {i: n + 1 for n, i in enumerate(site_sorted)}
    site_avg = int(sum(c['score'] for c in computed.values()) / total + 0.5)

    #  MOVEMENT - how each person is tracking against the last build.
    #  Yesterday's board is read BEFORE today's is written, or everyone
    #  would be compared against themselves and hold their spot forever.
    _MOVE = {}
    try:
        import mygear_movement as _mv
        _today_key = _mv.today_key()
        _prev_day, _prev = _mv.previous_day(BASE, _today_key)
        for _i in computed:
            m = _mv.movement(_prev, _i, computed[_i]['score'], site_rank[_i])
            if m:
                _MOVE[_i] = m
        _days = _mv.save(BASE, _today_key,
                         {_i: {'score': computed[_i]['score'],
                               'rank': site_rank[_i]} for _i in computed})
        if _prev_day:
            _up = sum(1 for m in _MOVE.values() if m['places'] > 0)
            print('  Movement: compared against {} - {} people moved up, '
                  '{} day(s) of history kept.'.format(_prev_day, _up, _days))
        else:
            print('  Movement: first scoreboard saved - arrows start '
                  'appearing on tomorrow\'s build.')
    except Exception as _e:
        print('  NOTE: movement not available ({}) - scores still build.'
              .format(_e))
    comp_of = {i: people[i]['company'].strip().upper() for i in order}
    crew_avg, crew_rank, crew_total = {}, {}, {}
    for co in set(comp_of.values()):
        members = [i for i in order if comp_of[i] == co]
        avg = int(sum(computed[i]['score'] for i in members) / len(members) + 0.5)
        ms = sorted(members, key=lambda i: (-computed[i]['score'], order.index(i)))
        for n, i in enumerate(ms):
            crew_avg[i], crew_rank[i], crew_total[i] = avg, n + 1, len(members)

    #  Crew standings (Andrew, 28 Jul 2026: "how they score against
    #  other - their company or even onsite"). Every crew ranked by its
    #  average score so a card can say "your crew sits #2 of 12 on
    #  site". A crew of one still ranks - the wording stays kind.
    _co_avg = {}
    for co in set(comp_of.values()):
        members = [i for i in order if comp_of[i] == co]
        _co_avg[co] = int(sum(computed[i]['score'] for i in members) /
                          len(members) + 0.5)
    _co_rank = {co: n + 1 for n, (co, _a) in enumerate(
        sorted(_co_avg.items(), key=lambda kv: (-kv[1], kv[0])))}
    #  Top crew for the landing pulse: judged only among crews of 3+ so
    #  one bloke with one perfect return doesn't outrank a whole squad.
    _big = [co for co in _co_avg
            if sum(1 for i in order if comp_of[i] == co) >= 3]
    top_crew = (max(_big, key=lambda co: _co_avg[co]) if _big
                else (max(_co_avg, key=_co_avg.get) if _co_avg else ''))

    # ---------------- payloads --------------------------------------------
    DATA, warn_dupes = {}, {}
    tot_items = rad_out_tot = gas_out_tot = dmg_tot = 0
    for idno in order:
        p = people[idno]; s = stats[p['name']]; c = computed[idno]
        norm = re.sub(r'\s+', '', idno)
        if norm in warn_dupes:
            print('  WARNING: IDs {} and {} normalise the same - only the first '
                  'is served.'.format(warn_dupes[norm], idno)); continue
        warn_dupes[norm] = idno
        # first goes straight into the story card via innerHTML, so it is
        # HTML-escaped here - a name with a stray < or & can never break or
        # inject into the page. The story's own <b> tags are added below,
        # around this already-safe value.
        name = p['name']; first = html.escape(name.split(' ')[0] or name, quote=False)
        initials = ''.join(w[0] for w in name.split()[:2]).upper() or 'K2'
        company = re.sub(r'\s+', ' ', p['company']).title()
        items, mixc, aging = [], {}, {'g': 0, 'a': 0, 'r': 0}
        #  Longest-held first, then A-Z within the same day count, so two
        #  items out the same three days always sit in the same order.
        #  (Andrew, 29 Jul 2026: "starting longest days a-z")
        for it in sorted(p['items'],
                         key=lambda x: (-((today - x['start']).days
                                          if x['start'] else -1),
                                        (x['desc'] or '').upper())):
            days = max(0, (today - it['start']).days) if it['start'] else '-'
            # RENTAL_STOCK wins on asset number + current description
            asset, desc = it['item'], it['desc']
            if it.get('barcode') and it['barcode'] in rental_bc:
                a2, d2 = rental_bc[it['barcode']]
                asset, desc = (a2 or asset), (d2 or desc)
            cat = category(desc)
            mixc[cat] = mixc.get(cat, 0) + 1
            if isinstance(days, int):
                aging['g' if days <= 2 else ('a' if days <= 4 else 'r')] += 1
            # 'b' = the compliance chips for this line (already-escaped HTML
            # from equipment_compliance). Empty string when the master says
            # this item has nothing to declare - so the card stays clean.
            items.append({'d': desc, 'n': asset, 'days': days,
                          'c': CAT_COLOURS[cat],
                          'b': EC.badges_html(asset, desc, wrap=False),
                          # the orange Plant ID pill - the number the crews
                          # actually say out loud (A. Fisher, 25 Jul 2026)
                          'pid': EC.plant_id(asset),
                          # the photo key - shows WHAT the item looks like
                          # on the card once its variant photo is collected
                          'v': VAR_OF_ITEM.get(asset, ''),
                          'va': ALT_OF_ITEM.get(asset, ''),
                          # replacement cost, printed on the A4 in the
                          # highlighter - null shows TBC, never $0
                          'r': _repl_cost(asset, desc)})
        tot_items += len(items)
        # tie-break matches the established cards: count desc, then name Z->A
        mix = [[k, v, CAT_COLOURS[k]] for k, v in
               sorted(sorted(mixc.items(), key=lambda kv: kv[0], reverse=True),
                      key=lambda kv: -kv[1])]
        sd, ret, had, out_now, score = c['same'], c['ret'], c['had'], c['out'], c['score']
        if sd > 0: rating = {'stars': 5, 'label': 'Same-day returner'}
        elif ret > 0: rating = {'stars': 2, 'label': 'Returns, but not same-day'}
        else: rating = {'stars': 0, 'label': 'No returns logged yet'}
        # badges - thresholds documented here, all data-driven
        badges = []
        #  third slot is the icon slug in Art\icons\ (Andrew's elite
        #  pack, 2 Aug 2026). The character stays as the fallback for
        #  any build where the artwork is missing.
        if out_now >= 8: badges.append(['▣', 'Big Kit', 'big-kit'])
        if len(mixc) >= 4: badges.append(['✦', 'All-Rounder', 'all-rounder'])
        if len(s['txn_keys']) >= 8: badges.append(['⚡', 'Store Regular', 'store-regular'])
        if mixc.get('Rigging'): badges.append(['⛓', 'Rigger', 'rigger'])
        if mixc.get('Electrical'): badges.append(['⭐', 'Powerhouse', 'powerhouse'])
        if sd >= 1: badges.append(['↺', 'Same-Day Legend', 'same-day-legend'])
        if crew_rank[idno] == 1 and sd >= 1: badges.append(['\U0001f3c6', 'Top of the Store', 'top-of-store'])
        if site_rank[idno] == 1 and sd >= 1: badges.append(['\U0001f525', 'Site Legend', 'site-legend'])
        # approved wording (22 Jul 2026) - do not reintroduce "Good on ya"
        if sd > 0:
            story = ("Nice work, " + first + " — you've brought back <b>" +
                     str(ret) + " of " + str(had) + "</b> you've had, <b>" + str(sd) +
                     " same-day</b>. Ripper — that's what keeps the yard humming "
                     "and the next crew equipped. Thanks for doing your bit.")
        elif ret > 0:
            story = (first + ", you've brought back <b>" + str(ret) + " of " +
                     str(had) + "</b> so far. Sending gear back the same day "
                     "you're finished keeps your record clean and the gear moving "
                     "— an easy one when you're passing the store.")
        else:
            story = (first + ", you've had " + str(out_now) +
                     (" item" if out_now == 1 else " items") + " out and none are "
                     "back yet — send one in and get your first same-day "
                     "return on the board. Thanks for doing your bit.")
        visits, txns = len(s['dates']), len(s['txn_keys'])
        if txns > 0:
            story += (" You've used the store " +
                      ("once" if visits == 1 else str(visits) + " times") +
                      " this shut — " + str(txns) +
                      (" transaction" if txns == 1 else " transactions") + " all up.")
        cons_n = sum(s['cons'].values())
        rad_t, rad_b = s['radio']; gas_t, gas_b = s['gas']; oth_t, oth_b = s['oth']
        rad_out_tot += rad_t - rad_b; gas_out_tot += gas_t - gas_b
        dmg_tot += len(s['dmg'])
        # What this person is carrying, counted up. The per-line chips tell
        # you what each item needs; this tells them the size of the job in
        # one sentence, so a bloke with six tagged leads reads it once
        # instead of six times.
        _cn = EC.summarise(
            [{'asset': i['n'], 'description': i['d']} for i in items])
        payload = {
            'name': name, 'company': company, 'id': idno, 'initials': initials,
            #  both scannable numbers: the card ID and SiteIQ's hirer ID
            'hid': p.get('hid', ''),
            'stats': {'items': out_now, 'types': len({i['d'] for i in items}),
                      'returned': ret, 'sameday': sd,
                      'rex': sum(i['r'] for i in items if i.get('r')),
                      'rex_tbc': sum(1 for i in items if not i.get('r'))},
            'rating': rating, 'score': score, 'hasReturns': ret > 0,
            'rank': {'comp': crew_rank[idno], 'compTotal': crew_total[idno],
                     'site': site_rank[idno], 'siteTotal': total,
                     'pct': int(site_rank[idno] / total * 100 + 0.5),
                     #  the whole crew's standing on site, for the
                     #  "your crew sits #N of M" line on the card
                     'crewPos': _co_rank.get(comp_of[idno], 0),
                     'crewOf': len(_co_rank)},
            'cmp': {'you': score, 'crew': crew_avg[idno], 'site': site_avg},
            #  how they are tracking against yesterday - None on the very
            #  first day, so the card says nothing rather than guessing
            'move': _MOVE.get(idno),
            'mix': mix, 'story': story, 'items': items, 'aging': aging,
            'badges': badges, 'comp': _cn,
            'act': {'visits': visits, 'txns': txns, 'to': txn_to},
            'cons': {'n': cons_n,
                     'list': [{'d': k, 'q': v} for k, v in sorted(s['cons'].items())]},
            'radios': {'taken': rad_t, 'back': rad_b, 'out': rad_t - rad_b},
            'gas': {'taken': gas_t, 'back': gas_b, 'out': gas_t - gas_b},
            'oth': {'taken': oth_t, 'back': oth_b, 'out': oth_t - oth_b},
            'dmg': {'n': len(s['dmg']), 'list': s['dmg']},
        }
        DATA[lk_tag(norm)] = lk_enc(norm, json.dumps(payload, ensure_ascii=True,
                                                     separators=(',', ':')))
    DATA[lk_tag('SELFTEST')] = lk_enc('SELFTEST', '{"name":"SELFTEST"}')

    # store-only people we cannot serve (no EXTERNAL_ID in today's ON_HIRE)
    known = set(by_name)
    seen = {norm_name(r.get('HIRER_NAME')) for r in tc + eq} - {''}
    missing = sorted(n for n in seen if n not in known and not skip_pool(n))
    if missing:
        print('  NOTE: store activity but no card (no ID in today\'s ON_HIRE '
              'export): ' + ', '.join(missing))
        print('        They get a card as soon as they appear in an ON_HIRE '
              'pull with their EXTERNAL_ID.')

    #  The phone-first layer: the site guides, the save-to-phone button,
    #  the card scanner. jsQR rides along so the scanner works on an
    #  iPhone, where the browser has no barcode reader of its own - the
    #  page still opens with no signal once it's on the handset.
    _jsqr = ''
    _qp = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       'jsQR.min.js')
    if os.path.exists(_qp):
        with open(_qp, 'r', encoding='utf-8') as _fh:
            _jsqr = _fh.read()
    else:
        print('  NOTE: jsQR.min.js is missing, so the Scan button will only '
              'work on phones with a built-in reader. Typing the ID still '
              'works everywhere.')
    #  the store catalogue rides in as one more guide screen - same
    #  shelf button, same full-screen sheet, nothing new to learn
    _store_btn, _store_pane = '', ''
    if STOCK.get('hire') or STOCK.get('cons'):
        _sst = STOCK['stats']
        _store_btn = (
            "<button class='gbtn' onclick=\"openGuide('store')\">"
            "<svg viewBox='0 0 24 24' fill='none' stroke='currentColor' "
            "stroke-width='1.7' stroke-linecap='round' stroke-linejoin='round'>"
            "<path d='M4 8h16v11a1 1 0 0 1-1 1H5a1 1 0 0 1-1-1zM4 8l2-4h12l2 4"
            "M9 12h6'/></svg>"
            "<div><b>What's in the store</b><span>{h} ready to hire &middot; "
            "{c} consumable lines &mdash; search it</span></div>"
            "<em>&rsaquo;</em></button>").format(
                h=_sst.get('hireItems', 0), c=_sst.get('consLines', 0))
        _store_pane = ("<div id='g-store' class='gpane'>"
                       + mygear_store.pane(STOCK, asof) + "</div>")
    #  guides live on the corner pills (RADIO / GAS / CONTACTS) - the
    #  big duplicate cards came off the front page (Andrew, 31 Jul 2026:
    #  'we have those pills on the right side... hence why i asked')
    #  retired 1 Aug 2026 - the front door tiles replaced the shelf.
    #  Kept because it is one line and the next site may want it back.
    _shelf = mygear_ui.shelf_html(_store_btn)  # noqa: F841
    #  The site pulse - live numbers on the front door (Andrew, 28 Jul
    #  2026: "utilising all reports to provide as much data as we can").
    #  Every figure is computed above from today's exports; the counters
    #  animate up on load, neon so they read from a metre away.
    _sd_site = sum(c['same'] for c in computed.values())
    _esc_py = lambda s: (str(s).replace('&', '&amp;').replace('<', '&lt;')
                         .replace('>', '&gt;'))
    #  ACTIVE people, not carded people (Andrew, 29 Jul 2026: "total
    #  active people using the store, not total of people added into the
    #  database"). A card in the drawer is not a user - a transaction is.
    #  The site plant pool is a location, not a person, and is excluded
    #  or it would sit top of every board with hundreds of movements.
    _active = {norm_name(r.get('HIRER_NAME')) for r in tc + eq}
    _active = {n for n in _active if n and not skip_pool(n)}
    _firms = {str(r.get('EMPLOYER_NAME') or '').strip() for r in tc + eq}
    _firms |= {str(p.get('comp') or '').strip() for p in people.values()}
    _firms = {f for f in _firms if f and f.lower() not in ('', 'none')}
    _st = STOCK.get('stats', {})
    def _pu(v, label, sub=''):
        return ('<span class="pu"><b class="pv" data-to="{v}">0</b>{l}'
                '{s}</span>').format(v=v, l=label,
                                     s=('<i>' + sub + '</i>') if sub else '')
    pulse = ('<div class="pulse">'
             + _pu(len(_active), 'USING THE STORE',
                   '{} carded'.format(total))
             + _pu(len(_firms), 'COMPANIES')
             + _pu(tot_items, 'ITEMS OUT NOW')
             + (_pu(_st.get('hireItems', 0), 'READY TO HIRE',
                    '{} different things'.format(_st.get('hireLines', 0)))
                if _st.get('hireItems') else '')
             + _pu(_sd_site, 'SAME-DAY RETURNS')
             + ('<span class="pu"><b class="pvt">{t}</b>TOP CREW</span>'
                if top_crew else '')
             + '</div>').format(t=_esc_py(top_crew.title()) if top_crew else '')
    #  The front door's instrument strip - three numbers, one row. The
    #  full six-stat pulse still exists above for anyone who wants it
    #  back; the door only ever needed the three a worker reads.
    def _stat(art, n, label, tone=''):
        return ("<div class='s3" + (" " + tone if tone else "") + "'>"
                "<i style=\"background-image:url('art/" + art + ".jpg')\"></i>"
                "<div><b>" + "{:,}".format(n) + "</b><span>" + label
                + "</span></div></div>")

    _strip = ('<div class="stats3">'
              + _stat('st_users', len(_active), 'USERS')
              + _stat('st_onhire', tot_items, 'ON HIRE', 'amb')
              + (_stat('st_ready', _st.get('hireItems', 0), 'READY TO HIRE',
                       'grn') if _st.get('hireItems') else '')
              + '</div>')
    _gl_dir = os.path.join(BASE, 'Gear_Lookup')
    os.makedirs(_gl_dir, exist_ok=True)
    #  the front door's artwork (Andrew's own renders, 1 Aug 2026) - copied
    #  beside the page so phones load it off the store Wi-Fi like the
    #  thumbnails, not baked into the HTML where it would bloat every load
    try:
        import shutil as _sh
        _art_src = os.path.join(BASE, 'Art')
        _art_dst = os.path.join(_gl_dir, 'art')
        if os.path.isdir(_art_src):
            #  walk it - the icon pack lives in Art\icons\ and a flat
            #  listdir silently skipped every one of them (found 2 Aug
            #  2026: the cards fell back to line drawings and nobody
            #  would have known why)
            _n_art = 0
            for _root, _dirs, _files in os.walk(_art_src):
                _rel = os.path.relpath(_root, _art_src)
                _dst = _art_dst if _rel == '.' else os.path.join(_art_dst, _rel)
                os.makedirs(_dst, exist_ok=True)
                for _a in sorted(_files):
                    if _a.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
                        _sh.copyfile(os.path.join(_root, _a),
                                     os.path.join(_dst, _a))
                        _n_art += 1
            print('  Front door artwork: {} file(s) copied.'.format(_n_art))
        else:
            print('  Front door artwork: Art\\ folder not found - the page '
                  'still builds, the cards just wear their plain look.')
    except (KeyboardInterrupt, SystemExit):
        raise
    except BaseException as _e:
        print('  Front door artwork skipped ({!r})'.format(_e))
    _stores_tag = ''
    #  THE STORES TEAM PAGE - the counter's own view, behind a code.
    #  A separate file on purpose: the crew page stays light, and the
    #  staff view can carry who-has-what without putting it in front of
    #  900 people. The code lives in stores_code.txt (protected from
    #  updates); change the file, re-run 04, done.
    try:
        import mygear_stores
        _code_p = os.path.join(BASE, 'stores_code.txt')
        _code = '2026'
        if os.path.isfile(_code_p):
            with io.open(_code_p, encoding='utf-8') as _fh:
                _code = (_fh.read().strip() or _code)
        else:
            with io.open(_code_p, 'w', encoding='utf-8') as _fh:
                _fh.write('2026\n')
            print('  Stores code file created: stores_code.txt (code 2026 - '
                  'change it and re-run).')
        _sk = find_export('STOCKTAKE*.xlsx', 'STOCKTAKE', required=False)
        #  SALES_STOCK feeds the consumables pane. Optional on purpose -
        #  a store with no consumables still gets a board.
        _sales = find_export('SALES_STOCK*.xlsx', 'SALES_STOCK',
                             required=False)
        _sd = mygear_stores.read(rental_path, _sk, MASTER,
                                 txn_path=txn_path,
                                 sales_path=_sales, base=BASE)
        #  the manager layer - money, under its own code
        _mgr_p = os.path.join(BASE, 'manager_code.txt')
        _mgr = 'army8686ARRA'
        if os.path.isfile(_mgr_p):
            with io.open(_mgr_p, encoding='utf-8') as _fh:
                _mgr = (_fh.read().strip() or _mgr)
        else:
            with io.open(_mgr_p, 'w', encoding='utf-8') as _fh:
                _fh.write('army8686ARRA\n')
        _pr = mygear_stores._pricing(onhire_path, MASTER)
        with io.open(os.path.join(_gl_dir, 'stores.html'), 'w',
                     encoding='utf-8') as _fh:
            _fh.write(mygear_stores.build(_sd, _code, asof,
                                          pricing=_pr, mgr_code=_mgr))
        #  the front-door override answers to the code AND its numeric
        #  phone-keypad twin - the crew ID box only offers a number
        #  keypad on phones, so a letters-only code could never be
        #  typed there (caught 29 Jul 2026: "no option to enter
        #  letters"). NOIS answers to 6647.
        _stores_tag = mygear_stores.tag(_code.upper())
        _alias = mygear_stores.keypad_alias(_code)
        if _alias:
            _stores_tag += ',' + mygear_stores.tag(_alias)
        #  Say which code this build actually answers to - masked, but
        #  enough to see AT A GLANCE that the laptop is running "2***"
        #  when everyone has been told the code starts with N. A whole
        #  evening was lost to exactly that on 29 Jul 2026: the
        #  protected stores_code.txt had been created with the default
        #  on one machine while the code everyone knew lived on
        #  another. The file wins, so the build must show the file.
        print('  Stores door: code {}{} ({} chars, from stores_code.txt)'
              '{}'.format(_code[0], '*' * (len(_code) - 1), len(_code),
                          ' | phone keypad twin {}{} works too'.format(
                              _alias[0], '*' * (len(_alias) - 1))
                          if _alias else ''))
        _t = _sd['tiles']
        print('  Stores team page: {} on the shelf | {} out | {} to chase '
              '| stocktake {}% | {} not counted | {} arriving'.format(
                  _t['avail'], _t['onhire'], _t['chase'], _t['stockPct'],
                  _t['stale'], _t['arrivals']))
        if _pr:
            print('  Manager layer: ${:,.2f}/day on hire, {} zero-rate '
                  'line(s) flagged.'.format(_pr['perDay'], _pr['zeroN']))
        #  A filter that removes lines silently is a filter nobody
        #  trusts. If a count ever looks light, this line already said
        #  why. (Andrew's hidden items, 2 Aug 2026.)
        try:
            if _SDAY is not None:
                print('  Shutdown clock: ' + _SDAY.label())
        except Exception:
            pass
        try:
            import hidden_stock as _HS
            if _HS.count():
                print('  Hidden items: {} rule(s) applied from '
                      'HIDDEN_ITEMS.txt - those lines are left out of '
                      'every report.'.format(_HS.count()))
        except Exception:
            pass
    except Exception as _e:
        #  This is not a side dish failing - if this block dies, the
        #  stores DOOR dies with it: no code of any kind will open the
        #  board, and the front page just says "no gear found". Shout.
        print('  ' + '!' * 62)
        print('  WARNING: the stores page FAILED to build ({}).'.format(_e))
        print('  The stores door is DEAD this build - NO code will open')
        print('  the board until this is fixed and 04 is run again.')
        print('  The crew page itself is unaffected.')
        print('  ' + '!' * 62)

    #  GEAR PICTURES, BEFORE THE PAGE IS WRITTEN, NOT AFTER.
    #  This used to run at the very end, which meant the page could
    #  never know which photos existed - so every tile shipped an
    #  <img> and found out by 404 whether there was anything behind
    #  it. On the rig: 60 cards in a bay, 56 still holding an image
    #  that had not failed yet, and 56 empty holes on the phone until
    #  the store Wi-Fi returned each miss. Andrew photographed exactly
    #  that (2 Aug 2026) - a socket card with nothing in it.
    #
    #  Now the shrink runs FIRST and the page is handed the list of
    #  thumbnails that actually exist, so a tile with no photo draws
    #  its monogram immediately and asks the network for nothing.
    #  None = the build could NOT tell (the shrink threw, old suite).
    #  [] = it looked and there are genuinely none. The page treats the
    #  two differently: unknown means ask the network the way it always
    #  did, empty means draw every monogram now. Conflating them is how
    #  the first cut of this fix quietly did nothing on a store with no
    #  photos yet - which is exactly the store this was written for.
    _THUMBSET = None
    try:
        import mygear_thumbs
        _n, _made, _ready = mygear_thumbs.refresh(BASE)
        _reg = len(mygear_thumbs.variant_register(BASE))
        _tdir = os.path.join(BASE, 'Gear_Lookup', 'thumbs')
        if os.path.isdir(_tdir):
            _THUMBSET = sorted(f[:-4] for f in os.listdir(_tdir)
                               if f.lower().endswith('.jpg'))
        print('  Gear pictures: {} of {} variants have a photo{} - run '
              '56_PHOTO_HUNT for the wanted list.'.format(
                  _ready, _reg,
                  ' ({} shrunk this build)'.format(_made) if _made else ''))
        if _reg and _ready < _reg:
            print('    The other {} draw a two-letter tile instead - no '
                  'empty boxes, and no wasted lookups on the store '
                  'Wi-Fi.'.format(_reg - _ready))
    except Exception as _e:
        #  No manifest = every tile falls back to asking, exactly as it
        #  did before. Degraded, not broken.
        print('  Gear pictures: skipped this build ({})'.format(_e))

    page = (TEMPLATE
            .replace('__DATA__', json.dumps(DATA))
            .replace('__THUMBS__', json.dumps(_THUMBSET))
            .replace('__PULSE__', pulse)
            .replace('__TOPIC__', _topic_banner())
            .replace('__STATSTRIP__', _strip)
            .replace('__TILES__', _front_tiles())
            .replace('__HERO__', _hero(_sst if _store_pane else {},
                                       bool(_store_pane)))
            .replace('__STORESTAG__', _stores_tag)
            .replace('__ASOF__', asof or 'last refresh')
            .replace('__FDAY__', _flame_tag())
            .replace('__UICSS__', mygear_font.FONT_CSS + mygear_ui.CSS
                     + mygear_guides.TT_CSS)
            .replace('//__QRJS__//',
                     __import__('mygear_stores')._QR_JS
                     + '\nfunction qr(t,px){if(t==null||t===\'\'){return \'\';}'
                       'return QRL.svg(String(t),px||46);}')
            .replace('__IDROW__', mygear_ui.ID_ROW)
            .replace('__SHEET__', mygear_ui.sheet_html(
                mygear_guides.guides_html() + _store_pane))
            .replace('__STORECSS__', mygear_store.CSS if _store_pane else '')
            .replace('__STOREJS__',
                     ('var STORE=' + json.dumps(
                         STOCK['hire'] + STOCK['cons'],
                         separators=(',', ':')) + ';\n'
                      #  the CURRENT tag colour word, stamped from the
                      #  compliance master (Jul-Aug = BLUE) so a new
                      #  quarter needs a rebuild, never a code change
                      + mygear_store.JS
                          .replace('__TAGC__', (EC.tag_colour()[0] or ''))
                          .replace('__TAGX__', EC.tag_hex() or '#8A97A8'))
                     if _store_pane else 'var STORE=[];')
            .replace('__UIJS__', (_jsqr + '\n' if _jsqr else '')
                     + mygear_ui.JS))
    out_dir = os.path.join(BASE, 'Gear_Lookup')
    os.makedirs(out_dir, exist_ok=True)
    #  Gear_Lookup is SERVED to every phone on the store Wi-Fi, so the
    #  office People List (the whole site's names and hire IDs in one
    #  file) must never sit in it. It now builds into People_List\, and
    #  this sweep - run every morning with the page - removes any copy an
    #  older build left behind. (Security, 28 Jul 2026.)
    for _office in ('People_List.csv', 'People_List.html', 'Needs_An_ID.csv'):
        _p = os.path.join(out_dir, _office)
        if os.path.isfile(_p):
            try:
                os.remove(_p)
                print('  cleaned out of the served folder: ' + _office)
            except OSError:
                pass
    out = os.path.join(out_dir, 'index.html')
    with io.open(out, 'w', encoding='utf-8') as f:
        f.write(page)
    print('My Gear scorecard page built: {} people, {} items on hire.'.format(
        len(warn_dupes), tot_items))
    print('  Radios still out across site: {} | Gas monitors out: {} | '
          'Damage charges: {}{}'.format(
              rad_out_tot, gas_out_tot, dmg_tot + len(dmg_unattached),
              ' ({} awaiting a card)'.format(len(dmg_unattached))
              if dmg_unattached else ''))
    print('  Data as at: ' + (asof or 'unknown') + ' | Output: ' + out)

def _topic_banner():
    """Today's toolbox talk on the front door (Andrew, 1 Aug 2026), now
    wearing a photograph. The art follows the topic where we have one
    that fits - the radio shot on radio day, the gas monitor on gas day -
    and the tool montage the rest of the time. Never allowed to take the
    daily build down."""
    try:
        import safety_conversation as _SC
        tp = _SC.topic_for()
        title = str(tp["title"])
        low = title.lower()
        art = ('radio' if 'radio' in low else
               'gas' if 'gas' in low else
               'tables' if ('sling' in low or 'lift' in low or 'rig' in low
                            or 'torque' in low) else
               'store')
        return ("<div class='talk'>"
                "<i style=\"background-image:url('art/" + art + ".jpg')\"></i>"
                "<div><span>TOOLBOX TALK &middot; {d} OF {of}</span>"
                "<b>{t}</b><p>{k}</p></div></div>").format(
            d=tp["day"], of=tp["of"], t=html.escape(title),
            k=html.escape(str(tp.get("takeaway") or "")))
    except Exception as _e:
        print("  Toolbox topic skipped ({})".format(_e))
        return ""


#  THE FRONT DOOR TILES (Andrew, 1 Aug 2026: "include the contacts.
#  radio and gas monitor as well as viewing what we have in the store").
#  Everything a worker needs is ON the door now - and it fixed a real
#  hole: the trade tables guide had no way in at all, because the crew
#  page never called guide_buttons().
_FD_ICONS = {
    'store': ("<path d='M4 8h16v11a1 1 0 0 1-1 1H5a1 1 0 0 1-1-1zM4 8l2-4h12"
              "l2 4M9 12h6'/>"),
    'radio': ("<rect x='7' y='2' width='10' height='20' rx='2'/>"
              "<path d='M11 6h2M10 18h4'/>"),
    'gas': "<circle cx='12' cy='12' r='9'/><path d='M12 7v5l3 3'/>",
    'contacts': ("<path d='M22 16.9v3a2 2 0 0 1-2.2 2 19.8 19.8 0 0 1-8.6-3.1"
                 " 19.5 19.5 0 0 1-6-6A19.8 19.8 0 0 1 2.1 4.2 2 2 0 0 1 4.1 2"
                 "h3a2 2 0 0 1 2 1.7c.1 1 .4 2 .7 2.9a2 2 0 0 1-.5 2.1L8.1 9.9"
                 "a16 16 0 0 0 6 6l1.2-1.2a2 2 0 0 1 2.1-.5c.9.3 1.9.6 2.9.7a2"
                 " 2 0 0 1 1.7 2z'/>"),
    'tables': "<path d='M3 5h18v14H3zM3 10h18M9 5v14M15 5v14'/>",
}


def _fd_tile(key, chip, title, sub, opens):
    """Andrew's navigation cards (1 Aug 2026): the photograph IS the card,
    a category chip says what kind of thing it is, and the bottom line
    states exactly what opens before a thumb goes anywhere near it."""
    return ("<button class='navt' type='button' onclick=\"openGuide('"
            + key + "')\" style=\"background-image:url('art/" + key
            + ".jpg')\">"
            "<span class='ntchip'>" + chip + "</span>"
            "<span class='ntgo'>&rsaquo;</span>"
            "<span class='ntb'><b>" + title + "</b>"
            "<span>" + sub + "</span>"
            "<em>" + opens + " &rarr;</em></span></button>")


def _front_tiles():
    """Radio, gas, the contact board and the trade tables. The store is
    not here - it is the hero card at the top of the door."""
    return (_fd_tile('radio', 'GUIDE', 'Two-way radio',
                     'Channels, calls &amp; care', 'OPEN RADIO GUIDE')
            + _fd_tile('gas', 'GUIDE', 'Gas monitor',
                       'Check it. Wear it.', 'OPEN GAS GUIDE')
            + _fd_tile('contacts', 'DIRECTORY', 'Contact board',
                       'Every number. Tap to call.', 'OPEN CONTACTS')
            + _fd_tile('tables', 'REFERENCE', 'Trade tables',
                       'Spanners &middot; taps &middot; WLL',
                       'OPEN TRADE TABLES'))


def _hero(stats, has_store):
    """The store, across the top, wearing its own photograph."""
    if not has_store:
        return ""
    n = stats.get('hireItems', 0)
    c = stats.get('consLines', 0)
    return ("<button class='hero' type='button' onclick=\"openGuide('store')\">"
            "<span class='hchip'>EXPLORE</span>"
            "<span class='hb'><b>What&rsquo;s in the store</b>"
            "<span>{:,} ready to hire &nbsp;&middot;&nbsp; {:,} consumable "
            "lines</span></span>"
            "<span class='hgo'>&rsaquo;</span></button>").format(n, c)


TEMPLATE = r'''<!doctype html><!-- MY GEAR HQ · the Coates digital tool store · designed and built by Andrew Fisher --><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<link rel="icon" href="data:image/svg+xml,%3Csvg xmlns=%22http://www.w3.org/2000/svg%22 viewBox=%220 0 16 16%22%3E%3Ccircle cx=%228%22 cy=%228%22 r=%227%22 fill=%22%23F36F21%22/%3E%3C/svg%3E"><title>MY GEAR · Coates Tool Store</title><style>
:root{--org:#F26222;--org2:#C44C28;--bright:#FFA24D;--ink:#0A0E14;--panel:#151C27;--panel2:#1B2330;--line:#28323F;--tx:#EAF0F7;--mut:#8A97A8;--soft:#C3CDDA;--grn:#35D68A;--amb:#F0B429;--red:#FF5A4D;--glow:rgba(242,98,34,.5)}
*{box-sizing:border-box;margin:0;padding:0;-webkit-print-color-adjust:exact;print-color-adjust:exact}
body{background:radial-gradient(900px 500px at 80% -10%,rgba(242,98,34,.16),transparent 60%),linear-gradient(180deg,#0B1017,#0A0E14);color:var(--tx);font-family:"Segoe UI",Arial,sans-serif;min-height:100vh}
.wrap{max-width:560px;margin:0 auto;padding:20px 16px 60px}
.brand{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:8px}
.brand .logo{font-size:30px;font-weight:900;color:var(--org);letter-spacing:.5px;line-height:1}.brand .logo b{display:block;color:#c9cfd8;font-size:11px;font-weight:600;letter-spacing:.6px;margin-top:3px}
.brand .siteiq{text-align:right;color:var(--org);font-weight:800;font-size:11px;letter-spacing:1px;margin-top:6px}
.hero{background:linear-gradient(160deg,#18202C,#0E141D);border:1px solid var(--line);border-top:5px solid var(--org);border-radius:16px;padding:22px;text-align:center;margin-bottom:16px}
.hero h1{font-size:23px;color:#fff;margin-bottom:4px}
.hero p{color:var(--soft);font-size:13.5px;line-height:1.5}
.idbox{margin-top:16px}
input{font-size:24px;padding:14px;width:100%;text-align:center;background:var(--panel);border:2px solid var(--line);border-radius:12px;color:#fff;font-family:inherit;letter-spacing:2px}
input:focus{outline:none;border-color:var(--org);box-shadow:0 0 0 4px rgba(242,98,34,.16)}
.btn{margin-top:12px;width:100%;background:linear-gradient(135deg,var(--org),var(--org2));color:#fff;border:none;border-radius:12px;padding:15px;font-size:16px;font-weight:800;cursor:pointer;font-family:inherit}
.btn:active{transform:translateY(1px)}
.err{color:var(--amb);font-size:13px;margin-top:10px;min-height:18px}
.card{display:none}
/* ---- THE ROLLER DOOR (Andrew's own pack, 2 Aug 2026) ------------
   His motion, lifted as designed: the latch recoil, the motor take-up,
   the two load pauses, the slow creep into the housing and the end-stop.
   The camera pushes into the store as the shutter clears, a shadow and
   an orange light leak ride the bottom rail, and a fixed steel head box
   swallows the door.

   IT PLAYS EVERY TIME the store is opened - Andrew's call. No latch, no
   memory. The escape hatches are there on every single run instead: SKIP
   OPENING bottom right, Escape on a keyboard, and a phone set to reduced
   motion goes straight in without ever seeing it.
------------------------------------------------------------------ */
.dw{position:fixed;inset:0;z-index:80;display:none;background:#060a10;
 overflow:hidden}
.dw.on{display:block}
.dw-stage,.dw-int,.dw-rig,.dw-shut,.dw-vig{position:absolute;inset:0;
 width:100%;height:100%}
.dw-stage{z-index:1;overflow:hidden;background:#060a10}
.dw-int{object-fit:cover;transform:scale(1.015);transform-origin:50% 60%;
 filter:brightness(.58) saturate(.82) contrast(1.08)}
.dw-light{position:absolute;inset:44% -20% -12%;opacity:0;mix-blend-mode:screen;
 background:radial-gradient(ellipse at 50% 100%,rgba(255,104,30,.34),transparent 63%)}
.dw-rig{z-index:5;will-change:transform;transform:translate3d(0,0,0)}
.dw-shut{display:block;object-fit:cover;filter:contrast(1.06) brightness(.9)}
.dw-shadow{position:absolute;left:-6%;right:-6%;bottom:-2px;height:8%;
 transform:translateY(48%);filter:blur(5px);
 background:linear-gradient(to bottom,rgba(0,0,0,.08),rgba(0,0,0,.96));
 box-shadow:0 10px 26px rgba(0,0,0,.85)}
.dw-leak{position:absolute;left:3%;right:3%;bottom:-3px;height:18px;opacity:0;
 mix-blend-mode:screen;background:linear-gradient(to bottom,
 rgba(255,118,44,.92) 0 2px,rgba(255,83,18,.28) 3px,transparent 100%)}
.dw-thresh{position:absolute;z-index:3;left:4%;right:4%;bottom:-2%;height:28%;
 opacity:0;mix-blend-mode:screen;background:radial-gradient(ellipse at 50% 100%,
 rgba(255,89,20,.62),rgba(255,89,20,.14) 34%,transparent 70%)}
.dw-head{position:absolute;z-index:6;left:0;right:0;top:0;height:54px;
 background:linear-gradient(180deg,#161d28,#0b1119 62%,#070c13);
 box-shadow:0 10px 26px rgba(0,0,0,.72),inset 0 -2px 0 rgba(255,255,255,.05)}
.dw-vig{z-index:7;pointer-events:none;box-shadow:inset 0 0 120px rgba(0,0,0,.75)}
/* the skip is now the everyday exit, not a rare one - big enough for a
   gloved thumb and clear of the notch on a phone */
.dw-skip{position:absolute;z-index:9;right:14px;background:rgba(8,13,20,.78);
 top:14px;top:calc(14px + env(safe-area-inset-top));
 border:1px solid #2A3547;color:#DCE3EC;font:inherit;font-size:12px;font-weight:800;
 letter-spacing:1.4px;border-radius:999px;padding:12px 18px;cursor:pointer;
 -webkit-backdrop-filter:blur(6px);backdrop-filter:blur(6px)}
.dw-cap{position:absolute;z-index:8;left:16px;right:16px;bottom:34px;
 opacity:0;animation:dwcap .7s 1.1s ease both}
.dw-cap span{display:block;font-size:10px;font-weight:800;letter-spacing:3px;
 color:#F26222}
.dw-cap b{display:block;font-size:24px;font-weight:800;color:#F5F7FB;margin-top:5px;
 text-shadow:0 4px 18px rgba(0,0,0,.8)}
@keyframes dwcap{from{opacity:0;transform:translateY(12px)}to{opacity:1;transform:none}}
.dw.run .dw-rig{animation:dw-lift 6.8s both}
.dw.run .dw-int{animation:dw-cam 6.8s cubic-bezier(.2,.72,.16,1) both}
.dw.run .dw-light{animation:dw-wake 6.8s ease-out both}
.dw.run .dw-shadow{animation:dw-shadow 6.8s ease-out both}
.dw.run .dw-leak{animation:dw-leak 6.8s ease-out both}
.dw.run .dw-thresh{animation:dw-thresh 6.8s ease-out both}
@keyframes dw-lift{
 0%{transform:translate3d(0,0,0);animation-timing-function:cubic-bezier(.42,0,.76,.44)}
 2.8%{transform:translate3d(0,.35%,0)}
 5.5%{transform:translate3d(0,-.8%,0);animation-timing-function:cubic-bezier(.2,.5,.22,1)}
 9%{transform:translate3d(0,-.35%,0)}
 14%{transform:translate3d(0,-4%,0)}
 27%{transform:translate3d(0,-21%,0)}
 32%{transform:translate3d(0,-22%,0)}
 47%{transform:translate3d(0,-43%,0)}
 52%{transform:translate3d(0,-44%,0)}
 68%{transform:translate3d(0,-69%,0)}
 73%{transform:translate3d(0,-70%,0)}
 89%{transform:translate3d(0,-94%,0);animation-timing-function:cubic-bezier(.12,.8,.24,1)}
 96%{transform:translate3d(0,-102%,0)}
 100%{transform:translate3d(0,-104.5%,0)}}
@keyframes dw-cam{
 0%,10%{transform:scale(1.015);filter:brightness(.58) saturate(.82) contrast(1.08)}
 34%{transform:scale(1.026) translateY(-.15%);filter:brightness(.7) saturate(.9) contrast(1.07)}
 72%{transform:scale(1.048) translateY(-.48%);filter:brightness(.88) saturate(1) contrast(1.04)}
 100%{transform:scale(1.065) translateY(-.8%);filter:brightness(.96) saturate(1.04) contrast(1.02)}}
@keyframes dw-wake{0%,18%{opacity:0}62%{opacity:.25}100%{opacity:.42}}
@keyframes dw-shadow{0%,8%{opacity:1;transform:translateY(48%) scaleY(1)}
 60%{opacity:.58;transform:translateY(44%) scaleY(.72)}
 100%{opacity:.24;transform:translateY(40%) scaleY(.48)}}
@keyframes dw-leak{0%,7%{opacity:0}15%{opacity:.46}72%{opacity:.78}100%{opacity:.38}}
@keyframes dw-thresh{0%,8%{opacity:0;transform:scaleX(.72)}
 24%{opacity:.2;transform:scaleX(.84)}62%,100%{opacity:.5;transform:scaleX(1)}}
@media (prefers-reduced-motion: reduce){.dw{display:none!important}}

/* ---- THE PERSONAL GEAR BAY (Andrew's own pack, 2 Aug 2026) -------
   His second door, and the thinking behind it is his: the stores team
   comes through the big roller shutter because they are walking into
   the WHOLE store. A bloke's ID only unlocks HIS OWN gear, so he comes
   through a smaller secured mesh gate. Two different doors for two
   different rights, which is exactly right.

   The motion is his: a short latch recoil, the two panels sliding
   apart under weight with a load pause at 58-72%, the bay lighting
   waking front to back, warm floor light spreading as the gap widens,
   and a personal docket rising with his name and counts before the
   report takes the screen.

   Every panel here is CSS. The only photograph is his bay image, so
   the whole thing costs one 107 KB WebP.

   IT PLAYS EVERY TIME, same call he made on the roller door. SKIP
   ENTRY and Escape are on screen for every run, and a phone set to
   reduce motion never sees it.
------------------------------------------------------------------ */
.wg{position:fixed;inset:0;z-index:82;display:none;background:#050a10;
 overflow:hidden}
.wg.on{display:block}
.wg-stage,.wg-bay,.wg-wake,.wg-floor,.wg-vig{position:absolute;inset:0;
 width:100%;height:100%}
.wg-stage{z-index:1;overflow:hidden;background:#050a10}
.wg-bay{object-fit:cover;object-position:center;transform:scale(1.018);
 filter:brightness(.43) saturate(.76) contrast(1.12)}
.wg.run .wg-bay{animation:wg-cam 3.52s cubic-bezier(.18,.78,.18,1) both}
.wg-wake{opacity:0;mix-blend-mode:screen;background:
 linear-gradient(180deg,rgba(255,126,50,.12),transparent 25%),
 radial-gradient(ellipse at 50% 44%,rgba(255,111,33,.22),transparent 53%)}
.wg.run .wg-wake{animation:wg-wake 3.3s ease-out both}
.wg-floor{inset:56% -12% -8%;width:auto;height:auto;opacity:0;
 transform:scaleX(.55);mix-blend-mode:screen;
 background:radial-gradient(ellipse at 50% 100%,rgba(255,89,22,.65),
  rgba(255,89,22,.13) 40%,transparent 72%)}
.wg.run .wg-floor{animation:wg-floor 3.4s ease-out both}
/* the two mesh panels - steel frame, expanded mesh, centre brace */
.wg-panel{position:absolute;z-index:7;top:0;bottom:0;width:51.5%;
 overflow:hidden;will-change:transform;background:linear-gradient(90deg,
  rgba(6,9,13,.9),rgba(18,25,34,.82) 9%,rgba(8,12,17,.68) 94%,
  rgba(1,3,5,.92));
 box-shadow:inset 0 0 0 7px #171F2A,inset 0 0 0 9px rgba(102,119,139,.38),
  0 0 24px rgba(0,0,0,.72)}
.wg-left{left:0}
.wg-right{right:0;transform:scaleX(-1)}
.wg-mesh{position:absolute;inset:8px;opacity:.82;
 background-color:rgba(6,10,15,.74);background-size:20px 14px;
 background-image:
  linear-gradient(54deg,transparent 43%,rgba(101,115,132,.78) 44%,
   rgba(31,40,51,.94) 49%,rgba(120,133,149,.68) 53%,transparent 54%),
  linear-gradient(-54deg,transparent 43%,rgba(96,111,128,.68) 44%,
   rgba(27,36,47,.94) 49%,rgba(111,126,143,.66) 53%,transparent 54%);
 box-shadow:inset 0 0 55px rgba(0,0,0,.78)}
.wg-brace{position:absolute;top:10%;bottom:10%;left:47%;width:7px;
 border-radius:4px;box-shadow:0 0 12px rgba(0,0,0,.72);
 background:linear-gradient(90deg,#080B0F,#657287 35%,#1B2430 63%,#05070A)}
.wg-seam{position:absolute;z-index:8;top:0;bottom:0;left:calc(50% - 3px);
 width:6px;box-shadow:0 0 16px #000;
 background:linear-gradient(90deg,#080B10,#758195 47%,#101722 55%,#020304)}
.wg-shadow{position:absolute;z-index:6;inset:0;pointer-events:none;
 opacity:.86;background:linear-gradient(90deg,rgba(0,0,0,.65),transparent 22%,
  transparent 78%,rgba(0,0,0,.65))}
.wg.on .wg-panel{animation:wg-recoil .48s cubic-bezier(.2,.8,.2,1) both}
.wg.on .wg-right{animation-name:wg-recoil-r}
.wg.run .wg-left{animation:wg-open-l 3.52s cubic-bezier(.32,.05,.18,1) both}
.wg.run .wg-right{animation:wg-open-r 3.52s cubic-bezier(.32,.05,.18,1) both}
.wg.run .wg-seam{animation:wg-seam .34s ease-out both}
.wg.run .wg-shadow{animation:wg-shadowout 3.1s ease-out both}
/* The sign over the bay, in the door's materials. It sits BELOW the
   two buttons on purpose - at 3.2% the SKIP pill landed on top of it
   and the whole corner read as clutter. */
.wg-sign{position:absolute;z-index:11;left:7%;right:7%;display:grid;
 top:70px;top:calc(70px + env(safe-area-inset-top));
 grid-template-columns:auto 1fr;align-items:end;gap:4px 12px;
 padding:11px 13px 10px;border:1px solid #263143;border-radius:12px;
 background:linear-gradient(160deg,#121A27,#0C121C);
 box-shadow:0 10px 22px rgba(0,0,0,.54)}
.wg-sign span{grid-row:1/3;align-self:center;color:#F26222;font-size:10px;
 font-weight:800;letter-spacing:3px;writing-mode:vertical-rl;
 transform:rotate(180deg)}
.wg-sign strong{font-size:14px;font-weight:800;color:#F5F7FB;letter-spacing:1.2px}
.wg-sign i{display:block;width:100%;height:2px;
 background:linear-gradient(90deg,#F26222,transparent)}
/* the name plate - the beat that says WE FOUND YOU before the gate moves */
.wg-plate{position:absolute;z-index:12;left:8%;right:8%;bottom:7%;
 display:flex;align-items:flex-start;gap:11px;padding:14px 15px;
 border:1px solid #263143;border-radius:18px;
 background:linear-gradient(160deg,#121A27,#0C121C);
 box-shadow:0 18px 34px rgba(0,0,0,.58);
 opacity:0;transform:translateY(14px);
 transition:opacity .3s ease,transform .3s ease}
.wg.on .wg-plate{opacity:1;transform:translateY(0)}
.wg.run .wg-plate{opacity:0;transform:translateY(18px)}
.wg-lamp{width:12px;height:12px;flex:none;margin-top:4px;border-radius:50%;
 border:1px solid #65E8A8;background:#27B974;
 box-shadow:0 0 0 4px rgba(53,199,132,.1),0 0 15px rgba(53,199,132,.82)}
.wg-plate small{display:block;color:#8794A6;font-size:9.5px;font-weight:800;
 letter-spacing:2.4px}
.wg-plate strong{display:block;margin-top:4px;font-size:17px;font-weight:800;
 color:#F5F7FB;letter-spacing:-.2px}
.wg-plate em{display:block;margin-top:3px;font-style:normal;color:#8794A6;
 font-size:11.5px;letter-spacing:.3px}
/* the docket - his counts, in the report's own tile materials */
.wg-docket{position:absolute;z-index:13;left:8%;right:8%;bottom:7%;
 padding:15px 16px;pointer-events:none;opacity:0;
 transform:translateY(70px) scale(.9);
 border:1px solid rgba(242,98,34,.55);border-radius:18px;
 background:linear-gradient(160deg,#121A27,#0C121C);
 box-shadow:0 25px 52px rgba(0,0,0,.64),0 0 30px rgba(242,98,34,.13)}
.wg.run .wg-docket{animation:wg-docket 1.15s 2.15s cubic-bezier(.16,1,.3,1) both}
.wg-docket small{display:block;color:#F26222;font-size:9.5px;font-weight:800;
 letter-spacing:2.4px}
.wg-docket b.who{display:block;margin-top:5px;font-size:16px;font-weight:800;
 color:#F5F7FB;letter-spacing:-.2px}
.wg-dnums{display:grid;grid-template-columns:repeat(3,1fr);gap:6px;
 margin-top:12px}
.wg-dnums span{padding:9px 5px;color:#8794A6;font-size:8.5px;font-weight:800;
 letter-spacing:.6px;text-align:center;border:1px solid #263143;
 border-radius:11px;background:#0B111A}
.wg-dnums b{display:block;margin-bottom:3px;color:#F5F7FB;font-size:17px}
.wg-dnums .a b{color:#FFB020}.wg-dnums .r b{color:#FF5A4D}
.wg-dnums .g b{color:#35D68A}
.wg-vig{z-index:14;pointer-events:none;background:radial-gradient(ellipse at 50% 46%,
 transparent 42%,rgba(0,0,0,.55) 100%)}
.wg-btns{position:absolute;z-index:16;right:14px;display:flex;gap:8px;
 top:14px;top:calc(14px + env(safe-area-inset-top))}
.wg-skip,.wg-snd{background:rgba(8,13,20,.78);border:1px solid #2A3547;
 color:#DCE3EC;font:inherit;font-size:12px;font-weight:800;letter-spacing:1.4px;
 border-radius:999px;padding:12px 18px;cursor:pointer;
 -webkit-backdrop-filter:blur(6px);backdrop-filter:blur(6px)}
.wg-snd{padding:12px 14px;letter-spacing:1px}
.wg-snd.off{color:#6C7A8C}
@keyframes wg-recoil{0%{transform:translateX(0)}55%{transform:translateX(-.7%)}
 100%{transform:translateX(0)}}
@keyframes wg-recoil-r{0%{transform:scaleX(-1) translateX(0)}
 55%{transform:scaleX(-1) translateX(-.7%)}
 100%{transform:scaleX(-1) translateX(0)}}
@keyframes wg-open-l{0%{transform:translateX(0)}8%{transform:translateX(-1.2%)}
 16%{transform:translateX(-.6%)}58%{transform:translateX(-70%)}
 72%{transform:translateX(-72%)}100%{transform:translateX(-107%)}}
@keyframes wg-open-r{0%{transform:scaleX(-1) translateX(0)}
 8%{transform:scaleX(-1) translateX(-1.2%)}
 16%{transform:scaleX(-1) translateX(-.6%)}
 58%{transform:scaleX(-1) translateX(-70%)}
 72%{transform:scaleX(-1) translateX(-72%)}
 100%{transform:scaleX(-1) translateX(-107%)}}
@keyframes wg-seam{from{opacity:1}to{opacity:0}}
@keyframes wg-shadowout{from{opacity:.86}to{opacity:.06}}
@keyframes wg-cam{0%,8%{transform:scale(1.018);
  filter:brightness(.43) saturate(.76) contrast(1.12)}
 52%{transform:scale(1.045) translateY(-.3%);
  filter:brightness(.72) saturate(.92) contrast(1.08)}
 100%{transform:scale(1.07) translateY(-.7%);
  filter:brightness(.9) saturate(1.02) contrast(1.04)}}
@keyframes wg-wake{0%,12%{opacity:0}58%{opacity:.24}100%{opacity:.42}}
@keyframes wg-floor{0%,12%{opacity:0;transform:scaleX(.55)}
 64%{opacity:.34}100%{opacity:.46;transform:scaleX(1)}}
@keyframes wg-docket{from{opacity:0;transform:translateY(70px) scale(.9)}
 to{opacity:1;transform:translateY(0) scale(1)}}
@media (prefers-reduced-motion: reduce){.wg{display:none!important}}

/* ---- THE REPORT, IN THE DOOR'S MATERIALS (2 Aug 2026) ------------
   Andrew: "elite style matching and designing like we have done with
   the main page. any icons added. must be generated in such a way it
   matches our style on the main page."
   So: the same panel gradient, the same #263143 borders, the same 18px
   corners, and the orange bar down the LEFT the way the door wears it -
   the card used to wear it across the top, which read as a different
   product. Icons are drawn the same way as the door's tiles: a stroke
   glyph in a rounded orange-tinted square.
------------------------------------------------------------------ */
.pcard{position:relative;background:linear-gradient(160deg,#121A27,#0C121C);
 border:1px solid #263143;border-radius:18px;padding:18px 16px 16px;
 margin-bottom:14px;overflow:hidden}
.pcard:before{content:"";position:absolute;left:0;top:16px;bottom:16px;
 width:4px;background:#F26222;border-radius:0 4px 4px 0}
.ph{display:flex;justify-content:space-between;align-items:flex-start;gap:12px}
.ph .nm{font-size:21px;font-weight:800;color:#F5F7FB;letter-spacing:-.2px}
.ph .co{color:#8794A6;font-size:11.5px;margin-top:3px;letter-spacing:.3px}
.ph .idb{background:#0B111A;border:1px solid #2A3547;border-radius:10px;
 padding:6px 11px;font-weight:800;color:#F26222;font-size:11.5px;
 letter-spacing:1px;white-space:nowrap}
.rate{display:flex;align-items:center;gap:12px;background:rgba(242,98,34,.09);border:1px solid rgba(242,98,34,.4);border-radius:12px;padding:12px 15px;margin:15px 0}
.rate .stars{font-size:26px;letter-spacing:3px;color:var(--org);line-height:1}
.rate .rl b{color:#fff;font-size:15px}.rate .rl div{color:var(--soft);font-size:11.5px;margin-top:2px}
.stats{display:grid;grid-template-columns:repeat(2,1fr);gap:8px;margin-bottom:6px}
.st{display:flex;align-items:center;gap:9px;background:linear-gradient(160deg,#121A27,#0C121C);
 border:1px solid #263143;border-radius:14px;padding:10px;text-align:left;min-width:0}
.st .sti{width:34px;height:34px;border-radius:11px;flex:none;display:flex;
 align-items:center;justify-content:center;background:rgba(242,98,34,.14);
 overflow:hidden}
.st .sti svg{width:18px;height:18px;color:#F26222}
.st.good .sti{background:rgba(53,214,138,.14)}
.st.good .sti svg{color:#35D68A}
/* Andrew's photographic tiles: the artwork IS the tile, so the orange
   wash comes off and a hairline keeps it seated in the card. They pop
   in on a short stagger - enough to feel built, not enough to wait for. */
.st .sti.art{background:#0B111A;border:1px solid #2A3547;width:38px;height:38px;
 border-radius:12px}
.st .sti.art img{width:100%;height:100%;object-fit:cover;display:block;
 animation:icpop .42s cubic-bezier(.2,.9,.3,1.4) both}
.st:nth-child(2) .sti.art img{animation-delay:.06s}
.st:nth-child(3) .sti.art img{animation-delay:.12s}
.st:nth-child(4) .sti.art img{animation-delay:.18s}
@keyframes icpop{from{opacity:0;transform:scale(.62)}to{opacity:1;transform:none}}
.badge .bimg{width:22px;height:22px;border-radius:7px;object-fit:cover;
 flex:none;margin-right:1px}
.clr .clrimg{width:52px;height:52px;border-radius:14px;object-fit:cover;
 flex:none;border:1px solid #2A3547;animation:icpop .5s cubic-bezier(.2,.9,.3,1.4) both}
.clr .ci{width:38px;height:38px;font-size:16px}
@media (prefers-reduced-motion: reduce){
 .st .sti.art img,.clr .clrimg{animation:none}}
.st .sv{min-width:0}
.st .v{font-size:20px;font-weight:850;color:#F5F7FB;line-height:1.1}
.st .l{font-size:8.5px;color:#8794A6;text-transform:uppercase;letter-spacing:.6px;
 margin-top:2px;font-weight:700;line-height:1.25}
.story{background:linear-gradient(160deg,#121A27,#0C121C);border:1px solid #263143;border-left:3px solid #F26222;border-radius:0 14px 14px 0;padding:13px 15px;margin:14px 0;font-size:13px;line-height:1.6;color:#DCE3EC}
.story b{color:var(--bright)}
h3.sec{font-size:10px;color:#6B7789;text-transform:uppercase;letter-spacing:2.4px;font-weight:800;margin:18px 0 9px}
/* ---- THE INSTRUMENT PANEL (Andrew's pack, 2 Aug 2026) -----------
   "The panel should feel like physical store equipment" - his rule,
   and it is the right one. So: a power rail that energises once when
   the report lands, every section numbered like a module on a rack,
   and the hire ages read as three lit cells instead of a thin bar.
   No blur, no fog, no particles, no looping pulses. Everything below
   settles once and then sits still, the way a real gauge does.
------------------------------------------------------------------ */
.pcard{counter-reset:mod}
h3.sec:before{counter-increment:mod;
 content:"MODULE " counter(mod,decimal-leading-zero);display:block;
 color:#F26222;font-size:9px;letter-spacing:2.4px;margin-bottom:4px}
.prail{position:relative;height:3px;border-radius:3px;margin:13px 0 0;
 background:#1A2331;overflow:hidden}
.prail:after{content:"";position:absolute;inset:0;transform:scaleX(0);
 transform-origin:left;animation:prail .9s .12s cubic-bezier(.16,1,.3,1) both;
 background:linear-gradient(90deg,#F26222,#FFA24D 55%,rgba(255,162,77,0))}
@keyframes prail{to{transform:scaleX(1)}}
.prsub{margin-top:7px;font-size:9px;font-weight:800;letter-spacing:1.8px;
 color:#6B7789;text-transform:uppercase}
/* the day number is the bit a bloke on site actually navigates by */
.prsub .fd{color:#F26222}
/* the hire age tracker - clear / watch / return */
.hage{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:9px}
.hcell{border:1px solid #263143;border-radius:14px;padding:11px 5px 9px;
 text-align:center;background:linear-gradient(160deg,#121A27,#0C121C)}
.hcell small{display:block;font-size:8px;font-weight:800;letter-spacing:1.2px;
 color:#8794A6}
.hcell b{display:block;margin:5px 0 3px;font-size:24px;font-weight:800;
 color:#F5F7FB;line-height:1}
.hcell em{display:block;font-style:normal;font-size:8px;font-weight:800;
 letter-spacing:1.6px;color:#6B7789}
/* green is the good column so it always reads green. Amber and red only
   light up when they have something in them - a zero glowing red is a
   lie about a bloke who is square with the store. */
.hcell.g{border-color:rgba(53,214,138,.45);
 background:linear-gradient(160deg,rgba(53,214,138,.13),#0C121C)}
.hcell.g em{color:#35D68A}
.hcell.a.hot{border-color:rgba(240,180,41,.5);
 background:linear-gradient(160deg,rgba(240,180,41,.14),#0C121C)}
.hcell.a.hot em{color:#F0B429}
.hcell.r.hot{border-color:rgba(255,90,77,.5);
 background:linear-gradient(160deg,rgba(255,90,77,.14),#0C121C)}
.hcell.r.hot em{color:#FF5A4D}
.hold{font-size:9px;font-weight:800;letter-spacing:1.6px;color:#8794A6;
 text-align:right;margin:-4px 0 9px}
/* the shift docket - his stamp, wrapped round the story that was
   already there rather than repeating it underneath */
.sdock{display:flex;gap:12px;align-items:flex-start;margin:14px 0;padding:13px 15px;
 border:1px solid #263143;border-left:3px solid #F26222;border-radius:0 14px 14px 0;
 background:linear-gradient(160deg,#121A27,#0C121C)}
.stamp{flex:none;width:66px;padding:9px 5px;border:2px solid #F26222;
 border-radius:10px;transform:rotate(-3deg);text-align:center;color:#F26222;
 font-size:9px;font-weight:900;letter-spacing:1.2px;line-height:1.35}
.sdbody{min-width:0;flex:1}
.sdbody .sdlab{font-size:9px;font-weight:800;letter-spacing:2.2px;color:#6B7789;
 margin-bottom:5px}
.sdtx{font-size:13px;line-height:1.6;color:#DCE3EC}
.sdtx b{color:var(--bright)}
.sdnums{display:grid;grid-template-columns:repeat(3,1fr);gap:7px;margin-top:11px}
.sdnums span{padding:9px 4px;text-align:center;border:1px solid #263143;
 border-radius:11px;background:#0B111A;font-size:8px;font-weight:800;
 letter-spacing:.9px;color:#8794A6}
.sdnums b{display:block;margin-bottom:3px;font-size:17px;color:#F5F7FB}
/* NEXT ACTION - only ever fires on a real 5+ day item */
.nact{display:flex;gap:11px;align-items:center;margin-top:10px;padding:12px 14px;
 border:1px solid rgba(255,90,77,.5);border-left:3px solid #FF5A4D;
 border-radius:0 14px 14px 0;
 background:linear-gradient(90deg,rgba(255,90,77,.12),rgba(255,90,77,.02))}
.nact .nal{font-size:9px;font-weight:800;letter-spacing:2.2px;color:#FF5A4D}
.nact .nab{display:block;margin-top:4px;font-size:14.5px;font-weight:800;
 color:#F5F7FB;line-height:1.3}
.nact .nas{display:block;margin-top:3px;font-size:11.5px;color:#8794A6}
.nact .nax{flex:none;width:34px;height:34px;border-radius:11px;display:flex;
 align-items:center;justify-content:center;font-size:19px;font-weight:800;
 color:#fff;background:#E5443A}
/* ---- THE REPORT BUILDS ITSELF AS YOU GO DOWN IT (2 Aug 2026) -----
   Each block waits below the fold, then lifts into place when it
   arrives. Twelve pixels and a fade - enough to feel alive, not enough
   to make anybody wait. A phone on reduce motion, or a browser with no
   IntersectionObserver, gets .rv-in put on everything immediately, so
   nothing can ever be left invisible behind an animation that did not
   run. THAT is the rule this whole thing is built on.
------------------------------------------------------------------ */
[data-rv]{opacity:0;transform:translateY(12px);
 transition:opacity .5s cubic-bezier(.16,1,.3,1),
            transform .5s cubic-bezier(.16,1,.3,1)}
[data-rv].rv-in{opacity:1;transform:none}
@media (prefers-reduced-motion: reduce){
 [data-rv]{opacity:1!important;transform:none!important;transition:none!important}}
/* search and filter a bloke's own gear */
.gtools{margin:2px 0 10px}
.gsrch{width:100%;box-sizing:border-box;padding:12px 14px;border-radius:14px;
 border:1px solid #263143;background:#0B111A;color:#F5F7FB;font:inherit;
 font-size:14px;outline:none;-webkit-appearance:none}
.gsrch::placeholder{color:#6B7789}
.gsrch:focus{border-color:#F26222;box-shadow:0 0 0 3px rgba(242,98,34,.13)}
/* four across, always. As a wrapping flex row RETURN dropped onto a
   line of its own at 412px and the set stopped reading as one control */
.glanes{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-top:8px}
.gl{min-width:0;display:inline-flex;align-items:center;justify-content:center;
 gap:5px;padding:9px 4px;border:1px solid #263143;border-radius:11px;
 background:#0B111A;color:#8794A6;font:800 9px/1 inherit;letter-spacing:.6px;
 cursor:pointer}
.gl b{color:#F5F7FB;font-size:11px}
.gl.on{border-color:#F26222;color:#F5F7FB;background:#1B1410}
.gl.g.on{border-color:#35D68A;background:rgba(53,214,138,.12)}
.gl.a.on{border-color:#F0B429;background:rgba(240,180,41,.12)}
.gl.r.on{border-color:#FF5A4D;background:rgba(255,90,77,.12)}
.gnone{margin-top:10px;padding:12px 14px;border:1px solid #263143;
 border-left:3px solid #F0B429;border-radius:0 12px 12px 0;background:#0B111A;
 color:#8794A6;font-size:12.5px;line-height:1.55}
/* ---- THE SECTION RAIL --------------------------------------------
   Same shape and same materials as the stores dock, on purpose: the
   crew page and the counter page should feel like one system.
------------------------------------------------------------------ */
.rnav{position:fixed;left:0;right:0;bottom:0;z-index:58;display:none;
 grid-template-columns:repeat(5,1fr);gap:2px;
 padding:6px 6px calc(6px + env(safe-area-inset-bottom));
 border-top:1px solid #263143;background:rgba(9,14,21,.94);
 -webkit-backdrop-filter:blur(10px);backdrop-filter:blur(10px)}
.rnav.on{display:grid}
.rnav button{position:relative;display:flex;flex-direction:column;
 align-items:center;gap:4px;padding:8px 2px 6px;border:0;border-radius:12px;
 background:transparent;color:#8794A6;font:800 8.5px/1 inherit;
 letter-spacing:1.1px;cursor:pointer}
.rnav button svg{width:19px;height:19px;fill:none;stroke:currentColor;
 stroke-width:1.8;stroke-linecap:round;stroke-linejoin:round}
.rnav button.on{color:#F26222}
.rnav button:active{background:#111A27}
.rnav button b{position:absolute;top:2px;right:calc(50% - 20px);min-width:15px;
 padding:1px 4px;border-radius:99px;background:#F26222;color:#fff;
 font-size:9px;line-height:1.5;text-align:center}
.rnav button b:empty{display:none}
body.hasrnav #result{padding-bottom:70px}
@media print{.rnav,.gtools{display:none!important}
 [data-rv]{opacity:1!important;transform:none!important}
 /* BIG ON THE PHONE, TIGHT ON PAPER. (Andrew, 3 Aug 2026: "i want
    workers photo to be the same as stores, and when it prints it
    stays tight.")
    The photo is 132px on screen so it matches the store and stores
    boards - the same tool the same size wherever a bloke looks at it.
    On paper that costs 32px a row across every item a worker holds,
    and this report gets printed and emailed, so it goes back to 88
    for print only. Screen size is untouched; the two never argue
    because they are two different media. */
 .item .ith{width:88px!important;height:88px!important;
  border-radius:10px!important}
 .item{padding:7px 10px!important;margin-bottom:5px!important}}
@media (min-width:900px){.rnav{max-width:560px;margin:0 auto;
 border-left:1px solid #263143;border-right:1px solid #263143;
 border-radius:14px 14px 0 0}}
.item{display:flex;justify-content:space-between;align-items:center;gap:10px;background:linear-gradient(160deg,#121A27,#0C121C);border:1px solid #263143;border-radius:14px;padding:10px 12px;margin-bottom:8px}
.item .d{font-size:13px;color:#eef2f8;font-weight:600}.item .n{font-size:11px;color:var(--mut);margin-top:1px}
/* the item's own picture - what the thing in your name LOOKS like.
   Photo when its variant shot is collected, two-letter tile until. */
/* 132px. ONE PICTURE SIZE ACROSS THE WHOLE SUITE. (Andrew,
   3 Aug 2026: "needs to be fixed in stores and workers".)
   The store board ran 132, the stores board 112 and the worker
   page 100 - three sizes for the same photo of the same tool,
   so the gear changed size as a bloke moved between screens.
   132 is the one Andrew pointed at. If this ever moves, it moves
   on all four: mygear_store, mygear_stores, BUILD_MY_GEAR and
   build_fleet_detail. */
.item .ith{flex:none;width:132px;height:132px;border-radius:14px;overflow:hidden;background:#1B2330;display:flex;align-items:center;justify-content:center}
.item .ith img{width:100%;height:100%;object-fit:cover;display:block}
.item .ith.mono{color:#8A97A8;font-weight:900;font-size:21px;letter-spacing:.5px}
.item .itxt{flex:1;min-width:0}
/* .cb = compliance chips under an item. Deliberately NOT .badge/.badges -
   those are the achievement pills above and must not change. The chips
   themselves are inline-styled so they survive print and email. */
.item .cb{margin-top:3px;line-height:1.6}
/* the orange Plant ID - the number the crews say out loud */
.item .n .pid{display:inline-block;background:#F26222;color:#fff;border-radius:9px;
  padding:1px 8px;margin-left:6px;font-weight:800;font-size:11px;letter-spacing:.3px}
.age{flex:none;min-width:52px;text-align:center;border-radius:8px;padding:5px 9px;font-weight:800;font-size:12px}
.age.g{background:rgba(53,214,138,.14);color:#8fe9b8}.age.a{background:rgba(240,180,41,.14);color:#f3d98c}.age.r{background:rgba(255,90,77,.14);color:#f6b3ab}
.clr{display:flex;align-items:center;gap:14px;border-radius:13px;padding:14px 16px;margin-top:8px}
.clr.open{background:rgba(242,98,34,.10);border:1px solid rgba(242,98,34,.45)}
.clr.done{background:rgba(53,214,138,.12);border:1px solid rgba(53,214,138,.5)}
.clr .ci{flex:none;width:46px;height:46px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:22px;font-weight:900}
.clr.open .ci{background:var(--org);color:#fff}.clr.done .ci{background:var(--grn);color:#04120a}
.clr .ct b{color:#fff;font-size:15px;display:block}.clr .ct span{color:var(--soft);font-size:12px}
.help{background:rgba(242,98,34,.08);border:1px solid rgba(242,98,34,.35);border-radius:11px;padding:12px 15px;margin-top:12px;font-size:12.8px;line-height:1.55;color:var(--soft)}
.help b{color:var(--bright)}
.ph{align-items:center}
.pavatar{flex:none;width:52px;height:52px;border-radius:14px;background:linear-gradient(135deg,var(--org),var(--org2));color:#fff;font-weight:900;font-size:20px;display:flex;align-items:center;justify-content:center;letter-spacing:1px}
.pmeta{flex:1}
.scorewrap{display:flex;align-items:center;gap:16px;background:linear-gradient(135deg,rgba(242,98,34,.15),rgba(242,98,34,.03));border:1px solid rgba(242,98,34,.4);border-radius:16px;padding:14px 16px;margin:15px 0}
/* THE RETURNS GAUGE (Andrew's pack). Same number the ring carried, read
   the way the rest of the panel reads - a dial with a needle that swings
   up once and then sits still. The needle's ANGLE is written as the SVG
   transform attribute, never as a CSS-only rotation, so a browser that
   will not animate it still parks the needle on the right number. The
   sweep is the animation; the reading is not. */
.gauge{width:104px;height:94px;flex:none}
.gtrack{stroke:#28323F}
.gtick{stroke:#3A4757;stroke-width:2;stroke-linecap:round}
.gtick.on{stroke:#6E7F94}
.gend{fill:#5A6B80;font-size:8px;font-weight:800;
 font-family:"Segoe UI",Arial,sans-serif;letter-spacing:.5px}
.gneedle{fill:#F26222}
.ghub{fill:#0B111A;stroke:#F26222;stroke-width:2}
.ghubdot{fill:#F26222}
.gnum{fill:#fff;font-size:27px;font-weight:900;
 font-family:"Segoe UI",Arial,sans-serif}
/* no returns yet = no needle and no reading. A needle parked on zero is
   a lie about a bloke who simply has not brought anything back yet. */
.gnum.none{font-size:23px;fill:#8794A6}
.scoremeta{flex:1}
.scorelab{font-weight:800;color:#fff;font-size:16px}
.rankline{color:var(--soft);font-size:12px;margin-top:3px}
.rate2{margin-top:9px;font-size:13px;color:#fff}.rate2 .stars{color:var(--org);letter-spacing:2px}.rate2 b{color:var(--bright)}
.badges{display:flex;flex-wrap:wrap;gap:7px;margin-bottom:14px}
.badge{display:inline-flex;align-items:center;gap:6px;background:linear-gradient(160deg,#121A27,#0C121C);border:1px solid #263143;border-left:3px solid #F26222;border-radius:999px;padding:6px 13px;font-size:11.5px;font-weight:700;color:#DCE3EC}
.badge .bi{font-size:14px}
.mix{display:flex;height:18px;border-radius:9px;overflow:hidden;background:var(--panel);border:1px solid var(--line)}
.mix .seg{height:100%}
.legend{display:flex;flex-wrap:wrap;gap:12px;margin-top:9px}
.legend .lg{font-size:11.5px;color:var(--soft)}
.legend .dot{display:inline-block;width:9px;height:9px;border-radius:50%;margin-right:5px;vertical-align:middle}
.confp{position:fixed;top:-14px;width:9px;height:15px;border-radius:2px;opacity:.95;z-index:99;pointer-events:none;animation:fall 2.3s linear forwards}
@keyframes fall{to{transform:translateY(110vh) rotate(540deg);opacity:0}}
.pcard,.story{animation:fup .45s ease both}
.scorewrap{animation:fup .5s ease both}
.badge{animation:bpop .45s ease both}
@keyframes bpop{0%{opacity:0;transform:scale(.55)}70%{transform:scale(1.08)}100%{opacity:1;transform:scale(1)}}
.cbar i,.mix .seg,.agingbar .ab{transform-origin:left;animation:growx .9s ease-out both}
@keyframes growx{from{transform:scaleX(0)}to{transform:scaleX(1)}}
.item{animation:fup .4s ease both}
@keyframes fup{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:none}}
.clr .ci{animation:bpop .55s ease both}
.gsweep{transition:stroke-dashoffset 1.1s cubic-bezier(.22,.75,.3,1)}
/* the needle settles the way a real one does - quick off the stop, slow
   into the number, no bounce. Motion only; see .gauge above. The pivot
   is the hub, stated in viewBox units, which is what transform-box buys. */
.gneedle{transform-box:view-box;transform-origin:60px 66px;
 transform:rotate(-90deg);
 transition:transform 1.15s cubic-bezier(.16,.9,.3,1)}
@media (prefers-reduced-motion: reduce){
 .gsweep,.gneedle{transition:none}}
.st.good{border-top-color:var(--grn)}
.cmp{display:flex;flex-direction:column;gap:9px;margin-bottom:4px}
.cmpr{display:flex;align-items:center;gap:10px}
.cl{width:80px;font-size:12px;color:var(--soft)}
.cbar{flex:1;height:13px;background:var(--panel);border:1px solid var(--line);border-radius:7px;overflow:hidden}
.cbar i{display:block;height:100%;background:linear-gradient(90deg,var(--org2),var(--bright));border-radius:7px}
.cbar i.me{background:linear-gradient(90deg,var(--org),var(--bright))}
.cvv{width:30px;text-align:right;font-weight:800;color:#fff;font-size:13px}
.agingbar{display:flex;height:14px;border-radius:8px;overflow:hidden;background:var(--panel);border:1px solid var(--line);margin-bottom:8px}
.agingbar .ab{height:100%}
.idot{flex:none;width:10px;height:10px;border-radius:50%}
.itxt{flex:1}
.actions{display:flex;gap:10px;margin-top:16px}
.actions button{flex:1;background:linear-gradient(160deg,#121A27,#0C121C);border:1px solid #263143;color:#DCE3EC;border-radius:13px;padding:13px;font-weight:700;font-size:13.5px;cursor:pointer;font-family:inherit}
.actions button.primary{background:linear-gradient(90deg,#FF681F,#E94713);color:#fff;border:none}
.ft{color:var(--mut);font-size:10.5px;text-align:center;margin-top:18px;line-height:1.7}
.ft .val{color:var(--org);font-weight:700}
.alert{border-radius:11px;padding:11px 14px;margin-top:9px;font-size:12.8px;line-height:1.5}
.alert.amb{background:rgba(240,180,41,.10);border:1px solid rgba(240,180,41,.45);color:#f3d98c}
.alert.red{background:rgba(255,90,77,.10);border:1px solid rgba(255,90,77,.5);color:#f6b3ab}
.alert b{color:#fff}
.okline{color:#8fe9b8;font-size:12px;margin-top:9px}
.subline{color:var(--mut);font-size:11.5px;margin-top:7px}
/* ---- ELITE PASS 2 (Andrew, 28 Jul 2026): neon numbers, the site
   pulse on the front door, shimmer on the title, and a contacts
   button that is never more than one thumb away. ---- */
.pulse{display:flex;flex-wrap:wrap;justify-content:center;gap:10px 18px;margin:12px 0 2px;animation:fup .6s .15s ease both}
.pulse .pu{display:flex;flex-direction:column;align-items:center;font-size:8.5px;letter-spacing:1.6px;color:var(--mut);font-weight:700;text-align:center}
.pulse .pu i{display:block;font-style:normal;font-size:8px;letter-spacing:.7px;color:#6E7A8A;margin-top:2px;text-transform:none}
.pulse .pv{font-size:22px;letter-spacing:0;color:#EFFF3D;font-weight:850;text-shadow:0 0 14px rgba(239,255,61,.4);line-height:1.15;font-variant-numeric:tabular-nums}
.pulse .pvt{font-size:13px;letter-spacing:.4px;color:#EFFF3D;font-weight:850;text-shadow:0 0 12px rgba(239,255,61,.35);line-height:1.65;max-width:170px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.mg b{background:linear-gradient(100deg,#F26222 25%,#FFB347 42%,#EFFF3D 50%,#FFB347 58%,#F26222 75%);background-size:240% 100%;-webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent;animation:mgshine 6s ease-in-out infinite}
@keyframes mgshine{0%,100%{background-position:90% 0}45%,55%{background-position:10% 0}}
/* ---- the MY GEAR masthead (the brand, 1 Aug 2026). Archivo Black
   carried inside the page, white MY, forged-orange GEAR, and the
   glow line off the rack tickets. K2 left the wordmark the same day -
   the site is an address now, not the identity. ---- */
.mgx{font-family:'Archivo Black','Arial Black',Arial,sans-serif;font-size:60px;line-height:1;display:flex;justify-content:center;align-items:baseline;gap:13px;letter-spacing:2px;margin:4px 0 0}
.mgxw{color:#f4f6f8;text-shadow:0 2px 8px rgba(0,0,0,.6)}
.mgxo{background:linear-gradient(180deg,#ffa25c 0%,#F26222 46%,#d94e12 100%);-webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent;filter:drop-shadow(0 0 14px rgba(242,98,34,.45)) drop-shadow(0 2px 6px rgba(0,0,0,.55))}
/* ---- THE FRONT DOOR (Andrew's own design, 1 Aug 2026) -----------
   His spec, followed: page gradient #070A10 -> #0B111A -> #080C13, an
   orange glow off the top, panels #121A27 -> #0C121C, the CTA running
   #FF681F -> #E94713, and every tile stating what it opens BEFORE a
   thumb goes near it. The photographs are his renders, sitting in
   art/ beside the page.
   Typing the ID is still the only way in - there is nothing to scan.
------------------------------------------------------------------ */
body{background:linear-gradient(160deg,#070A10 0%,#0B111A 55%,#080C13 100%)}
body:before{content:"";position:fixed;left:0;right:0;top:0;height:280px;
 background:radial-gradient(75% 100% at 50% 0,rgba(255,91,34,.12),transparent 70%);
 pointer-events:none;z-index:0}
.wrap{position:relative;z-index:1}

.mast{display:flex;align-items:baseline;gap:11px;margin:14px 0 0;
 font-family:'Archivo Black','Arial Black',Arial,sans-serif;font-size:44px;
 line-height:1;letter-spacing:.5px}
.mast .mw{color:#F5F7FB}
.mast .mo{background:linear-gradient(180deg,#ffa25c,#F26222 55%,#d94e12);
 -webkit-background-clip:text;background-clip:text;
 -webkit-text-fill-color:transparent;
 filter:drop-shadow(0 0 16px rgba(242,98,34,.4))}
.mast .mhq{font-size:15px;color:#F5F7FB;align-self:flex-start;margin-top:3px;
 letter-spacing:.5px}
.mast i{flex:1;height:1px;background:linear-gradient(90deg,
 rgba(242,98,34,.75),rgba(242,98,34,0));margin-left:4px;align-self:center}
.mkick{font-size:11px;font-weight:800;letter-spacing:4px;color:#8794A6;
 margin:7px 0 0}

/* the store, across the top, wearing its photograph */
.hero{display:block;position:relative;width:100%;text-align:left;margin:16px 0 0;
 border:1px solid #2A3547;border-radius:20px;overflow:hidden;cursor:pointer;
 min-height:140px;padding:16px 62px 16px 16px;color:#F5F7FB;font:inherit;
 background:#0A0E14 url('art/store.jpg') 72% center/cover no-repeat;
 transition:transform .12s,border-color .12s}
.hero:before{content:"";position:absolute;inset:0;background:linear-gradient(90deg,
 rgba(7,10,16,.97) 0%,rgba(7,10,16,.9) 34%,rgba(7,10,16,.5) 62%,
 rgba(7,10,16,.12) 88%)}
.hero:active{transform:scale(.99);border-color:#F26222}
.hero>*{position:relative}
.hchip{display:inline-block;background:#F26222;color:#fff;font-size:9.5px;
 font-weight:800;letter-spacing:2.2px;padding:5px 11px;border-radius:999px}
.hb{display:block;margin-top:12px}
.hb b{display:block;font-size:21px;font-weight:800;letter-spacing:-.2px}
.hb span{display:block;font-size:11.5px;color:#9EAABB;margin-top:5px}
.hgo{position:absolute;right:14px;top:50%;transform:translateY(-50%);
 width:34px;height:34px;border-radius:50%;border:1.5px solid #F26222;
 color:#F26222;font-size:22px;line-height:31px;text-align:center}

/* THE DOOR - typed ID, and nothing competing with it */
.door{position:relative;margin:16px 0 0;padding:16px 15px 15px;
 background:linear-gradient(160deg,#121A27,#0C121C);border:1px solid #263143;
 border-radius:18px}
.dbar{position:absolute;left:0;top:16px;bottom:16px;width:4px;
 background:#F26222;border-radius:0 4px 4px 0}
.dk{font-size:10px;font-weight:800;letter-spacing:2.2px;color:#F26222}
.dt{font-size:20px;font-weight:800;color:#F5F7FB;margin-top:6px}
.ds{font-size:12.5px;color:#8794A6;margin-top:3px}
.drow{display:flex;justify-content:space-between;align-items:baseline;
 margin:14px 0 6px}
.drow span{font-size:10px;font-weight:800;letter-spacing:2px;color:#8794A6}
.drow i{font-style:normal;font-size:10.5px;color:#6B7789}
.door .idrow{margin:0}
.door #idno{height:54px;font-size:17px;letter-spacing:1.5px;text-align:left;
 padding:0 16px;background:#0B111A;border-color:#2A3547;font-weight:700}
.door #idno::placeholder{letter-spacing:.5px;font-weight:600;color:#5B6675}
.door .scanbtn{display:none!important}
.door .btn{margin-top:12px;height:52px;border-radius:13px;font-size:15.5px;
 letter-spacing:.5px;background:linear-gradient(90deg,#FF681F,#E94713);
 position:relative;overflow:hidden}
.door .btn:after{content:"";position:absolute;top:0;bottom:0;width:45%;
 left:-60%;background:linear-gradient(90deg,transparent,rgba(255,255,255,.28),
 transparent);transform:skewX(-18deg)}
.door .btn.sweep:after{animation:sweep .55s ease}
@keyframes sweep{to{left:120%}}
.door .err{min-height:0;margin-top:8px}
.door .err:empty{display:none}
.door .subline{display:none}

/* what the box is doing, said plainly */
.dstate{display:none;margin-top:10px;font-size:12px;line-height:1.5;
 padding:9px 11px;border-radius:10px;border:1px solid #263143;background:#0B111A}
.dstate.on{display:block}
.dstate b{font-weight:800}
.dstate .sp{display:inline-block;width:12px;height:12px;margin-right:7px;
 border-radius:50%;border:2px solid #2A3547;border-top-color:#F26222;
 vertical-align:-2px;animation:spin .7s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}
.dstate.look{color:#9EAABB}
.dstate.bad{color:#FF8A7A;border-color:#4A2621;background:#180E0D}
.dstate.warn{color:#F0B429;border-color:#4a3a10;background:#191307}

/* the welcome - who you are and what you hold, before you go in */
.wel{display:none;margin:12px 0 0;padding:15px;border-radius:18px;
 border:1px solid #2C4A38;background:linear-gradient(160deg,#101C17,#0B1210)}
.wel.on{display:block}
.welh{display:flex;align-items:center;gap:12px}
.welav{width:46px;height:46px;border-radius:14px;flex:none;display:flex;
 align-items:center;justify-content:center;font-weight:900;font-size:16px;
 color:#0A0E14;background:linear-gradient(160deg,#4FE39B,#2BB673)}
.welh b{display:block;font-size:18px;font-weight:800;color:#F5F7FB}
.welh span{display:block;font-size:11.5px;color:#8794A6;margin-top:2px}
.welsum{display:flex;gap:8px;margin:13px 0 0}
.welsum div{flex:1;text-align:center;border-radius:11px;padding:9px 4px;
 background:#0B111A;border:1px solid #263143}
.welsum b{display:block;font-size:19px;font-weight:850;line-height:1.1}
.welsum span{display:block;font-size:8.5px;letter-spacing:1.2px;color:#8794A6;
 font-weight:700;margin-top:3px}
.welsum .g b{color:#35D68A}.welsum .a b{color:#F0B429}.welsum .r b{color:#FF5A4D}
.wel .btn{margin-top:12px}
.welback{display:block;width:100%;margin-top:8px;background:none;border:0;
 color:#6B7789;font:inherit;font-size:11.5px;text-decoration:underline;
 cursor:pointer;padding:4px}

/* the store in three numbers, each with its own shot */
.stats3{display:flex;gap:8px;margin:16px 0 0}
.s3{flex:1;display:flex;align-items:center;gap:8px;padding:9px 8px;
 background:linear-gradient(160deg,#121A27,#0C121C);border:1px solid #263143;
 border-radius:14px;min-width:0}
.s3 i{width:36px;height:36px;border-radius:11px;flex:none;
 background:#0B111A center/cover no-repeat;border:1px solid #2A3547}
.s3 div{min-width:0}
.s3 b{display:block;font-size:18px;font-weight:850;color:#F5F7FB;line-height:1.1}
.s3 span{display:block;font-size:8px;letter-spacing:.5px;color:#8794A6;
 font-weight:700;margin-top:2px;line-height:1.25}
.s3.grn b{color:#35D68A}
.s3.amb b{color:#F0B429}

/* today's talk */
.talk{display:flex;gap:13px;margin:16px 0 0;padding:13px;
 background:linear-gradient(160deg,#121A27,#0C121C);border:1px solid #263143;
 border-radius:18px}
.talk i{width:78px;height:78px;border-radius:14px;flex:none;
 background:#0B111A center/cover no-repeat;border:1px solid #2A3547}
.talk div{min-width:0}
.talk span{display:block;font-size:9.5px;font-weight:800;letter-spacing:2px;
 color:#F26222}
.talk b{display:block;font-size:16px;font-weight:800;color:#F5F7FB;margin-top:4px}
.talk p{font-size:11.5px;color:#8794A6;line-height:1.5;margin-top:4px}

/* the four cards */
.navh{font-size:10px;font-weight:800;letter-spacing:2.4px;color:#6B7789;
 margin:20px 0 9px}
.navg{display:grid;grid-template-columns:1fr 1fr;gap:9px}
.navt{position:relative;display:block;text-align:left;border:1px solid #2A3547;
 border-radius:18px;overflow:hidden;cursor:pointer;color:#F5F7FB;font:inherit;
 padding:0;aspect-ratio:1/1.06;background:#0A0E14 center/cover no-repeat;
 transition:transform .12s,border-color .12s}
.navt:active{transform:scale(.985);border-color:#F26222}
.navt:before{content:"";position:absolute;inset:0;background:linear-gradient(
 180deg,rgba(7,10,16,.55) 0%,rgba(7,10,16,.05) 34%,rgba(7,10,16,.92) 78%)}
.ntchip{position:absolute;top:10px;left:10px;background:#F26222;color:#fff;
 font-size:8px;font-weight:800;letter-spacing:1.6px;padding:4px 9px;
 border-radius:999px}
.ntgo{position:absolute;top:9px;right:9px;width:26px;height:26px;
 border-radius:50%;border:1.4px solid #F26222;color:#F26222;font-size:17px;
 line-height:23px;text-align:center;background:rgba(7,10,16,.5)}
.ntb{position:absolute;left:11px;right:11px;bottom:10px;
 border-left:3px solid #F26222;padding-left:9px}
.ntb b{display:block;font-size:14.5px;font-weight:800;line-height:1.2}
.ntb span{display:block;font-size:10.5px;color:#9EAABB;margin-top:2px}
.ntb em{display:block;font-style:normal;font-size:8.5px;font-weight:800;
 letter-spacing:1.1px;color:#F26222;margin-top:6px}

/* no app, no password, any phone */
.trust{display:flex;justify-content:space-between;gap:6px;margin:16px 0 0;
 padding:11px 12px;background:linear-gradient(160deg,#121A27,#0C121C);
 border:1px solid #263143;border-radius:14px}
.trust span{display:flex;align-items:center;gap:6px;font-size:11px;
 color:#9EAABB;font-weight:600}
.trust i{width:15px;height:15px;flex:none;border-radius:5px;
 background:center/11px 11px no-repeat}
.trust .tg{background-color:rgba(242,98,34,.16);background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23F26222' stroke-width='2.4' stroke-linecap='round'%3E%3Crect x='4' y='4' width='7' height='7' rx='1.6'/%3E%3Crect x='13' y='4' width='7' height='7' rx='1.6'/%3E%3Crect x='4' y='13' width='7' height='7' rx='1.6'/%3E%3Cpath d='M14 21 21 14'/%3E%3C/svg%3E")}
.trust .tk{background-color:rgba(242,98,34,.16);background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23F26222' stroke-width='2.4' stroke-linecap='round'%3E%3Crect x='4' y='10' width='16' height='11' rx='2'/%3E%3Cpath d='M8 10V7a4 4 0 0 1 7-2.6'/%3E%3C/svg%3E")}
.trust .tp{background-color:rgba(53,214,138,.16);background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%2335D68A' stroke-width='2.4' stroke-linecap='round'%3E%3Crect x='6' y='2.5' width='12' height='19' rx='2.5'/%3E%3Cpath d='M10.5 18.5h3'/%3E%3C/svg%3E")}
.fddoor .scanbtn{display:none!important}
/* THE FLOATING PILLS ARE GONE (Andrew, 2 Aug 2026: "remove the contacts
   pill from everywhere").
   They earned their place when the front door had no tiles. Then radio
   and gas became big cards on the door and only CONTACTS was left
   floating - and once the section rail landed along the bottom of the
   report, that pill was one more thing hovering over a bloke's own
   words. It is off every page now.
   NOTHING WAS LOST. The contact board is a full photo card on the front
   door, and inside a report the Site guides link opens the same board -
   contact board, radio and gas monitor. Both checked on the rig. */

/* THE CARD SITS IN THE SAME GUTTER AS THE DOOR (2 Aug 2026).
   #result has always lived outside .wrap, so the report ran full-bleed
   to the screen edge while the front door sat in a 16px gutter - the
   new orange bar down the card's left made that obvious, it was touching
   the glass. And .wrap's 60px bottom padding was left hanging above the
   card once the landing hid, which is where that dead band came from. */
#result{max-width:560px;margin:0 auto;padding:0 16px}
body.hascard .wrap{padding-bottom:0}
body.hascard #result{padding-bottom:78px}
.mgxline{width:78%;max-width:420px;height:3px;margin:10px auto 8px;border-radius:2px;background:linear-gradient(90deg,transparent,rgba(242,98,34,.9) 20%,#ffc891 50%,rgba(242,98,34,.9) 80%,transparent);box-shadow:0 0 14px 2px rgba(242,98,34,.55),0 0 40px 8px rgba(242,98,34,.25)}
/* The floating pill STACK (Andrew, 29 Jul 2026: "did you not add the
   floating pills for gas monitors and radios"). Contacts keeps the
   colour and the ring - it is the one that means "call us". Radio and
   gas sit above it as quiet dark pills: always in reach, never
   competing with the call button for attention. */
.neon{color:#EFFF3D !important;text-shadow:0 0 12px rgba(239,255,61,.45)}
.crewline{margin-top:9px;font-size:12.5px;color:var(--mut)}
.crewline b{color:#EFFF3D;text-shadow:0 0 10px rgba(239,255,61,.4)}
.lgheld{margin-top:10px;font-size:12.5px;color:var(--mut);background:var(--panel);border:1px solid var(--line);border-left:3px solid #EFFF3D;border-radius:0 9px 9px 0;padding:9px 12px;line-height:1.5}
.lgheld b{color:var(--tx)}
@media (prefers-reduced-motion: reduce){.pulse,.mg b{animation:none}}
__UICSS__
__STORECSS__
</style></head><body><div class="wrap">
<div class="brand"><div class="logo">coates<b>Equipped for anything</b></div><div class="siteiq">POWERED BY SITEIQ<br><span style="color:#8B9099;font-weight:600;letter-spacing:0">Cement Australia K2 &middot; Gladstone</span></div></div>
<div id="landing">

<div class="mast"><span class="mw">MY</span><span class="mo">GEAR</span><span class="mhq">HQ</span><i></i></div>
<div class="mkick">K2 DIGITAL TOOL STORE</div>

__HERO__

<div class="door">
<span class="dbar"></span>
<div class="dk">ACCESS MY GEAR</div>
<div class="dt">Enter your site ID</div>
<div class="ds">Use the number printed on your site card.</div>
<div class="drow"><span>SITE ID</span><i>Letters and numbers accepted</i></div>
__IDROW__
<button class="btn" id="gobtn" onclick="go()">OPEN MY GEAR HQ</button>
<div class="dstate" id="dstate"></div>
<div class="err" id="err"></div>
<div class="subline" id="scanhelp"></div>
</div>

<div id="welcome" class="wel"></div>

__STATSTRIP__

__TOPIC__

<div class="navh">GUIDES &amp; QUICK ACCESS</div>
<div class="navg">
__TILES__
</div>

<div class="trust">
<span><i class="tg"></i>No app</span>
<span><i class="tk"></i>No password</span>
<span><i class="tp"></i>Any phone</span>
</div>

<div class="ft"><b style="color:#F26222">Updated once a day, about 7:00 AM.</b> Anything taken or handed back since then shows on tomorrow's refresh.<br>Read-only SiteIQ snapshot as at __ASOF__ &middot; locked to your own ID — a wrong number shows nothing.<br><span class="val">MY GEAR HQ</span> · POWERED BY SITEIQ · Designed &amp; built by Andrew Fisher</div>
</div>
</div>
<div id="result" class="card"></div>
</div>
<div id="rnav" class="rnav">
 <button type="button" data-g="sum" onclick="rvGo('sum')"><svg viewBox="0 0 24 24"><path d="M4 19V9M9.5 19V5M15 19v-7M20.5 19v-4"/></svg>SUMMARY</button>
 <button type="button" data-g="score" onclick="rvGo('score')"><svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="8"/><path d="M12 8v4l3 2"/></svg>SCORE</button>
 <button type="button" data-g="gear" onclick="rvGo('gear')"><svg viewBox="0 0 24 24"><path d="m4 7 8-4 8 4-8 4-8-4ZM4 7v10l8 4 8-4V7M12 11v10"/></svg>GEAR<b id="rnavn"></b></button>
 <button type="button" data-g="store" onclick="rvGo('store')"><svg viewBox="0 0 24 24"><path d="M4 4v16M20 4v16M4 8h16M4 14h16M7 5.5h3M13 5.5h4"/></svg>STORE</button>
 <button type="button" data-g="help" onclick="rvGo('help')"><svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="9"/><path d="M9.5 9.5a2.5 2.5 0 1 1 3.4 2.3c-.6.3-.9.8-.9 1.4v.3M12 17h.01"/></svg>HELP</button>
</div>
<div id="doorway" class="dw" aria-hidden="true">
 <div class="dw-stage">
  <img class="dw-int" src="art/tool-store-interior.webp" alt="" onerror="dwBail()">
  <div class="dw-light"></div>
 </div>
 <div class="dw-thresh"></div>
 <div class="dw-rig">
  <img class="dw-shut" src="art/roller-shutter.webp" alt="" onerror="dwBail()">
  <div class="dw-shadow"></div>
  <div class="dw-leak"></div>
 </div>
 <div class="dw-head"></div>
 <div class="dw-vig"></div>
 <button type="button" class="dw-skip" onclick="dwSkip()">SKIP OPENING</button>
 <div class="dw-cap"><span>OPENING</span><b>What&rsquo;s in the store</b></div>
</div>
<div id="gatebay" class="wg" aria-hidden="true">
 <div class="wg-stage">
  <img class="wg-bay" src="art/personal-gear-bay.webp" alt="" onerror="wgBail()">
  <div class="wg-wake"></div>
  <div class="wg-floor"></div>
 </div>
 <div class="wg-sign"><span>MY GEAR HQ</span><strong>PERSONAL GEAR BAY</strong><i></i></div>
 <div class="wg-panel wg-left"><div class="wg-mesh"></div><div class="wg-brace"></div></div>
 <div class="wg-panel wg-right"><div class="wg-mesh"></div><div class="wg-brace"></div></div>
 <div class="wg-seam"></div>
 <div class="wg-shadow"></div>
 <div class="wg-plate" id="wg-plate">
  <span class="wg-lamp"></span>
  <div><small>PERSONAL ACCESS</small><strong id="wg-nameplate">&nbsp;</strong>
   <em id="wg-siteplate">&nbsp;</em></div>
 </div>
 <div class="wg-docket" id="wg-docket"></div>
 <div class="wg-vig"></div>
 <div class="wg-btns">
  <button type="button" class="wg-snd" id="wg-snd" onclick="wgSound()"
   title="Turn the entry sound on or off">SOUND</button>
  <button type="button" class="wg-skip" onclick="wgSkip()">SKIP ENTRY</button>
 </div>
</div>
__SHEET__
<script>//__QRJS__//
var DATA=__DATA__;
/* WHICH VARIANTS ACTUALLY HAVE A PHOTO. Stamped by the build from the
   thumbs folder itself, so a tile never has to ask the network to find
   out there is nothing there. Empty list = fall back to asking, which
   is what the page did before this existed. */
var THUMBS_LIST=__THUMBS__;   /* null = could not tell; [] = none */
var THUMBS=(function(){var o={},a=THUMBS_LIST||[],i;
  for(i=0;i<a.length;i++)o[a[i]]=1;return o;})();
function hasThumb(v){
  if(!v) return false;
  /* could not tell - ask, and let the onerror fallback do its job the
     way it always did. An EMPTY list is an answer, not an absence. */
  if(THUMBS_LIST===null||THUMBS_LIST===undefined) return true;
  return !!THUMBS[tsafe(v)];
}
/* the snapshot date, for the power rail's line under the header */
var ASOF="__ASOF__";
/* days from Flame Off - a date means nothing on a shutdown, a day
   number means everything (Andrew, 2 Aug 2026) */
var FDAY="__FDAY__";

/* THE CARD'S ICONS (2 Aug 2026). Same drawing rules as the front
   door's tiles - 24 viewBox, 1.8 stroke, round caps, currentColor - so
   the two screens look like they were drawn by the same hand. */
var STI={
 out:"<path d='M4 8h16v11a1 1 0 0 1-1 1H5a1 1 0 0 1-1-1z'/>"
    +"<path d='M4 8l2-4h12l2 4'/><path d='M12 18v-6M9.6 14.4 12 12l2.4 2.4'/>",
 types:"<rect x='3.5' y='3.5' width='7' height='7' rx='1.6'/>"
    +"<rect x='13.5' y='3.5' width='7' height='7' rx='1.6'/>"
    +"<rect x='3.5' y='13.5' width='7' height='7' rx='1.6'/>"
    +"<rect x='13.5' y='13.5' width='7' height='7' rx='1.6'/>",
 back:"<path d='M9 14 4 9l5-5'/><path d='M4 9h11a5 5 0 0 1 5 5v6'/>",
 bolt:"<path d='M13 2 4.8 13.6H11l-1 8.4 8.2-11.6H12z'/>"
};
/* THE ICONS (Andrew's elite pack, 2 Aug 2026). Photographic tiles out of
   art/icons/, with the old stroke glyph kept underneath as the fallback -
   if the artwork is ever missing the report still reads properly rather
   than showing a broken picture. */
function sticon(k,slug,good){
 if(!slug) return "<i class='sti"+(good?' g':'')+"'>"+stiSvg(k)+"</i>";
 return "<i class='sti art"+(good?' g':'')+"' data-k='"+k+"'>"
  +"<img src='art/icons/"+slug+".webp' alt='' loading='lazy' "
  +"onerror='stiFallback(this)'></i>";
}
function stiSvg(k){
 return "<svg viewBox='0 0 24 24' fill='none' stroke='currentColor' "
  +"stroke-width='1.8' stroke-linecap='round' stroke-linejoin='round'>"
  +(STI[k]||'')+"</svg>";
}
/* artwork missing? fall back to the stroke glyph rather than a broken
   picture - the report has to read properly on a machine that never got
   the Art folder. */
function stiFallback(img){
 var i=img.parentNode, good=i.className.indexOf(' g')>=0;
 i.className='sti'+(good?' g':'');
 i.innerHTML=stiSvg(i.getAttribute('data-k')||'');
}
function esc(s){return String(s==null?'':s).replace(/[&<>]/g,function(c){return{'&':'&amp;','<':'&lt;','>':'&gt;'}[c]})}
function ageCls(d){d=parseInt(d);if(isNaN(d))return'a';return d<=2?'g':(d<=4?'a':'r')}
/* the item's picture tile: its variant photo out of thumbs/, or a clean
   two-letter tile until the photo hunt collects it (31 Jul 2026) */
function thMono(n){
 var w=String(n||'').split(/[^A-Za-z0-9]+/).filter(function(x){return x});
 return ((w[0]||'?').charAt(0)+(w[1]||w[0]||'').charAt(0)).toUpperCase();
}
/* Windows can't save a file named SOCKET1/2DR11MM.jpg - photos for
   codes carrying \ / : * ? " < > | land with those swapped to _ ,
   so the lookup swaps the same way (matches safe_name in Python) */
function tsafe(v){return String(v).replace(/[/:*?"<>|]/g,'_')}
function wthumb(it){
 if(!it.v || !hasThumb(it.v)) return '<span class="ith mono">'+thMono(it.d)+'</span>';
 return '<span class="ith"><img src="thumbs/'+encodeURIComponent(tsafe(it.v))
  +'.jpg" loading="lazy" alt="" data-m="'+thMono(it.d)
  +(it.va&&it.va!==it.v?'" data-a="'+esc(it.va):'')
  +'" onerror="thx(this)"></span>';
}
function thx(img){
 var a=img.getAttribute('data-a');
 if(a){img.removeAttribute('data-a');
  img.src='thumbs/'+encodeURIComponent(tsafe(a))+'.jpg';return;}
 var s=img.parentNode;
 s.className='ith mono';
 s.textContent=img.getAttribute('data-m')||'?';
}
function xmur3(str){for(var i=0,h=1779033703^str.length;i<str.length;i++){h=Math.imul(h^str.charCodeAt(i),3432918353);h=h<<13|h>>>19}return function(){h=Math.imul(h^h>>>16,2246822507);h=Math.imul(h^h>>>13,3266489909);h^=h>>>16;return h>>>0}}
function mulberry32(a){return function(){a|=0;a=a+0x6D2B79F5|0;var t=Math.imul(a^a>>>15,1|a);t=t+Math.imul(t^t>>>7,61|t)^t;return((t^t>>>14)>>>0)/4294967296}}
function tag(id){return(xmur3(id+'|CoatesK2tag2026')()>>>0).toString(16)}
function dec(id,b64){var rnd=mulberry32(xmur3(id+'|CoatesK2gear2026')());var raw=atob(b64),o='';for(var i=0;i<raw.length;i++){o+=String.fromCharCode(raw.charCodeAt(i)^Math.floor(rnd()*256))}return o}
function stars(n){var s='';for(var i=0;i<5;i++)s+=(i<n?'★':'☆');return s}
function countUp(el,to,ms){var t0=null;function s(ts){if(!t0)t0=ts;var k=Math.min(1,(ts-t0)/ms);el.textContent=Math.round(k*to);if(k<1)requestAnimationFrame(s);else el.textContent=to}requestAnimationFrame(s)}
/* ------------------------------------------------------------------
   ANIMATE WHAT YOU ARE ACTUALLY LOOKING AT (2 Aug 2026).

   This used to fire every counter on the page the moment the card was
   drawn. On a phone that meant the four tiles up top animated where a
   bloke could see them and EVERYTHING ELSE - the scorecard, the
   comparison bars, the clearance count - finished counting while it
   was still two screens below him. He scrolled down to numbers that
   had already stopped moving. All that work, seen by nobody.

   Now each number waits until it is on screen and then counts. Same
   with the score ring and every section of the report. Scroll the card
   and it builds itself in front of you the whole way down.

   Everything degrades safely: no IntersectionObserver (an old tablet)
   or a phone set to reduce motion, and the whole report is simply
   there, finished, instantly. Nothing is ever hidden behind an
   animation that did not run.
------------------------------------------------------------------ */
function rvReduced(){
  try{ return !!(window.matchMedia &&
    matchMedia('(prefers-reduced-motion: reduce)').matches); }catch(e){ return false; }
}
/* THE GAUGE'S READING. Everything that makes the dial say a number goes
   through here - the lit arc and the needle both - so there is exactly
   one way for the reading to arrive, whether it swings in, is forced by
   the backstop, or is written straight out on a reduced-motion phone.
   data-rot carries the pivot as well as the angle, so the geometry is
   stated once, in the markup that owns it. */
/* Chrome will NOT transition an SVG transform ATTRIBUTE - set it and the
   needle teleports. The swing only happens through the CSS transform
   property, and that needs transform-box:view-box to pivot on the hub
   instead of the needle's own middle. So: ask the browser whether it
   understands transform-box. If it does, drive the CSS property and the
   needle swings. If it does not, write the attribute and the needle
   simply appears on the number. Both paths read the same; only one of
   them moves. Caught on the rig - the first cut animated nothing. */
var G_CSS=(function(){
  try{ return !!(window.CSS && CSS.supports
    && CSS.supports('transform-box','view-box')); }catch(e){ return false; }
})();
function gSet(el){
  if(!el || el.getAttribute('data-set')) return;
  el.setAttribute('data-set','1');
  var off=el.getAttribute('data-off'), rot=el.getAttribute('data-rot');
  if(off!==null) el.style.strokeDashoffset=off;
  if(rot!==null){
    el.setAttribute('transform','rotate('+rot+')');
    if(G_CSS) el.style.transform='rotate('+parseFloat(rot)+'deg)';
  }
}
function gSetAll(root){
  var e=(root||document).querySelectorAll('.gsweep,.gneedle'),i;
  for(i=0;i<e.length;i++) gSet(e[i]);
}
function rvShowAll(root){
  var i,e;
  e=(root||document).querySelectorAll('[data-rv]');
  for(i=0;i<e.length;i++) e[i].className+=' rv-in';
  e=(root||document).querySelectorAll('[data-to]');
  for(i=0;i<e.length;i++) e[i].textContent=parseInt(e[i].getAttribute('data-to'))||0;
  gSetAll(root);
}
function animate(root){
  root=root||document;
  if(!('IntersectionObserver' in window) || rvReduced()){ rvShowAll(root); return; }
  /* the numbers */
  var nio=new IntersectionObserver(function(en){
    for(var i=0;i<en.length;i++){
      if(!en[i].isIntersecting) continue;
      var el=en[i].target; nio.unobserve(el);
      countUp(el, parseInt(el.getAttribute('data-to'))||0, 900);
    }
  },{rootMargin:'0px 0px -12% 0px',threshold:.01});
  var ns=root.querySelectorAll('[data-to]');
  for(var i=0;i<ns.length;i++) nio.observe(ns[i]);
  /* the sections */
  var rio=new IntersectionObserver(function(en){
    for(var j=0;j<en.length;j++){
      if(!en[j].isIntersecting) continue;
      rio.unobserve(en[j].target);
      en[j].target.className+=' rv-in';
    }
  /* rootMargin 0 on the bottom, deliberately. Shrinking it by 8% read
     nicely in the middle of the report and then stranded the LAST block
     on the page: once you have scrolled as far as the page goes, an
     element sitting inside that shrunken strip can never enter it, so
     the help panel stayed invisible forever. Caught on the rig. */
  },{rootMargin:'0px',threshold:.01});
  var rs=root.querySelectorAll('[data-rv]');
  for(var k=0;k<rs.length;k++) rio.observe(rs[k]);
  /* the gauge reads when you reach it, not before */
  var gel=root.querySelectorAll('.gsweep,.gneedle');
  var gio=null;
  if(gel.length){
    gio=new IntersectionObserver(function(en){
      for(var m=0;m<en.length;m++){
        if(!en[m].isIntersecting) continue;
        gio.unobserve(en[m].target);
        gSet(en[m].target);
      }
    },{threshold:.3});
    for(var g=0;g<gel.length;g++) gio.observe(gel[g]);
  }
  /* ---- the backstops ---------------------------------------------
     A throttled webview can hand out an IntersectionObserver that
     never calls back. That is not "it did not animate" - the card sits
     there with every tile reading 0 and the needle on the stop, which
     is the WRONG NUMBER on a bloke's own report. Found on the rig with
     a dead observer: the needle said 92 and the score under it said 0.

     Two backstops, and neither of them spoils the scroll:
       1. after 2.5s, fix anything that is ON SCREEN and still unwritten.
          If he can see it and it is wrong, it gets corrected. Blocks
          below the fold are left alone so they still build as he
          scrolls to them - that is the whole point of the reveal.
       2. once the page is scrolled to the bottom, fix the lot.
     ------------------------------------------------------------------ */
  var numFix=function(el){
    var want=parseInt(el.getAttribute('data-to'))||0;
    if(el.textContent.trim()!==String(want)) el.textContent=want;
  };
  var onScreen=function(el){
    try{ var b=el.getBoundingClientRect();
      return b.bottom>0 && b.top<(window.innerHeight||0); }
    catch(e){ return true; }
  };
  var fixAll=function(vis){
    var q,left=root.querySelectorAll('[data-rv]:not(.rv-in)');
    for(q=0;q<left.length;q++)
      if(!vis||onScreen(left[q])) left[q].className+=' rv-in';
    var nn=root.querySelectorAll('[data-to]');
    for(q=0;q<nn.length;q++) if(!vis||onScreen(nn[q])) numFix(nn[q]);
    var gg=root.querySelectorAll('.gsweep,.gneedle');
    for(q=0;q<gg.length;q++) if(!vis||onScreen(gg[q])) gSet(gg[q]);
  };
  setTimeout(function(){ fixAll(true); },2500);
  /* rootMargin 0 on the bottom, deliberately - see the sections above.
     A block a bloke cannot see is worse than a block that did not
     animate, so the bottom of the page settles everything left. */
  var sweep=function(){
    if(window.innerHeight+window.pageYOffset < document.body.scrollHeight-4) return;
    fixAll(false);
    if(!root.querySelectorAll('[data-rv]:not(.rv-in)').length)
      window.removeEventListener('scroll',sweep);
  };
  window.addEventListener('scroll',sweep,{passive:true});
}
function confetti(){var cs=['#F26222','#FFA24D','#FFD27A','#ffffff'];for(var i=0;i<70;i++){var d=document.createElement('div');d.className='confp';d.style.left=(Math.random()*100)+'vw';d.style.background=cs[i%4];d.style.animationDelay=(Math.random()*0.35)+'s';document.body.appendChild(d);(function(x){setTimeout(function(){x.remove()},2700)})(d)}}
function cmpbar(l,v,me){v=Math.round(v||0);return '<div class="cmpr"><span class="cl">'+l+'</span><span class="cbar"><i style="width:'+v+'%" '+(me?'class="me"':'')+'></i></span><span class="cvv">'+v+'</span></div>'}
/* THE DOOR'S STATES (Andrew, 1 Aug 2026: "searching, ID found, invalid
   ID and connection unavailable").

   Worth knowing WHEN each one can honestly fire. The whole register
   travels inside this page, so once it has loaded a worker can walk out
   of Wi-Fi range and still read his gear - there is no connection to
   lose at that point. But there IS one before that point: if the page
   comes down half-made, DATA arrives empty and nothing will ever match.
   That is a connection failure, and it says so. */
function idState(kind,msg){
 var d=document.getElementById('dstate'); if(!d) return;
 if(!kind){ d.className='dstate'; d.innerHTML=''; return; }
 d.className='dstate on '+kind;
 d.innerHTML=(kind==='look'?'<span class="sp"></span>':'')+msg;
}
function go(){
 var b=document.getElementById('gobtn');
 if(b){ b.classList.remove('sweep'); void b.offsetWidth; b.classList.add('sweep'); }
 var id0=(document.getElementById('idno').value||'').replace(/\s+/g,'');
 if(!id0){ idState('bad','Type the ID number off your site card.'); return; }
 if(typeof DATA!=='object'||DATA===null||!Object.keys(DATA).length){
   idState('warn','<b>Connection unavailable.</b> This page did not finish '
    +'loading from the store. Step closer to the store Wi-Fi and pull down '
    +'to refresh &mdash; or ask at the window.'); return; }
 idState('look','Checking the register&hellip;');
 setTimeout(goNow,220);
}
function goNow(){
 var id=(document.getElementById('idno').value||'').replace(/\s+/g,'');
 var err=document.getElementById('err');
 if(!id){idState('bad','Type the ID number off your site card.');return}
 /* STORES OVERRIDE - the counter types its own code into the same box
    and lands on the stores board. One door for everybody: a crew member
    types his hire ID, the stores team types theirs. The code is checked
    by hash, never stored in this page, and it is handed to the stores
    page through sessionStorage rather than the address bar - a code in
    a URL ends up in history and over someone's shoulder.
    (Andrew, 29 Jul 2026: "a stores override to get into the raw") */
 if(typeof STORES_TAG==='string' && STORES_TAG &&
    STORES_TAG.split(',').indexOf(
      (xmur3(id.toUpperCase()+'|CoatesK2storestag2026')()>>>0).toString(16))>=0){
   try{ sessionStorage.setItem('k2stores', id.toUpperCase()); }catch(e){}
   window.location.href='stores.html';
   return;
 }
 /* IDs with letters exist (18479CEM) and SiteIQ stores them upper-case.
    Try the ID exactly as typed, then upper-cased, so a bloke typing
    his card in lower case still lands on his own record. */
 var blob=null, cand=[id, id.toUpperCase()];
 for(var ci=0;ci<cand.length;ci++){
   if(DATA[tag(cand[ci])]){ id=cand[ci]; blob=DATA[tag(id)]; break; }
 }
 if(!blob){idState('bad','<b>That ID is not on the register.</b> Check the '
   +'number on your site card, or ask at the window &mdash; two metres away.');
   return}
 var p;try{p=JSON.parse(dec(id,blob))}catch(e){
   idState('bad','<b>That record would not open.</b> Show this at the window.');
   return}
 idState('');
 showWelcome(p);
}

/* THE WELCOME (Andrew, 1 Aug 2026: "after a valid ID, show the person's
   name and item count before opening their gear"). A beat between typing
   a number and a wall of detail - and it is the moment a bloke spots he
   typed someone else's ID. */
function showWelcome(p){
 window.PENDING=p;
 var st=p.stats||{}, ag=p.aging||{g:0,a:0,r:0};
 var w=document.getElementById('welcome');
 w.innerHTML='<div class="welh"><div class="welav">'+esc(p.initials||'K2')
  +'</div><div><b>'+esc(p.name)+'</b><span>'+esc(p.company)+' &middot; ID '
  +esc(p.id)+'</span></div></div>'
  +'<div class="welsum">'
  +'<div><b>'+(st.items||0)+'</b><span>IN YOUR NAME</span></div>'
  /* the traffic lights only earn their space when the export carried
     hire dates - four boxes of zeros tells a bloke nothing */
  +((ag.g+ag.a+ag.r)>0
    ? '<div class="g"><b>'+ag.g+'</b><span>0&ndash;2 DAYS</span></div>'
     +'<div class="a"><b>'+ag.a+'</b><span>3&ndash;4 DAYS</span></div>'
     +'<div class="r"><b>'+ag.r+'</b><span>5+ DAYS</span></div>'
    : '<div style="flex:3"><b style="font-size:13px;color:#8794A6;'
     +'font-weight:700">How long you have had them is inside</b>'
     +'<span>OPEN TO SEE THE DAYS</span></div>')
  +'</div>'
  +'<button class="btn" onclick="openGear()">OPEN MY GEAR</button>'
  +'<button class="welback" onclick="notMe()">That is not me &mdash; try '
  +'another ID</button>';
 w.className='wel on';
 document.querySelector('.door').style.display='none';
 w.scrollIntoView({block:'center'});
}
function notMe(){
 document.getElementById('welcome').className='wel';
 var d=document.querySelector('.door'); d.style.display='';
 idState(''); var i=document.getElementById('idno'); i.value=''; i.focus();
}
/* OPEN MY GEAR. The personal gear bay goes in this one seam - if the
   gate is unavailable for any reason the report lands exactly as it
   always did, so the animation can never stand between a bloke and his
   own record. (Andrew's Personal Gear Bay pack, 2 Aug 2026.) */
function openGear(){
 if(!window.PENDING) return;
 /* the gate covers the screen, but the button keeps keyboard focus -
    a second Enter would otherwise render the card behind the gate and
    again when it finished */
 if(typeof WG_BUSY!=='undefined' && WG_BUSY) return;
 if(typeof wgCan==='function' && wgCan()){ wgPlay(window.PENDING); return; }
 renderCard(window.PENDING);
}

/* THE RETURNS GAUGE (Andrew's pack, 2 Aug 2026). The ring said the same
   number; this says it the way the rest of the instrument panel speaks -
   a dial, a lit arc, five ticks and a needle that swings up once and
   stops. Nothing about the reading is new: p.score is untouched, and the
   printed A4 and the saved picture still carry it as a plain number.
   0 is hard left, 100 is hard right, so 1.8 degrees per point.
   A bloke with no returns yet gets NO needle and a dash - a needle
   resting on zero would accuse him of a score he has not been given. */
function gauge(p){
 var CX=60, CY=66, R=46, RAD=Math.PI/180;
 var has=!!p.hasReturns, v=Math.max(0,Math.min(100,Number(p.score)||0));
 var L=Math.PI*R;                    // the lit arc is half a circumference
 var off=has?(L-L*v/100):L;
 var ticks='';
 for(var t=0;t<=100;t+=25){
  var a=(180-1.8*t)*RAD, c=Math.cos(a), s=Math.sin(a);
  /* ticks sit just inside the rim, the needle reaches two thirds - the
     way a dial you would find on the wall of a workshop is laid out.
     They overlapped at first and the whole thing read as a smudge. */
  ticks+='<line class="gtick'+((has&&t<=v)?' on':'')
   +'" x1="'+(CX+34*c).toFixed(1)+'" y1="'+(CY-34*s).toFixed(1)
   +'" x2="'+(CX+40*c).toFixed(1)+'" y2="'+(CY-40*s).toFixed(1)+'"/>';
 }
 /* the needle parks on the stop and is swung by gSet. The angle is
    written as the SVG transform ATTRIBUTE, not a CSS-only rotation, so
    a browser that will not run the transition still lands it on the
    right number - the swing is decoration, the reading is not. */
 var needle=has?('<g class="gneedle" transform="rotate(-90 '+CX+' '+CY+')"'
   +' data-rot="'+(-90+1.8*v).toFixed(2)+' '+CX+' '+CY+'">'
   +'<polygon points="'+(CX-3.1)+','+CY+' '+CX+','+(CY-30)+' '
   +(CX+3.1)+','+CY+'"/></g>'
   +'<circle class="ghub" cx="'+CX+'" cy="'+CY+'" r="5.5"/>'
   +'<circle class="ghubdot" cx="'+CX+'" cy="'+CY+'" r="1.8"/>'):'';
 var num=has
   ?'<text class="gnum" x="60" y="103" text-anchor="middle" data-to="'
     +Math.round(v)+'">0</text>'
   :'<text class="gnum none" x="60" y="103" text-anchor="middle">&#8212;</text>';
 var arc='d="M14 66A46 46 0 0 1 106 66" fill="none" stroke-width="9"'
   +' stroke-linecap="round"';
 return '<svg class="gauge" viewBox="0 0 120 108" role="img" aria-label="'
  +'Returns score '+(has?Math.round(v)+' out of 100':'not scored yet')+'">'
  +'<defs><linearGradient id="gg" x1="0" y1="0" x2="1" y2="0">'
  +'<stop offset="0" stop-color="#FFA24D"/>'
  +'<stop offset="1" stop-color="#F26222"/></linearGradient></defs>'
  +'<path class="gtrack" '+arc+'/>'
  +'<path class="gsweep" '+arc+' stroke="url(#gg)" stroke-dasharray="'
  +L.toFixed(2)+'" stroke-dashoffset="'+L.toFixed(2)+'" data-off="'
  +off.toFixed(2)+'"/>'
  +ticks
  +'<text class="gend" x="13" y="80" text-anchor="middle">0</text>'
  +'<text class="gend" x="107" y="80" text-anchor="middle">100</text>'
  +needle+num+'</svg>';
}

function renderCard(p){
 var err=document.getElementById('err');
 window.LASTP=p;   // the print sheet builds from the same decoded payload
 var st=p.stats,r=p.rating,rk=p.rank||{comp:1,compTotal:1,pct:100};
 var ring=gauge(p);
 var badges=(p.badges||[]).map(function(b,bi){
   var ic=b[2]?'<img class="bimg" src="art/icons/'+b[2]+'.webp" alt="" '
     +'loading="lazy" onerror="this.outerHTML=\'<span class=&quot;bi&quot;>'
     +b[0]+'</span>\'">':'<span class="bi">'+b[0]+'</span>';
   return '<span class="badge" style="animation-delay:'+(0.25+bi*0.13).toFixed(2)
     +'s">'+ic+esc(b[1])+'</span>'}).join('');
 var tot=(p.mix||[]).reduce(function(a,m){return a+m[1]},0)||1;
 var segs=(p.mix||[]).map(function(m){return '<div class="seg" style="width:'+(100*m[1]/tot)+'%;background:'+m[2]+'"></div>'}).join('');
 var legend=(p.mix||[]).map(function(m){return '<span class="lg"><span class="dot" style="background:'+m[2]+'"></span>'+esc(m[0])+' '+m[1]+'</span>'}).join('');
 var rankline;
 if(!p.hasReturns){rankline='Your score starts with your first return &mdash; easy points from your next drop-off'}
 else if(!(st.sameday>0)){rankline=st.returned+' returned so far &mdash; same-day returns are what lift your score'}
 else{rankline='<b class="neon">#'+rk.comp+'</b> of '+rk.compTotal+' in your crew &middot; top '+rk.pct+'% on site'}
 var html='<div class="pcard"><div class="ph"><div class="pavatar">'+esc(p.initials||'K2')+'</div><div class="pmeta"><div class="nm">'+esc(p.name)+'</div><div class="co">'+esc(p.company)+'</div></div><div class="idb">ID '+esc(p.id)+'</div></div>'
 /* the power rail energises once as the report lands - his idea, and
    it is the thing that makes the card read as equipment rather than
    a web page. The line under it says what this snapshot IS. */
 /* one line, one line only - a wrapped subtitle under a power rail
    looks like a mistake. The date only earns its place when the export
    actually carried one. */
 +'<div class="prail"></div><div class="prsub">Cement Australia K2 &middot; Gladstone'
 +((typeof ASOF==='string'&&ASOF&&ASOF!=='last refresh')?' &middot; '+esc(ASOF):'')
 +((typeof FDAY==='string'&&FDAY)?' &middot; <b class="fd">'+esc(FDAY)+'</b>':'')
 +'</div>'
 +'<div data-sec="sum"></div>'
 +'<div class="scorewrap" data-rv>'+ring+'<div class="scoremeta"><div class="scorelab">Returns Score</div><div class="rankline">'+rankline+'</div><div class="rate2"><span class="stars">'+stars(r.stars)+'</span> <b>'+esc(r.label)+'</b></div></div></div>'
 +(badges?'<div class="badges" data-rv>'+badges+'</div>':'')
 +'<div class="stats" data-rv>'
 +'<div class="st">'+sticon('out','items-out')+'<div class="sv"><div class="v'+(st.items>0?' neon':'')+'" data-to="'+st.items+'">0</div><div class="l">Items out</div></div></div>'
 +'<div class="st">'+sticon('types','item-types')+'<div class="sv"><div class="v" data-to="'+st.types+'">0</div><div class="l">Item types</div></div></div>'
 +'<div class="st good">'+sticon('back','returned',1)+'<div class="sv"><div class="v" data-to="'+st.returned+'">0</div><div class="l">Returned</div></div></div>'
 +'<div class="st good">'+sticon('bolt','same-day',1)+'<div class="sv"><div class="v'+(st.sameday>0?' neon':'')+'" data-to="'+st.sameday+'">0</div><div class="l">Same-day</div></div></div></div>'
 +(p.cmp?'<h3 class="sec" data-sec="score">How you compare</h3><div class="cmp" data-rv>'+cmpbar('You',p.cmp.you,1)+cmpbar('Your crew',p.cmp.crew,0)+cmpbar('Site avg',p.cmp.site,0)+(rk.crewPos?'<div class="crewline">'+esc(p.company)+' sits <b>#'+rk.crewPos+'</b> of '+rk.crewOf+' crews on site</div>':'')+'</div>':'')
 +(segs?'<h3 class="sec">Your kit mix</h3><div class="mix" data-rv>'+segs+'</div><div class="legend">'+legend+'</div>':'')
 /* THE SHIFT DOCKET (his module 04). The story was already written and
    already true - this frames it as a docket off the counter and puts
    the three numbers it mentions in prose where they can be read at a
    glance instead of hunted for in a sentence. */
 +'<div class="sdock" data-rv><div class="stamp">SHIFT<br>DOCKET</div><div class="sdbody">'
 +'<div class="sdlab">PERSONAL RETURN SUMMARY</div>'
 +'<div class="sdtx">'+p.story+'</div>'
 /* His docket carries three counters. Ours does NOT: store visits and
    transactions are already the store scorecard below, and returned is
    already a stat tile above. Printing them a third time inside the
    docket made the card look padded. The stamp and the summary are the
    value here - the numbers have their own homes. */
 +'</div></div>';
 /* THE HIRE AGE TRACKER (his module 03). This replaces the thin
    three-colour bar that used to sit here: same three numbers, but now
    they are readable without a legend and each one says what it wants
    a bloke to DO - clear, watch, return. */
 var ag=p.aging||{g:0,a:0,r:0};
 var oldest=null;
 (p.items||[]).forEach(function(it){var d=parseInt(it.days);
   if(!isNaN(d)&&(!oldest||d>oldest.d)) oldest={d:d,nm:it.d}});
 /* A bloke who is square with the store gets NOTHING here - no heading,
    no three empty cells. The old thin bar drew nothing so it hid this;
    the tracker's cells are big enough that three zeros would read as a
    fault. Return clearance below already tells him he is cleared. */
 if(st.items>0){
  html+='<h3 class="sec" data-sec="gear">Your gear on hire now</h3>'
  +'<div class="hage" data-rv>'
  +'<div class="hcell g"><small>0&ndash;2 DAYS</small><b>'+ag.g+'</b><em>CLEAR</em></div>'
  +'<div class="hcell a'+(ag.a>0?' hot':'')+'"><small>3&ndash;4 DAYS</small><b>'+ag.a+'</b><em>WATCH</em></div>'
  +'<div class="hcell r'+(ag.r>0?' hot':'')+'"><small>5+ DAYS</small><b>'+ag.r+'</b><em>RETURN</em></div>'
  +'</div>'
  +(oldest?'<div class="hold">OLDEST &middot; '+oldest.d+' DAY'+(oldest.d===1?'':'S')+'</div>':'');
  /* SEARCH AND FILTER YOUR OWN LIST. A bloke with five items scrolls;
     a bloke with thirty scrolls past the one he came to find. The box
     only appears once the list is long enough to be worth it - eight
     items - so a small kit stays clean, and the three lanes are the
     same CLEAR / WATCH / RETURN the tracker above already taught him.
     Type nothing and tap nothing and the list behaves exactly as it
     always did. (2 Aug 2026.) */
  if(p.items.length>=8){
   html+='<div class="gtools" data-rv>'
    +'<input class="gsrch" id="gitq" type="search" inputmode="search"'
    +' placeholder="Find something in your own gear" oninput="gfilt()"'
    +' autocomplete="off" autocapitalize="none" spellcheck="false">'
    +'<div class="glanes">'
    +'<button type="button" class="gl on" data-l="all" onclick="gfilt(\'all\')">ALL <b>'+p.items.length+'</b></button>'
    +'<button type="button" class="gl g" data-l="g" onclick="gfilt(\'g\')">CLEAR <b>'+ag.g+'</b></button>'
    +'<button type="button" class="gl a" data-l="a" onclick="gfilt(\'a\')">WATCH <b>'+ag.a+'</b></button>'
    +'<button type="button" class="gl r" data-l="r" onclick="gfilt(\'r\')">RETURN <b>'+ag.r+'</b></button>'
    +'</div><div class="gnone" id="gnone" hidden>Nothing in your gear matches that. '
    +'Try part of the name or the item number &mdash; or ask at the window.</div></div>';
  }
 }
 // it.b is the compliance chips, built and escaped in Python before it was
 // encrypted into the payload. It goes in as HTML on purpose - do NOT wrap
 // it in esc() or the crew see markup instead of the tag colour. Every
 // other field still goes through esc().
 p.items.forEach(function(it,ii){
  var _d=parseInt(it.days), _ln=isNaN(_d)?'u':(_d<=2?'g':(_d<=4?'a':'r'));
  html+='<div class="item" data-l="'+_ln+'" data-q="'
   +esc(((it.d||'')+' '+(it.n||'')+' '+(it.pid||'')).toLowerCase())+'"'
   +' style="animation-delay:'+Math.min(ii*0.05,0.65).toFixed(2)+'s">'+wthumb(it)+'<div class="itxt"><div class="d">'+esc(it.d)+'</div><div class="n"><span class="idot" style="background:'+(it.c||"#8A97A8")+';width:8px;height:8px;display:inline-block;border-radius:50%;margin-right:5px;vertical-align:0"></span>Item '+esc(it.n)+(it.pid?'<span class="pid">ID '+esc(it.pid)+'</span>':'')+'</div>'+(it.b?'<div class="cb">'+it.b+'</div>':'')+'</div><div class="age '+ageCls(it.days)+'">'+(it.days==='-'?'—':it.days+'d')+'</div></div>'});
 // The one item that's been out longest gets its own line - a story,
 // not a nag. Only shows from 3 days out, so a fresh kit stays clean.
 // From 5 days it becomes NEXT ACTION in his instrument style: red is
 // reserved for a genuine 5+ day return and nothing else, so when a
 // bloke sees red on this card it always means the same thing.
 var lgh=oldest;
 if(lgh&&lgh.d>=5){html+='<div class="nact"><div><span class="nal">NEXT ACTION</span>'
   +'<b class="nab">'+ag.r+' item'+(ag.r===1?'':'s')+' now 5+ days out</b>'
   +'<span class="nas">Oldest: '+esc(lgh.nm)+' &middot; '+lgh.d+' days. Finished with it? '
   +'Straight to the counter and it\'s off your list.</span></div>'
   +'<span class="nax">!</span></div>'}
 else if(lgh&&lgh.d>=3){html+='<div class="lgheld">Longest out: <b>'+esc(lgh.nm)+'</b> &mdash; <span class="neon" style="font-weight:800">'+lgh.d+' days</span>. Finished with it? Straight to the counter and it\'s off your list.</div>'}
 // One plain-English summary of what the whole list obliges them to do -
 // the chips say it per item, this says it once so nothing gets skimmed.
 if(p.comp&&p.comp.any){var cbits=[];
  if(p.comp.electrical)cbits.push(p.comp.electrical+' electrical');
  if(p.comp.rigging)cbits.push(p.comp.rigging+' rigging');
  if(p.comp.logbook)cbits.push(p.comp.logbook+' needing a daily pre-start and logbook entry');
  if(p.comp.ret)cbits.push(p.comp.ret+' due back today');
  html+='<div class="alert amb"><b>Check before you use it.</b> Of the gear in your name: '+cbits.join(', ')+'. Anything showing a tag colour needs a current tag with a readable date &mdash; if it hasn\'t got one, don\'t use it, bring it to the store and we\'ll sort it.</div>'}
 var cleared=st.items===0;
 html+='<h3 class="sec" data-sec="store">Return clearance</h3><div class="clr '+(cleared?'done':'open')+'" data-rv>'
  +'<img class="clrimg" src="art/icons/'+(cleared?'cleared':'still-to-clear')+'.webp" '
  +'alt="" loading="lazy" onerror="this.style.display=\'none\'">'
  +'<div class="ci">'+(cleared?'&#10003;':st.items)+'</div><div class="ct">'+(cleared?'<b>Cleared — all gear returned</b><span>Nothing on hire in your name. Legend — thanks for bringing it all back.</span>':'<b>'+st.items+' still to clear</b><span>Bring these back to the tool store and you\'re fully cleared of tools on hire.</span>')+'</div></div>';
 if(p.act){
 html+='<h3 class="sec">Your store scorecard</h3><div class="stats" data-rv>'
 +'<div class="st">'+sticon('out','store-visits')+'<div class="sv"><div class="v" data-to="'+p.act.visits+'">0</div><div class="l">Store visits</div></div></div>'
 +'<div class="st">'+sticon('types','transactions')+'<div class="sv"><div class="v" data-to="'+p.act.txns+'">0</div><div class="l">Transactions</div></div></div>'
 +'<div class="st">'+sticon('back','consumables')+'<div class="sv"><div class="v" data-to="'+((p.cons&&p.cons.n)||0)+'">0</div><div class="l">Consumables</div></div></div>'
 +'<div class="st">'+sticon('bolt','radio-gear')+'<div class="sv"><div class="v" data-to="'+((p.radios&&p.radios.taken)||0)+'">0</div><div class="l">Radio gear</div></div></div></div>';
 var xl=[];
 if(p.radios&&p.radios.taken){xl.push('Radio gear: '+p.radios.back+' of '+p.radios.taken+' back')}
 if(p.gas&&p.gas.taken){xl.push('Gas monitors: '+p.gas.back+' of '+p.gas.taken+' back')}
 if(p.oth&&p.oth.taken){xl.push('Client gear through you: '+p.oth.back+' of '+p.oth.taken+' back')}
 if(p.cons&&p.cons.n&&p.cons.list&&p.cons.list.length){xl.push(p.cons.list.map(function(c){return esc(c.d)+(c.q>1?' x'+c.q:'')}).join(', '))}
 if(p.act.to){xl.push('Store history from SiteIQ transactions to '+esc(p.act.to))}
 if(xl.length){html+='<div class="subline">'+xl.join(' &middot; ')+'</div>'}
 if(p.radios&&p.radios.out>0){html+='<div class="alert amb"><b>&#128251; '+p.radios.out+' piece'+(p.radios.out===1?'':'s')+' of radio kit still in your name.</b> Radios and batteries are gold and the next crew needs them charged &mdash; over the counter and they&rsquo;re off your list in seconds.</div>'}
 if(p.gas&&p.gas.out>0){html+='<div class="alert red"><b>&#9888;&#65039; '+p.gas.out+' gas monitor'+(p.gas.out===1?'':'s')+' not back.</b> Monitors keep you alive and need charge and calibration between shifts &mdash; that&rsquo;s Life Saving Rule 5 territory. Counter today, please.</div>'}
 if(p.dmg&&p.dmg.n>0){html+='<div class="alert amb"><b>'+p.dmg.n+' damage charge'+(p.dmg.n===1?'':'s')+' recorded in your name'+(p.dmg.list&&p.dmg.list.length?' ('+esc(p.dmg.list.join(', '))+')':'')+'.</b> Doesn&rsquo;t look right? See us at the counter &mdash; easy fixed.</div>'}
 else{html+='<div class="okline">&#10003; No damage recorded in your name this shutdown &mdash; gear looked after.</div>'}
 }
 html+='<div class="help" data-sec="help" data-rv><b>We\'re here to help.</b> Finished with something? Bring it back and we\'ll clean it and get it ready for the next crew. Need gear, or something\'s not right? Tell the tool store team, day or night — we\'ll sort it, no fuss.</div>';
  html+='<div class="guidelink" onclick="openGuide(\'contacts\')"><b>Site guides</b><span>Contact board &middot; radio &middot; gas monitor</span><em>&rsaquo;</em></div>';
 // THE BARCODES CAME OFF THE PHONE SCREEN (Andrew, 2 Aug 2026: "can we
 // remove this part"). Two white blocks in the middle of a dark report
 // is a hard stop for the eye, and in practice the counter scans the
 // bloke's actual Coates card, not a picture of one on his phone.
 // They are STILL ON THE PRINTED A4, where a scanner genuinely needs
 // them - see barcodePair() in the print sheet. Nothing else changed:
 // the ID still shows in the header badge and on the welcome card.
 // A Scan button HERE too (Andrew, 27 Jul 2026): once you are in, the
 // only way to look up another card was Done then re-type. At a busy
 // counter that is the wrong answer - scan the next bloke's card from
 // right here. Hidden unless the browser will actually allow a camera,
 // same rule as the landing page.
 html+='<div class="actions"><button class="primary" onclick="saveCard()">&#128241; Save picture</button><button onclick="printCard()">&#128424;&#65039; Print A4</button><button id="cardscan" onclick="startScan()" hidden>&#9635; Scan another</button><button onclick="reset()">Done</button></div>'
 +'<div class="subline" style="text-align:center">Save puts your report on your phone as a picture — no signal needed. Print gives you one clean A4 page; pick the store Wi-Fi printer in the print menu.</div>'
 +'<div class="ft">Coates · K2 Shutdown 2026 · Gladstone &middot; a keepsake of your shutdown<div class="val">Care Deeply · Customer Focused · Be Our Best · One Team · Competitive Spirit</div>POWERED BY SITEIQ · Author: Andrew Fisher</div></div>';
 document.getElementById('result').innerHTML=html;
 document.body.classList.add('hascard');
 // the card's own Scan button lives inside that HTML, so reveal it now
 try{ revealScanControls(); }catch(e){}
 document.getElementById('landing').style.display='none';
 document.getElementById('result').style.display='block';
 window.scrollTo(0,0);
 /* two frames so the browser has laid the card out before anything is
    asked to move - a transition set on the same frame as the markup
    just snaps */
 requestAnimationFrame(function(){requestAnimationFrame(function(){
   animate(document.getElementById('result'));
   rvNav();
 })});
 if((p.score||0)>=85||st.items===0||st.sameday>0) confetti();
}
/* ------------------------------------------------------------------
   THE SECTION RAIL. The report is five screens long on a phone. This
   is the same idea as the stores dock, in the same materials, so the
   suite has one way of moving around and not two.

   It tracks where you actually are rather than only where you tapped,
   so scrolling with a thumb keeps it honest. GEAR carries the number
   of items in his name, because that is the one a bloke jumps to.
------------------------------------------------------------------ */
/* the filter itself. It only ever hides rows - it never rebuilds the
   list, so the compliance chips, photos and QR codes already drawn stay
   exactly as they were. */
var GLANE='all';
function gfilt(lane){
  if(lane) GLANE=lane;
  var q=(document.getElementById('gitq')||{value:''}).value
         .toLowerCase().replace(/^\s+|\s+$/g,'');
  var rows=document.querySelectorAll('#result .item'), shown=0;
  for(var i=0;i<rows.length;i++){
    var r=rows[i];
    var okL=(GLANE==='all')||(r.getAttribute('data-l')===GLANE);
    var okQ=!q||(r.getAttribute('data-q')||'').indexOf(q)>=0;
    if(okL&&okQ){ r.style.display=''; shown++; } else { r.style.display='none'; }
  }
  var bs=document.querySelectorAll('#result .gl');
  for(var j=0;j<bs.length;j++)
    bs[j].className='gl '+(bs[j].getAttribute('data-l')==='all'?'':bs[j].getAttribute('data-l'))
      +(bs[j].getAttribute('data-l')===GLANE?' on':'');
  var none=document.getElementById('gnone');
  if(none) none.hidden=(shown>0);
}
var RVN_IO=null;
function rvNav(){
  var bar=document.getElementById('rnav');
  if(!bar) return;
  var n=document.getElementById('rnavn');
  if(n){ var st=(window.LASTP&&window.LASTP.stats)||{};
         n.textContent=st.items>0?st.items:''; }
  /* a bloke who is fully cleared has no gear list to jump to */
  var gb=bar.querySelector('[data-g="gear"]');
  if(gb) gb.style.display=document.querySelector('[data-sec="gear"]')?'':'none';
  bar.className='rnav on';
  document.body.classList.add('hasrnav');
  rvMark('sum');
  if(RVN_IO){ window.removeEventListener('scroll',RVN_IO); RVN_IO=null; }
  /* Tracked with an IntersectionObserver first and it was wrong: the
     markers are headings and one empty div, so they are only a few
     pixels tall. Watching for them to cross a band in the middle of
     the screen meant whichever heading passed through LAST stayed lit
     no matter where you actually were - scroll back to the top and the
     rail still said STORE.
     A marker's position is the honest question, so ask it directly:
     the section you are in is the last marker above the line. */
  var tick=false;
  RVN_IO=function(){
    if(tick) return; tick=true;
    requestAnimationFrame(function(){
      tick=false;
      var m=document.querySelectorAll('[data-sec]'), line=window.innerHeight*0.34,
          cur=null;
      for(var i=0;i<m.length;i++){
        if(m[i].offsetParent===null && m[i].getAttribute('data-sec')!=='sum') continue;
        if(m[i].getBoundingClientRect().top<=line) cur=m[i].getAttribute('data-sec');
      }
      /* at the very bottom the last section is the one you are reading,
         whatever the maths says about a line a third of the way down */
      if(window.innerHeight+window.pageYOffset >= document.body.scrollHeight-4 && m.length)
        cur=m[m.length-1].getAttribute('data-sec');
      rvMark(cur||'sum');
    });
  };
  window.addEventListener('scroll',RVN_IO,{passive:true});
  RVN_IO();
}
function rvMark(k){
  var bs=document.querySelectorAll('#rnav button');
  for(var i=0;i<bs.length;i++)
    bs[i].className=(bs[i].getAttribute('data-g')===k)?'on':'';
}
function rvGo(k){
  var t=document.querySelector('[data-sec="'+k+'"]');
  if(!t) return;
  rvMark(k);
  try{ t.scrollIntoView({behavior:'smooth',block:'start'}); }
  catch(e){ t.scrollIntoView(); }
}
function reset(){document.body.classList.remove('hascard');
 /* the rail belongs to a card - it goes when the card goes */
 document.body.classList.remove('hasrnav');
 var _rn=document.getElementById('rnav'); if(_rn)_rn.className='rnav';
 if(RVN_IO){window.removeEventListener('scroll',RVN_IO);RVN_IO=null} GLANE='all';document.getElementById('welcome').className='wel';var _d=document.querySelector('.door');if(_d)_d.style.display='';idState('');window.PENDING=null;document.getElementById('result').style.display='none';document.getElementById('result').innerHTML='';document.getElementById('landing').style.display='block';document.getElementById('idno').value='';document.getElementById('idno').focus()}
document.getElementById('idno').addEventListener('keydown',function(e){if(e.key==='Enter')go()});
var STORES_TAG='__STORESTAG__';
__STOREJS__
__UIJS__
// self-test
try{var ok=JSON.parse(dec('SELFTEST',DATA[tag('SELFTEST')])).name==='SELFTEST';if(!ok)console.warn('selftest fail')}catch(e){console.warn('selftest err',e)}
</script></body></html>'''

if __name__ == '__main__':
    try:
        build()
    except SystemExit:
        raise
    except Exception as e:
        print('PROBLEM: the build fell over - ' + str(e))
        print('Nothing was overwritten. Check the exports are the standard '
              'SiteIQ pulls and run again.')
        sys.exit(1)
