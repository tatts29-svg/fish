#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | PEOPLE LIST - everyone on site, and who still has no card
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 28 Jul 2026): "how about i get a full full list, as this
#  won't work doing it this way."
#
#  He is right. A My Gear card is built from ON_HIRE, keyed on
#  EXTERNAL_ID - so it only exists for someone who has held gear under
#  their own ID at least once. Everyone else shows up as a line of names
#  in a NOTE and can do nothing about it.
#
#  And there is no second source to fall back on: neither
#  TRANSACTION_CHARGES nor CUSTOMER_CONTRACTOR_EQUIP carries a hirer ID.
#  Checked, 28 Jul 2026. ON_HIRE is the only place an ID appears.
#
#  So this pulls every person out of every export we have, puts them in
#  one list, and marks plainly who has an ID and who doesn't. The ones
#  without are the work: paste them into the ID Cards sheet of
#  Coates_Report_Recipients.xlsx and they get a card on the next build.
#
#  Writes three things:
#     People_List.csv    - the lot, for sorting and filtering
#     Needs_An_ID.csv    - just the gap, ready to paste into ID Cards
#     People_List.html   - the printable version
# =====================================================================
import csv
import datetime as dt
import io
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "Gear_Lookup")


def _openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print(" openpyxl isn't installed - run:  py -m pip install openpyxl")
        sys.exit(1)


def newest(pattern):
    import glob
    hits = [p for p in glob.glob(os.path.join(HERE, "Data_SiteIQ", pattern))
            if not os.path.basename(p).startswith("~")]
    hits += [p for p in glob.glob(os.path.join(HERE, pattern))
             if not os.path.basename(p).startswith("~")]
    return max(hits, key=os.path.getmtime) if hits else None


def norm(s):
    """ON_HIRE writes 'First - Last', the transaction sheets write
    'First Last'. Same person, two spellings - flatten both."""
    s = str(s or "").replace("–", "-").strip()
    s = " ".join(s.replace(" - ", " ").split())
    return s


def sheet(openpyxl, path, want):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[want] if want in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}, []
    hdr = {str(v).strip(): i for i, v in enumerate(rows[0]) if v}
    return hdr, rows[1:]


def main():
    print("=" * 72)
    print(" COATES | PEOPLE LIST - everyone on site, and who has no card")
    print("=" * 72)
    openpyxl = _openpyxl()
    os.makedirs(OUT, exist_ok=True)
    people = {}          # normalised name -> record

    def rec(name):
        n = norm(name)
        if not n:
            return None
        return people.setdefault(n, {
            "name": n, "company": "", "id": "", "onhire": 0,
            "activity": 0, "damage": 0, "source": set(),
            "status": ""})

    # ---- 1. ON_HIRE: the only place an EXTERNAL_ID ever appears -------
    p = newest("ON_HIRE*.xlsx")
    if not p:
        print(" No ON_HIRE export found. Run 28_PULL_SITEIQ_EXPORTS.bat first.")
        return 1
    hdr, rows = sheet(openpyxl, p, "ON_HIRE")
    for r in rows:
        def g(c):
            i = hdr.get(c)
            return str(r[i] or "").strip() if i is not None and i < len(r) else ""
        e = rec(g("HIRER_NAME"))
        if not e:
            continue
        e["id"] = e["id"] or g("EXTERNAL_ID")
        e["company"] = e["company"] or g("COMPANY")
        e["onhire"] += 1
        e["source"].add("on hire")
    print(" ON_HIRE            : {} people holding gear".format(
        sum(1 for e in people.values() if e["onhire"])))

    # ---- 1b. the site roster - everyone, with their hire ID ----------
    #  This is the one that closes the gap. ON_HIRE only shows people
    #  holding gear today; HIRERS_ID.xlsx is the whole site.
    rp = newest("HIRERS_ID*.xlsx")
    roster = 0
    if rp:
        rw = openpyxl.load_workbook(rp, read_only=True, data_only=True)
        rr = list(rw[rw.sheetnames[0]].iter_rows(values_only=True))
        rh = {}
        for row in rr[:5]:
            c = [str(v).strip().lower() if v is not None else '' for v in row]
            if "name" in c and "external id" in c:
                rh = {v: i for i, v in enumerate(c) if v}
                break
        if rh:
            ni, ei = rh["name"], rh["external id"]
            ci, si = rh.get("employer"), rh.get("status")
            for row in rr:
                nm = str(row[ni] or "").strip() if ni < len(row) else ""
                idn = str(row[ei] or "").strip() if ei < len(row) else ""
                if not nm or nm.lower() == "name":
                    continue
                e = rec(nm)
                if not e:
                    continue
                e["source"].add("site roster")
                if idn and not e["id"]:
                    e["id"] = idn
                    roster += 1
                if ci is not None and ci < len(row) and not e["company"]:
                    e["company"] = str(row[ci] or "").strip()
                if si is not None and si < len(row):
                    st = str(row[si] or "").strip()
                    if st and st.lower() != "active":
                        e["status"] = st
        rw.close()
    print(" Site roster        : {} more with an ID (HIRERS_ID.xlsx)"
          .format(roster))

    # ---- 2. the roster memory - IDs seen in any earlier pull ----------
    cache = os.path.join(HERE, "MY_GEAR_PEOPLE.csv")
    remembered = 0
    if os.path.isfile(cache):
        with io.open(cache, encoding="utf-8", errors="replace", newline="") as f:
            for row in csv.reader(f):
                if len(row) >= 3 and row[0]:
                    e = rec(row[1])
                    if e and not e["id"]:
                        e["id"] = row[0].strip()
                        e["company"] = e["company"] or row[2].strip()
                        e["source"].add("earlier pull")
                        remembered += 1
    print(" Roster memory      : {} more with an ID from an earlier pull"
          .format(remembered))

    # ---- 3. IDs Andrew has typed in by hand ---------------------------
    book = os.path.join(HERE, "Coates_Report_Recipients.xlsx")
    typed = 0
    if os.path.isfile(book):
        try:
            hdr2, rows2 = sheet(openpyxl, book, "ID Cards")
            ci = hdr2.get("ID No")
            ni = hdr2.get("Hirer Name")
            for r in rows2:
                if ci is None or ni is None:
                    break
                idno = str(r[ci] or "").strip()
                nm = str(r[ni] or "").strip()
                if not idno or not nm or "DEMO" in str(r[-1] or "").upper():
                    continue
                e = rec(nm)
                if e and not e["id"]:
                    e["id"] = idno
                    e["source"].add("typed in")
                    typed += 1
        except Exception:
            pass
    print(" ID Cards sheet     : {} typed in by hand".format(typed))

    # ---- 4. everyone who touched the store in the period --------------
    p = newest("TRANSACTION*.xlsx")
    if p:
        for tab in ("TRANSACTION_CHARGES", "CUSTOMER_CONTRACTOR_EQUIP"):
            try:
                hdr3, rows3 = sheet(openpyxl, p, tab)
            except Exception:
                continue
            hi = hdr3.get("HIRER_NAME") or hdr3.get("HIRER")
            di = hdr3.get("DAMAGE ($)")
            ei = hdr3.get("EMPLOYER_NAME")
            if hi is None:
                continue
            for r in rows3:
                e = rec(r[hi])
                if not e:
                    continue
                e["activity"] += 1
                e["source"].add("store activity")
                if ei is not None and not e["company"]:
                    e["company"] = str(r[ei] or "").strip()
                if di is not None:
                    try:
                        e["damage"] += float(r[di] or 0)
                    except (TypeError, ValueError):
                        pass
    print(" Store activity     : {} people seen at the counter".format(
        sum(1 for e in people.values() if e["activity"])))

    # ---- not people ---------------------------------------------------
    def is_person(n):
        low = n.lower()
        return not (low.startswith("site plant equipment")
                    or low.startswith("admin costs")
                    or low.startswith("pool ")
                    or low in ("", "-"))

    rows_out = [e for e in people.values() if is_person(e["name"])]
    dropped = len(people) - len(rows_out)
    have = [e for e in rows_out if e["id"]]
    gap = [e for e in rows_out if not e["id"]]

    #  the gap first - that is the only part that needs doing
    rows_out.sort(key=lambda e: (bool(e["id"]), -e["activity"], e["name"]))

    def w(path, items, cols):
        with io.open(path, "w", encoding="utf-8-sig", newline="") as f:
            wr = csv.writer(f)
            wr.writerow(cols)
            for e in items:
                wr.writerow([e["name"], e["company"], e["id"] or "",
                             e["onhire"], e["activity"],
                             ("%.2f" % e["damage"]) if e["damage"] else "",
                             "yes" if e["id"] else "NO CARD",
                             ", ".join(sorted(e["source"]))][:len(cols)])

    cols = ["Name", "Company", "Hire ID", "On hire now", "Store activity",
            "Damage $", "Has a card", "Seen in"]
    w(os.path.join(OUT, "People_List.csv"), rows_out, cols)
    #  Needs_An_ID is shaped to paste STRAIGHT into the ID Cards sheet:
    #  ID No first (blank for him to fill), then the name.
    with io.open(os.path.join(OUT, "Needs_An_ID.csv"), "w",
                 encoding="utf-8-sig", newline="") as f:
        wr = csv.writer(f)
        wr.writerow(["ID No", "Hirer Name", "Notes"])
        for e in sorted(gap, key=lambda x: (-x["activity"], x["name"])):
            note = "{} store transaction(s)".format(e["activity"])
            if e["damage"]:
                note += " | ${:,.2f} damage".format(e["damage"])
            if e["company"]:
                note = e["company"] + " | " + note
            wr.writerow(["", e["name"], note])

    html = build_html(rows_out, have, gap)
    hp = os.path.join(OUT, "People_List.html")
    with io.open(hp, "w", encoding="utf-8") as f:
        f.write(html)

    print("")
    print("-" * 72)
    print(" TOTAL people       : {}".format(len(rows_out)))
    print("   have a card      : {}".format(len(have)))
    print("   NO card yet      : {}   <-- these are the ones to fix"
          .format(len(gap)))
    if dropped:
        print("   not people       : {} (cost centres, pools) - left out"
              .format(dropped))
    print("")
    print(" Written to Gear_Lookup\\ :")
    print("   People_List.csv    every person, every source")
    print("   Needs_An_ID.csv    just the gap - paste STRAIGHT into the")
    print("                      'ID Cards' sheet of")
    print("                      Coates_Report_Recipients.xlsx, fill in the")
    print("                      ID No column, then run 04_RUN_MY_GEAR.bat")
    print("   People_List.html   the printable version")
    return 0


def build_html(rows, have, gap):
    def esc(s):
        return (str(s).replace("&", "&amp;").replace("<", "&lt;")
                .replace(">", "&gt;"))
    body = []
    for e in rows:
        card = ("<span style='color:#1B7F3B;font-weight:700'>{}</span>"
                .format(esc(e["id"])) if e["id"] else
                "<span style='color:#C62828;font-weight:800'>NO CARD</span>")
        body.append(
            "<tr><td>{n}</td><td>{c}</td><td class='n'>{id}</td>"
            "<td class='n'>{oh}</td><td class='n'>{a}</td>"
            "<td class='n'>{d}</td></tr>".format(
                n=esc(e["name"]), c=esc(e["company"]), id=card,
                oh=e["onhire"] or "-", a=e["activity"] or "-",
                d=("${:,.0f}".format(e["damage"]) if e["damage"] else "-")))
    #  CSS is full of { } and .format() reads those as fields - that is
    #  exactly how a page of headline numbers printed as "{v}" this
    #  morning. So the stylesheet is concatenated, never formatted, and
    #  the only .format() runs on the small block that has real fields.
    css = ("@page{size:A4 portrait;margin:12mm}"
           "*{box-sizing:border-box;margin:0;padding:0}"
           "body{font-family:'Segoe UI',Calibri,Arial,sans-serif;color:#1D1D1B;"
           "font-size:9.5pt}"
           "h1{font-size:19pt;color:#F26222;letter-spacing:-.5px}"
           ".sub{color:#5A6472;margin:2mm 0 5mm;font-size:10pt}"
           ".tiles{display:grid;grid-template-columns:repeat(3,1fr);gap:4mm;"
           "margin-bottom:6mm}"
           ".t{border:.4mm solid #D6DBE2;border-top:1.2mm solid #F26222;"
           "border-radius:3mm;padding:4mm}"
           ".t b{display:block;font-size:20pt;line-height:1}"
           ".t span{font-size:8pt;letter-spacing:1.2px;color:#5A6472;"
           "text-transform:uppercase}"
           "table{width:100%;border-collapse:collapse}"
           "th{background:#1D1D1B;color:#fff;font-size:8pt;letter-spacing:1px;"
           "text-transform:uppercase;padding:2.5mm;text-align:left}"
           "td{padding:2mm 2.5mm;border-bottom:.2mm solid #E6E9EE}"
           "td.n,th.n{text-align:right}"
           "tr:nth-child(even) td{background:#F7F8FA}"
           ".note{margin-top:5mm;font-size:9pt;color:#5A6472;"
           "border-left:1.2mm solid #F26222;padding-left:4mm}")
    tiles = ("<div class='tiles'>"
             "<div class='t'><b>{tot}</b><span>People</span></div>"
             "<div class='t'><b>{h}</b><span>Have a card</span></div>"
             "<div class='t'><b style='color:#C62828'>{g}</b>"
             "<span>No card yet</span></div></div>").format(
                 tot=len(rows), h=len(have), g=len(gap))
    return ("<!DOCTYPE html><html><head><meta charset='utf-8'>"
            "<title>Coates K2 - People List</title><style>" + css +
            "</style></head><body>"
            "<h1>Everyone on site &mdash; and who still has no card</h1>"
            "<div class='sub'>Cement Australia K2 Shutdown 2026 &middot; "
            "Gladstone &middot; " + dt.date.today().strftime("%d %b %Y") +
            " &middot; Author: Andrew Fisher</div>"
            + tiles +
            "<table><tr><th>Name</th><th>Company</th><th class='n'>Hire ID</th>"
            "<th class='n'>On hire</th><th class='n'>Store activity</th>"
            "<th class='n'>Damage</th></tr>"
            + "".join(body) + "</table>"
            "<div class='note'>A card is built from the ON_HIRE export, keyed "
            "on EXTERNAL_ID - the only place an ID appears. Anyone who has "
            "never held gear under their own ID has no ID for us to use. "
            "Put their number in the <b>ID Cards</b> sheet of "
            "Coates_Report_Recipients.xlsx and they get a card on the next "
            "build.</div></body></html>")


if __name__ == "__main__":
    sys.exit(main())
