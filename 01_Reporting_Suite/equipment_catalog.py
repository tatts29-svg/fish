#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | HVAC & POWER EQUIPMENT CATALOG - what we can get from fleet
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  HVAC_POWER_EQUIPMENT_CATALOG.xlsx is the register of fleet catalog
#  MODELS (taken in from Andrew's extract, 20 Aug 2026): packaged air
#  conditioners, fluid chillers, air handlers, and the medium/large
#  diesel generators. One tab, keyed on MODEL - the SAME code SiteIQ
#  calls PRODUCT_VARIANT in every export and the contracted-rates
#  import calls Product Variant ID. That one code is how gear is
#  ordered, how it lands in the exports, and how its rate is set.
#
#    CATEGORY_CODE | TYPE_CODE | TYPE_DESC | PRICING_GROUP |
#    PRICING_GROUP_DESC | MODEL | MODEL_DESC | COMMON_NO |
#    MODEL_TYPE | ACTIVE_FLAG
#
#  This is MODEL-level (what exists in the fleet), not asset-level -
#  asset identity stays with K2_MASTER_EQUIPMENT_PRICING.xlsx and its
#  ITEM_NUMBERs. No rates live here either: a pricing group is the
#  billing family a model charges under, not a dollar figure.
#
#  Add a row in the register and the next build of the catalog page
#  (57_HVAC_POWER_CATALOG.bat) carries it. Same manners as the master
#  file: found beside the suite (root or Data_SiteIQ\), newest wins,
#  and loading is optional-safe - no file, empty catalog, and anything
#  that asked just carries on without it.
# =====================================================================

import os
import glob
import re
import datetime as dt

CATALOG_PATTERN = "HVAC_POWER_EQUIPMENT_CATALOG*.xlsx"
_HERE = os.path.dirname(os.path.abspath(__file__))


class Catalog(object):
    def __init__(self):
        self.by_model = {}
        self.path = None
        self.mtime = None

    @property
    def loaded(self):
        return bool(self.by_model)

    def rec(self, model):
        """The catalog row for a MODEL / PRODUCT_VARIANT code, or None.
        Forgiving on case and stray spaces - export cells arrive with
        both."""
        if model is None:
            return None
        return self.by_model.get(str(model).strip().upper())

    def desc(self, model, fallback=""):
        r = self.rec(model)
        if r and r["desc"]:
            return r["desc"]
        return fallback

    def category(self, model):
        r = self.rec(model)
        return r["category"] if r else ""

    def type_desc(self, model):
        r = self.rec(model)
        return r["type_desc"] if r else ""

    def group(self, model):
        r = self.rec(model)
        return r["group"] if r else ""

    def group_desc(self, model):
        r = self.rec(model)
        return r["group_desc"] if r else ""

    def models(self):
        """Every record, catalog order preserved."""
        return list(self.by_model.values())

    def categories(self):
        """Category codes in the order they first appear (HVAC, POWER)."""
        seen = []
        for r in self.by_model.values():
            if r["category"] and r["category"] not in seen:
                seen.append(r["category"])
        return seen

    def shared_groups(self):
        """Pricing groups carrying more than one model - real in the
        source (12985 arrived holding both the 350kVA and the 365kVA
        genset), so the page can say so instead of looking wrong."""
        counts = {}
        for r in self.by_model.values():
            if r["group"]:
                counts.setdefault(r["group"], []).append(r["model"])
        return {g: ms for g, ms in counts.items() if len(ms) > 1}


def _clean(v):
    return "" if v is None else str(v).strip()


def _norm_header(v):
    """MODEL_DESC, 'Model Desc' and 'Common No.' all answer to the same
    name - Andrew's extract had spaced headers, the register uses
    underscores, and a future re-export could be either."""
    return re.sub(r"[^A-Z0-9]+", "_", _clean(v).upper()).strip("_")


_CAP = re.compile(r"(\d+)\s*(KVA|KW)", re.I)


def capacity(model_or_desc):
    """(number, unit) parsed out of a model code or description -
    ACPAC20KW -> (20, 'kW'), GENERATOR1250KVAD -> (1250, 'kVA').
    Sort-order only, never a data field. None when nothing matches."""
    m = _CAP.search(_clean(model_or_desc))
    if not m:
        return None
    return int(m.group(1)), ("kVA" if m.group(2).upper() == "KVA" else "kW")


def load(base_dir=None, quiet=False):
    """Load the catalog register. Empty Catalog (never None) when the
    file is absent - callers carry on as if it never existed."""
    base = base_dir or _HERE
    hits = []
    for d in (base, os.path.join(base, "Data_SiteIQ")):
        hits += [p for p in glob.glob(os.path.join(d, CATALOG_PATTERN))
                 if not os.path.basename(p).startswith("~$")]
    c = Catalog()
    if not hits:
        if not quiet:
            print("  Catalog register : (not found - HVAC & power catalog "
                  "not loaded)")
        return c
    path = max(hits, key=os.path.getmtime)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)

        #  The register carries a title band above the headers, so the
        #  header row is FOUND, not assumed - a re-export with no band
        #  loads just the same.
        hdr = None
        for _ in range(12):
            try:
                row = next(rows)
            except StopIteration:
                break
            names = [_norm_header(v) for v in row]
            if "MODEL" in names:
                hdr = names
                break
        if hdr is None:
            if not quiet:
                print("  Catalog register : found but no MODEL column in "
                      "the first 12 rows - skipped ({})."
                      .format(os.path.basename(path)))
            wb.close()
            return c

        def col(name):
            return hdr.index(name) if name in hdr else None

        c_cat = col("CATEGORY_CODE")
        c_tc = col("TYPE_CODE")
        c_td = col("TYPE_DESC")
        c_pg = col("PRICING_GROUP")
        c_pgd = col("PRICING_GROUP_DESC")
        c_mod = col("MODEL")
        c_md = col("MODEL_DESC")
        c_cn = col("COMMON_NO")
        c_mt = col("MODEL_TYPE")
        c_af = col("ACTIVE_FLAG")

        for row in rows:
            def g(ci):
                return _clean(row[ci]) if ci is not None and ci < len(row) else ""

            model = g(c_mod).upper()
            if not model:
                continue
            c.by_model[model] = {
                "model": model,
                "category": g(c_cat).upper(),
                "type_code": g(c_tc),
                "type_desc": g(c_td),
                "group": g(c_pg),
                "group_desc": g(c_pgd),
                "desc": g(c_md),
                "common_no": g(c_cn),
                "model_type": g(c_mt),
                "active": g(c_af).lower() == "active",
            }
        wb.close()
        c.path = path
        c.mtime = dt.datetime.fromtimestamp(os.path.getmtime(path))
        if not quiet:
            cats = {}
            for r in c.by_model.values():
                cats[r["category"]] = cats.get(r["category"], 0) + 1
            print("  Catalog register : {}  ({:,} models | {})".format(
                os.path.basename(path), len(c.by_model),
                " | ".join("{} {}".format(n, k)
                           for k, n in sorted(cats.items()))))
    except Exception as e:
        if not quiet:
            print("  Catalog register : couldn't read it ({}) - HVAC & "
                  "power catalog not loaded.".format(e))
        c.by_model = {}
    return c
