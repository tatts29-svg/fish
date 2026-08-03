#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | K2 MASTER EQUIPMENT FILE - one file feeds everything
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  K2_MASTER_EQUIPMENT_PRICING.xlsx is the single source for equipment
#  identity across the whole suite AND the K2 Excel (per A. Fisher,
#  24 Jul 2026): keyed on ITEM_NUMBER (the asset number - "use the item
#  number for everything"), one tab, carrying:
#
#    STORAGE_UNIT | ITEM_NUMBER | PLANT_ID | ITEM_DESCRIPTION |
#    PRODUCT_VARIANT | NEW_DESCRIPTION | REPLACEMENT_COST_AUD |
#    REPLACEMENT_PRICE_SOURCE | EQUIPMENT_CATEGORY |
#    ELECTRICAL_TAG | RIGGING_TAG | LOGBOOK_REQUIRED | RETURN_REQUIREMENT
#
#  Add or edit a row there and the next run updates every report and the
#  Excel true-up alike:
#    * NEW_DESCRIPTION becomes the DISPLAYED name wherever the item
#      appears. The original description is kept alongside so SiteIQ
#      matching, billing counts and pattern rules never break.
#    * REPLACEMENT_COST_AUD (with its SOURCE) prices the item by exact
#      asset identity - stronger than any description match. The old
#      description-keyed schedule stays as the fallback; anything priced
#      nowhere shows TBC and rides the daily gap list. Never guessed.
#    * ELECTRICAL_TAG / RIGGING_TAG / LOGBOOK_REQUIRED / RETURN_REQUIREMENT
#      (added 25 Jul 2026) put the compliance line UNDER the description
#      on every report, every email and My Gear. Y turns a flag on, blank
#      turns it off - no data in the column, nothing in the report. The
#      wording and the current tag colour live in equipment_compliance.py.
#
#  Portable: found beside the suite (root or Data_SiteIQ\), newest wins.
#  Loading is optional-safe: no file -> every report runs exactly as it
#  did before the master existed.
# =====================================================================

import os
import glob
import re
import datetime as dt

MASTER_PATTERN = "K2_MASTER_EQUIPMENT_PRICING*.xlsx"
_HERE = os.path.dirname(os.path.abspath(__file__))


DECISIONS_FILE = "RENAME_DECISIONS.txt"


def _vkey(v):
    """A PRODUCT_VARIANT reduced to something joinable.

    SiteIQ writes it two ways for the same asset: RENTAL_STOCK gives the
    code GIRDERTROLLEY2T, STOCKTAKE gives the words "Girder Trolley -
    2t". Stripping everything but letters and digits makes them the same
    key, so one rule covers the asset wherever it turns up.
    """
    return re.sub(r"[^A-Z0-9]", "", _clean(v).upper())


def _load_decisions(base):
    """Andrew's locked-in naming calls. See RENAME_DECISIONS.txt.

    Two kinds of line, both keyed on SiteIQ's own wording:
      "<siteiq wording>  =>  <the name it reads>"   a rename
      "<siteiq wording>  =>  KEEP"                  approve the master's
                                                    answer and stop
                                                    asking about it

    Returns (renames, approved). Missing file = both empty and the suite
    behaves exactly as it did before the file existed.
    """
    ren, keep = {}, set()
    p = os.path.join(base or _HERE, DECISIONS_FILE)
    if not os.path.isfile(p):
        return ren, keep
    try:
        with open(p, encoding="utf-8", errors="replace") as fh:
            for line in fh:
                line = line.strip()
                if not line or line.startswith("#") or "=>" not in line:
                    continue
                left, right = line.split("=>", 1)
                v = right.strip()
                #  "<wording> | <PRODUCT_VARIANT>" pins the rule to ONE
                #  of two products SiteIQ describes identically. Same
                #  bar as HIDDEN_ITEMS.txt uses to pin a rule to one
                #  shelf. Without it a rule is about the wording alone.
                var = ""
                if "|" in left:
                    left, var = left.split("|", 1)
                k = _dkey(left)
                if not k or not v:
                    continue
                if var.strip():
                    ren[(k, _vkey(var))] = v
                elif v.upper() == "KEEP":
                    keep.add(k)
                else:
                    ren[k] = v
    except OSError:
        pass
    return ren, keep


class Master(object):
    def __init__(self):
        #  Andrew's locked-in calls, applied over the top of the
        #  spreadsheet - see RENAME_DECISIONS.txt
        self.decisions = {}
        self.approved = set()
        #  variant-pinned decisions, resolved to exact asset numbers
        self.by_item_decision = {}
        self.pinned_rules = {}
        self.by_item = {}
        #  SiteIQ's old wording -> Andrew's new wording, built from the
        #  master file's own two description columns. See _index_desc.
        self.by_desc = {}
        self.desc_split = {}
        #  wordings SiteIQ uses for more than one product - never
        #  carried by wording alone. See _index_desc.
        self.desc_ambig = {}
        self.path = None
        self.mtime = None
        self.n_renames = 0
        self.n_priced = 0
        self.n_elec = 0
        self.n_rig = 0
        self.n_log = 0
        self.n_ret = 0

    @property
    def loaded(self):
        return bool(self.by_item)

    def rec(self, item_number):
        if item_number is None:
            return None
        return self.by_item.get(str(item_number).strip())

    def disp(self, item_number, fallback):
        """Displayed name: the master's NEW_DESCRIPTION when this asset
        has one, otherwise whatever the export said. Never blank.

        Two lookups, in order of how much they know:

        1. THIS ASSET NUMBER. Exact, and it always wins - if Andrew has
           named asset 1258193 specifically, that is its name.

        2. THIS WORDING. The master covers 4,400-odd asset numbers, not
           all 5,380, and the gap is not tidy: of the assets SiteIQ
           calls "Bow Shackle - Alloy 3.25t", 13 are in his file and 17
           are not. Keyed on the number alone, thirteen shackles read
           "Shackle - Bow - Alloy - 3.25 t" and seventeen identical
           shackles sat next to them still reading "Bow Shackle" - the
           same product under two names on one screen, which is the
           exact complaint the master file exists to end (3 Aug 2026).

           So an asset with no entry of its own inherits the name his
           file gives that wording elsewhere. Nothing is invented: the
           new name is one he wrote, for gear SiteIQ describes with the
           same words.

        Step 2 is fenced twice, because a name carried too far is worse
        than a raw one. It never fires for a wording SiteIQ uses for
        more than one product (`desc_ambig` - three different boards are
        all "Distribution Board"), and it never fires where his file has
        renamed the same wording two irreconcilable ways (`desc_split`).
        Both stay on SiteIQ's words and get counted in the build.
        A rename nobody chose is not a tidier name.
        """
        #  A LOCKED-IN DECISION OUTRANKS EVERYTHING, including the
        #  asset's own row. That is the point of it: "any air hoses that
        #  are 3/4 fall under the one name" is a call about the whole
        #  register, not about whichever assets happen to be listed by
        #  number - and 57 of the 84 hoses were not (Andrew, 3 Aug 2026).
        if item_number is not None:
            _d = self.by_item_decision.get(str(item_number).strip())
            if _d:
                return _d
        k = _dkey(fallback)
        if k and k in self.decisions:
            return self.decisions[k]
        r = self.rec(item_number)
        if r and r["new_desc"]:
            return r["new_desc"]
        if k and k in self.by_desc:
            return self.by_desc[k]
        return fallback

    def price(self, item_number):
        """Replacement cost by exact asset identity, or None."""
        r = self.rec(item_number)
        if r and r["repl"] is not None and r["repl"] > 0:
            return r["repl"]
        return None

    def price_source(self, item_number):
        r = self.rec(item_number)
        return r["source"] if r else ""

    def plant_id(self, item_number):
        r = self.rec(item_number)
        return r["plant_id"] if r else ""

    def category(self, item_number):
        r = self.rec(item_number)
        return r["category"] if r else ""

    # -- compliance, added 25 Jul 2026 -------------------------------
    def electrical(self, item_number):
        return _flag(self.rec(item_number), "electrical")

    def rigging(self, item_number):
        return _flag(self.rec(item_number), "rigging")

    def logbook(self, item_number):
        return _flag(self.rec(item_number), "logbook")

    def return_note(self, item_number):
        r = self.rec(item_number)
        return r["ret"] if r else ""

    @property
    def n_compliance(self):
        return self.n_elec + self.n_rig + self.n_log + self.n_ret

    def _ruled(self, item_number, raw):
        """Has Andrew already settled this one? Then stop asking."""
        if item_number is not None and \
                str(item_number).strip() in self.by_item_decision:
            return True
        k = _dkey(raw)
        return k in self.approved or k in self.decisions

    def open_ambiguities(self):
        """Wordings SiteIQ uses for two products that nobody has ruled on.

        desc_ambig is the FENCE and it stays armed either way - one name
        must never carry across two products by wording alone. But once
        every variant behind a wording has a pinned decision, the
        question is answered, and a build that keeps asking it trains
        people to scroll past the ones that are still open
        (3 Aug 2026 - the girder trolleys).
        """
        out = {}
        decided = set(vk for _dk, vk in self.pinned_rules)
        for k, variants in self.desc_ambig.items():
            left = [v for v in variants if _vkey(v) not in decided]
            if left:
                out[k] = left
        return out

    def name_tally(self, pairs):
        """Where each asset's displayed name came from.

        Returns {'item': n, 'wording': n, 'siteiq': n} over any iterable
        of (item_number, siteiq_description). The middle bucket is the
        one worth watching: those assets are not in the master by number
        and take their name from what Andrew called the same wording
        elsewhere. Correct, and it should still be said out loud.
        """
        t = {"item": 0, "wording": 0, "siteiq": 0}
        for item, raw in pairs:
            r = self.rec(item)
            if r and r["new_desc"]:
                t["item"] += 1
            elif _dkey(raw) in self.by_desc:
                t["wording"] += 1
            else:
                t["siteiq"] += 1
        return t

    def size_merges(self, pairs):
        """Renames that put two DIFFERENT sizes under one name.

        Most merges in the file are the point of it: SiteIQ writes the
        same fan three ways - "Exhaust Fan - Electric 300mm", "Exhaust
        Fan 300MM", "Exhaust Fan - Electric - 300mm" - and one name for
        all three is exactly the tidy-up. 57 names merge something, and
        56 of them are that.

        The one to catch is the merge where a MEASUREMENT moves: a 19 mm
        air hose and a 25 mm air hose both landing on "19 mm Air Hose".
        That is not tidier wording, it is two sizes wearing one label,
        and the bloke who orders off the label gets the wrong hose.

        Returns [(new_name, [(raw, count), ...])] worst first. Reports
        only; his file still wins the name, same as everywhere.
        """
        seen = {}
        for item, raw in pairs:
            raw = _clean(raw)
            if not raw:
                continue
            new = self.disp(item, raw)
            if not new or new == raw:
                continue
            if self._ruled(item, raw):
                continue
            seen.setdefault(new, {}).setdefault(raw, 0)
            seen[new][raw] += 1
        out = []
        for new, raws in seen.items():
            if len(raws) < 2:
                continue
            #  the measurements each raw description declares. If two of
            #  them declare different ones, the merge lost a size.
            sizes = [_sizes(r) for r in raws]
            real = [s for s in sizes if s]
            if len(real) < 2 or all(s == real[0] for s in real):
                continue
            out.append((new, sorted(raws.items(), key=lambda x: -x[1])))
        out.sort(key=lambda x: -sum(c for _, c in x[1]))
        return out

    def rating_conflicts(self, pairs):
        """Renames where the rated capacity itself moved.

        The change-of-description file is Andrew's own wording and it is
        meant to win - "Cumalong - 3.0t" reading "3 t Lever Block" on
        every screen is the whole point of it. But a lever block's
        tonnage is not wording. If SiteIQ has an asset at 1.6 t and the
        master renames it to 1.5 t, one of the two is wrong about a
        rated capacity on lifting gear, and quietly picking the master's
        number would put a figure on a print that nobody chose.

        So the swap still happens - his file wins, same as everywhere
        else - and the build says out loud which ones moved and by how
        much. A number that changed on its own is the one somebody
        re-derives by hand at the worst moment.

        `pairs` is any iterable of (item_number, siteiq_description).
        Returns [(raw, renamed, count, from_t, to_t)], biggest first.
        """
        seen = {}
        for item, raw in pairs:
            raw = _clean(raw)
            if not raw:
                continue
            new = self.disp(item, raw)
            if not new or new == raw:
                continue
            a, b = _tonnes(raw), _tonnes(new)
            if not (a and b) or a == b:
                continue
            #  ...unless he has already looked at it and ruled. A
            #  decision recorded in RENAME_DECISIONS.txt is an answer,
            #  and a build that keeps asking a question already answered
            #  trains people to scroll past the ones that are new.
            #  ...and a variant-pinned decision is an answer too. The
            #  girder trolley whose capacity "moves" 1 t -> 2 t moves
            #  because Andrew said so; reporting his own ruling back to
            #  him as an open question is noise (3 Aug 2026).
            if self._ruled(item, raw):
                continue
            k = (raw, new)
            if k in seen:
                seen[k][2] += 1
            else:
                seen[k] = [raw, new, 1, sorted(a), sorted(b)]
        out = [tuple(v) for v in seen.values()]
        out.sort(key=lambda x: -x[2])
        return out


def _dkey(desc):
    """A description reduced to something joinable: case and run-on
    spaces are noise, everything else is kept. Deliberately NOT
    punctuation-stripped - "Spanner - Combo 1" and "Spanner - Combo 1/2"
    must stay apart."""
    return " ".join(_clean(desc).lower().split())


def _common_stem(names):
    """The leading fields every one of these names shares, or "".

    Field-wise, never character-wise: "1 t Lever Block" and "10 t Lever
    Block" share the characters "1" but nothing meaningful, and a stem
    of "1" would be worse than no stem at all. Splitting on " - " first
    means a stem is always a run of whole fields.

    At least two fields required, so a stem is a name and not a
    category - "Chain Block" alone, standing in for six capacities,
    would lose the one detail that matters about a chain block.
    """
    parts = [n.split(" - ") for n in names]
    stem = []
    for i in range(min(len(p) for p in parts)):
        f = parts[0][i]
        if any(p[i] != f for p in parts):
            break
        stem.append(f)
    if len(stem) < 2:
        return ""
    return " - ".join(stem).strip()


def _register_variants(base, item_map=None):
    """SiteIQ's own wording -> the set of PRODUCT_VARIANT codes using it.

    Pass `item_map` (a dict) to have it filled with
    item number -> (wording key, normalised variant) at the same time,
    off the same read. That is what lets a decision be written against
    ONE of two products SiteIQ describes identically - see
    _apply_variant_decisions.

    Read off the fleet register, which is the only place that knows a
    "Distribution Board" can be a DISTBOARDLIFEGUARD16, a
    DISTBOARDLIFEGUARD17 or a DISTBOARDLIFEGUARD4-10A. The master file
    cannot tell them apart - it files all three under a flat DISTBOARD -
    so the check has to come from the register. Best effort: no export,
    no map, and the caller falls back to naming by number only.
    """
    out = {}
    #  GLOB, like every other export reader in the suite. SiteIQ's raw
    #  downloads carry their timestamp - "RENTAL_STOCK_23_07_2026 08_50
    #  AM 1.xlsx" - and an exact filename finds none of them. On a
    #  machine that has not renamed its pulls this returned nothing, so
    #  the one-wording-two-products fence never armed and the build's
    #  CHECK line about it vanished without a word. A guard that turns
    #  itself off quietly is worse than no guard (caught on review,
    #  3 Aug 2026).
    for pat in ("RENTAL_STOCK*.xlsx", "SALES_STOCK*.xlsx"):
        for d in (os.path.join(base, "Data_SiteIQ"), base):
            hits = [h for h in glob.glob(os.path.join(d, pat))
                    if not os.path.basename(h).startswith("~$")]
            if not hits:
                continue
            p = max(hits, key=os.path.getmtime)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(p, read_only=True,
                                            data_only=True)
                ws = wb[wb.sheetnames[-1]]
                rows = ws.iter_rows(values_only=True)
                hdr = [_clean(c).upper() for c in next(rows)]
                ix = {h: i for i, h in enumerate(hdr) if h}
                cd = ix.get("ITEM_DESCRIPTION", ix.get("SKU_DESCRIPTION"))
                cv = ix.get("PRODUCT_VARIANT")
                ci = ix.get("ITEM_NUMBER", ix.get("SKU_NUMBER"))
                if cd is None or cv is None:
                    wb.close()
                    break
                for r in rows:
                    if not r:
                        continue
                    k = _dkey(r[cd])
                    v = _clean(r[cv]).upper()
                    if k and v:
                        out.setdefault(k, set()).add(v)
                    if item_map is not None and ci is not None and k:
                        it = _clean(r[ci])
                        if it:
                            item_map[it] = (k, _vkey(v))
                wb.close()
            except Exception:
                pass
            break
    return out


def _apply_variant_decisions(m, base):
    """Turn "<wording> | <VARIANT> => <name>" rules into asset numbers.

    Two girder trolleys are both described "Girder Trolley - 1t" in
    SiteIQ and told apart only by variant - GIRDERTROLLEY1T and
    GIRDERTROLLEY2T - so one of them is a 2 t trolley reading 1 t on the
    shelf (Andrew, 3 Aug 2026: "all girder trolleys the correct way is
    girder trolley 1T or 2T").

    A rule keyed on wording alone cannot separate them. Resolving it to
    ASSET NUMBERS here means everything downstream keeps working
    unchanged - disp() already answers by asset number, and that is the
    strongest key there is.
    """
    pinned = [k for k in m.decisions if isinstance(k, tuple)]
    if not pinned:
        return
    item_map = {}
    _register_variants(base, item_map)
    for item, (dk, vk) in item_map.items():
        name = m.decisions.get((dk, vk))
        if name:
            m.by_item_decision[item] = name
    #  the pinned rules have done their job; leave only plain-wording
    #  rules in decisions so nothing downstream trips over a tuple key
    m.pinned_rules = {k: m.decisions.pop(k) for k in pinned}


def _index_desc(m, base=None):
    """Build the wording index, and record where it cannot decide.

    The names themselves come only from the master file's own
    ITEM_DESCRIPTION and NEW_DESCRIPTION columns, so the index can never
    invent a name Andrew did not write. The register is read for one
    purpose: to find the wordings that cover more than one product.
    """
    fleet = _register_variants(base) if base else {}
    votes = {}
    for rec in m.by_item.values():
        old, new = rec["orig_desc"], rec["new_desc"]
        if not (old and new) or _dkey(old) == _dkey(new):
            continue
        votes.setdefault(_dkey(old), {}).setdefault(new, 0)
        votes[_dkey(old)][new] += 1
    for k, names in votes.items():
        #  ONE WORDING, SEVERAL PRODUCTS - leave it alone.
        #
        #  SiteIQ calls three different boards "Distribution Board" and
        #  tells them apart only by variant: DISTBOARDLIFEGUARD16,
        #  DISTBOARDLIFEGUARD17, DISTBOARDLIFEGUARD4-10A. Andrew has
        #  named one of them, "Lifeguard 16". Carrying that name to the
        #  other two by wording alone would put "Lifeguard 16" on a
        #  Lifeguard 17 - a wrong name, printed confidently, on gear a
        #  sparky picks up off a shelf (3 Aug 2026).
        #
        #  His entry still names his asset. The others keep SiteIQ's
        #  words until he names them, and the build says how many.
        if len(fleet.get(k, ())) > 1:
            m.desc_ambig[k] = sorted(fleet[k])
            continue
        if len(names) == 1:
            m.by_desc[k] = next(iter(names))
            continue
        #  One wording, several new names. Usually that is not two
        #  products at all - it is one product named per unit, the way
        #  the gas monitors carry their own serial: "Multi-Gas Detector
        #  - Honeywell BW Flex - Serial GM206396". Thirty of those
        #  disagree only about the last field.
        #
        #  So take what they all agree on, cut back to a whole field:
        #  "Multi-Gas Detector - Honeywell BW Flex". His words, his
        #  order, and no serial invented for a unit he never listed.
        stem = _common_stem(names)
        if stem:
            m.by_desc[k] = stem
        else:
            #  genuinely two products behind one set of words. Left on
            #  SiteIQ's wording and named out loud rather than silently
            #  picking a winner.
            m.desc_split[k] = sorted(names.items(), key=lambda x: -x[1])


#  a rated capacity, in tonnes: "1.6T", "3.0t", "10 tonne". Deliberately
#  strict - a bare "6" in "6M Drop" is a length, not a load, and reading
#  it as one would invent conflicts that are not there.
_TONNE_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(?:t\b|tonne)", re.I)


def _tonnes(s):
    return set(float(x) for x in _TONNE_RE.findall(s or ""))


#  Any declared measurement: 300mm, 20m, 10A, 240V, 18V, 5AH, 1.5t.
#  Normalised so "300mm" and "300 MM" are one thing and a merge is not
#  reported for a spacing difference.
_SIZE_RE = re.compile(
    r"(\d+(?:\.\d+)?)\s*(mm|cm|m|kg|t|tonne|a|amp|v|volt|ah|w|lm|in|inch)\b",
    re.I)
_SIZE_ALIAS = {"tonne": "t", "amp": "a", "volt": "v", "inch": "in"}


def _sizes(s):
    out = set()
    for num, unit in _SIZE_RE.findall(s or ""):
        u = unit.lower()
        out.add((float(num), _SIZE_ALIAS.get(u, u)))
    return out


def _clean(v):
    return "" if v is None else str(v).strip()


_TRUE = ("y", "yes", "true", "1", "x", "req", "required")


def _isyes(v):
    return _clean(v).lower() in _TRUE


def _flag(rec, key):
    return bool(rec and rec.get(key))


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace("$", "").replace(",", ""))
    except Exception:
        return None


def load(base_dir=None, quiet=False):
    """Load the master file. Empty Master (never None) when absent -
    callers just fall back to pre-master behaviour."""
    base = base_dir or _HERE
    hits = []
    for d in (base, os.path.join(base, "Data_SiteIQ")):
        hits += [p for p in glob.glob(os.path.join(d, MASTER_PATTERN))
                 if not os.path.basename(p).startswith("~$")]
    m = Master()
    if not hits:
        if not quiet:
            print("  Master file : (not found - descriptions and prices run "
                  "as before)")
        return m
    path = max(hits, key=os.path.getmtime)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        hdr = [_clean(c).upper() for c in next(rows)]

        def col(name):
            return hdr.index(name) if name in hdr else None

        c_item = col("ITEM_NUMBER")
        c_orig = col("ITEM_DESCRIPTION")
        c_new = col("NEW_DESCRIPTION")
        c_repl = col("REPLACEMENT_COST_AUD")
        c_src = col("REPLACEMENT_PRICE_SOURCE")
        c_pid = col("PLANT_ID")
        c_cat = col("EQUIPMENT_CATEGORY")
        c_su = col("STORAGE_UNIT")
        c_elec = col("ELECTRICAL_TAG")
        c_rig = col("RIGGING_TAG")
        c_log = col("LOGBOOK_REQUIRED")
        c_ret = col("RETURN_REQUIREMENT")
        if c_item is None:
            if not quiet:
                print("  Master file : found but no ITEM_NUMBER column - "
                      "skipped ({}).".format(os.path.basename(path)))
            wb.close()
            return m
        for row in rows:
            item = _clean(row[c_item]) if c_item < len(row) else ""
            if not item:
                continue

            def g(ci):
                return _clean(row[ci]) if ci is not None and ci < len(row) else ""

            rec = {
                "item": item,
                "orig_desc": g(c_orig),
                "new_desc": g(c_new),
                "repl": _num(row[c_repl]) if c_repl is not None and c_repl < len(row) else None,
                "source": g(c_src),
                "plant_id": g(c_pid),
                "category": g(c_cat),
                "su": g(c_su),
                "electrical": _isyes(g(c_elec)),
                "rigging": _isyes(g(c_rig)),
                "logbook": _isyes(g(c_log)),
                "ret": g(c_ret),
            }
            m.by_item[item] = rec
            if rec["new_desc"] and rec["new_desc"] != rec["orig_desc"]:
                m.n_renames += 1
            if rec["repl"] is not None and rec["repl"] > 0:
                m.n_priced += 1
            if rec["electrical"]:
                m.n_elec += 1
            if rec["rigging"]:
                m.n_rig += 1
            if rec["logbook"]:
                m.n_log += 1
            if rec["ret"]:
                m.n_ret += 1
        wb.close()
        m.decisions, m.approved = _load_decisions(base)
        _apply_variant_decisions(m, base)
        _index_desc(m, base)
        m.path = path
        m.mtime = dt.datetime.fromtimestamp(os.path.getmtime(path))
        if not quiet:
            print("  Master file : {}  ({:,} items | {:,} priced | {:,} "
                  "renamed)".format(os.path.basename(path), len(m.by_item),
                                    m.n_priced, m.n_renames))
            _held = len(m.desc_ambig) + len(m.desc_split)
            print("  Renaming    : {:,} wording(s) carry to assets not "
                  "listed by number{}".format(
                      len(m.by_desc),
                      "; {:,} held back - SiteIQ uses those words for "
                      "more than one product".format(_held)
                      if _held else ""))
            print("  Compliance  : {:,} electrical | {:,} rigging | {:,} "
                  "logbook | {:,} return-daily".format(
                      m.n_elec, m.n_rig, m.n_log, m.n_ret))
    except Exception as e:
        if not quiet:
            print("  Master file : couldn't read it ({}) - descriptions and "
                  "prices run as before.".format(e))
        m.by_item = {}
    return m
