#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | K2 - PLANT LIFE ON SITE
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Andrew, 3 Aug 2026:
#    "this is only for equipment that went onhire to that site plant
#     barcode. so we can have start date... then we want to know the
#     date of the first ever hire out to someone if there was one. then
#     we want to know record of it coming off hire. then a record of it
#     then becoming status departed in stock take. this keeps a record
#     of what and when."
#
#  Four dates, one row per asset, in the order they actually happen:
#
#    1  ON SITE      first on hire to the site plant barcode
#                    (PLANTONHIRE - Coates' own holding account). This
#                    is the day the job took delivery of it.
#    2  FIRST OUT    the first time a NAMED PERSON took it. Some gear
#                    never gets there, and that is the finding, not a
#                    gap - a machine that sat on the holding account for
#                    three weeks and never went out is a machine the
#                    job did not need.
#    3  BACK         the latest hire that has ended - it came off hire.
#    4  DEPARTED     the stocktake recorded it leaving site
#                    (LAST_SIGHTED_ACTION = Departure).
#
#  WHERE EACH DATE COMES FROM, AND WHY IT IS NOT GUESSED
#
#  The transaction feeds carry both sides: a row whose HIRER_NAME is
#  the site plant account is the job holding it, and a row with a
#  person's name is that person holding it. Same asset number, two
#  different kinds of row, and the dates come straight off them.
#
#  SiteIQ spells the account TWO ways - "Site Plant Equipment - ." in
#  ON_HIRE and "Site Plant Equipment ." in the transaction sheets - so
#  the match is on the words, not the punctuation. Keyed on the exact
#  string, a third of the history goes missing and nothing says so.
#
#  Nothing here is inferred. A date that is not in an export is shown
#  as blank and counted as blank; the page says how many of each.
# =====================================================================

from __future__ import print_function

import os
import re
import glob
import datetime as dt

HERE = os.path.dirname(os.path.abspath(__file__))

#  The site's own holding account. Barcode PLANTONHIRE on the ON_HIRE
#  export; the transaction sheets carry only the name.
PLANT_BARCODE = "PLANTONHIRE"
PLANT_WORDS = "siteplantequipment"

#  Sheets that carry a hire with a clock on it. FIXED_CHARGES and
#  TRANSACTION_WITHOUT_CHARGES are empty on this job but are read
#  anyway - an empty sheet costs nothing and a full one that nobody
#  read is how two thirds of the movements went missing once already.
TXN_SHEETS = ("TRANSACTION_CHARGES", "CUSTOMER_CONTRACTOR_EQUIP",
              "TRANSACTION_WITHOUT_CHARGES")


def _is_plant(name):
    """The holding account, however SiteIQ spelt it this morning."""
    return PLANT_WORDS in re.sub(r"[^a-z]", "", str(name or "").lower())


def _txt(v):
    return "" if v is None else str(v).strip()


def _date(v):
    """A date out of whatever SiteIQ wrote - a real date, or text."""
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = _txt(v)
    if not s:
        return None
    s = s.split(" ")[0].replace(".", "/").replace("-", "/")
    for f in ("%d/%m/%Y", "%Y/%m/%d", "%d/%m/%y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(s, f).date()
        except ValueError:
            pass
    return None


#  "03/08/2026 06:30 AM" - how SiteIQ writes the stocktake clock, as
#  TEXT, not as a date cell.
_DT_FORMATS = ("%d/%m/%Y %I:%M %p", "%d/%m/%Y %H:%M",
               "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M")


def _stamp(v):
    """Date AND time where the export carries one.

    The departure record carries a real clock - 06:30 AM - and the first
    cut of this threw it away and stamped midnight instead, which put
    "00:00" on the page next to every departure. A made-up time reads
    exactly like a real one and there is no way to tell from the screen
    which you are looking at. Returns None for the time half when the
    export genuinely has no clock, and the page then shows the date
    alone (3 Aug 2026).
    """
    if isinstance(v, dt.datetime):
        return v
    s = _txt(v)
    for f in _DT_FORMATS:
        try:
            return dt.datetime.strptime(s, f)
        except ValueError:
            pass
    d = _date(v)
    #  date only - hand back the date itself so the caller can tell the
    #  difference between "08:00" and "we do not know the time"
    return d if d else None


def _find(pattern, base=None):
    base = base or HERE
    hits = []
    for d in (os.path.join(base, "Data_SiteIQ"), base):
        hits += [p for p in glob.glob(os.path.join(d, pattern))
                 if not os.path.basename(p).startswith("~$")]
    return max(hits, key=os.path.getmtime) if hits else None


def _sheet(path, name):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[name] if name in wb.sheetnames else None
    if ws is None:
        wb.close()
        return {}, []
    rows = ws.iter_rows(values_only=True)
    try:
        hdr = [_txt(c) for c in next(rows)]
    except StopIteration:
        wb.close()
        return {}, []
    ix = {h: i for i, h in enumerate(hdr) if h}
    out = [r for r in rows if r and any(c is not None for c in r)]
    wb.close()
    return ix, out


def build(base=None, master=None):
    """One row per asset the job took on the plant account.

    Returns {'rows': [...], 'stats': {...}}. Empty and harmless when the
    exports are not there.
    """
    base = base or HERE
    txn = _find("TRANSACTIONS*.xlsx", base)
    stk = _find("STOCKTAKE*.xlsx", base)
    onh = _find("ON_HIRE*.xlsx", base)
    if not txn:
        return {"rows": [], "stats": {"note": "no TRANSACTIONS export"}}

    plant_on, first_out, came_back, desc, fam = {}, {}, {}, {}, {}
    #  every asset the transaction feed mentions at all, so a departure
    #  can be reported for gear that never touched the plant account
    seen_any = set()
    n_rows = 0
    for sh in TXN_SHEETS:
        ix, rs = _sheet(txn, sh)
        if not ix:
            continue
        c_it = ix.get("SKU/ITEM_NUMBER", ix.get("ITEM_NUMBER"))
        c_wh = ix.get("HIRER_NAME")
        c_ds = ix.get("SKU/ITEM DESCRIPTION", ix.get("ITEM_DESCRIPTION"))
        c_fm = ix.get("PRODUCT_FAMILY")
        c_s = ix.get("TRAN_START_DATE")
        c_e = ix.get("TRAN_END_DATE")
        if c_it is None or c_wh is None:
            continue
        for r in rs:
            item = _txt(r[c_it])
            if not item:
                continue
            n_rows += 1
            seen_any.add(item)
            who = _txt(r[c_wh])
            start = _date(r[c_s]) if c_s is not None else None
            end = _date(r[c_e]) if c_e is not None else None
            if c_ds is not None and item not in desc:
                desc[item] = _txt(r[c_ds])
            if c_fm is not None and item not in fam:
                fam[item] = _txt(r[c_fm])
            if _is_plant(who):
                #  EARLIEST wins: the day the job first took it.
                if start and (item not in plant_on or start < plant_on[item]):
                    plant_on[item] = start
            elif who:
                if start and (item not in first_out
                              or start < first_out[item][0]):
                    first_out[item] = (start, who)
                #  LATEST end wins: the most recent time it came back.
                if end and (item not in came_back or end > came_back[item][0]):
                    came_back[item] = (end, who)

    #  currently still on the plant account - so a row can say "still
    #  on the job" rather than leave the last column blank and let a
    #  reader assume it has gone
    still = set()
    if onh:
        ix, rs = _sheet(onh, "ON_HIRE")
        c_it = ix.get("ITEM_NUMBER")
        c_bc = ix.get("HIRER_BARCODE")
        c_wh = ix.get("HIRER_NAME")
        c_dt = ix.get("ON_HIRE_DATE")
        for r in rs:
            if c_it is None:
                break
            item = _txt(r[c_it])
            bc = _txt(r[c_bc]) if c_bc is not None else ""
            who = _txt(r[c_wh]) if c_wh is not None else ""
            if item and (bc.upper() == PLANT_BARCODE or _is_plant(who)):
                still.add(item)
                d0 = _date(r[c_dt]) if c_dt is not None else None
                #  ON_HIRE carries the live start; the transaction feed
                #  only goes back as far as its reporting period, so
                #  this fills the gap for anything still out.
                if d0 and (item not in plant_on or d0 < plant_on[item]):
                    plant_on[item] = d0

    departed, sighted, stk_desc, stk_unit = {}, {}, {}, {}
    if stk:
        ix, rs = _sheet(stk, "STOCKTAKE")
        c_it = ix.get("SKU_NUMBER", ix.get("ITEM_OR_CONSUMABLE"))
        c_ac = ix.get("LAST_SIGHTED_ACTION")
        c_dt = ix.get("LAST_SIGHTED_DATE_TIME")
        c_by = ix.get("LAST_SIGHTED_BY")
        for r in rs:
            if c_it is None or c_ac is None:
                break
            item = _txt(r[c_it])
            if not item:
                continue
            act = _txt(r[c_ac])
            when = _stamp(r[c_dt]) if c_dt is not None else None
            by = _txt(r[c_by]) if c_by is not None else ""
            #  the stocktake names the gear too, and for anything that
            #  never touched the transaction feed it is the ONLY name
            #  there is - 82 departed assets came out blank without it
            c_ds = ix.get("DESCRIPTION")
            c_un = ix.get("STORAGE_UNIT")
            if c_ds is not None and item not in stk_desc:
                stk_desc[item] = _txt(r[c_ds])
            if c_un is not None and item not in stk_unit:
                stk_unit[item] = _txt(r[c_un])
            if act.lower().startswith("depart"):
                departed[item] = (when, by)
            elif when:
                sighted[item] = (when, act, by)

    #  ---- IS IT REALLY GONE? ------------------------------------------
    #  Andrew, 5 Aug 2026, on the store search saying nothing about
    #  departed gear. Fixing that turned up the mirror image here.
    #
    #  The rule above is "any LAST_SIGHTED_ACTION starting with depart".
    #  That is too loose on its own: on 5 Aug it called six assets
    #  departed that were still sitting on the register, because gear
    #  that goes out and comes back keeps a departure stamp behind it.
    #  Telling a storeman a machine went home when it is on the shelf is
    #  the same class of wrong as saying nothing when it really has.
    #
    #  departures.py already settles this properly - it checks the
    #  status AND the action, and it compares the departure time against
    #  when the register was pulled, because the two exports are taken
    #  hours apart. So ask it rather than keep a second, looser copy of
    #  the rule here that can drift away from it.
    #
    #  Best effort: if it cannot answer, nothing is dropped and this
    #  behaves exactly as it always did.
    try:
        import departures as _dep
        #  include_hidden: a held line's LIFECYCLE is still a fact.
        _live = _dep.load(on_register=_dep.on_register_keys(),
                          include_hidden=True)
        if _live:
            _dropped = [k for k in departed if _dep._key(k) not in _live]
            for k in _dropped:
                del departed[k]
            if _dropped:
                #  Say what they ARE, not just what they are not. Every
                #  one of these is a consumable whose status now reads
                #  In Stock or Stock Low - it carries an old departure
                #  stamp and is back on the shelf.
                print('  Plant life: {} line(s) carry an old departure stamp '
                      'but their status'.format(len(_dropped)))
                print('              now says they are back in stock - not '
                      'called departed.')
                print('              ({}{})'.format(
                    ', '.join(sorted(_dropped)[:5]),
                    ' ...' if len(_dropped) > 5 else ''))
    except Exception:
        pass

    def iso(d):
        return d.isoformat() if d else ""

    #  THE PLANT ACCOUNT IS NO LONGER THE GATE (Andrew, 3 Aug 2026:
    #  "the plant id is not needed"). Anything that has DEPARTED gets a
    #  row whether or not it ever sat on the plant barcode - a departure
    #  is the end of that asset's life on this job and the record should
    #  not depend on whose account it happened to be on. The plant ones
    #  are still marked, so they can still be read on their own.
    population = set(plant_on) | set(departed)
    rows = []
    for item in sorted(population):
        fo = first_out.get(item)
        cb = came_back.get(item)
        dp = departed.get(item)
        name = desc.get(item, "") or stk_desc.get(item, "")
        if master is not None:
            try:
                name = master.disp(item, name) or name
            except Exception:
                pass
        #  WHERE IT IS UP TO. Read backwards - the furthest thing that
        #  has happened is the stage it is at.
        if dp:
            stage = "departed"
        elif item in still:
            stage = "onsite"
        elif cb:
            stage = "back"
        elif fo:
            stage = "out"
        else:
            stage = "onsite"
        rows.append({
            "i": item,
            "pl": 1 if item in plant_on else 0,
            "n": name,
            "f": fam.get(item, "") or stk_unit.get(item, ""),
            "p": iso(plant_on.get(item)),
            "o": iso(fo[0]) if fo else "",
            "ow": (fo[1][:30] if fo else ""),
            "b": iso(cb[0]) if cb else "",
            "bw": (cb[1][:30] if cb else ""),
            #  the clock only when the export actually carried one
            "d": (dp[0].strftime("%Y-%m-%d %H:%M")
                  if dp and isinstance(dp[0], dt.datetime)
                  else (dp[0].isoformat() if dp and dp[0] else "")),
            "dw": (dp[1][:26] if dp else ""),
            "s": stage,
            #  days it sat on the account before anybody took it - the
            #  number that answers "did we need it this early"
            "w": ((fo[0] - plant_on[item]).days
                  if (fo and item in plant_on) else None),
        })

    #  newest arrivals first, then the ones that never went out, so the
    #  page opens on the question worth asking
    rows.sort(key=lambda r: (r["o"] != "", r["p"]), reverse=True)

    plants = [r for r in rows if r["pl"]]
    never = [r for r in plants if not r["o"]]
    waits = [r["w"] for r in rows if r["w"] is not None]
    stats = {
        "total": len(rows),
        "plant": len(plants),
        "offPlant": len(rows) - len(plants),
        #  "never went out" only means anything for gear the job took on
        #  the account - an asset that only appears here because it
        #  departed was never on the account to sit on it
        "everOut": sum(1 for r in plants if r["o"]),
        "never": len(never),
        "back": sum(1 for r in rows if r["b"]),
        "departed": sum(1 for r in rows if r["d"]),
        "still": len(still),
        "avgWait": round(sum(waits) / float(len(waits)), 1) if waits else None,
        "txnRows": n_rows,
        "haveStocktake": bool(stk),
    }
    return {"rows": rows, "stats": stats}


if __name__ == "__main__":
    import json
    r = build()
    print(json.dumps(r["stats"], indent=2))
    for x in r["rows"][:12]:
        print("  {i:<10} {p:<11} {o:<11} {b:<11} {d:<17} {s:<9} {n}"
              .format(**x))
