#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | MISSING PICTURES - the photo gap list, as a spreadsheet
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 31 Jul 2026): "a script to find missing picture then
#  goes into a excel so then I know what to look for and how it need
#  to be attached".
#
#  56_PHOTO_HUNT is the clickable browser version of this hunt; this
#  is the same truth as an Excel you can sort, filter, print and tick
#  off on a clipboard. One row per picture the register still needs:
#  what the thing is, which aisle it lives in, how many items hide
#  behind the one photo, the EXACT filename to save it as, and a
#  search link. A second sheet lists what's already covered - and by
#  which file - so the radio/gas near-enough names can be seen doing
#  their job.
#
#  Counting "covered" uses the SAME matching as the build itself
#  (exact name, near-enough name, and the radio/gas word match), so
#  this list never asks for a photo the build already has.
# =====================================================================
import datetime as dt
import os
import re
import sys

import mygear_thumbs

HERE = os.path.dirname(os.path.abspath(__file__))

ORANGE = 'F26222'
INK = '1D1D1B'
GREY = '8A94A2'


def main():
    print('=' * 66)
    print(' COATES | MISSING PICTURES - the photo gap list')
    print('=' * 66)
    reg = mygear_thumbs.variant_register(HERE)
    if not reg:
        print(' No RENTAL_STOCK / SALES_STOCK exports found - run the')
        print(' morning downloads first.')
        return 1
    safe = mygear_thumbs.safe_name
    #  the same claim logic the build uses - exact, near-enough, and
    #  the radio/gas word match - so DONE here means DONE on the page
    claimed = mygear_thumbs.alias_photos(
        dict(mygear_thumbs._photo_files(HERE)), reg.keys(),
        loose={c: v.get('n', '') for c, v in reg.items() if v.get('drv')})

    #  "covered" is not "on the page": the page shows the SHRUNK thumb
    #  out of Gear_Lookup\thumbs, built by 04. A photo sitting in
    #  Photos\ that hasn't shrunk yet still shows a two-letter tile -
    #  so this list says which it is, honestly ("some pictures dont
    #  show and it says i have all images", 31 Jul 2026).
    tdir = os.path.join(HERE, 'Gear_Lookup', 'thumbs')

    #  ------------------------------------------------------------------
    #  COUNT WHAT THE PAGE READS, WHICH IS THE THUMBS FOLDER.
    #
    #  (Andrew, 4 Aug 2026: "these are in the thumbs location with the
    #  rest and your stating they are missing - all thumbs are in there.")
    #  He was right and this report had the same fault the build's own
    #  counter had on 3 Aug: it measured coverage off Photos\ , the
    #  SOURCE folder. A machine whose thumbnails arrived some other way -
    #  copied from the other laptop, restored, handed over - has a full
    #  thumbs folder and a thin Photos folder, and this called every one
    #  of them missing while the phone drew them perfectly.
    #
    #  The page reads Gear_Lookup\thumbs. So that is what decides
    #  COVERED, and a source photo not yet shrunk is the only other way
    #  to be covered (it becomes a thumbnail on the next 04).
    #  ------------------------------------------------------------------
    thumb_files = {}
    if os.path.isdir(tdir):
        for f in os.listdir(tdir):
            if f.lower().endswith('.jpg') and not f.startswith('_'):
                thumb_files[os.path.splitext(f)[0]] = os.path.join(tdir, f)
    #  the SAME claim rules the build uses, run over the thumbnails, so a
    #  family picture is seen covering its whole family here too
    thumb_claim = mygear_thumbs.alias_photos(
        dict(thumb_files), reg.keys(),
        loose={c: v.get('n', '') for c, v in reg.items() if v.get('drv')})
    used = set()

    missing, covered = [], []
    unshrunk = 0
    for code, e in reg.items():
        row = {'code': code, 'n': e['n'] or code, 'f': e['f'] or '',
               'q': e['q'],
               'k': 'Consumable' if e['k'] == 'cons' else
                    ('Radio / gas fleet' if e.get('drv') else 'Hire gear')}
        #  on the page NOW (a thumbnail claims it), or on the page after
        #  the next 04 (a source photo claims it but has not shrunk yet)
        t_hit = thumb_claim.get(safe(code))
        p_hit = claimed.get(safe(code))
        if t_hit is None and p_hit is None:
            missing.append(row)
        else:
            if t_hit is not None:
                #  the STEM, not the filename - thumb_files is keyed by
                #  stem, and comparing the two made every thumbnail look
                #  like an orphan
                used.add(os.path.splitext(os.path.basename(t_hit))[0])
            row['by'] = os.path.basename(t_hit or p_hit)
            row['sh'] = t_hit is not None
            if not row['sh']:
                unshrunk += 1
            covered.append(row)
    #  ---- and the other half of the question ---------------------------
    #  A PICTURE THAT ANSWERS TO NOTHING. He has 1,081 thumbnails and
    #  items still saying NO PHOTO YET - both can be true at once, if a
    #  file is named after what the thing IS rather than after its code.
    #  Nothing has ever told him which ones those are, so the work went
    #  in and the picture never appeared. This is that list.
    orphans = sorted(n for n in thumb_files if n not in used)
    #  BUT A SPARE COPY IS NOT A GAP. Two of the thumbnails on this
    #  machine are the radio and the gas monitor saved under their full
    #  product names as well as under their codes. Nothing claims the
    #  long-named copy - but the item IS on the page, drawn by its
    #  code-named twin. Calling that "will never appear" sends a bloke
    #  hunting a problem that is not there, which is its own kind of
    #  lie. So each one is asked the only question that matters: is the
    #  thing it is a picture of on the page anyway?
    #  compare with the punctuation OUT, the way the matcher does -
    #  MOTOROLA_DP4801E_TWO_WAY_RADIO_WITH_HANDPIECE and
    #  MOTOROLADP4801ETWOWAYRADIO are the same words with underscores
    #  in between, and safe() keeps those
    def _flat(x):
        return re.sub(r'[^A-Z0-9]+', '', str(x or '').upper())
    _covered_codes = set(_flat(m['code']) for m in covered)

    def _spare(stem):
        u = _flat(stem)
        for c in _covered_codes:
            if c and (c in u or u in c):
                return True
        return False
    orphan_rows = [{'f': n, 'spare': _spare(n)} for n in orphans]
    dead = [o for o in orphan_rows if not o['spare']]
    #  biggest wins first: the photo that pictures 24 bollards beats
    #  the one that pictures a single spanner
    missing.sort(key=lambda r: (-r['q'], r['n'].upper()))
    covered.sort(key=lambda r: r['n'].upper())

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    thin = Border(bottom=Side(style='thin', color='DDDDDD'))
    h_font = Font(bold=True, color='FFFFFF', size=11)
    h_fill = PatternFill('solid', fgColor=ORANGE)

    def brand(ws, title, sub):
        ws['A1'] = 'COATES | ' + title
        ws['A1'].font = Font(bold=True, size=16, color=INK)
        ws['A2'] = ('Cement Australia K2 Shutdown 2026 - Gladstone | '
                    'POWERED BY SITEIQ | Author: Andrew Fisher')
        ws['A2'].font = Font(size=9, color=GREY)
        ws['A3'] = sub
        ws['A3'].font = Font(size=10, italic=True, color=INK)

    def header(ws, r, cols):
        for i, (label, width) in enumerate(cols, start=1):
            c = ws.cell(row=r, column=i, value=label)
            c.font = h_font
            c.fill = h_fill
            c.alignment = Alignment(vertical='center')
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = ws.cell(row=r + 1, column=1)
        ws.auto_filter.ref = '{}{}:{}{}'.format(
            'A', r, get_column_letter(len(cols)), r)

    #  ---- sheet 1: the gaps ------------------------------------------
    ws = wb.active
    ws.title = 'MISSING ({})'.format(len(missing))
    brand(ws, 'MISSING PICTURES',
          '{} pictures still wanted, {} already covered - built {}'.format(
              len(missing), len(covered),
              dt.datetime.now().strftime('%d %b %Y %H:%M')))
    ws['A5'] = ('HOW TO ATTACH A PICTURE: 1) find the row, 2) get a photo - '
                'a phone shot of the gear on the shelf is perfect, or use the '
                'search link - 3) save it into the Photos folder with the '
                'SAVE AS name (close counts: spaces and underscores are fine, '
                'extra words on the end are fine), 4) run 04 and it appears '
                'in My Gear everywhere that item shows.')
    ws['A5'].font = Font(size=10, color=INK)
    ws['A5'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('A5:G5')
    ws.row_dimensions[5].height = 42
    HR = 7
    header(ws, HR, [('WHAT IT IS', 52), ('AISLE / FAMILY', 20),
                    ('ITEMS BEHIND IT', 16), ('KIND', 16),
                    ('SAVE THE PHOTO AS', 42), ('SEARCH', 16),
                    ('DONE?', 9)])
    try:
        from urllib.parse import quote_plus
    except ImportError:
        from urllib import quote_plus
    r = HR + 1
    for m in missing:
        ws.cell(row=r, column=1, value=m['n']).border = thin
        ws.cell(row=r, column=2, value=m['f']).border = thin
        qc = ws.cell(row=r, column=3, value=m['q'])
        qc.border = thin
        if m['q'] >= 10:
            qc.font = Font(bold=True, color=ORANGE)
        ws.cell(row=r, column=4, value=m['k']).border = thin
        fc = ws.cell(row=r, column=5, value=safe(m['code']) + '.jpg')
        fc.font = Font(name='Consolas', size=10)
        fc.border = thin
        lc = ws.cell(row=r, column=6, value='Google Images')
        lc.hyperlink = ('https://www.google.com/search?tbm=isch&q='
                        + quote_plus(m['n']))
        lc.font = Font(color='0563C1', underline='single')
        lc.border = thin
        ws.cell(row=r, column=7, value='').border = thin
        r += 1

    #  ---- sheet 2: covered, and by which file ------------------------
    w2 = wb.create_sheet('COVERED ({})'.format(len(covered)))
    brand(w2, 'PICTURES ALREADY COVERED',
          'Every register code with a photo, the file covering it, and '
          'whether the PAGE is actually showing it yet.'
          + (' {} row(s) say RUN 04 - the photo is in the folder but not '
             'shrunk onto the page yet.'.format(unshrunk) if unshrunk
             else ''))
    header(w2, 5, [('WHAT IT IS', 52), ('AISLE / FAMILY', 20),
                   ('ITEMS BEHIND IT', 16), ('KIND', 16),
                   ('COVERED BY (file in Photos)', 48),
                   ('ON THE PAGE?', 22)])
    #  the not-on-the-page rows first, so the problem is at the top
    covered.sort(key=lambda m: (m['sh'], m['n'].upper()))
    r = 6
    for m in covered:
        w2.cell(row=r, column=1, value=m['n']).border = thin
        w2.cell(row=r, column=2, value=m['f']).border = thin
        w2.cell(row=r, column=3, value=m['q']).border = thin
        w2.cell(row=r, column=4, value=m['k']).border = thin
        bc = w2.cell(row=r, column=5, value=m['by'])
        bc.font = Font(name='Consolas', size=10)
        bc.border = thin
        sc = w2.cell(row=r, column=6,
                     value='YES' if m['sh'] else 'RUN 04 - not shrunk yet')
        sc.font = (Font(bold=True, color='2BB673') if m['sh']
                   else Font(bold=True, color='E23B2E'))
        sc.border = thin
        r += 1

    #  ---- sheet 3: pictures that answer to nothing --------------------
    if orphans:
        w3 = wb.create_sheet('NOT LANDING ({})'.format(len(dead)))
        brand(w3, 'PICTURES THAT ANSWER TO NOTHING',
              '{} thumbnail(s) no register code claims. {} of them are '
              'SPARE COPIES - the thing IS on the page, drawn by another '
              'file - and {} are doing no work at all: the file is named '
              'after what the thing IS, not after its code.'.format(
                  len(orphans), len(orphans) - len(dead), len(dead)))
        w3['A5'] = ('TO FIX ONE: find what it is a picture of in the MISSING '
                    'sheet, copy the SAVE AS name from there, and rename this '
                    'file to it. Then run 04. Or leave it - nothing is broken, '
                    'it is just a photo doing no work.')
        w3['A5'].font = Font(size=10, color=INK)
        w3.merge_cells('A5:C5')
        w3.row_dimensions[5].height = 46
        w3['A5'].alignment = Alignment(wrap_text=True, vertical='top')
        header(w3, 7, [('FILE IN thumbs\\', 56),
                       ('WHAT IT LOOKS LIKE IT MIGHT BE', 46),
                       ('DOES IT MATTER?', 40)])
        r = 8
        #  the ones doing no work first - the spares are just noise
        orphan_rows.sort(key=lambda o: (o['spare'], o['f'].upper()))
        for o in orphan_rows:
            n = o['f']
            c = w3.cell(row=r, column=1, value=n + '.jpg')
            c.font = Font(name='Consolas', size=10)
            c.border = thin
            #  a plain-English guess, from the register, so he is not
            #  reading raw codes to work out what a file was meant to be
            guess = ''
            up = safe(n).upper()
            for code, e in reg.items():
                nm = (e['n'] or '').upper()
                if nm and (safe(code).upper() in up
                           or up[:10] in nm.replace(' ', '')):
                    guess = e['n']
                    break
            w3.cell(row=r, column=2, value=guess or '-').border = thin
            #  worded so neither line can mislead. "Rename it and it
            #  appears" would be a promise this cannot keep - the file
            #  might be a second picture of something already covered
            #  under a different product name, and renaming it would
            #  only make a duplicate.
            sc = w3.cell(row=r, column=3, value=(
                'No - the same thing is on the page under another name'
                if o['spare']
                else 'Nothing claims this file - check MISSING for the '
                     'name to save it as'))
            sc.font = (Font(color=GREY) if o['spare']
                       else Font(bold=True, color='E23B2E'))
            sc.border = thin
            r += 1

    out = os.path.join(HERE, 'Missing_Pictures_{}.xlsx'.format(
        dt.date.today().isoformat()))
    wb.save(out)
    mygear_thumbs.photos_dir(HERE)
    print(' Register codes            : {}'.format(len(reg)))
    print(' ON THE PAGE right now     : {}'.format(
        len(covered) - unshrunk))
    print(' Still missing a picture   : {}'.format(len(missing)))
    if unshrunk:
        print(' In Photos, not shrunk yet : {}  <-- run 04'.format(unshrunk))
    if dead:
        print(' Photos doing no work      : {}  <-- see NOT LANDING; a'
              .format(len(dead)))
        print(' (named wrong, never shown)      rename fixes each one')
    if len(orphans) - len(dead):
        print(' Spare copies (harmless)   : {}'.format(
            len(orphans) - len(dead)))
    print('')
    if not missing and not dead:
        print(' EVERY register code has a picture on the page. Nothing')
        print(' is waiting and nothing is going to waste.')
    elif not missing:
        print(' Every register code has a picture on the page. The')
        print(' NOT LANDING list is spare work, not a gap.')
    else:
        print(' {} of {} register codes have a picture on the page.'.format(
            len(covered) - unshrunk, len(reg)))
        print(' The MISSING sheet names the other {} - one row each,'.format(
            len(missing)))
        print(' biggest wins first, with the exact filename to save as.')
    print('')
    print(' Written to : {}'.format(out))
    print('')
    print(' Sort or filter the MISSING sheet, snap the photos, save them')
    print(' into Photos\\ under the SAVE AS names, run 04. Run me again')
    print(' any time - the list re-counts itself.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
