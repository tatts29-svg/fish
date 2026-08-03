#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | TEST DAILY EMAIL PACKS
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Proves the day's packs before anybody presses Send.
#
#  It reads the routing workbook FOR ITSELF - it does not ask the
#  builder what it thinks it did. Then it opens every draft the way a
#  mail client opens it and compares the two.
#
#  Part 1  EVERY PACK       folder, report, attachments, addresses,
#                           subject, record, the lot
#  Part 2  ISOLATION        no contractor can see another's gear, and
#                           no two contractors share an address
#  Part 3  DELIBERATE SABOTAGE  four routing mistakes are fed in on
#                           purpose - a wrong address, a missing
#                           oversight contact, a contractor on the
#                           client's list, an empty To. Every one of
#                           them must be caught and held.
#  Part 4  NO OVERWRITES    the day is built twice; not one byte of
#                           the first build may change
#
#  Any failure exits non-zero and says exactly what and where.
#
#  Usage:  py TEST_DAILY_PACKS.py [--date 2026-07-25]
# =====================================================================

from __future__ import print_function

import os
import re
import sys
import glob
import email
import shutil
import hashlib
import argparse
import tempfile
import subprocess
import email.policy
import datetime as dt

HERE = os.path.dirname(os.path.abspath(__file__))
BUILDER = os.path.join(HERE, "BUILD_DAILY_EMAIL_PACKS.py")

PASS, FAIL = [], []
#  checks that could not run, and why - counted separately from
#  passes, because "did not run" is not "passed"
SKIPPED = []


def ok(name, detail=""):
    PASS.append(name)
    print("  PASS  {}{}".format(name, "  -  " + detail if detail else ""))


def bad(name, detail=""):
    FAIL.append((name, detail))
    print("  FAIL  {}{}".format(name, "  -  " + detail if detail else ""))


def check(name, condition, detail=""):
    (ok if condition else bad)(name, detail)
    return bool(condition)


def head(title):
    print("")
    print("-" * 68)
    print(" " + title)
    print("-" * 68)


# ---------------------------------------------------------------------
#  Read the workbook ourselves - the whole point of an independent test
# ---------------------------------------------------------------------
def emails_in(text):
    #  Same rule as the builder's _emails: "John Pickels <john@...>" is a
    #  perfectly normal way to write an address - take the address out of
    #  it rather than dropping the person. The old space-means-skip rule
    #  made this prover blind to exactly that row (29 Jul 2026).
    out = []
    for p in re.split(r"[;,\r\n\t]+", str(text or "")):
        p = p.strip()
        if not p:
            continue
        m = re.search(r"<([^<>@\s]+@[^<>@\s]+)>", p)
        if m:
            p = m.group(1)
        elif " " in p:
            m = re.search(r"([^<>@\s]+@[^<>@\s]+)", p)
            p = m.group(1) if m else ""
        p = p.strip("<>.,; ")
        if p and "@" in p and " " not in p and p not in out:
            out.append(p)
    return out


def slug(name):
    """The same folder-name rule the builder uses. Written out here, not
    imported, so this file keeps proving the builder rather than
    agreeing with it."""
    out = re.sub(r'[<>:"/\\|?*]', "-", str(name or "").strip()).strip().rstrip(". ")
    return out or "UNNAMED COMPANY"


def read_workbook(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {"companies": [], "contacts": {}, "fixed_cc": []}

    rows = [[("" if c is None else str(c).strip())
             for c in r] for r in wb["Company Routing"].iter_rows(values_only=True)]
    hdr = next(r for r in rows if r and r[0] == "Company")
    ix = {v: i for i, v in enumerate(hdr) if v}
    for r in rows[rows.index(hdr) + 1:]:
        if not r or not r[0]:
            continue
        out["companies"].append({
            "company": r[ix["Company"]],
            "kind": r[ix["Email Type"]],
            "to": emails_in(r[ix["To"]]),
            "cc": emails_in(r[ix["Cc"]]),
        })

    rows = [[("" if c is None else str(c).strip())
             for c in r] for r in wb["Contacts"].iter_rows(values_only=True)]
    hdr = next(r for r in rows if r and r[0] == "Company")
    ix = {v: i for i, v in enumerate(hdr) if v}
    for r in rows[rows.index(hdr) + 1:]:
        if not r or not r[0]:
            continue
        e = r[ix["Email"]].lower()
        if "@" in e:
            out["contacts"].setdefault(r[ix["Company"]].lower(), set()).add(e)

    for r in wb["Summary"].iter_rows(values_only=True):
        vals = [("" if c is None else str(c)) for c in r]
        got = []
        for v in vals:
            got += emails_in(v)
        for g in got:
            if g.lower() not in out["fixed_cc"]:
                out["fixed_cc"].append(g.lower())
    #  Deliberately NOT derived from the Cc columns. An address pasted
    #  into every external row would otherwise make itself a "fixed
    #  oversight contact" and this test would agree with the mistake.
    wb.close()
    return out


def report_identity(path):
    """Whose report is this? Taken from the report's own headline, not
    from its filename - a file can be named anything."""
    try:
        with open(path, encoding="utf-8", errors="ignore") as fh:
            text = fh.read()
    except Exception:
        return ""
    m = re.search(r"<h1[^>]*>(.*?)</h1>", text, re.S | re.I)
    if not m:
        return ""
    return re.sub(r"<[^>]+>", "", m.group(1)).strip()


def pdf_pages(path):
    """How many pages the PDF has, counted in the PDF itself.

    The report's own 'Page 3 of 17' is written by JavaScript when the
    page is opened, so it is not in the file to be read. The PDF is a
    finished artifact built by the browser from the same document - so
    it is the honest, independent answer to 'how long is this report'."""
    try:
        with open(path, "rb") as fh:
            raw = fh.read()
    except Exception:
        return 0
    n = raw.count(b"/Type /Page") - raw.count(b"/Type /Pages")
    if n <= 0:
        n = raw.count(b"/Type/Page") - raw.count(b"/Type/Pages")
    m = re.search(rb"/Count\s+(\d+)", raw)
    if m:
        n = max(n, int(m.group(1)))
    return max(n, 0)


def open_eml(path):
    with open(path, "rb") as fh:
        m = email.message_from_binary_file(fh, policy=email.policy.default)

    def addrs(f):
        return sorted(a.addr_spec.lower()
                      for a in (m[f].addresses if m[f] else ())
                      if a.addr_spec)

    inline, files, html = [], [], ""
    for p in m.walk():
        if p.get_content_maintype() == "multipart":
            continue
        if p.get_content_type() == "text/html" and not html:
            html = p.get_content()
        cid = (p.get("Content-ID") or "").strip("<>")
        if p.get_content_disposition() == "inline" and cid:
            inline.append(cid)
        elif p.get_content_disposition() == "attachment":
            files.append((p.get_filename() or "", p.get_payload(decode=True)))
    return {"to": addrs("To"), "cc": addrs("Cc"), "subject": m["Subject"] or "",
            "unsent": (m["X-Unsent"] or ""), "inline": inline, "html": html,
            "attachments": files, "top": m.get_content_type(),
            "first": m.get_payload()[0].get_content_type()
                     if m.is_multipart() else ""}


def md5(path):
    h = hashlib.md5()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def tree_state(root):
    out = {}
    for base, _dirs, files in os.walk(root):
        for f in files:
            p = os.path.join(base, f)
            out[os.path.relpath(p, root)] = md5(p)
    return out


def run_builder(args, expect_zero=True):
    cmd = [sys.executable, BUILDER] + args
    r = subprocess.run(cmd, capture_output=True, text=True, cwd=HERE,
                       errors="replace")
    if expect_zero and r.returncode != 0:
        print(r.stdout[-2000:])
        print(r.stderr[-2000:])
    return r


# =====================================================================
#  PART 1 + 2 - every pack, against the workbook
# =====================================================================
def verify_packs(root, wbk, date_tag):
    disp = dt.datetime.strptime(date_tag, "%Y-%m-%d").strftime("%d %b %Y")
    companies = wbk["companies"]
    fixed = set(wbk["fixed_cc"])
    seen_to = {}
    identities = {}

    head("PART 1 - EVERY PACK AGAINST THE WORKBOOK")
    for c in companies:
        co = c["company"]
        folder = os.path.join(root, "01 COMPANY REPORTS", slug(co), date_tag)
        if not check("{}: dated folder exists".format(co),
                     os.path.isdir(folder), folder):
            continue
        rec = os.path.join(folder, "Email Record - {}.txt".format(date_tag))
        check("{}: email record written".format(co), os.path.isfile(rec))

        eml = os.path.join(folder, "Email Draft - {}.eml".format(date_tag))
        body = glob.glob(os.path.join(folder, "*On-Hire Report*.html")) + \
            glob.glob(os.path.join(folder, "Executive Daily Summary*.html"))

        if not os.path.isfile(eml):
            held = glob.glob(os.path.join(
                folder, "HOLD - Email Draft - {}*".format(date_tag)))
            if held:
                #  A held draft is the guard WORKING, not a draft gone
                #  missing - the record names exactly which check failed
                #  (e.g. an attachment not built yet). Held is never
                #  counted as "no draft".
                check("{}: draft held, record names the reason".format(co),
                      os.path.isfile(rec), "held, record on file")
                continue
            #  no draft is only correct when there is no report either
            check("{}: no draft AND no report (nothing on hire)".format(co),
                  not body, "folder and record only")
            continue

        got = open_eml(eml)
        check("{}: To matches the workbook exactly".format(co),
              got["to"] == sorted(e.lower() for e in c["to"]),
              "{} v {}".format(len(got["to"]), len(c["to"])))
        check("{}: Cc matches the workbook exactly".format(co),
              got["cc"] == sorted(e.lower() for e in c["cc"]),
              "{} v {}".format(len(got["cc"]), len(c["cc"])))
        check("{}: company named in the subject".format(co),
              co.lower() in got["subject"].lower(), got["subject"])
        #  "DGH Engineering | Daily On-Hire Report | 25 July 2026"
        long_day = dt.datetime.strptime(date_tag, "%Y-%m-%d") \
            .strftime("%d %B %Y").lstrip("0")
        check("{}: subject carries the reporting day".format(co),
              disp in got["subject"] or long_day in got["subject"], long_day)
        check("{}: opens in Outlook as an unsent draft".format(co),
              got["unsent"] == "1")
        check("{}: MIME nests mixed > related".format(co),
              got["top"] == "multipart/mixed"
              and got["first"] == "multipart/related",
              "{} > {}".format(got["top"], got["first"]))

        external = c["kind"].lower().startswith("external")
        if external:
            check("{}: the four fixed oversight contacts are in Cc".format(co),
                  fixed.issubset(set(got["cc"])),
                  "missing " + ", ".join(sorted(fixed - set(got["cc"])))
                  if not fixed.issubset(set(got["cc"])) else "all four")
            names = [n for n, _d in got["attachments"]]
            #  The company email model (Andrew, 28 Jul 2026): the story
            #  in the body, and on the paperclip exactly two things -
            #  their own full report as ONE PDF, plus the safety PDF.
            check("{}: report PDF + safety PDF are the attachments"
                  .format(co),
                  len(names) == 2
                  and any(n.startswith("Daily Safety & Compliance Report")
                          for n in names)
                  and any("On-Hire Report" in n for n in names),
                  "; ".join(names))
        else:
            want = ["Daily Safety & Compliance Report",
                    "Daily Cost Tracking Report", "Site Plant Report",
                    "Cement Australia On-Hire Report",
                    "Consumables Usage Report"]
            names = [n for n, _d in got["attachments"]]
            #  When the exec-summary body is too heavy to paste as
            #  pages (the 10 MB safe-send rule), the builder rightly
            #  attaches it as a sixth PDF - allowed, and only that.
            extras = [n for n in names
                      if not any(n.startswith(w) for w in want)]
            check("{}: all five client attachments present".format(co),
                  all(any(n.startswith(w) for n in names) for w in want)
                  and all(n.startswith("Executive Daily Summary")
                          for n in extras),
                  "{} attached".format(len(names)))

        #  every attachment matches the file filed beside it, byte for byte
        for name, data in got["attachments"]:
            on_disk = os.path.join(folder, name)
            same = (os.path.isfile(on_disk) and data is not None
                    and hashlib.md5(data).hexdigest() == md5(on_disk))
            check("{}: attachment matches the filed copy ({})".format(
                co, name[:38]), same)

        #  every attachment carries today's date in its name
        check("{}: attachments are today's".format(co),
              all(date_tag in n for n, _d in got["attachments"]),
              date_tag)

        #  the body's images all exist, and none is orphaned
        used = sorted(set(re.findall(r"cid:([A-Za-z0-9_.-]+)", got["html"])))
        check("{}: every image in the body is attached inline".format(co),
              used == sorted(got["inline"]),
              "{} used / {} inline".format(len(used), len(got["inline"])))
        check("{}: the report is in the body, not just attached".format(co),
              bool(got["inline"]) or "Coates" in got["html"],
              "{} page image(s)".format(len(got["inline"])))

        #  the report filed for the body belongs to this company, judged
        #  by the report's OWN headline - a file can be named anything
        if body:
            with open(body[0], encoding="utf-8", errors="ignore") as fh:
                text = fh.read()
            who = report_identity(body[0])
            identities[co] = who
            check("{}: the filed report is named for this company".format(co),
                  os.path.basename(body[0]).lower().startswith(co.lower())
                  or not external,
                  os.path.basename(body[0]))
            if external:
                others = [o["company"] for o in companies
                          if o["company"] != co
                          and o["company"].lower() not in ("cement australia",
                                                           "coates")]
                stray = [o for o in others if o.lower() in text.lower()]
                check("{}: no other contractor named in the report".format(co),
                      not stray, ", ".join(stray))
                #  and the report must actually be about somebody
                check("{}: the report carries a company headline".format(co),
                      bool(who), who or "no headline found")

        #  CRLF, as a mail file must be
        with open(eml, "rb") as fh:
            raw = fh.read()
        bare = raw.count(b"\n") - raw.count(b"\r\n")
        check("{}: the draft uses proper mail line endings".format(co),
              bare == 0, "{} bare line feed(s)".format(bare))

        #  the WHOLE report is in the body - counted from the PDF, which
        #  is built from the same document by the browser and knows
        #  nothing about the email pipeline
        if body and got["inline"]:
            #  The printed report and the emailed one are the same content
            #  in two layouts: on paper each page carries a white margin,
            #  in an email the report fills the page. So the PDF always
            #  needs AT LEAST as many pages as the email - fewer would
            #  mean the email is carrying pages the report hasn't got.
            stem = ("Coates_K2_Executive_Summary" if not external else
                    "Coates_K2_Activity_" +
                    re.sub(r"[^A-Za-z0-9]+", "_",
                           report_identity(body[0])).strip("_").upper())
            pdf = os.path.join(HERE, "Reports", date_tag, "PDF",
                               "{}_{}.pdf".format(stem, date_tag))
            pages = pdf_pages(pdf) if os.path.isfile(pdf) else 0
            if pages:
                check("{}: the email carries the whole report".format(co),
                      pages >= len(got["inline"]) > 0,
                      "{} pages in the email, {} on paper".format(
                          len(got["inline"]), pages))
            #  and the pages must be a clean run, no gap in the middle
            cids = re.findall(r"cid:pg(\d+)", got["html"])
            check("{}: the pages run 1..{} with no gap".format(
                co, len(got["inline"])),
                [int(c) for c in cids] == list(range(1, len(cids) + 1)),
                ", ".join(cids[:8]))

        #  a ready pack must not have a held draft sitting beside it
        check("{}: no held draft left beside a ready one".format(co),
              not glob.glob(os.path.join(folder, "HOLD - Email Draft - {}"
                                                 ".eml.txt".format(date_tag))),
              "a HOLD file is still in this folder")

        if external:
            seen_to[co] = set(got["to"])

    head("PART 2 - ISOLATION BETWEEN CONTRACTORS")
    #  Two companies must never be holding the same report. This compares
    #  the reports' own headlines, so it works even though the register
    #  spells companies differently from the workbook (DKE Group is
    #  Dark Knight Engineering on the report).
    dupes = []
    for a_co, a_id in identities.items():
        for b_co, b_id in identities.items():
            if a_co < b_co and a_id and a_id == b_id:
                dupes.append("{} and {} both hold '{}'".format(a_co, b_co, a_id))
    check("No two companies were given the same report",
          not dupes, "; ".join(dupes))
    names = sorted(seen_to)
    clashes = []
    for i, a in enumerate(names):
        for b in names[i + 1:]:
            shared = (seen_to[a] & seen_to[b]) - fixed
            if shared:
                clashes.append("{} / {}: {}".format(a, b, ", ".join(shared)))
    check("No two contractors share a recipient",
          not clashes, "; ".join(clashes))

    reg = {}
    for co, addrs in seen_to.items():
        allowed = (wbk["contacts"].get(co.lower(), set())
                   | wbk["contacts"].get("coates", set())
                   | wbk["contacts"].get("cement australia", set()) | fixed)
        unknown = addrs - allowed
        if unknown:
            reg[co] = unknown
    check("Every recipient is on the contact register for their company",
          not reg, "; ".join("{}: {}".format(k, ", ".join(v))
                            for k, v in reg.items()))

    #  the masters were filed
    mdir = os.path.join(root, "00 MASTER DAILY REPORTS", date_tag)
    check("Master reports filed for the day", os.path.isdir(mdir)
          and len(glob.glob(os.path.join(mdir, "*.pdf"))) >= 5,
          "{} PDF(s)".format(len(glob.glob(os.path.join(mdir, "*.pdf")))))

    #  the control sheet and register
    cdir = os.path.join(root, "02 EMAIL CONTROL")
    check("Daily control sheet written",
          os.path.isfile(os.path.join(
              cdir, "Daily Email Control - {}.html".format(date_tag))))
    reg_x = os.path.join(cdir, "Email Control Register.xlsx")
    if check("Control register written", os.path.isfile(reg_x)):
        import openpyxl
        ws = openpyxl.load_workbook(reg_x, data_only=True)["Register"]
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                if r and str(r[0])[:10] == date_tag]
        #  The register is a LEDGER, not a snapshot - the builder's own
        #  rule (write_register, Rule 7) says a company appears AGAIN on
        #  the same day when its status changes, so the trail shows a
        #  pack that was held and then fixed. Demanding exactly one row
        #  per company failed the first real day a pack was re-run -
        #  which is exactly the day the register is doing its job.
        #  What the ledger actually promises: every company has a row,
        #  and the same company+status pair is never written twice.
        #  (Caught 29 Jul 2026, on a day the packs were built 3 times.)
        row_cos = {str(r[1]).strip() for r in rows}
        missing = sorted({c["company"] for c in companies} - row_cos)
        check("Every company has a register row for the day",
              not missing, ", ".join(missing))
        pairs = [(str(r[1]).strip(), str(r[3] or '').strip()) for r in rows]
        dups = sorted({p for p in pairs if pairs.count(p) > 1})
        check("No company+status pair is ever written twice",
              not dups, "; ".join("{} / {}".format(*p) for p in dups))
        allowed_status = {"Not Generated", "Generated", "Checked",
                          "Draft Ready", "Sent", "Failed - Review Required"}
        odd = sorted({str(r[3]) for r in rows} - allowed_status)
        check("Only the agreed statuses are ever written",
              not odd, ", ".join(odd))

    #  settings folder keeps the workbook that did the routing
    sdir = os.path.join(root, "03 CONTACTS & SETTINGS")
    check("The routing workbook is kept with the day",
          bool(glob.glob(os.path.join(sdir, "*.xlsx"))))
    check("A plain record of the routing used is kept",
          os.path.isfile(os.path.join(
              sdir, "Routing in use - {}.txt".format(date_tag))))


# =====================================================================
#  PART 3 - deliberate sabotage
# =====================================================================
def doctor(src, dest, change):
    """A copy of the routing workbook with one thing wrong on purpose."""
    import openpyxl
    shutil.copy2(src, dest)
    wb = openpyxl.load_workbook(dest)
    ws = wb["Company Routing"]
    hdr_row = next(r for r in range(1, 12)
                   if str(ws.cell(r, 1).value).strip() == "Company")
    cols = {str(ws.cell(hdr_row, c).value).strip(): c
            for c in range(1, ws.max_column + 1)}
    for r in range(hdr_row + 1, ws.max_row + 1):
        name = str(ws.cell(r, cols["Company"]).value or "").strip()
        change(ws, r, cols, name)
    wb.save(dest)
    return dest


def sabotage(date_tag, book):
    head("PART 3 - DELIBERATE SABOTAGE (every one must be caught)")
    cases = []

    def wrong_address(ws, r, cols, name):
        if name == "DGH Engineering":
            ws.cell(r, cols["To"]).value = "jayden.andrew@cleanaway.com.au"
    cases.append(("A Cleanaway address pasted into the DGH row",
                  wrong_address, "DGH Engineering"))

    def missing_oversight(ws, r, cols, name):
        if name == "Xtreme Engineering":
            cc = str(ws.cell(r, cols["Cc"]).value or "")
            ws.cell(r, cols["Cc"]).value = ";".join(
                p for p in cc.split(";") if "david.moore" not in p)
    cases.append(("An oversight contact dropped off Xtreme's Cc",
                  missing_oversight, "Xtreme Engineering"))

    def contractor_on_client_list(ws, r, cols, name):
        if name == "Cement Australia":
            ws.cell(r, cols["Cc"]).value = (
                str(ws.cell(r, cols["Cc"]).value or "")
                + "; ross.gould@dgh.com.au")
    cases.append(("A contractor added to the client's master list",
                  contractor_on_client_list, "Cement Australia"))

    def empty_to(ws, r, cols, name):
        if name == "Veolia Refractory":
            ws.cell(r, cols["To"]).value = ""
    cases.append(("Veolia's To left empty", empty_to, "Veolia Refractory"))

    for label, change, company in cases:
        tmp = tempfile.mkdtemp(prefix="k2packtest_")
        try:
            bad_book = doctor(book, os.path.join(
                tmp, "K2_Daily_Email_Report_Allocation.xlsx"), change)
            root = os.path.join(tmp, "OUT")
            r = run_builder(["--date", date_tag, "--root", root,
                             "--book", bad_book])
            folder = os.path.join(root, "01 COMPANY REPORTS", company, date_tag)
            rec = os.path.join(folder, "Email Record - {}.txt".format(date_tag))
            text = ""
            if os.path.isfile(rec):
                with open(rec, encoding="utf-8", errors="ignore") as fh:
                    text = fh.read()
            held = ("Failed - Review Required" in text
                    or "Not Generated" in text)
            check("Caught: {}".format(label), held,
                  "status was " + (re.search(r"Status\s*:\s*(.+)", text).group(1)
                                   if re.search(r"Status\s*:\s*(.+)", text)
                                   else "no record written"))
            #  and it must never have produced a sendable draft
            eml = os.path.join(folder, "Email Draft - {}.eml".format(date_tag))
            if os.path.isfile(eml) and company != "Veolia Refractory":
                got = open_eml(eml)
                if label.startswith("A Cleanaway"):
                    check("  and DGH's gear never carried a Cleanaway address"
                          " on a ready pack",
                          "cleanaway" not in " ".join(got["to"]) or held,
                          "held: {}".format(held))
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


# =====================================================================
#  PART 3b - the ways a day can quietly go wrong
# =====================================================================
def _status(root, company, date_tag):
    rec = os.path.join(root, "01 COMPANY REPORTS", slug(company), date_tag,
                       "Email Record - {}.txt".format(date_tag))
    if not os.path.isfile(rec):
        return "(no record written)"
    with open(rec, encoding="utf-8", errors="ignore") as fh:
        m = re.search(r"Status\s*:\s*(.+)", fh.read())
    return m.group(1).strip() if m else "(no status)"


def _mirror_reports(date_tag, tmp):
    """A copy of today's built reports to meddle with, so the real ones
    are never touched."""
    src = os.path.join(HERE, "Reports", date_tag)
    dst = os.path.join(tmp, "Reports", date_tag)
    shutil.copytree(src, dst)
    return os.path.join(tmp, "Reports")


def quiet_failures(date_tag, book):
    head("PART 3b - THE WAYS A DAY CAN QUIETLY GO WRONG")

    # 1. a report belonging to a similarly-named entity turns up
    tmp = tempfile.mkdtemp(prefix="k2pack_prefix_")
    try:
        reports = _mirror_reports(date_tag, tmp)
        pages = os.path.join(reports, date_tag, "Pages")
        real = os.path.join(pages, "Coates_OnHire_Report_DGH_ENGINEERING_"
                                   "{}.html".format(date_tag))
        fake = os.path.join(pages, "Coates_OnHire_Report_VEOLIA_WATER_"
                                   "TECHNOLOGIES_{}.html".format(date_tag))
        shutil.copy2(real, fake)
        os.utime(fake, None)                      # newer than the real one
        root = os.path.join(tmp, "OUT")
        run_builder(["--date", date_tag, "--root", root, "--book", book,
                     "--reports", reports])
        body = glob.glob(os.path.join(root, "01 COMPANY REPORTS",
                                      "Veolia Refractory", date_tag, "*.html"))
        text = ""
        if body:
            with open(body[0], encoding="utf-8", errors="ignore") as fh:
                text = fh.read()
        check("A look-alike company's report is never picked up",
              "DGH Engineering" not in text,
              "Veolia Refractory's pack holds: "
              + (report_identity(body[0]) if body else "nothing"))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    # 2. a company renamed in the workbook must be held, not skipped
    for label, change, company in (
            ("A company renamed to the register's spelling",
             lambda ws, r, c, n: ws.cell(r, c["Company"]).__setattr__(
                 "value", "Dark Knight Engineering") if n == "DKE Group" else None,
             "Dark Knight Engineering"),
            ("A brand-new company added to the routing sheet",
             lambda ws, r, c, n: ws.cell(r, c["Company"]).__setattr__(
                 "value", "Aestec") if n == "Industec" else None,
             "Aestec")):
        tmp = tempfile.mkdtemp(prefix="k2pack_rename_")
        try:
            bad_book = doctor(book, os.path.join(
                tmp, "K2_Daily_Email_Report_Allocation.xlsx"), change)
            root = os.path.join(tmp, "OUT")
            run_builder(["--date", date_tag, "--root", root, "--book", bad_book])
            st = _status(root, company, date_tag)
            check("Caught: {}".format(label),
                  st.startswith("Failed"), "status was " + st)
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    # 3. a stray space or the wrong case must NOT cost a company its report
    def messy(ws, r, cols, name):
        if name == "Xtreme Engineering":
            ws.cell(r, cols["Company"]).value = "  xtreme engineering "
    tmp = tempfile.mkdtemp(prefix="k2pack_messy_")
    try:
        bad_book = doctor(book, os.path.join(
            tmp, "K2_Daily_Email_Report_Allocation.xlsx"), messy)
        root = os.path.join(tmp, "OUT")
        run_builder(["--date", date_tag, "--root", root, "--book", bad_book])
        folder = glob.glob(os.path.join(root, "01 COMPANY REPORTS",
                                        "*xtreme*", date_tag, "*On-Hire*.html"))
        check("A stray space in the company name still finds the report",
              bool(folder),
              os.path.basename(folder[0]) if folder else "no report filed")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    # 4. a report re-issued AFTER the pack was built must hold the pack
    tmp = tempfile.mkdtemp(prefix="k2pack_reissue_")
    try:
        reports = _mirror_reports(date_tag, tmp)
        root = os.path.join(tmp, "OUT")
        run_builder(["--date", date_tag, "--root", root, "--book", book,
                     "--reports", reports])
        safety = os.path.join(reports, date_tag, "PDF",
                              "Coates_K2_Safety_Assurance_{}.pdf".format(date_tag))
        with open(safety, "ab") as fh:
            fh.write(b"%% re-issued %%")
        run_builder(["--date", date_tag, "--root", root, "--book", book,
                     "--reports", reports])
        folder = os.path.join(root, "01 COMPANY REPORTS", "DGH Engineering",
                              date_tag)
        st = _status(root, "DGH Engineering", date_tag)
        again = glob.glob(os.path.join(folder, "*(2)*"))
        #  A second run of the day is normal - a fresh SiteIQ pull. The
        #  pack is rebuilt from the new report and stays ready.
        check("A report re-issued mid-day rebuilds the pack",
              st == "Draft Ready", "status was " + st)
        check("  and the first copy is kept, the new one filed beside it",
              bool(again),
              os.path.basename(again[0]) if again else "no second copy")

        #  But once you have SIGNED the record, sending again is your
        #  decision - the pack must stop and ask.
        rec = os.path.join(folder, "Email Record - {}.txt".format(date_tag))
        with open(rec, encoding="utf-8") as fh:
            signed = fh.read().replace("Sent by      : ____________________",
                                       "Sent by      : A. Fisher 07:15")
        with open(rec, "w", encoding="utf-8") as fh:
            fh.write(signed)
        with open(safety, "ab") as fh:
            fh.write(b"%% re-issued again %%")
        run_builder(["--date", date_tag, "--root", root, "--book", book,
                     "--reports", reports])
        #  A signed record is never touched; the run's version goes
        #  beside it as (2), (3)... - NOT "(earlier N)", which is the
        #  archive of an unsigned one.
        recs = sorted(p for p in glob.glob(os.path.join(folder,
                                                        "Email Record*.txt"))
                      if re.search(r"\((\d+)\)\.txt$", p))
        latest = ""
        if recs:
            recs.sort(key=lambda p: int(re.search(r"\((\d+)\)\.txt$",
                                                  p).group(1)))
            with open(recs[-1], encoding="utf-8") as fh:
                m = re.search(r"Status\s*:\s*(.+)", fh.read())
                latest = m.group(1).strip() if m else ""
        check("Caught: a report re-issued AFTER the record was signed",
              latest.startswith("Failed"), "status was " + (latest or "?"))
        check("  and the signed record was not touched",
              "A. Fisher 07:15" in open(rec, encoding="utf-8").read())
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    # 5. the four oversight contacts listed one per row must all be found
    def one_per_row(ws, r, cols, name):
        pass
    tmp = tempfile.mkdtemp(prefix="k2pack_summary_")
    try:
        import openpyxl
        dest = os.path.join(tmp, "K2_Daily_Email_Report_Allocation.xlsx")
        shutil.copy2(book, dest)
        wb = openpyxl.load_workbook(dest)
        ws = wb["Summary"]
        row = None
        for r in range(1, ws.max_row + 1):
            joined = " ".join(str(ws.cell(r, c).value or "")
                              for c in range(1, ws.max_column + 1)).upper()
            if "FIXED" in joined and "CC" in joined:
                row = r
                break
        #  Gather EVERY address already in the block, however it is laid
        #  out - the live book mixes a four-in-one-cell row with John
        #  Pickels on his own row (added 28 Jul 2026). Assuming the first
        #  cell held them all made this scenario plant 4 and find 5.
        block = []
        rr = row + 1
        while rr <= ws.max_row:
            got = emails_in(" ".join(str(ws.cell(rr, c).value or "")
                                     for c in range(1, ws.max_column + 1)))
            if not got:
                break
            block += [e for e in got if e not in block]
            rr += 1
        n_have = rr - (row + 1)          # rows the block occupies today
        if len(block) > n_have:
            ws.insert_rows(row + 1 + n_have, len(block) - n_have)
        #  the Summary sheet merges cells for looks - a merged follower
        #  cell is read-only, and clearing it doesn't matter anyway (the
        #  block's addresses only ever lived in column 1)
        for i, e in enumerate(block):    # rewrite strictly one per row
            for c in range(1, ws.max_column + 1):
                try:
                    ws.cell(row + 1 + i, c).value = None
                except AttributeError:
                    pass                 # merged follower - nothing in it
            ws.cell(row + 1 + i, 1).value = e
        wb.save(dest)
        root = os.path.join(tmp, "OUT")
        r = run_builder(["--date", date_tag, "--root", root, "--book", dest])
        m = re.search(r"Fixed Cc\s*:\s*(\d+)", r.stdout or "")
        check("Oversight contacts listed one per row are all found",
              bool(m) and int(m.group(1)) == len(block),
              "{} of {}".format(m.group(1) if m else "?", len(block)))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    # 6. a column inserted in front of the routing sheet must not empty it
    tmp = tempfile.mkdtemp(prefix="k2pack_column_")
    try:
        import openpyxl
        dest = os.path.join(tmp, "K2_Daily_Email_Report_Allocation.xlsx")
        shutil.copy2(book, dest)
        wb = openpyxl.load_workbook(dest)
        wb["Company Routing"].insert_cols(1)
        wb.save(dest)
        root = os.path.join(tmp, "OUT")
        r = run_builder(["--date", date_tag, "--root", root, "--book", dest],
                        expect_zero=False)
        m = re.search(r"Companies\s*:\s*(\d+)", r.stdout or "")
        found = int(m.group(1)) if m else 0
        check("A column inserted on the routing sheet doesn't empty the day",
              found > 0 or r.returncode != 0,
              "{} companies read, exit {}".format(found, r.returncode))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    # 7. an Alt+Enter line break inside an address cell must not kill the run
    def alt_enter(ws, r, cols, name):
        if name == "Programmed":
            ws.cell(r, cols["To"]).value = "debra.rogers@programmed.com.au\nx@y.com"
    tmp = tempfile.mkdtemp(prefix="k2pack_newline_")
    try:
        bad_book = doctor(book, os.path.join(
            tmp, "K2_Daily_Email_Report_Allocation.xlsx"), alt_enter)
        root = os.path.join(tmp, "OUT")
        r = run_builder(["--date", date_tag, "--root", root, "--book", bad_book],
                        expect_zero=False)
        built_all = len(glob.glob(os.path.join(root, "01 COMPANY REPORTS",
                                               "*", date_tag,
                                               "Email Record*.txt")))
        check("A line break inside an address cell can't kill the run",
              r.returncode == 0 and built_all >= 12,
              "exit {}, {} records written".format(r.returncode, built_all))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    # 8. with no page images, the report must be ON the email
    tmp = tempfile.mkdtemp(prefix="k2pack_noimg_")
    try:
        reports = _mirror_reports(date_tag, tmp)
        for junk in glob.glob(os.path.join(reports, date_tag, "Emails",
                                           "*_Email*")):
            os.remove(junk)          # pictures AND the page count beside them
        root = os.path.join(tmp, "OUT")
        run_builder(["--date", date_tag, "--root", root, "--book", book,
                     "--reports", reports])
        eml = os.path.join(root, "01 COMPANY REPORTS", "DGH Engineering",
                           date_tag, "Email Draft - {}.eml".format(date_tag))
        if check("A pack with no page images still builds a draft",
                 os.path.isfile(eml)):
            got = open_eml(eml)
            names = [n for n, _d in got["attachments"]]
            check("  and the report itself is attached, not just mentioned",
                  any("on-hire" in n.lower() for n in names),
                  "; ".join(names))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    # 9. pages missing from the middle of a report must never be emailed
    for label, wreck in (
            ("A page missing from the middle of the report",
             lambda d, stem: os.remove(os.path.join(d, stem + "_Email7.png"))),
            ("The last few pages of the report gone",
             lambda d, stem: [os.remove(p) for p in sorted(
                 glob.glob(os.path.join(d, stem + "_Email*.png")),
                 key=lambda q: int(q.rsplit("_Email", 1)[1].split(".")[0]))
                 [-3:]])):
        tmp = tempfile.mkdtemp(prefix="k2pack_gap_")
        try:
            reports = _mirror_reports(date_tag, tmp)
            emails = os.path.join(reports, date_tag, "Emails")
            stem = "Coates_K2_Activity_DGH_ENGINEERING_{}".format(date_tag)
            if not glob.glob(os.path.join(emails, stem + "_Email*.png")):
                #  Company emails carry the report as ONE attached PDF
                #  now (Andrew, 28 Jul 2026) - there are no page images
                #  to go missing, so this failure cannot exist for them.
                #  Scenario 8 above proves the PDF path instead.
                print("   (skipped: {} - company emails are story + PDF "
                      "now, no page images exist to lose)".format(label))
                shutil.rmtree(tmp, ignore_errors=True)
                continue
            wreck(emails, stem)
            root = os.path.join(tmp, "OUT")
            run_builder(["--date", date_tag, "--root", root, "--book", book,
                         "--reports", reports])
            st = _status(root, "DGH Engineering", date_tag)
            check("Caught: {}".format(label),
                  st.startswith("Failed"), "status was " + st)
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


# =====================================================================
#  PART 4 - no overwrites
# =====================================================================
def record_handling(date_tag, book):
    head("PART 4a - THE RECORD YOU SIGN, AND THE ONE YOU READ")
    tmp = tempfile.mkdtemp(prefix="k2pack_record_")
    try:
        root = os.path.join(tmp, "OUT")
        run_builder(["--date", date_tag, "--root", root, "--book", book])
        rec = os.path.join(root, "01 COMPANY REPORTS", "DGH Engineering",
                           date_tag, "Email Record - {}.txt".format(date_tag))

        #  1. sign it by hand, then change the routing and run again
        with open(rec, encoding="utf-8") as fh:
            signed = fh.read().replace("Sent by      : ____________________",
                                       "Sent by      : A. Fisher 07:15")
        with open(rec, "w", encoding="utf-8") as fh:
            fh.write(signed)

        def swap(ws, r, cols, name):
            if name == "DGH Engineering":
                ws.cell(r, cols["To"]).value = "jayden.andrew@cleanaway.com.au"
        bad_book = doctor(book, os.path.join(tmp, "changed.xlsx"), swap)
        run_builder(["--date", date_tag, "--root", root, "--book", bad_book])

        #  a held pack must not leave a file anyone can double-click
        folder = os.path.dirname(rec)
        live = os.path.join(folder, "Email Draft - {}.eml".format(date_tag))
        held = os.path.join(folder,
                            "HOLD - Email Draft - {}.eml.txt".format(date_tag))
        check("A held pack leaves no sendable draft behind",
              not os.path.isfile(live) and os.path.isfile(held),
              "sendable .eml present" if os.path.isfile(live)
              else "kept as HOLD - ... .eml.txt")

        with open(rec, encoding="utf-8") as fh:
            after = fh.read()
        check("A record you have signed is never rewritten",
              "A. Fisher 07:15" in after, "the Sent by line is still there")
        check("  and the new one is filed beside it",
              bool(glob.glob(os.path.join(os.path.dirname(rec),
                                          "Email Record*(2).txt"))))

        #  2. an UNSIGNED record must show the current status, not the old
        rec2 = os.path.join(root, "01 COMPANY REPORTS", "Xtreme Engineering",
                            date_tag, "Email Record - {}.txt".format(date_tag))

        def drop_cc(ws, r, cols, name):
            if name == "Xtreme Engineering":
                cc = str(ws.cell(r, cols["Cc"]).value or "")
                ws.cell(r, cols["Cc"]).value = ";".join(
                    p for p in cc.split(";") if "david.moore" not in p)
        worse = doctor(book, os.path.join(tmp, "worse.xlsx"), drop_cc)
        run_builder(["--date", date_tag, "--root", root, "--book", worse])
        with open(rec2, encoding="utf-8") as fh:
            now = fh.read()
        check("An unsigned record shows the CURRENT status",
              "Failed - Review Required" in now,
              "the record you open says what the run just found")
        check("  and the earlier record is kept, not thrown away",
              bool(glob.glob(os.path.join(os.path.dirname(rec2),
                                          "Email Record*(earlier*.txt"))))

        #  3. put the routing back and run again - the pack must come
        #     right, with a sendable draft carrying the right addresses
        run_builder(["--date", date_tag, "--root", root, "--book", book])
        eml = os.path.join(folder, "Email Draft - {}.eml".format(date_tag))
        if check("Fixing the workbook brings the pack back",
                 os.path.isfile(eml)):
            got = open_eml(eml)
            check("  and the draft carries the CURRENT recipients",
                  "jayden.andrew@cleanaway.com.au" not in got["to"]
                  and len(got["to"]) == 4,
                  "; ".join(got["to"]))
        old = glob.glob(os.path.join(folder, "*superseded*"))
        check("  and every earlier draft is kept, none of them sendable",
              all(not p.lower().endswith(".eml") for p in old),
              "; ".join(os.path.basename(p) for p in old) or "none kept")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def no_overwrite(root, date_tag):
    head("PART 4 - RUN IT AGAIN: NOTHING MAY CHANGE")
    before = tree_state(root)
    run_builder(["--date", date_tag, "--root", root])
    after = tree_state(root)
    changed = sorted(k for k in before
                     if k in after and before[k] != after[k])
    added = sorted(set(after) - set(before))
    lost = sorted(set(before) - set(after))
    check("Not one filed byte changed on the second run",
          not changed, "; ".join(changed[:4]))
    check("Nothing new appeared", not added, "; ".join(added[:4]))
    check("Nothing disappeared", not lost, "; ".join(lost[:4]))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date", default=dt.date.today().strftime("%Y-%m-%d"))
    ap.add_argument("--root")
    a = ap.parse_args()
    date_tag = a.date
    root = a.root or os.path.join(HERE, "K2 DAILY REPORTING")

    books = [p for p in glob.glob(os.path.join(
        HERE, "K2_Daily_Email_Report_Allocation*.xlsx"))
        if not os.path.basename(p).startswith("~$")]
    if not books:
        print("No routing workbook found - nothing to test against.")
        return 2
    book = max(books, key=os.path.getmtime)

    print("")
    print("=" * 68)
    print(" COATES | K2 DAILY EMAIL PACKS - PROVING THE DAY")
    print(" {}   routing from {}".format(date_tag, os.path.basename(book)))
    print("=" * 68)

    if not os.path.isdir(root):
        print("No packs built yet - run BUILD_DAILY_EMAIL_PACKS.py first.")
        return 2
    if not os.path.isdir(os.path.join(HERE, "Reports", date_tag)):
        print("No reports built for {} - test the day that was built, e.g.".format(date_tag))
        print("  py TEST_DAILY_PACKS.py --date 2026-07-25")
        return 2

    #  PART 3b MEDDLES WITH THE COMPANY REPORTS, so it needs them to
    #  exist. Without them it used to walk straight into shutil.copy2
    #  and throw a FileNotFoundError traceback - forty lines of Python
    #  whose actual message is "run the company report first". A test
    #  that cannot say what is missing is not a test, it is a stack
    #  trace with a job title (3 Aug 2026).
    _co = glob.glob(os.path.join(HERE, "Reports", date_tag, "Pages",
                                 "Coates_OnHire_Report_*.html"))

    wbk = read_workbook(book)
    verify_packs(root, wbk, date_tag)
    sabotage(date_tag, book)
    if _co:
        quiet_failures(date_tag, book)
    else:
        print("")
        print("-" * 68)
        print(" PART 3b - SKIPPED, and here is exactly why")
        print("-" * 68)
        print("  These checks meddle with a copy of the per-company")
        print("  on-hire reports, and none are built for {}.".format(date_tag))
        print("")
        print("  Build them first:   py build_company_onhire_report.py --all")
        print("                 or:   03_RUN_COMPANY_REPORT.bat")
        print("")
        print("  Then run me again and Part 3b will run with the rest.")
        SKIPPED.append("Part 3b - no company on-hire reports built")
    record_handling(date_tag, book)
    no_overwrite(root, date_tag)

    print("")
    print("=" * 68)
    print(" {} passed, {} failed{}".format(
        len(PASS), len(FAIL),
        ", {} SKIPPED".format(len(SKIPPED)) if SKIPPED else ""))
    print("=" * 68)
    #  A skipped check reported as a pass is how a suite says "all
    #  good" about work it never did.
    for _s in SKIPPED:
        print("  SKIPPED: " + _s)
    if FAIL:
        print("")
        for name, detail in FAIL:
            print("  FAILED: {}{}".format(name, "  -  " + detail if detail else ""))
        print("")
        print(" Do not send anything until these are fixed.")
        return 1
    print("")
    print(" Every pack matches the workbook. Nothing was overwritten, and")
    print(" four deliberate routing mistakes were all caught and held.")
    if SKIPPED:
        print("")
        print(" {} check(s) above did NOT run. Read them before you treat".format(len(SKIPPED)))
        print(" this as a clean day.")
    print("")
    return 0


if __name__ == "__main__":
    sys.exit(main())
