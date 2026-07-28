#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | SET UP THE COMPLIANCE FLAGS
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Adds the four compliance columns to K2_MASTER_EQUIPMENT_PRICING.xlsx
#  and fills them in from the fleet we actually hold:
#
#      ELECTRICAL_TAG | RIGGING_TAG | LOGBOOK_REQUIRED | RETURN_REQUIREMENT
#
#  The rules below were written against the real 5,225-row master, not a
#  guess at what a hire fleet looks like. Every rule is listed so it can
#  be argued with - which is the point. Get one wrong, change the cell in
#  the spreadsheet, and the next report run picks it up.
#
#  SAFE TO RE-RUN, every day. An assessed row always carries Y or N in
#  the three flag columns, never blank - so "already done" tells itself
#  apart from "new gear, not looked at yet". A re-run only touches gear
#  that has arrived since, and nothing you have corrected by hand is ever
#  overwritten. Nothing new? It leaves the file completely alone.
#  Use --force to re-do the lot from scratch.
#
#      py SETUP_COMPLIANCE_FLAGS.py              first run / new gear
#      py SETUP_COMPLIANCE_FLAGS.py --dry-run    show me, change nothing
#      py SETUP_COMPLIANCE_FLAGS.py --force      redo everything
#
#  A timestamped backup is written beside the file before every save.
# =====================================================================

import os
import re
import sys
import glob
import shutil
import datetime as dt
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
PATTERN = "K2_MASTER_EQUIPMENT_PRICING*.xlsx"

NEW_COLS = ["ELECTRICAL_TAG", "RIGGING_TAG", "LOGBOOK_REQUIRED",
            "RETURN_REQUIREMENT"]

# ---------------------------------------------------------------------
#  RETURN WORDING - what the badge says under the description
# ---------------------------------------------------------------------
RET_GAS = "Return daily for bump test and recharge"
RET_RADIO = "Return daily for charging and channel check"
RET_RADIO_BATT = "Return daily for charging"
#  Chargers live in the store and stay there - the crews never take
#  them out, so never ask for one back. (A. Fisher, 25 Jul 2026)
RET_BATTERY = "Return tool and batteries daily"

# ---------------------------------------------------------------------
#  THE RULES
# ---------------------------------------------------------------------

# --- RIGGING -----------------------------------------------------------
#  Everything in the Lifting category is genuine rigging gear - checked
#  against all 68 distinct descriptions: shackles, slings, chain and
#  lever blocks, girder clamps and trolleys, eye bolts, lifting bags,
#  bull bags, Tirfor winch. No exceptions in there.
RIG_CATS = {"Lifting"}
#  Height safety gear is inspection-tagged the same way and is the
#  highest-consequence item to get wrong, so it carries the tag badge
#  too. Ladders and steps are access equipment - not tagged rigging.
RIG_HEIGHT_RE = re.compile(
    r"harness|lanyard|anchor strap|retractable|srl|fall arrest|rope grab",
    re.I)
RIG_HEIGHT_NOT_RE = re.compile(r"ladder|safety step|platform step", re.I)

# --- ELECTRICAL (test and tag, AS/NZS 3760) ----------------------------
#  Whole categories where every item is a portable electrical appliance,
#  lead, board or RCD.
ELEC_CATS = {"Electrical", "Plant - Distribution & Leads"}
#  ...except the stands and hooks that live in the Electrical category
#  but carry no current themselves.
ELEC_CAT_NOT_RE = re.compile(r"lead stand|cable hook", re.I)
#  Elsewhere in the fleet: flag only where the description actually says
#  it is mains powered. Battery gear is deliberately excluded - a
#  Milwaukee light is not a test-and-tag item.
ELEC_DESC_RE = re.compile(
    r"\b240\s?v\b|\b415\s?v\b|\b32\s?v\b|\belectric(al)?\b|"
    r"plasma cutter|floodlight|lead light|transformer|"
    r"earth leakage|\brcd\b|distribution board|extension lead|"
    #  A charger, a hotbox and a mains power supply all plug into the
    #  wall - they are test-and-tag items even though the tool they
    #  serve runs on a battery. (A. Fisher, 25 Jul 2026)
    r"\bcharger\b|hotbox|hot box|power supply", re.I)
#  Checked BEFORE the exclusion list below, so a "Milwaukee Battery
#  Charger" still gets its blue tag.
ELEC_ALWAYS_RE = re.compile(r"\bcharger\b|hotbox|hot box|power supply", re.I)
ELEC_DESC_NOT_RE = re.compile(
    r"battery|milwaukee|bosch 18|cordless|\bair\b|hydraulic hose|"
    r"lead stand|cable hook|ducting|welding extension lead|"
    r"welding earth lead|welding handpiece|welding lead tail|"
    r"caddy welding|welding hose|conveyor belt|"
    # ride-on, towed and self-powered plant - inspected through the log
    # book and its own service regime, not the test-and-tag register
    r"pallet truck|forklift|scissor|generator|lighting tower|"
    r"\bboom\b|telehandler|skid steer|excavator|dozer|roller|dumper|"
    r"water blaster|compressor|air vantage|welding vantage|"
    r"welder\s*-?\s*motorised", re.I)

# --- LOGBOOK (daily pre-start on operated plant) -----------------------
#  Categories where the machine is operated and carries a Coates log book.
LOG_CATS = {
    "Plant - Booms", "Plant - Scissor Lifts", "Plant - Forklifts",
    "Plant - Generators", "Plant - Lighting", "Plant - Welders",
    "Plant - Earthmoving", "Plant - Air & Compressors",
    "Plant - Material Handling", "Plant - Surface Preparation",
}
#  Attachments, hoses, buckets and frames ride under those categories but
#  are not operated plant and have no log book of their own.
LOG_NOT_RE = re.compile(
    r"bucket|attachment|bull hose|\bhose\b|frame|hopper|cage|kibble|"
    r"clamp|excavator trailer|brick/block", re.I)
#  Operated plant that sits under some other category.
LOG_DESC_RE = re.compile(
    r"forklift|telehandler|skid steer|excavator|dozer|roller|dumper|"
    r"scissor lift|boom lift|knuckle boom|straight boom|"
    r"trailer mounted boom|\bewp\b|generator|lighting tower|"
    r"light telescopic|lighting pod|compressor|water blaster|"
    r"pallet truck|conveyor|"
    r"welder\s*-?\s*motorised|welder diesel|welding (air )?vantage", re.I)

# --- RETURN DAILY ------------------------------------------------------
GAS_RE = re.compile(
    r"gas (monitor|detector)|multi-gas|gasalert|microclip|altair|quattro|"
    r"ventis|draeger|drager|x-am|odalog", re.I)
RADIO_RE = re.compile(r"two-way radio|\bradio\b", re.I)
#  A spare handset battery comes back to charge - it has no channel to set.
RADIO_BATT_RE = re.compile(r"radio battery", re.I)
RADIO_NOT_RE = re.compile(r"radio cover|antenna|charger rack", re.I)
BATT_RE = re.compile(
    r"milwaukee|bosch 18|cordless|\b\d+\s?ah battery\b|18 ?v battery|"
    #  Anything else that comes back to charge: intrinsically safe lights,
    #  USB-chargeable gauges, purge monitors, battery power supplies.
    r"battery light|battery power|chargeable|purge monitor|battery fan",
    re.I)


def norm(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()


def _isyes(v):
    return norm(v).lower() in ("y", "yes", "true", "1", "x", "req", "required")


def rules_for(desc, cat):
    """Return (electrical, rigging, logbook, return_note) for one asset."""
    d, c = norm(desc), norm(cat)

    # ---- rigging ----
    rigging = False
    if c in RIG_CATS:
        rigging = True
    elif c == "Working at Height":
        rigging = bool(RIG_HEIGHT_RE.search(d)) and not RIG_HEIGHT_NOT_RE.search(d)

    # ---- electrical ----
    if c in ELEC_CATS:
        electrical = not bool(ELEC_CAT_NOT_RE.search(d))
    elif ELEC_ALWAYS_RE.search(d):
        electrical = True
    else:
        electrical = (bool(ELEC_DESC_RE.search(d))
                      and not ELEC_DESC_NOT_RE.search(d))

    # ---- logbook ----
    if c in LOG_CATS:
        logbook = not bool(LOG_NOT_RE.search(d))
    else:
        logbook = (bool(LOG_DESC_RE.search(d))
                   and not LOG_NOT_RE.search(d))
    #  Mains-powered plant - a 240 V scarifier, a 240 V conveyor - carries
    #  BOTH: a pre-start log book and a current electrical tag. Diesel and
    #  battery plant is filtered out of the electrical rule above, so
    #  nothing here needs an override.

    # ---- return daily ----
    ret = ""
    if GAS_RE.search(d) or c == "Gas Detection":
        ret = RET_GAS
    elif RADIO_BATT_RE.search(d):
        ret = RET_RADIO_BATT
    elif RADIO_RE.search(d) and not RADIO_NOT_RE.search(d):
        ret = RET_RADIO
    elif BATT_RE.search(d) or c == "Battery Tools":
        ret = RET_BATTERY

    return electrical, rigging, logbook, ret


def find_master():
    hits = [p for p in glob.glob(os.path.join(HERE, PATTERN))
            + glob.glob(os.path.join(HERE, "Data_SiteIQ", PATTERN))
            if not os.path.basename(p).startswith("~$")]
    if not hits:
        sys.exit("  Couldn't find {} beside this script.".format(PATTERN))
    return max(hits, key=os.path.getmtime)


def main():
    force = "--force" in sys.argv
    dry = "--dry-run" in sys.argv
    path = find_master()
    print("=" * 68)
    print("  COATES K2 | COMPLIANCE FLAGS")
    print("  Author: Andrew Fisher | POWERED BY SITEIQ")
    print("=" * 68)
    print("  Master file : {}".format(os.path.basename(path)))

    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]

    hdr = {}
    for c in range(1, ws.max_column + 1):
        v = norm(ws.cell(1, c).value).upper()
        if v:
            hdr[v] = c
    for need in ("ITEM_NUMBER", "ITEM_DESCRIPTION", "EQUIPMENT_CATEGORY"):
        if need not in hdr:
            sys.exit("  Master is missing the {} column.".format(need))

    # add any missing compliance columns on the end
    added = []
    nxt = ws.max_column + 1
    for name in NEW_COLS:
        if name not in hdr:
            # reuse a trailing blank header cell if there is one
            while nxt > 1 and not norm(ws.cell(1, nxt - 1).value):
                nxt -= 1
            ws.cell(1, nxt, name)
            hdr[name] = nxt
            added.append(name)
            nxt += 1
    print("  Columns     : {}".format(
        ("added " + ", ".join(added)) if added else "already present"))

    c_item = hdr["ITEM_NUMBER"]
    c_desc = hdr["ITEM_DESCRIPTION"]
    c_new = hdr.get("NEW_DESCRIPTION")
    c_cat = hdr["EQUIPMENT_CATEGORY"]
    c_e, c_r, c_l, c_t = (hdr["ELECTRICAL_TAG"], hdr["RIGGING_TAG"],
                          hdr["LOGBOOK_REQUIRED"], hdr["RETURN_REQUIREMENT"])

    tally = Counter()
    examples = defaultdict(Counter)
    skipped = 0

    for r in range(2, ws.max_row + 1):
        if not norm(ws.cell(r, c_item).value):
            continue
        tally["rows"] += 1
        existing = [norm(ws.cell(r, c).value) for c in (c_e, c_r, c_l, c_t)]
        #  A row counts as assessed only when ALL THREE Y/N flag columns
        #  are filled. Testing "any" instead would let a hand-typed note
        #  in the RETURN column mark a brand-new 240 V grinder as done and
        #  quietly leave its blue tag off the report - the exact failure
        #  this whole change exists to prevent.
        if all(existing[:3]) and not force:
            skipped += 1
            hit = False
            for val, key in zip(existing, ("elec", "rig", "log")):
                if _isyes(val):
                    tally[key] += 1
                    hit = True
            #  RETURN_REQUIREMENT is free text, not a Y/N flag - anything
            #  in it that isn't an explicit N counts.
            if existing[3] and existing[3].upper() != "N":
                tally["ret"] += 1
                hit = True
            if hit:
                tally["any"] += 1
            continue
        tally["touched"] += 1

        desc = norm(ws.cell(r, c_new).value) if c_new else ""
        desc = desc or norm(ws.cell(r, c_desc).value)
        cat = norm(ws.cell(r, c_cat).value)
        e, g, l, t = rules_for(desc, cat)

        if not dry:
            #  Y or N, never blank - see the note above. RETURN_REQUIREMENT
            #  is free text, so it stays empty when there is nothing to say.
            #  Y or N, never blank - see the note above. RETURN_REQUIREMENT
            #  is free text, so it stays empty when there is nothing to say.
            #  Assign through .value: ws.cell(r, c, None) is a no-op in
            #  openpyxl, so --force would never be able to clear a stale
            #  return note.
            ws.cell(r, c_e).value = "Y" if e else "N"
            ws.cell(r, c_r).value = "Y" if g else "N"
            ws.cell(r, c_l).value = "Y" if l else "N"
            ws.cell(r, c_t).value = t or None

        if e:
            tally["elec"] += 1
            examples["elec"][desc] += 1
        if g:
            tally["rig"] += 1
            examples["rig"][desc] += 1
        if l:
            tally["log"] += 1
            examples["log"][desc] += 1
        if t:
            tally["ret"] += 1
            examples[t][desc] += 1
        if e or g or l or t:
            tally["any"] += 1

    # header styling to match the rest of the sheet
    if not dry:
        try:
            ref = ws.cell(1, c_item)
            for c in (c_e, c_r, c_l, c_t):
                cell = ws.cell(1, c)
                cell.font = Font(name=ref.font.name or "Calibri",
                                 bold=True, size=ref.font.size or 11,
                                 color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="F26222")
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center", wrap_text=True)
            from openpyxl.utils import get_column_letter
            for c, w in ((c_e, 14), (c_r, 12), (c_l, 17), (c_t, 46)):
                ws.column_dimensions[get_column_letter(c)].width = w
        except Exception as ex:
            print("  (header styling skipped: {})".format(ex))

    print("-" * 68)
    print("  Rows        : {:,}".format(tally["rows"]))
    if skipped:
        print("  Left alone  : {:,} already had flags "
              "(use --force to redo)".format(skipped))
    print("  Electrical  : {:,}  blue tag".format(tally["elec"]))
    print("  Rigging     : {:,}  blue tag".format(tally["rig"]))
    print("  Logbook     : {:,}  daily pre-start".format(tally["log"]))
    print("  Return daily: {:,}".format(tally["ret"]))
    print("  Lines with a compliance check: {:,}".format(tally["any"]))
    print("-" * 68)

    for key, title in (("elec", "ELECTRICAL"), ("rig", "RIGGING"),
                       ("log", "LOGBOOK")):
        if examples[key]:
            print("  {} - {} distinct, top of the list:".format(
                title, len(examples[key])))
            for d, n in examples[key].most_common(8):
                print("     {:5d}  {}".format(n, d[:58]))
    for t in (RET_GAS, RET_RADIO, RET_BATTERY):
        if examples[t]:
            print("  RETURN: {} - {} distinct".format(t, len(examples[t])))
            for d, n in examples[t].most_common(4):
                print("     {:5d}  {}".format(n, d[:58]))

    if dry:
        print("\n  DRY RUN - nothing written.")
        return

    #  Nothing new to flag and no new columns to add? Leave the file
    #  completely alone. Safe to sit in the daily run without piling up
    #  backups or touching a workbook Andrew may have open notes against.
    if not added and not tally["touched"]:
        print("\n  Nothing to do - every asset already has its flags set.")
        print("  File untouched. New gear in the master gets flagged")
        print("  automatically on the next run.")
        return

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = os.path.join(os.path.dirname(path),
                       "_backup_{}_{}".format(stamp, os.path.basename(path)))
    shutil.copy2(path, bak)
    wb.save(path)
    print("\n  Backup      : {}".format(os.path.basename(bak)))
    print("  Saved       : {}".format(os.path.basename(path)))
    print("\n  Anything wrong? Change the cell in the spreadsheet and run")
    print("  the reports again - blank means the badge does not show.")


if __name__ == "__main__":
    main()
