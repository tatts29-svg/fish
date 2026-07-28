#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Coates K2 Charge Reporter.

One local tool (127.0.0.1) to log a chargeable shutdown event - Damage,
Consumables sold to the client, Fuel, or Transport - fill a short form, and
get a professional Coates PDF record with the story behind it, a register
entry, an unsent Outlook draft, and a Costing Feed row that drops the charge
into the right cost line. The main K2 workbook is never touched: charges reach
the costing through CHARGE_REGISTER.xlsx (Costing Feed sheet), which the kit
rolls up into Daily Tracking.

All paths are relative to this script so the whole folder can be moved to
another computer without editing a username or drive.
"""

import argparse
import base64
import datetime as dt
import email.message
import email.policy
import html
import io
import json
import mimetypes
import os
import re
import shutil
import subprocess
import sys
import threading
import webbrowser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import unquote, urlparse
from zoneinfo import ZoneInfo

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("This reporter needs openpyxl. Run:  py -m pip install openpyxl")
    sys.exit(1)

KIT_DIR = os.path.dirname(os.path.abspath(__file__))
REGISTER_PATH = os.path.join(KIT_DIR, "CHARGE_REGISTER.xlsx")
PHOTO_DIR = os.path.join(KIT_DIR, "Charge_Photos")
OUTPUT_DIR = os.path.join(KIT_DIR, "Output_Charge_Reports")
PROJECT = "Cement Australia K2 Shutdown 2026 - Gladstone"
ORANGE = "#F26222"
DARK = "#1D1D1B"
MAX_BODY = 55 * 1024 * 1024
MAX_PHOTOS = 3
MAX_PHOTO_BYTES = 12 * 1024 * 1024

SUBMISSION_LOCK = threading.Lock()

try:
    BRISBANE = ZoneInfo("Australia/Brisbane")
except Exception:
    BRISBANE = dt.timezone(dt.timedelta(hours=10), "AEST")


# ---------------------------------------------------------------------------
#  Field + category configuration.  A "field" drives the form input, the
#  register column and the report detail row.  Field dict keys:
#    name, label, type(text|textarea|select|date|time|money|number|file),
#    options(list, for select), req(bool), hint(str), width(register col width)
# ---------------------------------------------------------------------------

def F(name, label, type="text", options=None, req=False, hint="", width=20):
    return {"name": name, "label": label, "type": type, "options": options or [],
            "req": req, "hint": hint, "width": width}


# common fields at the top of every form
LEAD = [
    F("company", "Client / company", req=True, hint="Who the charge is to.", width=26),
    F("date", "Date", type="date", req=True),
    F("time", "Time", type="time", req=True),
    F("raised_by", "Raised by", req=True, width=22),
    F("contact", "Contact"),
    F("site_ref", "Site / client reference", width=20),
]
# common fields at the foot of every form
TAIL_STORY = F("story", "What happened / why - the story", type="textarea", req=True,
               hint="Plain facts. This becomes the narrative on the record.", width=48)
TAIL_OWNER = F("owner", "Coates owner", width=20)
TAIL_PHOTOS = F("photos", "Photo / docket evidence", type="file",
                hint="Up to 3 images, 12 MB each. Optional but recommended.")

STATUS_OPTS = ["For approval", "Estimate", "Quote", "Approved", "Invoiced", "No charge"]

CATEGORIES = {
    "damage": {
        "label": "Damage / Breakdown",
        "prefix": "DMG",
        "doc": "Damage Advice & Request for Direction",
        "cost_line": "Damage Recovery",
        "amount": "approved_cost",
        "status": "cost_status",
        "photos_prompt": True,
        "fields": [
            F("asset_no", "Asset no. (item number)", hint="ITEM_NUMBER - matches RENTAL_STOCK where available.", width=18),
            F("plant_id", "Plant ID", width=16),
            F("description", "Item description", width=40),
            F("location", "Location"),
            F("person", "Person using / returning item", width=24),
            F("found_during", "Found during", type="select",
              options=["", "Return inspection", "Issue inspection", "While in use",
                       "Stocktake", "Maintenance inspection", "Other"]),
            F("event_type", "Event type", type="select",
              options=["", "Damage", "Breakdown / failure", "Missing component",
                       "Failed inspection", "Loss / cannot locate", "Other"]),
            F("classification", "Preliminary classification", type="select",
              options=["No determination - under review", "Breakdown / mechanical failure",
                       "Fair wear and tear", "Accidental damage",
                       "Misuse suspected - review required", "Transport damage", "Other"], width=26),
            F("operational_status", "Operational status", type="select",
              options=["", "Out of service - isolated", "Restricted use pending review",
                       "Safe and operational", "Returned to Coates", "Cannot locate"]),
            F("immediate_action", "Immediate action taken", type="textarea"),
            F("tagged", "Out-of-service tag", type="select", options=["", "Yes", "No", "Not required"]),
            F("off_hired", "Off-hired", type="select", options=["", "Yes", "No", "Not applicable"]),
            F("safety", "Safety / injury", type="select",
              options=["", "No injury or safety event reported",
                       "Potential safety concern - separate review required",
                       "Injury / event reported - safety process initiated"], width=26),
            F("witnesses", "Witnesses"),
            F("supervisor", "Supervisor notified"),
            F("replacement_cost", "Replacement-cost exposure", type="money",
              hint="Reference exposure only; not a charge."),
            F("repair_cost", "Repair / quoted cost", type="money"),
            F("approved_cost", "Approved recoverable cost", type="money",
              hint="Only after approval - this is what flows to costing."),
            F("cost_status", "Cost status", type="select",
              options=["", "Estimate", "Quote received", "Invoice received", "Approved", "No charge"]),
            F("approval_ref", "Approval / PO reference"),
        ],
    },
    "consumables": {
        "label": "Consumables sold",
        "prefix": "CON",
        "doc": "Consumables Supply & Charge Record",
        "cost_line": "Consumables Sales",
        "amount": "charge_total",
        "status": "status",
        "photos_prompt": False,
        "fields": [
            F("description", "Consumables supplied", req=True, width=40),
            F("quantity", "Quantity"),
            F("unit", "Unit", type="select",
              options=["Each", "Box", "Pack", "Litre", "Metre", "Roll", "Kg", "Set", "Other"]),
            F("unit_price", "Unit price", type="money"),
            F("charge_total", "Total charge", type="money", req=True,
              hint="The amount charged to the client - flows to costing."),
            F("area", "Area / work front"),
            F("ordered_by", "Ordered / authorised by", width=22),
            F("po_ref", "PO / approval reference"),
            F("status", "Charge status", type="select", options=STATUS_OPTS),
        ],
    },
    "fuel": {
        "label": "Fuel",
        "prefix": "FUEL",
        "doc": "Fuel Supply & Charge Record",
        "cost_line": "Fuel",
        "amount": "charge_total",
        "status": "status",
        "photos_prompt": False,
        "fields": [
            F("fuel_type", "Fuel type", type="select",
              options=["Diesel", "Unleaded", "AdBlue", "Oil / lubricant", "Other"]),
            F("asset_no", "Machine / asset fuelled", width=22),
            F("litres", "Litres", type="number"),
            F("rate", "Rate per litre", type="money"),
            F("charge_total", "Total charge", type="money", req=True,
              hint="The amount charged to the client - flows to costing."),
            F("location", "Location"),
            F("supplied_by", "Supplied by"),
            F("docket_ref", "Docket / PO reference"),
            F("status", "Charge status", type="select", options=STATUS_OPTS),
        ],
    },
    "transport": {
        "label": "Transport",
        "prefix": "TRN",
        "doc": "Transport Charge Record",
        "cost_line": "Transport",
        "amount": "charge_total",
        "status": "status",
        "photos_prompt": False,
        "fields": [
            F("move_from", "From", req=True),
            F("move_to", "To", req=True),
            F("load_desc", "Load / what was moved", req=True, width=40),
            F("direction", "Direction", type="select", options=["One way", "Return"]),
            F("truck_rego", "Truck / rego"),
            F("distance", "Distance (km) or hours"),
            F("charge_total", "Total charge", type="money", req=True,
              hint="The amount charged to the client - flows to costing."),
            F("docket_ref", "Docket / PO reference"),
            F("status", "Charge status", type="select", options=STATUS_OPTS),
        ],
    },
}

CATEGORY_ORDER = ["damage", "consumables", "fuel", "transport"]


def sheet_name(key):
    return re.sub(r"[\\/*?:\[\]]", "-", CATEGORIES[key]["label"])[:31]


def category_fields(key):
    """Full ordered field list for a category: lead + specific + story/owner/photos."""
    cfg = CATEGORIES[key]
    return LEAD + cfg["fields"] + [TAIL_STORY, TAIL_OWNER, TAIL_PHOTOS]


def register_headers(key):
    cfg = CATEGORIES[key]
    detail = [f for f in category_fields(key) if f["type"] != "file"]
    return (["Charge ID"] + [f["label"] for f in detail] +
            ["Amount to Costing", "Cost Line", "Photo Files", "Report HTML",
             "Report PDF", "Outlook Draft", "Submitted At", "Last Updated"])


# ---------------------------------------------------------------------------
#  small helpers
# ---------------------------------------------------------------------------

def now_local():
    return dt.datetime.now(BRISBANE).replace(tzinfo=None)


def clean(value, limit=4000):
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", str(value or "")).strip()[:limit]


def money(value):
    text = clean(value).replace("$", "").replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        raise ValueError("Costs must be numbers, for example 1250.00.")
    if number < 0:
        raise ValueError("Costs cannot be negative.")
    return round(number, 2)


def fmt_money(value):
    return "Pending" if value is None else "${:,.2f}".format(value)


def shown(value):
    return clean(value) or "Not recorded"


def safe_name(value, fallback="item"):
    value = re.sub(r"[^A-Za-z0-9._-]+", "_", clean(value)).strip("._")
    return value[:90] or fallback


def ensure_dirs():
    os.makedirs(PHOTO_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)


# ---------------------------------------------------------------------------
#  register: one workbook, a sheet per category + a Costing Feed roll-up
# ---------------------------------------------------------------------------

COSTING_HEADERS = ["Date", "Charge ID", "Category", "Cost Line", "Company",
                   "Description", "Amount", "Status", "Reference", "Report PDF",
                   "Submitted At"]


def _style_header(ws, ncol):
    orange = PatternFill("solid", fgColor="F26222")
    white = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = orange
        cell.font = white
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = "A1:{}1".format(get_column_letter(ncol))


def ensure_register():
    if os.path.isfile(REGISTER_PATH):
        return
    wb = openpyxl.Workbook()
    # category sheets
    first = True
    for key in CATEGORY_ORDER:
        cfg = CATEGORIES[key]
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = sheet_name(key)
        headers = register_headers(key)
        ws.append(headers)
        detail = [f for f in category_fields(key) if f["type"] != "file"]
        widths = [20] + [f["width"] for f in detail] + [16, 20, 34, 40, 40, 40, 18, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        _style_header(ws, len(headers))
    # costing feed roll-up
    cf = wb.create_sheet("Costing Feed")
    cf.append(COSTING_HEADERS)
    for i, w in enumerate([13, 20, 18, 20, 26, 46, 15, 14, 22, 40, 18], 1):
        cf.column_dimensions[get_column_letter(i)].width = w
    _style_header(cf, len(COSTING_HEADERS))
    wb.save(REGISTER_PATH)


def next_charge_id(key, event_date):
    ensure_register()
    prefix = "{}-{}-".format(CATEGORIES[key]["prefix"], event_date.strftime("%Y%m%d"))
    wb = openpyxl.load_workbook(REGISTER_PATH, read_only=True, data_only=True)
    ws = wb[sheet_name(key)]
    seq = 0
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        v = clean(row[0])
        if v.startswith(prefix):
            try:
                seq = max(seq, int(v.rsplit("-", 1)[1]))
            except ValueError:
                pass
    wb.close()
    return "{}{:03d}".format(prefix, seq + 1)


def append_register(key, charge_id, values, amount, system):
    try:
        wb = openpyxl.load_workbook(REGISTER_PATH)
        cfg = CATEGORIES[key]
        ws = wb[sheet_name(key)]
        detail = [f for f in category_fields(key) if f["type"] != "file"]
        row = [charge_id]
        for f in detail:
            val = values.get(f["name"])
            if f["type"] == "money":
                row.append(val if val is not None else None)
            elif f["type"] == "date" and isinstance(val, dt.date):
                row.append(val)
            else:
                row.append(clean(val) if val not in (None,) else "")
        row += [amount, cfg["cost_line"], system["photos"], system["html"],
                system["pdf"], system["eml"], system["submitted"], system["submitted"]]
        ws.append(row)
        r = ws.max_row
        for cell in ws[r]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        # money formats
        for i, f in enumerate(detail, 2):
            if f["type"] == "money":
                ws.cell(r, i).number_format = '$#,##0.00'
        ws.cell(r, 1 + len(detail) + 1).number_format = '$#,##0.00'  # Amount to Costing

        # costing feed
        cf = wb["Costing Feed"]
        cf.append([values.get("date"), charge_id, cfg["label"], cfg["cost_line"],
                   clean(values.get("company")), _summary_desc(key, values),
                   amount, clean(values.get(cfg["status"])) or "For approval",
                   _reference(key, values), system["pdf"], system["submitted"]])
        cr = cf.max_row
        cf.cell(cr, 7).number_format = '$#,##0.00'
        for cell in cf[cr]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        wb.save(REGISTER_PATH)
        wb.close()
    except PermissionError:
        raise RuntimeError("CHARGE_REGISTER.xlsx is open in Excel. Close it and submit again.")


def _summary_desc(key, v):
    if key == "damage":
        return clean(v.get("description")) or clean(v.get("asset_no")) or "Damage event"
    if key == "consumables":
        return clean(v.get("description")) or "Consumables"
    if key == "fuel":
        return "{} {}".format(clean(v.get("fuel_type")) or "Fuel",
                              ("- " + clean(v.get("asset_no"))) if clean(v.get("asset_no")) else "").strip()
    if key == "transport":
        return "{} to {} - {}".format(clean(v.get("move_from")), clean(v.get("move_to")),
                                      clean(v.get("load_desc")))
    return "Charge"


def _reference(key, v):
    for name in ("approval_ref", "po_ref", "docket_ref", "site_ref"):
        if clean(v.get(name)):
            return clean(v.get(name))
    return ""


# ---------------------------------------------------------------------------
#  optional asset lookup (RENTAL_STOCK / PLANT_ID_REGISTER) - damage prefill
# ---------------------------------------------------------------------------

def load_asset_catalog():
    catalog = {}
    companies = set()
    stock = os.path.join(KIT_DIR, "RENTAL_STOCK.xlsx")
    plant_ids = {}
    plant_register = os.path.join(KIT_DIR, "PLANT_ID_REGISTER.xlsx")
    if os.path.isfile(plant_register):
        try:
            wb = openpyxl.load_workbook(plant_register, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            headers = [clean(v) for v in next(rows)]
            for row in rows:
                data = dict(zip(headers, row))
                pid = clean(data.get("Plant ID"))
                for key in (data.get("Asset No"), data.get("Item No"), data.get("ITEM_NUMBER")):
                    key = clean(key).upper()
                    if key and pid:
                        plant_ids[key] = pid
            wb.close()
        except Exception:
            plant_ids = {}
    if not os.path.isfile(stock):
        return catalog, []
    try:
        wb = openpyxl.load_workbook(stock, read_only=True, data_only=True)
        ws = wb["RENTAL_STOCK"]
        rows = ws.iter_rows(values_only=True)
        headers = [clean(v) for v in next(rows)]
        records = []
        for row in rows:
            data = dict(zip(headers, row))
            rec = {
                "asset": clean(data.get("ITEM_NUMBER")),   # ITEM_NUMBER is the identifier
                "item_number": clean(data.get("ITEM_NUMBER")),
                "sku": clean(data.get("SKU_NUMBER")),
                "description": clean(data.get("ITEM_DESCRIPTION")),
                "company": clean(data.get("COMPANY_NAME")),
                "person": clean(data.get("HIRER_NAME")),
                "status": clean(data.get("ITEM_STATUS")),
                "plant_id": (plant_ids.get(clean(data.get("ITEM_NUMBER")).upper())
                             or plant_ids.get(clean(data.get("SKU_NUMBER")).upper()) or ""),
            }
            if rec["company"]:
                companies.add(rec["company"])
            records.append(rec)
        matches = {}
        for rec in records:
            for k in {rec["asset"], rec["item_number"], rec["sku"]}:
                if k:
                    matches.setdefault(k.upper(), []).append(rec)
        for k, kr in matches.items():
            catalog[k] = kr[0] if len(kr) == 1 else {"ambiguous": True, "matches": len(kr)}
        wb.close()
    except Exception:
        return {}, []
    return catalog, sorted(companies, key=str.upper)


# ---------------------------------------------------------------------------
#  report (client-facing PDF/HTML) - generic, category aware
# ---------------------------------------------------------------------------

def data_uri(photo):
    return "data:{};base64,{}".format(photo["mime"], photo["data"])


def report_css():
    # Component styles only - the page frame (white A4 sheet, dark rounded
    # panel with the orange top bar, one panel per page) comes from the
    # suite's report_pages engine, which also guarantees the page rules:
    # nothing sliced, tables split at whole rows, headings keep with their
    # content. (A. Fisher, 24 Jul 2026)
    return """
*{box-sizing:border-box;-webkit-print-color-adjust:exact;print-color-adjust:exact}
html,body{margin:0;background:#ffffff}
body{font-family:'Segoe UI',system-ui,Arial,sans-serif;color:#E6E9EE}
.sheet{max-width:210mm;margin:0 auto;padding:11mm;background:#ffffff}
.panel{background:#0E1116;border-radius:16px;border-top:9px solid #F26222;padding:14mm 13mm 15mm;position:relative}
.top{padding-top:0;display:flex;justify-content:space-between;gap:20px;flex:none}
.brand{font-size:31px;font-weight:900;letter-spacing:-1px;color:#fff}.brand span{color:#F26222}
.kicker{font-size:10px;text-transform:uppercase;letter-spacing:2px;color:#F26222;font-weight:800}
h1{font-size:25px;margin:6px 0 2px;color:#fff}
.doc{text-align:right;font-size:11px;line-height:1.7;color:#8B9099}.doc b{color:#fff}
.tag{display:inline-block;background:#F26222;color:#fff;font-size:10px;font-weight:800;letter-spacing:1px;text-transform:uppercase;padding:4px 12px;border-radius:20px;margin-top:8px}
.intro{background:#171B22;border:1px solid #2A313C;border-left:6px solid #F26222;border-radius:0 10px 10px 0;padding:15px 17px;margin:18px 0;font-size:13px;line-height:1.65;color:#E6E9EE;page-break-inside:avoid}.intro b{color:#fff}
h2{font-size:13px;margin:20px 0 9px;color:#fff;border-bottom:2px solid #2A313C;padding-bottom:6px;text-transform:uppercase;letter-spacing:.6px}
.hero{display:flex;gap:11px;margin:14px 0;page-break-inside:avoid}
.hero div{flex:1;background:#171B22;border:1px solid #2A313C;border-left:4px solid #F26222;border-radius:10px;padding:13px 15px}
.hero .v{font-size:22px;font-weight:900;color:#fff;line-height:1.15}.hero .l{font-size:9px;text-transform:uppercase;letter-spacing:.7px;color:#8B9099;margin-top:4px}
table{width:100%;border-collapse:separate;border-spacing:0;font-size:11px;background:#171B22;border:1px solid #2A313C;border-radius:10px;overflow:hidden}
th{background:#20262F;color:#fff;text-align:left;padding:8px 10px}
td{padding:8px 10px;border-bottom:1px solid #232A34;vertical-align:top;color:#E6E9EE}
tr:last-child td{border-bottom:0}.label{width:32%;font-weight:700;color:#9198A3}.money{font-weight:800;color:#fff}
.photos{display:grid;grid-template-columns:repeat(2,1fr);gap:10px}.photos figure{margin:0;background:#171B22;border:1px solid #2A313C;border-radius:8px;padding:6px;page-break-inside:avoid}.photos img{display:block;width:100%;max-height:80mm;object-fit:contain;border-radius:4px;background:#0b0e12}.photos figcaption{font-size:9px;color:#8B9099;padding-top:5px}
.choice{background:#171B22;border:1px solid #2A313C;border-radius:6px;padding:9px 12px;margin:7px 0;font-size:11px;color:#E6E9EE}.box{display:inline-block;border:1.5px solid #8B9099;width:12px;height:12px;margin-right:9px;vertical-align:-2px;border-radius:2px}
.response{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-top:15px;page-break-inside:avoid}.line{border-bottom:1px solid #4A5260;height:26px;font-size:9px;color:#8B9099;padding-top:11px}
.limits{font-size:9px;color:#9198A3;background:#141922;border:1px solid #2A313C;border-radius:8px;padding:9px 12px;margin-top:16px;line-height:1.5}.limits b{color:#F26222}
.foot{font-size:9px;color:#8B9099;border-top:1px solid #2A313C;padding-top:8px;margin-top:15px;display:flex;justify-content:space-between}
"""


def narrative(key, v):
    when = "{} at {}".format(v["date"].strftime("%d %b %Y"), clean(v.get("time")) or "time not recorded")
    story = clean(v.get("story"))
    company = clean(v.get("company")) or "the client"
    if key == "damage":
        item = clean(v.get("description")) or "the item"
        asset = clean(v.get("asset_no")) or "asset pending"
        base = ("Coates is providing transparent notice that <b>{item}</b> ({asset}) was "
                "reported for review on {when}.").format(item=html.escape(item),
                                                         asset=html.escape(asset),
                                                         when=html.escape(when))
    elif key == "consumables":
        item = clean(v.get("description")) or "consumables"
        base = ("This record confirms consumables supplied to <b>{co}</b> on {when}: "
                "<b>{item}</b>.").format(co=html.escape(company), when=html.escape(when),
                                         item=html.escape(item))
    elif key == "fuel":
        ft = clean(v.get("fuel_type")) or "Fuel"
        base = ("This record confirms {ft} supplied on {when}"
                "{asset}.").format(ft=html.escape(ft), when=html.escape(when),
                                   asset=(" to " + html.escape(clean(v.get("asset_no"))))
                                   if clean(v.get("asset_no")) else "")
    else:  # transport
        base = ("This record confirms a transport movement on {when}: "
                "<b>{load}</b> from {frm} to {to}.").format(
                    when=html.escape(when), load=html.escape(clean(v.get("load_desc")) or "load"),
                    frm=html.escape(clean(v.get("move_from")) or "origin"),
                    to=html.escape(clean(v.get("move_to")) or "destination"))
    if story:
        base += " " + html.escape(story)
    return base


def build_report(key, charge_id, v, photos, amount, generated):
    cfg = CATEGORIES[key]
    detail = [f for f in category_fields(key) if f["type"] not in ("file",)]
    rows = []
    for f in detail:
        val = v.get(f["name"])
        if f["type"] == "money":
            disp = fmt_money(val)
        elif f["type"] == "date" and isinstance(val, dt.date):
            disp = val.strftime("%d %b %Y")
        else:
            disp = shown(val)
        rows.append("<tr><td class='label'>{}</td><td>{}</td></tr>".format(
            html.escape(f["label"]), html.escape(disp)))
    # one table with a proper header row: the page engine splits it at
    # whole-row boundaries and repeats this header on every continuation
    # page - a row is never cut in half.
    rows_html = ("<table><thead><tr><th style='width:32%'>Field</th>"
                 "<th>Detail</th></tr></thead><tbody>"
                 + "".join(rows) + "</tbody></table>")

    status = clean(v.get(cfg["status"])) or "For approval"
    hero = ("<div class='hero'><div><div class='v'>{amt}</div><div class='l'>{aml}</div></div>"
            "<div><div class='v' style='font-size:14px'>{status}</div><div class='l'>Charge status</div></div>"
            "<div><div class='v' style='font-size:14px'>{line}</div><div class='l'>Costing line</div></div></div>"
            ).format(amt=fmt_money(amount),
                     aml=("Approved recoverable" if key == "damage" else "Total charge"),
                     status=html.escape(status), line=html.escape(cfg["cost_line"]))

    photo_html = ""
    if photos:
        photo_html = "<h2>Evidence</h2><div class='photos'>" + "".join(
            "<figure><img src='{}' alt='evidence'><figcaption>{}</figcaption></figure>".format(
                data_uri(p), html.escape(p["original_name"])) for p in photos) + "</div>"

    if key == "damage":
        direction = ("<h2>How would you like us to proceed?</h2>"
                     "<div class='choice'><span class='box'></span>Arrange assessment and provide a repair quotation before work proceeds</div>"
                     "<div class='choice'><span class='box'></span>Proceed with repair, subject to the approval / PO reference below</div>"
                     "<div class='choice'><span class='box'></span>Proceed with replacement, subject to the approval / PO reference below</div>"
                     "<div class='choice'><span class='box'></span>Hold the item for a joint inspection or further review</div>")
        limits = ("<div class='limits'><b>Important:</b> replacement-cost exposure is the value at "
                  "risk, not automatically the amount to repair and not an approved charge. Only the "
                  "separately approved recoverable cost is carried into Coates cost reporting.</div>")
        confirm_lbl = ("Client representative / role", "Approval, PO or claim reference")
    else:
        direction = ("<h2>Confirmation</h2>"
                     "<div class='choice'><span class='box'></span>Charge confirmed for invoicing against the reference below</div>"
                     "<div class='choice'><span class='box'></span>Query - please contact before invoicing</div>")
        limits = ("<div class='limits'>This record documents a charge raised during the K2 shutdown. "
                  "The amount shown flows to the Coates costing under the cost line noted above once its "
                  "status is Approved or Invoiced.</div>")
        confirm_lbl = ("Client representative / role", "PO / approval reference")

    # ---- REAL PAGES (A. Fisher, 24 Jul 2026): the record is laid out by
    # the suite's page engine - white A4 sheets, dark Coates panel per
    # page, header on every page, tables split at whole rows with the
    # header repeated, nothing ever cut mid-element. Falls back to the
    # single-flow document only if report_pages.py can't be found.
    first_header = (
        "<header class='top'><div><div class='kicker'>Coates &middot; K2 Tool Store</div>"
        "<div class='brand'>SITE<span>IQ</span></div><h1>{doc}</h1><div class='proj'>{project}</div>"
        "<div class='tag'>{label}</div></div>"
        "<div class='doc'><b>{cid}</b><br>Prepared {gen}<br>Prepared by {owner}<br>"
        "Shutdown Manager &middot; Coates<br><span class='pgnum'></span></div></header>"
    ).format(doc=html.escape(cfg["doc"]), project=html.escape(PROJECT),
             label=html.escape(cfg["label"]), cid=html.escape(charge_id),
             gen=generated.strftime("%d %b %Y %H:%M"),
             owner=html.escape(clean(v.get("owner")) or "Andrew Fisher"))
    body = ("<section class='intro'>{narr}</section>"
            "{hero}"
            "<h2>Details</h2>{rows}{photos}"
            "{limits}{direction}").format(
        narr=narrative(key, v), hero=hero, rows=rows_html,
        photos=photo_html, limits=limits, direction=direction)
    footer = ("<footer class='foot'><span>{cid} &middot; Coates K2 charge record &middot; "
              "Author: <b>Andrew Fisher</b></span>"
              "<span>Generated by the Coates K2 Charge Reporter &middot; POWERED BY SITEIQ</span>"
              "</footer>").format(cid=html.escape(charge_id))
    try:
        import sys as _sys
        _parent = os.path.dirname(KIT_DIR)
        if _parent not in _sys.path:
            _sys.path.insert(0, _parent)
        import report_pages
        later_header = report_pages.compact_header(
            "K2 Tool Store &middot; Charge Record",
            "{} &middot; {}".format(html.escape(cfg["doc"]), html.escape(charge_id)),
            generated.strftime("%d %b %Y %H:%M"))
        return report_pages.paged_doc(
            cfg["doc"], report_css(), first_header, later_header, footer,
            body, doc_title="{} - Coates {}".format(charge_id, cfg["doc"]))
    except ImportError:
        return ("<!DOCTYPE html><html><head><meta charset='utf-8'>"
                "<title>{cid} - Coates {doc}</title>"
                "<style>{css}</style></head><body><main class='sheet'>"
                "<div class='panel'>{hdr}{body}{foot}</div></main></body></html>"
                ).format(cid=html.escape(charge_id), doc=html.escape(cfg["doc"]),
                         css=report_css(), hdr=first_header, body=body, foot=footer)


# ---------------------------------------------------------------------------
#  PDF + Outlook draft + photos
# ---------------------------------------------------------------------------

def find_edge():
    candidates = [
        os.path.expandvars(r"%ProgramFiles(x86)%\Microsoft\Edge\Application\msedge.exe"),
        os.path.expandvars(r"%ProgramFiles%\Microsoft\Edge\Application\msedge.exe"),
        shutil.which("msedge"), shutil.which("microsoft-edge"), shutil.which("google-chrome"),
    ]
    return next((p for p in candidates if p and os.path.isfile(p)), None)


def make_pdf(html_path, pdf_path):
    try:
        from weasyprint import HTML
        HTML(filename=html_path).write_pdf(pdf_path)
        return os.path.isfile(pdf_path) and os.path.getsize(pdf_path) > 0
    except Exception:
        pass
    edge = find_edge()
    if not edge:
        return False
    profile = os.path.join(os.environ.get("TEMP", OUTPUT_DIR), "coates_charge_pdf_{}".format(os.getpid()))
    url = "file:///" + os.path.abspath(html_path).replace("\\", "/")
    try:
        subprocess.run([edge, "--headless", "--disable-gpu", "--no-first-run",
                        "--user-data-dir=" + profile, "--no-pdf-header-footer",
                        "--print-to-pdf=" + os.path.abspath(pdf_path), url],
                       capture_output=True, timeout=180)
    except Exception:
        return False
    return os.path.isfile(pdf_path) and os.path.getsize(pdf_path) > 0


def make_eml(path, key, charge_id, v, amount, report_path, photos):
    cfg = CATEGORIES[key]
    company = clean(v.get("company")) or "client"
    subject_item = _summary_desc(key, v)
    msg = email.message.EmailMessage(policy=email.policy.SMTP)
    msg["Subject"] = "Coates K2 - {} {} - {} - {}".format(cfg["doc"], charge_id, subject_item, company)
    msg["X-Unsent"] = "1"
    amt = fmt_money(amount)
    if key == "damage":
        ask = ("Could you please let us know how you would like us to proceed: assessment and "
               "quotation, approved repair, replacement, or a joint review?")
    else:
        ask = ("Please confirm the charge for invoicing, or let us know if you have any query.")
    body = ("Hi,\n\n"
            "Please find attached Coates {doc} {cid} regarding {item} for {co}.\n\n"
            "Amount: {amt} ({line}, status: {status}).\n\n{ask}\n\n"
            "The attached record sets out the known facts and evidence where available.\n\n"
            "Thanks,\n{owner}\nShutdown Manager - Coates\n{project}\n").format(
        doc=cfg["doc"], cid=charge_id, item=subject_item, co=company, amt=amt,
        line=cfg["cost_line"], status=clean(v.get(cfg["status"])) or "For approval",
        ask=ask, owner=clean(v.get("owner")) or "Andrew Fisher", project=PROJECT)
    msg.set_content(body)
    with open(report_path, "rb") as h:
        data = h.read()
    if report_path.lower().endswith(".pdf"):
        msg.add_attachment(data, maintype="application", subtype="pdf",
                           filename=os.path.basename(report_path))
    else:
        msg.add_attachment(data, maintype="text", subtype="html",
                           filename=os.path.basename(report_path))
    for photo in photos:
        with open(photo["saved_path"], "rb") as h:
            data = h.read()
        main, sub = photo["mime"].split("/", 1)
        msg.add_attachment(data, maintype=main, subtype=sub,
                           filename=os.path.basename(photo["saved_path"]))
    with open(path, "wb") as h:
        h.write(msg.as_bytes())


def decode_photos(items, charge_id):
    items = items or []
    if len(items) > MAX_PHOTOS:
        raise ValueError("A maximum of 3 photos can be added to one report.")
    saved = []
    folder = os.path.join(PHOTO_DIR, charge_id)
    for index, item in enumerate(items, 1):
        mime = clean(item.get("type"), 80).lower()
        name = safe_name(item.get("name"), "photo_{}.jpg".format(index))
        if not mime.startswith("image/"):
            raise ValueError("Only image files can be added as photos.")
        try:
            raw = base64.b64decode(clean(item.get("data"), MAX_BODY), validate=True)
        except Exception:
            raise ValueError("One photo could not be read. Please choose the image again.")
        if not raw or len(raw) > MAX_PHOTO_BYTES:
            raise ValueError("Each photo must be between 1 byte and 12 MB.")
        os.makedirs(folder, exist_ok=True)
        stem, ext = os.path.splitext(name)
        if not ext:
            ext = mimetypes.guess_extension(mime) or ".jpg"
        p = os.path.join(folder, "{:02d}_{}{}".format(index, safe_name(stem), ext.lower()))
        with open(p, "wb") as h:
            h.write(raw)
        saved.append({"mime": mime, "data": base64.b64encode(raw).decode("ascii"),
                      "saved_path": p, "original_name": name})
    return saved


# ---------------------------------------------------------------------------
#  submit
# ---------------------------------------------------------------------------

def parse_values(key, payload):
    """Validate + coerce the incoming payload into a values dict."""
    v = {}
    fields = category_fields(key)
    # date/time
    try:
        v["date"] = dt.datetime.strptime(clean(payload.get("date")), "%Y-%m-%d").date()
    except ValueError:
        raise ValueError("Please select the date.")
    for f in fields:
        n = f["name"]
        if n in ("date", "photos"):
            continue
        raw = payload.get(n)
        if f["type"] == "money":
            v[n] = money(raw)
        else:
            v[n] = clean(raw)
    # required checks
    for f in fields:
        if f.get("req") and f["type"] != "file":
            val = v.get(f["name"])
            if val in (None, ""):
                raise ValueError("Please complete: {}".format(f["label"]))
    if not v.get("owner"):
        v["owner"] = "Andrew Fisher"
    return v


def submit(payload):
    generated = now_local()
    key = clean(payload.get("category")).lower()
    if key not in CATEGORIES:
        raise ValueError("Please choose what you are logging (Damage, Consumables, Fuel or Transport).")
    v = parse_values(key, payload)
    with SUBMISSION_LOCK:
        return submit_locked(key, payload, v, generated)


def submit_locked(key, payload, v, generated):
    cfg = CATEGORIES[key]
    charge_id = next_charge_id(key, v["date"])
    photos = decode_photos(payload.get("photos"), charge_id)
    amount = v.get(cfg["amount"])
    stem = "Coates_K2_{}_{}_{}".format(cfg["prefix"], charge_id, safe_name(clean(v.get("company"))))
    html_path = os.path.join(OUTPUT_DIR, stem + ".html")
    pdf_path = os.path.join(OUTPUT_DIR, stem + ".pdf")
    eml_path = os.path.join(OUTPUT_DIR, stem + "_OUTLOOK.eml")
    report = build_report(key, charge_id, v, photos, amount, generated)
    with io.open(html_path, "w", encoding="utf-8") as h:
        h.write(report)
    pdf_ok = make_pdf(html_path, pdf_path)
    attachment = pdf_path if pdf_ok else html_path
    make_eml(eml_path, key, charge_id, v, amount, attachment, photos)
    system = {
        "photos": "; ".join(os.path.relpath(p["saved_path"], KIT_DIR) for p in photos),
        "html": os.path.relpath(html_path, KIT_DIR),
        "pdf": os.path.relpath(pdf_path, KIT_DIR) if pdf_ok else "",
        "eml": os.path.relpath(eml_path, KIT_DIR),
        "submitted": generated,
    }
    append_register(key, charge_id, v, amount, system)
    return {"id": charge_id, "category": cfg["label"], "html": os.path.basename(html_path),
            "pdf": os.path.basename(pdf_path) if pdf_ok else "", "eml": os.path.basename(eml_path)}


# ---------------------------------------------------------------------------
#  form (server-rendered from the config) + server
# ---------------------------------------------------------------------------

def render_field(f):
    n, t = f["name"], f["type"]
    req = " <span class='req'>*</span>" if f.get("req") else ""
    hint = "<div class='hint'>{}</div>".format(html.escape(f["hint"])) if f["hint"] else ""
    label = "<label>{}{}</label>".format(html.escape(f["label"]), req)
    if t == "textarea":
        inp = "<textarea name='{}'></textarea>".format(n)
        cls = "full"
    elif t == "select":
        opts = "".join("<option>{}</option>".format(html.escape(o)) if o else "<option value=''>Pending / not set</option>"
                       for o in f["options"])
        inp = "<select name='{}'>{}</select>".format(n, opts)
        cls = ""
    elif t == "file":
        inp = "<input type='file' name='{}' accept='image/*' multiple>".format(n)
        cls = "full"
    elif t == "money":
        inp = "<div class='money'><input inputmode='decimal' name='{}'></div>".format(n)
        cls = ""
    elif t in ("date", "time", "number"):
        it = {"date": "date", "time": "time", "number": "number"}[t]
        inp = "<input type='{}' name='{}'>".format(it, n)
        cls = ""
    else:
        extra = " list='companies'" if n == "company" else (" id='asset'" if n == "asset_no" else "")
        inp = "<input name='{}'{}>".format(n, extra)
        cls = "full" if n in ("description", "load_desc") else ""
    return "<div class='{}'>{}{}{}</div>".format(cls, label, inp, hint)


def render_category_block(key):
    fields = CATEGORIES[key]["fields"] + [TAIL_STORY, TAIL_OWNER, TAIL_PHOTOS]
    body = "".join(render_field(f) for f in fields)
    return "<div class='catblock' data-cat='{}' style='display:none'><div class='grid'>{}</div></div>".format(key, body)


def build_form():
    lead = "".join(render_field(f) for f in LEAD)
    blocks = "".join(render_category_block(k) for k in CATEGORY_ORDER)
    cat_opts = "".join("<option value='{}'>{}</option>".format(k, html.escape(CATEGORIES[k]["label"]))
                       for k in CATEGORY_ORDER)
    return FORM_SHELL.replace("__LEAD__", lead).replace("__BLOCKS__", blocks).replace("__CATOPTS__", cat_opts)


FORM_SHELL = r"""<!DOCTYPE html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>
<title>Coates K2 Charge Reporter</title><style>
*{box-sizing:border-box}body{margin:0;background:#1D1D1B;font-family:'Segoe UI',Arial;color:#1D1D1B}
.wrap{max-width:900px;margin:0 auto;padding:22px 12px 50px}.card{background:#fff;border-radius:12px;border-top:9px solid #F26222;padding:22px;box-shadow:0 8px 30px #0007}
.brand{font-size:29px;font-weight:900}.brand span{color:#F26222}.eyebrow{color:#F26222;font-size:10px;font-weight:800;letter-spacing:2px;text-transform:uppercase}
h1{font-size:24px;margin:4px 0}.lead{color:#666;line-height:1.55;margin-bottom:16px}
.picker{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin:8px 0 18px}
.picker button{border:2px solid #ddd;background:#faf7f5;border-radius:9px;padding:12px 8px;font-weight:800;font-size:13px;cursor:pointer;color:#1D1D1B}
.picker button.on{border-color:#F26222;background:#FFF3EC;color:#F26222}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:13px}label{display:block;font-size:12px;font-weight:750;margin-bottom:4px}.full{grid-column:1/-1}
input,select,textarea{width:100%;padding:11px 10px;border:1px solid #aaa;border-radius:6px;font:inherit;background:#fff}textarea{min-height:80px;resize:vertical}
.req{color:#B3261E}.hint{font-size:10px;color:#777;margin-top:4px}.lookup{background:#FFF3EC;border-left:5px solid #F26222;padding:9px 11px;font-size:11px;display:none;margin-top:10px}
.money{position:relative}.money input{padding-left:26px}.money:before{content:'$';position:absolute;left:10px;top:11px;color:#777}
button.go{border:0;border-radius:7px;background:#F26222;color:#fff;font-weight:850;font-size:17px;padding:14px 20px;width:100%;margin-top:18px;cursor:pointer}button.go:disabled{opacity:.55}
.notice{background:#f5f5f5;padding:10px 12px;border-radius:5px;font-size:11px;line-height:1.5;margin-top:14px}
.status{display:none;margin-top:14px;padding:13px;border-radius:7px;line-height:1.55}.status.ok{display:block;background:#E9F7ED;border-left:6px solid #24823B}.status.bad{display:block;background:#FDECEC;border-left:6px solid #B3261E}
.actions a{display:inline-block;margin:8px 8px 0 0;padding:8px 10px;border-radius:5px;background:#1D1D1B;color:#fff;text-decoration:none;font-weight:700;font-size:12px}
h3{font-size:13px;margin:18px 0 6px;color:#F26222;text-transform:uppercase;letter-spacing:1px}
@media(max-width:650px){.grid,.picker{grid-template-columns:1fr}.full{grid-column:auto}.card{padding:17px}}
</style></head><body><div class='wrap'><div class='card'>
<div class='eyebrow'>Coates &middot; Cement Australia K2</div><div class='brand'>SITE<span>IQ</span></div><h1>Charge Reporter</h1>
<div class='lead'>Log a chargeable event once. It creates the register entry, a professional Coates record (PDF) with the story behind it, an unsent Outlook draft, and drops the charge into the right Costing line.</div>
<div class='picker' id='picker'>__PICKERBTNS__</div>
<form id='f'>
<input type='hidden' name='category' id='category'>
<h3>Common details</h3>
<div class='grid'>__LEAD__</div>
<div class='lookup' id='lookup'></div>
<div id='blocks'>__BLOCKS__</div>
<div class='notice'><b>Kept honest:</b> amounts flow to the Coates costing under their cost line once status is Approved or Invoiced. Nothing is sent automatically - the Outlook draft is yours to check and send.</div>
<button class='go' id='submit' type='submit' disabled>Choose a type above to begin</button>
<div id='status' class='status'></div></form></div></div>
<datalist id='companies'>__COMPANIES__</datalist>
<script>
const catalog=__CATALOG__;
const f=document.getElementById('f'),statusBox=document.getElementById('status'),catInput=document.getElementById('category'),submitBtn=document.getElementById('submit');
const LABELS=__LABELS__;
function setStatus(kind,html){statusBox.className='status '+kind;statusBox.innerHTML=html;statusBox.scrollIntoView({behavior:'smooth',block:'nearest'})}
function selectCat(key){catInput.value=key;
  document.querySelectorAll('.catblock').forEach(b=>{const on=b.dataset.cat===key;b.style.display=on?'block':'none';
    b.querySelectorAll('input,select,textarea').forEach(el=>el.disabled=!on)});
  document.querySelectorAll('#picker button').forEach(b=>b.classList.toggle('on',b.dataset.cat===key));
  submitBtn.disabled=false;submitBtn.textContent='Generate '+LABELS[key]+' record';window.__initDT()}
document.querySelectorAll('#picker button').forEach(b=>b.addEventListener('click',e=>{e.preventDefault();selectCat(b.dataset.cat)}));
const assetEl=()=>f.querySelector("[name=asset_no]");
document.addEventListener('change',e=>{if(!e.target.matches('[name=asset_no]'))return;const r=catalog[e.target.value.trim().toUpperCase()],box=document.getElementById('lookup');if(!r){box.style.display='none';return}box.style.display='block';if(r.ambiguous){box.textContent='Multiple RENTAL_STOCK items match ('+r.matches+'). Enter the unique number or confirm manually.';return}box.textContent='Known: '+(r.description||'')+' · '+(r.plant_id?'Plant ID '+r.plant_id+' · ':'')+(r.company||'')+' · '+(r.status||'');const set=(n,val)=>{const el=f.querySelector('[name='+n+']');if(el&&!el.value&&val)el.value=val};set('description',r.description);set('plant_id',r.plant_id);set('company',r.company);set('person',r.person)});
async function files(){const el=f.querySelector('[name=photos]');const selected=el?Array.from(el.files):[];if(selected.length>3)throw new Error('Maximum 3 photos.');const out=[];for(const file of selected){if(file.size>12*1024*1024)throw new Error(file.name+' is over 12 MB.');const data=await new Promise((ok,bad)=>{const r=new FileReader();r.onload=()=>ok(String(r.result).split(',')[1]);r.onerror=bad;r.readAsDataURL(file)});out.push({name:file.name,type:file.type,data})}return out}
f.addEventListener('submit',async e=>{e.preventDefault();if(!catInput.value){setStatus('bad','Choose what you are logging first.');return}submitBtn.disabled=true;submitBtn.textContent='Generating record…';try{const data={};new FormData(f).forEach((val,k)=>{data[k]=val});data.photos=await files();const res=await fetch('/submit',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(data)});const body=await res.json();if(!res.ok)throw new Error(body.error||'The record could not be generated.');setStatus('ok','<b>'+body.id+' created ('+body.category+').</b><div class="actions"><a target="_blank" href="/output/'+encodeURIComponent(body.html)+'">Open record</a>'+(body.pdf?'<a target="_blank" href="/output/'+encodeURIComponent(body.pdf)+'">Open PDF</a>':'')+'<a href="/output/'+encodeURIComponent(body.eml)+'">Outlook draft</a><a href="/register">Charge register</a></div><br>Saved to the register and the Costing Feed.');const keep=catInput.value;f.reset();catInput.value=keep;document.querySelector('[name=owner]')&&(document.querySelector('[name=owner]').value='Andrew Fisher');document.getElementById('lookup').style.display='none';window.__initDT&&window.__initDT()}catch(err){setStatus('bad','<b>Could not generate the record.</b><br>'+String(err.message||err))}finally{submitBtn.disabled=false;submitBtn.textContent='Generate '+LABELS[catInput.value]+' record'}});
window.__initDT=function(){const d=new Date();const p=n=>String(n).padStart(2,'0');f.querySelectorAll("[name=date]").forEach(el=>{if(!el.value)el.value=d.getFullYear()+'-'+p(d.getMonth()+1)+'-'+p(d.getDate())});f.querySelectorAll("[name=time]").forEach(el=>{if(!el.value)el.value=p(d.getHours())+':'+p(d.getMinutes())})};
document.querySelectorAll('.catblock input,.catblock select,.catblock textarea').forEach(el=>el.disabled=true);
window.__initDT();
</script></body></html>"""


class Handler(BaseHTTPRequestHandler):
    server_version = "CoatesChargeReporter/1.0"

    def log_message(self, fmt, *args):
        print("  " + (fmt % args))

    def send_bytes(self, data, content_type, status=200, filename=None):
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("X-Content-Type-Options", "nosniff")
        if filename:
            self.send_header("Content-Disposition", 'attachment; filename="{}"'.format(filename))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        route = urlparse(self.path).path
        if route == "/":
            catalog, companies = load_asset_catalog()
            options = "".join("<option value='{}'>".format(html.escape(c, quote=True)) for c in companies)
            picker = "".join("<button data-cat='{}'>{}</button>".format(k, html.escape(CATEGORIES[k]["label"]))
                             for k in CATEGORY_ORDER)
            labels = json.dumps({k: CATEGORIES[k]["label"] for k in CATEGORY_ORDER})
            page = (build_form().replace("__CATALOG__", json.dumps(catalog, ensure_ascii=True))
                    .replace("__COMPANIES__", options).replace("__PICKERBTNS__", picker)
                    .replace("__LABELS__", labels))
            self.send_bytes(page.encode("utf-8"), "text/html; charset=utf-8")
            return
        if route == "/register":
            with open(REGISTER_PATH, "rb") as h:
                self.send_bytes(h.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                filename=os.path.basename(REGISTER_PATH))
            return
        if route.startswith("/output/"):
            name = os.path.basename(unquote(route[len("/output/"):]))
            path = os.path.abspath(os.path.join(OUTPUT_DIR, name))
            if os.path.dirname(path) != os.path.abspath(OUTPUT_DIR) or not os.path.isfile(path):
                self.send_error(404)
                return
            with open(path, "rb") as h:
                ctype = mimetypes.guess_type(path)[0] or "application/octet-stream"
                self.send_bytes(h.read(), ctype, filename=name if name.lower().endswith(".eml") else None)
            return
        self.send_error(404)

    def do_POST(self):
        if urlparse(self.path).path != "/submit":
            self.send_error(404)
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            if length <= 0 or length > MAX_BODY:
                raise ValueError("The submission is empty or too large (maximum 55 MB).")
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            result = submit(payload)
            self.send_bytes(json.dumps(result).encode("utf-8"), "application/json")
        except (ValueError, RuntimeError) as exc:
            self.send_bytes(json.dumps({"error": str(exc)}).encode("utf-8"), "application/json", status=400)
        except Exception as exc:
            print("Unexpected error: {}: {}".format(type(exc).__name__, exc))
            self.send_bytes(json.dumps({"error": "Something unexpected went wrong: {}".format(exc)}).encode("utf-8"),
                            "application/json", status=500)


def run_server(port, open_browser=True):
    ensure_dirs()
    ensure_register()
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    url = "http://127.0.0.1:{}/".format(server.server_address[1])
    print("=" * 62)
    print(" COATES | K2 CHARGE REPORTER")
    print(" {}".format(PROJECT))
    print(" Local address: {}".format(url))
    print(" Close this window when finished.")
    print("=" * 62)
    if open_browser:
        threading.Timer(0.6, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


def main():
    parser = argparse.ArgumentParser(description="Coates K2 Charge Reporter")
    parser.add_argument("--port", type=int, default=8768)
    parser.add_argument("--no-browser", action="store_true")
    parser.add_argument("--init-only", action="store_true")
    args = parser.parse_args()
    ensure_dirs()
    ensure_register()
    if args.init_only:
        print(REGISTER_PATH)
        return
    run_server(args.port, open_browser=not args.no_browser)


if __name__ == "__main__":
    main()
