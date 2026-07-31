#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | INVOICE TRUE-UP - the monthly invoice, proven line by line
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 31 Jul 2026): the monthly-stream invoice (radios, gas
#  monitors, welders, fridges, transport) lands as a Coates PDF and a
#  BASEPLAN_CHARGES export. Before it goes anywhere, prove it: every
#  line recomputed from its own dates, quantity and rate on the 7-day
#  week, tied against what Baseplan billed, the streams re-added, GST
#  checked, and the radio/gas fleet counts crossed against the rental
#  stock register. First run (INV 24955507) tied 16 of 16 lines to
#  the cent and found two radios on site the invoice wasn't charging.
#
#  This is the SEPARATE monthly stream - it never shares a dollar with
#  the SiteIQ daily invoice that 54_RUN_INVOICE_BREAKDOWN covers.
#
#  HOW (Andrew, 31 Jul 2026: "a set folder where i drop the invoices
#  and a set folder for the baseplan excel"): two folders live next to
#  the buttons and make themselves on first run -
#      Invoices\   <- the Coates invoice PDFs go here
#      Baseplan\   <- the BASEPLAN_CHARGES exports go here
#  Drop the files in, run 58, newest of each wins. Downloads still
#  works as a fallback for the Baseplan export, so muscle memory
#  doesn't break anything.
# =====================================================================
import datetime as dt
import glob
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
INV_DIR = os.path.join(HERE, 'Invoices')
BP_DIR = os.path.join(HERE, 'Baseplan')


def make_folders():
    for d, blurb in (
        (INV_DIR, 'DROP THE COATES INVOICE PDFs IN HERE\n'
                  '=====================================\n'
                  'One file per invoice, e.g. INV24955507.PDF - the number\n'
                  'in the filename names the true-up. Newest file wins.\n'
                  'Then double-click 58_RUN_INVOICE_TRUEUP.bat.\n'),
        (BP_DIR, 'DROP THE BASEPLAN CHARGES EXPORTS IN HERE\n'
                 '==========================================\n'
                 'The charge-lines export out of Baseplan, as .xlsx.\n'
                 'Any name is fine in this folder - newest file wins.\n'
                 'Then double-click 58_RUN_INVOICE_TRUEUP.bat.\n')):
        os.makedirs(d, exist_ok=True)
        rd = os.path.join(d, '_READ_ME_FIRST.txt')
        if not os.path.isfile(rd):
            try:
                with open(rd, 'w', encoding='utf-8') as f:
                    f.write(blurb)
            except OSError:
                pass


def find_baseplan():
    """Newest charges export: the Baseplan folder first (any .xlsx),
    then the old haunts by the BASEPLAN_CHARGES name."""
    hits = [p for p in glob.glob(os.path.join(BP_DIR, '*.xlsx'))
            if not os.path.basename(p).startswith('~$')]
    if hits:
        return max(hits, key=os.path.getmtime)
    for d in (os.path.join(os.path.expanduser('~'), 'Downloads'),
              os.path.join(HERE, 'Data_SiteIQ'), HERE):
        if os.path.isdir(d):
            hits += [p for p in glob.glob(os.path.join(d, 'BASEPLAN_CHARGES*.xlsx'))
                     if not os.path.basename(p).startswith('~$')]
    return max(hits, key=os.path.getmtime) if hits else None


def _pdf_text(p):
    """The PDF's text, or '' if it won't read. pypdf rides INSIDE the
    suite (the pypdf\\ folder), so this works on a bare machine with
    nothing installed. BaseException on purpose: a half-installed pdf
    stack can die with a low-level panic that is not an Exception, and
    reading a PDF must never take the true-up down (seen 31 Jul 2026)."""
    try:
        sys.path.insert(0, HERE)
        import pypdf
        return ''.join(pg.extract_text() or ''
                       for pg in pypdf.PdfReader(p).pages)
    except (KeyboardInterrupt, SystemExit):
        raise
    except BaseException:
        return ''


def scan_invoices():
    """Every PDF in Invoices\\, read and sorted into its stream:
    's' = the one-line SiteIQ invoice (Served By: Siteiq Service),
    'm' = the monthly Baseplan one (radios, gas, welders, transport).
    Newest of each stream wins - old months can stay in the folder as
    the archive. pypdf's text order puts amounts BEFORE their labels,
    so the money regexes accept both ways round."""
    out = {'m': None, 's': None}
    pdfs = [p for p in glob.glob(os.path.join(INV_DIR, '*.pdf'))
            + glob.glob(os.path.join(INV_DIR, '*.PDF'))
            if not os.path.basename(p).startswith('~$')]
    for p in sorted(set(pdfs), key=os.path.getmtime, reverse=True):
        txt = _pdf_text(p)
        m = re.search(r'(\d{6,})', os.path.basename(p))
        number = m.group(1) if m else ''
        mn = re.search(r'Tax\s+Invoice\s+Number\s*(\d+)', txt)
        if mn:
            number = mn.group(1)
        total = ex = None
        for pat in (r'Total\s+Amount\s+Due\s*\$?([\d,]+\.\d{2})',
                    r'Invoice\s+Total\s*\$?([\d,]+\.\d{2})',
                    r'\$([\d,]+\.\d{2})\s*Invoice\s+Total'):
            mt = re.search(pat, txt)
            if mt:
                total = float(mt.group(1).replace(',', ''))
                break
        #  the ex-GST figure is the one that ties EXACTLY - Coates
        #  rounds GST at line level, so the inc-GST total can sit a few
        #  cents off a straight 10% and still be right
        for pat in (r'Price\s+Excluding\s+GST\s*\$?([\d,]+\.\d{2})',
                    r'\$([\d,]+\.\d{2})\s*Price\s+Excluding\s+GST'):
            mx = re.search(pat, txt)
            if mx:
                ex = float(mx.group(1).replace(',', ''))
                break
        kind = ('s' if re.search(r'SiteIQ\s+Billing|Siteiq\s+Service',
                                 txt, re.I) else 'm')
        period = None
        if kind == 's':
            md = re.search(r'(\d{2}/\d{2}/\d{4})\s+\d{2}:\d{2}\s+'
                           r'(\d{2}/\d{2}/\d{4})\s+(\d{2}):\d{2}', txt)
            if md:
                try:
                    d1 = dt.datetime.strptime(md.group(1), '%d/%m/%Y').date()
                    d2 = dt.datetime.strptime(md.group(2), '%d/%m/%Y').date()
                    #  the 5am cutover: a period ending 04:59 on the
                    #  31st covers charge days up to the 30th
                    if int(md.group(3)) < 5:
                        d2 -= dt.timedelta(days=1)
                    period = (d1, d2)
                except ValueError:
                    period = None
        if out[kind] is None:
            out[kind] = {'path': p, 'no': number, 'total': total,
                         'ex': ex, 'period': period}
    return out


def _dmkey(s):
    """dd/mm sorted by month then day - '05/08' comes AFTER '27/07'."""
    d, m = s.split('/')[:2]
    return (int(m), int(d))


def siteiq_days():
    """SiteIQ's own invoiced dollars, day by day, off the newest
    DAILY_SUMMARY export - read by the same code the cost snapshot
    trusts. Empty = no export on this machine."""
    try:
        sys.path.insert(0, HERE)
        import build_cost_snapshot as CS
        days, path, _asat = CS.read_daily_summary()
        return days, path
    except (KeyboardInterrupt, SystemExit):
        raise
    except BaseException:
        return {}, None


def fleet_counts():
    """Serialised radios and gas monitors on the rental stock register -
    the on-site truth the invoice quantities are checked against."""
    import openpyxl
    hits = []
    for d in (os.path.join(HERE, 'Data_SiteIQ'), HERE):
        if os.path.isdir(d):
            hits += [p for p in glob.glob(os.path.join(d, 'RENTAL_STOCK*.xlsx'))
                     if not os.path.basename(p).startswith('~$')]
    if not hits:
        return None, None
    wb = openpyxl.load_workbook(max(hits, key=os.path.getmtime),
                                read_only=True, data_only=True)
    ws = wb['RENTAL_STOCK'] if 'RENTAL_STOCK' in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    hd = {str(v or '').strip(): i for i, v in enumerate(next(rows))}
    ucol = hd.get('STORAGE_UNIT')
    radios = gas = 0
    for r in rows:
        u = str(r[ucol] or '').strip() if ucol is not None else ''
        if u == 'Radios':
            radios += 1
        elif u == 'Gas Monitors':
            gas += 1
    wb.close()
    return radios, gas


def main():
    print('=' * 66)
    print(' COATES | INVOICE TRUE-UP - the monthly invoice, proven')
    print('=' * 66)
    make_folders()
    bp_path = find_baseplan()
    if not bp_path:
        print(' No charges export found. Drop the Baseplan export into:')
        print('   ' + BP_DIR)
        print(' and run me again. (Downloads works too.)')
        return 1
    invs = scan_invoices()
    minv, sinv = invs['m'], invs['s']
    inv_path = minv['path'] if minv else None
    inv_no = minv['no'] if minv else ''
    inv_total = minv['total'] if minv else None
    print(' Charges  : ' + bp_path)
    for tag, iv in (('Monthly', minv), ('SiteIQ', sinv)):
        if iv:
            print(' {:8s}: {}{}'.format(
                tag, os.path.basename(iv['path']),
                '  (total ${:,.2f} read from the PDF)'.format(iv['total'])
                if iv['total'] else ''))
    if not minv and not sinv:
        print(' Invoice  : none in ' + INV_DIR + ' - true-up still runs;')
        print('            drop the PDFs there and the report names itself.')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(bp_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hd = {str(v or '').strip(): i for i, v in enumerate(rows[0])}
    need = ['Line', 'Description', 'Quantity', 'Rate 1', 'Billed Amount',
            'Billed Units', 'Start Date', 'Billed To Date']
    missing_cols = [c for c in need if c not in hd]
    if missing_cols:
        print(' That export is missing columns I need: '
              + ', '.join(missing_cols))
        print(' Export the charge lines again with the standard layout.')
        return 1

    lines = []
    for r in rows[1:]:
        def g(k):
            return r[hd[k]] if k in hd else None
        ln = str(g('Line') or '').strip()
        if not ln:
            continue
        try:
            qty = float(g('Quantity') or 0)
            rate = float(g('Rate 1') or 0)
            billed = float(g('Billed Amount') or 0)
            units = float(g('Billed Units') or 0)
        except (TypeError, ValueError):
            continue
        start, to = g('Start Date'), g('Billed To Date')
        days = None
        if isinstance(start, dt.datetime) and isinstance(to, dt.datetime):
            days = (to.date() - start.date()).days + 1
        item = str(g('Item') or '').strip()
        lines.append({'ln': int(float(ln)), 'desc': str(g('Description') or '').strip(),
                      'item': item, 'qty': qty, 'rate': rate, 'billed': billed,
                      'units': units, 'days': days,
                      'start': start.date().strftime('%d/%m') if isinstance(start, dt.datetime) else '',
                      'to': to.date().strftime('%d/%m') if isinstance(to, dt.datetime) else ''})
    if not lines:
        print(' No charge lines in that export - nothing to true up.')
        return 1
    lines.sort(key=lambda x: x['ln'])

    ties = checks = 0
    hire_total = other_total = 0.0
    runrate = 0.0
    for L in lines:
        if L['rate'] > 0 and L['days']:
            L['rec'] = round(L['qty'] * L['days'] * L['rate'], 2)
            L['status'] = 'TIES' if abs(L['rec'] - L['billed']) < 0.005 else 'CHECK'
            #  Baseplan's own billed-days column gets a look too - a
            #  days drift with a matching total would mean rate drift
            if L['units'] and abs(L['units'] - L['days']) > 0.01 and L['status'] == 'TIES':
                L['status'] = 'CHECK'
                L['note'] = 'billed units {} v {} days by the calendar'.format(
                    int(L['units']), L['days'])
            hire_total += L['billed']
            runrate += L['qty'] * L['rate']
        else:
            #  flat lines - transport, ice, MISC. No dates to recompute
            #  from; they tie to the agreed amount by definition and are
            #  listed so nothing rides through unread.
            L['rec'] = None
            L['status'] = 'FLAT'
            other_total += L['billed']
        if L['status'] == 'TIES':
            ties += 1
        elif L['status'] == 'CHECK':
            checks += 1
    ex_gst = round(hire_total + other_total, 2)
    gst = round(ex_gst * 0.10, 2)

    #  the fleet cross-check: register v billed quantities
    reg_radios, reg_gas = fleet_counts()
    def _blq(word):
        return sum(int(L['qty']) for L in lines if word in L['desc'].upper()
                   and L['rate'] > 0)
    bill_radios, bill_gas = _blq('RADIO'), _blq('GAS')

    out_wb = openpyxl.Workbook()
    ORANGE, INK, GREY = 'F26222', '1D1D1B', '8A94A2'
    GREEN, RED = '2BB673', 'E23B2E'
    thin = Border(bottom=Side(style='thin', color='DDDDDD'))
    hf = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill('solid', fgColor=ORANGE)
    mono = Font(name='Consolas', size=10)

    def brand(w, title, sub):
        w['A1'] = 'COATES | ' + title
        w['A1'].font = Font(bold=True, size=16, color=INK)
        w['A2'] = ('Cement Australia K2 Shutdown 2026 - Gladstone | POWERED '
                   'BY SITEIQ | Author: Andrew Fisher')
        w['A2'].font = Font(size=9, color=GREY)
        w['A3'] = sub
        w['A3'].font = Font(size=10, italic=True, color=INK)

    def header(w, r, cols):
        for i, (lab, wd) in enumerate(cols, 1):
            c = w.cell(row=r, column=i, value=lab)
            c.font = hf
            c.fill = hfill
            w.column_dimensions[get_column_letter(i)].width = wd
        w.freeze_panes = w.cell(row=r + 1, column=1)

    w1 = out_wb.active
    w1.title = 'TRUE-UP'
    verdict = ('TIES TO THE CENT, {} of {} recomputable lines'.format(ties, ties + checks)
               if not checks else
               '{} LINE(S) NEED A LOOK - marked CHECK in red'.format(checks))
    brand(w1, 'MONTHLY INVOICE TRUE-UP' + (' - INV ' + inv_no if inv_no else ''),
          'Built {} off {} - every dated line recomputed from its own dates, '
          'quantity and rate on the 7-day week. Verdict: {}.'.format(
              dt.datetime.now().strftime('%d %b %Y %H:%M'),
              os.path.basename(bp_path), verdict))
    header(w1, 5, [('LINE', 6), ('DESCRIPTION', 44), ('ITEM', 15), ('QTY', 7),
                   ('FROM', 9), ('TO', 9), ('DAYS', 7), ('RATE', 10),
                   ('BILLED $', 12), ('RECOMPUTED $', 13), ('VERDICT', 12),
                   ('NOTE', 34)])
    r = 6
    for L in lines:
        vals = [L['ln'], L['desc'], L['item'], L['qty'], L['start'], L['to'],
                L['days'] if L['days'] else '-', L['rate'] if L['rate'] else '-',
                L['billed'], L['rec'] if L['rec'] is not None else '-']
        for i, v in enumerate(vals, 1):
            c = w1.cell(row=r, column=i, value=v)
            c.border = thin
            if i in (8, 9, 10) and isinstance(v, float):
                c.number_format = '#,##0.00'
            if i == 3:
                c.font = mono
        vc = w1.cell(row=r, column=11, value=L['status'])
        vc.font = Font(bold=True, color=GREEN if L['status'] == 'TIES'
                       else (GREY if L['status'] == 'FLAT' else RED))
        vc.border = thin
        w1.cell(row=r, column=12, value=L.get('note', '')).border = thin
        r += 1
    r += 1
    should = round(ex_gst + gst, 2)
    total_rows = [('Hire charges (dated lines)', round(hire_total, 2)),
                  ('Other charges (flat lines)', round(other_total, 2)),
                  ('Price ex GST', ex_gst), ('GST (10%)', gst),
                  ('THE INVOICE SHOULD TOTAL', should)]
    if inv_total is not None:
        total_rows.append(('PRINTED INVOICE TOTAL (read from the PDF)',
                           inv_total))
    for lab, val in total_rows:
        w1.cell(row=r, column=2, value=lab).font = Font(bold=True)
        c = w1.cell(row=r, column=9, value=val)
        c.number_format = '#,##0.00'
        c.font = Font(bold=True, color=ORANGE if 'SHOULD' in lab else INK)
        if 'PRINTED' in lab:
            ok = abs(inv_total - should) < 0.01
            vc = w1.cell(row=r, column=11,
                         value='TIES' if ok else 'CHECK')
            vc.font = Font(bold=True, color=GREEN if ok else RED)
        r += 1
    w1.cell(row=r + 1, column=2,
            value='Check the bottom line against the printed PDF. Daily '
                  'run-rate of the dated lines still on hire: ${:,.2f} ex GST.'
                  .format(runrate)).font = Font(italic=True, size=10)

    #  ---- the SITEIQ sheet: the other stream's one-line invoice ------
    #  (Andrew, 31 Jul 2026: "next one is this invoice this is the
    #  siteiq one... where can i drop this one" - same folder, and the
    #  true-up adds SiteIQ's own day-by-day invoiced dollars for the
    #  printed period, so the one line on the PDF gets its workings.)
    sq_note = None
    days, ds_path = siteiq_days()
    ws3 = out_wb.create_sheet('SITEIQ INVOICE')
    s_title = ('SITEIQ INVOICE TRUE-UP'
               + (' - INV ' + sinv['no'] if sinv and sinv['no'] else ''))
    if not days:
        brand(ws3, s_title,
              'No DAILY_SUMMARY export found on this machine - pull the '
              'morning exports and run me again to build this sheet.')
        sq_note = ('SITEIQ SHEET DID NOT BUILD: no DAILY_SUMMARY export '
                   'found. Pull the morning exports and re-run 58.')
    else:
        period = sinv['period'] if sinv else None
        if period:
            span = [d for d in sorted(days) if period[0] <= d <= period[1]]
            missing = [period[0] + dt.timedelta(days=i)
                       for i in range((period[1] - period[0]).days + 1)
                       if period[0] + dt.timedelta(days=i) not in days]
            sub = ('The printed period {} to {} (5am cutover applied), '
                   'day by day off {}.').format(
                       period[0].strftime('%d/%m'),
                       period[1].strftime('%d/%m'),
                       os.path.basename(ds_path))
        else:
            span, missing = sorted(days), []
            sub = ('No SiteIQ invoice PDF (or no period read from it) - '
                   'every invoiced day off {}, with the running total to '
                   'hold against the printed one-liner.').format(
                       os.path.basename(ds_path))
        brand(ws3, s_title, sub)
        header(ws3, 5, [('CHARGE DAY', 14), ('SITEIQ INVOICED $', 18),
                        ('RUNNING TOTAL $', 18), ('NOTE', 40)])
        r3, run = 6, 0.0
        for d in span:
            run += days[d]['total']
            ws3.cell(row=r3, column=1,
                     value=d.strftime('%a %d %b')).border = thin
            c = ws3.cell(row=r3, column=2, value=round(days[d]['total'], 2))
            c.number_format = '#,##0.00'
            c.border = thin
            c2 = ws3.cell(row=r3, column=3, value=round(run, 2))
            c2.number_format = '#,##0.00'
            c2.border = thin
            ws3.cell(row=r3, column=4, value='').border = thin
            r3 += 1
        for d in missing:
            ws3.cell(row=r3, column=1,
                     value=d.strftime('%a %d %b')).border = thin
            mc = ws3.cell(row=r3, column=4,
                          value='NOT IN THE EXPORT - pull fresh exports')
            mc.font = Font(bold=True, color=RED)
            mc.border = thin
            r3 += 1
        s_ex = round(run, 2)
        r3 += 1
        #  the tie is on EX GST - the exact number both sides print.
        #  Coates rounds GST at line level, so inc-GST can sit cents
        #  off a straight 10% while being perfectly right.
        p_ex = sinv.get('ex') if sinv else None
        s_rows = [('THE INVOICED DAYS ADD TO (ex GST)', s_ex)]
        if p_ex is not None:
            s_rows.append(('PRINTED PRICE EX GST (read from the PDF)', p_ex))
        if sinv and sinv['total'] is not None:
            s_rows.append(('PRINTED INVOICE TOTAL inc GST', sinv['total']))
        for lab, val in s_rows:
            ws3.cell(row=r3, column=1, value=lab).font = Font(bold=True)
            c = ws3.cell(row=r3, column=3, value=val)
            c.number_format = '#,##0.00'
            c.font = Font(bold=True,
                          color=ORANGE if 'ADD TO' in lab else INK)
            if 'PRINTED PRICE' in lab:
                s_ok = not missing and abs(p_ex - s_ex) < 0.01
                vc = ws3.cell(row=r3, column=4,
                              value='TIES' if s_ok else 'CHECK')
                vc.font = Font(bold=True, color=GREEN if s_ok else RED)
            r3 += 1
        s_verdict, s_gap = None, 0.0
        s_period_str = ('{} to {}'.format(period[0].strftime('%d/%m'),
                                          period[1].strftime('%d/%m'))
                        if period else 'all invoiced days')
        if p_ex is not None:
            gap = round(p_ex - s_ex, 2)
            s_gap = gap
            s_verdict = ('CHECK' if (missing or abs(gap) >= 0.01)
                         else 'TIES')
            if missing:
                sq_note = ('SITEIQ INVOICE: {} charge day(s) of the printed '
                           'period are not in the DAILY_SUMMARY export - '
                           'pull fresh exports and re-run before calling '
                           'the ${:,.2f} gap real.'.format(
                               len(missing), abs(gap)))
            elif abs(gap) >= 0.01:
                sq_note = ('SITEIQ INVOICE DOES NOT MATCH: the PDF says '
                           '${:,.2f} ex GST, the day-by-day invoiced '
                           'dollars compute ${:,.2f} - a ${:,.2f} gap to '
                           'chase.'.format(p_ex, s_ex, abs(gap)))
            else:
                sq_note = ('SITEIQ INVOICE: ties - the one-line ${:,.2f} '
                           'ex GST is exactly the {} invoiced days added '
                           'up.'.format(p_ex, len(span)))

    w2 = out_wb.create_sheet('CHECKS')
    brand(w2, 'THINGS WORTH AN EYES-ON',
          'The questions this true-up leaves on the table.')
    header(w2, 5, [('#', 4), ('CHECK', 110)])
    notes = []
    if reg_radios is not None:
        for word, reg, bl in (('radios', reg_radios, bill_radios),
                              ('gas monitors', reg_gas, bill_gas)):
            if not bl:
                continue
            if reg == bl:
                notes.append('FLEET - {}: register {} v billed {} - ties '
                             'exactly.'.format(word, reg, bl))
            else:
                notes.append('FLEET - {}: the rental stock register carries '
                             '{} serialised units, the invoice bills {}. '
                             '{} unit(s) {} - free-of-charge spares, later '
                             'additions, or a billing miss: confirm with the '
                             'branch before the final invoice.'.format(
                                 word, reg, bl, abs(reg - bl),
                                 'on site not being charged' if reg > bl
                                 else 'billed beyond the register'))
    else:
        notes.append('FLEET: no RENTAL_STOCK export found, so the radio/gas '
                     'register cross-check did not run. Pull the morning '
                     'exports and run me again for the full check.')
    notes.append('PROGRESS INVOICE: dated lines keep charging at '
                 '${:,.2f} ex GST per day until off-hired - the final '
                 'invoice washes up the rest.'.format(runrate))
    if sq_note:
        notes.insert(0, sq_note)
    if inv_total is not None and abs(inv_total - should) >= 0.01:
        notes.insert(0, 'THE PRINTED MONTHLY INVOICE DOES NOT MATCH: the PDF '
                        'says ${:,.2f}, the charge lines compute to ${:,.2f} '
                        '- a ${:,.2f} gap to chase before anything is paid '
                        'or passed on.'.format(inv_total, should,
                                               abs(inv_total - should)))
    elif inv_total is None and not inv_path:
        notes.append('NO MONTHLY INVOICE PDF in the Invoices folder - the '
                     'computed total stands alone. Drop the PDF in and '
                     're-run to have it named and tied automatically.')
    if checks:
        notes.insert(0, '{} LINE(S) MARKED CHECK on the TRUE-UP sheet - the '
                        'recomputed dollars or days do not match what was '
                        'billed. Read those rows first.'.format(checks))
    r = 6
    for i, txt in enumerate(notes, 1):
        w2.cell(row=r, column=1, value=i).font = Font(bold=True, color=ORANGE)
        c = w2.cell(row=r, column=2, value=txt)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        w2.row_dimensions[r].height = 56
        r += 1

    nos = '_'.join('INV' + iv['no'] for iv in (minv, sinv)
                   if iv and iv['no'])
    out = os.path.join(HERE, 'Invoice_TrueUp_{}{}.xlsx'.format(
        nos + '_' if nos else '', dt.date.today().isoformat()))
    out_wb.save(out)

    #  ---- the invoice register: how this ties into the daily reports.
    #  Each proven invoice is logged here, and the morning cost
    #  snapshot's true-up section shows the register - so the daily
    #  pack says not just "tracked v invoiced" but "these are the
    #  actual invoices issued, and they tie" (Andrew, 31 Jul 2026).
    #  Lives in Invoices\ - update-proof, and each machine keeps its
    #  own. Best effort: a register hiccup never fails the true-up.
    try:
        import json
        reg_p = os.path.join(INV_DIR, 'invoice_register.json')
        reg = {}
        if os.path.isfile(reg_p):
            with open(reg_p, encoding='utf-8') as f:
                reg = json.load(f)
        dated = [L for L in lines if L['start']]
        m_verdict = ('CHECK' if checks or (
            inv_total is not None and abs(inv_total - should) >= 0.01)
            else 'TIES')
        if inv_no or minv:
            reg[inv_no or os.path.basename(bp_path)] = {
                'kind': 'Monthly (radios, gas, plant, transport)',
                'period': ('{} to {}'.format(
                    min((L['start'] for L in dated), key=_dmkey),
                    max((L['to'] for L in dated), key=_dmkey))
                    if dated else ''),
                'ex': ex_gst, 'inc': round(should, 2),
                'verdict': m_verdict,
                'gap': round((inv_total - should), 2) if inv_total else 0.0,
                'run': dt.date.today().isoformat()}
        if sinv and sinv['no'] and days and s_verdict:
            reg[sinv['no']] = {
                'kind': 'SiteIQ (the one-line invoice)',
                'period': s_period_str,
                'ex': s_ex,
                'inc': sinv['total'] if sinv['total'] else round(s_ex * 1.1, 2),
                'verdict': s_verdict, 'gap': s_gap,
                'run': dt.date.today().isoformat()}
        with open(reg_p, 'w', encoding='utf-8') as f:
            json.dump(reg, f, indent=1, sort_keys=True)
        print(' Register : {} invoice(s) logged - tomorrow\'s cost snapshot'
              .format(len(reg)))
        print('            shows them on the true-up page.')
    except (KeyboardInterrupt, SystemExit):
        raise
    except BaseException:
        pass
    print(' Lines    : {} ({} tie, {} to check, {} flat)'.format(
        len(lines), ties, checks, len(lines) - ties - checks))
    print(' Should   : ${:,.2f} inc GST'.format(should))
    if inv_total is not None:
        print(' Printed  : ${:,.2f} - {}'.format(
            inv_total, 'TIES' if abs(inv_total - should) < 0.01
            else 'DOES NOT MATCH - read the CHECKS sheet'))
    if sq_note:
        print(' SiteIQ   : ' + sq_note.split(':', 1)[-1].strip()[:58])
    print('')
    print(' Written to : ' + out)
    print('')
    print(' Check the SHOULD-TOTAL against the printed invoice PDF, read')
    print(' the CHECKS sheet, and anything marked CHECK in red first.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
