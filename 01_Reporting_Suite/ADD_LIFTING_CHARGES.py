#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | ADD LIFTING CHARGES - the approved slings & shackles
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 28 Jul 2026): lifting consumables sold to the client -
#  17 lines of round slings and shackles, approved by David Moore,
#  going into SiteIQ as ONE service charge called "Lifting Consumables"
#  for the confirmed total of $2,825.11.
#
#  This records the itemised paper trail in the charge register: one
#  register row per line (same pattern as every consumable charge),
#  all referenced "Lifting Consumables", so the register's 17 lines
#  add to exactly the single figure Andrew keys into SiteIQ.
#
#  The register is backed up first. Safe to run twice - if the Lifting
#  Consumables lines are already in, nothing is added again.
# =====================================================================
import datetime as dt
import os
import shutil
import sys
import time

HERE = os.path.dirname(os.path.abspath(__file__))
REG = os.path.join(HERE, "Coates_K2_Charge_Reporter", "CHARGE_REGISTER.xlsx")
BK = os.path.join(HERE, "Updates", "charge_backups")

BATCH_REF = "Lifting Consumables"
APPROVED_BY = "David Moore"
CHARGE_DATE = dt.date(2026, 7, 28)

#  Quantities, unit costs and line totals exactly as approved. The line
#  totals are the authoritative figures (unit prices shown are rounded
#  display values) - they sum to $2,825.11, confirmed 28 Jul 2026.
LINES = [
    ("Round Sling 1T Violet 1M", 38, 3.99, 151.63),
    ("Round Sling 1T Violet 2M", 22, 7.25, 159.50),
    ("Round Sling 1T Violet 3M", 11, 10.79, 118.66),
    ("Round Sling 1T Violet 4M", 10, 14.26, 142.63),
    ("Round Sling 2T Green 1M", 30, 6.25, 187.50),
    ("Round Sling 2T Green 2M", 6, 11.69, 70.13),
    ("Round Sling 2T Green 3M", 8, 17.72, 141.79),
    ("Round Sling 2T Green 4M", 7, 24.18, 169.28),
    ("Round Sling 3T Yellow 1M", 14, 8.64, 120.93),
    ("Round Sling 3T Yellow 2M", 8, 16.19, 129.49),
    ("Round Sling 3T Yellow 3M", 10, 23.88, 238.78),
    ("Round Sling 3T Yellow 4M", 7, 32.36, 226.49),
    ("Round Sling 3T Yellow 6M", 9, 44.46, 400.10),
    ("Shackles 2T", 20, 3.14, 62.80),
    ("Shackles 3.2T", 20, 5.50, 110.00),
    ("Shackles 4.7T", 20, 8.17, 163.40),
    ("Shackles 6T", 20, 11.60, 232.00),
]
TOTAL = round(sum(t for _d, _q, _u, t in LINES), 2)


def main():
    print("=" * 70)
    print(" COATES | ADD LIFTING CHARGES - slings & shackles, approved")
    print("=" * 70)
    try:
        import openpyxl
    except ImportError:
        print(" openpyxl isn't installed - run:  py -m pip install openpyxl")
        return 1
    if not os.path.isfile(REG):
        print(" Can't find the charge register - run me from the suite "
              "folder.")
        return 1
    wb = openpyxl.load_workbook(REG)
    if "Consumables sold" not in wb.sheetnames:
        print(" No 'Consumables sold' sheet in the register - nothing "
              "changed. Tell Claude.")
        return 1
    ws = wb["Consumables sold"]

    #  Idempotence: the batch reference is the fingerprint. Any existing
    #  row carrying it means this has already been recorded.
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(BATCH_REF.lower() in str(c or "").lower() for c in row):
            print("")
            print(" The Lifting Consumables lines are already in the "
                  "register.")
            print(" Nothing added - that is the point: running me twice "
                  "can never double-charge.")
            print(" The confirmed total remains ${:,.2f}.".format(TOTAL))
            return 0

    #  Next free CON id for the charge date.
    tag = CHARGE_DATE.strftime("%Y%m%d")
    used = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        cid = str(row[0] or "")
        if cid.startswith("CON-" + tag + "-"):
            try:
                used.add(int(cid.rsplit("-", 1)[1]))
            except ValueError:
                pass
    nxt = (max(used) + 1) if used else 1

    os.makedirs(BK, exist_ok=True)
    bpath = os.path.join(BK, "{}_pre_lifting_CHARGE_REGISTER.xlsx".format(
        time.strftime("%Y%m%d_%H%M%S")))
    shutil.copy2(REG, bpath)

    feed = wb["Costing Feed"] if "Costing Feed" in wb.sheetnames else None
    ids = []
    for i, (desc, qty, unit, total) in enumerate(LINES):
        cid = "CON-{}-{:03d}".format(tag, nxt + i)
        ids.append(cid)
        #  REAL date objects, never text - the register's older rows are
        #  Excel dates, and one text date in the same column crashed the
        #  whole 29 Jul morning build (str and date can't sort together).
        ws.append([cid, "Cement Australia",
                   dt.datetime(CHARGE_DATE.year, CHARGE_DATE.month,
                               CHARGE_DATE.day), "14:00",
                   "Andrew Fisher", "", BATCH_REF + " - K2",
                   desc, qty, "Each", unit, total])
        if feed is not None:
            feed.append([dt.datetime(CHARGE_DATE.year, CHARGE_DATE.month,
                                     CHARGE_DATE.day), cid,
                         "Consumables sold", "Consumables Sales",
                         "Cement Australia",
                         BATCH_REF + " - " + desc, total, "Approved",
                         APPROVED_BY, "", dt.datetime.now()])
    wb.save(REG)

    print("")
    print(" RECORDED - {} line(s), {} to {}.".format(
        len(LINES), ids[0], ids[-1]))
    print(" Client: Cement Australia | Approved by: {}".format(APPROVED_BY))
    print(" Register lines sum to exactly:  ${:,.2f}".format(TOTAL))
    print("")
    print(" That is the ONE figure for SiteIQ - a service charge named")
    print(" 'Lifting Consumables' for ${:,.2f}. The register carries the".format(TOTAL))
    print(" 17-line detail so the invoice and the paper trail always")
    print(" reconcile. Replaced register backed up to:")
    print("   " + os.path.relpath(bpath, HERE))
    return 0


if __name__ == "__main__":
    sys.exit(main())
