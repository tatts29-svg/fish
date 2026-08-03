#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | UTILISATION CONTROL - Andrew's hub, the page-2 design
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 3 Aug 2026): "where is my area with all my graphs.
#  visuals. time line of the shut. I only see print hub. Fleet entry.
#  open the counter entry ... come on mate. gotta do better than that."
#
#  Fair cop. His design page 2 - the Utilisation Control screen with
#  the date range, the performance bars, the tiles, the attention list
#  and the category bars - was shown, agreed, and then never built.
#  This is it, off the same engines as everything else:
#
#    shutdown_day    the clock and the flame-off date
#    mygear_intel    the register, days, cycles
#    fleet_detail    the utilisation definitions (never a second copy)
#    asset_facts     USE NEXT picks, serials, per-fleet rollups
#    whats_used      the rates, for $/day on charge
#    TRANSACTIONS    the day-by-day timeline of the shut
#
#  COATES INTERNAL. It carries money, so it writes to Reports\ with
#  the other internal pages and is never linked from anything that
#  goes on the store Wi-Fi. Its doors are the other internal pages,
#  which sit in the same folder.
# =====================================================================
import datetime as dt
import html
import io
import os
from collections import Counter

import asset_facts as AF
import forecast as FC
import fleet_detail as FD
import mygear_intel as MI
import serials as SR
import shutdown_day as SD
import whats_used as WU

BASE = os.path.dirname(os.path.abspath(__file__))


def esc(s):
    return html.escape('' if s is None else str(s), quote=True)


def _timeline(txn_path, start, end):
    """Issues and returns per day, straight off the charge feed.

    The timeline is the shut's heartbeat - flat days and big days are
    both worth seeing. Counted from TRAN_START_DATE (gear going out)
    and TRAN_END_DATE (gear coming home); a missing end date is an
    open hire, not a return, so it counts nowhere.
    """
    import openpyxl

    def au(v):
        """SiteIQ writes '31/07/2026' as TEXT. The first cut of this
        read sheet zero (REFERENCE_INFO) with isinstance(datetime)
        checks and produced a timeline of one movement - a flat line
        that looked like a dead shut. Parse the string, and read the
        sheets that actually carry transactions."""
        if isinstance(v, dt.datetime):
            return v.date()
        if isinstance(v, dt.date):
            return v
        s = str(v or '').strip()
        if '/' in s:
            try:
                d0, m0, y0 = s.split(' ')[0].split('/')
                return dt.date(int(y0), int(m0), int(d0))
            except (ValueError, IndexError):
                return None
        return None

    outs, ins = Counter(), Counter()
    wb = openpyxl.load_workbook(txn_path, read_only=True, data_only=True)
    for sn in ('TRANSACTION_CHARGES', 'TRANSACTION_WITHOUT_CHARGES',
               'CUSTOMER_CONTRACTOR_EQUIP'):
        if sn not in wb.sheetnames:
            continue
        rows = wb[sn].iter_rows(values_only=True)
        try:
            hdr = [str(c or '').strip() for c in next(rows)]
        except StopIteration:
            continue
        ix = {h: i for i, h in enumerate(hdr)}
        si, ei = ix.get('TRAN_START_DATE'), ix.get('TRAN_END_DATE')
        for r in rows:
            if not r:
                continue
            for col, book in ((si, outs), (ei, ins)):
                if col is None or col >= len(r):
                    continue
                d = au(r[col])
                if d and start <= d <= end:
                    book[d] += 1
    wb.close()
    days = []
    d = start
    while d <= end:
        days.append((d, outs.get(d, 0), ins.get(d, 0)))
        d += dt.timedelta(days=1)
    return days


def _bar(pct, cls=''):
    w = max(1, min(100, int(pct + 0.5)))
    return ("<div class='bar {c}'><i style='width:{w}%'></i>"
            "<b>{p}%</b></div>").format(c=cls, w=w, p=int(pct + 0.5))


def build(today=None):
    today = today or dt.date.today()
    rental = FC._newest('RENTAL_STOCK*.xlsx')
    txn = FC._newest('TRANSACTIONS*.xlsx')
    onhire = FC._newest('ON_HIRE*.xlsx')
    if not (rental and txn):
        print(' No RENTAL_STOCK / TRANSACTIONS export - nothing to build.')
        return 1
    data = MI.read(rental, txn, onhire)
    assets = list(data['assets'].values())
    af = AF.build(data)
    af.pop('iv', None)

    onsite = len(assets)
    out_now = sum(1 for a in assets if a.get('status') == MI.OUT_STATUS)
    ready = sum(1 for a in assets
                if (a.get('status') or '').lower() == 'available for hire')
    never = sum(1 for a in assets if not (a.get('cycles') or 0)
                and a.get('status') != MI.OUT_STATUS)

    #  the two utilisations, whole-store, same maths as Fleet Details:
    #  day-weighted so a spanner out an hour cannot equal a tower light
    #  out a fortnight
    span = None
    cl = cm = 0.0
    for a in assets:
        cl += a.get('clientDays') or 0.0
        cm += a.get('commercialDays') or 0.0
    try:
        f0 = dt.date.fromisoformat(data.get('sourceFrom'))
        t0 = dt.date.fromisoformat(data.get('sourceTo'))
        span = max(1, (t0 - f0).days + 1)
    except Exception:
        span = None
    #  Two bars, two honest denominators, each named in words on the
    #  page. The design mock's 64% is the first one: of the days gear
    #  sat on charge, the share a crew actually held it. The second is
    #  the whole-store load: of every asset-day on site, how many were
    #  charged at all - low, truthfully, because most of the register
    #  is tooling that has never been issued.
    client_pct = 100.0 * cl / cm if cm else 0.0
    denom = float(span or 1) * max(1, onsite)
    comm_pct = 100.0 * cm / denom

    #  money per day on charge - the manager number, same rates the
    #  crew page's manager layer uses
    #  same exclusions as the crew page's manager layer - tracked and
    #  client-owned gear carries no figure ANYWHERE, Andrew's 2 Aug
    #  rule, so this tile lands on the same dollar the phone shows
    import ownership as OWN
    seqs = OWN.zero_cost_sequences(assets)
    rates = WU.load_rates(BASE, txn_path=txn)
    per_day = 0.0
    unpriced = 0
    for a in assets:
        if a.get('status') != MI.OUT_STATUS:
            continue
        if OWN.stream(a, seqs)[0] in ('COATES_TRACKED', 'CLIENT'):
            continue
        r, _src = WU._rate_of(rates, a)
        if r:
            per_day += r
        else:
            unpriced += 1

    #  attention list - each line is real, current and actionable
    rotate = sum(1 for v in af['variant'].values()
                 if v.get('w') == 'USE NEXT' and (v.get('rd') or 0) > 0
                 and (v.get('out') or 0) > 0)
    ser_cov = SR.coverage(a.get('item') or '' for a in assets)
    all_out = sum(1 for v in af['variant'].values()
                  if (v.get('rd') or 0) == 0 and (v.get('out') or 0) > 0)

    #  category bars: client-issued share by aisle, day-weighted
    units = {}
    for a in assets:
        u = a.get('unit') or 'Unfiled'
        e = units.setdefault(u, [0.0, 0])
        e[0] += a.get('clientDays') or 0.0
        e[1] += 1
    cat = sorted(((u, 100.0 * d / (float(span or 1) * max(1, n)), n)
                  for u, (d, n) in units.items() if n >= 20),
                 key=lambda x: -x[1])[:8]

    #  the timeline, flame-off to today
    start = SD.FLAME_OFF - dt.timedelta(days=12)
    tl = _timeline(txn, start, today)
    peak = max([max(o, i) for _, o, i in tl] or [1]) or 1

    H = []
    H.append("<div class='wrap'>")
    H.append("<div class='bar0'><h1>Utilisation Control</h1>"
             "<span class='siq'>POWERED BY SITEIQ</span></div>")
    H.append("<div class='crumbs'>"
             "<a href='Coates_K2_Utilisation_Intelligence_{d}.html'>"
             "Utilisation Intelligence</a>"
             "<a href='Coates_K2_Money_And_Whats_Used_{d}.html'>"
             "Money &amp; What&rsquo;s Used</a>"
             "<a href='Coates_K2_Fleet_Details_{d}.html'>Fleet Details"
             "</a></div>".format(d=today.isoformat()))
    H.append("<p class='sub'>SHUTDOWN TO DATE &middot; {} &ndash; today "
             "&middot; <b>{}</b> &middot; built {}</p>".format(
                 SD.FLAME_OFF.strftime('%d %b'), esc(SD.label()),
                 today.strftime('%d %b %Y')))

    H.append("<div class='panel'><div class='ph'>Fleet performance</div>")
    H.append("<div class='pr'><span>OF CHARGED DAYS, CREW-HELD</span>"
             + _bar(client_pct, 'cy') + "</div>")
    H.append("<div class='pr'><span>OF ALL ASSET-DAYS, CHARGED</span>"
             + _bar(comm_pct, 'org') + "</div>")
    H.append("<div class='pnote'>Top bar: when gear was on charge, how "
             "much of that time a crew actually held it - the gap is "
             "paid-for gear sitting unused. Bottom bar: how hard the "
             "whole register is working - low is truthful, because most "
             "of the store is tooling that has never been issued."
             "</div></div>")

    H.append("<div class='tiles'>")
    for v, l, c in ((format(onsite, ','), 'Assets onsite', 'cy'),
                    (format(out_now, ','), 'Out with crews now', 'org'),
                    (format(never, ','), 'Never issued', 'am'),
                    ('${:,.0f}'.format(per_day), 'Per day on charge', 'org'),
                    (format(ready, ','), 'Ready to issue', 'gd'),
                    (format(unpriced, ','), 'On hire, no rate', 'am')):
        H.append("<div class='tile'><b class='{}'>{}</b><span>{}</span>"
                 "</div>".format(c, v, l))
    H.append("</div>")

    H.append("<div class='panel'><div class='ph'>Timeline of the shut "
             "&mdash; issues out, returns home</div>")
    H.append("<div class='tl'>")
    for d, o, i in tl:
        H.append("<div class='tld' title='{} - {} out, {} back'>"
                 "<i class='o' style='height:{}%'></i>"
                 "<i class='r' style='height:{}%'></i>{}</div>".format(
                     d.strftime('%d %b'), o, i,
                     max(2, int(100.0 * o / peak)),
                     max(2, int(100.0 * i / peak)),
                     "<em>{}</em>".format(d.day) if d.weekday() == 0
                     or d == SD.FLAME_OFF else ''))
    H.append("</div><div class='pnote'><i class='sw o'></i> out to a crew "
             "&nbsp; <i class='sw r'></i> returned &middot; flame off {}"
             " &middot; tallest day: {} movements</div></div>".format(
                 SD.FLAME_OFF.strftime('%d %b'), peak))

    H.append("<div class='panel'><div class='ph'>What needs attention</div>"
             "<div class='att'>")
    for n, txt, page in (
            (rotate, 'fleets have fresh gear on the shelf while worked '
             'units are out - rotate on next issue (USE NEXT on the '
             'counter screens)', None),
            (all_out, 'product lines are completely out with crews - '
             'nothing left to issue', None),
            (ser_cov['placeholder'], 'plant-numbered assets carry a '
             'placeholder instead of a serial - SERIALS_TO_FIX.csv, a '
             'Baseplan housekeeping job', None),
            (unpriced, 'assets on hire carry no rate on this export - '
             'invisible dollars until rated', None)):
        if n:
            H.append("<div class='ar'><b>{}</b><span>{}</span></div>"
                     .format(format(n, ','), txt))
    H.append("</div></div>")

    H.append("<div class='panel'><div class='ph'>Fleet by aisle &mdash; "
             "client-issued share</div>")
    for u, p, n in cat:
        H.append("<div class='pr'><span>{} ({})</span>{}</div>".format(
            esc(u.upper()), n, _bar(p, 'gd' if p >= 50 else
                                    ('am' if p >= 25 else 'cy'))))
    H.append("<div class='pnote'>Aisles with fewer than 20 assets are "
             "left off rather than shown as a spiky small-sample bar."
             "</div></div>")

    H.append("<div class='foot'>COATES INTERNAL &middot; carries revenue "
             "&middot; never on the store Wi-Fi, never to the client"
             "<br>Author: Andrew Fisher &middot; data to {}</div>"
             .format(esc(data.get('sourceTo') or today.isoformat())))
    H.append("</div>")

    page = ("<!doctype html><html lang='en'><head><meta charset='utf-8'>"
            "<meta name='viewport' content='width=device-width,"
            "initial-scale=1'><title>Utilisation Control - Coates K2"
            "</title><style>" + CSS + "</style></head><body>"
            + '\n'.join(H) + "</body></html>")

    out_dir = os.path.join(BASE, 'Reports', today.isoformat(), 'Pages')
    os.makedirs(out_dir, exist_ok=True)
    out = os.path.join(out_dir,
                       'Coates_K2_Utilisation_Control_{}.html'
                       .format(today.isoformat()))
    with io.open(out, 'w', encoding='utf-8') as fh:
        fh.write(page)
    print(' Written      : ' + out)
    print(' Client-issued: {:.0f}% | Commercial: {:.0f}% | ${:,.0f}/day '
          'on charge'.format(client_pct, comm_pct, per_day))
    print(' COATES INTERNAL - carries revenue. Not for the store Wi-Fi.')
    _open_for_andrew(out)
    return 0


def _open_for_andrew(path):
    """Windows only, and never when chained by 04."""
    if os.environ.get('K2_CHAINED'):
        return
    if hasattr(os, 'startfile'):
        try:
            os.startfile(path)
        except OSError:
            pass


CSS = """
*{margin:0;padding:0;box-sizing:border-box}
body{background:#0A0E14;color:#E9EEF5;
 font:400 14px/1.5 system-ui,Segoe UI,Roboto,sans-serif}
.wrap{max-width:760px;margin:0 auto;padding:0 12px 40px}
.bar0{display:flex;align-items:center;justify-content:space-between;
 background:#F26222;margin:0 -12px;padding:14px 16px}
.bar0 h1{font-size:19px;font-weight:900;color:#fff}
.siq{font-size:9px;font-weight:900;letter-spacing:2px;color:#FFD9C4}
.crumbs{display:flex;gap:8px;padding:12px 0 0;flex-wrap:wrap}
.crumbs a{color:#7FB1C8;text-decoration:none;font-size:11.5px;
 font-weight:800;border:1px solid #2A3340;border-radius:9px;padding:5px 10px}
.sub{color:#98A4B4;font-size:12px;margin:12px 0 14px;font-weight:700;
 letter-spacing:.4px}
.panel{background:#151A22;border:1px solid #2A3340;border-radius:14px;
 padding:14px;margin-bottom:12px}
.ph{font-size:11px;font-weight:900;letter-spacing:1.2px;color:#98A4B4;
 text-transform:uppercase;margin-bottom:10px}
.pr{display:flex;align-items:center;gap:10px;margin-bottom:8px}
.pr>span{flex:none;width:158px;font-size:10.5px;font-weight:800;
 letter-spacing:.5px;color:#C7CED8}
.bar{flex:1;display:flex;align-items:center;gap:8px}
.bar i{display:block;height:9px;border-radius:5px;background:#2BB673;
 min-width:2px;flex:none}
.bar b{font-size:13px;font-weight:900}
.bar.cy i{background:#2AA9C4}.bar.org i{background:#F26222}
.bar.am i{background:#F5A623}.bar.gd i{background:#2BB673}
.bar{position:relative}
.bar i{width:0}
.pnote{color:#6B7789;font-size:11px;margin-top:8px;line-height:1.5}
.tiles{display:grid;grid-template-columns:repeat(auto-fill,minmax(110px,1fr));
 gap:10px;margin-bottom:12px}
.tile{background:#151A22;border:1px solid #2A3340;border-radius:12px;
 padding:12px 10px;text-align:center}
.tile b{display:block;font-size:21px;font-weight:900}
.tile span{font-size:9.5px;font-weight:800;letter-spacing:.6px;
 color:#98A4B4;text-transform:uppercase}
.cy{color:#2AA9C4}.org{color:#F26222}.am{color:#F5A623}.gd{color:#2BB673}
.tl{display:flex;align-items:flex-end;gap:2px;height:110px;
 padding:4px 0 14px;overflow-x:auto}
.tld{flex:1;min-width:9px;height:100%;display:flex;align-items:flex-end;
 gap:1px;position:relative}
.tld i{flex:1;border-radius:2px 2px 0 0}
.tld i.o{background:#F26222}.tld i.r{background:#2BB673}
.tld em{position:absolute;bottom:-14px;left:0;right:0;text-align:center;
 font-size:8px;color:#6B7789;font-style:normal;font-weight:700}
.sw{display:inline-block;width:9px;height:9px;border-radius:2px;
 vertical-align:-1px}
.sw.o{background:#F26222}.sw.r{background:#2BB673}
.att .ar{display:flex;gap:12px;align-items:baseline;padding:7px 0;
 border-bottom:1px solid #1C232D;font-size:12.5px;line-height:1.5}
.att .ar:last-child{border-bottom:0}
.att .ar b{flex:none;min-width:44px;text-align:right;font-size:16px;
 font-weight:900;color:#F5A623}
.att .ar span{color:#C7CED8}
.foot{text-align:center;color:#5A6472;font-size:10.5px;margin-top:18px;
 line-height:1.7;letter-spacing:.4px}
@media print{body{background:#fff;color:#000}
 .panel,.tile{border-color:#bbb;background:#fff}
 .bar0{background:#fff;border-bottom:3px solid #F26222}
 .bar0 h1,.tile b,.att .ar b{color:#000}
 .crumbs{display:none}}
"""


if __name__ == '__main__':
    import sys
    sys.exit(build())
