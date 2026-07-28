#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | ARE WE IN SYNC? - one screen, one honest answer
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  THE QUESTION (Andrew, 27 Jul 2026): "the reports should always take
#  the updated data - how do I keep in sync?"
#
#  THE CHAIN. Everything flows one way, and each step reads the step
#  above it - so if the order is right, they cannot disagree:
#
#     SiteIQ exports  ->  Data_SiteIQ\  ->  the reports
#                                       ->  the workbook (via the align)
#
#  This checks the ORDER actually happened today: every export's own
#  "as at" time, when the reports were built, and when the workbook was
#  aligned. Anything built BEFORE the data it should have used gets
#  called out, because that is the only way these drift apart.
# =====================================================================
import datetime as dt
import glob
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
EXPORTS = ["RENTAL_STOCK", "ON_HIRE", "TRANSACTIONS", "SALES_STOCK",
           "STOCKTAKE", "DAILY_SUMMARY"]


def when_of(path):
    """The export's own REQUESTED_DATE/TIME - not the file date."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if "REFERENCE_INFO" not in wb.sheetnames:
                return None
            rows = []
            for i, r in enumerate(wb["REFERENCE_INFO"].iter_rows(values_only=True)):
                rows.append(r)
                if i >= 1:
                    break
            if len(rows) < 2:
                return None
            hdr = [str(v or "").strip().lower() for v in rows[0]]
            for i, h in enumerate(hdr):
                if "requested_date" in h and i < len(rows[1]):
                    v = rows[1][i]
                    if isinstance(v, dt.datetime):
                        return v
                    m = re.match(r"\s*(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{2})\s*([AP]M)?",
                                 str(v or ""), re.I)
                    if m:
                        d, mo, y, hh, mi, ap = m.groups()
                        hh = int(hh)
                        if ap and ap.upper() == "PM" and hh != 12:
                            hh += 12
                        if ap and ap.upper() == "AM" and hh == 12:
                            hh = 0
                        return dt.datetime(int(y), int(mo), int(d), hh, int(mi))
            return None
        finally:
            wb.close()
    except Exception:
        return None


def fmt(t):
    return t.strftime("%d %b %I:%M %p").replace(" 0", " ").lstrip("0") if t else "-"


def main():
    print("=" * 62)
    print(" COATES | ARE WE IN SYNC?")
    print("=" * 62)
    newest_data, problems = None, []

    print(" THE DATA - what SiteIQ says, and when it was pulled")
    for stem in EXPORTS:
        hits = []
        for d in (os.path.join(HERE, "Data_SiteIQ"), HERE):
            hits += [p for p in glob.glob(os.path.join(d, stem + "*.xlsx"))
                     if not os.path.basename(p).startswith("~$")]
        if not hits:
            print("   MISSING | {}".format(stem))
            problems.append("{} has never been pulled in.".format(stem))
            continue
        p = max(hits, key=os.path.getmtime)
        w = when_of(p) or dt.datetime.fromtimestamp(os.path.getmtime(p))
        if not newest_data or w > newest_data:
            newest_data = w
        age = (dt.datetime.now() - w).total_seconds() / 3600.0
        tag = "  OK    " if age < 24 else "  OLD   "
        if age >= 24:
            problems.append("{} is {:.0f} hours old - pull a fresh one."
                            .format(stem, age))
        print("  {}| {:<15} as at {}".format(tag, stem, fmt(w)))

    # ---- the reports -----------------------------------------------------
    print("")
    print(" THE REPORTS - when they were last built")
    days = sorted(glob.glob(os.path.join(HERE, "Reports", "20*")))
    built = None
    if days:
        pages = glob.glob(os.path.join(days[-1], "Pages", "*.html"))
        if pages:
            built = dt.datetime.fromtimestamp(
                max(os.path.getmtime(p) for p in pages))
            print("   {} | {} report page(s) in {}".format(
                "OK    " if newest_data and built >= newest_data else "STALE ",
                len(pages), os.path.basename(days[-1])))
            print("           built {}".format(fmt(built)))
    if not built:
        print("   NONE   | no report pages built yet")
        problems.append("No reports have been built yet - run "
                        "00_RUN_EVERYTHING.bat.")
    elif newest_data and built < newest_data:
        problems.append("The reports were built BEFORE the newest export "
                        "landed - run 00_RUN_EVERYTHING.bat again.")

    # ---- the workbook ----------------------------------------------------
    print("")
    print(" THE WORKBOOK - when it was last aligned")
    wbs = [p for p in glob.glob(os.path.join(HERE, "Cement_Australia_Report*K2*.xlsm"))
           if not os.path.basename(p).startswith("~$")]
    if len(wbs) > 1:
        print("   WARN   | {} workbooks here - the align picks whichever was"
              " touched last:".format(len(wbs)))
        for w in wbs:
            print("            " + os.path.basename(w))
        problems.append("More than one K2 workbook in this folder. Rename "
                        "all but the real one (put OLD_ on the front).")
    if wbs:
        w = max(wbs, key=os.path.getmtime)
        wt = dt.datetime.fromtimestamp(os.path.getmtime(w))
        ok = newest_data and wt >= newest_data
        print("   {} | {}".format("OK    " if ok else "STALE ",
                                  os.path.basename(w)))
        print("           aligned {}".format(fmt(wt)))
        if newest_data and wt < newest_data:
            problems.append("The workbook was last touched BEFORE the newest "
                            "export - run 13_ALIGN_EXCEL_DATA.bat.")
    else:
        print("   NONE   | no K2 workbook in this folder")

    # ---- verdict ---------------------------------------------------------
    print("=" * 62)
    if not problems:
        print(" IN SYNC. The exports, the reports and the workbook are all")
        print(" telling the same story - send with confidence.")
        return 0
    print(" OUT OF SYNC - {} thing(s) to sort:".format(len(problems)))
    for i, p in enumerate(problems, 1):
        print("   {}. {}".format(i, p))
    print("")
    print(" The order that always works:")
    print("   1. 28_PULL_SITEIQ_EXPORTS.bat   (newest data in)")
    print("   2. 00_RUN_EVERYTHING.bat        (reports + workbook off it)")
    print("   3. me again                     (proof they agree)")
    return 1


if __name__ == "__main__":
    sys.exit(main())
