#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | RECIPIENTS - every report email pre-addressed automatically
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Reads Coates_Report_Recipients*.xlsx (the same address book the
#  MAKE_OUTLOOK_DRAFTS script uses) and answers: who gets this report?
#  Every report kit calls resolve() when it writes its .eml, so the
#  moment an email address goes into the book with Include = Yes, that
#  person is on the To/CC of every matching report - automatically.
#
#  The rules, identical to MAKE_OUTLOOK_DRAFTS.ps1:
#    Include  = Yes to receive; No keeps a person on file, unaddressed
#    Add As   = To or CC
#    Company  = ALL (every report), or a company name (that company's
#               reports only)
#    Reports  = ALL, or a comma list of tags (COMPANY, EXEC, DEMOB,
#               CONSUMABLES, DAILY, SAFETY, RETURNS, WEEKLY, RIGHTSIZE,
#               CLOSEOUT, COST, PLANT ...)
#
#  Missing or empty book = empty To/CC - the kit never guesses an
#  email address.
# =====================================================================

import os
import glob
import re

_CACHE = {}


#  ---- COMPANIES THAT HAVE FINISHED ON SITE ---------------------------
#  Andrew, 5 Aug 2026: "yes finished with walz and ish24."
#
#  NOT DELETED FROM THE BOOK, AND ON PURPOSE. Their rows stay exactly
#  where they are - names, addresses, who was who - because a demob is
#  not the end of the paperwork. A damage claim, a missing item or an
#  invoice query weeks from now needs the contact, and hunting it back
#  out of a deleted spreadsheet row is how those go unanswered.
#
#  What stops is the addressing: nothing is drafted TO them again. The
#  build says so by name every run rather than quietly leaving them out,
#  because a company that silently stops receiving reports looks
#  identical to a company that was forgotten - and one of those is a
#  problem.
#
#  Take a line out and they are back on the next run.
#
#  company (as the book spells it) -> (why, who said so and when)
FINISHED = {
    "Walz Construction": ("Finished on site", "Andrew Fisher, 5 Aug 2026"),
    "ISH24": ("Finished on site", "Andrew Fisher, 5 Aug 2026"),
}


def _key(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


_FINISHED_KEYS = {_key(k): v for k, v in FINISHED.items()}


def is_finished(company):
    """(why, who) if this company has demobbed, else None."""
    return _FINISHED_KEYS.get(_key(company))


_FIN_EMAIL_CACHE = {}


def _finished_emails(base_dir):
    """Every address the book files under a finished company.

    Read once off the book itself rather than typed here, so the day a
    demobbed contractor's address changes there is nothing to keep in
    step. Empty set when nobody has finished."""
    if not FINISHED:
        return set()
    key = os.path.abspath(base_dir or ".")
    if key in _FIN_EMAIL_CACHE:
        return _FIN_EMAIL_CACHE[key]
    out = set()
    try:
        for r in _load_book(base_dir):
            if is_finished(r["company"]):
                e = r["email"].strip().lower()
                if e:
                    out.add(e)
    except Exception:
        out = set()
    _FIN_EMAIL_CACHE[key] = out
    return out


def finished_note():
    """One honest line per demobbed company, for the run to print."""
    if not FINISHED:
        return []
    return ["{} - {} ({})".format(n, w, who)
            for n, (w, who) in sorted(FINISHED.items())]


def _load_book(base_dir):
    """Rows from the Recipients sheet with Include = Yes and an email.
    Looks in base_dir, then one folder up (same as the drafts script)."""
    key = os.path.abspath(base_dir)
    if key in _CACHE:
        return _CACHE[key]
    rows = []
    path = None
    for d in (base_dir, os.path.dirname(os.path.abspath(base_dir))):
        hits = sorted(glob.glob(os.path.join(d, "Coates_Report_Recipients*.xlsx")),
                      key=os.path.getmtime, reverse=True)
        hits = [h for h in hits if not os.path.basename(h).startswith("~$")]
        if hits:
            path = hits[0]
            break
    if path:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if "Recipients" in wb.sheetnames:
                ws = wb["Recipients"]
                hdr = None
                for r in ws.iter_rows(values_only=True):
                    vals = [str(v).strip() if v is not None else "" for v in r]
                    if hdr is None:
                        if vals and vals[0].lower() == "company" and "email" in [v.lower() for v in vals]:
                            hdr = {v.lower(): i for i, v in enumerate(vals)}
                        continue

                    def cell(name):
                        i = hdr.get(name)
                        return vals[i] if i is not None and i < len(vals) else ""

                    email_addr = cell("email")
                    if not email_addr or "@" not in email_addr:
                        continue
                    if cell("include").strip().lower() != "yes":
                        continue
                    rows.append({"company": cell("company"),
                                 "email": email_addr,
                                 "addas": cell("add as"),
                                 "reports": cell("reports")})
            wb.close()
        except Exception:
            rows = []          # unreadable book = unaddressed, never a crash
    _CACHE[key] = rows
    return rows


# Coates-internal reports: Reports = ALL is NOT enough to be addressed on
# these unless the address is @coates.com.au - a client contact put on ALL
# ("send them everything") must never receive an internal figure sheet.
# Listing the tag explicitly (Reports = BILLING) is a deliberate act and
# is always honoured. Mirrored in MAKE_OUTLOOK_DRAFTS.ps1.
INTERNAL_TAGS = {"BILLING"}

# Personal emails (a person's own gear scorecard): addressed to that person
# from the People (Email) sheet. Reports = ALL never adds the address book
# to someone's personal email - only an explicit MYGEAR tag joins these.
PERSONAL_TAGS = {"MYGEAR"}


def resolve(base_dir, company="", report=""):
    """Return ('a@x; b@y', 'c@z') To and CC strings for this report.
    company = the company the report is about ('' for site-wide reports);
    report  = the report tag (COMPANY, EXEC, DEMOB, ...)."""
    to, cc = [], []
    #  A company that has finished on site is not addressed again -
    #  neither on its own report nor by riding a blanket ALL row.
    if company and is_finished(company):
        return "", ""
    internal = str(report).upper() in INTERNAL_TAGS
    personal = str(report).upper() in PERSONAL_TAGS
    for r in _load_book(base_dir):
        comp = (r["company"] or "").strip()
        #  ...and their people do not get site-wide reports either.
        #
        #  Checking the row's own company is not enough, and this very
        #  nearly shipped: the book files a person under their company
        #  for COMPANY/DEMOB, and AGAIN under "ALL" for SAFETY. Skipping
        #  the company rows alone left Walz's man on the site-wide
        #  safety pack - a contractor who has left the job still being
        #  sent the whole site's overdue gear. So the addresses are
        #  matched, not just the labels.
        if _finished_emails(base_dir) and (
                r["email"].strip().lower() in _finished_emails(base_dir)):
            continue
        if is_finished(comp):
            continue
        comp_ok = comp == "" or comp.upper() == "ALL" or (
            bool(company) and comp.upper() == str(company).upper())
        if not comp_ok:
            continue
        reps = (r["reports"] or "").strip()
        tags = [t.strip().upper() for t in reps.split(",")]
        explicit = bool(report) and str(report).upper() in tags
        rep_ok = reps == "" or reps.upper() == "ALL" or explicit
        if not rep_ok:
            continue
        if internal and not explicit and not r["email"].strip().lower().endswith("@coates.com.au"):
            continue           # internal report + blanket ALL + outside address
        if personal and not explicit:
            continue           # personal email - the book never piles on via ALL
        (cc if r["addas"].strip().lower().startswith("cc") else to).append(r["email"])
    # dedupe, To wins over CC - same as the drafts script
    seen = set()
    to_u = [a for a in to if not (a.upper() in seen or seen.add(a.upper()))]
    cc_u = [a for a in cc if not (a.upper() in seen or seen.add(a.upper()))]
    return "; ".join(to_u), "; ".join(cc_u)
