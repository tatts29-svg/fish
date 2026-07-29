#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | UTILISATION - plant, tool store and consumables, one model
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 29 Jul 2026): "utilisation of plant gear and tooling.
#  want a insanely over powered simplified Utilisation report that
#  covers everything.. plant.. toolstore and consumables"
#
#  ---------------------------------------------------------------
#  WHAT "USED" HONESTLY MEANS, AND THE TRAP IN GETTING THERE
#  ---------------------------------------------------------------
#  The obvious way to ask "has this ever gone out?" is to look for the
#  barcode in TRANSACTIONS. Do that and the radios come back at ZERO
#  PER CENT - 217 handsets, 63 of them out on hire this minute, and
#  the report says not one has ever moved.
#
#  The reason: the TRANSACTION_CHARGES sheet only holds lines that
#  CHARGED. Radios go out free-issue on this job, so they never appear
#  in it. A second sheet, CUSTOMER_CONTRACTOR_EQUIP, records the
#  movement without the money - and 332 barcodes appear there and
#  nowhere else. Read both and the radios come back at 80%.
#
#  There is a third gap. TRANSACTIONS covers a report period (13 Jul
#  onward here). Anything hired before that and still sitting on
#  continuous hire has no line in either sheet. So an item currently
#  ON HIRE counts as used no matter what the transaction sheets say.
#
#      used  =  charged  OR  moved-without-charge  OR  out right now
#
#  Get any one of those three wrong and the number flatters or damns
#  the wrong category. That is why it is written down here.
#
#  ---------------------------------------------------------------
#  TWO DIFFERENT NUMBERS, BOTH CALLED "UTILISATION"
#  ---------------------------------------------------------------
#    * OUT NOW    - on hire this minute / everything we hold. A
#                   snapshot. Answers "how hard is the fleet working
#                   today".
#    * EVER USED  - anything that has moved once since the shut
#                   started / everything we hold. Answers "what did we
#                   mobilise that has never earned its freight".
#
#  They answer different questions and the page shows both, labelled.
#  Quoting one as the other is how a fleet gets cut in the wrong place.
# =====================================================================
import collections
import datetime as dt
import os

SHUT_START = dt.date(2026, 7, 13)      # 13 Jul 05:00, per the SiteIQ period


def _num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return 0.0


def _date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = ('' if v is None else str(v)).strip()
    if not s:
        return None
    for f in ('%d/%m/%Y %I:%M %p', '%d/%m/%Y %H:%M', '%d/%m/%Y', '%Y-%m-%d'):
        try:
            return dt.datetime.strptime(s, f).date()
        except ValueError:
            continue
    return None


def _sheet(path, name):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []
    hdr = [(str(h).strip() if h is not None else '') for h in rows[0]]
    out = []
    for r in rows[1:]:
        d = {}
        for i, h in enumerate(hdr):
            if h:
                d[h] = r[i] if i < len(r) else None
        out.append(d)
    return out


def _txt(d, k):
    v = d.get(k)
    return '' if v is None else str(v).strip()


#  Storage units that are plant rather than tool store. Kept as a list
#  of markers because SiteIQ has no "is this plant" flag - the storage
#  unit is the only honest signal we have, and it is the one the store
#  itself uses when it walks the yard.
PLANT_MARKERS = ('SITE PLANT', 'PLANT', 'BARRIER', 'CHUTE', 'LAYDOWN')


def _is_plant(unit):
    u = (unit or '').upper()
    return any(m in u for m in PLANT_MARKERS)


def read(rental_path, txn_path, today=None):
    """Utilisation across the whole holding. None if there is no stock."""
    if not rental_path or not os.path.isfile(rental_path):
        return None
    today = today or dt.date.today()
    stock = _sheet(rental_path, 'RENTAL_STOCK')
    if not stock:
        return None

    #  --- the three signals of "this has been used" ---
    charged, moved = set(), set()
    shifts = collections.defaultdict(float)
    if txn_path and os.path.isfile(txn_path):
        for r in _sheet(txn_path, 'TRANSACTION_CHARGES'):
            b = _txt(r, 'LATEST_BARCODE')
            if b:
                charged.add(b)
                shifts[b] += _num(r.get('SHIFTS'))
        for r in _sheet(txn_path, 'CUSTOMER_CONTRACTOR_EQUIP'):
            b = _txt(r, 'LATEST_BARCODE')
            if b:
                moved.add(b)

    days_in = max(1, (today - SHUT_START).days)

    units = collections.defaultdict(lambda: {
        'n': 0, 'out': 0, 'used': 0, 'avail': 0, 'other': 0,
        'plant': False, 'shifts': 0.0, 'longest': 0, 'idleItems': []})
    fams = collections.defaultdict(lambda: {'n': 0, 'out': 0, 'used': 0})
    total = {'n': 0, 'out': 0, 'used': 0, 'avail': 0, 'other': 0}
    statuses = collections.Counter()
    never = []
    longest = []

    for r in stock:
        bc = _txt(r, 'ITEM_BARCODE')
        unit = _txt(r, 'STORAGE_UNIT') or 'Unassigned'
        fam = _txt(r, 'PRODUCT_FAMILY') or 'Other'
        st = _txt(r, 'ITEM_STATUS')
        statuses[st] += 1
        is_out = (st == 'On Hire')
        #  the three-way test - see the header comment
        is_used = is_out or (bc in charged) or (bc in moved)

        u = units[unit]
        u['n'] += 1
        u['plant'] = _is_plant(unit)
        u['shifts'] += shifts.get(bc, 0.0)
        f = fams[fam]
        f['n'] += 1
        total['n'] += 1
        if is_out:
            u['out'] += 1
            f['out'] += 1
            total['out'] += 1
            d = _date(r.get('ON_HIRE_DATE'))
            days = (today - d).days if d else None
            if days is not None:
                u['longest'] = max(u['longest'], days)
                longest.append({'n': _txt(r, 'ITEM_DESCRIPTION'),
                                'u': unit, 'd': days,
                                'w': _txt(r, 'HIRER_NAME'),
                                'co': _txt(r, 'COMPANY_NAME')})
        elif st == 'Available for Hire':
            u['avail'] += 1
            total['avail'] += 1
        else:
            u['other'] += 1
            total['other'] += 1
        if is_used:
            u['used'] += 1
            f['used'] += 1
            total['used'] += 1
        else:
            never.append({'n': _txt(r, 'ITEM_DESCRIPTION'), 'u': unit,
                          'f': fam, 'st': st})

    def pct(a, b):
        return (100.0 * a / b) if b else 0.0

    unit_rows = []
    for name, v in units.items():
        unit_rows.append({
            'name': name, 'n': v['n'], 'out': v['out'], 'used': v['used'],
            'avail': v['avail'], 'other': v['other'], 'plant': v['plant'],
            'outPct': pct(v['out'], v['n']), 'usedPct': pct(v['used'], v['n']),
            'never': v['n'] - v['used'], 'longest': v['longest'],
            'shifts': v['shifts'],
        })
    unit_rows.sort(key=lambda x: -x['n'])

    fam_rows = [{'name': k, 'n': v['n'], 'out': v['out'], 'used': v['used'],
                 'outPct': pct(v['out'], v['n']),
                 'usedPct': pct(v['used'], v['n']),
                 'never': v['n'] - v['used']}
                for k, v in fams.items()]
    fam_rows.sort(key=lambda x: -x['n'])

    never_by_unit = collections.Counter(x['u'] for x in never)
    longest.sort(key=lambda x: -x['d'])

    plant = {'n': 0, 'out': 0, 'used': 0, 'avail': 0, 'other': 0}
    store = {'n': 0, 'out': 0, 'used': 0, 'avail': 0, 'other': 0}
    for row in unit_rows:
        t = plant if row['plant'] else store
        for k in ('n', 'out', 'used', 'avail', 'other'):
            t[k] += row[k]
    for t in (plant, store):
        t['outPct'] = pct(t['out'], t['n'])
        t['usedPct'] = pct(t['used'], t['n'])
        t['never'] = t['n'] - t['used']

    return {
        'asof': today.isoformat(),
        'shutStart': SHUT_START.isoformat(),
        'daysIn': days_in,
        'total': total,
        'outPct': pct(total['out'], total['n']),
        'usedPct': pct(total['used'], total['n']),
        'never': total['n'] - total['used'],
        'neverList': never,
        'neverByUnit': never_by_unit.most_common(),
        'units': unit_rows,
        'families': fam_rows,
        'plant': plant,
        'store': store,
        'longest': longest[:40],
        'statuses': statuses.most_common(),
        'charged': len(charged),
        'movedOnly': len(moved - charged),
        'shiftsTotal': sum(shifts.values()),
    }
