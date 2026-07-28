#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | SWITCH ON - who actually receives their company's report
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 27 Jul 2026): the contacts import deliberately leaves
#  every new person switched OFF - putting 65 people onto live client
#  emails must never be a side effect of an import. But that leaves the
#  actual decision stranded in a spreadsheet column, and a company with
#  219 items on hire sat there with nobody receiving its report.
#
#  So: pick a company, its contacts go on. Pick it again, they go off.
#  Nothing else in the book is touched.
#
#  It shows how much gear each company is holding right next to the
#  count, because that is the thing that decides who matters. The
#  company with 219 items and nobody on it is the one to fix first.
# =====================================================================
import os
import shutil
import sys
import time

HERE = os.path.dirname(os.path.abspath(__file__))
BOOK = os.path.join(HERE, "Coates_Report_Recipients.xlsx")
BK = os.path.join(HERE, "Updates", "contact_backups")


def gear_by_company():
    """How many items each company is holding - the reason to care."""
    try:
        import build_company_onhire_report as B
        items = B.load_rental()[0]
    except Exception:
        return {}, None
    n = {}
    for i in items:
        c = i.get("company")
        if c:
            d = B.display_company(c)
            n[d] = n.get(d, 0) + 1
    return n, B


def load(ws, hdr):
    rows = {}
    for i in range(hdr + 1, ws.max_row + 1):
        v = [str(ws.cell(row=i, column=j).value or "").strip()
             for j in range(1, 8)]
        if not any(v):
            continue
        rows[i] = {"company": v[0], "name": v[1], "email": v[2],
                   "include": v[4] or "No"}
    return rows


def tally(rows):
    by = {}
    for r in rows.values():
        if not r["email"]:
            continue
        c = r["company"] or "(none)"
        e = by.setdefault(c, {"on": 0, "off": 0})
        e["on" if r["include"].lower() == "yes" else "off"] += 1
    return by


def board(rows, gear, order=None):
    """Companies in the book, worst gap first.

    ORDER IS FIXED once and never re-sorted. It ranks by "nobody on, most
    gear", so switching a company on would otherwise move it down the
    list and every number below it would shift up - you pick 1, then 4,
    and 4 is no longer the company you were reading a second ago. That
    mis-fires client emails at the wrong contractor. The numbers stay
    put for the whole session; only the counts beside them move."""
    by = tally(rows)
    if order is not None:
        names = [c for c in order if c in by] + \
                [c for c in sorted(by) if c not in order]
        return names, by

    def rank(c):
        e = by[c]
        #  nobody on + most gear = the biggest hole, straight to the top
        return (0 if e["on"] == 0 else 1, -gear.get(c, 0), c)
    return sorted(by, key=rank), by


def main():
    print("=" * 68)
    print(" COATES | SWITCH ON - who receives their company's report")
    print("=" * 68)
    try:
        import openpyxl
    except ImportError:
        print(" openpyxl isn't installed - run:  py -m pip install openpyxl")
        return 1
    if not os.path.isfile(BOOK):
        print(" Coates_Report_Recipients.xlsx isn't in this folder.")
        return 1

    gear, B = gear_by_company()
    wb = openpyxl.load_workbook(BOOK)
    ws = wb["Recipients"]
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        v = [str(x).strip().lower() if x is not None else "" for x in row]
        if v and v[0] == "company" and "email" in v:
            hdr = i
            break
    if not hdr:
        print(" Couldn't find the header row in the Recipients sheet.")
        return 1
    rows = load(ws, hdr)

    #  Straight from the command line:  34_SWITCH_ON.bat "DGH Engineering"
    args = [a for a in sys.argv[1:] if a.strip()]
    changed = []

    def flip(company, want_on):
        n = 0
        for r in rows.values():
            if r["company"] == company and r["email"]:
                new = "Yes" if want_on else "No"
                if r["include"] != new:
                    r["include"] = new
                    changed.append((company, r["name"], r["email"], new))
                    n += 1
        return n

    fixed = None
    while True:
        names, by = board(rows, gear, fixed)
        if fixed is None:
            fixed = list(names)          # locked for the whole session
        print("")
        print(" {:<3} {:<38} {:>8}  {:>9}".format(
            "", "COMPANY", "ON / OFF", "ON HIRE"))
        for n, c in enumerate(names, 1):
            e = by[c]
            g = gear.get(c)
            flag = "   <-- nobody on" if e["on"] == 0 and g else ""
            print(" {:<3} {:<38} {:>3} / {:<4} {:>9}{}".format(
                n, c[:38], e["on"], e["off"],
                "{} items".format(g) if g else "-", flag))
        print("")
        if args:
            pick = args.pop(0)
        else:
            print(" Type a number to switch that company ON.")
            print(" Add ' off' to switch it off  (e.g.  4 off ).")
            try:
                pick = input(" Number (or Enter when finished): ").strip()
            except EOFError:
                pick = ""
        if not pick:
            break
        want_on = True
        if pick.lower().endswith(" off"):
            want_on, pick = False, pick[:-4].strip()
        elif pick.lower().endswith(" on"):
            pick = pick[:-3].strip()
        comp = None
        if pick.isdigit() and 1 <= int(pick) <= len(names):
            comp = names[int(pick) - 1]
        else:
            for c in names:
                if c.lower() == pick.lower():
                    comp = c
                    break
        if not comp:
            print(" ! Didn't recognise '{}' - type one of the numbers.".format(pick))
            continue
        n = flip(comp, want_on)
        print(" {} {} contact(s) at {}".format(
            "Switched ON " if want_on else "Switched OFF", n, comp)
            if n else " Nothing to change at {} - already {}.".format(
                comp, "on" if want_on else "off"))

    if not changed:
        print("")
        print(" No changes made. Nothing written.")
        return 0

    os.makedirs(BK, exist_ok=True)
    shutil.copy2(BOOK, os.path.join(
        BK, "Coates_Report_Recipients_%s_pre_switch.xlsx"
            % time.strftime("%Y%m%d_%H%M%S")))
    for i, r in rows.items():
        ws.cell(row=i, column=5).value = r["include"]
    try:
        wb.save(BOOK)
    except PermissionError:
        print("")
        print(" ! The book is open in Excel. Close it and run me again.")
        print(" ! Nothing has been changed.")
        return 1

    print("")
    print(" CHANGED:")
    for c, n, e, s in changed:
        print("   {:<4} {:<26} {:<20} {}".format(s.upper(), c[:26], n[:20], e))
    print("")
    print(" Backup  : Updates\\contact_backups\\")

    #  Prove it: ask the reports themselves who they would address now.
    try:
        import recipients
        recipients._CACHE.clear()
        print("")
        print(" Who each report goes to now:")
        for c in sorted(set(gear) | {r["company"] for r in rows.values()}):
            if c not in gear:
                continue
            to, _ = recipients.resolve(HERE, c, "COMPANY")
            k = len([x for x in to.split(";") if x.strip()])
            print("   {:<38} {} addressed".format(c[:38], k))
    except Exception:
        pass
    print("")
    print(" Done. Run your reports as normal - they pick this up straight away.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
