#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | FORECAST v ACTUAL - is the job spending what it was quoted?
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 2 Aug 2026): "see attached forecast, tooling was
#  98,674.32, plant equipment is 287,926.37, dates is 20/07/2026 to
#  11/08/2026. we need to take the radio costs and gas monitor costs
#  out of the plant as this was charged through the branch."
#
#  THE FORECAST IS A HUMAN NUMBER, so it lives here where a human can
#  change it. Nothing in the exports carries it and nothing should
#  pretend to derive it.
#
#  ---------------------------------------------------------------
#  THE TRAP THIS FILE EXISTS TO AVOID
#  ---------------------------------------------------------------
#  Set Andrew's forecast against SiteIQ's hire figure on its own and
#  the job looks miles under budget - every time, for the whole
#  shutdown. It isn't. The forecast was written for ALL the equipment
#  on this job, and this job bills through two streams:
#
#      SiteIQ    - the store's own rental register
#      Baseplan  - branch and sub-hire gear, on its own invoice
#
#  Radios and gas monitors are the obvious Baseplan lines and Andrew
#  named them, so they come OUT of the plant forecast. But the welders,
#  the forklift, the compressor and the crib gear are Baseplan too, and
#  they were never taken out of the forecast - so they have to be added
#  to the ACTUAL instead. Take one side off and leave the other in and
#  the variance is just the size of the gear you forgot.
#
#  So: radios and gas leave the forecast. The rest of Baseplan joins
#  the actual. Both sides are shown line by line so the arithmetic can
#  be argued with.
#
#  NO BASEPLAN TOTAL IS EVER MIXED INTO A SITEIQ TOTAL. They are added
#  only at the very last line, and that line says so.
#
#  ---------------------------------------------------------------
#  TO DATE v COMMITTED
#  ---------------------------------------------------------------
#  Baseplan's export carries a Prebill Amount - what the line will cost
#  by its expected off-hire, not what it has cost so far. Mixing a
#  to-date figure with a to-term figure would flatter the spend, so
#  both are worked out and both are shown: to-date against to-date, and
#  committed only where the projection is.
#
#  ---------------------------------------------------------------
#  WHAT IS NOT EQUIPMENT HIRE
#  ---------------------------------------------------------------
#  Transport, labour and service are real money and they are not
#  equipment. An equipment forecast does not cover them, so they are
#  listed and excluded rather than dropped silently.
# =====================================================================
import datetime as dt
import glob
import os

import day_rates as DR
import mygear_intel as MI

BASE = os.path.dirname(os.path.abspath(__file__))

#  --------------------------------------------------------------
#  THE FORECAST. Andrew's figures. Change them here and every
#  comparison follows.
#  --------------------------------------------------------------
FORECAST = {
    'tooling': 98674.32,
    'plant': 287926.37,
}
FROM = dt.date(2026, 7, 20)      # the day the gear started going out
TO = dt.date(2026, 8, 11)

#  Baseplan lines that are branch-charged and must come OUT of the
#  plant forecast before it is compared with anything SiteIQ reports.
#  Named by Andrew, not guessed - and priced from the Baseplan export.
BRANCH_CHARGED = ('Radio 2 Way Hand Held Uhf', 'Gas Monitors')

#  Baseplan sales analysis codes. FRI is freight, MIS is miscellaneous
#  sales - neither is equipment hire.
NOT_HIRE_CODES = ('GLST-FRI', 'GLST-MIS')


def days():
    return (TO - FROM).days + 1


# =====================================================================
#  BASEPLAN - the second invoice stream
# =====================================================================
def _cell(row, ix, name):
    i = ix.get(name)
    if i is None or i >= len(row):
        return None
    return row[i]


def _date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def _f(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _is_branch(desc):
    d = DR._norm(desc)
    return any(DR._norm(b) == d for b in BRANCH_CHARGED)


def baseplan_lines(here=None):
    """Every Baseplan line, told apart rather than lumped together.

    Returns a list of dicts with enough on each one that the page can
    say WHY a line counted or did not: hire or not, branch or not,
    what it costs a day, what it has cost so far and what it is
    committed to by its expected off-hire.
    """
    here = here or BASE
    hits = [p for p in glob.glob(os.path.join(here, 'Data_Baseplan', '*.xlsx'))
            if not os.path.basename(p).startswith('~')]
    if not hits:
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(max(hits, key=os.path.getmtime),
                                    read_only=True, data_only=True)
    except Exception:
        return []
    out = []
    for sn in wb.sheetnames:
        rows = list(wb[sn].iter_rows(values_only=True))
        if not rows:
            continue
        hdr = [str(c).strip() if c is not None else '' for c in rows[0]]
        ix = {h: i for i, h in enumerate(hdr) if h}
        if 'Prebill Amount' not in ix:
            continue
        for r in rows[1:]:
            if not r:
                continue
            desc = _cell(r, ix, 'Description')
            desc = str(desc).strip() if desc else ''
            if not desc:
                continue
            code = str(_cell(r, ix, 'Sales Analysis Code') or '').strip()
            itype = _cell(r, ix, 'Item Type')
            qty = _f(_cell(r, ix, 'Quantity')) or 1.0
            rate = _f(_cell(r, ix, 'Rate 1'))
            start = _date(_cell(r, ix, 'Start Date'))
            term = (_date(_cell(r, ix, 'Expected Term Date'))
                    or _date(_cell(r, ix, 'Term Date')))
            hire = (itype == 1 and code.upper() not in NOT_HIRE_CODES
                    and rate > 0)
            out.append({
                'desc': desc, 'code': code, 'qty': qty, 'rate': rate,
                'perDay': qty * rate,
                'start': start, 'term': term,
                'committed': _f(_cell(r, ix, 'Prebill Amount')),
                'hire': hire,
                'branch': _is_branch(desc),
            })
    try:
        wb.close()
    except Exception:
        pass
    return out


def baseplan_to_date(line, upto):
    """What this line has cost by `upto` - never past its own off-hire.

    Days are counted inclusive of both ends, the same way the rest of
    this suite counts a hire day, so a one-day hire is one day.
    """
    if not line['hire'] or not line['start'] or not upto:
        return 0.0
    end = min(upto, line['term']) if line['term'] else upto
    if end < line['start']:
        return 0.0
    n = (end - line['start']).days + 1
    return line['perDay'] * n


def branch_daily(here=None):
    """What the branch-charged lines cost a day, off Baseplan itself."""
    out = []
    for ln in baseplan_lines(here):
        if ln['branch'] and ln['rate'] > 0:
            out.append((ln['desc'], ln['qty'], ln['rate'], ln['perDay'],
                        ln['committed']))
    if out:
        return out
    #  no Baseplan export on this machine - fall back to the rate card
    #  so the sheet still prints, and say where the number came from.
    base = DR.from_baseplan(here or BASE)
    for name in BRANCH_CHARGED:
        rec = base['byDesc'].get(DR._norm(name))
        if rec:
            q = rec.get('qty') or 0
            out.append((rec.get('desc') or name, q, rec['rate'],
                        q * rec['rate'], 0.0))
    return out


# =====================================================================
#  SITEIQ - the store's own register
# =====================================================================
def read_daily_summary(path):
    """SiteIQ's own revenue by day. Used as a cross-check, not as the
    comparison - it carries no plant/tooling split."""
    import openpyxl
    if not path or not os.path.isfile(path):
        return None
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if 'DAILY_SUMMARY' not in wb.sheetnames:
        wb.close()
        return None
    rows = list(wb['DAILY_SUMMARY'].iter_rows(values_only=True))
    wb.close()
    if not rows:
        return None
    hdr = [str(c).strip().upper() if c is not None else '' for c in rows[0]]
    try:
        hi = hdr.index('HIRE')
    except ValueError:
        return None
    daily, total = [], 0.0
    for r in rows[1:]:
        lab = str(r[0] or '')
        v = r[hi] if hi < len(r) else 0
        try:
            v = float(v or 0)
        except (TypeError, ValueError):
            v = 0.0
        if lab.upper().startswith('REPORT TOTAL'):
            total = v
            continue
        d = None
        for tok in lab.replace(':', ' ').split():
            try:
                d = dt.datetime.strptime(tok.strip(), '%d/%m/%Y').date()
                break
            except ValueError:
                continue
        if d:
            daily.append((d, v))
    daily.sort()
    return {'daily': daily, 'total': total}


def siteiq(data):
    """The store's hire spend, split the way the project pays for it.

    Off-register money is plant hired outside this rental register -
    real gear on this job, just not this fleet - so it sits with plant
    and is named separately so nobody thinks the store holds it.
    Service and admin are not equipment hire and do not come in here.
    """
    t = data['totals']
    return {
        'tooling': t['tooling']['revenue'],
        'plantOnRegister': t['plant']['revenue'],
        'offRegister': t['revenueOffRegister'],
        'plant': t['plant']['revenue'] + t['revenueOffRegister'],
        'admin': t['revenueService'],
        'from': data['sourceFrom'], 'to': data['sourceTo'],
        'days': data['sourceDays'],
    }


# =====================================================================
#  THE COMPARISON
# =====================================================================
def compare(data, ds=None, here=None):
    """The forecast, the actual and the honest arithmetic between them."""
    n = days()
    lines_bp = baseplan_lines(here)
    branch = branch_daily(here)
    branch_day = sum(b[3] for b in branch)
    branch_window = branch_day * n
    branch_committed = sum(b[4] for b in branch)

    plant_ex = FORECAST['plant'] - branch_window
    fc_total = FORECAST['tooling'] + plant_ex

    out = {
        'from': FROM, 'to': TO, 'days': n,
        'forecast': dict(FORECAST),
        'branch': branch, 'branchDaily': branch_day,
        'branchWindow': branch_window,
        'branchCommitted': branch_committed,
        'plantExBranch': plant_ex,
        'forecastComparable': fc_total,
        'perDayNeeded': fc_total / n if n else 0.0,
        'perDayTooling': FORECAST['tooling'] / n if n else 0.0,
        'perDayPlant': plant_ex / n if n else 0.0,
    }

    if not data:
        out['problem'] = 'no transaction export to compare against'
        return out

    sq = siteiq(data)
    out['siteiq'] = sq
    out['splitGap'] = data['totals'].get('splitGap') or 0.0
    asof = dt.date.fromisoformat(sq['to']) if sq['to'] else None
    out['asof'] = asof

    #  BASEPLAN, radios and gas left out because they already left the
    #  forecast. Counting them on both sides would cancel to a lie.
    keep = [l for l in lines_bp if l['hire'] and not l['branch']]
    out['baseplanRead'] = len(lines_bp)
    out['baseplanLines'] = keep
    out['baseplanNotHire'] = [l for l in lines_bp if not l['hire']]
    out['baseplanToDate'] = sum(baseplan_to_date(l, asof) for l in keep)
    out['baseplanCommitted'] = sum(l['committed'] for l in keep)
    out['baseplanNotHireTotal'] = sum(l['committed']
                                      for l in out['baseplanNotHire'])

    #  the window the actual actually covers. Days before 20 Jul are
    #  not what was forecast and days after the export are NOT DATA.
    a_from = dt.date.fromisoformat(sq['from']) if sq['from'] else FROM
    elapsed = ((asof - max(a_from, FROM)).days + 1) if asof else 0
    elapsed = max(elapsed, 0)
    out['elapsed'] = elapsed
    out['daysLeft'] = max(0, (TO - asof).days) if asof else n

    out['actualTooling'] = sq['tooling']
    out['actualPlant'] = sq['plant'] + out['baseplanToDate']
    out['actual'] = out['actualTooling'] + out['actualPlant']

    out['rateTooling'] = (sq['tooling'] / elapsed) if elapsed else 0.0
    out['ratePlant'] = (out['actualPlant'] / elapsed) if elapsed else 0.0
    out['rate'] = (out['actual'] / elapsed) if elapsed else 0.0

    #  WHERE IT LANDS, at the pace so far. Baseplan is already
    #  committed to its off-hire date, so it is not projected - it is
    #  added at its commitment and only SiteIQ's run rate is carried
    #  forward.
    left = out['daysLeft']
    sq_rate = ((sq['tooling'] + sq['plant']) / elapsed) if elapsed else 0.0
    out['projTooling'] = sq['tooling'] + out['rateTooling'] * left
    out['projSiteIQ'] = (sq['tooling'] + sq['plant']) + sq_rate * left
    out['proj'] = out['projSiteIQ'] + out['baseplanCommitted']
    out['projPlant'] = out['proj'] - out['projTooling']
    out['varTooling'] = out['projTooling'] - FORECAST['tooling']
    out['varPlant'] = out['projPlant'] - plant_ex
    out['var'] = out['proj'] - fc_total

    #  --------------------------------------------------------------
    #  DAILY_SUMMARY - SiteIQ's own revenue by day.
    #
    #  It cannot split plant from tooling, so it cannot BE the line by
    #  line comparison. What it CAN do is two things nothing else here
    #  can: show what SiteIQ actually billed on its last complete day,
    #  and show whether the charge lines have caught up.
    #
    #  They had not. On this pull the charge lines run three days past
    #  the daily summary and those three days carry a fraction of the
    #  money the summary's last day carries - charges get raised after
    #  the fact, so the newest days on a charge export are always thin.
    #  An average taken across them reads low for a reason that has
    #  nothing to do with how much gear is out. So the lag is measured
    #  and printed, and a second projection is run at the last day
    #  SiteIQ actually billed.
    #  --------------------------------------------------------------
    if ds and ds['daily']:
        last = ds['daily'][-1][0]
        out['dsTo'] = last
        out['dsTotal'] = ds['total']
        out['dsStale'] = bool(asof and last < asof)
        out['dsCheck'] = (sq['tooling'] + sq['plantOnRegister']
                          + sq['offRegister'])
        out['dsInWindow'] = sum(v for d, v in ds['daily']
                                if FROM <= d <= min(TO, last))
        out['dsLatestDay'] = ds['daily'][-1][1]
        out['dsBefore'] = sum(v for d, v in ds['daily'] if d < FROM)

        #  what the charge lines claim for the days beyond the summary
        lag = (asof - last).days if (asof and asof > last) else 0
        out['lagDays'] = lag
        if lag:
            out['lagMoney'] = out['dsCheck'] - out['dsInWindow'] \
                - out['dsBefore']
            out['lagRate'] = out['lagMoney'] / lag
        #  PROJECTION AT THE LAST BILLED DAY. Built entirely off the
        #  daily summary so it never mixes a thin charge day into a
        #  billed one, then Baseplan added at its commitment.
        ahead = max(0, (TO - last).days)
        out['projLatestDays'] = ahead
        out['projLatest'] = (out['dsInWindow']
                             + out['dsLatestDay'] * ahead
                             + out['baseplanCommitted'])
        out['varLatest'] = out['projLatest'] - fc_total
    return out


def lines(c):
    """The comparison as printable lines - used by 67_ and the sheet."""
    L = []
    A = L.append
    m = '${:>12,.2f}'.format
    A('=' * 66)
    A(' FORECAST v ACTUAL   {} to {}   {} days'.format(
        c['from'].strftime('%d %b'), c['to'].strftime('%d %b %Y'),
        c['days']))
    A('=' * 66)
    A('')
    A('  FORECAST AS QUOTED')
    A('    tooling                       ' + m(c['forecast']['tooling']))
    A('    plant equipment               ' + m(c['forecast']['plant']))
    if c['branch']:
        A('')
        A('  LESS BRANCH-CHARGED, which SiteIQ will never show:')
        for nm, qty, rate, per, com in c['branch']:
            A('    {:<26} {:>4.0f} x ${:>7.2f} = ${:>9,.2f}/day'.format(
                nm[:26], qty, rate, per))
        A('    over {} days                  -{}'.format(
            c['days'], m(c['branchWindow']).strip()))
        if c['branchCommitted']:
            A('    (Baseplan has committed {} to these two lines'.format(
                '${:,.2f}'.format(c['branchCommitted'])))
            A('     over their own hire dates - close, not identical,')
            A('     because they did not all start on 20 Jul)')
        A('    plant less branch             ' + m(c['plantExBranch']))
    A('')
    A('    COMPARABLE FORECAST           ' + m(c['forecastComparable']))
    A('    tooling needs                 ' + m(c['perDayTooling']) + ' /day')
    A('    plant needs                   ' + m(c['perDayPlant']) + ' /day')
    A('    together                      ' + m(c['perDayNeeded']) + ' /day')
    if c.get('problem'):
        A('')
        A('  ' + c['problem'])
        return L

    sq = c['siteiq']
    A('')
    A('  ACTUAL, {} to {} ({} days of it)'.format(
        dt.date.fromisoformat(sq['from']).strftime('%d %b'),
        dt.date.fromisoformat(sq['to']).strftime('%d %b %Y'), c['elapsed']))
    A('')
    A('    SITEIQ, off the charge lines')
    A('      tooling                     ' + m(sq['tooling']))
    A('      plant on this register      ' + m(sq['plantOnRegister']))
    A('      plant off this register     ' + m(sq['offRegister']))
    if c['baseplanLines']:
        A('')
        A('    BASEPLAN, its own invoice - plant the forecast paid for')
        A('    but SiteIQ will never show. Radios and gas are NOT here;')
        A('    they came off the forecast instead.')
        for l in sorted(c['baseplanLines'],
                        key=lambda x: -x['committed'])[:12]:
            A('      {:<28} {:>3.0f} x ${:>7.2f}  ${:>9,.2f}'.format(
                l['desc'][:28], l['qty'], l['rate'], l['committed']))
        A('      to date                   ' + m(c['baseplanToDate']))
        A('      committed to off-hire     ' + m(c['baseplanCommitted']))
    A('')
    A('    TOOLING SPENT                 ' + m(c['actualTooling'])
      + '  ' + '${:,.2f}'.format(c['rateTooling']) + '/day')
    A('    PLANT SPENT                   ' + m(c['actualPlant'])
      + '  ' + '${:,.2f}'.format(c['ratePlant']) + '/day')
    A('    BOTH STREAMS TOGETHER         ' + m(c['actual'])
      + '  ' + '${:,.2f}'.format(c['rate']) + '/day')
    A('')
    A('  NOT EQUIPMENT HIRE, so not in any line above')
    A('    admin - transport, labour, service  ' + m(sq['admin']))
    if c.get('baseplanNotHireTotal'):
        A('    Baseplan transport and sales        '
          + m(c['baseplanNotHireTotal']))
    A('')
    A('  WHERE IT LANDS - TWO PACES, AND THE GAP BETWEEN THEM IS')
    A('  THE POINT. Baseplan is added to both at what it is already')
    A('  committed to; it does not need projecting.')
    A('')
    A('  1. AT THE AVERAGE SO FAR, {} days still to run'.format(
        c['daysLeft']))
    A('    tooling           ' + m(c['projTooling'])
      + '   {:+,.2f}'.format(c['varTooling']))
    A('    plant             ' + m(c['projPlant'])
      + '   {:+,.2f}'.format(c['varPlant']))
    A('    together          ' + m(c['proj'])
      + '   {:+,.2f} against forecast'.format(c['var']))
    A('    Reads LOW. It averages in the first days of the job, when')
    A('    little was out, and the newest days on the charge export,')
    A('    which are thin because the charges are still being raised.')
    if c.get('projLatest') is not None:
        A('')
        A('  2. AT THE LAST DAY SITEIQ ACTUALLY BILLED')
        A('    {} billed {} on {}'.format(
            'SiteIQ', '${:,.2f}'.format(c['dsLatestDay']),
            c['dsTo'].strftime('%d %b')))
        A('    held flat for the {} days after it'.format(
            c['projLatestDays']))
        A('    together          ' + m(c['projLatest'])
          + '   {:+,.2f} against forecast'.format(c['varLatest']))
        A('    No plant/tooling split on this one - the daily summary')
        A('    does not carry one and this page will not invent it.')
        A('    It holds the busiest billed day FLAT. More gear out and')
        A('    it lands higher; gear coming home and it lands lower.')
    A('')
    A('    THE TRUTH IS BETWEEN THEM. Neither is quoted as THE number.')
    A('    Read the gap to forecast as room still on the job, not as a')
    A('    saving already banked.')
    if c.get('dsTo'):
        A('')
        A('  HOW FAR BEHIND THE CHARGE LINES ARE')
        A('    daily summary to {}      {}'.format(
            c['dsTo'].strftime('%d %b'), m(c['dsTotal']).strip()))
        A('    charge lines to {}       {}'.format(
            c['asof'].strftime('%d %b') if c['asof'] else '?',
            m(c['dsCheck']).strip()))
        if c.get('lagDays'):
            A('    So the charge export runs {} days past the summary'
              .format(c['lagDays']))
            A('    and puts {} on those {} days - {}/day,'.format(
                '${:,.2f}'.format(c['lagMoney']), c['lagDays'],
                '${:,.2f}'.format(c['lagRate'])))
            A('    against {} on the last day the summary'.format(
                '${:,.2f}'.format(c['dsLatestDay'])))
            A('    covers. Charges land after the fact. That gap is')
            A('    billing catching up, not gear going home.')
        elif abs(c['dsTotal'] - c['dsCheck']) < 0.01:
            A('    Same day, and they agree to the cent.')
        else:
            A('    THEY DISAGREE by ${:,.2f}. Same day, different'.format(
                abs(c['dsTotal'] - c['dsCheck'])))
            A('    answer - worth a look before this page is quoted.')
    A('')
    A('  WHAT THIS PAGE CANNOT SEE')
    A('    Any Baseplan spend that is not in the export on this')
    A('    machine. {} lines were read. If the branch has raised'.format(
        c.get('baseplanRead', 0)))
    A('    anything else against this job, it is not counted here and')
    A('    the actual is understated by exactly that much.')
    if c.get('splitGap'):
        A('')
        A('    PLANT AND TOOLING DO NOT ADD BACK to the store total -')
        A('    ${:,.2f} is in neither book. The split above is'.format(
            abs(c['splitGap'])))
        A('    not safe to quote until that is found.')
    return L


def _newest(pattern):
    hits = [q for q in glob.glob(os.path.join(BASE, 'Data_SiteIQ', pattern))
            if not os.path.basename(q).startswith('~')]
    hits += [q for q in glob.glob(os.path.join(BASE, pattern))
             if not os.path.basename(q).startswith('~')]
    return max(hits, key=os.path.getmtime) if hits else None


def run():
    data = None
    rental, txn = _newest('RENTAL_STOCK*.xlsx'), _newest('TRANSACTIONS*.xlsx')
    if rental and txn:
        data = MI.read(rental, txn, _newest('ON_HIRE*.xlsx'))
    ds = read_daily_summary(_newest('DAILY_SUMMARY*.xlsx'))
    return compare(data, ds)


if __name__ == '__main__':
    print('\n'.join(lines(run())))
