#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | K2 DAILY REPORTING - the folders, the packs, the checks
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  Builds the daily reporting structure A. Fisher specified (25 Jul
#  2026), files every report where it belongs, and refuses to mark
#  anything ready to send until it has proved the pack is right.
#
#      K2 DAILY REPORTING
#      |-- 00 MASTER DAILY REPORTS \ <date> \      the site-wide masters
#      |-- 01 COMPANY REPORTS \ <Company> \ <date> \
#      |-- 02 EMAIL CONTROL \                      the daily status trail
#      \-- 03 CONTACTS & SETTINGS \                the routing workbook
#
#  THE ROUTING SOURCE IS THE WORKBOOK, NOT THIS FILE
#  -------------------------------------------------
#  K2_Daily_Email_Report_Allocation.xlsx decides who is in To, who is in
#  Cc, what goes in the body and what gets attached. Not one email
#  address is written into this script. Change the workbook and the next
#  run follows it - which is the whole point of having it.
#
#  THE COMPANY NAME IS THE CONTROL KEY
#  -----------------------------------
#  "A DGH report must never be sent using the Cleanaway recipient list."
#  The workbook spells companies its own way and SiteIQ spells them
#  another ("DKE Group" v "Dark Knight Engineering"). Every one of those
#  pairings is written down in ALIASES below, deliberately and visibly -
#  never fuzzy-matched, because a near-miss here sends one contractor's
#  gear list to another. Anything not in that map is reported as
#  unmatched and gets NO email.
#
#  NOTHING IS EVER OVERWRITTEN
#  ---------------------------
#  Rule 7. A dated folder that already holds a report is left exactly as
#  it is; the run says so and moves on.
# =====================================================================

import os
import re
import glob
import shutil
import hashlib
import datetime as dt

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT_NAME = "K2 DAILY REPORTING"
F_MASTER = "00 MASTER DAILY REPORTS"
F_COMPANY = "01 COMPANY REPORTS"
F_CONTROL = "02 EMAIL CONTROL"
F_SETTINGS = "03 CONTACTS & SETTINGS"

ALLOCATION_PATTERN = "K2_Daily_Email_Report_Allocation*.xlsx"

#  The statuses, exactly as specified - no others are ever written.
ST_NOT_GENERATED = "Not Generated"
ST_GENERATED = "Generated"
ST_CHECKED = "Checked"
ST_DRAFT_READY = "Draft Ready"
ST_SENT = "Sent"
ST_FAILED = "Failed - Review Required"


# ---------------------------------------------------------------------
#  COMPANY NAME MAP - the control key, written down on purpose
# ---------------------------------------------------------------------
#  workbook name  ->  the company as SiteIQ spells it
#  None = we know the company is on the routing list but holds no gear
#  in the on-hire register, so there is no report to send today.
ALIASES = {
    "Cement Australia": "Cement Australia Holdings Pty Ltd",
    "DGH Engineering": "DGH Engineering",
    "DKE Group": "Dark Knight Engineering",
    "High Risk Solutions": "High Risk Solutions",
    "Programmed": "Programmed",
    "Veolia Refractory": "Veolia",
    "Xtreme Engineering": "Xtreme Engineering",
    #  THESE FOUR WERE MAPPED TO None - "nothing on hire in the current
    #  register". That was true the day it was written and stopped being
    #  true without anybody touching the file: on 3 Aug 2026 Sync Lift
    #  held 12 items, Universal Cranes 9, Tasman Rope Access 7 and
    #  Industec 2. Thirty items across four contractors whose daily pack
    #  said "no gear on hire" and sent no email at all - no chase list,
    #  no record, and nothing anywhere saying they had been skipped.
    #
    #  The routing workbook spells them short; SiteIQ spells them long,
    #  or in capitals. Mapped properly now, and check_stale_aliases()
    #  below reads the live export every run so this can never rot
    #  quietly again.
    "Industec": "Industec Filtration",
    "Synclift": "Sync Lift Engineering",
    "Tasman Rope Access": "TASMAN ROPE ACCESS",
    "Universal Cranes": "UNIVERSAL CRANES",
    #  FOUR COMPANIES THE PACKS HAD NEVER HEARD OF (3 Aug 2026).
    #
    #  Andrew sent a photograph of a contact card: three days of
    #  Coates_K2_Activity_BOSCH_REXROTH_*.pdf sitting against
    #  gerhardus.vanderwesthuizen@boschrexroth.com.au, timestamped 7:45,
    #  8:16 and 7:53 in the morning. The suite BUILDS that report - it
    #  just had nobody to address it to, because Bosch was not on this
    #  map or in the routing workbook. So he had been attaching it by
    #  hand every morning, which is also why his own name kept turning
    #  up in the To: he opens a draft addressed to himself and adds the
    #  contractor.
    #
    #  Four companies holding gear were in that state - 12 items and 6
    #  people between them. Small enough to go unnoticed, which is
    #  exactly why they did.
    "Bosch Rexroth": "BOSCH REXROTH",
    "Industrial Rescue": "Industrial Rescue & Emergency Training",
    "Filtercare": "Filtercare",
    #  QWest Crane Hire - NO PACK BY INSTRUCTION (4 Aug 2026). Kept
    #  mapped so their rows still carry the right name on every
    #  site-wide report; see NO_PACK below for why, and for what they
    #  are still holding.
    "QWest Crane Hire": "QWEST CRANE HIRE",
    #  NOTHING ON HIRE TODAY, AND THEY STILL GET THE REPORT.
    #  Aestec and Cutrite hold zero items right now, but both have an
    #  Activity & Accountability report built and Andrew has been
    #  sending both by hand - Cutrite at 7:57 and Aestec yesterday. The
    #  report covers what MOVED, not only what is still out, so a
    #  company that returned everything last week still has a page worth
    #  reading. Mapped to their SiteIQ spelling so the report is found;
    #  the pack builder decides what to do with an empty on-hire list.
    "Aestec": "AESTEC",
    "Cutrite": "CUTRITE",
    #  Genuinely nothing on hire today - they get a folder and a record,
    #  never an email with no data. The guard watches this one too.
    "Cleanaway": None,
}


#  ---- COMPANIES THAT GET NO PACK, BY INSTRUCTION ---------------------
#  Andrew, 4 Aug 2026: "qwest no longer needed."
#
#  DELETING THE LINE WOULD HAVE BEEN THE WRONG FIX. QWest is still on
#  the register holding gear - two tool lanyards with Allan Watt, out
#  since 29 Jul. Quietly dropping a company off the pack list stops
#  anybody ever being told about those two again: the pack was the only
#  thing that named them to QWest, and a company nobody reports on is a
#  company nobody chases.
#
#  So they come off the packs and STAY on the books. The run prints them
#  by name every morning with whatever they are still holding, and the
#  site-wide reports carry their lines exactly as before. When the
#  lanyards come back, the line here can go.
#
#  company (as spelled in ALIASES) -> (why, who said so and when)
NO_PACK = {
    "QWest Crane Hire": ("No longer needed - no pack, no draft, no held "
                         "folder", "Andrew Fisher, 4 Aug 2026"),
}


def _key(name):
    """Tidy a company name for looking it up - spaces and capitals only.
    This is NOT fuzzy matching: 'Xtreme Engineering ' still has to be
    Xtreme Engineering. It just stops a stray space costing somebody
    their report."""
    return re.sub(r"\s+", " ", _txt(name)).strip().lower()


_ALIAS_KEYS = None


def resolve_company(name):
    """(the company as SiteIQ spells it, is it on the map at all).

    Two different answers that used to look the same: a company we know
    holds nothing on hire, and a company nobody has ever mapped. The
    first is a normal day. The second means somebody renamed a row in
    the workbook and that company's report is about to go missing
    without a word being said."""
    global _ALIAS_KEYS
    if _ALIAS_KEYS is None:
        _ALIAS_KEYS = {_key(k): v for k, v in ALIASES.items()}
    k = _key(name)
    if k not in _ALIAS_KEYS:
        return None, False
    return _ALIAS_KEYS[k], True


def held_by(onhire_path):
    """{company key: how many lines they are holding} from the live
    export. Best effort - an unreadable export returns {} and the
    caller says "not known" rather than "nothing", because those two
    are not the same answer and only one of them is safe."""
    out = {}
    if not onhire_path or not os.path.isfile(onhire_path):
        return out
    try:
        import openpyxl
        wb = openpyxl.load_workbook(onhire_path, read_only=True,
                                    data_only=True)
        ws = wb["ON_HIRE"] if "ON_HIRE" in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        hdr = [_txt(c) for c in next(rows)]
        ix = {h: i for i, h in enumerate(hdr) if h}
        if "COMPANY" not in ix:
            wb.close()
            return out
        for r in rows:
            if not r:
                continue
            c = _txt(r[ix["COMPANY"]]).strip()
            if c:
                out[_key(c)] = out.get(_key(c), 0) + 1
                _SEEN_SPELLING[_key(c)] = c
        wb.close()
    except Exception:
        return {}
    return out


def check_stale_aliases(onhire_path):
    """Companies mapped to "holds nothing" that are holding something.

    A hardcoded snapshot of who has gear is right until the day it is
    not, and nothing about it looks wrong in the meantime - the pack
    says "no gear on hire", files a tidy record, and the contractor
    never gets a chase list. Four of them had gone stale by 3 Aug 2026
    with thirty items between them.

    So the map is checked against the live export on every run. Returns
    [(workbook name, siteiq name, items)], loudest first, empty when the
    map is honest. Best effort: no export, nothing to say.
    """
    if not onhire_path or not os.path.isfile(onhire_path):
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(onhire_path, read_only=True,
                                    data_only=True)
        ws = wb["ON_HIRE"] if "ON_HIRE" in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        hdr = [_txt(c) for c in next(rows)]
        ix = {h: i for i, h in enumerate(hdr) if h}
        if "COMPANY" not in ix:
            wb.close()
            return []
        held = {}
        for r in rows:
            if not r:
                continue
            c = _txt(r[ix["COMPANY"]]).strip()
            if c:
                held[_key(c)] = held.get(_key(c), 0) + 1
                _SEEN_SPELLING[_key(c)] = c
        wb.close()
    except Exception:
        return []
    out = []
    for wb_name, mapped in ALIASES.items():
        if mapped is not None:
            continue
        #  the workbook's own spelling, and the near-misses SiteIQ uses
        #  for it - "Synclift" for "Sync Lift Engineering"
        kk = _key(wb_name)
        squash = kk.replace(" ", "")
        for hk, n in held.items():
            if hk == kk or hk.replace(" ", "").startswith(squash) \
                    or squash.startswith(hk.replace(" ", "")):
                out.append((wb_name, _SEEN_SPELLING.get(hk, hk), n))
                break
    out.sort(key=lambda x: -x[2])
    return out


_SEEN_SPELLING = {}


#  The master reports, by the names A. Fisher uses, mapped to what the
#  suite actually writes. Order matters - it is the attachment order on
#  the Cement master email.
MASTERS = [
    ("Executive Daily Summary", "Coates_K2_Executive_Summary"),
    ("Daily Safety & Compliance Report", "Coates_K2_Safety_Assurance"),
    ("Daily Cost Tracking Report", "Coates_K2_Cost_Tracking_Snapshot"),
    ("Site Plant Report", "Coates_SitePlant_Report"),
    ("Cement Australia On-Hire Report",
     "Coates_OnHire_Report_CEMENT_AUSTRALIA_HOLDINGS_PTY_LTD"),
    ("Consumables Usage Report", "Coates_K2_Consumables_Report"),
]

#  The site-wide safety report. It goes to Cement Australia and stays in
#  the store folder. It is NOT a contractor attachment and must not
#  become one again: it is one file for the whole site, so it names every
#  contractor's overdue gear, asset numbers and days out. Attached to a
#  contractor email it hands them everyone else's accountability list.
#  (Off contractor email 3 Aug 2026 - A. Fisher.)
SAFETY_PDF = "Daily Safety & Compliance Report"

#  The filed names, exactly as the workbook's Folder Rules sheet sets
#  them out. The built file is copied in under the name the pack calls
#  it, so the folder reads like the spec and not like a script.
FILED_NAME = {
    "Executive Daily Summary": "Executive Daily Summary - {d}",
    "Daily Safety & Compliance Report": "Daily Safety & Compliance Report - {d}",
    "Daily Cost Tracking Report": "Daily Cost Tracking Report - {d}",
    "Site Plant Report": "Site Plant Report - {d}",
    "Cement Australia On-Hire Report": "Cement Australia On-Hire Report - {d}",
    "Consumables Usage Report": "Consumables Usage Report - {d}",
}

#  Cement's five attachments, in the order the Report Allocation sheet
#  lists them.
CEMENT_ATTACHMENTS = [
    "Daily Safety & Compliance Report",
    "Daily Cost Tracking Report",
    "Site Plant Report",
    "Cement Australia On-Hire Report",
    "Consumables Usage Report",
]

#  An email nobody receives is not a report. The limit lives in
#  email_images.MAX_EMAIL_MB (10 MB - plenty of receiving gateways stop
#  there, well under Exchange's 25) so the packs, the report emitter and
#  the checks all hold the same line. The fallback only exists so this
#  module still imports if email_images is somehow missing.
try:
    from email_images import MAX_EMAIL_MB
except Exception:
    MAX_EMAIL_MB = 10.0


def _txt(v):
    return "" if v is None else str(v).strip()


def _slug(name):
    """The company folder name - the workbook's spelling, kept readable
    and safe on Windows."""
    out = re.sub(r'[<>:"/\\|?*]', "-", _txt(name)).strip().rstrip(". ")
    #  A name that tidies away to nothing would drop a company's files
    #  loose in the parent folder - never that.
    return out or "UNNAMED COMPANY"


def _emails(s):
    """Addresses out of a spreadsheet cell. Split on anything a person
    might have typed between them - including the line break you get
    from Alt+Enter, which used to survive into the email header and kill
    the whole run."""
    out = []
    for part in re.split(r"[;,\r\n\t]+", _txt(s)):
        p = part.strip()
        if not p:
            continue
        #  "Ross Gould <ross.gould@dgh.com.au>" is a perfectly normal way
        #  to write an address - take the address out of it rather than
        #  dropping the person. (25 Jul 2026)
        m = re.search(r"<([^<>@\s]+@[^<>@\s]+)>", p)
        if m:
            p = m.group(1)
        elif " " in p:
            m = re.search(r"([^<>@\s]+@[^<>@\s]+)", p)
            p = m.group(1) if m else ""
        p = p.strip("<>.,; ")
        if p and "@" in p and " " not in p and p not in out:
            out.append(p)
    return out


# ---------------------------------------------------------------------
#  1. THE WORKBOOK - the only place recipients come from
# ---------------------------------------------------------------------
def load_allocation(base=None, path=None):
    """Company Routing + Contacts, straight off the master workbook.
    Pass path to read a specific workbook (used to try a routing change
    before it goes live)."""
    if not path:
        base = base or HERE
        hits = [p for p in glob.glob(os.path.join(base, ALLOCATION_PATTERN))
                if not os.path.basename(p).startswith("~$")]
        if not hits:
            return None
        path = max(hits, key=os.path.getmtime)
    if not os.path.isfile(path):
        return None
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    def sheet(name):
        """The header row is the one carrying 'Company' - found wherever
        it sits, not assumed to be column A, so inserting a column in
        front of it doesn't silently empty the whole run."""
        if name not in wb.sheetnames:
            return [], {}
        rows = [[_txt(c) for c in r]
                for r in wb[name].iter_rows(values_only=True)]
        hdr = next((r for r in rows if r and "Company" in r), None)
        if not hdr:
            return [], {}
        ix = {v: i for i, v in enumerate(hdr) if v}
        c = ix["Company"]
        body = [r for r in rows[rows.index(hdr) + 1:]
                if r and len(r) > c and r[c]]
        return body, ix

    #  The four fixed oversight contacts come off the Summary sheet, not
    #  out of this script - so adding a fifth is a spreadsheet edit.
    #  Every row under the heading is read, not just the first one that
    #  happens to hold an address: listing the four one per row is the
    #  obvious way to type them, and reading only the first row meant the
    #  check passed on one contact instead of four. (25 Jul 2026)
    fixed = []
    if "Summary" in wb.sheetnames:
        rows = [[_txt(c) for c in r]
                for r in wb["Summary"].iter_rows(values_only=True)]
        for i, r in enumerate(rows):
            head = " ".join(r).upper()
            if "FIXED" in head and "CC" in head:
                for nxt in rows[i + 1:]:
                    got = []
                    for cell in nxt:
                        got += _emails(cell)
                    if not got:
                        if fixed or any(_txt(c) for c in nxt):
                            break        # blank row, or the next heading
                        continue
                    for e in got:
                        if e.lower() not in [f.lower() for f in fixed]:
                            fixed.append(e)
                break

    routing, rix = sheet("Company Routing")
    contacts, cix = sheet("Contacts")
    out = {"path": path, "companies": [], "contacts": [], "fixed_cc": fixed}
    for r in routing:
        def g(k):
            i = rix.get(k)
            return r[i] if i is not None and i < len(r) else ""
        out["companies"].append({
            "company": g("Company"),
            "kind": g("Email Type"),
            "to": _emails(g("To")),
            "cc": _emails(g("Cc")),
            "body": g("Pasted in Email Body"),
            "attach": [a.strip() for a in re.split(r";", g("PDF Attachments"))
                       if a.strip()],
            "note": g("Status"),
        })
    for r in contacts:
        def gc(k):
            i = cix.get(k)
            return r[i] if i is not None and i < len(r) else ""
        out["contacts"].append({
            "company": gc("Company"), "name": gc("Contact"),
            "role": gc("Role"), "email": gc("Email"),
            "phone": gc("Phone"), "allocation": gc("Allocation"),
            "verification": gc("Verification"),
        })
    out["fixed_cc"] = fixed
    out["fixed_cc_source"] = "Summary sheet"
    if not fixed:
        #  Only if the Summary sheet gave us nothing: anyone in the Cc of
        #  EVERY external company is doing the same job. This is a
        #  fallback, and the run says so - it is never used to decide
        #  that an address is allowed to be there (see New-2, 25 Jul
        #  2026: an address pasted into every Cc row would otherwise
        #  make itself legitimate).
        ext = [c for c in out["companies"]
               if _txt(c["kind"]).lower().startswith("external")]
        if ext:
            common = set(e.lower() for e in ext[0]["cc"])
            for c in ext[1:]:
                common &= set(e.lower() for e in c["cc"])
            out["fixed_cc"] = sorted(common)
            out["fixed_cc_source"] = ("worked out from the Cc lists - the "
                                      "Summary sheet gave none")

    wb.close()
    return out


# ---------------------------------------------------------------------
#  2. THE FOLDERS
# ---------------------------------------------------------------------
def ensure_tree(root, companies, date_tag):
    """Every folder the day needs. Permanent company folders, one new
    dated folder each - and never a company folder we cannot name."""
    made = []
    for d in (root,
              os.path.join(root, F_MASTER),
              os.path.join(root, F_MASTER, date_tag),
              os.path.join(root, F_COMPANY),
              os.path.join(root, F_CONTROL),
              os.path.join(root, F_SETTINGS)):
        if not os.path.isdir(d):
            os.makedirs(d)
            made.append(d)
    for c in companies:
        for d in (os.path.join(root, F_COMPANY, _slug(c)),
                  os.path.join(root, F_COMPANY, _slug(c), date_tag)):
            if not os.path.isdir(d):
                os.makedirs(d)
                made.append(d)
    return made




def write_once(path, text, ignore=("Created ",), scrub=()):
    """Write a text file that is never overwritten - rule 7, and the
    reason it matters: you may have written 'Sent by' on yesterday's
    record by hand, and a re-run must not rub it out.

    Same content as what's already there (bar the build time)? Leave it
    alone. Genuinely different - because you fixed the routing and ran
    again? Write it beside the first as (2), (3)... so both the original
    and the correction are on the record.

    Returns (path written, note)."""
    def strip(s):
        #  scrub first - an HTML page is one long line, so dropping whole
        #  lines can't remove a build time buried in the middle of it.
        for pat in scrub:
            s = re.sub(pat, "", s)
        return "\n".join(ln for ln in s.splitlines()
                         if not any(ln.startswith(p) for p in ignore))

    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as fh:
                if strip(fh.read()) == strip(text):
                    return path, "unchanged - left as it was"
        except Exception:
            return path, "already filed - left untouched"
        stem, ext = os.path.splitext(path)
        for n in range(2, 60):
            alt = "{} ({}){}".format(stem, n, ext)
            if not os.path.exists(alt):
                with open(alt, "w", encoding="utf-8") as fh:
                    fh.write(text)
                return alt, "changed since the first run - written as ({})".format(n)
            #  the same correction filed again is NOT another correction -
            #  without this, one genuine re-issue meant every later run
            #  filed an identical (3), (4), (5)... (29 Jul 2026, and the
            #  same rule _place has always had)
            try:
                with open(alt, "r", encoding="utf-8") as fh:
                    if strip(fh.read()) == strip(text):
                        return alt, "already filed as ({})".format(n)
            except Exception:
                pass
        return path, "already filed - left untouched"
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(text)
    return path, ""


def _md5(path):
    h = hashlib.md5()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def _place(src, dest_dir, dest_name):
    """Copy a report in, but NEVER over the top of one already filed -
    rule 7. Returns (filed_path, note, re-issued?).

    Same file already there: leave it, say so, carry on.
    A DIFFERENT file under the same name: the report was re-issued after
    the pack was built. The first one stays exactly where it is and the
    new one is filed beside it as (2) - and the caller is told, so the
    pack is held rather than going out with yesterday's picture on a
    draft that says today. (25 Jul 2026)"""
    dest = os.path.join(dest_dir, dest_name)
    if os.path.exists(dest):
        try:
            if _md5(src) == _md5(dest):
                return dest, "already filed - left untouched", False
        except Exception:
            return dest, "already filed - left untouched", False
        stem, ext = os.path.splitext(dest)
        for n in range(2, 60):
            alt = "{} ({}){}".format(stem, n, ext)
            if not os.path.exists(alt):
                shutil.copy2(src, alt)
                return alt, "re-issued today - filed as ({})".format(n), True
            try:
                if _md5(src) == _md5(alt):
                    return alt, "already filed as ({})".format(n), True
            except Exception:
                pass
        return dest, "already filed - left untouched", False
    shutil.copy2(src, dest)
    return dest, "", False


# ---------------------------------------------------------------------
#  3. THE PRE-SEND CHECKS - all ten, run before anything is ready
# ---------------------------------------------------------------------
#  Companies that legitimately appear on any report - Coates prints it
#  and Cement Australia owns the site, so neither is evidence of a
#  crossed wire.
NEUTRAL = ("coates", "cement australia")


def contact_index(book):
    """Who is on the register, and under which company. Straight off the
    Contacts sheet - the only place a name and an address are tied
    together."""
    by_company = {}
    for c in book.get("contacts") or []:
        e = _txt(c.get("email")).lower()
        if "@" not in e:
            continue
        by_company.setdefault(_key(c.get("company")), set()).add(e)
    return by_company


def domains_of(by_company, names):
    out = set()
    for n in names:
        for e in by_company.get(_txt(n).lower(), ()):
            out.add(e.split("@")[-1])
    return out


def _unescape(s):
    """HTML back to the words a person reads. &amp; -> & and so on."""
    try:
        import html as _html
        return _html.unescape(s)
    except Exception:
        return (s.replace("&amp;", "&").replace("&nbsp;", " ")
                 .replace("&#38;", "&").replace("&quot;", '"')
                 .replace("&lt;", "<").replace("&gt;", ">"))


def _body_names_other_companies(body_file, own_company, own_data_company):
    """Read the report that is about to be pasted into the email and look
    for ANOTHER contractor's name in it. Checking the filename is not
    enough - this reads the report itself, because this is the check that
    stands between a DGH gear list and a Cleanaway inbox."""
    if not body_file or not os.path.isfile(body_file):
        return ["(there is no report to read)"]
    try:
        with open(body_file, "r", encoding="utf-8", errors="ignore") as fh:
            #  UNESCAPE BEFORE COMPARING. The report writes the company
            #  name as HTML - "Industrial Rescue &amp; Emergency
            #  Training" - and this check was looking for the raw "&".
            #  It never matched, so the report "did not name" the very
            #  company it was about, and the pack was held for ever.
            #
            #  ANY company with an ampersand hits this. The dangerous
            #  part is not the held pack: it is that the obvious way to
            #  make a stuck pack go out is to loosen this check, and
            #  this check is the thing standing between a DGH gear list
            #  and a Cleanaway inbox. Unescaping makes it STRICTER -
            #  both halves now compare real text against real text.
            #  (3 Aug 2026.)
            text = _unescape(fh.read()).lower()
    except Exception as exc:
        #  Can't read it, can't clear it. A check never passes by default.
        return ["(could not read the report: {})".format(exc)]
    #  It must be THIS company's report, not merely free of anyone else's
    #  name - a blank or wrong-entity report would otherwise sail through.
    if _txt(own_data_company).lower() not in text:
        return ["(this report does not name {})".format(own_data_company)]
    return _others_in(text, own_company, own_data_company)


def _others_in(text, own_company, own_data_company):
    """Which OTHER contractors does this text name?"""
    mine = {_txt(own_company).lower(), _txt(own_data_company).lower()}
    found = []
    for name, data_name in ALIASES.items():
        for cand in (name, data_name):
            c = _txt(cand).lower()
            if not c or c in mine:
                continue
            if any(n in c for n in NEUTRAL):
                continue
            if c in text and name not in found:
                found.append(name)
    return found


#  A PDF built by the browser stores its words as glyph numbers in a
#  subset font, so reading the PDF bytes finds no names even when the
#  page is full of them. Every PDF this suite attaches is printed from
#  an HTML page, and that page is the readable twin - so the check reads
#  the twin. If a PDF turns up with no twin to read, the check FAILS.
#  It never says "clean" about a document it could not open. (3 Aug 2026)
def _unversioned(stem):
    """'Report - 2026-08-03 (2)' -> 'Report - 2026-08-03'. A report
    re-issued mid-day is filed beside the first copy as (2), (3)... so
    anything matching on the filed name has to see through that."""
    return re.sub(r"\s*\(\d+\)$", "", _txt(stem))


def _twin_of(pdf_path, extra_dirs=()):
    """The HTML page a filed PDF was printed from, if it can be found."""
    stem = os.path.splitext(os.path.basename(pdf_path))[0]
    here = os.path.dirname(pdf_path)
    for s in (stem, _unversioned(stem)):
        for d in [here] + [x for x in extra_dirs if x]:
            cand = os.path.join(d, s + ".html")
            if os.path.isfile(cand):
                return cand
    return None


def _pack_names_other_companies(pack, own_company, own_data_company,
                                extra_dirs=()):
    """EVERY document on this email - the one in the body and every one
    on the paperclip - read for another contractor's name.

    Until 3 Aug 2026 this only read the body. The site-wide safety report
    was attached to all eleven contractor emails and named every other
    contractor's overdue gear, item, asset number and days out - and the
    pack still recorded 'No other company's information in the pack -
    clean', because nothing ever opened the attachment."""
    strays = list(_body_names_other_companies(
        pack.get("body_file"), own_company, own_data_company))
    body = os.path.abspath(pack.get("body_file") or "")
    for att in pack.get("attachments") or []:
        if os.path.abspath(att) == body:
            continue
        if att.lower().endswith((".html", ".htm")):
            src = att
        else:
            src = _twin_of(att, extra_dirs)
        if not src:
            strays.append("(nothing could read {} - not cleared)".format(
                os.path.basename(att)))
            continue
        try:
            with open(src, "r", encoding="utf-8", errors="ignore") as fh:
                text = _unescape(fh.read()).lower()
        except Exception as exc:
            strays.append("(could not read {}: {})".format(
                os.path.basename(att), exc))
            continue
        for n in _others_in(text, own_company, own_data_company):
            label = "{} (in {})".format(n, os.path.basename(att))
            if label not in strays:
                strays.append(label)
    return strays


def run_checks(pack, date_tag):
    """Every pre-send check, run before anything is called ready. Returns
    a list of (check name, passed, detail). One failure and the pack
    stays in Draft - Check Required; it never reaches Draft Ready."""
    out = []

    def check(name, ok, detail=""):
        out.append((name, bool(ok), detail))

    co = pack["company"]
    external = pack["kind"].lower().startswith("external")

    #  1. the company is real and on the register
    check("Company matched to the on-hire register",
          bool(pack["data_company"]),
          pack["data_company"] or
          "'{}' holds nothing on hire - no report to send".format(co))

    #  2. the folder belongs to this company and this day, AND every
    #     file filed in it carries today's date - the folder can't be
    #     right while holding yesterday's paperwork.
    right_folder = pack["folder"].endswith(os.path.join(_slug(co), date_tag))
    wrong_dated = [os.path.basename(f)
                   for f in [pack["body_file"]] + list(pack["attachments"])
                   if f and date_tag not in os.path.basename(f)]
    check("Folder is this company's, and everything in it is today's",
          right_folder and not wrong_dated,
          ", ".join(wrong_dated) if wrong_dated
          else os.path.join(_slug(co), date_tag))

    #  3. the subject names the company
    check("Company named in the subject line",
          _slug(co).lower() in pack["subject"].lower()
          or co.lower() in pack["subject"].lower(),
          pack["subject"])

    #  4. somebody is actually receiving it
    check("At least one confirmed address in To",
          bool(pack["to"]), "{} in To".format(len(pack["to"])))

    #  5. the four fixed oversight contacts
    if external:
        fixed = [e.lower() for e in pack["fixed_cc"]]
        missing = [e for e in fixed if e not in
                   [c.lower() for c in pack["cc"]]]
        check("Four fixed oversight contacts in Cc",
              bool(fixed) and not missing,
              "missing: " + ", ".join(missing) if missing
              else "{} in Cc".format(len(pack["cc"])))
    else:
        check("Cement master routing used (its own To and Cc)",
              bool(pack["to"]) and bool(pack["cc"]),
              "{} To / {} Cc".format(len(pack["to"]), len(pack["cc"])))

    #  6. the report in the body is the one built for today, and it was
    #     not re-issued after this pack was filed
    body = os.path.basename(pack["body_file"] or "")
    reissued = pack.get("reissued") or []
    check("Report in the body is today's, and still the current one",
          bool(body) and date_tag in body and not reissued,
          "re-issued since this pack was built: " + ", ".join(reissued)
          if reissued else (body or "nothing generated"))

    #  7. there is something to paste in
    check("Report ready for the email body",
          bool(pack["body_file"]),
          os.path.basename(pack["body_file"] or "") or "nothing generated")

    #  8. the attachment
    if external:
        #  ONE attachment, and it is their own report. Not "their report
        #  is in there somewhere" - the site safety report used to be in
        #  there too, and it named everyone else's overdue gear.
        names = [os.path.basename(a) for a in pack["attachments"]]
        own = _unversioned(os.path.splitext(os.path.basename(
            pack["body_file"] or ""))[0])
        #  Compared without the (2), (3)... a mid-day re-issue adds -
        #  that copy is still their own report, just the newer print.
        ours = [n for n in names
                if own and _unversioned(os.path.splitext(n)[0]) == own]
        check("Their own report is the only attachment",
              bool(own) and len(names) == 1 and len(ours) == 1,
              "attached: " + ", ".join(names) if names
              else "nothing attached")
    else:
        names = [os.path.basename(a).lower() for a in pack["attachments"]]
        missing = [w for w in CEMENT_ATTACHMENTS
                   if not any(n.startswith(w.lower()) for n in names)]
        check("All five Cement attachments present",
              not missing,
              "missing: " + ", ".join(missing) if missing
              else "{} attached".format(len(names)))

    #  9. NOBODY ELSE'S GEAR IN THIS PACK - the control-key check.
    #     A contractor sees their own gear and nobody else's. The client
    #     master is the site-wide summary and names every contractor by
    #     design, so for that pack the same principle is checked at the
    #     other end: no contractor is on the distribution list.
    if external:
        stray = _pack_names_other_companies(pack, co, pack["data_company"])
        check("No other company's information in the pack",
              not stray,
              "found: " + ", ".join(stray) if stray
              else "body and {} attachment(s) all read - clean".format(
                  len(pack["attachments"])))
    else:
        outsiders = sorted(pack.get("outsiders") or [])
        check("Client master goes to Cement and Coates only",
              not outsiders,
              "outside address: " + ", ".join(outsiders) if outsiders
              else "{} recipients, all client or Coates".format(
                  len(pack["to"]) + len(pack["cc"])))

    #  9b. Every address is a person on the contact register, against
    #      this company. Nothing is ever guessed, and an address pasted
    #      into the wrong row cannot slip through.
    unknown = sorted(pack.get("unknown_addresses") or [])
    check("Every address on the contact register for this company",
          not unknown,
          "not on the register: " + ", ".join(unknown) if unknown
          else "{} address(es) confirmed".format(
              len(pack["to"]) + len(pack["cc"])))

    #  10. the attachments are real, once each
    names = [os.path.basename(a) for a in pack["attachments"]]
    gone = [n for n, a in zip(names, pack["attachments"])
            if not os.path.exists(a)]
    check("Every attachment on disk, listed once",
          len(names) == len(set(names)) and not gone,
          "missing: " + ", ".join(gone) if gone
          else ("duplicate attachment" if len(names) != len(set(names))
                else "{} file(s)".format(len(names))))

    #  Extra: the draft, opened the way a mail client opens it, is the
    #  email we meant to write. Writing a file is not proof.
    rb = pack.get("readback") or []
    check("Draft reads back exactly as addressed",
          not rb, "; ".join(rb) if rb else "To, Cc, body and attachments "
                                           "all confirmed")

    #  Extra, because an email nobody receives is not a report:
    mb = pack.get("size_mb") or 0.0
    check("Email size within the mail server's limit",
          mb <= MAX_EMAIL_MB,
          "{:.1f} MB (limit {:.0f} MB)".format(mb, MAX_EMAIL_MB))

    #  Extra: the PDF must not predate the report it came from.
    stale = pack.get("stale") or []
    check("Every PDF built from today's report",
          not stale,
          "stale: " + ", ".join(stale) if stale else "all current")

    return out


def failures(results):
    return ["{} ({})".format(n, d) if d else n
            for n, ok, d in results if not ok]


# ---------------------------------------------------------------------
#  4. THE EMAIL RECORD - the trail left behind
# ---------------------------------------------------------------------
def record_signed(folder, date_tag):
    """Has somebody written on today's record for this company? Once the
    'Sent by' line is filled in, this day has been acted on and nothing
    the script does may quietly replace it."""
    path = os.path.join(folder, "Email Record - {}.txt".format(date_tag))
    if not os.path.isfile(path):
        return False
    try:
        with open(path, "r", encoding="utf-8") as fh:
            for line in fh:
                if line.strip().startswith("Sent by"):
                    value = line.split(":", 1)[1] if ":" in line else ""
                    return bool(value.strip().strip("_").strip())
    except Exception:
        return True                 # can't read it - treat it as precious
    return False


def write_record(pack, date_tag, generated):
    lines = []
    w = lines.append
    w("COATES | K2 SHUTDOWN 2026 - EMAIL RECORD")
    w("Cement Australia K2 - Gladstone")
    w("Author: Andrew Fisher | POWERED BY SITEIQ")
    w("=" * 66)
    w("Company       : {}".format(pack["company"]))
    w("Package       : {}".format(pack["kind"]))
    w("Report date   : {}".format(date_tag))
    w("Created       : {}".format(generated.strftime("%d %b %Y %H:%M")))
    w("Status        : {}".format(pack["status"]))
    w("")
    w("SUBJECT")
    w("  " + pack["subject"])
    w("")
    w("TO ({})".format(len(pack["to"])))
    for e in pack["to"]:
        w("  " + e)
    w("")
    w("CC ({})".format(len(pack["cc"])))
    for e in pack["cc"]:
        w("  " + e)
    w("")
    w("PASTED INTO THE EMAIL BODY")
    w("  " + (os.path.basename(pack["body_file"]) if pack["body_file"]
              else "(nothing - no report generated)"))
    w("")
    w("ATTACHMENTS ({})".format(len(pack["attachments"])))
    for a in pack["attachments"]:
        w("  " + os.path.basename(a))
    if not pack["attachments"]:
        w("  (none)")
    w("")
    if pack.get("checks"):
        w("PRE-SEND CHECKS")
        for name, ok, detail in pack["checks"]:
            w("  [{}] {}{}".format("PASS" if ok else "FAIL", name,
                                   "  -  " + detail if detail else ""))
        w("")
    if pack["errors"]:
        w("WHY THIS IS NOT READY TO SEND")
        for e in pack["errors"]:
            w("  - " + e)
        w("")
    w("SENT CONFIRMATION")
    w("  Sent by      : ____________________")
    w("  Date / time  : ____________________")
    w("  Update the status above to 'Sent' once it has gone.")
    w("")
    w("-" * 66)
    w("Routing taken from {}".format(os.path.basename(pack["book"])))
    w("Nothing in this record was typed by hand - change the workbook and")
    w("the next run follows it.")
    path = os.path.join(pack["folder"], "Email Record - {}.txt".format(date_tag))
    text = "\n".join(lines) + "\n"

    #  Two different things live in this one file, and they need opposite
    #  treatment:
    #
    #    what YOU wrote  - the 'Sent by' line, filled in by hand. Sacred.
    #                      Never touched, never moved, never overwritten.
    #    what the RUN wrote - status, addresses, checks. This has to be
    #                      current, because this is the file you open to
    #                      decide whether to send.
    #
    #  So: once you've signed it, a later run leaves it exactly as it is
    #  and writes its version beside it as (2). Until then, the record
    #  stays current and the earlier one is kept as (earlier N).
    signed = False
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as fh:
                old = fh.read()
            same = "\n".join(l for l in old.splitlines()
                             if not l.startswith("Created ")) == \
                   "\n".join(l for l in text.splitlines()
                             if not l.startswith("Created "))
            if same:
                pack["record_note"] = "unchanged - left as it was"
                return path
            for line in old.splitlines():
                if not line.strip().startswith("Sent by"):
                    continue
                #  Anything written after the colon that isn't the blank
                #  line itself means somebody signed this. Typing over
                #  the start of the underscores still counts.
                value = line.split(":", 1)[1] if ":" in line else ""
                if value.strip().strip("_").strip():
                    signed = True
                break
        except Exception:
            signed = True                 # can't read it - don't touch it

    if signed:
        stem, ext = os.path.splitext(path)
        for n in range(2, 60):
            alt = "{} ({}){}".format(stem, n, ext)
            if not os.path.exists(alt):
                with open(alt, "w", encoding="utf-8") as fh:
                    fh.write(text)
                pack["record_note"] = ("this record is signed - the new one "
                                       "is beside it as ({})".format(n))
                return alt
        pack["record_note"] = "signed - left as it was"
        return path

    if os.path.exists(path):
        stem, ext = os.path.splitext(path)
        moved = False
        for n in range(1, 500):
            old_path = "{} (earlier {}){}".format(stem, n, ext)
            if not os.path.exists(old_path):
                try:
                    os.rename(path, old_path)
                    moved = True
                    pack["record_note"] = ("rewritten - the earlier record is "
                                           "kept as (earlier {})".format(n))
                except OSError:
                    pass
                break
        if not moved:
            #  The earlier record could not be put safely aside, so this
            #  one goes beside it. A record is never written over.
            for n in range(2, 500):
                alt = "{} ({}){}".format(stem, n, ext)
                if not os.path.exists(alt):
                    with open(alt, "w", encoding="utf-8") as fh:
                        fh.write(text)
                    pack["record_note"] = "written beside the earlier record"
                    return alt
            pack["record_note"] = "could not be written - the earlier one kept"
            return path
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(text)
    return path
