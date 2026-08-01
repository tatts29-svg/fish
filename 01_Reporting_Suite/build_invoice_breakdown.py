#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  COATES | INVOICE BREAKDOWN - what the SiteIQ invoice is made of
#  Cement Australia K2 Shutdown 2026 - Gladstone
#  Author: Andrew Fisher | POWERED BY SITEIQ
#
#  WHY (Andrew, 30 Jul 2026): "generally its a one line invoice so i
#  just want to supply clarity professionally on whats it made of...
#  product group in plant and tooling so they get a nice outlay."
#  SiteIQ renders the month's invoice as a single line; this report is
#  the professional companion that goes with it - the invoice broken
#  into its streams, the hire split into Plant / Tooling / Gear, and
#  each of those laid out by SiteIQ's own product groups. A progress
#  invoice this month; the final invoice follows next month.
#
#  EVERY DOLLAR IS SITEIQ'S OWN: streams come straight from
#  DAILY_SUMMARY's invoiced columns, the hire split and product groups
#  from SiteIQ's charge lines - the same partition the cost snapshot
#  PROVES ties to the cent, day by day. Nothing estimated, nothing
#  invented. The month-end rule applies: SiteIQ registers the last day
#  of a month into the NEXT month's invoice.
#
#  The separate monthly invoice (radios, gas monitors, welders) never
#  appears here - the two streams never share a dollar.
# =====================================================================
import datetime as dt
import os
import re
import sys

import openpyxl

import report_paths
import build_cost_snapshot as CS

HERE = os.path.dirname(os.path.abspath(__file__))
money0 = CS.money0


def registers_in(d):
    """The billing month a charge day lands in. THE MONTH-END RULE
    (A. Fisher, 24 Jul 2026): SiteIQ does not invoice the last day of a
    month within that month - it registers into the next month."""
    nxt = d + dt.timedelta(days=1)
    if nxt.month != d.month:
        return dt.date(nxt.year, nxt.month, 1)
    return dt.date(d.year, d.month, 1)


def au_date(v):
    s = str(v or "").strip().split()[0] if str(v or "").strip() else ""
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def read_family_split(wb, inv_days):
    """The hire dollars inside the invoice window, by bucket and by
    SiteIQ's own PRODUCT_FAMILY - the product groups Andrew wants the
    client to see. Same spread and same classifier as read_hire_daily
    (ONE definition across the suite), restricted to the invoice's
    days, so the family sums tie to the bucket sums by construction.
    Returns {bucket: {family: {"amt": $, "items": set(barcodes)}}}."""
    bucket_of = CS.hire_classifier(wb)
    tx = CS._gfind("TRANSACTIONS*.xlsx")
    if not tx:
        return None
    try:
        twb = openpyxl.load_workbook(max(tx, key=os.path.getmtime),
                                     read_only=True, data_only=True)
        if "TRANSACTION_CHARGES" not in twb.sheetnames:
            twb.close()
            return None
        rows = list(twb["TRANSACTION_CHARGES"].iter_rows(values_only=True))
        twb.close()
    except Exception:
        return None
    if not rows:
        return None
    H = {str(v).strip().upper(): i for i, v in enumerate(rows[0]) if v}
    for need in ("HIRE_CHARGE ($)", "LATEST_BARCODE", "SHIFT_CHARGE_DATE FROM"):
        if need not in H:
            return None
    days = set(inv_days)
    out = {"plant": {}, "tool": {}, "gear": {}}
    for r in rows[1:]:
        try:
            hire = float(r[H["HIRE_CHARGE ($)"]] or 0)
        except (TypeError, ValueError):
            hire = 0.0
        if not hire:
            continue
        bc = str(r[H["LATEST_BARCODE"]] or "").strip().upper()
        if bc.startswith("SUB"):
            continue          # welders: separate invoice, separate stream
        f = au_date(r[H["SHIFT_CHARGE_DATE FROM"]])
        t = au_date(r[H.get("SHIFT_CHARGE_DATE_TO", -1)]) or f
        if not f:
            continue
        fam = str(r[H.get("PRODUCT_FAMILY", -1)] or "").strip().title() or "Other"
        bucket = bucket_of(bc)
        ndays = max(1, (t - f).days + 1)
        per = hire / ndays
        for i in range(ndays):
            d = f + dt.timedelta(days=i)
            if d not in days:
                continue
            slot = out[bucket].setdefault(fam, {"amt": 0.0, "items": set()})
            slot["amt"] += per
            if bc:
                slot["items"].add(bc)
    return out


#  DAILY_SUMMARY's invoiced columns -> the streams the client reads.
#  Order is the story order: the hire first, then people, then the rest.
STREAMS = (
    ("hire", "Plant, tooling &amp; equipment hire"),
    ("labour", "Personnel / labour &mdash; progress claims"),
    ("service", "Service claims &mdash; accommodation &amp; consumables"),
    ("transport", "Transport"),
    ("consumable", "Consumables"),
    ("fuel", "Fuel"),
    ("damage", "Damage recovery"),
    ("disposal", "Disposal"),
    ("other", "Other charges"),
    ("waiver", "Waivers"),
    ("stamp", "Stamp duty"),
)


def name_key(s):
    """One key for 'Matthew Wooldridge' AND 'WOOLDRIDGE MATTHEW': the
    name's words, uppercased and sorted. The charge sheet and the
    register write the same person both ways round, and an unmatched
    name dumped everyone into 'Not named' (caught 1 Aug 2026)."""
    return " ".join(sorted(re.sub(r"[^A-Z ]", "", str(s or "").upper())
                           .split()))


def hirer_company_map():
    """Who works for whom, off the newest ON_HIRE and RENTAL_STOCK
    exports - the charge lines name the hirer, the client wants the
    COMPANY totals (Andrew, 1 Aug 2026). Keyed by name_key so the
    word order of a name never matters."""
    out = {}
    for pat, sheet_name, wcol, ccol in (
            ("RENTAL_STOCK*.xlsx", "RENTAL_STOCK", "HIRER_NAME",
             "COMPANY_NAME"),
            ("ON_HIRE*.xlsx", "ON_HIRE", "HIRER_NAME", "COMPANY")):
        hits = CS._gfind(pat)
        if not hits:
            continue
        try:
            wb2 = openpyxl.load_workbook(max(hits, key=os.path.getmtime),
                                         read_only=True, data_only=True)
            ws = (wb2[sheet_name] if sheet_name in wb2.sheetnames
                  else wb2.active)
            rows = list(ws.iter_rows(values_only=True))
            wb2.close()
        except Exception:
            continue
        if not rows:
            continue
        ix = {str(v or "").strip().upper(): i for i, v in enumerate(rows[0])}
        if wcol not in ix or ccol not in ix:
            continue
        for r in rows[1:]:
            w = name_key(r[ix[wcol]])
            c = str(r[ix[ccol]] or "").strip()
            if w and c and w not in out:
                out[w] = c
    return out


def read_charge_lines(inv_days, comap):
    """Every hire charge line inside the invoice window, folded per
    item-and-hirer with the hirer's company on it - the full cost-check
    detail the client may ask for (Andrew, 1 Aug 2026: "they may want a
    full breakdown to check over costs... like a story - this company
    did this and that, and here is the breakdown and total"). Same
    source sheet, same day-proration and same welder exclusion as the
    product-group split, so the appendix total ties to the hire split
    BY CONSTRUCTION, not by luck."""
    tx = CS._gfind("TRANSACTIONS*.xlsx")
    if not tx:
        return None
    try:
        twb = openpyxl.load_workbook(max(tx, key=os.path.getmtime),
                                     read_only=True, data_only=True)
        if "TRANSACTION_CHARGES" not in twb.sheetnames:
            twb.close()
            return None
        rows = list(twb["TRANSACTION_CHARGES"].iter_rows(values_only=True))
        twb.close()
    except Exception:
        return None
    if not rows:
        return None
    H = {str(v).strip().upper(): i for i, v in enumerate(rows[0]) if v}
    for need in ("HIRE_CHARGE ($)", "LATEST_BARCODE",
                 "SHIFT_CHARGE_DATE FROM"):
        if need not in H:
            return None

    def gv(r, k):
        i = H.get(k)
        return "" if i is None else str(r[i] or "").strip()

    #  WHAT THE ITEM IS (Andrew, 1 Aug 2026: "we need to have a
    #  description of what the item is"): the charge sheet doesn't
    #  always carry a description, so item numbers resolve through the
    #  rental register and then the master's renames - the same clean
    #  names the client sees everywhere else in the suite.
    namemap, bybc = {}, {}
    rs = CS._gfind("RENTAL_STOCK*.xlsx")
    if rs:
        try:
            rwb = openpyxl.load_workbook(max(rs, key=os.path.getmtime),
                                         read_only=True, data_only=True)
            rws = (rwb["RENTAL_STOCK"] if "RENTAL_STOCK" in rwb.sheetnames
                   else rwb.active)
            rrows = list(rws.iter_rows(values_only=True))
            rwb.close()
            rix = {str(v or "").strip().upper(): i
                   for i, v in enumerate(rrows[0])}

            def rg(r, k):
                i = rix.get(k)
                return "" if i is None else str(r[i] or "").strip()
            for r in rrows[1:]:
                k = rg(r, "ITEM_NUMBER")
                v = rg(r, "ITEM_DESCRIPTION")
                b = rg(r, "ITEM_BARCODE").upper()
                if k and v and k not in namemap:
                    namemap[k] = v
                #  the charge sheet's LATEST_BARCODE is the register's
                #  ITEM_BARCODE - and on the real export the barcode IS
                #  often the only identity a charge line carries
                #  (caught 1 Aug 2026: item numbers blank, descriptions
                #  showing as numbers)
                if b and b not in bybc:
                    bybc[b] = (k, v)
        except Exception:
            pass
    try:
        import master_equipment
        MM = master_equipment.load(HERE, quiet=True)
    except Exception:
        MM = None

    def identify(r, bc):
        """(item_number, plain description) for a charge line - the
        charge sheet's own columns first, then the register by item
        number or barcode, then the master's rename on top."""
        it = gv(r, "ITEM_NUMBER")
        desc = gv(r, "ITEM_DESCRIPTION")
        if (not it or not desc) and bc in bybc:
            rit, rdesc = bybc[bc]
            it = it or rit
            desc = desc or rdesc
        if not it and bc and not bc.startswith("BC"):
            it = bc               # SiteIQ often barcodes BY item number
        desc = desc or namemap.get(it, "")
        if MM is not None and it:
            try:
                desc = MM.disp(it, desc or ("Item " + it))
            except Exception:
                pass
        return it, (desc or ("Item " + it if it else bc))
    days = set(inv_days)
    fold = {}
    for r in rows[1:]:
        try:
            hire = float(r[H["HIRE_CHARGE ($)"]] or 0)
        except (TypeError, ValueError):
            hire = 0.0
        if not hire:
            continue
        bc = gv(r, "LATEST_BARCODE").upper()
        if bc.startswith("SUB"):
            continue          # welders: separate invoice, separate stream
        f = au_date(r[H["SHIFT_CHARGE_DATE FROM"]])
        t = au_date(r[H.get("SHIFT_CHARGE_DATE_TO", -1)]) or f
        if not f:
            continue
        ndays = max(1, (t - f).days + 1)
        per = hire / ndays
        inwin = [f + dt.timedelta(days=i) for i in range(ndays)
                 if f + dt.timedelta(days=i) in days]
        if not inwin:
            continue
        who = gv(r, "HIRER_NAME") or "Site"
        co = (gv(r, "COMPANY") or gv(r, "COMPANY_NAME")
              or comap.get(name_key(who), "") or "Not named")
        #  SITE PLANT is idle-capable gear that lives on site by design
        #  - kept OUT of the company splits and shown on its own at the
        #  very end (Andrew, 1 Aug 2026)
        sp = ("SITE PLANT" in who.upper() or "SITE PLANT" in co.upper())
        it, desc = identify(r, bc)
        e = fold.setdefault((bc, who), {
            "n": desc, "i": it, "w": who, "co": co, "sp": sp,
            "fam": gv(r, "PRODUCT_FAMILY").title() or "Other",
            "f": inwin[0], "t": inwin[-1], "d": 0, "amt": 0.0})
        e["amt"] += per * len(inwin)
        e["d"] += len(inwin)
        e["f"] = min(e["f"], inwin[0])
        e["t"] = max(e["t"], inwin[-1])
    return sorted(fold.values(), key=lambda e: (e["co"].upper(), -e["amt"]))


def appendix_pages(lines):
    """The charge register the way Andrew specified it (1 Aug 2026):
    every company and its split of the total, companies in
    ALPHABETICAL order, and inside each company the lines by ITEM
    NUMBER - description, start date, end date, hire rate and line
    total on every row. Chunked so the PDF paginates cleanly."""
    total = sum(e["amt"] for e in lines)
    #  site plant rides at the very end, never inside a company split
    splant = [e for e in lines if e.get("sp")]
    coline = [e for e in lines if not e.get("sp")]
    sp_amt = sum(e["amt"] for e in splant)
    cos = {}
    for e in coline:
        cos.setdefault(e["co"], []).append(e)
    order = sorted(cos, key=lambda c: c.upper())

    def by_item(e):
        #  numeric item numbers sort as numbers, oddballs after them
        try:
            return (0, int(e["i"]))
        except (TypeError, ValueError):
            return (1, str(e["i"]))

    #  ---- page: who the hire went to ------------------------------
    srows = []
    for co in order:
        ls = cos[co]
        amt = sum(e["amt"] for e in ls)
        srows.append(
            ("<tr><td style='font-weight:700'>{co}</td>"
             "<td style='text-align:right'>{n}</td>"
             "<td style='text-align:right'>{w}</td>"
             "<td style='text-align:right'>{d:,}</td>"
             "<td style='text-align:right'>${a:,.2f}</td>"
             "<td style='text-align:right'>{p:.1f}%</td></tr>").format(
                co=CS.html.escape(co), n=len(ls),
                w=len(set(e["w"] for e in ls)),
                d=int(sum(e["d"] for e in ls)), a=amt,
                p=100.0 * amt / total if total else 0.0))
    summary = (
        "<div class='card'><h2>Who the hire went to</h2>"
        "<div class='cap'>Every company's split of the invoice's hire "
        "dollars, A to Z. Each company then gets its own pages: every "
        "charge line by item number - what it is, who had it, start "
        "and end date, the day rate and the line total - so any cost "
        "can be checked line by line. A line spanning the month-end "
        "carries only its in-window days, so these companies add to "
        "the hire total exactly. Welders, radios and gas monitors bill "
        "on their own monthly invoice and are itemised there.</div>"
        "<table><thead><tr><th>Company</th>"
        "<th style='text-align:right'>Lines</th>"
        "<th style='text-align:right'>People</th>"
        "<th style='text-align:right'>Charge-days</th>"
        "<th style='text-align:right'>Amount</th>"
        "<th style='text-align:right'>Share</th></tr></thead><tbody>"
        + "".join(srows) +
        (("<tr><td style='color:#8A94A2'>Site plant &mdash; on site by "
          "design, kept separate (last pages)</td>"
          "<td style='text-align:right'>{n}</td><td></td>"
          "<td style='text-align:right'>{d:,}</td>"
          "<td style='text-align:right'>${a:,.2f}</td>"
          "<td style='text-align:right'>{p:.1f}%</td></tr>").format(
             n=len(splant), d=int(sum(e["d"] for e in splant)), a=sp_amt,
             p=100.0 * sp_amt / total if total else 0.0)
         if splant else "") +
        ("<tr><td style='font-weight:700;border-top:2px solid " + CS.INK
         + "'>Total hire</td><td></td><td></td><td></td>"
         "<td style='text-align:right;font-weight:700;border-top:2px "
         "solid " + CS.INK + "'>${a:,.2f}</td><td></td></tr>")
        .format(a=total)
        + "</tbody></table></div>")
    pages = [("The full charge register &middot; who the hire went to",
              summary)]

    #  ---- each company's chapter: lines by ITEM NUMBER --------------
    PER = 24
    for co in order:
        ls = sorted(cos[co], key=by_item)
        amt = sum(e["amt"] for e in ls)
        biggest = max(ls, key=lambda e: e["amt"])
        story = (
            "<div class='cap'><b>{co}</b> &mdash; {n} hire line{s} "
            "through {w} of their people, {d:,} charge-days &mdash; "
            "<b>${a:,.2f}</b>, {p:.1f}% of the hire on this invoice. "
            "Biggest single line: {big} at ${ba:,.2f}. Lines below in "
            "item-number order.</div>"
        ).format(co=CS.html.escape(co), n=len(ls),
                 s="" if len(ls) == 1 else "s",
                 w=len(set(e["w"] for e in ls)),
                 d=int(sum(e["d"] for e in ls)),
                 a=amt, p=100.0 * amt / total if total else 0.0,
                 big=CS.html.escape(biggest["n"][:46]), ba=biggest["amt"])
        chunks = [ls[i:i + PER] for i in range(0, len(ls), PER)]
        for ci, chunk in enumerate(chunks):
            body = ["<div class='card'><h2>" + CS.html.escape(co)
                    + ("" if ci == 0 else " (continued)") + "</h2>"]
            if ci == 0:
                body.append(story)
            body.append("<table><thead><tr><th>Item no</th>"
                        "<th>Description</th>"
                        "<th>Hirer</th><th>Start</th><th>End</th>"
                        "<th style='text-align:right'>Days</th>"
                        "<th style='text-align:right'>Day rate</th>"
                        "<th style='text-align:right'>Total</th></tr>"
                        "</thead><tbody>")
            for e in chunk:
                rate = e["amt"] / e["d"] if e["d"] else 0.0
                body.append(
                    ("<tr><td style='white-space:nowrap'>{i}</td>"
                     "<td>{n}</td><td>{w}</td>"
                     "<td style='white-space:nowrap'>{f}</td>"
                     "<td style='white-space:nowrap'>{t}</td>"
                     "<td style='text-align:right'>{d}</td>"
                     "<td style='text-align:right'>${r:,.2f}</td>"
                     "<td style='text-align:right'>${a:,.2f}</td></tr>")
                    .format(i=(CS.html.escape(e["i"]) if e["i"]
                               else "&mdash;"),
                            n=CS.html.escape(e["n"][:48]),
                            w=CS.html.escape(e["w"][:22]),
                            f=e["f"].strftime("%d %b %Y"),
                            t=e["t"].strftime("%d %b %Y"),
                            d=e["d"], r=rate, a=e["amt"]))
            if ci == len(chunks) - 1:
                body.append(
                    ("<tr><td colspan='7' style='font-weight:700;color:"
                     + CS.INK + ";border-top:2px solid " + CS.INK +
                     "'>{co} total &mdash; {p:.1f}% of the hire on this "
                     "invoice</td>"
                     "<td style='text-align:right;font-weight:700;color:"
                     + CS.INK + ";border-top:2px solid " + CS.INK +
                     "'>${a:,.2f}</td></tr>").format(
                        co=CS.html.escape(co), a=amt,
                        p=100.0 * amt / total if total else 0.0))
            body.append("</tbody></table></div>")
            pages.append(("The full charge register &middot; "
                          + CS.html.escape(co), "".join(body)))

    #  ---- site plant, at the very end as asked ----------------------
    if splant:
        ls = sorted(splant, key=by_item)
        chunks = [ls[i:i + PER] for i in range(0, len(ls), PER)]
        for ci, chunk in enumerate(chunks):
            body = ["<div class='card'><h2>Site plant &mdash; on site "
                    "by design" + ("" if ci == 0 else " (continued)")
                    + "</h2>"]
            if ci == 0:
                body.append(
                    ("<div class='cap'>The barriers, chutes, hoppers and "
                     "site infrastructure that live on site for the whole "
                     "shut &mdash; on hire by design, idle-capable, and "
                     "deliberately kept OUT of the company splits above. "
                     "{n} line{s}, {d:,} charge-days, "
                     "<b>${a:,.2f}</b> ({p:.1f}% of the hire on this "
                     "invoice). Lines by item number.</div>").format(
                        n=len(ls), s="" if len(ls) == 1 else "s",
                        d=int(sum(e["d"] for e in ls)), a=sp_amt,
                        p=100.0 * sp_amt / total if total else 0.0))
            body.append("<table><thead><tr><th>Item no</th>"
                        "<th>Description</th>"
                        "<th>Hirer</th><th>Start</th><th>End</th>"
                        "<th style='text-align:right'>Days</th>"
                        "<th style='text-align:right'>Day rate</th>"
                        "<th style='text-align:right'>Total</th></tr>"
                        "</thead><tbody>")
            for e in chunk:
                rate = e["amt"] / e["d"] if e["d"] else 0.0
                body.append(
                    ("<tr><td style='white-space:nowrap'>{i}</td>"
                     "<td>{n}</td><td>{w}</td>"
                     "<td style='white-space:nowrap'>{f}</td>"
                     "<td style='white-space:nowrap'>{t}</td>"
                     "<td style='text-align:right'>{d}</td>"
                     "<td style='text-align:right'>${r:,.2f}</td>"
                     "<td style='text-align:right'>${a:,.2f}</td></tr>")
                    .format(i=(CS.html.escape(e["i"]) if e["i"]
                               else "&mdash;"),
                            n=CS.html.escape(e["n"][:48]),
                            w=CS.html.escape(e["w"][:22]),
                            f=e["f"].strftime("%d %b %Y"),
                            t=e["t"].strftime("%d %b %Y"),
                            d=e["d"], r=rate, a=e["amt"]))
            if ci == len(chunks) - 1:
                body.append(
                    ("<tr><td colspan='7' style='font-weight:700;color:"
                     + CS.INK + ";border-top:2px solid " + CS.INK +
                     "'>Site plant total</td>"
                     "<td style='text-align:right;font-weight:700;color:"
                     + CS.INK + ";border-top:2px solid " + CS.INK +
                     "'>${a:,.2f}</td></tr>").format(a=sp_amt))
            body.append("</tbody></table></div>")
            pages.append(("The full charge register &middot; site plant",
                          "".join(body)))
    return pages


def fam_rows(fams, total, keep=9):
    """Family table rows for one bucket: biggest first, an honest
    'everything else' line for the tail, a share bar on every row."""
    items = sorted(fams.items(), key=lambda kv: -kv[1]["amt"])
    shown, rest_amt, rest_items, rest_n = items[:keep], 0.0, set(), 0
    for f, v in items[keep:]:
        rest_amt += v["amt"]
        rest_items |= v["items"]
        rest_n += 1
    rows = ""
    for f, v in shown:
        pc = (v["amt"] / total * 100.0) if total else 0.0
        rows += ("<tr><td>{f}</td><td class='n'>{n}</td><td class='n'>{a}</td>"
                 "<td style='width:130px'><div class='bar' style='height:9px;margin:4px 0'>"
                 "<i style='width:{w:.0f}%'></i></div></td>"
                 "<td class='n' style='color:" + CS.MUTED + "'>{p:.0f}%</td></tr>").format(
            f=f, n=len(v["items"]), a=money0(v["amt"]), w=min(100.0, pc), p=pc)
    if rest_amt > 0.005:
        pc = (rest_amt / total * 100.0) if total else 0.0
        rows += ("<tr><td style='color:" + CS.MUTED + "'>{n} more product group(s)</td>"
                 "<td class='n'>{i}</td><td class='n'>{a}</td>"
                 "<td style='width:130px'><div class='bar' style='height:9px;margin:4px 0'>"
                 "<i style='width:{w:.0f}%'></i></div></td>"
                 "<td class='n' style='color:" + CS.MUTED + "'>{p:.0f}%</td></tr>").format(
            n=rest_n, i=len(rest_items), a=money0(rest_amt), w=min(100.0, pc), p=pc)
    rows += ("<tr><td style='color:" + CS.INK + ";font-weight:700'>Total</td>"
             "<td class='n'></td><td class='n' style='color:" + CS.INK + ";font-weight:700'>{a}</td>"
             "<td></td><td></td></tr>").format(a=money0(total))
    return rows


def fam_card(title, cap, fams, total):
    return ("<div class='card' style='margin-top:14px'><h2>" + title + "</h2>"
            "<div class='cap'>" + cap + "</div>"
            "<table class='tight'><thead><tr><th>Product group</th>"
            "<th style='text-align:right'>Items</th>"
            "<th style='text-align:right'>Amount</th><th></th>"
            "<th style='text-align:right'>Share</th></tr></thead><tbody>"
            + fam_rows(fams, total) + "</tbody></table></div>")


def main():
    print("=" * 70)
    print(" COATES | INVOICE BREAKDOWN - what the SiteIQ invoice is made of")
    print("=" * 70)
    ds, ds_path, ds_asat = CS.read_daily_summary()
    if not ds:
        print(" DAILY_SUMMARY export not found - download it from SiteIQ,")
        print(" save it over the top, and run this again.")
        return 1

    #  which invoice? the latest billing month in the data, unless told
    #  (python build_invoice_breakdown.py 2026-07)
    marg = next((a for a in sys.argv[1:]
                 if re.match(r"^\d{4}-\d{2}$", str(a))), None)
    months = sorted(set(registers_in(d) for d in ds))
    if marg:
        month = dt.date(int(marg[:4]), int(marg[5:7]), 1)
    else:
        month = months[-1]
        #  more than one billing month in the data = ask which invoice
        #  (Andrew, 1 Aug 2026: "i dont have an option for July, only
        #  says august" - the month-end rule had rolled 31 Jul into
        #  August's invoice and taken 'latest' with it)
        if len(months) > 1:
            try:
                if sys.stdin is not None and sys.stdin.isatty():
                    print(" Billing months in the data:")
                    for i, m in enumerate(months, 1):
                        nd = sum(1 for d in ds if registers_in(d) == m)
                        print("   {}  {:<16} ({} invoiced day{})".format(
                            i, m.strftime("%B %Y"), nd,
                            "" if nd == 1 else "s"))
                    ans = input(" Which invoice? (Enter = {}) : ".format(
                        month.strftime("%B %Y"))).strip()
                    if ans.isdigit() and 1 <= int(ans) <= len(months):
                        month = months[int(ans) - 1]
            except (EOFError, KeyboardInterrupt):
                pass
    inv_days = sorted(d for d in ds if registers_in(d) == month)
    if not inv_days:
        print(" No invoiced days register into {:%B %Y} yet.".format(month))
        return 1
    mname = month.strftime("%B %Y")
    print(" Invoice month : {}  ({} invoiced day(s), {} - {})".format(
        mname, len(inv_days), inv_days[0].strftime("%d %b"),
        inv_days[-1].strftime("%d %b")))

    #  the full charge register is OPT-IN - the standard client pack
    #  stays lean; when the client asks to check costs, run with FULL
    #  (or answer F here) and every line rides along, company by
    #  company (Andrew, 1 Aug 2026)
    full = any(str(a).upper() in ("FULL", "F") for a in sys.argv[1:])
    if not full:
        try:
            if sys.stdin is not None and sys.stdin.isatty():
                ans = input(" Enter = summary pack | F = add the FULL "
                            "charge register, company by company : ")
                full = str(ans).strip().upper() in ("F", "FULL")
        except (EOFError, KeyboardInterrupt):
            full = False

    #  ---- the streams, straight from SiteIQ's invoiced columns --------
    sums = {k: sum(ds[d][k] for d in inv_days) for k, _l in STREAMS}
    total = sum(ds[d]["total"] for d in inv_days)
    alloc = sum(sums.values())
    gap = total - alloc
    if abs(gap) >= 0.01:
        #  never hide a cent: whatever DAILY_SUMMARY's total carries
        #  beyond the named columns gets its own honest line
        sums["_unalloc"] = gap
    tie = "PROVEN" if abs(gap) < 0.01 else "gap ${:,.2f} shown as its own line".format(gap)
    print(" Invoice total : {}   (streams sum {} - {})".format(
        money0(total), money0(alloc), tie))
    for k, lab in STREAMS:
        if sums.get(k):
            print("   {:<12} {}".format(k, money0(sums[k])))

    #  ---- the hire split & the product groups -------------------------
    wb = openpyxl.load_workbook(CS.find_workbook(), data_only=True)
    hire_daily = CS.read_hire_daily(wb) or {}
    tied, stale = [], []
    for d in inv_days:
        sp = hire_daily.get(d)
        if sp and abs((sp["plant"] + sp["tool"] + sp["gear"]) - ds[d]["hire"]) < 1:
            tied.append(d)
        elif ds[d]["hire"]:
            stale.append(d)
    split = {k: sum(hire_daily[d][k] for d in tied) for k in ("plant", "tool", "gear")}
    stale_amt = sum(ds[d]["hire"] for d in stale)
    print(" Hire split    : plant {} | tooling {} | gear {}  ({}/{} day(s) tie to the cent)".format(
        money0(split["plant"]), money0(split["tool"]), money0(split["gear"]),
        len(tied), len(tied) + len(stale)))
    if stale:
        print(" !! {} day(s) not yet posted by SiteIQ ({}) - {} of hire can't be".format(
            len(stale), ", ".join(d.strftime("%d %b") for d in stale), money0(stale_amt)))
        print("    split yet. Re-download TRANSACTIONS after 09:30 and run this again.")
    fams = read_family_split(wb, tied) or {"plant": {}, "tool": {}, "gear": {}}

    appx = None
    if full:
        appx = read_charge_lines(tied, hirer_company_map())
        if appx:
            print(" Register      : {} charge lines across {} companies "
                  "ride along".format(
                      len(appx), len(set(e["co"] for e in appx))))
        else:
            print(" Register      : charge lines not readable this run - "
                  "summary only")

    #  ---- the claims booked into this invoice --------------------------
    svc = CS.read_service_invoiced()
    claim_lines = [(w, de, a) for w, de, a in svc.get("rows", [])
                   if au_date(w) and registers_in(au_date(w)) == month]

    #  =================== the pages =====================================
    now = dt.datetime.now()
    asat = now.strftime("%d %b %Y - %H:%M")
    hire_sum = sums.get("hire", 0.0)
    claims_sum = total - hire_sum

    #  ---- sheet 1 - what the invoice is made of ------------------------
    srows = ""
    for k, lab in STREAMS:
        v = sums.get(k, 0.0)
        if not v:
            continue
        pc = v / total * 100.0 if total else 0.0
        srows += ("<tr><td>{l}</td><td class='n'>{a}</td>"
                  "<td style='width:150px'><div class='bar' style='height:10px;margin:4px 0'>"
                  "<i style='width:{w:.0f}%'></i></div></td>"
                  "<td class='n' style='color:" + CS.MUTED + "'>{p:.0f}%</td></tr>").format(
            l=lab, a=money0(v), w=min(100.0, pc), p=pc)
    if sums.get("_unalloc"):
        srows += ("<tr><td style='color:" + CS.MUTED + "'>Other invoiced adjustments</td>"
                  "<td class='n'>{a}</td><td></td><td></td></tr>").format(
            a=money0(sums["_unalloc"]))
    srows += ("<tr><td style='color:" + CS.INK + ";font-weight:700'>Invoice total</td>"
              "<td class='n' style='color:" + CS.INK + ";font-weight:700'>{a}</td>"
              "<td></td><td></td></tr>").format(a=money0(total))

    roll_day = (month - dt.timedelta(days=1))          # last day of prior month
    next_m = (month + dt.timedelta(days=32)).replace(day=1)
    last_of_month = next_m - dt.timedelta(days=1)
    page1 = (
        "<div class='banner'><span class='pill'>Progress invoice</span>"
        "<p>SiteIQ renders the {m} invoice as a <b>single line</b>. This breakdown is the "
        "companion that shows what that line is made of &mdash; every dollar grouped by stream, "
        "straight from SiteIQ's own daily invoicing, covering <b>{d0} &ndash; {d1}</b> "
        "({n} invoiced day(s)). A <b>final invoice</b> follows next month as the shutdown "
        "closes out.</p></div>"
        "<div class='kpis' style='grid-template-columns:repeat(3,1fr)'>"
        "<div class='kpi b'><div class='v'>{tot}</div><div class='l'>{m} invoice &middot; all in</div>"
        "<div class='s'>every stream on this invoice</div></div>"
        "<div class='kpi'><div class='v'>{h}</div><div class='l'>Hire</div>"
        "<div class='s'>plant, tooling &amp; equipment</div></div>"
        "<div class='kpi'><div class='v'>{c}</div><div class='l'>Claims &amp; services</div>"
        "<div class='s'>labour, accommodation, transport &amp; the rest</div></div></div>"
        "<div class='card'><h2>What the invoice is made of</h2>"
        "<div class='cap'>Each stream as SiteIQ invoiced it &mdash; the rows sum to the "
        "invoice total to the cent.</div>"
        "<table><thead><tr><th>Stream</th><th style='text-align:right'>Amount</th><th></th>"
        "<th style='text-align:right'>Share</th></tr></thead><tbody>" + srows + "</tbody></table>"
        "<div class='footnote'>The month-end rule: SiteIQ registers the <b>last day of a month</b> "
        "into the next month's invoice &mdash; {lst}'s hire sits in next month's final invoice"
        + (", and {rd}'s sits in this one".format(rd=roll_day.strftime("%d %b"))
           if roll_day in inv_days else "") +
        ". Radios, gas monitors and the welders are invoiced separately and are "
        "<b>not in this figure</b> &mdash; the two invoices never share a dollar.</div></div>"
    ).format(m=mname, d0=inv_days[0].strftime("%d %b"),
             d1=inv_days[-1].strftime("%d %b %Y"), n=len(inv_days),
             tot=money0(total), h=money0(hire_sum), c=money0(claims_sum),
             lst=last_of_month.strftime("%d %b"))

    #  ---- sheet 2 - inside the hire: the split + plant groups ----------
    hire_tied = split["plant"] + split["tool"] + split["gear"]

    def pct(v):
        return (v / hire_tied * 100.0) if hire_tied else 0.0
    stale_note = ""
    if stale:
        stale_note = (" {a} of hire across {n} day(s) is invoiced but its charge lines "
                      "haven't posted yet (SiteIQ posts them about 09:30 the next morning) "
                      "&mdash; it is in every total above, just not split below yet.").format(
            a=money0(stale_amt), n=len(stale))
    page2 = (
        "<div class='kpis' style='grid-template-columns:repeat(3,1fr)'>"
        "<div class='kpi b'><div class='v'>{p}</div><div class='l'>Plant</div>"
        "<div class='s'>{pp:.0f}% of the hire</div></div>"
        "<div class='kpi'><div class='v'>{t}</div><div class='l'>Tooling</div>"
        "<div class='s'>{tp:.0f}% of the hire</div></div>"
        "<div class='kpi'><div class='v'>{g}</div><div class='l'>General gear</div>"
        "<div class='s'>{gp:.0f}% of the hire</div></div></div>"
    ).format(p=money0(split["plant"]), pp=pct(split["plant"]),
             t=money0(split["tool"]), tp=pct(split["tool"]),
             g=money0(split["gear"]), gp=pct(split["gear"]))
    page2 += fam_card(
        "Plant &mdash; by product group",
        "The site's plant on charge &mdash; plant registers and the plant storage units "
        "(site plant, rubbish chutes, barriers) &mdash; grouped by SiteIQ's own product "
        "families." + stale_note,
        fams["plant"], split["plant"])
    page2 += ("<div class='footnote' style='margin-top:14px'>Every dollar on this page is "
              "SiteIQ's own charge line, split by the store's proven plant identity &mdash; "
              "the same partition the cost snapshot ties to the invoice to the cent, "
              "day by day. Items = individual pieces of gear that earned within the group.</div>")

    #  ---- sheet 3 - tooling & general gear groups ----------------------
    page3 = fam_card(
        "Tooling &mdash; by product group",
        "The tool aisles (Tooling, Hydraulics &amp; Hi-Torque). The number runs small "
        "because the hand-tool aisles go out largely free-issue &mdash; that is the real "
        "shape of the hire, not missing data.",
        fams["tool"], split["tool"])
    page3 += fam_card(
        "General gear &mdash; by product group",
        "The rest of the store &mdash; rigging, electrical, welding, air, lighting and "
        "the like, grouped by SiteIQ's own product families.",
        fams["gear"], split["gear"])

    #  ---- sheet 4 - the claims & services on this invoice --------------
    page4 = ""
    if claim_lines:
        lrows = ""
        for w, de, a in claim_lines:
            d = au_date(w)
            lrows += ("<tr><td style='white-space:nowrap'>{w}</td><td>{d}</td>"
                      "<td class='n'>${a:,.2f}</td></tr>").format(
                w=d.strftime("%d %b %Y"), d=CS.html.escape(de[:70]), a=a)
        lrows += ("<tr><td></td><td style='color:" + CS.INK + ";font-weight:700'>Total claims "
                  "on this invoice</td><td class='n' style='color:" + CS.INK +
                  ";font-weight:700'>${a:,.2f}</td></tr>").format(
            a=sum(a for _w, _de, a in claim_lines))
        page4 = (
            "<div class='card'><h2>The claims &amp; services on this invoice</h2>"
            "<div class='cap'>Booked through SiteIQ as service claims &mdash; to the cent, "
            "the same lines that appear on the invoice.</div>"
            "<table><thead><tr><th>Booked</th><th>Claim line</th>"
            "<th style='text-align:right'>Amount</th></tr></thead><tbody>" + lrows +
            "</tbody></table>"
            "<div class='footnote'>Labour and accommodation are invoiced in lump progress "
            "claims covering a stretch of days at a time; consumable claims are one-off "
            "supply items. Neither is ever double counted against the daily hire.</div></div>")

    #  ---- assemble ------------------------------------------------------
    sections = [("What the invoice is made of", page1),
                ("Inside the hire &middot; plant by product group", page2),
                ("Inside the hire &middot; tooling &amp; general gear", page3)]
    if page4:
        sections.append(("Claims &amp; services on this invoice", page4))
    if appx:
        sections.extend(appendix_pages(appx))
    total_pages = len(sections)
    footer = ("<div class='foot'><span>Data source: <b>SiteIQ</b> &middot; Extract {a} &middot; "
              "Progress invoice breakdown &mdash; final invoice next month &middot; every figure "
              "is SiteIQ's own invoiced line.</span>"
              "<span style='white-space:nowrap'>Author: <b>Andrew Fisher</b> &middot; "
              "POWERED BY SITEIQ</span></div>").format(a=asat)
    sheets = []
    for i, (sub, inner) in enumerate(sections, 1):
        hd = ("<div class='hd2'><div>"
              "<div class='brand'>COATES<span>Equipped for anything</span></div>"
              "<h1>Invoice Breakdown &mdash; {m}</h1><div class='rp'>{sub}</div>"
              "</div><div><div class='siteiq'>POWERED BY SITEIQ</div>"
              "<div class='meta'>As at <b>{a}</b><br>Page {p} of {tp}</div></div></div>"
              ).format(m=mname, sub=sub, a=asat, p=i, tp=total_pages)
        sheets.append("<section class='sheet'><div class='panel'>" + hd +
                      "<div class='rule'></div>" + inner + "<div class='grow'></div>" +
                      (footer if i == total_pages else "") + "</div></section>")
    #  concatenate (never .format) around the CSS - it has its own { }
    page = ("<!DOCTYPE html><html><head><meta charset='utf-8'>"
            "<meta name='viewport' content='width=device-width,initial-scale=1'>"
            "<title>Coates K2 - Invoice Breakdown</title><style>"
            + CS.PANEL_CSS + CS.LIGHT_CSS + "</style></head><body>"
            + "".join(sheets) + "</body></html>")

    dirs = report_paths.out_dirs(HERE)
    stamp = dt.datetime.now().strftime("%Y-%m-%d")
    out = os.path.join(dirs["pages"], "Coates_K2_Invoice_Breakdown_{}_{}.html".format(
        month.strftime("%B%Y"), stamp))
    with open(out, "w", encoding="utf-8") as f:
        f.write(page)
    print(" Saved: {}".format(out))
    pdf_out = os.path.join(dirs["pdf"], os.path.basename(out).replace(".html", ".pdf"))
    if CS.make_pdf(out, pdf_out):
        print(" PDF  : {}".format(pdf_out))
        print("")
        print(" Attach the PDF alongside the SiteIQ invoice - it is the")
        print(" clarity page: one sheet of what the line is made of, then")
        print(" the product-group outlay for plant, tooling and gear.")
    if stale:
        print("")
        print(" REMINDER: re-download TRANSACTIONS + DAILY_SUMMARY after")
        print(" 09:30, run this again, and the not-yet-posted day(s) fill in.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
