#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
K2 COMPANY ON-HIRE REPORT KIT
Cement Australia K2 Shutdown 2026 - Gladstone
Author: Andrew Fisher | POWERED BY SITEIQ

Produces a branded per-company on-hire report (PDF + HTML + Outlook .eml draft)
from the SiteIQ exports sitting in the SAME folder as this script.

This kit NEVER modifies the K2 Excel workbook or any source file. Read-only.

Run:
    double-click 03_RUN_COMPANY_REPORT.bat          (interactive menu)
    python build_company_onhire_report.py --company "BELTEC"
    python build_company_onhire_report.py --all
"""

import argparse
import base64
import datetime as dt
import email.message
import email.policy
import fnmatch
import glob
import html as html_mod
import io
import json
import os
import re
import shutil
import subprocess
import sys

try:
    import openpyxl
except ImportError:
    print("This kit needs the 'openpyxl' Python package and it is not installed.")
    print("On a machine with internet, run:  pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# COMPANY NAME NORMALISATION - edit this dict as new aliases appear.
# Keys and values are compared UPPER-TRIMMED. Applied to BOTH the rental
# (RENTAL_STOCK) and consumables (SALES_STOCK) data, so one company's gear
# and consumables always land on the same report.
# ---------------------------------------------------------------------------
COMPANY_ALIASES = {
    "BELTECH": "BELTEC",
    "DAWSONS ENGINEERING": "DAWSONS",
    "DAWSONS RELINERS": "DAWSONS",
    "MMG AUSTRALIA LIMITED": "MMG",
    #  The Charge Reporter logs the client as plain "Cement Australia";
    #  the hire register knows them as the Holdings entity. Without this
    #  bridge a charge logged at the counter never reached their report -
    #  CON-20260722-001 ($1,090.50 of Ledlensers) vanished exactly this
    #  way. (Found 26 Jul 2026.)
    "CEMENT AUSTRALIA": "CEMENT AUSTRALIA HOLDINGS PTY LTD",
    #  "dke is darknight" (Andrew, 27 Jul 2026). SiteIQ has logged this
    #  outfit both ways - the 26 Jul exports said DKE GROUP, the current
    #  ones say DARK KNIGHT ENGINEERING. Without the bridge they read as
    #  two different contractors: the report splits in two, yesterday's
    #  folder stops matching today's, and the four dkegroup.com.au
    #  contacts sit against a name no report is ever built under.
    "DKE GROUP": "DARK KNIGHT ENGINEERING",
    "DKE": "DARK KNIGHT ENGINEERING",
    #  "ish is ish24" (Andrew, 27 Jul 2026) - Outlook has the space in,
    #  the register doesn't. Same rescue crew either way.
    "ISH 24": "ISH24",
    "ISH": "ISH24",
}

# Acronym-style names that must stay ALL CAPS in display (extend as needed)
DISPLAY_ACRONYMS = {"MMG", "JLG", "BHP", "CBS", "UGL", "CPB", "AMJ", "RKS",
                    "HIS", "RDS", "IMIS", "ISH24", "AMAST", "DGH", "CQ", "I&C"}

# ---------------------------------------------------------------------------
# KNOWN COMPANIES - the K2 contractor register (per Andrew, 17 Jul 2026),
# with the display spelling used on every report: Capital first letter of
# every word, acronyms kept upper. Anything appearing in SiteIQ that is NOT
# on this list gets flagged (console + internal summary) - usually a typo at
# the counter or an uninducted company. NO merging is done here - possible
# duplicates (e.g. Schneider v Schneider Electrical) are Andrew's final-step
# call via COMPANY_ALIASES above.
# ---------------------------------------------------------------------------
KNOWN_COMPANIES = {
    "ABSAFE": "Absafe",
    "ALL FABRICATION ENGINEERING": "All Fabrication Engineering",
    "AMAST": "AMAST",
    "AMJ": "AMJ",
    "BELLIS": "Bellis",
    "BELTEC": "Beltec",
    "BLITZ ENGINEERING": "Blitz Engineering",
    "BOSCH REXROTH": "Bosch Rexroth",
    "CEMENT AUSTRALIA HOLDINGS PTY LTD": "Cement Australia Holdings Pty Ltd",
    "CIVMEC": "Civmec",
    "CLEANAWAY": "Cleanaway",
    "COATES": "Coates",
    "CONTROL SYSTEMS TECHNOLOGY": "Control Systems Technology",
    "CONVEYING & HOISTING": "Conveying & Hoisting",
    "COREFIELD": "Corefield",
    "CQ POWER PUMPING": "CQ Power Pumping",
    "CUTRITE": "Cutrite",
    "DARK KNIGHT ENGINEERING": "Dark Knight Engineering",
    "DAWSONS": "Dawsons",
    "DEMEX": "Demex",
    "DGH ENGINEERING": "DGH Engineering",
    "DOWN TO EARTH": "Down To Earth",
    "EFFECTIVE WORK": "Effective Work",
    "FILTERCARE": "Filtercare",
    "HAYWARDS": "Haywards",
    "HIGH RISK SOLUTIONS": "High Risk Solutions",
    "HIS": "HIS",
    "I&C ELECTRICAL": "I&C Electrical",
    "IMIS": "IMIS",
    "ISH24": "ISH24",
    "LEE CRANE HIRE": "Lee Crane Hire",
    "LOCLUR ENG/SWP": "Loclur Eng/SWP",
    "MACHINE MONITOR": "Machine Monitor",
    "MECHA PTY LTD": "Mecha Pty Ltd",
    "MMG": "MMG",
    "NILSEN": "Nilsen",
    "OZARC": "Ozarc",
    "PROGRAMMED": "Programmed",
    "QWEST CRANE HIRE": "Qwest Crane Hire",
    "RDS": "RDS",
    "REMA TIP TOP": "Rema Tip Top",
    "RKS": "RKS",
    "ROTARY KILN SERVICES": "Rotary Kiln Services",
    "SANDVIK": "Sandvik",
    "SAVOUR THE FLAVOUR": "Savour The Flavour",
    "SCHNEIDER": "Schneider",
    "SCHNEIDER ELECTRICAL": "Schneider Electrical",
    "SEIMENS": "Seimens",
    "SIMMONS CIVIL": "Simmons Civil",
    "SYNC LIFT ENGINEERING": "Sync Lift Engineering",
    "TASMAN ROPE ACCESS": "Tasman Rope Access",
    "UNIVERSAL CRANES": "Universal Cranes",
    "VEOLIA": "Veolia",
    "WALTZ CONSTRUCTION": "Waltz Construction",
    "WALZ GROUP": "Walz Group",
    "XTREME ENGINEERING": "Xtreme Engineering",
}

# ---------------------------------------------------------------------------
# RADIO CHARGING (directed by Andrew, 17 Jul 2026): radio HANDSETS charge
# $12.81 each per day whether on hire to a company or not, starting
# 17 Jul 2026 - no backdating. Handsets only: the Radios storage unit also
# holds IMPRES batteries and radio covers, and those do NOT attract the
# charge. Hire costs appear ONLY on the internal all-companies summary,
# never on the client-facing per-company reports.
# ---------------------------------------------------------------------------
RADIO_DAILY_RATE = 12.81
RADIO_CHARGE_START = dt.date(2026, 7, 17)
# 70 of the 72 handsets are billable; 2 are held as unbilled spares in case
# a unit stops working (Andrew's direction, 17 Jul 2026). The spares are
# DISCLOSED on the internal summary, never silently hidden.
RADIO_BILLABLE_CAP = 70
RADIO_HANDSET_KEYWORDS = ("DP4801", "TWO-WAY RADIO", "TWO WAY RADIO",
                          "HANDHELD RADIO", "PORTABLE RADIO", "UHF RADIO")

GAS_MONITOR_KEYWORDS = ("GAS MONITOR", "GASMONITOR", "MULTI-GAS", "MULTI GAS",
                        "MULTIGAS", "GAS DETECTOR", "BW FLEX", "BWFLEX",
                        "HONEYWELL BW", "DRAEGER", "DRAGER", "X-AM", "XAM")
GAS_MONITOR_ACCESSORY_KEYWORDS = ("CHARGER", "CHARGING", "DOCK", "CALIBRATION",
                                  "BUMP STATION", "BUMP TEST STATION",
                                  "REGULATOR", "TUBING", "TEST GAS",
                                  "GAS CYLINDER", "SENSOR", "FILTER", "CLIP",
                                  "CASE")


def is_gas_monitor_desc(description):
    """True for an actual gas monitor unit - not a charger, bump station,
    calibration accessory or spare part."""
    d = (description or "").upper()
    if any(k in d for k in GAS_MONITOR_ACCESSORY_KEYWORDS):
        return False
    return any(k in d for k in GAS_MONITOR_KEYWORDS)

# ---------------------------------------------------------------------------
# SITE PLANT (directed by Andrew, 17 Jul 2026): plant lives in the
# "Cement Site Plant" storage unit. Site convention: when plant comes BACK
# from a contractor it is put on hire to hirer "Site Plant Equipment",
# company Coates. So anything on hire to Site Plant Equipment is by
# definition sitting available on site and still being charged - that is
# the potential cost saving the Site Plant report surfaces.
# ---------------------------------------------------------------------------
PLANT_STORAGE_UNIT = "CEMENT SITE PLANT"

# Site infrastructure: barriers, crowd control, rubbish chutes, hoppers,
# brackets. On site for the duration of the shutdown by design - NOT counted
# as idle plant, not a saving opportunity, not an issue. Mentioned by name
# and total quantity only.
INFRA_STORAGE_UNITS = {"CEMENT BARRIERS", "CEMENT RUBBISH CHUTES"}

# Gas monitors: whole fleet charges $29.75/unit/day from 24 Jul 2026 until
# the last day the Daily Tracking forecast carries a gas value, on hire or
# not (Andrew's direction, 17 Jul 2026).
GAS_RATE = 29.75
GAS_START = dt.date(2026, 7, 24)

# Demob: the gear return push starts 3-4 days out. K2 finishes 11 Aug 2026 -
# edit here if the schedule moves.
SHUTDOWN_END = dt.date(2026, 8, 11)

# Dead placeholder SKUs (per Andrew, 19 Jul): these exact SKU_NUMBERs must
# NEVER appear on any report, in any run, regardless of what quantities any
# export happens to carry. Blocklisted by number, not by heuristic.
DEAD_SKUS = {
    "CEME9016-590678", "CEME9016-590688", "CEME9016-590700",
    "CEME9016-590716", "CEME9016-590719", "CEME9016-590721",
    "CEME9016-590725",
}


def is_site_plant_equipment(hirer):
    """Match the 'Site Plant Equipment' hirer however it's been typed
    ('Site Plant Equipment', 'Site - Plant Equipment', extra spaces...)."""
    return "SITEPLANTEQUIPMENT" in re.sub(r"[^A-Z]", "", hirer.upper())


def is_radio_handset(item):
    """True for an actual radio handset - not a battery, cover or charger.
    Checks the export's original description as well as any master rename,
    so the fleet-level billing rules can never be broken by a rename."""
    if item["storage_unit"].upper() != "RADIOS":
        return False
    d = (item["description"] + " " + item.get("desc0", "")).upper()
    return any(k in d for k in RADIO_HANDSET_KEYWORDS)


#  Gear charged on the SEPARATE INVOICE (billed monthly outside
#  SiteIQ), told apart the way Andrew showed on 28 Jul 2026 - by the
#  BARCODE, never by the description:
#
#    sub-hired (separate invoice):  ITEM_BARCODE SUBHARVEY001 / FK505
#    Coates, charged via SiteIQ:    ITEM_BARCODE 1312691 / COA~5564082
#
#  The dashed ITEM_NUMBER (16816-...) is NOT enough on its own - every
#  site-created item carries it, including the Cement radios, gas
#  monitors and the purchased slings (verified against the live
#  register, 28 Jul 2026). Description words are not used either -
#  there are Coates welders and forklifts legitimately charged through
#  SiteIQ in the same storage unit. Radios and gas monitors stay
#  fleet-level rules. The Separate Invoice Tracker (49) carries every
#  sub-hired dollar. NO COUNTS live in this rule - what is sub-hired
#  comes off the register and the Baseplan pull, run by run.
SEP_BARCODES = ("FK505",)          # supplier codes without the SUB prefix

#  The stream stamp (Andrew, 28 Jul 2026: "make sure with any costs its
#  clearly viewed and nothing overlaps onto another figure"). Printed
#  wherever SiteIQ dollars appear, so no figure can ever be read as the
#  other stream's - the Separate Invoice Tracker carries the mirror
#  statement on its own banner.
STREAM_NOTE_SITEIQ = (
    "<div class='note' style='border-left:4px solid #F26222'>"
    "<b>COST STREAM: SiteIQ invoicing only.</b> Every dollar on this "
    "page bills through SiteIQ. The separate monthly invoice - radios, "
    "gas monitors, sub-hired welders and forklift, compressor, crib "
    "gear and its own transport - is tracked on its own page (button "
    "49) and none of its dollars appear here. The two streams never "
    "share a figure.</div>")


def is_sep_invoice(item):
    bc = str(item.get("barcode") or "").strip().upper()
    return bc.startswith("SUB") or bc in SEP_BARCODES

SHUTDOWN_NAME = "Cement Australia K2 Shutdown 2026 - Gladstone"
COATES_ORANGE = "#F26222"   # sampled from the official Coates Way cog artwork
COATES_DARK = "#1D1D1B"

KIT_DIR = os.path.dirname(os.path.abspath(__file__))
# One folder per day - Reports\<YYYY-MM-DD>\ - so each day's company reports,
# demob packs and draft manifests live together (see report_paths.py).
import report_paths
import master_equipment
# The compliance line that sits UNDER every description - blue tag for
# electrical and rigging, daily pre-start and logbook for plant, return
# requirements for gas, radios and battery gear. Driven entirely by the
# four new columns in the master file: no data in the column, nothing in
# the report. (A. Fisher, 25 Jul 2026)
import equipment_compliance as EC
# One safety conversation a day, picked off the report date so the store,
# the crews and the client are all talking about the same thing.
import safety_conversation
import logbook_poster
# The two-level Equipment Activity & Accountability report: company
# position on page one, then one page per hirer as the evidence behind
# it. Company totals are BUILT as the sum of the hirer pages, so they
# cannot drift. (A. Fisher, 25 Jul 2026)
import activity_model
import activity_report
# The report pasted into the email body as images, in order - same look
# as the Cost Tracking Snapshot email (see email_images.py).
import email_images
# The page frame and the executive light theme switch live together in
# report_pages - INK_STRONG is the emphasis colour that inline table
# cells use so they read on either theme.
import report_pages as _RP
INK_STRONG = "#14181F" if getattr(_RP, "EXEC_LIGHT", False) else "#fff"
# Every report email pre-addressed from Coates_Report_Recipients.xlsx
# (see recipients.py) - fill the book in once, every email finds its people.
import recipients
_OUT = report_paths.out_dirs(KIT_DIR)   # organised day folder (A. Fisher)
OUTPUT_DIR = _OUT["day"]                # registers + DATA_SOURCES at the root
PAGES_DIR = _OUT["pages"]               # the reports as framed HTML pages
PDF_DIR = _OUT["pdf"]                   # print copies
EMAILS_DIR = _OUT["emails"]             # .eml drafts, page images, manifests

#  ONE DRAFT PER COMPANY PER MORNING.
#  Set True to bring back the old flat On-Hire report's own Outlook
#  draft. It is False because the Activity & Accountability report is
#  what a contractor gets, and two drafts for the same company on the
#  same morning is the double-up Andrew reported on 5 Aug 2026. The
#  page and the PDF are still built either way - only the draft stops.
COMPANY_ONHIRE_DRAFTS = False


class KitError(Exception):
    """A problem we explain in plain English - never a raw traceback."""


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def safe_print(s):
    """Windows consoles can be cp1252 - never let a fancy character in a
    company name crash an unattended run."""
    try:
        print(s)
    except UnicodeEncodeError:
        print(s.encode("ascii", errors="replace").decode("ascii"))


def clean_text(v):
    """Trim a cell value to clean text ('' for blank/None)."""
    if v is None:
        return ""
    return str(v).strip()


def to_num(v):
    """Parse a number from a cell that may be text like '$1,234.50'.
    Returns None when the cell is blank or not a number - we never invent
    a zero for a missing figure (a confident wrong report is worse than a gap)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_date_au(v):
    """Parse an Australian day-first date. SiteIQ exports these as TEXT like
    '16/07/2026' - parsing them as US month-first would give a plausible,
    WRONG report, so we parse explicitly and never rely on a default."""
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%Y-%m-%d",
                "%d/%m/%Y %I:%M %p", "%d/%m/%Y %H:%M"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def fmt_date(d):
    """Dates print as '16 Jul 2026' (Coates standard)."""
    return d.strftime("%d %b %Y") if d else "-"


def fmt_money(v):
    return "-" if v is None else "${:,.2f}".format(v)


def esc(s):
    return html_mod.escape(str(s), quote=True)


def customer_owned(it):
    """Is this the customer's own gear sitting in our store?

    The OWNER column says who bought it. Anything that isn't Coates is
    somebody else's property - Cement's bollards and Diphoterine, UGL's
    ratchet, Ampol's lead stands. We book it in and out and count it at
    stocktake exactly like everything else, but it is not our money.

    Blank owner means Coates. The export writes the owner out three
    different ways ("Coates", "COATES", "COATES HIRE"), so this tests for
    the word rather than an exact match."""
    o = (it.get("owner") or "").strip().upper() if isinstance(it, dict) else \
        str(it or "").strip().upper()
    return bool(o) and "COATES" not in o


def repl_display(r):
    """What goes in the replacement-cost column.

    Three different things, and they must not look alike. A priced item
    shows the money. An item we own but haven't priced yet shows TBC -
    a promise to fill it in. The customer's own gear shows CUSTOMER
    OWNED, because there is nothing to fill in and TBC would read as a
    figure still to come."""
    if r.get("cust_owned"):
        return ("<span style='color:#6B7280;font-weight:700;"
                "letter-spacing:.4px'>CUSTOMER OWNED</span>")
    if r.get("repl") is not None:
        return "<span class='repl'>{}</span>".format(fmt_money(r["repl"]))
    return "<span style='color:#8B9099'>TBC</span>"


def canonical_company(raw):
    """Upper-trim then apply the alias table."""
    key = clean_text(raw).upper()
    return COMPANY_ALIASES.get(key, key)


def display_company(canonical):
    """Display name per Andrew's standard: the register spelling if known;
    otherwise Capital first letter of EVERY word, acronyms kept upper."""
    if canonical in KNOWN_COMPANIES:
        return KNOWN_COMPANIES[canonical]
    words = []
    for w in canonical.split():
        words.append(w if w in DISPLAY_ACRONYMS else w.capitalize())
    return " ".join(words)


def is_known_company(canonical):
    return canonical in KNOWN_COMPANIES


def safe_filename(name):
    s = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_")
    return s or "COMPANY"


# ---------------------------------------------------------------------------
# Source file discovery - newest match wins, case-insensitive, and spaces /
# underscores in filenames are treated as the same thing (the rates file has
# shipped under both spellings).
# ---------------------------------------------------------------------------

def _norm_name(name):
    return re.sub(r"[ _]+", "", name.lower())


def find_newest(pattern, what, required=True):
    """Find the newest file in the kit folder whose normalised name matches
    the normalised pattern. Skips Excel lock files (~$...)."""
    norm_pat = _norm_name(pattern)
    hits = []
    for d in (os.path.join(KIT_DIR, "Data_SiteIQ"), KIT_DIR):
        if not os.path.isdir(d):
            continue
        for fn in os.listdir(d):
            if fn.startswith("~$"):
                continue
            full = os.path.join(d, fn)
            if not os.path.isfile(full):
                continue
            if fnmatch.fnmatch(_norm_name(fn), norm_pat):
                hits.append(full)
    if not hits:
        if required:
            raise KitError(
                "Couldn't find {w}.\n"
                "  Looked for files matching: {p}\n"
                "  In folder: {d}\n"
                "Drop the SiteIQ export in that folder and run again.".format(
                    w=what, p=pattern, d=KIT_DIR))
        return None
    return max(hits, key=os.path.getmtime)


def sheet_rows(path, sheet_name, what):
    """Open a worksheet read-only and return (headers, row dicts)."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise KitError(
            "Couldn't open {w}:\n  {p}\n  ({e})\n"
            "If the file is open in Excel, close it and run again.".format(
                w=what, p=path, e=e))
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise KitError(
            "{w} ({p}) has no sheet called '{s}'.\n"
            "Sheets found: {have}".format(
                w=what, p=os.path.basename(path), s=sheet_name,
                have=", ".join(wb.sheetnames)))
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [clean_text(h) for h in next(rows)]
    except StopIteration:
        wb.close()
        raise KitError("{w} sheet '{s}' is empty.".format(w=what, s=sheet_name))
    out = []
    for r in rows:
        d = {}
        for i, h in enumerate(headers):
            if h:
                d[h] = r[i] if i < len(r) else None
        out.append(d)
    wb.close()
    return headers, out


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def load_rental():
    path = find_newest("RENTAL_STOCK*.xlsx", "the RENTAL_STOCK export")
    headers, raw = sheet_rows(path, "RENTAL_STOCK", "the RENTAL_STOCK export")

    # NOTE: the export's variant header carries a trailing space
    # ('PRODUCT_VARIANT ') - find it defensively.
    variant_col = None
    for h in headers:
        if h.strip().upper() == "PRODUCT_VARIANT":
            variant_col = h
            break

    items = []
    plant = []
    infra = []
    radio_fleet = 0
    gas_fleet = 0
    radio_on_hire = 0
    gas_on_hire = 0
    milw_batt_fleet = 0
    milw_tool_fleet = 0
    for r in raw:
        su_upper = clean_text(r.get("STORAGE_UNIT")).upper()
        # Site infrastructure (barriers/chutes/hoppers): on site for the
        # duration - reported by name and qty only, never as idle plant.
        if su_upper in INFRA_STORAGE_UNITS:
            infra.append(MASTER.disp(clean_text(r.get("ITEM_NUMBER")),
                                     clean_text(r.get("ITEM_DESCRIPTION"))))
        # Site plant fleet: EVERY row in the plant storage unit, hired out
        # or not - the Site Plant report needs the whole picture.
        if su_upper == PLANT_STORAGE_UNIT:
            plant.append({
                # Display name from the MASTER file (item-number keyed);
                # the export's original rides along as desc0 for matching.
                "description": MASTER.disp(clean_text(r.get("ITEM_NUMBER")),
                                           clean_text(r.get("ITEM_DESCRIPTION"))),
                "desc0": clean_text(r.get("ITEM_DESCRIPTION")),
                # ITEM_NUMBER is the asset IDENTIFIER people key on (per
                # Andrew, 22 Jul - supersedes SKU_NUMBER). ITEM_BARCODE is
                # kept ONLY as an internal join key for transaction/charge
                # attribution - it is never displayed on any report.
                "asset": clean_text(r.get("ITEM_NUMBER")),
                "barcode": clean_text(r.get("ITEM_BARCODE")),
                "hirer": clean_text(r.get("HIRER_NAME")),
                "company_raw": clean_text(r.get("COMPANY_NAME")),
                "status": clean_text(r.get("ITEM_STATUS")),
                "on_hire_date": parse_date_au(r.get("ON_HIRE_DATE")),
                "family": clean_text(r.get("PRODUCT_FAMILY")),
                "product": clean_text(r.get("PRODUCT")),
                # variant header carries a trailing space in the export
                "variant": clean_text(r.get(variant_col)).upper() if variant_col else "",
                #  Who actually owns the thing. The export has carried this
                #  all along and the loader used to drop it, so every
                #  report treated a customer's own gear as Coates stock.
                #  (Andrew, 28 Jul 2026, on Cement's 24 bollards: "no
                #  replacement costs on these.")
                "owner": clean_text(r.get("OWNER")),
            })
        # Radio fleet count includes every radio HANDSET in the export, on
        # hire or not - handsets charge $12.81/day regardless (see
        # RADIO_CHARGE_START). Batteries and covers in the same storage
        # unit do not count.
        if is_radio_handset({
                "storage_unit": clean_text(r.get("STORAGE_UNIT")),
                "description": clean_text(r.get("ITEM_DESCRIPTION"))}):
            radio_fleet += 1
        # Gas monitor fleet: every unit in the GAS MONITORS storage unit
        # bills at GAS_RATE/day from GAS_START, on hire or not.
        if su_upper == "GAS MONITORS":
            gas_fleet += 1
        # Milwaukee fleet (Daily Hit List): every Milwaukee line in the
        # export, on hire or not - batteries counted separately from the
        # battery-powered tools they drive.
        _d_up = clean_text(r.get("ITEM_DESCRIPTION")).upper()
        if "MILWAUKEE" in _d_up:
            if "BATTER" in _d_up:
                milw_batt_fleet += 1
            else:
                milw_tool_fleet += 1
        company_raw = clean_text(r.get("COMPANY_NAME"))
        hirer_raw = clean_text(r.get("HIRER_NAME"))
        # ON HIRE definition (per the K2 workbook): COMPANY_NAME is non-blank.
        # We deliberately do NOT filter on ITEM_STATUS here - that field has
        # previously included returned items and inflated on-hire counts.
        if not company_raw:
            continue
        # Idle pool (Andrew, 21 Jul 2026): gear "on hire" to hirer "Site Plant
        # Equipment" (Coates) is internal plant sitting available and still on
        # charge - NOT genuine on-hire to an outside crew. It belongs to the
        # Site Plant report only (which reads the separate 'plant' list above),
        # and must never enter the company on-hire models, the item counts, or
        # the "Working hardest" ranking - where 130-odd identical idle units
        # would otherwise swamp the list and read as the hardest-worked gear.
        if is_site_plant_equipment(hirer_raw):
            continue
        # Genuinely with a hirer right now (not the idle pool) - the "how
        # many of the fleet are actually out" count for the exec brief.
        if is_radio_handset({
                "storage_unit": clean_text(r.get("STORAGE_UNIT")),
                "description": clean_text(r.get("ITEM_DESCRIPTION"))}):
            radio_on_hire += 1
        if su_upper == "GAS MONITORS":
            gas_on_hire += 1
        items.append({
            "company": canonical_company(company_raw),
            "hirer": clean_text(r.get("HIRER_NAME")),
            # Display name from the MASTER file (item-number keyed); the
            # export's original rides along as desc0 - billing counts and
            # pattern rules stay anchored to SiteIQ's own words.
            "description": MASTER.disp(clean_text(r.get("ITEM_NUMBER")),
                                       clean_text(r.get("ITEM_DESCRIPTION"))),
            "desc0": clean_text(r.get("ITEM_DESCRIPTION")),
            # Displayed asset identifier = ITEM_NUMBER (Andrew, 22 Jul).
            # barcode kept internal-only for charge attribution, never shown.
            "asset": clean_text(r.get("ITEM_NUMBER")),
            "barcode": clean_text(r.get("ITEM_BARCODE")),
            "status": clean_text(r.get("ITEM_STATUS")),
            "storage_unit": clean_text(r.get("STORAGE_UNIT")),
            "on_hire_date": parse_date_au(r.get("ON_HIRE_DATE")),
            #  The rest of what the export tells us about this asset. The
            #  activity pack shows the lot - if SiteIQ knows it, the
            #  person holding the gear can see it. (A. Fisher, 25 Jul 2026)
            "on_hire_time": clean_text(r.get("ON_HIRE_TIME")),
            "family": clean_text(r.get("PRODUCT_FAMILY")),
            "product": clean_text(r.get("PRODUCT")),
            "toolstore": clean_text(r.get("TOOLSTORE")),
            #  Who owns it. The export has carried OWNER all along and the
            #  loader dropped it here, so every report counted a
            #  customer's own gear as Coates stock. (Andrew, 28 Jul 2026,
            #  on Cement's 24 bollards: "no replacement costs on these.")
            "owner": clean_text(r.get("OWNER")),
            "variant": clean_text(r.get(variant_col)).upper() if variant_col else "",
        })
    mtime = dt.datetime.fromtimestamp(os.path.getmtime(path))
    return (items, path, mtime, radio_fleet, plant, infra, gas_fleet,
            radio_on_hire, gas_on_hire, milw_batt_fleet, milw_tool_fleet)


# Populated once in main() from the ON_HIRE export - EXTERNAL_ID by hirer
# name (upper). Empty when the export isn't present, so every consumer
# falls back to an honest '-'.
HIRER_ID_BY_NAME = {}


def hirer_id_for(name):
    """The person's EXTERNAL_ID for a hirer name - '' when the export is
    absent, the name is unknown, or two people share the name (never guess)."""
    return HIRER_ID_BY_NAME.get(clean_text(name).upper(), "")


def load_on_hire():
    """Per-person export (ON_HIRE*.xlsx): every row is one item on hire to
    one PERSON, and EXTERNAL_ID is the number on that person's ID card.
    Per Andrew (19 Jul): EXTERNAL_ID is the ONLY person identifier we ever
    use - never HIRER_ID, never a hand-typed list. Returns None when the
    export isn't in the kit folder (reports then show '-' for Hirer ID)."""
    path = find_newest("ON_HIRE*.xlsx", "the ON_HIRE (per-person) export",
                       required=False)
    if not path:
        return None
    _, raw = sheet_rows(path, "ON_HIRE", "the ON_HIRE export")
    people = {}          # EXTERNAL_ID -> {"name", "company_raw", "items"}
    by_barcode = {}      # ITEM_BARCODE (upper) -> EXTERNAL_ID
    by_name = {}         # HIRER_NAME (upper) -> EXTERNAL_ID ('' if ambiguous)
    skipped = 0
    for r in raw:
        idno = clean_text(r.get("EXTERNAL_ID"))
        # numeric cells can come back as '10009866.0'
        if idno.endswith(".0") and idno[:-2].isdigit():
            idno = idno[:-2]
        name = clean_text(r.get("HIRER_NAME"))
        if not idno:
            skipped += 1        # a person with no card number - never guess
            continue
        rec = people.setdefault(idno, {
            "name": name,
            "company_raw": clean_text(r.get("COMPANY")),
            "items": []})
        bc = clean_text(r.get("ITEM_BARCODE")).upper()
        rec["items"].append({
            "item_number": clean_text(r.get("ITEM_NUMBER")),
            "barcode": bc,
            "description": clean_text(r.get("ITEM_DESCRIPTION")),
            "on_hire_date": parse_date_au(r.get("START_DATE")),
        })
        if bc:
            by_barcode[bc] = idno
        if name:
            key = name.upper()
            if key in by_name and by_name[key] != idno:
                by_name[key] = ""   # two people, one name - never guess
            else:
                by_name[key] = idno
    # As-at: SiteIQ stamps the request time on REFERENCE_INFO; the file
    # time is the fallback.
    asof = dt.datetime.fromtimestamp(os.path.getmtime(path))
    try:
        _, ref = sheet_rows(path, "REFERENCE_INFO", "the ON_HIRE export")
        for rr in ref:
            for h, v in rr.items():
                if "REQUESTED_DATE" in h.upper() and clean_text(v):
                    try:
                        asof = dt.datetime.strptime(
                            clean_text(v), "%d/%m/%Y %I:%M %p")
                    except ValueError:
                        pass
    except KitError:
        pass
    if skipped:
        safe_print("  NOTE: {} ON_HIRE row(s) had no EXTERNAL_ID - those "
                   "items show '-' for Hirer ID.".format(skipped))
    ambiguous = sorted(k for k, v in by_name.items() if v == "")
    if ambiguous:
        safe_print("  NOTE: one name, two EXTERNAL_IDs: "
                   + ", ".join(ambiguous)
                   + " - name-only matches show '-' rather than guess.")
    return {"path": path, "asof": asof, "people": people,
            "by_barcode": by_barcode, "by_name": by_name}


def load_rates():
    path = find_newest("ImportSample_Contracted Rates and Prices*.xlsx",
                       "the Contracted Rates and Prices file")
    _, raw = sheet_rows(path, "Coates Equipment", "the Contracted Rates file")
    rates = {}
    for r in raw:
        variant = clean_text(r.get("Product Variant ID")).upper()
        if not variant or variant in rates:
            continue
        rate = to_num(r.get("Base Rate"))
        # Blank or zero Base Rate = unpriced. Shown as '-' on the report,
        # never $0, never estimated, excluded from value totals.
        rates[variant] = rate if (rate is not None and rate > 0) else None
    return rates, path


# The master equipment file (K2_MASTER_EQUIPMENT_PRICING.xlsx) - loaded
# once in main(); item-number-keyed display names + replacement prices.
# Empty-safe: without the file, everything behaves exactly as before.
MASTER = master_equipment.Master()


class ReplacementIndex(dict):
    """The replacement schedule with an honest second chance: an item that
    misses on the exact description is retried as a TOKEN SET (same words,
    any order/punctuation - 'Bow - Shackle - Alloy - 3.25t' matches
    'BOW SHACKLE - ALLOY 3.25T'). Every word must match, so a different
    size or model can never borrow another item's cost; ambiguous token
    sets (two schedule rows, two costs) are never used. Items genuinely
    absent from the schedule show TBC - the gap list names them
    so the schedule can be completed with real prices."""

    def _tokens(self, s):
        return frozenset(w for w in re.split(r"[^A-Z0-9./]+", str(s).upper()) if w)

    def build_token_index(self):
        self._tok = {}
        drop = set()
        for k, v in self.items():
            t = self._tokens(k)
            if t in self._tok and self._tok[t] != v:
                drop.add(t)
            else:
                self._tok[t] = v
        for t in drop:
            self._tok.pop(t, None)

    def get(self, key, default=None):
        hit = dict.get(self, key)
        if hit is not None:
            return hit
        tok = getattr(self, "_tok", None)
        if tok is None:
            return default
        return tok.get(self._tokens(key), default)


def load_replacement():
    """Replacement cost per item, joined on the rental stock DESCRIPTION
    (the schedule's 'Rental Stock Description' column matches RENTAL_STOCK's
    ITEM_DESCRIPTION - that is the clean join key in this file)."""
    path = find_newest("Tooling_Replacement_Costs_Gladstone_K2*.xlsx",
                       "the Tooling Replacement Costs schedule", required=False)
    if not path:
        return {}, None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = None
        costs = {}
        for row in ws.iter_rows(values_only=True):
            vals = [clean_text(c) for c in row]
            if headers is None:
                # header row is a few rows down in this schedule - find it
                if "Rental Stock Description" in vals:
                    headers = vals
                continue
            d = dict(zip(headers, row))
            desc = clean_text(d.get("Rental Stock Description")).upper()
            cost = to_num(d.get("Unit Replacement Cost (AUD)"))
            if desc and cost is not None and cost > 0 and desc not in costs:
                costs[desc] = cost
        wb.close()
        if headers is None:
            print("NOTE: replacement cost schedule found but its layout wasn't")
            print("      recognised - replacement costs omitted from this run.")
            return {}, None
        idx = ReplacementIndex(costs)
        idx.build_token_index()
        return idx, path
    except KitError:
        raise
    except Exception as e:
        print("NOTE: couldn't read the replacement cost schedule ({}).".format(e))
        print("      Replacement costs omitted from this run.")
        return {}, None


def load_transactions():
    """Transaction history (optional): who took what through the tool store
    and whether it came back same day. Company attribution follows the K2
    workbook: LATEST_BARCODE joined to the on-hire company in RENTAL_STOCK."""
    path = find_newest("TRANSACTIONS*.xlsx", "the TRANSACTIONS export",
                       required=False)
    if not path:
        return [], None
    tx = []
    any_sheet_ok = False
    for sheet in ("TRANSACTION_CHARGES", "CUSTOMER_CONTRACTOR_EQUIP"):
        try:
            _, raw = sheet_rows(path, sheet, "the TRANSACTIONS export")
        except KitError:
            continue
        any_sheet_ok = True
        for r in raw:
            desc = clean_text(r.get("SKU/ITEM DESCRIPTION"))
            hirer = clean_text(r.get("HIRER_NAME"))
            barcode = clean_text(r.get("LATEST_BARCODE")).upper()
            # EMPLOYER_NAME is the company that RAN the transaction - the
            # authoritative attribution. The old barcode->current-on-hire
            # join dropped every returned item's history and could credit
            # a re-hired item's past to the wrong company.
            employer = clean_text(r.get("EMPLOYER_NAME"))
            start = parse_date_au(r.get("TRAN_START_DATE"))
            end = parse_date_au(r.get("TRAN_END_DATE"))
            if not (hirer or barcode or desc):
                continue
            tx.append({"hirer": hirer, "barcode": barcode, "description": desc,
                       "company": canonical_company(employer) if employer else "",
                       "start": start, "end": end,
                       # times ride along for the peak-hours story (v3 skin)
                       "start_time": clean_text(r.get("TRAN_START_TIME")),
                       "end_time": clean_text(r.get("TRAN_END_TIME")),
                       "tx_id": clean_text(r.get("TRANSACTION_ID"))})
    if not any_sheet_ok:
        # Present but unreadable = data-feed failure, not a genuine nil.
        print("NOTE: TRANSACTIONS export present but unreadable - "
              "transaction story reported as unavailable, not zero.")
        return [], None
    return tx, path


def load_stocktake():
    """Stocktake coverage: when was each item last physically sighted."""
    path = find_newest("STOCKTAKE*.xlsx", "the STOCKTAKE export", required=False)
    if not path:
        return None, None
    try:
        _, raw = sheet_rows(path, "STOCKTAKE", "the STOCKTAKE export")
    except KitError:
        print("NOTE: STOCKTAKE export present but unreadable - coverage "
              "reported as unavailable, not zero.")
        return None, None
    total = 0
    sighted3 = 0
    counters = set()
    entries = []
    for r in raw:
        su = clean_text(r.get("STORAGE_UNIT"))
        if not su or su.upper() in ("ON HIRE", "SHUTDOWN", "TURNAROUND",
                            # on site for the duration - not stocktake
                            # units (per Andrew, 19 Jul)
                            "CEMENT SITE PLANT", "CEMENT BARRIERS",
                            "CEMENT RUBBISH CHUTES", "LAYDOWN"):
            continue
        total += 1
        d = parse_date_au(r.get("LAST_SIGHTED_DATE_TIME"))
        entries.append(d)
        by = clean_text(r.get("LAST_SIGHTED_BY"))
        if by:
            counters.add(by)
    # Anchor the 3-day window to the EXPORT's own as-at date (its newest
    # sighting), not the run date - re-running an old export must not
    # collapse the coverage figure to zero.
    dates = [d for d in entries if d]
    anchor = max(dates) if dates else dt.date.today()
    for d in entries:
        if d and (anchor - d).days <= 3:
            sighted3 += 1
    return {"total": total, "sighted3": sighted3,
            "pct": (sighted3 / total) if total else 0,
            "counters": len(counters)}, path


def load_daily_tracking_summary(shift_day):
    """Read Forecast v Actual from the K2 workbook's Daily Tracking
    (read-only - the kit never writes to the workbook)."""
    path = find_newest("Cement_Australia_Report*K2*.xlsm",
                       "the K2 report workbook", required=False)
    if not path:
        return None, None
    cats = [("Plant Equipment", "Plant Equipment Hire F/C", "Plant Equipment Hire Actual"),
            ("Tooling", "Tooling Hire F/C", "Tooling Hire Actual"),
            ("Radios", "Radio Hire F/C", "Radio Hire Actual"),
            ("Gas Monitors", "Gas Monitor Hire F/C", "Gas Monitor Hire Actual"),
            ("Personnel / Labour", "Personnel / Labour F/C", "Personnel / Labour Actual"),
            ("Accommodation", "Accommodation F/C", "Accommodation Actual"),
            ("Transport", "Transport F/C", "Transport Actual"),
            ("Damage Recovery", "Damage Recovery F/C", "Damage Recovery Actual")]
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Daily Tracking"]
        rows = ws.iter_rows(values_only=True)
        hdr = [clean_text(h) for h in next(rows)]
        ix = {h: i for i, h in enumerate(hdr)}
        today_row = None
        fc_to_date = 0.0
        act_to_date = 0.0
        fc_full = 0.0
        for r in rows:
            d = r[ix["Date"]]
            if isinstance(d, dt.datetime):
                d = d.date()
            elif not isinstance(d, dt.date):
                d = parse_date_au(d)
            if d is None:
                continue
            fc = to_num(r[ix["Daily Forecast Total"]]) or 0.0
            act = to_num(r[ix["Daily Actual Total"]]) or 0.0
            fc_full += fc
            if d <= shift_day:
                fc_to_date += fc
                act_to_date += act
            if d == shift_day:
                today_row = [(name,
                              to_num(r[ix[fch]]),
                              to_num(r[ix[ach]]))
                             for name, fch, ach in cats]
        wb.close()
        return {"today": today_row, "fc_to_date": fc_to_date,
                "act_to_date": act_to_date, "fc_full": fc_full}, path
    except Exception as e:
        print("NOTE: couldn't read Daily Tracking from the workbook ({}).".format(e))
        return None, None


# Categories where SiteIQ's real invoice (DAILY_SUMMARY) can be compared
# against our own Actual - confirmed against the Contracted Rates card:
# radios, gas monitors and subhire have no rate entry there, so they
# structurally never invoice through this channel and are never compared.
DAILY_SUMMARY_CATEGORIES = {
    "HIRE": "Hire (Plant + Tooling)",
    "LABOUR": "Labour",
    "TRANSPORT": "Transport",
    "DAMAGE": "Damage Recovery",
}


def load_daily_summary(shift_day):
    """Read SiteIQ's real invoiced totals for shift_day from DAILY_SUMMARY,
    for the categories that can legitimately be compared (see
    DAILY_SUMMARY_CATEGORIES). Rolling-window export, read-only. Returns
    None if the export isn't present or has no row for shift_day."""
    path = find_newest("DAILY_SUMMARY*.xlsx", "the DAILY_SUMMARY export",
                       required=False)
    if not path:
        return None, None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["DAILY_SUMMARY"]
        rows = ws.iter_rows(values_only=True)
        hdr = [clean_text(h) for h in next(rows)]
        ix = {h: i for i, h in enumerate(hdr)}
        wanted_date = shift_day.strftime("%d/%m/%Y")
        result = None
        for r in rows:
            label = clean_text(r[0])
            m = re.match(r"^TOTAL FOR:\s*(\d{2}/\d{2}/\d{4})", label)
            if not m or m.group(1) != wanted_date:
                continue
            result = {}
            for col, disp in DAILY_SUMMARY_CATEGORIES.items():
                if col in ix:
                    result[col] = to_num(r[ix[col]]) or 0.0
            break
        wb.close()
        return result, path
    except Exception as e:
        print("NOTE: couldn't read DAILY_SUMMARY ({}).".format(e))
        return None, None


def load_gear_dispatch_summary():
    """Read the SiteIQ TRANSACTIONS export for radio handset and gas
    monitor dispatch history: how many distinct units have gone out in
    this rolling window, how many of those are still out (open - no
    TRAN_END_DATE on their latest transaction), how many came back same
    trip. Rolling window, not full shutdown-to-date. Returns None if the
    export isn't present."""
    path = find_newest("TRANSACTIONS*.xlsx", "the TRANSACTIONS export",
                       required=False)
    if not path:
        return None, None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ref_ws = wb["REFERENCE_INFO"]
        period_row = next(ref_ws.iter_rows(min_row=2, max_row=2, values_only=True))
        period_text = clean_text(period_row[0])
        parts = period_text.split(" - ")
        rep_start = dt.datetime.strptime(parts[0].strip(), "%d/%m/%Y %H:%M:%S").date()
        rep_end = dt.datetime.strptime(parts[1].strip(), "%d/%m/%Y %H:%M:%S").date()

        # barcode -> {"dispatches": n, "open": bool}
        radio_units = {}
        gas_units = {}
        for sheet_name in ("TRANSACTION_CHARGES", "CUSTOMER_CONTRACTOR_EQUIP"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            hdr = [clean_text(h) for h in next(rows)]
            ix = {h: i for i, h in enumerate(hdr)}
            desc_i = ix.get("SKU/ITEM DESCRIPTION")
            bar_i = ix.get("LATEST_BARCODE")
            end_i = ix.get("TRAN_END_DATE")
            tid_i = ix.get("TRANSACTION_ID")
            if desc_i is None or bar_i is None:
                continue
            for r in rows:
                desc = clean_text(r[desc_i])
                bar = clean_text(r[bar_i]).upper()
                if not bar:
                    continue
                is_open = end_i is None or r[end_i] in (None, "")
                target = None
                if is_radio_handset({"storage_unit": "RADIOS", "description": desc}):
                    target = radio_units
                elif is_gas_monitor_desc(desc):
                    target = gas_units
                if target is None:
                    continue
                rec = target.setdefault(bar, {"dispatches": 0, "open": False})
                rec["dispatches"] += 1
                if is_open:
                    rec["open"] = True
        wb.close()

        def summarize(units):
            dispatched = len(units)
            still_out = sum(1 for v in units.values() if v["open"])
            returned = dispatched - still_out
            return {"dispatched": dispatched, "still_out": still_out,
                    "returned": returned}

        return {"period_start": rep_start, "period_end": rep_end,
                "radios": summarize(radio_units),
                "gas_monitors": summarize(gas_units)}, path
    except Exception as e:
        print("NOTE: couldn't read TRANSACTIONS for gear dispatch history ({}).".format(e))
        return None, None


def load_plant_categories():
    """Read Andrew's own plant categories (Misc, Welders, Booms, Forklifts...)
    from the K2 workbook's Plant Equipment register - READ ONLY, the kit
    never writes to the workbook. Returns {variant: {category, rate}}."""
    path = find_newest("Cement_Australia_Report*K2*.xlsm",
                       "the K2 report workbook", required=False)
    if not path:
        return {}, None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Plant Equipment" not in wb.sheetnames:
            wb.close()
            return {}, None
        ws = wb["Plant Equipment"]
        rows = ws.iter_rows(values_only=True)
        hdr = [clean_text(h) for h in next(rows)]
        if "Category" not in hdr or "Product Variant" not in hdr:
            wb.close()
            return {}, None
        i_cat = hdr.index("Category")
        i_var = hdr.index("Product Variant")
        i_rate = hdr.index("Daily Rate") if "Daily Rate" in hdr else None
        cats = {}
        for r in rows:
            var = clean_text(r[i_var] if i_var < len(r) else None).upper()
            cat = clean_text(r[i_cat] if i_cat < len(r) else None)
            if var and cat and var not in cats:
                rate = to_num(r[i_rate]) if (i_rate is not None and i_rate < len(r)) else None
                cats[var] = {"category": cat,
                             "rate": rate if (rate is not None and rate > 0) else None}
        wb.close()
        return cats, path
    except Exception as e:
        print("NOTE: couldn't read plant categories from the K2 workbook ({}).".format(e))
        print("      Plant will be grouped as 'Not in Plant Register'.")
        return {}, None


def plant_category_info(variant, cats):
    """Exact variant match first; the register truncates some variant codes,
    so fall back to a prefix match (8+ chars) either direction."""
    if variant in cats:
        return cats[variant]
    for v, info in cats.items():
        if len(v) >= 8 and (variant.startswith(v) or v.startswith(variant)):
            return info
    return None


def load_sales():
    path = find_newest("SALES_STOCK*.xlsx", "the SALES_STOCK export",
                       required=False)
    if not path:
        return [], None
    try:
        _, raw = sheet_rows(path, "SALES_STOCK", "the SALES_STOCK export")
    except KitError:
        print("NOTE: SALES_STOCK export present but unreadable - consumables "
              "reported as unavailable, not zero.")
        return [], None
    moves = []
    for r in raw:
        # PERMANENT EXCLUSION (per Andrew): blocklisted dead SKUs are never
        # counted or shown anywhere, on any report, ever - by SKU number.
        if clean_text(r.get("SKU_NUMBER")) in DEAD_SKUS:
            continue
        # Belt and braces: all-zero placeholder rows with no sales date are
        # also excluded, whatever their SKU number.
        if (not clean_text(r.get("SALES_DATE"))
                and all((to_num(r.get(c)) or 0) == 0 for c in (
                    "AVAILABLE_QUANTITY", "MINIMUM_QUANTITY",
                    "REORDER_QUANTITY", "SALES_QUANTITY",
                    "RETURN_QUANTITY"))):
            continue
        company_raw = clean_text(r.get("COMPANY_NAME"))
        qty = to_num(r.get("SALES_QUANTITY"))
        # Movement rows = company attributed AND quantity actually taken
        if not company_raw or qty is None or qty <= 0:
            continue
        moves.append({
            "company": canonical_company(company_raw),
            "hirer": clean_text(r.get("HIRER")),
            "sku": clean_text(r.get("SKU_DESCRIPTION")),
            "qty": qty,
            "date": parse_date_au(r.get("SALES_DATE")),
            "time": clean_text(r.get("SALES_TIME")),
        })
    return moves, path


# ---------------------------------------------------------------------------
# Per-company model
# ---------------------------------------------------------------------------

def build_company_model(company, items, rates, repl, sales, today,
                        sales_loaded=True, tx=None, tx_loaded=False,
                        barcode_company=None):
    rows = []
    for it in (i for i in items if i["company"] == company):
        days = None
        if it["on_hire_date"]:
            days = max(0, (today - it["on_hire_date"]).days)
        rate = rates.get(it["variant"]) if it["variant"] else None
        charge_days = (days + 1) if days is not None else None
        if is_radio_handset(it):
            # Handsets bill ONCE, at fleet level (70 x $12.81/day on the
            # internal summary, on hire or not). Never per company - that
            # would count the same handset-days twice in the same pack.
            rate = None
            charge_days = None
        if is_sep_invoice(it):
            # Sub-hired gear (dashed item number / SUB barcode) is
            # charged on the SEPARATE INVOICE - its dollars live on the
            # Separate Invoice Tracker (49), never in SiteIQ-stream
            # totals, or the same machine-day counts on both streams.
            # Coates welders/forklifts with plain numeric IDs keep
            # their SiteIQ rates - they ARE SiteIQ-charged. The gear
            # itself still tracks here for issue, return and stocktake.
            rate = None
            charge_days = None
        #  Customer-owned gear carries no replacement cost, and the reason
        #  is simply that it isn't ours to replace. Charging Cement a
        #  replacement figure for bollards they bought themselves - and
        #  that we invoiced them for last week - would be wrong on their
        #  own report. It still tracks for issue, return and stocktake.
        own = customer_owned(it)
        cost = None if own else MASTER.price(it.get("asset"))
        if cost is None and not own:
            cost = (repl.get((it.get("desc0") or it["description"]).upper())
                    if repl else None)
        rows.append({**it, "days": days, "rate": rate,
                     "charge_days": charge_days, "repl": cost,
                     "cust_owned": own})

    # Oldest first - this is the chase list
    rows.sort(key=lambda r: (-(r["days"] if r["days"] is not None else -1),
                             r["description"]))

    hirers = sorted({r["hirer"] for r in rows if r["hirer"]})
    days_known = [r["days"] for r in rows if r["days"] is not None]
    daily_value = sum(r["rate"] for r in rows if r["rate"] is not None)
    # Spend to date = rate x charged days, inclusive of the on-hire day
    # (day 0 bills as 1 day) - same basis as the K2 workbook and Andrew's
    # radio direction ("starts today" means today is a charged day).
    # These hire figures appear on the INTERNAL summary only.
    spend = sum(r["rate"] * r["charge_days"] for r in rows
                if r["rate"] is not None and r["charge_days"] is not None)
    repl_exposure = sum(r["repl"] for r in rows if r["repl"] is not None)
    # Handsets are deliberately rate-less here (fleet-level billing), and
    # separate-invoice gear (welders, forklifts) deliberately carries its
    # value on the other stream - neither is "unpriced" in the
    # missing-a-rate sense.
    unpriced = sum(1 for r in rows
                   if r["rate"] is None and not is_radio_handset(r)
                   and not is_sep_invoice(r))

    aging = {"0-2": 0, "3-4": 0, "5+": 0, "unknown": 0}
    for r in rows:
        if r["days"] is None:
            aging["unknown"] += 1
        elif r["days"] <= 2:
            aging["0-2"] += 1
        elif r["days"] <= 4:
            aging["3-4"] += 1
        else:
            aging["5+"] += 1

    # ---- transaction story: people, volume, return behaviour -------------
    my_tx = []
    if tx:
        # EMPLOYER_NAME first (authoritative, survives returns/re-hires);
        # barcode->current-company map only for rows with no employer.
        my_tx = [t for t in tx
                 if (t.get("company") == company
                     or (not t.get("company") and barcode_company and t["barcode"]
                         and barcode_company.get(t["barcode"]) == company))]
    same_day = sum(1 for t in my_tx
                   if t["start"] and t["end"] and t["end"] == t["start"])
    kept_longer = sum(1 for t in my_tx
                      if t["start"] and t["end"] and t["end"] > t["start"])
    open_tx = sum(1 for t in my_tx if t["start"] and not t["end"])
    tx_people = {t["hirer"] for t in my_tx if t["hirer"]}
    sales_people = {m["hirer"] for m in sales
                    if m["company"] == company and m["hirer"]}
    onhire_people = {r["hirer"] for r in rows if r["hirer"]}
    people = sorted(onhire_people | tx_people | sales_people)
    item_types = {r["description"] for r in rows if r["description"]}
    item_types |= {t["description"] for t in my_tx if t["description"]}

    # ---- high-care gear currently in this company's name -----------------
    def _is_gas(r):
        return r["storage_unit"].upper() == "GAS MONITORS"

    def _is_milwaukee(r):
        # Original description OR the master rename - Andrew branding an
        # item Milwaukee in the master pulls it into the care list; nothing
        # already Milwaukee can ever drop out.
        return ("MILWAUKEE" in r["description"].upper()
                or "MILWAUKEE" in r.get("desc0", "").upper())

    high_care = {
        "milwaukee": [r for r in rows if _is_milwaukee(r)],
        "radios": [r for r in rows if is_radio_handset(r)],
        "gas": [r for r in rows if _is_gas(r)],
    }

    cons = {}
    cons_draws = 0
    for m in (m for m in sales if m["company"] == company):
        # A draw is ONE transaction regardless of quantity taken - qty 100
        # is still one transaction (Andrew's rule; keeps the count honest).
        cons_draws += 1
        key = (m["sku"], m["hirer"])
        c = cons.setdefault(key, {"sku": m["sku"], "hirer": m["hirer"],
                                  "hirer_id": hirer_id_for(m["hirer"]),
                                  "qty": 0.0, "last": None})
        c["qty"] += m["qty"]
        if m["date"] and (c["last"] is None or m["date"] > c["last"]):
            c["last"] = m["date"]
    # Alphabetical by SKU - clean and tidy, the Coates standard
    consumables = sorted(cons.values(), key=lambda c: (c["sku"], c["hirer"]))

    return {
        "company": company,
        "display": display_company(company),
        "rows": rows,
        "hirers": hirers,
        "n_items": len(rows),
        "n_hirers": len(hirers),
        "oldest_days": max(days_known) if days_known else None,
        "daily_value": daily_value if any(r["rate"] is not None for r in rows) else None,
        "spend": spend if any(r["rate"] is not None and r["days"] is not None
                              for r in rows) else None,
        "repl_exposure": repl_exposure if any(r["repl"] is not None for r in rows) else None,
        "unpriced": unpriced,
        "aging": aging,
        "consumables": consumables,
        "cons_qty": sum(c["qty"] for c in consumables),
        "sales_loaded": sales_loaded,
        "tx_loaded": tx_loaded,
        "n_tx": len(my_tx),
        "cons_draws": cons_draws,
        "n_tx_total": len(my_tx) + cons_draws,
        "same_day": same_day,
        "kept_longer": kept_longer,
        "open_tx": open_tx,
        "people": people,
        "item_types": len(item_types),
        "high_care": high_care,
    }


# ---------------------------------------------------------------------------
# HTML rendering - branding lives in the template so it can't be forgotten
# ---------------------------------------------------------------------------

BASE_CSS = """
  /* Layout v2 - 24 Jul 2026: bigger type, lifted greys, roomier tiles and
     tables across EVERY report in the suite - same design, much clearer.
     Story cards (.intro) lead each report; scorecard tiles carry the
     figures; tables carry the detail. */
  * { box-sizing: border-box; margin: 0; padding: 0;
      -webkit-print-color-adjust: exact; print-color-adjust: exact; }
  html, body { background: #0E1116; }
  body { font-family: Calibri, 'Segoe UI', Arial, sans-serif; color: #E6E9EE;
         font-size: 11pt; line-height: 1.5; }
  .page { max-width: 200mm; margin: 0 auto; padding: 6mm 5mm; background: #0E1116; }
  .band { background: #171B22; color: #fff; padding: 16px 20px;
          border-bottom: 3px solid __ORANGE__; border-radius: 12px 12px 0 0; }
  .band .kicker { color: __ORANGE__; font-weight: bold; letter-spacing: 2px;
                  font-size: 9pt; text-transform: uppercase; }
  .band h1 { font-size: 20pt; margin: 2px 0 1px 0; color: #fff; }
  .band .sub { color: #c9cfd8; font-size: 10.5pt; }
  .band .meta { color: #A9B1BD; font-size: 9pt; margin-top: 6px; }
  .band .meta b { color: #fff; }
  .siteiq { float: right; text-align: right; color: #fff; font-weight: bold;
            font-size: 9pt; letter-spacing: 1px; padding-top: 4px; }
  .siteiq .o { color: __ORANGE__; }
  h2 { font-size: 13pt; color: #fff; border-bottom: 2px solid #2A313C;
       padding-bottom: 4px; margin: 20px 0 10px 0; }
  table.tiles { width: 100%; border-collapse: separate; border-spacing: 8px 0;
                table-layout: fixed; margin-top: 12px; }
  table.tiles td { background: #171B22; color: #fff; text-align: center;
                   padding: 14px 6px; border-radius: 10px; border: 1px solid #2A313C;
                   border-top: 3px solid __ORANGE__; }
  /* Status tints (the look Andrew picked, 28 Jul 2026): a tile only
     wears colour when its number IS a judgement - red action, amber
     watch, green healthy. Plain facts stay plain; the old cycle of
     colours-by-position was decoration and is gone. */
  table.tiles td.tg { background:#12271A; border-color:#3FB950; border-top:3px solid #3FB950; }
  table.tiles td.tg .v { color:#3FB950; }
  table.tiles td.ta { background:#2A2210; border-color:#fab219; border-top:3px solid #fab219; }
  table.tiles td.ta .v { color:#fab219; }
  table.tiles td.tr { background:#2A1512; border-color:#F85149; border-top:3px solid #F85149; }
  table.tiles td.tr .v { color:#F85149; }
  table.tiles .v { font-size: 17pt; font-weight: bold; display: block; color: #fff; }
  .hb { display: flex; align-items: center; gap: 10px; margin: 7px 0; }
  .hb .hl { width: 225px; color: #A9B1BD; font-size: 12px; }
  .hb .ht { flex: 1; background: #20262e; border-radius: 6px; overflow: hidden; }
  .hb .hf { height: 13px; border-radius: 6px;
            background: linear-gradient(90deg, __ORANGE__, #EFFF3D); }
  .hb .hv { width: 120px; text-align: right; color: #fff; font-size: 12px; font-weight: 600; }
  table.tiles .l { font-size: 8.5pt; color: #A9B1BD; text-transform: uppercase;
                   letter-spacing: 0.5px; display: block; margin-top: 4px; }
  table.data { width: 100%; border-collapse: collapse; margin-top: 4px; }
  table.data th { background: #20262F; color: #fff; text-align: left;
                  padding: 7px 8px; font-size: 9pt; text-transform: uppercase;
                  letter-spacing: 0.4px; }
  table.data td { padding: 7px 8px; border-bottom: 1px solid #232A34;
                  font-size: 10.5pt; vertical-align: top; color: #D7DBE2; }
  table.data tr:nth-child(even) td { background: #141922; }
  /* .num is applied by _tbl to every column after the first, whatever it
     holds - so a long description or header in one of those columns used
     to be unbreakable and pushed the last column clean off the sheet
     (Exec Summary, RightSize, Consumables and the all-companies summary
     all lost a column that way). Numbers and money have no spaces in
     them, so they still never break; only long text does, and only when
     the column is squeezed - which beats being invisible.
     (A. Fisher, 25 Jul 2026) */
  table.data .num { text-align: right; white-space: normal; }
  table.data th { white-space: normal; }
  table.aging { border-collapse: collapse; margin-top: 4px; }
  table.aging td, table.aging th { border: 1px solid #2A313C; padding: 8px 16px;
                                   text-align: center; font-size: 10.5pt; color: #D7DBE2; }
  table.aging th { background: #20262F; font-size: 9pt;
                   text-transform: uppercase; color: #fff; }
  .g { background: #14301b; color: #cfe8d4; } .a { background: #33280c; color: #f3e2b0; }
  .r { background: #3a1a17; color: #f3c4bb; }
  /*  Neither a match nor an exception: a gap somebody DECIDED on.
      Amber would say "go and fix this", green would say "nothing to
      see" - both wrong for a hold. Grey says "known, and here is
      who said so".  (4 Aug 2026)  */
  .n { background: #232B36; color: #c3ccd8; }
  .note { font-size: 9.5pt; color: #A9B1BD; margin-top: 8px; line-height: 1.5; }
  .intro { background: #171B22; border-left: 4px solid __ORANGE__;
           padding: 14px 18px; margin-top: 14px; font-size: 11pt; color: #D7DBE2;
           line-height: 1.65; border-radius: 0 10px 10px 0; }
  .intro b { color: #fff; }
  /* Story emphasis (26 Jul 2026): the bold words and numbers inside a
     story card carry the colour - the eye lands on the numbers first,
     which is the whole point of a story card. */
  .limits { font-size: 8.5pt; color: #A9B1BD; border-top: 1px solid #2A313C;
            margin-top: 16px; padding-top: 7px; line-height: 1.5; }
  /* Replacement money wears the highlighter - Andrew's rule (26 Jul 2026):
     when a replacement cost is on a page, you SEE it. Dark ink on neon
     reads on white pages, dark pages and email screenshots alike. */
  /* No white-space:nowrap here - money has no spaces so it never breaks
     internally, and nowrap was pushing the cost column clean off the
     sheet on busy tables (the exact bug the .num note above records).
     Found again 26 Jul 2026 by the overflow sweep. */
  .repl { background: #EFFF3D; color: #101317; font-weight: 800;
          padding: 0 5px; border-radius: 4px; }
  .footer { background: #171B22; color: #c9cfd8; padding: 14px 20px;
            margin-top: 18px; font-size: 9pt; border-top: 3px solid __ORANGE__;
            border-radius: 0 0 12px 12px; }
  .footer .values { color: __ORANGE__; font-weight: bold; margin-top: 6px; }
  .footer .auth { color: #fff; margin-top: 6px; }
  @page { size: A4 portrait; margin: 11mm 10mm; }
  @media print {
    .page { max-width: none; padding: 4mm 3mm; }
    table.data thead { display: table-header-group; }
    table.data tr { page-break-inside: avoid; }
    h2 { page-break-after: avoid; }
  }
""".replace("__ORANGE__", COATES_ORANGE).replace("__DARK__", COATES_DARK)


def header_html(title, company_line, asof, generated,
                asof_label="RENTAL_STOCK export file time"):
    return (
        '<div class="band">'
        '<div class="siteiq">POWERED BY <span class="o">SITEIQ</span><br>'
        '<span style="font-weight:normal;color:#bbb;">Equipped for anything</span></div>'
        '<div class="kicker">COATES &middot; {title}</div>'
        '<h1>{company}</h1>'
        '<div class="sub">{shutdown}</div>'
        '<div class="meta">Generated: <b>{gen}</b> &nbsp;|&nbsp; '
        'Data as at: <b>{asof}</b> ({lbl}) &nbsp;|&nbsp; '
        'Author: <b>Andrew Fisher</b></div>'
        '</div>'
    ).format(title=esc(title), company=esc(company_line),
             shutdown=esc(SHUTDOWN_NAME),
             gen=generated.strftime("%d %b %Y %H:%M"),
             asof=asof.strftime("%d %b %Y %H:%M"), lbl=esc(asof_label))


def footer_html(source_line):
    return (
        '<div class="footer">'
        'All returns go through the Coates tool store, handed to the stores team '
        'and double-scanned. Nothing is returned unless it is scanned.'
        '<div class="values">Care Deeply &middot; Customer Focused &middot; '
        'Be Our Best &middot; One Team &middot; Competitive Spirit</div>'
        '<div class="auth">Author: Andrew Fisher &nbsp;|&nbsp; '
        'POWERED BY <span style="color:{o};font-weight:bold;">SITEIQ</span> '
        '&nbsp;|&nbsp; {src}</div>'
        '</div>'
    ).format(o=COATES_ORANGE, src=esc(source_line))


def paged_report_doc(title, company_line, asof, generated, source_line,
                     body_html, doc_title=None, extra_css="",
                     asof_label="RENTAL_STOCK export file time"):
    """Every report rendered as REAL pages - the snapshot-style framed
    sheets (see report_pages.py). Whole blocks only; long tables split at
    row boundaries with the header repeated; nothing ever snipped."""
    import report_pages
    asat = asof.strftime("%d %b %Y %H:%M") if asof else "-"
    #  The compliance KEY rides at the top of EVERY page, under the
    #  header - so each item below only needs a dot and a word, and the
    #  reader never has to hunt for what a colour means.
    legend = EC.key_strip_html(on=asof)
    return report_pages.paged_doc(
        esc(title), BASE_CSS + extra_css,
        header_html(title, company_line, asof, generated,
                    asof_label=asof_label) + legend,
        report_pages.compact_header(esc(title), esc(company_line), asat)
        + legend,
        footer_html(source_line),
        body_html,
        doc_title=doc_title or ("Coates K2 - " + title),
        team_html=v3_team_strip())


def honest_limits(have_repl, sales_loaded, internal=False, tx_loaded=None):
    """Small print built from what THIS run actually contains - never
    describe a data feed the report doesn't carry."""
    parts = []
    if internal:
        parts.append(
            'Rates are contracted base rates; unpriced items are excluded from '
            'value totals - never estimated, never $0. Spend to date is charged '
            'inclusive of the on-hire day. Radios with no contracted rate charge '
            '${:.2f} each per day (on hire or not) from {} - no backdating.'.format(
                RADIO_DAILY_RATE, RADIO_CHARGE_START.strftime("%d %b %Y")))
    else:
        parts.append(
            'This report shows replacement cost exposure only - hire rates and '
            'hire spend are not shown on company reports.')
    if have_repl:
        parts.append(
            'Replacement costs come from the K2 replacement cost schedule where '
            'an item matches by description; unmatched items show TBC and are '
            'excluded from the exposure total.')
    else:
        parts.append(
            'No replacement cost schedule was found for this run, so replacement '
            'costs are not shown - that is a data gap, not a nil value.')
    if sales_loaded:
        parts.append(
            'Consumables are quantities only - no consumables price feed is '
            'loaded, so no dollar values are shown or implied for them.')
    else:
        parts.append(
            'No SALES_STOCK export was found for this run - consumables are '
            'unavailable, not zero.')
    if tx_loaded is False:
        parts.append(
            'No TRANSACTIONS export was found for this run - transaction and '
            'return-behaviour counts are unavailable, not zero.')
    elif tx_loaded:
        parts.append(
            'Transaction counts come from the SiteIQ TRANSACTIONS export '
            '(attributed per item) plus consumables draws - a draw '
            'counts as ONE transaction regardless of quantity taken.')
    parts.append(
        'All data reflects SiteIQ as at the data-as-at stamp above - this '
        'report is a point-in-time snapshot.')
    return " ".join(parts)


def intro_html(m):
    """The company's own story, straight from their scorecard data - then
    the standing ask. Same story-first pattern as the Cost Tracking
    Snapshot; nothing here is ever estimated."""
    bits = ["<b>{n} item{s}</b> out with the K2 tool store in {c}'s name".format(
        n=m["n_items"], s="" if m["n_items"] == 1 else "s", c=esc(m["display"]))]
    if m["people"]:
        bits.append("across <b>{p} {pp}</b>".format(
            p=len(m["people"]), pp="person" if len(m["people"]) == 1 else "people"))
    if m["item_types"]:
        bits.append("covering {t} item type{ts}".format(
            t=m["item_types"], ts="" if m["item_types"] == 1 else "s"))
    story = "Right now there {vb} ".format(
        vb="is" if m["n_items"] == 1 else "are") + ", ".join(bits) + "."
    if m["oldest_days"] is not None:
        story += (" The oldest item has been out <b>{d} day{ds}</b> - the "
                  "chase list below runs oldest first.").format(
            d=m["oldest_days"], ds="" if m["oldest_days"] == 1 else "s")
    if m["repl_exposure"] is not None:
        story += (" If everything on this list went missing today, the "
                  "replacement cost would be "
                  "<span class='repl'>{e}</span>{only}.").format(
            e=fmt_money(m["repl_exposure"]),
            only=("" if m.get("charges")
                  else " - that's the only dollar figure on this report"))
    return (
        '<div class="intro story">{story}</div>'
        '<div class="intro">Coates is on site to keep your crew equipped - '
        'this report is how we help you keep control of it. One ask that '
        'saves everyone money: <b>when gear is finished with, bring it '
        'back</b>. Returned gear is double-scanned off your record, checked '
        'and maintained, and ready for the next crew - that is Best Service '
        '&amp; Value working for you.</div>').format(story=story)


def tiles_html(m):
    # Client-facing tiles: NO hire costs (Andrew's direction, 17 Jul 2026) -
    # replacement exposure is the only dollar figure a company sees.
    tiles = [
        (str(m["n_items"]), "Outstanding On Hire"),
        (str(len(m["people"])), "People Using Tool Store"),
        (str(m["n_tx_total"]) if (m["tx_loaded"] or m["sales_loaded"]) else "-",
         "Transactions"),
        (str(m["item_types"]), "Item Types Used"),
        ("-" if m["oldest_days"] is None else str(m["oldest_days"]),
         "Oldest On-Hire (days)"),
    ]
    if m["repl_exposure"] is not None:
        tiles.append(("<span class='repl'>{}</span>".format(
            fmt_money(m["repl_exposure"])), "Replacement Cost Exposure"))
    #  Values are generated here (counts and money only) - render raw so
    #  the highlighter span survives; labels still escape.
    cells = "".join('<td><span class="v">{}</span><span class="l">{}</span></td>'
                    .format(v, esc(l)) for v, l in tiles)
    return '<table class="tiles"><tr>{}</tr></table>'.format(cells)


def returns_html(m):
    if not m["tx_loaded"]:
        return ('<h2>How the gear is moving</h2>'
                '<div class="note">Transaction history is not in this SiteIQ '
                'export yet - same-day and multi-day return counts will appear '
                'here as transactions are recorded.</div>')
    if m["n_tx"] == 0:
        return ('<h2>How the gear is moving</h2>'
                '<div class="note">No completed tool store transactions are '
                'recorded against {c} yet in this export.</div>'.format(
                    c=esc(m["display"])))
    return (
        '<h2>How the gear is moving</h2>'
        '<table class="aging"><tr>'
        '<th>Returned same day</th><th>Kept longer, returned</th>'
        '<th>Still on hire now</th></tr>'
        '<tr><td class="g">{sd}</td><td class="a">{kl}</td>'
        '<td class="r">{oh}</td></tr></table>'
        '<div class="note">Same-day returns keep your record clean and the '
        'gear maintained - thank your crew for those. Anything still out is '
        'listed below, oldest first.</div>').format(
        sd=m["same_day"], kl=m["kept_longer"], oh=m["n_items"])


def high_care_html(m):
    hc = m["high_care"]
    if not (hc["milwaukee"] or hc["radios"] or hc["gas"]):
        return ('<h2>High-care gear</h2>'
                '<div class="note">No Milwaukee gear, radios or gas monitors '
                'are currently in {c}\'s name - nothing high-care to track.'
                '</div>'.format(c=esc(m["display"])))
    blocks = []
    if hc["gas"]:
        blocks.append(
            '<b>Gas monitors ({n}):</b> every monitor is bump-tested and '
            'calibrated before issue - that is life-safety equipment under the '
            'Coates Life Saving Rules (SEQ-GL-009). Return them to the tool '
            'store daily for charging and checks; never swap between crews '
            'without going through the store.'.format(n=len(hc["gas"])))
    if hc["milwaukee"]:
        blocks.append(
            '<b>Milwaukee gear ({n}):</b> high replacement value tooling - '
            'see the replacement cost column below. When it is not in a hand, '
            'it belongs in the tool store, not a crib room.'.format(
                n=len(hc["milwaukee"])))
    if hc["radios"]:
        blocks.append(
            '<b>Radios ({n}):</b> charged daily and pooled site-wide - return '
            'to the store for overnight charging so the next shift picks up a '
            'working unit.'.format(n=len(hc["radios"])))
    return ('<h2>High-care gear in your name</h2>'
            + "".join('<div class="note">{}</div>'.format(b) for b in blocks))


def aging_html(m):
    a = m["aging"]
    unknown = ('<th>Date unknown</th>' if a["unknown"] else "")
    unknown_v = ('<td>{}</td>'.format(a["unknown"]) if a["unknown"] else "")
    return (
        '<h2>Aging - days on hire</h2>'
        '<table class="aging"><tr>'
        '<th>0-2 days (on track)</th><th>3-4 days (at risk)</th>'
        '<th>5+ days (action needed)</th>{u}</tr>'
        '<tr><td class="g">{g}</td><td class="a">{am}</td>'
        '<td class="r">{r}</td>{uv}</tr></table>'
        '<div class="note">K2 workbook accountability thresholds for reference: '
        'overdue at 2 days, high at 3, critical at 4+. Anything 3 days or older '
        'warrants follow-up - the detail below is sorted oldest first so the '
        'chase list is at the top.</div>'
        + stackbar([("Fresh 0-2d", a["0-2"], "#3FB950"),
                    ("Watch 3-4d", a["3-4"], "#fab219"),
                    ("Chase 5d+", a["5+"], "#F85149")])
        + _aging_bars(a)
    ).format(g=a["0-2"], am=a["3-4"], r=a["5+"], u=unknown, uv=unknown_v)


def _aging_bars(a):
    total = (a["0-2"] + a["3-4"] + a["5+"]) or 1
    return ("".join((
        hbar("Fresh - out 0-2 days", "{} items".format(a["0-2"]),
             a["0-2"] / total, "#3FB950"),
        hbar("Watch - out 3-4 days", "{} items".format(a["3-4"]),
             a["3-4"] / total, "#fab219"),
        hbar("Chase - out 5+ days", "{} items".format(a["5+"]),
             a["5+"] / total, "#F85149"))))


def detail_html(m, have_repl):
    # Client-facing detail: no Daily Rate column (hire costs are internal
    # only) - REPLACEMENT COST is the story of this table, so it gets the
    # room it deserves: merged cells (hirer + ID; date + days; storage unit
    # tucked under the description), money never wrapped, priced rows in
    # white, and an exposure total row anchoring the column. Spread over
    # extra pages by the engine as needed - never squashed. (A. Fisher)
    head = ("<tr><th>Description</th><th>Asset No</th>"
            "<th>Hirer &middot; ID</th><th>On Hire</th>"
            + ("<th class='num'>Replacement Cost</th>" if have_repl else "")
            + "</tr>")
    body = []
    exposure = 0.0
    priced = unpriced = 0
    for r in m["rows"]:
        hirer = esc(re.sub(r"\s*-\s*", " ", r["hirer"] or "")) or "-"
        hid = esc(r.get("hirer_id") or "")
        cells = [
            "<td>{d}</td>".format(
                d=desc_disp(r["description"], r.get("asset"),
                            sub=esc(r["storage_unit"]) or "&nbsp;")),
            "<td style='white-space:nowrap'>{}</td>".format(asset_disp(r.get("asset"))),
            "<td style='white-space:nowrap'>{h}{i}</td>".format(
                h=hirer, i=("<br><span style='color:#8B9099;font-size:8.5pt'>"
                            + hid + "</span>") if hid else ""),
            "<td style='white-space:nowrap'>{d}<br>"
            "<span style='color:#8B9099;font-size:8.5pt'>{dy}</span></td>".format(
                d=fmt_date(r["on_hire_date"]),
                dy=("out {} day{}".format(r["days"], "" if r["days"] == 1 else "s")
                    if r["days"] is not None else "-")),
        ]
        if have_repl:
            if r["repl"] is not None:
                exposure += r["repl"]
                priced += 1
            elif not r.get("cust_owned"):
                #  only OUR unpriced gear is a gap to chase - the
                #  customer's own gear is not waiting on a price
                unpriced += 1
            cells.append("<td class='num' style='white-space:nowrap'>"
                         + repl_display(r) + "</td>")
        body.append("<tr>" + "".join(cells) + "</tr>")
    total_row = ""
    note = ""
    if have_repl:
        total_row = ("<tr><td colspan='4' style='color:" + INK_STRONG +
                     ";font-weight:700'>"
                     "Replacement cost exposure - if nothing came back"
                     "</td><td class='num' style='white-space:nowrap'>"
                     "<span class='repl'>{}</span></td></tr>").format(
                         fmt_money(exposure))
        if unpriced:
            note = ("<div class='note'>{n} item(s) show TBC - not yet in the "
                    "replacement cost schedule, so they are excluded from the "
                    "exposure total rather than guessed. They are being "
                    "priced.</div>").format(n=unpriced)
    return ('<h2>Outstanding - on hire to you, not yet returned (oldest first)</h2>'
            '<table class="data"><thead>{h}</thead><tbody>{b}{t}</tbody></table>{n}'
            ).format(h=head, b="".join(body), t=total_row, n=note)


def consumables_html(m):
    if not m["sales_loaded"]:
        # No export = unknown, not zero. Say so honestly - never present an
        # absent data feed as a real nil result.
        return ('<h2>Consumables taken</h2>'
                '<div class="note">No SALES_STOCK export was found for this run, '
                'so consumables movements are UNAVAILABLE - this is a data gap, '
                'not a nil result. Drop the SALES_STOCK export in the kit folder '
                'and rerun to include them.</div>')
    if not m["consumables"]:
        return ('<h2>Consumables taken</h2>'
                '<div class="note">No consumables movements are recorded against '
                'this company in the SALES_STOCK export.</div>')
    draw_word = "draw" if m["cons_draws"] == 1 else "draws"
    rows = "".join(
        "<tr><td>{}</td><td>{}</td><td>{}</td><td class='num'>{:g}</td><td>{}</td></tr>".format(
            esc(c["sku"]), esc(c["hirer"]) or "-",
            esc(c.get("hirer_id") or hirer_id_for(c["hirer"])) or "-",
            c["qty"], fmt_date(c["last"]))
        for c in m["consumables"])
    return ('<h2>Consumables taken ({d} {dw})</h2>'
            '<table class="data"><thead><tr><th>SKU Description</th><th>Hirer</th>'
            '<th>Hirer ID</th>'
            '<th class="num">Total Qty Taken</th><th>Last Taken</th></tr></thead>'
            '<tbody>{r}</tbody></table>'
            '<div class="note">Each draw counts as one transaction regardless of '
            'quantity taken. Quantities only - no consumables price feed is '
            'loaded, so no dollar values are shown for consumables.</div>'
            ).format(d=m["cons_draws"], dw=draw_word, r=rows)


def render_company_html(m, asof, generated, have_repl, source_line):
    m["charges"] = charges_for_company(m["display"])
    return (
        paged_report_doc(
            "Company On-Hire Report", m["display"], asof, generated,
            source_line,
            intro_html(m) + v3_company_kpis(m)
            + returns_html(m) + aging_html(m)
            + high_care_html(m) + detail_html(m, have_repl)
            + EC.legend_html()
            + consumables_html(m) + company_charges_html(m)
            + EC.store_story_html()
            + v3_team_page()
            + "<div class='limits'><b>Honest limits:</b> {}</div>".format(
                honest_limits(have_repl, m["sales_loaded"], internal=False,
                              tx_loaded=m["tx_loaded"])),
            doc_title="Coates On-Hire Report - " + m["display"]))


def render_summary_html(models, asof, generated, source_line, radio_fleet=0,
                        have_repl=True, gas_fleet=0):
    # Alphabetical by company on all reports - the Coates standard.
    models = sorted(models, key=lambda m: m["display"].upper())
    rows = []
    for i, m in enumerate(models, 1):
        rows.append(
            "<tr><td class='num'>{i}</td><td>{c}</td><td class='num'>{n}</td>"
            "<td class='num'>{h}</td><td class='num'>{o}</td>"
            "<td class='num'>{dv}</td><td class='num'>{sp}</td>"
            "<td class='num'>{cq}</td></tr>".format(
                i=i, c=esc(m["display"]), n=m["n_items"], h=m["n_hirers"],
                o="-" if m["oldest_days"] is None else m["oldest_days"],
                dv=fmt_money(m["daily_value"]), sp=fmt_money(m["spend"]),
                cq=("{:g}".format(m["cons_qty"]) if m["sales_loaded"] else "-")))
    unknown = [m["display"] for m in models if not is_known_company(m["company"])]
    unknown_note = ""
    if unknown:
        unknown_note = ('<div class="note"><b>Not on the K2 contractor '
                        'register:</b> {} - check the spelling at the counter '
                        'or confirm induction, then add to the register or the '
                        'alias table.</div>').format(esc(", ".join(unknown)))
    table = (unknown_note +
             '<h2>All companies - alphabetical</h2>'
             '<table class="data"><thead><tr><th class="num">#</th><th>Company</th>'
             '<th class="num">Items On Hire</th><th class="num">Unique Hirers</th>'
             '<th class="num">Oldest (days)</th><th class="num">Est. Daily Value</th>'
             '<th class="num">Est. Spend To Date</th>'
             '<th class="num">Consumables Qty</th></tr></thead>'
             '<tbody>{}</tbody></table>').format("".join(rows))
    totals = ('<div class="note">Totals: {n} items on hire across {c} companies; '
              'est. daily value {dv}; est. spend to date {sp}. This summary is '
              'the INTERNAL Coates view - hire costs do not appear on the '
              'per-company reports.</div>').format(
        n=sum(m["n_items"] for m in models), c=len(models),
        dv=fmt_money(sum(m["daily_value"] or 0 for m in models)),
        sp=fmt_money(sum(m["spend"] or 0 for m in models)))
    totals += STREAM_NOTE_SITEIQ
    radio_days = max(1, (dt.date.today() - RADIO_CHARGE_START).days + 1)
    radio_line = ""
    if radio_fleet:
        billable = min(radio_fleet, RADIO_BILLABLE_CAP)
        spares = radio_fleet - billable
        spare_txt = ""
        if spares > 0:
            spare_txt = (' {sp} handset{pl2} held as unbilled spare{pl2} in '
                         'case a unit stops working - disclosed here, '
                         'excluded from the charge.').format(
                sp=spares, pl2="" if spares == 1 else "s")
        radio_line = (
            '<h2>Radio fleet daily charge - on the separate invoice</h2>'
            '<div class="note"><b>Charged on the separate invoice, not '
            'through SiteIQ</b> - shown here so the whole cost picture is '
            'on one page, never added to the SiteIQ totals above. '
            'Cross-checked line-for-line by the Separate Invoice Tracker '
            '(button 49). {n} radio handsets in RENTAL_STOCK (batteries '
            'and covers excluded); <b>{b} billable</b> at {r} each per day, '
            'charged whether on hire or not, from {s} (no backdating): '
            '<b>{dv} per day</b>; accrued to date ({d} charged day{pl}): '
            '<b>{acc}</b>.{spare}</div>').format(
            n=radio_fleet, b=billable, r=fmt_money(RADIO_DAILY_RATE),
            s=RADIO_CHARGE_START.strftime("%d %b %Y"),
            dv=fmt_money(billable * RADIO_DAILY_RATE),
            d=radio_days, pl="" if radio_days == 1 else "s",
            acc=fmt_money(billable * RADIO_DAILY_RATE * radio_days),
            spare=spare_txt)
    radio_line += (
        '<div class="note"><b>Sub-hired gear is the same deal</b> - the '
        'units whose barcodes carry the SUB prefix (the SUBHARVEY '
        'welders) or a supplier code (forklift FK505) are <b>charged '
        'on the separate invoice</b>: listed and tracked in this pack, '
        'no rate in the SiteIQ totals, dollars on the Separate Invoice '
        'Tracker (button 49). Coates welders and forklifts with plain '
        'numeric barcodes ARE SiteIQ-charged and keep their rates. '
        'Nothing counted twice, nothing stripped wrongly.</div>')
    gas_line = ""
    if gas_fleet:
        gas_daily = gas_fleet * GAS_RATE
        if dt.date.today() >= GAS_START:
            gas_days = (dt.date.today() - GAS_START).days + 1
            gas_line = (
                '<h2>Gas monitor fleet daily charge - on the separate '
                'invoice</h2>'
                '<div class="note"><b>Charged on the separate invoice, not '
                'through SiteIQ</b> - shown for the whole picture, never '
                'added to the SiteIQ totals above; the Separate Invoice '
                'Tracker (button 49) carries and cross-checks it. '
                '{n}-unit gas monitor fleet; billed at {r} '
                'each per day, charged whether on hire or not, from {s} until '
                'the last forecast day in Daily Tracking: <b>{dv} per day</b>; '
                'accrued to date ({d} charged day{pl}): <b>{acc}</b>.'
                '</div>').format(
                n=gas_fleet, r=fmt_money(GAS_RATE),
                s=GAS_START.strftime("%d %b %Y"), dv=fmt_money(gas_daily),
                d=gas_days, pl="" if gas_days == 1 else "s",
                acc=fmt_money(gas_daily * gas_days))
        else:
            gas_line = (
                '<h2>Gas monitor fleet daily charge</h2>'
                '<div class="note">{n}-unit gas monitor fleet; billing starts '
                '{s} at {r} each per day ({dv} per day), charged whether on '
                'hire or not, until the last forecast day in Daily Tracking. '
                'Nothing accrued yet.</div>').format(
                n=gas_fleet, s=GAS_START.strftime("%d %b %Y"),
                r=fmt_money(GAS_RATE), dv=fmt_money(gas_daily))
    sales_loaded = all(m["sales_loaded"] for m in models) if models else True
    return (
        paged_report_doc(
            "On-Hire Summary (Internal)", "All Companies", asof, generated,
            source_line,
            table + totals + radio_line + gas_line
            + v3_team_page()
            + "<div class='limits'><b>Honest limits:</b> {}</div>".format(
                honest_limits(have_repl, sales_loaded, internal=True)),
            doc_title="Coates On-Hire Summary - All Companies"))


# ---------------------------------------------------------------------------
# Site Plant report - in use by Andrew's categories, idle-and-charged at the
# bottom (the potential cost saving view)
# ---------------------------------------------------------------------------

def build_plant_model(plant, rates, cats, today, infra=None):
    in_use, idle = [], []
    uncharged_rows = []
    for p in plant:
        info = plant_category_info(p["variant"], cats) if cats else None
        p["category"] = info["category"] if info else "Not in Plant Register"
        # rate source: contracted rates file first, plant register fallback
        rate = rates.get(p["variant"])
        if rate is None and info:
            rate = info["rate"]
        if is_sep_invoice(p):
            # Sub-hired plant (dashed item number / SUB barcode - the
            # SUBHARVEY welders and forklift FK505) is charged on the
            # SEPARATE INVOICE - no dollar from it joins a SiteIQ-stream
            # total. Coates plant with plain numeric IDs keeps its
            # rates. Tracked for use/idle/audit as always; the Separate
            # Invoice Tracker (49) carries the sub-hire money.
            #
            # No count is assumed anywhere here - this tests the barcode
            # on the row in front of it. Worth knowing though: Baseplan
            # bills 8 welders and the SiteIQ pull evidences 7, because
            # SUBHARVEY002 has no row in it. That gap is a decision, not
            # a fault - see HELD_BACK in build_baseplan_costs.py.
            rate = None
        p["rate"] = rate
        if is_site_plant_equipment(p["hirer"]):
            # Back on site, still on hire = still being charged. The saving.
            days = (max(0, (today - p["on_hire_date"]).days)
                    if p["on_hire_date"] else None)
            p["days_sitting"] = days
            p["idle_charge"] = (rate * (days + 1)
                                if (rate is not None and days is not None) else None)
            idle.append(p)
        elif p["hirer"] or p["company_raw"]:
            p["company"] = (display_company(canonical_company(p["company_raw"]))
                            if p["company_raw"] else "-")
            p["days"] = (max(0, (today - p["on_hire_date"]).days)
                         if p["on_hire_date"] else None)
            in_use.append(p)
        else:
            # No hirer recorded - not in use, charging not yet started.
            uncharged_rows.append(p)
    in_use.sort(key=lambda p: (p["category"], p["description"]))
    idle.sort(key=lambda p: (-(p["days_sitting"] if p["days_sitting"] is not None
                               else -1), p["category"], p["description"]))
    # Aggregate the not-in-use fleet by category/description/status - this is
    # the true-utilisation view and the mechanic's preventative maintenance
    # window (service it while it's idle, not in a contractor's hands).
    agg = {}
    for p in uncharged_rows:
        key = (p["category"], p["description"], p["status"])
        agg[key] = agg.get(key, 0) + 1
    available_groups = sorted(
        ({"category": k[0], "description": k[1], "status": k[2], "qty": n}
         for k, n in agg.items()),
        key=lambda g: (g["category"], -g["qty"], g["description"]))
    # Itemised idle list for the printable audit checklist: every asset not
    # in a contractor's hands, with its asset number, so the yard walk can
    # tick each one off. Anything not sighted and not on hire went out
    # without going on hire - that is the compliance catch.
    audit_rows = []
    for p in idle:
        audit_rows.append({**p, "state": "Idle - on charge (Site Plant Equipment)"})
    for p in uncharged_rows:
        audit_rows.append({**p, "state": "Not on hire"})
    audit_rows.sort(key=lambda p: (p["category"], p["description"], p["barcode"]))

    # Site infrastructure: name + total qty only (on site for the duration)
    infra_agg = {}
    for desc in (infra or []):
        infra_agg[desc] = infra_agg.get(desc, 0) + 1
    infra_groups = sorted(({"description": k, "qty": n}
                           for k, n in infra_agg.items()),
                          key=lambda g: g["description"])
    return {
        "in_use": in_use, "idle": idle,
        "uncharged": len(uncharged_rows),
        "available_groups": available_groups,
        "audit_rows": audit_rows,
        "fleet": len(plant),
        "infra_groups": infra_groups,
        "infra_total": len(infra or []),
        "saving_per_day": sum(p["rate"] for p in idle if p["rate"] is not None),
        "idle_unpriced": sum(1 for p in idle if p["rate"] is None),
    }


def render_plant_html(pm, asof, generated, source_line, cats_found):
    n_avail = len(pm["idle"]) + pm["uncharged"]
    util = (len(pm["in_use"]) / pm["fleet"]) if pm["fleet"] else 0
    tiles = [
        (str(pm["fleet"]), "Plant Fleet Onsite"),
        (str(len(pm["in_use"])), "In Use (On Hire)"),
        (str(n_avail), "Available On Site"),
        ("{:.0%}".format(util), "True Utilisation"),
        (fmt_money(pm["saving_per_day"]) if pm["idle"] else "$0.00",
         "Potential Saving / Day"),
    ]
    tiles_h = '<table class="tiles"><tr>{}</tr></table>'.format(
        "".join('<td><span class="v">{}</span><span class="l">{}</span></td>'
                .format(esc(v), esc(l)) for v, l in tiles))
    tiles_h += STREAM_NOTE_SITEIQ

    # ---- in use, grouped by Andrew's register categories ----
    rows = []
    last_cat = None
    for p in pm["in_use"]:
        if p["category"] != last_cat:
            rows.append('<tr><td colspan="7" style="background:{};color:#fff;'
                        'font-weight:bold;">{}</td></tr>'.format(
                            COATES_DARK, esc(p["category"])))
            last_cat = p["category"]
        rows.append(
            "<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td>"
            "<td>{}</td>"
            "<td class='num'>{}</td></tr>".format(
                desc_disp(p["description"], p.get("asset")),
                asset_disp(p.get("asset")),
                esc(p["hirer"]) or "-",
                esc(p.get("hirer_id")) or "-",
                esc(p["company"]),
                fmt_date(p["on_hire_date"]),
                "-" if p["days"] is None else p["days"]))
    if pm["in_use"]:
        in_use_h = ('<h2>Plant in use - by category, who has it and for whom</h2>'
                    '<table class="data"><thead><tr><th>Description</th>'
                    '<th>Asset No</th><th>On Hire To</th><th>Hirer ID</th>'
                    '<th>Company</th>'
                    '<th>On-Hire Date</th><th class="num">Days</th></tr></thead>'
                    '<tbody>{}</tbody></table>').format("".join(rows))
    else:
        in_use_h = ('<h2>Plant in use</h2><div class="note">No site plant is '
                    'currently on hire to a contractor.</div>')

    # ---- available and still charged (the saving) ----
    if pm["idle"]:
        irows = "".join(
            "<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td>"
            "<td class='num'>{}</td><td class='num'>{}</td>"
            "<td class='num'>{}</td></tr>".format(
                esc(p["category"]),
                desc_disp(p["description"], p.get("asset")),
                asset_disp(p.get("asset")),
                fmt_date(p["on_hire_date"]),
                "-" if p["days_sitting"] is None else p["days_sitting"],
                fmt_money(p["rate"]), fmt_money(p["idle_charge"]))
            for p in pm["idle"])
        unpriced_note = ""
        if pm["idle_unpriced"]:
            unpriced_note = (' {} item(s) have no rate on file and are excluded '
                             'from the saving figure.').format(pm["idle_unpriced"])
        idle_h = (
            '<h2>Available on site - still being charged (potential saving)</h2>'
            '<table class="data"><thead><tr><th>Category</th><th>Description</th>'
            '<th>Asset No</th><th>Back On Site Since</th>'
            '<th class="num">Days Sitting</th><th class="num">Rate / Day</th>'
            '<th class="num">Charged While Idle</th></tr></thead>'
            '<tbody>{r}</tbody></table>'
            '<div class="note"><b>Off-hiring this gear saves {s} per day.</b> '
            'Sorted longest-sitting first - that is the conversation list.'
            '{u}</div>').format(r=irows, s=fmt_money(pm["saving_per_day"]),
                                u=unpriced_note)
    else:
        idle_h = (
            '<h2>Available on site - still being charged (potential saving)</h2>'
            '<div class="note">Nothing is on hire to Site Plant Equipment yet - '
            'charging under the return-to-site convention has not started. '
            'From now: when plant goes out it is hired to the person using it; '
            'when it comes back the tool store puts it on hire to '
            '<b>Site Plant Equipment (Coates)</b>. Anything in that name is, '
            'by definition, sitting available and still being charged - it will '
            'appear here with days sitting and the daily saving available.</div>')

    # ---- not-in-use fleet: true utilisation + the maintenance window ----
    uncharged_note = ""
    if pm["available_groups"]:
        arows = []
        last_cat = None
        for g in pm["available_groups"]:
            if g["category"] != last_cat:
                arows.append('<tr><td colspan="3" style="background:{};'
                             'color:#fff;font-weight:bold;">{}</td></tr>'.format(
                                 COATES_DARK, esc(g["category"])))
                last_cat = g["category"]
            arows.append(
                "<tr><td>{}</td><td>{}</td><td class='num'>{:g}</td></tr>".format(
                    desc_disp(g["description"]), esc(g["status"]) or "-",
                    g["qty"]))
        uncharged_note = (
            '<h2>Available on site - not in use ({n} items): the maintenance '
            'window</h2>'
            '<div class="note">If it\'s not in use, it comes back to the tool '
            'store. This list is the true-utilisation picture of the site plant '
            'fleet, and it is the Coates mechanic\'s preventative maintenance '
            'worklist - service gear while it sits idle, so it never fails in a '
            'contractor\'s hands (Life Saving Rule 5 - Tools and Equipment). '
            'Charging for these items has not started; as each one moves, the '
            'double-scan puts it on hire to its user, and on return to '
            '<b>Site Plant Equipment (Coates)</b>.</div>'
            '<table class="data"><thead><tr><th>Description</th><th>Status</th>'
            '<th class="num">Qty</th></tr></thead><tbody>{r}</tbody></table>'
        ).format(n=pm["uncharged"], r="".join(arows))

    # ---- site infrastructure: on site for the duration, name + qty only ----
    infra_h = ""
    if pm["infra_groups"]:
        irows2 = "".join(
            "<tr><td>{}</td><td class='num'>{:g}</td></tr>".format(
                desc_disp(g["description"]), g["qty"])
            for g in pm["infra_groups"])
        infra_h = (
            '<h2>Site infrastructure - barriers, chutes &amp; hoppers '
            '({n} items)</h2>'
            '<div class="note">Barriers, crowd control, rubbish chutes, '
            'hoppers and brackets stay on site for the duration of the '
            'shutdown by design - they are not idle plant, not a saving '
            'opportunity and not an issue. Listed for the record:</div>'
            '<table class="data"><thead><tr><th>Description</th>'
            '<th class="num">Total Qty</th></tr></thead>'
            '<tbody>{r}</tbody></table>').format(n=pm["infra_total"], r=irows2)

    cats_note = "" if cats_found else (
        '<div class="note">NOTE: the K2 workbook Plant Equipment register was '
        'not found this run, so categories fall back to "Not in Plant '
        'Register".</div>')

    limits = (
        'Idle plant is identified strictly by the site convention: on hire to '
        'hirer "Site Plant Equipment", company Coates - nothing else is counted '
        'as available-and-charged. Rates are contracted base rates (plant '
        'register rate used only where no contracted rate exists); items with '
        'no rate show a dash and are excluded from saving totals - never '
        'estimated, never $0. Charged-while-idle is inclusive of the return '
        'day. Categories come from the K2 Plant Equipment register. All data '
        'reflects SiteIQ as at the data-as-at stamp above - this report is a '
        'point-in-time snapshot.')

    lsr = ('<div class="note"><b>Life Saving Rules (SEQ-GL-009):</b> '
           'Rule 5 - Tools and Equipment: damaged or defective plant is tagged '
           'out and removed from service, never reissued. Rule 6 - Critical '
           'Risk Non-Negotiables: mobile plant work follows the non-negotiable '
           'controls for that Critical Risk.</div>')

    return (
        paged_report_doc(
            "Site Plant Report", "Site Plant Equipment", asof, generated,
            source_line,
            tiles_h + in_use_h + idle_h + uncharged_note + infra_h
            + cats_note + lsr
            + "<div class='limits'><b>Honest limits:</b> {}</div>".format(limits),
            doc_title="Coates Site Plant Report - K2"))


def generate_plant(pm, asof, generated, source_line, date_tag, cats_found):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stem = "Coates_SitePlant_Report_{}".format(date_tag)
    html_path = os.path.join(PAGES_DIR, stem + ".html")
    pdf_path = os.path.join(PDF_DIR, stem + ".pdf")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(render_plant_html(pm, asof, generated, source_line, cats_found))
    pdf_ok = html_to_pdf(html_path, pdf_path)
    print("  Site Plant report: {} in use, {} available-charged | HTML{}".format(
        len(pm["in_use"]), len(pm["idle"]),
        " + PDF" if pdf_ok else " (no PDF engine)"))
    return pdf_ok


# ---------------------------------------------------------------------------
# Gear Return / Demob push pack - "Let's Bring It Home"
# ---------------------------------------------------------------------------

def render_demob_html(models, stocktake, asof, generated, source_line):
    today = dt.date.today()
    days_left = max(0, (SHUTDOWN_END - today).days)

    all_rows = []
    for m in models.values():
        for r in m["rows"]:
            all_rows.append({**r, "company": m["company"],
                             "display": m["display"]})
    by_co = {}
    for r in all_rows:
        co = by_co.setdefault(r["display"], {})
        h = co.setdefault(r["hirer"] or "(no name recorded)",
                          {"items": 0, "oldest": None})
        h["items"] += 1
        if r["days"] is not None and (h["oldest"] is None or
                                      r["days"] > h["oldest"]):
            h["oldest"] = r["days"]
    aged = sorted((r for r in all_rows if r["days"] is not None
                   and r["days"] >= 5), key=lambda r: -r["days"])
    people_holding = len({(r["display"], r["hirer"]) for r in all_rows})
    n_items = len(all_rows)

    v3strip = v3_kpis([
        ("clock", "{:,}".format(days_left), "Days to demob",
         "<span style='color:{c};font-size:8pt;font-weight:700'>{d}</span>".format(
             c=V3_AMBER if days_left <= 7 else V3_STEEL,
             d=SHUTDOWN_END.strftime("%d %b"))),
        ("box", "{:,}".format(n_items), "Items still out", ""),
        ("people", "{:,}".format(people_holding), "People holding gear", ""),
        ("alert", "{:,}".format(len(aged)), "Out 5 days or more",
         "<span style='color:{c};font-size:8pt;font-weight:700'>chase list "
         "inside</span>".format(c=V3_BAD if aged else V3_GOOD)),
    ])
    intro = (
        "The job is winding down, and the next few days are about one thing: "
        "getting every bit of gear home, checked and accounted for. Nothing "
        "in this pack is an accusation - it is a checklist, so demob day is "
        "boring in the best possible way. Every return takes seconds at the "
        "tool store, is double-scanned straight off your record, and gets "
        "checked by Coates on the spot - gear you send back today is gear "
        "you are not answering for later. {n} item{s} {are} still out across "
        "{c} compan{cy} and {p} {pw}; the lists below show exactly who has "
        "what, so supervisors can walk the crib room, not the whole site."
    ).format(n=n_items, s="" if n_items == 1 else "s",
             are="is" if n_items == 1 else "are", c=len(by_co),
             cy="y" if len(by_co) == 1 else "ies", p=people_holding,
             pw="person" if people_holding == 1 else "people")

    tiles = [
        (str(n_items), "Items Still Out"),
        (str(len(by_co)), "Companies Holding"),
        (str(people_holding), "People Holding"),
        (str(len(aged)), "Aged 5+ Days"),
        (str(days_left), "Days To Demob"),
    ]
    tiles_h = '<table class="tiles"><tr>{}</tr></table>'.format(
        "".join('<td><span class="v">{}</span><span class="l">{}</span></td>'
                .format(esc(v), esc(l)) for v, l in tiles))
    if by_co:
        cnts = {co: sum(v["items"] for v in ppl.values())
                for co, ppl in by_co.items()}
        mx = max(cnts.values()) or 1
        _pal = ["#F26222", "#fab219", "#3FB950", "#8B9099", "#C44C28", "#FFD27A"]
        _top = sorted(cnts.items(), key=lambda kv: -kv[1])[:5]
        _rest = sum(n for _c, n in sorted(cnts.items(), key=lambda kv: -kv[1])[5:])
        segs = [(c.split()[0], n, _pal[i % len(_pal)]) for i, (c, n) in enumerate(_top)]
        if _rest:
            segs.append(("Others", _rest, "#5b6472"))
        tiles_h += stackbar(segs, height=30)
        tiles_h += ("<div class='note' style='margin-top:12px'><b>Still out, "
                    "by company</b> - the push list at a glance.</div>"
                    + "".join(hbar(co, "{} items".format(n), n / mx)
                              for co, n in sorted(cnts.items(),
                                                  key=lambda kv: -kv[1])[:10]))

    if aged:
        rows = "".join(
            "<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td>"
            "<td class='num'>{}</td>"
            "</tr>".format(desc_disp(r["description"], r.get("asset")),
                           esc(r["display"]),
                           esc(r["hirer"]) or "-",
                           esc(r.get("hirer_id")) or "-", r["days"])
            for r in aged[:30])
        concern_h = (
            '<h2>Worth a conversation first - out 5 days or more</h2>'
            '<table class="data"><thead><tr><th>Item</th><th>Company</th>'
            '<th>Person</th><th>Hirer ID</th>'
            '<th class="num">Days Out</th></tr></thead>'
            '<tbody>{r}</tbody></table>'
            '<div class="note">Long days out usually just means the gear is '
            'working hard - a quick check-in confirms it is still needed or '
            'sends it home. Either answer is a good answer.</div>').format(r=rows)
    else:
        concern_h = ('<h2>Worth a conversation first</h2>'
                     '<div class="note">Nothing has been out 5 days or more - '
                     'a clean sheet going into demob.</div>')

    sections = []
    for co in sorted(by_co, key=str.upper):
        hirers = by_co[co]
        cnt = sum(h["items"] for h in hirers.values())
        # class='pgbrk': each company STARTS A NEW PAGE in the paged
        # layout (Andrew's rule - a new company never carries on halfway
        # down another company's page)
        sections.append(
            '<tr class="pgbrk"><td colspan="4" style="background:{dk};color:#fff;'
            'font-weight:bold;">{co} - {n} item{s} across {p} {pw}</td></tr>'
            .format(dk=COATES_DARK, co=esc(co), n=cnt,
                    s="" if cnt == 1 else "s", p=len(hirers),
                    pw="person" if len(hirers) == 1 else "people"))
        for hname in sorted(hirers, key=str.upper):
            h = hirers[hname]
            sections.append(
                "<tr><td>{}</td><td>{}</td><td class='num'>{}</td>"
                "<td class='num'>{}</td></tr>".format(
                    esc(hname), esc(hirer_id_for(hname)) or "-", h["items"],
                    "-" if h["oldest"] is None else h["oldest"]))
    byco_h = ('<h2>Who has what - by company and person</h2>'
              '<table class="data"><thead><tr><th>Person</th>'
              '<th>Hirer ID</th>'
              '<th class="num">Items</th><th class="num">Oldest (Days)</th>'
              '</tr></thead><tbody>{}</tbody></table>'
              '<div class="note">Line-item detail per company is in the '
              'company on-hire reports - hand each person their own list.'
              '</div>').format("".join(sections))

    st_bit = ("Daily stocktake coverage is evidenced in SiteIQ.")
    if stocktake:
        st_bit = ("{sv:,} of {tot:,} tool store items physically verified in "
                  "the last 3 days ({pct:.0%}) by {cnt} counters - so when "
                  "the final reconciliation runs, it is confirming what we "
                  "already know.".format(
                      sv=stocktake["sighted3"], tot=stocktake["total"],
                      pct=stocktake["pct"], cnt=stocktake["counters"]))
    process_h = (
        '<h2>Why this works - the process behind a clean demob</h2>'
        '<div class="note"><b>Double scan - two scans, two looks, one '
        'standard.</b> Every return is scanned twice at the counter, looked '
        'at both times, and comes straight off the company\'s record - no '
        'paperwork, no disputes, done in seconds.</div>'
        '<div class="note"><b>We check the gear itself, not just the paperwork.</b> '
        'Every returned item is looked over by Coates as it lands: complete, '
        'clean, working. Nothing goes back on the shelf unless it is ready '
        'for hire - clean, compliant, in date, not damaged. Harnesses and '
        'slings get extra care on inspection - safety-critical gear is '
        'checked thoroughly, never glanced at. Anything damaged or defective '
        'is tagged out through the Out of Service process on the spot - Life '
        'Saving Rule 5, Tools and Equipment (SEQ-GL-009) - so nothing dodgy '
        'goes back on a shelf or out to the next job.</div>'
        '<div class="note"><b>Stocktakes back it up.</b> {st}</div>'
        '<div class="note"><b>The ask.</b> If it\'s not in a hand, bring it '
        'to the tool store today. Care Deeply, One Team - let\'s land this '
        'shutdown with a zero-loss sheet everyone can be proud of.</div>'
    ).format(st=st_bit)

    limits = (
        'This pack reflects SiteIQ as at the data-as-at stamp - a '
        'point-in-time snapshot; anything returned after the extract is '
        'already off the list. Days out counts calendar days since the item '
        'went on hire. No charges appear here - this pack exists purely to '
        'make demob clean.')

    hero = ('<div style="background:{o};color:#fff;font-weight:bold;'
            'text-align:center;padding:8px;font-size:11.5pt;'
            '-webkit-print-color-adjust:exact;print-color-adjust:exact;">'
            '{n} item{ns} still out &bull; {c} compan{cs} &bull; demob in {d} '
            'day{ds} - if it\'s not in use, bring it home</div>').format(
        o=COATES_ORANGE, n=n_items, ns="" if n_items == 1 else "s",
        c=len(by_co), cs="y" if len(by_co) == 1 else "ies",
        d=days_left, ds="" if days_left == 1 else "s")

    return (
        paged_report_doc(
            "Gear Return - Demob Push", "Let's Bring It Home", asof,
            generated, source_line,
            hero + '<div class="intro story">' + esc(intro) + '</div>'
            + v3strip
            + tiles_h + concern_h + byco_h + process_h
            + v3_team_page()
            + "<div class='limits'><b>Honest limits:</b> {}</div>".format(limits),
            doc_title="Coates K2 Gear Return - Demob Push"))


def demob_eml_body(models, asof):
    today = dt.date.today()
    days_left = max(0, (SHUTDOWN_END - today).days)
    n_items = sum(m["n_items"] for m in models.values())
    return (
        "Hi,\n\n"
        "With demob {dl} day{s} away, attached is the gear return pack - "
        "everything still on hire across the job, listed by company and by "
        "person, current as at {asof}.\n\n"
        "The picture right now: {n} items out across {c} companies. The "
        "pack flags anything out 5 days or more for a quick check-in first "
        "- usually it just means the gear is working hard, and either "
        "answer (still needed, or coming home) is a good answer.\n\n"
        "Every return is double-scanned off the record in seconds and "
        "checked by our team on the spot, so the earlier gear comes back, "
        "the cleaner everyone's final reconciliation. If anything on the "
        "list has already been returned, flag it and we'll square it away "
        "same day.\n\n"
        "Let's land this one with a zero-loss sheet.\n\n"
        "Thanks,\n"
        "Andrew Fisher\n"
        "Shutdown Manager - Coates\n"
        "{site}\n").format(
        dl=days_left, s="" if days_left == 1 else "s",
        asof=asof.strftime("%d %b %Y %H:%M"), n=n_items,
        c=len(models), site=SHUTDOWN_NAME)


def generate_demob(models, stocktake, asof, generated, source_line, date_tag):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stem = "Coates_K2_GearReturn_Demob_{}".format(date_tag)
    html_path = os.path.join(PAGES_DIR, stem + ".html")
    pdf_path = os.path.join(PDF_DIR, stem + ".pdf")
    eml_path = os.path.join(EMAILS_DIR, stem + "_OUTLOOK.eml")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(render_demob_html(models, stocktake, asof, generated,
                                  source_line))
    pdf_ok = html_to_pdf(html_path, pdf_path)
    subject = "Coates K2 Gear Return - Demob {} - {}".format(
        SHUTDOWN_END.strftime("%d %b"), dt.date.today().strftime("%d %b %Y"))
    attach = pdf_path if pdf_ok else html_path
    # the demob pack itself, pasted into the email body as images in order
    r_to, r_cc = recipients.resolve(KIT_DIR, "", "DEMOB")
    imgs = email_images.pages_that_fit(
        email_images.capture_report_images(html_path, EMAILS_DIR, stem),
        "Demob pack")
    if imgs:
        msg = email_images.build_image_eml(subject, imgs, to=r_to, cc=r_cc)
        email_images.save_eml(msg, eml_path)
        cids = ["pg{}".format(i) for i in range(1, len(imgs) + 1)]
        write_draft_manifest(eml_path, subject, email_images.images_body(cids),
                             [], report="DEMOB",
                             images=email_images.manifest_images(imgs))
        n_items = sum(m["n_items"] for m in models.values())
        print("  Demob pack: {} items still out | HTML{} | report in email "
              "body ({} images) | Outlook draft".format(
                  n_items, " + PDF" if pdf_ok else " (no PDF engine)", len(imgs)))
        return pdf_ok
    msg = email.message.EmailMessage(policy=email.policy.SMTP)
    msg["Subject"] = subject
    if r_to:
        msg["To"] = email_images._addr_list(r_to)
    if r_cc:
        msg["Cc"] = email_images._addr_list(r_cc)
    msg["X-Unsent"] = "1"
    text = demob_eml_body(models, asof)
    msg.set_content(text)
    # like-for-like: the full PDF design rebuilt Outlook-safe in the body
    _html = email_demob_html(models, stocktake, asof, generated)
    msg.add_alternative(_html, subtype="html")
    write_draft_manifest(eml_path, str(msg["Subject"]), _html,
                         [pdf_path if pdf_ok else html_path])
    with open(attach, "rb") as f:
        data = f.read()
    if attach.endswith(".pdf"):
        msg.add_attachment(data, maintype="application", subtype="pdf",
                           filename=os.path.basename(attach))
    else:
        msg.add_attachment(data, maintype="text", subtype="html",
                           filename=os.path.basename(attach))
    with open(eml_path, "wb") as f:
        f.write(msg.as_bytes())
    n_items = sum(m["n_items"] for m in models.values())
    print("  Demob pack: {} items still out | HTML{} | Outlook draft".format(
        n_items, " + PDF" if pdf_ok else " (no PDF engine)"))
    return pdf_ok


# ---------------------------------------------------------------------------
# Plant Audit List - printable idle checklist (compliance + maintenance)
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Consumables Report (site-wide, client-facing): the tool store keeps the
# site's consumables flowing - draws, people, top items and shelf assurance.
# Quantities only - consumables carry no charges on any client report.
# ---------------------------------------------------------------------------

def load_sales_stock_levels():
    """Shelf position for the consumables report - qty only, never priced."""
    path = find_newest("SALES_STOCK*.xlsx", "the SALES_STOCK export",
                       required=False)
    if not path:
        return [], None
    try:
        _, raw = sheet_rows(path, "SALES_STOCK", "the SALES_STOCK export")
    except KitError:
        return [], None
    lines = []
    for r in raw:
        if clean_text(r.get("SKU_NUMBER")) in DEAD_SKUS:
            continue
        desc = clean_text(r.get("SKU_DESCRIPTION"))
        if not desc:
            continue
        avail = to_num(r.get("AVAILABLE_QUANTITY")) or 0
        minq = to_num(r.get("MINIMUM_QUANTITY")) or 0
        sold = to_num(r.get("SALES_QUANTITY")) or 0
        if (avail == 0 and minq == 0 and sold == 0
                and not clean_text(r.get("SALES_DATE"))):
            continue    # all-zero placeholder - never shown
        lines.append({"desc": desc, "avail": avail, "min": minq})
    return lines, path


def build_consumables_site_model(sales, stock_lines):
    by_co = {}
    item_stats = {}
    people = set()
    for m in sales:
        co = m["company"]
        e = by_co.setdefault(co, {"draws": 0, "qty": 0.0, "items": {}})
        e["draws"] += 1                      # a draw = ONE transaction
        e["qty"] += m["qty"]
        e["items"][m["sku"]] = e["items"].get(m["sku"], 0) + 1
        ist = item_stats.setdefault(m["sku"], {"draws": 0, "qty": 0.0})
        ist["draws"] += 1
        ist["qty"] += m["qty"]
        if m["hirer"]:
            people.add(m["hirer"])
    companies = sorted(by_co, key=lambda c: display_company(c).upper())
    top = sorted(item_stats.items(),
                 key=lambda kv: (-kv[1]["draws"], -kv[1]["qty"]))[:10]
    stocked = sum(1 for l in stock_lines if l["avail"] > 0)
    managed = sum(1 for l in stock_lines if l["min"] > 0)
    restock = sum(1 for l in stock_lines
                  if l["min"] > 0 and l["avail"] <= l["min"])
    return {
        "by_co": by_co, "companies": companies, "top": top,
        "moves": sales,          # raw movement rows - the per-person truth
        "draws": sum(e["draws"] for e in by_co.values()),
        "qty": sum(e["qty"] for e in by_co.values()),
        "people": len(people), "item_types": len(item_stats),
        "stock_lines": len(stock_lines), "stocked": stocked,
        "managed": managed, "restock": restock,
    }


def _cons_qty(v):
    return "{:,.0f}".format(v)


def render_consumables_html(cm, asof, generated, source_line, sales_loaded):
    if cm["draws"]:
        story = (
            "The tool store is not just tools - it keeps the whole site's "
            "consumables flowing. So far this shutdown, crews have made "
            "<b>{d:,} consumable draws</b> ({q} items issued) across "
            "<b>{c} compan{cs}</b> and <b>{p} people</b>, covering {t} "
            "different consumable lines. Every draw goes through the counter "
            "and the double scan - two scans, two looks, one standard - so "
            "every number here is a scanned fact. A draw counts as ONE "
            "transaction no matter the quantity taken, and consumables carry "
            "no charges on this report - quantities only.").format(
            d=cm["draws"], q=_cons_qty(cm["qty"]), c=len(cm["companies"]),
            cs="y" if len(cm["companies"]) == 1 else "ies",
            p=cm["people"], t=cm["item_types"])
    else:
        story = (
            "The shelf is ready before the rush: <b>{s} consumable lines are "
            "stocked on site</b>, {m} of them under min/max management, "
            "checked every day as part of the daily stocktake. No consumable "
            "draws are recorded in SiteIQ yet - as crews start drawing from "
            "the store, every draw is double-scanned and this report builds "
            "the usage story company by company, automatically. A draw "
            "counts as ONE transaction no matter the quantity, and "
            "consumables carry no charges on this report - quantities "
            "only.").format(s=cm["stocked"], m=cm["managed"])
    glance = (
        '<h2>At a glance</h2><table class="data"><thead><tr>'
        '<th class="num">Draws</th><th class="num">Items Issued</th>'
        '<th class="num">Companies</th><th class="num">People</th>'
        '<th class="num">Consumable Types Used</th></tr></thead><tbody>'
        '<tr><td class="num">{d:,}</td><td class="num">{q}</td>'
        '<td class="num">{c}</td><td class="num">{p}</td>'
        '<td class="num">{t}</td></tr></tbody></table>').format(
        d=cm["draws"], q=_cons_qty(cm["qty"]), c=len(cm["companies"]),
        p=cm["people"], t=cm["item_types"])
    rows = []
    for co in cm["companies"]:
        e = cm["by_co"][co]
        top_item = (max(e["items"].items(), key=lambda kv: kv[1])[0]
                    if e["items"] else "-")
        rows.append("<tr><td>{}</td><td class='num'>{}</td>"
                    "<td class='num'>{}</td><td>{}</td></tr>".format(
                        esc(display_company(co)), e["draws"],
                        _cons_qty(e["qty"]), esc(top_item)))
    _pal = [ORANGE_HEX, V3_GOOD, V3_AMBER, V3_BLUE, "#B85CD6", "#5AC8FA",
            "#F97583", "#D6A35C"]
    _segs = sorted(((display_company(co), cm["by_co"][co]["qty"])
                    for co in cm["companies"]), key=lambda t: -t[1])
    _donut = ""
    if _segs:
        _s = [(k, v, _pal[i % len(_pal)]) for i, (k, v) in
              enumerate(_segs[:7]) if v > 0]
        _rest = sum(v for _k, v in _segs[7:])
        if _rest:
            _s.append(("Other", _rest, V3_STEEL))
        if _s:
            _donut = ("<h2>Where the consumables went</h2>"
                      + v3_donut(_s, total_label="ITEMS DRAWN"))
    co_html = (_donut + '<h2>Draws by company (alphabetical)</h2>'
               '<table class="data"><thead><tr><th>Company</th>'
               '<th class="num">Draws</th><th class="num">Items Issued</th>'
               '<th>Most-Drawn Item</th></tr></thead><tbody>{}</tbody>'
               '</table>').format("".join(rows))
    if cm["companies"]:
        mxd = max(cm["by_co"][co]["draws"] for co in cm["companies"]) or 1
        co_html += ("<div class='note' style='margin-top:10px'><b>The same "
                    "picture, drawn</b></div>"
                    + "".join(hbar(display_company(co),
                                   "{} draws".format(cm["by_co"][co]["draws"]),
                                   cm["by_co"][co]["draws"] / mxd)
                              for co in sorted(cm["companies"],
                                               key=lambda c: -cm["by_co"][c]["draws"])[:10]))
    # WHO drew - every draw lands on a named person (A. Fisher, 24 Jul 2026)
    ppl = {}
    for mv in cm.get("moves") or []:
        who = mv.get("hirer") or ""
        if not who:
            continue
        k = (who, mv["company"])
        e = ppl.setdefault(k, {"draws": 0, "qty": 0, "items": {}})
        e["draws"] += 1
        e["qty"] += mv["qty"]
        e["items"][mv["sku"]] = e["items"].get(mv["sku"], 0) + mv["qty"]
    person_html = ""
    if ppl:
        prows = ""
        for (who, co), e in sorted(ppl.items(), key=lambda kv: -kv[1]["draws"])[:18]:
            top_i = max(e["items"].items(), key=lambda kv: kv[1])[0] if e["items"] else "-"
            prows += ("<tr><td>{w}</td><td>{c}</td><td class='num'>{d}</td>"
                      "<td class='num'>{q}</td><td>{t}</td></tr>").format(
                w=esc(who), c=esc(display_company(co)), d=e["draws"],
                q=_cons_qty(e["qty"]), t=esc(top_i[:40]))
        person_html = ('<h2>Who drew what - by person</h2>'
                       "<div class='note'>Every draw is double-scanned to a "
                       "named person at the counter - the store always knows "
                       "who has what. The busiest drawers first.</div>"
                       '<table class="data"><thead><tr><th>Person</th>'
                       '<th>Company</th><th class="num">Draws</th>'
                       '<th class="num">Items</th><th>Most-Drawn Item</th>'
                       '</tr></thead><tbody>' + prows + '</tbody></table>')
    trows = "".join(
        "<tr><td>{}</td><td class='num'>{}</td><td class='num'>{}</td></tr>"
        .format(esc(sku), st["draws"], _cons_qty(st["qty"]))
        for sku, st in cm["top"])
    top_html = ('<h2>Working hardest - top consumables by draws</h2>'
                '<table class="data"><thead><tr><th>Consumable</th>'
                '<th class="num">Draws</th><th class="num">Qty Issued</th>'
                '</tr></thead><tbody>{}</tbody></table>').format(trows)
    sds = load_sds_register()
    shelf = (
        '<h2>Shelf assurance - stocked before crews notice</h2>'
        '<div class="note">{s} consumable lines are stocked on site right '
        'now, {m} of them under min/max management. {r} line{rp} at or '
        'below minimum {rv} flagged for restock - watched every day as part '
        'of the daily stocktake, so the shelf is ready before anyone has to '
        'ask. Nothing goes on the shelf unless it is ready for use. '
        'Life Saving Rules (SEQ-GL-009) apply - Rule 5, Tools and Equipment: '
        'nothing damaged or out of date is ever issued, consumables '
        'included.{sds}</div>'
    ).format(s=cm["stocked"], m=cm["managed"], r=cm["restock"],
             rp="" if cm["restock"] == 1 else "s",
             rv="is" if cm["restock"] == 1 else "are",
             sds=(" Every chemical line on this shelf is covered by the "
                  "Coates SDS register - {n} substances, each with a "
                  "current Safety Data Sheet and a QR link at the shelf "
                  "(full detail in the Safety &amp; High-Care Assurance "
                  "pack).".format(n=len(sds["items"])) if sds else ""))
    # bought-in items charged on (Charge Reporter register) - separate from
    # the store's own shelf stock, which stays quantities-only
    bought = [c for c in load_charge_feed() if c["line"] == "Consumables Sales"]
    if bought:
        bought_html = (
            "<h2>Bought-in &amp; charged on</h2>"
            "<div class='note'>Items purchased for the site and charged on - "
            "logged in the Coates charge register, separate from the store's "
            "own shelf stock above. These carry recorded charge amounts; the "
            "shelf itself stays quantities-only.</div>"
            + charges_rows_tbl(bought, show_company=True))
    else:
        bought_html = ""
    limits = (
        'A draw counts as one transaction regardless of quantity taken. '
        'Figures come from the SiteIQ SALES_STOCK export as at the '
        'data-as-at stamp. No consumables price feed is loaded - quantities '
        'only{exc}. '.format(
            exc=(', with one exception: the bought-in & charged-on section '
                 'shows recorded charge amounts from the Coates charge '
                 'register' if bought else
                 ', no dollar values anywhere on this report'))
        + ('' if sales_loaded else
           'The SALES_STOCK export was unavailable for this run - draw '
           'figures are unavailable, not zero. '))
    return (
        paged_report_doc(
            "Consumables Report", "Site-Wide - All Companies", asof,
            generated, source_line,
            "<div class='intro story'>" + story + "</div>"
            + glance + co_html + person_html + top_html + shelf + bought_html
            + "<div class='limits'><b>Honest limits:</b> {}</div>".format(limits),
            doc_title="Coates K2 Consumables Report"))


def email_consumables_html(cm, asof, generated, sales_loaded):
    inner = []
    if cm["draws"]:
        lead = ("The tool store keeps the site's consumables flowing - "
                "every draw double-scanned, every number a scanned fact. "
                "A draw counts as one transaction no matter the quantity, "
                "and consumables carry no charges on this report - "
                "quantities only.")
    else:
        lead = ("The shelf is ready before the rush: {s} consumable lines "
                "stocked on site, {m} under min/max management, checked "
                "daily. As crews start drawing, every draw is "
                "double-scanned and this report builds the usage story "
                "automatically - quantities only, never charges.".format(
                    s=cm["stocked"], m=cm["managed"]))
    inner.append(_e_note(esc(lead), "The story."))
    inner.append(_e_tiles([
        ("{:,}".format(cm["draws"]), "Draws"),
        (_cons_qty(cm["qty"]), "Items issued"),
        (str(len(cm["companies"])), "Companies"),
        (str(cm["people"]), "People"),
        (str(cm["item_types"]), "Types used"),
    ]))
    inner.append(_e_h2("Draws by company (alphabetical)"))
    rows = []
    for co in cm["companies"]:
        e = cm["by_co"][co]
        top_item = (max(e["items"].items(), key=lambda kv: kv[1])[0]
                    if e["items"] else "-")
        rows.append([display_company(co), str(e["draws"]),
                     _cons_qty(e["qty"]), top_item])
    inner.append(_e_tbl(["Company", "Draws", "Items Issued",
                         "Most-Drawn Item"], rows))
    inner.append(_e_h2("Working hardest - top consumables"))
    inner.append(_e_tbl(["Consumable", "Draws", "Qty Issued"],
                        [[sku, str(st["draws"]), _cons_qty(st["qty"])]
                         for sku, st in cm["top"]]))
    inner.append(_e_note(esc(
        "{s} consumable lines stocked on site, {m} under min/max "
        "management, {r} flagged for restock - watched daily as part of "
        "the stocktake so the shelf is ready before anyone asks. Life "
        "Saving Rules (SEQ-GL-009) apply - Rule 5: nothing damaged or out "
        "of date is ever issued, consumables included.".format(
            s=cm["stocked"], m=cm["managed"], r=cm["restock"])),
        "Shelf assurance."))
    meta = ('Generated: {g} | Data as at: {a} (SALES_STOCK export) | '
            'Author: Andrew Fisher').format(
        g=generated.strftime("%d %b %Y %H:%M") if generated else "-",
        a=asof.strftime("%d %b %Y %H:%M") if asof else "-")
    if cm["draws"]:
        hero = "{d:,} draws &bull; {c} compan{cs} &bull; {p} people".format(
            d=cm["draws"], c=len(cm["companies"]),
            cs="y" if len(cm["companies"]) == 1 else "ies", p=cm["people"])
    else:
        hero = ("{s} consumable lines stocked &bull; {m} min/max managed "
                "&bull; ready before the rush").format(
            s=cm["stocked"], m=cm["managed"])
    limits = ('A draw = one transaction regardless of quantity. Quantities '
              'only - no consumables pricing is loaded. Point-in-time '
              'snapshot of the SiteIQ export.')
    return _e_frame("COATES - K2 CONSUMABLES REPORT", "Cement Australia K2",
                    meta, hero, ''.join(inner), limits)


def generate_consumables(sales, sales_loaded, asof, generated, source_line,
                         date_tag):
    stock_lines, _ = load_sales_stock_levels()
    cm = build_consumables_site_model(sales, stock_lines)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stem = "Coates_K2_Consumables_Report_{}".format(date_tag)
    html_path = os.path.join(PAGES_DIR, stem + ".html")
    pdf_path = os.path.join(PDF_DIR, stem + ".pdf")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(render_consumables_html(cm, asof, generated, source_line,
                                        sales_loaded))
    pdf_ok = html_to_pdf(html_path, pdf_path)
    eml_path = os.path.join(EMAILS_DIR, stem + "_OUTLOOK.eml")
    subject = ("Coates K2 - Consumables Report - {} - {:,} draws"
               .format(asof.strftime("%d %b %Y") if asof else "",
                       cm["draws"]))
    attach_path = pdf_path if pdf_ok else html_path
    # the consumables report itself, pasted into the email body as images
    r_to, r_cc = recipients.resolve(KIT_DIR, "", "CONSUMABLES")
    imgs = email_images.pages_that_fit(
        email_images.capture_report_images(html_path, EMAILS_DIR, stem),
        "Consumables report")
    if imgs:
        msg = email_images.build_image_eml(subject, imgs, to=r_to, cc=r_cc)
        email_images.save_eml(msg, eml_path)
        cids = ["pg{}".format(i) for i in range(1, len(imgs) + 1)]
        write_draft_manifest(eml_path, subject, email_images.images_body(cids),
                             [], report="CONSUMABLES",
                             images=email_images.manifest_images(imgs))
        print("  Consumables report: {} draws | HTML{} | report in email "
              "body ({} images) | Outlook draft".format(
                  cm["draws"], " + PDF" if pdf_ok else " (PDF skipped)", len(imgs)))
        return
    msg = email.message.EmailMessage(policy=email.policy.SMTP)
    msg["Subject"] = subject
    if r_to:
        msg["To"] = email_images._addr_list(r_to)
    if r_cc:
        msg["Cc"] = email_images._addr_list(r_cc)
    msg["X-Unsent"] = "1"
    msg.set_content("This report is best viewed in HTML. The PDF is "
                    "attached.\n")
    body = email_consumables_html(cm, asof, generated or dt.datetime.now(),
                                  sales_loaded)
    msg.add_alternative(body, subtype="html")
    if os.path.isfile(attach_path):
        with open(attach_path, "rb") as f:
            data = f.read()
        if attach_path.lower().endswith(".pdf"):
            msg.add_attachment(data, maintype="application", subtype="pdf",
                               filename=os.path.basename(attach_path))
        else:
            msg.add_attachment(data, maintype="text", subtype="html",
                               filename=os.path.basename(attach_path))
    with open(eml_path, "wb") as f:
        f.write(msg.as_bytes())
    write_draft_manifest(eml_path, str(msg["Subject"]), body, [attach_path],
                         report="CONSUMABLES")
    print("  Consumables report: {} draws | HTML{} | Outlook draft".format(
        cm["draws"], " + PDF" if pdf_ok else " (PDF skipped)"))



# ---------------------------------------------------------------------------
# Supplementary report suite (Andrew, 19 Jul): Daily One-Pager, Safety &
# High-Care Assurance, Returns Performance, Weekly Roll-Up, Plant Right-Size,
# End-of-Shut Close-Out. One shared emitter; each report is a story builder.
# Client wording rule: the client shift day is 6:00am to 6:00am - always.
# ---------------------------------------------------------------------------

SHUTDOWN_START = dt.date(2026, 7, 12)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def load_tracking_rows():
    """Every Daily Tracking row (read-only) for trend/roll-up reports."""
    try:
        path = find_newest("Cement_Australia_Report*K2*.xlsm",
                           "the K2 report workbook")
    except KitError:
        return [], None
    try:
        _, raw = sheet_rows(path, "Daily Tracking", "the K2 workbook")
    except KitError:
        return [], None
    rows = []
    for r in raw:
        d = parse_date_au(r.get("Date"))
        if not d:
            continue
        rows.append({
            "date": d,
            "fc": to_num(r.get("Daily Forecast Total")),
            "ac": to_num(r.get("Daily Actual Total")),
            "labour_ac": to_num(r.get("Personnel / Labour Actual")),
            "accom_ac": to_num(r.get("Accommodation Actual")),
            "plant_ac": to_num(r.get("Plant Equipment Hire Actual")),
            "tool_ac": to_num(r.get("Tooling Hire Actual")),
            "radio_ac": to_num(r.get("Radio Hire Actual")),
            "gas_ac": to_num(r.get("Gas Monitor Hire Actual")),
        })
    rows.sort(key=lambda x: x["date"])
    return rows, path


def load_stocktake_daily():
    """Sightings per day from the STOCKTAKE export (counted units only)."""
    path = find_newest("STOCKTAKE*.xlsx", "the STOCKTAKE export",
                       required=False)
    if not path:
        return {}
    try:
        _, raw = sheet_rows(path, "STOCKTAKE", "the STOCKTAKE export")
    except KitError:
        return {}
    per_day = {}
    for r in raw:
        su = clean_text(r.get("STORAGE_UNIT")).upper()
        if not su or su in ("ON HIRE", "SHUTDOWN", "TURNAROUND",
                            "CEMENT SITE PLANT", "CEMENT BARRIERS",
                            "CEMENT RUBBISH CHUTES", "LAYDOWN"):
            continue
        d = parse_date_au(r.get("LAST_SIGHTED_DATE_TIME"))
        if d:
            per_day[d] = per_day.get(d, 0) + 1
    return per_day


#  EMAIL ONLY (Andrew, 28 Jul 2026): "only do the final Email production
#  of reports not the prep."
#
#  To be straight about what this can and cannot do: the email body IS
#  HTML, and every page image in it is a photograph of that HTML. There
#  is no email without it. What there IS, is a pile of files that exist
#  only to get there - the standalone page in Pages\, the print PDF, the
#  page-count sidecar. None of those are the email. Those go.
#
#  The HTML still has to touch disk for a moment, because the browser
#  that screenshots it and prints it can only open a file. So it is
#  written to a scratch folder and deleted the instant it has been used.
#  What is left in Reports\<today>\Emails is the email and nothing else.
EMAIL_ONLY = False


def _drop_prep(html_path, emails_dir, stem):
    """Bin the working files the moment the email exists. Only ever runs
    in email-only mode, and only ever touches this report's own scratch
    copy - never the Emails folder's .eml, .body.html, .draft.json or
    page images, which are what MAKE_OUTLOOK_DRAFTS builds the native
    Outlook draft from."""
    if not EMAIL_ONLY:
        return
    for p in (html_path,
              os.path.join(emails_dir, "{}_Email.pages".format(stem))):
        try:
            if p and os.path.isfile(p):
                os.remove(p)
        except OSError:
            pass
    #  and take the scratch folder with it once the last report is done
    try:
        os.rmdir(os.path.dirname(html_path))
    except OSError:
        pass          # still holding the next report's page - fine


def _oversight_cc():
    """The fixed oversight names CC'd on every company report email -
    read off the routing workbook's Summary sheet, the same block the
    daily packs read (add a person there, never here). Cached for the
    run; empty string if the workbook isn't on this machine - the email
    still goes, just without the CC."""
    if _oversight_cc._memo is None:
        got = []
        try:
            wb = openpyxl.load_workbook(
                os.path.join(KIT_DIR,
                             "K2_Daily_Email_Report_Allocation.xlsx"),
                read_only=True, data_only=True)
            if "Summary" in wb.sheetnames:
                rows = [[("" if c is None else str(c).strip()) for c in r]
                        for r in wb["Summary"].iter_rows(values_only=True)]
                grab = False
                for r in rows:
                    head = " ".join(r).upper()
                    if not grab and "FIXED" in head and "CC" in head:
                        grab = True
                        continue
                    if grab:
                        found = re.findall(r"[\w.+-]+@[\w.-]+\.\w+",
                                           " ".join(r))
                        if not found:
                            if got or any(r):
                                break
                            continue
                        for e in found:
                            if e.lower() not in [x.lower() for x in got]:
                                got.append(e)
            wb.close()
        except Exception:
            got = []
        _oversight_cc._memo = ", ".join(got)
    return _oversight_cc._memo


_oversight_cc._memo = None


def emit_report(stem, title, company_line, body_html, hero, inner_email,
                limits, subject, asof, generated, source_line,
                report_tag="", pdf_attach_only=False, cc_extra=""):
    """Shared emitter: PDF + framed email + native-draft manifest.

    pdf_attach_only is THE COMPANY EMAIL RULE (Andrew, 28 Jul 2026):
    "give them a story of what their full report is... say see attached
    full report... thats the only report they get." No page pictures at
    all - a friendly note in the body, the finished PDF on the
    paperclip. Cuts each company from minutes of browser captures to
    seconds, and the drafts open and send instantly."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    work = os.path.join(OUTPUT_DIR, "_work") if EMAIL_ONLY else PAGES_DIR
    os.makedirs(work, exist_ok=True)
    html_path = os.path.join(work, stem + ".html")
    pdf_path = os.path.join(PDF_DIR, stem + ".pdf")
    doc = paged_report_doc(
        title, company_line, asof, generated, source_line,
        body_html
        + v3_team_page()
        + "<div class='limits'><b>Honest limits:</b> {}</div>".format(limits),
        doc_title="Coates K2 - " + title)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(doc)
    #  The print PDF is a prep file when the report goes in the body as
    #  images - it isn't even attached. Only build it if it will be used.
    pdf_ok = False if EMAIL_ONLY else html_to_pdf(html_path, pdf_path)
    meta = ('Generated: {g} | Data as at: {a} | Author: Andrew '
            'Fisher').format(
        g=generated.strftime("%d %b %Y %H:%M") if generated else "-",
        a=asof.strftime("%d %b %Y %H:%M") if asof else "-")
    eml_path = os.path.join(EMAILS_DIR, stem + "_OUTLOOK.eml")
    attach_path = pdf_path if pdf_ok else html_path
    # The report itself, pasted into the email body as images in order -
    # same look as the Cost Tracking Snapshot email. If no browser is on
    # this machine, fall back to the framed summary body as before.
    r_to, r_cc = recipients.resolve(KIT_DIR, "", report_tag)
    if cc_extra:
        r_cc = ", ".join(x for x in (r_cc, cc_extra) if x)
    if pdf_attach_only:
        #  Story + PDF, never pictures. Make sure the PDF exists even in
        #  email-only mode, and sweep any page images an earlier run of
        #  the same day left behind so the pack builder can't pick a
        #  stale picture email over the new PDF one.
        if not pdf_ok:
            pdf_ok = html_to_pdf(html_path, pdf_path)
            attach_path = pdf_path if pdf_ok else html_path
        import glob as _glob
        for leftover in _glob.glob(os.path.join(EMAILS_DIR,
                                                stem + "_Email*")):
            try:
                os.remove(leftover)
            except OSError:
                pass
        inner_email = ([_e_intro(
            "G'day team &mdash; thanks for looking after the gear. "
            "Here's your position at a glance, and <b>your full report "
            "is attached as a PDF</b> &mdash; every number, and a page "
            "for each of your people. Reckon something on it has "
            "already come back? Reply or see us at the counter and "
            "we'll run the same-day double-check return so your list "
            "stays right.")]
            + list(inner_email)
            + [_e_note("Thanks for doing your bit &mdash; it keeps the "
                       "whole site moving. Care Deeply &middot; Customer "
                       "Focused &middot; Be Our Best &middot; One Team "
                       "&middot; Competitive Spirit",
                       "One team.")])
        imgs = []
    else:
        imgs = email_images.capture_report_images(html_path, EMAILS_DIR,
                                                  stem)
    #  Pages go in the body only when the whole message will actually
    #  arrive. DGH's 21-page report built a 17.3 MB draft - over what
    #  plenty of corporate gateways accept, so it bounces AFTER the send
    #  button. Too heavy = drop to the PDF-attached form below: same
    #  report, same recipients, always fits. (28 Jul 2026)
    if imgs:
        est_mb = email_images.email_weight_mb(imgs)
        if est_mb > email_images.MAX_EMAIL_MB:
            print("  {}: {} pages would make a {:.0f} MB email - over the "
                  "{:.0f} MB safe-send limit, so the PDF rides the "
                  "paperclip instead.".format(
                      title, len(imgs), est_mb, email_images.MAX_EMAIL_MB))
            imgs = []
    if imgs:
        msg = email_images.build_image_eml(subject, imgs, to=r_to, cc=r_cc)
        email_images.save_eml(msg, eml_path)
        cids = ["pg{}".format(i) for i in range(1, len(imgs) + 1)]
        write_draft_manifest(eml_path, subject,
                             email_images.images_body(cids), [],
                             report=report_tag,
                             images=email_images.manifest_images(imgs))
        _drop_prep(html_path, EMAILS_DIR, stem)
        print("  {}: {} | report in email body ({} images) | Outlook "
              "draft".format(
                  title,
                  "email only" if EMAIL_ONLY else
                  ("HTML + PDF" if pdf_ok else "HTML (PDF skipped)"),
                  len(imgs)))
        return
    #  No browser on this machine, so there are no page images and the
    #  email needs something attached. That is the one case where the
    #  PDF earns its place even in email-only mode.
    if EMAIL_ONLY and not pdf_ok:
        pdf_ok = html_to_pdf(html_path, pdf_path)
        attach_path = pdf_path if pdf_ok else html_path
    body = _e_frame("COATES - " + title.upper(), company_line, meta, hero,
                    ''.join(inner_email), limits)
    msg = email.message.EmailMessage(policy=email.policy.SMTP)
    msg["Subject"] = subject
    if r_to:
        msg["To"] = email_images._addr_list(r_to)
    if r_cc:
        msg["Cc"] = email_images._addr_list(r_cc)
    msg["X-Unsent"] = "1"
    msg.set_content("This report is best viewed in HTML. The PDF is "
                    "attached.\n")
    msg.add_alternative(body, subtype="html")
    if os.path.isfile(attach_path):
        with open(attach_path, "rb") as f:
            data = f.read()
        if attach_path.lower().endswith(".pdf"):
            msg.add_attachment(data, maintype="application", subtype="pdf",
                               filename=os.path.basename(attach_path))
        else:
            msg.add_attachment(data, maintype="text", subtype="html",
                               filename=os.path.basename(attach_path))
    with open(eml_path, "wb") as f:
        f.write(msg.as_bytes())
    write_draft_manifest(eml_path, subject, body, [attach_path],
                         report=report_tag)
    if not (EMAIL_ONLY and attach_path == html_path):
        _drop_prep(html_path, EMAILS_DIR, stem)   # keep it if it IS the attachment
    print("  {}: {} | Outlook draft".format(
        title,
        "email only" if EMAIL_ONLY else
        ("HTML + PDF" if pdf_ok else "HTML (PDF skipped)")))


def _tbl(headers, rows):
    h = "".join("<th{}>{}</th>".format(
        " class='num'" if i else "", esc(x)) for i, x in enumerate(headers))
    b = "".join("<tr>" + "".join(
        "<td{}>{}</td>".format(" class='num'" if i else "", c)
        for i, c in enumerate(r)) + "</tr>" for r in rows)
    return ("<table class='data'><thead><tr>{}</tr></thead>"
            "<tbody>{}</tbody></table>").format(h, b)


def tiles_html(cells):
    """One row of KPI tiles, the suite standard. cells = (value, label)
    or (value, label, status) with status 'g'/'a'/'r'. Colour ONLY where
    the number is a judgement (the look Andrew picked, 28 Jul 2026):
    red action, amber watch, green healthy - a plain fact stays a plain
    tile. Values go in as given (some carry entities or the neon money
    highlight); labels are escaped here."""
    tds = ""
    for c in cells:
        v, l = c[0], c[1]
        st = c[2] if len(c) > 2 and c[2] in ("g", "a", "r") else ""
        tds += ("<td{cls}><span class='v'>{v}</span>"
                "<span class='l'>{l}</span></td>").format(
            cls=" class='t{}'".format(st) if st else "", v=v, l=esc(l))
    return "<table class='tiles'><tr>" + tds + "</tr></table>"


def hbar(label, val_text, frac, colour="#F26222"):
    """One horizontal bar row - the suite's standard visual (same look as
    the Cost Tracking Snapshot's plan bars). frac is clamped 0..1; a zero
    still shows a sliver so the row reads as measured, not missing."""
    return ("<div class='hb'><div class='hl'>{l}</div>"
            "<div class='ht'><div class='hf' style='width:{w:.1f}%;"
            "background:{c}'></div></div>"
            "<div class='hv'>{v}</div></div>").format(
        l=esc(label), w=max(2.0, min(100.0, (frac or 0) * 100)), c=colour,
        v=val_text)


def ring(pct, label, sub="", colour=None, size=118):
    """A progress ring (conic-gradient) - the suite's eye-catcher for a
    single percentage. Prints faithfully in Edge/Chrome."""
    pct = max(0.0, min(1.0, pct or 0))
    c = colour or rag_colour(pct)
    return ("<div style='display:inline-block;text-align:center;margin:6px 14px'>"
            "<div style='width:{s}px;height:{s}px;border-radius:50%;"
            "background:conic-gradient({c} {deg:.0f}deg, #20262e 0deg);"
            "display:flex;align-items:center;justify-content:center'>"
            "<div style='width:{i}px;height:{i}px;border-radius:50%;background:#171B22;"
            "display:flex;flex-direction:column;align-items:center;justify-content:center'>"
            "<span style='color:#fff;font-size:{fs}px;font-weight:800'>{v:.0f}%</span>"
            "<span style='color:#8B9099;font-size:8px;text-transform:uppercase;"
            "letter-spacing:.5px'>{sub}</span></div></div>"
            "<div style='color:#A9B1BD;font-size:10.5px;margin-top:6px;max-width:{s}px'>{l}</div>"
            "</div>").format(s=size, i=size - 26, c=c, deg=pct * 360,
                             v=pct * 100, fs=size // 5, sub=esc(sub),
                             l=esc(label))


def stackbar(segments, height=26):
    """One 100% stacked bar: segments = [(label, value, colour)]. Values
    label inside when wide enough; zero-total shows an empty track."""
    total = sum(v for _l, v, _c in segments) or 1
    cells = ""
    for lab, v, c in segments:
        w = v / total * 100
        txt = ("{} {}".format(lab, v) if w >= 14 else "")
        cells += ("<div style='width:{w:.1f}%;background:{c};display:flex;"
                  "align-items:center;justify-content:center;color:{ink};"
                  "font-weight:700;font-size:10.5px;white-space:nowrap;"
                  "overflow:hidden'>{t}</div>").format(w=w, c=c, t=esc(txt),
                                                       ink=ink_on(c))
    return ("<div style='display:flex;border-radius:7px;overflow:hidden;"
            "height:{h}px;margin:8px 0'>{c}</div>").format(h=height, c=cells)



# ===========================================================================
# V3 VISUAL SYSTEM (24 Jul 2026, per Andrew's approved style boards):
# charcoal + Coates orange + RAG accents, icon KPI tiles with vs-yesterday
# deltas, health rings with sub-score checklists, donuts, SVG trend charts,
# peak-hour heat grids, contractor leaderboards, alert panels and the
# shutdown timeline with the TODAY marker. Shared by EVERY report so the
# whole family reads as one product. All inline styles/SVG - prints, PDFs
# and email images exactly as drawn, fully offline.
# ===========================================================================

ORANGE_HEX = COATES_ORANGE
V3_GOOD = "#3FB950"
V3_AMBER = "#fab219"
V3_BAD = "#F85149"
V3_STEEL = "#8B9099"
V3_BLUE = "#4C8DD6"
V3_CARD = "#171B22"
V3_LINE = "#2A313C"
#  Ink for V3 text that sits ON THE PAGE (not on a dark card). White ink
#  on the old dark pages; crisp dark ink on the executive white ones -
#  the donut legends and score checklists were white-on-white without it.
#  (Found by the computed-style sweep, 26 Jul 2026.)
V3_PAGE_INK = INK_STRONG
V3_PAGE_BODY = "#33393F" if getattr(_RP, "EXEC_LIGHT", False) else "#D7DBE2"


def ink_on(colour):
    """Dark ink on light/amber/steel fills, white on everything else -
    white-on-amber was a squint everywhere it appeared."""
    return ("#14181F" if str(colour).upper() in
            ("#FAB219", "#F2B01E", "#EFFF3D", "#8B9099", "#3FB950")
            else "#fff")

_ICON_PATHS = {
    "wrench": "M14.7 6.3a4.5 4.5 0 0 0-6 5.6L3 17.6 6.4 21l5.7-5.7a4.5 4.5 0 0 0 5.6-6l-2.8 2.8-2.5-.6-.6-2.5z",
    "people": "M9 11a3.5 3.5 0 1 0-3.5-3.5A3.5 3.5 0 0 0 9 11zm7 1a3 3 0 1 0-3-3 3 3 0 0 0 3 3zM2.5 19c0-2.8 3.6-4.2 6.5-4.2s6.5 1.4 6.5 4.2V20h-13zm14.6 1H21v-.8c0-1.9-2-3.2-4.2-3.6a5.6 5.6 0 0 1 .3 4.4z",
    "swap": "M7 7h10l-2.5-2.5L16 3l5 5-5 5-1.5-1.5L17 9H7zm10 8H7l2.5 2.5L8 19l-5-5 5-5 1.5 1.5L7 13h10z",
    "radio": "M6 3h2v4h8a2 2 0 0 1 2 2v10a2 2 0 0 1-2 2H8a2 2 0 0 1-2-2zm3 7a3 3 0 1 0 3 3 3 3 0 0 0-3-3zm6 0h2v2h-2zm0 4h2v2h-2z",
    "gas": "M12 2s5 5.5 5 9.5a5 5 0 0 1-10 0C7 7.5 12 2 12 2zm0 18a7 7 0 0 0 7-7h2a9 9 0 0 1-18 0h2a7 7 0 0 0 7 7z",
    "battery": "M16 4h-2V2h-4v2H8a2 2 0 0 0-2 2v14a2 2 0 0 0 2 2h8a2 2 0 0 0 2-2V6a2 2 0 0 0-2-2zm-3 14h-2v-3H9l4-6v3h2z",
    "shield": "M12 2l8 3v6c0 5-3.4 9.3-8 11-4.6-1.7-8-6-8-11V5zm-1 13l6-6-1.4-1.4L11 12.2l-2-2L7.6 11.6z",
    "clock": ("M12 2a10 10 0 1 0 10 10A10 10 0 0 0 12 2zm0 2a8 8 0 1 1"
              "-8 8 8 8 0 0 1 8-8zm1 3h-2v6l5 3 1-1.7-4-2.3z"),
    "box": "M12 2l9 4.5v11L12 22l-9-4.5v-11zm0 2.2L5.7 7 12 9.8 18.3 7zM5 8.6v7l6 3v-7zm14 0l-6 3v7l6-3z",
    "truck": "M3 5h11v4h3.5L21 12.5V17h-2a2.5 2.5 0 0 1-5 0h-3a2.5 2.5 0 0 1-5 0H3zm13.5 8H20l-1.8-2H16.5z",
    "chart": "M4 20V10h3v10zm6.5 0V4h3v16zM17 20v-7h3v7z",
    "alert": "M12 2L1 21h22zm1 14h-2v2h2zm0-7h-2v5h2z",
    "tick": "M9.6 17.4L4.8 12.6l1.7-1.7 3.1 3.1 7.9-7.9 1.7 1.7z",
    "store": "M4 4h16v3l-1 4h-1v9H6v-9H5L4 7zm4 9v5h3v-5zm5 0v5h3v-5z",
    "dollar": "M12 2a10 10 0 1 0 10 10A10 10 0 0 0 12 2zm1 15.5V19h-2v-1.5a4 4 0 0 1-2.8-2l1.7-1a2.6 2.6 0 0 0 2.3 1.3c.9 0 1.7-.4 1.7-1.2s-.9-1.1-2-1.4c-1.6-.5-3.4-1-3.4-3a2.9 2.9 0 0 1 2.5-2.8V6h2v1.4a3.6 3.6 0 0 1 2.5 1.8l-1.6 1a2 2 0 0 0-1.9-1.1c-.9 0-1.5.4-1.5 1.1s.9 1 2.1 1.4c1.6.5 3.3 1.1 3.3 3.1a3 3 0 0 1-2.9 2.8z",
}


def pill(status):
    """RAG status pill - the mock's little coloured badges."""
    s = str(status).upper()
    if s in ("PASS", "GREEN", "CURRENT", "OK", "READY", "EXCELLENT", "GOOD",
             "MATCHES", "ALL HOME"):
        bg, fg = "#14301b", "#7fe09a"
    elif s.startswith("CHECK") or s in ("WATCH", "PENDING", "PROVISIONAL",
                                        "FOLLOW UP", "AMBER", "VERIFY",
                                        "STRONG"):
        bg, fg = "#33280c", "#f3d47f"
    elif s.startswith("FAIL") or s in ("CRITICAL", "ESCALATE", "RED",
                                       "MISSING", "OVERDUE", "ACTION"):
        bg, fg = "#3a1a17", "#f8a89c"
    else:
        bg, fg = "#232A34", "#c9cfd8"
    return ("<span style='display:inline-block;padding:2px 10px;"
            "border-radius:10px;font-size:8.5pt;font-weight:bold;"
            "background:{bg};color:{fg}'>{t}</span>").format(
        bg=bg, fg=fg, t=esc(status))


def v3_company_kpis(m):
    """v3 icon KPI strip for a company pack - counts only, client-safe."""
    late = sum(1 for r in m["rows"] if r["days"] is not None and r["days"] >= 4)
    oldest = max([r["days"] for r in m["rows"] if r["days"] is not None] or [0])
    expo = m.get("repl_exposure")
    expo_h = ("<span class='repl'>{}</span>".format(fmt_money(expo))
              if expo is not None else None)
    hc = m.get("high_care") or {}
    hc_n = sum(len(v) for v in hc.values()) if isinstance(hc, dict) else 0
    return v3_kpis([
        ("box", "{:,}".format(m["n_items"]), "Items on hire now", ""),
        ("people", "{:,}".format(len(m["people"])), "People holding gear", ""),
        ("alert", "{:,}".format(late), "Out 4 days or more",
         ("<span style='color:{c};font-size:8pt;font-weight:700'>oldest {o}d"
          "</span>").format(c=V3_BAD if late else V3_GOOD, o=oldest)),
        ("shield", expo_h if expo_h is not None else "TBC",
         "Replacement exposure", ""),
        ("swap", "{:,}".format(m.get("n_tx_total") or 0),
         "Transactions this shut", ""),
        ("wrench", "{:,}".format(m.get("item_types") or 0),
         "Item types used", ""),
        ("tick", "{:,}".format(m.get("same_day") or 0),
         "Same-day returns", ""),
        ("battery", "{:,}".format(hc_n), "High-care items out", ""),
    ])


#  (name, role, what they actually do for you - Andrew's words,
#  26 Jul 2026). Emails follow the Coates standard: first.last@coates.com.au
V3_TEAM_DAY = [
    ("Thomas Maher", "Dayshift Lead",
     "Runs the day-to-day of days - your first call on day shift"),
    ("Helen O'Grady", "Dayshift Tool Store Controller",
     "Expertise, solutions and help at the counter"),
    ("Lucas Armstrong", "Day Shift Mechanic",
     "Looks after every breakdown on site - report it, he's on it"),
]
V3_TEAM_NIGHT = [
    ("Cody Mitchell", "Nightshift Tool Store Controller",
     "Runs the nights - your first call after dark"),
    ("Adam Boyce", "Night Shift Tool Store Controller",
     "Expertise, solutions and help at the counter"),
]


def team_email(name):
    """first.last@coates.com.au off the roster name - apostrophes and
    middle names collapse the way the mail system does it."""
    parts = re.sub(r"[^A-Za-z ]", "", name).lower().split()
    return (parts[0] + "." + parts[-1] + "@coates.com.au") if parts else ""


def asset_disp(asset):
    """Asset number, with its allocated Plant ID beside it wherever one
    exists in the master file (A. Fisher, 25 Jul 2026 - IDs on show)."""
    a = clean_text(asset)
    if not a:
        return "-"
    pid = MASTER.plant_id(a)
    if pid:
        return ("<span style='white-space:nowrap'>{a} "
                "<span style='background:{o};color:#fff;border-radius:9px;"
                "padding:1px 8px;font-weight:800;font-size:9pt'>ID {p}"
                "</span></span>").format(a=esc(a), o=ORANGE_HEX, p=esc(pid))
    return esc(a)


def desc_disp(desc, asset=None, sub=None):
    """A description cell with its compliance line underneath it.

    The item name, then - only where the master file says so - the blue
    tag, the daily pre-start and logbook line, and any return-daily
    requirement. This is the drop-in replacement for a bare
    esc(r["description"]) anywhere a hire line is rendered, so the
    obligation travels with the gear onto every report, every email and
    My Gear alike. Nothing is guessed: a blank column means a blank line.
    (A. Fisher, 25 Jul 2026)

    `sub` is the small grey line some tables already show under the name
    (the storage unit, usually) - passed through so nothing is lost.
    """
    out = esc(desc) or "-"
    if sub:
        out += ("<br><span style='color:#8B9099;font-size:8.5pt'>{}</span>"
                .format(sub))
    return out + EC.badges_html(asset, desc)


def compliance_block(rows, item_key="asset", desc_key="description"):
    """The counts band a report opens with - what is out there right now
    that has to be tagged, pre-started or brought back today."""
    try:
        return EC.summary_band_html(EC.summarise(rows, item_key, desc_key))
    except Exception:
        return ""


def _v3_person(name, role, duty="", boss=False):
    """A proper CONTACT CARD (Andrew, 26 Jul 2026): highlighter border
    like the store cards, role in the highlighter colour, their Coates
    email, and one line on what they actually do for the crews."""
    parts = name.replace("O'", "O").split()
    ini = (parts[0][:1] + (parts[-1][:1] if len(parts) > 1 else "")).upper()
    return ("<td style='background:{card};border:1px solid {line};"
            "border-left:4px solid #EFFF3D;"
            "border-top:3px solid {top};border-radius:10px;padding:12px 8px;"
            "text-align:center'>"
            "<div style='width:44px;height:44px;border-radius:50%;margin:0 auto;"
            "background:{av};display:flex;align-items:center;justify-content:center;"
            "color:#fff;font-weight:800;font-size:15px'>{i}</div>"
            "<div style='color:#fff;font-weight:700;font-size:10pt;"
            "margin-top:6px'>{n}</div>"
            "<div style='color:#EFFF3D;font-size:8pt;text-transform:uppercase;"
            "letter-spacing:.4px;margin-top:2px;font-weight:800'>{r}</div>"
            "<div style='color:#9fd0ff;font-size:7.5pt;margin-top:4px;"
            "line-height:1.4'>{e}</div>"
            + ("<div style='color:{m};font-size:7.5pt;line-height:1.45;"
               "margin-top:5px'>{d}</div>" if duty else "")
            + "</td>").format(
        card=V3_CARD, line=V3_LINE, top=ORANGE_HEX if boss else V3_LINE,
        av=ORANGE_HEX if boss else "#2b3442", i=ini, n=esc(name), r=esc(role),
        #  The address stacks at the @ so five cards across never bleed
        #  into each other. (Caught on the 26 Jul 2026 visual check.)
        e=esc(team_email(name)).replace("@", "<br>@"), d=esc(duty),
        m="#c3cad4")


def v3_team_strip():
    """Meet the team, on EVERY page (A. Fisher, 25 Jul 2026). The full
    team page still rides at the back of each pack; this is the running
    footer so a client never reads a page without knowing who is behind
    the counter."""
    day = " &middot; ".join("<b>{}</b> {}".format(esc(n), esc(r))
                            for n, r, _d in V3_TEAM_DAY)
    night = " &middot; ".join("<b>{}</b> {}".format(esc(n), esc(r))
                              for n, r, _d in V3_TEAM_NIGHT)
    return ("<div class='tl'>Your Coates tool store team</div>"
            "<div class='tn'><b>Andrew Fisher</b> Shutdown Manager QLD &amp; "
            "NT &nbsp;|&nbsp; <span style='color:#F26222;font-weight:700'>"
            "DAY</span> {d} &nbsp;|&nbsp; <span style='color:#F26222;"
            "font-weight:700'>NIGHT</span> {n} &nbsp;|&nbsp; "
            "<span style='color:#8B9099;font-weight:800;letter-spacing:1px'>"
            "MY GEAR HQ</span></div>").format(d=day, n=night)


def v3_team_page():
    """Meet the team - on EVERY report (A. Fisher, 25 Jul 2026). The crew
    running the K2 toolstore, day and night, with the three promises."""
    row1 = ("<table style='width:44%;margin:0 auto;border-collapse:separate;"
            "border-spacing:8px 0'><tr>"
            + _v3_person("Andrew Fisher", "Shutdown Manager",
                         "Oversees the whole store and the whole shutdown - "
                         "anything at all, start here", boss=True)
            + "</tr></table>")
    cells = "".join(_v3_person(n, r, d) for n, r, d in
                    V3_TEAM_DAY + V3_TEAM_NIGHT)
    row2 = ("<table style='width:100%;border-collapse:separate;"
            "border-spacing:8px 0;table-layout:fixed;margin-top:8px'><tr>"
            + cells + "</tr></table>")
    chips = ("<div style='text-align:center;margin-top:10px'>"
             + "".join("<span style='display:inline-block;background:{card};"
                       "border:1px solid {line};border-radius:14px;"
                       "padding:5px 12px;margin:3px 4px;font-size:9pt;"
                       "color:#D7DBE2'><b style='color:{g}'>&#10003;</b> {t}"
                       "</span>".format(card=V3_CARD, line=V3_LINE, g=V3_GOOD,
                                        t=t)
                       for t in ("The right gear to the right crew",
                                 "Looked after, maintained &amp; ready",
                                 "Checked and right - every time")) + "</div>")
    return ("<h2>Meet the tool store team</h2>"
            "<div class='note'>The crew running your K2 tool store - getting "
            "the right gear to the right crew, keeping it looked after and "
            "ready, and making sure everything that leaves the store is "
            "right. Day and night, so the shutdown keeps moving. Something "
            "not on hand or not right? Tell us and we'll sort it.</div>"
            + row1 + row2 + chips)


def v3_icon(name, size=22, colour=ORANGE_HEX):
    p = _ICON_PATHS.get(name, _ICON_PATHS["box"])
    return ("<svg width='{s}' height='{s}' viewBox='0 0 24 24' "
            "style='vertical-align:middle'><path d='{p}' fill='{c}' "
            "fill-rule='evenodd'/></svg>").format(s=size, p=p, c=colour)


def v3_delta(cur, prev, unit="", invert=False, label="vs yesterday",
             neutral=False):
    """The mock's little arrow: green when the move is good, red when bad.
    invert=True for measures where UP is bad (outstanding, overdue)."""
    if prev is None:
        return "<span style='color:{m};font-size:8pt'>&mdash;</span>".format(m=V3_STEEL)
    try:
        d = cur - prev
    except Exception:
        return ""
    if abs(d) < 0.0005:
        return ("<span style='color:{m};font-size:8pt'>&#9644; no change</span>"
                ).format(m=V3_STEEL)
    up = d > 0
    good = (not up) if invert else up
    col = V3_STEEL if neutral else (V3_GOOD if good else V3_BAD)
    arrow = "&#9650;" if up else "&#9660;"
    mag = ("{:,.0f}".format(abs(d)) if abs(d) >= 1 or d == int(d)
           else "{:,.2f}".format(abs(d)))
    return ("<span style='color:{c};font-size:8pt;font-weight:700'>{a} {m}{u} "
            "<span style='color:{s};font-weight:400'>{l}</span></span>").format(
        c=col, a=arrow, m=mag, u=esc(unit), s=V3_STEEL, l=esc(label))


def v3_kpis(items, per_row=4):
    """Icon KPI tiles - the mock's headline row. items = [(icon, value,
    label, delta_html_or_empty)]. Renders per_row to a table row."""
    out = ""
    for i in range(0, len(items), per_row):
        chunk = items[i:i + per_row]
        tds = ""
        for ic, val, lab, dl in chunk:
            tds += ("<td style='background:{card};border:1px solid {line};"
                    "border-radius:10px;padding:12px 8px;text-align:center;"
                    "vertical-align:top'>"
                    "<div>{icon}</div>"
                    "<div style='font-size:17pt;font-weight:800;color:#fff;"
                    "margin-top:2px'>{v}</div>"
                    "<div style='font-size:7.5pt;color:{mut};text-transform:"
                    "uppercase;letter-spacing:.5px;margin-top:2px'>{l}</div>"
                    "<div style='margin-top:3px'>{d}</div></td>").format(
                card=V3_CARD, line=V3_LINE, icon=v3_icon(ic, 22),
                v=val, mut=V3_STEEL, l=esc(lab), d=dl or "")
        out += ("<table style='width:100%;border-collapse:separate;"
                "border-spacing:8px 0;table-layout:fixed;margin-top:10px'>"
                "<tr>" + tds + "</tr></table>")
    return out


def v3_donut(segments, title="", total_label="TOTAL", size=150):
    """Multi-segment donut (conic-gradient) with a legend - the mock's
    on-hire wheel. segments = [(label, value, colour)]."""
    total = sum(v for _l, v, _c in segments)
    if total <= 0:
        return ""
    stops, acc = [], 0.0
    for _l, v, c in segments:
        a0 = acc / total * 360
        acc += v
        a1 = acc / total * 360
        stops.append("{c} {a0:.1f}deg {a1:.1f}deg".format(c=c, a0=a0, a1=a1))
    legend = "".join(
        ("<div style='margin:3px 0;font-size:9.5pt;color:" + V3_PAGE_BODY +
         "'>"
         "<span style='display:inline-block;width:10px;height:10px;"
         "border-radius:2px;background:{c};margin-right:7px'></span>"
         "{l} &nbsp;<b style='color:" + V3_PAGE_INK + "'>{v:,.0f}</b>"
         "<span style='color:{m}'> &nbsp;{p:.0f}%</span></div>").format(
            c=c, l=esc(l), v=v, m=V3_STEEL, p=v / total * 100)
        for l, v, c in segments)
    return ("<table style='width:100%;border-collapse:collapse'><tr>"
            "<td style='width:{w}px;padding:6px'>"
            "<div style='width:{s}px;height:{s}px;border-radius:50%;"
            "background:conic-gradient({stops});display:flex;align-items:center;"
            "justify-content:center'>"
            "<div style='width:{i}px;height:{i}px;border-radius:50%;"
            "background:{card};display:flex;flex-direction:column;"
            "align-items:center;justify-content:center'>"
            "<span style='color:#fff;font-size:16pt;font-weight:800'>{t:,.0f}</span>"
            "<span style='color:{m};font-size:7pt;text-transform:uppercase'>{tl}</span>"
            "</div></div></td>"
            "<td style='vertical-align:middle;padding-left:14px'>{ttl}{leg}</td>"
            "</tr></table>").format(
        w=size + 12, s=size, stops=",".join(stops), i=size - 40, card=V3_CARD,
        t=total, m=V3_STEEL, tl=esc(total_label),
        ttl=("<div style='font-size:10pt;font-weight:700;color:" +
             V3_PAGE_INK + ";margin-bottom:4px'>" + esc(title) +
             "</div>") if title else "",
        leg=legend)


def v3_trend(series, width=700, height=150, colour=ORANGE_HEX, unit="",
             fill=True):
    """SVG line chart with dots + end label - the mock's 7-day trends.
    series = [(label, value)] in order. Pure SVG, prints perfectly."""
    if not series:
        return ""
    vals = [v for _l, v in series]
    vmax = max(vals) or 1
    vmin = min(0, min(vals))
    pad_l, pad_r, pad_t, pad_b = 8, 52, 12, 22
    W, H = width, height
    span = (vmax - vmin) or 1
    n = len(series)
    def xy(i, v):
        x = pad_l + (W - pad_l - pad_r) * (i / max(1, n - 1))
        y = pad_t + (H - pad_t - pad_b) * (1 - (v - vmin) / span)
        return x, y
    pts = [xy(i, v) for i, (_l, v) in enumerate(series)]
    line = " ".join("{:.1f},{:.1f}".format(x, y) for x, y in pts)
    area = ""
    if fill:
        area = ("<polygon points='{p0x:.1f},{base:.1f} {line} {p1x:.1f},{base:.1f}' "
                "fill='{c}' opacity='0.12'/>").format(
            p0x=pts[0][0], p1x=pts[-1][0], base=H - pad_b, line=line, c=colour)
    dots = "".join("<circle cx='{:.1f}' cy='{:.1f}' r='3.2' fill='{}'/>".format(
        x, y, colour) for x, y in pts)
    labels = ""
    step = max(1, n // 8)
    for i, (lab, _v) in enumerate(series):
        if i % step and i != n - 1:
            continue
        x, _ = pts[i]
        labels += ("<text x='{x:.1f}' y='{y}' fill='{m}' font-size='8.5' "
                   "text-anchor='middle'>{l}</text>").format(
            x=x, y=H - 6, m=V3_STEEL, l=esc(lab))
    endx, endy = pts[-1]
    endval = ("<text x='{x:.1f}' y='{y:.1f}' fill='#fff' font-size='13' "
              "font-weight='700'>{v:,.0f}{u}</text>").format(
        x=min(endx + 8, W - pad_r + 4), y=max(endy + 4, 14),
        v=vals[-1], u=esc(unit))
    grid = "".join("<line x1='{pl}' y1='{y:.1f}' x2='{xr}' y2='{y:.1f}' "
                   "stroke='{lc}' stroke-width='0.6'/>".format(
        pl=pad_l, xr=W - pad_r, y=pad_t + (H - pad_t - pad_b) * f, lc=V3_LINE)
        for f in (0, 0.5, 1))
    return ("<svg width='100%' viewBox='0 0 {W} {H}' style='background:{card};"
            "border:1px solid {lc};border-radius:10px;margin:6px 0'>{grid}{area}"
            "<polyline points='{line}' fill='none' stroke='{c}' "
            "stroke-width='2.4' stroke-linejoin='round'/>{dots}{labels}{endval}"
            "</svg>").format(W=W, H=H, card=V3_CARD, lc=V3_LINE, grid=grid,
                             area=area, line=line, c=colour, dots=dots,
                             labels=labels, endval=endval)


def v3_bars_pair(series_a, series_b, name_a, name_b, col_a=None, col_b=None,
                 width=700, height=170):
    """Grouped daily bars - issues v returns per day. series = [(label, v)]."""
    col_a = col_a or ORANGE_HEX
    col_b = col_b or V3_GOOD
    n = len(series_a)
    if not n:
        return ""
    vmax = max([v for _l, v in series_a] + [v for _l, v in series_b] + [1])
    pad_l, pad_t, pad_b = 8, 14, 24
    W, H = width, height
    plot_h = H - pad_t - pad_b
    group_w = (W - pad_l * 2) / n
    bar_w = min(22.0, group_w / 2.6)
    bars, labels = "", ""
    for i in range(n):
        lab, va = series_a[i]
        _lb, vb = series_b[i]
        gx = pad_l + i * group_w + group_w / 2
        for off, v, c in ((-bar_w - 1, va, col_a), (1, vb, col_b)):
            h = plot_h * (v / vmax)
            bars += ("<rect x='{x:.1f}' y='{y:.1f}' width='{w:.1f}' "
                     "height='{h:.1f}' rx='2' fill='{c}'/>").format(
                x=gx + off, y=pad_t + plot_h - h, w=bar_w, h=max(h, 1.5), c=c)
            if v > 0:
                bars += ("<text x='{x:.1f}' y='{y:.1f}' fill='#fff' "
                         "font-size='8' text-anchor='middle'>{v:,.0f}</text>"
                         ).format(x=gx + off + bar_w / 2,
                                  y=max(pad_t + plot_h - h - 3, 10), v=v)
        labels += ("<text x='{x:.1f}' y='{y}' fill='{m}' font-size='8.5' "
                   "text-anchor='middle'>{l}</text>").format(
            x=gx, y=H - 8, m=V3_STEEL, l=esc(lab))
    legend = ("<rect x='{x1}' y='4' width='9' height='9' rx='2' fill='{ca}'/>"
              "<text x='{x1t}' y='12' fill='#D7DBE2' font-size='9'>{na}</text>"
              "<rect x='{x2}' y='4' width='9' height='9' rx='2' fill='{cb}'/>"
              "<text x='{x2t}' y='12' fill='#D7DBE2' font-size='9'>{nb}</text>"
              ).format(x1=W - 190, x1t=W - 177, ca=col_a, na=esc(name_a),
                       x2=W - 110, x2t=W - 97, cb=col_b, nb=esc(name_b))
    return ("<svg width='100%' viewBox='0 0 {W} {H}' style='background:{card};"
            "border:1px solid {lc};border-radius:10px;margin:6px 0'>"
            "{bars}{labels}{legend}</svg>").format(
        W=W, H=H, card=V3_CARD, lc=V3_LINE, bars=bars, labels=labels,
        legend=legend)


def v3_heat(hour_counts, title="Peak times - when the counter runs hottest"):
    """Hour-of-day heat strip from real transaction times. hour_counts =
    {hour: n}. Green -> amber -> red like the mock."""
    if not hour_counts:
        return ""
    hours = list(range(4, 19))
    mx = max(hour_counts.get(h, 0) for h in hours) or 1
    cells = ""
    for h in hours:
        v = hour_counts.get(h, 0)
        f = v / mx
        c = ("#20262e" if v == 0 else V3_GOOD if f < 0.45
             else V3_AMBER if f < 0.8 else V3_BAD)
        cells += ("<td style='padding:2px'><div style='background:{c};"
                  "border-radius:4px;height:34px;display:flex;align-items:center;"
                  "justify-content:center;color:{tc};font-size:8.5pt;"
                  "font-weight:700'>{v}</div>"
                  "<div style='text-align:center;color:{m};font-size:7pt;"
                  "margin-top:2px'>{h:02d}</div></td>").format(
            c=c, tc=ink_on(c) if v else V3_STEEL, v=v or "", m=V3_STEEL,
            h=h)
    return ("<div style='background:{card};border:1px solid {lc};"
            "border-radius:10px;padding:10px 12px;margin:8px 0'>"
            "<div style='font-size:9.5pt;font-weight:700;color:#fff;"
            "margin-bottom:4px'>{t}</div>"
            "<table style='width:100%;border-collapse:collapse;"
            "table-layout:fixed'><tr>{c}</tr></table>"
            "<div style='font-size:8pt;color:{m};margin-top:4px'>"
            "<span style='color:{g}'>&#9632;</span> steady &nbsp;"
            "<span style='color:{a}'>&#9632;</span> busy &nbsp;"
            "<span style='color:{r}'>&#9632;</span> peak &nbsp;&mdash; "
            "issue + return events by hour, whole shutdown to date</div></div>"
            ).format(card=V3_CARD, lc=V3_LINE, t=esc(title), c=cells,
                     m=V3_STEEL, g=V3_GOOD, a=V3_AMBER, r=V3_BAD)


def v3_scorebar(score, width=110):
    f = max(0.0, min(1.0, score / 100.0))
    c = V3_GOOD if score >= 90 else V3_AMBER if score >= 70 else V3_BAD
    return ("<span style='display:inline-block;width:{w}px;background:#20262e;"
            "border-radius:5px;overflow:hidden;vertical-align:middle'>"
            "<span style='display:block;width:{p:.0f}%;background:{c};"
            "height:9px'></span></span>").format(w=width, p=f * 100, c=c)


def v3_alerts(alerts):
    """The mock's alerts panel. alerts = [(sev 'red'|'amber'|'green'|'blue',
    headline, detail)]."""
    if not alerts:
        return ""
    dots = {"red": V3_BAD, "amber": V3_AMBER, "green": V3_GOOD, "blue": V3_BLUE}
    rows = "".join(
        ("<div style='margin:7px 0;font-size:9.5pt;line-height:1.45'>"
         "<span style='color:{c};font-size:11pt'>&#9679;</span> "
         "<b style='color:#fff'>{h}</b>"
         "<div style='color:{m};margin-left:16px'>{d}</div></div>").format(
            c=dots.get(sev, V3_STEEL), h=esc(head), m=V3_STEEL, d=esc(det))
        for sev, head, det in alerts)
    return ("<div style='background:{card};border:1px solid {lc};"
            "border-left:4px solid {o};border-radius:10px;"
            "padding:10px 14px;margin:8px 0'>"
            "<div style='font-size:10pt;font-weight:700;color:{o};"
            "text-transform:uppercase;letter-spacing:.5px'>Alerts &amp; "
            "actions</div>{r}</div>").format(card=V3_CARD, lc=V3_LINE,
                                             o=ORANGE_HEX, r=rows)


def v3_timeline(today, roster_curve=None, gear_curve=None, width=708,
                height=190):
    """The shutdown timeline: phase bands (mobilisation / shutdown / demob),
    the manpower and gear-out curves, peaks flagged, TODAY marker."""
    d0 = dt.date(2026, 7, 12)
    d1 = dt.date(2026, 8, 15)
    total = (d1 - d0).days or 1
    W, H = width, height
    pad_t, pad_b = 30, 30
    plot_h = H - pad_t - pad_b
    def X(d):
        return (d - d0).days / total * (W - 16) + 8
    phases = [("MOBILISATION", d0, dt.date(2026, 7, 19), "#2b3442"),
              ("SHUTDOWN - 23 DAYS", dt.date(2026, 7, 20), dt.date(2026, 8, 11),
               "#33261a"),
              ("DEMOB", dt.date(2026, 8, 12), d1, "#243328")]
    bands = "".join(
        ("<rect x='{x0:.1f}' y='{t}' width='{w:.1f}' height='{h}' fill='{c}' "
         "rx='4'/><text x='{xm:.1f}' y='{ty}' fill='{m}' font-size='8' "
         "text-anchor='middle' letter-spacing='1'>{n}</text>").format(
            x0=X(p0), t=pad_t, w=X(p1) - X(p0) + (W - 16) / total, h=plot_h,
            c=c, xm=(X(p0) + X(p1)) / 2, ty=pad_t - 6, m=V3_STEEL, n=n)
        for n, p0, p1, c in phases)
    def curve(series, colour, label):
        if not series:
            return ""
        vmax = max(v for _d, v in series) or 1
        pts = ["{:.1f},{:.1f}".format(X(d), pad_t + plot_h * (1 - v / vmax))
               for d, v in series]
        peak_d, peak_v = max(series, key=lambda t: t[1])
        px, py = X(peak_d), pad_t + plot_h * (1 - peak_v / vmax)
        return ("<polyline points='{p}' fill='none' stroke='{c}' "
                "stroke-width='2'/>"
                "<circle cx='{px:.1f}' cy='{py:.1f}' r='3.4' fill='{c}'/>"
                "<text x='{px:.1f}' y='{ly:.1f}' fill='{c}' font-size='8' "
                "text-anchor='middle'>peak {l} {d}</text>").format(
            p=" ".join(pts), c=colour, px=px, py=py,
            ly=max(py - 7, pad_t + 8), l=label, d=peak_d.strftime("%d %b"))
    curves = (curve(roster_curve or [], V3_BLUE, "crew")
              + curve(gear_curve or [], ORANGE_HEX, "gear"))
    tx = X(today)
    marker = ("<line x1='{x:.1f}' y1='{t}' x2='{x:.1f}' y2='{b}' "
              "stroke='#fff' stroke-width='2' stroke-dasharray='5,3'/>"
              "<rect x='{bx:.1f}' y='{by}' width='58' height='16' rx='8' "
              "fill='#fff'/><text x='{cx:.1f}' y='{ty}' fill='#0E1116' "
              "font-size='9' font-weight='700' text-anchor='middle'>TODAY</text>"
              ).format(x=tx, t=pad_t - 2, b=pad_t + plot_h + 2,
                       bx=min(max(tx - 29, 2), W - 60), by=pad_t + plot_h + 6,
                       cx=min(max(tx, 31), W - 31), ty=pad_t + plot_h + 18)
    ticks = ""
    for d in (dt.date(2026, 7, 12), dt.date(2026, 7, 20), dt.date(2026, 8, 1),
              dt.date(2026, 8, 11), dt.date(2026, 8, 15)):
        ticks += ("<text x='{x:.1f}' y='{y}' fill='{m}' font-size='8' "
                  "text-anchor='middle'>{l}</text>").format(
            x=X(d), y=H - 4, m=V3_STEEL, l=d.strftime("%d %b"))
    legend = ("<circle cx='{x1}' cy='11' r='4' fill='{b}'/>"
              "<text x='{x1t}' y='14' fill='#D7DBE2' font-size='9'>Coates crew "
              "(rostered hours)</text>"
              "<circle cx='{x2}' cy='11' r='4' fill='{o}'/>"
              "<text x='{x2t}' y='14' fill='#D7DBE2' font-size='9'>Gear on hire"
              "</text>").format(x1=W - 300, x1t=W - 292, b=V3_BLUE,
                                x2=W - 130, x2t=W - 122, o=ORANGE_HEX)
    return ("<svg width='100%' viewBox='0 0 {W} {H}' style='background:{card};"
            "border:1px solid {lc};border-radius:10px;margin:8px 0'>"
            "{bands}{curves}{marker}{ticks}{legend}</svg>").format(
        W=W, H=H, card=V3_CARD, lc=V3_LINE, bands=bands, curves=curves,
        marker=marker, ticks=ticks, legend=legend)



# ---- V3 ANALYTICS: real stories computed from the exports ----------------

def tx_analytics(tx, today):
    """Everything the transactions history can honestly tell: daily issue /
    return series, yesterday's movement, people served, hold times, same-day
    rate, peak hours. Times come straight off SiteIQ's stamps."""
    if not tx:
        return None
    days = {}
    hours = {}
    holds = []
    same_day = 0
    completed = 0
    d_lo = min((t["start"] for t in tx if t["start"]), default=today)
    for t in tx:
        s, e = t.get("start"), t.get("end")
        if s:
            d = days.setdefault(s, {"iss": 0, "ret": 0, "people": set()})
            d["iss"] += 1
            d["people"].add(t["hirer"])
            tm = str(t.get("start_time") or "")
            m = re.match(r"\s*(\d{1,2}):(\d{2})\s*([AP]M)?", tm, re.I)
            if m:
                h = int(m.group(1)) % 12
                if (m.group(3) or "").upper() == "PM":
                    h += 12
                hours[h] = hours.get(h, 0) + 1
        if e:
            d = days.setdefault(e, {"iss": 0, "ret": 0, "people": set()})
            d["ret"] += 1
            d["people"].add(t["hirer"])
            tm = str(t.get("end_time") or "")
            m = re.match(r"\s*(\d{1,2}):(\d{2})\s*([AP]M)?", tm, re.I)
            if m:
                h = int(m.group(1)) % 12
                if (m.group(3) or "").upper() == "PM":
                    h += 12
                hours[h] = hours.get(h, 0) + 1
        if s and e:
            completed += 1
            holds.append((e - s).days)
            if e == s:
                same_day += 1
    yday = today - dt.timedelta(days=1)
    y = days.get(yday, {"iss": 0, "ret": 0, "people": set()})
    t0 = days.get(today, {"iss": 0, "ret": 0, "people": set()})
    dby = days.get(yday - dt.timedelta(days=1), {"iss": 0, "ret": 0,
                                                 "people": set()})
    span = [d_lo + dt.timedelta(days=i)
            for i in range((min(today, max(days) if days else today) - d_lo).days + 1)]
    series_iss = [(d.strftime("%d %b"), days.get(d, {}).get("iss", 0)) for d in span]
    series_ret = [(d.strftime("%d %b"), days.get(d, {}).get("ret", 0)) for d in span]
    served = [(d.strftime("%d %b"), len(days.get(d, {}).get("people", set())))
              for d in span]
    net, acc = [], 0
    for d in span:
        acc += days.get(d, {}).get("iss", 0) - days.get(d, {}).get("ret", 0)
        net.append((d, acc))
    hold3 = sum(1 for h in holds if h <= 3)
    return {
        "hold3_pct": (hold3 / completed) if completed else None,
        "days": days, "span": span,
        "iss_series": series_iss, "ret_series": series_ret,
        "served_series": served, "net_curve": net,
        "hours": hours,
        "yesterday": {"iss": y["iss"], "ret": y["ret"],
                      "people": len(y["people"])},
        "day_before": {"iss": dby["iss"], "ret": dby["ret"],
                       "people": len(dby["people"])},
        "today": {"iss": t0["iss"], "ret": t0["ret"],
                  "people": len(t0["people"])},
        "same_day": same_day, "completed": completed,
        "same_day_pct": (same_day / completed) if completed else None,
        "avg_hold": (sum(holds) / len(holds)) if holds else None,
        "people_total": len({t["hirer"] for t in tx if t["hirer"]}),
        "window": (d_lo, max(days) if days else today),
    }


def company_return_stats(tx, models, today):
    """Per-company accountability from the transaction history: issued,
    returned, outstanding NOW (register truth), return %, average hold,
    same-day %, late (4d+) count."""
    per = {}
    for t in tx:
        comp = t.get("company") or ""
        if not comp or is_site_plant_equipment(t.get("hirer", "")):
            continue
        c = per.setdefault(comp, {"iss": 0, "ret": 0, "holds": [],
                                  "same": 0, "done": 0})
        if t.get("start"):
            c["iss"] += 1
        if t.get("end"):
            c["ret"] += 1
        if t.get("start") and t.get("end"):
            c["done"] += 1
            c["holds"].append((t["end"] - t["start"]).days)
            if t["end"] == t["start"]:
                c["same"] += 1
    out = []
    for disp_key, m in models.items():
        comp = m["display"]
        c = per.get(comp, {"iss": 0, "ret": 0, "holds": [], "same": 0,
                           "done": 0})
        outstanding = m["n_items"]
        late = sum(1 for r in m["rows"]
                   if r["days"] is not None and r["days"] >= 4)
        issued_total = max(c["iss"], c["ret"] + outstanding)
        ret_pct = (c["ret"] / issued_total) if issued_total else None
        out.append({
            "company": comp, "issued": issued_total, "returned": c["ret"],
            "outstanding": outstanding, "ret_pct": ret_pct,
            "same_pct": (c["same"] / c["done"]) if c["done"] else None,
            "avg_hold": (sum(c["holds"]) / len(c["holds"])) if c["holds"] else None,
            "late": late,
        })
    return out


def contractor_scores(stats, charges_by_company):
    """The accountability score, 0-100, formula printed on the page:
    return performance 40 + timeliness 25 + damage record 20 + hold time 15.
    Only ever computed from what actually happened."""
    out = []
    for s in stats:
        ret = (s["ret_pct"] or 0.0) * 40.0
        late_pen = min(25.0, 5.0 * s["late"])
        timely = 25.0 - late_pen
        dmg_n = charges_by_company.get(s["company"], 0)
        damage = max(0.0, 20.0 - 10.0 * dmg_n)
        if s["avg_hold"] is None:
            hold = 15.0 if s["outstanding"] == 0 else 7.5
        else:
            hold = 15.0 * max(0.0, min(1.0, 1.0 - max(0.0, s["avg_hold"] - 1) / 6.0))
        score = max(0.0, min(100.0, ret + timely + damage + hold))
        band = ("EXCELLENT" if score >= 90 else "GOOD" if score >= 75
                else "WATCH" if score >= 55 else "ACTION")
        out.append({**s, "score": score, "band": band, "damages": dmg_n,
                    "parts": {"return": ret, "timely": timely,
                              "damage": damage, "hold": hold}})
    out.sort(key=lambda x: -x["score"])
    return out


def health_scorecard(x, txa, scores, stocktake_pct, stockouts_big,
                     damage_n, gas_overdue):
    """The client health score - six measures, every formula stated and
    calibrated to what each thing IS: gas monitors are the safety-critical
    daily item; radios and Milwaukee ride the returns discipline; a
    stockout only counts when a genuinely stocked line (fleet of 3+) has
    every unit out; damages counts damage charges, not consumable sales."""
    safety = max(0.0, 100.0 - 25.0 * gas_overdue)
    compliance = (stocktake_pct or 0.0) * 100.0
    if txa and txa["same_day_pct"] is not None:
        returns = (txa["same_day_pct"] * 50.0
                   + (txa["hold3_pct"] or 0.0) * 50.0)
    else:
        returns = None
    availability = max(0.0, 100.0 - 6.0 * stockouts_big)
    total_out = x["n_items"] or 1
    late_items = sum(s["late"] for s in scores)
    outstanding = max(0.0, 100.0 - (late_items / total_out) * 100.0)
    damages = max(0.0, 100.0 - 15.0 * damage_n)
    parts = [("Safety", safety,
              "100 less 25 per gas monitor out overnight - bump testing is "
              "daily ({} out today)".format(gas_overdue)),
             ("Compliance", compliance,
              "stocktake coverage - share of store items physically sighted "
              "inside the cycle"),
             ("Returns", returns,
              "half same-day rate, half back-inside-3-days rate, across {} "
              "completed hire cycles".format(txa["completed"] if txa else 0)),
             ("Tool availability", availability,
              "100 less 6 per stocked line (fleet of 3+) with every unit out "
              "({} today)".format(stockouts_big)),
             ("Outstanding", outstanding,
              "share of on-hire gear inside the 3-day bar ({} of {} items "
              "4d+)".format(late_items, total_out)),
             ("Damages", damages,
              "100 less 15 per damage charge this shut ({} logged)".format(
                  damage_n))]
    known = [(n, v, w) for n, v, w in parts if v is not None]
    overall = sum(v for _n, v, _w in known) / len(known) if known else 0.0
    verdict = ("EXCELLENT" if overall >= 90 else "STRONG" if overall >= 80
               else "WATCH" if overall >= 65 else "ACTION")
    return {"overall": overall, "verdict": verdict, "parts": parts}


def load_roster_curve():
    """Coates crew rostered hours per day from the K2 workbook's costing
    model (read-only). Optional-safe: no workbook, no curve."""
    try:
        import glob as _g
        c = (_g.glob(os.path.join(KIT_DIR, "Cement_Australia_Report*K2*.xlsm"))
             or [])
        if not c:
            return []
        path = max(c, key=os.path.getmtime)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Shutdown Costing"]
        per = {}
        for row in ws.iter_rows(min_row=16, max_row=190, values_only=True):
            d, hrs = row[0], row[8]
            if isinstance(d, dt.datetime):
                d = d.date()
            if isinstance(d, dt.date):
                try:
                    per[d] = per.get(d, 0.0) + float(hrs or 0)
                except Exception:
                    pass
        wb.close()
        return sorted(per.items())
    except Exception:
        return []


def rag_colour(frac, good=0.95, warn=0.80):
    """Standard RAG colour for a 0..1 measure."""
    return ("#3FB950" if frac >= good else
            "#fab219" if frac >= warn else "#F85149")


# ---- CHARGE REPORTER FEED: logged charges surface across the reports -----
# One register (Coates_K2_Charge_Reporter\CHARGE_REGISTER.xlsx) feeds every
# report - log a charge once and it shows up wherever it belongs: the
# company's own report, the exec summary, the daily brief, the consumables
# report and My Gear. Never summed with SiteIQ invoiced figures - when a
# charge reaches DAILY_SUMMARY it becomes the invoiced truth and these
# stay the register view. (A. Fisher, 24 Jul 2026)
_CHARGE_FEED = None


def load_charge_feed():
    """Costing Feed rows from the Charge Reporter register. Empty-safe."""
    global _CHARGE_FEED
    if _CHARGE_FEED is not None:
        return _CHARGE_FEED
    rows = []
    path = os.path.join(KIT_DIR, "Coates_K2_Charge_Reporter",
                        "CHARGE_REGISTER.xlsx")
    if os.path.isfile(path):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if "Costing Feed" in wb.sheetnames:
                feed = list(wb["Costing Feed"].iter_rows(values_only=True))
                hdr = {str(v).strip(): i for i, v in enumerate(feed[0]) if v}
                for r in feed[1:]:
                    def g(name):
                        i = hdr.get(name)
                        return r[i] if i is not None and i < len(r) else None
                    if not g("Charge ID"):
                        continue
                    #  The Date cell arrives as a real Excel date OR as
                    #  text, depending on what wrote the row - 50_ADD_
                    #  LIFTING_CHARGES wrote text on 28 Jul 2026 and the
                    #  mixed sort crashed the whole morning build
                    #  (str < date). Parse whatever shape turns up;
                    #  junk parses to None and sorts first, never crashes.
                    when = parse_date_au(g("Date"))
                    try:
                        amt = float(g("Amount") or 0)
                    except (TypeError, ValueError):
                        amt = 0.0
                    status = clean_text(g("Status")) or "For approval"
                    rows.append({
                        "when": when, "cid": clean_text(g("Charge ID")),
                        "category": clean_text(g("Category")),
                        "line": clean_text(g("Cost Line")),
                        "company": clean_text(g("Company")),
                        "desc": clean_text(g("Description")),
                        "amount": amt, "status": status,
                        "firm": status.lower() in ("approved", "invoiced")})
            wb.close()
        except Exception:
            rows = []          # unreadable register = no charges shown, never a crash
    rows.sort(key=lambda x: (x["when"] or dt.date.min, x["cid"]))
    _CHARGE_FEED = rows
    return rows


def charges_for_company(display_name):
    """Feed rows recorded against this company (alias-tolerant)."""
    want = canonical_company(display_name)
    return [c for c in load_charge_feed()
            if canonical_company(c["company"]) == want]


def load_damage_people():
    """Damage register rows keyed for person matching (My Gear): each row
    (person_raw, item, date, amount, status, cid, company)."""
    out = []
    path = os.path.join(KIT_DIR, "Coates_K2_Charge_Reporter",
                        "CHARGE_REGISTER.xlsx")
    if not os.path.isfile(path):
        return out
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Damage - Breakdown" in wb.sheetnames:
            rows = list(wb["Damage - Breakdown"].iter_rows(values_only=True))
            hdr = {str(v).strip(): i for i, v in enumerate(rows[0]) if v}
            for r in rows[1:]:
                def g(name):
                    i = hdr.get(name)
                    return r[i] if i is not None and i < len(r) else None
                if not g("Charge ID"):
                    continue
                when = g("Date")
                if isinstance(when, dt.datetime):
                    when = when.date()
                try:
                    amt = float(g("Approved recoverable cost") or 0)
                except (TypeError, ValueError):
                    amt = 0.0
                out.append({"person": clean_text(g("Person using / returning item")),
                            "item": clean_text(g("Item description")),
                            "when": when, "amount": amt,
                            "status": clean_text(g("Cost status")),
                            "cid": clean_text(g("Charge ID")),
                            "company": clean_text(g("Client / company"))})
        wb.close()
    except Exception:
        return []
    return out


def charges_rows_tbl(rows, show_company=False):
    """The standard charges table used everywhere the feed appears."""
    heads = ["Date", "Charge", "What"] + (["Company"] if show_company else []) \
        + ["Amount", "Status"]
    body = []
    for c in rows:
        cells = [c["when"].strftime("%d %b") if c["when"] else "-",
                 esc(c["category"]), esc(c["desc"])]
        if show_company:
            cells.append(esc(c["company"]))
        cells += [fmt_money(c["amount"]), esc(c["status"])]
        body.append(cells)
    return _tbl(heads, body)


def company_charges_html(m):
    """Client-facing charges section for one company's report - shown only
    when that company has charges recorded. Dollar figures here are the
    recorded charge amounts (damage recovery / items charged on), never
    hire rates."""
    rows = m.get("charges") or []
    if not rows:
        return ""
    total = sum(c["amount"] for c in rows if c["firm"])
    return (
        "<h2>Charges recorded this shutdown</h2>"
        "<div class='intro'>Recorded against {co} in the Coates charge "
        "register - each has its own advice document with the full story "
        "and evidence. Firm (approved) charges total <b>{t}</b>. Anything "
        "you believe is already resolved, flag it and we'll put it through "
        "the review process same day.</div>".format(
            co=esc(m["display"]), t=fmt_money(total))
        + charges_rows_tbl(rows))


# ---- CEMENT STORE: stock on hand, what's used, what's low, what we need ---

def load_store_shelf():
    """The store's shelf straight from SALES_STOCK: stock rows (no company
    attached) keyed by item, with on-hand / minimum / SiteIQ reorder qty.
    Movement rows (company draws) are load_sales()'s job."""
    path = find_newest("SALES_STOCK*.xlsx", "the SALES_STOCK export",
                       required=False)
    if not path:
        return {}, None
    try:
        _, raw = sheet_rows(path, "SALES_STOCK", "the SALES_STOCK export")
    except KitError:
        return {}, None
    shelf = {}
    for r in raw:
        if clean_text(r.get("SKU_NUMBER")) in DEAD_SKUS:
            continue
        desc = clean_text(r.get("SKU_DESCRIPTION"))
        if not desc:
            continue
        if clean_text(r.get("COMPANY_NAME")):
            continue                      # a draw, not a shelf line
        avail = to_num(r.get("AVAILABLE_QUANTITY")) or 0
        minq = to_num(r.get("MINIMUM_QUANTITY")) or 0
        reorder = to_num(r.get("REORDER_QUANTITY")) or 0
        if avail == 0 and minq == 0 and reorder == 0 \
                and not clean_text(r.get("SALES_DATE")):
            continue                      # all-zero placeholder row
        e = shelf.setdefault(desc, {"avail": 0.0, "min": 0.0, "reorder": 0.0})
        e["avail"] += avail
        e["min"] = max(e["min"], minq)
        e["reorder"] = max(e["reorder"], reorder)
    return shelf, path


def build_store_model(shelf, sales):
    """One row per consumable line: shelf position + scanned usage.
    Status is fact-based: OUT (nothing on hand where it matters), LOW
    (at or under the minimum), OK. Nothing estimated."""
    used = {}
    for m in sales:
        u = used.setdefault(m["sku"], {"draws": 0, "qty": 0.0,
                                       "cos": set(), "last": None})
        u["draws"] += 1
        u["qty"] += m["qty"]
        if m["company"]:
            u["cos"].add(display_company(m["company"]))
        if m["date"] and (u["last"] is None or m["date"] > u["last"]):
            u["last"] = m["date"]
    rows = []
    for desc in sorted(set(shelf) | set(used), key=str.upper):
        s = shelf.get(desc, {"avail": 0.0, "min": 0.0, "reorder": 0.0})
        u = used.get(desc, {"draws": 0, "qty": 0.0, "cos": set(), "last": None})
        if s["min"] > 0 and s["avail"] <= 0:
            status = "OUT"
        elif s["avail"] <= 0 and u["qty"] > 0:
            status = "OUT"                # used dry (no min set on the line)
        elif s["min"] > 0 and s["avail"] <= s["min"]:
            status = "LOW"
        else:
            status = "OK"
        rows.append({"desc": desc, "avail": s["avail"], "min": s["min"],
                     "reorder": s["reorder"], "draws": u["draws"],
                     "used": u["qty"], "cos": u["cos"], "last": u["last"],
                     "status": status,
                     "topup": max(0.0, s["min"] - s["avail"]) if s["min"] > 0 else 0.0})
    need = [r for r in rows if r["status"] in ("OUT", "LOW")]
    need.sort(key=lambda r: (0 if r["status"] == "OUT" else 1,
                             -(r["used"]), r["desc"].upper()))
    return {"rows": rows, "need": need,
            "stocked": sum(1 for r in rows if r["avail"] > 0),
            "managed": sum(1 for r in rows if r["min"] > 0),
            "low": sum(1 for r in rows if r["status"] == "LOW"),
            "out": sum(1 for r in rows if r["status"] == "OUT"),
            "draws": sum(r["draws"] for r in rows),
            "used_qty": sum(r["used"] for r in rows)}


def _qty(v):
    return "{:g}".format(v)


def generate_store_report(sales, sales_loaded, asof, generated, source_line,
                          date_tag):
    shelf, shelf_path = load_store_shelf()
    sm = build_store_model(shelf, sales)
    st_cls = {"OK": "g", "LOW": "a", "OUT": "r"}

    story = (
        "This one is for the Cement Australia store team - your consumables "
        "shelf as the tool store sees it, straight from scanned draws and the "
        "live SiteIQ stock position. <b>{stocked} lines are stocked</b> on the "
        "shelf right now, {managed} of them under min/max management. Crews "
        "have made <b>{draws:,} draws</b> ({qty} items issued) this shutdown - "
        "every one through the counter and the double scan. "
        "<b>{need} line{ns} {vb} attention</b>: {out} out, {low} running at or "
        "under minimum - the reorder list below runs worst-first so the top "
        "of the list is today's order.").format(
        stocked=sm["stocked"], managed=sm["managed"], draws=sm["draws"],
        qty=_qty(sm["used_qty"]), need=len(sm["need"]),
        ns="" if len(sm["need"]) == 1 else "s",
        vb="needs" if len(sm["need"]) == 1 else "need",
        out=sm["out"], low=sm["low"])
    body = "<div class='intro story'>" + story + "</div>"
    # v3: shelf KPIs + reorder alerts, straight off the store model
    body += v3_kpis([
        ("store", "{:,}".format(sm["stocked"]), "Lines on the shelf", ""),
        ("box", "{:,}".format(sm["managed"]), "Under min/max management", ""),
        ("alert", "{:,}".format(sm["low"] + sm["out"]), "Low or out",
         "<span style='color:{c};font-size:8pt;font-weight:700'>{t}</span>".format(
             c=V3_BAD if (sm["low"] + sm["out"]) else V3_GOOD,
             t="reorder worst-first below" if (sm["low"] + sm["out"])
             else "shelf holding")),
        ("swap", "{:,}".format(sm["draws"]), "Draws through the counter", ""),
    ])

    tiles = [(str(sm["stocked"]), "Lines Stocked"),
             (str(sm["managed"]), "Min/Max Managed"),
             (str(sm["out"]), "Out Of Stock", "r" if sm["out"] else "g"),
             (str(sm["low"]), "At / Under Min", "a" if sm["low"] else "g"),
             ("{:,}".format(sm["draws"]), "Draws This Shut"),
             (_qty(sm["used_qty"]), "Items Issued")]
    body += tiles_html(tiles)

    # -- what we need: the reorder list, worst first -----------------------
    if sm["need"]:
        rows = "".join(
            "<tr><td>{d}</td><td class='num'>{a}</td><td class='num'>{m}</td>"
            "<td class='num'>{t}</td><td class='num'>{r}</td>"
            "<td class='num'>{u}</td><td class='{c}' style='text-align:center'>{s}</td></tr>".format(
                d=esc(r["desc"]), a=_qty(r["avail"]), m=_qty(r["min"]),
                t=_qty(r["topup"]) if r["topup"] > 0 else "-",
                r=_qty(r["reorder"]) if r["reorder"] > 0 else "-",
                u=_qty(r["used"]), c=st_cls[r["status"]], s=r["status"])
            for r in sm["need"])
        body += ("<h2>What we need - reorder list (worst first)</h2>"
                 "<table class='data'><thead><tr><th>Item</th>"
                 "<th class='num'>On Hand</th><th class='num'>Min</th>"
                 "<th class='num'>To Get Back To Min</th>"
                 "<th class='num'>SiteIQ Reorder Qty</th>"
                 "<th class='num'>Used This Shut</th><th>Status</th></tr></thead>"
                 "<tbody>" + rows + "</tbody></table>"
                 "<div class='note'>'To get back to min' is simple arithmetic "
                 "(minimum minus on hand) - the call on how much to order is "
                 "yours. A dash in SiteIQ Reorder Qty means no reorder "
                 "quantity is set on that line in SiteIQ.</div>"
                 + "<div class='note' style='margin-top:10px'><b>Shelf v "
                   "minimum, drawn</b> - how far under the line each one "
                   "sits.</div>"
                 + "".join(hbar(r["desc"][:42],
                                "{} of min {}".format(_qty(r["avail"]),
                                                      _qty(r["min"])),
                                (r["avail"] / r["min"]) if r["min"] else 0,
                                rag_colour((r["avail"] / r["min"])
                                           if r["min"] else 0,
                                           good=1.0, warn=0.5))
                           for r in sm["need"][:10]))
    else:
        body += ("<h2>What we need - reorder list</h2>"
                 "<div class='intro'>Nothing is out or under minimum right "
                 "now - the shelf is holding. The full list below shows the "
                 "position line by line.</div>")

    # -- what's been used --------------------------------------------------
    used_rows = sorted([r for r in sm["rows"] if r["used"] > 0],
                       key=lambda r: (-r["used"], r["desc"].upper()))
    if used_rows:
        rows = "".join(
            "<tr><td>{d}</td><td class='num'>{dr}</td><td class='num'>{q}</td>"
            "<td class='num'>{a}</td><td>{co}</td></tr>".format(
                d=esc(r["desc"]), dr=r["draws"], q=_qty(r["used"]),
                a=_qty(r["avail"]),
                co=esc(", ".join(sorted(r["cos"])[:3])
                       + (" +{}".format(len(r["cos"]) - 3) if len(r["cos"]) > 3 else "")))
            for r in used_rows)
        body += ("<h2>What's been used (most first)</h2>"
                 "<table class='data'><thead><tr><th>Item</th>"
                 "<th class='num'>Draws</th><th class='num'>Qty Issued</th>"
                 "<th class='num'>On Hand Now</th><th>Drawn By</th></tr></thead>"
                 "<tbody>" + rows + "</tbody></table>")
    else:
        body += ("<h2>What's been used</h2>"
                 "<div class='note'>No consumable draws recorded in SiteIQ "
                 "yet this shutdown.</div>")

    # -- the full shelf ----------------------------------------------------
    shelf_rows = sorted([r for r in sm["rows"] if r["avail"] > 0 or r["min"] > 0],
                        key=lambda r: r["desc"].upper())
    if shelf_rows:
        rows = "".join(
            "<tr><td>{d}</td><td class='num'>{a}</td><td class='num'>{m}</td>"
            "<td class='{c}' style='text-align:center'>{s}</td></tr>".format(
                d=esc(r["desc"]), a=_qty(r["avail"]), m=_qty(r["min"]),
                c=st_cls[r["status"]], s=r["status"])
            for r in shelf_rows)
        body += ("<h2>The full shelf - every stocked line</h2>"
                 "<table class='data'><thead><tr><th>Item</th>"
                 "<th class='num'>On Hand</th><th class='num'>Min</th>"
                 "<th>Status</th></tr></thead><tbody>" + rows + "</tbody></table>")

    hero = ("{s} lines stocked &bull; {o} out &bull; {l} at/under min &bull; "
            "{d:,} draws this shut").format(s=sm["stocked"], o=sm["out"],
                                            l=sm["low"], d=sm["draws"])
    inner = [_e_note(esc("Shelf position and usage for the Cement store - "
                         "the reorder list runs worst-first."), "Store stock."),
             _e_h2("Needs attention"),
             _e_tbl(["Item", "On Hand", "Min", "Status"],
                    [[r["desc"], _qty(r["avail"]), _qty(r["min"]), r["status"]]
                     for r in sm["need"][:12]] or [["-", "-", "-", "OK"]])]
    limits = ("Quantities only - no consumables pricing is loaded, so no "
              "dollar values appear. Shelf position and usage are from the "
              "SALES_STOCK export at the data-as-at stamp; draws are scanned "
              "facts. 'To get back to min' is minimum minus on hand, never a "
              "forecast. Lines with no minimum set can't be flagged LOW - set "
              "minimums in SiteIQ to put them under management."
              + ("" if sales_loaded else " NOTE: no SALES_STOCK export was "
                 "found - this run shows structure only."))
    emit_report("Coates_K2_Cement_Store_Stock_{}".format(date_tag),
                "Cement Store - Stock & Reorder", "Cement Australia Store Team",
                body, hero, inner, limits,
                "Coates K2 - Cement Store Stock & Reorder - {} - {} line{} "
                "need attention".format(
                    asof.strftime("%d %b %Y") if asof else "",
                    len(sm["need"]), "" if len(sm["need"]) == 1 else "s"),
                asof, generated, source_line, report_tag="STORE")


# ---- CONSUMABLE REQUESTS: what crews asked for at the counter -------------

REQUEST_STATUSES = ["New", "Ordered", "Arrived", "Declined"]


def load_requests():
    """The counter's request log - CONSUMABLE_REQUESTS.xlsx, filled in by
    the tool store as crews ask for things. Tolerant reader: any row with
    an item counts; unknown status reads as New; the example row is
    skipped. Missing file = no requests, plainly said, never an error."""
    path = find_newest("CONSUMABLE_REQUESTS*.xlsx", "the requests register",
                       required=False)
    if not path:
        return [], None
    reqs = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        hdr = None
        for row in ws.iter_rows(values_only=True):
            vals = [clean_text(v) for v in row]
            if hdr is None:
                if vals and vals[0].lower() == "date":
                    hdr = {v.lower(): i for i, v in enumerate(vals)}
                continue

            def cell(name):
                i = hdr.get(name)
                return vals[i] if i is not None and i < len(vals) else ""

            item = cell("item requested") or cell("item")
            if not item:
                continue
            status = cell("status").title() or "New"
            if status.lower() == "example":
                continue
            if status not in REQUEST_STATUSES:
                status = "New"
            reqs.append({"date": parse_date_au(cell("date")),
                         "who": cell("requested by"),
                         "company": cell("company"),
                         "item": item,
                         "qty": cell("qty"),
                         "status": status,
                         "notes": cell("notes")})
        wb.close()
    except Exception:
        return [], path
    return reqs, path


def generate_requests_report(asof, generated, source_line, date_tag):
    reqs, req_path = load_requests()
    by = {s: [r for r in reqs if r["status"] == s] for s in REQUEST_STATUSES}

    if reqs:
        story = (
            "What crews have asked the tool store for over and above the "
            "shelf - logged at the counter as it happens, so the store sees "
            "real demand, not guesswork. <b>{n} request{ns}</b> so far: "
            "<b>{new} open</b> needing a call, {ordered} on order, {arrived} "
            "arrived and closed out{dec}. Open requests lead - they're "
            "today's conversation.").format(
            n=len(reqs), ns="" if len(reqs) == 1 else "s",
            new=len(by["New"]), ordered=len(by["Ordered"]),
            arrived=len(by["Arrived"]),
            dec=", {} declined".format(len(by["Declined"])) if by["Declined"] else "")
    else:
        story = (
            "What crews ask the tool store for over and above the shelf - "
            "logged at the counter in CONSUMABLE_REQUESTS.xlsx as it "
            "happens. Nothing is logged yet. When someone asks for "
            "something we don't stock, jot it in the register (item, who, "
            "company, qty) and it appears here automatically - so the "
            "store sees real demand, not guesswork.")
    body = "<div class='intro story'>" + story + "</div>"

    tiles = [(str(len(by["New"])), "Open - Action Needed",
              "a" if by["New"] else "g"),
             (str(len(by["Ordered"])), "On Order"),
             (str(len(by["Arrived"])), "Arrived / Closed"),
             (str(len(reqs)), "Total Requests")]
    body += tiles_html(tiles)

    def req_table(rows):
        return ("<table class='data'><thead><tr><th>Date</th><th>Item</th>"
                "<th class='num'>Qty</th><th>Requested By</th><th>Company</th>"
                "<th>Notes</th></tr></thead><tbody>" + "".join(
                    "<tr><td>{d}</td><td>{i}</td><td class='num'>{q}</td>"
                    "<td>{w}</td><td>{c}</td><td>{n}</td></tr>".format(
                        d=r["date"].strftime("%d %b") if r["date"] else "-",
                        i=esc(r["item"]), q=esc(r["qty"] or "-"),
                        w=esc(r["who"] or "-"), c=esc(r["company"] or "-"),
                        n=esc(r["notes"] or "")) for r in rows)
                + "</tbody></table>")

    if by["New"]:
        body += ("<h2>Open requests - action needed</h2>"
                 + req_table(sorted(by["New"],
                                    key=lambda r: r["date"] or dt.date.min)))
    if by["Ordered"]:
        body += "<h2>On order - coming</h2>" + req_table(by["Ordered"])
    if by["Arrived"]:
        body += "<h2>Arrived - closed out</h2>" + req_table(by["Arrived"])
    if by["Declined"]:
        body += "<h2>Declined - with reasons</h2>" + req_table(by["Declined"])

    hero = "{new} open &bull; {o} on order &bull; {a} arrived".format(
        new=len(by["New"]), o=len(by["Ordered"]), a=len(by["Arrived"]))
    inner = [_e_note(esc("Requests logged at the tool store counter - open "
                         "items first."), "Requests."),
             _e_h2("Open requests"),
             _e_tbl(["Date", "Item", "Qty", "Who"],
                    [[r["date"].strftime("%d %b") if r["date"] else "-",
                      r["item"], str(r["qty"] or "-"), r["who"] or "-"]
                     for r in by["New"][:12]] or [["-", "None open", "-", "-"]])]
    limits = ("Requests are logged by hand at the counter in "
              "CONSUMABLE_REQUESTS.xlsx - this report shows exactly what's "
              "in that register, nothing inferred. Empty register = no "
              "requests logged, not necessarily no requests made."
              + ("" if req_path else " The register file wasn't found next "
                 "to the kit - it ships with the suite as "
                 "CONSUMABLE_REQUESTS.xlsx."))
    emit_report("Coates_K2_Consumable_Requests_{}".format(date_tag),
                "Consumable Requests", "What The Crews Have Asked For",
                body, hero, inner, limits,
                "Coates K2 - Consumable Requests - {} - {} open".format(
                    asof.strftime("%d %b %Y") if asof else "", len(by["New"])),
                asof, generated, source_line, report_tag="REQUESTS")


# ---- STOCKTAKE SCORECARD + DAILY RUN SHEET --------------------------------
# Two outputs, two audiences (A. Fisher, 24 Jul 2026):
#   1. The Stocktake Scorecard - the client assurance piece: every storage
#      unit scored on the 3-day sighting cycle, the formula for what gets
#      counted next, and the on-hire assurance check.
#   2. The Daily Run Sheet - the floor worklist: today's due items with
#      tick boxes, worst area first. Printed, walked, scanned into SiteIQ.
# On-hire gear is counted as SEEN (it was sighted when it was issued) but
# never forgotten: anything on hire and unsighted for 7+ days makes the
# eyes-on list. Departed gear (pending branch receipt) is not counted or
# shown. Site plant, barriers, chutes and laydown live in the plant world
# (the Plant Audit List), not the tool-store stocktake - the standing
# convention from 19 Jul 2026.

STOCKTAKE_CYCLE_DAYS = 3
SHUT_END = dt.date(2026, 8, 11)      # "shut finishes around the 11/08/2026"
_ST_EXCLUDED_SU = {"", "ON HIRE", "SHUTDOWN", "TURNAROUND",
                   "CEMENT SITE PLANT", "CEMENT BARRIERS",
                   "CEMENT RUBBISH CHUTES", "LAYDOWN"}


def load_stocktake_model():
    """The full stocktake position by storage unit, from the STOCKTAKE
    export. Returns None if the export is missing."""
    path = find_newest("STOCKTAKE*.xlsx", "the STOCKTAKE export",
                       required=False)
    if not path:
        return None
    try:
        _, raw = sheet_rows(path, "STOCKTAKE", "the STOCKTAKE export")
    except KitError:
        print("NOTE: STOCKTAKE export unreadable - stocktake scorecard "
              "skipped, not zeroed.")
        return None

    def sight_dt(r):
        s = clean_text(r.get("LAST_SIGHTED_DATE_TIME"))
        for f in ("%d/%m/%Y %I:%M %p", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
            try:
                return dt.datetime.strptime(s, f)
            except ValueError:
                continue
        return None

    rows = []
    for r in raw:
        su = clean_text(r.get("STORAGE_UNIT"))
        if su.upper() in _ST_EXCLUDED_SU:
            continue
        status = clean_text(r.get("SIGHTED_STATUS"))
        if status == "Pending Branch Receipt":
            continue                      # departed - not counted, not shown
        rows.append({"su": su, "status": status,
                     "barcode": clean_text(r.get("LATEST_BARCODE")),
                     "desc": clean_text(r.get("DESCRIPTION")),
                     "when": sight_dt(r),
                     "by": clean_text(r.get("LAST_SIGHTED_BY")),
                     "action": clean_text(r.get("LAST_SIGHTED_ACTION")),
                     "sku": clean_text(r.get("SKU_NUMBER")),
                     "qty": to_num(r.get("SIGHTED_QUANTITY"))})
    if not rows:
        return None
    dates = [x["when"] for x in rows if x["when"]]
    anchor = max(dates) if dates else dt.datetime.now()
    for x in rows:
        x["days"] = (anchor.date() - x["when"].date()).days if x["when"] else 999

    units = {}
    onhire_bands = {"0-3": 0, "4-7": 0, "8+": 0}
    eyes_on = []
    counters24 = set()
    counted24 = 0
    for x in rows:
        if x["when"] and (anchor - x["when"]).days < 1:
            counted24 += 1
            if x["by"]:
                counters24.add(x["by"])
        u = units.setdefault(x["su"], {"name": x["su"], "items": [],
                                       "onhire": 0, "n": 0, "cyc": 0,
                                       "due": 0, "oldest": 0})
        if x["status"] == "On Hire":
            u["onhire"] += 1
            band = ("0-3" if x["days"] <= 3 else
                    "4-7" if x["days"] <= 7 else "8+")
            onhire_bands[band] += 1
            if x["days"] > 7:
                eyes_on.append(x)
        else:
            u["n"] += 1
            u["items"].append(x)
            if x["days"] <= STOCKTAKE_CYCLE_DAYS:
                u["cyc"] += 1
            else:
                u["due"] += 1
            u["oldest"] = max(u["oldest"], x["days"])
    for u in units.values():
        u["pct"] = (u["cyc"] / u["n"]) if u["n"] else 1.0
        u["items"].sort(key=lambda x: -x["days"])
    eyes_on.sort(key=lambda x: -x["days"])
    total = sum(u["n"] for u in units.values())
    cyc = sum(u["cyc"] for u in units.values())
    due = sum(u["due"] for u in units.values())
    onh = sum(u["onhire"] for u in units.values())
    today = dt.date.today()
    days_left = max(0, (SHUT_END - today).days)
    # worst first: lowest coverage, then most due
    ranked = sorted(units.values(), key=lambda u: (u["pct"], -u["due"]))
    return {"path": path, "anchor": anchor, "units": ranked,
            "total": total, "cyc": cyc, "due": due, "onhire": onh,
            "pct": (cyc / total) if total else 1.0,
            "bands": onhire_bands, "eyes_on": eyes_on,
            "counters24": sorted(counters24), "counted24": counted24,
            "days_left": days_left,
            "cycles_left": days_left // STOCKTAKE_CYCLE_DAYS,
            "steady": -(-total // STOCKTAKE_CYCLE_DAYS),
            "rows": rows}


def _st_rag(u):
    if u["pct"] >= 0.95:
        return ("#3FB950", "On cycle")
    if u["pct"] >= 0.80:
        return ("#fab219", "Slipping &mdash; count today")
    return ("#F85149", "Off cycle &mdash; priority count")


def generate_stocktake(asof, generated, source_line, date_tag):
    sm = load_stocktake_model()
    if not sm:
        print("  ! STOCKTAKE export not found - drop STOCKTAKE.xlsx into "
              "this folder and run again.")
        return
    pct = sm["pct"]
    worst = [u for u in sm["units"] if u["due"]]
    story = (
        "Every item in the tool store gets physically sighted at least "
        "every {cyc} days - that's the stocktake wheel, and it turns right "
        "through to the end of the shut on {end}. Right now <b>{p:.0%} of "
        "the {tot:,} countable items are on the cycle</b>; {due} are due a "
        "count, led by {lead}. {c24} item{c24s} were sighted in the last "
        "24 hours by {who}. Gear on hire ({onh} items) counts as sighted "
        "- it was physically in someone's hands when it was issued - and "
        "the on-hire assurance check below makes sure nothing rides that "
        "rule quietly. With {dl} day{dls} to go there "
        "{cw} {cl} full pass{cls} of the store left before handover."
    ).format(cyc=STOCKTAKE_CYCLE_DAYS, end=SHUT_END.strftime("%d %b %Y"),
             p=pct, tot=sm["total"], due=sm["due"],
             lead=(esc(worst[0]["name"]) + " ({} due)".format(worst[0]["due"])
                   if worst else "nothing - every unit is current"),
             c24=sm["counted24"], c24s="" if sm["counted24"] == 1 else "s",
             who=(", ".join(esc(w) for w in sm["counters24"][:4])
                  or "the store team"),
             onh=sm["onhire"], dl=sm["days_left"],
             dls="" if sm["days_left"] == 1 else "s",
             cw="is" if sm["cycles_left"] == 1 else "are",
             cl=sm["cycles_left"],
             cls="" if sm["cycles_left"] == 1 else "es")
    body = "<div class='intro story'>" + story + "</div>"
    body += ("<div style='text-align:center'>"
             + ring(pct, "Whole store on the 3-day cycle", "on cycle")
             + ring(sm["cyc"] / sm["total"] if sm["total"] else 0,
                    "{:,} of {:,} items sighted".format(sm["cyc"], sm["total"]),
                    "sighted", colour="#F26222")
             + ring(1 - (sm["due"] / sm["total"] if sm["total"] else 0),
                    "{} due a count today".format(sm["due"]), "current",
                    colour=rag_colour(1 - (sm["due"] / sm["total"] if sm["total"] else 0)))
             + "</div>")
    tiles = [("{:.0%}".format(pct), "Store On The 3-Day Cycle",
              "g" if pct >= 0.95 else "a" if pct >= 0.80 else "r"),
             ("{:,}/{:,}".format(sm["cyc"], sm["total"]), "Items Sighted (3 Days)"),
             (str(sm["due"]), "Due A Count Today",
              "g" if sm["due"] == 0 else "a"),
             (str(sm["days_left"]), "Days To Shut End")]
    body += tiles_html(tiles)

    # ---- scorecards by storage unit ----------------------------------
    rows_h = ""
    for u in sm["units"]:
        colour, status = _st_rag(u)
        rows_h += ("<tr><td>{n}</td><td class='num'>{i}</td>"
                   "<td class='num'>{c}</td><td class='num'>{p:.0%}</td>"
                   "<td class='num'>{o}</td><td class='num'>{d}</td>"
                   "<td><span style='display:inline-block;width:9px;height:9px;"
                   "border-radius:50%;background:{col};margin-right:7px'></span>"
                   "{st}</td></tr>").format(
            n=esc(u["name"]), i=u["n"], c=u["cyc"], p=u["pct"],
            o="{}d".format(u["oldest"]) if u["n"] else "-", d=u["due"],
            col=colour, st=status)
    body += (
        "<h2>Storage unit scorecards &mdash; where we stand</h2>"
        "<table class='data'><thead><tr><th>Storage unit</th>"
        "<th class='num'>Items</th><th class='num'>On cycle</th>"
        "<th class='num'>Coverage</th><th class='num'>Oldest sighting</th>"
        "<th class='num'>Due now</th><th>Status</th></tr></thead><tbody>"
        + rows_h + "</tbody></table>"
        + "<div class='note' style='margin-top:10px'><b>Coverage, drawn</b> "
          "- the same scores as bars, worst first.</div>"
        + "".join(hbar(u["name"], "{:.0%}".format(u["pct"]), u["pct"],
                       rag_colour(u["pct"])) for u in sm["units"])
        + "<div class='note'>Coverage counts the items sighted inside the "
        "last {cyc} days against everything that lives in that unit right "
        "now. On-hire and departed gear is excluded from the count - "
        "on-hire gear was sighted when it was issued, and departed gear "
        "has left site.</div>".format(cyc=STOCKTAKE_CYCLE_DAYS))

    # ---- the formula + what's next -----------------------------------
    nxt = ""
    for i, u in enumerate([u for u in sm["units"] if u["due"]], 1):
        nxt += ("<tr><td class='num'>{i}</td><td>{n}</td>"
                "<td class='num'>{d}</td><td class='num'>{o}d</td>"
                "<td>{go}</td></tr>").format(
            i=i, n=esc(u["name"]), d=u["due"], o=u["oldest"],
            go="Today" if i <= 3 else "This cycle")
    if not nxt:
        nxt = ("<tr><td colspan='5'>Nothing due - every unit is inside "
               "the cycle. Today's wheel keeps it that way.</td></tr>")
    body += (
        "<h2>What gets counted next &mdash; the formula</h2>"
        "<div class='intro'><b>The formula, in plain terms:</b> an item is "
        "DUE when its last sighting is more than {cyc} days old. The run "
        "order is worst unit first (lowest coverage), oldest items first "
        "inside each unit. Clearing today's {due} due items puts the whole "
        "store back on the wheel; after that, roughly a third of the shelf "
        "(~{steady:,} items) each day keeps it there through to "
        "{end}.</div>"
        "<table class='data'><thead><tr><th class='num'>Priority</th>"
        "<th>Storage unit</th><th class='num'>Items due</th>"
        "<th class='num'>Oldest</th><th>When</th></tr></thead><tbody>"
        + nxt + "</tbody></table>"
        "<div class='note'>The Daily Run Sheet in today's Reports folder "
        "lists every due item with a tick box, in this order - print it, "
        "walk it, scan the counts into SiteIQ.</div>").format(
        cyc=STOCKTAKE_CYCLE_DAYS, due=sm["due"], steady=sm["steady"],
        end=SHUT_END.strftime("%d %b"))

    # ---- on-hire assurance -------------------------------------------
    eo = ""
    for x in sm["eyes_on"][:12]:
        eo += ("<tr><td>{d}</td><td>{su}</td><td class='num'>{days}d</td>"
               "<td>{by}</td></tr>").format(
            d=desc_disp(x["desc"]), su=esc(x["su"]), days=x["days"],
            by=esc(x["by"] or "-"))
    body += (
        "<h2>On-hire assurance &mdash; nothing rides the rule quietly</h2>"
        "<div class='intro'>{onh} items are out on hire. On-hire gear "
        "counts as sighted - it was physically handled when issued and "
        "every return is double-scanned - but the check here is for gear "
        "that could have been missed: anything on hire with no sighting "
        "for more than 7 days earns an eyes-on with the hirer at the next "
        "natural touchpoint. Right now: <b>{b03} sighted inside 3 days, "
        "{b47} at 4-7 days, {b8} past 7 days</b>.</div>").format(
        onh=sm["onhire"], b03=sm["bands"]["0-3"], b47=sm["bands"]["4-7"],
        b8=sm["bands"]["8+"])
    if eo:
        body += ("<table class='data'><thead><tr><th>Item</th>"
                 "<th>Storage unit</th><th class='num'>Since sighted</th>"
                 "<th>Last sighted by</th></tr></thead><tbody>" + eo
                 + "</tbody></table>"
                 "<div class='note'>These aren't lost - they're on hire "
                 "and working. The ask is simply an eyes-on next time "
                 "we're near that crew, so the record stays fresh.</div>")
    else:
        body += ("<div class='note'>Nothing past 7 days - every on-hire "
                 "item has been sighted recently through issues, returns "
                 "and site movement.</div>")

    limits = (
        "Sightings come from the SiteIQ STOCKTAKE export as at the newest "
        "sighting in the file ({anchor}); a sighting is any physical scan "
        "- stocktake, arrival, issue or return. On-hire items are excluded "
        "from the countable base (sighted at issue) and covered by the "
        "assurance check instead. Departed gear (pending branch receipt) "
        "is not counted or shown. Site plant, barriers, chutes and laydown "
        "belong to the plant world and are audited on the Plant Audit "
        "List, not here. Shut end {end} is the working date - the cycle "
        "plan moves with it.").format(
        anchor=sm["anchor"].strftime("%d %b %Y %H:%M"),
        end=SHUT_END.strftime("%d %b %Y"))

    hero = "{p:.0%} on cycle &bull; {due} due today &bull; {cl} passes left".format(
        p=pct, due=sm["due"], cl=sm["cycles_left"])
    inner = [_e_note(esc("Stocktake position - every unit scored on the "
                         "3-day sighting cycle."), "Stocktake."),
             _e_h2("Where we stand"),
             _e_tbl(["Unit", "Coverage", "Due"],
                    [[u["name"], "{:.0%}".format(u["pct"]), str(u["due"])]
                     for u in sm["units"]])]
    emit_report("Coates_K2_Stocktake_Scorecard_{}".format(date_tag),
                "Stocktake Scorecard", "Every Item, Every 3 Days",
                body, hero, inner, limits,
                "Coates K2 - Stocktake Scorecard - {} - {:.0%} on cycle".format(
                    generated.strftime("%d %b %Y"), pct),
                asof, generated, source_line, report_tag="STOCKTAKE")
    generate_stocktake_runsheet(sm, generated, date_tag)


def generate_stocktake_runsheet(sm, generated, date_tag):
    """The floor worklist: today's due items with tick boxes, worst area
    first. Printed and walked - the scorecard is the client story, this is
    the day's work."""
    today_s = dt.date.today().strftime("%d %b %Y")
    due_units = [u for u in sm["units"] if u["due"]]
    total_due = sum(u["due"] for u in due_units)
    if due_units:
        intro = ("<div class='intro story'>Today's count: <b>{n} item{ns} "
                 "across {m} area{ms}</b>, worst first. Tick each item as "
                 "it's physically sighted, then scan the counts into SiteIQ "
                 "at the store desk - the scorecard refreshes from the "
                 "export, not from this sheet. Anything you can't find, "
                 "circle it and bring the sheet to Andrew.</div>").format(
            n=total_due, ns="" if total_due == 1 else "s",
            m=len(due_units), ms="" if len(due_units) == 1 else "s")
    else:
        intro = ("<div class='intro story'>Nothing due today - the whole "
                 "store is inside the {c}-day cycle. Keep the wheel "
                 "turning: today's pass is roughly a third of the shelf "
                 "(~{s:,} items), any order you like.</div>").format(
            c=STOCKTAKE_CYCLE_DAYS, s=sm["steady"])
    body = intro
    for u in due_units:
        rows = "".join(
            ("<tr><td style='width:34px;text-align:center'>"
             "<span style='display:inline-block;width:14px;height:14px;"
             "border:1.6px solid #8B9099;border-radius:3px'></span></td>"
             "<td>{bc}</td><td>{d}</td><td class='num'>{days}d</td>"
             "<td>{by}</td></tr>").format(
                bc=esc(x["barcode"] or "-"), d=desc_disp(x["desc"]),
                days=x["days"], by=esc(x["by"] or "-"))
            for x in u["items"] if x["days"] > STOCKTAKE_CYCLE_DAYS)
        body += ("<h2>{n} &mdash; {d} item{ds} due (oldest {o}d)</h2>"
                 "<table class='data'><thead><tr><th></th><th>Barcode</th>"
                 "<th>Description</th><th class='num'>Last sighted</th>"
                 "<th>By</th></tr></thead><tbody>" + rows
                 + "</tbody></table>").format(
            n=esc(u["name"]), d=u["due"], ds="" if u["due"] == 1 else "s",
            o=u["oldest"])
    doc = paged_report_doc(
        "Daily Stocktake Run Sheet", today_s, sm["anchor"], generated,
        "STOCKTAKE.xlsx as at " + sm["anchor"].strftime("%d %b %Y %H:%M"),
        body, doc_title="Coates K2 Stocktake Run Sheet",
        asof_label="STOCKTAKE export, newest sighting")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stem = "Coates_K2_Stocktake_RunSheet_{}".format(date_tag)
    html_path = os.path.join(PAGES_DIR, stem + ".html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(doc)
    pdf_ok = html_to_pdf(html_path, os.path.join(PDF_DIR, stem + ".pdf"))
    print("  Stocktake Run Sheet ({} due): HTML{} - print and walk it".format(
        total_due, " + PDF" if pdf_ok else " (print from the browser)"))


# ---- STORES STOCKTAKE TEAM REPORT -----------------------------------------
# The team's side of the stocktake story (A. Fisher, 25 Jul 2026): the
# compliance scorecard tells the CLIENT the store is under control - THIS
# pack tells the TEAM what they have achieved. What's been counted, when,
# by whom, at what pace, day shift v night shift, and a fair split of
# what's left. Positive by design, honest by rule: every figure computed
# from the STOCKTAKE export, methods printed on the page, nothing invented.

# Shift windows per Andrew (25 Jul 2026): DAY 06:00-18:00, NIGHT 18:00-06:00.
_DAY_COL = ORANGE_HEX
_NIGHT_COL = "#4A90D9"


def _hr_of(tmstr):
    """Hour (0-23) off a SiteIQ time stamp like '6:23 AM' / '18:07'; None
    when there's no usable stamp - unstamped events are counted but never
    guessed into a shift."""
    m = re.match(r"\s*(\d{1,2}):(\d{2})\s*([AP]M)?", str(tmstr or ""), re.I)
    if not m:
        return None
    h = int(m.group(1))
    ap = (m.group(3) or "").upper()
    if ap:
        h = h % 12 + (12 if ap == "PM" else 0)
    return h if 0 <= h <= 23 else None


def _shift_window(h):
    return "day" if 6 <= h < 18 else "night"


def _txn_hour(s):
    """The hour out of a TRANSACTIONS time cell ('01:07 PM', '13:07').
    None when the cell is blank or unreadable - an unstamped event is
    left out of the shift split and the page says so, never guessed."""
    s = str(s or "").strip().upper()
    if not s:
        return None
    for f in ("%I:%M %p", "%I:%M%p", "%H:%M:%S", "%H:%M"):
        try:
            return dt.datetime.strptime(s, f).hour
        except ValueError:
            continue
    return None


def _shift_day(d, h):
    """The date a shift event BELONGS to (A. Fisher, 25 Jul 2026): night
    shift runs 18:00 into 06:00 next morning and belongs to the day it
    STARTS - so a 02:00 event on the 24th is the 23rd's night shift."""
    if h is not None and h < 6:
        return d - dt.timedelta(days=1)
    return d


def _sun_ico(size=20):
    return ("<svg width='{s}' height='{s}' viewBox='0 0 24 24'>"
            "<path fill='{c}' d='M12 7a5 5 0 100 10 5 5 0 000-10zm0-6h0l1 4h-2l1-4zm0 22l-1-4h2l-1 4zM1 12l4-1v2l-4-1zm22 0l-4 1v-2l4 1zM4.2 4.2l3.5 2.1-1.4 1.4L4.2 4.2zm15.6 15.6l-3.5-2.1 1.4-1.4 2.1 3.5zM19.8 4.2l-2.1 3.5-1.4-1.4 3.5-2.1zM4.2 19.8l2.1-3.5 1.4 1.4-3.5 2.1z'/>"
            "</svg>").format(s=size, c=_DAY_COL)


def _moon_ico(size=20):
    return ("<svg width='{s}' height='{s}' viewBox='0 0 24 24'>"
            "<path fill='{c}' d='M21 12.8A9 9 0 1111.2 3a7 7 0 009.8 9.8z'/>"
            "</svg>").format(s=size, c=_NIGHT_COL)


def _hourbars24(hours, title="", h_px=200):
    """The 24-hour counter clock: one bar per hour, day-shift hours in
    Coates orange, night-shift hours in blue, peak labelled, quiet hours
    dimmed. hours = {hour: count}. Pure SVG - prints perfectly."""
    mx = max([hours.get(h, 0) for h in range(24)] + [1])
    W, H = 660, h_px
    pad_l, pad_b, pad_t = 10, 34, 26
    bw = (W - pad_l * 2) / 24.0
    bars = ""
    for h in range(24):
        v = hours.get(h, 0)
        bh = 0 if mx == 0 else (H - pad_b - pad_t) * v / mx
        x = pad_l + h * bw
        col = _DAY_COL if _shift_window(h) == "day" else _NIGHT_COL
        op = "1" if v else ".25"
        bars += ("<rect x='{x:.1f}' y='{y:.1f}' width='{w:.1f}' "
                 "height='{bh:.1f}' rx='3' fill='{c}' opacity='{o}'/>"
                 ).format(x=x + 1.5, y=H - pad_b - max(bh, 2.5),
                          w=bw - 3, bh=max(bh, 2.5), c=col, o=op)
        if v == mx and mx > 0:
            bars += ("<text x='{x:.1f}' y='{y:.1f}' fill='#fff' "
                     "font-size='11' font-weight='800' text-anchor='middle'>"
                     "{v}</text>").format(x=x + bw / 2,
                                          y=H - pad_b - bh - 7, v=v)
        if h % 3 == 0:
            bars += ("<text x='{x:.1f}' y='{y}' fill='#8B93A1' "
                     "font-size='9.5' text-anchor='middle'>{h:02d}:00"
                     "</text>").format(x=x + bw / 2, y=H - pad_b + 14, h=h)
    # day/night band under the axis
    band = ""
    for h in range(24):
        x = pad_l + h * bw
        col = _DAY_COL if _shift_window(h) == "day" else _NIGHT_COL
        band += ("<rect x='{x:.1f}' y='{y}' width='{w:.1f}' height='4' "
                 "fill='{c}' opacity='.8'/>").format(
            x=x + 1.5, y=H - pad_b + 20, w=bw - 3, c=col)
    legend = ("<div style='text-align:center;font-size:8.5pt;color:#A9B1BD;"
              "margin-top:2px'>{sun} <b style='color:{dc}'>DAY 06:00-18:00"
              "</b> &nbsp;&nbsp; {moon} <b style='color:{nc}'>NIGHT "
              "18:00-06:00</b></div>").format(
        sun=_sun_ico(13), moon=_moon_ico(13), dc=_DAY_COL, nc=_NIGHT_COL)
    t = ("<div style='font-size:9pt;color:#A9B1BD;margin:2px 0 4px'>{}</div>"
         .format(esc(title)) if title else "")
    # (plain div - 'newpage' anywhere in a class name would trigger the
    #  paginator's fresh-page rule and strand the section heading)
    return ("<div>" + t
            + "<svg width='{W}' height='{H}' viewBox='0 0 {W} {H}'>"
              "<rect x='0' y='0' width='{W}' height='{H}' rx='10' "
              "fill='#12161C'/>".format(W=W, H=H)
            + bars + band + "</svg>" + legend + "</div>")


def _quiet_stretch(hours):
    """Longest run of consecutive quiet hours (counts <= 10% of peak),
    scanned around the clock. Returns (start_h, end_h, run_len) or None."""
    mx = max([hours.get(h, 0) for h in range(24)] + [0])
    if mx == 0:
        return None
    thr = mx * 0.10
    best = None
    run = 0
    for i in range(48):
        h = i % 24
        if hours.get(h, 0) <= thr:
            run += 1
            if run <= 24 and (best is None or run > best[2]):
                best = ((h - run + 1) % 24, (h + 1) % 24, run)
        else:
            run = 0
    return best


def generate_stocktake_team(asof, generated, source_line, date_tag):
    sm = load_stocktake_model()
    if not sm:
        print("  ! STOCKTAKE export not found - stocktake TEAM report "
              "skipped.")
        return
    # counting work = sightings logged through the stocktake screen; the
    # export keeps each item's LATEST sighting, so trends read "items whose
    # newest count landed that day" - said plainly in the limits.
    counts = [x for x in sm["rows"]
              if x["when"] and x["action"].lower() == "stocktake"]
    if not counts:
        print("  ! No stocktake-action sightings in the export - team "
              "report skipped, not invented.")
        return
    anchor = sm["anchor"]
    today = anchor.date()

    # Shift credit comes from the TIME STAMP, never the roster: work done
    # 06:00-18:00 is day-shift work whoever did it, 18:00-06:00 is night's
    # and belongs to the day it starts. (25 Jul 2026: roster-based credit
    # called day-hours counting "night shift" before nights had even
    # started - stamps only, always.)
    by_person = {}
    by_shift = {"day": 0, "night": 0}
    by_day = {}
    by_hour = {}
    for x in counts:
        p = x["by"] or "(not stamped)"
        e = by_person.setdefault(p, {"n": 0, "days": {}, "latest": None,
                                     "day": 0, "night": 0})
        e["n"] += 1
        d = x["when"].date()
        e["days"][d] = e["days"].get(d, 0) + 1
        if e["latest"] is None or x["when"] > e["latest"]:
            e["latest"] = x["when"]
        w = _shift_window(x["when"].hour)
        by_shift[w] += 1
        e[w] += 1
        by_day[d] = by_day.get(d, 0) + 1
        by_hour[x["when"].hour] = by_hour.get(x["when"].hour, 0) + 1

    # pace: counting sessions = runs of counts by one person with gaps
    # under 5 minutes; pace = counts per active hour inside sessions
    per_p = {}
    for x in counts:
        per_p.setdefault(x["by"] or "?", []).append(x["when"])
    sess_secs = 0.0
    sess_counts = 0
    for times in per_p.values():
        times.sort()
        run_start = times[0]
        prev = times[0]
        n_run = 1
        for t in times[1:]:
            if (t - prev).total_seconds() <= 300:
                prev = t
                n_run += 1
                continue
            if n_run >= 3:
                sess_secs += (prev - run_start).total_seconds()
                sess_counts += n_run
            run_start = prev = t
            n_run = 1
        if n_run >= 3:
            sess_secs += (prev - run_start).total_seconds()
            sess_counts += n_run
    pace_hr = (sess_counts / (sess_secs / 3600.0)) if sess_secs > 0 else 0
    sec_per = (sess_secs / sess_counts) if sess_counts else 0

    last14 = [(today - dt.timedelta(days=i)) for i in range(13, -1, -1)]
    trend = [(d.strftime("%d %b") if d.day in (1, 15) or i % 2 == 0 else "",
              by_day.get(d, 0)) for i, d in enumerate(last14)]
    best_day, best_n = None, 0
    for d, n in by_day.items():
        if n > best_n:
            best_day, best_n = d, n
    last3 = [by_day.get(today - dt.timedelta(days=i), 0) for i in (0, 1, 2)]
    rate3 = sum(last3) / 3.0
    finish_days = (sm["due"] / rate3) if rate3 > 0 else None

    total_counts = len(counts)
    pct = sm["pct"]
    story = (
        "This one is yours. The client sees a scorecard that says the "
        "store is under control - here is the work that MAKES it true. "
        "<b>{tc:,} counts</b> have gone through the stocktake screen, "
        "{c24} of them in the last 24 hours, and right now <b>{p:.0%} of "
        "the {tot:,} countable items are inside the {cyc}-day cycle</b>. "
        "Your biggest single day so far: <b>{bn} counts on {bd}</b>. "
        "The wheel doesn't turn itself - it's turned by the people on "
        "this page."
    ).format(tc=total_counts, c24=sm["counted24"], p=pct, tot=sm["total"],
             cyc=STOCKTAKE_CYCLE_DAYS, bn=best_n,
             bd=best_day.strftime("%d %b") if best_day else "-")
    body = "<div class='intro story'>" + story + "</div>"
    body += v3_kpis([
        ("tick", "{:,}".format(total_counts), "Counts logged this shut", ""),
        ("clock", "{:,}".format(sm["counted24"]), "In the last 24 hours", ""),
        ("chart", "{:.0%}".format(pct), "Store on the {}-day cycle".format(
            STOCKTAKE_CYCLE_DAYS), ""),
        ("alert", "{:,}".format(sm["due"]), "Left to count today",
         "<span style='color:{c};font-size:8pt;font-weight:700'>{t}</span>".format(
             c=V3_GOOD if sm["due"] == 0 else ORANGE_HEX,
             t="all caught up" if sm["due"] == 0
             else "the 70/30 deal below")),
    ])
    body += ("<div style='text-align:center'>"
             + ring(pct, "Whole store on the cycle", "on cycle")
             + ring(min(1.0, total_counts / max(1, sm["total"])),
                    "{:,} counts v {:,} items".format(total_counts,
                                                      sm["total"]),
                    "coverage", colour=ORANGE_HEX)
             + "</div>")

    # ---- THE DAILY THREE -------------------------------------------------
    #  (Andrew, 28 Jul 2026): "the 3 things we want stock taked daily.
    #  gas monitors. radios. radio batteries. and we must show this data.
    #  and whats been missed." Life-safety and comms gear is counted
    #  EVERY DAY, not on the 3-day wheel. On-hire units count as sighted
    #  at issue (store convention) - the daily walk is the shelf.
    def _d3(label, pred):
        shelf = [x for x in sm["rows"]
                 if pred(x) and x["status"] != "On Hire"]
        out_n = sum(1 for x in sm["rows"]
                    if pred(x) and x["status"] == "On Hire")
        done = [x for x in shelf if x["days"] <= 0]
        missed = sorted((x for x in shelf if x["days"] >= 1),
                        key=lambda x: -x["days"])
        return {"label": label, "shelf": shelf, "out": out_n,
                "done": done, "missed": missed}

    def _is_radio_batt(x):
        return "BATTER" in x["desc"].upper()

    d3 = [
        _d3("Gas monitors",
            lambda x: x["su"].upper() == "GAS MONITORS"),
        _d3("Radios",
            lambda x: x["su"].upper() == "RADIOS"
            and not _is_radio_batt(x)),
        _d3("Radio batteries",
            lambda x: x["su"].upper() == "RADIOS" and _is_radio_batt(x)),
    ]
    d3_missed = sum(len(g["missed"]) for g in d3)
    body += "<h2>The daily three &mdash; counted every single day</h2>"
    body += ("<div class='intro story'><b>Three groups are on a DAILY "
             "count, not the {c}-day wheel: gas monitors, radios and "
             "radio batteries.</b> Gas monitors keep people alive, radios "
             "keep the site talking, and batteries are what walks. Every "
             "shelf unit in these three gets eyes on it every day - gear "
             "out with a crew counts as sighted at issue. Below is "
             "today's position and <b>exactly what was missed</b>."
             "</div>").format(c=STOCKTAKE_CYCLE_DAYS)
    kcells = []
    for g in d3:
        n_shelf = len(g["shelf"])
        ok = not g["missed"]
        kcells.append((
            "tick" if ok else "alert",
            "{}/{}".format(len(g["done"]), n_shelf) if n_shelf
            else "0/0",
            "{} sighted today".format(g["label"]),
            "<span style='color:{c};font-size:8pt;font-weight:700'>{t}"
            "</span>".format(
                c=V3_GOOD if ok else V3_BAD,
                t=("CLEAN - all counted" if ok and n_shelf else
                   "nothing on the shelf" if not n_shelf else
                   "{} MISSED".format(len(g["missed"]))))))
    body += v3_kpis(kcells, per_row=3)
    for g in d3:
        n_shelf = len(g["shelf"])
        body += ("<h3 style='margin:14px 0 4px'>{l} &mdash; {n} on the "
                 "shelf &middot; {o} out with crews</h3>").format(
            l=esc(g["label"]), n=n_shelf, o=g["out"])
        if n_shelf:
            body += hbar("Sighted in the last 24h",
                         "{} of {}".format(len(g["done"]), n_shelf),
                         len(g["done"]) / n_shelf,
                         "#3FB950" if not g["missed"] else "#F2B01E")
        if g["missed"]:
            rows_h = "".join(
                "<tr><td>{d}</td><td>{b}</td><td class='num'>{dy}d ago</td>"
                "<td>{by}</td></tr>".format(
                    d=esc(x["desc"]), b=esc(x["barcode"] or "-"),
                    dy=x["days"], by=esc(x["by"] or "-"))
                for x in g["missed"][:12])
            more = len(g["missed"]) - 12
            body += ("<div class='note' style='border-left:4px solid "
                     "#F85149'><b>Missed - not sighted in the last 24 "
                     "hours:</b></div>"
                     "<table class='data'><thead><tr><th>Item</th>"
                     "<th>Barcode</th><th class='num'>Last sighted</th>"
                     "<th>By</th></tr></thead><tbody>" + rows_h
                     + "</tbody></table>")
            if more > 0:
                body += ("<div class='note'>&hellip; and {} more in this "
                         "group - the Daily Run Sheet lists the lot."
                         "</div>").format(more)
        elif n_shelf:
            body += ("<div class='note' style='border-left:4px solid "
                     "#3FB950'><b>Clean</b> - every shelf unit sighted "
                     "inside 24 hours. Say so at pre-start.</div>")
    if d3_missed == 0 and any(len(g["shelf"]) for g in d3):
        body += ("<div class='intro story'><b>The daily three are clean "
                 "today</b> - every gas monitor, radio and radio battery "
                 "on the shelf has been sighted inside 24 hours. That is "
                 "the standard - hold it.</div>")

    # ---- what you've done: trend + shifts + people -----------------------
    body += "<h2>What you've done &mdash; day by day</h2>"
    body += ("<div class='intro'>Counts landing per day over the last "
             "fortnight. Every bar is barcodes scanned, shelves walked and "
             "gear physically sighted.</div>")
    body += v3_trend(trend)
    body += ("<div class='note'>The export keeps each item's NEWEST count "
             "only - as items get re-counted, earlier days fade from this "
             "chart. Recent days are the true read; the total above counts "
             "everything currently on record.</div>")
    segs = [("Day window 06:00-18:00", by_shift["day"], _DAY_COL),
            ("Night window 18:00-06:00", by_shift["night"], _NIGHT_COL)]
    body += v3_donut([s for s in segs if s[1] > 0],
                     title="When the wheel turns - counts by shift "
                           "window (SiteIQ stamps)", total_label="COUNTS")
    if by_shift["night"] == 0:
        body += ("<div class='note'>Every count so far is stamped in day "
                 "hours - the night window hasn't put counts on the board "
                 "yet. The moment it does, it shows here in blue.</div>")

    # ---- DAY v NIGHT - the power board -----------------------------------
    #  (Andrew, 28 Jul 2026): "what shift is busier what shift does the
    #  most transactions. lets really show some high detail." Two decks:
    #  the stocktake wheel AND the counter, each split day/night off the
    #  SiteIQ stamps. Unstamped counter events are left out of the split
    #  and said out loud - never guessed onto a shift.
    tx_all, _txp = load_transactions()
    ctr = {"day": {"ev": 0, "out": 0, "back": 0, "hours": {}},
           "night": {"ev": 0, "out": 0, "back": 0, "hours": {}}}
    ctr_by_day = {}
    unstamped = 0
    for t in tx_all:
        for when, tm, kind in ((t["start"], t["start_time"], "out"),
                               (t["end"], t["end_time"], "back")):
            if not when:
                continue
            h = _txn_hour(tm)
            if h is None:
                unstamped += 1
                continue
            w = _shift_window(h)
            ctr[w]["ev"] += 1
            ctr[w][kind] += 1
            ctr[w]["hours"][h] = ctr[w]["hours"].get(h, 0) + 1
            sd = _shift_day(when, h)
            e = ctr_by_day.setdefault(sd, {"day": 0, "night": 0})
            e[w] += 1

    def _peak(hours):
        if not hours:
            return "-"
        h = max(hours, key=hours.get)
        return "{:02d}:00 ({})".format(h, hours[h])

    body += "<h2>Day v night &mdash; the power board</h2>"
    tot_ct = by_shift["day"] + by_shift["night"]
    tot_ev = ctr["day"]["ev"] + ctr["night"]["ev"]
    deck = ""
    for w, nm, col in (("day", "DAY SHIFT 06:00&ndash;18:00", _DAY_COL),
                       ("night", "NIGHT SHIFT 18:00&ndash;06:00",
                        _NIGHT_COL)):
        c = ctr[w]
        dhrs = {h: n for h, n in by_hour.items() if _shift_window(h) == w}
        deck += (
            "<td style='width:50%;vertical-align:top;background:{cd};"
            "border:1px solid {ln};border-top:4px solid {cl};"
            "border-radius:12px;padding:12px 14px'>"
            "<div style='font-weight:800;letter-spacing:1.5px;color:{cl};"
            "font-size:10pt'>{nm}</div>"
            "<table style='width:100%;margin-top:8px'><tr>"
            "<td><span style='font-size:20pt;font-weight:800;color:{ik}'>"
            "{cnt:,}</span><br><span style='font-size:7.5pt;color:{mut};"
            "letter-spacing:1px'>COUNTS LOGGED ({cp:.0%})</span></td>"
            "<td><span style='font-size:20pt;font-weight:800;color:{ik}'>"
            "{ev:,}</span><br><span style='font-size:7.5pt;color:{mut};"
            "letter-spacing:1px'>COUNTER EVENTS ({ep:.0%})</span></td>"
            "</tr><tr>"
            "<td style='padding-top:7px'><span style='font-size:13pt;"
            "font-weight:800;color:{ik}'>{go:,} / {gb:,}</span><br>"
            "<span style='font-size:7.5pt;color:{mut};letter-spacing:1px'>"
            "GEAR OUT / GEAR BACK</span></td>"
            "<td style='padding-top:7px'><span style='font-size:13pt;"
            "font-weight:800;color:{ik}'>{pkc}</span><br>"
            "<span style='font-size:7.5pt;color:{mut};letter-spacing:1px'>"
            "BUSIEST COUNT HOUR</span></td>"
            "</tr><tr>"
            "<td colspan='2' style='padding-top:7px'><span style="
            "'font-size:13pt;font-weight:800;color:{ik}'>{pke}</span><br>"
            "<span style='font-size:7.5pt;color:{mut};letter-spacing:1px'>"
            "BUSIEST COUNTER HOUR</span></td>"
            "</tr></table></td>").format(
            cd=V3_CARD, ln=V3_LINE, cl=col, nm=nm, ik="#fff",
            mut="#8B9099",
            cnt=by_shift[w], cp=(by_shift[w] / tot_ct) if tot_ct else 0,
            ev=c["ev"], ep=(c["ev"] / tot_ev) if tot_ev else 0,
            go=c["out"], gb=c["back"], pkc=_peak(dhrs),
            pke=_peak(c["hours"]))
    body += ("<table style='width:100%;border-collapse:separate;"
             "border-spacing:8px 0'><tr>" + deck + "</tr></table>")
    #  the verdict, in words - who is busier at what
    if tot_ct or tot_ev:
        cw = "day" if by_shift["day"] >= by_shift["night"] else "night"
        ew = "day" if ctr["day"]["ev"] >= ctr["night"]["ev"] else "night"
        cwp = (by_shift[cw] / tot_ct) if tot_ct else 0
        ewp = (ctr[ew]["ev"] / tot_ev) if tot_ev else 0
        body += ("<div class='intro story'><b>The verdict:</b> "
                 "<b>{cn} shift turns the wheel</b> - {cwp:.0%} of all "
                 "stocktake counts - and <b>{en} shift owns the counter"
                 "</b> with {ewp:.0%} of counter events (gear out and "
                 "back). Different shifts, different strengths - the "
                 "70/30 count deal below plays to exactly that."
                 "</div>").format(
            cn=cw.title(), cwp=cwp, en=ew.title(), ewp=ewp)
    #  counter events, day v night, last 7 shift-days - the busy chart
    if ctr_by_day:
        last7 = sorted(ctr_by_day)[-7:]
        sa = [(d.strftime("%d %b"), ctr_by_day[d]["day"]) for d in last7]
        sb = [(d.strftime("%d %b"), ctr_by_day[d]["night"]) for d in last7]
        body += ("<div class='note'><b>The counter, day v night, last 7 "
                 "shift-days</b> - every issue and every return with a "
                 "time stamp on it.</div>")
        body += v3_bars_pair(sa, sb, "Day shift", "Night shift",
                             col_a=_DAY_COL, col_b=_NIGHT_COL)
    if unstamped:
        body += ("<div class='note'>{:,} counter event(s) carry no time "
                 "stamp in the export and are left out of the day/night "
                 "split - counted in totals elsewhere, never guessed "
                 "onto a shift.</div>").format(unstamped)

    prs = sorted(by_person.items(), key=lambda kv: -kv[1]["n"])
    rows_h = ""
    for name, e in prs:
        bd = max(e["days"].items(), key=lambda kv: kv[1])
        if e["day"] and e["night"]:
            when = ("days + nights ({d:,}/{n:,})".format(d=e["day"],
                                                         n=e["night"]))
        elif e["night"]:
            when = "night hours"
        else:
            when = "day hours"
        rows_h += ("<tr><td>{n}</td><td>{s}</td><td class='num'>{c:,}</td>"
                   "<td class='num'>{days}</td><td>{best}</td>"
                   "<td>{latest}</td></tr>").format(
            n=esc(name), s=when, c=e["n"], days=len(e["days"]),
            best="{} on {}".format(bd[1], bd[0].strftime("%d %b")),
            latest=e["latest"].strftime("%d %b %H:%M"))
    body += (
        "<h2>The counters &mdash; credit where it's earned</h2>"
        "<table class='data'><thead><tr><th>Name</th><th>When they "
        "counted</th><th class='num'>Counts</th><th class='num'>Days "
        "active</th><th>Best day</th><th>Latest count</th></tr></thead>"
        "<tbody>" + rows_h + "</tbody></table>"
        + "".join(hbar(n, "{:,}".format(e["n"]),
                       e["n"] / max(1, prs[0][1]["n"])) for n, e in prs))

    # ---- when the counting happens + pace --------------------------------
    body += "<h2>When the counting happens &mdash; and the pace</h2>"
    body += v3_heat(by_hour, title="Counting by hour of day - every "
                                   "stocktake sighting this shut")
    if sess_counts:
        body += (
            "<div class='intro'>Inside counting sessions the pace runs "
            "<b>{ph:.0f} counts an hour</b> - about one every "
            "{sp:.0f} seconds. (A session = one person counting with no "
            "break longer than 5 minutes; {sc:,} counts across "
            "{sh:.1f} hours of session time.)</div>").format(
            ph=pace_hr, sp=sec_per, sc=sess_counts, sh=sess_secs / 3600.0)

    # ---- where we're heading ---------------------------------------------
    body += "<h2>Where we're heading</h2>"
    lead = ""
    if sm["due"] == 0:
        lead = ("<b>Nothing is due right now</b> - the whole store is "
                "inside the cycle. Roughly {steady:,} items a day keeps "
                "it that way through to {end}.")
    elif finish_days is not None and finish_days <= 1.5:
        lead = ("<b>{due} items are due</b> and at your 3-day average of "
                "{r:.0f} counts a day that's cleared inside a day - then "
                "~{steady:,} a day holds the wheel through to {end}.")
    elif finish_days is not None:
        lead = ("<b>{due} items are due.</b> At your 3-day average of "
                "{r:.0f} counts a day the backlog clears in about "
                "{fd:.0f} days; ~{steady:,} a day after that holds the "
                "wheel through to {end}.")
    else:
        lead = ("<b>{due} items are due.</b> No counts have landed in the "
                "last 3 days - the split below gets the wheel turning "
                "again; ~{steady:,} a day holds it through to {end}.")
    body += "<div class='intro story'>" + lead.format(
        due=sm["due"], r=rate3, steady=sm["steady"],
        fd=finish_days if finish_days else 0,
        end=SHUT_END.strftime("%d %b")) + "</div>"
    body += tiles_html([
        (str(sm["due"]), "Due Now", "g" if sm["due"] == 0 else "a"),
        ("{:.0f}/day".format(rate3), "Your 3-Day Average",
         "g" if rate3 >= sm["steady"] else "a"),
        ("{:,}/day".format(sm["steady"]), "Keeps The Wheel Turning"),
        (str(sm["days_left"]), "Days To Shut End")])

    # ---- the next 24 hours: who counts what -------------------------------
    # (A. Fisher, 25 Jul 2026): the 24-hour plan, fresh every run - what's
    # DUE now plus what FALLS DUE before this time tomorrow, dealt fairly
    # across day and night shift: whole storage units, biggest workload
    # first, always onto the lighter pile.
    plan_units = []
    for u in sm["units"]:
        job_items = [x for x in u["items"]
                     if x["days"] >= STOCKTAKE_CYCLE_DAYS]
        falling = sum(1 for x in job_items
                      if x["days"] == STOCKTAKE_CYCLE_DAYS)
        w = len(job_items)
        if w > 0:
            plan_units.append({"name": u["name"], "due": u["due"],
                               "falling": falling, "w": w,
                               "oldest": u["oldest"],
                               "items": sorted(job_items,
                                               key=lambda x: -x["days"])})
    due_total = sum(p["due"] for p in plan_units)
    fall_total = sum(p["falling"] for p in plan_units)
    work_total = due_total + fall_total
    #  THE 70/30 DEAL (Andrew, 28 Jul 2026): nights carry 70% of the
    #  wheel, days 30. Nights have the quiet hours and full shelves;
    #  days have the counter traffic - the split matches the work to
    #  the hours that can actually do it. Whole units still deal
    #  biggest-first, each onto whichever pile is furthest under ITS
    #  OWN share - so the final split lands as close to 70/30 as whole
    #  storage units allow.
    SHIFT_SHARE = {"Day shift": 0.30, "Night shift": 0.70}
    alloc = {"Day shift": [], "Night shift": []}
    tally = {"Day shift": 0, "Night shift": 0}
    for p in sorted(plan_units, key=lambda p: -p["w"]):
        pick = min(tally,
                   key=lambda k: (tally[k] + p["w"]) / SHIFT_SHARE[k])
        alloc[pick].append(p)
        tally[pick] += p["w"]
    body += "<h2>The next 24 hours &mdash; who counts what</h2>"
    if not plan_units:
        body += ("<div class='intro story'><b>The wheel is clear for the "
                 "next 24 hours</b> - nothing due, nothing falling due "
                 "before this time tomorrow. Keep it that way: roughly "
                 "{steady:,} sightings across the day on the 70/30 deal "
                 "- nights carry about 70%, days about 30 - oldest "
                 "sightings first. The Daily Run Sheet orders the walk "
                 "for you.</div>").format(
            steady=sm["steady"])
    else:
        body += ("<div class='intro story'>Fresh from this run's export: "
                 "<b>{d:,} item{ds} due now</b> plus <b>{f:,} falling due "
                 "before this time tomorrow</b> = <b>{w:,} items to sight "
                 "in the next 24 hours</b>. Dealt below as whole storage "
                 "units on <b>the 70/30 deal - nights carry 70% of the "
                 "wheel, days 30</b>: nights have the quiet hours and "
                 "full shelves, days have the counter. Biggest workload "
                 "first, each unit onto whichever shift is furthest "
                 "under its share.</div>").format(
            d=due_total, ds="" if due_total == 1 else "s",
            f=fall_total, w=work_total)
        body += v3_kpis([
            ("alert", "{:,}".format(due_total), "Due now", ""),
            ("clock", "{:,}".format(fall_total), "Falling due in 24h", ""),
            ("tick", "{:,}".format(work_total), "The 24-hour job", ""),
            ("people", "{:,} / {:,}".format(tally["Day shift"],
                                            tally["Night shift"]),
             "Day / night share", ""),
        ])
        # fairness meter: one bar, orange v blue
        dpct = tally["Day shift"] / float(work_total) if work_total else 0
        body += (
            "<div style='margin:10px 0 4px;font-size:9.5pt;color:#A9B1BD'>"
            "<b>The 70/30 meter</b> - target: nights 70%, days 30%. Whole "
            "units can't split to the item, this is as close as the "
            "areas deal:</div>"
            "<div style='display:flex;height:30px;border-radius:9px;"
            "overflow:hidden;font-size:10pt;font-weight:800'>"
            "<div style='width:{dp:.1%};background:{dc};display:flex;"
            "align-items:center;justify-content:center;color:#fff'>"
            "DAY {dn:,} ({dp:.0%})</div>"
            "<div style='flex:1;background:{nc};display:flex;align-items:"
            "center;justify-content:center;color:#fff'>NIGHT {nn:,} "
            "({np2:.0%})</div></div>").format(
            dp=dpct, dc=_DAY_COL, dn=tally["Day shift"], nc=_NIGHT_COL,
            nn=tally["Night shift"], np2=1 - dpct)
        for shift, col, ico in (("Day shift", _DAY_COL, _sun_ico(16)),
                                ("Night shift", _NIGHT_COL, _moon_ico(16))):
            rows_h = ""
            for p in alloc[shift]:
                rows_h += ("<tr><td>{u}</td><td class='num'>{d}</td>"
                           "<td class='num'>{f}</td><td class='num'><b>{w}"
                           "</b></td><td class='num'>{o}d</td></tr>"
                           ).format(u=esc(p["name"]), d=p["due"],
                                    f=p["falling"], w=p["w"], o=p["oldest"])
            if not rows_h:
                rows_h = ("<tr><td colspan='5'>No units this window - "
                          "the other shift's pile covers today's wheel."
                          "</td></tr>")
            body += (
                "<div style='margin-top:12px;font-weight:800;font-size:"
                "12pt'>{ico} <span style='color:{col}'>{s} job card</span>"
                " &mdash; {t:,} item{ts}</div>"
                "<table class='data'><thead><tr><th>Storage unit</th>"
                "<th class='num'>Due now</th><th class='num'>Falling due"
                "</th><th class='num'>To sight</th><th class='num'>Oldest"
                "</th></tr></thead><tbody>" + rows_h + "</tbody></table>"
            ).format(ico=ico, col=col, s=shift, t=tally[shift],
                     ts="" if tally[shift] == 1 else "s")
            # name the ACTUAL items so the crew knows exactly what to
            # find and where (A. Fisher, 25 Jul 2026) - all of them when
            # the pile is small, the oldest 24 when it's big
            all_items = [(p["name"], x) for p in alloc[shift]
                         for x in p["items"]]
            all_items.sort(key=lambda t: -t[1]["days"])
            show = all_items[:24]
            if show:
                rows_i = ""
                for su_name, x in show:
                    rows_i += ("<tr><td>{d}</td><td class='num'>{b}</td>"
                               "<td>{u}</td><td class='num'>{ls}</td>"
                               "<td class='num'>{dy}d</td></tr>").format(
                        d=desc_disp(x["desc"]),
                        b=esc(x["barcode"] or "-"), u=esc(su_name),
                        ls=(x["when"].strftime("%d %b")
                            if x["when"] else "never"),
                        dy=x["days"] if x["days"] < 999 else "-")
                more = len(all_items) - len(show)
                body += (
                    "<div style='margin-top:6px;font-size:9.5pt;"
                    "color:#A9B1BD'>The actual items, oldest sighting "
                    "first{cap}:</div>"
                    "<table class='data'><thead><tr><th>Item</th>"
                    "<th class='num'>Barcode</th><th>Where it lives</th>"
                    "<th class='num'>Last sighted</th><th class='num'>Age"
                    "</th></tr></thead><tbody>" + rows_i
                    + "</tbody></table>").format(
                    cap=(" (top 24 of {:,} - the Daily Run Sheet has the "
                         "lot, tick boxes and all".format(len(all_items))
                         + ")") if more > 0 else "")
        body += ("<div class='note'>Whole units, biggest first, always "
                 "onto the lighter pile - swap by agreement, and if a "
                 "unit's mid-count at handover, the next shift picks up "
                 "where the last one left off: the wheel doesn't care "
                 "whose torch it is. The Daily Run Sheet lists every due "
                 "item with a tick box in walk order - print one per "
                 "shift. This plan rebuilds from the freshest export "
                 "every time the report runs.</div>")

    # ================= THE SHELF AND THE FIELD, RIGHT NOW =================
    # (A. Fisher, 25 Jul 2026): the team pack stands on ALL FIVE exports -
    # STOCKTAKE, SALES_STOCK, TRANSACTIONS, RENTAL_STOCK and ON_HIRE.
    # RENTAL_STOCK gives the fleet position (out working v ready on the
    # shelf); ON_HIRE gives who's holding gear right now.
    fleet_su = {}
    infra_out = 0
    rs_bc = {}
    try:
        rs_path = find_newest("RENTAL_STOCK*.xlsx", "the RENTAL_STOCK "
                              "export", required=False)
        if rs_path:
            _, rs_raw = sheet_rows(rs_path, "RENTAL_STOCK",
                                   "the RENTAL_STOCK export")
            for r in rs_raw:
                su = clean_text(r.get("STORAGE_UNIT"))
                stat = clean_text(r.get("ITEM_STATUS"))
                _bc = clean_text(r.get("ITEM_BARCODE")).upper()
                if _bc:
                    rs_bc[_bc] = clean_text(r.get("ITEM_NUMBER"))
                if su.upper() in _ST_EXCLUDED_SU or not su:
                    if stat == "On Hire":
                        infra_out += 1
                    continue
                e = fleet_su.setdefault(su, {"fleet": 0, "out": 0,
                                             "ready": 0})
                e["fleet"] += 1
                if stat == "On Hire":
                    e["out"] += 1
                elif stat == "Available for Hire":
                    e["ready"] += 1
    except KitError:
        fleet_su = {}
    holders = {}
    try:
        oh_path = find_newest("ON_HIRE*.xlsx", "the ON_HIRE export",
                              required=False)
        if oh_path:
            _, oh_raw = sheet_rows(oh_path, "ON_HIRE", "the ON_HIRE export")
            for r in oh_raw:
                nm = clean_text(r.get("HIRER_NAME"))
                if not nm or is_site_plant_equipment(nm):
                    continue
                holders[nm] = holders.get(nm, 0) + 1
    except KitError:
        holders = {}
    if fleet_su:
        f_tot = sum(e["fleet"] for e in fleet_su.values())
        f_out = sum(e["out"] for e in fleet_su.values())
        f_rdy = sum(e["ready"] for e in fleet_su.values())
        body += "<h2>The shelf and the field &mdash; right now</h2>"
        deep = max(holders.items(), key=lambda kv: kv[1]) if holders else None
        body += ("<div class='intro story'>The live position your counting "
                 "keeps honest: of <b>{f:,} store fleet items</b>, "
                 "<b>{o:,} are out working</b>{ppl} and <b>{r:,} sit ready "
                 "on the shelf</b> - and a deep, counted shelf is the whole "
                 "promise: nobody waits for a tool at this window. "
                 "({inf:,} site infrastructure items - plant, barriers, "
                 "chutes - are out on charge besides.)</div>").format(
            f=f_tot, o=f_out,
            ppl=(" with <b>{:,} people</b> holding gear in their own name"
                 .format(len(holders)) if holders else ""),
            r=f_rdy, inf=infra_out)
        body += v3_kpis([
            ("store", "{:,}".format(f_tot), "Store fleet items", ""),
            ("swap", "{:,}".format(f_out), "Out working now", ""),
            ("box", "{:,}".format(f_rdy), "Ready on the shelf", ""),
            ("people", "{:,}".format(len(holders)), "People holding gear",
             ("<span style='color:{c};font-size:8pt;font-weight:700'>"
              "deepest list: {n} items</span>".format(
                  c=ORANGE_HEX, n=deep[1]) if deep else "")),
        ])
        rows_h = ""
        for su, e in sorted(fleet_su.items(), key=lambda kv: -kv[1]["fleet"])[:8]:
            rows_h += ("<tr><td>{s}</td><td class='num'>{f:,}</td>"
                       "<td class='num'>{o:,}</td><td class='num'>{r:,}</td>"
                       "<td class='num'>{p:.0%}</td></tr>").format(
                s=esc(su), f=e["fleet"], o=e["out"], r=e["ready"],
                p=(e["ready"] / e["fleet"]) if e["fleet"] else 0)
        body += ("<table class='data'><thead><tr><th>Storage unit</th>"
                 "<th class='num'>Fleet</th><th class='num'>Out working</th>"
                 "<th class='num'>Ready</th><th class='num'>Ready rate</th>"
                 "</tr></thead><tbody>" + rows_h + "</tbody></table>"
                 "<div class='note'>Fleet and status straight from "
                 "RENTAL_STOCK; holders from ON_HIRE by name. The ready "
                 "shelf is exactly what the counting wheel keeps honest - "
                 "{p:.0%} of it inside the cycle right now.</div>".format(
                     p=pct))

    # ================= THE STORE THROUGH THE DOOR =========================
    # (A. Fisher, 25 Jul 2026): transactions in the team pack - who came
    # over, what went out and when, day shift v night shift scored, quiet
    # hours highlighted, consumables included, safety always on the page.
    _txall = load_transactions()[0] or []
    _seen_t = set()
    txc = []
    for t in _txall:
        if is_site_plant_equipment(t.get("hirer", "")):
            continue
        k = (t.get("tx_id") or (t.get("barcode"), t.get("start"),
                                t.get("end"), t.get("hirer")))
        if k in _seen_t:
            continue
        _seen_t.add(k)
        txc.append(t)
    sales_mv = load_sales()[0] or []

    # one event = one person at the counter: an issue, a return or a draw
    events = []          # (date, hour|None, hirer, kind, barcode)
    for t in txc:
        if t.get("start"):
            events.append((t["start"], _hr_of(t.get("start_time")),
                           t.get("hirer") or "", "iss",
                           t.get("barcode") or ""))
        if t.get("end"):
            events.append((t["end"], _hr_of(t.get("end_time")),
                           t.get("hirer") or "", "ret",
                           t.get("barcode") or ""))
    for mv in sales_mv:
        if mv.get("date"):
            events.append((mv["date"], _hr_of(mv.get("time")),
                           mv.get("hirer") or "", "draw", ""))

    if events:
        kinds = {"iss": 0, "ret": 0, "draw": 0}
        S = {"day": dict(kinds), "night": dict(kinds)}
        S_ppl = {"day": set(), "night": set()}
        hours_k = {"iss": {}, "ret": {}, "draw": {}}
        val_shift = {"day": 0.0, "night": 0.0}
        val_missing = 0
        hours24 = {}
        visits = set()
        people = set()
        unstamped = 0
        for d, h, who, kind, bc in events:
            kinds[kind] += 1
            if who:
                people.add(who)
            if h is None:
                unstamped += 1
                continue
            w = _shift_window(h)
            S[w][kind] += 1
            if who:
                S_ppl[w].add(who)
                # a visit belongs to the shift's OWN day - a 02:00 visit
                # is the previous day's night shift
                visits.add((who, _shift_day(d, h), w))
            hours24[h] = hours24.get(h, 0) + 1
            hours_k[kind][h] = hours_k[kind].get(h, 0) + 1
            if kind == "iss":
                _it = rs_bc.get((bc or "").upper())
                _pr = MASTER.price(_it) if _it else None
                if _pr:
                    val_shift[w] += _pr
                else:
                    val_missing += 1
        draw_qty = sum((mv.get("qty") or 0) for mv in sales_mv
                       if mv.get("date"))

        body += "<h2>The store through the door &mdash; this shut</h2>"
        body += ("<div class='intro story'>The stocktake keeps the shelves "
                 "honest - THIS is the traffic those shelves serve. "
                 "<b>{i:,} items issued</b>, <b>{r:,} returned</b> and "
                 "<b>{c:,} consumable draws</b> ({q:g} items) have gone "
                 "over the counter, for <b>{p:,} different people</b> "
                 "across <b>{v:,} store visits</b>. Every one of those is "
                 "a crew member equipped and back on the tools.</div>"
                 ).format(i=kinds["iss"], r=kinds["ret"], c=kinds["draw"],
                          q=draw_qty, p=len(people), v=len(visits))
        body += v3_kpis([
            ("swap", "{:,}".format(kinds["iss"]), "Items issued", ""),
            ("tick", "{:,}".format(kinds["ret"]), "Items returned", ""),
            ("box", "{:,}".format(kinds["draw"]), "Consumable draws", ""),
            ("people", "{:,}".format(len(people)), "People served",
             "<span style='color:{c};font-size:8pt;font-weight:700'>"
             "{v:,} visits</span>".format(c=ORANGE_HEX, v=len(visits))),
        ])
        body += ("<div class='note'>A visit = one person at the counter in "
                 "one shift window on one day - morning and night runs "
                 "count twice, fair's fair. {u}</div>").format(
            u=("{:,} event(s) carry no time stamp - counted in the totals, "
               "left out of the shift split.".format(unstamped))
            if unstamped else "Every event carries a SiteIQ time stamp.")

        # ---- the scoreboard ----------------------------------------------
        day_ev = sum(S["day"].values())
        night_ev = sum(S["night"].values())
        rows_sc = [
            ("Items issued", S["day"]["iss"], S["night"]["iss"], False),
            ("Items returned", S["day"]["ret"], S["night"]["ret"], False),
            ("Consumable draws", S["day"]["draw"], S["night"]["draw"],
             False),
            ("People served", len(S_ppl["day"]), len(S_ppl["night"]),
             False),
            ("Store visits",
             sum(1 for _w, _d, w in visits if w == "day"),
             sum(1 for _w, _d, w in visits if w == "night"), False),
            ("Gear value issued (replacement $)",
             val_shift["day"], val_shift["night"], True),
        ]
        rd = sum(1 for _l, a, b, _m in rows_sc if a > b)
        rn = sum(1 for _l, a, b, _m in rows_sc if b > a)
        body += "<h2>Day shift v night shift &mdash; the scoreboard</h2>"
        trs = ""
        for lab, a, b, money in rows_sc:
            wd = a > b
            wn = b > a
            fa = "${:,.0f}".format(a) if money else "{:,}".format(a)
            fb = "${:,.0f}".format(b) if money else "{:,}".format(b)
            trs += (
                "<tr><td>{l}</td>"
                "<td class='num' style='{sa}'>{a}{pa}</td>"
                "<td class='num' style='{sb}'>{b}{pb}</td></tr>").format(
                l=esc(lab), a=fa, b=fb,
                sa=("color:{};font-weight:800".format(_DAY_COL) if wd else ""),
                sb=("color:{};font-weight:800".format(_NIGHT_COL) if wn else ""),
                pa=(" <span style='background:{};color:#fff;border-radius:8px;"
                    "padding:1px 7px;font-size:8pt;font-weight:800'>takes it"
                    "</span>".format(_DAY_COL) if wd else ""),
                pb=(" <span style='background:{};color:#fff;border-radius:8px;"
                    "padding:1px 7px;font-size:8pt;font-weight:800'>takes it"
                    "</span>".format(_NIGHT_COL) if wn else ""))
        body += ("<table class='data'><thead><tr><th></th>"
                 "<th class='num'>{sun} DAY SHIFT <span style='color:#8B93A1;"
                 "font-weight:400'>06:00-18:00</span></th>"
                 "<th class='num'>{moon} NIGHT SHIFT <span style='color:#8B93A1;"
                 "font-weight:400'>18:00-06:00</span></th></tr></thead>"
                 "<tbody>" + trs + "</tbody></table>").format(
            sun=_sun_ico(15), moon=_moon_ico(15))
        if day_ev and night_ev:
            if rd > rn:
                verdict = ("<b style='color:{dc}'>Day shift takes the round, "
                           "{rd}-{rn}.</b> Night crew - {ne:,} counter events "
                           "to {de:,} is the cruisier run. Enjoy it while it "
                           "lasts.").format(dc=_DAY_COL, rd=rd, rn=rn,
                                            ne=night_ev, de=day_ev)
            elif rn > rd:
                verdict = ("<b style='color:{nc}'>Night shift takes the "
                           "round, {rn}-{rd}.</b> Day crew - that's a "
                           "gauntlet thrown under lights.").format(
                    nc=_NIGHT_COL, rn=rn, rd=rd)
            else:
                verdict = ("<b>Dead heat, {rd}-{rn}.</b> Both windows "
                           "pulling their weight.").format(rd=rd, rn=rn)
        else:
            q = "night" if night_ev == 0 else "day"
            verdict = ("Every stamped counter event this shut sits in {w} "
                       "hours - the {q} window is quiet on the counter so "
                       "far, so the round goes uncontested.").format(
                w="day" if q == "night" else "night", q=q)
        body += ("<div class='intro story'>{v} Either way: one store, one "
                 "team - the scoreboard's for bragging rights, the gear's "
                 "for the job.</div>").format(v=verdict)
        if val_missing:
            body += ("<div class='note'>Gear value = master-file "
                     "replacement cost of items issued in each window; "
                     "{m:,} issued item(s) carry no master price yet and "
                     "are left out of the dollars, never guessed.</div>"
                     ).format(m=val_missing)
        # last-7-days paired bars, day v night - each event credited to
        # the shift's OWN day (night belongs to the day it starts)
        last7 = [today - dt.timedelta(days=i) for i in range(6, -1, -1)]
        day_daily = {}
        night_daily = {}
        for d, h, _who, _k, _bc in events:
            if h is None:
                continue
            sd = _shift_day(d, h)
            (day_daily if _shift_window(h) == "day" else
             night_daily)[sd] = (day_daily if _shift_window(h) == "day"
                                 else night_daily).get(sd, 0) + 1
        body += v3_bars_pair(
            [(d.strftime("%a"), day_daily.get(d, 0)) for d in last7],
            [(d.strftime("%a"), night_daily.get(d, 0)) for d in last7],
            "Day shift", "Night shift", _DAY_COL, _NIGHT_COL)

        # ---- the 24-hour clock -------------------------------------------
        body += "<h2>When the counter runs hot &mdash; and when it sleeps</h2>"
        body += _hourbars24(hours24, "Every stamped counter event this "
                                     "shut - issues, returns and "
                                     "consumable draws, by hour of day")
        tops = sorted(hours24.items(), key=lambda kv: -kv[1])[:3]
        chips = "".join(
            "<span style='display:inline-block;background:#1A2027;"
            "border:1px solid {c};color:#fff;border-radius:999px;"
            "padding:4px 14px;margin:4px 6px 0 0;font-size:10pt'>"
            "<b style='color:{c}'>{h:02d}:00</b> &middot; {n} events"
            "</span>".format(
                c=_DAY_COL if _shift_window(h) == "day" else _NIGHT_COL,
                h=h, n=n) for h, n in tops)
        body += ("<div class='intro'><b>The rush:</b> {chips}</div>"
                 ).format(chips=chips)
        qs = _quiet_stretch(hours24)
        if qs:
            body += ("<div class='intro'><b>The lull:</b> {a:02d}:00 to "
                     "{b:02d}:00 barely moves - that's the window for "
                     "restocks, counts and getting ahead of the next rush. "
                     "The quiet hours are working hours too, just quieter "
                     "ones.</div>").format(a=qs[0], b=qs[1])

        # ---- hour by hour: the shift ledger ------------------------------
        # (A. Fisher, 25 Jul 2026): the system stamps everything - so show
        # everything. Each metric by the hour, then the full ledger with
        # day and night subtotals. Good competitive data, no guesswork.
        body += "<h2>Hour by hour &mdash; the shift ledger</h2>"
        body += ("<div class='intro'>Four stories, one clock: stocktake "
                 "counts, gear out, gear back and consumable sales, each "
                 "by the hour it was stamped. Orange bars are day-window "
                 "hours, blue are night's.</div>")
        body += _hourbars24(by_hour, "Stocktake counts by hour", h_px=140)
        body += _hourbars24(hours_k["iss"], "Gear issued by hour",
                            h_px=140)
        body += _hourbars24(hours_k["ret"], "Gear returned by hour",
                            h_px=140)
        body += _hourbars24(hours_k["draw"],
                            "Consumable sales by hour", h_px=140)
        led = ""
        subt = {"day": [0, 0, 0, 0], "night": [0, 0, 0, 0]}
        for h in list(range(6, 24)) + list(range(0, 6)):
            w = _shift_window(h)
            vals = [by_hour.get(h, 0), hours_k["iss"].get(h, 0),
                    hours_k["ret"].get(h, 0), hours_k["draw"].get(h, 0)]
            for i, v in enumerate(vals):
                subt[w][i] += v
            tot = sum(vals)
            if h == 6:
                led += ("<tr><td colspan='6' style='color:{c};font-weight"
                        ":800;letter-spacing:1px'>{ico} DAY WINDOW "
                        "06:00-18:00</td></tr>").format(c=_DAY_COL,
                                                        ico=_sun_ico(13))
            if h == 18:
                led += ("<tr><td colspan='6' style='color:{c};font-weight"
                        ":800;letter-spacing:1px'>{ico} NIGHT WINDOW "
                        "18:00-06:00 (belongs to the day it starts)</td>"
                        "</tr>").format(c=_NIGHT_COL, ico=_moon_ico(13))
            dim = " style='color:#5A6472'" if tot == 0 else ""
            led += ("<tr{dim}><td class='num'>{h:02d}:00</td>"
                    "<td class='num'>{c}</td><td class='num'>{i}</td>"
                    "<td class='num'>{r}</td><td class='num'>{s}</td>"
                    "<td class='num'><b>{t}</b></td></tr>").format(
                dim=dim, h=h, c=vals[0], i=vals[1], r=vals[2],
                s=vals[3], t=tot)
            if h == 17:
                led += ("<tr style='color:{c};font-weight:800'>"
                        "<td>DAY TOTAL</td><td class='num'>{c0:,}</td>"
                        "<td class='num'>{c1:,}</td><td class='num'>{c2:,}"
                        "</td><td class='num'>{c3:,}</td><td class='num'>"
                        "{ct:,}</td></tr>").format(
                    c=_DAY_COL, c0=subt["day"][0], c1=subt["day"][1],
                    c2=subt["day"][2], c3=subt["day"][3],
                    ct=sum(subt["day"]))
        led += ("<tr style='color:{c};font-weight:800'>"
                "<td>NIGHT TOTAL</td><td class='num'>{c0:,}</td>"
                "<td class='num'>{c1:,}</td><td class='num'>{c2:,}</td>"
                "<td class='num'>{c3:,}</td><td class='num'>{ct:,}</td>"
                "</tr>").format(
            c=_NIGHT_COL, c0=subt["night"][0], c1=subt["night"][1],
            c2=subt["night"][2], c3=subt["night"][3],
            ct=sum(subt["night"]))
        body += ("<table class='data'><thead><tr><th class='num'>Hour</th>"
                 "<th class='num'>Counts</th><th class='num'>Issued</th>"
                 "<th class='num'>Returned</th><th class='num'>Sales</th>"
                 "<th class='num'>All</th></tr></thead><tbody>" + led
                 + "</tbody></table>")
        opp = []
        if night_ev == 0:
            opp.append("The open goal: the night window has the counter "
                       "to itself and a clean board - even a modest night "
                       "of counting and returns puts blue on every chart "
                       "in this pack.")
        elif day_ev and (max(day_ev, night_ev) / float(day_ev + night_ev)
                         > 0.70):
            lead_w = "day" if day_ev > night_ev else "night"
            opp.append("Balance check: {w} window is carrying {p:.0%} of "
                       "the counter - worth a look at what can shift "
                       "across the handover.".format(
                           w=lead_w,
                           p=max(day_ev, night_ev) / float(day_ev
                                                           + night_ev)))
        if opp:
            body += ("<div class='intro'><b>Opportunity:</b> "
                     + " ".join(opp) + "</div>")

        # ---- consumables by shift ----------------------------------------
        cons_shift = {"day": {}, "night": {}}
        for mv in sales_mv:
            if not mv.get("date"):
                continue
            h = _hr_of(mv.get("time"))
            if h is None:
                continue
            w = _shift_window(h)
            k = mv.get("sku") or "?"
            cons_shift[w][k] = cons_shift[w].get(k, 0) + 1
        body += "<h2>Consumables through the window</h2>"
        nd = sum(cons_shift["day"].values())
        nn = sum(cons_shift["night"].values())
        body += ("<div class='intro'>Not just the tooling - the shelf "
                 "moves too. <b style='color:{dc}'>{nd:,} draws in day "
                 "hours</b> v <b style='color:{nc}'>{nn:,} at night</b>."
                 "</div>").format(dc=_DAY_COL, nd=nd, nc=_NIGHT_COL, nn=nn)
        for w, col, ico in (("day", _DAY_COL, _sun_ico(14)),
                            ("night", _NIGHT_COL, _moon_ico(14))):
            top = sorted(cons_shift[w].items(), key=lambda kv: -kv[1])[:3]
            if top:
                mxv = top[0][1]
                body += ("<div style='margin-top:8px;font-weight:700'>"
                         "{ico} What {w} shift lives on</div>").format(
                    ico=ico, w=w)
                for k, n in top:
                    body += hbar(k[:44], "{} draws".format(n),
                                 n / float(mxv), col)

        # ---- safety never knocks off -------------------------------------
        gas_mv = sum(1 for t in txc
                     if "GAS MONITOR" in (t.get("description") or "").upper())
        rad_mv = sum(1 for t in txc
                     if is_radio_handset({"storage_unit": "RADIOS",
                                          "description": t.get("description")
                                          or ""}))
        body += "<h2>Safety never knocks off</h2>"
        body += (
            "<div class='intro story'><b style='color:{o}'>The store's Life "
            "Saving Rule - Tools and Equipment:</b> \"I will only use tools "
            "and equipment that are fit for purpose and for which I am "
            "trained, competent and authorised. I will ensure that damaged "
            "or defective equipment is tagged out and removed from "
            "service.\" That rule lives at this counter: damaged gear? Over "
            "the counter, tagged out, no questions - that's the system "
            "working, not a telling-off.</div>"
            "<div class='intro'>The high-care kit keeps its own rules: "
            "<b>{g:,} gas monitor movement{gs}</b> through the counter this "
            "shut - every monitor on the daily bump-test rule before it "
            "works - and <b>{r:,} radio movement{rs}</b>, serial-tracked in "
            "the holder's name. And Fit for Work stands every shift: right "
            "PPE, fit to be on the tools, day window or night.</div>"
        ).format(o=ORANGE_HEX, g=gas_mv, gs="" if gas_mv == 1 else "s",
                 r=rad_mv, rs="" if rad_mv == 1 else "s")

    limits = (
        "Counts are sightings logged with the Stocktake action in the "
        "SiteIQ STOCKTAKE export. The export keeps each item's LATEST "
        "sighting only, so day-by-day figures read 'items whose newest "
        "count landed that day' - a fair floor on the real effort, never "
        "an inflation. ALL shift credit follows SiteIQ's own time stamps "
        "against Andrew's windows - work stamped 06:00-18:00 is day-shift "
        "work whoever did it, 18:00-06:00 is night's, and a night shift "
        "belongs to the day it STARTS. The roster on the team page is who "
        "people are, never how work is credited. Session pace counts runs "
        "of 3+ counts with gaps under 5 minutes. Store traffic comes from "
        "the TRANSACTIONS export "
        "(duplicates dropped, the site idle pool excluded) plus consumable "
        "draws from SALES_STOCK; shift splits use SiteIQ's own time "
        "stamps against Andrew's windows - day 06:00-18:00, night "
        "18:00-06:00 - and events without a stamp are counted in totals "
        "but never guessed into a shift. The live shelf-and-field "
        "position reads RENTAL_STOCK (fleet + status, site infrastructure "
        "split out) and ON_HIRE (current holders by name) - the pack "
        "stands on all five SiteIQ exports.")
    hero = "{tc:,} counts &bull; {p:.0%} on cycle &bull; {due} to go".format(
        tc=total_counts, p=pct, due=sm["due"])
    inner = [_e_note(esc("The team's stocktake story - what's been counted, "
                         "by whom, and the fair split of what's left."),
                     "Stores team."),
             _e_h2("The counters"),
             _e_tbl(["Name", "When", "Counts"],
                    [[n, ("days + nights" if e["day"] and e["night"]
                          else "night hours" if e["night"]
                          else "day hours"), "{:,}".format(e["n"])]
                     for n, e in prs[:8]])]
    emit_report("Coates_K2_Stocktake_TEAM_{}".format(date_tag),
                "Stores Stocktake - Team Report",
                "The People Turning The Wheel",
                body, hero, inner, limits,
                "Coates K2 - Stocktake Team Report - {} - {:,} counts, "
                "{:.0%} on cycle".format(
                    generated.strftime("%d %b %Y"), total_counts, pct),
                asof, generated, source_line, report_tag="STOCKTAKE")


# ---- STORE CONSUMABLES WATCH ----------------------------------------------
# The stores team's shelf-health pack (A. Fisher, 25 Jul 2026): what we
# hold, what's moving, how long it lasts at today's pace, which lines to
# watch, and PROOF the checking is happening - check coverage and the
# counted-v-system match rate from the STOCKTAKE export's consumable
# lines. Positive for a team doing it right; honest about every method.

def generate_store_health(sales, sales_loaded, asof, generated, source_line,
                          date_tag):
    shelf, shelf_path = load_store_shelf()
    if not shelf:
        print("  ! SALES_STOCK export not found - store consumables watch "
              "skipped.")
        return
    sm_st = load_stocktake_model()
    today = dt.date.today()

    # burn per line: average daily draws over the last 7 days with the
    # shut running; cover = on-hand / burn
    daily = {}
    span_days = 7
    for mv in sales:
        if not mv.get("date"):
            continue
        age = (today - mv["date"]).days
        if 0 <= age < span_days:
            k = mv["sku"].upper()
            daily[k] = daily.get(k, 0.0) + (mv["qty"] or 0)
    lines = []
    for desc, e in shelf.items():
        k = desc.upper()
        burn = daily.get(k, 0.0) / float(span_days)
        cover = (e["avail"] / burn) if burn > 0 else None
        lines.append({"desc": desc, "avail": e["avail"], "min": e["min"],
                      "burn": burn, "cover": cover})
    stocked = [l for l in lines if l["avail"] > 0 or l["min"] > 0]
    movers = sorted([l for l in lines if l["burn"] > 0],
                    key=lambda l: -l["burn"])
    watch = sorted(
        [l for l in movers
         if (l["cover"] is not None and l["cover"] < 3)
         or (l["min"] > 0 and l["avail"] <= l["min"])],
        key=lambda l: (l["cover"] if l["cover"] is not None else 99))
    out_now = [l for l in stocked if l["avail"] <= 0 and l["burn"] > 0]
    healthy = [l for l in movers if l not in watch]

    # stock checks on consumable lines, from the stocktake export
    checks = []
    if sm_st:
        for x in sm_st["rows"]:
            if x["su"].upper().startswith("CONSUMABLE") and x["when"]:
                checks.append(x)
    checked7 = [x for x in checks
                if (sm_st["anchor"].date() - x["when"].date()).days <= 7]
    chk_pct = (len(checked7) / len(checks)) if checks else None
    # counted v system: the system figure keeps moving with every draw
    # AFTER the check, so a small gap is the shelf living, not a miss.
    # Only a gap that's big in both units AND proportion earns a recount.
    match_rows = []
    n_recount = 0
    for x in checks:
        sysq = None
        for desc, e in shelf.items():
            if desc.upper() == x["desc"].upper():
                sysq = e["avail"]
                break
        if sysq is not None and x["qty"] is not None:
            gap = (x["qty"] or 0) - sysq
            big = abs(gap) >= max(5.0, 0.2 * max(x["qty"] or 0, sysq))
            if big:
                n_recount += 1
            match_rows.append((x, sysq, gap, big))

    story = (
        "This one is for the people behind the counter. A shelf the crews "
        "never think about is a shelf being run WELL - and that's what "
        "this page shows. <b>{st:,} consumable lines are stocked</b>, "
        "{mv} of them moving this week. {chk} {out}"
    ).format(
        st=len(stocked), mv=len(movers),
        chk=("<b>{}</b> of the consumable lines in the stocktake wheel "
             "were checked inside the last 7 days{mr}.".format(
                 "{:.0%}".format(chk_pct) if chk_pct is not None else "-",
                 mr=(", and every counted figure sits within normal "
                     "draw-drift of the system" if n_recount == 0 else
                     ", with {} line{} worth a recount below".format(
                         n_recount, "" if n_recount == 1 else "s"))
                 if match_rows else "") if checks else
             "Consumable lines aren't in the stocktake export yet - the "
             "check metrics below switch on the day they are."),
        out=("Nothing that moves is out of stock. That's the job, done "
             "well." if not out_now else
             "<b>{} moving line{} at zero</b> - top of the watch list "
             "below.".format(len(out_now),
                             "" if len(out_now) == 1 else "s")))
    body = "<div class='intro story'>" + story + "</div>"
    body += v3_kpis([
        ("store", "{:,}".format(len(stocked)), "Lines stocked", ""),
        ("swap", "{:,}".format(len(movers)), "Moving this week", ""),
        ("tick", "{:.0%}".format(chk_pct) if chk_pct is not None else "-",
         "Checked in last 7 days", ""),
        ("alert", "{:,}".format(len(watch)), "On the watch list",
         "<span style='color:{c};font-size:8pt;font-weight:700'>{t}</span>".format(
             c=V3_BAD if watch else V3_GOOD,
             t="worst first below" if watch else "shelf holding")),
    ])

    # ---- stock checks -----------------------------------------------------
    body += "<h2>Stock checks &mdash; proof the shelf is watched</h2>"
    if checks:
        body += ("<div style='text-align:center'>"
                 + ring(chk_pct or 0, "{} of {} lines checked, last 7 days"
                        .format(len(checked7), len(checks)), "checked")
                 + "</div>")
        rows_h = ""
        for x, sysq, gap, big in sorted(match_rows,
                                        key=lambda t: t[0]["when"],
                                        reverse=True)[:14]:
            if big:
                call = ("<span style='color:{}'>worth a recount</span>"
                        .format(V3_AMBER))
            elif abs(gap) < 0.5:
                call = ("<span style='color:{}'>&#10003; spot on</span>"
                        .format(V3_GOOD))
            else:
                call = ("<span style='color:{}'>drift {:+g} - draws since "
                        "the check</span>".format(V3_GOOD, -gap))
            rows_h += ("<tr><td>{d}</td><td class='num'>{w}</td>"
                       "<td>{by}</td><td class='num'>{q}</td>"
                       "<td class='num'>{s}</td><td>{r}</td></tr>").format(
                d=esc(x["desc"]), w=x["when"].strftime("%d %b %H:%M"),
                by=esc(x["by"] or "-"),
                q="{:g}".format(x["qty"]) if x["qty"] is not None else "-",
                s="{:g}".format(sysq), r=call)
        body += ("<table class='data'><thead><tr><th>Line</th>"
                 "<th class='num'>Checked</th><th>By</th>"
                 "<th class='num'>Counted</th><th class='num'>System now</th>"
                 "<th>Call</th></tr></thead><tbody>" + rows_h
                 + "</tbody></table>"
                 "<div class='note'>'System now' keeps moving with every "
                 "draw AFTER the check, so a small gap is the shelf "
                 "living, not a miss. Only a gap that's big in both units "
                 "and proportion earns a recount - and a recount isn't a "
                 "fail, it's the system working.</div>")
    else:
        body += ("<div class='intro'>The stocktake export carries no "
                 "consumable lines yet. The moment consumable counts are "
                 "scanned through the stocktake screen, this page scores "
                 "the checking automatically - coverage, match rate, and "
                 "who checked what.</div>")

    # ---- days of cover ----------------------------------------------------
    body += "<h2>Days of cover &mdash; how long the shelf lasts</h2>"
    body += ("<div class='intro'>Cover = what's on hand divided by the "
             "line's average daily draws over the last {d} days. Under "
             "3 days lands a line on the watch list; lines with no draws "
             "this week aren't scored - no movement, no risk.</div>"
             ).format(d=span_days)
    if watch:
        rows_h = ""
        for l in watch[:14]:
            cv = ("{:.1f}d".format(l["cover"])
                  if l["cover"] is not None else "-")
            rows_h += ("<tr><td>{d}</td><td class='num'>{a:g}</td>"
                       "<td class='num'>{m:g}</td><td class='num'>{b:.1f}"
                       "</td><td class='num'>{c}</td><td>{f}</td></tr>"
                       ).format(
                d=esc(l["desc"]), a=l["avail"], m=l["min"], b=l["burn"],
                c=cv,
                f=("<span style='color:{}'>order today</span>".format(V3_BAD)
                   if (l["cover"] is not None and l["cover"] < 1.5)
                   or l["avail"] <= 0 else
                   "<span style='color:{}'>watch</span>".format(V3_AMBER)))
        body += ("<table class='data'><thead><tr><th>Line</th>"
                 "<th class='num'>On hand</th><th class='num'>Min</th>"
                 "<th class='num'>Draws/day</th><th class='num'>Cover</th>"
                 "<th>Call</th></tr></thead><tbody>" + rows_h
                 + "</tbody></table>")
        body += "".join(
            hbar(l["desc"][:40],
                 "{:.1f}d".format(l["cover"]) if l["cover"] is not None
                 else "0 on hand",
                 min(1.0, (l["cover"] or 0) / 5.0),
                 rag_colour(min(1.0, (l["cover"] or 0) / 5.0)))
            for l in watch[:8])
    else:
        body += ("<div class='intro'><b>Every moving line holds 3+ days "
                 "of cover.</b> That's a shelf run properly - the fast "
                 "movers below are the ones that keep earning it.</div>")
    if healthy:
        body += ("<div class='note'>{h} moving line{hs} sit comfortably "
                 "above the line - checked, stocked, no action needed."
                 "</div>").format(h=len(healthy),
                                  hs="" if len(healthy) == 1 else "s")

    # ---- the fast movers --------------------------------------------------
    body += "<h2>The fast movers &mdash; what the crews live on</h2>"
    body += "".join(
        hbar(l["desc"][:40], "{:.1f}/day".format(l["burn"]),
             l["burn"] / max(0.001, movers[0]["burn"]))
        for l in movers[:10]) or ("<div class='intro'>No draws recorded "
                                  "this week.</div>")

    # ---- when the shelf gets hit ------------------------------------------
    hrs_d = {}
    for mv in sales:
        if mv.get("date"):
            h = _hr_of(mv.get("time"))
            if h is not None:
                hrs_d[h] = hrs_d.get(h, 0) + 1
    if hrs_d:
        body += "<h2>When the shelf gets hit</h2>"
        body += _hourbars24(hrs_d, "Consumable draws by hour of day - "
                                   "every scanned draw this shut")
        qs2 = _quiet_stretch(hrs_d)
        if qs2:
            body += ("<div class='intro'>The shelf's quiet window runs "
                     "<b>{a:02d}:00 to {b:02d}:00</b> - that's the restock "
                     "window. Fill the hooks in the lull and the rush never "
                     "meets an empty one. And the PPE on these hooks IS the "
                     "Fit for Work rule in stock form - gloves, glasses and "
                     "plugs on the shelf keep every shift working safe."
                     "</div>").format(a=qs2[0], b=qs2[1])

    # ---- the watch method -------------------------------------------------
    top3 = ", ".join(esc(l["desc"]) for l in movers[:3]) or "the fast movers"
    body += (
        "<h2>Keeping an eye on it &mdash; the three glances</h2>"
        "<div class='intro'><b>Morning glance:</b> the watch list above - "
        "anything under 3 days of cover gets ordered before smoko. "
        "<b>Midday glance:</b> the fast movers ({t3}) - they burn "
        "hardest and empty fastest. <b>Close-of-shift glance:</b> "
        "tomorrow's first-hour lines - gloves, blades, tape - stocked "
        "tonight so the 06:00 rush never meets an empty hook. SiteIQ "
        "minimums are already set on {mm:,} line{ms}; this page just "
        "makes the walk faster.</div>").format(
        t3=top3, mm=len([l for l in lines if l["min"] > 0]),
        ms="" if len([l for l in lines if l["min"] > 0]) == 1 else "s")

    limits = (
        "Shelf position and minimums come from the SALES_STOCK export; "
        "draws are scanned counter movements from the same export. Burn "
        "and cover use the last {d} days of draws - lines with no draws "
        "this week are not scored. Check coverage and the match rate come "
        "from consumable lines in the STOCKTAKE export; counted-v-system "
        "gaps allow for draws made after the check, and only a gap big in "
        "both units and proportion is flagged for a recount. Where no "
        "consumable checks exist the page says so instead of guessing. "
        "Nothing here is an automatic order or charge.").format(d=span_days)
    hero = "{st:,} lines &bull; {w} on watch &bull; {c} checked 7d".format(
        st=len(stocked), w=len(watch),
        c="{:.0%}".format(chk_pct) if chk_pct is not None else "-")
    inner = [_e_note(esc("The shelf behind the counter - cover, checks "
                         "and the watch list."), "Stores team."),
             _e_h2("Watch list"),
             _e_tbl(["Line", "On hand", "Cover"],
                    [[l["desc"][:36], "{:g}".format(l["avail"]),
                      "{:.1f}d".format(l["cover"])
                      if l["cover"] is not None else "-"]
                     for l in watch[:8]] or [["Nothing on watch", "-", "-"]])]
    emit_report("Coates_K2_Store_Consumables_WATCH_{}".format(date_tag),
                "Store Consumables - Stock Watch",
                "The Shelf Behind The Counter",
                body, hero, inner, limits,
                "Coates K2 - Store Consumables Watch - {} - {} lines, "
                "{} on watch".format(generated.strftime("%d %b %Y"),
                                     len(stocked), len(watch)),
                asof, generated, source_line, report_tag="STORE")


# ---- MY GEAR MAIL: each person's own scorecard, emailed to them -----------
# "Like My Gear works", but delivered (A. Fisher, 24 Jul 2026): everyone
# with gear on hire AND an email in the People (Email) sheet gets their own
# one-page scorecard drafted to them personally - their list oldest-first,
# the standing ask, and any damage recorded in their name. No email = simply
# skipped, never guessed. Nothing sends without the 06 YES, same as always.
# NO dollar figures on personal emails - the list and the ask only.

def _person_key(s):
    return re.sub(r"\s+", " ", re.sub(r"\s*-\s*", " ",
                                      str(s or "").strip())).upper()


def load_people_emails():
    """Hirer Name -> email from the recipients book's People (Email) sheet
    (Send OK = Yes and a real address only)."""
    out = {}
    hits = sorted(glob.glob(os.path.join(KIT_DIR, "Coates_Report_Recipients*.xlsx")),
                  key=os.path.getmtime, reverse=True)
    hits = [h for h in hits if not os.path.basename(h).startswith("~$")]
    if not hits:
        return out
    try:
        wb = openpyxl.load_workbook(hits[0], read_only=True, data_only=True)
        if "People (Email)" in wb.sheetnames:
            rows = list(wb["People (Email)"].iter_rows(values_only=True))
            hdr = None
            for r in rows:
                vals = [clean_text(v) for v in r]
                if hdr is None:
                    if vals and vals[0].lower() == "hirer name":
                        hdr = {v.lower(): i for i, v in enumerate(vals)}
                    continue

                def cell(name):
                    i = hdr.get(name)
                    return vals[i] if i is not None and i < len(vals) else ""
                email_addr = cell("email")
                if ("@" in email_addr
                        and cell("send ok").strip().lower() == "yes"
                        and cell("hirer name")):
                    out[_person_key(cell("hirer name"))] = email_addr
        wb.close()
    except Exception:
        return {}
    return out


def load_person_history():
    """Every transaction each person has run through the tool store - with
    the TIMES - straight from the TRANSACTIONS export. Keyed by normalised
    hirer name. Empty-safe."""
    path = find_newest("TRANSACTIONS*.xlsx", "the TRANSACTIONS export",
                       required=False)
    out = {}
    if not path:
        return out
    for sheet in ("TRANSACTION_CHARGES", "CUSTOMER_CONTRACTOR_EQUIP"):
        try:
            _, raw = sheet_rows(path, sheet, "the TRANSACTIONS export")
        except KitError:
            continue
        for r in raw:
            who = clean_text(r.get("HIRER_NAME"))
            if not who or is_site_plant_equipment(who):
                continue
            k = _person_key(who)
            desc = clean_text(r.get("SKU/ITEM DESCRIPTION"))
            sd = parse_date_au(r.get("TRAN_START_DATE"))
            ed = parse_date_au(r.get("TRAN_END_DATE"))
            st = clean_text(r.get("TRAN_START_TIME"))
            et = clean_text(r.get("TRAN_END_TIME"))
            cat = clean_text(r.get("PRODUCT_CATEGORY"))
            try:
                qty = int(float(r.get("QUANTITY") or 1))
            except (TypeError, ValueError):
                qty = 1
            key = (clean_text(r.get("SKU/ITEM_NUMBER")) or desc,
                   sd.isoformat() if sd else "", sheet)
            out.setdefault(k, {}).setdefault("rows", []).append({
                "desc": desc, "start": sd, "end": ed, "start_t": st,
                "end_t": et, "cat": cat, "qty": max(1, qty), "key": key})
    for k, h in out.items():
        rows = h["rows"]
        # de-duplicate the same movement appearing on both sheets
        seen = set()
        uniq = []
        for r in rows:
            if r["key"] in seen:
                continue
            seen.add(r["key"])
            uniq.append(r)
        h["rows"] = uniq
        stamps = []
        for r in uniq:
            for d, t in ((r["start"], r["start_t"]), (r["end"], r["end_t"])):
                if d:
                    stamps.append((d, t))
        h["first"] = min(stamps) if stamps else None
        h["last"] = max(stamps) if stamps else None
        h["days_active"] = len({d for d, _t in stamps})
        hire_rows = [r for r in uniq if r["cat"] != "Consumable"]
        h["returns"] = sum(1 for r in hire_rows if r["end"])
        h["sameday"] = sum(1 for r in hire_rows
                           if r["end"] and r["start"] == r["end"])
        h["hire_n"] = len(hire_rows)
        h["radio"] = [sum(1 for r in hire_rows if "RADIO" in r["desc"].upper()),
                      sum(1 for r in hire_rows
                          if "RADIO" in r["desc"].upper() and r["end"])]
        h["gas"] = [sum(1 for r in hire_rows if is_gas_monitor_desc(r["desc"])),
                    sum(1 for r in hire_rows
                        if is_gas_monitor_desc(r["desc"]) and r["end"])]
        cons = {}
        for r in uniq:
            if r["cat"] == "Consumable":
                nm = re.sub(r"\s*-\s*K2/?\d*$", "", r["desc"]).strip(" -")[:38]
                cons[nm] = cons.get(nm, 0) + r["qty"]
        h["cons"] = sorted(cons.items(), key=lambda kv: -kv[1])
    return out


def _fmt_stamp(stamp):
    """(date, 'HH:MM[:SS]') -> '13 Jul 2026 at 06:42' - honest to the data:
    no time recorded shows the date alone."""
    d, t = stamp
    t = (t or "").strip()
    if t:
        t2 = t[:5] if re.match(r"^\d{2}:\d{2}", t) else t
        return "{} at {}".format(d.strftime("%d %b %Y"), t2)
    return d.strftime("%d %b %Y")


def generate_mygear_mail(models, asof, generated, source_line, date_tag):
    emails = load_people_emails()
    dmg_people = load_damage_people()
    history = load_person_history()
    people = {}
    for m in models.values():
        for r in m["rows"]:
            who = clean_text(r.get("hirer"))
            if not who or "SITE PLANT EQUIPMENT" in who.upper():
                continue
            k = _person_key(who)
            p = people.setdefault(k, {"name": who, "company": m["display"],
                                      "items": []})
            p["items"].append(r)
    if not people:
        print("  My Gear mail: no one currently has gear on hire.")
        return
    drafted = 0
    skipped = []
    for k, p in sorted(people.items()):
        email_addr = emails.get(k)
        if not email_addr:
            skipped.append(p["name"])
            continue
        first = esc(p["name"].split()[0].title())
        display = esc(re.sub(r"\s*-\s*", " ", p["name"]))
        h = history.get(k) or {}
        items = sorted(p["items"],
                       key=lambda r: -(r["days"] if r["days"] is not None else -1))
        oldest = next((r["days"] for r in items if r["days"] is not None), None)
        out_now = len(items)
        rets = h.get("returns", 0)
        same = h.get("sameday", 0)
        had = out_now + rets
        score = None
        if rets > 0 and had:
            score = int(100 * (0.75 * (same / rets) + 0.25 * (rets / had)) + 0.5)
        hid = hirer_id_for(p["name"])

        # ---- the story, start to finish (times and all) ---------------
        bits = []
        bits.append("<b>{n}</b>, this is your complete K2 shutdown story "
                    "with the Coates tool store - every issue, every "
                    "return, every scan in your name, start to "
                    "finish.".format(n=first))
        if h.get("first"):
            bits.append("Your record opens on <b>{f}</b>{l}, across "
                        "<b>{d} active day{ds}</b> through the store."
                        .format(f=_fmt_stamp(h["first"]),
                                l=(" and your latest movement is {}".format(
                                    _fmt_stamp(h["last"]))
                                   if h.get("last") and h["last"] != h["first"]
                                   else ""),
                                d=h.get("days_active", 0),
                                ds="" if h.get("days_active") == 1 else "s"))
        bits.append("Right now <b>{n} item{ns}</b> {vb} out in your "
                    "name{old}.".format(
                        n=out_now, ns="" if out_now == 1 else "s",
                        vb="is" if out_now == 1 else "are",
                        old=(", the oldest out {d} day{ds}".format(
                            d=oldest, ds="" if oldest == 1 else "s")
                            if oldest is not None else "")))
        if rets:
            bits.append("You have returned <b>{r} item{rs}</b>, {sd} of "
                        "them the same day you took them - that is the "
                        "standard we promote, and it is how your score "
                        "below is earned.".format(
                            r=rets, rs="" if rets == 1 else "s", sd=same))
        story = "<div class='intro story'>" + " ".join(bits) + "</div>"

        body = story
        tiles = [("{}".format(score) if score is not None else "&mdash;",
                  "Returns Score / 100",
                  "" if score is None else
                  "g" if score >= 75 else "a" if score >= 55 else "r"),
                 (str(out_now), "In Your Name Now"),
                 ("{} &middot; {} same-day".format(rets, same), "Returned"),
                 (str(h.get("days_active", 0)), "Active Days"),
                 (esc(hid) if hid else "&mdash;", "Your Hirer ID")]
        body += tiles_html(tiles)

        # ---- everything through the store -----------------------------
        thru = []
        if h.get("hire_n"):
            thru.append("<b>{}</b> hire items through your hands".format(h["hire_n"]))
        if h.get("radio", [0, 0])[0]:
            thru.append("<b>{}</b> radio{} ({} returned)".format(
                h["radio"][0], "" if h["radio"][0] == 1 else "s", h["radio"][1]))
        if h.get("gas", [0, 0])[0]:
            thru.append("<b>{}</b> gas monitor{} ({} returned)".format(
                h["gas"][0], "" if h["gas"][0] == 1 else "s", h["gas"][1]))
        cons = h.get("cons") or []
        if cons:
            thru.append("<b>{}</b> consumable line{}".format(
                len(cons), "" if len(cons) == 1 else "s"))
        if thru or cons:
            body += ("<h2>Everything through the store in your name</h2>"
                     "<div class='intro'>" + (", ".join(thru) or "") +
                     ".</div>")
            if cons:
                body += ("<table class='data'><thead><tr>"
                         "<th>Consumable</th><th class='num'>Qty</th></tr>"
                         "</thead><tbody>" + "".join(
                             "<tr><td>{}</td><td class='num'>{}</td></tr>".format(
                                 esc(nm), q) for nm, q in cons[:12])
                         + "</tbody></table>")
        # ---- current gear ---------------------------------------------
        if items:
            body += ("<h2>Your gear right now - oldest first</h2>"
                     "<table class='data'><thead><tr><th>Description</th>"
                     "<th>Asset No</th><th>On Hire Since</th>"
                     "<th class='num'>Days</th></tr></thead><tbody>" + "".join(
                         "<tr><td>{d}</td><td>{a}</td><td>{s}</td>"
                         "<td class='num'>{dy}</td></tr>".format(
                             d=desc_disp(r["description"],
                                         r.get("asset")),
                             a=asset_disp(r.get("asset")),
                             s=(r["on_hire_date"].strftime("%d %b %Y")
                                if r.get("on_hire_date") else "-"),
                             dy=r["days"] if r["days"] is not None else "-")
                         for r in items) + "</tbody></table>"
                     "<div class='note'>Finished with something? Bring it "
                     "back to the store - it comes off this list on the "
                     "spot with the double scan, and it is ready for the "
                     "next crew. Anything here you believe has already "
                     "come back, tell us at the counter and we will run "
                     "the double-check return process the same day.</div>")
        # ---- the score, worked openly ---------------------------------
        body += ("<h2>Your returns score - worked out in the open</h2>"
                 "<div class='intro'>Returns Score = 100 &times; (0.75 "
                 "&times; same-day rate + 0.25 &times; returned rate). "
                 + ("Yours: {sd}/{r} same-day and {r}/{had} returned "
                    "= <b>{sc}/100</b>. ".format(sd=same, r=rets, had=had,
                                                 sc=score)
                    if score is not None else
                    "Yours starts building with your first return. ")
                 + "Same-day returns lift it fastest - and every return "
                   "is double-scanned, checked and maintained, so the "
                   "next crew gets gear that is ready to work. That is "
                   "Best Service &amp; Value working for all of us.</div>")
        # ---- safety, the Coates Way -----------------------------------
        dmg = [d for d in dmg_people
               if _person_key(re.sub(r"\s*\(.*?\)\s*", " ", d["person"])) == k]
        body += ("<h2>Safety - the Coates Way</h2>"
                 "<div class='intro'>Everything you draw from this store "
                 "is issued ready: gas monitors bump-tested and fully "
                 "charged, battery gear on a full charge, nothing damaged "
                 "or out of date ever issued. Life Saving Rule 5 - Tools "
                 "and Equipment - works both ways: if any tool feels off, "
                 "bring it straight in. We tag it out, it is never "
                 "reissued, and we put a fresh one in your hand. No "
                 "paperwork on you, no fuss - that is what we are here "
                 "for.</div>")
        if dmg:
            body += ("<div class='note'>For the record: " + "; ".join(
                "{} ({})".format(esc(d["item"]),
                                 d["when"].strftime("%d %b %Y")
                                 if d["when"] else "-") for d in dmg)
                + " is recorded in your name. If that does not look "
                  "right, see us at the counter - easy fixed, no "
                  "drama.</div>")
        # ---- here to support ------------------------------------------
        body += ("<h2>Here to support you</h2>"
                 "<div class='intro'>Care Deeply &middot; Customer Focused "
                 "&middot; Be Our Best &middot; One Team &middot; "
                 "Competitive Spirit - the Coates Way is not a poster, it "
                 "is how this store runs. If you need something we do not "
                 "stock, ask - we log every request and chase it. Scan "
                 "the QR poster at the store window any time to see this "
                 "list live; it updates with every issue and return. "
                 "Thanks for working with us, {n} - see you at the "
                 "counter.</div>".format(n=first))

        doc = paged_report_doc(
            "My Gear - Your Shutdown Story", display, asof, generated,
            source_line, body,
            doc_title="Coates K2 - My Gear - " + p["name"])
        stem = "Coates_K2_MyGear_{}_{}".format(safe_filename(p["name"]), date_tag)
        html_path = os.path.join(PAGES_DIR, stem + ".html")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(doc)
        imgs = email_images.capture_report_images(html_path, EMAILS_DIR, stem)
        #  THIS ONE TRIMS, IT DOES NOT DROP. Everywhere else too-heavy
        #  means "PDF on the paperclip instead" - but a person's My Gear
        #  mail has no attachment path behind it, so emptying the list
        #  here would silently send them nothing at all. A bloke with a
        #  lot of gear out is exactly the one who needs the mail. So the
        #  pages that fit go in the body, and the message says plainly
        #  how many were left off and where to read the rest.
        if imgs and email_images.email_weight_mb(imgs) \
                > email_images.MAX_EMAIL_MB:
            _keep = list(imgs)
            while _keep and email_images.email_weight_mb(_keep) \
                    > email_images.MAX_EMAIL_MB:
                _keep.pop()
            print("  ! {}: {} pages is over the {:.0f} MB safe-send limit "
                  "- first {} page(s) in the body, full list on the Gear "
                  "Lookup page.".format(
                      p["name"], len(imgs), email_images.MAX_EMAIL_MB,
                      len(_keep)))
            imgs = _keep
        subject = "Coates K2 - your shutdown story - {} item{} in your name".format(
            out_now, "" if out_now == 1 else "s")
        if imgs:
            msg = email_images.build_image_eml(subject, imgs, to=email_addr)
            eml_path = os.path.join(EMAILS_DIR, stem + "_OUTLOOK.eml")
            email_images.save_eml(msg, eml_path)
            cids = ["pg{}".format(i) for i in range(1, len(imgs) + 1)]
            write_draft_manifest(eml_path, subject,
                                 email_images.images_body(cids), [],
                                 report="MYGEAR", to=email_addr,
                                 images=email_images.manifest_images(imgs))
            drafted += 1
        else:
            print("  ! {}: page images unavailable - HTML written, no email.".format(
                p["name"]))
    print("  My Gear mail: {} of {} people drafted ({} with gear, no email"
          " in People (Email): {}{})".format(
              drafted, len(people), len(skipped),
              ", ".join(skipped[:5]) or "none",
              "..." if len(skipped) > 5 else ""))


# ---- 1. DAILY ONE-PAGER ---------------------------------------------------

def generate_daily_brief(models, pm, tracking_rows, stocktake, asof,
                         generated, source_line, date_tag):
    today = dt.date.today()
    shift = today - dt.timedelta(days=1) if dt.datetime.now().hour < 5 \
        else today
    yday = next((r for r in tracking_rows if r["date"] == shift), None)
    day_no = (shift - SHUTDOWN_START).days + 1
    days_left = max(0, (SHUTDOWN_END - shift).days)
    # The same conversation the Safety Assurance report carries today, in
    # one line - the brief is what the crews actually read at handover.
    # Report date, not the shift being reported on, so both documents
    # always show the same topic.
    try:
        _conv_day = dt.datetime.strptime(date_tag, "%Y-%m-%d").date()
    except Exception:
        _conv_day = dt.date.today()
    chase = []
    for m in models.values():
        for r in m["rows"][:3]:
            if r["days"] is not None and r["days"] >= 3:
                chase.append((r["days"], m["display"],
                              desc_disp(r["description"], r.get("asset")),
                              asset_disp(r.get("asset"))))
    chase.sort(reverse=True)
    chase = chase[:3]
    story = (
        "Day {dn} of the K2 shutdown, {dl} day{dls} to go. "
        "{oh} items are on hire across {co} companies. ".format(
            dn=day_no, dl=days_left, dls="" if days_left == 1 else "s",
            oh=sum(m["n_items"] for m in models.values()),
            co=len(models))
        + ("Yesterday's client day (6:00am to 6:00am) closed at {ac} actual "
           "against {fc} forecast. ".format(
               ac=fmt_money(yday["ac"]), fc=fmt_money(yday["fc"]))
           if yday and yday["ac"] is not None and yday["fc"] is not None
           else "Yesterday's actuals land as they are entered - forecast "
                "v actual follows in the executive pack. ")
        + ("Stocktake coverage: {sv}/{st} counted items sighted inside the "
           "window by {cnt} counters. ".format(
               sv=stocktake["sighted3"], st=stocktake["total"],
               cnt=stocktake["counters"]) if stocktake else "")
        + "The asks below are today's three conversations - everything else "
          "is tracking.")
    body = "<div class='intro story'>{}</div>".format(story)
    body += safety_conversation.panel_html(_conv_day, compact=True)
    # v3: yesterday's counter movement, straight off the transaction stamps
    _tx_all = load_transactions()[0] or []
    _seen_db = set()
    _txc = []
    for t in _tx_all:
        if is_site_plant_equipment(t.get("hirer", "")):
            continue
        k = (t.get("tx_id") or (t.get("barcode"), t.get("start"),
                                t.get("end"), t.get("hirer")))
        if k in _seen_db:
            continue
        _seen_db.add(k)
        _txc.append(t)
    _txa = tx_analytics(_txc, today)
    if _txa:
        y, db = _txa["yesterday"], _txa["day_before"]
        body += v3_kpis([
            ("swap", "{:,}".format(y["iss"]), "Issued yesterday",
             v3_delta(y["iss"], db["iss"], neutral=True,
                      label="vs day before")),
            ("tick", "{:,}".format(y["ret"]), "Returned yesterday",
             v3_delta(y["ret"], db["ret"], label="vs day before")),
            ("people", "{:,}".format(y["people"]), "People served",
             v3_delta(y["people"], db["people"], neutral=True,
                      label="vs day before")),
            ("box", "{:,}".format(sum(m["n_items"]
                                      for m in models.values())),
             "On hire right now", ""),
        ])
    if chase:
        body += "<h2>Today's three asks (oldest gear first)</h2>"
        # `it` is already desc_disp() output - escaped name plus badge
        # markup. Escaping it again prints the raw HTML on the page and
        # blows the table off the sheet. Every other _tbl caller in this
        # file passes pre-escaped cells for the same reason.
        body += _tbl(["Days Out", "Company", "Item", "Asset No"],
                     [[d, esc(co), it, an]
                      for d, co, it, an in chase])
    else:
        body += ("<h2>Today's three asks</h2><div class='note'>Nothing over "
                 "3 days old - the gear is moving the way it should. Keep "
                 "the double scan humming.</div>")
    if pm["idle"]:
        body += ("<div class='note'><b>Plant:</b> {} idle-but-charged "
                 "asset(s) on site - off-hiring saves {} per day.</div>"
                 ).format(len(pm["idle"]), fmt_money(pm["saving_per_day"]))
    ch_rows = load_charge_feed()
    if ch_rows:
        latest = ch_rows[-1]
        body += ("<div class='note'><b>Charges captured to date:</b> {n} "
                 "charge(s), {t} firm - latest {cat} {d} ({co}). Full detail "
                 "in the executive pack and each company's report.</div>"
                 ).format(n=len(ch_rows),
                          t=fmt_money(sum(c["amount"] for c in ch_rows if c["firm"])),
                          cat=esc(latest["category"].lower()),
                          d=latest["when"].strftime("%d %b") if latest["when"] else "-",
                          co=esc(latest["company"]))
    inner = [_e_note(esc(story), "Morning brief."),
             _e_h2("Today's three asks")]
    if chase:
        inner.append(_e_tbl(["Days", "Company", "Item", "Asset No"],
                            [[str(d), co, it, an]
                             for d, co, it, an in chase]))
    else:
        inner.append(_e_note("Nothing over 3 days old - gear is moving "
                             "the way it should."))
    limits = ("One page by design. Figures from SiteIQ exports and the K2 "
              "workbook as at the data-as-at stamp; a client day runs "
              "6:00am to 6:00am. Detail lives in the companion reports.")
    emit_report("Coates_K2_Daily_Brief_{}".format(date_tag),
                "Daily One-Pager", "K2 Morning Brief - Day {}".format(day_no),
                body, "Day {dn} &bull; {dl} days to go".format(
                    dn=day_no, dl=days_left),
                inner, limits,
                "Coates K2 - Daily Brief - {} - day {} of the shut".format(
                    shift.strftime("%d %b %Y"), day_no),
                asof, generated, source_line, report_tag="DAILY")


# ---- 3. RETURNS PERFORMANCE ----------------------------------------------

def generate_returns_report(models, tx, tx_loaded, asof, generated,
                            source_line, date_tag):
    per_co = {}
    hero_people = {}
    for t in tx:
        co = t.get("company")
        if not co:
            continue
        # Internal Coates idle-pool movements (plant parked in the Site
        # Plant Equipment pool) are not a real client's return behaviour -
        # they'd otherwise show up as a "Coates" row with hundreds of
        # "still open" transactions, and confuse a client-facing leaderboard
        # with our own internal housekeeping. Same exclusion as everywhere
        # else in the kit (Andrew, 21 Jul direction on the idle pool).
        if co == "COATES" or is_site_plant_equipment(t.get("hirer") or ""):
            continue
        e = per_co.setdefault(co, {"done": 0, "same": 0, "open": 0})
        if t["start"] and t["end"]:
            e["done"] += 1
            if t["end"] == t["start"]:
                e["same"] += 1
                if t["hirer"]:
                    hero_people[t["hirer"]] = hero_people.get(t["hirer"], 0) + 1
        elif t["start"]:
            e["open"] += 1
    # Companies with gear on hire but no tool-store transaction yet still
    # belong on the board as 0-transaction rows - never silently missing
    # (a reader comparing this with the demob pack must find every company).
    for co, m in models.items():
        if co not in per_co and m.get("n_items"):
            per_co[co] = {"done": 0, "same": 0, "open": 0}
    total_done = sum(e["done"] for e in per_co.values())
    total_same = sum(e["same"] for e in per_co.values())
    if total_done:
        story = (
            "Gear that comes back the day it went out keeps the whole site "
            "moving - and this crew is doing it. {sd} of {td} completed "
            "tool store transactions came back the same day ({pct:.0%}). "
            "Every return is inspected, double-scanned and stocktaken back "
            "to its shelf - the return isn't complete until we've looked. "
            "This is a leaderboard, not a blame list: the names below are "
            "the standard-setters.").format(
            sd=total_same, td=total_done,
            pct=total_same / total_done)
    else:
        story = (
            "The returns story builds itself as gear starts coming home - "
            "every return inspected, double-scanned and stocktaken back to "
            "its shelf. No completed returns are recorded in SiteIQ yet; "
            "as they land, this report shows same-day performance by "
            "company and names the standard-setters.")
    body = "<div class='intro story'>{}</div>".format(story)
    # v3: the movement trend + when the counter runs hottest
    _seen_r = set()
    _txc = []
    for t in tx:
        if is_site_plant_equipment(t.get("hirer", "")):
            continue
        k = (t.get("tx_id") or (t.get("barcode"), t.get("start"),
                                t.get("end"), t.get("hirer")))
        if k in _seen_r:
            continue
        _seen_r.add(k)
        _txc.append(t)
    _txa = tx_analytics(_txc, dt.date.today())
    if _txa:
        body += "<h2>The movement, day by day</h2>"
        body += v3_bars_pair(_txa["iss_series"][-10:],
                             _txa["ret_series"][-10:], "Issued", "Returned")
        body += v3_heat(_txa["hours"])
        body += ("<div class='note'>Issue and return events from SiteIQ's "
                 "own stamps, site plant pool excluded - the counter's real "
                 "workload.</div>")
    if total_done:
        body += ("<div style='text-align:center'>"
                 + ring(total_same / total_done, "Same-day returns across the "
                        "shut", "the standard", colour="#3FB950",
                        size=132) + "</div>")
    rows = []
    for co in sorted(per_co, key=lambda c: display_company(c).upper()):
        e = per_co[co]
        pct = "{:.0%}".format(e["same"] / e["done"]) if e["done"] else "-"
        rows.append([esc(display_company(co)), e["done"], e["same"], pct,
                     e["open"]])
    if rows:
        body += "<h2>Return behaviour by company (alphabetical)</h2>"
        body += _tbl(["Company", "Completed", "Same-Day", "Same-Day %",
                      "Still Open"], rows)
        ranked = sorted((c for c in per_co if per_co[c]["done"]),
                        key=lambda c: -(per_co[c]["same"] / per_co[c]["done"]))
        if ranked:
            body += ("<div class='note' style='margin-top:10px'><b>Same-day "
                     "rate, drawn</b> - best first; green is the standard "
                     "we promote.</div>"
                     + "".join(hbar(display_company(c),
                                    "{:.0%}".format(per_co[c]["same"] / per_co[c]["done"]),
                                    per_co[c]["same"] / per_co[c]["done"],
                                    rag_colour(per_co[c]["same"] / per_co[c]["done"],
                                               good=0.75, warn=0.5))
                               for c in ranked[:10]))
    heroes = sorted(hero_people.items(), key=lambda kv: -kv[1])[:5]
    if heroes:
        body += "<h2>Same-day standard-setters</h2>"
        body += _tbl(["Person", "Hirer ID", "Same-Day Returns"],
                     [[esc(h), esc(hirer_id_for(h)) or "-", n]
                      for h, n in heroes])
    inner = [_e_note(esc(story), "The story.")]
    if rows:
        inner.append(_e_h2("Return behaviour by company"))
        inner.append(_e_tbl(["Company", "Completed", "Same-Day", "%", "Open"],
                            [[r[0], str(r[1]), str(r[2]), r[3], str(r[4])]
                             for r in rows]))
    if heroes:
        inner.append(_e_h2("Same-day standard-setters"))
        inner.append(_e_tbl(["Person", "Hirer ID", "Same-Day Returns"],
                            [[h, hirer_id_for(h) or "-", str(n)]
                             for h, n in heroes]))
    limits = ("Transactions attributed by the company that ran them "
              "(EMPLOYER_NAME). Same-day = returned on the issue date. "
              "Gear that went on hire without a tool store transaction "
              "(e.g. plant delivered direct) counts on the on-hire and "
              "demob reports, not in these transaction figures - such "
              "companies appear here as 0-transaction rows. "
              + ("" if tx_loaded else "TRANSACTIONS export unavailable this "
                 "run - figures unavailable, not zero. ")
              + "A positive report by design: it names good behaviour, "
                "never bad.")
    emit_report("Coates_K2_Returns_Performance_{}".format(date_tag),
                "Returns Performance", "Same-Day Heroes - All Companies",
                body,
                "{sd} same-day of {td} completed".format(
                    sd=total_same, td=total_done) if total_done
                else "the returns story starts with the first return",
                inner, limits,
                "Coates K2 - Returns Performance - {}".format(
                    asof.strftime("%d %b %Y") if asof else ""),
                asof, generated, source_line, report_tag="RETURNS")


# ---- 5. PLANT UTILISATION & RIGHT-SIZE -----------------------------------

def generate_rightsize_report(pm, asof, generated, source_line, date_tag):
    cat = {}
    for p in pm["in_use"]:
        c = cat.setdefault(p["category"], {"use": 0, "idle": 0, "unch": 0})
        c["use"] += 1
    for p in pm["idle"]:
        c = cat.setdefault(p["category"], {"use": 0, "idle": 0, "unch": 0})
        c["idle"] += 1
    for g in pm["available_groups"]:
        c = cat.setdefault(g["category"], {"use": 0, "idle": 0, "unch": 0})
        c["unch"] += g["qty"]
    saving_bit = ("Off-hiring today's idle list saves <b>{} per "
                  "day</b>.".format(fmt_money(pm["saving_per_day"]))
                  if pm["idle"] else
                  "Nothing is sitting idle on charge today - the plant "
                  "fleet is earning its keep.")
    story = (
        "True utilisation, told straight: {fleet} site plant assets - "
        "{use} in a contractor's hands, {idle} back on site but still "
        "charged (the saving list), {unch} not yet charged and available. "
        "Idle gear back in the store is also the Coates mechanic's "
        "preventative maintenance window - service it while it's idle, "
        "not in a crew's hands. {saving}").format(
        fleet=pm["fleet"], use=len(pm["in_use"]), idle=len(pm["idle"]),
        unch=pm["uncharged"], saving=saving_bit)
    body = "<div class='intro story'>{}</div>".format(story)
    rows = []
    for cname in sorted(cat):
        c = cat[cname]
        tot = c["use"] + c["idle"] + c["unch"]
        util = "{:.0%}".format(c["use"] / tot) if tot else "-"
        rows.append([esc(cname), tot, c["use"], c["idle"], c["unch"], util])
    body += "<h2>Utilisation by category</h2>"
    body += _tbl(["Category", "Fleet", "In Use", "Idle (charged)",
                  "Not Charged", "In-Use %"], rows)
    if pm["idle"]:
        body += "<h2>Right-size candidates - longest sitting first</h2>"
        body += _tbl(["Category", "Description", "Asset No", "Days Sitting",
                      "Saving / Day"],
                     [[esc(p["category"]),
                       desc_disp(p["description"], p.get("asset")),
                       asset_disp(p.get("asset")),
                       "-" if p["days_sitting"] is None else p["days_sitting"],
                       fmt_money(p["rate"]) if p["rate"] is not None else "-"]
                      for p in pm["idle"]])
    else:
        body += ("<h2>Right-size candidates</h2><div class='note'>Nothing "
                 "is sitting idle on charge right now - the plant fleet is "
                 "earning its keep.</div>")
    inner = [_e_note(esc(story), "True utilisation."),
             _e_h2("Utilisation by category"),
             _e_tbl(["Category", "Fleet", "In Use", "Idle", "Not Charged",
                     "In-Use %"],
                    [[r[0], str(r[1]), str(r[2]), str(r[3]), str(r[4]), r[5]]
                     for r in rows])]
    limits = ("Categories from the Plant Equipment register; rates are "
              "contracted or register rates; unpriced items are excluded "
              "from savings, never estimated. Barriers, chutes and hoppers "
              "are site infrastructure - never counted as idle.")
    emit_report("Coates_K2_Plant_RightSize_{}".format(date_tag),
                "Plant Utilisation & Right-Size", "Site Plant - K2",
                body,
                ("{use} in use &bull; {idle} idle &bull; {save}/day on "
                 "the table".format(
                     use=len(pm["in_use"]), idle=len(pm["idle"]),
                     save=fmt_money(pm["saving_per_day"])) if pm["idle"]
                 else "{use} in use &bull; fleet earning its keep".format(
                     use=len(pm["in_use"]))),
                inner, limits,
                "Coates K2 - Plant Right-Size - {} - {} idle".format(
                    asof.strftime("%d %b %Y") if asof else "",
                    len(pm["idle"])),
                asof, generated, source_line, report_tag="RIGHTSIZE")



# ---- 2. SAFETY & HIGH-CARE ASSURANCE (client-facing) ---------------------

def load_sds_register():
    """The Coates SDS register for the shutdown store - every chemical on
    the shelf with a current Safety Data Sheet. Read as-is from Andrew's
    register workbook; absent file = section simply doesn't appear."""
    path = find_newest("Coates_SDS_Register*.xlsx", "the SDS register",
                       required=False)
    if not path:
        return None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        print("NOTE: SDS register present but unreadable - section skipped.")
        return None
    hdr = None
    items = []
    asat = ""
    for r in rows:
        vals = [clean_text(v) for v in r]
        joined = " ".join(vals)
        if not asat and joined.strip().lower().startswith("as at:"):
            asat = joined.split("|")[0].replace("As at:", "").strip()
        if hdr is None:
            if vals and vals[0] == "No." and "Product / Substance" in vals:
                hdr = {v: i for i, v in enumerate(vals) if v}
            continue
        def cell(name):
            i = hdr.get(name)
            return vals[i] if i is not None and i < len(vals) else ""
        if not cell("No.") or not cell("No.").rstrip(".").isdigit():
            continue
        items.append({"product": cell("Product / Substance"),
                      "maker": cell("Manufacturer / Supplier"),
                      "use": cell("Type & Use"),
                      "sds_date": cell("SDS Date"),
                      "review": cell("Next Review Due")})
    if not items:
        return None
    return {"path": path, "items": items, "asat": asat}


def diphoterine_safety_html(models):
    """DIPHOTERINE EMERGENCY RINSE - the store package section of the
    daily safety pack (Andrew's initiative, 27 Jul 2026). Everything
    here reads off the Diphoterine_Store_Package folder (SDS data, QR
    codes, the unit register) plus the live hire rows - so the day these
    units start going out, the holders appear here by name with expiry
    dates, automatically. No package folder, no section - never a
    guess."""
    import base64
    import json as _json
    pkg = os.path.join(KIT_DIR, "Diphoterine_Store_Package")
    jpath = os.path.join(pkg, "Diphoterine SDS Source Data.json")
    if not os.path.isfile(jpath):
        return ""
    try:
        with io.open(jpath, encoding="utf-8") as f:
            src_data = _json.load(f)
    except Exception:
        return ""
    products = src_data.get("products") or []

    # ---- the unit register: barcode -> batch / expiry -------------------
    units = []
    try:
        wb = openpyxl.load_workbook(
            os.path.join(pkg, "Diphoterine Issue & Return Register - K2 - "
                              "27 Jul 2026.xlsx"),
            read_only=True, data_only=True)
        rows = list(wb["Item Master"].iter_rows(values_only=True))
        hi = next(i for i, r in enumerate(rows)
                  if r and str(r[0]).strip() == "SKU_NUMBER")
        hdr = [str(v).strip() if v else "" for v in rows[hi]]
        for r in rows[hi + 1:]:
            d = dict(zip(hdr, r))
            if d.get("ITEM_BARCODE"):
                units.append(d)
        wb.close()
    except Exception:
        units = []
    by_bc = {str(u.get("ITEM_BARCODE") or "").strip().upper(): u
             for u in units}

    # ---- live holders off today's hire rows ------------------------------
    holders = []
    for m in models.values():
        for r in m["rows"]:
            du = (r["description"] + " " + r.get("desc0", "")).upper()
            #  Match the loaded barcodes (DIPH50K2..., DIPHDAP...) as well
            #  as the word, so a SiteIQ rename can never hide a unit from
            #  the who-has-it table. (28 Jul 2026)
            if ("DIPHOTERINE" in du or
                    str(r.get("barcode") or "").strip().upper()
                    .startswith("DIPH")):
                holders.append((m, r))

    expiries = sorted(str(u.get("EXPIRY") or "")[:10] for u in units
                      if u.get("EXPIRY"))
    soonest = expiries[0] if expiries else "-"

    body = "<h2>Diphoterine emergency rinse - chemical splash response</h2>"
    body += ("<div class='intro'><b>New capability at the store.</b> "
             "<b>{n}</b> single-use emergency chemical-splash rinse units "
             "are held for issue to Cement Australia crews - sterile, "
             "sealed, date-controlled first-aid gear for corrosive splash, "
             "checked against the physical stock. These are not tools: the "
             "return check is <b>unused, sealed and in date</b> - anything "
             "else is an incident, not a return.</div>").format(
        n=len(units) or 32)
    body += _tiles_diph([
        (str(len(units) or 32), "Units held at the store"),
        (str(len(holders)), "Out with crews now"),
        (esc(soonest), "Soonest expiry"),
        ("{} current".format(len(products)), "Australian SDS verified"),
    ])

    if holders:
        body += ("<div class='note'><b>Who has them right now</b> - every "
                 "unit out is in someone's name, with its expiry.</div>")
        h = ("<table class='data'><thead><tr><th>Who</th><th>Company</th>"
             "<th>Unit</th><th>Batch / expiry</th><th class='num'>Out</th>"
             "</tr></thead><tbody>")
        for m, r in sorted(holders, key=lambda p: -(p[1]["days"] or 0)):
            u = by_bc.get(str(r.get("barcode") or "").strip().upper(), {})
            be = ("{} / {}".format(u.get("BATCH"),
                                   str(u.get("EXPIRY"))[:10])
                  if u else "-")
            days = r["days"]
            h += ("<tr><td><b>{p}</b></td><td>{c}</td><td>{d}</td>"
                  "<td style='white-space:nowrap'>{be}</td>"
                  "<td class='num'>{o}</td></tr>").format(
                p=esc(activity_model._clean_name(r["hirer"]) or "-"),
                c=esc(m["display"]), d=esc(r["description"]),
                be=esc(be),
                o=("-" if days is None else "today" if days == 0
                   else "{}d".format(days)))
        body += h + "</tbody></table>"
    else:
        body += ("<div class='note'><b>Nothing out yet - all units on the "
                 "shelf, sealed and ready.</b> The moment one is issued "
                 "through SiteIQ it appears here by name, with its batch "
                 "and expiry.</div>")

    # ---- SDS + QR cards ---------------------------------------------------
    qr_files = {
        "SIEW": "SDS_QR_Diphoterine_Eye_Wash_50mL.png",
        "MICRO DAP": "SDS_QR_Diphoterine_Micro_DAP_100mL.png",
    }
    cards = ""
    for p in products:
        qf = next((f for k, f in qr_files.items()
                   if k in p.get("product", "").upper()), None)
        qr = ""
        if qf and os.path.isfile(os.path.join(pkg, qf)):
            with open(os.path.join(pkg, qf), "rb") as f:
                qr = ("<img src='data:image/png;base64,{}' style='width:"
                      "26mm;height:26mm'>").format(
                    base64.b64encode(f.read()).decode())
        cards += (
            "<td style='width:50%;vertical-align:top;background:#F6F7F9;"
            "border:1px solid #E3E6EB;border-left:4px solid " + EC.NEON +
            ";border-radius:8px;padding:10px 12px'>"
            "<table style='width:100%'><tr>"
            "<td style='vertical-align:top'>"
            "<div style='font-size:9pt;font-weight:800;color:#14181F'>{n}"
            "</div>"
            "<div style='font-size:8pt;color:#33393F;margin-top:3px;"
            "line-height:1.5'>{u}</div>"
            "<div style='font-size:7.5pt;color:#5B6472;margin-top:4px'>"
            "SDS rev {v} &middot; {d} &middot; review due {rv}</div>"
            "<div style='font-size:7.5pt;font-weight:800;color:#14181F;"
            "margin-top:3px'>Scan the QR - the SDS downloads straight "
            "to your phone.</div></td>"
            "<td style='width:28mm;text-align:right;vertical-align:top'>"
            "{q}</td></tr></table></td>").format(
            n=esc(p.get("product", "")), u=esc(p.get("use", "")),
            v=esc(str(p.get("ver", "-"))), d=esc(p.get("sds_date", "-")),
            rv=esc(p.get("review", "-")), q=qr)
    if cards:
        body += ("<table style='width:100%;border-collapse:separate;"
                 "border-spacing:6px'><tr>" + cards + "</tr></table>")

    # ---- the process, in the package's own words -------------------------
    body += (
        "<div class='note'><b>The process at the counter:</b> sign one out "
        "at the store (sign-out sheet at the counter), carry it sealed, and "
        "hand it back the same way - <b>unused, sealed and in date</b>. "
        "<span class='repl'>ANY USE = INCIDENT</span> - a used unit means a "
        "chemical splash has happened: first aid comes first, then notify "
        "the Cement Australia site supervisor and the Coates tool store "
        "immediately so the unit is replaced and the incident is recorded. "
        "<b>Not for hydrofluoric acid</b> or fluorides in acidic medium - "
        "Diphoterine has limited action there; confirm the chemical with "
        "the Cement Australia supervisor if in doubt. Full SOP, issue "
        "register and quick-scan board live at the store.</div>")
    return body


def _tiles_diph(cells):
    return tiles_html(cells)


def generate_safety_report(models, radio_fleet, gas_fleet, stocktake, asof,
                           generated, source_line, date_tag):
    milw = []
    gas_out = []
    radio_out = []
    for m in models.values():
        for r in m["rows"]:
            du = (r["description"] + " " + r.get("desc0", "")).upper()
            if "MILWAUKEE" in du:
                milw.append((m["display"], r))
            if r["storage_unit"].upper() == "GAS MONITORS":
                gas_out.append((m["display"], r))
            if is_radio_handset(r):
                radio_out.append((m["display"], r))
    rb = min(radio_fleet, RADIO_BILLABLE_CAP)
    story = (
        "Safety-critical gear gets a different level of care, and this "
        "report is the evidence. The {gf}-unit gas monitor fleet is pooled "
        "through the store - every monitor bump tested and fully charged "
        "before it is issued, never a partial charge, never an out-of-date "
        "unit. {rf} radio handsets are managed the same way ({rb} in the "
        "working pool, {rs} held as spares), with batteries charge-checked "
        "on every return - flat means on charge before ready. Milwaukee "
        "tooling is tracked piece by piece for its high replacement value. "
        "Life Saving Rules (SEQ-GL-009) sit under all of it - Rule 5, Tools "
        "and Equipment: anything damaged or defective is tagged out, "
        "photographed and removed from service on the spot. Never left on "
        "the shelf. Never reissued.").format(
        gf=gas_fleet or 30, rf=radio_fleet, rb=rb,
        rs=max(0, radio_fleet - rb))
    body = "<div class='intro story'>{}</div>".format(story)

    # Today's safety conversation - one topic a day off a 20-day cycle.
    # Keyed to the REPORT date, not the data-as-at stamp: the conversation
    # is for the crew standing in front of you this morning, and a stale
    # export must never shift which topic everyone is talking about.
    # (A. Fisher, 25 Jul 2026)
    try:
        _conv_day = dt.datetime.strptime(date_tag, "%Y-%m-%d").date()
    except Exception:
        _conv_day = dt.date.today()
    body += "<h2>Today's safety conversation</h2>"
    body += safety_conversation.panel_html(_conv_day)

    # What is out there right now that carries an obligation.
    _all_rows = [r for m in models.values() for r in m["rows"]]
    _cn = EC.summarise(_all_rows)
    if _cn["any"]:
        body += "<h2>Compliance on hire &mdash; what has to be checked</h2>"
        body += EC.summary_band_html(_cn)
        body += ("<div class='note'>Every hire line on every report in this "
                 "pack now carries its own requirement underneath the item "
                 "name. That removes the &lsquo;nobody told me&rsquo; "
                 "conversation at the counter, and it gives the Coates team "
                 "the detail they need to check gear over before it goes "
                 "back out. Blank means the master file has nothing to "
                 "declare for that item &mdash; never a guess.</div>")
        body += EC.legend_html(_conv_day)

    body += "<h2>High-care gear in the field right now</h2>"
    body += _tbl(["Type", "Items Out", "Companies Holding"],
                 [["Gas monitors", len(gas_out),
                   len({c for c, _ in gas_out})],
                  ["Radio handsets", len(radio_out),
                   len({c for c, _ in radio_out})],
                  ["Milwaukee tooling", len(milw),
                   len({c for c, _ in milw})]])
    if milw:
        body += "<h2>Milwaukee register - who holds what</h2>"
        body += _tbl(["Company", "Item", "Asset No", "Days Out"],
                     [[esc(c), desc_disp(r["description"],
                                          r.get("asset")),
                       asset_disp(r.get("asset")),
                       "-" if r["days"] is None else r["days"]]
                      for c, r in sorted(
                          milw, key=lambda cr: (cr[0], cr[1]["description"],
                                                str(cr[1].get("asset") or "")))])
    if stocktake:
        body += ("<div class='note'><b>Eyes on the fleet.</b> {sv} of {st} "
                 "counted items physically sighted inside the window by "
                 "{cnt} counters - daily stocktakes, no exceptions, and the "
                 "count doubles as on-hire verification.</div>").format(
            sv=stocktake["sighted3"], st=stocktake["total"],
            cnt=stocktake["counters"])
    body += diphoterine_safety_html(models)
    sds = load_sds_register()
    if sds:
        body += (
            "<h2>Hazardous substances &mdash; the SDS register</h2>"
            "<div class='intro'>Every chemical on the shutdown store shelf "
            "is registered with a current Safety Data Sheet &mdash; "
            "<b>{n} substances</b>, each verified to the manufacturer's "
            "official SDS with a QR link at the shelf (as at {asat}). "
            "Nothing goes on the shelf without one. This is doing our part "
            "before anyone has to ask.</div>".format(
                n=len(sds["items"]), asat=esc(sds["asat"] or "register date"))
            + _tbl(["Product", "Manufacturer", "SDS Date", "Review Due"],
                   [[esc(i["product"]), esc(i["maker"]),
                     esc(i["sds_date"]), esc(i["review"])]
                    for i in sds["items"]])
            + "<div class='note'>Register: {f} &mdash; sources and verified "
              "links on its Sources &amp; Notes sheet.</div>".format(
                  f=esc(os.path.basename(sds["path"]))))
    # The Log Books site notice, rebuilt as vector HTML - it lands as a
    # full page in the PDF and rides into the Outlook body as an image.
    body += logbook_poster.poster_html()

    inner = [_e_note(esc(story), "The assurance."),
             _RawHTML(safety_conversation.email_html(_conv_day)),
             _e_h2("High-care gear in the field"),
             _e_tbl(["Type", "Items Out", "Companies"],
                    [["Gas monitors", str(len(gas_out)),
                      str(len({c for c, _ in gas_out}))],
                     ["Radio handsets", str(len(radio_out)),
                      str(len({c for c, _ in radio_out}))],
                     ["Milwaukee tooling", str(len(milw)),
                      str(len({c for c, _ in milw}))]])]
    limits = ("Counts from the SiteIQ RENTAL_STOCK export as at the "
              "data-as-at stamp. This is an assurance report - no hire "
              "costs appear anywhere on it.")
    emit_report("Coates_K2_Safety_Assurance_{}".format(date_tag),
                "Safety & High-Care Assurance", "Client Assurance Pack - K2",
                body,
                "{g} gas &bull; {r} radios &bull; {m} Milwaukee in the "
                "field".format(g=len(gas_out), r=len(radio_out),
                               m=len(milw)),
                inner, limits,
                "Coates K2 - Safety & High-Care Assurance - {}".format(
                    asof.strftime("%d %b %Y") if asof else ""),
                asof, generated, source_line, report_tag="SAFETY")


# ---- 4. WEEKLY EXECUTIVE ROLL-UP -----------------------------------------

def generate_weekly_rollup(models, tracking_rows, st_daily, asof, generated,
                           source_line, date_tag):
    weeks = {}
    for r in tracking_rows:
        if r["date"] < SHUTDOWN_START:
            continue
        wk = (r["date"] - SHUTDOWN_START).days // 7 + 1
        w = weeks.setdefault(wk, {"fc": 0.0, "ac": 0.0, "days_in": 0,
                                  "sights": 0})
        if r["fc"] is not None:
            w["fc"] += r["fc"]
        # A day only counts as ENTERED when a category actual cell holds a
        # value - the Daily Actual Total formula reads 0 on untouched
        # future rows and must never masquerade as a real nil.
        entered = any(r[k] is not None for k in
                      ("plant_ac", "tool_ac", "radio_ac", "gas_ac",
                       "labour_ac", "accom_ac"))
        if entered and r["ac"] is not None:
            w["ac"] += r["ac"]
            w["days_in"] += 1
    for d, n in st_daily.items():
        if d >= SHUTDOWN_START:
            wk = (d - SHUTDOWN_START).days // 7 + 1
            if wk in weeks:
                weeks[wk]["sights"] += n
    wk_now = (dt.date.today() - SHUTDOWN_START).days // 7 + 1
    story = (
        "Week {wk} of the K2 shutdown. The roll-up below tracks every week "
        "against forecast - actuals build as each client day (6:00am to "
        "6:00am) is entered, and stocktake sightings show the checking "
        "cadence behind the numbers. {oh} items are on hire across {co} "
        "companies right now.").format(
        wk=wk_now, oh=sum(m["n_items"] for m in models.values()),
        co=len(models))
    body = "<div class='intro story'>{}</div>".format(story)
    rows = []
    for wk in sorted(weeks):
        w = weeks[wk]
        wstart = SHUTDOWN_START + dt.timedelta(days=(wk - 1) * 7)
        wend = min(wstart + dt.timedelta(days=6), SHUTDOWN_END)
        var = w["ac"] - w["fc"] if w["days_in"] else None
        rows.append([
            "Week {} ({} - {})".format(wk, wstart.strftime("%d %b"),
                                       wend.strftime("%d %b")),
            fmt_money(w["fc"]),
            fmt_money(w["ac"]) if w["days_in"] else "-",
            fmt_money(var) if var is not None else "-",
            w["days_in"], "{:,}".format(w["sights"])])
    body += "<h2>Week by week - forecast v actual</h2>"
    body += _tbl(["Week", "Forecast", "Actual", "Variance",
                  "Days Entered", "Stocktake Sightings"], rows)
    body += ("<div class='note'>Variance reads actual minus forecast for "
             "the days entered so far - a week with missing days shows what "
             "is known, never a guessed total.</div>")
    inner = [_e_note(esc(story), "Week {} roll-up.".format(wk_now)),
             _e_h2("Week by week"),
             _e_tbl(["Week", "Forecast", "Actual", "Variance", "Days",
                     "Sightings"],
                    [[r[0], r[1], r[2], r[3], str(r[4]), r[5]]
                     for r in rows])]
    limits = ("Weeks run from the shutdown start (12 Jul 2026). Figures "
              "from the K2 workbook Daily Tracking (read-only) and the "
              "STOCKTAKE export; missing days are shown as gaps, never "
              "zeros. A client day runs 6:00am to 6:00am.")
    emit_report("Coates_K2_Weekly_Rollup_{}".format(date_tag),
                "Weekly Executive Roll-Up", "Week {} - K2".format(wk_now),
                body,
                "week {wk} &bull; {oh} items on hire".format(
                    wk=wk_now,
                    oh=sum(m["n_items"] for m in models.values())),
                inner, limits,
                "Coates K2 - Weekly Roll-Up - Week {}".format(wk_now),
                asof, generated, source_line, report_tag="WEEKLY")


# ---- 7. END-OF-SHUT CLOSE-OUT PACK ---------------------------------------

def generate_closeout_report(models, pm, tx, sales, stocktake, tracking_rows,
                             asof, generated, source_line, date_tag):
    today = dt.date.today()
    days_left = max(0, (SHUTDOWN_END - today).days)
    still_out = sum(m["n_items"] for m in models.values())
    fc_td = sum(r["fc"] for r in tracking_rows
                if r["fc"] is not None and r["date"] <= today)
    ac_td = sum(r["ac"] for r in tracking_rows
                if r["ac"] is not None and r["date"] <= today)
    # Exclude internal Coates idle-pool movements (plant parked in Site
    # Plant Equipment) - they're not a genuine service transaction and
    # would otherwise pad this client-facing headline number.
    done_tx = sum(1 for t in tx
                  if t["start"] and t["end"] and t.get("company") != "COATES"
                  and not is_site_plant_equipment(t.get("hirer") or ""))
    draws = len(sales)
    if days_left > 0:
        headline = ("PROGRESS TOWARD CLOSE-OUT - {d} day{ds} to go"
                    .format(d=days_left, ds="" if days_left == 1 else "s"))
        story = (
            "The finish line is {d} day{ds} away, and the close-out "
            "position builds here daily. {so} items remain on hire across "
            "{co} companies - the zero-loss sheet is written in the last "
            "week, and it starts with this list going to zero. Spend to "
            "date reads {ac} against {fc} forecast for the days entered. "
            "Every issue and return this shutdown went through the double "
            "scan - the close-out numbers will be scanned facts, not "
            "estimates.").format(
            d=days_left, ds="" if days_left == 1 else "s", so=still_out,
            co=len(models), ac=fmt_money(ac_td), fc=fmt_money(fc_td))
    else:
        headline = "CLOSE-OUT - THE ZERO-LOSS SHEET"
        story = (
            "The K2 shutdown is complete. Final position: {so} item{sos} "
            "remaining on hire across {co} compan{cos}. Spend closed at "
            "{ac} against {fc} forecast. Every movement this shutdown ran "
            "through the double scan - these numbers are scanned facts."
        ).format(so=still_out, sos="" if still_out == 1 else "s",
                 co=len(models), cos="y" if len(models) == 1 else "ies",
                 ac=fmt_money(ac_td), fc=fmt_money(fc_td))
    body = "<div class='note'><b>{}</b><br>{}</div>".format(headline, story)
    rows = []
    for m in sorted(models.values(), key=lambda m: m["display"].upper()):
        if m["n_items"]:
            oldest = max((r["days"] or 0) for r in m["rows"]) \
                if m["rows"] else 0
            rows.append([esc(m["display"]), m["n_items"], oldest])
    if rows:
        body += "<h2>The list to zero - items still out</h2>"
        body += _tbl(["Company", "Items Still Out", "Oldest (days)"], rows)
    else:
        body += ("<h2>The list to zero</h2><div class='note'>Zero items "
                 "outstanding - the zero-loss sheet is real. Care Deeply, "
                 "One Team: thank you.</div>")
    body += "<h2>The shutdown in numbers (so far)</h2>"
    body += _tbl(["Measure", "Value"],
                 [["Tool store transactions completed", "{:,}".format(done_tx)],
                  ["Consumable draws", "{:,}".format(draws)],
                  ["Companies served", len(models)],
                  ["Site plant fleet managed", pm["fleet"]],
                  ["Stocktake items on watch",
                   stocktake["total"] if stocktake else "-"]])
    inner = [_e_note(esc(story), headline + "."),
             _e_h2("The list to zero")]
    if rows:
        inner.append(_e_tbl(["Company", "Still Out", "Oldest (days)"],
                            [[r[0], str(r[1]), str(r[2])] for r in rows]))
    else:
        inner.append(_e_note("Zero items outstanding - the zero-loss sheet "
                             "is real. Thank you."))
    limits = ("A living pack: run it any day for the progress view; the "
              "final run after the last return is the close-out of record. "
              "All figures from SiteIQ exports and the K2 workbook as at "
              "the data-as-at stamp; missing days shown as gaps, never "
              "zeros.")
    emit_report("Coates_K2_CloseOut_{}".format(date_tag),
                "End-of-Shut Close-Out", headline.title(),
                body,
                "{so} items to zero &bull; {d} day{ds} to go".format(
                    so=still_out, d=days_left,
                    ds="" if days_left == 1 else "s") if days_left
                else "the zero-loss close-out",
                inner, limits,
                "Coates K2 - Close-Out Pack - {}".format(
                    today.strftime("%d %b %Y")),
                asof, generated, source_line, report_tag="CLOSEOUT")



# ---- GEAR LOOKUP PAGE (QR at the window) ----------------------------------
# Self-contained page: person enters the ID number off their card - their
# EXTERNAL_ID, straight from the SiteIQ ON_HIRE export, the only person
# identifier we ever use - and ONLY their on-hire list appears. Every
# person's list is encrypted with a key derived from their own ID (XTEA
# cipher, identical JS/Python implementations, with a build-time self-test
# vector the page verifies on every load). Wrong or unknown ID decrypts
# nothing. Snapshot of the morning refresh, stamped.

LOOKUP_SALT_KEY = "CoatesK2GearLookup2026"
LOOKUP_SALT_TAG = "CoatesK2TagSalt2026"


def _xtea_encrypt_block(v0, v1, key):
    delta = 0x9E3779B9
    total = 0
    for _ in range(32):
        v0 = (v0 + ((((v1 << 4) ^ (v1 >> 5)) + v1) ^
                    (total + key[total & 3]))) & 0xFFFFFFFF
        total = (total + delta) & 0xFFFFFFFF
        v1 = (v1 + ((((v0 << 4) ^ (v0 >> 5)) + v0) ^
                    (total + key[(total >> 11) & 3]))) & 0xFFFFFFFF
    return v0, v1


def _lk_pad(data):
    b = data.encode("utf-8")
    b += b"\x80" + b"\x00" * ((15 - (len(b) % 16)) % 16)
    return b


def _lk_hash(text, iv0, iv1):
    """Davies-Meyer over XTEA: 64-bit digest as two u32s."""
    a, b = iv0, iv1
    data = _lk_pad(text)
    for i in range(0, len(data), 16):
        k = [int.from_bytes(data[i + j:i + j + 4], "big") for j in
             (0, 4, 8, 12)]
        ea, eb = _xtea_encrypt_block(a, b, k)
        a, b = ea ^ a, eb ^ b
    return a, b


def _lk_key(idno):
    a, b = _lk_hash(LOOKUP_SALT_KEY + "|" + idno, 0x436F6174, 0x65734B32)
    c, d = _lk_hash(idno + "|" + LOOKUP_SALT_KEY, 0x4C6F6F6B, 0x75703236)
    return [a, b, c, d]


def _lk_tag(idno):
    a, b = _lk_hash(LOOKUP_SALT_TAG + "|" + idno, 0x54616731, 0x54616732)
    return "{:08x}{:08x}".format(a, b)


def _lk_encrypt(idno, plaintext):
    key = _lk_key(idno)
    tag = _lk_tag(idno)
    nonce = int(tag[:8], 16)
    data = plaintext.encode("utf-8")
    out = bytearray()
    for i in range(0, len(data), 8):
        k0, k1 = _xtea_encrypt_block(nonce, i // 8, key)
        ks = k0.to_bytes(4, "big") + k1.to_bytes(4, "big")
        chunk = data[i:i + 8]
        out += bytes(cb ^ kb for cb, kb in zip(chunk, ks))
    return base64.b64encode(bytes(out)).decode("ascii")


def _lk_norm(idno):
    """One normalisation, applied identically in Python and the page JS:
    what the person types matches what we encrypted, however they type it."""
    s = re.sub(r"\s+", "", clean_text(idno)).upper()
    # ASCII digits only, to mirror the JS /^[0-9]+$/ exactly - str.isdigit()
    # is Unicode-aware and would diverge on exotic digit characters.
    if re.fullmatch(r"[0-9]+", s):
        s = s.lstrip("0") or "0"
    return s


def generate_lookup(models, onhire, asof, generated, date_tag):
    """RETIRED 26 Jul 2026 (Andrew's call): this was the OLD My Gear
    start-up page, and having two lookalike pages on the store laptop
    was a mix-up waiting to happen. The NEW My Gear (index.html - the
    one with the barcode reader and the Print A4 button, built by
    BUILD_MY_GEAR.py) is the only page now. lookup.html still gets
    written, but only as an instant redirect - so any old QR code or
    bookmark lands on the real thing instead of a dead link."""
    out_dir = os.path.join(BASE_DIR, "Gear_Lookup")
    os.makedirs(out_dir, exist_ok=True)
    with io.open(os.path.join(out_dir, "lookup.html"), "w",
                 encoding="utf-8") as f:
        f.write(
            "<!DOCTYPE html><html><head><meta charset='utf-8'>"
            "<meta http-equiv='refresh' content='0;url=index.html'>"
            "<title>My Gear - Coates K2</title></head>"
            "<body style='font-family:Segoe UI,Arial;background:#101317;"
            "color:#fff;text-align:center;padding-top:80px'>"
            "<p>This page has moved to <a href='index.html' "
            "style='color:#F26222'>My Gear</a> - taking you there now.</p>"
            "<script>location.replace('index.html')</script></body></html>")
    print("  Gear lookup: old page retired - lookup.html now redirects to "
          "the new My Gear (index.html, barcode + print).")
    return


def _generate_lookup_RETIRED(models, onhire, asof, generated, date_tag):
    """The old page builder, kept for reference only - never called."""
    if not onhire:
        print("  Gear lookup NOT rebuilt: no ON_HIRE export in the kit "
              "folder. Pull the ON_HIRE report from SiteIQ (it carries "
              "EXTERNAL_ID), drop it in this folder and rerun --lookup.")
        return
    # Asset numbers and register company names come from the rental view
    # when the exact unit (barcode) matches; the ON_HIRE row is fallback.
    by_bc = {}
    for m in models.values():
        for r in m["rows"]:
            bc = clean_text(r["barcode"]).upper()
            if bc and bc not in by_bc:
                by_bc[bc] = (r.get("asset") or "", m["display"])
    lk_asof = onhire["asof"] or asof
    asof_txt = (lk_asof.strftime("%d %b %Y %I:%M %p").lstrip("0")
                if lk_asof else "last refresh")
    today = dt.date.today()
    entries = {}
    seen_norm = {}
    n_items = 0
    for idno, person in onhire["people"].items():
        norm = _lk_norm(idno)
        if norm in seen_norm:
            safe_print("  WARNING: EXTERNAL_IDs {} and {} normalise to the "
                       "same lookup key - only the first is served.".format(
                           seen_norm[norm], idno))
            continue
        seen_norm[norm] = idno
        rows = []
        for it in person["items"]:
            asset, disp = by_bc.get(it["barcode"], ("", ""))
            days = (max(0, (today - it["on_hire_date"]).days)
                    if it["on_hire_date"] else "-")
            if not disp:
                disp = (display_company(canonical_company(
                    person["company_raw"]))
                    if person["company_raw"] else "-")
            rows.append([asset or it["item_number"] or "-",
                         it["description"], days,
                         fmt_date(it["on_hire_date"]), disp,
                         EC.badges_html(asset or it["item_number"],
                                        it["description"], wrap=False)])
        rows.sort(key=lambda r: -(r[2] if isinstance(r[2], int) else -1))
        n_items += len(rows)
        payload = json.dumps({"name": person["name"] or ("ID " + idno),
                              "idno": idno, "rows": rows,
                              "asof": asof_txt}, ensure_ascii=True)
        entries[_lk_tag(norm)] = _lk_encrypt(norm, payload)
    # build-time self-test vector the page verifies on every load
    st_payload = json.dumps({"name": "SELFTEST", "idno": "", "rows": [],
                             "asof": asof_txt})
    st_tag = _lk_tag("SELFTEST")
    entries[st_tag] = _lk_encrypt("SELFTEST", st_payload)

    #  MY GEAR OWNS Gear_Lookup\index.html. This older lookup page used to
    #  write the same file, so running --lookup on its own silently replaced
    #  My Gear on the tool store - the crews' signs say MY GEAR and they'd
    #  have got a different page. It writes lookup.html now and the two can
    #  never stand on each other again. (A. Fisher, 25 Jul 2026)
    out_dir = os.path.join(BASE_DIR, "Gear_Lookup")
    os.makedirs(out_dir, exist_ok=True)
    page = (LOOKUP_TEMPLATE
            .replace("__DATA__", json.dumps(entries))
            .replace("__ASOF__", asof_txt)
            .replace("__STTAG__", st_tag))
    with io.open(os.path.join(out_dir, "lookup.html"), "w",
                 encoding="utf-8") as f:
        f.write(page)
    print("  Gear lookup page: {} people, {} items on hire (source: {}) "
          "| Gear_Lookup\\lookup.html".format(
              len(seen_norm), n_items, os.path.basename(onhire["path"])))


LOOKUP_TEMPLATE = r"""<!DOCTYPE html><html><head><meta charset='utf-8'><link rel="icon" href="data:image/svg+xml,%3Csvg xmlns=%22http://www.w3.org/2000/svg%22 viewBox=%220 0 16 16%22%3E%3Ccircle cx=%228%22 cy=%228%22 r=%227%22 fill=%22%23F36F21%22/%3E%3C/svg%3E">
<meta name='viewport' content='width=device-width,initial-scale=1'>
<meta name='theme-color' content='#1D1D1B'>
<meta name='format-detection' content='telephone=no'>
<title>SiteIQ My Gear - Coates K2</title><style>
*{-webkit-tap-highlight-color:transparent}
body{font-family:'Segoe UI',Arial,sans-serif;background:#1D1D1B;color:#fff;
margin:0;display:flex;flex-direction:column;align-items:center;
min-height:100vh;-webkit-font-smoothing:antialiased}
.card{background:#fff;color:#1D1D1B;border-radius:12px;max-width:560px;
width:94%;margin:22px 0;padding:22px 20px;border-top:8px solid #F26222;
box-shadow:0 4px 18px rgba(0,0,0,.35);box-sizing:border-box}
h1{font-size:22px;margin:0 0 4px 0}
.pipe{color:#F26222}
.eyebrow{color:#F26222;font-weight:800;font-size:10px;letter-spacing:2px;
text-transform:uppercase}
h1.brand{font-size:27px;margin:3px 0 0 0;letter-spacing:-.5px}
.oj{color:#F26222}
.tagline{color:#8a8a8a;font-size:12px;font-weight:600;margin-top:2px}
.brandrule{height:3px;background:#F26222;border-radius:2px;
margin:10px 0 2px 0;width:52px}
.sub{color:#F26222;font-weight:700;font-size:13px}
.asof{color:#777;font-size:11px;margin-top:6px}
input{font-size:26px;padding:12px;width:100%;box-sizing:border-box;
border:2px solid #1D1D1B;border-radius:6px;margin-top:16px;
text-align:center;letter-spacing:2px}
button{font-size:18px;font-weight:700;background:#F26222;color:#fff;
border:0;border-radius:6px;padding:12px;width:100%;margin-top:10px}
table{border-collapse:collapse;width:100%;margin-top:14px;font-size:13px}
th{background:#1D1D1B;color:#fff;text-align:left;padding:6px 8px;
font-size:10px;text-transform:uppercase}
td{padding:6px 8px;border-bottom:1px solid #eee}
.days-hot{color:#B3261E;font-weight:700}
.cb{margin-top:4px;line-height:1.7}
.msg{margin-top:14px;font-size:13px;line-height:1.6}
.story{background:#FFF3EC;border-left:6px solid #F26222;padding:12px 14px;
margin-top:14px;font-size:14px;line-height:1.65;border-radius:0 8px 8px 0}
button.spk,button.shr{background:#1D1D1B;font-size:15px;margin-top:14px;
width:auto;display:inline-block;padding:11px 16px;margin-right:8px}
@media print{body{background:#fff}.foot,#entry,button,.noprint{
display:none!important}
.card{box-shadow:none;margin:0;max-width:none;width:100%;
border-radius:0;border-top:8px solid #F26222}}
.team{margin-top:16px;font-size:13px;color:#F26222;font-weight:700;
text-align:center;line-height:1.6}
.foot{color:#bbb;font-size:10px;margin:6px 0 22px 0;text-align:center;
line-height:1.7}
.err{color:#B3261E;font-weight:700;margin-top:12px;line-height:1.6}
.ok{color:#1E7B34}
@media(max-width:520px){
.card{padding:18px 14px;margin:12px 0}
table tr:first-child{display:none}
table,tbody,tr{display:block;width:100%}
tr{border:1px solid #eee;border-left:5px solid #F26222;border-radius:8px;
padding:10px 12px;margin-top:10px;background:#fff;
box-shadow:0 1px 3px rgba(0,0,0,.07);box-sizing:border-box}
td{display:flex;justify-content:space-between;align-items:baseline;
gap:12px;border:0;padding:3px 0;font-size:14px;text-align:right}
td::before{content:attr(data-l);font-weight:700;color:#888;
text-transform:uppercase;font-size:9px;letter-spacing:.6px;
flex-shrink:0;text-align:left}
}
</style></head><body>
<div class='card'>
<div class='eyebrow'>Coates &middot; K2 Tool Store</div>
<h1 class='brand'>SiteIQ <span class='oj'>My Gear</span></h1>
<div class='tagline'>Scan the window. See your gear. Your list,
nobody else&rsquo;s.</div>
<div class='brandrule'></div>
<div class='asof'>Position as at __ASOF__ &mdash; already returned
something? Then you're ahead of this page. Thank you.</div>
<div id='entry'>
<input id='idno' inputmode='numeric' autocomplete='off'
 placeholder='Enter your ID number'>
<button onclick='go()'>Show my gear</button>
<div id='err' class='err'></div>
</div>
<div id='result'></div>
</div>
<div class='foot'><b style='color:#F26222'>SiteIQ My Gear</b> &bull;
Two scans. Two looks. One standard. &bull; POWERED BY SITEIQ &bull;
Author: Andrew Fisher<br>
View-only &bull; nothing to log into &bull; no live connection into any
system &mdash; a window, not a door<br>
<span id='st'></span></div>
<script>
var DATA=__DATA__;
function u32(n){return n>>>0}
function enc(v0,v1,k){var d=0x9E3779B9,s=0,i;v0=u32(v0);v1=u32(v1);
for(i=0;i<32;i++){
v0=u32(v0+(u32((u32(v1<<4)^(v1>>>5))+v1)^u32(s+k[s&3])));
s=u32(s+d);
v1=u32(v1+(u32((u32(v0<<4)^(v0>>>5))+v0)^u32(s+k[(s>>>11)&3])));}
return[v0,v1]}
function pad(str){var b=[],i,c;for(i=0;i<str.length;i++){c=str.charCodeAt(i);
if(c<128)b.push(c);else if(c<2048){b.push(192|(c>>6),128|(c&63))}
else{b.push(224|(c>>12),128|((c>>6)&63),128|(c&63))}}
b.push(128);while(b.length%16)b.push(0);return b}
function hsh(str,i0,i1){var a=u32(i0),b=u32(i1),d=pad(str),i,j,k,r;
for(i=0;i<d.length;i+=16){k=[];for(j=0;j<16;j+=4){
k.push(u32((d[i+j]<<24)|(d[i+j+1]<<16)|(d[i+j+2]<<8)|d[i+j+3]))}
r=enc(a,b,k);a=u32(r[0]^a);b=u32(r[1]^b)}return[a,b]}
function hx(n){var s=n.toString(16);while(s.length<8)s='0'+s;return s}
function tag(id){var r=hsh('__TS__|'+id,0x54616731,0x54616732);
return hx(r[0])+hx(r[1])}
function key(id){var r1=hsh('__KS__|'+id,0x436F6174,0x65734B32),
r2=hsh(id+'|__KS__',0x4C6F6F6B,0x75703236);
return[r1[0],r1[1],r2[0],r2[1]]}
function b64b(s){var raw=atob(s),o=[],i;for(i=0;i<raw.length;i++)
o.push(raw.charCodeAt(i));return o}
function dec(id,b64){var k=key(id),t=tag(id),non=parseInt(t.substr(0,8),16),
c=b64b(b64),o=[],i,j,ks,r;
for(i=0;i<c.length;i+=8){r=enc(non,i/8,k);
ks=[(r[0]>>>24)&255,(r[0]>>>16)&255,(r[0]>>>8)&255,r[0]&255,
(r[1]>>>24)&255,(r[1]>>>16)&255,(r[1]>>>8)&255,r[1]&255];
for(j=0;j<8&&i+j<c.length;j++)o.push(c[i+j]^ks[j])}
try{return decodeURIComponent(o.map(function(x){
var h=x.toString(16);return '%'+(h.length<2?'0'+h:h)}).join(''))}
catch(e){return null}}
var timer=null,SPK=null,LAST=null;
function norm(s){s=(s||'').replace(/\s+/g,'').toUpperCase();
if(/^[0-9]+$/.test(s)){s=s.replace(/^0+/,'')||'0'}return s}
function speakIt(){if(!window.speechSynthesis||!SPK)return;
bump();speechSynthesis.cancel();
var u=new SpeechSynthesisUtterance(SPK);u.rate=0.95;
var vs=speechSynthesis.getVoices(),i;
for(i=0;i<vs.length;i++){
if(vs[i].lang&&vs[i].lang.indexOf('en-AU')===0){u.voice=vs[i];break}}
speechSynthesis.speak(u)}
function clearAll(){document.getElementById('result').innerHTML='';
document.getElementById('idno').value='';
document.getElementById('err').textContent='';
document.getElementById('entry').style.display='block';
if(window.speechSynthesis)speechSynthesis.cancel();SPK=null;LAST=null;
var ov=document.getElementById('snapov');
if(ov)document.body.removeChild(ov)}
function bump(){if(timer)clearTimeout(timer);
timer=setTimeout(clearAll,45000)}
function trm(s,n){s=String(s);return s.length>n?s.slice(0,n-1)+'…':s}
function snapIt(){if(!LAST)return;bump();
var p=LAST,rows=p.rows,W=1080,pad=54;
var H=300+(rows.length?rows.length*150:96)+150;
var c=document.createElement('canvas');c.width=W;c.height=H;
var g=c.getContext('2d');
g.fillStyle='#fff';g.fillRect(0,0,W,H);
g.fillStyle='#F26222';g.fillRect(0,0,W,16);
g.fillStyle='#F26222';g.font='bold 24px Segoe UI,Arial';
g.fillText('C O A T E S   ·   K 2   T O O L   S T O R E',pad,64);
g.fillStyle='#1D1D1B';g.font='bold 56px Segoe UI,Arial';
g.fillText('SiteIQ',pad,124);
var bw=g.measureText('SiteIQ ').width;
g.fillStyle='#F26222';
g.fillText('My Gear',pad+bw,124);
g.fillStyle='#777777';g.font='26px Segoe UI,Arial';
g.fillText('As at '+p.asof,pad,166);
g.fillStyle='#F26222';g.fillRect(pad,184,90,6);
g.fillStyle='#1D1D1B';g.font='bold 42px Segoe UI,Arial';
g.fillText(trm(p.name,26)+(p.idno?'  •  ID '+p.idno:''),pad,246);
var y=286;
rows.forEach(function(r){
g.fillStyle='#FAFAFA';g.fillRect(pad,y,W-2*pad,132);
g.fillStyle='#F26222';g.fillRect(pad,y,10,132);
g.fillStyle='#1D1D1B';g.font='bold 30px Segoe UI,Arial';
g.fillText(trm(r[1],54),pad+34,y+46);
g.fillStyle='#555555';g.font='26px Segoe UI,Arial';
g.fillText('Asset '+r[0]+'   •   on hire '+r[3]+
(typeof r[2]==='number'?'   •   out '+r[2]+
(r[2]===1?' day':' days'):''),pad+34,y+88);
g.fillStyle='#888888';g.font='24px Segoe UI,Arial';
g.fillText('For '+trm(r[4],60),pad+34,y+118);
y+=150});
if(!rows.length){g.fillStyle='#1E7B34';
g.font='bold 34px Segoe UI,Arial';
g.fillText('Nothing on hire - a clean sheet!',pad,y+46);y+=96}
g.fillStyle='#1D1D1B';g.fillRect(0,H-120,W,120);
g.fillStyle='#ffffff';g.font='26px Segoe UI,Arial';
g.fillText('Finished with gear? Over the counter - off your name in seconds.',pad,H-70);
g.fillStyle='#F26222';g.font='bold 24px Segoe UI,Arial';
g.fillText('SiteIQ My Gear  |  Two scans. Two looks. One standard.  |  POWERED BY SITEIQ',pad,H-32);
var url=c.toDataURL('image/png');
var ov=document.createElement('div');ov.id='snapov';
ov.style.cssText='position:fixed;top:0;left:0;right:0;bottom:0;'+
'background:rgba(0,0,0,.93);z-index:99;overflow:auto;'+
'text-align:center;padding:16px';
ov.innerHTML='<img src="'+url+'" style="width:94%;max-width:520px;'+
'border-radius:8px;margin-top:6px">'+
'<div style="color:#fff;font-size:13px;margin:10px 0">Press and hold '+
'the image to save it to your phone - or use Download.</div>'+
'<a download="SiteIQ_My_Gear.png" href="'+url+'" '+
'style="display:inline-block;background:#F26222;color:#fff;'+
'font-weight:700;border-radius:6px;padding:12px 18px;'+
'text-decoration:none;margin:4px">Download image</a>'+
'<button onclick="var o=document.getElementById(\'snapov\');'+
'if(o)document.body.removeChild(o)" style="width:auto;'+
'background:#fff;color:#1D1D1B;padding:12px 18px;margin:4px">'+
'Close</button>';
document.body.appendChild(ov)}
function printIt(){if(!LAST)return;bump();window.print()}
function go(){var id=norm(document.getElementById('idno').value);
var errEl=document.getElementById('err');errEl.textContent='';
if(!id){errEl.textContent='Type your ID number first.';return}
var t=tag(id);if(!(t in DATA)){
errEl.innerHTML='<span class="ok"><b>Nothing on hire against that ID '+
'as at the last refresh - a clean sheet!</b></span><br>Typo? Have '+
'another go. Doesn&rsquo;t sound right? See us at the counter - '+
'happy to help.';
return}
var txt=dec(id,DATA[t]);var p=null;
try{p=JSON.parse(txt)}catch(e){}
if(!p){errEl.textContent='No gear found for that ID - see the counter.';return}
LAST=p;
var first=((p.name||'').split(' - ')[0]||'').split(' ')[0]||p.name;
var h='<h1 style="margin-top:10px">Good to see you, '+first+'</h1>';
if(p.idno){h+='<div class="sub">ID '+p.idno+'</div>'}
if(p.rows.length){var n=p.rows.length;
h+='<div class="story"><b>Your gear at a glance</b> - '+
n+(n===1?' item':' items')+' in your name, and every one was looked '+
'over before it hit your hand: clean, compliant, in date, ready for '+
'work. That&rsquo;s Life Saving Rule 5 - Tools and Equipment '+
'(SEQ-GL-009) - looking after you on every shift. Checking your list '+
'is part of working safe - thank you for taking the look.</div>';
h+='<table><tr><th>Asset No</th><th>Item</th>'+
'<th>Days</th><th>On Hire</th><th>For</th></tr>';
p.rows.forEach(function(r){
h+='<tr><td data-l="Asset No"><b>'+r[0]+'</b></td>'+
'<td data-l="Item">'+r[1]+(r[5]?'<div class="cb">'+r[5]+'</div>':'')+'</td>'+
'<td data-l="Days Out" class="'+
(typeof r[2]==='number'&&r[2]>=3?'days-hot':'')+'">'+r[2]+
'</td><td data-l="On Hire">'+r[3]+'</td>'+
'<td data-l="For">'+r[4]+'</td></tr>'});
h+='</table><div class="msg"><b>Still using it?</b> Perfect - '+
'that&rsquo;s gear working hard, exactly what it&rsquo;s here for. '+
'<b>Finished with any of it?</b> Straight over the counter - two scans, '+
'two looks, off your name in seconds, checked on the spot and back on '+
'the shelf ready for the next crew. Every return you make keeps the '+
'whole site moving.</div>';
h+='<div class="team">Care Deeply &bull; One Team<br>The K2 gear story '+
'stays clean because of people like you. Thank you.</div>';
var say='Good to see you, '+first+'. You have '+n+
(n===1?' item':' items')+' in your name, all checked and ready for '+
'work. ';
if(n<=2){p.rows.forEach(function(r){say+=r[1]+
(typeof r[2]==='number'?', out '+r[2]+(r[2]===1?' day':' days'):'')+
'. '})}
else{var mx=null;p.rows.forEach(function(r){
if(typeof r[2]==='number'&&(mx===null||r[2]>mx))mx=r[2]});
say+='Your full list is on the screen'+
(mx!==null?', and the longest out is '+mx+(mx===1?' day':' days'):'')+
'. '}
say+='Still using it? All good. Finished with any of it? Straight over '+
'the counter, off your name in seconds. Working safe, working '+
'together. Thank you. One team!';
SPK=say}
else{h+='<div class="msg ok"><b>Nothing on hire in your name - a clean '+
'sheet!</b> That&rsquo;s the gold standard, and it doesn&rsquo;t happen '+
'by accident. Thanks for keeping the gear moving for the whole '+
'team.</div>';
h+='<div class="team">Care Deeply &bull; One Team<br>Need gear? '+
'We&rsquo;ll set you up at the counter - checked, tagged and ready.'+
'</div>';
SPK='Good to see you, '+first+'. Nothing on hire in your name - a '+
'clean sheet, and that is the gold standard. Thank you for keeping '+
'the gear moving safely. One team!'}
h+='<div class="noprint">';
if(window.speechSynthesis){
h+='<button class="spk" onclick="speakIt()">&#128266; Read it to me'+
'</button>'}
h+='<button class="shr" onclick="snapIt()">&#128247; Save as image'+
'</button>'+
'<button class="shr" onclick="printIt()">&#128424;&#65039; Save as PDF'+
'</button></div>';
h+='<div class="asof noprint">This screen clears itself shortly.</div>';
document.getElementById('result').innerHTML=h;
document.getElementById('entry').style.display='none';
if(timer)clearTimeout(timer);timer=setTimeout(clearAll,45000)}
document.getElementById('idno').addEventListener('keydown',
function(e){if(e.key==='Enter')go()});
(function(){var t='__STTAG__',ok=false;
try{var x=dec('SELFTEST',DATA[t]);ok=!!(x&&JSON.parse(x).name==='SELFTEST')}
catch(e){}
document.getElementById('st').textContent= ok?
'lookup engine verified':'engine check failed - see the counter for your list';
})();
</script></body></html>"""
LOOKUP_TEMPLATE = LOOKUP_TEMPLATE.replace("__KS__", LOOKUP_SALT_KEY)
LOOKUP_TEMPLATE = LOOKUP_TEMPLATE.replace("__TS__", LOOKUP_SALT_TAG)


def render_audit_html(pm, asof, generated, source_line):
    rows = []
    last_cat = None
    for p in pm["audit_rows"]:
        if p["category"] != last_cat:
            rows.append('<tr><td colspan="6" style="background:{};color:#fff;'
                        'font-weight:bold;">{}</td></tr>'.format(
                            COATES_DARK, esc(p["category"])))
            last_cat = p["category"]
        rows.append(
            "<tr><td style='font-weight:bold;white-space:nowrap;'>{}</td>"
            "<td>{}</td><td>{}</td>"
            "<td class='tick'>&#9744;</td><td class='tick'>&#9744;</td>"
            "<td style='min-width:38mm;'>&nbsp;</td></tr>".format(
                asset_disp(p.get("asset")),
                desc_disp(p["description"], p.get("asset")),
                esc(p["state"])))
    n = len(pm["audit_rows"])
    audit_css = (
        " td.tick { text-align:center; font-size:13pt; }"
        " .sign { margin-top:14px; font-size:10pt; page-break-inside:avoid; }"
        " .sign table { width:100%; border-collapse:collapse; font-size:10pt; }"
        " .sign td.l { white-space:nowrap; width:1%; padding:0 4px 0 0; }"
        " .sign td.l+td.l, .sign td.w+td.l { padding-left:14px; }"
        " .sign td.w { border-bottom:1px solid #555; }")
    body = (
        "<div class='intro story'><b>The audit, plain and simple:</b> walk the list "
        "and tick every asset you physically sight. Anything you cannot sight "
        "that is not on hire has left the yard without going on hire - "
        "investigate it today, same shift. While you walk: idle gear is the "
        "Coates mechanic's preventative maintenance window - tick Condition OK "
        "or note what it needs. The right way, every time.</div>"
        "<h2>Idle plant checklist - {n} assets, by category</h2>"
        "<table class='data'><thead><tr><th>Asset Number</th><th>Description</th>"
        "<th>State</th><th>Sighted</th><th>Condition OK</th>"
        "<th>Maintenance / Notes</th></tr></thead><tbody>{rows}</tbody></table>"
        "<div class='note'>Life Saving Rule 5 - Tools and Equipment "
        "(SEQ-GL-009): anything found damaged or defective is tagged out and "
        "removed from service on the spot - never left on the shelf, never "
        "reissued. Items on hire to contractors are NOT on this list - they "
        "are on the company on-hire reports.</div>"
        "<div class='sign'><table>"
        "<tr><td class='l'>Audited by:</td><td class='w'>&nbsp;</td>"
        "<td class='l'>Signature:</td><td class='w'>&nbsp;</td>"
        "<td class='l'>Date / time:</td><td class='w'>&nbsp;</td></tr>"
        "<tr><td colspan='6' style='height:16px;border:0;'>&nbsp;</td></tr>"
        "<tr><td class='l'>Mechanic review:</td><td class='w'>&nbsp;</td>"
        "<td class='l'>Signature:</td><td class='w'>&nbsp;</td>"
        "<td class='l'>Date / time:</td><td class='w'>&nbsp;</td></tr>"
        "</table></div>"
        "<div class='limits'><b>Honest limits:</b> This list reflects SiteIQ "
        "as at the data-as-at stamp above. Idle = no hirer recorded, or on "
        "hire to Site Plant Equipment (Coates) under the return-to-site "
        "convention. Sight the asset number, not the description.</div>"
    ).format(n=n, rows="".join(rows))
    return paged_report_doc(
        "Plant Audit List", "Idle Plant Checklist", asof, generated,
        source_line, body, doc_title="Coates K2 Plant Audit List",
        extra_css=audit_css)


def generate_audit(pm, asof, generated, source_line, date_tag):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stem = "Coates_K2_Plant_Audit_List_{}".format(date_tag)
    html_path = os.path.join(PAGES_DIR, stem + ".html")
    pdf_path = os.path.join(PDF_DIR, stem + ".pdf")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(render_audit_html(pm, asof, generated, source_line))
    pdf_ok = html_to_pdf(html_path, pdf_path)
    print("  Plant audit list: {} idle assets | HTML{}".format(
        len(pm["audit_rows"]), " + PDF" if pdf_ok else " (no PDF engine)"))
    return pdf_ok


# ---------------------------------------------------------------------------
# Executive Summary - the whole K2 story on one pack
# ---------------------------------------------------------------------------

def build_exec_model(models, pm, tx, sales, stocktake, tracking, radio_fleet,
                     shift_day, tx_loaded, sales_loaded, gas_fleet=0,
                     radio_on_hire=0, gas_on_hire=0, confirmed=None,
                     gear_dispatch=None):
    all_models = list(models.values())
    n_items = sum(m["n_items"] for m in all_models)
    people = set()
    for m in all_models:
        people.update(m["people"])
    n_tx = sum(m["n_tx_total"] for m in all_models)
    same_day = sum(m["same_day"] for m in all_models)
    kept = sum(m["kept_longer"] for m in all_models)
    repl_total = sum(m["repl_exposure"] or 0 for m in all_models)

    # most-used: current on-hire item types by count, across all companies
    type_count = {}
    for m in all_models:
        for r in m["rows"]:
            if r["description"]:
                type_count[r["description"]] = type_count.get(r["description"], 0) + 1
    top_items = sorted(type_count.items(), key=lambda x: (-x[1], x[0]))[:5]

    sku_qty = {}
    for mv in sales:
        sku_qty[mv["sku"]] = sku_qty.get(mv["sku"], 0) + mv["qty"]
    top_cons = sorted(sku_qty.items(), key=lambda x: (-x[1], x[0]))[:5]

    return {
        "shift_day": shift_day,
        "models": models,          # per-company models - v3 risk/score pages
        "tx_rows": tx,             # raw transaction history - v3 analytics
        "n_items": n_items,
        "by_company": sorted(((m["display"], m["n_items"]) for m in all_models),
                             key=lambda t: -t[1]),
        "n_companies": len(all_models),
        "n_people": len(people),
        "n_tx": n_tx, "same_day": same_day, "kept": kept,
        "tx_loaded": tx_loaded, "sales_loaded": sales_loaded,
        "repl_total": repl_total,
        "top_items": top_items, "top_cons": top_cons,
        "pm": pm,
        "stocktake": stocktake,
        "tracking": tracking,
        "radio_fleet": radio_fleet,
        "gas_fleet": gas_fleet,
        "radio_on_hire": radio_on_hire,
        "gas_on_hire": gas_on_hire,
        "confirmed": confirmed,
        "gear_dispatch": gear_dispatch,
    }


def _exec_legacy_sections(x, asof, generated, source_line):
    pm = x["pm"]
    st = x["stocktake"]
    tr = x["tracking"]
    shift_txt = x["shift_day"].strftime("%d %b %Y")

    # ---- the position, told straight --------------------------------------
    bits = []
    bits.append("Coates has {n} item{s} on hire across {c} compan{cy}, with "
                "{p} people from site crews using the tool store.".format(
                    n=x["n_items"], s="" if x["n_items"] == 1 else "s",
                    c=x["n_companies"],
                    cy="y" if x["n_companies"] == 1 else "ies",
                    p=x["n_people"]))
    if st:
        bits.append("{sv} of {tot} tool store items ({pct:.0%}) have been "
                    "physically sighted in the last 3 days by {cnt} different "
                    "counters - every item, every day, evidenced in SiteIQ.".format(
                        sv=st["sighted3"], tot=st["total"], pct=st["pct"],
                        cnt=st["counters"]))
    if pm["idle"]:
        bits.append("Idle plant worth {v} per day is flagged for off-hire - "
                    "found by our own return-to-site discipline.".format(
                        v=fmt_money(pm["saving_per_day"])))
    else:
        bits.append("No plant is sitting idle on charge - the return-to-site "
                    "convention keeps every charged item accounted for.")
    gd = x.get("gear_dispatch")
    if gd:
        days = (gd["period_end"] - gd["period_start"]).days
        for label, fleet_key, stats in (
                ("radio handsets", "radio_fleet", gd["radios"]),
                ("gas monitors", "gas_fleet", gd["gas_monitors"])):
            fleet = x.get(fleet_key) or 0
            never = fleet - stats["dispatched"]
            if stats["dispatched"] == 0:
                bits.append(
                    "{lbl_cap}: none of the {fleet} have left the shelf in the "
                    "last {d} days - the whole fleet sits ready and untouched.".format(
                        lbl_cap=label.capitalize(), fleet=fleet, d=days))
            elif stats["still_out"] == 0:
                bits.append(
                    "{lbl_cap}: {disp} of {fleet} have gone out in the last {d} "
                    "days, and every one came back the same trip - none still "
                    "out. {never} {lbl} haven't left the shelf yet.".format(
                        lbl_cap=label.capitalize(), lbl=label, disp=stats["dispatched"],
                        fleet=fleet, d=days, never=never))
            else:
                bits.append(
                    "{lbl_cap}: {disp} of {fleet} have gone out in the last {d} "
                    "days - {ret} back already, {out} still with a crew, not "
                    "yet returned. {never} {lbl} haven't left the shelf.".format(
                        lbl_cap=label.capitalize(), lbl=label, disp=stats["dispatched"],
                        fleet=fleet, d=days, ret=stats["returned"],
                        out=stats["still_out"], never=never))
    elif x.get("gas_fleet") or x.get("radio_fleet"):
        bits.append("From yesterday's count, {ro} of {rf} radio handsets and "
                    "{go} of {gf} gas monitors were out with crews - the rest "
                    "sit ready in the store, charged and bump-tested for the "
                    "next shift.".format(
                        ro=x.get("radio_on_hire", 0), rf=x.get("radio_fleet", 0),
                        go=x.get("gas_on_hire", 0), gf=x.get("gas_fleet", 0)))
    bits.append("Every movement in or out is double-scanned. Nothing leaves "
                "unrecorded; nothing comes back unchecked.")
    if x.get("repl_total"):
        bits.append("Replacement exposure riding across every company: "
                    "<span class='repl'>{}</span>.".format(
                        fmt_money(x["repl_total"])))
    position = " ".join(bits)

    tiles = [
        (str(x["n_items"]), "Items On Hire"),
        (str(x["n_companies"]), "Companies"),
        (str(x["n_people"]), "People Using Store"),
        (str(x["n_tx"]) if (x["tx_loaded"] or x["sales_loaded"]) else "-",
         "Transactions"),
        ("{:.0%}".format(st["pct"]) if st else "-", "Fleet Verified (3d)"),
        ("<span class='repl'>{}</span>".format(fmt_money(x["repl_total"]))
         if x.get("repl_total") else
         (fmt_money(pm["saving_per_day"]) if pm["idle"] else "$0.00"),
         "Replacement Exposure" if x.get("repl_total") else "Idle Plant / Day"),
    ]
    #  Tile values are all generated in this file (money/counts/spans) -
    #  never user text - so they render raw; labels still escape.
    tiles_h = '<table class="tiles"><tr>{}</tr></table>'.format(
        "".join('<td><span class="v">{}</span><span class="l">{}</span></td>'
                .format(v, esc(l)) for v, l in tiles))
    # page 1 gets colour of its own: the three bars that tell the shut's
    # story at a glance (the full visual set follows on the glance page)
    strip = ""
    if tr and tr.get("fc_full"):
        run_frac = (tr.get("act_to_date") or 0) / tr["fc_full"]
        strip += hbar("Shut plan {} - actual so far".format(fmt_money(tr["fc_full"])),
                      fmt_money(tr.get("act_to_date") or 0), run_frac, "#3FB950")
    if st:
        strip += hbar("Store verified on the 3-day cycle",
                      "{:.0%}".format(st["pct"]), st["pct"],
                      rag_colour(st["pct"]))
    fleet = pm.get("fleet") or 0
    if fleet:
        dep = len(pm.get("in_use") or [])
        strip += hbar("Site plant out with crews", "{} of {}".format(dep, fleet),
                      dep / fleet)
    if strip:
        tiles_h += ("<div style='margin-top:12px'>" + strip + "</div>")

    # ---- forecast v actual ------------------------------------------------
    if tr and tr["today"]:
        rows = []
        t_fc = 0.0
        t_ac = 0.0
        any_actual = False
        for name, fc, ac in tr["today"]:
            fcv = fc or 0.0
            acv = ac if ac is not None else None
            t_fc += fcv
            if acv is not None:
                t_ac += acv
                any_actual = True
            var = (acv - fcv) if acv is not None else None
            rows.append("<tr><td>{}</td><td class='num'>{}</td>"
                        "<td class='num'>{}</td><td class='num'>{}</td></tr>".format(
                            esc(name), fmt_money(fcv),
                            fmt_money(acv) if acv is not None else "-",
                            fmt_money(var) if var is not None else "-"))
        fva = ('<h2>Forecast v actual - {d}</h2>'
               '<table class="data"><thead><tr><th>Category</th>'
               '<th class="num">Forecast</th><th class="num">Actual</th>'
               '<th class="num">Variance</th></tr></thead><tbody>{r}'
               '<tr><td><b>Day total</b></td><td class="num"><b>{tf}</b></td>'
               '<td class="num"><b>{ta}</b></td><td class="num"><b>{tv}</b></td>'
               '</tr></tbody></table>'
               '<div class="note">Shutdown to date: actual {atd} against '
               '{ftd} forecast; full-shutdown forecast {ff}. A shift day runs '
               '6:00am to 6:00am. Manual categories (accommodation, transport, '
               'damage) show once entered.</div>').format(
            d=shift_txt, r="".join(rows), tf=fmt_money(t_fc),
            # No actuals entered yet = an absent feed, never a $0.00 nil.
            ta=fmt_money(t_ac) if any_actual else "-",
            tv=fmt_money(t_ac - t_fc) if any_actual else "-",
            atd=fmt_money(tr["act_to_date"]), ftd=fmt_money(tr["fc_to_date"]),
            ff=fmt_money(tr["fc_full"]))
    else:
        fva = ('<h2>Forecast v actual</h2><div class="note">Daily Tracking '
               'was not readable this run or has no row for {d}.</div>'.format(
                   d=shift_txt))

    # ---- SiteIQ billing confirmation ---------------------------------------
    # PREFILL always writes today's Actual immediately at 5am from the live
    # snapshot - it never waits on SiteIQ's real invoice. This section is
    # the separate catch-up: once DAILY_SUMMARY has a confirmed total for
    # this day, show it alongside our own estimate (never overwriting the
    # Plant/Tooling detail above - just confirming or flagging it).
    confirmed = x.get("confirmed")
    if confirmed:
        actual_by_name = {}
        if tr and tr["today"]:
            actual_by_name = {name: (ac or 0.0) for name, fc, ac in tr["today"]}
        ours_by_cat = {
            "HIRE": (actual_by_name.get("Plant Equipment", 0.0) +
                    actual_by_name.get("Tooling", 0.0)),
            "LABOUR": actual_by_name.get("Personnel / Labour", 0.0),
            "TRANSPORT": actual_by_name.get("Transport", 0.0),
            "DAMAGE": actual_by_name.get("Damage Recovery", 0.0),
        }
        conf_rows = []
        for col, label in DAILY_SUMMARY_CATEGORIES.items():
            if col not in confirmed:
                continue
            theirs = confirmed[col]
            ours = ours_by_cat.get(col, 0.0)
            variance = ours - theirs
            if abs(variance) < 5:
                note = "Matches"
            elif theirs == 0 and ours > 0:
                note = "Not yet invoiced by SiteIQ"
            else:
                note = "Variance - worth a look"
            conf_rows.append(
                "<tr><td>{}</td><td class='num'>{}</td><td class='num'>{}</td>"
                "<td class='num'>{}</td><td>{}</td></tr>".format(
                    esc(label), fmt_money(ours), fmt_money(theirs),
                    fmt_money(variance), esc(note)))
        conf_html = ('<h2>SiteIQ billing confirmation - {d}</h2>'
                    '<table class="data"><thead><tr><th>Category</th>'
                    '<th class="num">Our Actual</th><th class="num">SiteIQ Invoiced</th>'
                    '<th class="num">Variance</th><th>Note</th></tr></thead><tbody>{r}'
                    '</tbody></table>'
                    '<div class="note">Radios, gas monitors and subhire never appear '
                    'here - confirmed against the Contracted Rates card, they have no '
                    'rate entry so SiteIQ cannot invoice them through this channel. '
                    'Accommodation isn\'t compared yet - no confirmed column mapping.'
                    '</div>').format(d=shift_txt, r="".join(conf_rows))
        fva = fva + conf_html

    # ---- movement and returns --------------------------------------------
    if x["tx_loaded"] and (x["same_day"] or x["kept"] or x["n_items"]):
        mv = ('<h2>Movement and returns</h2>'
              '<table class="aging"><tr><th>Returned same day</th>'
              '<th>Kept longer, returned</th><th>Still on hire</th></tr>'
              '<tr><td class="g">{sd}</td><td class="a">{k}</td>'
              '<td class="r">{oh}</td></tr></table>'
              '<div class="note">Same-day returns are the standard we promote '
              'with every crew: bring it back when it\'s not in use. Gear that '
              'comes back gets checked, maintained and back on the shelf for '
              'the next job.</div>').format(sd=x["same_day"], k=x["kept"],
                                            oh=x["n_items"])
    else:
        mv = ('<h2>Movement and returns</h2>'
              '<div class="note">Transaction history builds as SiteIQ records '
              'it - same-day and multi-day return counts appear here. The '
              'standard stands from day one: bring it back when it\'s not in '
              'use.</div>')

    # ---- what's working hardest ------------------------------------------
    hardest = ""
    if x["top_items"]:
        rws = "".join("<tr><td>{}</td><td class='num'>{}</td></tr>".format(
            esc(d), n) for d, n in x["top_items"])
        hardest += ('<h2>Working hardest - hire gear</h2>'
                    '<table class="data"><thead><tr><th>Item</th>'
                    '<th class="num">On Hire Now</th></tr></thead>'
                    '<tbody>{}</tbody></table>').format(rws)
    if x["top_cons"]:
        rws = "".join("<tr><td>{}</td><td class='num'>{:g}</td></tr>".format(
            esc(d), n) for d, n in x["top_cons"])
        hardest += ('<h2>Working hardest - consumables</h2>'
                    '<table class="data"><thead><tr><th>SKU</th>'
                    '<th class="num">Qty Taken</th></tr></thead>'
                    '<tbody>{}</tbody></table>'
                    '<div class="note">Quantities only - no consumables price '
                    'feed is loaded. Utilisation evidence sizes the next '
                    'order.</div>').format(rws)
    if not hardest:
        hardest = ('<h2>Working hardest</h2><div class="note">Usage rankings '
                   'build as hire and consumables movement is recorded - the '
                   'shutdown proper starts them moving.</div>')

    # ---- plant utilisation ------------------------------------------------
    util = (len(pm["in_use"]) / pm["fleet"]) if pm["fleet"] else 0
    plant_h = ('<h2>Site plant - utilisation and idle value</h2>'
               '<div class="note">{iu} of {fl} site plant items in use '
               '({u:.0%} true utilisation). {idle} Barriers, chutes and '
               'hoppers ({inf} items) stay on site for the duration by '
               'design. The not-in-use fleet doubles as the Coates '
               'mechanic\'s preventative maintenance window - gear is '
               'serviced while it sits, not after it fails.</div>').format(
        iu=len(pm["in_use"]), fl=pm["fleet"], u=util,
        idle=("Idle-on-charge plant is worth {} per day - the off-hire "
              "conversation list.".format(fmt_money(pm["saving_per_day"]))
              if pm["idle"] else
              "Nothing is idle on charge under the return-to-site convention."),
        inf=pm["infra_total"])

    # ---- the standards behind the numbers --------------------------------
    standards = (
        '<h2>The standards behind the numbers</h2>'
        '<div class="note"><b>Double scan - two scans, two looks, one '
        'standard.</b> Every issue and every return goes through the tool '
        'store: scan, look. Scan, look. The scan is not complete until we '
        'have looked - so the record above is evidence, not estimate. Right '
        'first time. Every item. Every time.</div>'
        '<div class="note"><b>Checking and stocktakes.</b> Stock takes are '
        'done daily - no exceptions - with nothing in the store going more '
        'than 30 days unscanned. '
        + ('{sv} of {tot} items physically verified in the last 3 days '
           '({pct:.0%}) by {cnt} counters. '.format(
               sv=st["sighted3"], tot=st["total"], pct=st["pct"],
               cnt=st["counters"]) if st else
           'Stocktake coverage reports from the SiteIQ export each day. ')
        + 'Discrepancies are addressed on the spot, and the stock take '
        'doubles as our on-hire verification - anything physically in the '
        'store still showing as on hire gets corrected the same day. It is '
        'not just a counting exercise; it is another checkpoint for catching '
        'problems before a hirer does.</div>'
        '<div class="note"><b>Safety-critical assurance.</b> {rf} radio '
        'handsets ({rb} billable, {rs} spare{rsp} held) and the {gf}-unit '
        'gas monitor '
        'fleet are pooled, charged and maintained through the store - every '
        'gas monitor bump tested and fully charged before it is issued, and '
        'every battery item issued on a full charge, never a partial. '
        'Nothing leaves the store that we would not use ourselves. Life Saving '
        'Rules (SEQ-GL-009) apply: Rule 5 - Tools and Equipment - damaged or '
        'defective gear is tagged out and removed from service, never '
        'reissued.</div>'
        '<div class="note"><b>Pride in what we do.</b> Best Service &amp; '
        'Value is the objective; the disciplines above are how the K2 tool '
        'store delivers it, every shift.</div>').format(
            rf=x["radio_fleet"],
            rb=min(x["radio_fleet"], RADIO_BILLABLE_CAP),
            rs=max(0, x["radio_fleet"] - RADIO_BILLABLE_CAP),
            rsp="" if max(0, x["radio_fleet"] - RADIO_BILLABLE_CAP) == 1 else "s",
            gf=x.get("gas_fleet") or 30)

    limits = (
        'Forecast and actual figures come from the K2 workbook Daily Tracking '
        '(read-only). Counts come from SiteIQ exports as at the data-as-at '
        'stamp; a consumables draw counts as one transaction regardless of '
        'quantity. Unpriced items are excluded from value figures, never '
        'estimated. Radio billing: 70 of 72 handsets at $12.81/day from '
        '17 Jul 2026, 2 disclosed unbilled spares. Gas billing: 30-unit fleet '
        'at $29.75/day from 24 Jul 2026 to the forecast end, on hire or not. '
        'This pack is a point-in-time snapshot.')

    # ---- the picture at a glance: the story drawn, not just told ---------
    _vbar = hbar

    glance = "<h2>The picture at a glance</h2>"
    rings = ""
    if st:
        rings += ring(st["pct"], "Store verified on the 3-day cycle", "sighted")
    fleet0 = pm.get("fleet") or 0
    if fleet0:
        dep0 = len(pm.get("in_use") or [])
        rings += ring(dep0 / fleet0, "Site plant out with crews", "deployed",
                      colour="#F26222")
    if x.get("n_tx") and x.get("same_day") is not None and x["n_tx"]:
        rings += ring(x["same_day"] / x["n_tx"], "Same-day returns",
                      "of transactions", colour="#3FB950")
    if tr and tr.get("fc_full"):
        rings += ring((tr.get("act_to_date") or 0) / tr["fc_full"],
                      "Shut plan consumed so far", "of plan", colour="#8B9099")
    if rings:
        glance += ("<div style='text-align:center;margin:4px 0 10px'>"
                   + rings + "</div>")
    co_rank = (x.get("by_company") or [])[:8]
    if co_rank:
        mx = max(n for _c, n in co_rank) or 1
        glance += ("<div class='note' style='margin-bottom:4px'><b>Gear on "
                   "hire by company</b> - who is carrying the tool store's "
                   "work right now.</div>"
                   + "".join(_vbar(c, "{} items".format(n), n / mx)
                             for c, n in co_rank))
    fleet = pm["fleet"] or 0
    if fleet:
        dep = len(pm["in_use"])
        frac = dep / fleet
        glance += ("<div class='note' style='margin:10px 0 4px'><b>Site plant "
                   "working</b></div>"
                   + _vbar("Deployed with crews", "{} of {}".format(dep, fleet),
                           frac, "#3FB950" if frac >= 0.5 else "#F26222"))
    if st:
        frac = st["pct"]
        glance += ("<div class='note' style='margin:10px 0 4px'><b>Stocktake "
                   "coverage</b> - sighted inside the 3-day cycle.</div>"
                   + _vbar("Store on the cycle", "{:.0%}".format(frac), frac,
                           "#3FB950" if frac >= 0.95 else
                           ("#fab219" if frac >= 0.80 else "#F85149")))
    if tr and tr.get("fc_to_date"):
        fc_td = tr["fc_to_date"]
        ac_td = tr.get("act_to_date") or 0.0
        fc_full = tr.get("fc_full") or 0.0
        mx = max(fc_td, ac_td, fc_full) or 1
        glance += ("<div class='note' style='margin:10px 0 4px'><b>Forecast v "
                   "actual</b> - the whole-shut plan first, then the days "
                   "already run against it.</div>"
                   + (_vbar("PLAN - whole shut", fmt_money(fc_full),
                            fc_full / mx, "#F26222") if fc_full else "")
                   + _vbar("Forecast to date", fmt_money(fc_td), fc_td / mx,
                           "#8B9099")
                   + _vbar("Actual to date", fmt_money(ac_td), ac_td / mx,
                           "#3FB950" if ac_td <= fc_td else "#F85149"))
    if x.get("radio_fleet"):
        rf_ = x["radio_fleet"]
        ro_ = x.get("radio_on_hire") or 0
        glance += ("<div class='note' style='margin:10px 0 4px'><b>Safety-"
                   "critical fleets</b> - out with crews v ready on the "
                   "shelf.</div>"
                   + _vbar("Radio handsets out", "{} of {}".format(ro_, rf_),
                           (ro_ / rf_) if rf_ else 0))
        if x.get("gas_fleet"):
            gf_ = x["gas_fleet"]
            go_ = x.get("gas_on_hire") or 0
            glance += _vbar("Gas monitors out", "{} of {}".format(go_, gf_),
                            (go_ / gf_) if gf_ else 0)

    # ---- charges captured (Charge Reporter register) ----------------------
    ch_rows = load_charge_feed()
    if ch_rows:
        ch_firm = sum(c["amount"] for c in ch_rows if c["firm"])
        charges_h = (
            "<h2>Charges captured this shutdown</h2>"
            "<div class='note'>Logged in the Coates charge register - damage, "
            "consumables sold, fuel and transport - each with its own advice "
            "document and evidence. Firm (approved) charges total <b>{t}</b> "
            "across {n} charge(s). These are register records; once a charge "
            "is invoiced through SiteIQ, the invoiced figure is the billing "
            "truth - the two are never added together.</div>".format(
                t=fmt_money(ch_firm), n=len(ch_rows))
            + charges_rows_tbl(ch_rows, show_company=True))
    else:
        charges_h = ""

    return {"position": position, "tiles": tiles_h, "glance": glance,
            "fva": fva, "mv": mv, "hardest": hardest, "plant": plant_h,
            "charges": charges_h, "standards": standards, "limits": limits}



def render_exec_html(x, asof, generated, source_line):
    """THE EXECUTIVE PACK, v3 (24 Jul 2026, Andrew's approved style boards):
    scorecard + timeline up front, yesterday's movement, the daily trends,
    the risk dashboard with names, accountability + contractor scores, store
    productivity, utilisation - then the proven sections (forecast v actual,
    billing confirmation, working hardest, site plant, charges, standards,
    glance). Every figure is arithmetic on real SiteIQ events; every score
    formula is printed beside its number."""
    today = x["shift_day"]
    models = x.get("models") or {}
    tx_all = x.get("tx_rows") or []
    # crew activity only: the site-plant pool's internal handovers would
    # otherwise swamp the counter story (300-odd items on 20 Jul)
    tx = []
    _seen = set()
    for t in tx_all:
        if is_site_plant_equipment(t.get("hirer", "")):
            continue
        key = (t.get("tx_id") or
               (t.get("barcode"), t.get("start"), t.get("end"),
                t.get("hirer")))
        if key in _seen:
            continue
        _seen.add(key)
        tx.append(t)
    txa = tx_analytics(tx, today)
    stats = company_return_stats(tx, models, today)
    ch_rows = load_charge_feed()
    dmg_by_co = {}
    for c in ch_rows:
        if c["category"].lower().startswith("damage"):
            dmg_by_co[c["company"]] = dmg_by_co.get(c["company"], 0) + 1
    scores = contractor_scores(stats, dmg_by_co)

    # ---- named high-care lists (risk dashboard) --------------------------
    gas_out, radio_out, milw_out, crit_out = [], [], [], []
    for m in models.values():
        for r in m["rows"]:
            days = r["days"] if r["days"] is not None else 0
            rec = (m["display"], r["hirer"],
                   esc(r["description"])
                   + (" <span style='color:{m};font-size:8.5pt'>{a}</span>"
                      .format(m=V3_STEEL, a=asset_disp(r.get("asset")))
                      if MASTER.plant_id(r.get("asset")) else "")
                   + EC.badges_html(r.get("asset"), r["description"]),
                   r["on_hire_date"], days)
            if r["storage_unit"].upper() == "GAS MONITORS":
                if days >= 1:
                    gas_out.append(rec)
            elif is_radio_handset(r):
                if days >= 1:
                    radio_out.append(rec)
            elif ("MILWAUKEE" in r["description"].upper()
                  or "MILWAUKEE" in r.get("desc0", "").upper()):
                if days >= 1:
                    milw_out.append(rec)
            if days >= 4:
                crit_out.append(rec)
    for lst in (gas_out, radio_out, milw_out, crit_out):
        lst.sort(key=lambda t: (-t[4], t[0]))

    # ---- fleet availability by equipment group (stockouts) ---------------
    stockouts = []
    try:
        path = find_newest("RENTAL_STOCK*.xlsx", "the RENTAL_STOCK export",
                           required=False)
        if path:
            _hdr, raw = sheet_rows(path, "RENTAL_STOCK",
                                   "the RENTAL_STOCK export")
            groups = {}
            for r in raw:
                su = clean_text(r.get("STORAGE_UNIT")).upper()
                if su in INFRA_STORAGE_UNITS or su == PLANT_STORAGE_UNIT:
                    continue
                key = clean_text(r.get("ITEM_DESCRIPTION"))
                if not key:
                    continue
                g = groups.setdefault(key, {"total": 0, "out": 0})
                g["total"] += 1
                if clean_text(r.get("COMPANY_NAME")):
                    g["out"] += 1
            stockouts = sorted(
                ((k, g["total"], g["out"]) for k, g in groups.items()
                 if g["out"] > 0 and g["out"] >= g["total"]),
                key=lambda t: -t[2])
    except Exception:
        stockouts = []

    st = x["stocktake"]
    st7 = st["pct"] if st else None    # 3-day sighting share (cycle truth)
    charges_n = len(ch_rows)
    damage_n = sum(1 for c in ch_rows
                   if c["category"].lower().startswith("damage"))
    stockouts_big = [s for s in stockouts if s[1] >= 3]
    hs = health_scorecard(x, txa, scores, st7, len(stockouts_big), damage_n,
                          len([g for g in gas_out if g[4] >= 1]))

    day_n = (today - dt.date(2026, 7, 20)).days + 1
    stage = ("SHUTDOWN DAY {} OF 23".format(day_n)
             if 1 <= day_n <= 23 else
             "MOBILISATION" if day_n < 1 else "DEMOB / CLOSE-OUT")

    # ================= PAGE 1 - THE SHUTDOWN SCORECARD ====================
    p1 = ("<div class='intro story'><b>{stage} &nbsp;|&nbsp; THE SHUTDOWN "
          "SCORECARD.</b> One page, one honest answer: how healthy is the "
          "K2 toolstore operation today? Every score below is arithmetic on "
          "real SiteIQ events - the formula sits beside each number, so the "
          "score can be challenged, checked and trusted.</div>").format(
        stage=esc(stage))
    ring_col = (V3_GOOD if hs["overall"] >= 90 else V3_AMBER
                if hs["overall"] >= 65 else V3_BAD)
    checklist = ""
    for name, val, why in hs["parts"]:
        if val is None:
            vtxt = "<span style='color:{m}'>n/a</span>".format(m=V3_STEEL)
            bar = ""
        else:
            vtxt = ("<b style='color:" + V3_PAGE_INK +
                    "'>{:.0f}/100</b>".format(val))
            bar = v3_scorebar(val)
        checklist += ("<div style='margin:7px 0;font-size:10pt;"
                      "color:" + V3_PAGE_BODY + "'>{tick} {n} &nbsp;{b} &nbsp;{v}"
                      "<div style='color:{m};font-size:8pt;margin-left:24px'>"
                      "{w}</div></div>").format(
            tick=v3_icon("tick", 15, V3_GOOD if (val or 0) >= 90 else
                         V3_AMBER if (val or 0) >= 65 else V3_BAD),
            n=esc(name), b=bar, v=vtxt, m=V3_STEEL, w=esc(why))
    p1 += ("<table style='width:100%;border-collapse:collapse;margin-top:6px'>"
           "<tr><td style='width:230px;vertical-align:top;text-align:center'>"
           + ring(hs["overall"] / 100.0, "Shutdown health score",
                  hs["verdict"], colour=ring_col, size=170)
           + "</td><td style='vertical-align:top;padding-left:16px'>"
           + checklist + "</td></tr></table>")
    p1 += "<h3>The shutdown at a glance - where we are</h3>"
    roster = load_roster_curve()
    gear_curve = []
    if txa:
        open_by_day = []
        for d in txa["span"]:
            n_open = sum(1 for t in tx if t.get("start") and t["start"] <= d
                         and (not t.get("end") or t["end"] > d))
            open_by_day.append((d, n_open))
        gear_curve = open_by_day
    p1 += v3_timeline(today, roster_curve=roster, gear_curve=gear_curve)
    p1 += ("<div class='note'>Phase bands with the crew curve (rostered "
           "hours from the costing model) and the gear-on-hire curve "
           "(rebuilt from every SiteIQ transaction). The dashed line is "
           "today - {d}.</div>").format(d=today.strftime("%d %b %Y"))

    # ================= PAGE 2 - YESTERDAY AT THE STORE ====================
    p2 = "<h2>Yesterday at the store</h2>"
    if txa:
        y, db, t0 = txa["yesterday"], txa["day_before"], txa["today"]
        p2 += ("<div class='note'>Issue and return events from SiteIQ's "
               "transaction stamps - {d}. Deltas compare the day before.</div>"
               ).format(d=(today - dt.timedelta(days=1)).strftime("%A %d %b"))
        p2 += v3_kpis([
            ("swap", "{:,}".format(y["iss"]), "Issued yesterday",
             v3_delta(y["iss"], db["iss"], neutral=True,
                      label="vs day before")),
            ("tick", "{:,}".format(y["ret"]), "Returned yesterday",
             v3_delta(y["ret"], db["ret"], label="vs day before")),
            ("people", "{:,}".format(y["people"]), "People served",
             v3_delta(y["people"], db["people"], neutral=True,
                      label="vs day before")),
            ("box", "{:,}".format(x["n_items"]), "On hire right now",
             v3_delta(x["n_items"],
                      (gear_curve[-2][1] if len(gear_curve) > 1 else None),
                      invert=True, label="vs yesterday")),
        ])
        p2 += v3_kpis([
            ("wrench", "{:,}".format(t0["iss"]), "Issued so far today", ""),
            ("clock", "{:,}".format(t0["ret"]), "Returned so far today", ""),
            ("people", "{:,}".format(x["n_people"]),
             "People holding gear now", ""),
            ("shield", "{:,}".format(len(crit_out)),
             "Items out 4 days or more",
             "<span style='color:{c};font-size:8pt;font-weight:700'>"
             "named on the risk page</span>".format(
                 c=V3_BAD if crit_out else V3_GOOD)),
        ])
        alerts = []
        if gas_out:
            alerts.append(("red", "{} gas monitor{} out overnight".format(
                len(gas_out), "" if len(gas_out) == 1 else "s"),
                "Daily bump test due - names on the risk page. Chase at "
                "prestart."))
        if crit_out:
            alerts.append(("amber", "{} item{} at 4+ days".format(
                len(crit_out), "" if len(crit_out) == 1 else "s"),
                "Critical returns - escalated with replacement cost "
                "visibility."))
        if stockouts_big:
            alerts.append(("amber", "{} stocked line{} at zero "
                           "availability".format(
                               len(stockouts_big),
                               "" if len(stockouts_big) == 1 else "s"),
                           "Every unit of a 3+ fleet out with live demand - "
                           "restock or cross-hire call."))
        if charges_n:
            alerts.append(("blue", "{} charge{} logged this shut".format(
                charges_n, "" if charges_n == 1 else "s"),
                "Damage and consumable charges - detail on the charges "
                "page."))
        if not alerts:
            alerts.append(("green", "No open alerts",
                           "Returns, availability and high-care gear all "
                           "inside the bars."))
        p2 += v3_alerts(alerts)
        yday = today - dt.timedelta(days=1)
        by_co = {}
        for t in tx:
            comp = t.get("company") or ""
            if not comp:
                continue
            c = by_co.setdefault(comp, {"iss": 0, "ret": 0})
            if t.get("start") == yday:
                c["iss"] += 1
            if t.get("end") == yday:
                c["ret"] += 1
        movers = [(k, v) for k, v in by_co.items() if v["iss"] or v["ret"]]
        movers.sort(key=lambda kv: -(kv[1]["iss"] + kv[1]["ret"]))
        if movers:
            p2 += "<h3>Who moved gear yesterday</h3>"
            p2 += ("<table class='data'><thead><tr><th>Company</th>"
                   "<th class='num'>Took out</th><th class='num'>Brought "
                   "back</th><th class='num'>Net</th></tr></thead><tbody>")
            for comp, v in movers[:8]:
                net = v["iss"] - v["ret"]
                p2 += ("<tr><td>{c}</td><td class='num'>{i:,}</td>"
                       "<td class='num'>{r:,}</td><td class='num' "
                       "style='color:{nc};font-weight:700'>{n:+,}</td></tr>"
                       ).format(c=esc(comp), i=v["iss"], r=v["ret"],
                                nc=V3_GOOD if net <= 0 else V3_AMBER, n=net)
            p2 += "</tbody></table>"
            p2 += ("<div class='note'>Site plant pool handovers are "
                   "excluded across this page and the trends - these are "
                   "counter movements by site crews.</div>")
    else:
        p2 += ("<div class='note'>TRANSACTIONS export unavailable - "
               "yesterday's movement reported as unavailable, not zero.</div>")

    # ================= PAGE 3 - THE DAILY TREND ===========================
    p3 = "<h2>The daily trend - issues, returns, gear out</h2>"
    if txa:
        LOOK = 10
        p3 += "<h3>Issued v returned - last {} days</h3>".format(
            min(LOOK, len(txa["iss_series"])))
        p3 += v3_bars_pair(txa["iss_series"][-LOOK:], txa["ret_series"][-LOOK:],
                           "Issued", "Returned")
        p3 += "<h3>Gear on hire - the whole shutdown</h3>"
        p3 += v3_trend([(d.strftime("%d %b"), v) for d, v in gear_curve],
                       colour=ORANGE_HEX)
        p3 += "<h3>People served per day</h3>"
        p3 += v3_trend(txa["served_series"], colour=V3_BLUE, fill=False)
        drift = (gear_curve[-1][1] - x["n_items"]) if gear_curve else 0
        p3 += ("<div class='note'>Counted from SiteIQ transaction stamps "
               "({a} - {b}), site plant pool excluded. 'Gear on hire' counts "
               "every open transaction{dr}.</div>").format(
            a=txa["window"][0].strftime("%d %b"),
            b=txa["window"][1].strftime("%d %b"),
            dr=("; it sits {:+,} against the live register's {:,} - "
                "movements SiteIQ recorded without a return stamp yet"
                .format(drift, x["n_items"]) if drift else
                " - and matches the live register exactly"))
    else:
        p3 += "<div class='note'>Transaction history unavailable.</div>"

    # ================= PAGE 4 - RISK DASHBOARD ============================
    p4 = "<h2>Risk dashboard - high-care gear, named</h2>"
    p4 += ("<div class='intro'>Gas monitors come back <b>daily</b> for bump "
           "testing, charging and inspection - one out overnight is a "
           "safety-critical follow-up, not a hire matter. Radios and "
           "Milwaukee battery gear carry the same daily-return standard. "
           "Names are listed so the right person gets the right "
           "conversation.</div>")
    def _risk_tbl(title, rows_, icon_name, bad_days):
        h_ = ("<h3>{ic} {t} - {n}</h3>").format(
            ic=v3_icon(icon_name, 16), t=esc(title),
            n=("<span style='color:{c};font-weight:700'>{n} out</span>"
               .format(c=V3_BAD if rows_ else V3_GOOD,
                       n=len(rows_)) if rows_ else
               "<span style='color:{c};font-weight:700'>all home</span>"
               .format(c=V3_GOOD)))
        if not rows_:
            # A heading on its own reads as a broken page. Say the good
            # news in words. (A. Fisher, 25 Jul 2026)
            return h_ + ("<div class='note'>Every {t} is back in the tool "
                         "store - nothing out with a crew right now.</div>"
                         ).format(t=esc(title).lower().rstrip("s"))
        grouped = {}
        for comp, hirer, desc, ohd, days in rows_:
            k = (comp, hirer, desc, days)
            g = grouped.setdefault(k, {"qty": 0, "ohd": ohd})
            g["qty"] += 1
            if ohd and (g["ohd"] is None or ohd < g["ohd"]):
                g["ohd"] = ohd
        h_ += ("<table class='data'><thead><tr><th>Company</th><th>Person"
               "</th><th>Item</th><th class='num'>Qty</th><th>Out since</th>"
               "<th class='num'>Days</th><th>Call</th></tr></thead><tbody>")
        for (comp, hirer, desc, days), g in sorted(
                grouped.items(), key=lambda kv: (-kv[0][3], kv[0][0],
                                                 kv[0][1], kv[0][2])):
            h_ += ("<tr><td>{c}</td><td>{p}</td><td>{d}</td>"
                   "<td class='num'><b>{q}</b></td><td>{o}</td>"
                   "<td class='num'><b>{dy}</b></td><td>{call}</td></tr>"
                   ).format(c=esc(comp), p=esc(hirer), d=desc,
                            q=g["qty"],
                            o=g["ohd"].strftime("%d %b") if g["ohd"] else "-",
                            dy=days,
                            call=pill("OVERDUE") if days >= bad_days
                            else "<span style='color:{m}'>due back today"
                                 "</span>".format(m=V3_STEEL))
        return h_ + "</tbody></table>"
    p4 += _risk_tbl("Gas monitors out", gas_out, "gas", 1)
    p4 += _risk_tbl("Radios out", radio_out, "radio", 2)
    p4 += _risk_tbl("Milwaukee battery gear out", milw_out, "battery", 2)
    p4 += _risk_tbl("Anything out 4 days or more", crit_out, "alert", 4)

    # ============ PAGE 5 - ACCOUNTABILITY BY COMPANY ======================
    p5 = "<h2>Accountability by company</h2>"
    p5 += ("<div class='note'>Issued and returned from the transaction "
           "history; outstanding from the live register. Return % = returned "
           "of everything issued to date.</div>")
    p5 += ("<table class='data'><thead><tr><th>Company</th>"
           "<th class='num'>Issued</th><th class='num'>Returned</th>"
           "<th class='num'>Outstanding</th><th class='num'>Return %</th>"
           "<th class='num'>Avg hold</th><th class='num'>4d+</th>"
           "</tr></thead><tbody>")
    for s in sorted(stats, key=lambda t: -t["outstanding"]):
        p5 += ("<tr><td>{c}</td><td class='num'>{i:,}</td>"
               "<td class='num'>{r:,}</td><td class='num'><b>{o:,}</b></td>"
               "<td class='num'>{rp}</td><td class='num'>{ah}</td>"
               "<td class='num'>{lt}</td></tr>").format(
            c=esc(s["company"]), i=s["issued"], r=s["returned"],
            o=s["outstanding"],
            rp=("{:.0%}".format(s["ret_pct"]) if s["ret_pct"] is not None
                else "-"),
            ah=("{:.1f}d".format(s["avg_hold"]) if s["avg_hold"] is not None
                else "-"),
            lt=s["late"] or "")
    p5 += "</tbody></table>"
    p5 += "<h3>Contractor performance - the leaderboard</h3>"
    p5 += ("<table class='data'><thead><tr><th style='width:26px'>#</th>"
           "<th>Company</th><th class='num'>Return %</th>"
           "<th class='num'>4d+</th><th class='num'>Damages</th>"
           "<th class='num'>Score</th><th></th><th>Band</th></tr></thead>"
           "<tbody>")
    for i, s in enumerate(scores, 1):
        p5 += ("<tr><td><b style='color:{oc}'>{i}</b></td><td>{c}</td>"
               "<td class='num'>{rp}</td><td class='num'>{lt}</td>"
               "<td class='num'>{dm}</td>"
               "<td class='num'><b style='font-size:12pt'>{sc:.0f}</b></td>"
               "<td>{bar}</td><td>{band}</td></tr>").format(
            oc=ORANGE_HEX, i=i, c=esc(s["company"]),
            rp=("{:.0%}".format(s["ret_pct"]) if s["ret_pct"] is not None
                else "-"),
            lt=s["late"] or "", dm=s["damages"] or "",
            sc=s["score"], bar=v3_scorebar(s["score"]),
            band=pill(s["band"]))
    p5 += "</tbody></table>"
    if scores:
        top = scores[0]
        p5 += ("<div class='note'>{tr} <b style='color:" + V3_PAGE_INK + "'>Top performer: "
               "{c}</b> - {sc:.0f}/100. Score = return performance (40) + "
               "timeliness (25, less 5 per item at 4+ days) + damage record "
               "(20, less 10 per damage charge) + hold time (15, full marks "
               "at a day-and-back). Arithmetic on real events - nothing "
               "subjective.</div>").format(
            tr=v3_icon("shield", 14, V3_GOOD), c=esc(top["company"]),
            sc=top["score"])

    # ============ PAGE 6 - STORE PRODUCTIVITY =============================
    p6 = "<h2>Store productivity - the counter story</h2>"
    if txa:
        avg_items = (len([t for t in tx if t.get("start")])
                     / max(1, sum(len(v["people"]) for v in
                                  txa["days"].values())))
        p6 += v3_kpis([
            ("people", "{:,}".format(txa["people_total"]),
             "People served this shut", ""),
            ("swap", "{:,}".format(txa["completed"]),
             "Completed hire cycles", ""),
            ("clock", ("{:.1f} days".format(txa["avg_hold"])
                       if txa["avg_hold"] is not None else "-"),
             "Average time gear stays out", ""),
            ("tick", ("{:.0%}".format(txa["same_day_pct"])
                      if txa["same_day_pct"] is not None else "-"),
             "Back same day", ""),
        ])
        p6 += v3_heat(txa["hours"])
        p6 += ("<div class='note'>SiteIQ stamps every issue and return to "
               "the minute - the grid shows when the counter runs hottest, "
               "so crews can plan around the {pk}00 rush. SiteIQ does not "
               "record queue wait times; peak-hour load is the honest "
               "pressure measure.</div>").format(
            pk="{:02d}:".format(max(txa["hours"],
                                    key=lambda h: txa["hours"][h]))
            if txa["hours"] else "morning ")
    else:
        p6 += "<div class='note'>Transaction history unavailable.</div>"

    # ============ PAGE 7 - UTILISATION & AVAILABILITY =====================
    p7 = "<h2>Utilisation and availability</h2>"
    cats = {}
    for m in models.values():
        for r in m["rows"]:
            cat = (MASTER.category(r.get("asset"))
                   or r["storage_unit"].title() or "Other")
            cats[cat] = cats.get(cat, 0) + 1
    if cats:
        pal = [ORANGE_HEX, V3_GOOD, V3_AMBER, V3_BLUE, "#B85CD6", "#5AC8FA",
               "#F97583", "#8B9099", "#D6A35C", "#6AD68F", "#C0CA33",
               "#FF8A65"]
        segs = sorted(cats.items(), key=lambda t: -t[1])
        segments = [(k, v, pal[i % len(pal)]) for i, (k, v) in
                    enumerate(segs[:11])]
        rest = sum(v for _k, v in segs[11:])
        if rest:
            segments.append(("Other", rest, V3_STEEL))
        p7 += "<h3>What's on hire, by equipment category</h3>"
        p7 += v3_donut(segments, total_label="ON HIRE")
    if stockouts:
        p7 += ("<h3>Zero availability with live demand - {} group{}</h3>"
               ).format(len(stockouts), "" if len(stockouts) == 1 else "s")
        p7 += ("<table class='data'><thead><tr><th>Equipment</th>"
               "<th class='num'>Fleet</th><th class='num'>Every one out</th>"
               "</tr></thead><tbody>")
        for k, tot, out in stockouts[:14]:
            p7 += ("<tr><td>{k}</td><td class='num'>{t:,}</td>"
                   "<td class='num'><b>{o:,}</b></td></tr>").format(
                k=esc(k), t=tot, o=out)
        p7 += "</tbody></table>"
        if len(stockouts) > 14:
            p7 += ("<div class='note'>Top 14 shown of {} groups at zero "
                   "availability.</div>").format(len(stockouts))
        p7 += ("<div class='note'>Every unit of these is out with a crew "
               "right now - the demand evidence for restock, cross-hire or "
               "next shutdown's order.</div>")

    # ---- proven sections (v2) ride behind the new story ------------------
    legacy = _exec_legacy_sections(x, asof, generated, source_line)
    body = (p1 + p2 + p3 + p4 + p5 + p6 + p7
            + legacy["fva"] + legacy["hardest"] + legacy["plant"]
            + legacy["charges"] + legacy["standards"] + legacy["glance"]
            + v3_team_page()
            + "<div class='limits'><b>Honest limits:</b> {} Movement detail "
            "now lives on the trend and yesterday pages, computed from "
            "SiteIQ's own transaction stamps.</div>".format(legacy["limits"]))
    return paged_report_doc(
        "Executive Summary", "Cement Australia K2", asof, generated,
        source_line, body, doc_title="Coates K2 Executive Summary")


def exec_eml_body(x, asof):
    return (
        "Hi,\n\n"
        "Attached is the Coates executive summary for the K2 shutdown, "
        "current as at {asof}.\n\n"
        "The short version: {n} items on hire across {c} companies, {p} "
        "people using the tool store, and the daily forecast-v-actual "
        "position inside. Every movement is double-scanned and the stocktake "
        "discipline is evidenced in the pack.\n\n"
        "Happy to walk through it any time.\n\n"
        "Thanks,\n"
        "Andrew Fisher\n"
        "Shutdown Manager - Coates\n"
        "{shutdown}\n").format(
        asof=asof.strftime("%d %b %Y %H:%M"), n=x["n_items"],
        c=x["n_companies"], p=x["n_people"], shutdown=SHUTDOWN_NAME)


def generate_exec(x, asof, generated, source_line, date_tag):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stem = "Coates_K2_Executive_Summary_{}".format(date_tag)
    html_path = os.path.join(PAGES_DIR, stem + ".html")
    pdf_path = os.path.join(PDF_DIR, stem + ".pdf")
    eml_path = os.path.join(EMAILS_DIR,
                            "Coates_K2_Executive_Summary_OUTLOOK_{}.eml".format(date_tag))
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(render_exec_html(x, asof, generated, source_line))
    pdf_ok = html_to_pdf(html_path, pdf_path)

    subject_date = dt.datetime.strptime(date_tag, "%Y-%m-%d").strftime("%d %b %Y")
    subject = "Coates K2 Executive Summary - {}".format(subject_date)
    attach = pdf_path if pdf_ok else html_path
    # the executive summary itself, pasted into the email body as images
    r_to, r_cc = recipients.resolve(KIT_DIR, "", "EXEC")
    _stem = os.path.splitext(os.path.basename(html_path))[0]
    imgs = email_images.capture_report_images(html_path, EMAILS_DIR, _stem)
    #  Same safe-send rule as every other email: pages in the body only
    #  when the whole message will arrive. The exec summary is 33 pages
    #  and landed at 10.2 MB on 28 Jul 2026 - just over the line. Too
    #  heavy = the PDF-attached form below: same report, always fits.
    if imgs and email_images.email_weight_mb(imgs) > email_images.MAX_EMAIL_MB:
        print("  Executive Summary: {} pages would make a {:.0f} MB email "
              "- over the {:.0f} MB safe-send limit, so the PDF rides the "
              "paperclip instead.".format(
                  len(imgs), email_images.email_weight_mb(imgs),
                  email_images.MAX_EMAIL_MB))
        imgs = []
    if imgs:
        msg = email_images.build_image_eml(subject, imgs, to=r_to, cc=r_cc)
        email_images.save_eml(msg, eml_path)
        cids = ["pg{}".format(i) for i in range(1, len(imgs) + 1)]
        write_draft_manifest(eml_path, subject, email_images.images_body(cids),
                             [], report="EXEC",
                             images=email_images.manifest_images(imgs))
        print("  Executive Summary: HTML{} | report in email body ({} images) "
              "| Outlook draft".format(
                  " + PDF" if pdf_ok else " (no PDF engine)", len(imgs)))
        return pdf_ok
    msg = email.message.EmailMessage(policy=email.policy.SMTP)
    msg["Subject"] = subject
    if r_to:
        msg["To"] = email_images._addr_list(r_to)
    if r_cc:
        msg["Cc"] = email_images._addr_list(r_cc)
    msg["X-Unsent"] = "1"
    text = exec_eml_body(x, asof)
    msg.set_content(text)
    # like-for-like: the full PDF design rebuilt Outlook-safe in the body
    _html = email_exec_html(x, asof, generated)
    msg.add_alternative(_html, subtype="html")
    write_draft_manifest(eml_path, str(msg["Subject"]), _html,
                         [pdf_path if pdf_ok else html_path])
    with open(attach, "rb") as f:
        data = f.read()
    if attach.endswith(".pdf"):
        msg.add_attachment(data, maintype="application", subtype="pdf",
                           filename=os.path.basename(attach))
    else:
        msg.add_attachment(data, maintype="text", subtype="html",
                           filename=os.path.basename(attach))
    with open(eml_path, "wb") as f:
        f.write(msg.as_bytes())
    print("  Executive summary: HTML{} | Outlook draft".format(
        " + PDF" if pdf_ok else " (no PDF engine)"))
    return pdf_ok


# ---------------------------------------------------------------------------
# PDF engines: WeasyPrint if installed, else Edge headless, else HTML only
# ---------------------------------------------------------------------------

_PDF_ENGINE = None  # resolved once per run: "weasyprint" | "edge:<path>" | ""


def _find_edge():
    """Edge first - see browser_engine.py for the order and why."""
    try:
        import browser_engine
        return browser_engine.find_browser()
    except Exception:
        pass
    for c in (os.path.expandvars(
                  r"%ProgramFiles(x86)%\Microsoft\Edge\Application\msedge.exe"),
              os.path.expandvars(
                  r"%ProgramFiles%\Microsoft\Edge\Application\msedge.exe"),
              shutil.which("msedge")):
        if c and os.path.isfile(c):
            return c
    return None


def resolve_pdf_engine():
    global _PDF_ENGINE
    if _PDF_ENGINE is not None:
        return _PDF_ENGINE
    try:
        import weasyprint  # noqa: F401
        _PDF_ENGINE = "weasyprint"
        return _PDF_ENGINE
    except Exception:
        pass
    edge = _find_edge()
    _PDF_ENGINE = ("edge:" + edge) if edge else ""
    return _PDF_ENGINE


def _pdf_via_edge(edge, html_path, pdf_path):
    import time
    url = "file:///" + os.path.abspath(html_path).replace("\\", "/")
    # A stale PDF from an earlier run must never masquerade as this run's
    # output - remove it first so success genuinely means Edge wrote it.
    try:
        if os.path.isfile(pdf_path):
            os.remove(pdf_path)
    except OSError:
        pass
    _pdf_via_edge._n = getattr(_pdf_via_edge, "_n", 0) + 1
    for attempt in range(3):
        # Unique profile per attempt: with the user's default profile, a
        # running Edge makes headless print exit 0 WITHOUT writing the PDF.
        #  The browser profile goes in the TEMP folder. TEMP is a Windows
        #  variable - everywhere else this was quietly filling the day's
        #  Reports folder with hundreds of profile folders (800 MB of
        #  them found on 26 Jul 2026).
        import tempfile
        profile = os.path.join(os.environ.get("TEMP") or tempfile.gettempdir(),
                               "coates_edge_pdf_{}_{}_{}".format(
                                   os.getpid(), _pdf_via_edge._n, attempt))
        cmd = [edge, "--headless", "--disable-gpu", "--no-first-run",
               "--user-data-dir={}".format(profile), "--no-pdf-header-footer",
               "--print-to-pdf={}".format(os.path.abspath(pdf_path)), url]
        if os.name != "nt":
            #  Linux build servers run as root, where the sandbox refuses
            #  to start. Never applied on a Coates Windows machine.
            cmd.insert(1, "--no-sandbox")
        try:
            subprocess.run(cmd, capture_output=True, timeout=180)
        except subprocess.TimeoutExpired as e:
            #  A stall is not a verdict - on the OneDrive laptop the first
            #  print can hang while files hydrate (29 Jul 2026: two pages
            #  lost to single stalls). A timeout gets the same fresh-profile
            #  retry as an empty PDF; only the third strike gives up.
            if attempt < 2:
                print("  Edge PDF timed out - fresh profile, trying again...")
                continue
            print("  Edge PDF conversion failed: {}".format(e))
            return False
        except Exception as e:
            print("  Edge PDF conversion failed: {}".format(e))
            return False
        if os.path.isfile(pdf_path) and os.path.getsize(pdf_path) > 0:
            return True
        time.sleep(1.5)
    return False


def html_to_pdf(html_path, pdf_path):
    """Returns True if the PDF was produced."""
    global _PDF_ENGINE
    engine = resolve_pdf_engine()
    if engine == "weasyprint":
        try:
            import weasyprint
            weasyprint.HTML(filename=html_path).write_pdf(pdf_path)
            return os.path.isfile(pdf_path) and os.path.getsize(pdf_path) > 0
        except Exception as e:
            print("  WeasyPrint failed ({}) - trying Edge instead.".format(e))
            edge = _find_edge()
            _PDF_ENGINE = ("edge:" + edge) if edge else ""
            engine = _PDF_ENGINE
    if engine.startswith("edge:"):
        return _pdf_via_edge(engine[5:], html_path, pdf_path)
    #  No engine on this machine. The HTML was just rewritten, so any PDF
    #  still sitting there is from an EARLIER run - it looks current and
    #  isn't. Say so loudly: this is how a day-old PDF ends up attached to
    #  a client email. (Seen 24 Jul 2026 - HTML 13:59, PDF 13:44.)
    try:
        if os.path.isfile(pdf_path) and \
                os.path.getmtime(pdf_path) < os.path.getmtime(html_path) - 2:
            print("  ! No PDF engine - {} is from an earlier run and is now "
                  "OUT OF DATE. Install/close Edge and run again before "
                  "sending it.".format(os.path.basename(pdf_path)))
    except OSError:
        pass
    return False


# ---------------------------------------------------------------------------
# Outlook .eml draft
# ---------------------------------------------------------------------------

def email_company_html(m, asof, generated, have_repl):
    """Company on-hire report rebuilt in Outlook's dialect - the email IS the
    report. Mirrors render_company_html section for section: partnership
    intro, client-facing tiles (replacement exposure is the only dollar
    figure a company sees), movement story, aging bands, high-care gear,
    the oldest-first outstanding list and consumables. Degrades honestly:
    missing feeds are named as data gaps, never shown as zeros."""
    display = m.get("display") or "-"
    rows = m.get("rows") or []
    aging = m.get("aging") or {}
    hc = m.get("high_care") or {}
    gas = hc.get("gas") or []
    milwaukee = hc.get("milwaukee") or []
    radios = hc.get("radios") or []
    tx_loaded = bool(m.get("tx_loaded"))
    sales_loaded = bool(m.get("sales_loaded"))
    repl_exposure = m.get("repl_exposure")
    oldest = m.get("oldest_days")
    inner = []

    # ---- partnership intro box (mirrors intro_html) ----------------------
    inner.append(_e_intro(
        'Coates is on site to keep your crew equipped - this report is how '
        'we help you keep control of it. It shows what is out in {c}\'s '
        'name, how the tool store is being used, and what each item would '
        'cost to replace. One ask that saves everyone money: <b>when gear '
        'is finished with, bring it back</b>. Returned gear is '
        'double-scanned off your record, checked and maintained, and ready '
        'for the next crew - that is Best Service &amp; Value working for '
        'you.'.format(c=esc(display))))

    # ---- tiles (mirrors tiles_html - NO hire costs client-facing) --------
    tiles = [
        (str(m.get("n_items", 0)), "Outstanding On Hire"),
        (str(len(m.get("people") or [])), "People Using Tool Store"),
        (str(m.get("n_tx_total", 0)) if (tx_loaded or sales_loaded) else "-",
         "Transactions"),
        (str(m.get("item_types", 0)), "Item Types Used"),
        ("-" if oldest is None else str(oldest), "Oldest On-Hire (days)"),
    ]
    if repl_exposure is not None:
        tiles.append((fmt_money(repl_exposure), "Replacement Cost Exposure"))
    inner.append(_e_tiles(tiles))

    # ---- how the gear is moving (mirrors returns_html) -------------------
    inner.append(_e_h2("How the gear is moving"))
    if not tx_loaded:
        inner.append(_e_note(esc(
            'Transaction history is not in this SiteIQ export yet - '
            'same-day and multi-day return counts will appear here as '
            'transactions are recorded.')))
    elif m.get("n_tx", 0) == 0:
        inner.append(_e_note(esc(
            'No completed tool store transactions are recorded against '
            '{c} yet in this export.'.format(c=display))))
    else:
        inner.append(_e_aging([
            ("Returned same day", str(m.get("same_day", 0))),
            ("Kept longer, returned", str(m.get("kept_longer", 0))),
            ("Still on hire now", str(m.get("n_items", 0)))]))
        inner.append(_e_note(esc(
            'Same-day returns keep your record clean and the gear '
            'maintained - thank your crew for those. Anything still out is '
            'listed below, oldest first.')))

    # ---- aging bands (mirrors aging_html) --------------------------------
    inner.append(_e_h2("Aging - days on hire"))
    inner.append(_e_aging([
        ("0-2 days (on track)", str(aging.get("0-2", 0))),
        ("3-4 days (at risk)", str(aging.get("3-4", 0))),
        ("5+ days (action needed)", str(aging.get("5+", 0)))]))
    unknown = aging.get("unknown", 0)
    unknown_bit = ""
    if unknown:
        unknown_bit = (' Plus {n} item{s} with no recorded on-hire date, '
                       'shown with a dash in the detail below.').format(
            n=unknown, s="" if unknown == 1 else "s")
    inner.append(_e_note(esc(
        'K2 workbook accountability thresholds for reference: overdue at 2 '
        'days, high at 3, critical at 4+. Anything 3 days or older warrants '
        'follow-up - the detail below is sorted oldest first so the chase '
        'list is at the top.' + unknown_bit)))

    # ---- high-care gear (mirrors high_care_html) -------------------------
    if not (milwaukee or radios or gas):
        inner.append(_e_h2("High-care gear"))
        inner.append(_e_note(esc(
            'No Milwaukee gear, radios or gas monitors are currently in '
            '{c}\'s name - nothing high-care to track.'.format(c=display))))
    else:
        inner.append(_e_h2("High-care gear in your name"))
        if gas:
            inner.append(_e_note(esc(
                'every monitor is bump-tested and calibrated before issue - '
                'that is life-safety equipment under the Coates Life Saving '
                'Rules (SEQ-GL-009). Return them to the tool store daily for '
                'charging and checks; never swap between crews without going '
                'through the store.'),
                "Gas monitors ({n}):".format(n=len(gas))))
        if milwaukee:
            repl_bit = (' - see the replacement cost column below'
                        if have_repl else '')
            inner.append(_e_note(esc(
                'high replacement value tooling{rb}. When it is not in a '
                'hand, it belongs in the tool store, not a crib '
                'room.'.format(rb=repl_bit)),
                "Milwaukee gear ({n}):".format(n=len(milwaukee))))
        if radios:
            inner.append(_e_note(esc(
                'charged daily and pooled site-wide - return to the store '
                'for overnight charging so the next shift picks up a '
                'working unit.'),
                "Radios ({n}):".format(n=len(radios))))

    # ---- outstanding detail, oldest first (mirrors detail_html) ----------
    inner.append(_e_h2(
        "Outstanding - on hire to you, not yet returned (oldest first)"))
    if not rows:
        inner.append(_e_note(esc(
            'Nothing outstanding - every item hired in {c}\'s name has been '
            'returned and double-scanned off your record. A clean '
            'sheet.'.format(c=display))))
    else:
        heads = ["Description", "Asset No", "Hirer", "Hirer ID",
                 "Storage Unit", "On-Hire Date", "Days"]
        if have_repl:
            heads.append("Replacement Cost")
        body = []
        for r in rows:
            cells = [
                _e_desc(r.get("description"), r.get("asset")),
                asset_disp(r.get("asset")),
                r.get("hirer") or "-",
                r.get("hirer_id") or "-",
                r.get("storage_unit") or "-",
                fmt_date(r.get("on_hire_date")),
                "-" if r.get("days") is None else str(r.get("days")),
            ]
            if have_repl:
                cells.append(fmt_money(r.get("repl")))
            body.append(cells)
        inner.append(_e_tbl(heads, body))

    # ---- consumables (mirrors consumables_html) --------------------------
    consumables = m.get("consumables") or []
    if not sales_loaded:
        # No export = unknown, not zero. Say so honestly - never present an
        # absent data feed as a real nil result.
        inner.append(_e_h2("Consumables taken"))
        inner.append(_e_note(esc(
            'No SALES_STOCK export was found for this run, so consumables '
            'movements are UNAVAILABLE - this is a data gap, not a nil '
            'result. Drop the SALES_STOCK export in the kit folder and '
            'rerun to include them.')))
    elif not consumables:
        inner.append(_e_h2("Consumables taken"))
        inner.append(_e_note(esc(
            'No consumables movements are recorded against this company in '
            'the SALES_STOCK export.')))
    else:
        draws = m.get("cons_draws", 0)
        inner.append(_e_h2("Consumables taken ({d} {dw})".format(
            d=draws, dw="draw" if draws == 1 else "draws")))
        inner.append(_e_tbl(
            ["SKU Description", "Hirer", "Hirer ID", "Total Qty Taken",
             "Last Taken"],
            [[c.get("sku") or "-", c.get("hirer") or "-",
              c.get("hirer_id") or hirer_id_for(c.get("hirer")) or "-",
              "{:g}".format(c.get("qty") or 0), fmt_date(c.get("last"))]
             for c in consumables]))
        inner.append(_e_note(esc(
            'Each draw counts as one transaction regardless of quantity '
            'taken. Quantities only - no consumables price feed is loaded, '
            'so no dollar values are shown for consumables.')))

    # ---- hero, meta, limits, frame ---------------------------------------
    n = m.get("n_items", 0)
    hero = "{n} item{s} on hire".format(n=n, s="" if n == 1 else "s")
    if oldest is not None and oldest > 0:
        hero += " &bull; oldest out {d} day{s}".format(
            d=oldest, s="" if oldest == 1 else "s")
    elif oldest == 0:
        hero += " &bull; all of it went out today"
    if repl_exposure is not None:
        hero += (" &bull; <span style=\"background:#EFFF3D;color:#101317;"
                 "padding:0 4px;border-radius:4px;\">{v}</span> "
                 "replacement exposure".format(v=fmt_money(repl_exposure)))
    meta = ('Generated: {gen} | Data as at: {asof} (RENTAL_STOCK export '
            'file time) | Author: Andrew Fisher').format(
        gen=generated.strftime("%d %b %Y %H:%M") if generated else "-",
        asof=asof.strftime("%d %b %Y %H:%M") if asof else "-")
    limits = honest_limits(have_repl, sales_loaded, internal=False,
                           tx_loaded=tx_loaded)
    return _e_frame("COATES - COMPANY ON-HIRE REPORT", display, meta, hero,
                    ''.join(inner), limits)

# ---------------------------------------------------------------------------
# Executive Summary - the whole K2 story as one stand-alone Outlook email
# Mirrors render_exec_html(): same sections, same wording, same honest
# limits. Uses the kit's Outlook-safe email engine (_e_frame and friends) -
# every style inline, tables only, no external resources.
# ---------------------------------------------------------------------------

def email_exec_html(x, asof, generated):
    """Render the executive summary as a self-contained Outlook-safe email.
    x = build_exec_model() output. Degrades honestly when tracking,
    stocktake or transaction feeds are missing - never raises, never
    invents a number."""
    pm = x.get("pm") or {}
    st = x.get("stocktake")
    tr = x.get("tracking")
    fleet = pm.get("fleet") or 0
    in_use = pm.get("in_use") or []
    idle = pm.get("idle") or []
    saving = pm.get("saving_per_day") or 0.0
    n_items = x.get("n_items") or 0
    n_cos = x.get("n_companies") or 0
    n_people = x.get("n_people") or 0
    shift_txt = fmt_date(x.get("shift_day"))
    inner = []

    # ---- the position, told straight (mirrors the PDF bits logic) ---------
    bits = []
    bits.append("Coates has {n} item{s} on hire across {c} compan{cy}, with "
                "{p} people from site crews using the tool store.".format(
                    n=n_items, s="" if n_items == 1 else "s",
                    c=n_cos, cy="y" if n_cos == 1 else "ies", p=n_people))
    if st:
        bits.append("{sv} of {tot} tool store items ({pct:.0%}) have been "
                    "physically sighted in the last 3 days by {cnt} different "
                    "counters - every item, every day, evidenced in "
                    "SiteIQ.".format(
                        sv=st.get("sighted3") or 0, tot=st.get("total") or 0,
                        pct=st.get("pct") or 0.0,
                        cnt=st.get("counters") or 0))
    if idle:
        bits.append("Idle plant worth {v} per day is flagged for off-hire - "
                    "found by our own return-to-site discipline.".format(
                        v=fmt_money(saving)))
    else:
        bits.append("No plant is sitting idle on charge - the return-to-site "
                    "convention keeps every charged item accounted for.")
    bits.append("Every movement in or out is double-scanned. Nothing leaves "
                "unrecorded; nothing comes back unchecked.")
    inner.append(_e_intro(esc(" ".join(bits))))

    # ---- six tiles --------------------------------------------------------
    inner.append(_e_tiles([
        (str(n_items), "Items On Hire"),
        (str(n_cos), "Companies"),
        (str(n_people), "People Using Store"),
        (str(x.get("n_tx") or 0)
         if (x.get("tx_loaded") or x.get("sales_loaded")) else "-",
         "Transactions"),
        ("{:.0%}".format(st.get("pct") or 0.0) if st else "-",
         "Fleet Verified (3d)"),
        (fmt_money(saving) if idle else "$0.00", "Idle Plant / Day"),
    ]))

    # ---- forecast v actual ------------------------------------------------
    if tr and tr.get("today"):
        rows = []
        t_fc = 0.0
        t_ac = 0.0
        any_actual = False
        for name, fc, ac in tr["today"]:
            fcv = fc or 0.0
            acv = ac if ac is not None else None
            t_fc += fcv
            if acv is not None:
                t_ac += acv
                any_actual = True
            var = (acv - fcv) if acv is not None else None
            rows.append([name, fmt_money(fcv),
                         fmt_money(acv) if acv is not None else "-",
                         fmt_money(var) if var is not None else "-"])
        rows.append(["Day total", fmt_money(t_fc),
                     fmt_money(t_ac) if any_actual else "-",
                     fmt_money(t_ac - t_fc) if any_actual else "-"])
        inner.append(_e_h2("Forecast v actual - {}".format(shift_txt)))
        inner.append(_e_tbl(["Category", "Forecast", "Actual", "Variance"],
                            rows))
        inner.append(_e_note(esc(
            "Shutdown to date: actual {atd} against {ftd} forecast; "
            "full-shutdown forecast {ff}. A shift day runs 6:00am to 6:00am. "
            "Manual categories (accommodation, transport, damage) show once "
            "entered.".format(atd=fmt_money(tr.get("act_to_date")),
                              ftd=fmt_money(tr.get("fc_to_date")),
                              ff=fmt_money(tr.get("fc_full"))))))
    else:
        inner.append(_e_h2("Forecast v actual"))
        inner.append(_e_note(esc(
            "Daily Tracking was not readable this run or has no row for "
            "{}.".format(shift_txt))))

    # ---- movement and returns --------------------------------------------
    inner.append(_e_h2("Movement and returns"))
    if x.get("tx_loaded") and (x.get("same_day") or x.get("kept") or n_items):
        inner.append(_e_aging([
            ("Returned same day", str(x.get("same_day") or 0)),
            ("Kept longer, returned", str(x.get("kept") or 0)),
            ("Still on hire", str(n_items))]))
        inner.append(_e_note(esc(
            "Same-day returns are the standard we promote with every crew: "
            "bring it back when it's not in use. Gear that comes back gets "
            "checked, maintained and back on the shelf for the next job.")))
    else:
        inner.append(_e_note(esc(
            "Transaction history builds as SiteIQ records it - same-day and "
            "multi-day return counts appear here. The standard stands from "
            "day one: bring it back when it's not in use.")))

    # ---- what's working hardest ------------------------------------------
    top_items = x.get("top_items") or []
    top_cons = x.get("top_cons") or []
    if top_items:
        inner.append(_e_h2("Working hardest - hire gear"))
        inner.append(_e_tbl(["Item", "On Hire Now"],
                            [[d, str(n)] for d, n in top_items]))
    if top_cons:
        inner.append(_e_h2("Working hardest - consumables"))
        inner.append(_e_tbl(["SKU", "Qty Taken"],
                            [[d, "{:g}".format(n)] for d, n in top_cons]))
        inner.append(_e_note(esc(
            "Quantities only - no consumables price feed is loaded. "
            "Utilisation evidence sizes the next order.")))
    if not top_items and not top_cons:
        inner.append(_e_h2("Working hardest"))
        inner.append(_e_note(esc(
            "Usage rankings build as hire and consumables movement is "
            "recorded - the shutdown proper starts them moving.")))

    # ---- site plant - utilisation and idle value --------------------------
    util = (float(len(in_use)) / fleet) if fleet else 0.0
    idle_txt = ("Idle-on-charge plant is worth {} per day - the off-hire "
                "conversation list.".format(fmt_money(saving)) if idle else
                "Nothing is idle on charge under the return-to-site "
                "convention.")
    inner.append(_e_h2("Site plant - utilisation and idle value"))
    inner.append(_e_note(esc(
        "{iu} of {fl} site plant items in use ({u:.0%} true utilisation). "
        "{idle} Barriers, chutes and hoppers ({inf} items) stay on site for "
        "the duration by design. The not-in-use fleet doubles as the Coates "
        "mechanic's preventative maintenance window - gear is serviced while "
        "it sits, not after it fails.".format(
            iu=len(in_use), fl=fleet, u=util, idle=idle_txt,
            inf=pm.get("infra_total") or 0))))

    # ---- the standards behind the numbers --------------------------------
    inner.append(_e_h2("The standards behind the numbers"))
    inner.append(_e_note(esc(
        "Every issue and every return goes through the tool store: scan, "
        "look. Scan, look. The scan is not complete until we have looked - "
        "so the record above is evidence, not estimate. Right first time. "
        "Every item. Every time."),
        "Double scan - two scans, two looks, one standard."))
    if st:
        chk = ("{sv} of {tot} items physically verified in the last 3 days "
               "({pct:.0%}) by {cnt} counters. ".format(
                   sv=st.get("sighted3") or 0, tot=st.get("total") or 0,
                   pct=st.get("pct") or 0.0, cnt=st.get("counters") or 0))
    else:
        chk = "Stocktake coverage reports from the SiteIQ export each day. "
    inner.append(_e_note(esc(
        "Stock takes are done daily - no exceptions - with nothing in the "
        "store going more than 30 days unscanned. " + chk +
        "Discrepancies are addressed on the spot, and the stock take doubles "
        "as our on-hire verification - anything physically in the store "
        "still showing as on hire gets corrected the same day. It is not "
        "just a counting exercise; it is another checkpoint for catching "
        "problems before a hirer does."),
        "Checking and stocktakes."))
    _rf = x.get("radio_fleet") or 0
    _rb = min(_rf, RADIO_BILLABLE_CAP)
    _rs = max(0, _rf - RADIO_BILLABLE_CAP)
    inner.append(_e_note(esc(
        "{rf} radio handsets ({rb} billable, {rs} spare{sp} held) and the "
        "{gf}-unit "
        "gas monitor fleet are pooled, charged and maintained through the "
        "store - every gas monitor bump tested and fully charged before it "
        "is issued, and every battery item issued on a full charge, never a "
        "partial. Nothing leaves the store that we would not use ourselves. "
        "Life Saving Rules (SEQ-GL-009) apply: Rule 5 - Tools and Equipment "
        "- damaged or defective gear is tagged out and removed from service, "
        "never reissued.".format(rf=_rf, rb=_rb, rs=_rs,
                                 sp="" if _rs == 1 else "s",
                                 gf=x.get("gas_fleet") or 30)),
        "Safety-critical assurance."))
    inner.append(_e_note(esc(
        "Best Service & Value is the objective; the disciplines above are "
        "how the K2 tool store delivers it, every shift."),
        "Pride in what we do."))

    # ---- hero, limits, frame ---------------------------------------------
    hero = "{n} item{ns} on hire &bull; {c} compan{cs} &bull; {pv}".format(
        n=n_items, ns="" if n_items == 1 else "s",
        cs="y" if n_cos == 1 else "ies", c=n_cos,
        pv=("{:.0%} of fleet verified in 3 days".format(st.get("pct") or 0.0)
            if st else "every movement double-scanned"))
    limits = (
        "Forecast and actual figures come from the K2 workbook Daily "
        "Tracking (read-only). Counts come from SiteIQ exports as at the "
        "data-as-at stamp; a consumables draw counts as one transaction "
        "regardless of quantity. Unpriced items are excluded from value "
        "figures, never estimated. Radio billing: 70 of 72 handsets at "
        "$12.81/day from 17 Jul 2026, 2 disclosed unbilled spares. Gas "
        "billing: 30-unit fleet at $29.75/day from 24 Jul 2026 to the "
        "forecast end, on hire or not. This pack is a point-in-time "
        "snapshot.")
    meta = ("Data as at: {a} (RENTAL_STOCK export file time) | "
            "Generated: {g} | Author: Andrew Fisher".format(
                a=asof.strftime("%d %b %Y %H:%M") if asof else "-",
                g=generated.strftime("%d %b %Y %H:%M") if generated else "-"))
    return _e_frame("COATES - K2 EXECUTIVE SUMMARY", "Cement Australia K2",
                    meta, hero, "".join(inner), limits)

def email_demob_html(models, stocktake, asof, generated):
    """Outlook-safe, stand-alone email edition of the Gear Return / Demob
    push pack - mirrors render_demob_html() section for section using the
    kit's _e_* email rendering engine. Never raises on empty/missing data."""
    today = dt.date.today()
    days_left = max(0, (SHUTDOWN_END - today).days)

    all_rows = []
    for m in (models or {}).values():
        disp = m.get("display") or m.get("company") or "-"
        for r in (m.get("rows") or []):
            all_rows.append({"description": r.get("description"),
                             "hirer": r.get("hirer"),
                             "days": r.get("days"),
                             "display": disp})
    by_co = {}
    for r in all_rows:
        co = by_co.setdefault(r["display"], {})
        h = co.setdefault(r["hirer"] or "(no name recorded)",
                          {"items": 0, "oldest": None})
        h["items"] += 1
        if r["days"] is not None and (h["oldest"] is None or
                                      r["days"] > h["oldest"]):
            h["oldest"] = r["days"]
    aged = sorted((r for r in all_rows if r["days"] is not None
                   and r["days"] >= 5), key=lambda r: -r["days"])
    people_holding = len({(r["display"], r["hirer"]) for r in all_rows})
    n_items = len(all_rows)

    asof_txt = (asof.strftime("%d %b %Y %H:%M")
                if hasattr(asof, "strftime") else "-")
    gen_txt = (generated.strftime("%d %b %Y %H:%M")
               if hasattr(generated, "strftime") else "-")

    inner = []

    # -- the winding-down intro box (same wording as the PDF pack) ----------
    intro = (
        "The job is winding down, and the next few days are about one thing: "
        "getting every bit of gear home, checked and accounted for. Nothing "
        "in this pack is an accusation - it is a checklist, so demob day is "
        "boring in the best possible way. Every return takes seconds at the "
        "tool store, is double-scanned straight off your record, and gets "
        "checked by Coates on the spot - gear you send back today is gear "
        "you are not answering for later. {n} item{s} {are} still out across "
        "{c} compan{cy} and {p} {pw}; the lists below show exactly who has "
        "what, so supervisors can walk the crib room, not the whole site."
    ).format(n=n_items, s="" if n_items == 1 else "s",
             are="is" if n_items == 1 else "are", c=len(by_co),
             cy="y" if len(by_co) == 1 else "ies", p=people_holding,
             pw="person" if people_holding == 1 else "people")
    inner.append(_e_intro(esc(intro)))

    # -- five tiles ---------------------------------------------------------
    inner.append(_e_tiles([(str(n_items), "Items Still Out"),
                           (str(len(by_co)), "Companies Holding"),
                           (str(people_holding), "People Holding"),
                           (str(len(aged)), "Aged 5+ Days"),
                           (str(days_left), "Days To Demob")]))

    # -- worth a conversation first - out 5 days or more --------------------
    if aged:
        inner.append(_e_h2("Worth a conversation first - out 5 days or more"))
        inner.append(_e_tbl(
            ["Item", "Company", "Person", "Hirer ID", "Days Out"],
            [[_e_desc(r["description"], r.get("asset")),
              r["display"], r["hirer"] or "-",
              hirer_id_for(r["hirer"]) or "-",
              str(r["days"])] for r in aged[:20]]))
        if len(aged) > 20:
            inner.append(_e_note(esc("Plus {} more in the attached PDF."
                                     .format(len(aged) - 20))))
        inner.append(_e_note(esc(
            "Long days out usually just means the gear is working hard - a "
            "quick check-in confirms it is still needed or sends it home. "
            "Either answer is a good answer.")))
    else:
        inner.append(_e_h2("Worth a conversation first"))
        inner.append(_e_note(esc(
            "Nothing has been out 5 days or more - a clean sheet going into "
            "demob.")))

    # -- who has what - by company and person -------------------------------
    inner.append(_e_h2("Who has what - by company and person"))
    rows = []
    for co in sorted(by_co, key=str.upper):
        hirers = by_co[co]
        cnt = sum(h["items"] for h in hirers.values())
        rows.append("{} - {} item{} across {} {}".format(
            co, cnt, "" if cnt == 1 else "s", len(hirers),
            "person" if len(hirers) == 1 else "people"))
        for hname in sorted(hirers, key=str.upper):
            h = hirers[hname]
            rows.append([hname, hirer_id_for(hname) or "-", str(h["items"]),
                         "-" if h["oldest"] is None else str(h["oldest"])])
    if rows:
        inner.append(_e_tbl(["Person", "Hirer ID", "Items", "Oldest (Days)"],
                            rows))
        inner.append(_e_note(esc(
            "Line-item detail per company is in the company on-hire reports "
            "- hand each person their own list.")))
    else:
        inner.append(_e_note(esc(
            "Nothing is on hire - every item is home through the tool "
            "store.")))

    # -- why this works - the process behind a clean demob ------------------
    inner.append(_e_h2("Why this works - the process behind a clean demob"))
    inner.append(_e_note(esc(
        "Every return is scanned twice at the counter, looked at both "
        "times, and comes straight off the company's record - no paperwork, "
        "no disputes, done in seconds."),
        "Double scan - two scans, two looks, one standard."))
    inner.append(_e_note(esc(
        "Every returned item is looked over by Coates as it lands: "
        "complete, clean, working. Anything damaged or defective is tagged "
        "out and removed from service on the spot - Life Saving Rule 5, "
        "Tools and Equipment (SEQ-GL-009) - so nothing dodgy goes back on a "
        "shelf or out to the next job."),
        "We check the gear itself, not just the paperwork."))
    st_bit = "Daily stocktake coverage is evidenced in SiteIQ."
    if stocktake:
        try:
            st_bit = ("{sv:,} of {tot:,} tool store items physically "
                      "verified in the last 3 days ({pct:.0%}) by {cnt} "
                      "counters - so when the final reconciliation runs, it "
                      "is confirming what we already know.").format(
                sv=stocktake["sighted3"], tot=stocktake["total"],
                pct=stocktake["pct"], cnt=stocktake["counters"])
        except Exception:
            pass
    inner.append(_e_note(esc(st_bit), "Stocktakes back it up."))
    inner.append(_e_note(esc(
        "If it's not in a hand, bring it to the tool store today. Care "
        "Deeply, One Team - let's land this shutdown with a zero-loss sheet "
        "everyone can be proud of."), "The ask."))

    # -- hero strip, honest limits, frame -----------------------------------
    hero = ("{n} item{ns} still out &bull; {c} compan{cs} &bull; demob in {d} "
            "day{ds} - if it's not in use, bring it home").format(
        n=n_items, ns="" if n_items == 1 else "s",
        c=len(by_co), cs="y" if len(by_co) == 1 else "ies",
        d=days_left, ds="" if days_left == 1 else "s")
    limits = (
        "This pack reflects SiteIQ as at {} - a point-in-time snapshot; "
        "anything returned after the extract is already off the list. Days "
        "out counts calendar days since the item went on hire. No charges "
        "appear here - this pack exists purely to make demob clean."
    ).format(asof_txt)
    meta = ("Demob: {} ({} day{} away) | Data as at: {} | Generated: {} | "
            "Author: Andrew Fisher").format(
        SHUTDOWN_END.strftime("%d %b %Y"), days_left,
        "" if days_left == 1 else "s", asof_txt, gen_txt)
    return _e_frame("COATES - GEAR RETURN - DEMOB PUSH",
                    "Let's Bring It Home", meta, hero, "".join(inner), limits)


def eml_body(m, asof):
    d = m["oldest_days"]
    if d is None:
        oldest = ""
    elif d == 0:
        oldest = " The oldest went on hire today."
    else:
        oldest = " The oldest has been out {} day{}.".format(
            d, "" if d == 1 else "s")
    item_word = "item" if m["n_items"] == 1 else "items"
    npeople = len(m["people"])
    people_bit = ""
    if npeople:
        people_bit = ", and {} {} from your crew {} using the tool store".format(
            npeople, "person" if npeople == 1 else "people",
            "is" if npeople == 1 else "are")
    tx_bit = ""
    if m["n_tx_total"]:
        tx_bit = (" We've handled {} transactions through the tool store for "
                  "you so far").format(m["n_tx_total"])
        if m["same_day"]:
            tx_bit += (" - {} of them returned same day, which keeps your "
                       "record clean and the gear maintained. Thank your crew "
                       "for those.").format(m["same_day"])
        else:
            tx_bit += "."
    return (
        "Hi,\n\n"
        "Attached is your Coates on-hire report for {company} on the K2 "
        "shutdown, current as at {asof}.\n\n"
        "The quick picture: {n} {iw} on hire in your name{people}.{oldest}"
        "{tx}\n\n"
        "The report lists everything outstanding, oldest first, with "
        "replacement values, plus the consumables your crew has drawn. One "
        "thing that helps everyone: when gear is finished with, send it back "
        "to the tool store. Every return is double-scanned off your record on "
        "the spot, and returned gear gets checked and maintained so it's "
        "right for the next job - that's what we're here for.\n\n"
        "If anything on the list has already come back, flag it and we'll "
        "square it away same day through the double-check return process.\n\n"
        "We're on site to keep your crew equipped - if you need anything, "
        "just ask.\n\n"
        "Thanks,\n"
        "Andrew Fisher\n"
        "Shutdown Manager - Coates\n"
        "{shutdown}\n"
    ).format(company=m["display"], asof=asof.strftime("%d %b %Y %H:%M"),
             n=m["n_items"], iw=item_word, people=people_bit, oldest=oldest,
             tx=tx_bit, shutdown=SHUTDOWN_NAME)


# ---------------------------------------------------------------------------
# Like-for-like email rendering: the PDF design rebuilt in Outlook's dialect
# (Word engine) - nested tables, inline styles, explicit row striping.
# ---------------------------------------------------------------------------

_F = "font-family:Calibri,Arial,sans-serif;"


def _e_h2(text):
    return ('<tr><td style="padding:16px 20px 0 20px;"><table role="presentation" '
            'width="100%" cellpadding="0" cellspacing="0"><tr><td style="{f}'
            'font-size:15px;font-weight:bold;color:{dk};'
            'border-bottom:2px solid {o};padding-bottom:3px;">{t}</td></tr>'
            '</table></td></tr>').format(f=_F, dk=COATES_DARK, o=COATES_ORANGE,
                                         t=esc(text))


def _e_note(html_text, bold_lead=""):
    lead = ('<b>{}</b> '.format(esc(bold_lead))) if bold_lead else ""
    return ('<tr><td style="padding:6px 20px 0 20px;{f}font-size:11px;'
            'color:#555555;">{l}{t}</td></tr>').format(f=_F, l=lead, t=html_text)


def _e_intro(html_text):
    return ('<tr><td style="padding:12px 20px 0 20px;"><table '
            'role="presentation" width="100%" cellpadding="0" cellspacing="0">'
            '<tr><td style="{f}background:#fdf0e8;border-left:4px solid {o};'
            'padding:10px 14px;font-size:12px;color:#222222;">{t}</td></tr>'
            '</table></td></tr>').format(f=_F, o=COATES_ORANGE, t=html_text)


class _RawHTML(str):
    """A table cell that is ALREADY HTML - _e_tbl lets it through without
    escaping. Used so the compliance badges render in the Outlook body the
    same way they do on the report. (A. Fisher, 25 Jul 2026)"""


def _e_desc(desc, asset=None):
    """Description + its compliance line, safe to drop into _e_tbl."""
    return _RawHTML(esc(desc) or "-") if not EC.has_any(asset, desc) else \
        _RawHTML((esc(desc) or "-") + EC.badges_html(asset, desc))


def _e_tbl(headers, rows):
    """Outlook-safe striped table. rows = list of cell-lists; a plain string
    entry renders as a full-width dark section header row. A cell wrapped in
    _RawHTML is passed through unescaped."""
    out = ['<tr><td style="padding:6px 20px 0 20px;">'
           '<table role="presentation" width="100%" cellpadding="0" '
           'cellspacing="0" style="border-collapse:collapse;">']
    out.append('<tr>' + ''.join(
        '<td style="{f}background:{dk};color:#ffffff;font-size:10px;'
        'font-weight:bold;text-transform:uppercase;padding:5px 7px;'
        'text-align:left;">{h}</td>'.format(f=_F, dk=COATES_DARK, h=esc(h))
        for h in headers) + '</tr>')
    stripe = False
    for r in rows:
        if isinstance(r, str):
            out.append('<tr><td colspan="{n}" style="{f}background:{dk};'
                       'color:#ffffff;font-size:11px;font-weight:bold;'
                       'padding:5px 7px;">{t}</td></tr>'.format(
                           n=len(headers), f=_F, dk=COATES_DARK, t=esc(r)))
            stripe = False
            continue
        bg = "#f5f4f2" if stripe else "#ffffff"
        stripe = not stripe
        is_total = str(r[0]).strip().lower().startswith(("day total", "total"))
        if is_total:
            bg = "#ffffff"

        def _align(i, c):
            # Right-align money/percent/pure numbers only - never dates
            # (contain /) or asset/barcode codes (mixed alphanumerics).
            if not i:
                return "left"
            s = str(c).strip()
            if s.startswith("$") or s.endswith("%") or s == "-":
                return "right"
            return ("right" if re.fullmatch(r"-?[\d,]+(\.\d+)?", s)
                    and len(s) <= 6 else "left")
        out.append('<tr>' + ''.join(
            '<td style="{f}font-size:11px;color:#222222;background:{bg};'
            'padding:4px 7px;{bd}border-bottom:1px solid #e0e0e0;'
            '{w}text-align:{a};">{c}</td>'.format(
                f=_F, bg=bg, c=(c if isinstance(c, _RawHTML) else esc(c)),
                bd="border-top:2px solid #1D1D1B;" if is_total else "",
                w="font-weight:bold;" if is_total else "",
                a=_align(i, c))
            for i, c in enumerate(r)) + '</tr>')
    out.append('</table></td></tr>')
    return ''.join(out)


def _e_tiles(tiles):
    cells = ''.join(
        '<td align="center" width="{w}%" style="background:{dk};'
        'padding:10px 4px;border-top:3px solid {o};">'
        '<span style="{f}font-size:19px;font-weight:bold;color:#ffffff;">{v}'
        '</span><br><span style="{f}font-size:8px;color:#cccccc;'
        'text-transform:uppercase;letter-spacing:0.5px;">{l}</span></td>'.format(
            w=100 // max(1, len(tiles)), dk=COATES_DARK, o=COATES_ORANGE, f=_F,
            v=('<span style="background:#EFFF3D;color:#101317;padding:0 5px;'
               'border-radius:4px;">{}</span>'.format(esc(v))
               if "replacement" in str(l).lower() else esc(v)), l=esc(l))
        for v, l in tiles)
    return ('<tr><td style="padding:10px 20px 0 20px;"><table '
            'role="presentation" width="100%" cellpadding="0" cellspacing="4">'
            '<tr>{c}</tr></table></td></tr>').format(c=cells)


def _e_aging(cells3):
    (l1, v1), (l2, v2), (l3, v3) = cells3
    def cell(label, val, bg):
        return ('<td align="center" style="{f}border:1px solid #cccccc;'
                'padding:6px 14px;"><span style="font-size:9px;'
                'text-transform:uppercase;color:#555555;">{l}</span><br>'
                '<span style="font-size:15px;font-weight:bold;'
                'background:{bg};padding:1px 10px;">{v}</span></td>'.format(
                    f=_F, l=esc(label), v=esc(val), bg=bg))
    return ('<tr><td style="padding:6px 20px 0 20px;"><table '
            'role="presentation" cellpadding="0" cellspacing="0"><tr>{a}{b}{c}'
            '</tr></table></td></tr>').format(
        a=cell(l1, v1, "#e6f4e6"), b=cell(l2, v2, "#fdf3d7"),
        c=cell(l3, v3, "#fbe3dc"))


def _e_frame(kicker, heading, sub_meta, hero, inner, limits):
    #  The report goes inside ONE Coates border (email_images.frame_html).
    #  This table used to draw its own thin dark outline, which would have
    #  read as a frame inside a frame. (A. Fisher, 25 Jul 2026)
    return email_images.frame_html(
        '<table role="presentation" width="100%" cellpadding="0" '
        'cellspacing="0" style="background:#ffffff;">'
        '<tr><td style="background:{dk};padding:14px 20px;">'
        '<table role="presentation" width="100%" cellpadding="0" '
        'cellspacing="0"><tr>'
        '<td><span style="{f}font-size:10px;letter-spacing:2px;color:{o};'
        'font-weight:bold;">{k}</span><br>'
        '<span style="{f}font-size:22px;font-weight:bold;color:#ffffff;">{h}'
        '</span><br><span style="{f}font-size:11px;color:#dddddd;">{s}</span>'
        '<br><span style="{f}font-size:9px;color:#bbbbbb;">{m}</span></td>'
        '<td align="right" valign="top" style="{f}font-size:10px;'
        'font-weight:bold;color:#ffffff;">POWERED BY <span style="color:{o};">'
        'SITEIQ</span><br><span style="font-weight:normal;color:#bbbbbb;">'
        'Equipped for anything</span></td>'
        '</tr></table></td></tr>'
        '<tr><td style="background:{o};padding:8px;{f}font-size:13px;'
        'font-weight:bold;color:#ffffff;text-align:center;">{hero}</td></tr>'
        '{inner}'
        '<tr><td style="padding:10px 20px 14px 20px;{f}font-size:9px;'
        'color:#666666;border-top:1px solid #dddddd;"><b>Honest limits:</b> '
        '{lim}</td></tr>'
        '<tr><td style="background:{dk};padding:10px 20px;{f}font-size:10px;'
        'color:#dddddd;">All returns go through the Coates tool store, handed '
        'to the stores team and double-scanned. Nothing is returned unless it '
        'is scanned.<br><span style="color:{o};font-weight:bold;font-size:9px;">'
        'Care Deeply &middot; Customer Focused &middot; Be Our Best &middot; '
        'One Team &middot; Competitive Spirit</span><br>'
        '<span style="color:#ffffff;font-size:9px;">Author: Andrew Fisher '
        '&nbsp;|&nbsp; POWERED BY <span style="color:{o};font-weight:bold;">'
        'SITEIQ</span></span></td></tr>'
        '</table>'
        .format(dk=COATES_DARK, o=COATES_ORANGE, f=_F, k=esc(kicker),
                h=esc(heading), s=esc(SHUTDOWN_NAME), m=esc(sub_meta),
                hero=hero, inner=inner, lim=esc(limits)), width=680)


def outlook_html(kicker, heading, paragraphs, stats, footer_note):
    """Outlook-safe HTML email body: nested tables, inline styles only,
    Calibri, no CSS3, no external resources - renders correctly in the
    Word-based Outlook engine and stays under clipping limits."""
    stat_cells = "".join(
        '<td align="center" style="background:#1D1D1B;padding:10px 6px;">'
        '<span style="font-family:Calibri,Arial,sans-serif;font-size:20px;'
        'font-weight:bold;color:#ffffff;">{v}</span><br>'
        '<span style="font-family:Calibri,Arial,sans-serif;font-size:9px;'
        'color:#cccccc;">{l}</span></td>'.format(v=esc(v), l=esc(l).upper())
        for v, l in stats)
    paras = "".join(
        '<p style="font-family:Calibri,Arial,sans-serif;font-size:11pt;'
        'color:#222222;margin:0 0 12px 0;">{}</p>'.format(esc(p))
        for p in paragraphs)
    return email_images.frame_html(
        '<table role="presentation" width="100%" cellpadding="0" '
        'cellspacing="0" style="background:#ffffff;">'
        '<tr><td style="background:#1D1D1B;padding:16px 20px;">'
        '<span style="font-family:Calibri,Arial,sans-serif;font-size:10px;'
        'letter-spacing:2px;color:{o};font-weight:bold;">{k}</span><br>'
        '<span style="font-family:Calibri,Arial,sans-serif;font-size:22px;'
        'font-weight:bold;color:#ffffff;">{h}</span><br>'
        '<span style="font-family:Calibri,Arial,sans-serif;font-size:10px;'
        'color:#bbbbbb;">{s} &nbsp;|&nbsp; Author: Andrew Fisher &nbsp;|&nbsp; '
        'POWERED BY <span style="color:{o};font-weight:bold;">SITEIQ</span>'
        '</span></td></tr>'
        '<tr><td style="background:{o};height:4px;font-size:0;">&nbsp;</td></tr>'
        '<tr><td style="padding:16px 20px 4px 20px;">'
        '<table role="presentation" width="100%" cellpadding="0" '
        'cellspacing="4"><tr>{stats}</tr></table></td></tr>'
        '<tr><td style="padding:12px 20px 4px 20px;">{paras}</td></tr>'
        '<tr><td style="padding:4px 20px 16px 20px;">'
        '<p style="font-family:Calibri,Arial,sans-serif;font-size:9pt;'
        'color:#666666;margin:0;">{f}</p></td></tr>'
        '<tr><td style="background:#1D1D1B;padding:10px 20px;">'
        '<span style="font-family:Calibri,Arial,sans-serif;font-size:9px;'
        'color:{o};font-weight:bold;">Care Deeply &middot; Customer Focused '
        '&middot; Be Our Best &middot; One Team &middot; Competitive Spirit'
        '</span></td></tr>'
        '</table>'
        .format(o=COATES_ORANGE, k=esc(kicker), h=esc(heading),
                s=esc(SHUTDOWN_NAME), stats=stat_cells, paras=paras,
                f=esc(footer_note)), width=640)


def write_draft_manifest(eml_path, subject, html, attachments,
                         company="", report="", images=None,
                         to="", cc=""):
    """Sidecar for MAKE_OUTLOOK_DRAFTS: lets Outlook build a NATIVE draft
    (full To address-book search) from the same subject/body/attachments.
    company/report tags let the Coates_Report_Recipients address book
    pre-fill To and CC automatically. images = [{'cid','file'}] entries
    for reports pasted into the body as inline images - the drafts script
    attaches each with its Content-ID so they render in place."""
    stem = eml_path[:-4] if eml_path.lower().endswith(".eml") else eml_path
    body_path = stem + ".body.html"
    with open(body_path, "w", encoding="utf-8") as f:
        f.write(html)
    payload = {"subject": subject,
               "company": company, "report": report,
               "body": os.path.basename(body_path),
               "attachments": [os.path.relpath(a, os.path.dirname(eml_path))
                               for a in attachments if a]}
    if to:
        payload["to"] = to
    if cc:
        payload["cc"] = cc
    if images:
        payload["images"] = images
    with open(stem + ".draft.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=1)


def write_eml(m, asof, attach_path, eml_path, date_tag, generated=None,
              have_repl=True, html_path=None):
    # Subject carries the Coates display date ('17 Jul 2026'); the ISO
    # date_tag is for filenames only.
    subject_date = dt.datetime.strptime(date_tag, "%Y-%m-%d").strftime("%d %b %Y")
    subject = ("Coates On-Hire Report - {} - K2 Shutdown - {} - "
               "{} item{} on hire").format(
        m["display"], subject_date, m["n_items"],
        "" if m["n_items"] == 1 else "s")
    # The report itself, pasted into the email body as images in order -
    # same look as the Cost Tracking Snapshot email. Falls back to the
    # rebuilt Outlook-safe body if no browser is on this machine.
    r_to, r_cc = recipients.resolve(KIT_DIR, m["display"], "COMPANY")
    if html_path:
        stem = os.path.splitext(os.path.basename(html_path))[0]
        imgs = email_images.capture_report_images(html_path, EMAILS_DIR, stem)
        #  too heavy = the PDF-attached form below. Same report, same
        #  people, always arrives.
        imgs = email_images.pages_that_fit(imgs, m["display"])
        if imgs:
            msg = email_images.build_image_eml(subject, imgs, to=r_to, cc=r_cc)
            email_images.save_eml(msg, eml_path)
            cids = ["pg{}".format(i) for i in range(1, len(imgs) + 1)]
            write_draft_manifest(eml_path, subject,
                                 email_images.images_body(cids), [],
                                 company=m["display"], report="COMPANY",
                                 images=email_images.manifest_images(imgs))
            return
    msg = email.message.EmailMessage(policy=email.policy.SMTP)
    msg["Subject"] = subject
    if r_to:
        msg["To"] = email_images._addr_list(r_to)
    if r_cc:
        msg["Cc"] = email_images._addr_list(r_cc)
    msg["X-Unsent"] = "1"   # opens in Outlook as an unsent draft
    text = eml_body(m, asof)
    msg.set_content(text)
    # like-for-like: the full PDF design rebuilt Outlook-safe in the body;
    # the attached PDF remains the print/forward artifact.
    _html = email_company_html(m, asof, generated or dt.datetime.now(), have_repl)
    msg.add_alternative(_html, subtype="html")
    write_draft_manifest(eml_path, subject, _html, [attach_path],
                         company=m["display"], report="COMPANY")
    if attach_path and os.path.isfile(attach_path):
        with open(attach_path, "rb") as f:
            data = f.read()
        if attach_path.lower().endswith(".pdf"):
            msg.add_attachment(data, maintype="application", subtype="pdf",
                               filename=os.path.basename(attach_path))
        else:
            msg.add_attachment(data, maintype="text", subtype="html",
                               filename=os.path.basename(attach_path))
    with open(eml_path, "wb") as f:
        f.write(msg.as_bytes())


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------

def generate_for_company(m, asof, generated, have_repl, source_line, date_tag):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stem = "Coates_OnHire_Report_{}_{}".format(safe_filename(m["display"]).upper(),
                                               date_tag)
    html_path = os.path.join(PAGES_DIR, stem + ".html")
    pdf_path = os.path.join(PDF_DIR, stem + ".pdf")
    eml_path = os.path.join(EMAILS_DIR,
                            "Coates_OnHire_Report_OUTLOOK_{}_{}.eml".format(
                                safe_filename(m["display"]).upper(), date_tag))

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(render_company_html(m, asof, generated, have_repl, source_line))

    pdf_ok = html_to_pdf(html_path, pdf_path)

    #  ---- NO SECOND DRAFT FOR THE SAME COMPANY -----------------------
    #  Andrew, 5 Aug 2026: "i feel there are double ups happening."
    #
    #  He was right, and it is in the folder he actually works out of.
    #  On 3 Aug, Reports\<date>\Emails held 19 Activity drafts AND 15
    #  On-Hire drafts - fifteen companies with TWO drafts each, same
    #  morning, same company, covering the same gear.
    #
    #  The Activity report is the one a contractor gets - he said so on
    #  1 Aug ("contractors i only want to see their activity report as
    #  an attachment"), and the print hub has described this one as
    #  "the old flat one - the store's own copy, not what a contractor
    #  gets any more" ever since. It just kept drafting anyway.
    #
    #  So the PAGE and the PDF still build - it is the store's own copy
    #  and 37_PICK_A_REPORT still hands it to him - but it no longer
    #  puts a second draft in front of him every morning.
    if COMPANY_ONHIRE_DRAFTS:
        write_eml(m, asof, attach_path(pdf_ok, pdf_path, html_path),
                  eml_path, date_tag, generated=generated,
                  have_repl=have_repl, html_path=html_path)
    elif os.path.isfile(eml_path):
        #  A draft from before this changed would otherwise sit there
        #  looking current. Take it with us rather than leave a stale
        #  one behind.
        try:
            os.remove(eml_path)
        except OSError:
            pass

    safe_print("  {}: {} items | HTML{}{}".format(
        m["display"], m["n_items"], " + PDF" if pdf_ok else " (no PDF engine)",
        " | Outlook draft" if COMPANY_ONHIRE_DRAFTS
        else " | no draft (the Activity report is the one that goes out)"))
    return pdf_ok


def attach_path(pdf_ok, pdf_path, html_path):
    return pdf_path if pdf_ok else html_path


def generate_repl_gap_list(models, date_tag):
    """The honest fix for missing replacement costs: every on-hire item
    with no price anywhere (master file OR schedule), named in one sheet
    with its ITEM NUMBER(s) - add a row per item number to
    K2_MASTER_EQUIPMENT_PRICING.xlsx with the cost, and the next run
    flows it through the reports AND the Excel. The kit never guesses."""
    mlist = list(models.values()) if isinstance(models, dict) else list(models)
    gaps = {}
    for m in mlist:
        for r in m["rows"]:
            if r.get("repl") is None:
                d = r["description"]
                g = gaps.setdefault(d, {"n": 0, "cos": set(), "assets": set()})
                g["n"] += 1
                g["cos"].add(m["display"])
                if r.get("asset"):
                    g["assets"].add(r["asset"])
    if not gaps:
        print("  Replacement costs: every on-hire item is priced (master "
              "file / schedule) - no gaps.")
        return
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out = os.path.join(OUTPUT_DIR,
                       "Replacement_Costs_GAP_LIST_{}.xlsx".format(date_tag))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gap List"
    ws.append(["ITEMS ON HIRE WITH NO REPLACEMENT COST YET"])
    ws.append(["Add a row per ITEM NUMBER to K2_MASTER_EQUIPMENT_PRICING"
               ".xlsx (ITEM_NUMBER + ITEM_DESCRIPTION + REPLACEMENT_COST_"
               "AUD, source 'priced by Andrew') - the next run flows it "
               "through the reports AND the Excel. One file feeds "
               "everything. Author: Andrew Fisher | POWERED BY SITEIQ"])
    ws.append([])
    ws.append(["Rental Stock Description", "Item Number(s)",
               "Items On Hire Now", "Companies Holding",
               "Unit Replacement Cost (AUD) - TO PRICE"])
    for d in sorted(gaps):
        ws.append([d, ", ".join(sorted(gaps[d]["assets"])),
                   gaps[d]["n"], ", ".join(sorted(gaps[d]["cos"]))[:80], None])
    for col, w in zip("ABCDE", (52, 26, 14, 46, 28)):
        ws.column_dimensions[col].width = w
    wb.save(out)
    total = sum(g["n"] for g in gaps.values())
    print("  Replacement costs: {} on-hire item(s) across {} description(s) "
          "have no schedule price - GAP LIST written: {}".format(
              total, len(gaps), os.path.basename(out)))


def generate_summary(models, asof, generated, source_line, date_tag,
                     radio_fleet=0, have_repl=True, gas_fleet=0):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stem = "Coates_OnHire_Summary_ALL_COMPANIES_{}".format(date_tag)
    html_path = os.path.join(PAGES_DIR, stem + ".html")
    pdf_path = os.path.join(PDF_DIR, stem + ".pdf")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(render_summary_html(models, asof, generated, source_line,
                                    radio_fleet=radio_fleet,
                                    have_repl=have_repl,
                                    gas_fleet=gas_fleet))
    pdf_ok = html_to_pdf(html_path, pdf_path)
    print("  Combined summary: HTML{}".format(" + PDF" if pdf_ok else
                                              " (no PDF engine)"))
    return pdf_ok


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

# ===========================================================================
# DAILY HIT LIST - "who do we need to go and see, and where is this stuff"
# (Andrew, 26 Jul 2026.) The store's walking list for the day: gas monitors
# and radios that must come back EVERY shift for cleaning, charging and bump
# testing; Milwaukee batteries and tools out by name; and anything that has
# sat on hire 7+ days. Replacement value rides on everything so the size of
# the ask is never a mystery. Tone is the counter's, not a courtroom's:
# bring it back, we'll check it, clean it and have it safe for the next
# person - that is the whole deal.
# ===========================================================================
def generate_hitlist_report(models, radio_fleet, gas_fleet, milw_batt_fleet,
                            milw_tool_fleet, asof, generated, source_line,
                            date_tag):
    def _rows_where(pred):
        out = []
        for m in models.values():
            if m["company"] == "COATES":
                continue          # the site-plant pool chases nobody
            for r in m["rows"]:
                if pred(r, m):
                    out.append((m, r))
        return out

    def _repl_total(pairs):
        return sum(r["repl"] for _m, r in pairs if r["repl"] is not None)

    def _tiles(cells):
        return tiles_html(cells)

    def _chase_tbl(pairs, show_days=True):
        """Person | Company | Item | Out | Replacement - oldest first, the
        walking order."""
        pairs = sorted(pairs, key=lambda p: -(p[1]["days"] or 0))
        h = ("<table class='data'><thead><tr><th>Who to see</th>"
             "<th>Company</th><th>Item</th><th class='num'>Out</th>"
             "<th class='num'>Replacement</th></tr></thead><tbody>")
        b = ""
        for m, r in pairs:
            days = r["days"]
            b += ("<tr><td style='white-space:nowrap'><b>{p}</b></td>"
                  "<td>{c}</td><td>{d}</td>"
                  "<td class='num' style='white-space:nowrap'>{o}</td>"
                  "<td class='num' style='white-space:nowrap'>{v}</td></tr>"
                  ).format(
                p=esc(activity_model._clean_name(r["hirer"]) or "-"),
                c=esc(m["display"]),
                d=EC.desc_html(r["description"], r.get("asset")),
                o=("-" if days is None else "today" if days == 0
                   else "{}d".format(days)),
                v=repl_display(r))
        return h + b + "</tbody></table>"

    def _company_rollup(pairs):
        """Company | Units | Who has them (name xN, oldest Nd) | Replacement."""
        per = {}
        for m, r in pairs:
            e = per.setdefault(m["display"], {"n": 0, "repl": 0.0,
                                              "who": {}})
            e["n"] += 1
            if r["repl"] is not None:
                e["repl"] += r["repl"]
            w = e["who"].setdefault(
                activity_model._clean_name(r["hirer"]) or "-",
                {"n": 0, "old": 0})
            w["n"] += 1
            if r["days"] is not None and r["days"] > w["old"]:
                w["old"] = r["days"]
        rows = []
        for co in sorted(per, key=lambda c: -per[c]["n"]):
            e = per[co]
            #  "(0d)" on a walking list reads silly - gear taken today
            #  just shows the name; an age tag only appears once it has one.
            who = " &middot; ".join(
                "<b>{}</b> &times;{}{}".format(
                    esc(n), w["n"],
                    " ({}d)".format(w["old"]) if w["old"] else "")
                for n, w in sorted(e["who"].items(),
                                   key=lambda kv: -kv[1]["n"]))
            rows.append((esc(co), str(e["n"]), who,
                         "<span class='repl'>{}</span>".format(
                             fmt_money(e["repl"]))))
        h = ("<table class='data'><thead><tr><th>Company</th>"
             "<th class='num'>Units</th><th>Who has them</th>"
             "<th class='num'>Replacement</th></tr></thead><tbody>")
        b = "".join("<tr><td>{}</td><td class='num'>{}</td><td>{}</td>"
                    "<td class='num' style='white-space:nowrap'>{}</td></tr>"
                    .format(*r) for r in rows)
        return h + b + "</tbody></table>"

    def _store_note(html):
        return ("<div class='intro' style='margin-top:8px'>{}</div>"
                .format(html))

    gas = _rows_where(lambda r, m: r["storage_unit"].upper() == "GAS MONITORS")
    radios = _rows_where(lambda r, m: is_radio_handset(r))
    #  Radio batteries chase with the handsets (Andrew, 28 Jul 2026:
    #  "radios and radio batteries") - a handset home without its
    #  batteries is only half a return. Covers and chargers stay off.
    radio_batt = _rows_where(
        lambda r, m: r["storage_unit"].upper() == "RADIOS"
        and not is_radio_handset(r)
        and "BATTER" in (r["description"] + " "
                         + r.get("desc0", "")).upper())
    radios_all = radios + radio_batt
    milw = _rows_where(lambda r, m: "MILWAUKEE" in r["description"].upper()
                       or "MILWAUKEE" in r.get("desc0", "").upper())
    milw_batt = [(m, r) for m, r in milw
                 if "BATTER" in r["description"].upper()
                 or "BATTER" in r.get("desc0", "").upper()]
    #  The rest of the Milwaukee fleet only makes the list once it has
    #  been out 3 days (Andrew, 28 Jul 2026: "other milwaukee lets do
    #  after 3 days. lets leave it at that rest can stay off") - a tool
    #  in daily use isn't chased; one that has gone quiet is. The old
    #  everything-out-7-days section is gone for the same reason: the
    #  list was collecting too much and the real asks drowned.
    milw_tools_all = [(m, r) for m, r in milw if (m, r) not in milw_batt]
    milw_tools = [(m, r) for m, r in milw_tools_all
                  if (r["days"] or 0) >= 3]

    body = ("<div class='intro'><b>The day's walking list.</b> Everything "
            "on these pages is out with a named person right now. The ask "
            "at the counter is always the same and it is always friendly: "
            "<b>bring it back to the store</b> - the team will clean it, "
            "check it and have it safe and ready for the next person. "
            "That is how we look after the crews, the client and the "
            "gear, all at once. The list is kept deliberately tight: "
            "<b>gas monitors, radios and radio batteries, and Milwaukee "
            "batteries every day</b>, plus <b>Milwaukee tools once "
            "they've been out 3 days</b>. Everything else stays off it "
            "so the real asks stand out.</div>")

    # ---- gas monitors ----------------------------------------------------
    body += "<h2>Gas monitors - back every day, no exceptions</h2>"
    blk = _tiles([(gas_fleet, "Fleet on site"), (len(gas), "Out with crews"),
                    (max(0, gas_fleet - len(gas)), "On the shelf ready"),
                    ("<span class='repl'>{}</span>".format(
                        fmt_money(_repl_total(gas))), "Replacement value out")])
    if gas_fleet:
        blk += hbar("Out with crews", "{} of {}".format(len(gas), gas_fleet),
                     len(gas) / gas_fleet)
        blk += hbar("Ready at the store",
                     "{} of {}".format(max(0, gas_fleet - len(gas)), gas_fleet),
                     max(0, gas_fleet - len(gas)) / gas_fleet, "#3FB950")
    body += "<div>" + blk + "</div>"
    body += ("<div class='note'>Life-safety equipment under the Coates "
             "Life Saving Rules. Every monitor comes back to the store "
             "daily to be cleaned, charged and <b>bump tested</b> before "
             "it goes out again - that is what makes it safe for the next "
             "person who clips it on.</div>")
    if gas:
        body += _chase_tbl(gas)
        body += _company_rollup(gas)
    else:
        body += "<div class='note'>Nothing out - the whole fleet is home.</div>"
    body += _store_note(
        "<b>At the store:</b> HOW-TO cards on using your gas monitor are "
        "on the counter, and the <b>bump-testing register</b> is kept in "
        "the store - every test recorded, every unit accounted for.")

    # ---- radios + radio batteries ---------------------------------------
    body += "<h2>Radios and radio batteries - same deal, back daily</h2>"
    blk = _tiles([(radio_fleet, "Handset fleet"), (len(radios), "Handsets out"),
                    (len(radio_batt), "Radio batteries out"),
                    ("<span class='repl'>{}</span>".format(
                        fmt_money(_repl_total(radios_all))), "Replacement value out")])
    if radio_fleet:
        blk += hbar("Out with crews",
                     "{} of {}".format(len(radios), radio_fleet),
                     len(radios) / radio_fleet)
        blk += hbar("Ready at the store",
                     "{} of {}".format(max(0, radio_fleet - len(radios)),
                                       radio_fleet),
                     max(0, radio_fleet - len(radios)) / radio_fleet,
                     "#3FB950")
    body += "<div>" + blk + "</div>"
    body += ("<div class='note'>Handsets <b>and their batteries</b> come "
             "back to the store overnight for cleaning, checks and "
             "charging - the next shift picks up a working, charged unit "
             "every time. A handset home without its batteries is only "
             "half a return.</div>")
    if radios_all:
        body += _chase_tbl(radios_all)
        body += _company_rollup(radios_all)
    else:
        body += "<div class='note'>Nothing out - the whole fleet is home.</div>"
    body += _store_note(
        "<b>At the store:</b> radio HOW-TO cards and <b>channel cards</b> "
        "are available at the counter - who is on what channel, one card, "
        "no guesswork.")

    # ---- milwaukee batteries --------------------------------------------
    body += "<h2>Milwaukee batteries - first, always</h2>"
    blk = _tiles([(milw_batt_fleet, "On site total"),
                    (len(milw_batt), "Out with crews"),
                    (max(0, milw_batt_fleet - len(milw_batt)), "On charge ready"),
                    ("<span class='repl'>{}</span>".format(
                        fmt_money(_repl_total(milw_batt))), "Replacement value out")])
    if milw_batt_fleet:
        blk += hbar("Out with crews",
                     "{} of {}".format(len(milw_batt), milw_batt_fleet),
                     len(milw_batt) / milw_batt_fleet)
        blk += hbar("On charge ready",
                     "{} of {}".format(max(0, milw_batt_fleet - len(milw_batt)),
                                       milw_batt_fleet),
                     max(0, milw_batt_fleet - len(milw_batt)) / milw_batt_fleet,
                     "#3FB950")
    body += "<div>" + blk + "</div>"
    if milw_batt:
        body += _chase_tbl(milw_batt)
        body += _company_rollup(milw_batt)
    else:
        body += "<div class='note'>No batteries out under a name right now.</div>"

    # ---- milwaukee tools - only once they've been out 3 days -------------
    body += "<h2>Milwaukee battery-powered tools - out 3 days or more</h2>"
    blk = _tiles([(milw_tool_fleet, "On site total"),
                    (len(milw_tools_all), "Out with crews"),
                    (len(milw_tools), "Out 3+ days - on the list"),
                    ("<span class='repl'>{}</span>".format(
                        fmt_money(_repl_total(milw_tools))), "Replacement value on the list")])
    if milw_tool_fleet:
        blk += hbar("Out with crews",
                     "{} of {}".format(len(milw_tools_all), milw_tool_fleet),
                     len(milw_tools_all) / milw_tool_fleet)
        blk += hbar("In the store ready",
                     "{} of {}".format(max(0, milw_tool_fleet - len(milw_tools_all)),
                                       milw_tool_fleet),
                     max(0, milw_tool_fleet - len(milw_tools_all)) / milw_tool_fleet,
                     "#3FB950")
    body += "<div>" + blk + "</div>"
    body += ("<div class='note'>A tool in daily use is doing its job - it "
             "isn't chased. Once one has sat out <b>3 days</b> it makes "
             "this list for a friendly check-in.</div>")
    if milw_tools:
        body += _chase_tbl(milw_tools)
        body += _company_rollup(milw_tools)
    else:
        body += ("<div class='note'>Nothing out 3 days or more - the "
                 "Milwaukee fleet is turning over nicely.</div>")
    body += _store_note(
        "<b>At the store:</b> high replacement value tooling - when it is "
        "not in a hand it belongs in the tool store, not a crib room. "
        "Care cards available at the counter.")

    # ---- the money picture, one glance -------------------------------
    cats = [("Gas monitors", _repl_total(gas), "#3FB950"),
            ("Radios + batteries", _repl_total(radios_all), "#4C8DD6"),
            ("Milwaukee batteries", _repl_total(milw_batt), "#F2B01E"),
            ("Milwaukee tools 3d+", _repl_total(milw_tools), "#F26222")]
    mx = max([v for _l, v, _c in cats] or [1]) or 1
    body += ("<h2>Replacement value out right now - by category</h2>"
             "<div class='note'>The size of what is riding on today's "
             "walking list, category by category.</div>")
    for lab, v, colr in cats:
        body += hbar(lab, fmt_money(v), v / mx, colr)
    body += EC.store_story_html()

    unpriced = sum(1 for pairs in (gas, radios_all, milw_batt, milw_tools)
                   for _m, r in pairs
                   if r["repl"] is None and not r.get("cust_owned"))
    limits = ("Counts come from today's RENTAL_STOCK export; fleet totals "
              "count every unit in the export, on hire or not. Replacement "
              "values come from the K2 replacement cost schedule; {} "
              "item(s) on these lists show TBC and are excluded from "
              "value totals rather than guessed. Gas monitor and radio "
              "return expectations are store practice, not a charge."
              .format(unpriced))

    emit_report(
        "Coates_K2_Daily_Hit_List_{}".format(date_tag),
        "Daily Hit List", "Coates Tool Store - who we need to see today",
        body,
        "{g} gas &bull; {r} radios + batteries &bull; {mb} Milwaukee "
        "batteries &bull; {mt} Milwaukee tools 3d+".format(
            g=len(gas), r=len(radios_all), mb=len(milw_batt),
            mt=len(milw_tools)),
        [_e_note(esc(
            "{} gas monitors, {} radios and radio batteries, and {} "
            "Milwaukee batteries are out with named people right now; {} "
            "Milwaukee tool(s) have been out 3 days or more. The walking "
            "list with names, companies and replacement values is "
            "attached.".format(
                len(gas), len(radios_all), len(milw_batt),
                len(milw_tools))), "Today's hit list.")],
        limits,
        "Coates K2 - Daily Hit List - {}".format(
            generated.strftime("%d %B %Y") if generated else date_tag),
        asof, generated, source_line, report_tag="HITLIST")
    print("  Daily Hit List: gas {} | radios+batt {} | Milwaukee batt {} "
          "| tools 3d+ {}".format(
              len(gas), len(radios_all), len(milw_batt), len(milw_tools)))


def run(selected_company=None, do_all=False, do_plant=False, do_exec=False,
        do_audit=False, do_demob=False, do_cons=False, do_daily=False,
        do_store=False, do_requests=False, do_stocktake=False,
        do_mygear_mail_f=False,
        do_safety=False, do_returns=False, do_weekly=False,
        do_rightsize=False, do_closeout=False, do_lookup=False,
        do_stocktake_team=False, do_store_health=False,
        do_activity=False, do_hitlist=False):
    print("=" * 62)
    print(" COATES | K2 COMPANY ON-HIRE REPORT")
    print(" {} ".format(SHUTDOWN_NAME))
    print(" Author: Andrew Fisher | POWERED BY SITEIQ")
    print("=" * 62)

    global MASTER
    MASTER = master_equipment.load(KIT_DIR)
    # Same one file drives the compliance badges - bind it once here and
    # every report picks it up.
    EC.bind(MASTER)
    (items, rental_path, asof, radio_fleet, plant, infra, gas_fleet,
     radio_on_hire, gas_on_hire, milw_batt_fleet,
     milw_tool_fleet) = load_rental()
    rates, rates_path = load_rates()
    repl, repl_path = load_replacement()
    if repl_path is None:
        print("NOTE: no Tooling_Replacement_Costs_Gladstone_K2 schedule found -")
        print("      replacement costs will be marked unavailable on the reports.")
    sales, sales_path = load_sales()
    if not sales_path:
        print("NOTE: no SALES_STOCK export found - consumables section will")
        print("      show as unavailable on the reports.")
    tx, tx_path = load_transactions()
    if not tx_path:
        print("NOTE: no TRANSACTIONS export found - transaction counts will")
        print("      show as unavailable on the reports.")

    # Per-person IDs: EXTERNAL_ID from the ON_HIRE export is the ONLY
    # person identifier we ever use (Andrew, 19 Jul). Exact item barcode
    # match first, hirer name second, honest '-' when neither lands.
    onhire = load_on_hire()
    if onhire:
        HIRER_ID_BY_NAME.update(onhire["by_name"])
        oh_bc = onhire["by_barcode"]
        for it in items:
            it["hirer_id"] = (oh_bc.get(it["barcode"].upper())
                              or hirer_id_for(it["hirer"]))
        for p in plant:
            p["hirer_id"] = (oh_bc.get(p["barcode"].upper())
                             or hirer_id_for(p["hirer"]))
        print("Per-person IDs: {} people matched from {}".format(
            len(onhire["people"]), os.path.basename(onhire["path"])))
    else:
        print("NOTE: no ON_HIRE export found - the Hirer ID column will")
        print("      show '-' and the gear lookup page cannot be built.")

    today = dt.date.today()
    generated = dt.datetime.now()
    date_tag = today.strftime("%Y-%m-%d")
    source_line = "Source: {} | Extract (file time): {}".format(
        os.path.basename(rental_path), asof.strftime("%d %b %Y %H:%M"))
    if onhire:
        source_line += " | Hirer IDs: {} (EXTERNAL_ID)".format(
            os.path.basename(onhire["path"]))
    else:
        source_line += (" | ON_HIRE export not found - Hirer ID "
                        "unavailable this run")

    companies = sorted({i["company"] for i in items})
    if not companies:
        raise KitError(
            "No companies have items on hire in {}.\n"
            "(An item counts as on hire when COMPANY_NAME is filled in.)".format(
                os.path.basename(rental_path)))

    # barcode -> company map for transaction attribution (K2 workbook basis)
    barcode_company = {}
    for it in items:
        bc = it["barcode"].upper()
        if bc and bc not in barcode_company:
            barcode_company[bc] = it["company"]

    models = {c: build_company_model(c, items, rates, repl, sales, today,
                                     sales_loaded=sales_path is not None,
                                     tx=tx, tx_loaded=tx_path is not None,
                                     barcode_company=barcode_company)
              for c in companies}
    have_repl = bool(repl) or MASTER.n_priced > 0

    engine = resolve_pdf_engine()
    if engine == "weasyprint":
        print("PDF engine: WeasyPrint")
    elif engine.startswith("edge:"):
        print("PDF engine: Microsoft Edge (headless)")
    else:
        print("NOTE: no PDF engine found (WeasyPrint not installed, Edge not")
        print("      found). Reports will be produced as HTML only - open the")
        print("      HTML in a browser and print to PDF from there if needed.")

    def do_company(c):
        generate_for_company(models[c], asof, generated, have_repl,
                             source_line, date_tag)

    def do_plant_report():
        cats, cats_path = load_plant_categories()
        if cats_path:
            print("  plant categories from: {}".format(os.path.basename(cats_path)))
        pm = build_plant_model(plant, rates, cats, today, infra=infra)
        generate_plant(pm, asof, generated, source_line, date_tag, bool(cats))

    def do_exec_report():
        # shift day runs 05:00 -> 05:00 next morning
        now = dt.datetime.now()
        shift_day = (now.date() - dt.timedelta(days=1)) if now.hour < 5 else now.date()
        cats, _ = load_plant_categories()
        pm = build_plant_model(plant, rates, cats, today, infra=infra)
        stocktake, st_path = load_stocktake()
        if not st_path:
            print("NOTE: no STOCKTAKE export found - coverage stats unavailable.")
        tracking, trk_path = load_daily_tracking_summary(shift_day)
        confirmed, ds_path = load_daily_summary(shift_day)
        if ds_path:
            print("  SiteIQ confirmed billing from: {}".format(os.path.basename(ds_path)))
        gear_dispatch, gd_path = load_gear_dispatch_summary()
        if gd_path:
            print("  Radio/gas dispatch history from: {}".format(os.path.basename(gd_path)))
        x = build_exec_model(models, pm, tx, sales, stocktake, tracking,
                             radio_fleet, shift_day,
                             tx_path is not None, sales_path is not None,
                             gas_fleet=gas_fleet, radio_on_hire=radio_on_hire,
                             gas_on_hire=gas_on_hire, confirmed=confirmed,
                             gear_dispatch=gear_dispatch)
        generate_exec(x, asof, generated, source_line, date_tag)

    def do_audit_report():
        cats, cats_path = load_plant_categories()
        if cats_path:
            print("  plant categories from: {}".format(os.path.basename(cats_path)))
        pm = build_plant_model(plant, rates, cats, today, infra=infra)
        generate_audit(pm, asof, generated, source_line, date_tag)

    def do_demob_report():
        stocktake, _ = load_stocktake()
        generate_demob(models, stocktake, asof, generated, source_line,
                       date_tag)

    def do_cons_report():
        generate_consumables(sales, sales_path is not None, asof, generated,
                             source_line, date_tag)

    def do_store_report():
        generate_store_report(sales, sales_path is not None, asof, generated,
                              source_line, date_tag)

    def do_requests_report():
        generate_requests_report(asof, generated, source_line, date_tag)

    def do_stocktake_report():
        generate_stocktake(asof, generated, source_line, date_tag)

    def do_mygear_mail():
        generate_mygear_mail(models, asof, generated, source_line, date_tag)

    def _pm():
        cats, _ = load_plant_categories()
        return build_plant_model(plant, rates, cats, today, infra=infra)

    def do_daily_report():
        stocktake, _ = load_stocktake()
        trk, _ = load_tracking_rows()
        generate_daily_brief(models, _pm(), trk, stocktake, asof, generated,
                             source_line, date_tag)

    def do_safety_report():
        stocktake, _ = load_stocktake()
        generate_safety_report(models, radio_fleet, gas_fleet, stocktake,
                               asof, generated, source_line, date_tag)

    def do_activity_report():
        """Level 1 company summary + Level 2 hirer pages, one pack per
        company. Site Plant Equipment (Coates' own gear) is excluded -
        this report is about what contractors hold."""
        am = activity_model.build(models, today=dt.date.today(),
                                  hirer_id_for=hirer_id_for,
                                  exclude_company=["COATES"])
        bad = activity_model.reconcile(am)
        if bad:
            print("  Activity report NOT written - the totals do not tie. "
                  "Nothing goes to a client that does not add up.")
            return
        p0, p1 = am["period"]
        _no_activity = []
        for co in am["companies"]:
            if not co["hirers"]:
                #  NOT SILENTLY. Andrew, 5 Aug 2026: "any company that
                #  has nothing onhire should still get a report."
                #  A company with nothing on hire DOES get one - that
                #  is the all-clear story in activity_report.py. This
                #  skip is narrower and different: nobody from this
                #  company touched the counter in the period at all, so
                #  there is no activity to write a story from. That is
                #  a fair reason to skip and a terrible reason to say
                #  nothing, because from the outside it looks identical
                #  to a company that was forgotten.
                _no_activity.append(co["display"])
                continue
            body = activity_report.render(co, am["period"])
            slug = re.sub(r"[^A-Za-z0-9]+", "_", co["display"]).strip("_").upper()
            c = co["cards"]
            emit_report(
                "Coates_K2_Activity_{}_{}".format(slug, date_tag),
                "Equipment Activity & Accountability",
                co["display"], body,
                "{h} hirers &bull; {i} issued &bull; {r} returned &bull; "
                "{s} still on hire".format(h=co["n_hirers"], i=c["issued"],
                                           r=c["returned"], s=c["still"]),
                [_e_note(esc(
                    "{} hirers, {} items issued and {} returned this period. "
                    "{} items are still on hire. Every hirer has their own "
                    "page in the attached pack.".format(
                        co["n_hirers"], c["issued"], c["returned"],
                        c["still"])), "The position.")],
                "Activity from the SiteIQ transaction export for the period "
                "shown; on-hire position from the rental register. "
                "Consumables are usage data and are never counted as "
                "equipment on hire. No consumable pricing exists in any "
                "export, so units and product types are shown without a "
                "dollar value.",
                "Coates K2 - {} - Equipment Activity & Accountability".format(
                    co["display"]),
                asof, generated, source_line, report_tag="ACTIVITY",
                pdf_attach_only=True, cc_extra=_oversight_cc())
        if _no_activity:
            print("  Activity & Accountability: {} company/companies had "
                  "NOBODY at the counter".format(len(_no_activity)))
            print("     in this period, so there is no activity to write a "
                  "story from and no")
            print("     report was built for them: {}".format(
                ", ".join(sorted(_no_activity))))
            print("     They are not forgotten - widen the period or check "
                  "they are still on site.")
        print("  Activity & Accountability: {} company pack(s), {} hirer "
              "pages, period {} to {}".format(
                  len(am["companies"]),
                  sum(len(c["hirers"]) for c in am["companies"]),
                  p0.strftime("%d %b") if p0 else "?",
                  p1.strftime("%d %b %Y") if p1 else "?"))

    def do_hitlist_report():
        generate_hitlist_report(models, radio_fleet, gas_fleet,
                                milw_batt_fleet, milw_tool_fleet,
                                asof, generated, source_line, date_tag)

    def do_returns_report():
        generate_returns_report(models, tx, tx_path is not None, asof,
                                generated, source_line, date_tag)

    def do_weekly_report():
        trk, _ = load_tracking_rows()
        generate_weekly_rollup(models, trk, load_stocktake_daily(), asof,
                               generated, source_line, date_tag)

    def do_rightsize_report():
        generate_rightsize_report(_pm(), asof, generated, source_line,
                                  date_tag)

    def do_closeout_report():
        stocktake, _ = load_stocktake()
        trk, _ = load_tracking_rows()
        generate_closeout_report(models, _pm(), tx, sales, stocktake, trk,
                                 asof, generated, source_line, date_tag)

    def do_lookup_page():
        generate_lookup(models, onhire, asof, generated, date_tag)

    # Flags are ADDITIVE: --exec --plant --audit --demob and --company/--all
    # can be combined in one run - every requested report is generated
    # (previously each flag early-returned and silently dropped the rest).
    ran_flag = False
    if do_exec:
        do_exec_report()
        ran_flag = True
    if do_audit:
        do_audit_report()
        ran_flag = True
    if do_demob:
        do_demob_report()
        ran_flag = True
    if do_plant:
        do_plant_report()
        ran_flag = True
    if do_cons:
        do_cons_report()
        ran_flag = True
    if do_store:
        do_store_report()
        ran_flag = True
    if do_requests:
        do_requests_report()
        ran_flag = True
    if do_stocktake:
        do_stocktake_report()
        ran_flag = True
    if do_stocktake_team:
        generate_stocktake_team(asof, generated, source_line, date_tag)
        ran_flag = True
    if do_store_health:
        generate_store_health(sales, sales_path is not None, asof,
                              generated, source_line, date_tag)
        ran_flag = True
    if do_mygear_mail_f:
        do_mygear_mail()
        ran_flag = True
    if do_daily:
        do_daily_report()
        ran_flag = True
    if do_safety:
        do_safety_report()
        ran_flag = True
    if do_returns:
        do_returns_report()
    if do_hitlist:
        do_hitlist_report()
        ran_flag = True
    if do_activity:
        do_activity_report()
        ran_flag = True
    if do_weekly:
        do_weekly_report()
        ran_flag = True
    if do_rightsize:
        do_rightsize_report()
        ran_flag = True
    if do_closeout:
        do_closeout_report()
        ran_flag = True
    if do_lookup:
        do_lookup_page()
        ran_flag = True
    if ran_flag and not (do_all or selected_company):
        print("\nDone. Outputs are in:\n  {}".format(OUTPUT_DIR))
        # finish with REAL Outlook drafts - report pages in the compose
        # body, addressed from the recipients book, sitting in Drafts
        email_images.make_native_drafts(KIT_DIR)
        return

    unknown = sorted(display_company(c) for c in companies
                     if not is_known_company(c))
    if unknown:
        print("NOTE: not on the K2 contractor register (check spelling or")
        print("      induction): " + ", ".join(unknown))

    if do_all:
        print("\nGenerating reports for all {} companies...".format(len(companies)))
        for c in sorted(companies, key=lambda c: display_company(c).upper()):
            do_company(c)
        generate_summary(list(models.values()), asof, generated,
                         source_line, date_tag, radio_fleet=radio_fleet,
                         have_repl=have_repl, gas_fleet=gas_fleet)
        if have_repl:
            generate_repl_gap_list(models, date_tag)
        print("\nDone. Outputs are in:\n  {}".format(OUTPUT_DIR))
        # finish with REAL Outlook drafts - report pages in the compose
        # body, addressed from the recipients book, sitting in Drafts
        email_images.make_native_drafts(KIT_DIR)
        return

    if selected_company:
        want = canonical_company(selected_company)
        match = None
        if want in models:
            match = want
        else:
            partial = [c for c in companies if want in c]
            if len(partial) == 1:
                match = partial[0]
            elif len(partial) > 1:
                raise KitError(
                    "'{}' matches more than one company: {}.\n"
                    "Be more specific.".format(
                        selected_company,
                        ", ".join(display_company(p) for p in partial)))
        if not match:
            raise KitError(
                "No company matching '{}' has items on hire.\n"
                "Companies with gear on hire right now: {}".format(
                    selected_company,
                    ", ".join(display_company(c) for c in companies)))
        do_company(match)
        print("\nDone. Outputs are in:\n  {}".format(OUTPUT_DIR))
        # finish with REAL Outlook drafts - report pages in the compose
        # body, addressed from the recipients book, sitting in Drafts
        email_images.make_native_drafts(KIT_DIR)
        return

    # ---------------- interactive menu (alphabetical - the standard) -------
    order = sorted(companies, key=lambda c: display_company(c).upper())
    while True:
        print("\nCompanies with items on hire (as at {}):".format(
            asof.strftime("%d %b %Y %H:%M")))
        for i, c in enumerate(order, 1):
            n = models[c]["n_items"]
            safe_print("  {:2d}. {} ({} item{})".format(i, display_company(c),
                                                        n, "" if n == 1 else "s"))
        print("   A. All companies (plus combined summary)")
        print("   P. Site Plant report (in use v available-and-charged)")
        print("   E. Executive Summary (the whole K2 story)")
        print("   L. Plant audit list (printable idle checklist)")
        print("   D. Gear return / demob push pack (Let's Bring It Home)")
        print("   C. Consumables report (site-wide draws, qty only)")
        print("   T. Cement Store stock & reorder (for the store team)")
        print("   N. Consumable requests log (what crews asked for)")
        print("   K. Stocktake scorecard + daily run sheet (print & walk)")
        print("   M. My Gear mail - each person's gear list emailed to them")
        print("   B. Daily one-pager (morning brief)")
        print("   S. Safety & high-care assurance pack")
        print("   R. Returns performance (same-day heroes)")
        print("   W. Weekly executive roll-up")
        print("   U. Plant utilisation & right-size")
        print("   F. End-of-shut close-out pack")
        print("   G. Gear lookup page (QR at the window - scan, enter ID)")
        print("   H. Daily hit list (gas, radios + batteries, Milwaukee; "
              "tools at 3d+)")
        print("   Q. Quit")
        try:
            choice = input("Pick a number, A, P, E, L, D or Q: ").strip().upper()
        except EOFError:
            # No console input available (piped / headless run) - exit
            # cleanly rather than crash. Use --all or --company instead.
            print("\nNo console input available. Use --all or --company")
            print('(e.g. python build_company_onhire_report.py --all).')
            return
        if choice == "Q":
            return
        if choice == "P":
            do_plant_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "E":
            do_exec_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "L":
            do_audit_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "D":
            do_demob_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "C":
            do_cons_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "T":
            do_store_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "N":
            do_requests_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "K":
            do_stocktake_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "M":
            do_mygear_mail()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "B":
            do_daily_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "S":
            do_safety_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "H":
            do_hitlist_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "R":
            do_returns_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "W":
            do_weekly_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "U":
            do_rightsize_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "F":
            do_closeout_report()
            print("Outputs are in: {}".format(OUTPUT_DIR))
            continue
        if choice == "G":
            do_lookup_page()
            continue
        if choice == "A":
            for c in order:
                do_company(c)
            generate_summary(list(models.values()), asof, generated,
                             source_line, date_tag, radio_fleet=radio_fleet,
                             have_repl=have_repl)
            print("\nDone. Outputs are in:\n  {}".format(OUTPUT_DIR))
            # finish with REAL Outlook drafts - report pages in the compose
            # body, addressed from the recipients book, sitting in Drafts
            email_images.make_native_drafts(KIT_DIR)
            continue
        try:
            idx = int(choice)
            if not 1 <= idx <= len(order):
                raise ValueError
        except ValueError:
            print("Didn't catch that - type a number from the list, A or Q.")
            continue
        do_company(order[idx - 1])
        print("Outputs are in: {}".format(OUTPUT_DIR))


def main():
    ap = argparse.ArgumentParser(
        description="K2 Company On-Hire Report kit (Coates | Andrew Fisher)")
    ap.add_argument("--company", help='Company name, e.g. --company "BELTEC"')
    ap.add_argument("--all", action="store_true",
                    help="Generate every company plus a combined summary")
    ap.add_argument("--plant", action="store_true",
                    help="Generate the Site Plant report (in use v "
                         "available-and-charged)")
    ap.add_argument("--exec", dest="exec_", action="store_true",
                    help="Generate the Executive Summary pack")
    ap.add_argument("--audit", action="store_true",
                    help="Generate the printable plant audit list")
    ap.add_argument("--demob", action="store_true",
                    help="Generate the gear return / demob push pack")
    ap.add_argument("--store", action="store_true",
                    help="Cement Store stock & reorder report")
    ap.add_argument("--requests", action="store_true",
                    help="Consumable requests log report")
    ap.add_argument("--stocktake", action="store_true",
                    help="Stocktake scorecard + daily run sheet")
    ap.add_argument("--stocktake-team", action="store_true",
                    dest="stocktake_team",
                    help="Stores stocktake TEAM report - the people "
                         "turning the wheel")
    ap.add_argument("--store-health", action="store_true",
                    dest="store_health",
                    help="Store consumables watch - cover, checks and "
                         "the watch list")
    ap.add_argument("--mygear-mail", action="store_true", dest="mygear_mail",
                    help="Email each person their own gear scorecard")
    ap.add_argument("--consumables", action="store_true",
                    help="Generate the site-wide consumables report")
    ap.add_argument("--daily", action="store_true",
                    help="Daily one-pager morning brief")
    ap.add_argument("--safety", action="store_true",
                    help="Safety & high-care assurance pack")
    ap.add_argument("--returns", action="store_true",
                    help="Returns performance report")
    ap.add_argument("--activity", action="store_true",
                    help="Equipment Activity & Accountability pack - company "
                         "summary plus one page per hirer")
    ap.add_argument("--weekly", action="store_true",
                    help="Weekly executive roll-up")
    ap.add_argument("--rightsize", action="store_true",
                    help="Plant utilisation & right-size report")
    ap.add_argument("--closeout", action="store_true",
                    help="End-of-shut close-out pack")
    ap.add_argument("--lookup", action="store_true",
                    help="Gear lookup page (QR at the window)")
    ap.add_argument("--hitlist", action="store_true",
                    help="Daily hit list - gas, radios + batteries, "
                         "Milwaukee batteries; other Milwaukee at 3d+")
    ap.add_argument("--email-only", dest="email_only", action="store_true",
                    help="Produce only the emails - no Pages\\ HTML, no "
                         "print PDFs, no working files left behind")
    args = ap.parse_args()
    if args.email_only:
        globals()["EMAIL_ONLY"] = True
        print("EMAIL ONLY: producing the emails, skipping the prep files.")
    try:
        run(selected_company=args.company, do_all=args.all,
            do_plant=args.plant, do_exec=args.exec_, do_audit=args.audit,
            do_demob=args.demob, do_cons=args.consumables,
            do_store=args.store, do_requests=args.requests,
            do_stocktake=args.stocktake, do_mygear_mail_f=args.mygear_mail,
            do_daily=args.daily, do_safety=args.safety,
            do_returns=args.returns, do_weekly=args.weekly,
            do_rightsize=args.rightsize,
            do_closeout=args.closeout, do_lookup=args.lookup,
            do_stocktake_team=args.stocktake_team,
            do_store_health=args.store_health,
            do_activity=args.activity, do_hitlist=args.hitlist)
    except KitError as e:
        print("\nPROBLEM:\n{}".format(e))
        sys.exit(1)
    except KeyboardInterrupt:
        print("\nStopped.")
        sys.exit(1)
    except Exception as e:
        # A malformed export should never dump a raw traceback at 05:00 in a
        # toolstore - say what happened in plain English instead.
        print("\nPROBLEM: something unexpected went wrong while building the")
        print("report: {}: {}".format(type(e).__name__, e))
        print("Usual causes: a source file is open in Excel (close it), or an")
        print("export is malformed (re-download it from SiteIQ) - then run again.")
        sys.exit(1)


if __name__ == "__main__":
    main()
